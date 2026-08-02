"""Raw input ingestion for generic audit reconciliation workpapers.

This module is intentionally generic. It does not know customer names,
counterparties, invoice numbers, banks, or factor operators. Engagement-specific
details must be passed through ``assumptions``.
"""

from __future__ import annotations

import sys as _bootstrap_sys

_bootstrap_sys.dont_write_bytecode = True
_bootstrap_sys.pycache_prefix = (
    r"Z:\__audit_reconciliation_no_bytecode__"
    if _bootstrap_sys.platform == "win32"
    else "/dev/null/audit-reconciliation"
)

import os as _bootstrap_os

_BOOTSTRAP_PATH = _bootstrap_os.path.join(
    _bootstrap_os.path.dirname(_bootstrap_os.path.abspath(__file__)),
    "implementation_bootstrap.py",
)
_BOOTSTRAP_NAMESPACE = {
    "__file__": _BOOTSTRAP_PATH,
    "__name__": "_audit_reconciliation_implementation_bootstrap",
}
_bootstrap_stat = _bootstrap_os.lstat(_BOOTSTRAP_PATH)
if _bootstrap_stat.st_mode & 0o170000 != 0o100000 or _bootstrap_stat.st_nlink != 1:
    raise RuntimeError(
        "implementation bootstrap must be an ordinary single-link regular file"
    )
_bootstrap_descriptor = _bootstrap_os.open(
    _BOOTSTRAP_PATH,
    _bootstrap_os.O_RDONLY | getattr(_bootstrap_os, "O_NOFOLLOW", 0),
)
try:
    _bootstrap_open_stat = _bootstrap_os.fstat(_bootstrap_descriptor)
    _bootstrap_identity = (
        _bootstrap_stat.st_dev,
        _bootstrap_stat.st_ino,
        _bootstrap_stat.st_size,
        _bootstrap_stat.st_mtime_ns,
        _bootstrap_stat.st_nlink,
    )
    if _bootstrap_identity != (
        _bootstrap_open_stat.st_dev,
        _bootstrap_open_stat.st_ino,
        _bootstrap_open_stat.st_size,
        _bootstrap_open_stat.st_mtime_ns,
        _bootstrap_open_stat.st_nlink,
    ):
        raise RuntimeError("implementation bootstrap changed before it was read")
    with _bootstrap_os.fdopen(
        _bootstrap_descriptor,
        "rb",
        closefd=False,
    ) as _bootstrap_handle:
        _bootstrap_source = _bootstrap_handle.read()
    _bootstrap_after_stat = _bootstrap_os.fstat(_bootstrap_descriptor)
    if (
        _bootstrap_identity
        != (
            _bootstrap_after_stat.st_dev,
            _bootstrap_after_stat.st_ino,
            _bootstrap_after_stat.st_size,
            _bootstrap_after_stat.st_mtime_ns,
            _bootstrap_after_stat.st_nlink,
        )
        or len(_bootstrap_source) != _bootstrap_after_stat.st_size
    ):
        raise RuntimeError("implementation bootstrap changed while it was read")
finally:
    _bootstrap_os.close(_bootstrap_descriptor)
# Execute only the pre-opened single-link bootstrap source.
exec(  # nosec B102
    compile(_bootstrap_source, _BOOTSTRAP_PATH, "exec"),
    _BOOTSTRAP_NAMESPACE,
)
_BOOTSTRAP_ROOTS = _BOOTSTRAP_NAMESPACE["activate_implementation_boundary"](
    (
        "locale_support",
        "reconciliation_helpers",
        "accountant_report",
        "review_session",
        "workpaper_outputs",
    )
)
_SCRIPTS_DIR = _bootstrap_os.path.dirname(_bootstrap_os.path.abspath(__file__))
if _SCRIPTS_DIR not in _bootstrap_sys.path:
    _bootstrap_sys.path.insert(0, _SCRIPTS_DIR)

import hashlib
import io
import json
import os
import re
import sys
import zipfile
from collections import defaultdict
from collections.abc import Mapping
from dataclasses import asdict, dataclass
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from functools import lru_cache
from pathlib import Path
from typing import Any, Callable, Iterable


def _add_vera_assurance_module_path() -> None:
    """Use the shared module root admitted by the pre-import boundary."""

    module_root = Path(str(_BOOTSTRAP_ROOTS["shared_assurance"])).parent
    if str(module_root) not in sys.path:
        sys.path.insert(0, str(module_root))


_add_vera_assurance_module_path()

from vera_assurance import (  # noqa: E402
    AssuranceContractError,
    MoneyValidationError,
    build_client_engagement_context,
    build_source_qualification,
    build_studio_client_folder_binding,
    decimal_text,
    parse_canonical_decimal,
    parse_localized_decimal,
    validate_studio_client_folder_binding,
)

try:
    import fitz  # type: ignore
except (
    Exception
) as exc:  # pragma: no cover - optional dependency import failures are reported at runtime
    fitz = None  # type: ignore
    FITZ_IMPORT_ERROR: Exception | None = exc
else:
    FITZ_IMPORT_ERROR = None

try:
    from openpyxl import load_workbook
except (
    Exception
) as exc:  # pragma: no cover - optional dependency import failures are reported at runtime
    load_workbook = None  # type: ignore
    OPENPYXL_IMPORT_ERROR: Exception | None = exc
else:
    OPENPYXL_IMPORT_ERROR = None

try:
    import pdfplumber
except (
    Exception
) as exc:  # pragma: no cover - optional dependency import failures are reported at runtime
    pdfplumber = None  # type: ignore
    PDFPLUMBER_IMPORT_ERROR: Exception | None = exc
else:
    PDFPLUMBER_IMPORT_ERROR = None

try:
    from .audit_assurance import (
        SUPPORTED_SOURCE_ADAPTER_VERSIONS,
        AssuranceRunError,
        build_reviewed_source_decisions,
        build_source_receipts,
        finalize_assurance_run,
        reviewed_date_convention,
        reviewed_money_convention,
        validate_receipt_set,
    )
    from .build_missing_evidence_requests import (
        build_missing_evidence_request_pack,
        write_missing_evidence_workbook,
    )
    from .locale_support import (
        any_keyword_in,
        configured_language,
        keyword_tuple,
        language_candidates,
        normalize_language,
    )
    from .reconciliation_helpers import (
        checks_pass,
        clean_text,
        document_key,
        parse_date,
        parse_decimal,
        reconcile_open_items,
        reconciliation_checks,
    )
    from .reconciliation_workflow import build_reconciliation_artifacts
    from .review_session import write_review_session_artifacts
except ImportError:  # pragma: no cover - direct import support
    scripts_dir = Path(__file__).resolve().parent
    if str(scripts_dir) not in sys.path:
        sys.path.insert(0, str(scripts_dir))

    from audit_assurance import (  # type: ignore
        SUPPORTED_SOURCE_ADAPTER_VERSIONS,
        AssuranceRunError,
        build_reviewed_source_decisions,
        build_source_receipts,
        finalize_assurance_run,
        reviewed_date_convention,
        reviewed_money_convention,
        validate_receipt_set,
    )
    from build_missing_evidence_requests import (  # type: ignore
        build_missing_evidence_request_pack,
        write_missing_evidence_workbook,
    )
    from locale_support import (  # type: ignore
        any_keyword_in,
        configured_language,
        keyword_tuple,
        language_candidates,
        normalize_language,
    )
    from reconciliation_helpers import (  # type: ignore
        checks_pass,
        clean_text,
        document_key,
        parse_date,
        parse_decimal,
        reconcile_open_items,
        reconciliation_checks,
    )
    from reconciliation_workflow import build_reconciliation_artifacts  # type: ignore
    from review_session import write_review_session_artifacts  # type: ignore


DATE_DMY4_RE = re.compile(r"\b\d{2}/\d{2}/\d{4}\b")
DATE_DMY2_RE = re.compile(r"\b\d{2}/\d{2}/\d{2}\b")
AMOUNT_IT_RE = re.compile(r"-?\d{1,3}(?:[.,]\d{3})*[.,]\d{2}-?|-?\d+[.,]\d{2}-?")
OPEN_ITEM_DOC_RE = re.compile(r"\b\d{2}[A-Z]{2}\d{2}/\d{3,}\b")
LEDGER_DOC_LINE_RE = re.compile(
    r"(?P<date>\d{2}/\d{2}/\d{4})\s+(?P<doc>[A-Z0-9./-]{2,})\b",
    re.I,
)
LEDGER_SETTLEMENT_RE = re.compile(
    r"\b(?:N\.?|NO\.?)\s*(?P<doc>[A-Z0-9./-]+)\s+(?:del|dated?|du|fecha)\s+(?P<date>\d{6,8})\b",
    re.I,
)
GIT_WORKSPACE_ROOT = Path(__file__).resolve().parents[3]
WORKFLOW_ID = "audit-reconciliation"


def _is_relative_to(candidate: Path, parent: Path) -> bool:
    try:
        candidate.relative_to(parent)
    except ValueError:
        return False
    return True


def validate_run_output_dir(output_dir: str | Path, *, input_dir: str | Path) -> Path:
    """Return a resolved output directory, rejecting Git/GitHub Pages locations."""

    resolved = Path(output_dir).expanduser().resolve()
    workspace_root = GIT_WORKSPACE_ROOT.resolve()
    if _is_relative_to(resolved, workspace_root):
        recommended = Path(input_dir).expanduser().resolve().parent / "output"
        raise ValueError(
            "Audit Reconciliation output_dir must be outside the Git workspace; "
            f"got {resolved}. Use a sibling output directory such as {recommended}."
        )
    return resolved


def validate_run_cache_dir(cache_dir: str | Path, *, input_dir: str | Path) -> Path:
    """Return a resolved cache directory, rejecting repo-local run caches."""

    resolved = Path(cache_dir).expanduser().resolve()
    workspace_root = GIT_WORKSPACE_ROOT.resolve()
    if _is_relative_to(resolved, workspace_root):
        recommended = (
            Path(input_dir).expanduser().resolve().parent
            / "output"
            / ".audit_reconciliation_cache"
        )
        raise ValueError(
            "Audit Reconciliation cache_dir must be outside the Git workspace; "
            f"got {resolved}. Use an output-local cache such as {recommended}."
        )
    return resolved


def load_studio_client_folder_binding(
    value: Mapping[str, Any] | str | Path,
) -> dict[str, Any]:
    """Load and validate a Studio Archive client-folder binding."""

    payload: object = value
    if isinstance(value, (str, Path)):
        try:
            payload = (
                json.loads(sys.stdin.read())
                if str(value) == "-"
                else json.loads(Path(value).expanduser().read_text(encoding="utf-8"))
            )
        except (OSError, UnicodeDecodeError, json.JSONDecodeError) as exc:
            raise ValueError(f"Client-folder binding is unreadable: {exc}") from exc
    if isinstance(payload, Mapping) and "client_folder" in payload:
        payload = payload["client_folder"]
    try:
        binding = validate_studio_client_folder_binding(payload)
    except AssuranceContractError as exc:
        raise ValueError(f"Client-folder binding is invalid: {exc}") from exc
    for label in ("archive_root", "client_root"):
        path = Path(str(binding[label]))
        try:
            current = path.lstat()
        except OSError as exc:
            raise ValueError(f"Client-folder {label} is unavailable: {exc}") from exc
        if path.is_symlink() or not path.is_dir() or current.st_nlink < 1:
            raise ValueError(f"Client-folder {label} must be a real directory")
    return binding


def prepare_client_engagement_context(
    *,
    client_folder: Mapping[str, Any] | str | Path,
    engagement_id: str,
    input_dir: str | Path,
    workspace_root: str | Path,
    run_id: str | None = None,
) -> dict[str, Any]:
    """Resolve one client-bound Audit run and its managed output directory."""

    binding = load_studio_client_folder_binding(client_folder)
    input_path = Path(input_dir).expanduser()
    try:
        input_stat = input_path.lstat()
        resolved_input = input_path.resolve(strict=True)
    except OSError as exc:
        raise ValueError(f"Audit input directory is unavailable: {exc}") from exc
    if (
        input_path.is_symlink()
        or not resolved_input.is_dir()
        or input_stat.st_nlink < 1
    ):
        raise ValueError("Audit input directory must be a real directory")
    workspace_path = Path(workspace_root).expanduser()
    if not workspace_path.is_absolute():
        raise ValueError("Audit workspace_root must be an absolute path")
    if workspace_path.is_symlink():
        raise ValueError("Audit workspace_root cannot be a symbolic link")
    resolved_workspace = workspace_path.resolve(strict=False)
    active_run_id = run_id or (
        f"{WORKFLOW_ID}-{engagement_id}-"
        + datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S%fZ")
    )
    try:
        context = build_client_engagement_context(
            studio_client_folder=binding,
            engagement_id=engagement_id,
            workflow_id=WORKFLOW_ID,
            run_id=active_run_id,
            input_dir=resolved_input,
            workspace_root=resolved_workspace,
        )
    except AssuranceContractError as exc:
        raise ValueError(f"Client engagement is invalid: {exc}") from exc
    validate_run_output_dir(context["output_dir"], input_dir=resolved_input)
    return context


JOURNAL_HEADER_RE = re.compile(
    r"^(?P<date>\d{2}/\d{2}/\d{4})\s+(?P<causale>[A-Z][A-Z ]+?)(?:\s+A\s+-|\s+\d+\s+-|\s{2,}|$)"
)
JOURNAL_ACCOUNT_RE = re.compile(
    r"^\s*(?P<line>\d{1,8})\s+(?P<account>\d+\s*/\s*\d+\s*/\s*\d+)\s+"
)
BANK_ROW_RE = re.compile(
    r"^(?P<date>\d{2}/\d{2}/\d{2})\s+"
    r"(?P<value_date>\d{2}/\d{2}/\d{2})\s+"
    r"(?P<amount>\d{1,3}(?:[.,]\d{3})*[.,]\d{2})\s+"
    r"(?P<description>.+)$"
)
PAYMENT_ORDER_HEADER_RE = re.compile(
    r"\b(?:Distinta|Payment\s+Order|Payment\s+Batch|Remittance\s+Order|Ordre\s+de\s+Paiement|Lot\s+de\s+Paiement|Orden\s+de\s+Pago|Remesa\s+de\s+Pago|Lote\s+de\s+Pago)"
    r"\s+0*(?P<batch>\d+)\s+(?:Del|Dated?|Date|Du|Fecha)\s+(?P<date>\d{2}/\d{2}/\d{4})",
    re.I,
)
PAYMENT_ORDER_TOTAL_RE = re.compile(
    r"\b(?:Totale\s+Distinta|Total\s+Payment\s+Order|Total\s+Batch|Batch\s+Total|Total\s+Ordre|Total\s+Remise|Total\s+Lot|Total\s+Orden|Total\s+Remesa|Total\s+Lote)"
    r"\s+(?P<amount>-?\d{1,3}(?:[.,]\d{3})*[.,]\d{2}-?|-?\d+[.,]\d{2}-?)",
    re.I,
)
PAYMENT_ORDER_LINE_RE = re.compile(
    r"\b(?P<counterparty_doc>\d{1,7}[-/]\d{2})\s+"
    r"(?:(?:Fattura|Invoice|Facture|Factura)\s+)?"
    r"(?P<counterparty_date>\d{2}/\d{2}/\d{4})\s+"
    r"(?P<document_no>\d{1,7}[-/]\d{2})\s+"
    r"(?P<document_date>\d{2}/\d{2}/\d{4})\s+"
    r"(?P<withholding>-?\d{1,3}(?:[.,]\d{3})*[.,]\d{2}-?|-?\d+[.,]\d{2}-?)\s+"
    r"(?P<invoice_amount>-?\d{1,3}(?:[.,]\d{3})*[.,]\d{2}-?|-?\d+[.,]\d{2}-?)\s+"
    r"(?P<withholding_amount>-?\d{1,3}(?:[.,]\d{3})*[.,]\d{2}-?|-?\d+[.,]\d{2}-?)",
    re.I,
)
PAYMENT_BATCH_RE = re.compile(
    r"\b(?:DIST(?:INTA)?\.?\s*(?:PAG(?:AMENTO|\.TO)?|PG)?\.?|PAYMENT\s+BATCH|BATCH|REMITTANCE|REMESA|LOTE|LOT)\s*(?:NR\.?|NO\.?)?\s*(?P<ref>\d{1,5}(?:\s*-\s*\d{1,5})?)",
    re.I,
)
PDF_PAGE_CACHE_VERSION = "raw_pdf_pages_v2"
OPENING_ENTRY_TERMS = (
    "apertura esercizio",
    "riapertura",
    "saldo iniziale",
    "opening balance",
    "opening entry",
    "balance brought forward",
    "a-nouveau",
    "à-nouveau",
    "apertura ejercicio",
    "saldo inicial",
)
BANK_ACCOUNT_TERMS = (
    "banca",
    "banco",
    "bank",
    "banque",
    "kontoauszug",
    "bankkonto",
    "conto corrente",
    "c/c",
)
SUPPORTED_SOURCE_ROLES = {
    "open_items",
    "counterparty_open_items",
    "ledger",
    "journal",
    "bank_statement",
    "payment_order",
    "factoring_statement",
    "compensation_support",
}
LEGACY_ADAPTER_FAMILY = "legacy_it_accounting_export_v1"
DEFAULT_SOURCE_ADAPTERS = {
    (".pdf", "open_items"): "open_items_text_v1",
    (".pdf", "bank_statement"): "bank_statement_text_v1",
    (".xlsx", "journal"): "journal_header_columns_v1",
    (".xlsm", "journal"): "journal_header_columns_v1",
    (".zip", "payment_order"): "payment_order_html_zip_v1",
}
SUPPORTED_ADAPTER_FAMILIES = {
    *SUPPORTED_SOURCE_ADAPTER_VERSIONS,
}


@dataclass
class SourcePage:
    source_file: str
    source_role: str
    source_page: int
    extraction_method: str
    text_length: int
    line_count: int
    text: str


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def source_inventory(
    input_dir: str | Path, assumptions: dict[str, Any] | None = None
) -> list[dict[str, Any]]:
    root = Path(input_dir)
    language = configured_language(assumptions, purpose="document")
    rows: list[dict[str, Any]] = []
    for path in sorted(
        p for p in root.iterdir() if p.is_file() and not p.name.startswith(".")
    ):
        resolution = resolve_source_role(
            path,
            assumptions=assumptions,
            language=language,
        )
        rows.append(
            {
                "source_file": path.name,
                "source_role": resolution["source_role"],
                "suggested_source_role": resolution["suggested_source_role"],
                "source_role_candidates": ";".join(
                    resolution["source_role_candidates"]
                ),
                "source_role_status": resolution["status"],
                "suffix": path.suffix.lower(),
                "size_bytes": path.stat().st_size,
                "sha256": sha256_file(path),
            }
        )
    return rows


def infer_source_role_candidates(
    path: str | Path,
    sample_text: str = "",
    language: object | None = None,
) -> list[str]:
    """Return filename/text role suggestions without selecting source meaning.

    Keyword routing is deterministic only as a reviewer aid. Source-role
    selection is semantic and must come from ``reviewed_source_roles`` before
    a parser may emit accounting rows.
    """

    name = Path(path).name.lower()
    text = f"{name} {sample_text}".lower()
    candidates: list[str] = []
    for role in (
        "open_items",
        "bank_statement",
        "journal",
        "payment_order",
        "ledger",
        "factoring_statement",
    ):
        if any(
            any_keyword_in(text, keyword_tuple(candidate, "role_keywords", role))
            for candidate in language_candidates(language)
        ):
            candidates.append(role)
    if "factoring_statement" not in candidates and any(
        any_keyword_in(text, keyword_tuple(candidate, "evidence_keywords", "factoring"))
        for candidate in language_candidates(language)
    ):
        candidates.append("factoring_statement")
    return candidates


def infer_source_role(
    path: str | Path, sample_text: str = "", language: object | None = None
) -> str:
    """Return an unambiguous advisory suggestion, never a reviewed role."""

    candidates = infer_source_role_candidates(path, sample_text, language)
    return candidates[0] if len(candidates) == 1 else "unknown"


def _reviewed_source_decision(
    assumptions: dict[str, Any] | None,
    path: str | Path,
) -> dict[str, Any] | None:
    """Return a prevalidated current-source decision generated at intake."""

    decisions = (assumptions or {}).get("_reviewed_source_decision_receipts")
    if not isinstance(decisions, dict):
        return None
    source_path = Path(path)
    for key in (source_path.name, source_path.as_posix(), str(source_path)):
        value = decisions.get(key)
        if isinstance(value, dict):
            return value
    return None


def resolve_source_role(
    path: str | Path,
    *,
    assumptions: dict[str, Any] | None = None,
    sample_text: str = "",
    language: object | None = None,
) -> dict[str, Any]:
    """Resolve a source role only from reviewed input and retain suggestions."""

    candidates = infer_source_role_candidates(path, sample_text, language)
    decision = _reviewed_source_decision(assumptions, path)
    content = decision.get("content") if isinstance(decision, dict) else None
    reviewed_role = (
        str(content.get("role") or "").lower() if isinstance(content, dict) else ""
    )
    if reviewed_role in SUPPORTED_SOURCE_ROLES:
        return {
            "source_role": reviewed_role,
            "suggested_source_role": (
                candidates[0] if len(candidates) == 1 else "unknown"
            ),
            "source_role_candidates": candidates,
            "status": "reviewed",
        }
    return {
        "source_role": "unknown",
        "suggested_source_role": candidates[0] if len(candidates) == 1 else "unknown",
        "source_role_candidates": candidates,
        "status": (
            "needs_review"
            if candidates or not reviewed_role
            else "unsupported_source_layout"
        ),
    }


def source_adapter_family(
    path: str | Path,
    source_role: str,
    assumptions: dict[str, Any] | None,
) -> str:
    """Return an explicit legacy adapter or a mechanically bounded default."""

    decision = _reviewed_source_decision(assumptions, path)
    content = decision.get("content") if isinstance(decision, dict) else None
    if isinstance(content, dict):
        reviewed = str(content.get("adapter_family") or "")
        if reviewed:
            return reviewed
    return ""


def _requested_source_adapter_families(
    paths: Iterable[Path],
    assumptions: dict[str, Any],
) -> dict[str, str]:
    """Read adapter choices from explicit decisions without granting authority."""

    supplied = assumptions.get("reviewed_source_decisions")
    mapping = supplied if isinstance(supplied, dict) else {}
    families: dict[str, str] = {}
    for path in paths:
        value = mapping.get(path.name)
        if not isinstance(value, dict):
            value = mapping.get(path.as_posix())
        if not isinstance(value, dict):
            families[path.name] = ""
        elif value.get("schema_version") == "vera.reviewed_decision_receipt.v1":
            families[path.name] = str(value.get("adapter_id") or "")
        else:
            families[path.name] = str(value.get("adapter_family") or "")
    return families


def build_file_source_qualification(
    path: Path,
    *,
    resolution: dict[str, Any],
    adapter_family: str,
    reviewed_decision: dict[str, Any] | None,
    source_artifact_ref: str,
    candidate_row_count: int,
    emitted_row_count: int,
) -> dict[str, Any]:
    """Build the fail-closed source record consumed by Vera assurance gates."""

    source_ref = source_artifact_ref
    reviewed = resolution["status"] == "reviewed"
    supported_adapter = adapter_family in SUPPORTED_ADAPTER_FAMILIES
    layout_supported = candidate_row_count > 0
    if adapter_family and not supported_adapter:
        status = "unsupported_source_layout"
    elif not reviewed:
        status = (
            "unsupported_source_layout"
            if resolution["status"] == "unsupported_source_layout"
            else "needs_review"
        )
    elif not supported_adapter or not layout_supported:
        status = "unsupported_source_layout"
    else:
        status = "qualified"

    def control(
        control_id: str,
        control_status: str,
        detail: str,
    ) -> dict[str, Any]:
        return {
            "control_id": control_id,
            "required": True,
            "status": control_status,
            "evidence_refs": [source_ref],
            "detail": detail,
        }

    if status == "needs_review":
        role_control = "not_assessed"
        adapter_control = "not_assessed"
        layout_control = "not_assessed"
    else:
        role_control = "passed" if reviewed else "failed"
        adapter_control = "passed" if supported_adapter else "failed"
        layout_control = "passed" if layout_supported else "failed"
    limitations = []
    if not reviewed:
        limitations.append(
            "Filename and text role suggestions are advisory; a reviewed source role is required."
        )
    if adapter_family and not supported_adapter:
        limitations.append(
            "No supported source-layout adapter was declared or detected."
        )
    if reviewed and supported_adapter and not layout_supported:
        limitations.append(
            "The declared adapter emitted no mechanically qualified rows."
        )
    return build_source_qualification(
        qualification_id=f"qualification.{sha256_file(path)[:20]}",
        adapter_id=adapter_family or "unresolved.adapter",
        adapter_version="2",
        source_family=(
            f"{path.suffix.lower().lstrip('.') or 'file'}."
            f"{resolution['source_role']}"
        ),
        status=status,
        source_artifact_refs=[source_ref],
        controls=[
            control(
                "reviewed_source_role",
                role_control,
                f"Reviewed role: {resolution['source_role']}.",
            ),
            control(
                "supported_layout_adapter",
                adapter_control,
                f"Adapter family: {adapter_family or 'unresolved'}.",
            ),
            control(
                "qualified_rows_emitted",
                layout_control,
                f"Candidate rows: {candidate_row_count}.",
            ),
        ],
        candidate_row_count=candidate_row_count,
        emitted_row_count=emitted_row_count if status == "qualified" else 0,
        reviewed_mapping_ref=(
            str(reviewed_decision["decision_id"]) if reviewed_decision else None
        ),
        limitations=limitations,
    )


def configure_ocr_environment(cache_dir: Path) -> None:
    (cache_dir / "paddlex").mkdir(parents=True, exist_ok=True)
    (cache_dir / "matplotlib").mkdir(parents=True, exist_ok=True)
    os.environ.setdefault("PADDLE_PDX_CACHE_HOME", str(cache_dir / "paddlex"))
    # Deterministic source selection: BOS is Paddle's direct model host and
    # avoids Hugging Face/Xet range failures observed during OCR bootstrap.
    os.environ.setdefault("PADDLE_PDX_MODEL_SOURCE", "bos")
    os.environ.setdefault("PADDLE_PDX_DISABLE_MODEL_SOURCE_CHECK", "True")
    os.environ.setdefault("MPLCONFIGDIR", str(cache_dir / "matplotlib"))
    os.environ.setdefault("FLAGS_use_mkldnn", "0")


def _ocr_language(language: object | None) -> str:
    text = clean_text(language).lower().replace("_", "-")
    code = text.split("-", 1)[0]
    return code if code in {"de", "en", "fr", "it", "es"} else "en"


def _shared_ocr_text_from_image_bytes(
    image_bytes: bytes,
    *,
    lang: str,
    text_recognition_model_name: str | None = None,
) -> str | None:
    try:
        from modules.slides.ocr import (  # type: ignore
            extract_raw_ocr_from_image_bytes,
            extract_text_from_raw_ocr_result,
        )
    except Exception:
        return None

    raw = extract_raw_ocr_from_image_bytes(
        image_bytes,
        lang=lang,
        preprocess_profile="document_scan",
        allow_preprocess_fallback=True,
        text_recognition_model_name=text_recognition_model_name,
    )
    return extract_text_from_raw_ocr_result(raw)


@lru_cache(maxsize=8)
def _get_local_paddle_ocr(
    lang: str, text_recognition_model_name: str | None = None
) -> object:
    try:
        from paddleocr import PaddleOCR  # type: ignore
    except Exception as exc:  # pragma: no cover - depends on optional OCR install
        raise RuntimeError(
            "PaddleOCR is required for scanned PDF OCR. Install the plugin "
            "optional OCR dependencies from requirements-ocr.txt."
        ) from exc

    modern_kwargs: dict[str, object] = {
        "lang": lang,
        "use_doc_orientation_classify": False,
        "use_doc_unwarping": False,
        "use_textline_orientation": False,
    }
    if text_recognition_model_name:
        modern_kwargs["text_recognition_model_name"] = text_recognition_model_name
    try:
        return PaddleOCR(**modern_kwargs)
    except TypeError:
        legacy_kwargs: dict[str, object] = {
            "lang": lang,
            "show_log": False,
            "use_angle_cls": False,
        }
        return PaddleOCR(**legacy_kwargs)


def _raw_ocr_text(raw: object) -> str:
    texts: list[str] = []

    def collect(value: object) -> None:
        if value is None:
            return
        if isinstance(value, dict):
            for key in ("rec_texts", "texts"):
                nested = value.get(key)
                if isinstance(nested, list):
                    for item in nested:
                        if isinstance(item, str) and clean_text(item):
                            texts.append(clean_text(item))
                    return
            text = value.get("text")
            if isinstance(text, str) and clean_text(text):
                texts.append(clean_text(text))
                return
            for nested in value.values():
                collect(nested)
            return
        if isinstance(value, (list, tuple)):
            if (
                len(value) >= 2
                and isinstance(value[1], (list, tuple))
                and value[1]
                and isinstance(value[1][0], str)
            ):
                text = clean_text(value[1][0])
                if text:
                    texts.append(text)
                return
            for nested in value:
                collect(nested)

    collect(raw)
    return "\n".join(texts)


def _local_paddle_ocr_text_from_image_bytes(
    image_bytes: bytes,
    *,
    lang: str,
    text_recognition_model_name: str | None = None,
) -> str:
    import numpy as np  # type: ignore
    from PIL import Image  # type: ignore

    image = Image.open(io.BytesIO(image_bytes)).convert("RGB")
    engine = _get_local_paddle_ocr(
        lang, text_recognition_model_name=text_recognition_model_name
    )
    image_array = np.asarray(image)
    if hasattr(engine, "ocr"):
        raw = engine.ocr(image_array, cls=True)
    elif hasattr(engine, "predict"):
        raw = engine.predict(image_array)
    else:
        raise RuntimeError("No compatible PaddleOCR inference method is available.")
    return _raw_ocr_text(raw)


def _ocr_page_text(
    pdf_path: Path,
    page_index: int,
    cache_dir: Path,
    dpi_scale: float = 2.0,
    language: object | None = None,
) -> str:
    if fitz is None:
        detail = f": {FITZ_IMPORT_ERROR}" if FITZ_IMPORT_ERROR else ""
        raise RuntimeError(
            "PyMuPDF (fitz) is required for OCR on scanned PDFs. "
            "Install the plugin base dependencies from requirements.txt"
            f"{detail}"
        )
    configure_ocr_environment(cache_dir)
    doc = fitz.open(pdf_path)
    page = doc[page_index]
    pix = page.get_pixmap(matrix=fitz.Matrix(dpi_scale, dpi_scale), alpha=False)
    image_bytes = pix.tobytes("png")
    lang = _ocr_language(language)
    shared_text = _shared_ocr_text_from_image_bytes(
        image_bytes,
        lang=lang,
        text_recognition_model_name="PP-OCRv5_server_rec",
    )
    if shared_text is not None:
        return shared_text
    return _local_paddle_ocr_text_from_image_bytes(
        image_bytes,
        lang=lang,
        text_recognition_model_name="PP-OCRv5_server_rec",
    )


def _pdf_page_cache_path(
    path: Path, cache_dir: Path, *, ocr_scanned: bool, dpi_scale: float
) -> Path:
    content_hash = sha256_file(path)
    cache_key = hashlib.sha256(
        json.dumps(
            {
                "version": PDF_PAGE_CACHE_VERSION,
                "source_file": path.name,
                "content_sha256": content_hash,
                "ocr_scanned": ocr_scanned,
                "dpi_scale": dpi_scale,
            },
            sort_keys=True,
        ).encode("utf-8")
    ).hexdigest()
    return cache_dir / "pdf_pages" / f"{cache_key}.json"


def _read_pdf_page_cache(cache_path: Path, source_name: str) -> list[SourcePage] | None:
    if not cache_path.exists():
        return None
    payload = json.loads(cache_path.read_text(encoding="utf-8"))
    if payload.get("version") != PDF_PAGE_CACHE_VERSION:
        return None
    rows = payload.get("pages")
    if not isinstance(rows, list):
        return None
    pages: list[SourcePage] = []
    for row in rows:
        if not isinstance(row, dict):
            return None
        page = SourcePage(
            **{field: row.get(field, "") for field in SourcePage.__dataclass_fields__}
        )
        page.source_file = source_name
        pages.append(page)
    return pages


def _write_pdf_page_cache(cache_path: Path, pages: list[SourcePage]) -> None:
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = cache_path.with_suffix(".tmp")
    tmp_path.write_text(
        json.dumps(
            {
                "version": PDF_PAGE_CACHE_VERSION,
                "pages": [asdict(page) for page in pages],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    tmp_path.replace(cache_path)


def extract_pdf_pages(
    path: Path,
    cache_dir: Path,
    *,
    source_role: str = "unknown",
    ocr_scanned: bool = True,
    use_cache: bool = True,
    dpi_scale: float = 2.0,
    language: object | None = None,
    progress_every_pages: int = 10,
    progress_callback: Callable[[dict[str, Any]], None] | None = None,
) -> list[SourcePage]:
    if pdfplumber is None:
        detail = f": {PDFPLUMBER_IMPORT_ERROR}" if PDFPLUMBER_IMPORT_ERROR else ""
        raise RuntimeError(f"pdfplumber is required for PDF extraction{detail}")
    cache_path = _pdf_page_cache_path(
        path, cache_dir, ocr_scanned=ocr_scanned, dpi_scale=dpi_scale
    )
    if use_cache:
        cached_pages = _read_pdf_page_cache(cache_path, path.name)
        if cached_pages is not None:
            for page in cached_pages:
                page.source_role = source_role
            if progress_callback:
                progress_callback(
                    {
                        "event": "pdf_cache_hit",
                        "source_file": path.name,
                        "page_count": len(cached_pages),
                    }
                )
            return cached_pages
    pages: list[SourcePage] = []
    with pdfplumber.open(path) as pdf:
        page_count = len(pdf.pages)
        if progress_callback:
            progress_callback(
                {
                    "event": "pdf_file_start",
                    "source_file": path.name,
                    "page_count": page_count,
                }
            )
        total_text_length = 0
        ocr_page_count = 0
        progress_every = max(1, int(progress_every_pages or 1))
        for index, page in enumerate(pdf.pages):
            text = page.extract_text() or ""
            method = "pdf_text"
            if ocr_scanned and len(text.strip()) < 40:
                ocr_page_count += 1
                if progress_callback:
                    progress_callback(
                        {
                            "event": "ocr_page_start",
                            "source_file": path.name,
                            "source_page": index + 1,
                            "page_count": page_count,
                        }
                    )
                text = _ocr_page_text(
                    path,
                    index,
                    cache_dir,
                    dpi_scale=dpi_scale,
                    language=language,
                )
                method = "paddle_ocr"
                if progress_callback:
                    progress_callback(
                        {
                            "event": "ocr_page_done",
                            "source_file": path.name,
                            "source_page": index + 1,
                            "page_count": page_count,
                            "text_length": len(text),
                        }
                    )
            lines = [line for line in text.splitlines() if clean_text(line)]
            total_text_length += len(text)
            pages.append(
                SourcePage(
                    source_file=path.name,
                    source_role=source_role,
                    source_page=index + 1,
                    extraction_method=method,
                    text_length=len(text),
                    line_count=len(lines),
                    text=text,
                )
            )
            source_page = index + 1
            if progress_callback and (
                source_page == page_count
                or source_page == 1
                or source_page % progress_every == 0
            ):
                progress_callback(
                    {
                        "event": "pdf_page_done",
                        "source_file": path.name,
                        "source_page": source_page,
                        "page_count": page_count,
                        "extraction_method": method,
                        "text_length": len(text),
                        "line_count": len(lines),
                    }
                )
        if progress_callback:
            progress_callback(
                {
                    "event": "pdf_file_done",
                    "source_file": path.name,
                    "page_count": page_count,
                    "ocr_page_count": ocr_page_count,
                    "text_length": total_text_length,
                }
            )
    if use_cache:
        _write_pdf_page_cache(cache_path, pages)
    return pages


def money_convention_for_source(
    assumptions: dict[str, Any],
    source_file: str,
) -> dict[str, Any] | None:
    """Return the reviewed convention bound to one exact source artifact."""

    return reviewed_money_convention(
        _reviewed_source_decision(assumptions, source_file)
    )


def date_convention_for_source(
    assumptions: dict[str, Any],
    source_file: str,
) -> dict[str, Any] | None:
    """Return the source-bound reviewed date order."""

    return reviewed_date_convention(_reviewed_source_decision(assumptions, source_file))


def source_perimeter_for_source(
    assumptions: dict[str, Any],
    source_file: str,
) -> dict[str, Any] | None:
    """Return the exact reviewed accounting perimeter for one source."""

    decision = _reviewed_source_decision(assumptions, source_file)
    content = decision.get("content") if isinstance(decision, dict) else None
    perimeter = content.get("perimeter") if isinstance(content, dict) else None
    return dict(perimeter) if isinstance(perimeter, dict) else None


def apply_source_perimeter(
    rows: Iterable[dict[str, Any]],
    *,
    source_file: str,
    assumptions: dict[str, Any],
) -> list[dict[str, Any]]:
    """Bind prepared rows to the reviewed entity/party/value perimeter."""

    perimeter = source_perimeter_for_source(assumptions, source_file)
    money = money_convention_for_source(assumptions, source_file)
    if perimeter is None or money is None:
        return []
    normalized = []
    for row in rows:
        current = dict(row)
        current["entity_ref"] = perimeter["entity_ref"]
        current["party_ref"] = perimeter["party_ref"]
        current["currency"] = perimeter["currency"]
        current["unit"] = perimeter["unit"]
        current["direction_policy"] = perimeter["direction_policy"]
        current["allocation_policy"] = perimeter["allocation_policy"]
        current["reported_unit"] = money["reported_unit"]
        current["reported_increment"] = money["reported_increment"]
        normalized.append(current)
    return normalized


def apply_source_perimeters(
    rows: Iterable[dict[str, Any]],
    assumptions: dict[str, Any],
) -> list[dict[str, Any]]:
    """Apply each row's exact reviewed source perimeter."""

    normalized = []
    for row in rows:
        source_file = clean_text(row.get("source_file")).split("!", 1)[0]
        normalized.extend(
            apply_source_perimeter(
                [row],
                source_file=source_file,
                assumptions=assumptions,
            )
        )
    return normalized


def parse_money(
    value: object,
    *,
    convention: dict[str, Any] | None = None,
) -> Decimal | None:
    """Parse exact money, abstaining on floats or ambiguous punctuation."""

    text = clean_text(value)
    if text.endswith("-"):
        parsed = parse_money(text[:-1], convention=convention)
        return -parsed if parsed is not None else None
    if isinstance(value, float):
        return None
    try:
        parsed = parse_localized_decimal(
            value,
            decimal_separator=(
                convention.get("decimal_separator") if convention else None
            ),
            thousands_separator=(
                convention.get("thousands_separator") if convention else None
            ),
            allow_float=False,
        )
        if convention is not None and "reported_increment" in convention:
            if convention["reported_increment"] != "0.01":
                return None
            increment = parse_canonical_decimal(
                convention["reported_increment"],
                label="reported_increment",
            )
            if increment <= 0 or parsed % increment:
                return None
        return parsed
    except MoneyValidationError:
        return None


def amount_string(
    value: object,
    *,
    convention: dict[str, Any] | None = None,
) -> str:
    parsed = parse_money(value, convention=convention)
    return decimal_text(parsed) if parsed is not None else ""


def iso_date(
    value: object,
    *,
    convention: dict[str, Any] | None = None,
) -> str:
    """Parse native/ISO dates or a reviewed day/month source order."""

    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean_text(value)
    if not text:
        return ""
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
        try:
            return datetime.strptime(text, "%Y-%m-%d").date().isoformat()
        except ValueError:
            return ""
    order = convention.get("order") if isinstance(convention, dict) else None
    if order not in {"day_first", "month_first"}:
        return ""
    date_format = "%d/%m/%Y" if order == "day_first" else "%m/%d/%Y"
    short_format = "%d/%m/%y" if order == "day_first" else "%m/%d/%y"
    compact_format = "%d%m%Y" if order == "day_first" else "%m%d%Y"
    compact_short = "%d%m%y" if order == "day_first" else "%m%d%y"
    normalized = text.replace(".", "/").replace("-", "/")
    for pattern, date_pattern in (
        (r"\d{1,2}/\d{1,2}/\d{4}", date_format),
        (r"\d{1,2}/\d{1,2}/\d{2}", short_format),
        (r"\d{8}", compact_format),
        (r"\d{6}", compact_short),
    ):
        if re.fullmatch(pattern, normalized):
            try:
                return datetime.strptime(normalized, date_pattern).date().isoformat()
            except ValueError:
                return ""
    return ""


def normalize_open_item_document(raw_doc: str, doc_date: str) -> str:
    text = clean_text(raw_doc).upper().replace(" ", "")
    match = re.match(r"^(?P<yy>\d{2})(?P<kind>[A-Z]{2})\d{2}/0*(?P<num>\d+)$", text)
    if match:
        return document_key(
            f"{int(match.group('num'))}-{match.group('kind')}", doc_date
        )
    return document_key(text, doc_date)


def parse_open_items(
    pages: list[SourcePage], assumptions: dict[str, Any]
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for page in pages:
        if page.source_role != "open_items":
            continue
        money_convention = money_convention_for_source(assumptions, page.source_file)
        date_convention = date_convention_for_source(
            assumptions,
            page.source_file,
        )
        perimeter = source_perimeter_for_source(assumptions, page.source_file)
        side = str((perimeter or {}).get("direction_policy") or "")
        if side not in {
            "customer",
            "supplier",
            "receivable",
            "payable",
            "debit",
            "credit",
        }:
            continue
        lines = [
            clean_text(line) for line in page.text.splitlines() if clean_text(line)
        ]
        line_index = 0
        while line_index < len(lines):
            line = lines[line_index]
            if not OPEN_ITEM_DOC_RE.fullmatch(line):
                line_index += 1
                continue
            doc_no = line
            doc_date = ""
            amount = ""
            balance = ""
            if line_index + 3 < len(lines):
                doc_date = iso_date(
                    lines[line_index + 1],
                    convention=date_convention,
                )
                if clean_text(lines[line_index + 1]) and not doc_date:
                    return []
                amount = amount_string(
                    lines[line_index + 2],
                    convention=money_convention,
                )
                balance = amount_string(
                    lines[line_index + 3],
                    convention=money_convention,
                )
            if doc_date and amount:
                document_no = doc_no
                rows.append(
                    {
                        "record_id": f"open:{page.source_file}:p{page.source_page}:l{line_index + 1}",
                        "source_file": page.source_file,
                        "source_page": page.source_page,
                        "source_row": line_index + 1,
                        "source_value_row": line_index + 3,
                        "source_role": "open_items",
                        "source_side": side,
                        "expected_side": side,
                        "document_no": document_no,
                        "document_date": doc_date,
                        "posting_date": doc_date,
                        "amount": amount,
                        "balance": balance or amount,
                        "currency": assumptions.get("currency", "EUR"),
                        "description": doc_no,
                        "evidence_type": "open_item",
                        "document_key": normalize_open_item_document(doc_no, doc_date),
                    }
                )
                line_index += 4
                continue
            line_index += 1
    return rows


def parse_ledger_or_factoring_pages(
    pages: list[SourcePage], assumptions: dict[str, Any]
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    counterparty_keywords = [
        str(v).lower() for v in assumptions.get("counterparty_keywords", [])
    ]
    factor_keywords = [
        str(v).lower() for v in assumptions.get("factoring_operator_keywords", [])
    ]
    language = configured_language(assumptions, purpose="document")
    invoice_terms = keyword_tuple(language, "evidence_keywords", "invoice")
    closure_terms = keyword_tuple(language, "evidence_keywords", "closure")
    compensation_terms = keyword_tuple(
        language, "evidence_keywords", "compensation"
    ) + keyword_tuple(language, "evidence_keywords", "netting")
    for page in pages:
        if page.source_role not in {"ledger", "factoring_statement"}:
            continue
        if (
            source_adapter_family(page.source_file, page.source_role, assumptions)
            != LEGACY_ADAPTER_FAMILY
        ):
            continue
        money_convention = money_convention_for_source(assumptions, page.source_file)
        date_convention = date_convention_for_source(
            assumptions,
            page.source_file,
        )
        lines = [
            clean_text(line) for line in page.text.splitlines() if clean_text(line)
        ]
        current_header = ""
        for idx, line in enumerate(lines, start=1):
            lower = line.lower()
            if any_keyword_in(
                lower, invoice_terms + closure_terms + compensation_terms
            ):
                current_header = line
            doc_match = LEDGER_DOC_LINE_RE.search(line) or LEDGER_SETTLEMENT_RE.search(
                line
            )
            if not doc_match:
                continue
            doc_no = clean_text(doc_match.group("doc"))
            doc_date = iso_date(
                doc_match.group("date"),
                convention=date_convention,
            )
            if not doc_date:
                return []
            amounts = AMOUNT_IT_RE.findall(line)
            source_value_row = idx
            if not amounts and idx < len(lines):
                amounts = AMOUNT_IT_RE.findall(lines[idx])
                source_value_row = idx + 1
            amount = amount_string(
                amounts[-2] if len(amounts) >= 2 else (amounts[-1] if amounts else ""),
                convention=money_convention,
            )
            if not amount:
                continue
            text_window = " ".join(lines[max(0, idx - 3) : min(len(lines), idx + 3)])
            classification_text = f"{current_header} {line}".lower()
            evidence_type = "internal_booking"
            if any_keyword_in(classification_text, closure_terms + compensation_terms):
                evidence_type = "internal_closure"
            if any_keyword_in(classification_text, compensation_terms):
                evidence_type = "compensation"
            if page.source_role == "factoring_statement":
                evidence_type = "external_factoring"
            elif any(
                keyword and keyword in text_window.lower()
                for keyword in factor_keywords
            ):
                evidence_type = "factoring_bridge"
            counterparty_context = f"{page.source_file} {text_window}".lower()
            if counterparty_keywords and not any(
                keyword in counterparty_context for keyword in counterparty_keywords
            ):
                if evidence_type == "internal_booking":
                    continue
            rows.append(
                {
                    "record_id": f"evidence:{page.source_file}:p{page.source_page}:l{idx}",
                    "source_file": page.source_file,
                    "source_page": page.source_page,
                    "source_row": idx,
                    "source_value_row": source_value_row,
                    "source_role": page.source_role,
                    "document_no": doc_no,
                    "document_date": doc_date,
                    "posting_date": doc_date,
                    "amount": amount,
                    "currency": assumptions.get("currency", "EUR"),
                    "description": text_window or current_header,
                    "evidence_type": evidence_type,
                    "document_key": document_key(doc_no, doc_date),
                }
            )
    return rows


LEDGER_BALANCE_RE = re.compile(
    r"\b20\d{2}\s+"
    r"(?P<amount>-?\d{1,3}(?:[.,]\d{3})*[.,]\d{2}-?|-?\d+[.,]\d{2}-?)\s+"
    r"(?P<balance>-?\d{1,3}(?:[.,]\d{3})*[.,]\d{2}-?|-?\d+[.,]\d{2}-?)\s*"
    r"(?P<sign>[+-])"
)


def parse_ledger_account_header(text: str) -> tuple[str, str]:
    for line in text.splitlines():
        match = re.search(
            r"\bConto:\s*(?P<account>\d+\s*/\s*\d+\s*/\s*\d+)\s+(?P<name>.+)$",
            clean_text(line),
            re.I,
        )
        if match:
            return clean_text(match.group("account")), clean_text(match.group("name"))
    return "", ""


def signed_balance(amount: Decimal, sign: str) -> Decimal:
    return amount if sign == "+" else -amount


def last_ledger_balance(
    text: str,
    *,
    convention: dict[str, Any] | None = None,
) -> tuple[Decimal | None, str]:
    matches = list(LEDGER_BALANCE_RE.finditer(text))
    if not matches:
        return None, ""
    match = matches[-1]
    balance = parse_money(match.group("balance"), convention=convention)
    if balance is None:
        return None, ""
    sign = match.group("sign")
    return signed_balance(balance, sign), sign


def first_ledger_balance_after(
    text: str,
    marker: str,
    *,
    convention: dict[str, Any] | None = None,
) -> tuple[Decimal | None, str]:
    lower = text.lower()
    idx = lower.find(marker.lower())
    if idx < 0:
        return None, ""
    match = LEDGER_BALANCE_RE.search(text[idx : idx + 800])
    if not match:
        return None, ""
    balance = parse_money(match.group("balance"), convention=convention)
    if balance is None:
        return None, ""
    sign = match.group("sign")
    return signed_balance(balance, sign), sign


def parse_ledger_balance_pages(
    pages: list[SourcePage], assumptions: dict[str, Any]
) -> list[dict[str, Any]]:
    counterparty_keywords = [
        clean_text(keyword).lower()
        for keyword in assumptions.get("counterparty_keywords", [])
        if clean_text(keyword)
    ]
    grouped: dict[tuple[str, str, str], list[SourcePage]] = defaultdict(list)
    for page in pages:
        if page.source_role != "ledger":
            continue
        if (
            source_adapter_family(page.source_file, page.source_role, assumptions)
            != LEGACY_ADAPTER_FAMILY
        ):
            continue
        account, account_name = parse_ledger_account_header(page.text)
        if not account:
            continue
        if counterparty_keywords and not any(
            keyword in account_name.lower() for keyword in counterparty_keywords
        ):
            continue
        grouped[(page.source_file, account, account_name)].append(page)

    rows: list[dict[str, Any]] = []
    for (source_file, account, account_name), account_pages in grouped.items():
        account_pages = sorted(account_pages, key=lambda item: item.source_page)
        text = "\n".join(page.text for page in account_pages)
        money_convention = money_convention_for_source(assumptions, source_file)
        opening, opening_sign = first_ledger_balance_after(
            text,
            "apertura esercizio",
            convention=money_convention,
        )
        lower = text.lower()
        cutoff_idx = lower.find("chiusura esercizio")
        if cutoff_idx < 0:
            cutoff_idx = lower.find("dare avere totali")
        closing_text = text[:cutoff_idx] if cutoff_idx >= 0 else text
        closing, closing_sign = last_ledger_balance(
            closing_text,
            convention=money_convention,
        )
        if opening is None and closing is None:
            continue
        rows.append(
            {
                "source_file": source_file,
                "source_role": "ledger",
                "source_pages": f"{account_pages[0].source_page}-{account_pages[-1].source_page}",
                "account": account,
                "account_name": account_name,
                "opening_balance_signed_debit_minus_credit": f"{(opening or Decimal('0.00')):.2f}",
                "opening_balance_sign": opening_sign,
                "closing_balance_signed_debit_minus_credit": f"{(closing or Decimal('0.00')):.2f}",
                "closing_balance_sign": closing_sign,
                "currency": assumptions.get("currency", "EUR"),
                "basis": "Ledger opening and last running balance before closing/totals.",
            }
        )

    total_opening = sum(
        (
            parse_money(row["opening_balance_signed_debit_minus_credit"])
            or Decimal("0.00")
        )
        for row in rows
    )
    total_closing = sum(
        (
            parse_money(row["closing_balance_signed_debit_minus_credit"])
            or Decimal("0.00")
        )
        for row in rows
    )
    if rows:
        rows.insert(
            0,
            {
                "source_file": "TOTAL",
                "source_role": "ledger",
                "source_pages": "",
                "account": "TOTAL",
                "account_name": "All matched counterparty ledgers",
                "opening_balance_signed_debit_minus_credit": f"{total_opening:.2f}",
                "opening_balance_sign": "+" if total_opening >= 0 else "-",
                "closing_balance_signed_debit_minus_credit": f"{total_closing:.2f}",
                "closing_balance_sign": "+" if total_closing >= 0 else "-",
                "currency": assumptions.get("currency", "EUR"),
                "basis": "Sum of matched ledger balances.",
            },
        )
    return rows


def parse_bank_statement_pages(
    pages: list[SourcePage], assumptions: dict[str, Any]
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    counterparty_keywords = [
        str(v).lower() for v in assumptions.get("counterparty_keywords", [])
    ]
    factor_keywords = [
        str(v).lower() for v in assumptions.get("factoring_operator_keywords", [])
    ]
    for page in pages:
        if page.source_role != "bank_statement":
            continue
        lines = [
            clean_text(line) for line in page.text.splitlines() if clean_text(line)
        ]
        current = ""
        current_start = 0
        for idx, line in enumerate(lines, start=1):
            if DATE_DMY2_RE.match(line):
                if current:
                    rows.extend(
                        _bank_row_from_text(
                            page,
                            current,
                            current_start,
                            assumptions,
                            counterparty_keywords,
                            factor_keywords,
                        )
                    )
                current = line
                current_start = idx
            elif current:
                current += " " + line
        if current:
            rows.extend(
                _bank_row_from_text(
                    page,
                    current,
                    current_start,
                    assumptions,
                    counterparty_keywords,
                    factor_keywords,
                )
            )
    return rows


def _bank_row_from_text(
    page: SourcePage,
    text: str,
    source_row: int,
    assumptions: dict[str, Any],
    counterparty_keywords: list[str],
    factor_keywords: list[str] | None = None,
) -> list[dict[str, Any]]:
    match = BANK_ROW_RE.search(text)
    if not match:
        return []
    description = clean_text(match.group("description"))
    factor_keywords = factor_keywords or []
    lower_description = description.lower()
    if (
        counterparty_keywords
        and not any(keyword in lower_description for keyword in counterparty_keywords)
        and not any(keyword in lower_description for keyword in factor_keywords)
    ):
        return []
    date_convention = date_convention_for_source(
        assumptions,
        page.source_file,
    )
    posting_date = iso_date(
        match.group("date"),
        convention=date_convention,
    )
    value_date = iso_date(
        match.group("value_date"),
        convention=date_convention,
    )
    if not posting_date or not value_date:
        return []
    amount = amount_string(
        match.group("amount"),
        convention=money_convention_for_source(assumptions, page.source_file),
    )
    if not amount:
        return []
    batch_ids = extract_payment_batch_ids(description)
    doc_refs = extract_invoice_refs(
        description,
        posting_date,
        date_convention=date_convention,
    )
    if not doc_refs:
        doc_refs = [("", "")]
    rows: list[dict[str, Any]] = []
    for doc_no, doc_date in doc_refs:
        rows.append(
            {
                "record_id": f"bank:{page.source_file}:p{page.source_page}:l{source_row}:{doc_no or 'unallocated'}",
                "source_file": page.source_file,
                "source_page": page.source_page,
                "source_row": source_row,
                "source_role": "bank_statement",
                "document_no": doc_no,
                "document_date": doc_date,
                "posting_date": posting_date,
                "value_date": value_date,
                "amount": amount,
                "bank_amount": amount,
                "batch_id": batch_ids[0] if batch_ids else "",
                "batch_ids": ";".join(batch_ids),
                "group_id": batch_ids[0] if batch_ids else "",
                "group_ids": ";".join(batch_ids),
                "currency": assumptions.get("currency", "EUR"),
                "description": description,
                "evidence_type": (
                    "external_bank" if doc_no else "unallocated_external_bank"
                ),
                "document_key": document_key(doc_no, doc_date) if doc_no else "",
            }
        )
    return rows


def extract_invoice_refs(
    text: str,
    fallback_date: str = "",
    *,
    date_convention: dict[str, Any] | None = None,
) -> list[tuple[str, str]]:
    refs: list[tuple[str, str]] = []
    for match in re.finditer(
        r"\b(?:N\.?|NO\.?|NUM(?:ERO)?|FATT\.?|FATTURA|INVOICE|INV\.?|FACTURE|FACTURA)?\s*"
        r"(\d{1,7}(?:[-/][A-Z0-9]{1,8})?)\s+(?:del|dated?|du|fecha)\s+(\d{6,8})\b",
        text,
        re.I,
    ):
        parsed_date = iso_date(
            match.group(2),
            convention=date_convention,
        )
        if parsed_date:
            refs.append((match.group(1), parsed_date))
    for match in re.finditer(r"\b(\d{1,7}[-/](?:FE|NE|FF|V\d+))\b", text, re.I):
        refs.append((match.group(1), fallback_date))
    for match in re.finditer(
        r"\b(?:FATT(?:URA|URE|\.?)|INVOICE|INV\.?|FACTURE|FACTURA)\s*(?:N\.?|NO\.?)?\s*(\d{1,7})(?![-/]\d)\b",
        text,
        re.I,
    ):
        refs.append((match.group(1), fallback_date))
    for match in re.finditer(
        r"\bFT\.?\s*(?:N\.?)?\s*(\d{1,7}(?:[-/]\d{2})?)(?=\s|[-–—]|$)", text, re.I
    ):
        refs.append((match.group(1), fallback_date))
    seen: set[tuple[str, str]] = set()
    out: list[tuple[str, str]] = []
    for doc_no, doc_date in refs:
        key = (clean_text(doc_no), doc_date)
        if key not in seen:
            seen.add(key)
            out.append(key)
    return out


def expand_numeric_range(value: str, *, max_span: int = 50) -> list[int]:
    value = clean_text(value).replace(" ", "")
    if "-" not in value:
        return [int(value)] if value.isdigit() else []
    start_text, end_text = value.split("-", 1)
    if not (start_text.isdigit() and end_text.isdigit()):
        return []
    start = int(start_text)
    end = int(end_text)
    if end < start or end - start > max_span:
        return []
    return list(range(start, end + 1))


def extract_payment_batch_ids(text: str) -> list[str]:
    ids: list[str] = []
    seen: set[str] = set()
    for match in PAYMENT_BATCH_RE.finditer(text):
        for number in expand_numeric_range(match.group("ref")):
            key = f"distinta:{number}"
            if key not in seen:
                seen.add(key)
                ids.append(key)
    return ids


def parse_journal_xlsx(path: Path, assumptions: dict[str, Any]) -> list[dict[str, Any]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required for XLSX journal extraction")
    rows: list[dict[str, Any]] = []
    language = configured_language(assumptions, purpose="document")
    closure_terms = keyword_tuple(language, "evidence_keywords", "closure")
    compensation_terms = keyword_tuple(
        language, "evidence_keywords", "compensation"
    ) + keyword_tuple(language, "evidence_keywords", "netting")
    adapter_family = source_adapter_family(path, "journal", assumptions)
    money_convention = money_convention_for_source(assumptions, path.name)
    date_convention = date_convention_for_source(assumptions, path.name)
    workbook = load_workbook(path, read_only=True, data_only=True)
    for sheet in workbook.worksheets:
        layout = journal_layout_for_sheet(sheet)
        if (
            adapter_family != LEGACY_ADAPTER_FAMILY
            and not layout.get("debit_col")
            and not layout.get("credit_col")
        ):
            continue
        current_date = ""
        current_causale = ""
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = [clean_text(value) for value in row]
            joined = " ".join(value for value in values if value)
            if not joined:
                continue
            header = JOURNAL_HEADER_RE.search(joined)
            if header:
                current_date = iso_date(
                    header.group("date"),
                    convention=date_convention,
                )
                if not current_date:
                    workbook.close()
                    return []
                current_causale = clean_text(header.group("causale"))
                continue
            line_no = values[0] if values else ""
            if not re.fullmatch(r"\d{1,8}", line_no):
                continue
            account = next(
                (
                    value
                    for value in values
                    if re.fullmatch(r"\d+\s*/\s*\d+\s*/\s*\d+", value)
                ),
                "",
            )
            text_values = [
                value
                for value in values
                if value and value != line_no and value != account
            ]
            description = " ".join(text_values)
            doc_refs = extract_invoice_refs(
                description,
                current_date,
                date_convention=date_convention,
            )
            if not doc_refs:
                doc_match = re.search(r"\bn([A-Z0-9./-]{2,})\b", description, re.I)
                if doc_match:
                    doc_refs = [(doc_match.group(1), current_date)]
            if not doc_refs:
                continue
            if adapter_family == LEGACY_ADAPTER_FAMILY:
                amounts = [
                    parse_money(value, convention=money_convention) for value in row
                ]
                numeric_amounts = [value for value in amounts if value is not None]
                amount = decimal_text(numeric_amounts[-1]) if numeric_amounts else ""
            else:
                debit, credit = journal_amount_sides(
                    row,
                    layout,
                    convention=money_convention,
                )
                if debit and credit:
                    continue
                exact_amount = debit or credit
                amount = decimal_text(abs(exact_amount)) if exact_amount else ""
            if not amount:
                continue
            lower_text = f"{current_causale} {description}".lower()
            evidence_type = "internal_accounting"
            if any_keyword_in(lower_text, closure_terms + compensation_terms):
                evidence_type = "internal_closure"
            if any_keyword_in(lower_text, compensation_terms):
                evidence_type = "compensation"
            if any(
                str(keyword).lower() in lower_text
                for keyword in assumptions.get("factoring_operator_keywords", [])
            ):
                evidence_type = "factoring_bridge"
            for doc_no, doc_date in doc_refs:
                rows.append(
                    {
                        "record_id": f"journal:{path.name}:{sheet.title}:r{row_index}:{doc_no}",
                        "source_file": path.name,
                        "source_sheet": sheet.title,
                        "source_row": row_index,
                        "source_role": "journal",
                        "document_no": doc_no,
                        "document_date": doc_date or current_date,
                        "posting_date": current_date,
                        "amount": amount,
                        "currency": assumptions.get("currency", "EUR"),
                        "description": f"{current_causale} {description}".strip(),
                        "evidence_type": evidence_type,
                        "document_key": document_key(doc_no, doc_date or current_date),
                    }
                )
    return rows


def journal_money_cell(
    value: object,
    *,
    convention: dict[str, Any] | None = None,
) -> Decimal | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, float):
        return None
    if isinstance(value, (int, Decimal)):
        return Decimal(value)
    text = clean_text(value)
    if not text or not AMOUNT_IT_RE.fullmatch(text):
        return None
    return parse_money(text, convention=convention)


def journal_layout_for_sheet(sheet: Any) -> dict[str, int]:
    layouts: dict[tuple[int, int, int], int] = defaultdict(int)
    for row in sheet.iter_rows(values_only=True):
        values = [clean_text(value) for value in row]
        layout = journal_layout_from_header_values(values)
        if layout:
            layouts[
                (layout["operation_col"], layout["debit_col"], layout["credit_col"])
            ] += 1
    if not layouts:
        return {"operation_col": 1, "debit_col": 0, "credit_col": 0}
    operation_col, debit_col, credit_col = sorted(
        layouts.items(), key=lambda item: (-item[1], item[0])
    )[0][0]
    return {
        "operation_col": operation_col,
        "debit_col": debit_col,
        "credit_col": credit_col,
    }


def journal_layout_from_header_values(values: list[str]) -> dict[str, int]:
    operation_col = 0
    debit_col = 0
    credit_col = 0
    for idx, value in enumerate(values, start=1):
        lower = value.lower()
        if "descrizione dell'operazione" in lower or "operation description" in lower:
            operation_col = idx
        if lower in {"dare", "debit"}:
            debit_col = idx
        if lower in {"avere", "credit"}:
            credit_col = idx
    if operation_col and debit_col and credit_col:
        return {
            "operation_col": operation_col,
            "debit_col": debit_col,
            "credit_col": credit_col,
        }
    return {}


def journal_amount_sides(
    row: tuple[Any, ...],
    layout: dict[str, int],
    *,
    adapter_family: str = "",
    convention: dict[str, Any] | None = None,
) -> tuple[Decimal, Decimal]:
    debit = Decimal("0.00")
    credit = Decimal("0.00")
    debit_col = layout.get("debit_col", 0)
    credit_col = layout.get("credit_col", 0)
    operation_col = layout.get("operation_col", 1)
    midpoint = (
        ((debit_col + credit_col) / 2)
        if adapter_family == LEGACY_ADAPTER_FAMILY and debit_col and credit_col
        else None
    )
    for idx, value in enumerate(row, start=1):
        if idx <= operation_col:
            continue
        amount = journal_money_cell(value, convention=convention)
        if amount is None:
            continue
        if idx == credit_col:
            credit += amount
        elif idx == debit_col:
            debit += amount
        elif midpoint is not None:
            if idx > midpoint:
                credit += amount
            else:
                debit += amount
    return debit, credit


def journal_row_text_values(
    row: tuple[Any, ...], start_col: int, end_col: int
) -> list[str]:
    return [
        clean_text(value)
        for idx, value in enumerate(row, start=1)
        if start_col <= idx < end_col and clean_text(value)
    ]


def parse_journal_rollforward_xlsx(
    path: Path, assumptions: dict[str, Any]
) -> list[dict[str, Any]]:
    if load_workbook is None:
        raise RuntimeError(
            "openpyxl is required for XLSX journal roll-forward extraction"
        )
    counterparty_keywords = [
        clean_text(keyword).lower()
        for keyword in assumptions.get("counterparty_keywords", [])
        if clean_text(keyword)
    ]
    if not counterparty_keywords:
        return []
    workbook = load_workbook(path, read_only=True, data_only=True)
    adapter_family = source_adapter_family(path, "journal", assumptions)
    money_convention = money_convention_for_source(assumptions, path.name)
    date_convention = date_convention_for_source(assumptions, path.name)
    rows: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        layout = journal_layout_for_sheet(sheet)
        current_date = ""
        current_causale = ""
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = [clean_text(value) for value in row]
            joined = " ".join(value for value in values if value)
            if not joined:
                continue
            row_layout = journal_layout_from_header_values(values)
            if row_layout:
                layout = row_layout
                continue
            operation_col = layout.get("operation_col", 1)
            debit_col = layout.get("debit_col", 0)
            credit_col = layout.get("credit_col", 0)
            header = JOURNAL_HEADER_RE.search(joined)
            if header:
                current_date = iso_date(
                    header.group("date"),
                    convention=date_convention,
                )
                if not current_date:
                    workbook.close()
                    return []
                current_causale = clean_text(header.group("causale"))
                continue
            line_no = values[0] if values else ""
            if not re.fullmatch(r"\d{1,8}", line_no):
                continue
            account_cell = next(
                (
                    (idx, value)
                    for idx, value in enumerate(values, start=1)
                    if re.fullmatch(r"\d+\s*/\s*\d+\s*/\s*\d+", value)
                ),
                None,
            )
            if not account_cell:
                continue
            account_col, account = account_cell
            account_name_values = journal_row_text_values(
                row, account_col + 1, operation_col
            )
            account_name = account_name_values[0] if account_name_values else ""
            description_values = journal_row_text_values(
                row,
                operation_col,
                min([col for col in (debit_col, credit_col) if col] or [len(row) + 1]),
            )
            description = " ".join(description_values)
            if assumptions.get("rollforward_match_descriptions", False):
                match_text = f"{account_name} {description} {current_causale}".lower()
            else:
                match_text = account_name.lower()
            if not any(keyword in match_text for keyword in counterparty_keywords):
                continue
            debit, credit = journal_amount_sides(
                row,
                layout,
                adapter_family=adapter_family,
                convention=money_convention,
            )
            if debit == Decimal("0.00") and credit == Decimal("0.00"):
                continue
            movement_text = f"{current_causale} {description}".lower()
            movement_type = (
                "opening"
                if any(term in movement_text for term in OPENING_ENTRY_TERMS)
                else "period_movement"
            )
            signed = debit - credit
            rows.append(
                {
                    "record_id": f"journal_rollforward:{path.name}:{sheet.title}:r{row_index}",
                    "source_file": path.name,
                    "source_sheet": sheet.title,
                    "source_row": row_index,
                    "source_role": "journal",
                    "posting_date": current_date,
                    "causale": current_causale,
                    "account": account,
                    "account_name": account_name,
                    "description": description,
                    "movement_type": movement_type,
                    "debit_amount": decimal_text(debit),
                    "credit_amount": decimal_text(credit),
                    "signed_debit_minus_credit": decimal_text(signed),
                    "currency": assumptions.get("currency", "EUR"),
                    "debit_column": debit_col,
                    "credit_column": credit_col,
                    "operation_column": operation_col,
                }
            )
    return rows


def summarize_journal_rollforward(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    buckets: dict[tuple[str, str], dict[str, Any]] = {}

    def add_to_bucket(key: tuple[str, str], row: dict[str, Any]) -> None:
        bucket = buckets.setdefault(
            key,
            {
                "account": key[0],
                "account_name": key[1],
                "rows": 0,
                "opening_debit": Decimal("0.00"),
                "opening_credit": Decimal("0.00"),
                "period_debit": Decimal("0.00"),
                "period_credit": Decimal("0.00"),
            },
        )
        movement_type = clean_text(row.get("movement_type"))
        debit = parse_money(row.get("debit_amount")) or Decimal("0.00")
        credit = parse_money(row.get("credit_amount")) or Decimal("0.00")
        bucket["rows"] += 1
        if movement_type == "opening":
            bucket["opening_debit"] += debit
            bucket["opening_credit"] += credit
        else:
            bucket["period_debit"] += debit
            bucket["period_credit"] += credit

    for row in rows:
        add_to_bucket(
            (clean_text(row.get("account")), clean_text(row.get("account_name"))), row
        )
        add_to_bucket(("TOTAL", "All matched counterparty journal accounts"), row)

    summary: list[dict[str, Any]] = []
    for bucket in buckets.values():
        opening_net = bucket["opening_debit"] - bucket["opening_credit"]
        period_net = bucket["period_debit"] - bucket["period_credit"]
        closing_net = opening_net + period_net
        summary.append(
            {
                "account": bucket["account"],
                "account_name": bucket["account_name"],
                "rows": bucket["rows"],
                "opening_debit": f"{bucket['opening_debit']:.2f}",
                "opening_credit": f"{bucket['opening_credit']:.2f}",
                "opening_net_debit_minus_credit": f"{opening_net:.2f}",
                "period_debit": f"{bucket['period_debit']:.2f}",
                "period_credit": f"{bucket['period_credit']:.2f}",
                "period_net_debit_minus_credit": f"{period_net:.2f}",
                "closing_net_debit_minus_credit": f"{closing_net:.2f}",
            }
        )
    return sorted(
        summary,
        key=lambda row: (
            row["account"] != "TOTAL",
            row["account"],
            row["account_name"],
        ),
    )


def is_bank_like_ledger_row(row: dict[str, Any]) -> bool:
    text = f"{row.get('source_file', '')} {row.get('account_name', '')}".lower()
    return any(term in text for term in BANK_ACCOUNT_TERMS)


def rollforward_counterparty_keywords(
    ledger_balance_rows: list[dict[str, Any]],
    assumptions: dict[str, Any],
) -> list[str]:
    """Infer conservative journal filter keywords from non-bank ledger accounts."""

    seen: set[str] = set()
    keywords: list[str] = []

    def add(value: object) -> None:
        keyword = clean_text(value).lower()
        if len(keyword) < 3 or keyword in seen:
            return
        seen.add(keyword)
        keywords.append(keyword)

    for keyword in assumptions.get("counterparty_keywords", []):
        add(keyword)
    if keywords:
        return keywords

    for row in ledger_balance_rows:
        if clean_text(row.get("account")) == "TOTAL" or is_bank_like_ledger_row(row):
            continue
        add(row.get("account_name"))
    return keywords


def matched_ledger_balance_rows(
    ledger_balance_rows: list[dict[str, Any]],
    journal_rollforward_summary: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    journal_accounts = {
        clean_text(row.get("account"))
        for row in journal_rollforward_summary
        if clean_text(row.get("account")) != "TOTAL"
    }
    if journal_accounts:
        return [
            row
            for row in ledger_balance_rows
            if clean_text(row.get("account")) in journal_accounts
            and clean_text(row.get("account")) != "TOTAL"
        ]
    return [
        row
        for row in ledger_balance_rows
        if clean_text(row.get("account")) != "TOTAL"
        and not is_bank_like_ledger_row(row)
    ]


def rollforward_decimal(value: object) -> Decimal:
    return parse_money(value) or Decimal("0.00")


def rollforward_status(
    opening_diff: Decimal | None, closing_diff: Decimal | None, tolerance: Decimal
) -> str:
    if opening_diff is None or closing_diff is None:
        return "MISSING_JOURNAL_OR_LEDGER"
    if abs(opening_diff) <= tolerance and abs(closing_diff) <= tolerance:
        return "PASS"
    return "DIFFERENCE"


def rollforward_check_note(status: str) -> str:
    if status == "PASS":
        return "Saldo iniziale e saldo finale del giornale riconciliano con il mastro entro tolleranza."
    if status == "MISSING_JOURNAL_OR_LEDGER":
        return "Manca il conto nel giornale o nel mastro; verificare parole chiave e layout dei file."
    return (
        "Il saldo ricostruito dal giornale non coincide con il saldo finale del mastro."
    )


def build_account_rollforward_check(
    ledger_balance_rows: list[dict[str, Any]],
    journal_rollforward_summary: list[dict[str, Any]],
    assumptions: dict[str, Any],
) -> list[dict[str, Any]]:
    """Compare journal roll-forward totals to ledger opening/closing balances."""

    if not ledger_balance_rows and not journal_rollforward_summary:
        return []

    raw_tolerance = assumptions.get("rollforward_amount_tolerance")
    if raw_tolerance is None:
        raw_tolerance = assumptions.get("amount_tolerance", "0.01")
    tolerance = parse_money(raw_tolerance)
    if tolerance is None or tolerance < 0:
        raise ValueError("roll-forward tolerance must be a non-negative Decimal")
    journal_by_account = {
        clean_text(row.get("account")): row for row in journal_rollforward_summary
    }
    ledger_rows = matched_ledger_balance_rows(
        ledger_balance_rows, journal_rollforward_summary
    )
    rows: list[dict[str, Any]] = []

    compared_accounts: set[str] = set()
    for ledger in ledger_rows:
        account = clean_text(ledger.get("account"))
        compared_accounts.add(account)
        journal = journal_by_account.get(account)
        ledger_opening = rollforward_decimal(
            ledger.get("opening_balance_signed_debit_minus_credit")
        )
        ledger_closing = rollforward_decimal(
            ledger.get("closing_balance_signed_debit_minus_credit")
        )
        if journal:
            journal_opening = rollforward_decimal(
                journal.get("opening_net_debit_minus_credit")
            )
            journal_period = rollforward_decimal(
                journal.get("period_net_debit_minus_credit")
            )
            journal_closing = rollforward_decimal(
                journal.get("closing_net_debit_minus_credit")
            )
            opening_diff: Decimal | None = (journal_opening - ledger_opening).quantize(
                Decimal("0.01")
            )
            closing_diff: Decimal | None = (journal_closing - ledger_closing).quantize(
                Decimal("0.01")
            )
            status = rollforward_status(opening_diff, closing_diff, tolerance)
        else:
            journal_opening = Decimal("0.00")
            journal_period = Decimal("0.00")
            journal_closing = Decimal("0.00")
            opening_diff = None
            closing_diff = None
            status = "MISSING_JOURNAL_OR_LEDGER"
        rows.append(
            {
                "account": account,
                "account_name": clean_text(ledger.get("account_name")),
                "ledger_source_file": clean_text(ledger.get("source_file")),
                "ledger_source_pages": clean_text(ledger.get("source_pages")),
                "journal_rows": int(journal.get("rows", 0)) if journal else 0,
                "ledger_opening_balance": f"{ledger_opening:.2f}",
                "journal_opening_balance": f"{journal_opening:.2f}",
                "opening_difference_journal_minus_ledger": (
                    "" if opening_diff is None else f"{opening_diff:.2f}"
                ),
                "journal_period_net_movement": f"{journal_period:.2f}",
                "journal_recalculated_closing": f"{journal_closing:.2f}",
                "ledger_closing_balance": f"{ledger_closing:.2f}",
                "closing_difference_journal_minus_ledger": (
                    "" if closing_diff is None else f"{closing_diff:.2f}"
                ),
                "status": status,
                "review_note": rollforward_check_note(status),
            }
        )

    for journal in journal_rollforward_summary:
        account = clean_text(journal.get("account"))
        if account == "TOTAL" or account in compared_accounts:
            continue
        status = "MISSING_JOURNAL_OR_LEDGER"
        rows.append(
            {
                "account": account,
                "account_name": clean_text(journal.get("account_name")),
                "ledger_source_file": "",
                "ledger_source_pages": "",
                "journal_rows": int(journal.get("rows", 0)),
                "ledger_opening_balance": "",
                "journal_opening_balance": journal.get(
                    "opening_net_debit_minus_credit", "0.00"
                ),
                "opening_difference_journal_minus_ledger": "",
                "journal_period_net_movement": journal.get(
                    "period_net_debit_minus_credit", "0.00"
                ),
                "journal_recalculated_closing": journal.get(
                    "closing_net_debit_minus_credit", "0.00"
                ),
                "ledger_closing_balance": "",
                "closing_difference_journal_minus_ledger": "",
                "status": status,
                "review_note": rollforward_check_note(status),
            }
        )

    if rows:
        journal_total = journal_by_account.get("TOTAL", {})
        ledger_opening_total = sum(
            (
                rollforward_decimal(row.get("ledger_opening_balance"))
                for row in rows
                if row.get("ledger_opening_balance")
            ),
            Decimal("0.00"),
        )
        ledger_closing_total = sum(
            (
                rollforward_decimal(row.get("ledger_closing_balance"))
                for row in rows
                if row.get("ledger_closing_balance")
            ),
            Decimal("0.00"),
        )
        journal_opening_total = rollforward_decimal(
            journal_total.get("opening_net_debit_minus_credit")
        )
        journal_period_total = rollforward_decimal(
            journal_total.get("period_net_debit_minus_credit")
        )
        journal_closing_total = rollforward_decimal(
            journal_total.get("closing_net_debit_minus_credit")
        )
        opening_diff_total = (journal_opening_total - ledger_opening_total).quantize(
            Decimal("0.01")
        )
        closing_diff_total = (journal_closing_total - ledger_closing_total).quantize(
            Decimal("0.01")
        )
        status = rollforward_status(opening_diff_total, closing_diff_total, tolerance)
        rows.insert(
            0,
            {
                "account": "TOTAL",
                "account_name": "Conti confrontati",
                "ledger_source_file": "",
                "ledger_source_pages": "",
                "journal_rows": int(journal_total.get("rows", 0) or 0),
                "ledger_opening_balance": f"{ledger_opening_total:.2f}",
                "journal_opening_balance": f"{journal_opening_total:.2f}",
                "opening_difference_journal_minus_ledger": f"{opening_diff_total:.2f}",
                "journal_period_net_movement": f"{journal_period_total:.2f}",
                "journal_recalculated_closing": f"{journal_closing_total:.2f}",
                "ledger_closing_balance": f"{ledger_closing_total:.2f}",
                "closing_difference_journal_minus_ledger": f"{closing_diff_total:.2f}",
                "status": status,
                "review_note": rollforward_check_note(status),
            },
        )
    return rows


def journal_evidence_type(text: str, assumptions: dict[str, Any]) -> str:
    lower_text = text.lower()
    language = configured_language(assumptions, purpose="document")
    closure_terms = keyword_tuple(language, "evidence_keywords", "closure")
    compensation_terms = keyword_tuple(
        language, "evidence_keywords", "compensation"
    ) + keyword_tuple(language, "evidence_keywords", "netting")
    evidence_type = "internal_accounting"
    if any_keyword_in(lower_text, closure_terms + compensation_terms):
        evidence_type = "internal_closure"
    if any_keyword_in(lower_text, compensation_terms):
        evidence_type = "compensation"
    if any(
        str(keyword).lower() in lower_text
        for keyword in assumptions.get("factoring_operator_keywords", [])
    ):
        evidence_type = "factoring_bridge"
    return evidence_type


def parse_journal_pages(
    pages: list[SourcePage], assumptions: dict[str, Any]
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for page in pages:
        if page.source_role != "journal":
            continue
        if (
            source_adapter_family(page.source_file, page.source_role, assumptions)
            != LEGACY_ADAPTER_FAMILY
        ):
            continue
        money_convention = money_convention_for_source(assumptions, page.source_file)
        date_convention = date_convention_for_source(
            assumptions,
            page.source_file,
        )
        current_date = ""
        current_causale = ""
        lines = [
            clean_text(line) for line in page.text.splitlines() if clean_text(line)
        ]
        for idx, line in enumerate(lines, start=1):
            header = JOURNAL_HEADER_RE.search(line)
            if header:
                current_date = iso_date(
                    header.group("date"),
                    convention=date_convention,
                )
                if not current_date:
                    return []
                current_causale = clean_text(header.group("causale"))
                continue
            account = ""
            match = JOURNAL_ACCOUNT_RE.search(line)
            description = line
            if match:
                account = clean_text(match.group("account"))
                description = clean_text(line[match.end() :])
            doc_refs = extract_invoice_refs(
                description,
                current_date,
                date_convention=date_convention,
            )
            if not doc_refs:
                doc_match = re.search(r"\bn([A-Z0-9./-]{2,})\b", description, re.I)
                if doc_match:
                    doc_refs = [(doc_match.group(1), current_date)]
            if not doc_refs:
                continue
            amounts = AMOUNT_IT_RE.findall(description)
            amount = amount_string(
                amounts[-1] if amounts else "",
                convention=money_convention,
            )
            if not amount:
                continue
            full_description = f"{current_causale} {description}".strip()
            evidence_type = journal_evidence_type(full_description, assumptions)
            for doc_no, doc_date in doc_refs:
                rows.append(
                    {
                        "record_id": f"journal_pdf:{page.source_file}:p{page.source_page}:l{idx}:{doc_no}",
                        "source_file": page.source_file,
                        "source_page": page.source_page,
                        "source_row": idx,
                        "source_role": "journal",
                        "account": account,
                        "document_no": doc_no,
                        "document_date": doc_date or current_date,
                        "posting_date": current_date,
                        "amount": amount,
                        "currency": assumptions.get("currency", "EUR"),
                        "description": full_description,
                        "evidence_type": evidence_type,
                        "document_key": document_key(doc_no, doc_date or current_date),
                    }
                )
    return rows


def parse_payment_order_zip(
    path: Path, assumptions: dict[str, Any]
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    money_convention = money_convention_for_source(assumptions, path.name)
    date_convention = date_convention_for_source(assumptions, path.name)
    with zipfile.ZipFile(path) as zf:
        for member in sorted(zf.namelist()):
            if member.endswith("/") or not member.lower().endswith(
                (".doc", ".html", ".htm", ".txt")
            ):
                continue
            raw = zf.read(member).decode("utf-8", errors="replace")
            text = re.sub(r"<[^>]+>", " ", raw)
            text = clean_text(text)
            header = PAYMENT_ORDER_HEADER_RE.search(text)
            batch_no = str(int(header.group("batch"))) if header else ""
            batch_id = f"distinta:{batch_no}" if batch_no else ""
            order_date = (
                iso_date(
                    header.group("date"),
                    convention=date_convention,
                )
                if header
                else ""
            )
            if header and not order_date:
                return []
            valuta = ""
            valuta_match = re.search(
                r"\b(?:Valuta|Value\s+Date|Date\s+de\s+Valeur|Fecha\s+Valor)\s+(\d{2}/\d{2}/\d{4})",
                text,
                re.I,
            )
            if valuta_match:
                valuta = iso_date(
                    valuta_match.group(1),
                    convention=date_convention,
                )
                if not valuta:
                    return []
            total_matches = list(PAYMENT_ORDER_TOTAL_RE.finditer(text))
            batch_total = (
                amount_string(
                    total_matches[-1].group("amount"),
                    convention=money_convention,
                )
                if total_matches
                else ""
            )
            for idx, match in enumerate(PAYMENT_ORDER_LINE_RE.finditer(text), start=1):
                doc_no = clean_text(match.group("document_no"))
                doc_date = iso_date(
                    match.group("document_date"),
                    convention=date_convention,
                )
                counterparty_doc = clean_text(match.group("counterparty_doc"))
                counterparty_date = iso_date(
                    match.group("counterparty_date"),
                    convention=date_convention,
                )
                if not doc_date or not counterparty_date:
                    return []
                invoice_amount = amount_string(
                    match.group("invoice_amount"),
                    convention=money_convention,
                )
                if not invoice_amount:
                    continue
                rows.append(
                    {
                        "record_id": f"payment_order:{path.name}:{member}:{idx}",
                        "source_file": f"{path.name}!{member}",
                        "source_role": "payment_order",
                        "document_no": doc_no,
                        "document_date": doc_date,
                        "posting_date": valuta or order_date,
                        "value_date": valuta,
                        "payment_order_date": order_date,
                        "counterparty_document_no": counterparty_doc,
                        "counterparty_document_date": counterparty_date,
                        "amount": invoice_amount,
                        "batch_total": batch_total,
                        "group_total": batch_total,
                        "batch_id": batch_id,
                        "group_id": batch_id,
                        "currency": assumptions.get("currency", "EUR"),
                        "description": text[:1000],
                        "evidence_type": "payment_order_bridge",
                        "document_key": document_key(doc_no, doc_date),
                    }
                )
    return rows


def extract_normalized_records(
    input_dir: str | Path,
    assumptions: dict[str, Any] | None = None,
    *,
    output_dir: str | Path | None = None,
) -> dict[str, Any]:
    active = dict(assumptions or {})
    root = Path(input_dir)
    out_dir = validate_run_output_dir(
        Path(output_dir) if output_dir else root.parent / "output",
        input_dir=root,
    )
    cache_dir = validate_run_cache_dir(
        active.get("cache_dir")
        or active.get("ocr_cache_dir")
        or out_dir / ".audit_reconciliation_cache",
        input_dir=root,
    )
    cache_dir.mkdir(parents=True, exist_ok=True)

    files = sorted(
        p for p in root.iterdir() if p.is_file() and not p.name.startswith(".")
    )
    supplied_source_receipts = active.get("source_artifact_receipts")
    if supplied_source_receipts is None:
        source_receipts = build_source_receipts(root, files)
    elif isinstance(supplied_source_receipts, list):
        source_receipts = validate_receipt_set(
            {"source": root.resolve()},
            supplied_source_receipts,
        )
        if {str(item["path"]) for item in source_receipts} != {
            path.relative_to(root).as_posix() for path in files
        }:
            raise AssuranceRunError(
                "source receipts do not cover the exact current source-file set"
            )
    else:
        raise AssuranceRunError("source_artifact_receipts must be a list")
    requested_adapters = _requested_source_adapter_families(files, active)
    decisions_by_path, decision_errors = build_reviewed_source_decisions(
        input_root=root,
        source_receipts=source_receipts,
        adapter_families=requested_adapters,
        assumptions=active,
    )
    active["_reviewed_source_decision_receipts"] = {
        **decisions_by_path,
        **{Path(path).name: value for path, value in decisions_by_path.items()},
    }
    inventory = source_inventory(root, active)
    all_pages: list[SourcePage] = []
    evidence_rows: list[dict[str, Any]] = []
    journal_rollforward_rows: list[dict[str, Any]] = []
    journal_paths: list[Path] = []
    extraction_errors: list[dict[str, Any]] = []
    language = configured_language(active, purpose="document")
    source_resolutions = {
        path.name: resolve_source_role(
            path,
            assumptions=active,
            language=language,
        )
        for path in files
    }
    source_adapters = {
        path.name: (
            source_adapter_family(
                path,
                source_resolutions[path.name]["source_role"],
                active,
            )
            or requested_adapters[path.name]
        )
        for path in files
    }
    candidate_rows_by_source: dict[str, int] = defaultdict(int)
    spreadsheet_roles = {
        source_resolutions[path.name]["source_role"]
        for path in files
        if path.suffix.lower() in {".xlsx", ".xlsm", ".xls", ".csv"}
        and source_resolutions[path.name]["status"] == "reviewed"
    }
    prefer_spreadsheet_for_roles = set(active.get("prefer_spreadsheet_for_roles", []))

    def progress_callback(event: dict[str, Any]) -> None:
        if not active.get("verbose_extraction"):
            return
        if event.get("event") == "pdf_file_start":
            print(
                "[audit-reconciliation] PDF start "
                f"{event.get('source_file')} pages={event.get('page_count')}",
                flush=True,
            )
        elif event.get("event") == "pdf_page_done":
            print(
                "[audit-reconciliation] PDF progress "
                f"{event.get('source_file')} "
                f"page {event.get('source_page')}/{event.get('page_count')} "
                f"method={event.get('extraction_method')} "
                f"text_chars={event.get('text_length')}",
                flush=True,
            )
        elif event.get("event") == "pdf_file_done":
            print(
                "[audit-reconciliation] PDF done "
                f"{event.get('source_file')} pages={event.get('page_count')} "
                f"ocr_pages={event.get('ocr_page_count')} "
                f"text_chars={event.get('text_length')}",
                flush=True,
            )
        elif event.get("event") == "ocr_page_start":
            print(
                "[audit-reconciliation] OCR page "
                f"{event.get('source_page')}/{event.get('page_count')} "
                f"{event.get('source_file')}",
                flush=True,
            )
        elif event.get("event") == "ocr_page_done":
            print(
                "[audit-reconciliation] OCR done "
                f"{event.get('source_page')}/{event.get('page_count')} "
                f"{event.get('source_file')} "
                f"text_chars={event.get('text_length')}",
                flush=True,
            )
        elif event.get("event") == "pdf_cache_hit":
            print(
                "[audit-reconciliation] cache hit "
                f"{event.get('source_file')} pages={event.get('page_count')}",
                flush=True,
            )

    for path in files:
        try:
            resolution = source_resolutions[path.name]
            role_from_name = resolution["source_role"]
            adapter_family = source_adapters[path.name]
            if active.get("verbose_extraction"):
                print(
                    f"[audit-reconciliation] extracting {path.name} as {role_from_name}",
                    flush=True,
                )
            if adapter_family and adapter_family not in SUPPORTED_ADAPTER_FAMILIES:
                extraction_errors.append(
                    {
                        "source_file": path.name,
                        "status": "unsupported_source_layout",
                        "reason": (
                            decision_errors.get(path.name)
                            or decision_errors.get(path.as_posix())
                            or (
                                "The requested adapter is not in the supported "
                                "source-layout allowlist."
                            )
                        ),
                    }
                )
                continue
            if resolution["status"] != "reviewed":
                extraction_errors.append(
                    {
                        "source_file": path.name,
                        "status": resolution["status"],
                        "reason": (
                            decision_errors.get(path.name)
                            or decision_errors.get(path.as_posix())
                            or "Source-role suggestions are advisory. Record a "
                            "reviewed source decision before parsing."
                        ),
                        "source_role_candidates": resolution["source_role_candidates"],
                    }
                )
                continue
            if adapter_family not in SUPPORTED_ADAPTER_FAMILIES:
                extraction_errors.append(
                    {
                        "source_file": path.name,
                        "status": "unsupported_source_layout",
                        "reason": (
                            "No supported mechanically bounded adapter is available. "
                            "Declare a supported reviewed_source_adapters entry or "
                            "provide a supported structured export."
                        ),
                    }
                )
                continue
            if (
                path.suffix.lower() == ".pdf"
                and role_from_name in prefer_spreadsheet_for_roles
                and role_from_name in spreadsheet_roles
            ):
                extraction_errors.append(
                    {
                        "source_file": path.name,
                        "status": "skipped",
                        "reason": f"Skipped duplicate {role_from_name} PDF because spreadsheet source is available.",
                    }
                )
                if active.get("verbose_extraction"):
                    print(
                        f"[audit-reconciliation] skipped {path.name}: spreadsheet source available",
                        flush=True,
                    )
                continue
            if path.suffix.lower() == ".pdf":
                pages = extract_pdf_pages(
                    path,
                    cache_dir,
                    source_role=role_from_name,
                    language=language,
                    progress_every_pages=int(
                        active.get("pdf_progress_every_pages", 10)
                    ),
                    progress_callback=progress_callback,
                )
                all_pages.extend(pages)
                parsed_evidence = [
                    *parse_journal_pages(pages, active),
                    *parse_ledger_or_factoring_pages(pages, active),
                    *parse_bank_statement_pages(pages, active),
                ]
                parsed_evidence = apply_source_perimeter(
                    parsed_evidence,
                    source_file=path.name,
                    assumptions=active,
                )
                evidence_rows.extend(parsed_evidence)
                candidate_rows_by_source[path.name] += len(parsed_evidence)
            elif (
                path.suffix.lower() in {".xlsx", ".xlsm"}
                and role_from_name == "journal"
            ):
                parsed_evidence = parse_journal_xlsx(path, active)
                parsed_evidence = apply_source_perimeter(
                    parsed_evidence,
                    source_file=path.name,
                    assumptions=active,
                )
                evidence_rows.extend(parsed_evidence)
                candidate_rows_by_source[path.name] += len(parsed_evidence)
                journal_paths.append(path)
            elif path.suffix.lower() == ".zip" and role_from_name == "payment_order":
                parsed_evidence = parse_payment_order_zip(path, active)
                parsed_evidence = apply_source_perimeter(
                    parsed_evidence,
                    source_file=path.name,
                    assumptions=active,
                )
                evidence_rows.extend(parsed_evidence)
                candidate_rows_by_source[path.name] += len(parsed_evidence)
            else:
                extraction_errors.append(
                    {
                        "source_file": path.name,
                        "status": "unsupported_source_layout",
                        "reason": (
                            f"No parser is registered for {path.suffix.lower()} "
                            f"with reviewed role {role_from_name}."
                        ),
                    }
                )
            if active.get("verbose_extraction"):
                print(
                    f"[audit-reconciliation] done {path.name}: open_items={len(open_items) if 'open_items' in locals() else 'pending'} evidence_rows={len(evidence_rows)} pages={len(all_pages)}",
                    flush=True,
                )
        except (
            Exception
        ) as exc:  # keep run auditable instead of hiding extraction failures
            extraction_errors.append(
                {"source_file": path.name, "error": f"{type(exc).__name__}: {exc}"}
            )
            if active.get("verbose_extraction"):
                print(
                    f"[audit-reconciliation] error {path.name}: {type(exc).__name__}: {exc}",
                    flush=True,
                )

    open_items = apply_source_perimeters(
        parse_open_items(all_pages, active),
        active,
    )
    ledger_balance_rows = apply_source_perimeters(
        parse_ledger_balance_pages(all_pages, active),
        active,
    )
    for row in (*open_items, *ledger_balance_rows):
        candidate_rows_by_source[clean_text(row.get("source_file"))] += 1
    rollforward_keywords = rollforward_counterparty_keywords(
        ledger_balance_rows, active
    )
    if journal_paths and rollforward_keywords:
        rollforward_assumptions = {
            **active,
            "counterparty_keywords": rollforward_keywords,
        }
        for journal_path in journal_paths:
            journal_rollforward_rows.extend(
                apply_source_perimeter(
                    parse_journal_rollforward_xlsx(
                        journal_path,
                        rollforward_assumptions,
                    ),
                    source_file=journal_path.name,
                    assumptions=active,
                )
            )
    journal_rollforward_summary = summarize_journal_rollforward(
        journal_rollforward_rows
    )
    account_rollforward_check = build_account_rollforward_check(
        ledger_balance_rows,
        journal_rollforward_summary,
        active,
    )
    normalized_records = [*open_items, *evidence_rows]
    page_rows = [asdict(page) for page in all_pages]
    source_qualifications = [
        build_file_source_qualification(
            path,
            resolution=source_resolutions[path.name],
            adapter_family=source_adapters[path.name],
            reviewed_decision=_reviewed_source_decision(active, path),
            source_artifact_ref=next(
                str(receipt["artifact_id"])
                for receipt in source_receipts
                if Path(str(receipt["path"])).name == path.name
            ),
            candidate_row_count=candidate_rows_by_source[path.name],
            emitted_row_count=candidate_rows_by_source[path.name],
        )
        for path in files
    ]
    qualification_by_source_ref = {
        qualification["source_artifact_refs"][0]: qualification
        for qualification in source_qualifications
    }
    for inventory_row, path in zip(inventory, files):
        source_receipt = next(
            receipt
            for receipt in source_receipts
            if Path(str(receipt["path"])).name == path.name
        )
        qualification = qualification_by_source_ref[str(source_receipt["artifact_id"])]
        inventory_row["source_artifact_id"] = source_receipt["artifact_id"]
        inventory_row["source_adapter_family"] = source_adapters[path.name]
        inventory_row["source_qualification_id"] = qualification["qualification_id"]
        inventory_row["source_qualification_status"] = qualification["status"]
        if qualification["status"] == "unsupported_source_layout" and not any(
            row.get("source_file") == path.name
            and row.get("status") == "unsupported_source_layout"
            for row in extraction_errors
        ):
            extraction_errors.append(
                {
                    "source_file": path.name,
                    "status": "unsupported_source_layout",
                    "reason": "; ".join(qualification["limitations"])
                    or "The declared adapter did not emit qualified rows.",
                }
            )
    validate_receipt_set({"source": root.resolve()}, source_receipts)
    return {
        "source_inventory": inventory,
        "source_root": str(root.resolve()),
        "source_artifact_receipts": source_receipts,
        "reviewed_source_decision_receipts": list(decisions_by_path.values()),
        "source_qualifications": source_qualifications,
        "source_pages": page_rows,
        "open_items": open_items,
        "evidence_rows": evidence_rows,
        "ledger_balance_rows": ledger_balance_rows,
        "account_rollforward_check": account_rollforward_check,
        "journal_rollforward_rows": journal_rollforward_rows,
        "journal_rollforward_summary": journal_rollforward_summary,
        "normalized_records": normalized_records,
        "extraction_errors": extraction_errors,
        "cache_dir": str(cache_dir),
    }


def write_json(path: Path, payload: Any) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8"
    )
    return path


def run_raw_input_reconciliation(
    *,
    input_dir: str | Path,
    client_folder: Mapping[str, Any] | str | Path,
    engagement_id: str,
    workspace_root: str | Path,
    assumptions: dict[str, Any] | None = None,
    title: str | None = None,
    narrative: str = "",
    language: str = "it",
    run_id: str | None = None,
    expected_predecessor_checkpoint: str | None = None,
) -> dict[str, Any]:
    client_engagement = prepare_client_engagement_context(
        client_folder=client_folder,
        engagement_id=engagement_id,
        input_dir=input_dir,
        workspace_root=workspace_root,
        run_id=run_id,
    )
    requested_language = normalize_language(
        (assumptions or {}).get("locale") or language
    )
    active = {
        "scope_year": None,
        "cutoff_date": None,
        "report_language": requested_language,
        "document_language": requested_language,
        "currency": "EUR",
        "post_cutoff_events_excluded": True,
        "payment_orders_are_bank_evidence": False,
        "factoring_pro_soluto_closes_item": True,
        "compensation_requires_bank": False,
        **(assumptions or {}),
    }
    out_dir = validate_run_output_dir(
        client_engagement["output_dir"], input_dir=input_dir
    )
    out_dir.mkdir(parents=True, exist_ok=True)

    extracted = extract_normalized_records(input_dir, active, output_dir=out_dir)
    review_rows = None
    review_rows_path = active.get("review_rows_path")
    if review_rows_path:
        review_rows = json.loads(Path(review_rows_path).read_text(encoding="utf-8"))

    result = build_reconciliation_artifacts(
        output_dir=out_dir,
        open_items=extracted["open_items"],
        evidence_rows=extracted["evidence_rows"],
        assumptions=active,
        source_inventory=extracted["source_inventory"],
        source_qualifications=extracted.get("source_qualifications", []),
        source_artifact_root=extracted.get("source_root"),
        source_artifact_receipts=extracted.get("source_artifact_receipts", []),
        reviewed_source_decision_receipts=extracted.get(
            "reviewed_source_decision_receipts", []
        ),
        extraction_errors=extracted["extraction_errors"],
        normalized_records=extracted["normalized_records"],
        ledger_balance_rows=extracted["ledger_balance_rows"],
        account_rollforward_check=extracted.get("account_rollforward_check", []),
        aggregate_rollforward_rows=extracted["journal_rollforward_rows"],
        aggregate_rollforward_summary=extracted["journal_rollforward_summary"],
        metadata={
            "Input folder": str(input_dir),
            "Studio client ID": client_engagement["studio_client_folder"][
                "studio_client_id"
            ],
            "Engagement ID": client_engagement["engagement_id"],
            "Run timestamp": datetime.now().isoformat(timespec="seconds"),
        },
        client_engagement=client_engagement,
        run_id=client_engagement["run_id"],
        title=title,
        narrative=narrative,
        language=active.get("report_language", requested_language),
        excel_name="riconciliazione_audit.xlsx",
        word_name="relazione_riconciliazione_audit.docx",
        fail_on_check_errors=False,
        review_rows=review_rows,
        challenged_rows=active.get("challenged_rows"),
        review_seed=active.get("review_seed", "audit-reconciliation-review"),
        review_high_value_count=int(active.get("review_high_value_count", 10)),
        review_random_count=int(active.get("review_random_count", 20)),
        require_completed_review=bool(active.get("require_completed_review", True)),
        defer_assurance_finalization=True,
        expected_predecessor_checkpoint=expected_predecessor_checkpoint,
    )
    missing_evidence_pack = build_missing_evidence_request_pack(
        result["reconciliation_rows"],
        source_inventory=extracted["source_inventory"],
        normalized_records=extracted["normalized_records"],
        entity_name=active.get("entity_name") or active.get("company_name") or "",
        counterparty_name=active.get("counterparty_name")
        or active.get("counterparty")
        or "",
        cutoff_date=active.get("cutoff_date"),
        language=active.get("report_language", requested_language),
    )
    missing_evidence_requests_path = write_missing_evidence_workbook(
        out_dir / "richieste_mirate_evidenze.xlsx",
        missing_evidence_pack,
    )
    review_status_counts: dict[str, int] = defaultdict(int)
    for row in result["review_rows"]:
        review_status_counts[
            clean_text(row.get("review_status")).upper() or "MISSING"
        ] += 1

    source_pages_path = out_dir / "source_pages.json"
    manifest = {
        "client_engagement": client_engagement,
        "input_dir": str(input_dir),
        "output_dir": str(out_dir),
        "cache_dir": extracted["cache_dir"],
        "source_pages_path": str(source_pages_path),
        "assumptions": active,
        "counts": {
            "source_files": len(extracted["source_inventory"]),
            "source_qualifications": len(extracted.get("source_qualifications", [])),
            "unsupported_source_layouts": sum(
                1
                for row in extracted.get("source_qualifications", [])
                if row.get("status") == "unsupported_source_layout"
            ),
            "source_roles_needing_review": sum(
                1
                for row in extracted.get("source_qualifications", [])
                if row.get("status") == "needs_review"
            ),
            "source_pages": len(extracted["source_pages"]),
            "open_items": len(extracted["open_items"]),
            "evidence_rows": len(extracted["evidence_rows"]),
            "ledger_balance_rows": len(extracted["ledger_balance_rows"]),
            "account_rollforward_check_rows": len(
                extracted.get("account_rollforward_check", [])
            ),
            "journal_rollforward_rows": len(extracted["journal_rollforward_rows"]),
            "journal_rollforward_summary_rows": len(
                extracted["journal_rollforward_summary"]
            ),
            "reconciliation_rows": len(result["reconciliation_rows"]),
            "bank_allocation_candidates": len(result["bank_allocation_candidates"]),
            "relationship_allocation_ledgers": len(
                result.get("relationship_allocation_ledgers", [])
            ),
            "external_evidence_rows": len(result["external_evidence_detail"]),
            "external_evidence_summary_rows": len(result["external_evidence_summary"]),
            "post_cutoff_candidates": len(result["post_cutoff_candidates"]),
            "aging_summary_rows": len(result["aging_summary"]),
            "review_signal_rows": len(result["review_signals"]),
            "evidence_concentration_rows": len(result["evidence_concentration"]),
            "document_source_map_rows": len(result["document_source_map"]),
            "reversal_candidate_rows": len(result["reversal_candidates"]),
            "cutoff_window_movement_rows": len(result["cutoff_window_movements"]),
            "review_rows": len(result["review_rows"]),
            "review_status_counts": dict(sorted(review_status_counts.items())),
            "missing_evidence_request_rows": sum(
                len(rows) for rows in missing_evidence_pack.request_sections.values()
            ),
            "extraction_errors": len(extracted["extraction_errors"]),
        },
        "checks": result["checks"],
        "checks_pass": result["checks_pass"],
        "excel_path": result["excel_path"],
        "accountant_report_path": result["accountant_report_path"],
        "word_path": result["word_path"],
        "missing_evidence_requests_path": str(missing_evidence_requests_path),
    }
    write_json(out_dir / "run_manifest.json", manifest)
    write_json(source_pages_path, extracted["source_pages"])
    write_json(out_dir / "normalized_records.json", extracted["normalized_records"])
    write_json(
        out_dir / "account_rollforward_check.json",
        extracted.get("account_rollforward_check", []),
    )
    write_json(out_dir / "codex_review_packet.json", result["review_rows"])
    existing_review_session = result.get("review_session") or {}
    review_session = write_review_session_artifacts(
        out_dir,
        run_id=str(existing_review_session.get("run_id") or ""),
        run_intake_path=Path(existing_review_session["run_intake_path"]),
        result={
            **result,
            "assumptions": active,
            "missing_evidence_requests_path": str(missing_evidence_requests_path),
        },
        source_inventory=extracted["source_inventory"],
        source_paths=[input_dir],
        missing_evidence_requests_path=missing_evidence_requests_path,
        language=active.get("report_language", requested_language),
    )
    manifest["review_session"] = {
        "run_id": review_session.run_id,
        "run_intake_path": str(review_session.run_intake_path),
        "review_payload_path": str(review_session.review_payload_path),
        "ui_decisions_path": str(review_session.ui_decisions_path),
        "review_html_path": str(review_session.review_html_path),
        "final_artifacts_path": str(review_session.final_artifacts_path),
        "review_item_count": review_session.review_item_count,
    }
    manifest["assurance"] = {
        "receipts_path": str(out_dir / "assurance_receipts.json"),
        "gates_path": str(out_dir / "assurance_gates.json"),
        "final_output_inventory_path": str(out_dir / "final_output_inventory.json"),
        "final_output_boundary": str(out_dir / "assurance_final_outputs"),
        "canonical_data_path": str(
            out_dir / "assurance_final_outputs" / "reconciliation_results.json"
        ),
    }
    write_json(out_dir / "run_manifest.json", manifest)
    result["assurance"] = finalize_assurance_run(
        output_dir=out_dir,
        context=result["assurance_context"],
        reconciliation_rows=result["reconciliation_rows"],
        allocation_ledgers=result.get("relationship_allocation_ledgers", []),
        checks=result["checks"],
        review_rows=result["review_rows"],
        source_qualifications=extracted.get("source_qualifications", []),
        source_processing=result["source_processing"],
        analyses=result["analyses"],
        declared_outputs=[
            Path(result["excel_path"]),
            Path(result["accountant_report_path"]),
            Path(result["word_path"]),
            missing_evidence_requests_path,
        ],
        workbook_name=Path(result["excel_path"]).name,
    )
    return {
        **result,
        **extracted,
        "client_engagement": client_engagement,
        "manifest": manifest,
        "missing_evidence_requests_path": str(missing_evidence_requests_path),
        "missing_evidence_request_pack": missing_evidence_pack,
    }


def _cli_parser() -> Any:
    import argparse

    parser = argparse.ArgumentParser(
        description="Run one client-bound Audit Reconciliation workflow."
    )
    parser.add_argument("--client-folder-binding", type=Path, required=True)
    parser.add_argument("--engagement-id", required=True)
    parser.add_argument("--input-dir", type=Path, required=True)
    parser.add_argument("--workspace-root", type=Path, required=True)
    parser.add_argument("--assumptions-json", type=Path, required=True)
    parser.add_argument("--title")
    parser.add_argument("--narrative", default="")
    parser.add_argument("--language", default="it")
    parser.add_argument("--run-id")
    parser.add_argument("--expected-predecessor-checkpoint")
    return parser


def main(argv: list[str] | None = None) -> int:
    """Execute the client-bound raw runner from a portable JSON binding."""

    args = _cli_parser().parse_args(argv)
    try:
        assumptions = json.loads(args.assumptions_json.read_text(encoding="utf-8"))
        if not isinstance(assumptions, dict):
            raise ValueError("Assumptions JSON must contain one object")
        result = run_raw_input_reconciliation(
            input_dir=args.input_dir,
            client_folder=args.client_folder_binding,
            engagement_id=args.engagement_id,
            workspace_root=args.workspace_root,
            assumptions=assumptions,
            title=args.title,
            narrative=args.narrative,
            language=args.language,
            run_id=args.run_id,
            expected_predecessor_checkpoint=args.expected_predecessor_checkpoint,
        )
    except (OSError, ValueError) as exc:
        sys.stderr.write(f"Audit Reconciliation failed: {exc}\n")
        return 1
    summary = {
        "status": "ready_for_review",
        "client_engagement": result["client_engagement"],
        "run_manifest_path": str(
            Path(result["client_engagement"]["output_dir"]) / "run_manifest.json"
        ),
        "artifact_card_path": str(
            Path(result["client_engagement"]["output_dir"]) / "artifact_card.md"
        ),
    }
    sys.stdout.write(json.dumps(summary, ensure_ascii=False, sort_keys=True) + "\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
