"""Build targeted evidence requests from a reconciliation workbook."""

from __future__ import annotations

import argparse
import logging
import sys
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from .locale_support import normalize_language
except ImportError:  # pragma: no cover - direct import support
    scripts_dir = Path(__file__).resolve().parent
    if str(scripts_dir) not in sys.path:
        sys.path.insert(0, str(scripts_dir))
    from locale_support import normalize_language  # type: ignore


__all__ = [
    "MissingEvidenceRequestPack",
    "build_missing_evidence_request_pack",
    "load_reconciliation_context",
    "write_missing_evidence_workbook",
]


LOGGER = logging.getLogger(__name__)
DEFAULT_EXCEL_NAME = "richieste_mirate_evidenze.xlsx"
HEADER_FILL = "17365D"

SECTION_ORDER = (
    "reconciled_strong",
    "probable_payment",
    "accounting_support_needed",
    "missing_evidence_needed",
    "open_balance_confirmation",
    "unresolved",
)
REQUEST_FIELDS = (
    "row_id",
    "side",
    "document",
    "document_date",
    "amount",
    "operational_category",
    "operational_criterion",
    "operational_owner",
    "available_evidence",
    "targeted_missing_item",
    "existing_reference",
    "requested_action",
    "evidence_description",
)
INSTRUCTION_FIELDS = ("principle", "guidance")
SUMMARY_FIELDS = ("section", "rows", "amount_total")
INVENTORY_FIELDS = (
    "evidence_category",
    "normalized_rows",
    "source_files",
    "operational_use",
)


TEXT: dict[str, dict[str, Any]] = {
    "it": {
        "sheet_names": {
            "instructions": "istruzioni",
            "summary": "sintesi",
            "inventory": "evidenze_disponibili",
            "reconciled_strong": "riconciliate_forti",
            "probable_payment": "pagamenti_probabili",
            "accounting_support_needed": "scritture_da_supportare",
            "missing_evidence_needed": "evidenze_da_integrare",
            "open_balance_confirmation": "saldi_aperti",
            "unresolved": "non_risolte",
        },
        "headers": {
            "principle": "principio",
            "guidance": "indicazione",
            "section": "sezione",
            "rows": "righe",
            "amount_total": "importo_totale",
            "evidence_category": "categoria_evidenza",
            "normalized_rows": "righe_normalizzate",
            "source_files": "file_sorgente",
            "operational_use": "uso_operativo",
            "row_id": "id_riga_workpaper",
            "side": "lato",
            "document": "documento",
            "document_date": "data_documento",
            "amount": "importo",
            "operational_category": "categoria_operativa",
            "operational_criterion": "criterio_operativo",
            "operational_owner": "destinatario_operativo",
            "available_evidence": "evidenza_gia_disponibile",
            "targeted_missing_item": "dato_mancante_mirato",
            "existing_reference": "riferimento_esistente",
            "requested_action": "azione_richiesta",
            "evidence_description": "descrizione_evidenza",
        },
        "sections": {
            "reconciled_strong": "Riconciliate con evidenza forte",
            "probable_payment": "Pagamenti probabili da allocare",
            "accounting_support_needed": "Scritture contabili da supportare",
            "missing_evidence_needed": "Evidenze da integrare",
            "open_balance_confirmation": "Saldi aperti da confermare",
            "unresolved": "Non risolte",
        },
        "criteria": {
            "reconciled_strong": "Evidenza esterna o catena contabile-bancaria sufficiente.",
            "probable_payment": "Movimento bancario o distinta probabile, allocazione non ancora confermata.",
            "accounting_support_needed": "Scrittura contabile di chiusura priva di supporto esterno sufficiente.",
            "missing_evidence_needed": "Indizio presente ma prova conclusiva mancante.",
            "open_balance_confirmation": "Saldo aperto supportato dalle evidenze interne.",
            "unresolved": "Collegamento insufficiente nelle evidenze acquisite.",
        },
        "side": {
            "customer": "cliente",
            "receivable": "cliente",
            "client": "cliente",
            "supplier": "fornitore",
            "payable": "fornitore",
            "vendor": "fornitore",
            "unknown": "non indicato",
        },
        "owners": {
            "default_entity": "soggetto revisionato",
            "default_counterparty": "controparte",
            "counterparty_confirmation": "{entity}; eventuale conferma saldo {counterparty}",
        },
        "inventory_usage": {
            "open_items": "Popolazione delle partite aperte o da riconciliare.",
            "ledger": "Supporto di saldo, partitario, mastro e registrazioni interne.",
            "journal": "Scritture contabili, chiusure interne, giroconti e movimenti di periodo.",
            "bank_statement": "Incassi e pagamenti effettivi da estratto conto.",
            "payment_order": "Distinte o batch utili per collegare banca e fatture.",
            "factoring_statement": "Evidenza factor, anticipo o operatore esterno.",
            "compensation_support": "Supporto di compensazione o netting documentato.",
            "unknown": "Fonte acquisita ma non classificata automaticamente.",
            "default": "Evidenza acquisita nel fascicolo.",
        },
        "available": {
            "reconciled_strong": "La riga e gia riconciliata nel workpaper con evidenza forte.",
            "probable_payment": "Estratto banca e candidato di match gia presenti nel workpaper.",
            "accounting_support_needed": "All.A e scrittura contabile gia rilevate.",
            "missing_evidence_needed": "Indizio o documento ponte gia rilevato, ma non conclusivo.",
            "open_balance_confirmation": "All.A e supporto interno di saldo gia presenti.",
            "unresolved": "Riga All.A presente, senza collegamento sufficiente nelle evidenze acquisite.",
            "reference": " Riferimento: {reference}.",
        },
        "missing": {
            "reconciled_strong": "Nessun dato mancante per la richiesta mirata.",
            "probable_payment": "Conferma allocazione fattura-per-fattura del movimento bancario o della distinta gia identificata.",
            "accounting_support_needed": "Supporto esterno o spiegazione documentata della scrittura contabile: banca, factor, compensazione, giroconto o chiusura non-cash.",
            "missing_evidence_needed": "Documento necessario per trasformare l'indizio in evidenza conclusiva.",
            "open_balance_confirmation": "Conferma che la partita era effettivamente aperta al {cutoff}, oppure prova puntuale della chiusura non rilevata.",
            "unresolved": "Riferimento sorgente o prova puntuale di chiusura non rilevata.",
        },
        "actions": {
            "reconciled_strong": "Nessuna richiesta operativa; conservare il riferimento nel workpaper.",
            "probable_payment": "Confermare o correggere la mappatura proposta, indicando per ogni fattura importo allocato, distinta/SEPA e data pagamento.",
            "accounting_support_needed": "Inviare solo il supporto della contropartita o spiegare la natura della chiusura non-cash; non rimandare mastro/giornale se invariati.",
            "missing_evidence_needed": "Indicare quale documento gia acquisito chiude la riga, oppure fornire il solo documento mancante indicato.",
            "open_balance_confirmation": "Se aperta: confermare il saldo. Se chiusa: indicare documento di pagamento/incasso/factor/compensazione e data.",
            "unresolved": "Fornire riferimento fattura/partitario o indicare se la riga e errata; se chiusa, allegare il documento di chiusura.",
        },
        "instructions": {
            "no_reask": "Non richiedere documenti gia acquisiti",
            "no_reask_body": "Il workpaper principale contiene gia All.A, mastri, giornale, estratti banca/factor, distinte e normalizzazioni disponibili.",
            "targeted": "Richiesta mirata",
            "targeted_body": "Per ogni riga chiedere solo il tassello mancante indicato: allocazione, supporto esterno, conferma aperto o riferimento sorgente.",
            "owner": "Destinatario",
            "owner_body": "{entity} deve prima confermare/correggere usando il fascicolo. {counterparty} serve solo per eventuale conferma saldo o riconciliazione esterna al {cutoff}.",
        },
        "log_path": "Workbook richieste mirate: {path}",
    },
    "fr": {
        "sheet_names": {
            "instructions": "instructions",
            "summary": "synthese",
            "inventory": "preuves_disponibles",
            "reconciled_strong": "rapprochees_preuve_forte",
            "probable_payment": "paiements_probables",
            "accounting_support_needed": "ecritures_a_justifier",
            "missing_evidence_needed": "preuves_a_completer",
            "open_balance_confirmation": "soldes_ouverts",
            "unresolved": "non_resolues",
        },
        "headers": {
            "principle": "principe",
            "guidance": "indication",
            "section": "section",
            "rows": "lignes",
            "amount_total": "montant_total",
            "evidence_category": "categorie_preuve",
            "normalized_rows": "lignes_normalisees",
            "source_files": "fichier_source",
            "operational_use": "usage_operationnel",
            "row_id": "id_ligne_workpaper",
            "side": "sens",
            "document": "document",
            "document_date": "date_document",
            "amount": "montant",
            "operational_category": "categorie_operationnelle",
            "operational_criterion": "critere_operationnel",
            "operational_owner": "destinataire_operationnel",
            "available_evidence": "preuve_deja_disponible",
            "targeted_missing_item": "element_manquant_cible",
            "existing_reference": "reference_existante",
            "requested_action": "action_demandee",
            "evidence_description": "description_preuve",
        },
        "sections": {
            "reconciled_strong": "Rapprochees avec preuve forte",
            "probable_payment": "Paiements probables a affecter",
            "accounting_support_needed": "Ecritures comptables a justifier",
            "missing_evidence_needed": "Preuves a completer",
            "open_balance_confirmation": "Soldes ouverts a confirmer",
            "unresolved": "Non resolues",
        },
        "criteria": {
            "reconciled_strong": "Preuve externe ou chaine comptable-bancaire suffisante.",
            "probable_payment": "Mouvement bancaire ou remise probable, affectation non encore confirmee.",
            "accounting_support_needed": "Ecriture comptable de cloture sans support externe suffisant.",
            "missing_evidence_needed": "Indice present mais preuve concluante manquante.",
            "open_balance_confirmation": "Solde ouvert supporte par les preuves internes.",
            "unresolved": "Lien insuffisant dans les preuves disponibles.",
        },
        "side": {
            "customer": "client",
            "receivable": "client",
            "client": "client",
            "supplier": "fournisseur",
            "payable": "fournisseur",
            "vendor": "fournisseur",
            "unknown": "non indique",
        },
        "owners": {
            "default_entity": "entite auditee",
            "default_counterparty": "contrepartie",
            "counterparty_confirmation": "{entity}; confirmation eventuelle du solde {counterparty}",
        },
        "inventory_usage": {
            "open_items": "Population des postes ouverts ou a rapprocher.",
            "ledger": "Support de solde, grand livre, auxiliaire et ecritures internes.",
            "journal": "Ecritures comptables, clotures internes, virements internes et mouvements de periode.",
            "bank_statement": "Encaissements et paiements effectifs selon releve bancaire.",
            "payment_order": "Remises ou lots utiles pour relier banque et factures.",
            "factoring_statement": "Preuve factor, avance ou operateur externe.",
            "compensation_support": "Support de compensation ou netting documente.",
            "unknown": "Source recue mais non classee automatiquement.",
            "default": "Preuve disponible dans le dossier.",
        },
        "available": {
            "reconciled_strong": "La ligne est deja rapprochee dans le workpaper avec une preuve forte.",
            "probable_payment": "Releve bancaire et candidat de rapprochement deja presents dans le workpaper.",
            "accounting_support_needed": "Poste ouvert et ecriture comptable deja identifies.",
            "missing_evidence_needed": "Indice ou document de liaison deja identifie, mais non concluant.",
            "open_balance_confirmation": "Poste ouvert et support interne de solde deja presents.",
            "unresolved": "Poste ouvert present, sans lien suffisant dans les preuves disponibles.",
            "reference": " Reference: {reference}.",
        },
        "missing": {
            "reconciled_strong": "Aucun element manquant pour la demande ciblee.",
            "probable_payment": "Confirmation de l'affectation facture par facture du mouvement bancaire ou de la remise deja identifiee.",
            "accounting_support_needed": "Support externe ou explication documentee de l'ecriture comptable: banque, factor, compensation, virement interne ou cloture non cash.",
            "missing_evidence_needed": "Document necessaire pour transformer l'indice en preuve concluante.",
            "open_balance_confirmation": "Confirmation que le poste etait effectivement ouvert au {cutoff}, ou preuve ponctuelle d'une cloture non detectee.",
            "unresolved": "Reference source ou preuve ponctuelle de cloture non detectee.",
        },
        "actions": {
            "reconciled_strong": "Aucune demande operationnelle; conserver la reference dans le workpaper.",
            "probable_payment": "Confirmer ou corriger l'affectation proposee, avec montant affecte, remise/SEPA et date de paiement par facture.",
            "accounting_support_needed": "Envoyer seulement le support de contrepartie ou expliquer la nature de la cloture non cash; ne pas renvoyer grand livre/journal inchanges.",
            "missing_evidence_needed": "Indiquer quel document deja acquis cloture la ligne, ou fournir uniquement le document manquant indique.",
            "open_balance_confirmation": "Si ouvert: confirmer le solde. Si cloture: indiquer document de paiement/encaissement/factor/compensation et date.",
            "unresolved": "Fournir la reference facture/auxiliaire ou indiquer si la ligne est erronee; si cloturee, joindre le document de cloture.",
        },
        "instructions": {
            "no_reask": "Ne pas redemander les documents deja obtenus",
            "no_reask_body": "Le workpaper principal contient deja postes ouverts, grands livres, journal, releves bancaires/factor, remises et normalisations disponibles.",
            "targeted": "Demande ciblee",
            "targeted_body": "Pour chaque ligne, demander uniquement l'element manquant indique: affectation, support externe, confirmation du solde ouvert ou reference source.",
            "owner": "Destinataire",
            "owner_body": "{entity} doit d'abord confirmer/corriger avec le dossier. {counterparty} sert seulement pour une eventuelle confirmation de solde ou reconciliation externe au {cutoff}.",
        },
        "log_path": "Workbook demandes ciblees: {path}",
    },
    "de": {
        "sheet_names": {
            "instructions": "anweisungen",
            "summary": "zusammenfassung",
            "inventory": "verfuegbare_evidenzen",
            "reconciled_strong": "abgestimmt_stark",
            "probable_payment": "wahrscheinliche_zahlungen",
            "accounting_support_needed": "buchungen_zu_belegen",
            "missing_evidence_needed": "evidenzen_ergaenzen",
            "open_balance_confirmation": "offene_salden",
            "unresolved": "nicht_geloest",
        },
        "headers": {
            "principle": "prinzip",
            "guidance": "hinweis",
            "section": "abschnitt",
            "rows": "zeilen",
            "amount_total": "betrag_summe",
            "evidence_category": "evidenz_kategorie",
            "normalized_rows": "normalisierte_zeilen",
            "source_files": "quelldateien",
            "operational_use": "operative_verwendung",
            "row_id": "workpaper_zeilen_id",
            "side": "seite",
            "document": "dokument",
            "document_date": "dokumentdatum",
            "amount": "betrag",
            "operational_category": "operative_kategorie",
            "operational_criterion": "operatives_kriterium",
            "operational_owner": "operativer_empfaenger",
            "available_evidence": "bereits_verfuegbare_evidenz",
            "targeted_missing_item": "gezielt_fehlendes_element",
            "existing_reference": "bestehende_referenz",
            "requested_action": "angeforderte_aktion",
            "evidence_description": "evidenz_beschreibung",
        },
        "sections": {
            "reconciled_strong": "Abgestimmt mit starker Evidenz",
            "probable_payment": "Wahrscheinliche Zahlungen zuzuordnen",
            "accounting_support_needed": "Buchungen zu belegen",
            "missing_evidence_needed": "Evidenzen zu ergaenzen",
            "open_balance_confirmation": "Offene Salden zu bestaetigen",
            "unresolved": "Nicht geloest",
        },
        "criteria": {
            "reconciled_strong": "Externe Evidenz oder ausreichende Buchungs-Bank-Kette.",
            "probable_payment": "Wahrscheinliche Bankbewegung oder Zahlungsliste, Zuordnung noch nicht bestaetigt.",
            "accounting_support_needed": "Buchhalterische Ausbuchung ohne ausreichenden externen Beleg.",
            "missing_evidence_needed": "Hinweis vorhanden, aber abschliessende Evidenz fehlt.",
            "open_balance_confirmation": "Offener Saldo durch interne Evidenz gestuetzt.",
            "unresolved": "Unzureichende Verbindung in den verfuegbaren Evidenzen.",
        },
        "side": {
            "customer": "kunde",
            "receivable": "kunde",
            "client": "kunde",
            "supplier": "lieferant",
            "payable": "lieferant",
            "vendor": "lieferant",
            "unknown": "nicht angegeben",
        },
        "owners": {
            "default_entity": "gepruefte einheit",
            "default_counterparty": "gegenpartei",
            "counterparty_confirmation": "{entity}; ggf. Saldobestaetigung {counterparty}",
        },
        "inventory_usage": {
            "open_items": "Grundgesamtheit der offenen oder abzustimmenden Posten.",
            "ledger": "Saldo-, Hauptbuch-, Nebenbuch- und interne Buchungsnachweise.",
            "journal": "Buchungen, interne Ausbuchungen, Umbuchungen und Periodenbewegungen.",
            "bank_statement": "Tatsaechliche Ein- und Auszahlungen laut Bankauszug.",
            "payment_order": "Zahlungslisten oder Zahlungslaeufe zur Verbindung von Bank und Rechnungen.",
            "factoring_statement": "Evidenz des Factors, Vorschusses oder externen Operators.",
            "compensation_support": "Dokumentierter Aufrechnungs- oder Netting-Nachweis.",
            "unknown": "Quelle erfasst, aber nicht automatisch klassifiziert.",
            "default": "Im Dossier verfuegbare Evidenz.",
        },
        "available": {
            "reconciled_strong": "Die Zeile ist im Workpaper bereits mit starker Evidenz abgestimmt.",
            "probable_payment": "Bankauszug und wahrscheinlicher Match-Kandidat sind im Workpaper bereits vorhanden.",
            "accounting_support_needed": "Offener Posten und Buchung sind bereits identifiziert.",
            "missing_evidence_needed": "Hinweis oder Brueckenbeleg ist bereits identifiziert, aber nicht abschliessend.",
            "open_balance_confirmation": "Offener Posten und interne Saldounterstuetzung sind bereits vorhanden.",
            "unresolved": "Offener Posten vorhanden, ohne ausreichende Verbindung in den verfuegbaren Evidenzen.",
            "reference": " Referenz: {reference}.",
        },
        "missing": {
            "reconciled_strong": "Kein fehlendes Element fuer die gezielte Anfrage.",
            "probable_payment": "Bestaetigung der rechnungsweisen Zuordnung der bereits identifizierten Bankbewegung oder Zahlungsliste.",
            "accounting_support_needed": "Externer Beleg oder dokumentierte Erklaerung der Buchung: Bank, Factor, Aufrechnung, Umbuchung oder nicht zahlungswirksame Schliessung.",
            "missing_evidence_needed": "Dokument, das den Hinweis in abschliessende Evidenz ueberfuehrt.",
            "open_balance_confirmation": "Bestaetigung, dass der Posten am {cutoff} tatsaechlich offen war, oder konkreter Nachweis einer nicht erkannten Schliessung.",
            "unresolved": "Quellreferenz oder konkreter Nachweis einer nicht erkannten Schliessung.",
        },
        "actions": {
            "reconciled_strong": "Keine operative Anfrage; Referenz im Workpaper beibehalten.",
            "probable_payment": "Vorgeschlagene Zuordnung bestaetigen oder korrigieren, je Rechnung mit zugeordnetem Betrag, Zahlungsliste/SEPA und Zahlungsdatum.",
            "accounting_support_needed": "Nur den Gegenbeleg senden oder die nicht zahlungswirksame Schliessung erklaeren; unveraendertes Hauptbuch/Journal nicht erneut senden.",
            "missing_evidence_needed": "Angeben, welches bereits vorliegende Dokument die Zeile schliesst, oder nur das angeforderte fehlende Dokument liefern.",
            "open_balance_confirmation": "Wenn offen: Saldo bestaetigen. Wenn geschlossen: Zahlungs-/Einzugs-/Factor-/Aufrechnungsdokument und Datum angeben.",
            "unresolved": "Rechnungs-/Nebenbuchreferenz liefern oder angeben, ob die Zeile falsch ist; falls geschlossen, Schliessungsbeleg beifuegen.",
        },
        "instructions": {
            "no_reask": "Bereits erhaltene Dokumente nicht erneut anfordern",
            "no_reask_body": "Das Haupt-Workpaper enthaelt bereits offene Posten, Hauptbuecher, Journal, Bank-/Factor-Ausweise, Zahlungslisten und verfuegbare Normalisierungen.",
            "targeted": "Gezielte Anfrage",
            "targeted_body": "Je Zeile nur das angegebene fehlende Element anfordern: Zuordnung, externer Beleg, Bestaetigung offener Saldo oder Quellreferenz.",
            "owner": "Empfaenger",
            "owner_body": "{entity} soll zuerst anhand des Dossiers bestaetigen/korrigieren. {counterparty} wird nur fuer eine moegliche Saldobestaetigung oder externe Abstimmung zum {cutoff} benoetigt.",
        },
        "log_path": "Workbook gezielte Anfragen: {path}",
    },
    "es": {
        "sheet_names": {
            "instructions": "instrucciones",
            "summary": "resumen",
            "inventory": "evidencias_disponibles",
            "reconciled_strong": "conciliadas_prueba_fuerte",
            "probable_payment": "pagos_probables",
            "accounting_support_needed": "asientos_por_justificar",
            "missing_evidence_needed": "evidencias_por_completar",
            "open_balance_confirmation": "saldos_abiertos",
            "unresolved": "sin_resolver",
        },
        "headers": {
            "principle": "principio",
            "guidance": "indicación",
            "section": "sección",
            "rows": "líneas",
            "amount_total": "importe_total",
            "evidence_category": "categoría_evidencia",
            "normalized_rows": "líneas_normalizadas",
            "source_files": "archivos_fuente",
            "operational_use": "uso_operativo",
            "row_id": "id_línea_papel_trabajo",
            "side": "lado",
            "document": "documento",
            "document_date": "fecha_documento",
            "amount": "importe",
            "operational_category": "categoría_operativa",
            "operational_criterion": "criterio_operativo",
            "operational_owner": "destinatario_operativo",
            "available_evidence": "evidencia_disponible",
            "targeted_missing_item": "elemento_pendiente_dirigido",
            "existing_reference": "referencia_existente",
            "requested_action": "acción_solicitada",
            "evidence_description": "descripción_evidencia",
        },
        "sections": {
            "reconciled_strong": "Conciliadas con evidencia sólida",
            "probable_payment": "Pagos probables por asignar",
            "accounting_support_needed": "Asientos contables por justificar",
            "missing_evidence_needed": "Evidencias por completar",
            "open_balance_confirmation": "Saldos abiertos por confirmar",
            "unresolved": "Sin resolver",
        },
        "criteria": {
            "reconciled_strong": "Evidencia externa o cadena contable-bancaria suficiente.",
            "probable_payment": "Movimiento bancario o remesa probable cuya asignación todavía no está confirmada.",
            "accounting_support_needed": "Asiento contable de cierre sin respaldo externo suficiente.",
            "missing_evidence_needed": "Existe un indicio, pero falta evidencia concluyente.",
            "open_balance_confirmation": "Saldo abierto respaldado por evidencia interna.",
            "unresolved": "Vínculo insuficiente en las evidencias disponibles.",
        },
        "side": {
            "customer": "cliente",
            "receivable": "cliente",
            "client": "cliente",
            "supplier": "proveedor",
            "payable": "proveedor",
            "vendor": "proveedor",
            "unknown": "no indicado",
        },
        "owners": {
            "default_entity": "entidad auditada",
            "default_counterparty": "contraparte",
            "counterparty_confirmation": "{entity}; posible confirmación del saldo por {counterparty}",
        },
        "inventory_usage": {
            "open_items": "Población de partidas abiertas o por conciliar.",
            "ledger": "Respaldo de saldos, mayores, auxiliares y registros internos.",
            "journal": "Asientos contables, cierres internos, traspasos y movimientos del periodo.",
            "bank_statement": "Cobros y pagos efectivos según el extracto bancario.",
            "payment_order": "Órdenes o lotes de pago útiles para vincular bancos y facturas.",
            "factoring_statement": "Evidencia del factor, anticipo u operador externo.",
            "compensation_support": "Respaldo documentado de compensación o netting.",
            "unknown": "Fuente adquirida, pero no clasificada automáticamente.",
            "default": "Evidencia disponible en el expediente.",
        },
        "available": {
            "reconciled_strong": "La línea ya está conciliada en el papel de trabajo con evidencia sólida.",
            "probable_payment": "El extracto bancario y el candidato de coincidencia ya constan en el papel de trabajo.",
            "accounting_support_needed": "La partida abierta y el asiento contable ya están identificados.",
            "missing_evidence_needed": "Ya se identificó un indicio o documento puente, pero no es concluyente.",
            "open_balance_confirmation": "La partida abierta y el respaldo interno del saldo ya están disponibles.",
            "unresolved": "La partida abierta está disponible, pero no existe un vínculo suficiente en las evidencias adquiridas.",
            "reference": " Referencia: {reference}.",
        },
        "missing": {
            "reconciled_strong": "No falta ningún elemento para la solicitud dirigida.",
            "probable_payment": "Confirmación de la asignación factura por factura del movimiento bancario o del lote de pagos ya identificado.",
            "accounting_support_needed": "Respaldo externo o explicación documentada del asiento contable: banco, factor, compensación, traspaso o cierre sin efectivo.",
            "missing_evidence_needed": "Documento necesario para convertir el indicio en evidencia concluyente.",
            "open_balance_confirmation": "Confirmación de que la partida seguía abierta a {cutoff}, o evidencia concreta de un cierre no detectado.",
            "unresolved": "Referencia fuente o evidencia concreta de un cierre no detectado.",
        },
        "actions": {
            "reconciled_strong": "No se requiere ninguna solicitud operativa; conserve la referencia en el papel de trabajo.",
            "probable_payment": "Confirme o corrija la asignación propuesta e indique el importe asignado, el lote o SEPA y la fecha de pago de cada factura.",
            "accounting_support_needed": "Envíe únicamente el justificante de contrapartida o explique el cierre sin efectivo; no vuelva a enviar el mayor o el diario si no han cambiado.",
            "missing_evidence_needed": "Indique qué documento ya adquirido cierra la línea o proporcione únicamente el documento pendiente indicado.",
            "open_balance_confirmation": "Si está abierta, confirme el saldo. Si está cerrada, indique el documento de pago, cobro, factor o compensación y la fecha.",
            "unresolved": "Proporcione la referencia de la factura o del auxiliar, o indique si la línea es errónea; si está cerrada, adjunte el documento de cierre.",
        },
        "instructions": {
            "no_reask": "No vuelva a solicitar documentos ya adquiridos",
            "no_reask_body": "El papel de trabajo principal ya contiene las partidas abiertas, los mayores, el diario, los extractos bancarios o del factor, los lotes de pagos y las normalizaciones disponibles.",
            "targeted": "Solicitud dirigida",
            "targeted_body": "Para cada línea, solicite solo el elemento pendiente indicado: asignación, respaldo externo, confirmación de saldo abierto o referencia fuente.",
            "owner": "Destinatario",
            "owner_body": "{entity} debe confirmar o corregir primero con el expediente. {counterparty} solo se necesita para una posible confirmación de saldo o conciliación externa a {cutoff}.",
        },
        "log_path": "Libro de solicitudes dirigidas: {path}",
    },
    "en": {
        "sheet_names": {
            "instructions": "instructions",
            "summary": "summary",
            "inventory": "available_evidence",
            "reconciled_strong": "reconciled_strong",
            "probable_payment": "probable_payments",
            "accounting_support_needed": "accounting_support_needed",
            "missing_evidence_needed": "evidence_to_complete",
            "open_balance_confirmation": "open_balances",
            "unresolved": "unresolved",
        },
        "headers": {
            "principle": "principle",
            "guidance": "guidance",
            "section": "section",
            "rows": "rows",
            "amount_total": "amount_total",
            "evidence_category": "evidence_category",
            "normalized_rows": "normalized_rows",
            "source_files": "source_files",
            "operational_use": "operational_use",
            "row_id": "workpaper_row_id",
            "side": "side",
            "document": "document",
            "document_date": "document_date",
            "amount": "amount",
            "operational_category": "operational_category",
            "operational_criterion": "operational_criterion",
            "operational_owner": "operational_owner",
            "available_evidence": "available_evidence",
            "targeted_missing_item": "targeted_missing_item",
            "existing_reference": "existing_reference",
            "requested_action": "requested_action",
            "evidence_description": "evidence_description",
        },
        "sections": {
            "reconciled_strong": "Reconciled with strong evidence",
            "probable_payment": "Probable payments to allocate",
            "accounting_support_needed": "Accounting entries to support",
            "missing_evidence_needed": "Evidence to complete",
            "open_balance_confirmation": "Open balances to confirm",
            "unresolved": "Unresolved",
        },
        "criteria": {
            "reconciled_strong": "External evidence or sufficient accounting-bank chain.",
            "probable_payment": "Probable bank movement or batch, allocation not yet confirmed.",
            "accounting_support_needed": "Accounting closing entry without sufficient external support.",
            "missing_evidence_needed": "An indicator exists but conclusive evidence is missing.",
            "open_balance_confirmation": "Open balance supported by internal evidence.",
            "unresolved": "Insufficient link in available evidence.",
        },
        "side": {
            "customer": "customer",
            "receivable": "customer",
            "client": "customer",
            "supplier": "supplier",
            "payable": "supplier",
            "vendor": "supplier",
            "unknown": "not specified",
        },
        "owners": {
            "default_entity": "audited entity",
            "default_counterparty": "counterparty",
            "counterparty_confirmation": "{entity}; possible balance confirmation from {counterparty}",
        },
        "inventory_usage": {
            "open_items": "Open-item or disputed population.",
            "ledger": "Balance, subledger, ledger and internal posting support.",
            "journal": "Accounting entries, internal closings, transfers and period movements.",
            "bank_statement": "Actual receipts and payments from bank statement.",
            "payment_order": "Payment orders or batches used to connect bank and invoices.",
            "factoring_statement": "Factoring, advance or external operator evidence.",
            "compensation_support": "Documented set-off or netting support.",
            "unknown": "Source acquired but not automatically classified.",
            "default": "Evidence acquired in the file.",
        },
        "available": {
            "reconciled_strong": "The row is already reconciled in the workpaper with strong evidence.",
            "probable_payment": "Bank statement and match candidate already exist in the workpaper.",
            "accounting_support_needed": "Open item and accounting entry already identified.",
            "missing_evidence_needed": "Bridge document or indicator already identified, but not conclusive.",
            "open_balance_confirmation": "Open item and internal balance support already exist.",
            "unresolved": "Open item exists, with no sufficient link in the acquired evidence.",
            "reference": " Reference: {reference}.",
        },
        "missing": {
            "reconciled_strong": "No missing item for the targeted request.",
            "probable_payment": "Invoice-by-invoice allocation confirmation for the bank movement or payment batch already identified.",
            "accounting_support_needed": "External support or documented explanation for the accounting entry: bank, factor, set-off, transfer or non-cash closing.",
            "missing_evidence_needed": "Document required to turn the indicator into conclusive evidence.",
            "open_balance_confirmation": "Confirmation that the item was still open at {cutoff}, or specific evidence of an undetected closing.",
            "unresolved": "Source reference or specific evidence of an undetected closing.",
        },
        "actions": {
            "reconciled_strong": "No operational request; retain the reference in the workpaper.",
            "probable_payment": "Confirm or correct the proposed allocation, with allocated amount, batch/SEPA and payment date for each invoice.",
            "accounting_support_needed": "Send only the counterparty support or explain the non-cash closing; do not resend unchanged ledger/journal.",
            "missing_evidence_needed": "Indicate which already-acquired document closes the row, or provide only the missing document requested.",
            "open_balance_confirmation": "If open: confirm the balance. If closed: indicate payment/receipt/factor/set-off document and date.",
            "unresolved": "Provide invoice/subledger reference or indicate if the row is wrong; if closed, attach the closing document.",
        },
        "instructions": {
            "no_reask": "Do not request documents already acquired",
            "no_reask_body": "The main workpaper already contains available open items, ledgers, journal, bank/factor statements, payment batches and normalizations.",
            "targeted": "Targeted request",
            "targeted_body": "For each row, request only the indicated missing item: allocation, external support, open-balance confirmation or source reference.",
            "owner": "Recipient",
            "owner_body": "{entity} should first confirm/correct using the file. {counterparty} is only for possible balance confirmation or external reconciliation at {cutoff}.",
        },
        "log_path": "Targeted request workbook: {path}",
    },
}


@dataclass(frozen=True)
class MissingEvidenceRequestPack:
    """Targeted evidence requests and the source inventory behind them."""

    language: str
    instructions: list[dict[str, Any]]
    summary: list[dict[str, Any]]
    evidence_inventory: list[dict[str, Any]]
    request_sections: dict[str, list[dict[str, Any]]]


def clean_text(value: object) -> str:
    return " ".join(str(value or "").strip().split())


def text_for(language: object) -> dict[str, Any]:
    code = normalize_language(language)
    return TEXT.get(code, TEXT["en"])


def parse_decimal(value: object) -> Decimal:
    text = clean_text(value).replace(" ", "")
    if not text:
        return Decimal("0.00")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0.00")


def format_decimal(value: object) -> str:
    return f"{parse_decimal(value):.2f}"


def side_label(value: object, language: object) -> str:
    side = clean_text(value).lower()
    labels = text_for(language)["side"]
    return labels.get(side, labels["unknown"])


def load_sheet_rows(
    workbook_path: str | Path,
    sheet_name: str,
    *,
    retain_workbook_row: bool = False,
) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    header_row = next(sheet.iter_rows(min_row=1, max_row=1), None)
    if header_row is None:
        return []
    headers = [clean_text(cell.value) for cell in header_row]
    rows: list[dict[str, Any]] = []
    for workbook_row, values in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        if not any(value not in (None, "") for value in values):
            continue
        row = dict(zip(headers, values))
        if retain_workbook_row:
            row["workbook_row"] = workbook_row
        rows.append(row)
    return rows


def load_reconciliation_context(
    workbook_path: str | Path,
) -> dict[str, list[dict[str, Any]]]:
    """Load the standard sheets needed for targeted evidence requests."""

    return {
        "reconciliation_rows": load_sheet_rows(
            workbook_path, "Reconciliation detail", retain_workbook_row=True
        ),
        "source_inventory": load_sheet_rows(workbook_path, "Source inventory"),
        "normalized_records": load_sheet_rows(workbook_path, "Normalized records"),
    }


def source_files_by_role(
    source_inventory: list[dict[str, Any]],
    normalized_records: list[dict[str, Any]],
) -> dict[str, set[str]]:
    files_by_role: dict[str, set[str]] = {}
    for row in [*source_inventory, *normalized_records]:
        role = clean_text(row.get("source_role"))
        source_file = clean_text(row.get("source_file"))
        if role and source_file:
            files_by_role.setdefault(role, set()).add(source_file)
    return files_by_role


def record_counts_by_role(normalized_records: list[dict[str, Any]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in normalized_records:
        role = clean_text(row.get("source_role")) or "unknown"
        counts[role] = counts.get(role, 0) + 1
    return counts


def evidence_usage(role: str, language: object) -> str:
    inventory_usage = text_for(language)["inventory_usage"]
    return inventory_usage.get(role, inventory_usage["default"])


def evidence_inventory_rows(
    source_inventory: list[dict[str, Any]],
    normalized_records: list[dict[str, Any]],
    *,
    language: object,
) -> list[dict[str, Any]]:
    files_by_role = source_files_by_role(source_inventory, normalized_records)
    counts_by_role = record_counts_by_role(normalized_records)
    roles = sorted(set(files_by_role) | set(counts_by_role))
    return [
        {
            "evidence_category": evidence_usage(role, language).split(".", 1)[0],
            "normalized_rows": counts_by_role.get(role, 0),
            "source_files": "; ".join(sorted(files_by_role.get(role, set()))),
            "operational_use": evidence_usage(role, language),
        }
        for role in roles
    ]


def operational_owner(
    section: str,
    *,
    entity_name: str,
    counterparty_name: str,
    language: object,
) -> str:
    texts = text_for(language)
    owners = texts["owners"]
    entity = clean_text(entity_name) or owners["default_entity"]
    counterparty = clean_text(counterparty_name) or owners["default_counterparty"]
    if section == "open_balance_confirmation":
        return owners["counterparty_confirmation"].format(
            entity=entity, counterparty=counterparty
        )
    return entity


def section_for_row(row: dict[str, Any]) -> str:
    status = clean_text(row.get("reconciliation_status"))
    rule = clean_text(row.get("rule_applied"))
    if status == "closed":
        return "reconciled_strong"
    if status == "probable_payment":
        return "probable_payment"
    if status == "needs_evidence" and rule == "internal_closure_without_external":
        return "accounting_support_needed"
    if status == "needs_evidence":
        return "missing_evidence_needed"
    if status == "open_supported":
        return "open_balance_confirmation"
    if status == "unresolved":
        return "unresolved"
    return ""


def existing_reference(row: dict[str, Any]) -> str:
    for field in (
        "probable_bank_reference",
        "matched_evidence_reference",
        "evidence_reference",
        "source_reference",
    ):
        value = clean_text(row.get(field))
        if value:
            return value
    source_file = clean_text(row.get("source_file"))
    workbook_row = clean_text(row.get("workbook_row"))
    if source_file and workbook_row:
        return f"{source_file}; workpaper row {workbook_row}"
    return source_file


def existing_description(row: dict[str, Any]) -> str:
    for field in (
        "probable_bank_description",
        "matched_evidence_description",
        "evidence_description",
        "description",
    ):
        value = clean_text(row.get(field))
        if value:
            return value
    return ""


def available_text(
    section: str,
    row: dict[str, Any],
    *,
    language: object,
) -> str:
    texts = text_for(language)
    available = texts["available"][section]
    reference = existing_reference(row)
    if reference:
        available += texts["available"]["reference"].format(reference=reference)
    return available


def missing_text(
    section: str,
    *,
    cutoff_date: object,
    language: object,
) -> str:
    cutoff = clean_text(cutoff_date) or "cut-off"
    return text_for(language)["missing"][section].format(cutoff=cutoff)


def requested_action(section: str, *, language: object) -> str:
    return text_for(language)["actions"][section]


def request_row(
    row: dict[str, Any],
    *,
    section: str,
    entity_name: str,
    counterparty_name: str,
    cutoff_date: object,
    language: object,
) -> dict[str, Any]:
    texts = text_for(language)
    return {
        "row_id": clean_text(row.get("record_id")),
        "side": side_label(row.get("expected_side"), language),
        "document": clean_text(row.get("document_no")),
        "document_date": clean_text(row.get("document_date")),
        "amount": format_decimal(row.get("amount")),
        "operational_category": texts["sections"][section],
        "operational_criterion": texts["criteria"][section],
        "operational_owner": operational_owner(
            section,
            entity_name=entity_name,
            counterparty_name=counterparty_name,
            language=language,
        ),
        "available_evidence": available_text(section, row, language=language),
        "targeted_missing_item": missing_text(
            section, cutoff_date=cutoff_date, language=language
        ),
        "existing_reference": existing_reference(row),
        "requested_action": requested_action(section, language=language),
        "evidence_description": existing_description(row),
    }


def build_instructions(
    *,
    entity_name: str,
    counterparty_name: str,
    cutoff_date: object,
    language: object,
) -> list[dict[str, Any]]:
    texts = text_for(language)
    instructions = texts["instructions"]
    owners = texts["owners"]
    entity = clean_text(entity_name) or owners["default_entity"]
    counterparty = clean_text(counterparty_name) or owners["default_counterparty"]
    cutoff = clean_text(cutoff_date) or "cut-off"
    return [
        {
            "principle": instructions["no_reask"],
            "guidance": instructions["no_reask_body"],
        },
        {
            "principle": instructions["targeted"],
            "guidance": instructions["targeted_body"],
        },
        {
            "principle": instructions["owner"],
            "guidance": instructions["owner_body"].format(
                entity=entity, counterparty=counterparty, cutoff=cutoff
            ),
        },
    ]


def build_summary(
    request_sections: dict[str, list[dict[str, Any]]],
    *,
    language: object,
) -> list[dict[str, Any]]:
    texts = text_for(language)
    rows: list[dict[str, Any]] = []
    for section in SECTION_ORDER:
        section_rows = request_sections.get(section, [])
        amount_total = sum(parse_decimal(row.get("amount")) for row in section_rows)
        rows.append(
            {
                "section": texts["sections"][section],
                "rows": len(section_rows),
                "amount_total": f"{amount_total:.2f}",
            }
        )
    return rows


def build_missing_evidence_request_pack(
    reconciliation_rows: list[dict[str, Any]],
    *,
    source_inventory: list[dict[str, Any]] | None = None,
    normalized_records: list[dict[str, Any]] | None = None,
    entity_name: str = "",
    counterparty_name: str = "",
    cutoff_date: object = "",
    language: object = "it",
) -> MissingEvidenceRequestPack:
    """Build a localized operational request pack from technical workpaper rows."""

    output_language = normalize_language(language)
    request_sections: dict[str, list[dict[str, Any]]] = {
        section: [] for section in SECTION_ORDER
    }
    for row in reconciliation_rows:
        section = section_for_row(row)
        if not section:
            continue
        request_sections[section].append(
            request_row(
                row,
                section=section,
                entity_name=entity_name,
                counterparty_name=counterparty_name,
                cutoff_date=cutoff_date,
                language=output_language,
            )
        )
    return MissingEvidenceRequestPack(
        language=output_language,
        instructions=build_instructions(
            entity_name=entity_name,
            counterparty_name=counterparty_name,
            cutoff_date=cutoff_date,
            language=output_language,
        ),
        summary=build_summary(request_sections, language=output_language),
        evidence_inventory=evidence_inventory_rows(
            source_inventory or [],
            normalized_records or [],
            language=output_language,
        ),
        request_sections=request_sections,
    )


def sheet_headers(
    rows: list[dict[str, Any]],
    fields: tuple[str, ...],
) -> list[str]:
    headers = list(fields)
    extras = sorted({key for row in rows for key in row if key not in fields})
    headers.extend(extras)
    return headers


def write_sheet(
    workbook: Workbook,
    name: str,
    rows: list[dict[str, Any]],
    fields: tuple[str, ...],
    *,
    language: object,
) -> None:
    texts = text_for(language)
    sheet = workbook.create_sheet(title=name[:31])
    keys = sheet_headers(rows, fields)
    display_headers = [texts["headers"].get(key, key) for key in keys]
    sheet.append(display_headers)
    for row in rows:
        sheet.append([row.get(key, "") for key in keys])
    fill = PatternFill(fill_type="solid", fgColor=HEADER_FILL)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for column_cells in sheet.columns:
        max_length = (
            max(
                len(clean_text(cell.value))
                for cell in column_cells
                if cell.value is not None
            )
            if any(cell.value is not None for cell in column_cells)
            else 8
        )
        adjusted_width = min(max(max_length + 2, 10), 60)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = (
            adjusted_width
        )
        for cell in column_cells:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.freeze_panes = "A2"


def write_missing_evidence_workbook(
    path: str | Path,
    pack: MissingEvidenceRequestPack,
) -> Path:
    """Write the localized targeted request workbook."""

    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    texts = text_for(pack.language)
    sheet_names = texts["sheet_names"]
    write_sheet(
        workbook,
        sheet_names["instructions"],
        pack.instructions,
        INSTRUCTION_FIELDS,
        language=pack.language,
    )
    write_sheet(
        workbook,
        sheet_names["summary"],
        pack.summary,
        SUMMARY_FIELDS,
        language=pack.language,
    )
    write_sheet(
        workbook,
        sheet_names["inventory"],
        pack.evidence_inventory,
        INVENTORY_FIELDS,
        language=pack.language,
    )
    for section in SECTION_ORDER:
        write_sheet(
            workbook,
            sheet_names[section],
            pack.request_sections.get(section, []),
            REQUEST_FIELDS,
            language=pack.language,
        )
    workbook.save(output_path)
    return output_path


def default_output_dir(workbook_path: str | Path) -> Path:
    return Path(workbook_path).resolve().parent


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Build localized targeted evidence requests from a reconciliation workbook."
    )
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output-dir", type=Path)
    parser.add_argument("--output-name", default=DEFAULT_EXCEL_NAME)
    parser.add_argument("--entity-name", default="")
    parser.add_argument("--counterparty-name", default="")
    parser.add_argument("--cutoff-date", default="")
    parser.add_argument("--language", default="it")
    args = parser.parse_args(argv)

    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    language = normalize_language(args.language)
    context = load_reconciliation_context(args.workbook)
    pack = build_missing_evidence_request_pack(
        context["reconciliation_rows"],
        source_inventory=context["source_inventory"],
        normalized_records=context["normalized_records"],
        entity_name=args.entity_name,
        counterparty_name=args.counterparty_name,
        cutoff_date=args.cutoff_date,
        language=language,
    )
    output_dir = args.output_dir or default_output_dir(args.workbook)
    output_path = write_missing_evidence_workbook(output_dir / args.output_name, pack)
    LOGGER.info(text_for(language)["log_path"].format(path=output_path))
    for row in pack.summary:
        LOGGER.info(
            "%s: %s / %s",
            row["section"],
            row["rows"],
            row["amount_total"],
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
