"""Build an operational accountant-facing reconciliation workbook."""

from __future__ import annotations

from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

__all__ = [
    "build_accountant_report_rows",
    "write_accountant_report_workbook",
]


HEADER_FILL = "17365D"
LIGHT_BLUE = "E8EEF7"
FILL_BY_CONFIDENCE = {
    "Alta": "D9EAD3",
    "Media": "FFF2CC",
    "Bassa": "FCE4D6",
    "Non trovato": "F4CCCC",
}

MAIN_HEADERS = [
    "id dettaglio",
    "partita",
    "data documento",
    "importo",
    "saldo",
    "data pagamento",
    "banca / fonte incasso",
    "modalita di pagamento",
    "compensazione",
    "non pagata",
    "non capita",
    "stato riscontro",
    "confidenza",
    "evidenza usata",
    "differenza importo",
    "azione richiesta",
    "riferimento fonte",
]

DETAIL_HEADERS = [
    "id dettaglio",
    "partita",
    "tipo evidenza",
    "confidenza",
    "data evidenza",
    "importo evidenza",
    "differenza",
    "fonte / banca",
    "descrizione evidenza",
    "azione richiesta",
    "riferimento fonte",
]


def clean_text(value: object) -> str:
    """Return a trimmed string for optional source values."""

    return "" if value is None else str(value).strip()


def split_values(value: object) -> list[str]:
    """Split semicolon-separated source fields without empty entries."""

    return [part.strip() for part in clean_text(value).split(";") if part.strip()]


def parse_decimal(value: object) -> Decimal | None:
    """Parse accounting amounts stored as strings, floats or Decimals."""

    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, Decimal):
        return value.quantize(Decimal("0.01"))
    if isinstance(value, (int, float)):
        return Decimal(str(value)).quantize(Decimal("0.01"))

    text = clean_text(value)
    if not text:
        return None
    text = text.replace("EUR", "").replace("euro", "").replace(" ", "")
    text = text.replace("(", "-").replace(")", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return Decimal(text).quantize(Decimal("0.01"))
    except InvalidOperation:
        return None


def format_decimal(value: Decimal | None) -> str:
    """Format Decimal amounts for Excel display."""

    if value is None:
        return ""
    return f"{value:.2f}"


def amount_difference(open_amount: object, evidence_amount: object) -> str:
    """Return absolute evidence minus open amount difference when available."""

    left = parse_decimal(open_amount)
    right = parse_decimal(evidence_amount)
    if left is None or right is None:
        return ""
    return format_decimal(right.copy_abs() - left.copy_abs())


def source_reference(row: dict[str, Any]) -> str:
    """Build a compact source reference from standard source fields."""

    parts = []
    source = clean_text(row.get("source_file"))
    if source:
        parts.append(source)
    page = clean_text(row.get("source_page"))
    if page:
        parts.append(f"pag. {page}")
    source_row = clean_text(row.get("source_row"))
    if source_row:
        parts.append(f"riga {source_row}")
    record_id = clean_text(row.get("record_id"))
    if record_id:
        parts.append(record_id)
    return "; ".join(parts)


def source_label(row: dict[str, Any]) -> str:
    """Return a readable source label from source role and filename."""

    source_role = clean_text(row.get("source_role")).lower()
    source_file = clean_text(row.get("source_file"))
    evidence_type = clean_text(row.get("evidence_type")).lower()
    if source_role == "bank_statement" or evidence_type in {
        "external_bank",
        "unallocated_external_bank",
        "grouped_bank_unallocated",
        "unallocated_bank",
    }:
        return source_file or "Estratto banca"
    if source_role == "payment_order" or evidence_type in {
        "payment_order",
        "payment_order_bridge",
    }:
        return source_file or "Distinta pagamento"
    if source_role == "compensation_support" or evidence_type == "compensation":
        return source_file or "Supporto compensazione"
    if "factoring" in evidence_type:
        return source_file or "Factoring / anticipo"
    if source_role in {"journal", "ledger"} or evidence_type.startswith("internal"):
        return source_file or "Scrittura contabile"
    return source_file or "Fonte non determinata"


def evidence_type_label(evidence_type: object, source_role: object = "") -> str:
    """Map internal evidence types to operational Italian labels."""

    evidence = clean_text(evidence_type).lower()
    role = clean_text(source_role).lower()
    if role == "bank_statement" or evidence in {
        "external_bank",
        "unallocated_external_bank",
        "grouped_bank_unallocated",
        "unallocated_bank",
    }:
        return "Bonifico bancario"
    if evidence in {"payment_order", "payment_order_bridge"}:
        return "Distinta pagamento"
    if evidence == "compensation" or role == "compensation_support":
        return "Compensazione / giroconto"
    if "factoring" in evidence:
        return "Factoring / anticipo"
    if evidence in {
        "internal_closure",
        "internal_bank_closure",
        "closure_without_external",
        "internal_accounting",
        "internal_booking",
        "ledger_open_item",
        "open_balance",
    }:
        return "Scrittura contabile interna"
    if evidence:
        return evidence.replace("_", " ")
    return "Evidenza non classificata"


def is_compensation_like(*values: object) -> bool:
    """Return whether any value indicates compensation or netting."""

    blob = " ".join(clean_text(value).lower() for value in values)
    return any(token in blob for token in ("compens", "giroconto", "netting"))


def confidence_for_row(row: dict[str, Any], has_detail: bool) -> str:
    """Translate reconciliation status into an accountant-facing confidence."""

    status = clean_text(row.get("reconciliation_status"))
    evidence_level = clean_text(row.get("evidence_level"))
    evidence_type = clean_text(row.get("matched_evidence_type"))
    if status == "closed":
        return "Alta"
    if status == "probable_payment":
        return "Media"
    if status == "needs_evidence":
        if evidence_type in {
            "external_bank",
            "unallocated_external_bank",
            "payment_order",
            "payment_order_bridge",
            "factoring_bridge",
        }:
            return "Media"
        return "Bassa" if has_detail else "Non trovato"
    if status == "open_supported":
        return "Bassa"
    if status == "out_of_scope":
        return "N/A"
    if evidence_level and evidence_level != "none":
        return "Bassa"
    return "Non trovato"


def status_label(row: dict[str, Any]) -> str:
    """Translate reconciliation status to operational wording."""

    status = clean_text(row.get("reconciliation_status"))
    if status == "closed":
        return "Riscontro forte"
    if status == "probable_payment":
        return "Probabile, da verificare"
    if status == "needs_evidence":
        return "Evidenza parziale / da integrare"
    if status == "open_supported":
        return "Aperta da confermare"
    if status == "out_of_scope":
        return "Non applicabile"
    return "Non trovato"


def default_action(row: dict[str, Any], confidence: str) -> str:
    """Return the next operational action for an accountant-facing row."""

    missing = clean_text(row.get("missing_evidence"))
    if missing:
        return missing
    status = clean_text(row.get("reconciliation_status"))
    if status == "closed":
        return "Conservare il riferimento nel workpaper."
    if status == "probable_payment":
        return "Confermare allocazione fattura-per-fattura del movimento bancario o della distinta."
    if status == "open_supported":
        return "Confermare che la partita resta aperta o indicare la prova puntuale di chiusura."
    if status == "out_of_scope":
        return "Nessuna azione nel perimetro corrente."
    if confidence == "Non trovato":
        return "Richiedere evidenza di pagamento/incasso o spiegazione della partita aperta."
    return "Richiedere evidenza esterna o dettaglio di allocazione riga-documento."


def candidate_matches_row(candidate: dict[str, Any], row: dict[str, Any]) -> bool:
    """Return whether a bank candidate belongs to a reconciliation row."""

    record_id = clean_text(row.get("record_id"))
    if record_id and record_id in split_values(
        candidate.get("candidate_open_record_ids")
    ):
        return True
    document_key = clean_text(row.get("document_key"))
    return bool(
        document_key
        and document_key
        in {
            *split_values(candidate.get("candidate_document_keys")),
            *split_values(candidate.get("bank_reference_keys_found")),
        }
    )


def evidence_record_detail(
    *,
    row_id: str,
    open_row: dict[str, Any],
    evidence: dict[str, Any],
    confidence: str,
) -> dict[str, Any]:
    """Build one detail row from a normalized evidence record."""

    evidence_type = clean_text(evidence.get("evidence_type"))
    source_role = clean_text(evidence.get("source_role"))
    amount = evidence.get("amount") or evidence.get("bank_amount")
    description = (
        clean_text(evidence.get("description"))
        or clean_text(evidence.get("document_no"))
        or evidence_type_label(evidence_type, source_role)
    )
    return {
        "id dettaglio": row_id,
        "partita": clean_text(open_row.get("document_no"))
        or clean_text(open_row.get("document_key")),
        "tipo evidenza": evidence_type_label(evidence_type, source_role),
        "confidenza": confidence,
        "data evidenza": clean_text(
            evidence.get("bank_date")
            or evidence.get("posting_date")
            or evidence.get("value_date")
            or evidence.get("document_date")
        ),
        "importo evidenza": clean_text(amount),
        "differenza": amount_difference(open_row.get("amount"), amount),
        "fonte / banca": source_label(evidence),
        "descrizione evidenza": description[:500],
        "azione richiesta": default_action(open_row, confidence),
        "riferimento fonte": source_reference(evidence),
        "_rank": 20 if confidence == "Alta" else 40,
    }


def matched_reference_detail(
    *, row_id: str, row: dict[str, Any], confidence: str
) -> dict[str, Any] | None:
    """Build a detail row from matched reference fields on the main result."""

    matched_reference = clean_text(row.get("matched_evidence_reference"))
    matched_type = clean_text(row.get("matched_evidence_type"))
    supporting_reference = clean_text(row.get("supporting_bank_reference"))
    if not any(
        [
            matched_reference,
            matched_type,
            clean_text(row.get("matched_evidence_id")),
            supporting_reference,
        ]
    ):
        return None
    amount = clean_text(row.get("matched_evidence_amounts"))
    return {
        "id dettaglio": row_id,
        "partita": clean_text(row.get("document_no"))
        or clean_text(row.get("document_key")),
        "tipo evidenza": evidence_type_label(matched_type),
        "confidenza": confidence,
        "data evidenza": clean_text(row.get("supporting_bank_date")),
        "importo evidenza": amount,
        "differenza": amount_difference(row.get("amount"), amount),
        "fonte / banca": clean_text(row.get("supporting_bank_reference"))
        or "Fonte nel workpaper",
        "descrizione evidenza": clean_text(row.get("supporting_bank_description"))
        or matched_reference
        or matched_type,
        "azione richiesta": default_action(row, confidence),
        "riferimento fonte": supporting_reference or matched_reference,
        "_rank": 30,
    }


def bank_candidate_detail(
    *, row_id: str, row: dict[str, Any], candidate: dict[str, Any]
) -> dict[str, Any]:
    """Build one detail row from an advisory bank-allocation candidate."""

    amount_match = clean_text(candidate.get("candidate_amount_match")).upper() == "YES"
    raw_confidence = clean_text(candidate.get("candidate_confidence")).lower()
    confidence = "Alta" if amount_match and raw_confidence == "high" else "Media"
    return {
        "id dettaglio": row_id,
        "partita": clean_text(row.get("document_no"))
        or clean_text(row.get("document_key")),
        "tipo evidenza": "Banca candidata",
        "confidenza": confidence,
        "data evidenza": clean_text(candidate.get("bank_date")),
        "importo evidenza": clean_text(candidate.get("bank_amount")),
        "differenza": clean_text(candidate.get("amount_difference_bank_minus_open")),
        "fonte / banca": clean_text(candidate.get("bank_source_file"))
        or "Estratto banca",
        "descrizione evidenza": clean_text(candidate.get("bank_description"))[:500],
        "azione richiesta": clean_text(candidate.get("required_follow_up"))
        or "Verificare allocazione bancaria sulla specifica fattura.",
        "riferimento fonte": source_reference(
            {
                "source_file": candidate.get("bank_source_file"),
                "source_page": candidate.get("bank_source_page"),
                "source_row": candidate.get("bank_source_row"),
                "record_id": candidate.get("bank_record_id"),
            }
        ),
        "_rank": 10 if confidence == "Alta" else 35,
    }


def collect_detail_rows(
    *,
    row_id: str,
    row: dict[str, Any],
    evidence_by_id: dict[str, dict[str, Any]],
    evidence_by_document_key: dict[str, list[dict[str, Any]]],
    bank_allocation_candidates: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Collect the detail evidence rows for one reconciliation row."""

    details: list[dict[str, Any]] = []
    confidence = confidence_for_row(row, has_detail=True)
    matched_id = clean_text(row.get("matched_evidence_id"))
    if matched_id and matched_id in evidence_by_id:
        details.append(
            evidence_record_detail(
                row_id=row_id,
                open_row=row,
                evidence=evidence_by_id[matched_id],
                confidence=confidence,
            )
        )
    else:
        fallback = matched_reference_detail(
            row_id=row_id, row=row, confidence=confidence
        )
        if fallback:
            details.append(fallback)

    for candidate in bank_allocation_candidates:
        if candidate_matches_row(candidate, row):
            details.append(
                bank_candidate_detail(row_id=row_id, row=row, candidate=candidate)
            )

    document_key = clean_text(row.get("document_key"))
    if document_key and not details:
        for evidence in evidence_by_document_key.get(document_key, [])[:5]:
            details.append(
                evidence_record_detail(
                    row_id=row_id,
                    open_row=row,
                    evidence=evidence,
                    confidence=confidence_for_row(row, has_detail=True),
                )
            )

    if not details:
        details.append(
            {
                "id dettaglio": row_id,
                "partita": clean_text(row.get("document_no"))
                or clean_text(row.get("document_key")),
                "tipo evidenza": "Nessuna evidenza",
                "confidenza": "Non trovato",
                "data evidenza": "",
                "importo evidenza": "",
                "differenza": "",
                "fonte / banca": "",
                "descrizione evidenza": "Nessun riscontro deterministico nelle fonti indicizzate.",
                "azione richiesta": default_action(row, "Non trovato"),
                "riferimento fonte": "",
                "_rank": 99,
            }
        )

    unique: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str, str]] = set()
    for detail in sorted(
        details,
        key=lambda item: (
            int(item.get("_rank", 99)),
            clean_text(item.get("data evidenza")),
            clean_text(item.get("riferimento fonte")),
        ),
    ):
        key = (
            clean_text(detail.get("tipo evidenza")),
            clean_text(detail.get("data evidenza")),
            clean_text(detail.get("importo evidenza")),
            clean_text(detail.get("riferimento fonte")),
        )
        if key in seen:
            continue
        seen.add(key)
        unique.append(detail)
    return unique[:12]


def build_evidence_indexes(
    normalized_records: list[dict[str, Any]] | None,
) -> tuple[dict[str, dict[str, Any]], dict[str, list[dict[str, Any]]]]:
    """Index non-open evidence records by id and document key."""

    by_id: dict[str, dict[str, Any]] = {}
    by_document_key: dict[str, list[dict[str, Any]]] = {}
    for record in normalized_records or []:
        if clean_text(record.get("evidence_type")) == "open_item":
            continue
        record_id = clean_text(record.get("record_id"))
        if record_id:
            by_id[record_id] = record
        document_key = clean_text(record.get("document_key"))
        if document_key:
            by_document_key.setdefault(document_key, []).append(record)
    return by_id, by_document_key


def main_row_from_detail(
    *, row_id: str, row: dict[str, Any], details: list[dict[str, Any]]
) -> dict[str, Any]:
    """Build the main accountant-facing row from reconciliation data."""

    best = details[0]
    confidence = confidence_for_row(
        row,
        has_detail=clean_text(best.get("tipo evidenza")) != "Nessuna evidenza",
    )
    if confidence == "N/A":
        data_pagamento = "N/A"
        banca = "N/A"
        modalita = "Riga fuori perimetro"
        compensazione = "N/A"
        non_pagata = "N/A"
        non_capita = "N/A"
    else:
        data_pagamento = clean_text(best.get("data evidenza")) or "Non trovato"
        banca = clean_text(best.get("fonte / banca")) or "Nessuna evidenza trovata"
        modalita = clean_text(best.get("tipo evidenza")) or "Non trovato"
        compensation = is_compensation_like(
            row.get("rule_applied"),
            row.get("matched_evidence_type"),
            best.get("tipo evidenza"),
            best.get("descrizione evidenza"),
        )
        compensazione = "SI - vedi dettaglio" if compensation else "NO"
        has_evidence = confidence not in {"Non trovato", "N/A"}
        non_pagata = "NO" if has_evidence else "SI"
        ambiguous = (
            confidence in {"Media", "Bassa", "Non trovato"}
            or len(details) > 1
            or bool(clean_text(best.get("differenza")))
        )
        non_capita = "SI" if ambiguous else "NO"

    evidence_used = clean_text(best.get("descrizione evidenza"))
    return {
        "id dettaglio": row_id,
        "partita": clean_text(row.get("document_no"))
        or clean_text(row.get("document_key")),
        "data documento": clean_text(
            row.get("document_date") or row.get("posting_date")
        ),
        "importo": clean_text(row.get("amount")),
        "saldo": clean_text(row.get("balance")),
        "data pagamento": data_pagamento,
        "banca / fonte incasso": banca,
        "modalita di pagamento": modalita,
        "compensazione": compensazione,
        "non pagata": non_pagata,
        "non capita": non_capita,
        "stato riscontro": status_label(row),
        "confidenza": confidence,
        "evidenza usata": evidence_used[:500] if evidence_used else "Nessun riscontro",
        "differenza importo": clean_text(best.get("differenza")) or "N/A",
        "azione richiesta": default_action(row, confidence),
        "riferimento fonte": clean_text(best.get("riferimento fonte")),
    }


def build_accountant_report_rows(
    reconciliation_rows: list[dict[str, Any]],
    *,
    bank_allocation_candidates: list[dict[str, Any]] | None = None,
    normalized_records: list[dict[str, Any]] | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    """Build main, detail and legend rows for the operational workbook."""

    evidence_by_id, evidence_by_document_key = build_evidence_indexes(
        normalized_records
    )
    bank_candidates = bank_allocation_candidates or []
    main_rows: list[dict[str, Any]] = []
    detail_rows: list[dict[str, Any]] = []
    status_counts: dict[str, int] = {}
    confidence_counts: dict[str, int] = {}

    for index, row in enumerate(reconciliation_rows, start=1):
        row_id = f"R{index:04d}"
        details = collect_detail_rows(
            row_id=row_id,
            row=row,
            evidence_by_id=evidence_by_id,
            evidence_by_document_key=evidence_by_document_key,
            bank_allocation_candidates=bank_candidates,
        )
        main = main_row_from_detail(row_id=row_id, row=row, details=details)
        main_rows.append(main)
        for detail_index, detail in enumerate(details, start=1):
            visible = {key: detail.get(key, "") for key in DETAIL_HEADERS}
            visible["id dettaglio"] = f"{row_id}-{detail_index:03d}"
            detail_rows.append(visible)
        status = clean_text(main.get("stato riscontro"))
        confidence = clean_text(main.get("confidenza"))
        status_counts[status] = status_counts.get(status, 0) + 1
        confidence_counts[confidence] = confidence_counts.get(confidence, 0) + 1

    legend_rows = [
        {
            "campo": "Scopo",
            "valore": (
                "Scheda operativa riga-per-riga: data pagamento/incasso, fonte, "
                "modalita, compensazione, stato, confidenza e azione richiesta."
            ),
        },
        {"campo": "Righe", "valore": len(main_rows)},
        {
            "campo": "Alta",
            "valore": "Evidenza forte o chiusura deterministica con riferimento conservato.",
        },
        {
            "campo": "Media",
            "valore": "Banca, distinta o fonte collegabile ma con allocazione da verificare.",
        },
        {
            "campo": "Bassa",
            "valore": "Solo evidenza interna, ponte o supporto non conclusivo.",
        },
        {
            "campo": "Non trovato",
            "valore": "Nessuna evidenza utile nelle fonti disponibili.",
        },
    ]
    for key, value in sorted(status_counts.items()):
        legend_rows.append({"campo": f"stato: {key}", "valore": value})
    for key, value in sorted(confidence_counts.items()):
        legend_rows.append({"campo": f"confidenza: {key}", "valore": value})
    return main_rows, detail_rows, legend_rows


def write_rows_sheet(
    workbook: Workbook,
    sheet_name: str,
    rows: list[dict[str, Any]],
    headers: list[str],
    widths: dict[str, int],
) -> None:
    """Write a styled worksheet from dictionaries."""

    worksheet = workbook.create_sheet(sheet_name[:31])
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])
    style_worksheet(worksheet, widths)


def style_worksheet(worksheet, widths: dict[str, int]) -> None:
    """Apply filters, table style, widths and wrapping."""

    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        table_name = (
            "".join(ch for ch in worksheet.title if ch.isalnum())[:20] or "Report"
        )
        table = Table(displayName=f"{table_name}Table", ref=worksheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)
    for col_idx in range(1, worksheet.max_column + 1):
        letter = get_column_letter(col_idx)
        worksheet.column_dimensions[letter].width = widths.get(letter, 18)
        for cell in worksheet[letter]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def apply_main_confidence_fill(worksheet) -> None:
    """Color the operational columns by confidence."""

    header_positions = {cell.value: cell.column for cell in worksheet[1]}
    confidence_col = header_positions.get("confidenza")
    if not confidence_col:
        return
    for row_idx in range(2, worksheet.max_row + 1):
        confidence = clean_text(worksheet.cell(row_idx, confidence_col).value)
        fill_color = FILL_BY_CONFIDENCE.get(confidence)
        if not fill_color:
            continue
        fill = PatternFill("solid", fgColor=fill_color)
        for col_idx in range(confidence_col, worksheet.max_column + 1):
            worksheet.cell(row_idx, col_idx).fill = fill


def write_accountant_report_workbook(
    output_path: str | Path,
    reconciliation_rows: list[dict[str, Any]],
    *,
    bank_allocation_candidates: list[dict[str, Any]] | None = None,
    normalized_records: list[dict[str, Any]] | None = None,
) -> Path:
    """Write the standard accountant-facing operational workbook."""

    main_rows, detail_rows, legend_rows = build_accountant_report_rows(
        reconciliation_rows,
        bank_allocation_candidates=bank_allocation_candidates,
        normalized_records=normalized_records,
    )

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    write_rows_sheet(
        workbook,
        "Legenda",
        legend_rows,
        ["campo", "valore"],
        {"A": 28, "B": 90},
    )
    workbook["Legenda"]["A1"].fill = PatternFill("solid", fgColor=HEADER_FILL)
    workbook["Legenda"]["B1"].fill = PatternFill("solid", fgColor=HEADER_FILL)
    write_rows_sheet(
        workbook,
        "Scheda operativa",
        main_rows,
        MAIN_HEADERS,
        {
            "A": 14,
            "B": 22,
            "C": 14,
            "D": 14,
            "E": 14,
            "F": 16,
            "G": 30,
            "H": 24,
            "I": 22,
            "J": 12,
            "K": 12,
            "L": 28,
            "M": 14,
            "N": 56,
            "O": 16,
            "P": 60,
            "Q": 72,
        },
    )
    apply_main_confidence_fill(workbook["Scheda operativa"])
    write_rows_sheet(
        workbook,
        "Dettaglio riscontri",
        detail_rows,
        DETAIL_HEADERS,
        {
            "A": 16,
            "B": 22,
            "C": 26,
            "D": 14,
            "E": 16,
            "F": 16,
            "G": 14,
            "H": 32,
            "I": 66,
            "J": 60,
            "K": 76,
        },
    )
    workbook.save(path)
    return path
