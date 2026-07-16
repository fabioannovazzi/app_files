"""Build a reviewer-friendly sample from an audit reconciliation workbook."""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from .locale_support import normalize_language
except ImportError:  # pragma: no cover - direct import support
    scripts_dir = Path(__file__).resolve().parent
    if str(scripts_dir) not in sys.path:
        sys.path.insert(0, str(scripts_dir))
    from locale_support import normalize_language  # type: ignore


__all__ = [
    "ReviewSample",
    "build_review_sample",
    "load_reconciliation_rows",
    "write_review_sample_workbook",
    "write_review_request",
]


DEFAULT_STATUS = "open_supported"
DEFAULT_EXCEL_NAME = "campione_movimenti_da_controllare.xlsx"
DEFAULT_REQUEST_NAME = "testo_richiesta_controllo.md"
GROUP_TOTAL_RE = re.compile(r"group_open_amount_total=([-+]?\d+(?:[.,]\d+)?)")
GROUP_ROWS_RE = re.compile(r"group_rows=(\d+)")
REFERENCE_FIELD_RE = re.compile(r"(file|page|row)=([^;]+)")
SELECTED_HEADERS = [
    "riga_file_riconciliazione",
    "documento",
    "data_documento",
    "importo",
    "saldo",
    "lato",
    "lettura_operativa",
    "motivo_scelta",
    "file_origine",
    "pagina_origine",
    "riga_origine",
    "riscontro_trovato",
    "cosa_controllare",
    "risposta_attesa",
]
QUESTION_HEADERS = [
    "riga_file_riconciliazione",
    "documento",
    "domanda",
    "risposta_attesa",
    "perché_questa_riga",
]
CRITERIA_HEADERS = ["criterio", "descrizione"]


@dataclass(frozen=True)
class ReviewSample:
    """Rows and reviewer-facing notes selected for follow-up."""

    selected_rows: list[dict[str, Any]]
    related_rows: list[dict[str, Any]]
    review_questions: list[dict[str, Any]]
    selection_criteria: list[dict[str, Any]]


def clean_text(value: object) -> str:
    return " ".join(str(value or "").strip().split())


def parse_decimal(value: object) -> Decimal:
    text = clean_text(value).replace(" ", "")
    if not text:
        return Decimal("0.00")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0.00")


def parse_iso_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean_text(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def format_amount_it(value: object) -> str:
    amount = parse_decimal(value).quantize(Decimal("0.01"))
    sign = "-" if amount < 0 else ""
    raw = f"{abs(amount):,.2f}"
    return f"{sign}{raw.replace(',', 'X').replace('.', ',').replace('X', '.')}"


def side_label(value: object) -> str:
    side = clean_text(value).lower()
    if side in {"customer", "receivable", "client"}:
        return "cliente"
    if side in {"supplier", "payable", "vendor"}:
        return "fornitore"
    return side or "non indicato"


def is_grouped_row(row: dict[str, Any]) -> bool:
    rule = clean_text(row.get("rule_applied")).lower()
    evidence_amounts = clean_text(row.get("matched_evidence_amounts")).lower()
    return "grouped" in rule or "group_rows=" in evidence_amounts


def group_total(row: dict[str, Any]) -> Decimal:
    evidence_amounts = clean_text(row.get("matched_evidence_amounts"))
    match = GROUP_TOTAL_RE.search(evidence_amounts)
    if match:
        return parse_decimal(match.group(1))
    return abs(parse_decimal(row.get("amount")))


def group_rows_count(row: dict[str, Any]) -> str:
    match = GROUP_ROWS_RE.search(clean_text(row.get("matched_evidence_amounts")))
    return match.group(1) if match else ""


def row_sort_amount(row: dict[str, Any]) -> Decimal:
    return max(abs(parse_decimal(row.get("amount"))), group_total(row))


def load_reconciliation_rows(
    workbook_path: str | Path,
    *,
    sheet_name: str = "Reconciliation detail",
) -> list[dict[str, Any]]:
    """Load reconciliation detail rows and retain their Excel row number."""

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Workbook does not contain sheet {sheet_name!r}")
    sheet = workbook[sheet_name]
    header_row = next(sheet.iter_rows(min_row=1, max_row=1), None)
    if header_row is None:
        return []
    headers = [clean_text(cell.value) for cell in header_row]
    rows: list[dict[str, Any]] = []
    for workbook_row, values in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        row = dict(zip(headers, values))
        if any(value not in (None, "") for value in values):
            row["workbook_row"] = workbook_row
            rows.append(row)
    return rows


def first_not_selected(
    rows: list[dict[str, Any]],
    selected_keys: set[str],
    predicate: Callable[[dict[str, Any]], bool] | None = None,
) -> dict[str, Any] | None:
    for row in rows:
        key = clean_text(row.get("record_id")) or str(row.get("workbook_row"))
        if key in selected_keys:
            continue
        if predicate is not None and not predicate(row):
            continue
        return row
    return None


def selection_key(row: dict[str, Any]) -> str:
    return clean_text(row.get("record_id")) or str(row.get("workbook_row"))


def add_selected(
    selected: list[dict[str, Any]],
    selected_keys: set[str],
    row: dict[str, Any] | None,
    reason: str,
) -> None:
    if row is None:
        return
    key = selection_key(row)
    if key in selected_keys:
        return
    candidate = dict(row)
    candidate["selection_reason"] = reason
    selected.append(candidate)
    selected_keys.add(key)


def select_review_rows(
    rows: list[dict[str, Any]],
    *,
    status: str = DEFAULT_STATUS,
    count: int = 3,
) -> list[dict[str, Any]]:
    """Select useful rows for manual review, prioritizing material and grouped cases."""

    candidates = [
        row
        for row in rows
        if clean_text(row.get("reconciliation_status")).lower() == status.lower()
    ]
    if not candidates or count <= 0:
        return []

    by_amount = sorted(candidates, key=row_sort_amount, reverse=True)
    grouped = sorted(
        [row for row in candidates if is_grouped_row(row)],
        key=lambda row: (group_total(row), abs(parse_decimal(row.get("amount")))),
        reverse=True,
    )
    by_age = sorted(
        candidates,
        key=lambda row: parse_iso_date(row.get("document_date")) or date.max,
    )

    selected: list[dict[str, Any]] = []
    selected_keys: set[str] = set()
    add_selected(
        selected,
        selected_keys,
        first_not_selected(by_amount, selected_keys),
        "Importo più materiale tra le righe da verificare.",
    )
    if len(selected) < count:
        add_selected(
            selected,
            selected_keys,
            first_not_selected(grouped, selected_keys),
            "Caso con riscontro interno aggregato su più righe dello stesso documento.",
        )
    if len(selected) < count:
        selected_sides = {
            clean_text(row.get("expected_side")).lower() for row in selected
        }
        add_selected(
            selected,
            selected_keys,
            first_not_selected(
                by_amount,
                selected_keys,
                lambda row: clean_text(row.get("expected_side")).lower()
                not in selected_sides,
            ),
            "Copertura di un lato diverso del rapporto contabile.",
        )
    if len(selected) < count:
        add_selected(
            selected,
            selected_keys,
            first_not_selected(by_age, selected_keys),
            "Data documento più risalente, utile per verificare l'aging a cut-off.",
        )
    for row in by_amount:
        if len(selected) >= count:
            break
        add_selected(
            selected,
            selected_keys,
            row,
            "Ulteriore riga significativa per importo.",
        )
    return selected


def selected_related_rows(
    all_rows: list[dict[str, Any]],
    selected_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Return same-document rows needed to inspect grouped selections."""

    selected_row_numbers = {row.get("workbook_row") for row in selected_rows}
    grouped_keys = {
        clean_text(row.get("document_key"))
        for row in selected_rows
        if is_grouped_row(row) and clean_text(row.get("document_key"))
    }
    grouped_evidence_ids = {
        clean_text(row.get("matched_evidence_id"))
        for row in selected_rows
        if is_grouped_row(row) and clean_text(row.get("matched_evidence_id"))
    }
    related: list[dict[str, Any]] = []
    for row in all_rows:
        document_match = clean_text(row.get("document_key")) in grouped_keys
        evidence_match = (
            clean_text(row.get("matched_evidence_id")) in grouped_evidence_ids
        )
        if document_match or evidence_match:
            candidate = dict(row)
            candidate["is_selected_row"] = (
                "sì" if row.get("workbook_row") in selected_row_numbers else "no"
            )
            related.append(candidate)
    return sorted(related, key=lambda row: int(row.get("workbook_row") or 0))


def operational_status(row: dict[str, Any]) -> str:
    status = clean_text(row.get("reconciliation_status")).lower()
    if status == "open_supported" and is_grouped_row(row):
        return (
            "Risulta ancora aperta; il riscontro è stato trovato sommando più righe "
            "dello stesso documento."
        )
    if status == "open_supported":
        return "Risulta ancora aperta, ma trova riscontro nei mastrini."
    if status == "probable_payment":
        return "Esiste un pagamento bancario probabile; va confermata l'allocazione alla riga."
    if status == "needs_evidence":
        return "Serve evidenza aggiuntiva per decidere se la riga è chiusa."
    if status == "unresolved":
        return "Non è stato trovato un collegamento sufficiente con le evidenze disponibili."
    if status == "closed":
        return "Risulta chiusa sulla base delle evidenze disponibili."
    return status or "Stato non indicato."


def evidence_summary(row: dict[str, Any]) -> str:
    reference = clean_text(row.get("matched_evidence_reference"))
    if reference:
        fields = {
            key: clean_text(value)
            for key, value in REFERENCE_FIELD_RE.findall(reference)
        }
        parts = []
        if fields.get("file"):
            parts.append(fields["file"])
        if fields.get("page"):
            parts.append(f"pagina {fields['page']}")
        if fields.get("row"):
            parts.append(f"riga {fields['row']}")
        if parts:
            return "; ".join(parts)
        return reference.replace("page=", "pagina=").replace("row=", "riga=")
    return "Nessun riscontro specifico indicato."


def review_question(row: dict[str, Any]) -> str:
    document_no = clean_text(row.get("document_no")) or "il documento indicato"
    if is_grouped_row(row):
        rows_count = group_rows_count(row)
        total = format_amount_it(group_total(row))
        detail = f"le {rows_count} righe" if rows_count else "le righe collegate"
        return (
            f"Verificare se {detail} del documento {document_no} vanno lette insieme, "
            f"se il totale {total} era ancora aperto al cut-off e se esistono incassi, "
            "pagamenti, compensazioni, storni o giroconti che chiudono il gruppo in tutto o in parte."
        )
    return (
        f"Verificare se il documento {document_no} era ancora aperto al cut-off oppure "
        "se esistono incassi, pagamenti, compensazioni, storni o giroconti che lo chiudono "
        "in tutto o in parte."
    )


def expected_answer(row: dict[str, Any]) -> str:
    if is_grouped_row(row):
        return (
            "Confermare aperto per l'intero totale, indicare una chiusura totale/parziale, "
            "oppure segnalare che le righe non devono essere raggruppate."
        )
    return (
        "Confermare aperta, indicare una chiusura totale/parziale con data e riferimento, "
        "oppure segnalare un collegamento corretto diverso."
    )


def reviewer_row(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "riga_file_riconciliazione": row.get("workbook_row"),
        "documento": clean_text(row.get("document_no")),
        "data_documento": clean_text(row.get("document_date")),
        "importo": format_amount_it(row.get("amount")),
        "saldo": format_amount_it(row.get("balance") or row.get("amount")),
        "lato": side_label(row.get("expected_side")),
        "lettura_operativa": operational_status(row),
        "motivo_scelta": clean_text(row.get("selection_reason")),
        "file_origine": clean_text(row.get("source_file")),
        "pagina_origine": clean_text(row.get("source_page")),
        "riga_origine": clean_text(row.get("source_row")),
        "riscontro_trovato": evidence_summary(row),
        "cosa_controllare": review_question(row),
        "risposta_attesa": expected_answer(row),
    }


def review_question_row(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "riga_file_riconciliazione": row.get("workbook_row"),
        "documento": clean_text(row.get("document_no")),
        "domanda": review_question(row),
        "risposta_attesa": expected_answer(row),
        "perché_questa_riga": clean_text(row.get("selection_reason")),
    }


def build_review_sample(
    rows: list[dict[str, Any]],
    *,
    status: str = DEFAULT_STATUS,
    count: int = 3,
) -> ReviewSample:
    selected = select_review_rows(rows, status=status, count=count)
    related = selected_related_rows(rows, selected)
    criteria = [
        {
            "criterio": "Materialità",
            "descrizione": "Priorità alle righe con importo più alto.",
        },
        {
            "criterio": "Casi aggregati",
            "descrizione": "Include almeno un caso dove il riscontro nasce dalla somma di più righe, se presente.",
        },
        {
            "criterio": "Copertura",
            "descrizione": "Quando possibile, copre lati contabili o anzianità diverse.",
        },
        {
            "criterio": "Linguaggio operativo",
            "descrizione": "Le domande evitano codici tecnici e indicano cosa verificare nei documenti.",
        },
    ]
    return ReviewSample(
        selected_rows=[reviewer_row(row) for row in selected],
        related_rows=[reviewer_row(row) for row in related],
        review_questions=[review_question_row(row) for row in selected],
        selection_criteria=criteria,
    )


def sheet_headers(
    rows: list[dict[str, Any]], preferred: list[str] | None = None
) -> list[str]:
    if not rows:
        return ["messaggio"]
    preferred_headers = preferred or []
    extras = sorted(
        {key for row in rows for key in row.keys()} - set(preferred_headers)
    )
    return [
        header for header in preferred_headers if any(header in row for row in rows)
    ] + extras


def write_sheet(
    workbook: Workbook,
    name: str,
    rows: list[dict[str, Any]],
    *,
    preferred_headers: list[str] | None = None,
) -> None:
    sheet = workbook.create_sheet(name)
    headers = sheet_headers(rows, preferred_headers)
    sheet.append(headers)
    if rows:
        for row in rows:
            sheet.append([row.get(header) for header in headers])
    else:
        sheet.append(["Nessuna riga."])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.freeze_panes = "A2"
    for index, header in enumerate(headers, start=1):
        width = min(max(len(str(header)) + 2, 14), 64)
        if header in {
            "cosa_controllare",
            "risposta_attesa",
            "riscontro_trovato",
            "domanda",
        }:
            width = 64
        sheet.column_dimensions[get_column_letter(index)].width = width


def write_review_sample_workbook(path: str | Path, sample: ReviewSample) -> Path:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    write_sheet(
        workbook,
        "movimenti_da_controllare",
        sample.selected_rows,
        preferred_headers=SELECTED_HEADERS,
    )
    write_sheet(
        workbook,
        "righe_collegate",
        sample.related_rows,
        preferred_headers=SELECTED_HEADERS,
    )
    write_sheet(
        workbook,
        "domande_per_revisione",
        sample.review_questions,
        preferred_headers=QUESTION_HEADERS,
    )
    write_sheet(
        workbook,
        "criteri_di_scelta",
        sample.selection_criteria,
        preferred_headers=CRITERIA_HEADERS,
    )
    workbook.save(output_path)
    return output_path


def request_markdown(sample: ReviewSample, *, greeting: str = "Ciao,") -> str:
    lines = [
        greeting,
        "",
        "vi mando un piccolo campione di movimenti da controllare nella riconciliazione.",
        "",
        (
            "Ho scelto questi casi perché aiutano a capire se alcune partite che risultano ancora "
            "aperte sono davvero aperte al cut-off, oppure se devono essere considerate chiuse, "
            "parzialmente chiuse o collegate a un movimento diverso."
        ),
        "",
        (
            "In questi casi il controllo ha trovato un riscontro utile, ma serve confermare "
            "se chiude davvero la partita, se la chiude solo in parte o se va lasciata aperta."
        ),
        "",
    ]
    for index, row in enumerate(sample.selected_rows, start=1):
        lines.extend(
            [
                f"{index}) Documento {row.get('lato')} {row.get('documento')}",
                "",
                f"- Riga nel file di riconciliazione: {row.get('riga_file_riconciliazione')}",
                f"- Data documento: {row.get('data_documento')}",
                f"- Importo: {row.get('importo')}",
                f"- File di origine: {row.get('file_origine')}",
                f"- Riferimento nel file di origine: pagina {row.get('pagina_origine')}, riga {row.get('riga_origine')}",
                f"- Riscontro trovato: {row.get('riscontro_trovato')}",
                "",
                str(row.get("cosa_controllare")),
                "",
                f"Risposta utile: {row.get('risposta_attesa')}",
                "",
            ]
        )
    lines.extend(
        [
            (
                "In sintesi, serve capire se questi movimenti devono restare tra le partite aperte "
                "oppure se vanno spostati tra le chiuse, le parzialmente chiuse o le righe da rivedere "
                "con un collegamento diverso."
            ),
            "",
            "Grazie,",
        ]
    )
    return "\n".join(lines)


def write_review_request(
    path: str | Path, sample: ReviewSample, *, greeting: str = "Ciao,"
) -> Path:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        request_markdown(sample, greeting=greeting), encoding="utf-8"
    )
    return output_path


def default_output_dir(workbook_path: Path) -> Path:
    return workbook_path.parent


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Build an operational review sample from a reconciliation workbook."
    )
    parser.add_argument(
        "workbook", help="Path to riconciliazione_audit.xlsx or equivalent workbook."
    )
    parser.add_argument("--output-dir", help="Directory for generated review files.")
    parser.add_argument(
        "--count", type=int, default=3, help="Number of rows to select."
    )
    parser.add_argument(
        "--status", default=DEFAULT_STATUS, help="Technical status to sample from."
    )
    parser.add_argument(
        "--language",
        default="it",
        help="Output language code. Italian is currently implemented.",
    )
    parser.add_argument(
        "--greeting", default="Ciao,", help="Greeting for the Markdown request draft."
    )
    parser.add_argument(
        "--excel-name", default=DEFAULT_EXCEL_NAME, help="Generated workbook file name."
    )
    parser.add_argument(
        "--request-name",
        default=DEFAULT_REQUEST_NAME,
        help="Generated Markdown request file name.",
    )
    args = parser.parse_args(argv)

    language = normalize_language(args.language)
    if language != "it":
        raise ValueError(
            "build_review_sample currently renders reviewer-facing text in Italian only."
        )

    workbook_path = Path(args.workbook)
    output_dir = (
        Path(args.output_dir) if args.output_dir else default_output_dir(workbook_path)
    )
    rows = load_reconciliation_rows(workbook_path)
    sample = build_review_sample(rows, status=args.status, count=args.count)
    excel_path = write_review_sample_workbook(output_dir / args.excel_name, sample)
    request_path = write_review_request(
        output_dir / args.request_name, sample, greeting=args.greeting
    )
    print(f"Review sample workbook: {excel_path}")
    print(f"Review request draft: {request_path}")
    print(f"Selected rows: {len(sample.selected_rows)}")
    print(f"Related rows: {len(sample.related_rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
