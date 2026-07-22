from __future__ import annotations

import argparse
import json
import logging
import math
import re
import sys
import tempfile
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable, Sequence

import openpyxl
import polars as pl

try:
    from .review_session import (
        workbook_sheet_name,
        write_review_session_artifacts,
        write_run_intake,
    )
except ImportError:  # pragma: no cover - supports direct script imports
    import importlib.util

    _review_session_path = Path(__file__).resolve().parent / "review_session.py"
    _review_session_spec = importlib.util.spec_from_file_location(
        "mparanza_journal_sampling_review_session",
        _review_session_path,
    )
    assert _review_session_spec and _review_session_spec.loader
    _review_session = importlib.util.module_from_spec(_review_session_spec)
    sys.modules[_review_session_spec.name] = _review_session
    _review_session_spec.loader.exec_module(_review_session)
    workbook_sheet_name = _review_session.workbook_sheet_name
    write_review_session_artifacts = _review_session.write_review_session_artifacts
    write_run_intake = _review_session.write_run_intake

LOGGER = logging.getLogger(__name__)

SUPPORTED_SUFFIXES = {".csv", ".xls", ".xlsx", ".xlsm", ".pdf"}
SUPPORTED_LANGUAGES = ("it", "en", "fr", "de", "es")
CANONICAL_COLUMNS = [
    "entry_date",
    "movement_number",
    "line_number",
    "account",
    "account_desc",
    "line_desc",
    "debit",
    "credit",
    "amount_signed",
    "amount_abs",
    "source_file",
    "source_sheet",
    "source_page",
    "source_row",
]
DATE_FORMATS = (
    "%Y-%m-%d",
    "%d/%m/%Y",
    "%d-%m-%Y",
    "%d.%m.%Y",
    "%d/%m/%y",
    "%m/%d/%Y",
    "%Y/%m/%d",
)
DATE_TOKEN_RE = re.compile(
    r"\b(\d{1,2}[./-]\d{1,2}[./-]\d{2,4}|\d{4}[./-]\d{1,2}[./-]\d{1,2})\b"
)
ACCOUNT_RE = re.compile(r"^[A-Za-z]?\s*\d+(?:\s*[./-]\s*\d+){0,4}[A-Za-z]?$")
AMOUNT_TOKEN_RE = re.compile(
    r"(?<!\w)\(?-?\d{1,3}(?:[.,\s]\d{3})*(?:[.,]\d{2})\)?(?!\w)"
)


__all__ = [
    "CANONICAL_COLUMNS",
    "InspectionResult",
    "NormalizationResult",
    "SampleResult",
    "language_assumptions",
    "inspect_path",
    "normalize_language",
    "normalize_path",
    "run_sample",
    "write_json",
]


def normalize_language(
    language: object | None,
    *,
    default: str = "en",
    allow_auto: bool = False,
) -> str:
    """Normalize a language tag to one of the supported plugin locales."""

    text = str(language or default).strip().lower().replace("_", "-")
    code = text.split("-", 1)[0]
    if allow_auto and code == "auto":
        return "auto"
    return code if code in SUPPORTED_LANGUAGES else default


def language_assumptions(
    recipe: dict[str, Any],
    *,
    language: object | None = None,
    document_language: object | None = None,
) -> dict[str, str]:
    """Resolve working and source-document language assumptions."""

    working = normalize_language(language or recipe.get("language"), default="en")
    source = normalize_language(
        document_language or recipe.get("document_language") or "auto",
        default=working,
        allow_auto=True,
    )
    return {"language": working, "document_language": source}


@dataclass(frozen=True)
class InspectionResult:
    """Deterministic inspection output for one or more journal files."""

    files: list[dict[str, Any]]
    total_rows: int
    suggested_recipe: dict[str, Any]


@dataclass(frozen=True)
class NormalizationResult:
    """Normalized journal rows plus parser diagnostics."""

    frame: pl.DataFrame
    diagnostics: dict[str, Any]


@dataclass(frozen=True)
class SampleResult:
    """Sample rows plus reproducibility metadata."""

    frame: pl.DataFrame
    audit: dict[str, Any]


def configure_logging(verbose: bool = False) -> None:
    """Configure script logging without affecting imported use."""

    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(level=level, format="%(levelname)s: %(message)s")


def write_json(path: Path, payload: Any) -> None:
    """Write JSON with stable formatting."""

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )


def read_json(path: Path | None) -> dict[str, Any]:
    """Return a JSON object or an empty mapping when no file is provided."""

    if path is None:
        return {}
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError(f"Recipe must be a JSON object: {path}")
    return payload


def supported_files(input_path: Path) -> list[Path]:
    """Return supported journal files from a file or folder path."""

    path = input_path.expanduser()
    if path.is_file():
        return [path] if path.suffix.lower() in SUPPORTED_SUFFIXES else []
    if not path.exists():
        raise FileNotFoundError(f"Input path does not exist: {path}")
    files = [
        candidate
        for candidate in sorted(path.rglob("*"))
        if candidate.is_file()
        and candidate.suffix.lower() in SUPPORTED_SUFFIXES
        and not candidate.name.startswith("~$")
    ]
    return files


def _excel_column_name(index: int) -> str:
    idx = index + 1
    letters: list[str] = []
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        letters.append(chr(65 + rem))
    return "".join(reversed(letters))


def _unique_names(values: Sequence[Any]) -> list[str]:
    names: list[str] = []
    seen: dict[str, int] = {}
    for idx, value in enumerate(values):
        text = _clean_text(value)
        base = (
            text
            if text and text.lower() not in {"none", "nan"}
            else _excel_column_name(idx)
        )
        count = seen.get(base, 0)
        seen[base] = count + 1
        names.append(base if count == 0 else f"{base}_{count + 1}")
    return names


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\u00a0", " ").strip()


def _norm_label(value: Any) -> str:
    return re.sub(r"\s+", " ", _clean_text(value).lower())


def _nonempty_count(row: Sequence[Any]) -> int:
    return sum(1 for value in row if _clean_text(value))


def _read_excel_raw(path: Path) -> pl.DataFrame:
    try:
        return pl.read_excel(
            path,
            has_header=False,
            drop_empty_rows=False,
            drop_empty_cols=False,
        )
    except (ValueError, RuntimeError, OSError) as exc:
        LOGGER.info(
            "Polars Excel read failed for %s; trying openpyxl: %s", path.name, exc
        )

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=False)
    sheet = workbook.worksheets[0]
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    max_width = max((len(row) for row in rows), default=0)
    if max_width == 0:
        return pl.DataFrame()
    data = {
        f"column_{idx}": [row[idx] if idx < len(row) else None for row in rows]
        for idx in range(max_width)
    }
    return pl.DataFrame(data)


def _read_csv_raw(path: Path) -> pl.DataFrame:
    return pl.read_csv(path, has_header=False, infer_schema=False, ignore_errors=True)


def _raw_table(path: Path) -> pl.DataFrame:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv_raw(path)
    if suffix in {".xls", ".xlsx", ".xlsm"}:
        return _read_excel_raw(path)
    raise ValueError(f"Unsupported tabular file: {path}")


def _row_values(df: pl.DataFrame, idx: int) -> list[Any]:
    return list(df.row(idx))


def _suggest_header_rows(df: pl.DataFrame) -> list[int]:
    if df.is_empty():
        return [1]
    best_idx = 0
    best_score = -1
    limit = min(df.height, 20)
    header_tokens = (
        "data",
        "date",
        "datum",
        "conto",
        "compte",
        "konto",
        "account",
        "descrizione",
        "description",
        "beschreibung",
        "libelle",
        "libellé",
        "dare",
        "soll",
        "avere",
        "haben",
        "debit",
        "débit",
        "credit",
        "crédit",
        "amount",
        "montant",
        "betrag",
        "importo",
    )
    for idx in range(limit):
        row = _row_values(df, idx)
        score = 0
        for value in row:
            text = _norm_label(value)
            if not text:
                continue
            score += 1
            if any(token in text for token in header_tokens):
                score += 3
            elif any(ch.isalpha() for ch in text):
                score += 1
        if score > best_score:
            best_idx = idx
            best_score = score
    if best_idx > 0:
        previous = _row_values(df, best_idx - 1)
        current = _row_values(df, best_idx)
        fillable = 0
        for top, base in zip(previous, current):
            if _clean_text(top) and not _clean_text(base):
                fillable += 1
        if fillable >= 2:
            return [best_idx, best_idx + 1]
    return [best_idx + 1]


def _merge_header_rows(rows: Sequence[Sequence[Any]]) -> list[str]:
    if not rows:
        return []
    width = max((len(row) for row in rows), default=0)
    labels: list[str] = []
    for idx in range(width):
        parts = []
        for row in rows:
            value = _clean_text(row[idx] if idx < len(row) else "")
            if value and value.lower() not in {"none", "nan"}:
                parts.append(value)
        labels.append(" ".join(parts))
    return _unique_names(labels)


def _apply_header(df: pl.DataFrame, rows_1_indexed: Sequence[int]) -> pl.DataFrame:
    if df.is_empty():
        return df
    row_indexes = sorted({int(row) - 1 for row in rows_1_indexed})
    if not row_indexes or min(row_indexes) < 0:
        raise ValueError("Header rows must be 1-indexed positive integers.")
    if max(row_indexes) >= df.height:
        raise ValueError("Header row exceeds available rows.")
    labels = _merge_header_rows([_row_values(df, idx) for idx in row_indexes])
    body = df.slice(max(row_indexes) + 1)
    if body.width != len(labels):
        labels = _unique_names(labels[: body.width])
    body.columns = labels
    return _drop_empty_rows(_drop_empty_columns(body))


def _drop_empty_columns(df: pl.DataFrame) -> pl.DataFrame:
    if df.is_empty() or df.width == 0:
        return df
    keep: list[str] = []
    for col in df.columns:
        values = (
            df.get_column(col)
            .cast(pl.Utf8, strict=False)
            .fill_null("")
            .str.strip_chars()
        )
        if int((values != "").sum()) > 0:
            keep.append(col)
    return df.select(keep) if keep else pl.DataFrame()


def _drop_empty_rows(df: pl.DataFrame) -> pl.DataFrame:
    if df.is_empty() or df.width == 0:
        return df
    exprs = [
        pl.col(col).cast(pl.Utf8, strict=False).fill_null("").str.strip_chars() == ""
        for col in df.columns
    ]
    return df.filter(~pl.all_horizontal(exprs))


def _parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        serial = float(value)
        if 20000 <= serial <= 60000:
            return date(1899, 12, 30) + timedelta(days=int(serial))
        return None
    text = _clean_text(value)
    if not text:
        return None
    token_match = DATE_TOKEN_RE.search(text)
    token = token_match.group(1) if token_match else text
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(token, fmt).date()
        except ValueError:
            continue
    return None


def _parse_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if math.isnan(float(value)):
            return None
        return float(value)
    text = _clean_text(value)
    if not text:
        return None
    negative = text.startswith("-") or (text.startswith("(") and text.endswith(")"))
    text = re.sub(r"[^\d,.\-]", "", text)
    if not text or text == "-":
        return None
    if "," in text and "." in text:
        decimal = "," if text.rfind(",") > text.rfind(".") else "."
        thousands = "." if decimal == "," else ","
    elif "," in text:
        decimal = ","
        thousands = "."
    else:
        decimal = "."
        thousands = ","
    text = text.replace(thousands, "").replace(decimal, ".").replace("-", "")
    try:
        number = float(text)
    except ValueError:
        return None
    return -number if negative else number


def _amount_tokens(text: str) -> list[float]:
    values = []
    for match in AMOUNT_TOKEN_RE.finditer(text):
        value = _parse_number(match.group(0))
        if value is not None:
            values.append(value)
    return values


def _date_ratio(series: pl.Series) -> float:
    values = [_parse_date(value) for value in series.drop_nulls().head(100).to_list()]
    if not values:
        return 0.0
    return sum(value is not None for value in values) / len(values)


def _account_ratio(series: pl.Series) -> float:
    values = [_clean_text(value) for value in series.drop_nulls().head(100).to_list()]
    values = [value for value in values if value]
    if not values:
        return 0.0
    matches = sum(bool(ACCOUNT_RE.fullmatch(value)) for value in values)
    return matches / len(values)


def _amount_ratio(series: pl.Series) -> float:
    values = series.drop_nulls().head(100).to_list()
    if not values:
        return 0.0
    matches = sum(_parse_number(value) is not None for value in values)
    return matches / len(values)


def infer_mapping(df: pl.DataFrame) -> dict[str, str | None]:
    """Infer canonical field mapping from headers and simple column profiles."""

    mapping: dict[str, str | None] = {
        "date": None,
        "movement_number": None,
        "account": None,
        "account_desc": None,
        "line_desc": None,
        "debit": None,
        "credit": None,
        "amount": None,
    }
    for col in df.columns:
        label = _norm_label(col)
        if mapping["date"] is None and any(
            token in label for token in ("data", "date", "datum")
        ):
            mapping["date"] = col
        if mapping["movement_number"] is None and any(
            token in label
            for token in (
                "nr. reg",
                "n. reg",
                "movimento",
                "movement",
                "mouvement",
                "bewegung",
                "riga",
                "ligne",
                "zeile",
            )
        ):
            mapping["movement_number"] = col
        if mapping["account"] is None and (
            label in {"conto", "account", "compte", "konto"}
            or (
                any(token in label for token in ("conto", "account", "compte", "konto"))
                and "desc" not in label
                and "descr" not in label
                and "beschreibung" not in label
            )
        ):
            mapping["account"] = col
        if mapping["account_desc"] is None and (
            "descrizione conto" in label
            or "account desc" in label
            or "description compte" in label
            or "kontobeschreibung" in label
            or "konto beschreibung" in label
        ):
            mapping["account_desc"] = col
        if mapping["line_desc"] is None and (
            "descrizione oper" in label
            or "description" in label
            or "libelle" in label
            or "libellé" in label
            or "beschreibung" in label
            or "causale" in label
            or label == "descrizione"
        ):
            mapping["line_desc"] = col
        if mapping["debit"] is None and any(
            token in label for token in ("dare", "debit", "débit", "addebit", "soll")
        ):
            mapping["debit"] = col
        if mapping["credit"] is None and any(
            token in label
            for token in ("avere", "credit", "crédit", "accredit", "haben")
        ):
            mapping["credit"] = col
        if mapping["amount"] is None and any(
            token in label
            for token in ("amount", "importo", "saldo", "montant", "betrag")
        ):
            mapping["amount"] = col

    if mapping["date"] is None:
        candidates = [(col, _date_ratio(df.get_column(col))) for col in df.columns]
        mapping["date"] = (
            max(candidates, key=lambda item: item[1])[0]
            if candidates and max(score for _, score in candidates) >= 0.5
            else None
        )
    if mapping["account"] is None:
        candidates = [(col, _account_ratio(df.get_column(col))) for col in df.columns]
        mapping["account"] = (
            max(candidates, key=lambda item: item[1])[0]
            if candidates and max(score for _, score in candidates) >= 0.4
            else None
        )
    if (
        mapping["debit"] is None
        and mapping["credit"] is None
        and mapping["amount"] is None
    ):
        candidates = [(col, _amount_ratio(df.get_column(col))) for col in df.columns]
        amount_cols = [col for col, score in candidates if score >= 0.3]
        if len(amount_cols) >= 2:
            mapping["debit"] = amount_cols[-2]
            mapping["credit"] = amount_cols[-1]
        elif amount_cols:
            mapping["amount"] = amount_cols[-1]
    return mapping


def _recipe_for_file(recipe: dict[str, Any], path: Path) -> dict[str, Any]:
    files = recipe.get("files")
    if isinstance(files, dict):
        item = files.get(path.name) or files.get(path.as_posix())
        if isinstance(item, dict):
            merged = {key: value for key, value in recipe.items() if key != "files"}
            merged.update(item)
            return merged
    return recipe


def _field(mapping: dict[str, Any], name: str) -> str | None:
    value = mapping.get(name)
    return str(value) if value else None


def _normalize_record(
    source: dict[str, Any],
    *,
    source_file: str,
    source_sheet: str | None,
    source_page: int | None,
    source_row: int | None,
) -> dict[str, Any] | None:
    debit = _parse_number(source.get("debit"))
    credit = _parse_number(source.get("credit"))
    amount = _parse_number(source.get("amount"))
    if amount is not None and debit is None and credit is None:
        amount_signed = amount
        debit = amount if amount >= 0 else None
        credit = abs(amount) if amount < 0 else None
    else:
        amount_signed = float(debit or 0.0) - float(credit or 0.0)
    account = _clean_text(source.get("account"))
    line_desc = _clean_text(source.get("line_desc"))
    account_desc = _clean_text(source.get("account_desc"))
    if not account and not line_desc and amount_signed == 0:
        return None
    entry_date = _parse_date(source.get("entry_date"))
    return {
        "entry_date": entry_date.isoformat() if entry_date else None,
        "movement_number": _clean_text(source.get("movement_number")) or None,
        "line_number": _clean_text(source.get("line_number")) or None,
        "account": account or None,
        "account_desc": account_desc or None,
        "line_desc": line_desc or None,
        "debit": debit,
        "credit": credit,
        "amount_signed": amount_signed,
        "amount_abs": abs(amount_signed),
        "source_file": source_file,
        "source_sheet": source_sheet,
        "source_page": source_page,
        "source_row": source_row,
    }


def _normalize_tabular(
    df: pl.DataFrame, mapping: dict[str, Any], path: Path
) -> pl.DataFrame:
    records: list[dict[str, Any]] = []
    current_date: Any = None
    current_movement: Any = None
    for row_idx, row in enumerate(df.iter_rows(named=True), start=1):
        date_col = _field(mapping, "date")
        movement_col = _field(mapping, "movement_number")
        row_date = row.get(date_col) if date_col else None
        row_movement = row.get(movement_col) if movement_col else None
        if _parse_date(row_date) is not None:
            current_date = row_date
        if _clean_text(row_movement):
            current_movement = row_movement
        source = {
            "entry_date": current_date or row_date,
            "movement_number": current_movement or row_movement,
            "line_number": row.get(_field(mapping, "line_number") or ""),
            "account": row.get(_field(mapping, "account") or ""),
            "account_desc": row.get(_field(mapping, "account_desc") or ""),
            "line_desc": row.get(_field(mapping, "line_desc") or ""),
            "debit": row.get(_field(mapping, "debit") or ""),
            "credit": row.get(_field(mapping, "credit") or ""),
            "amount": row.get(_field(mapping, "amount") or ""),
        }
        record = _normalize_record(
            source,
            source_file=path.name,
            source_sheet=None,
            source_page=None,
            source_row=row_idx,
        )
        if record and (record["account"] or record["amount_abs"] > 0):
            records.append(record)
    return (
        pl.DataFrame(records, schema=CANONICAL_COLUMNS)
        if records
        else pl.DataFrame(schema=CANONICAL_COLUMNS)
    )


def _find_header_index(raw_df: pl.DataFrame, tokens: Sequence[str]) -> int | None:
    for idx in range(min(raw_df.height, 40)):
        labels = [_norm_label(value) for value in _row_values(raw_df, idx)]
        joined = " ".join(labels)
        if all(token in joined for token in tokens):
            return idx
    return None


def _first_number_in_range(
    row: Sequence[Any], start: int | None, width: int = 3
) -> float | None:
    if start is None:
        return None
    for idx in range(start, min(start + width, len(row))):
        value = _parse_number(row[idx])
        if value is not None and abs(value) > 0:
            return value
    return None


def _first_text_in_range(row: Sequence[Any], start: int | None, width: int = 2) -> str:
    if start is None:
        return ""
    for idx in range(start, min(start + width, len(row))):
        text = _clean_text(row[idx])
        if text:
            return text
    return ""


def _find_header_col(header: Sequence[Any], *tokens: str) -> int | None:
    for idx, value in enumerate(header):
        label = _norm_label(value)
        if any(token in label for token in tokens):
            return idx
    return None


def _parse_print_friendly_excel(
    path: Path, raw_df: pl.DataFrame
) -> tuple[pl.DataFrame, dict[str, Any]]:
    header_idx = _find_header_index(raw_df, ("dare", "avere"))
    if header_idx is None:
        return pl.DataFrame(schema=CANONICAL_COLUMNS), {
            "parser": "print_friendly_excel",
            "accepted": False,
        }

    header = _row_values(raw_df, header_idx)
    date_col = _find_header_col(header, "data reg", "data", "date", "datum")
    movement_col = _find_header_col(
        header,
        "nr. reg",
        "n. reg",
        "numero registrazione",
        "movement",
        "mouvement",
        "bewegung",
    )
    line_desc_col = _find_header_col(
        header, "descrizione", "description", "libelle", "libellé", "beschreibung"
    )
    account_col = _find_header_col(header, "conto", "account", "compte", "konto")
    account_desc_col = account_col + 1 if account_col is not None else None
    debit_col = _find_header_col(header, "dare", "debit", "débit", "soll")
    credit_col = _find_header_col(header, "avere", "credit", "crédit", "haben")
    records: list[dict[str, Any]] = []
    current_date: Any = None
    current_movement: Any = None
    current_desc = ""

    for row_idx in range(header_idx + 1, raw_df.height):
        row = _row_values(raw_df, row_idx)
        parsed_date = _parse_date(
            row[date_col] if date_col is not None and date_col < len(row) else None
        )
        if parsed_date is None:
            parsed_date = next(
                (
                    _parse_date(value)
                    for value in row[:4]
                    if _parse_date(value) is not None
                ),
                None,
            )
        if parsed_date is not None:
            current_date = parsed_date
        movement = _first_text_in_range(row, movement_col, width=1)
        if movement:
            current_movement = movement
        row_desc = _first_text_in_range(row, line_desc_col, width=1)
        if row_desc and not ACCOUNT_RE.fullmatch(row_desc):
            current_desc = row_desc
        account = _first_text_in_range(row, account_col, width=1)
        account_desc = _first_text_in_range(row, account_desc_col, width=1)
        debit = _first_number_in_range(row, debit_col, width=3)
        credit = _first_number_in_range(row, credit_col, width=3)
        if not account or (debit is None and credit is None):
            continue
        record = _normalize_record(
            {
                "entry_date": current_date,
                "movement_number": current_movement,
                "account": account,
                "account_desc": account_desc,
                "line_desc": current_desc or row_desc,
                "debit": debit,
                "credit": credit,
            },
            source_file=path.name,
            source_sheet=None,
            source_page=None,
            source_row=row_idx + 1,
        )
        if record:
            records.append(record)

    accepted = bool(records)
    diagnostics = {
        "parser": "print_friendly_excel",
        "accepted": accepted,
        "header_row": header_idx + 1,
        "row_count": len(records),
    }
    frame = (
        pl.DataFrame(records, schema=CANONICAL_COLUMNS)
        if records
        else pl.DataFrame(schema=CANONICAL_COLUMNS)
    )
    return frame, diagnostics


def _extract_pdf_text(path: Path) -> list[tuple[int, str]]:
    import pdfplumber

    lines: list[tuple[int, str]] = []
    with pdfplumber.open(path) as pdf:
        for page_idx, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            for line in text.splitlines():
                clean = line.strip()
                if clean:
                    lines.append((page_idx, clean))
    return lines


def _parse_text_pdf(path: Path) -> tuple[pl.DataFrame, dict[str, Any]]:
    records: list[dict[str, Any]] = []
    lines = _extract_pdf_text(path)
    current_date: date | None = None
    current_movement: str | None = None
    for line_idx, (page, line) in enumerate(lines, start=1):
        parsed_date = _parse_date(line)
        if parsed_date is not None:
            current_date = parsed_date
        account = next(
            (
                token
                for token in line.split()
                if ACCOUNT_RE.fullmatch(token) and _parse_date(token) is None
            ),
            "",
        )
        amounts = _amount_tokens(line)
        if not account or not amounts:
            continue
        debit = amounts[-2] if len(amounts) >= 2 else amounts[-1]
        credit = amounts[-1] if len(amounts) >= 2 else None
        desc = line
        if parsed_date is not None:
            desc = DATE_TOKEN_RE.sub("", desc).strip()
        record = _normalize_record(
            {
                "entry_date": current_date,
                "movement_number": current_movement,
                "account": account,
                "line_desc": desc,
                "debit": debit,
                "credit": credit,
            },
            source_file=path.name,
            source_sheet=None,
            source_page=page,
            source_row=line_idx,
        )
        if record:
            records.append(record)
    diagnostics = {
        "parser": "text_pdf",
        "accepted": bool(records),
        "row_count": len(records),
    }
    frame = (
        pl.DataFrame(records, schema=CANONICAL_COLUMNS)
        if records
        else pl.DataFrame(schema=CANONICAL_COLUMNS)
    )
    return frame, diagnostics


def normalize_file(
    path: Path, recipe: dict[str, Any] | None = None
) -> NormalizationResult:
    """Normalize one journal file with deterministic parser selection."""

    recipe = recipe or {}
    file_recipe = _recipe_for_file(recipe, path)
    suffix = path.suffix.lower()
    diagnostics: dict[str, Any] = {"source_file": path.name}
    if suffix == ".pdf":
        frame, parser_diag = _parse_text_pdf(path)
        diagnostics.update(parser_diag)
        diagnostics["missing_fields"] = _missing_fields(frame)
        return NormalizationResult(frame=frame, diagnostics=diagnostics)

    raw_df = _raw_table(path)
    if suffix in {".xls", ".xlsx", ".xlsm"} and file_recipe.get("parser") != "tabular":
        pf_frame, pf_diag = _parse_print_friendly_excel(path, raw_df)
        if pf_frame.height > 0:
            diagnostics.update(pf_diag)
            diagnostics["missing_fields"] = _missing_fields(pf_frame)
            return NormalizationResult(frame=pf_frame, diagnostics=diagnostics)

    header_rows = file_recipe.get("header_rows") or _suggest_header_rows(raw_df)
    table = _apply_header(raw_df, header_rows)
    mapping = (
        file_recipe.get("mapping")
        if isinstance(file_recipe.get("mapping"), dict)
        else infer_mapping(table)
    )
    frame = _normalize_tabular(table, mapping, path)
    diagnostics.update(
        {
            "parser": "tabular",
            "accepted": frame.height > 0,
            "header_rows": header_rows,
            "mapping": mapping,
            "raw_columns": table.columns,
            "row_count": frame.height,
            "missing_fields": _missing_fields(frame),
        }
    )
    return NormalizationResult(frame=frame, diagnostics=diagnostics)


def _missing_fields(frame: pl.DataFrame) -> list[str]:
    missing: list[str] = []
    for field in ("entry_date", "account", "amount_abs"):
        if field not in frame.columns or frame.is_empty():
            missing.append(field)
            continue
        series = frame.get_column(field)
        if series.null_count() == series.len():
            missing.append(field)
    return missing


def _confidence(diagnostics: dict[str, Any]) -> float:
    if not diagnostics.get("accepted"):
        return 0.0
    missing = set(diagnostics.get("missing_fields") or [])
    score = 0.9
    if "entry_date" in missing:
        score -= 0.2
    if "account" in missing:
        score -= 0.3
    if "amount_abs" in missing:
        score -= 0.4
    return max(score, 0.0)


def _recipe_entry(diag: dict[str, Any]) -> dict[str, Any]:
    entry: dict[str, Any] = {"parser": diag.get("parser")}
    if diag.get("header_rows"):
        entry["header_rows"] = diag["header_rows"]
    if diag.get("mapping"):
        entry["mapping"] = diag["mapping"]
    return entry


def inspect_path(
    input_path: Path,
    output_dir: Path | None = None,
    recipe_path: Path | None = None,
    *,
    language: object | None = None,
    document_language: object | None = None,
) -> InspectionResult:
    """Inspect supported files and optionally write inspection artifacts."""

    recipe = read_json(recipe_path)
    languages = language_assumptions(
        recipe, language=language, document_language=document_language
    )
    files = supported_files(input_path)
    inspections: list[dict[str, Any]] = []
    recipe_files: dict[str, Any] = {}
    total_rows = 0
    for file_path in files:
        result = normalize_file(file_path, recipe)
        preview = result.frame.head(20).to_dicts()
        diag = dict(result.diagnostics)
        diag.update(languages)
        diag["confidence"] = _confidence(diag)
        diag["preview"] = preview
        inspections.append(diag)
        recipe_files[file_path.name] = _recipe_entry(diag)
        total_rows += result.frame.height

    suggested_recipe = {
        "version": 1,
        "description": "Deterministic journal parsing recipe generated by Codex.",
        **languages,
        "files": recipe_files,
    }
    result = InspectionResult(
        files=inspections, total_rows=total_rows, suggested_recipe=suggested_recipe
    )
    if output_dir is not None:
        output_dir.mkdir(parents=True, exist_ok=True)
        write_json(
            output_dir / "inspection.json",
            {"files": inspections, "total_rows": total_rows, **languages},
        )
        write_json(output_dir / "suggested_recipe.json", suggested_recipe)
    return result


def normalize_path(
    input_path: Path,
    output_dir: Path,
    recipe_path: Path | None = None,
    *,
    language: object | None = None,
    document_language: object | None = None,
) -> NormalizationResult:
    """Normalize all supported files under a path and write canonical outputs."""

    recipe = read_json(recipe_path)
    languages = language_assumptions(
        recipe, language=language, document_language=document_language
    )
    frames: list[pl.DataFrame] = []
    diagnostics: list[dict[str, Any]] = []
    for file_path in supported_files(input_path):
        result = normalize_file(file_path, recipe)
        frames.append(result.frame)
        file_diag = dict(result.diagnostics)
        file_diag.update(languages)
        diagnostics.append(file_diag)
    frame = (
        pl.concat(frames, how="diagonal_relaxed")
        if frames
        else pl.DataFrame(schema=CANONICAL_COLUMNS)
    )
    frame = frame.select([col for col in CANONICAL_COLUMNS if col in frame.columns])
    output_dir.mkdir(parents=True, exist_ok=True)
    csv_path = output_dir / "normalized_journal.csv"
    frame.write_csv(csv_path)
    write_json(
        output_dir / "normalization_diagnostics.json",
        {
            "input": input_path.as_posix(),
            "row_count": frame.height,
            **languages,
            "files": diagnostics,
            "output_csv": csv_path.as_posix(),
        },
    )
    return NormalizationResult(
        frame=frame,
        diagnostics={"files": diagnostics, "row_count": frame.height, **languages},
    )


def _apply_filters(
    frame: pl.DataFrame,
    *,
    include_accounts: Sequence[str] = (),
    exclude_accounts: Sequence[str] = (),
    date_start: str | None = None,
    date_end: str | None = None,
    min_abs: float | None = None,
    keyword: str | None = None,
) -> pl.DataFrame:
    result = frame
    if include_accounts:
        result = result.filter(
            pl.col("account").cast(pl.Utf8).is_in(list(include_accounts))
        )
    if exclude_accounts:
        result = result.filter(
            ~pl.col("account").cast(pl.Utf8).is_in(list(exclude_accounts))
        )
    if date_start:
        result = result.filter(
            pl.col("entry_date").cast(pl.Date, strict=False)
            >= pl.lit(date_start).cast(pl.Date)
        )
    if date_end:
        result = result.filter(
            pl.col("entry_date").cast(pl.Date, strict=False)
            <= pl.lit(date_end).cast(pl.Date)
        )
    if min_abs is not None:
        result = result.filter(
            pl.col("amount_abs").cast(pl.Float64, strict=False) >= float(min_abs)
        )
    if keyword:
        lowered = keyword.lower()
        result = result.filter(
            pl.col("line_desc")
            .cast(pl.Utf8, strict=False)
            .fill_null("")
            .str.to_lowercase()
            .str.contains(lowered, literal=True)
        )
    return result


def _systematic_sample(frame: pl.DataFrame, size: int) -> pl.DataFrame:
    if frame.height == 0 or size <= 0:
        return frame.head(0)
    if size >= frame.height:
        return frame
    step = frame.height / size
    indexes = sorted({min(int(idx * step), frame.height - 1) for idx in range(size)})
    return (
        frame.with_row_index("__idx")
        .filter(pl.col("__idx").is_in(indexes))
        .drop("__idx")
    )


def _stratified_sample(
    frame: pl.DataFrame, size: int, group_column: str
) -> pl.DataFrame:
    if frame.height == 0 or size <= 0:
        return frame.head(0)
    if group_column not in frame.columns:
        raise ValueError(f"Stratified sampling group column not found: {group_column}")
    groups = frame.partition_by(group_column, maintain_order=True)
    per_group = max(1, math.ceil(size / len(groups)))
    parts = [
        group.sample(n=min(per_group, group.height), seed=42)
        for group in groups
        if group.height > 0
    ]
    return pl.concat(parts).head(size) if parts else frame.head(0)


def _mus_sample(frame: pl.DataFrame, size: int) -> pl.DataFrame:
    if frame.height == 0 or size <= 0:
        return frame.head(0)
    ordered = frame.with_row_index("__idx").sort("amount_abs", descending=True)
    total = float(ordered.get_column("amount_abs").sum() or 0.0)
    if total <= 0:
        return ordered.head(min(size, ordered.height)).drop("__idx")
    interval = total / min(size, ordered.height)
    thresholds = [(idx + 0.5) * interval for idx in range(min(size, ordered.height))]
    running = 0.0
    picked: list[int] = []
    threshold_idx = 0
    for row in ordered.iter_rows(named=True):
        running += float(row["amount_abs"] or 0.0)
        while threshold_idx < len(thresholds) and running >= thresholds[threshold_idx]:
            picked.append(int(row["__idx"]))
            threshold_idx += 1
    unique = sorted(dict.fromkeys(picked))
    return (
        frame.with_row_index("__idx")
        .filter(pl.col("__idx").is_in(unique))
        .drop("__idx")
        .head(size)
    )


def run_sample(
    normalized_csv: Path,
    output_dir: Path,
    *,
    method: str = "random",
    size: int = 25,
    group_column: str = "account",
    include_accounts: Sequence[str] = (),
    exclude_accounts: Sequence[str] = (),
    date_start: str | None = None,
    date_end: str | None = None,
    min_abs: float | None = None,
    keyword: str | None = None,
    language: object | None = None,
) -> SampleResult:
    """Run deterministic sampling from a normalized journal CSV."""

    method_key = method.strip().lower()
    language_code = normalize_language(language, default="en")
    output_dir.mkdir(parents=True, exist_ok=True)
    run_intake = write_run_intake(
        output_dir,
        normalized_csv=normalized_csv,
        method=method_key,
        size=size,
        group_column=group_column,
        include_accounts=include_accounts,
        exclude_accounts=exclude_accounts,
        date_start=date_start,
        date_end=date_end,
        min_abs=min_abs,
        keyword=keyword,
        language=language_code,
    )
    frame = pl.read_csv(normalized_csv, infer_schema_length=1000)
    population = _apply_filters(
        frame,
        include_accounts=include_accounts,
        exclude_accounts=exclude_accounts,
        date_start=date_start,
        date_end=date_end,
        min_abs=min_abs,
        keyword=keyword,
    )
    if method_key == "random":
        sample = (
            population.sample(n=min(size, population.height), seed=42)
            if population.height
            else population.head(0)
        )
    elif method_key == "systematic":
        sample = _systematic_sample(population, size)
    elif method_key == "stratified":
        sample = _stratified_sample(population, size, group_column)
    elif method_key in {"mus", "monetary unit", "monetary unit sampling"}:
        sample = _mus_sample(population, size)
    else:
        raise ValueError(f"Unsupported sampling method: {method}")

    sample_csv = output_dir / "journal_sample.csv"
    sample_xlsx = output_dir / "journal_sample.xlsx"
    sample.write_csv(sample_csv)
    try:
        sample.write_excel(
            sample_xlsx,
            worksheet=workbook_sheet_name(language_code),
        )
    except (ImportError, ModuleNotFoundError, RuntimeError, ValueError) as exc:
        LOGGER.warning("Could not write XLSX sample; CSV is available: %s", exc)

    audit = {
        "normalized_csv": normalized_csv.as_posix(),
        "language": language_code,
        "method": method_key,
        "seed": 42 if method_key == "random" else None,
        "requested_size": size,
        "population_size_before_filters": frame.height,
        "population_size_after_filters": population.height,
        "sample_size": sample.height,
        "filters": {
            "include_accounts": list(include_accounts),
            "exclude_accounts": list(exclude_accounts),
            "date_start": date_start,
            "date_end": date_end,
            "min_abs": min_abs,
            "keyword": keyword,
        },
        "outputs": {
            "csv": sample_csv.as_posix(),
            "xlsx": sample_xlsx.as_posix() if sample_xlsx.exists() else None,
        },
    }
    write_json(output_dir / "sampling_audit.json", audit)
    review_session = write_review_session_artifacts(
        output_dir,
        run_id=run_intake.run_id,
        run_intake_path=run_intake.path,
        sample=sample,
        audit=audit,
    )
    audit["review_session"] = {
        "run_id": review_session.run_id,
        "run_intake_path": str(review_session.run_intake_path),
        "review_payload_path": str(review_session.review_payload_path),
        "ui_decisions_path": str(review_session.ui_decisions_path),
        "final_artifacts_path": str(review_session.final_artifacts_path),
        "review_item_count": review_session.review_item_count,
    }
    write_json(output_dir / "sampling_audit.json", audit)
    return SampleResult(frame=sample, audit=audit)


def comma_list(value: str | None) -> list[str]:
    """Parse comma-separated CLI values."""

    if not value:
        return []
    return [item.strip() for item in value.split(",") if item.strip()]


def temp_output_dir(prefix: str) -> Path:
    """Create a temporary output directory for Codex scratch runs."""

    return Path(tempfile.mkdtemp(prefix=prefix))


def add_common_args(parser: argparse.ArgumentParser) -> None:
    parser.add_argument(
        "--language",
        help="Working/output language locale: it, en, fr, de, or es. Defaults to recipe or en.",
    )
    parser.add_argument(
        "--document-language",
        help="Source-document language locale: it, en, fr, de, es, or auto. Defaults to recipe or auto.",
    )
    parser.add_argument("--verbose", action="store_true", help="Enable debug logging.")
