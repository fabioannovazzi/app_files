from __future__ import annotations

import argparse
import json
import logging
import math
import re
import sys
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Sequence

import openpyxl
import polars as pl
from excel_sanitization import excel_safe_value

try:
    from .review_session import write_review_session_artifacts, write_run_intake
except ImportError:  # pragma: no cover - supports direct script imports
    import importlib.util

    _review_session_path = Path(__file__).resolve().parent / "review_session.py"
    _review_session_spec = importlib.util.spec_from_file_location(
        "mparanza_journal_bank_reconciliation_review_session",
        _review_session_path,
    )
    assert _review_session_spec and _review_session_spec.loader
    _review_session = importlib.util.module_from_spec(_review_session_spec)
    sys.modules[_review_session_spec.name] = _review_session
    _review_session_spec.loader.exec_module(_review_session)
    write_review_session_artifacts = _review_session.write_review_session_artifacts
    write_run_intake = _review_session.write_run_intake

LOGGER = logging.getLogger(__name__)

SUPPORTED_LANGUAGES = ("it", "en", "fr", "de")
SUPPORTED_INPUT_SUFFIXES = {".csv", ".xls", ".xlsx", ".xlsm", ".pdf"}
TRANSACTION_COLUMNS = [
    "side",
    "transaction_id",
    "transaction_date",
    "amount_signed",
    "amount_abs",
    "description",
    "beneficiary",
    "reference",
    "movement_number",
    "account",
    "source_file",
    "source_row",
]
NON_MOVEMENT_COLUMNS = [
    "side",
    "source_file",
    "source_row",
    "classification",
    "reason",
    "transaction_date",
    "amount_signed",
    "amount_abs",
    "description",
]
MATCH_COLUMNS = [
    "status",
    "stage",
    "bank_transaction_id",
    "journal_transaction_id",
    "bank_date",
    "journal_date",
    "date_diff_days",
    "bank_amount",
    "journal_amount",
    "amount_delta",
    "bank_description",
    "journal_description",
    "shared_references",
    "review_note",
]
DATE_FORMATS = (
    "%Y-%m-%d",
    "%d/%m/%Y",
    "%d-%m-%Y",
    "%d.%m.%Y",
    "%d/%m/%y",
    "%m/%d/%Y",
    "%Y/%m/%d",
)
DATE_TOKEN_RE = re.compile(
    r"\b(\d{1,2}[./-]\d{1,2}[./-]\d{2,4}|\d{4}[./-]\d{1,2}[./-]\d{1,2})\b"
)
AMOUNT_TOKEN_RE = re.compile(
    r"(?<!\w)\(?-?\d{1,3}(?:[.,\s]\d{3})*(?:[.,]\d{2})\)?(?!\w)"
)
WORD_RE = re.compile(r"[a-z0-9]+")
BANK_PDF_NON_MOVEMENT_PATTERNS = (
    (
        "balance",
        "Bank statement balance line, not an ordinary bank movement.",
        re.compile(
            r"\b(?:"
            r"saldo iniziale|saldo finale|saldo precedente|saldo contabile|"
            r"saldo disponibile|saldo al|saldo a inizio|saldo a fine|"
            r"saldo per|"
            r"opening balance|closing balance|initial balance|final balance|"
            r"previous balance|balance brought forward|balance carried forward|"
            r"solde initial|solde final|solde au|"
            r"anfangssaldo|endsaldo|anfangsbestand|schlussbestand|kontostand"
            r")\b"
        ),
    ),
    (
        "total",
        "Bank statement total line, not an ordinary bank movement.",
        re.compile(
            r"^(?:totale|totali|total|totaux|summe|gesamtbetrag)\s+"
            r"(?:(?:de|del|della|dei|degli|des|du|der|die|of)\s+)?"
            r"(?:entrate|uscite|accrediti|addebiti|versamenti|prelievi|"
            r"movimenti|competenze|spese|commissioni|dare|avere|a credito|"
            r"a debito|debits?|credits?|incoming|outgoing|transactions?|"
            r"operations?|charges?|fees?|interests?|interets?|"
            r"gutschriften|belastungen|gebuhren|gebuehren|soll|haben|zinsen)\b"
        ),
    ),
    (
        "scalare",
        "Bank statement scalare/interest summary line, not an ordinary bank movement.",
        re.compile(
            r"\b(?:"
            r"scalare|riassunto scalare|riepilogo scalare|numeri creditori|"
            r"numeri debitori|calcolo competenze|riepilogo competenze|"
            r"competenze scalari|interest scale|interest summary|"
            r"echelle d interets|echelle interets|zinsstaffel|zinsenstaffel"
            r")\b"
        ),
    ),
    (
        "conditions",
        "Bank statement conditions line, not an ordinary bank movement.",
        re.compile(
            r"\b(?:"
            r"condizioni economiche|condizioni applicate|riepilogo condizioni|"
            r"condizioni del conto|conditions applied|account conditions|"
            r"conditions economiques|conditions appliquees|"
            r"kontokonditionen|konditionen|wirtschaftliche bedingungen"
            r")\b"
        ),
    ),
)

__all__ = [
    "InspectionResult",
    "ReconciliationRunResult",
    "add_common_args",
    "configure_logging",
    "excel_safe_value",
    "inspect_inputs",
    "normalize_language",
    "run_reconciliation",
    "write_json",
]


@dataclass(frozen=True)
class InspectionResult:
    """Deterministic inspection output for bank and journal inputs."""

    bank: dict[str, Any]
    journal: dict[str, Any]
    sample: dict[str, Any]
    suggested_recipe: dict[str, Any]


@dataclass(frozen=True)
class ReconciliationRunResult:
    """Reconciliation output plus audit metadata."""

    matches: pl.DataFrame
    unmatched_bank: pl.DataFrame
    unmatched_journal: pl.DataFrame
    audit: dict[str, Any]


def configure_logging(verbose: bool = False) -> None:
    """Configure script logging without affecting imported use."""

    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(level=level, format="%(levelname)s: %(message)s")


def normalize_language(
    language: object | None,
    *,
    default: str = "en",
    allow_auto: bool = False,
) -> str:
    """Normalize a language tag to one supported plugin locale."""

    text = str(language or default).strip().lower().replace("_", "-")
    code = text.split("-", 1)[0]
    if allow_auto and code == "auto":
        return "auto"
    return code if code in SUPPORTED_LANGUAGES else default


def language_assumptions(
    recipe: dict[str, Any],
    *,
    language: object | None = None,
    document_language: object | None = None,
) -> dict[str, str]:
    """Resolve working and source-document language assumptions."""

    working = normalize_language(language or recipe.get("language"), default="en")
    source = normalize_language(
        document_language or recipe.get("document_language") or "auto",
        default=working,
        allow_auto=True,
    )
    return {"language": working, "document_language": source}


def write_json(path: Path, payload: Any) -> None:
    """Write JSON with stable formatting."""

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )


def read_json(path: Path | None) -> dict[str, Any]:
    """Return a JSON object or an empty mapping when no file is provided."""

    if path is None:
        return {}
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError(f"Recipe must be a JSON object: {path}")
    return payload


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\u00a0", " ").replace("\u202f", " ").strip()


def _norm_label(value: Any) -> str:
    text = unicodedata.normalize("NFKD", _clean_text(value))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower()
    return re.sub(r"\s+", " ", text)


def _word_tokens(value: Any) -> set[str]:
    return {
        token
        for token in WORD_RE.findall(_norm_label(value))
        if len(token) >= 3 and token not in {"the", "and", "per", "con", "eur"}
    }


def _excel_column_name(index: int) -> str:
    idx = index + 1
    letters: list[str] = []
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        letters.append(chr(65 + rem))
    return "".join(reversed(letters))


def _unique_names(values: Sequence[Any]) -> list[str]:
    names: list[str] = []
    seen: dict[str, int] = {}
    for idx, value in enumerate(values):
        text = _clean_text(value)
        base = (
            text
            if text and text.lower() not in {"none", "nan"}
            else _excel_column_name(idx)
        )
        count = seen.get(base, 0)
        seen[base] = count + 1
        names.append(base if count == 0 else f"{base}_{count + 1}")
    return names


def supported_files(input_path: Path) -> list[Path]:
    """Return supported files from a file or folder path."""

    path = input_path.expanduser()
    if path.is_file():
        return [path] if path.suffix.lower() in SUPPORTED_INPUT_SUFFIXES else []
    if not path.exists():
        raise FileNotFoundError(f"Input path does not exist: {path}")
    return [
        candidate
        for candidate in sorted(path.rglob("*"))
        if candidate.is_file()
        and candidate.suffix.lower() in SUPPORTED_INPUT_SUFFIXES
        and not candidate.name.startswith("~$")
    ]


def _read_excel_raw(path: Path) -> pl.DataFrame:
    try:
        return pl.read_excel(
            path,
            has_header=False,
            drop_empty_rows=False,
            drop_empty_cols=False,
        )
    except (ValueError, RuntimeError, OSError) as exc:
        LOGGER.info(
            "Polars Excel read failed for %s; trying openpyxl: %s", path.name, exc
        )

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=False)
    sheet = workbook.worksheets[0]
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    max_width = max((len(row) for row in rows), default=0)
    if max_width == 0:
        return pl.DataFrame()
    return pl.DataFrame(
        {
            f"column_{idx}": [row[idx] if idx < len(row) else None for row in rows]
            for idx in range(max_width)
        }
    )


def _read_csv_raw(path: Path) -> pl.DataFrame:
    return pl.read_csv(path, has_header=False, infer_schema=False, ignore_errors=True)


def _read_table_raw(path: Path) -> pl.DataFrame:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv_raw(path)
    if suffix in {".xls", ".xlsx", ".xlsm"}:
        return _read_excel_raw(path)
    raise ValueError(f"Unsupported tabular file: {path}")


def _drop_empty_columns(df: pl.DataFrame) -> pl.DataFrame:
    if df.is_empty() or df.width == 0:
        return df
    keep: list[str] = []
    for col in df.columns:
        values = (
            df.get_column(col)
            .cast(pl.Utf8, strict=False)
            .fill_null("")
            .str.strip_chars()
        )
        if int((values != "").sum()) > 0:
            keep.append(col)
    return df.select(keep) if keep else pl.DataFrame()


def _drop_empty_rows(df: pl.DataFrame) -> pl.DataFrame:
    if df.is_empty() or df.width == 0:
        return df
    exprs = [
        pl.col(col).cast(pl.Utf8, strict=False).fill_null("").str.strip_chars() == ""
        for col in df.columns
    ]
    return df.filter(~pl.all_horizontal(exprs))


def _row_values(df: pl.DataFrame, idx: int) -> list[Any]:
    return list(df.row(idx))


def _suggest_header_rows(df: pl.DataFrame) -> list[int]:
    if df.is_empty():
        return [1]
    tokens = (
        "data",
        "date",
        "datum",
        "amount",
        "importo",
        "montant",
        "betrag",
        "dare",
        "avere",
        "debit",
        "credit",
        "soll",
        "haben",
        "descrizione",
        "description",
        "libelle",
        "libellé",
        "beschreibung",
        "beneficiario",
        "beneficiary",
        "payee",
        "iban",
        "reference",
        "riferimento",
        "movimento",
        "movement",
        "conto",
        "account",
    )
    best_idx = 0
    best_score = -1
    for idx in range(min(df.height, 30)):
        row = _row_values(df, idx)
        score = sum(1 for value in row if _clean_text(value))
        for value in row:
            label = _norm_label(value)
            if any(token in label for token in tokens):
                score += 3
            elif any(ch.isalpha() for ch in label):
                score += 1
        if score > best_score:
            best_score = score
            best_idx = idx
    return [best_idx + 1]


def _merge_header_rows(rows: Sequence[Sequence[Any]]) -> list[str]:
    width = max((len(row) for row in rows), default=0)
    labels: list[str] = []
    for idx in range(width):
        parts = []
        for row in rows:
            value = _clean_text(row[idx] if idx < len(row) else "")
            if value and value.lower() not in {"none", "nan"}:
                parts.append(value)
        labels.append(" ".join(parts))
    return _unique_names(labels)


def _apply_header(df: pl.DataFrame, rows_1_indexed: Sequence[int]) -> pl.DataFrame:
    if df.is_empty():
        return df
    row_indexes = sorted({int(row) - 1 for row in rows_1_indexed})
    if not row_indexes or min(row_indexes) < 0:
        raise ValueError("Header rows must be 1-indexed positive integers.")
    if max(row_indexes) >= df.height:
        raise ValueError("Header row exceeds available rows.")
    labels = _merge_header_rows([_row_values(df, idx) for idx in row_indexes])
    body = df.slice(max(row_indexes) + 1)
    if body.width != len(labels):
        labels = _unique_names(labels[: body.width])
    body.columns = labels
    return _drop_empty_rows(_drop_empty_columns(body))


def _parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        serial = float(value)
        if 20000 <= serial <= 60000:
            return date(1899, 12, 30) + timedelta(days=int(serial))
        return None
    text = _clean_text(value)
    if not text:
        return None
    match = DATE_TOKEN_RE.search(text)
    token = match.group(1) if match else text
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(token, fmt).date()
        except ValueError:
            continue
    return None


def _parse_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if math.isnan(float(value)):
            return None
        return float(value)
    text = _clean_text(value)
    if not text:
        return None
    negative = text.startswith("-") or (text.startswith("(") and text.endswith(")"))
    text = re.sub(r"[^\d,.\-]", "", text)
    if not text or text == "-":
        return None
    if "," in text and "." in text:
        decimal = "," if text.rfind(",") > text.rfind(".") else "."
        thousands = "." if decimal == "," else ","
    elif "," in text:
        decimal = ","
        thousands = "."
    else:
        decimal = "."
        thousands = ","
    text = text.replace(thousands, "").replace(decimal, ".").replace("-", "")
    try:
        number = float(text)
    except ValueError:
        return None
    return -number if negative else number


def _date_ratio(series: pl.Series) -> float:
    values = [_parse_date(value) for value in series.drop_nulls().head(100).to_list()]
    if not values:
        return 0.0
    return sum(value is not None for value in values) / len(values)


def _amount_ratio(series: pl.Series) -> float:
    values = series.drop_nulls().head(100).to_list()
    if not values:
        return 0.0
    return sum(_parse_number(value) is not None for value in values) / len(values)


def _first_matching_column(df: pl.DataFrame, tokens: Sequence[str]) -> str | None:
    for col in df.columns:
        label = _norm_label(col)
        if any(token in label for token in tokens):
            return col
    return None


def infer_mapping(df: pl.DataFrame, side: str) -> dict[str, str | None]:
    """Infer transaction fields from headers and lightweight profiles."""

    mapping: dict[str, str | None] = {
        "date": _first_matching_column(df, ("data", "date", "datum", "booking")),
        "amount": _first_matching_column(
            df, ("amount", "importo", "montant", "betrag", "saldo", "total")
        ),
        "debit": _first_matching_column(df, ("dare", "debit", "addebito", "soll")),
        "credit": _first_matching_column(
            df, ("avere", "credit", "credito", "haben", "accredito")
        ),
        "description": _first_matching_column(
            df,
            (
                "descrizione",
                "description",
                "causale",
                "libelle",
                "libellé",
                "beschreibung",
                "narrative",
                "details",
            ),
        ),
        "beneficiary": _first_matching_column(
            df,
            (
                "beneficiario",
                "beneficiary",
                "payee",
                "payer",
                "cliente",
                "fornitore",
                "counterparty",
                "contrepartie",
                "beguenstigter",
                "begünstigter",
            ),
        ),
        "reference": _first_matching_column(
            df, ("reference", "riferimento", "document", "doc", "cro", "trn", "iban")
        ),
        "movement_number": _first_matching_column(
            df,
            (
                "movement",
                "movimento",
                "nr. reg",
                "n. reg",
                "registrazione",
                "journal",
                "beleg",
            ),
        ),
        "account": _first_matching_column(df, ("conto", "account", "iban", "konto")),
    }
    if mapping["date"] is None:
        candidates = [(col, _date_ratio(df.get_column(col))) for col in df.columns]
        if candidates and max(score for _, score in candidates) >= 0.5:
            mapping["date"] = max(candidates, key=lambda item: item[1])[0]
    if (
        mapping["amount"] is None
        and mapping["debit"] is None
        and mapping["credit"] is None
    ):
        candidates = [(col, _amount_ratio(df.get_column(col))) for col in df.columns]
        amount_cols = [col for col, score in candidates if score >= 0.3]
        if len(amount_cols) >= 2 and side == "journal":
            mapping["debit"] = amount_cols[-2]
            mapping["credit"] = amount_cols[-1]
        elif amount_cols:
            mapping["amount"] = amount_cols[-1]
    return mapping


def _mapping_for_file(recipe: dict[str, Any], side: str, path: Path) -> dict[str, Any]:
    side_recipe = recipe.get(side) if isinstance(recipe.get(side), dict) else {}
    files = side_recipe.get("files")
    if isinstance(files, dict):
        file_recipe = files.get(path.name) or files.get(path.as_posix())
        if isinstance(file_recipe, dict):
            merged = {
                key: value for key, value in side_recipe.items() if key != "files"
            }
            merged.update(file_recipe)
            return merged
    return side_recipe


def _mapped(row: dict[str, Any], mapping: dict[str, Any], key: str) -> Any:
    col = mapping.get(key)
    return row.get(str(col)) if col else None


def _amount_from_row(row: dict[str, Any], mapping: dict[str, Any]) -> float | None:
    amount = _parse_number(_mapped(row, mapping, "amount"))
    if amount is not None:
        return amount
    debit = _parse_number(_mapped(row, mapping, "debit"))
    credit = _parse_number(_mapped(row, mapping, "credit"))
    if debit is None and credit is None:
        return None
    return float(debit or 0.0) - float(credit or 0.0)


def _reference_tokens(*values: Any) -> set[str]:
    tokens: set[str] = set()
    for value in values:
        text = _norm_label(value)
        for token in re.findall(r"[a-z0-9]{4,}", text):
            if token.isdigit() and len(token) == 4 and token.startswith(("19", "20")):
                continue
            tokens.add(token)
    return tokens


def _transaction_frame(records: list[dict[str, Any]]) -> pl.DataFrame:
    if not records:
        return pl.DataFrame(schema={col: pl.Utf8 for col in TRANSACTION_COLUMNS})
    frame = pl.DataFrame(records, infer_schema_length=None)
    for col in TRANSACTION_COLUMNS:
        if col not in frame.columns:
            frame = frame.with_columns(pl.lit(None).alias(col))
    return frame.select(TRANSACTION_COLUMNS)


def _non_movement_frame(records: list[dict[str, Any]]) -> pl.DataFrame:
    if not records:
        return pl.DataFrame(schema={col: pl.Utf8 for col in NON_MOVEMENT_COLUMNS})
    frame = pl.DataFrame(records, infer_schema_length=None)
    for col in NON_MOVEMENT_COLUMNS:
        if col not in frame.columns:
            frame = frame.with_columns(pl.lit(None).alias(col))
    return frame.select(NON_MOVEMENT_COLUMNS)


def _non_movement_records(
    diagnostics: Sequence[dict[str, Any]],
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for diag in diagnostics:
        rows = diag.get("non_movement_rows")
        if isinstance(rows, list):
            records.extend(row for row in rows if isinstance(row, dict))
    return records


def _normalize_table(
    path: Path, side: str, recipe: dict[str, Any]
) -> tuple[pl.DataFrame, dict[str, Any]]:
    raw = _read_table_raw(path)
    file_recipe = _mapping_for_file(recipe, side, path)
    header_rows = file_recipe.get("header_rows") or _suggest_header_rows(raw)
    table = _apply_header(raw, header_rows)
    mapping = (
        file_recipe.get("mapping")
        if isinstance(file_recipe.get("mapping"), dict)
        else infer_mapping(table, side)
    )
    records: list[dict[str, Any]] = []
    for row_idx, row in enumerate(table.iter_rows(named=True), start=1):
        amount = _amount_from_row(row, mapping)
        parsed_date = _parse_date(_mapped(row, mapping, "date"))
        description = _clean_text(_mapped(row, mapping, "description"))
        beneficiary = _clean_text(_mapped(row, mapping, "beneficiary"))
        reference = _clean_text(_mapped(row, mapping, "reference"))
        movement = _clean_text(_mapped(row, mapping, "movement_number"))
        account = _clean_text(_mapped(row, mapping, "account"))
        if amount is None and not description and not reference and not movement:
            continue
        transaction_id = f"{path.stem}:{side}:{row_idx}"
        records.append(
            {
                "side": side,
                "transaction_id": transaction_id,
                "transaction_date": parsed_date.isoformat() if parsed_date else None,
                "amount_signed": amount,
                "amount_abs": abs(amount) if amount is not None else None,
                "description": description or None,
                "beneficiary": beneficiary or None,
                "reference": reference or None,
                "movement_number": movement or None,
                "account": account or None,
                "source_file": path.name,
                "source_row": row_idx,
            }
        )
    frame = _transaction_frame(records)
    diagnostics = {
        "source_file": path.name,
        "parser": "tabular",
        "header_rows": header_rows,
        "mapping": mapping,
        "raw_columns": table.columns,
        "row_count": frame.height,
        "preview": frame.head(20).to_dicts(),
        "missing_required_mapping": _missing_mapping(mapping),
    }
    return frame, diagnostics


def _extract_pdf_text(path: Path) -> str:
    import pdfplumber

    lines: list[str] = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            lines.extend(line.strip() for line in text.splitlines() if line.strip())
    return "\n".join(lines)


def _classify_bank_pdf_non_movement_line(line: str) -> tuple[str, str] | None:
    """Classify mechanically identifiable bank statement summary rows.

    These deterministic rules are intentionally narrow: they remove only explicit
    balance, total, scalare, and conditions lines that are statement metadata,
    while leaving ambiguous fee/payment descriptions in the matching population.
    """

    label = _norm_label(line)
    for classification, reason, pattern in BANK_PDF_NON_MOVEMENT_PATTERNS:
        if pattern.search(label):
            return classification, reason
    return None


def _count_classifications(rows: Sequence[dict[str, Any]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in rows:
        classification = _clean_text(row.get("classification"))
        if classification:
            counts[classification] = counts.get(classification, 0) + 1
    return counts


def _normalize_text_pdf(path: Path, side: str) -> tuple[pl.DataFrame, dict[str, Any]]:
    text = _extract_pdf_text(path)
    records: list[dict[str, Any]] = []
    non_movement_rows: list[dict[str, Any]] = []
    for line_idx, line in enumerate(text.splitlines(), start=1):
        parsed_date = _parse_date(line)
        amounts = [
            _parse_number(match.group(0)) for match in AMOUNT_TOKEN_RE.finditer(line)
        ]
        amounts = [value for value in amounts if value is not None]
        if not amounts:
            continue
        amount = amounts[-1]
        classification = (
            _classify_bank_pdf_non_movement_line(line) if side == "bank" else None
        )
        if classification is not None:
            class_name, reason = classification
            non_movement_rows.append(
                {
                    "side": side,
                    "source_file": path.name,
                    "source_row": line_idx,
                    "classification": class_name,
                    "reason": reason,
                    "transaction_date": (
                        parsed_date.isoformat() if parsed_date else None
                    ),
                    "amount_signed": amount,
                    "amount_abs": abs(amount),
                    "description": line,
                }
            )
            continue
        transaction_id = f"{path.stem}:{side}:{line_idx}"
        records.append(
            {
                "side": side,
                "transaction_id": transaction_id,
                "transaction_date": parsed_date.isoformat() if parsed_date else None,
                "amount_signed": amount,
                "amount_abs": abs(amount),
                "description": line,
                "beneficiary": None,
                "reference": None,
                "movement_number": None,
                "account": None,
                "source_file": path.name,
                "source_row": line_idx,
            }
        )
    frame = _transaction_frame(records)
    return frame, {
        "source_file": path.name,
        "parser": "text_pdf",
        "row_count": frame.height,
        "excluded_non_movement_row_count": len(non_movement_rows),
        "non_movement_classifications": _count_classifications(non_movement_rows),
        "non_movement_rows": non_movement_rows,
        "preview": frame.head(20).to_dicts(),
        "missing_required_mapping": [],
    }


def _normalize_files(
    input_path: Path, side: str, recipe: dict[str, Any]
) -> tuple[pl.DataFrame, list[dict[str, Any]]]:
    frames: list[pl.DataFrame] = []
    diagnostics: list[dict[str, Any]] = []
    for file_path in supported_files(input_path):
        if file_path.suffix.lower() == ".pdf":
            frame, diag = _normalize_text_pdf(file_path, side)
        else:
            frame, diag = _normalize_table(file_path, side, recipe)
        frames.append(frame)
        diagnostics.append(diag)
    if not frames:
        return _transaction_frame([]), diagnostics
    return pl.concat(frames, how="diagonal_relaxed"), diagnostics


def _missing_mapping(mapping: dict[str, Any]) -> list[str]:
    missing = []
    if not mapping.get("date"):
        missing.append("date")
    if not (mapping.get("amount") or mapping.get("debit") or mapping.get("credit")):
        missing.append("amount or debit/credit")
    return missing


def _recipe_side(diagnostics: list[dict[str, Any]]) -> dict[str, Any]:
    files = {}
    for diag in diagnostics:
        entry: dict[str, Any] = {"parser": diag.get("parser")}
        if diag.get("header_rows"):
            entry["header_rows"] = diag["header_rows"]
        if diag.get("mapping"):
            entry["mapping"] = diag["mapping"]
        files[str(diag["source_file"])] = entry
    return {"files": files}


def _read_sample_movements(sample_path: Path | None) -> set[str]:
    if sample_path is None:
        return set()
    frames, _ = _normalize_files(sample_path, "sample", {})
    if frames.is_empty():
        return set()
    values: set[str] = set()
    for row in frames.to_dicts():
        movement = _clean_text(row.get("movement_number"))
        reference = _clean_text(row.get("reference"))
        for value in (movement, reference):
            if value:
                values.add(value)
    if values:
        return values
    # Fallback for simple sample files where the first column is the movement id.
    first_file = (
        supported_files(sample_path)[0] if supported_files(sample_path) else None
    )
    if first_file and first_file.suffix.lower() != ".pdf":
        raw = _read_table_raw(first_file)
        table = _apply_header(raw, _suggest_header_rows(raw))
        if table.width:
            col = table.columns[0]
            return {
                _clean_text(value)
                for value in table.get_column(col).cast(pl.Utf8, strict=False).to_list()
                if _clean_text(value)
            }
    return set()


def inspect_inputs(
    bank_path: Path,
    journal_path: Path,
    output_dir: Path,
    recipe_path: Path | None = None,
    *,
    sample_path: Path | None = None,
    language: object | None = None,
    document_language: object | None = None,
) -> InspectionResult:
    """Inspect input files and write deterministic recipe artifacts."""

    recipe = read_json(recipe_path)
    languages = language_assumptions(
        recipe, language=language, document_language=document_language
    )
    bank_frame, bank_diag = _normalize_files(bank_path, "bank", recipe)
    journal_frame, journal_diag = _normalize_files(journal_path, "journal", recipe)
    sample_movements = _read_sample_movements(sample_path)
    suggested_recipe = {
        "version": 1,
        "description": "Deterministic journal-bank reconciliation recipe generated by Codex.",
        **languages,
        "bank": _recipe_side(bank_diag),
        "journal": _recipe_side(journal_diag),
        "matching": {
            "amount_tolerance": 1.0,
            "date_window_days": 7,
            "use_absolute_amounts": True,
            "stages": [
                "reference",
                "amount_date_unique",
                "beneficiary",
                "description_tokens",
                "amount_date_single",
            ],
        },
    }
    output_dir.mkdir(parents=True, exist_ok=True)
    write_json(
        output_dir / "inspection.json",
        {
            **languages,
            "bank": {"row_count": bank_frame.height, "files": bank_diag},
            "bank_pdf_non_movement_row_count": len(_non_movement_records(bank_diag)),
            "journal": {"row_count": journal_frame.height, "files": journal_diag},
            "sample": {
                "path": sample_path.as_posix() if sample_path else None,
                "movement_count": len(sample_movements),
                "movements": sorted(sample_movements)[:100],
            },
        },
    )
    write_json(output_dir / "suggested_recipe.json", suggested_recipe)
    return InspectionResult(
        bank={"row_count": bank_frame.height, "files": bank_diag},
        journal={"row_count": journal_frame.height, "files": journal_diag},
        sample={"movement_count": len(sample_movements)},
        suggested_recipe=suggested_recipe,
    )


def _date_diff_days(left: Any, right: Any) -> int | None:
    left_date = _parse_date(left)
    right_date = _parse_date(right)
    if left_date is None or right_date is None:
        return None
    return abs((left_date - right_date).days)


def _candidate_rows(
    bank_row: dict[str, Any],
    journal_rows: list[dict[str, Any]],
    used_journal: set[str],
    *,
    tolerance: float,
    date_window_days: int,
) -> list[dict[str, Any]]:
    candidates = []
    bank_amount = (
        float(bank_row["amount_abs"])
        if bank_row.get("amount_abs") is not None
        else None
    )
    for journal_row in journal_rows:
        journal_id = str(journal_row["transaction_id"])
        if journal_id in used_journal:
            continue
        journal_amount = (
            float(journal_row["amount_abs"])
            if journal_row.get("amount_abs") is not None
            else None
        )
        if bank_amount is None or journal_amount is None:
            continue
        amount_delta = abs(bank_amount - journal_amount)
        if amount_delta > tolerance:
            continue
        date_diff = _date_diff_days(
            bank_row.get("transaction_date"), journal_row.get("transaction_date")
        )
        if date_diff is not None and date_diff > date_window_days:
            continue
        candidates.append(
            {
                "row": journal_row,
                "amount_delta": amount_delta,
                "date_diff_days": date_diff,
                "shared_references": sorted(
                    _reference_tokens(
                        bank_row.get("reference"), bank_row.get("description")
                    )
                    & _reference_tokens(
                        journal_row.get("reference"), journal_row.get("description")
                    )
                ),
            }
        )
    return candidates


def _beneficiary_match(bank_row: dict[str, Any], journal_row: dict[str, Any]) -> bool:
    left = _word_tokens(bank_row.get("beneficiary")) or _word_tokens(
        bank_row.get("description")
    )
    right = _word_tokens(journal_row.get("beneficiary")) or _word_tokens(
        journal_row.get("description")
    )
    return bool(
        left and right and (left <= right or right <= left or len(left & right) >= 2)
    )


def _description_overlap(
    bank_row: dict[str, Any], journal_row: dict[str, Any]
) -> float:
    left = _word_tokens(bank_row.get("description"))
    right = _word_tokens(journal_row.get("description"))
    if not left or not right:
        return 0.0
    return len(left & right) / max(1, min(len(left), len(right)))


def _make_match_record(
    bank_row: dict[str, Any],
    journal_row: dict[str, Any],
    *,
    stage: str,
    amount_delta: float,
    date_diff_days: int | None,
    shared_references: Sequence[str],
) -> dict[str, Any]:
    return {
        "status": "matched",
        "stage": stage,
        "bank_transaction_id": bank_row["transaction_id"],
        "journal_transaction_id": journal_row["transaction_id"],
        "bank_date": bank_row.get("transaction_date"),
        "journal_date": journal_row.get("transaction_date"),
        "date_diff_days": date_diff_days,
        "bank_amount": bank_row.get("amount_signed"),
        "journal_amount": journal_row.get("amount_signed"),
        "amount_delta": amount_delta,
        "bank_description": bank_row.get("description"),
        "journal_description": journal_row.get("description"),
        "shared_references": ",".join(shared_references),
        "review_note": "",
    }


def _match_transactions(
    bank: pl.DataFrame,
    journal: pl.DataFrame,
    *,
    tolerance: float,
    date_window_days: int,
) -> tuple[pl.DataFrame, pl.DataFrame, pl.DataFrame, dict[str, int]]:
    bank_rows = bank.to_dicts()
    journal_rows = journal.to_dicts()
    used_bank: set[str] = set()
    used_journal: set[str] = set()
    matches: list[dict[str, Any]] = []
    stage_counts: dict[str, int] = {}

    def accept(bank_row: dict[str, Any], candidate: dict[str, Any], stage: str) -> None:
        matches.append(
            _make_match_record(
                bank_row,
                candidate["row"],
                stage=stage,
                amount_delta=float(candidate["amount_delta"]),
                date_diff_days=candidate["date_diff_days"],
                shared_references=candidate["shared_references"],
            )
        )
        used_bank.add(str(bank_row["transaction_id"]))
        used_journal.add(str(candidate["row"]["transaction_id"]))
        stage_counts[stage] = stage_counts.get(stage, 0) + 1

    for stage in (
        "reference",
        "amount_date_unique",
        "beneficiary",
        "description_tokens",
        "amount_date_single",
    ):
        for bank_row in bank_rows:
            if str(bank_row["transaction_id"]) in used_bank:
                continue
            candidates = _candidate_rows(
                bank_row,
                journal_rows,
                used_journal,
                tolerance=tolerance,
                date_window_days=date_window_days,
            )
            if not candidates:
                continue
            chosen: dict[str, Any] | None = None
            if stage == "reference":
                ref_candidates = [
                    item for item in candidates if item["shared_references"]
                ]
                if len(ref_candidates) == 1:
                    chosen = ref_candidates[0]
            elif stage == "amount_date_unique":
                if len(candidates) == 1 and candidates[0]["date_diff_days"] is not None:
                    chosen = candidates[0]
            elif stage == "beneficiary":
                beneficiary_candidates = [
                    item
                    for item in candidates
                    if _beneficiary_match(bank_row, item["row"])
                ]
                if len(beneficiary_candidates) == 1:
                    chosen = beneficiary_candidates[0]
            elif stage == "description_tokens":
                scored = [
                    (item, _description_overlap(bank_row, item["row"]))
                    for item in candidates
                ]
                scored = [item for item in scored if item[1] >= 0.5]
                if len(scored) == 1:
                    chosen = scored[0][0]
            elif stage == "amount_date_single" and len(candidates) == 1:
                chosen = candidates[0]
            if chosen is not None:
                accept(bank_row, chosen, stage)

    unmatched_bank = _transaction_frame(
        [row for row in bank_rows if str(row["transaction_id"]) not in used_bank]
    )
    unmatched_journal = _transaction_frame(
        [row for row in journal_rows if str(row["transaction_id"]) not in used_journal]
    )
    match_frame = (
        pl.DataFrame(matches, infer_schema_length=None)
        if matches
        else pl.DataFrame(schema={col: pl.Utf8 for col in MATCH_COLUMNS})
    )
    for col in MATCH_COLUMNS:
        if col not in match_frame.columns:
            match_frame = match_frame.with_columns(pl.lit(None).alias(col))
    return (
        match_frame.select(MATCH_COLUMNS),
        unmatched_bank,
        unmatched_journal,
        stage_counts,
    )


def _filter_journal_by_sample(
    journal: pl.DataFrame, movements: set[str]
) -> pl.DataFrame:
    if not movements or journal.is_empty():
        return journal
    if "movement_number" not in journal.columns:
        return journal.head(0)
    return journal.filter(
        pl.col("movement_number").cast(pl.Utf8, strict=False).is_in(sorted(movements))
    )


def _write_workbook(path: Path, sheets: dict[str, pl.DataFrame]) -> None:
    workbook = openpyxl.Workbook()
    default = workbook.active
    workbook.remove(default)
    for title, frame in sheets.items():
        sheet = workbook.create_sheet(title[:31])
        sheet.append([excel_safe_value(value) for value in frame.columns])
        for row in frame.iter_rows():
            sheet.append([excel_safe_value(value) for value in row])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _write_review_notes(path: Path, audit: dict[str, Any]) -> None:
    lines = [
        "# Journal-Bank Reconciliation Review Notes",
        "",
        f"- Language: {audit['language']}",
        f"- Bank rows: {audit['bank_row_count']}",
        f"- Journal rows: {audit['journal_row_count']}",
        f"- Matched rows: {audit['matched_count']}",
        f"- Unmatched bank rows: {audit['unmatched_bank_count']}",
        f"- Unmatched journal rows: {audit['unmatched_journal_count']}",
        "",
        "## Stage Counts",
    ]
    counts = audit.get("stage_counts", {})
    if counts:
        for stage, count in sorted(counts.items()):
            lines.append(f"- {stage}: {count}")
    else:
        lines.append("- none")
    lines.extend(
        [
            "",
            "## Review Policy",
            "The scripts only reconcile deterministic evidence. Codex must explain unresolved cases, inspect source rows where needed, and keep professional judgment explicit.",
            "",
        ]
    )
    path.write_text("\n".join(lines), encoding="utf-8")


def run_reconciliation(
    bank_path: Path,
    journal_path: Path,
    output_dir: Path,
    recipe_path: Path | None = None,
    *,
    sample_path: Path | None = None,
    tolerance: float = 1.0,
    date_window_days: int = 7,
    language: object | None = None,
    document_language: object | None = None,
) -> ReconciliationRunResult:
    """Run deterministic journal-to-bank reconciliation and write artifacts."""

    recipe = read_json(recipe_path)
    languages = language_assumptions(
        recipe, language=language, document_language=document_language
    )
    output_dir.mkdir(parents=True, exist_ok=True)
    run_intake = write_run_intake(
        output_dir,
        bank_path=bank_path,
        journal_path=journal_path,
        recipe_path=recipe_path,
        sample_path=sample_path,
        language=languages["language"],
        document_language=languages["document_language"],
        tolerance=tolerance,
        date_window_days=date_window_days,
    )
    bank, bank_diag = _normalize_files(bank_path, "bank", recipe)
    journal, journal_diag = _normalize_files(journal_path, "journal", recipe)
    bank_pdf_non_movements = _non_movement_frame(_non_movement_records(bank_diag))
    sample_movements = _read_sample_movements(sample_path)
    filtered_journal = _filter_journal_by_sample(journal, sample_movements)
    matches, unmatched_bank, unmatched_journal, stage_counts = _match_transactions(
        bank,
        filtered_journal,
        tolerance=float(tolerance),
        date_window_days=int(date_window_days),
    )
    paths = {
        "normalized_bank_csv": output_dir / "normalized_bank.csv",
        "normalized_journal_csv": output_dir / "normalized_journal.csv",
        "reconciliation_matches_csv": output_dir / "reconciliation_matches.csv",
        "unmatched_bank_csv": output_dir / "unmatched_bank.csv",
        "unmatched_journal_csv": output_dir / "unmatched_journal.csv",
        "bank_pdf_non_movement_rows_csv": output_dir / "bank_pdf_non_movement_rows.csv",
        "workbook_xlsx": output_dir / "journal_bank_reconciliation.xlsx",
        "audit_json": output_dir / "reconciliation_audit.json",
        "review_notes_md": output_dir / "review_notes.md",
    }
    bank.write_csv(paths["normalized_bank_csv"])
    filtered_journal.write_csv(paths["normalized_journal_csv"])
    matches.write_csv(paths["reconciliation_matches_csv"])
    unmatched_bank.write_csv(paths["unmatched_bank_csv"])
    unmatched_journal.write_csv(paths["unmatched_journal_csv"])
    bank_pdf_non_movements.write_csv(paths["bank_pdf_non_movement_rows_csv"])
    _write_workbook(
        paths["workbook_xlsx"],
        {
            "matches": matches,
            "unmatched_bank": unmatched_bank,
            "unmatched_journal": unmatched_journal,
            "bank_pdf_non_movements": bank_pdf_non_movements,
            "normalized_bank": bank,
            "normalized_journal": filtered_journal,
        },
    )
    audit = {
        **languages,
        "bank_path": bank_path.as_posix(),
        "journal_path": journal_path.as_posix(),
        "sample_path": sample_path.as_posix() if sample_path else None,
        "sample_movement_count": len(sample_movements),
        "bank_row_count": bank.height,
        "journal_row_count": filtered_journal.height,
        "matched_count": matches.height,
        "unmatched_bank_count": unmatched_bank.height,
        "unmatched_journal_count": unmatched_journal.height,
        "bank_pdf_non_movement_row_count": bank_pdf_non_movements.height,
        "bank_pdf_non_movement_classifications": _count_classifications(
            bank_pdf_non_movements.to_dicts()
        ),
        "stage_counts": stage_counts,
        "tolerance": tolerance,
        "date_window_days": date_window_days,
        "diagnostics": {"bank": bank_diag, "journal": journal_diag},
        "outputs": {key: value.as_posix() for key, value in paths.items()},
    }
    write_json(paths["audit_json"], audit)
    _write_review_notes(paths["review_notes_md"], audit)
    review_session = write_review_session_artifacts(
        output_dir,
        run_id=run_intake.run_id,
        run_intake_path=run_intake.path,
        matches=matches,
        unmatched_bank=unmatched_bank,
        unmatched_journal=unmatched_journal,
        bank_pdf_non_movements=bank_pdf_non_movements,
        audit=audit,
    )
    audit["review_session"] = {
        "run_id": review_session.run_id,
        "run_intake_path": str(review_session.run_intake_path),
        "review_payload_path": str(review_session.review_payload_path),
        "ui_decisions_path": str(review_session.ui_decisions_path),
        "final_artifacts_path": str(review_session.final_artifacts_path),
        "review_item_count": review_session.review_item_count,
    }
    write_json(paths["audit_json"], audit)
    return ReconciliationRunResult(
        matches=matches,
        unmatched_bank=unmatched_bank,
        unmatched_journal=unmatched_journal,
        audit=audit,
    )


def add_common_args(parser: argparse.ArgumentParser) -> None:
    parser.add_argument(
        "--language",
        help="Working/output language locale: it, en, fr, or de. Defaults to recipe or en.",
    )
    parser.add_argument(
        "--document-language",
        help="Source-document language locale: it, en, fr, de, or auto. Defaults to recipe or auto.",
    )
    parser.add_argument("--verbose", action="store_true", help="Enable debug logging.")
