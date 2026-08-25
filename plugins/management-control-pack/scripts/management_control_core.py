"""Deterministic core for Vera's connectorless management-control pack.

Fixed code is used only for mechanically verifiable work: stable file
inventory, explicit mappings, exact Decimal arithmetic, aging buckets,
reference closure, and artifact rendering. Source roles, accounting meaning,
business causes, and professional conclusions remain reviewed judgments.
"""

from __future__ import annotations

import csv
import hashlib
import html
import io
import json
import re
import zipfile
from collections import defaultdict
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass
from datetime import date, datetime
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from pathlib import Path, PurePosixPath
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

__all__ = [
    "COMMENTARY_SCHEMA",
    "PackContractError",
    "SourceTable",
    "build_inspection",
    "build_management_pack",
    "build_model_context",
    "finalize_commentary",
    "load_json",
    "load_source_tables",
    "render_html",
    "render_markdown",
    "sha256_file",
    "write_excel",
    "write_json",
]

WORKFLOW_ID = "management-control-pack"
RECIPE_SCHEMA = "vera.management_control_recipe.v1"
PACK_SCHEMA = "vera.management_control_pack.v1"
COMMENTARY_SCHEMA = "vera.management_control_commentary.v1"
CANONICAL_CATEGORIES = (
    "revenue",
    "cogs",
    "operating_expense",
    "other_operating",
    "depreciation_amortization",
    "interest",
    "tax",
    "other",
)
OPTIONAL_SECTIONS = (
    "budget_variance",
    "receivables_aging",
    "payables_aging",
    "cash_movement",
    "customer_concentration",
    "service_profitability",
)
MAX_SOURCE_BYTES = 200 * 1024 * 1024
MAX_ZIP_MEMBER_BYTES = 100 * 1024 * 1024
MAX_TABLES = 100
MAX_COLUMNS = 200
MAX_PREVIEW_ROWS = 10
MAX_MODEL_ROWS = 60
_IDENTIFIER_RE = re.compile(r"^[a-z][a-z0-9_.-]{0,119}$")


class PackContractError(ValueError):
    """Raised when an exact pack contract or reviewed mapping is invalid."""


@dataclass(frozen=True)
class SourceTable:
    """One stable tabular unit extracted from a supplied export."""

    table_id: str
    source_id: str
    source_label: str
    table_label: str
    source_sha256: str
    source_bytes: int
    columns: tuple[str, ...]
    rows: tuple[dict[str, Any], ...]


def _canonical_json_bytes(value: object) -> bytes:
    return (
        json.dumps(
            value,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        ).encode("utf-8")
        + b"\n"
    )


def write_json(path: Path, value: object) -> None:
    """Write canonical JSON, creating only the requested parent directory."""

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(_canonical_json_bytes(value))


def load_json(path: Path) -> dict[str, Any]:
    """Load one JSON object."""

    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise PackContractError(f"JSON must contain an object: {path.name}")
    return payload


def sha256_file(path: Path) -> str:
    """Return the SHA-256 digest of one regular file."""

    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _bounded_text(value: Any, *, maximum: int = 200) -> str:
    text = "" if value is None else str(value).strip()
    return text[:maximum]


def _json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return _bounded_text(value)


def _headers(values: Sequence[Any]) -> tuple[str, ...]:
    headers = tuple(_bounded_text(value, maximum=160) for value in values)
    if not headers or all(not item for item in headers):
        raise PackContractError("A table has no header row.")
    if any(not item for item in headers):
        raise PackContractError("Table headers must be non-empty.")
    if len(headers) > MAX_COLUMNS:
        raise PackContractError(f"A table exceeds {MAX_COLUMNS} columns.")
    if len(headers) != len(set(headers)):
        raise PackContractError("Table headers must be unique.")
    return headers


def _rows_from_matrix(
    matrix: Iterable[Sequence[Any]],
) -> tuple[tuple[str, ...], tuple[dict[str, Any], ...]]:
    iterator = iter(matrix)
    try:
        first = next(iterator)
    except StopIteration as exc:
        raise PackContractError("A table is empty.") from exc
    headers = _headers(first)
    rows: list[dict[str, Any]] = []
    for raw in iterator:
        values = list(raw[: len(headers)])
        values.extend([None] * (len(headers) - len(values)))
        if all(value is None or str(value).strip() == "" for value in values):
            continue
        rows.append(dict(zip(headers, values, strict=True)))
    return headers, tuple(rows)


def _csv_matrix(data: bytes) -> Iterable[Sequence[Any]]:
    text = data.decode("utf-8-sig")
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    return csv.reader(io.StringIO(text), dialect)


def _xlsx_tables(
    data: bytes,
) -> list[tuple[str, tuple[str, ...], tuple[dict[str, Any], ...]]]:
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    tables: list[tuple[str, tuple[str, ...], tuple[dict[str, Any], ...]]] = []
    try:
        for worksheet in workbook.worksheets:
            matrix = worksheet.iter_rows(values_only=True)
            try:
                headers, rows = _rows_from_matrix(matrix)
            except PackContractError as exc:
                if str(exc) == "A table is empty.":
                    continue
                raise
            tables.append((worksheet.title, headers, rows))
    finally:
        workbook.close()
    if not tables:
        raise PackContractError("The workbook contains no non-empty tables.")
    return tables


def _safe_zip_members(archive: zipfile.ZipFile) -> list[zipfile.ZipInfo]:
    members: list[zipfile.ZipInfo] = []
    for member in archive.infolist():
        if member.is_dir():
            continue
        relative = PurePosixPath(member.filename)
        if relative.is_absolute() or ".." in relative.parts:
            raise PackContractError("ZIP member path is unsafe.")
        suffix = relative.suffix.casefold()
        if suffix not in {".csv", ".xlsx", ".xlsm"}:
            continue
        if member.file_size > MAX_ZIP_MEMBER_BYTES:
            raise PackContractError("A ZIP member exceeds the supported size.")
        members.append(member)
    if not members:
        raise PackContractError("The ZIP contains no supported CSV or Excel export.")
    return sorted(members, key=lambda item: item.filename.casefold())


def _source_units(path: Path) -> list[tuple[str, bytes]]:
    if path.is_symlink() or not path.is_file():
        raise PackContractError(f"Input must be a regular file: {path}")
    size = path.stat().st_size
    if size > MAX_SOURCE_BYTES:
        raise PackContractError(f"Input exceeds {MAX_SOURCE_BYTES} bytes: {path.name}")
    suffix = path.suffix.casefold()
    data = path.read_bytes()
    if suffix in {".csv", ".xlsx", ".xlsm"}:
        return [(path.name, data)]
    if suffix == ".zip":
        units: list[tuple[str, bytes]] = []
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            for member in _safe_zip_members(archive):
                units.append((member.filename, archive.read(member)))
        return units
    raise PackContractError(f"Unsupported input type: {path.suffix or '<none>'}")


def load_source_tables(paths: Sequence[Path]) -> list[SourceTable]:
    """Read supported exports without assigning semantic source roles."""

    if not paths:
        raise PackContractError("At least one input export is required.")
    table_records: list[
        tuple[str, str, str, int, str, tuple[str, ...], tuple[dict[str, Any], ...]]
    ] = []
    source_index = 0
    for path in paths:
        for label, data in _source_units(path.expanduser().resolve(strict=True)):
            source_index += 1
            source_id = f"source_{source_index:03d}"
            digest = hashlib.sha256(data).hexdigest()
            suffix = PurePosixPath(label).suffix.casefold()
            if suffix == ".csv":
                headers, rows = _rows_from_matrix(_csv_matrix(data))
                table_records.append(
                    (
                        source_id,
                        label,
                        digest,
                        len(data),
                        PurePosixPath(label).stem,
                        headers,
                        rows,
                    )
                )
            else:
                for sheet, headers, rows in _xlsx_tables(data):
                    table_records.append(
                        (source_id, label, digest, len(data), sheet, headers, rows)
                    )
            if len(table_records) > MAX_TABLES:
                raise PackContractError(f"Input exceeds {MAX_TABLES} tables.")
    return [
        SourceTable(
            table_id=f"table_{index:03d}",
            source_id=source_id,
            source_label=source_label,
            table_label=table_label,
            source_sha256=source_sha256,
            source_bytes=source_bytes,
            columns=columns,
            rows=rows,
        )
        for index, (
            source_id,
            source_label,
            source_sha256,
            source_bytes,
            table_label,
            columns,
            rows,
        ) in enumerate(table_records, start=1)
    ]


def _column_types(table: SourceTable) -> dict[str, list[str]]:
    observed: dict[str, set[str]] = {column: set() for column in table.columns}
    for row in table.rows[:50]:
        for column in table.columns:
            value = row[column]
            if value is None or str(value).strip() == "":
                continue
            if isinstance(value, bool):
                kind = "boolean"
            elif isinstance(value, (date, datetime)):
                kind = "date"
            elif isinstance(value, (int, float, Decimal)):
                kind = "number"
            else:
                kind = "text"
            observed[column].add(kind)
    return {column: sorted(kinds) or ["empty"] for column, kinds in observed.items()}


def _inventory_payload(tables: Sequence[SourceTable]) -> dict[str, Any]:
    sources: dict[str, dict[str, Any]] = {}
    for table in tables:
        sources.setdefault(
            table.source_id,
            {
                "source_id": table.source_id,
                "sha256": table.source_sha256,
                "byte_count": table.source_bytes,
            },
        )
    table_items = [
        {
            "table_id": table.table_id,
            "source_id": table.source_id,
            "table_label": table.table_label,
            "columns": list(table.columns),
            "row_count": len(table.rows),
        }
        for table in tables
    ]
    content = {
        "schema_version": "vera.management_control_inventory.v1",
        "sources": list(sources.values()),
        "tables": table_items,
    }
    return {
        **content,
        "inventory_sha256": hashlib.sha256(_canonical_json_bytes(content)).hexdigest(),
    }


def build_inspection(
    tables: Sequence[SourceTable],
) -> tuple[dict[str, Any], dict[str, Any], dict[str, Any]]:
    """Build bounded model inspection, private control, and recipe skeleton."""

    inventory = _inventory_payload(tables)
    inspection = {
        "schema_version": "vera.management_control_inspection.v1",
        "workflow_id": WORKFLOW_ID,
        "inventory_sha256": inventory["inventory_sha256"],
        "table_count": len(tables),
        "tables": [
            {
                "table_id": table.table_id,
                "source_label": _bounded_text(table.source_label, maximum=160),
                "table_label": _bounded_text(table.table_label, maximum=160),
                "row_count": len(table.rows),
                "columns": [
                    {
                        "name": column,
                        "observed_types": _column_types(table)[column],
                    }
                    for column in table.columns
                ],
                "preview_rows": [
                    {column: _json_value(row[column]) for column in table.columns}
                    for row in table.rows[:MAX_PREVIEW_ROWS]
                ],
            }
            for table in tables
        ],
        "mapping_policy": "semantic_roles_require_review",
    }
    control = {
        **inventory,
        "source_locations": [
            {
                "source_id": table.source_id,
                "source_label": table.source_label,
            }
            for table in tables
        ],
    }
    # Deduplicate the private source labels without exposing absolute paths.
    seen_sources: set[str] = set()
    control["source_locations"] = [
        item
        for item in control["source_locations"]
        if not (
            item["source_id"] in seen_sources or seen_sources.add(item["source_id"])
        )
    ]
    recipe = {
        "schema_version": RECIPE_SCHEMA,
        "workflow_id": WORKFLOW_ID,
        "inventory_sha256": inventory["inventory_sha256"],
        "entity": "",
        "reporting_period": {"start": "", "end": "", "cutoff": ""},
        "currency": "",
        "fiscal_year_start_month": 1,
        "number_format": "dot_decimal",
        "date_format": "%Y-%m-%d",
        "tables": {},
        "category_roles": {},
        "category_multipliers": {},
        "aging_buckets": [30, 60, 90],
        "top_customers": 10,
        "control_totals": {},
        "control_tolerance": "0.01",
        "mapping_review": {
            "status": "not_reviewed",
            "reviewer": "",
            "reviewed_at": "",
        },
    }
    return inspection, control, recipe


def _text(value: Any, *, label: str, maximum: int = 200) -> str:
    text = _bounded_text(value, maximum=maximum)
    if not text:
        raise PackContractError(f"{label} is required.")
    return text


def _parse_decimal(value: Any, *, label: str, number_format: str) -> Decimal:
    if value is None or (isinstance(value, str) and not value.strip()):
        return Decimal("0")
    if isinstance(value, bool):
        raise PackContractError(f"{label} is not a monetary value.")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip().replace("\u00a0", "").replace(" ", "")
    if number_format == "comma_decimal":
        text = text.replace(".", "").replace(",", ".")
    elif number_format == "dot_decimal":
        text = text.replace(",", "")
    else:
        raise PackContractError("number_format must be dot_decimal or comma_decimal.")
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise PackContractError(f"{label} is not a valid decimal: {value}") from exc


def _decimal_text(value: Decimal) -> str:
    normalized = value.normalize()
    if normalized == normalized.to_integral():
        return format(normalized.quantize(Decimal("1")), "f")
    return format(normalized, "f")


def _ratio_text(value: Decimal) -> str:
    """Return a stable six-decimal ratio for review and display."""

    return _decimal_text(value.quantize(Decimal("0.000001"), rounding=ROUND_HALF_UP))


def _parse_date(value: Any, *, label: str, date_format: str) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value, label=label, maximum=80)
    try:
        return datetime.strptime(text, date_format).date()
    except ValueError as exc:
        raise PackContractError(
            f"{label} does not match reviewed date_format {date_format}: {text}"
        ) from exc


def _recipe_table(
    recipe: Mapping[str, Any],
    tables: Mapping[str, SourceTable],
    role: str,
    *,
    required: bool = False,
) -> tuple[SourceTable, dict[str, Any]] | None:
    mappings = recipe.get("tables")
    if not isinstance(mappings, dict):
        raise PackContractError("tables must be an object.")
    raw = mappings.get(role)
    if raw is None:
        if required:
            raise PackContractError(f"Required table role is missing: {role}")
        return None
    if not isinstance(raw, dict):
        raise PackContractError(f"Table mapping must be an object: {role}")
    table_id = _text(raw.get("table_id"), label=f"{role}.table_id", maximum=40)
    table = tables.get(table_id)
    if table is None:
        raise PackContractError(f"Unknown table_id for {role}: {table_id}")
    columns = raw.get("columns")
    if not isinstance(columns, dict):
        raise PackContractError(f"{role}.columns must be an object.")
    for logical, physical in columns.items():
        if not isinstance(logical, str) or not isinstance(physical, str):
            raise PackContractError(f"{role}.columns must map text to text.")
        if physical not in table.columns:
            raise PackContractError(f"{role} maps missing column: {physical}")
    return table, raw


def _column(
    mapping: Mapping[str, Any], logical: str, *, required: bool = False
) -> str | None:
    columns = mapping.get("columns")
    physical = columns.get(logical) if isinstance(columns, dict) else None
    if required and not isinstance(physical, str):
        raise PackContractError(f"Required logical column is missing: {logical}")
    return physical if isinstance(physical, str) else None


def _amount(
    row: Mapping[str, Any],
    mapping: Mapping[str, Any],
    *,
    label: str,
    number_format: str,
) -> Decimal:
    amount_column = _column(mapping, "amount")
    if amount_column:
        amount = _parse_decimal(
            row.get(amount_column), label=label, number_format=number_format
        )
    else:
        debit_column = _column(mapping, "debit")
        credit_column = _column(mapping, "credit")
        if not debit_column or not credit_column:
            raise PackContractError("Map amount or both debit and credit columns.")
        debit = _parse_decimal(
            row.get(debit_column), label=f"{label}.debit", number_format=number_format
        )
        credit = _parse_decimal(
            row.get(credit_column), label=f"{label}.credit", number_format=number_format
        )
        rule = mapping.get("amount_rule")
        if rule == "credit_minus_debit":
            amount = credit - debit
        elif rule == "debit_minus_credit":
            amount = debit - credit
        else:
            raise PackContractError(
                "Debit/credit mappings require amount_rule credit_minus_debit or debit_minus_credit."
            )
    multiplier = _parse_decimal(
        mapping.get("amount_multiplier", "1"),
        label=f"{label}.amount_multiplier",
        number_format="dot_decimal",
    )
    return amount * multiplier


def _review_recipe(recipe: Mapping[str, Any], inventory_sha256: str) -> dict[str, Any]:
    if (
        recipe.get("schema_version") != RECIPE_SCHEMA
        or recipe.get("workflow_id") != WORKFLOW_ID
    ):
        raise PackContractError("Recipe identity is invalid.")
    if recipe.get("inventory_sha256") != inventory_sha256:
        raise PackContractError("Recipe belongs to a different input inventory.")
    review = recipe.get("mapping_review")
    if not isinstance(review, dict) or review.get("status") != "reviewed":
        raise PackContractError(
            "Mappings must be explicitly reviewed before calculation."
        )
    _text(review.get("reviewer"), label="mapping_review.reviewer", maximum=160)
    _text(review.get("reviewed_at"), label="mapping_review.reviewed_at", maximum=80)
    currency = _text(recipe.get("currency"), label="currency", maximum=3).upper()
    if not re.fullmatch(r"[A-Z]{3}", currency):
        raise PackContractError("currency must be an ISO three-letter code.")
    period = recipe.get("reporting_period")
    if not isinstance(period, dict):
        raise PackContractError("reporting_period must be an object.")
    normalized = dict(recipe)
    normalized["currency"] = currency
    normalized["reporting_period"] = {
        key: _parse_date(
            period.get(key), label=f"reporting_period.{key}", date_format="%Y-%m-%d"
        )
        for key in ("start", "end", "cutoff")
    }
    if normalized["reporting_period"]["start"] > normalized["reporting_period"]["end"]:
        raise PackContractError("Reporting period start is after end.")
    return normalized


def _category_lookup(
    recipe: Mapping[str, Any], table_map: Mapping[str, SourceTable]
) -> dict[str, str]:
    role_map = recipe.get("category_roles")
    if not isinstance(role_map, dict) or not role_map:
        raise PackContractError("category_roles must contain reviewed mappings.")
    normalized: dict[str, str] = {}
    for source, canonical in role_map.items():
        source_text = _text(source, label="category source label", maximum=160)
        canonical_text = _text(canonical, label="canonical category", maximum=80)
        if canonical_text not in CANONICAL_CATEGORIES:
            raise PackContractError(f"Unsupported canonical category: {canonical_text}")
        normalized[source_text] = canonical_text

    account_map = _recipe_table(recipe, table_map, "account_map")
    if account_map is None:
        return normalized
    return normalized


def _account_categories(
    recipe: Mapping[str, Any], table_map: Mapping[str, SourceTable]
) -> dict[str, str]:
    account_map = _recipe_table(recipe, table_map, "account_map")
    if account_map is None:
        return {}
    table, mapping = account_map
    account_col = _column(mapping, "account_code", required=True)
    category_col = _column(mapping, "category", required=True)
    result: dict[str, str] = {}
    for index, row in enumerate(table.rows, start=2):
        account = _text(
            row.get(account_col), label=f"account_map row {index} account", maximum=160
        )
        category = _text(
            row.get(category_col),
            label=f"account_map row {index} category",
            maximum=160,
        )
        existing = result.get(account)
        if existing is not None and existing != category:
            raise PackContractError(f"Account {account} maps to multiple categories.")
        result[account] = category
    return result


def _normalized_category(source: str, recipe: Mapping[str, Any]) -> tuple[str, Decimal]:
    roles = recipe.get("category_roles")
    if not isinstance(roles, dict) or source not in roles:
        raise PackContractError(f"Unreviewed accounting category: {source}")
    canonical = str(roles[source])
    if canonical not in CANONICAL_CATEGORIES:
        raise PackContractError(f"Unsupported canonical category: {canonical}")
    multipliers = recipe.get("category_multipliers", {})
    if not isinstance(multipliers, dict):
        raise PackContractError("category_multipliers must be an object.")
    try:
        multiplier = Decimal(str(multipliers.get(source, "1")))
    except InvalidOperation as exc:
        raise PackContractError(f"Invalid category multiplier for {source}.") from exc
    return canonical, multiplier


def _metric(
    metrics: dict[str, dict[str, Any]],
    metric_id: str,
    label: str,
    value: Decimal,
    *,
    unit: str,
    section: str,
    period: str | None = None,
) -> None:
    if not _IDENTIFIER_RE.fullmatch(metric_id) or metric_id in metrics:
        raise PackContractError(f"Metric ID is invalid or duplicated: {metric_id}")
    item = {
        "metric_id": metric_id,
        "label": label,
        "value": _ratio_text(value) if unit == "ratio" else _decimal_text(value),
        "unit": unit,
        "section": section,
    }
    if period is not None:
        item["period"] = period
    metrics[metric_id] = item


def _control(role: str, actual: Decimal, recipe: Mapping[str, Any]) -> dict[str, Any]:
    declared = recipe.get("control_totals", {})
    if not isinstance(declared, dict):
        raise PackContractError("control_totals must be an object.")
    expected = declared.get(role)
    if expected is None:
        return {"role": role, "status": "not_assessed", "actual": _decimal_text(actual)}
    number_format = str(recipe.get("number_format", "dot_decimal"))
    expected_decimal = _parse_decimal(
        expected, label=f"control_totals.{role}", number_format=number_format
    )
    tolerance = _parse_decimal(
        recipe.get("control_tolerance", "0.01"),
        label="control_tolerance",
        number_format="dot_decimal",
    )
    difference = actual - expected_decimal
    return {
        "role": role,
        "status": "passed" if abs(difference) <= tolerance else "failed",
        "actual": _decimal_text(actual),
        "expected": _decimal_text(expected_decimal),
        "difference": _decimal_text(difference),
        "tolerance": _decimal_text(tolerance),
    }


def _pnl_section(
    recipe: Mapping[str, Any],
    table_map: Mapping[str, SourceTable],
    metrics: dict[str, dict[str, Any]],
) -> tuple[dict[str, Any], dict[str, Any]]:
    table, mapping = _recipe_table(recipe, table_map, "general_ledger", required=True)  # type: ignore[misc]
    date_col = _column(mapping, "date", required=True)
    category_col = _column(mapping, "category")
    account_col = _column(mapping, "account_code")
    if not category_col and not account_col:
        raise PackContractError(
            "General ledger requires category or account_code mapping."
        )
    account_categories = _account_categories(recipe, table_map)
    start = recipe["reporting_period"]["start"]
    end = recipe["reporting_period"]["end"]
    number_format = str(recipe.get("number_format", "dot_decimal"))
    date_format = str(recipe.get("date_format", "%Y-%m-%d"))
    monthly: dict[str, dict[str, Decimal]] = defaultdict(lambda: defaultdict(Decimal))
    control_total = Decimal("0")
    for index, row in enumerate(table.rows, start=2):
        row_date = _parse_date(
            row.get(date_col),
            label=f"general_ledger row {index} date",
            date_format=date_format,
        )
        if row_date < start or row_date > end:
            continue
        if category_col:
            source_category = _text(
                row.get(category_col),
                label=f"general_ledger row {index} category",
                maximum=160,
            )
        else:
            account = _text(
                row.get(account_col),
                label=f"general_ledger row {index} account",
                maximum=160,
            )
            source_category = account_categories.get(account, "")
            if not source_category:
                raise PackContractError(
                    f"Account {account} has no reviewed category mapping."
                )
        category, category_multiplier = _normalized_category(source_category, recipe)
        amount = (
            _amount(
                row,
                mapping,
                label=f"general_ledger row {index}",
                number_format=number_format,
            )
            * category_multiplier
        )
        control_total += amount
        monthly[row_date.strftime("%Y-%m")][category] += amount
    if not monthly:
        raise PackContractError(
            "General ledger has no rows inside the reporting period."
        )
    rows: list[dict[str, Any]] = []
    totals = defaultdict(Decimal)
    metric_labels = {
        "revenue": "Revenue",
        "gross_profit": "Gross profit",
        "ebitda": "EBITDA",
        "ebit": "EBIT",
        "net_result": "Net result",
    }
    for period in sorted(monthly):
        category_values = {
            category: monthly[period].get(category, Decimal("0"))
            for category in CANONICAL_CATEGORIES
        }
        revenue = category_values["revenue"]
        gross_profit = revenue + category_values["cogs"]
        ebitda = (
            gross_profit
            + category_values["operating_expense"]
            + category_values["other_operating"]
        )
        ebit = ebitda + category_values["depreciation_amortization"]
        net_result = sum(category_values.values(), Decimal("0"))
        derived = {
            "gross_profit": gross_profit,
            "ebitda": ebitda,
            "ebit": ebit,
            "net_result": net_result,
        }
        rows.append(
            {
                "period": period,
                **{key: _decimal_text(value) for key, value in category_values.items()},
                **{key: _decimal_text(value) for key, value in derived.items()},
            }
        )
        for key, value in {**category_values, **derived}.items():
            totals[key] += value
        for key in ("revenue", "gross_profit", "ebitda", "net_result"):
            _metric(
                metrics,
                f"pnl.{period}.{key}",
                f"{metric_labels[key]} {period}",
                {**category_values, **derived}[key],
                unit=recipe["currency"],
                section="monthly_pnl",
                period=period,
            )
    for key in ("revenue", "gross_profit", "ebitda", "ebit", "net_result"):
        _metric(
            metrics,
            f"pnl.total.{key}",
            metric_labels[key],
            totals[key],
            unit=recipe["currency"],
            section="monthly_pnl",
        )
    return {
        "status": "available",
        "rows": rows,
        "totals": {key: _decimal_text(value) for key, value in totals.items()},
    }, _control("general_ledger", control_total, recipe)


def _budget_section(
    recipe: Mapping[str, Any],
    table_map: Mapping[str, SourceTable],
    pnl: Mapping[str, Any],
    metrics: dict[str, dict[str, Any]],
) -> tuple[dict[str, Any], dict[str, Any]]:
    resolved = _recipe_table(recipe, table_map, "budget")
    if resolved is None:
        raise PackContractError("Budget export was not mapped.")
    table, mapping = resolved
    date_col = _column(mapping, "date", required=True)
    category_col = _column(mapping, "category")
    account_col = _column(mapping, "account_code")
    if not category_col and not account_col:
        raise PackContractError("Budget requires category or account_code mapping.")
    account_categories = _account_categories(recipe, table_map)
    number_format = str(recipe.get("number_format", "dot_decimal"))
    date_format = str(recipe.get("date_format", "%Y-%m-%d"))
    start = recipe["reporting_period"]["start"]
    end = recipe["reporting_period"]["end"]
    monthly: dict[str, dict[str, Decimal]] = defaultdict(lambda: defaultdict(Decimal))
    control_total = Decimal("0")
    for index, row in enumerate(table.rows, start=2):
        row_date = _parse_date(
            row.get(date_col), label=f"budget row {index} date", date_format=date_format
        )
        if row_date < start or row_date > end:
            continue
        if category_col:
            source_category = _text(
                row.get(category_col), label=f"budget row {index} category", maximum=160
            )
        else:
            account = _text(
                row.get(account_col), label=f"budget row {index} account", maximum=160
            )
            source_category = account_categories.get(account, "")
            if not source_category:
                raise PackContractError(
                    f"Budget account {account} has no reviewed category mapping."
                )
        category, multiplier = _normalized_category(source_category, recipe)
        amount = (
            _amount(
                row, mapping, label=f"budget row {index}", number_format=number_format
            )
            * multiplier
        )
        monthly[row_date.strftime("%Y-%m")][category] += amount
        control_total += amount
    actual_rows = {str(row["period"]): row for row in pnl["rows"]}
    rows: list[dict[str, Any]] = []
    for period in sorted(set(actual_rows) | set(monthly)):
        actual = Decimal(str(actual_rows.get(period, {}).get("ebitda", "0")))
        budget_values = monthly.get(period, {})
        budget = (
            budget_values.get("revenue", Decimal("0"))
            + budget_values.get("cogs", Decimal("0"))
            + budget_values.get("operating_expense", Decimal("0"))
            + budget_values.get("other_operating", Decimal("0"))
        )
        variance = actual - budget
        rows.append(
            {
                "period": period,
                "actual_ebitda": _decimal_text(actual),
                "budget_ebitda": _decimal_text(budget),
                "variance": _decimal_text(variance),
            }
        )
        _metric(
            metrics,
            f"budget.{period}.ebitda_variance",
            f"EBITDA variance {period}",
            variance,
            unit=recipe["currency"],
            section="budget_variance",
            period=period,
        )
    total_variance = sum((Decimal(row["variance"]) for row in rows), Decimal("0"))
    _metric(
        metrics,
        "budget.total.ebitda_variance",
        "Total EBITDA variance",
        total_variance,
        unit=recipe["currency"],
        section="budget_variance",
    )
    return {
        "status": "available",
        "rows": rows,
        "total_ebitda_variance": _decimal_text(total_variance),
    }, _control("budget", control_total, recipe)


def _aging_section(
    role: str,
    recipe: Mapping[str, Any],
    table_map: Mapping[str, SourceTable],
    metrics: dict[str, dict[str, Any]],
) -> tuple[dict[str, Any], dict[str, Any]]:
    resolved = _recipe_table(recipe, table_map, role)
    if resolved is None:
        raise PackContractError(
            f"{role.replace('_', ' ').title()} export was not mapped."
        )
    table, mapping = resolved
    party_logical = "customer_name" if role == "receivables" else "supplier_name"
    party_col = _column(mapping, party_logical, required=True)
    due_col = _column(mapping, "due_date", required=True)
    outstanding_col = _column(mapping, "outstanding_amount", required=True)
    cutoff = recipe["reporting_period"]["cutoff"]
    date_format = str(recipe.get("date_format", "%Y-%m-%d"))
    number_format = str(recipe.get("number_format", "dot_decimal"))
    buckets = recipe.get("aging_buckets", [30, 60, 90])
    if (
        buckets != sorted(set(buckets))
        or len(buckets) != 3
        or any(not isinstance(value, int) or value <= 0 for value in buckets)
    ):
        raise PackContractError(
            "aging_buckets must contain three increasing positive integers."
        )
    labels = (
        "not_due",
        f"1_{buckets[0]}",
        f"{buckets[0] + 1}_{buckets[1]}",
        f"{buckets[1] + 1}_{buckets[2]}",
        f"over_{buckets[2]}",
    )
    totals = defaultdict(Decimal)
    parties: dict[str, Decimal] = defaultdict(Decimal)
    control_total = Decimal("0")
    for index, row in enumerate(table.rows, start=2):
        due = _parse_date(
            row.get(due_col),
            label=f"{role} row {index} due_date",
            date_format=date_format,
        )
        amount = _parse_decimal(
            row.get(outstanding_col),
            label=f"{role} row {index} outstanding",
            number_format=number_format,
        )
        party = _text(
            row.get(party_col), label=f"{role} row {index} party", maximum=160
        )
        days = (cutoff - due).days
        if days <= 0:
            bucket = labels[0]
        elif days <= buckets[0]:
            bucket = labels[1]
        elif days <= buckets[1]:
            bucket = labels[2]
        elif days <= buckets[2]:
            bucket = labels[3]
        else:
            bucket = labels[4]
        totals[bucket] += amount
        parties[party] += amount
        control_total += amount
    total = sum(totals.values(), Decimal("0"))
    overdue = total - totals[labels[0]]
    prefix = "ar" if role == "receivables" else "ap"
    section = f"{role}_aging"
    _metric(
        metrics,
        f"{prefix}.total.outstanding",
        f"{role.title()} outstanding",
        total,
        unit=recipe["currency"],
        section=section,
    )
    _metric(
        metrics,
        f"{prefix}.total.overdue",
        f"{role.title()} overdue",
        overdue,
        unit=recipe["currency"],
        section=section,
    )
    top = [
        {"name": name, "outstanding": _decimal_text(amount)}
        for name, amount in sorted(
            parties.items(), key=lambda item: (-abs(item[1]), item[0])
        )[:20]
    ]
    return {
        "status": "available",
        "cutoff": cutoff.isoformat(),
        "buckets": [
            {"bucket": label, "amount": _decimal_text(totals[label])}
            for label in labels
        ],
        "total_outstanding": _decimal_text(total),
        "overdue": _decimal_text(overdue),
        "top_parties": top,
    }, _control(role, control_total, recipe)


def _cash_section(
    recipe: Mapping[str, Any],
    table_map: Mapping[str, SourceTable],
    metrics: dict[str, dict[str, Any]],
) -> tuple[dict[str, Any], dict[str, Any]]:
    resolved = _recipe_table(recipe, table_map, "bank")
    if resolved is None:
        raise PackContractError("Bank export was not mapped.")
    table, mapping = resolved
    date_col = _column(mapping, "date", required=True)
    amount_col = _column(mapping, "amount", required=True)
    balance_col = _column(mapping, "balance")
    account_col = _column(mapping, "account")
    date_format = str(recipe.get("date_format", "%Y-%m-%d"))
    number_format = str(recipe.get("number_format", "dot_decimal"))
    monthly: dict[str, dict[str, Decimal]] = defaultdict(lambda: defaultdict(Decimal))
    latest: dict[str, tuple[date, Decimal]] = {}
    control_total = Decimal("0")
    for index, row in enumerate(table.rows, start=2):
        row_date = _parse_date(
            row.get(date_col), label=f"bank row {index} date", date_format=date_format
        )
        amount = _parse_decimal(
            row.get(amount_col),
            label=f"bank row {index} amount",
            number_format=number_format,
        )
        period = row_date.strftime("%Y-%m")
        monthly[period]["inflow" if amount >= 0 else "outflow"] += amount
        monthly[period]["net"] += amount
        control_total += amount
        if balance_col:
            balance = _parse_decimal(
                row.get(balance_col),
                label=f"bank row {index} balance",
                number_format=number_format,
            )
            account = (
                _bounded_text(row.get(account_col), maximum=120)
                if account_col
                else "all_accounts"
            )
            if account not in latest or row_date >= latest[account][0]:
                latest[account] = (row_date, balance)
    rows = [
        {
            "period": period,
            "inflow": _decimal_text(values["inflow"]),
            "outflow": _decimal_text(values["outflow"]),
            "net": _decimal_text(values["net"]),
        }
        for period, values in sorted(monthly.items())
    ]
    net = sum((values["net"] for values in monthly.values()), Decimal("0"))
    latest_balance = (
        sum((item[1] for item in latest.values()), Decimal("0")) if latest else None
    )
    _metric(
        metrics,
        "cash.total.net_movement",
        "Net cash movement",
        net,
        unit=recipe["currency"],
        section="cash_movement",
    )
    if latest_balance is not None:
        _metric(
            metrics,
            "cash.latest.reported_balance",
            "Latest reported cash balance",
            latest_balance,
            unit=recipe["currency"],
            section="cash_movement",
        )
    return {
        "status": "available",
        "rows": rows,
        "net_movement": _decimal_text(net),
        "latest_reported_balance": (
            _decimal_text(latest_balance) if latest_balance is not None else None
        ),
        "balance_basis": (
            "latest_row_per_reviewed_account" if latest else "not_available"
        ),
    }, _control("bank", control_total, recipe)


def _sales_sections(
    recipe: Mapping[str, Any],
    table_map: Mapping[str, SourceTable],
    metrics: dict[str, dict[str, Any]],
) -> tuple[dict[str, Any], dict[str, Any], dict[str, Any]]:
    resolved = _recipe_table(recipe, table_map, "sales_lines")
    if resolved is None:
        raise PackContractError("Sales-line export was not mapped.")
    table, mapping = resolved
    date_col = _column(mapping, "date", required=True)
    customer_col = _column(mapping, "customer_name") or _column(mapping, "customer_id")
    service_col = _column(mapping, "service")
    revenue_col = _column(mapping, "revenue", required=True)
    direct_cost_col = _column(mapping, "direct_cost")
    if not customer_col and not service_col:
        raise PackContractError("Sales lines require customer or service mapping.")
    date_format = str(recipe.get("date_format", "%Y-%m-%d"))
    number_format = str(recipe.get("number_format", "dot_decimal"))
    start = recipe["reporting_period"]["start"]
    end = recipe["reporting_period"]["end"]
    customers: dict[str, Decimal] = defaultdict(Decimal)
    services: dict[str, dict[str, Decimal]] = defaultdict(lambda: defaultdict(Decimal))
    control_total = Decimal("0")
    for index, row in enumerate(table.rows, start=2):
        row_date = _parse_date(
            row.get(date_col), label=f"sales row {index} date", date_format=date_format
        )
        if row_date < start or row_date > end:
            continue
        revenue = _parse_decimal(
            row.get(revenue_col),
            label=f"sales row {index} revenue",
            number_format=number_format,
        )
        control_total += revenue
        if customer_col:
            customer = _text(
                row.get(customer_col), label=f"sales row {index} customer", maximum=160
            )
            customers[customer] += revenue
        if service_col:
            service = _text(
                row.get(service_col), label=f"sales row {index} service", maximum=160
            )
            services[service]["revenue"] += revenue
            if direct_cost_col:
                direct_cost = _parse_decimal(
                    row.get(direct_cost_col),
                    label=f"sales row {index} direct_cost",
                    number_format=number_format,
                )
                services[service]["direct_cost"] += direct_cost
    total_revenue = sum(customers.values(), Decimal("0"))
    top_n = recipe.get("top_customers", 10)
    if not isinstance(top_n, int) or not 1 <= top_n <= 50:
        raise PackContractError("top_customers must be between 1 and 50.")
    ranked_customers = sorted(customers.items(), key=lambda item: (-item[1], item[0]))
    customer_rows = [
        {
            "customer": name,
            "revenue": _decimal_text(value),
            "share": _ratio_text(value / total_revenue) if total_revenue else "0",
        }
        for name, value in ranked_customers[:top_n]
    ]
    top1_share = (
        ranked_customers[0][1] / total_revenue
        if ranked_customers and total_revenue
        else Decimal("0")
    )
    top5_share = (
        sum((value for _, value in ranked_customers[:5]), Decimal("0")) / total_revenue
        if total_revenue
        else Decimal("0")
    )
    concentration = (
        {
            "status": "available",
            "total_revenue": _decimal_text(total_revenue),
            "top_1_share": _ratio_text(top1_share),
            "top_5_share": _ratio_text(top5_share),
            "rows": customer_rows,
        }
        if customer_col
        else {"status": "unavailable", "reason": "Customer identity is not mapped."}
    )
    if customer_col:
        _metric(
            metrics,
            "customers.top1.share",
            "Top customer revenue share",
            top1_share,
            unit="ratio",
            section="customer_concentration",
        )
        _metric(
            metrics,
            "customers.top5.share",
            "Top five customer revenue share",
            top5_share,
            unit="ratio",
            section="customer_concentration",
        )
    if service_col and direct_cost_col:
        service_rows = []
        total_margin = Decimal("0")
        for service, values in sorted(
            services.items(), key=lambda item: (-item[1]["revenue"], item[0])
        ):
            revenue = values["revenue"]
            direct_cost = values["direct_cost"]
            margin = revenue - direct_cost
            total_margin += margin
            service_rows.append(
                {
                    "service": service,
                    "revenue": _decimal_text(revenue),
                    "direct_cost": _decimal_text(direct_cost),
                    "margin": _decimal_text(margin),
                    "margin_rate": _ratio_text(margin / revenue) if revenue else "0",
                }
            )
        profitability = {
            "status": "available",
            "rows": service_rows[:50],
            "total_margin": _decimal_text(total_margin),
        }
        _metric(
            metrics,
            "services.total.margin",
            "Service gross margin",
            total_margin,
            unit=recipe["currency"],
            section="service_profitability",
        )
    else:
        reason = "Service and authoritative direct-cost columns are both required."
        profitability = {"status": "unavailable", "reason": reason}
    return concentration, profitability, _control("sales_lines", control_total, recipe)


def build_management_pack(
    tables: Sequence[SourceTable], raw_recipe: Mapping[str, Any]
) -> dict[str, Any]:
    """Calculate the complete supported pack from one reviewed recipe."""

    inventory = _inventory_payload(tables)
    recipe = _review_recipe(raw_recipe, inventory["inventory_sha256"])
    table_map = {table.table_id: table for table in tables}
    _category_lookup(recipe, table_map)
    metrics: dict[str, dict[str, Any]] = {}
    controls: list[dict[str, Any]] = []
    limitations: list[str] = []
    sections: dict[str, dict[str, Any]] = {}
    coverage: list[dict[str, Any]] = []

    pnl, pnl_control = _pnl_section(recipe, table_map, metrics)
    sections["monthly_pnl"] = pnl
    controls.append(pnl_control)
    coverage.append({"section": "monthly_pnl", "status": "available"})

    optional_builders = (
        ("budget_variance", lambda: _budget_section(recipe, table_map, pnl, metrics)),
        (
            "receivables_aging",
            lambda: _aging_section("receivables", recipe, table_map, metrics),
        ),
        (
            "payables_aging",
            lambda: _aging_section("payables", recipe, table_map, metrics),
        ),
        ("cash_movement", lambda: _cash_section(recipe, table_map, metrics)),
    )
    for section_name, builder in optional_builders:
        try:
            section, control = builder()
        except PackContractError as exc:
            section = {"status": "unavailable", "reason": str(exc)}
            limitations.append(f"{section_name}: {exc}")
        else:
            controls.append(control)
        sections[section_name] = section
        coverage.append(
            {
                "section": section_name,
                "status": section["status"],
                **({"reason": section["reason"]} if "reason" in section else {}),
            }
        )

    try:
        concentration, profitability, sales_control = _sales_sections(
            recipe, table_map, metrics
        )
    except PackContractError as exc:
        concentration = {"status": "unavailable", "reason": str(exc)}
        profitability = {"status": "unavailable", "reason": str(exc)}
        limitations.extend(
            (f"customer_concentration: {exc}", f"service_profitability: {exc}")
        )
    else:
        controls.append(sales_control)
        for name, section in (
            ("customer_concentration", concentration),
            ("service_profitability", profitability),
        ):
            if section["status"] == "unavailable":
                limitations.append(f"{name}: {section['reason']}")
    for name, section in (
        ("customer_concentration", concentration),
        ("service_profitability", profitability),
    ):
        sections[name] = section
        coverage.append(
            {
                "section": name,
                "status": section["status"],
                **({"reason": section["reason"]} if "reason" in section else {}),
            }
        )

    failed_controls = [item for item in controls if item["status"] == "failed"]
    unavailable = [item for item in coverage if item["status"] != "available"]
    status = (
        "blocked"
        if failed_controls
        else ("partial" if unavailable else "ready_for_review")
    )
    if failed_controls:
        limitations.append("One or more declared source control totals failed.")
    source_lineage: dict[str, dict[str, Any]] = {}
    for table in tables:
        item = source_lineage.setdefault(
            table.source_id,
            {
                "source_id": table.source_id,
                "sha256": table.source_sha256,
                "byte_count": table.source_bytes,
                "table_ids": [],
            },
        )
        item["table_ids"].append(table.table_id)
    return {
        "schema_version": PACK_SCHEMA,
        "workflow_id": WORKFLOW_ID,
        "status": status,
        "report_status": "draft_pending_professional_review",
        "entity": _text(recipe.get("entity"), label="entity", maximum=200),
        "reporting_period": {
            key: recipe["reporting_period"][key].isoformat()
            for key in ("start", "end", "cutoff")
        },
        "currency": recipe["currency"],
        "inventory_sha256": inventory["inventory_sha256"],
        "recipe_sha256": hashlib.sha256(_canonical_json_bytes(raw_recipe)).hexdigest(),
        "coverage": coverage,
        "controls": controls,
        "metrics": metrics,
        "sections": sections,
        "limitations": limitations,
        "source_lineage": list(source_lineage.values()),
        "professional_boundary": (
            "Calculated facts and schema closure do not establish accounting correctness, "
            "source completeness, business causation, or professional approval."
        ),
    }


def build_model_context(pack: Mapping[str, Any]) -> dict[str, Any]:
    """Project a bounded post-calculation context without raw source rows."""

    sections = pack["sections"]
    projected: dict[str, Any] = {}
    for name, section in sections.items():
        if not isinstance(section, dict):
            continue
        item = {key: value for key, value in section.items() if key != "rows"}
        if isinstance(section.get("rows"), list):
            item["rows"] = section["rows"][:MAX_MODEL_ROWS]
        if isinstance(section.get("top_parties"), list):
            item["top_parties"] = section["top_parties"][:20]
        projected[name] = item
    return {
        "schema_version": "vera.management_control_model_context.v1",
        "workflow_id": WORKFLOW_ID,
        "status": pack["status"],
        "entity": pack["entity"],
        "reporting_period": pack["reporting_period"],
        "currency": pack["currency"],
        "coverage": pack["coverage"],
        "controls": pack["controls"],
        "metrics": list(pack["metrics"].values()),
        "sections": projected,
        "limitations": pack["limitations"],
        "source_lineage": pack["source_lineage"],
        "context_boundary": (
            "Bounded calculated results and top-ranked rows only; raw source populations "
            "and original filenames are not included in this post-calculation projection."
        ),
    }


def _table_markdown(rows: Sequence[Mapping[str, Any]], columns: Sequence[str]) -> str:
    if not rows:
        return "_No rows available._"
    header = "| " + " | ".join(columns) + " |"
    separator = "| " + " | ".join("---" for _ in columns) + " |"
    body = [
        "| " + " | ".join(str(row.get(column, "")) for column in columns) + " |"
        for row in rows
    ]
    return "\n".join((header, separator, *body))


def render_markdown(
    pack: Mapping[str, Any], commentary: Mapping[str, Any] | None = None
) -> str:
    """Render a compact, source-bound management report."""

    lines = [
        f"# Management Control Pack — {pack['entity']}",
        "",
        f"**Period:** {pack['reporting_period']['start']} to {pack['reporting_period']['end']}  ",
        f"**Cutoff:** {pack['reporting_period']['cutoff']}  ",
        f"**Currency:** {pack['currency']}  ",
        f"**Pack status:** `{pack['status']}`  ",
        f"**Review status:** `{pack['report_status']}`",
        "",
        "## Coverage",
        "",
        _table_markdown(pack["coverage"], ("section", "status", "reason")),
        "",
        "## Head metrics",
        "",
        _table_markdown(
            list(pack["metrics"].values())[:30], ("metric_id", "label", "value", "unit")
        ),
        "",
    ]
    detail_sections = (
        (
            "Monthly P&L",
            pack["sections"]["monthly_pnl"].get("rows", []),
            ("period", "revenue", "gross_profit", "ebitda", "net_result"),
        ),
        (
            "Budget variance",
            pack["sections"]["budget_variance"].get("rows", []),
            ("period", "actual_ebitda", "budget_ebitda", "variance"),
        ),
        (
            "Receivables aging",
            pack["sections"]["receivables_aging"].get("buckets", []),
            ("bucket", "amount"),
        ),
        (
            "Payables aging",
            pack["sections"]["payables_aging"].get("buckets", []),
            ("bucket", "amount"),
        ),
        (
            "Cash movement",
            pack["sections"]["cash_movement"].get("rows", []),
            ("period", "inflow", "outflow", "net"),
        ),
        (
            "Customer concentration",
            pack["sections"]["customer_concentration"].get("rows", []),
            ("customer", "revenue", "share"),
        ),
        (
            "Service profitability",
            pack["sections"]["service_profitability"].get("rows", []),
            ("service", "revenue", "direct_cost", "margin", "margin_rate"),
        ),
    )
    for title, rows, columns in detail_sections:
        lines.extend((f"## {title}", "", _table_markdown(rows, columns), ""))
    if commentary:
        labels = (
            ("observations", "Calculated observations"),
            ("hypotheses", "Hypotheses requiring evidence"),
            ("questions", "Questions"),
            ("limitations", "Limitations"),
        )
        for key, title in labels:
            lines.extend((f"## {title}", ""))
            items = commentary.get(key, [])
            if not items:
                lines.append("_None recorded._")
            for item in items:
                references = ", ".join(item.get("metric_ids", []))
                suffix = f" (`{references}`)" if references else ""
                lines.append(f"- {item['text']}{suffix}")
            lines.append("")
    lines.extend(
        (
            "## Controls and limitations",
            "",
            _table_markdown(
                pack["controls"], ("role", "status", "actual", "expected", "difference")
            ),
            "",
            *(f"- {item}" for item in pack["limitations"]),
            "",
            f"> {pack['professional_boundary']}",
            "",
        )
    )
    return "\n".join(lines)


def _html_table(rows: Sequence[Mapping[str, Any]], columns: Sequence[str]) -> str:
    if not rows:
        return '<p class="empty">No rows available.</p>'
    head = "".join(
        f"<th>{html.escape(column.replace('_', ' ').title())}</th>"
        for column in columns
    )
    body = "".join(
        "<tr>"
        + "".join(
            f"<td>{html.escape(str(row.get(column, '')))}</td>" for column in columns
        )
        + "</tr>"
        for row in rows
    )
    return f'<div class="table-wrap"><table><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table></div>'


def render_html(
    pack: Mapping[str, Any], commentary: Mapping[str, Any] | None = None
) -> str:
    """Render a self-contained management command centre."""

    head_ids = (
        "pnl.total.revenue",
        "pnl.total.gross_profit",
        "pnl.total.ebitda",
        "pnl.total.net_result",
        "ar.total.overdue",
        "cash.latest.reported_balance",
    )
    cards = []
    for metric_id in head_ids:
        metric = pack["metrics"].get(metric_id)
        if metric:
            cards.append(
                f'<article class="metric"><span>{html.escape(metric["label"])}</span>'
                f'<strong>{html.escape(metric["value"])}</strong><small>{html.escape(metric["unit"])}</small></article>'
            )
    commentary_html = ""
    if commentary:
        blocks = []
        for key, title in (
            ("observations", "Calculated observations"),
            ("hypotheses", "Hypotheses"),
            ("questions", "Questions"),
            ("limitations", "Limitations"),
        ):
            items = commentary.get(key, [])
            list_items = (
                "".join(f"<li>{html.escape(item['text'])}</li>" for item in items)
                or "<li>None recorded.</li>"
            )
            blocks.append(f"<article><h3>{title}</h3><ul>{list_items}</ul></article>")
        commentary_html = f'<section><p class="eyebrow">Professional review</p><h2>Interpretation</h2><div class="commentary">{"".join(blocks)}</div></section>'
    coverage_rows = pack["coverage"]
    pnl_rows = pack["sections"]["monthly_pnl"].get("rows", [])
    budget_rows = pack["sections"]["budget_variance"].get("rows", [])
    ar_rows = pack["sections"]["receivables_aging"].get("buckets", [])
    ap_rows = pack["sections"]["payables_aging"].get("buckets", [])
    cash_rows = pack["sections"]["cash_movement"].get("rows", [])
    customer_rows = pack["sections"]["customer_concentration"].get("rows", [])
    service_rows = pack["sections"]["service_profitability"].get("rows", [])
    status_class = (
        "blocked"
        if pack["status"] == "blocked"
        else ("partial" if pack["status"] == "partial" else "ready")
    )
    return f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Management Control Pack · {html.escape(pack['entity'])}</title>
<style>
:root{{--navy:#002060;--blue:#0070c0;--cyan:#00b0f0;--ink:#171816;--muted:#68727d;--line:#dbe2ea;--paper:#fff;--soft:#f4f7fa;--red:#9e2f2f;--amber:#9a6416;--green:#116149}}
*{{box-sizing:border-box}}body{{margin:0;background:var(--paper);color:var(--ink);font:15px/1.5 "Instrument Sans",Inter,Arial,sans-serif}}main{{width:min(1180px,calc(100% - 40px));margin:auto;padding:54px 0 80px}}header{{border-top:8px solid var(--navy);padding:38px 0 34px;border-bottom:1px solid var(--line)}}.eyebrow{{margin:0 0 10px;color:var(--blue);font-size:12px;font-weight:800;letter-spacing:.12em;text-transform:uppercase}}h1{{margin:0;font-size:clamp(36px,6vw,72px);line-height:.98;letter-spacing:-.055em}}h2{{font-size:30px;letter-spacing:-.035em}}h3{{font-size:17px}}.meta{{display:flex;gap:18px;flex-wrap:wrap;margin-top:24px;color:var(--muted)}}.status{{display:inline-flex;padding:6px 10px;border:1px solid currentColor;font-weight:800;text-transform:uppercase;font-size:11px;letter-spacing:.08em}}.status.ready{{color:var(--green)}}.status.partial{{color:var(--amber)}}.status.blocked{{color:var(--red)}}section{{padding:42px 0;border-bottom:1px solid var(--line)}}.metrics{{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:1px;background:var(--line);border:1px solid var(--line)}}.metric{{background:#fff;padding:22px;min-height:126px}}.metric span{{display:block;color:var(--muted)}}.metric strong{{display:block;margin-top:20px;font-size:28px;letter-spacing:-.03em}}.metric small{{color:var(--blue)}}.table-wrap{{overflow:auto;border:1px solid var(--line)}}table{{border-collapse:collapse;width:100%;min-width:620px}}th,td{{padding:11px 13px;border-bottom:1px solid var(--line);text-align:right;white-space:nowrap}}th:first-child,td:first-child{{text-align:left}}th{{background:var(--navy);color:#fff;font-size:11px;text-transform:uppercase;letter-spacing:.06em}}tbody tr:nth-child(even){{background:var(--soft)}}.grid{{display:grid;grid-template-columns:1fr 1fr;gap:26px}}.commentary{{display:grid;grid-template-columns:1fr 1fr;gap:18px}}.commentary article{{border-top:3px solid var(--cyan);padding:10px 18px 18px;background:var(--soft)}}.boundary{{margin-top:28px;padding:18px;border-left:3px solid var(--blue);background:var(--soft)}}.empty{{color:var(--muted)}}@media(max-width:800px){{.grid,.commentary{{grid-template-columns:1fr}}main{{width:min(100% - 24px,1180px)}}}}
</style></head><body><main><header><p class="eyebrow">Vera · Management Control Pack</p><h1>{html.escape(pack['entity'])}</h1><div class="meta"><span>{pack['reporting_period']['start']} → {pack['reporting_period']['end']}</span><span>Cutoff {pack['reporting_period']['cutoff']}</span><span>{pack['currency']}</span><span class="status {status_class}">{html.escape(pack['status'])}</span></div></header>
<section><p class="eyebrow">Head metrics</p><h2>Current picture</h2><div class="metrics">{"".join(cards)}</div></section>
<section><p class="eyebrow">Evidence coverage</p><h2>What this export supports</h2>{_html_table(coverage_rows, ('section','status','reason'))}</section>
<section><p class="eyebrow">Performance</p><h2>Monthly P&amp;L</h2>{_html_table(pnl_rows, ('period','revenue','gross_profit','ebitda','net_result'))}</section>
<section class="grid"><article><p class="eyebrow">Budget</p><h2>EBITDA variance</h2>{_html_table(budget_rows, ('period','actual_ebitda','budget_ebitda','variance'))}</article><article><p class="eyebrow">Working capital</p><h2>Receivables aging</h2>{_html_table(ar_rows, ('bucket','amount'))}</article></section>
<section class="grid"><article><p class="eyebrow">Working capital</p><h2>Payables aging</h2>{_html_table(ap_rows, ('bucket','amount'))}</article><article><p class="eyebrow">Liquidity</p><h2>Cash movement</h2>{_html_table(cash_rows, ('period','inflow','outflow','net'))}</article></section>
<section class="grid"><article><p class="eyebrow">Concentration</p><h2>Top customers</h2>{_html_table(customer_rows, ('customer','revenue','share'))}</article><article><p class="eyebrow">Profitability</p><h2>Services</h2>{_html_table(service_rows, ('service','revenue','direct_cost','margin','margin_rate'))}</article></section>
{commentary_html}<p class="boundary">{html.escape(pack['professional_boundary'])}</p></main></body></html>"""


def _append_sheet(
    workbook: Workbook,
    title: str,
    rows: Sequence[Mapping[str, Any]],
    columns: Sequence[str],
) -> None:
    worksheet = workbook.create_sheet(title=title[:31])
    worksheet.append(list(columns))
    numeric_columns = {
        "actual",
        "actual_ebitda",
        "amount",
        "budget_ebitda",
        "byte_count",
        "cogs",
        "difference",
        "direct_cost",
        "ebitda",
        "expected",
        "gross_profit",
        "inflow",
        "margin",
        "margin_rate",
        "net",
        "net_result",
        "operating_expense",
        "outflow",
        "revenue",
        "share",
        "tolerance",
        "variance",
    }
    ratio_columns = {"margin_rate", "share"}
    for row in rows:
        values = []
        for column in columns:
            value = row.get(column, "")
            if isinstance(value, (dict, list, tuple)):
                value = json.dumps(value, ensure_ascii=False, sort_keys=True)
            elif column in numeric_columns and isinstance(value, str) and value:
                try:
                    decimal_value = Decimal(value)
                except InvalidOperation:
                    pass
                else:
                    value = (
                        int(decimal_value)
                        if decimal_value == decimal_value.to_integral()
                        else float(decimal_value)
                    )
            values.append(value)
        worksheet.append(values)
    header_fill = PatternFill("solid", fgColor="002060")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    worksheet.freeze_panes = "A2"
    if worksheet.max_row > 1:
        worksheet.auto_filter.ref = worksheet.dimensions
    for index, column in enumerate(columns, start=1):
        if column in numeric_columns:
            number_format = "0.00%" if column in ratio_columns else "#,##0.00###"
            for row_index in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row_index, column=index).number_format = (
                    number_format
                )
        values = [
            str(column),
            *(
                str(worksheet.cell(row=row, column=index).value or "")
                for row in range(2, min(worksheet.max_row, 100) + 1)
            ),
        ]
        worksheet.column_dimensions[get_column_letter(index)].width = min(
            max(len(value) for value in values) + 2, 34
        )
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.print_title_rows = "1:1"


def write_excel(path: Path, pack: Mapping[str, Any]) -> None:
    """Write one reviewable workbook; canonical JSON remains the exact source."""

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary_rows = (
        ("Entity", pack["entity"]),
        ("Period start", pack["reporting_period"]["start"]),
        ("Period end", pack["reporting_period"]["end"]),
        ("Cutoff", pack["reporting_period"]["cutoff"]),
        ("Currency", pack["currency"]),
        ("Pack status", pack["status"]),
        ("Review status", pack["report_status"]),
    )
    for row in summary_rows:
        summary.append(row)
    summary.append(())
    summary.append(("Metric ID", "Label", "Value", "Unit"))
    for metric in pack["metrics"].values():
        raw_value = Decimal(metric["value"])
        excel_value = (
            int(raw_value) if raw_value == raw_value.to_integral() else float(raw_value)
        )
        summary.append(
            (metric["metric_id"], metric["label"], excel_value, metric["unit"])
        )
        summary.cell(row=summary.max_row, column=3).number_format = (
            "0.00%" if metric["unit"] == "ratio" else "#,##0.00###"
        )
    for cell in summary[9]:
        cell.fill = PatternFill("solid", fgColor="002060")
        cell.font = Font(color="FFFFFF", bold=True)
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 38
    summary.column_dimensions["C"].width = 18
    summary.column_dimensions["D"].width = 14
    summary.freeze_panes = "A9"
    summary.sheet_properties.pageSetUpPr.fitToPage = True
    summary.page_setup.orientation = "landscape"
    summary.page_setup.fitToWidth = 1
    summary.page_setup.fitToHeight = 1
    summary.print_title_rows = "9:9"
    section_specs = (
        (
            "Monthly P&L",
            pack["sections"]["monthly_pnl"].get("rows", []),
            (
                "period",
                "revenue",
                "cogs",
                "gross_profit",
                "operating_expense",
                "ebitda",
                "net_result",
            ),
        ),
        (
            "Budget variance",
            pack["sections"]["budget_variance"].get("rows", []),
            ("period", "actual_ebitda", "budget_ebitda", "variance"),
        ),
        (
            "AR aging",
            pack["sections"]["receivables_aging"].get("buckets", []),
            ("bucket", "amount"),
        ),
        (
            "AP aging",
            pack["sections"]["payables_aging"].get("buckets", []),
            ("bucket", "amount"),
        ),
        (
            "Cash",
            pack["sections"]["cash_movement"].get("rows", []),
            ("period", "inflow", "outflow", "net"),
        ),
        (
            "Customers",
            pack["sections"]["customer_concentration"].get("rows", []),
            ("customer", "revenue", "share"),
        ),
        (
            "Services",
            pack["sections"]["service_profitability"].get("rows", []),
            ("service", "revenue", "direct_cost", "margin", "margin_rate"),
        ),
        ("Coverage", pack["coverage"], ("section", "status", "reason")),
        (
            "Controls",
            pack["controls"],
            ("role", "status", "actual", "expected", "difference", "tolerance"),
        ),
        (
            "Lineage",
            pack["source_lineage"],
            ("source_id", "sha256", "byte_count", "table_ids"),
        ),
    )
    for title, rows, columns in section_specs:
        _append_sheet(workbook, title, rows, columns)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def finalize_commentary(
    pack: Mapping[str, Any], commentary: Mapping[str, Any]
) -> dict[str, Any]:
    """Validate shape and metric closure without judging semantic quality."""

    if commentary.get("schema_version") != COMMENTARY_SCHEMA:
        raise PackContractError("Commentary schema_version is invalid.")
    if commentary.get("workflow_id") != WORKFLOW_ID:
        raise PackContractError("Commentary workflow_id is invalid.")
    if (
        commentary.get("pack_sha256")
        != hashlib.sha256(_canonical_json_bytes(pack)).hexdigest()
    ):
        raise PackContractError("Commentary belongs to a different pack.")
    known_metrics = set(pack["metrics"])
    normalized: dict[str, Any] = {
        "schema_version": COMMENTARY_SCHEMA,
        "workflow_id": WORKFLOW_ID,
        "pack_sha256": commentary["pack_sha256"],
        "status": "draft_pending_professional_review",
    }
    for key in ("observations", "hypotheses", "questions", "limitations"):
        raw_items = commentary.get(key, [])
        if not isinstance(raw_items, list) or len(raw_items) > 100:
            raise PackContractError(f"{key} must be a bounded list.")
        items = []
        for index, raw in enumerate(raw_items, start=1):
            if not isinstance(raw, dict):
                raise PackContractError(f"{key} item {index} must be an object.")
            text = _text(
                raw.get("text"), label=f"{key} item {index} text", maximum=1200
            )
            refs = raw.get("metric_ids", [])
            if not isinstance(refs, list) or any(
                ref not in known_metrics for ref in refs
            ):
                raise PackContractError(
                    f"{key} item {index} references an unknown metric."
                )
            if key in {"observations", "hypotheses"} and not refs:
                raise PackContractError(
                    f"{key} item {index} requires at least one metric reference."
                )
            items.append({"text": text, "metric_ids": list(dict.fromkeys(refs))})
        normalized[key] = items
    return normalized
