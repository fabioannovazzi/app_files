"""Profile a tabular dataset into chart-useful mechanical role evidence."""

from __future__ import annotations

import argparse
import json
import math
import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

import polars as pl
from openpyxl import load_workbook

__all__ = ["build_dataset_profile", "profile_dataset", "main"]


def get_schema_and_column_names(
    frame: pl.DataFrame,
) -> tuple[list[str], dict[str, pl.DataType]]:
    """Return the stable column order and schema without repository dependencies."""

    schema = dict(frame.schema)
    return list(schema), schema


def get_row_count(frame: pl.DataFrame) -> int:
    """Return the Polars row count used throughout the standalone component."""

    return frame.height


VALUE_HINTS = {
    "amount",
    "cogs",
    "cost",
    "discount",
    "margin",
    "revenue",
    "sales",
    "salesamount",
    "turnover",
    "value",
}
VOLUME_HINTS = {"qty", "quantity", "unit", "units", "volume"}
COUNT_HINTS = {"count", "customers", "orders", "products", "visits"}
RATE_HINTS = {"asp", "averageprice", "avgprice", "conversion", "price", "rate", "ratio"}
SHARE_HINTS = {"percent", "percentage", "pct", "share"}
SCORE_HINTS = {"index", "rank", "score"}
IDENTIFIER_HINTS = {
    "barcode",
    "canonicalid",
    "code",
    "customerid",
    "id",
    "itemid",
    "parentid",
    "productid",
    "sku",
    "url",
}
PERIOD_HINTS = {"date", "month", "orderdate", "period", "quarter", "week", "year"}
ENTITY_HINTS = {
    "account",
    "barcode",
    "brand",
    "company",
    "customer",
    "item",
    "product",
    "sku",
    "variant",
}
SET_DIMENSION_HINTS = {
    "channel",
    "category",
    "class",
    "company",
    "group",
    "market",
    "region",
    "retailer",
    "scenario",
    "segment",
    "set",
    "store",
    "type",
}
RANK_OR_LANE_HINTS = {"band", "bucket", "class", "lane", "pareto", "rank", "tier"}
ORDERED_STAGE_HINTS = {"funnel", "phase", "step"}
STATEMENT_LINE_HINTS = {
    "account",
    "line",
    "lineitem",
    "p&l",
    "pnl",
    "rowkey",
    "statement",
}
STATEMENT_SCENARIO_HINTS = {"scenario"}


def _json_safe(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        return value
    return value


def _tokens(name: str) -> set[str]:
    with_boundaries = re.sub(r"(?<=[a-z0-9])(?=[A-Z])", " ", name)
    normalized = re.sub(r"[^a-z0-9]+", " ", with_boundaries.casefold())
    compact = re.sub(r"[^a-z0-9]+", "", name.casefold())
    return {token for token in normalized.split() if token} | {compact}


def _has_hint(name: str, hints: set[str]) -> bool:
    tokens = _tokens(name)
    return any(hint in tokens for hint in hints)


def _is_identifier_name(name: str) -> bool:
    """Identify mechanical ID columns without treating labels like "key benefits" as IDs."""

    tokens = _tokens(name)
    compact = re.sub(r"[^a-z0-9]+", "", name.casefold())
    if any(hint in tokens for hint in IDENTIFIER_HINTS):
        return True
    return compact.endswith("key") and any(
        stem in compact
        for stem in (
            "canonical",
            "category",
            "comp",
            "customer",
            "parent",
            "product",
            "variant",
        )
    )


def _is_temporal_dtype(dtype: pl.DataType) -> bool:
    return bool(dtype.is_temporal())


def _is_numeric_dtype(dtype: pl.DataType) -> bool:
    return bool(dtype.is_numeric())


def _is_string_dtype(dtype: pl.DataType) -> bool:
    return dtype in {pl.String, pl.Categorical, pl.Enum}


def _clean_headers(headers: list[Any]) -> list[str]:
    counts: defaultdict[str, int] = defaultdict(int)
    cleaned: list[str] = []
    for index, header in enumerate(headers, start=1):
        base = str(header).strip() if header not in (None, "") else f"column_{index}"
        counts[base] += 1
        cleaned.append(base if counts[base] == 1 else f"{base}_{counts[base]}")
    return cleaned


def _load_frame(
    path: Path, *, sheet_name: str | None
) -> tuple[pl.DataFrame, dict[str, Any]]:
    suffix = path.suffix.casefold()
    if suffix in {".parquet", ".pq"}:
        return pl.read_parquet(path), {"format": "parquet", "sheet_name": None}
    if suffix == ".csv":
        return pl.read_csv(path, infer_schema_length=10000), {
            "format": "csv",
            "sheet_name": None,
        }
    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook[sheet_name] if sheet_name else workbook.active
        rows = worksheet.iter_rows(values_only=True)
        headers = _clean_headers(list(next(rows)))
        values = [tuple(row) for row in rows]
        frame = pl.DataFrame(
            values, schema=headers, orient="row", infer_schema_length=None
        )
        return frame, {"format": suffix.lstrip("."), "sheet_name": worksheet.title}
    raise ValueError(f"Unsupported dataset format for {path}")


def _sample_values(series: pl.Series, limit: int = 8) -> list[Any]:
    values = series.drop_nulls().unique(maintain_order=True).head(limit).to_list()
    return [_json_safe(value) for value in values]


def _parse_period_value(value: Any) -> date | None:
    """Parse common machine-readable period labels for mechanical profiling."""

    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, int) and 1000 <= value <= 9999:
        return date(value, 1, 1)
    if isinstance(value, float) and value.is_integer() and 1000 <= value <= 9999:
        return date(int(value), 1, 1)
    if not isinstance(value, str):
        return None
    candidate = value.strip()
    if not candidate:
        return None
    try:
        return datetime.fromisoformat(candidate).date()
    except ValueError:
        pass
    quarter_match = re.fullmatch(
        r"(?:(?P<year_first>\d{4})[- /]?Q(?P<quarter_first>[1-4])|"
        r"Q(?P<quarter_second>[1-4])[- /]?(?P<year_second>\d{4}))",
        candidate,
        flags=re.IGNORECASE,
    )
    if quarter_match:
        year = int(
            quarter_match.group("year_first") or quarter_match.group("year_second")
        )
        quarter = int(
            quarter_match.group("quarter_first")
            or quarter_match.group("quarter_second")
        )
        return date(year, 1 + (quarter - 1) * 3, 1)
    week_match = re.fullmatch(
        r"(?P<year>\d{4})[- /]?W(?P<week>\d{1,2})",
        candidate,
        flags=re.IGNORECASE,
    )
    if week_match:
        try:
            return date.fromisocalendar(
                int(week_match.group("year")), int(week_match.group("week")), 1
            )
        except ValueError:
            return None
    for fmt in (
        "%Y-%m-%d",
        "%Y-%m",
        "%Y/%m/%d",
        "%Y/%m",
        "%d/%m/%Y",
        "%b-%Y",
        "%b %Y",
        "%B-%Y",
        "%B %Y",
        "%Y",
    ):
        try:
            return datetime.strptime(candidate, fmt).date()
        except ValueError:
            continue
    return None


def _parse_sample_dates(values: list[Any]) -> list[date]:
    parsed: list[date] = []
    for value in values:
        parsed_value = _parse_period_value(value)
        if parsed_value is not None:
            parsed.append(parsed_value)
    return parsed


def _full_period_bounds(series: pl.Series) -> tuple[date | None, date | None, int]:
    """Return exact bounds from all distinct parseable period values."""

    parsed = _parse_sample_dates(series.drop_nulls().unique().to_list())
    if not parsed:
        return None, None, 0
    return min(parsed), max(parsed), len(parsed)


def _period_parseability(dtype: pl.DataType, values: list[Any]) -> dict[str, Any]:
    """Return explicit period parse evidence without semantic interpretation."""

    sample_count = len([value for value in values if value not in (None, "")])
    parsed = _parse_sample_dates(values)
    parse_success_count = len(parsed)
    native_temporal = _is_temporal_dtype(dtype)
    parse_success_ratio = parse_success_count / sample_count if sample_count else 0.0
    is_parseable = native_temporal or (sample_count > 0 and parse_success_ratio >= 0.75)
    if native_temporal:
        parser = "native_temporal_dtype"
    elif is_parseable:
        parser = "sample_known_period_formats"
    elif parse_success_count:
        parser = "partial_period_sample_below_threshold"
    else:
        parser = "not_parseable_as_period_sample"
    return {
        "is_parseable": is_parseable,
        "parser": parser,
        "sample_count": sample_count,
        "parse_success_count": parse_success_count,
        "parse_success_ratio": parse_success_ratio,
        "parsed_min": _json_safe(min(parsed)) if parsed else None,
        "parsed_max": _json_safe(max(parsed)) if parsed else None,
        "inferred_grain": _infer_period_grain(values) if parsed else "unknown",
    }


def _infer_period_grain(values: list[Any]) -> str:
    non_empty = [str(value).strip() for value in values if value not in (None, "")]
    if non_empty and all(
        re.fullmatch(r"(?:\d{4}[- /]?Q[1-4]|Q[1-4][- /]?\d{4})", value, re.I)
        for value in non_empty
    ):
        return "quarter"
    if non_empty and all(
        re.fullmatch(r"\d{4}[- /]?W\d{1,2}", value, re.I) for value in non_empty
    ):
        return "week"
    if non_empty and all(re.fullmatch(r"\d{4}", value) for value in non_empty):
        return "year"
    month_patterns = (
        r"\d{4}[-/]\d{1,2}",
        r"(?:[A-Za-z]{3,9})[- ]\d{4}",
    )
    if non_empty and all(
        any(re.fullmatch(pattern, value) for pattern in month_patterns)
        for value in non_empty
    ):
        return "month"
    dates = sorted(set(_parse_sample_dates(values)))
    if len(dates) < 2:
        return "unknown"
    if all(value.day == 1 for value in dates):
        month_indexes = [value.year * 12 + value.month for value in dates]
        month_deltas = [
            right - left for left, right in zip(month_indexes, month_indexes[1:])
        ]
        if month_deltas and all(delta % 12 == 0 for delta in month_deltas):
            return "year"
        if month_deltas and all(delta % 3 == 0 for delta in month_deltas):
            return "quarter"
        if month_deltas and all(delta >= 1 for delta in month_deltas):
            return "month"
    deltas = sorted({(right - left).days for left, right in zip(dates, dates[1:])})
    if deltas == [1]:
        return "day"
    if deltas == [7]:
        return "week"
    if all(27 <= delta <= 32 for delta in deltas):
        return "month"
    if all(89 <= delta <= 92 for delta in deltas):
        return "quarter"
    if all(360 <= delta <= 370 for delta in deltas):
        return "year"
    return "irregular"


def _cardinality_class(row_count: int, distinct_count: int) -> str:
    if distinct_count == 0:
        return "empty"
    if distinct_count == 1:
        return "constant"
    if row_count and distinct_count / row_count > 0.8:
        return "high"
    if distinct_count <= 20:
        return "low"
    if distinct_count <= 200:
        return "medium"
    return "high"


def _metric_class_and_aggregation(name: str) -> tuple[str, str, list[str]]:
    reasons: list[str] = []
    if _has_hint(name, SHARE_HINTS):
        reasons.append("name suggests share or percent")
        return "share", "semantic_layer_defined", reasons
    if _has_hint(name, RATE_HINTS):
        reasons.append("name suggests price, rate, or ratio")
        return "rate", "semantic_layer_defined", reasons
    if _has_hint(name, SCORE_HINTS):
        reasons.append("name suggests score, rank, or index")
        return "score", "semantic_layer_defined", reasons
    if _has_hint(name, VOLUME_HINTS):
        reasons.append("name suggests additive volume")
        return "additive_volume", "sum", reasons
    if _has_hint(name, COUNT_HINTS):
        reasons.append("name suggests additive count")
        return "additive_count", "sum_or_count", reasons
    if _has_hint(name, VALUE_HINTS):
        reasons.append("name suggests additive value")
        return "additive_value", "sum", reasons
    reasons.append("numeric column without a strong metric-class hint")
    return "numeric_observation", "semantic_layer_defined", reasons


def _has_metric_name_hint(name: str) -> bool:
    return any(
        _has_hint(name, hints)
        for hints in (
            SHARE_HINTS,
            RATE_HINTS,
            SCORE_HINTS,
            VOLUME_HINTS,
            COUNT_HINTS,
            VALUE_HINTS,
        )
    )


def _numeric_like_share(values: list[Any]) -> float:
    if not values:
        return 0.0
    numeric = 0
    for value in values:
        if not isinstance(value, str):
            continue
        candidate = (
            value.strip()
            .replace("$", "")
            .replace("€", "")
            .replace("£", "")
            .replace(",", "")
            .replace("%", "")
        )
        if not candidate:
            continue
        try:
            float(candidate)
            numeric += 1
        except ValueError:
            continue
    return numeric / len(values)


def _classify_column(
    *,
    name: str,
    dtype: pl.DataType,
    row_count: int,
    distinct_count: int,
    sample_values: list[Any],
) -> dict[str, Any]:
    reasons: list[str] = []
    cardinality = _cardinality_class(row_count, distinct_count)
    period_evidence = _period_parseability(dtype, sample_values)
    if _is_temporal_dtype(dtype) or (
        _has_hint(name, PERIOD_HINTS) and period_evidence["is_parseable"]
    ):
        reasons.append("temporal type or parseable period-like values")
        return {
            "role": "period",
            "role_confidence": "high",
            "period_grain": _infer_period_grain(sample_values),
            "cardinality_class": cardinality,
            "inference_reasons": reasons,
        }
    if _is_identifier_name(name):
        reasons.append("name suggests identifier")
        return {
            "role": "identifier",
            "role_confidence": "high",
            "cardinality_class": cardinality,
            "inference_reasons": reasons,
        }
    if (
        _is_string_dtype(dtype)
        and _has_metric_name_hint(name)
        and _numeric_like_share(sample_values) >= 0.75
    ):
        metric_class, aggregation, metric_reasons = _metric_class_and_aggregation(name)
        reasons.extend(metric_reasons)
        reasons.append("string values look numeric after currency/percent cleanup")
        return {
            "role": "metric",
            "role_confidence": "medium",
            "metric_class": metric_class,
            "aggregation": aggregation,
            "requires_cast": True,
            "cardinality_class": cardinality,
            "inference_reasons": reasons,
        }
    if _is_numeric_dtype(dtype):
        metric_class, aggregation, metric_reasons = _metric_class_and_aggregation(name)
        reasons.extend(metric_reasons)
        return {
            "role": "metric",
            "role_confidence": (
                "medium" if metric_class == "numeric_observation" else "high"
            ),
            "metric_class": metric_class,
            "aggregation": aggregation,
            "cardinality_class": cardinality,
            "inference_reasons": reasons,
        }
    if _is_string_dtype(dtype) or dtype == pl.Boolean:
        reasons.append("categorical-compatible physical type")
        return {
            "role": "dimension",
            "role_confidence": "medium" if cardinality == "high" else "high",
            "cardinality_class": cardinality,
            "inference_reasons": reasons,
        }
    reasons.append("unsupported or mixed physical type")
    return {
        "role": "unknown",
        "role_confidence": "low",
        "cardinality_class": cardinality,
        "inference_reasons": reasons,
    }


def _profile_column(
    series: pl.Series, dtype: pl.DataType, row_count: int
) -> dict[str, Any]:
    null_count = int(series.null_count())
    distinct_count = int(series.n_unique())
    samples = _sample_values(series)
    classification = _classify_column(
        name=series.name,
        dtype=dtype,
        row_count=row_count,
        distinct_count=distinct_count,
        sample_values=samples,
    )
    profile: dict[str, Any] = {
        "physical_type": str(dtype),
        "null_count": null_count,
        "null_ratio": null_count / row_count if row_count else 0.0,
        "distinct_count": distinct_count,
        "sample_values": samples,
        "period_parseability": _period_parseability(dtype, samples),
        **classification,
    }
    non_null = series.drop_nulls()
    if get_row_count(non_null.to_frame()) > 0 and (
        _is_numeric_dtype(dtype) or _is_temporal_dtype(dtype)
    ):
        profile["min"] = _json_safe(non_null.min())
        profile["max"] = _json_safe(non_null.max())
    if profile["role"] == "period" and get_row_count(non_null.to_frame()) > 0:
        period_values = non_null.unique().sort().head(500).to_list()
        profile["ordered_values"] = [_json_safe(value) for value in period_values]
        profile["ordered_values_complete"] = distinct_count <= 500
        profile["period_grain"] = _infer_period_grain(period_values)
        parsed_min, parsed_max, parsed_distinct_count = _full_period_bounds(non_null)
        if parsed_min is not None and parsed_max is not None:
            profile["period_parseability"].update(
                {
                    "parsed_min": _json_safe(parsed_min),
                    "parsed_max": _json_safe(parsed_max),
                    "bounds_source": "full_column_distinct_values",
                    "parsed_distinct_count": parsed_distinct_count,
                }
            )
    if get_row_count(non_null.to_frame()) > 0 and _is_numeric_dtype(dtype):
        profile["sum"] = _json_safe(non_null.sum())
    return profile


def _profile_frame(
    frame: pl.DataFrame,
    *,
    dataset_id: str,
    source: dict[str, Any],
) -> dict[str, Any]:
    columns, schema = get_schema_and_column_names(frame)
    if schema is None:
        raise ValueError("Polars frame schema is unavailable.")
    row_count = get_row_count(frame)
    column_profiles = {
        column: _profile_column(frame.get_column(column), schema[column], row_count)
        for column in columns
    }
    role_index: dict[str, list[str]] = defaultdict(list)
    metric_class_index: dict[str, list[str]] = defaultdict(list)
    for column, profile in column_profiles.items():
        role = str(profile["role"])
        role_index[role].append(column)
        metric_class = profile.get("metric_class")
        if isinstance(metric_class, str):
            metric_class_index[metric_class].append(column)
    derived_metrics = _derived_metric_candidates(column_profiles)
    for metric_name, metric_profile in derived_metrics.items():
        metric_class_index[str(metric_profile["metric_class"])].append(metric_name)
    dimension_relationships = _dimension_relationships(frame, column_profiles)
    role_candidates = _canonical_role_candidates(
        column_profiles,
        specialized_candidates=_role_candidates(column_profiles),
    )
    return {
        "schema_version": "0.4",
        "dataset_id": dataset_id,
        "source": source,
        "row_count": row_count,
        "column_count": frame.width,
        "columns": column_profiles,
        "roles": dict(sorted(role_index.items())),
        "metric_classes": dict(sorted(metric_class_index.items())),
        "role_candidates": role_candidates,
        "role_candidate_columns": {
            role: [str(candidate["column"]) for candidate in candidates]
            for role, candidates in role_candidates.items()
        },
        "derived_metrics": derived_metrics,
        "dimension_relationships": dimension_relationships,
        "selector_profile": _selector_profile(
            column_profiles=column_profiles,
            role_candidates=role_candidates,
            derived_metrics=derived_metrics,
        ),
        "selector_boundary": (
            "This profile describes mechanically available dataset roles. It does "
            "not decide which analyses make business sense; a separate "
            "analysis-validity layer must do that."
        ),
        "review_status": {
            "metric_classification": "heuristic_needs_review",
            "analysis_validity": "not_included",
        },
    }


def _dimension_relationships(
    frame: pl.DataFrame,
    column_profiles: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    """Describe exact pairwise dimension dependence for mechanical chart checks."""

    dimensions = [
        column
        for column, profile in column_profiles.items()
        if profile.get("role") == "dimension"
        and int(profile.get("distinct_count") or 0) > 1
    ]
    relationships: list[dict[str, Any]] = []
    for left_index, left in enumerate(dimensions):
        for right in dimensions[left_index + 1 :]:
            pair = frame.select(left, right).drop_nulls()
            if pair.is_empty():
                continue
            left_count = int(pair.get_column(left).n_unique())
            right_count = int(pair.get_column(right).n_unique())
            joint_count = int(pair.unique().height)
            left_determines_right = joint_count == left_count
            right_determines_left = joint_count == right_count
            bijective_alias = (
                left_count > 1
                and left_count == right_count == joint_count
                and left_determines_right
                and right_determines_left
            )
            if bijective_alias:
                relationship = "bijective_alias"
            elif left_determines_right:
                relationship = "left_determines_right"
            elif right_determines_left:
                relationship = "right_determines_left"
            else:
                relationship = "cross_classifying"
            relationships.append(
                {
                    "left_column": left,
                    "right_column": right,
                    "non_null_pair_rows": pair.height,
                    "left_distinct_count": left_count,
                    "right_distinct_count": right_count,
                    "joint_distinct_count": joint_count,
                    "left_determines_right": left_determines_right,
                    "right_determines_left": right_determines_left,
                    "relationship": relationship,
                    "supports_multidimensional_path": not bijective_alias,
                }
            )
    return relationships


def _selector_profile(
    *,
    column_profiles: dict[str, dict[str, Any]],
    role_candidates: dict[str, list[dict[str, Any]]],
    derived_metrics: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    """Summarize chart-selection candidates without semantic approval."""

    period_candidates = [
        _period_candidate_record(column, profile)
        for column, profile in column_profiles.items()
        if profile.get("role") == "period"
    ]
    metric_candidates = [
        _metric_candidate_record(column, profile)
        for column, profile in column_profiles.items()
        if profile.get("role") == "metric"
    ]
    for metric_name, profile in derived_metrics.items():
        metric_candidates.append(_derived_metric_candidate_record(metric_name, profile))
    dimension_candidates = [
        _dimension_candidate_record(column, profile)
        for column, profile in column_profiles.items()
        if profile.get("role") == "dimension"
    ]
    identifier_candidates = [
        _dimension_candidate_record(column, profile)
        for column, profile in column_profiles.items()
        if profile.get("role") == "identifier"
    ]
    metric_candidates_by_class: defaultdict[str, list[dict[str, Any]]] = defaultdict(
        list
    )
    for candidate in metric_candidates:
        metric_candidates_by_class[str(candidate.get("metric_class"))].append(candidate)
    return {
        "period_candidates": sorted(
            period_candidates,
            key=lambda item: (
                item["grain"] == "unknown",
                item["null_ratio"],
                item["column"],
            ),
        ),
        "metric_candidates": sorted(
            metric_candidates,
            key=lambda item: (
                item["metric_class"],
                item["requires_cast"],
                item["null_ratio"],
                item["column"],
            ),
        ),
        "metric_candidates_by_class": {
            metric_class: sorted(
                candidates,
                key=lambda item: (
                    item["requires_cast"],
                    item["null_ratio"],
                    item["column"],
                ),
            )
            for metric_class, candidates in sorted(metric_candidates_by_class.items())
        },
        "dimension_candidates": sorted(
            dimension_candidates,
            key=lambda item: (
                item["cardinality_rank"],
                item["null_ratio"],
                item["column"],
            ),
        ),
        "identifier_candidates": sorted(
            identifier_candidates,
            key=lambda item: (
                item["null_ratio"],
                -int(item.get("distinct_count") or 0),
                item["column"],
            ),
        ),
        "role_candidate_counts": {
            role: len(candidates)
            for role, candidates in sorted(role_candidates.items())
        },
        "column_role_counts": {
            role: sum(
                1
                for profile in column_profiles.values()
                if str(profile.get("role")) == role
            )
            for role in sorted(
                {str(profile.get("role")) for profile in column_profiles.values()}
            )
        },
        "boundary": (
            "Mechanical selector profile only. Candidate rankings are based on "
            "types, names, nulls, and cardinality; semantic validity is not checked."
        ),
    }


def _base_selector_candidate(column: str, profile: dict[str, Any]) -> dict[str, Any]:
    return {
        "column": column,
        "source_role": profile.get("role"),
        "confidence": profile.get("role_confidence"),
        "null_ratio": profile.get("null_ratio"),
        "distinct_count": profile.get("distinct_count"),
        "cardinality_class": profile.get("cardinality_class"),
        "inference_reasons": profile.get("inference_reasons") or [],
    }


def _period_candidate_record(column: str, profile: dict[str, Any]) -> dict[str, Any]:
    record = _base_selector_candidate(column, profile)
    record.update(
        {
            "grain": profile.get("period_grain"),
            "period_parseability": profile.get("period_parseability") or {},
            "ordered_values": profile.get("ordered_values") or [],
            "ordered_values_complete": bool(
                profile.get("ordered_values_complete", False)
            ),
            "min": profile.get("min"),
            "max": profile.get("max"),
            "selector_use": "period_axis_or_filter",
        }
    )
    return record


def _metric_candidate_record(column: str, profile: dict[str, Any]) -> dict[str, Any]:
    record = _base_selector_candidate(column, profile)
    record.update(
        {
            "metric_class": profile.get("metric_class"),
            "aggregation": profile.get("aggregation"),
            "requires_cast": bool(profile.get("requires_cast", False)),
            "min": profile.get("min"),
            "max": profile.get("max"),
            "sum": profile.get("sum"),
            "selector_use": "source_metric",
        }
    )
    return record


def _derived_metric_candidate_record(
    metric_name: str, profile: dict[str, Any]
) -> dict[str, Any]:
    return {
        "column": metric_name,
        "source_role": "derived_metric",
        "confidence": "medium",
        "null_ratio": 0.0,
        "distinct_count": None,
        "cardinality_class": "derived",
        "metric_class": profile.get("metric_class"),
        "aggregation": profile.get("aggregation", "semantic_layer_defined"),
        "requires_cast": False,
        "produced_from": profile.get("produced_from") or [],
        "inference_reasons": profile.get("inference_reasons") or [],
        "selector_use": "derived_metric_candidate",
    }


def _dimension_candidate_record(column: str, profile: dict[str, Any]) -> dict[str, Any]:
    cardinality_class = str(profile.get("cardinality_class") or "")
    cardinality_rank = {
        "low": 0,
        "medium": 1,
        "constant": 2,
        "high": 3,
        "empty": 4,
    }.get(cardinality_class, 9)
    record = _base_selector_candidate(column, profile)
    record.update(
        {
            "cardinality_rank": cardinality_rank,
            "selector_use": "dimension_or_identifier",
        }
    )
    return record


def _candidate_record(
    column: str, profile: dict[str, Any], *, reason: str, confidence: str
) -> dict[str, Any]:
    return {
        "column": column,
        "source_role": profile.get("role"),
        "confidence": confidence,
        "distinct_count": profile.get("distinct_count"),
        "cardinality_class": profile.get("cardinality_class"),
        "null_ratio": profile.get("null_ratio"),
        "reason": reason,
    }


def _is_entity_candidate_name(column: str, source_role: str) -> bool:
    tokens = _tokens(column)
    if {"image", "url"} & tokens and "pdp" not in tokens:
        return False
    return source_role == "identifier" or _has_hint(column, ENTITY_HINTS)


def _is_ordered_stage_name(column: str) -> bool:
    """Recognize explicit process-stage fields without matching usage labels."""

    tokens = _tokens(column)
    compact = re.sub(r"[^a-z0-9]+", "", column.casefold())
    if "usage" in tokens:
        return False
    return compact in {"stage", "funnelstage", "salesstage"} or _has_hint(
        column, ORDERED_STAGE_HINTS
    )


def _is_explicit_set_dimension_name(column: str) -> bool:
    compact = re.sub(r"[^a-z0-9]+", "", column.casefold())
    return compact in {"set", "setid", "setname", "setgroup", "membershipset"}


def _role_candidates(
    column_profiles: dict[str, dict[str, Any]],
) -> dict[str, list[dict[str, Any]]]:
    """Collect mechanical role candidates; semantic validity is deliberately out of scope."""

    candidates: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for column, profile in column_profiles.items():
        role = str(profile.get("role"))
        distinct_count = int(profile.get("distinct_count") or 0)
        null_ratio = float(profile.get("null_ratio") or 0.0)
        cardinality = str(profile.get("cardinality_class") or "")
        if role == "dimension":
            candidates["direct_dimension"].append(
                _candidate_record(
                    column,
                    profile,
                    reason="categorical column available for grouping",
                    confidence=str(profile.get("role_confidence") or "medium"),
                )
            )
        if (
            role in {"dimension", "identifier"}
            and distinct_count > 1
            and null_ratio < 0.5
        ):
            if _is_entity_candidate_name(column, role):
                confidence = "high" if role == "identifier" else "medium"
                candidates["entity_key"].append(
                    _candidate_record(
                        column,
                        profile,
                        reason="stable entity candidate for cohort or set derivation",
                        confidence=confidence,
                    )
                )
        if role in {"dimension", "identifier"} and distinct_count > 1:
            if _is_entity_candidate_name(column, role):
                candidates["set_item"].append(
                    _candidate_record(
                        column,
                        profile,
                        reason="item candidate for set membership aggregation",
                        confidence="high" if role == "identifier" else "medium",
                    )
                )
        if role == "dimension" and 2 <= distinct_count <= 50:
            confidence = (
                "high"
                if _is_explicit_set_dimension_name(column)
                or _has_hint(column, SET_DIMENSION_HINTS)
                else "medium"
            )
            candidates["set_dimension"].append(
                _candidate_record(
                    column,
                    profile,
                    reason="low-cardinality grouping candidate for set membership",
                    confidence=confidence,
                )
            )
        if role in {"dimension", "metric"} and _has_hint(column, RANK_OR_LANE_HINTS):
            candidates["rank_or_lane"].append(
                _candidate_record(
                    column,
                    profile,
                    reason="name suggests rank, lane, bucket, class, or band",
                    confidence="medium",
                )
            )
        if role == "dimension" and _is_ordered_stage_name(column):
            compact = re.sub(r"[^a-z0-9]+", "", column.casefold())
            exact_funnel_name = compact in {"funnelstage", "salesstage"}
            candidates["ordered_stage"].append(
                _candidate_record(
                    column,
                    profile,
                    reason=(
                        "name explicitly identifies a funnel stage"
                        if exact_funnel_name
                        else "name suggests ordered process stage"
                    ),
                    confidence="high" if exact_funnel_name else "medium",
                )
            )
        if role == "dimension" and _has_hint(column, STATEMENT_LINE_HINTS):
            candidates["statement_line_item"].append(
                _candidate_record(
                    column,
                    profile,
                    reason="name suggests financial statement line item",
                    confidence="medium",
                )
            )
        if role == "dimension" and _has_hint(column, STATEMENT_SCENARIO_HINTS):
            candidates["statement_scenario"].append(
                _candidate_record(
                    column,
                    profile,
                    reason="name identifies a statement scenario field",
                    confidence="high",
                )
            )
    return {
        key: _sorted_role_candidates(key, value)
        for key, value in sorted(candidates.items())
    }


def _canonical_role_candidates(
    column_profiles: dict[str, dict[str, Any]],
    *,
    specialized_candidates: dict[str, list[dict[str, Any]]],
) -> dict[str, list[dict[str, Any]]]:
    """Expose canonical chart roles while retaining richer profile roles."""

    candidates = {
        role: list(records) for role, records in specialized_candidates.items()
    }
    periods: list[dict[str, Any]] = []
    metrics: list[dict[str, Any]] = []
    dimensions: list[dict[str, Any]] = []
    panels: list[dict[str, Any]] = []
    identifiers: list[dict[str, Any]] = []
    for column, profile in column_profiles.items():
        source_role = str(profile.get("role") or "")
        confidence = str(profile.get("role_confidence") or "medium")
        if source_role == "period":
            periods.append(
                _candidate_record(
                    column,
                    profile,
                    reason="parseable period column",
                    confidence=confidence,
                )
            )
        elif source_role == "metric":
            metric = _candidate_record(
                column,
                profile,
                reason="numeric metric candidate",
                confidence=confidence,
            )
            metric["metric_class"] = profile.get("metric_class")
            metrics.append(metric)
        elif source_role == "dimension":
            dimension = _candidate_record(
                column,
                profile,
                reason="categorical dimension candidate",
                confidence=confidence,
            )
            dimensions.append(dimension)
            distinct_count = int(profile.get("distinct_count") or 0)
            if 2 <= distinct_count <= 50:
                panels.append(dict(dimension))
        elif source_role == "identifier":
            identifiers.append(
                _candidate_record(
                    column,
                    profile,
                    reason="identifier candidate",
                    confidence=confidence,
                )
            )
    candidates.update(
        {
            "period_axis": periods,
            "period_filter": list(periods),
            "comparison_metric": metrics,
            "dimension_member": dimensions,
            "panel_dimension": panels,
            "identifier": identifiers,
        }
    )
    return {
        role: _sorted_role_candidates(role, records)
        for role, records in sorted(candidates.items())
    }


def _sorted_role_candidates(
    role: str, candidates: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    confidence_rank = {"high": 0, "medium": 1, "low": 2}

    def set_dimension_priority(column: str) -> int:
        tokens = _tokens(column)
        if _is_explicit_set_dimension_name(column):
            return -1
        if tokens & {
            "channel",
            "market",
            "region",
            "retailer",
            "scenario",
            "segment",
            "store",
        }:
            return 0
        if "category" in tokens and "secondary" not in tokens:
            return 1
        if tokens & {"class", "company", "group", "type"}:
            return 2
        return 3

    def sort_key(candidate: dict[str, Any]) -> tuple[Any, ...]:
        source_role = str(candidate.get("source_role") or "")
        confidence = str(candidate.get("confidence") or "")
        distinct_count = int(candidate.get("distinct_count") or 0)
        if role in {"entity_key", "set_item"}:
            return (
                0 if source_role == "identifier" else 1,
                confidence_rank.get(confidence, 9),
                -distinct_count,
                str(candidate.get("column") or ""),
            )
        if role == "set_dimension":
            return (
                set_dimension_priority(str(candidate.get("column") or "")),
                confidence_rank.get(confidence, 9),
                distinct_count,
                str(candidate.get("column") or ""),
            )
        return (
            confidence_rank.get(confidence, 9),
            str(candidate.get("column") or ""),
        )

    return sorted(candidates, key=sort_key)


def _derived_metric_candidates(
    column_profiles: dict[str, dict[str, Any]],
) -> dict[str, dict[str, Any]]:
    value_columns = [
        column
        for column, profile in column_profiles.items()
        if profile.get("metric_class") == "additive_value"
    ]
    volume_columns = [
        column
        for column, profile in column_profiles.items()
        if profile.get("metric_class") == "additive_volume"
    ]
    derived: dict[str, dict[str, Any]] = {}
    for value_column in value_columns[:5]:
        for volume_column in volume_columns[:5]:
            metric_name = f"{value_column}_per_{volume_column}"
            derived[metric_name] = {
                "role": "metric",
                "metric_class": "derived_rate",
                "aggregation": "derived_from_value_and_volume",
                "produced_from": [value_column, volume_column],
                "review_status": "needs_semantic_review",
                "inference_reasons": [
                    "additive value metric divided by additive volume metric"
                ],
            }
    return derived


def build_dataset_profile(
    path: Path,
    *,
    dataset_id: str,
    sheet_name: str | None = None,
) -> dict[str, Any]:
    resolved_path = path.expanduser().resolve()
    frame, source = _load_frame(resolved_path, sheet_name=sheet_name)
    source["path"] = str(resolved_path)
    return _profile_frame(frame, dataset_id=dataset_id, source=source)


def profile_dataset(
    path: Path,
    *,
    dataset_id: str | None = None,
    sheet_name: str | None = None,
) -> dict[str, Any]:
    """Return the standalone Reporting Engine dataset profile."""

    return build_dataset_profile(
        path,
        dataset_id=dataset_id or path.stem,
        sheet_name=sheet_name,
    )


def _write_profile(path: Path, profile: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(profile, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )


def main(argv: list[str] | None = None) -> int:
    """Write one mechanical dataset profile outside the component."""

    parser = argparse.ArgumentParser(
        description="Build chart-selection dataset profiles."
    )
    parser.add_argument("dataset", type=Path, help="Dataset file to profile.")
    parser.add_argument(
        "--dataset-id",
        "--dataset-contract-id",
        dest="dataset_id",
        help=(
            "Stable logical dataset identifier for recurring snapshots; defaults "
            "to filename for standalone profiling only."
        ),
    )
    parser.add_argument("--sheet-name", help="Excel sheet name for xlsx/xlsm inputs.")
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args(argv)

    profile = profile_dataset(
        args.dataset,
        dataset_id=args.dataset_id,
        sheet_name=args.sheet_name,
    )
    _write_profile(args.output, profile)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
