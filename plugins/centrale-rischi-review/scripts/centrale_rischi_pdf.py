#!/usr/bin/env python3
"""Normalize native-text Centrale Rischi PDF tables for professional review."""

from __future__ import annotations

import hashlib
import re
import unicodedata
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Mapping, Sequence

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

__all__ = [
    "PDF_CORPUS_EVALUATION_SCHEMA",
    "PDF_NORMALIZATION_SCHEMA",
    "evaluate_pdf_corpus",
    "normalize_pdf",
    "write_normalized_workbook",
]

PDF_NORMALIZATION_SCHEMA = "vera.centrale_rischi_pdf_normalization.v2"
PDF_CORPUS_EVALUATION_SCHEMA = "vera.centrale_rischi_pdf_corpus_evaluation.v2"

EXPOSURE_HEADERS = (
    "reference_month",
    "intermediary",
    "category",
    "location",
    "original_duration",
    "residual_duration",
    "currency",
    "import_export",
    "activity_type",
    "relationship_status",
    "guarantee_type",
    "borrower_role",
    "granted",
    "operational_granted",
    "used",
    "average_balance",
    "guaranteed_amount",
    "record_status",
    "valid_from",
    "valid_to",
    "source_page",
    "source_region",
    "source_row_locator",
    "extraction_confidence",
    "source_document_sha256",
)

GUARANTEE_HEADERS = (
    "reference_month",
    "intermediary",
    "category",
    "location",
    "guaranteed_party",
    "relationship_status",
    "guarantee_type",
    "guarantee_value",
    "guaranteed_amount",
    "record_status",
    "valid_from",
    "valid_to",
    "source_page",
    "source_region",
    "source_row_locator",
    "extraction_confidence",
    "source_document_sha256",
)

GUARANTOR_HEADERS = (
    "reference_month",
    "intermediary",
    "guarantor",
    "guarantee_value",
    "guaranteed_amount",
    "record_status",
    "valid_from",
    "valid_to",
    "source_page",
    "source_region",
    "source_row_locator",
    "extraction_confidence",
    "source_document_sha256",
)

CEDED_DEBTOR_HEADERS = (
    "reference_month",
    "intermediary",
    "ceded_debtor",
    "nominal_value",
    "record_status",
    "valid_from",
    "valid_to",
    "source_page",
    "source_region",
    "source_row_locator",
    "extraction_confidence",
    "source_document_sha256",
)

OTHER_RISK_HEADERS = (
    "reference_month",
    "intermediary",
    "category",
    "location",
    "original_duration",
    "residual_duration",
    "currency",
    "import_export",
    "activity_type",
    "relationship_status",
    "amount",
    "intrinsic_value",
    "record_status",
    "valid_from",
    "valid_to",
    "source_page",
    "source_region",
    "source_row_locator",
    "extraction_confidence",
    "source_document_sha256",
)

SUMMARY_TOTAL_HEADERS = (
    "reference_month",
    "intermediary",
    "summary_category",
    "granted",
    "operational_granted",
    "used",
    "guarantee_value",
    "guaranteed_amount",
    "intrinsic_value",
    "source_page",
    "source_region",
    "source_row_locator",
    "extraction_confidence",
    "source_document_sha256",
)

EVENT_HEADERS = (
    "intermediary",
    "event_date",
    "event_type",
    "event_cancelled",
    "source_page",
    "source_region",
    "source_row_locator",
    "extraction_confidence",
    "source_document_sha256",
)

REQUEST_HEADERS = (
    "intermediary",
    "request_date",
    "requested_period",
    "request_type",
    "request_reason_code",
    "request_reason",
    "validity_period",
    "notes",
    "source_page",
    "source_region",
    "source_row_locator",
    "extraction_confidence",
    "source_document_sha256",
)

SHEET_CONTRACTS: Mapping[str, tuple[str, ...]] = {
    "Esposizioni": EXPOSURE_HEADERS,
    "Garanzie ricevute": GUARANTEE_HEADERS,
    "Garanti intestatario": GUARANTOR_HEADERS,
    "Debitori ceduti": CEDED_DEBTOR_HEADERS,
    "Altre informazioni": OTHER_RISK_HEADERS,
    "Prospetto sintetico": SUMMARY_TOTAL_HEADERS,
    "Eventi inframensili": EVENT_HEADERS,
    "Richieste informazioni": REQUEST_HEADERS,
}

_HEADER_ALIASES = {
    "categoria": "category",
    "localizzazione": "location",
    "durata originaria": "original_duration",
    "durata residua": "residual_duration",
    "divisa": "currency",
    "import export": "import_export",
    "tipo attivita": "activity_type",
    "stato rapporto": "relationship_status",
    "tipo garanzia": "guarantee_type",
    "ruolo affidato": "borrower_role",
    "accordato": "granted",
    "accordato operativo": "operational_granted",
    "utilizzato": "used",
    "saldo medio": "average_balance",
    "importo garantito": "guaranteed_amount",
    "garantito": "guaranteed_party",
    "garante": "guarantor",
    "valore garanzia": "guarantee_value",
    "ceduto": "ceded_debtor",
    "valore nominale del credito ceduto": "nominal_value",
    "importo": "amount",
    "valore intrinseco": "intrinsic_value",
    "da": "valid_from",
    "a": "valid_to",
    "data evento": "event_date",
    "tipo evento": "event_type",
    "evento cancellato": "event_cancelled",
    "data della richiesta di informazione": "request_date",
    "periodo richiesto": "requested_period",
    "tipo richiesta di informazione": "request_type",
    "causale della richiesta": "request_reason_code",
    "descrizione causale": "request_reason",
    "periodo validita": "validity_period",
    "note": "notes",
}

_ITALIAN_MONTHS = {
    "gennaio": 1,
    "febbraio": 2,
    "marzo": 3,
    "aprile": 4,
    "maggio": 5,
    "giugno": 6,
    "luglio": 7,
    "agosto": 8,
    "settembre": 9,
    "ottobre": 10,
    "novembre": 11,
    "dicembre": 12,
}


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _clean_cell(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\u00a0", " ").split())


def _fold(value: Any) -> str:
    text = unicodedata.normalize("NFKD", _clean_cell(value))
    text = "".join(
        character for character in text if not unicodedata.combining(character)
    )
    return re.sub(r"[^a-z0-9]+", " ", text.casefold()).strip()


def _canonical_header(value: Any) -> str:
    folded = _fold(value)
    if folded in _HEADER_ALIASES:
        return _HEADER_ALIASES[folded]
    # Diagonal educational watermarks can cross a header cell. This removes
    # only isolated leading letters when the remaining text is an exact label.
    candidate = re.sub(r"^(?:[a-z]\s+){1,4}", "", folded)
    return _HEADER_ALIASES.get(candidate, folded.replace(" ", "_"))


def _collapse_blank_header_columns(rows: Sequence[Sequence[Any]]) -> list[list[str]]:
    if not rows:
        return []
    normalized = [[_clean_cell(cell) for cell in row] for row in rows]
    width = max(len(row) for row in normalized)
    padded = [row + [""] * (width - len(row)) for row in normalized]
    groups: list[list[int]] = []
    for index, header in enumerate(padded[0]):
        if not _fold(header) and groups:
            groups[-1].append(index)
        else:
            groups.append([index])
    collapsed: list[list[str]] = []
    for row in padded:
        collapsed.append(
            [_clean_cell(" ".join(row[index] for index in group)) for group in groups]
        )
    return collapsed


def _merge_fragmented_header_columns(rows: Sequence[Sequence[Any]]) -> list[list[str]]:
    """Join adjacent header fragments only when they form one exact known label.

    Exact alias reconstruction is deterministic and mechanically reviewable; it
    does not infer a semantic role from approximate text or document context.
    """

    if not rows:
        return []
    normalized = [[_clean_cell(cell) for cell in row] for row in rows]
    width = max(len(row) for row in normalized)
    padded = [row + [""] * (width - len(row)) for row in normalized]
    compact_aliases = {
        folded.replace(" ", "") for folded in _HEADER_ALIASES if len(folded) > 1
    }
    groups: list[list[int]] = []
    index = 0
    while index < width:
        matched: list[int] | None = None
        for size in (3, 2):
            if index + size > width:
                continue
            if not all(_fold(padded[0][part]) for part in range(index, index + size)):
                continue
            fragment = "".join(
                _fold(padded[0][part]).replace(" ", "")
                for part in range(index, index + size)
            )
            if fragment in compact_aliases:
                matched = list(range(index, index + size))
                break
        if matched is None:
            matched = [index]
        groups.append(matched)
        index = matched[-1] + 1
    merged: list[list[str]] = []
    for row_number, row in enumerate(padded):
        separator = "" if row_number == 0 else " "
        merged.append(
            [
                _clean_cell(separator.join(row[index] for index in group))
                for group in groups
            ]
        )
    return merged


def _classify_table(headers: set[str]) -> str | None:
    if {"event_date", "event_type", "event_cancelled"} <= headers:
        return "inframonthly_events"
    if {"request_date", "requested_period", "request_type"} <= headers:
        return "information_requests"
    if {"guarantor", "guarantee_value", "guaranteed_amount"} <= headers:
        return "guarantors"
    if {"guaranteed_party", "guarantee_value", "guaranteed_amount"} <= headers:
        return "guarantees_received"
    if {"ceded_debtor", "nominal_value"} <= headers:
        return "ceded_debtors"
    if "summary_category" in headers and headers & {
        "granted",
        "operational_granted",
        "used",
        "guarantee_value",
        "guaranteed_amount",
        "intrinsic_value",
    }:
        return "summary_totals"
    if "category" in headers and headers & {"amount", "intrinsic_value"}:
        return "other_risk_information"
    if {"category", "used"} <= headers:
        return "exposures"
    return None


def _reference_month(page_text: str) -> str:
    match = re.search(
        r"(?:DATA\s+DI\s+RIFERIMENTO|DATA\s+CONTABILE)\s*:\s*([A-Za-zÀ-ÿ]+)\s+(\d{4})",
        page_text,
        flags=re.IGNORECASE,
    )
    if match:
        month_number = _ITALIAN_MONTHS.get(_fold(match.group(1)))
        return f"{match.group(2)}-{month_number:02d}" if month_number else ""
    numeric = re.search(
        r"(?:DATA\s+DI\s+RIFERIMENTO|DATA\s+CONTABILE|ULTIMA\s+DATA\s+CONTABILE)\s*:\s*(\d{2})/(\d{2})/(\d{4})",
        page_text,
        flags=re.IGNORECASE,
    )
    return f"{numeric.group(3)}-{numeric.group(2)}" if numeric else ""


def _promote_summary_category_header(headers: Sequence[str]) -> list[str]:
    """Name one blank leading summary dimension only for an exact amount shape."""

    promoted = list(headers)
    if not promoted or promoted[0]:
        return promoted
    amount_headers = {
        "granted",
        "operational_granted",
        "used",
        "guarantee_value",
        "guaranteed_amount",
        "intrinsic_value",
    }
    remaining = {header for header in promoted[1:] if header}
    if remaining and remaining <= amount_headers:
        promoted[0] = "summary_category"
    return promoted


def _recover_terminal_valid_to(
    page: Any, table: Any, extracted_rows: Sequence[Sequence[Any]]
) -> list[list[Any]]:
    """Recover one clipped terminal ``A`` validity column from exact geometry.

    The recovery is intentionally narrow: the extracted table must end in the
    exact ``Da`` header, an exact ``A`` header must sit immediately to its
    right, and each recovered body value must be one exact Italian date within
    the corresponding table-row band. Ambiguous geometry is left unresolved.
    """

    rows = [list(row) for row in extracted_rows]
    if len(rows) < 2 or not rows[0]:
        return rows
    headers = [_canonical_header(value) for value in rows[0]]
    if "valid_from" not in headers or "valid_to" in headers:
        return rows
    right = float(table.bbox[2])
    if float(page.width) - right > 100:
        return rows
    table_rows = getattr(table, "rows", ())
    if len(table_rows) != len(rows):
        return rows
    words = page.extract_words() or []

    def row_band(row: Any) -> tuple[float, float] | None:
        cells = [cell for cell in row.cells if cell is not None]
        if not cells:
            return None
        return min(float(cell[1]) for cell in cells), max(
            float(cell[3]) for cell in cells
        )

    header_band = row_band(table_rows[0])
    if header_band is None:
        return rows
    header_matches = [
        word
        for word in words
        if float(word["x0"]) >= right - 1
        and header_band[0] - 1 <= float(word["top"])
        and float(word["bottom"]) <= header_band[1] + 1
        and _fold(word["text"]) == "a"
    ]
    if len(header_matches) != 1:
        return rows
    recovered: list[str] = []
    for table_row in table_rows[1:]:
        band = row_band(table_row)
        if band is None:
            return rows
        dates = [
            _clean_cell(word["text"])
            for word in words
            if float(word["x0"]) >= right - 1
            and band[0] - 1 <= float(word["top"])
            and float(word["bottom"]) <= band[1] + 1
            and re.fullmatch(r"\d{2}/\d{2}/\d{4}", _clean_cell(word["text"]))
        ]
        if len(dates) > 1:
            return rows
        recovered.append(dates[0] if dates else "")
    if not any(recovered):
        return rows
    rows[0].append("A")
    for row, value in zip(rows[1:], recovered):
        row.append(value)
    return rows


def _has_repeated_header_row(rows: Sequence[Sequence[str]]) -> bool:
    """Detect a merged multi-grid table from an exact repeated body header."""

    if len(rows) < 3:
        return False
    header_fields = {_canonical_header(value) for value in rows[0] if _fold(value)}
    for row in rows[1:]:
        repeated = {
            _canonical_header(value)
            for value in row
            if _canonical_header(value) in header_fields
        }
        if len(repeated) >= 3 and ({"category", "used"} & repeated):
            return True
    return False


def _unclassified_table(
    *,
    page_number: int,
    bbox: Sequence[float],
    headers: Sequence[str],
    rows: Sequence[Sequence[str]],
    reason: str,
) -> dict[str, Any]:
    """Return mechanical triage evidence without deciding semantic relevance."""

    analytical_headers = set().union(*SHEET_CONTRACTS.values())
    recognized_headers = sorted(set(headers) & analytical_headers)
    populated_body_rows = sum(
        sum(bool(_clean_cell(value)) for value in row) >= 2 for row in rows[1:]
    )
    data_candidate = len(recognized_headers) >= 2 and populated_body_rows > 0
    return {
        "source_page": page_number,
        "source_region": _region(bbox),
        "headers": list(headers),
        "row_count": max(len(rows) - 1, 0),
        "recognized_analytical_headers": recognized_headers,
        "populated_body_row_count": populated_body_rows,
        "review_priority": (
            "unsupported_data_candidate" if data_candidate else "layout_or_narrative"
        ),
        "reason": reason,
    }


def _page_lines(page: Any) -> list[tuple[float, str]]:
    words = sorted(page.extract_words() or [], key=lambda word: word["top"])
    lines: list[tuple[float, list[Mapping[str, Any]]]] = []
    for word in words:
        top = float(word["top"])
        if lines and abs(lines[-1][0] - top) <= 2.0:
            lines[-1][1].append(word)
        else:
            lines.append((top, [word]))
    return [
        (
            top,
            _line_text(line_words),
        )
        for top, line_words in lines
    ]


def _line_text(words: Sequence[Mapping[str, Any]]) -> str:
    ordered = sorted(words, key=lambda word: word["x0"])
    parts: list[str] = []
    previous_x1: float | None = None
    for word in ordered:
        if previous_x1 is not None and float(word["x0"]) - previous_x1 > 30:
            parts.append("|")
        parts.append(str(word["text"]))
        previous_x1 = float(word["x1"])
    return " ".join(parts)


def _nearest_intermediary(lines: Sequence[tuple[float, str]], table_top: float) -> str:
    matches: list[tuple[float, str]] = []
    for top, text in lines:
        if top >= table_top:
            continue
        match = re.search(
            r"\bIntermediario\s*:\s*(.+?)(?:\s+\|\s+|$)",
            text,
            re.IGNORECASE,
        )
        if match:
            matches.append((top, _clean_cell(match.group(1)).lstrip("| ")))
    return matches[-1][1] if matches else ""


def _nearest_standalone_intermediary(
    lines: Sequence[tuple[float, str]], table_top: float
) -> str:
    """Read the nearest exact uppercase standalone label used by request tables."""

    candidates: list[tuple[float, str]] = []
    for top, text in lines:
        cleaned = _clean_cell(text).strip("| ")
        if top >= table_top or table_top - top > 90:
            continue
        if not cleaned or _fold(cleaned) == "intermediario":
            continue
        if ":" in cleaned or "|" in cleaned or len(cleaned) > 80:
            continue
        if cleaned != cleaned.upper() or not any(
            character.isalpha() for character in cleaned
        ):
            continue
        candidates.append((top, cleaned))
    return candidates[-1][1] if candidates else ""


def _amount(value: str) -> str:
    text = _clean_cell(value).replace(" ", "")
    if not text:
        return "0"
    watermark_match = re.fullmatch(r"(?:[A-Za-z])+([+-]?\d[\d.,]*)", text)
    if watermark_match:
        text = watermark_match.group(1)
    if not re.fullmatch(r"[+-]?\d[\d.,]*", text):
        raise InvalidOperation
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        groups = text.split(",")
        text = (
            "".join(groups)
            if all(len(group) == 3 for group in groups[1:])
            else text.replace(",", ".")
        )
    elif "." in text:
        groups = text.split(".")
        if all(len(group) == 3 for group in groups[1:]):
            text = "".join(groups)
    number = Decimal(text)
    rendered = format(number, "f")
    return rendered.rstrip("0").rstrip(".") if "." in rendered else rendered


def _row_map(headers: Sequence[str], row: Sequence[str]) -> dict[str, str]:
    result: dict[str, str] = {}
    for header, value in zip(headers, row):
        if header and header not in result:
            result[header] = _clean_cell(value)
    return result


def _region(bbox: Sequence[float]) -> str:
    return ",".join(f"{float(value):.2f}" for value in bbox)


def _base_provenance(
    *,
    page_number: int,
    table_number: int,
    row_number: int,
    bbox: Sequence[float],
    source_hash: str,
) -> dict[str, str | int]:
    return {
        "source_page": page_number,
        "source_region": _region(bbox),
        "source_row_locator": f"page:{page_number}:table:{table_number}:row:{row_number}",
        "source_document_sha256": source_hash,
    }


def _exposure_row(
    source: Mapping[str, str],
    *,
    reference_month: str,
    intermediary: str,
    provenance: Mapping[str, str | int],
) -> tuple[dict[str, Any], list[str]]:
    issues: list[str] = []
    output: dict[str, Any] = {
        "reference_month": reference_month,
        "intermediary": intermediary,
        "category": source.get("category", ""),
        "location": source.get("location", ""),
        "original_duration": source.get("original_duration", ""),
        "residual_duration": source.get("residual_duration", ""),
        "currency": source.get("currency", ""),
        "import_export": source.get("import_export", ""),
        "activity_type": source.get("activity_type", ""),
        "relationship_status": source.get("relationship_status", ""),
        "guarantee_type": source.get("guarantee_type", ""),
        "borrower_role": source.get("borrower_role", ""),
        "record_status": (
            "previous" if "valid_from" in source or "valid_to" in source else "current"
        ),
        "valid_from": source.get("valid_from", ""),
        "valid_to": source.get("valid_to", ""),
        **provenance,
    }
    for field in (
        "granted",
        "operational_granted",
        "used",
        "average_balance",
        "guaranteed_amount",
    ):
        raw_value = source.get(field, "")
        try:
            output[field] = _amount(raw_value)
        except InvalidOperation:
            output[field] = raw_value
            if not (
                output["record_status"] == "previous"
                and _fold(raw_value) == "assenza di segnalazione"
            ):
                issues.append(f"invalid_amount:{field}")
    if not reference_month:
        issues.append("missing_reference_month")
    if not intermediary:
        issues.append("missing_intermediary")
    if output["record_status"] == "previous" and not (
        output["valid_from"] and output["valid_to"]
    ):
        issues.append("incomplete_previous_validity")
    output["extraction_confidence"] = "high" if not issues else "review_required"
    return output, issues


def _guarantee_row(
    source: Mapping[str, str],
    *,
    reference_month: str,
    intermediary: str,
    provenance: Mapping[str, str | int],
) -> tuple[dict[str, Any], list[str]]:
    issues: list[str] = []
    output: dict[str, Any] = {
        "reference_month": reference_month,
        "intermediary": intermediary,
        "category": source.get("category", ""),
        "location": source.get("location", ""),
        "guaranteed_party": source.get("guaranteed_party", ""),
        "relationship_status": source.get("relationship_status", ""),
        "guarantee_type": source.get("guarantee_type", ""),
        "record_status": (
            "previous" if "valid_from" in source or "valid_to" in source else "current"
        ),
        "valid_from": source.get("valid_from", ""),
        "valid_to": source.get("valid_to", ""),
        **provenance,
    }
    for field in ("guarantee_value", "guaranteed_amount"):
        raw_value = source.get(field, "")
        try:
            output[field] = _amount(raw_value)
        except InvalidOperation:
            output[field] = raw_value
            if not (
                output["record_status"] == "previous"
                and _fold(raw_value) == "assenza di segnalazione"
            ):
                issues.append(f"invalid_amount:{field}")
    if not reference_month:
        issues.append("missing_reference_month")
    if not intermediary:
        issues.append("missing_intermediary")
    output["extraction_confidence"] = "high" if not issues else "review_required"
    return output, issues


def _guarantor_row(
    source: Mapping[str, str],
    *,
    reference_month: str,
    intermediary: str,
    provenance: Mapping[str, str | int],
) -> tuple[dict[str, Any], list[str]]:
    issues: list[str] = []
    output: dict[str, Any] = {
        "reference_month": reference_month,
        "intermediary": intermediary,
        "guarantor": source.get("guarantor", ""),
        "record_status": (
            "previous" if "valid_from" in source or "valid_to" in source else "current"
        ),
        "valid_from": source.get("valid_from", ""),
        "valid_to": source.get("valid_to", ""),
        **provenance,
    }
    for field in ("guarantee_value", "guaranteed_amount"):
        raw_value = source.get(field, "")
        try:
            output[field] = _amount(raw_value)
        except InvalidOperation:
            output[field] = raw_value
            if not (
                output["record_status"] == "previous"
                and _fold(raw_value) == "assenza di segnalazione"
            ):
                issues.append(f"invalid_amount:{field}")
    if not reference_month:
        issues.append("missing_reference_month")
    if not intermediary:
        issues.append("missing_intermediary")
    if output["record_status"] == "previous" and not (
        output["valid_from"] and output["valid_to"]
    ):
        issues.append("incomplete_previous_validity")
    output["extraction_confidence"] = "high" if not issues else "review_required"
    return output, issues


def _ceded_debtor_row(
    source: Mapping[str, str],
    *,
    reference_month: str,
    intermediary: str,
    provenance: Mapping[str, str | int],
) -> tuple[dict[str, Any], list[str]]:
    issues: list[str] = []
    output: dict[str, Any] = {
        "reference_month": reference_month,
        "intermediary": intermediary,
        "ceded_debtor": source.get("ceded_debtor", ""),
        "record_status": (
            "previous" if "valid_from" in source or "valid_to" in source else "current"
        ),
        "valid_from": source.get("valid_from", ""),
        "valid_to": source.get("valid_to", ""),
        **provenance,
    }
    try:
        output["nominal_value"] = _amount(source.get("nominal_value", ""))
    except InvalidOperation:
        output["nominal_value"] = source.get("nominal_value", "")
        if not (
            output["record_status"] == "previous"
            and _fold(source.get("nominal_value", "")) == "assenza di segnalazione"
        ):
            issues.append("invalid_amount:nominal_value")
    if not reference_month:
        issues.append("missing_reference_month")
    if not intermediary:
        issues.append("missing_intermediary")
    if output["record_status"] == "previous" and not (
        output["valid_from"] and output["valid_to"]
    ):
        issues.append("incomplete_previous_validity")
    output["extraction_confidence"] = "high" if not issues else "review_required"
    return output, issues


def _other_risk_row(
    source: Mapping[str, str],
    *,
    reference_month: str,
    intermediary: str,
    provenance: Mapping[str, str | int],
) -> tuple[dict[str, Any], list[str]]:
    issues: list[str] = []
    output: dict[str, Any] = {
        "reference_month": reference_month,
        "intermediary": intermediary,
        "category": source.get("category", ""),
        "location": source.get("location", ""),
        "original_duration": source.get("original_duration", ""),
        "residual_duration": source.get("residual_duration", ""),
        "currency": source.get("currency", ""),
        "import_export": source.get("import_export", ""),
        "activity_type": source.get("activity_type", ""),
        "relationship_status": source.get("relationship_status", ""),
        "record_status": (
            "previous" if "valid_from" in source or "valid_to" in source else "current"
        ),
        "valid_from": source.get("valid_from", ""),
        "valid_to": source.get("valid_to", ""),
        **provenance,
    }
    for field in ("amount", "intrinsic_value"):
        if field not in source:
            output[field] = ""
            continue
        raw_value = source[field]
        try:
            output[field] = _amount(raw_value)
        except InvalidOperation:
            output[field] = raw_value
            issues.append(f"invalid_amount:{field}")
    if not reference_month:
        issues.append("missing_reference_month")
    if not intermediary:
        issues.append("missing_intermediary")
    output["extraction_confidence"] = "high" if not issues else "review_required"
    return output, issues


def _summary_total_row(
    source: Mapping[str, str],
    *,
    reference_month: str,
    intermediary: str,
    provenance: Mapping[str, str | int],
) -> tuple[dict[str, Any], list[str]]:
    issues: list[str] = []
    output: dict[str, Any] = {
        "reference_month": reference_month,
        "intermediary": intermediary,
        "summary_category": source.get("summary_category", ""),
        **provenance,
    }
    for field in (
        "granted",
        "operational_granted",
        "used",
        "guarantee_value",
        "guaranteed_amount",
        "intrinsic_value",
    ):
        if field not in source:
            output[field] = ""
            continue
        try:
            output[field] = _amount(source[field])
        except InvalidOperation:
            output[field] = source[field]
            issues.append(f"invalid_amount:{field}")
    if not reference_month:
        issues.append("missing_reference_month")
    if not intermediary:
        issues.append("missing_intermediary")
    output["extraction_confidence"] = "high" if not issues else "review_required"
    return output, issues


def normalize_pdf(
    path: Path,
    *,
    page_numbers: Sequence[int] | None = None,
    allow_no_supported_tables: bool = False,
) -> dict[str, Any]:
    """Extract native-text tables without assigning professional classifications."""

    if not path.is_file() or path.suffix.casefold() != ".pdf":
        raise ValueError(f"Expected one readable PDF: {path}")
    source_hash = _sha256_file(path)
    collections: dict[str, list[dict[str, Any]]] = {
        "exposures": [],
        "guarantees_received": [],
        "guarantors": [],
        "ceded_debtors": [],
        "other_risk_information": [],
        "summary_totals": [],
        "inframonthly_events": [],
        "information_requests": [],
    }
    issues: list[dict[str, Any]] = []
    unclassified: list[dict[str, Any]] = []
    native_character_count = 0
    page_count = 0
    with pdfplumber.open(path) as document:
        page_count = len(document.pages)
        if page_numbers is None:
            selected_page_numbers = tuple(range(1, page_count + 1))
        else:
            selected_page_numbers = tuple(sorted(set(page_numbers)))
            if not selected_page_numbers:
                raise ValueError("At least one PDF page must be selected.")
            invalid_pages = [
                number
                for number in selected_page_numbers
                if number < 1 or number > page_count
            ]
            if invalid_pages:
                raise ValueError(
                    "PDF page selection is outside the document: "
                    + ", ".join(map(str, invalid_pages))
                )
        active_reference_month = ""
        for page_number in selected_page_numbers:
            page = document.pages[page_number - 1]
            page_text = page.extract_text() or ""
            native_character_count += len(page_text.strip())
            page_reference_month = _reference_month(page_text)
            if page_reference_month:
                active_reference_month = page_reference_month
            reference_month = page_reference_month or active_reference_month
            lines = _page_lines(page)
            for table_number, table in enumerate(page.find_tables(), start=1):
                extracted_rows = _recover_terminal_valid_to(
                    page, table, table.extract()
                )
                rows = _merge_fragmented_header_columns(
                    _collapse_blank_header_columns(extracted_rows)
                )
                if not rows:
                    continue
                headers = _promote_summary_category_header(
                    [_canonical_header(value) for value in rows[0]]
                )
                if _has_repeated_header_row(rows):
                    unclassified.append(
                        _unclassified_table(
                            page_number=page.page_number,
                            bbox=table.bbox,
                            headers=headers,
                            rows=rows,
                            reason="repeated_header_in_body",
                        )
                    )
                    continue
                family = _classify_table(set(headers))
                if family is None:
                    unclassified.append(
                        _unclassified_table(
                            page_number=page.page_number,
                            bbox=table.bbox,
                            headers=headers,
                            rows=rows,
                            reason="unsupported_header_contract",
                        )
                    )
                    continue
                intermediary = _nearest_intermediary(lines, float(table.bbox[1]))
                if family == "information_requests" and not intermediary:
                    intermediary = _nearest_standalone_intermediary(
                        lines, float(table.bbox[1])
                    )
                for row_number, raw_row in enumerate(rows[1:], start=2):
                    source = _row_map(headers, raw_row)
                    if not any(source.values()):
                        continue
                    provenance = _base_provenance(
                        page_number=page.page_number,
                        table_number=table_number,
                        row_number=row_number,
                        bbox=table.bbox,
                        source_hash=source_hash,
                    )
                    row_issues: list[str] = []
                    if family == "exposures":
                        output, row_issues = _exposure_row(
                            source,
                            reference_month=reference_month,
                            intermediary=intermediary,
                            provenance=provenance,
                        )
                    elif family == "guarantees_received":
                        output, row_issues = _guarantee_row(
                            source,
                            reference_month=reference_month,
                            intermediary=intermediary,
                            provenance=provenance,
                        )
                    elif family == "guarantors":
                        output, row_issues = _guarantor_row(
                            source,
                            reference_month=reference_month,
                            intermediary=intermediary,
                            provenance=provenance,
                        )
                    elif family == "ceded_debtors":
                        output, row_issues = _ceded_debtor_row(
                            source,
                            reference_month=reference_month,
                            intermediary=intermediary,
                            provenance=provenance,
                        )
                    elif family == "other_risk_information":
                        output, row_issues = _other_risk_row(
                            source,
                            reference_month=reference_month,
                            intermediary=intermediary,
                            provenance=provenance,
                        )
                    elif family == "summary_totals":
                        output, row_issues = _summary_total_row(
                            source,
                            reference_month=reference_month,
                            intermediary=intermediary,
                            provenance=provenance,
                        )
                    else:
                        output = {
                            **source,
                            "intermediary": intermediary,
                            **provenance,
                        }
                        output["extraction_confidence"] = (
                            "high" if intermediary else "review_required"
                        )
                        if not intermediary:
                            row_issues.append("missing_intermediary")
                    collections[family].append(output)
                    if row_issues:
                        issues.append(
                            {
                                "source_row_locator": provenance["source_row_locator"],
                                "issues": row_issues,
                            }
                        )
    if native_character_count == 0:
        raise ValueError(
            "The PDF has no readable native text. Use the official digital Centrale Rischi report downloaded from the Banca d'Italia service."
        )
    if not any(collections.values()) and not allow_no_supported_tables:
        raise ValueError(
            "No supported Centrale Rischi table layout was found in the native-text PDF."
        )
    return {
        "schema_version": PDF_NORMALIZATION_SCHEMA,
        "workflow_id": "centrale-rischi-review",
        "source": {
            "source_document_sha256": source_hash,
            "source_kind": "native_pdf_extraction",
            "page_count": page_count,
            "selected_pages": list(selected_page_numbers),
        },
        "review_status": "pending_professional_review",
        "semantic_roles_assigned": False,
        "tables": collections,
        "issues": issues,
        "unclassified_tables": unclassified,
        "implementation_reason": (
            "Header-shape recognition, cell extraction, Italian-number parsing, record provenance and current-versus-previous separation are deterministic because they are mechanically reviewable. Duration meaning, risk family, materiality and professional conclusions remain reviewed judgments."
        ),
    }


def evaluate_pdf_corpus(
    path: Path, *, cases: Mapping[str, Sequence[int]] | None = None
) -> dict[str, Any]:
    """Measure parser coverage without creating or combining client analyses."""

    if not path.is_file() or path.suffix.casefold() != ".pdf":
        raise ValueError(f"Expected one readable PDF: {path}")
    with pdfplumber.open(path) as document:
        page_count = len(document.pages)
    if not page_count:
        raise ValueError("The PDF has no pages.")
    selected_cases: Mapping[str, Sequence[int]] = cases or {
        f"page_{page_number:03d}": (page_number,)
        for page_number in range(1, page_count + 1)
    }
    if not selected_cases:
        raise ValueError("At least one corpus case is required.")

    case_results: list[dict[str, Any]] = []
    for case_id, page_numbers in selected_cases.items():
        normalized_case_id = _clean_cell(case_id)
        if not normalized_case_id:
            raise ValueError("Corpus case IDs must be non-empty.")
        selected_pages = tuple(sorted(set(page_numbers)))
        try:
            normalization = normalize_pdf(
                path,
                page_numbers=selected_pages,
                allow_no_supported_tables=True,
            )
        except ValueError as exc:
            case_results.append(
                {
                    "case_id": normalized_case_id,
                    "pages": list(selected_pages),
                    "outcome": "not_recognized",
                    "row_counts": {
                        "exposures": 0,
                        "guarantees_received": 0,
                        "guarantors": 0,
                        "ceded_debtors": 0,
                        "other_risk_information": 0,
                        "summary_totals": 0,
                        "inframonthly_events": 0,
                        "information_requests": 0,
                    },
                    "issue_count": 0,
                    "unclassified_table_count": 0,
                    "unsupported_data_candidate_count": 0,
                    "layout_or_narrative_count": 0,
                    "diagnostic": str(exc),
                }
            )
            continue
        recognized = any(normalization["tables"].values())
        case_results.append(
            {
                "case_id": normalized_case_id,
                "pages": list(selected_pages),
                "outcome": "recognized" if recognized else "not_recognized",
                "row_counts": {
                    key: len(rows) for key, rows in normalization["tables"].items()
                },
                "issue_count": len(normalization["issues"]),
                "unclassified_table_count": len(normalization["unclassified_tables"]),
                "unsupported_data_candidate_count": sum(
                    item["review_priority"] == "unsupported_data_candidate"
                    for item in normalization["unclassified_tables"]
                ),
                "layout_or_narrative_count": sum(
                    item["review_priority"] == "layout_or_narrative"
                    for item in normalization["unclassified_tables"]
                ),
                "diagnostic": (
                    ""
                    if recognized
                    else "No supported Centrale Rischi table layout was found on the selected pages."
                ),
            }
        )
    return {
        "schema_version": PDF_CORPUS_EVALUATION_SCHEMA,
        "workflow_id": "centrale-rischi-review",
        "purpose": "parser_coverage_only",
        "analysis_generated": False,
        "source": {
            "source_document_sha256": _sha256_file(path),
            "page_count": page_count,
        },
        "case_count": len(case_results),
        "recognized_case_count": sum(
            item["outcome"] == "recognized" for item in case_results
        ),
        "cases": case_results,
        "implementation_reason": (
            "Explicit page selection, exact header contracts, row counts and mechanical unsupported-data triage are deterministic and reviewable. Case meaning, materiality and client-level interpretation are deliberately not inferred or combined."
        ),
    }


def write_normalized_workbook(path: Path, payload: Mapping[str, Any]) -> None:
    """Write separate review sheets from a PDF-normalization payload."""

    if payload.get("schema_version") != PDF_NORMALIZATION_SCHEMA:
        raise ValueError("Unsupported PDF-normalization payload.")
    tables = payload.get("tables")
    if not isinstance(tables, Mapping):
        raise ValueError("PDF-normalization tables are missing.")
    table_keys = {
        "Esposizioni": "exposures",
        "Garanzie ricevute": "guarantees_received",
        "Garanti intestatario": "guarantors",
        "Debitori ceduti": "ceded_debtors",
        "Altre informazioni": "other_risk_information",
        "Prospetto sintetico": "summary_totals",
        "Eventi inframensili": "inframonthly_events",
        "Richieste informazioni": "information_requests",
    }
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, contract in SHEET_CONTRACTS.items():
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(contract)
        rows = tables.get(table_keys[sheet_name], [])
        if not isinstance(rows, list):
            raise ValueError(f"Invalid normalized table: {table_keys[sheet_name]}")
        for row in rows:
            if not isinstance(row, Mapping):
                raise ValueError(
                    f"Invalid row in normalized table: {table_keys[sheet_name]}"
                )
            sheet.append([row.get(header, "") for header in contract])
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="002060")
        for column_cells in sheet.columns:
            width = min(
                42,
                max(12, max(len(str(cell.value or "")) for cell in column_cells) + 2),
            )
            sheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)
