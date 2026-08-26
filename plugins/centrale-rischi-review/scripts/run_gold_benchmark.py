#!/usr/bin/env python3
"""Run the reviewed Centrale Rischi gold corpus end to end."""

from __future__ import annotations

import argparse
import copy
import html
import json
import logging
import re
import sys
from pathlib import Path
from typing import Any, Mapping, Sequence

import pdfplumber
from openpyxl import load_workbook

SCRIPTS_DIR = Path(__file__).resolve().parent
PLUGIN_ROOT = SCRIPTS_DIR.parent
sys.path.insert(0, str(SCRIPTS_DIR))

from centrale_rischi_core import (  # noqa: E402
    CentraleRischiContractError,
    build_analysis,
    build_inspection,
    build_model_context,
    finalize_commentary,
    load_source_tables,
    render_html,
    render_markdown,
    sha256_file,
    write_excel,
    write_json,
)
from centrale_rischi_pdf import (  # noqa: E402
    normalize_pdf,
    write_normalized_workbook,
)

__all__ = ["main", "run_benchmark"]

LOGGER = logging.getLogger(__name__)
BENCHMARK_SCHEMA = "vera.centrale_rischi_gold_benchmark.v1"
SEMANTIC_REVIEW_SCHEMA = "vera.centrale_rischi_semantic_review.v2"
TABLE_KEYS = (
    "exposures",
    "guarantees_received",
    "guarantors",
    "ceded_debtors",
    "other_risk_information",
    "summary_totals",
    "inframonthly_events",
    "information_requests",
)
STABLE_ANALYSIS_KEYS = (
    "metrics",
    "original_term_summary",
    "residual_term_summary",
    "risk_category_summary",
    "category_movement_summary",
    "monthly_series",
)


def _source_argument(value: str) -> tuple[str, Path]:
    source_id, separator, path_text = value.partition("=")
    if not separator or not source_id.strip() or not path_text.strip():
        raise argparse.ArgumentTypeError("Expected SOURCE_ID=/absolute/report.pdf.")
    return source_id.strip(), Path(path_text).expanduser().resolve()


def _load_object(path: Path) -> dict[str, Any]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError(f"Expected one JSON object: {path}")
    return payload


def _row_counts(normalization: Mapping[str, Any]) -> dict[str, int]:
    tables = normalization.get("tables")
    if not isinstance(tables, Mapping):
        raise ValueError("Normalization tables are missing.")
    return {key: len(tables.get(key, [])) for key in TABLE_KEYS}


def _partial_row_match(actual: Mapping[str, Any], expected: Mapping[str, Any]) -> bool:
    return all(actual.get(key) == value for key, value in expected.items())


def _check_expected_rows(
    actual_rows: Sequence[Mapping[str, Any]],
    expected_rows: Sequence[Mapping[str, Any]],
) -> tuple[bool, list[dict[str, Any]]]:
    unused = list(actual_rows)
    missing: list[dict[str, Any]] = []
    for expected in expected_rows:
        match_index = next(
            (
                index
                for index, actual in enumerate(unused)
                if _partial_row_match(actual, expected)
            ),
            None,
        )
        if match_index is None:
            missing.append(dict(expected))
        else:
            unused.pop(match_index)
    return not missing, missing


def _metric_values(analysis: Mapping[str, Any]) -> dict[str, Any]:
    return {
        str(item["metric_id"]): item.get("value")
        for item in analysis.get("metrics", [])
    }


def _analysis_projection(
    analysis: Mapping[str, Any], *, omit_previous_count: bool = False
) -> dict[str, Any]:
    projection = {key: analysis[key] for key in STABLE_ANALYSIS_KEYS}
    if omit_previous_count:
        projection["metrics"] = [
            item
            for item in projection["metrics"]
            if item["metric_id"] != "cr.previous_record_count"
        ]
    return projection


def _reviewed_mapping(
    observed: set[str], profile: Mapping[str, Any], field: str
) -> dict[str, str]:
    mappings = profile.get(field)
    if not isinstance(mappings, Mapping):
        raise ValueError(f"Mapping profile is missing {field}.")
    missing = observed - set(map(str, mappings))
    if missing:
        raise ValueError(f"Gold mapping lacks {field} values: {sorted(missing)}")
    return {value: str(mappings[value]) for value in sorted(observed)}


def _build_case_analysis(
    normalization: Mapping[str, Any],
    case: Mapping[str, Any],
    profile: Mapping[str, Any],
    output_dir: Path,
) -> tuple[dict[str, Any], dict[str, Any], Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = output_dir / "centrale_rischi_normalized.xlsx"
    write_normalized_workbook(workbook_path, normalization)
    tables = load_source_tables([workbook_path])
    inspection, control, _ = build_inspection(tables)
    exposure_table = next(
        table for table in tables if table.table_label == "Esposizioni"
    )
    original_values = {str(row["original_duration"]) for row in exposure_table.rows}
    residual_values = {str(row["residual_duration"]) for row in exposure_table.rows}
    category_values = {str(row["category"]) for row in exposure_table.rows}
    source = normalization["source"]
    recipe = {
        "schema_version": "vera.centrale_rischi_recipe.v2",
        "workflow_id": "centrale-rischi-review",
        "inventory_sha256": inspection["inventory_sha256"],
        "entity": str(case["entity"]),
        "currency": "EUR",
        "analysis_mode": str(case.get("analysis_mode", "descriptive")),
        "analysis_objective": str(case["analysis_objective"]),
        "audience": "professional",
        "source_kind": "native_pdf_extraction",
        "source_document_sha256": str(source["source_document_sha256"]),
        "table_id": exposure_table.table_id,
        "columns": {
            "reference_month": "reference_month",
            "intermediary": "intermediary",
            "risk_category": "category",
            "original_duration": "original_duration",
            "residual_duration": "residual_duration",
            "granted": "granted",
            "operational_granted": "operational_granted",
            "used": "used",
            "guarantee_type": "guarantee_type",
            "guaranteed_amount": "guaranteed_amount",
            "prejudicial_event": "",
            "reporting_type": "",
            "relationship_status": "relationship_status",
            "record_status": "record_status",
            "valid_from": "valid_from",
            "valid_to": "valid_to",
            "source_page": "source_page",
            "source_region": "source_region",
            "source_row_locator": "source_row_locator",
            "extraction_confidence": "extraction_confidence",
        },
        "value_mappings": {
            "original_term": _reviewed_mapping(
                original_values, profile, "original_term"
            ),
            "residual_term": _reviewed_mapping(
                residual_values, profile, "residual_term"
            ),
            "exposure_family": _reviewed_mapping(
                category_values, profile, "exposure_family"
            ),
        },
        "control_totals": dict(case.get("control_totals", {})),
        "control_tolerance": "0.01",
        "mapping_review": {
            "status": "reviewed",
            "reviewer": "gold-corpus-reviewed-mapping",
            "reviewed_at": "2026-08-26T12:00:00+02:00",
        },
    }
    analysis = build_analysis(tables, recipe)
    write_json(output_dir / "inspection.json", inspection)
    write_json(output_dir / "inspection_control.json", control)
    write_json(output_dir / "reviewed_recipe.json", recipe)
    return analysis, recipe, workbook_path


def _artifact_checks(analysis: Mapping[str, Any], output_dir: Path) -> dict[str, Any]:
    model_context = build_model_context(analysis)
    markdown_path = output_dir / "centrale_rischi_facts.md"
    html_path = output_dir / "centrale_rischi_dashboard.html"
    xlsx_path = output_dir / "centrale_rischi_analysis.xlsx"
    analysis_path = output_dir / "centrale_rischi_analysis.json"
    context_path = output_dir / "model_context.json"
    write_json(analysis_path, analysis)
    write_json(context_path, model_context)
    markdown_path.write_text(render_markdown(analysis), encoding="utf-8")
    html_path.write_text(render_html(analysis), encoding="utf-8")
    write_excel(xlsx_path, analysis)

    rendered = html_path.read_text(encoding="utf-8")
    workbook = load_workbook(xlsx_path, data_only=False)
    try:
        formulas = sum(
            isinstance(cell.value, str) and cell.value.startswith("=")
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
        )
        sheet_names = workbook.sheetnames
        first_metric_numeric = workbook["KPI"]["C2"].data_type == "n"
    finally:
        workbook.close()
    context_text = json.dumps(model_context, ensure_ascii=False, sort_keys=True)
    expected_sheets = {
        "KPI",
        "Durata originaria",
        "Durata residua",
        "Categorie",
        "Variazione categorie",
        "Esposizioni",
        "Garanzie",
        "Garanzie ricevute",
        "Garanti intestatario",
        "Debitori ceduti",
        "Altre informazioni",
        "Prospetto sintetico",
        "Sconfinamenti",
        "Eventi inframensili",
        "Richieste informazioni",
        "Pregiudizievoli",
        "Serie mensile",
        "Controlli",
    }
    checks = {
        "html_self_contained": (
            rendered.startswith("<!doctype html>")
            and "<script" not in rendered.casefold()
            and "http://" not in rendered
            and "https://" not in rendered
        ),
        "html_required_sections": all(
            heading in rendered
            for heading in (
                "KPI",
                "Esposizioni per durata originaria",
                "Indicatori per categoria",
                "Limiti",
            )
        ),
        "xlsx_expected_sheets": set(sheet_names) == expected_sheets,
        "xlsx_no_formulas": formulas == 0,
        "xlsx_numeric_metric": first_metric_numeric,
        "context_no_absolute_paths": not re.search(
            r"(?:/Users/|/private/|[A-Za-z]:\\\\)", context_text
        ),
        "context_no_source_hashes": "source_document_sha256" not in context_text,
        "context_bounded": (
            len(model_context["monthly_series"]) <= 36
            and len(model_context["category_movement_summary"]) <= 50
            and len(model_context["previous_records"]) <= 20
            and len(model_context["top_overruns"]) <= 20
            and len(model_context["top_guarantees"]) <= 20
        ),
    }
    return {
        "passed": all(checks.values()),
        "checks": checks,
        "artifacts": {
            path.name: sha256_file(path)
            for path in (
                analysis_path,
                context_path,
                markdown_path,
                html_path,
                xlsx_path,
            )
        },
    }


def _metamorphic_checks(
    normalization: Mapping[str, Any],
    case: Mapping[str, Any],
    profile: Mapping[str, Any],
    base_analysis: Mapping[str, Any],
    output_dir: Path,
) -> dict[str, Any]:
    results: dict[str, Any] = {}

    reordered = copy.deepcopy(normalization)
    reordered["tables"]["exposures"] = list(reversed(reordered["tables"]["exposures"]))
    reordered_analysis, _, _ = _build_case_analysis(
        reordered, case, profile, output_dir / "row-order"
    )
    results["row_order_invariance"] = _analysis_projection(
        reordered_analysis
    ) == _analysis_projection(base_analysis)

    current_only = copy.deepcopy(normalization)
    previous_count = sum(
        row.get("record_status") == "previous"
        for row in current_only["tables"]["exposures"]
    )
    if previous_count:
        current_only["tables"]["exposures"] = [
            row
            for row in current_only["tables"]["exposures"]
            if row.get("record_status") != "previous"
        ]
        current_analysis, _, _ = _build_case_analysis(
            current_only, case, profile, output_dir / "without-previous"
        )
        results["previous_rows_do_not_change_current_metrics"] = _analysis_projection(
            current_analysis, omit_previous_count=True
        ) == _analysis_projection(base_analysis, omit_previous_count=True)
    else:
        results["previous_rows_do_not_change_current_metrics"] = "not_applicable"

    without_auxiliary = copy.deepcopy(normalization)
    auxiliary_count = 0
    for key in TABLE_KEYS:
        if key == "exposures":
            continue
        auxiliary_count += len(without_auxiliary["tables"][key])
        without_auxiliary["tables"][key] = []
    if auxiliary_count:
        auxiliary_analysis, _, _ = _build_case_analysis(
            without_auxiliary, case, profile, output_dir / "without-auxiliary"
        )
        results["auxiliary_populations_do_not_change_exposure_metrics"] = (
            _analysis_projection(auxiliary_analysis)
            == _analysis_projection(base_analysis)
        )
    else:
        results["auxiliary_populations_do_not_change_exposure_metrics"] = (
            "not_applicable"
        )
    return {
        "passed": all(
            value is True or value == "not_applicable" for value in results.values()
        ),
        "checks": results,
    }


def _check_context_facts(
    context: Mapping[str, Any], expected_facts: Sequence[Mapping[str, Any]]
) -> dict[str, Any]:
    def contains_partial(value: Any, expected: Mapping[str, Any]) -> bool:
        if isinstance(value, Mapping):
            if _partial_row_match(value, expected):
                return True
            return any(contains_partial(item, expected) for item in value.values())
        if isinstance(value, list):
            return any(contains_partial(item, expected) for item in value)
        return False

    missing = [
        dict(item) for item in expected_facts if not contains_partial(context, item)
    ]
    return {"passed": not missing, "missing": missing}


def _prepare_semantic_packet(
    manifest: Mapping[str, Any],
    output_dir: Path,
    analyses: Mapping[str, Mapping[str, Any]],
) -> dict[str, Any]:
    cases: list[dict[str, Any]] = []
    for contract in manifest.get("semantic_review_cases", []):
        case = dict(contract)
        case_id = str(case["case_id"])
        analysis = analyses.get(case_id)
        if analysis is None:
            case["commentary_status"] = "analysis_unavailable"
            cases.append(case)
            continue
        case_dir = output_dir / "analysis" / case_id
        analysis_path = case_dir / "centrale_rischi_analysis.json"
        context_path = case_dir / "model_context.json"
        commentary_path = case_dir / "centrale_rischi_commentary.json"
        case.update(
            {
                "analysis_sha256": sha256_file(analysis_path),
                "model_context_sha256": sha256_file(context_path),
            }
        )
        if not commentary_path.is_file():
            case["commentary_status"] = "missing"
            cases.append(case)
            continue
        commentary = finalize_commentary(analysis, _load_object(commentary_path))
        report_path = case_dir / "centrale_rischi_report.md"
        dashboard_path = case_dir / "centrale_rischi_dashboard_reviewed.html"
        report_path.write_text(render_markdown(analysis, commentary), encoding="utf-8")
        dashboard_path.write_text(render_html(analysis, commentary), encoding="utf-8")
        commentary_sha256 = sha256_file(commentary_path)
        reviewed_output_hashes = {
            report_path.name: sha256_file(report_path),
            dashboard_path.name: sha256_file(dashboard_path),
        }
        write_json(
            case_dir / "commentary_receipt.json",
            {
                "schema_version": "vera.centrale_rischi_commentary_receipt.v2",
                "workflow_id": "centrale-rischi-review",
                "status": "draft_pending_professional_review",
                "analysis_sha256": case["analysis_sha256"],
                "model_context_sha256": case["model_context_sha256"],
                "commentary_sha256": commentary_sha256,
                "outputs": [
                    {"name": name, "sha256": value}
                    for name, value in reviewed_output_hashes.items()
                ],
                "validation_boundary": (
                    "Schema and evidence-reference closure were validated; "
                    "semantic quality and professional approval remain separate."
                ),
            },
        )
        case.update(
            {
                "commentary_status": "evidence_refs_validated",
                "commentary_sha256": commentary_sha256,
                "commentary": commentary,
                "reviewed_output_hashes": reviewed_output_hashes,
            }
        )
        cases.append(case)
    return {
        "schema_version": "vera.centrale_rischi_semantic_review_packet.v2",
        "workflow_id": "centrale-rischi-review",
        "rubric": manifest.get("semantic_rubric", {}),
        "cases": cases,
        "boundary": (
            "This packet requires model or professional semantic judgment. "
            "No keyword rule assigns commentary quality."
        ),
    }


def _semantic_review_summary(
    packet: Mapping[str, Any], semantic_review_path: Path | None
) -> dict[str, Any]:
    cases = packet.get("cases", [])
    if semantic_review_path is None:
        return {
            "status": (
                "commentary_ready_review_pending"
                if all(
                    item.get("commentary_status") == "evidence_refs_validated"
                    for item in cases
                )
                else "commentary_pending"
            ),
            "reviews": [],
        }
    review = _load_object(semantic_review_path)
    if review.get("schema_version") != SEMANTIC_REVIEW_SCHEMA:
        raise ValueError("Unsupported semantic-review schema.")
    reviews = review.get("reviews")
    if not isinstance(reviews, list):
        raise ValueError("Semantic reviews must be a list.")
    required_ids = {str(item["case_id"]) for item in cases}
    supplied_ids = {str(item.get("case_id")) for item in reviews}
    missing = sorted(required_ids - supplied_ids)
    if missing:
        raise ValueError(f"Semantic review is missing cases: {missing}")
    extra = sorted(supplied_ids - required_ids)
    if extra:
        raise ValueError(f"Semantic review has unknown cases: {extra}")
    cases_by_id = {str(item["case_id"]): item for item in cases}
    missing_commentaries = sorted(
        case_id
        for case_id, case in cases_by_id.items()
        if case.get("commentary_status") != "evidence_refs_validated"
    )
    if missing_commentaries:
        raise ValueError(
            f"Semantic review cases lack validated commentary: {missing_commentaries}"
        )
    valid_scores = set(packet.get("rubric", {}).get("dimensions", {}))
    invalid_reviews = []
    for item in reviews:
        case_id = str(item.get("case_id"))
        scores = item.get("scores")
        blocking_findings = item.get("blocking_findings")
        expected_verdict = (
            "pass"
            if isinstance(scores, Mapping)
            and scores
            and all(isinstance(value, int) and value >= 3 for value in scores.values())
            and blocking_findings == []
            else "fail"
        )
        if (
            not isinstance(scores, Mapping)
            or set(scores) != valid_scores
            or any(
                not isinstance(value, int) or not 0 <= value <= 4
                for value in scores.values()
            )
            or item.get("reviewer_type") not in {"model", "professional"}
            or not isinstance(item.get("summary"), str)
            or not item["summary"].strip()
            or not isinstance(blocking_findings, list)
            or any(
                not isinstance(value, str) or not value.strip()
                for value in blocking_findings
            )
            or item.get("commentary_sha256")
            != cases_by_id.get(case_id, {}).get("commentary_sha256")
            or item.get("verdict") != expected_verdict
        ):
            invalid_reviews.append(case_id)
    if invalid_reviews:
        raise ValueError(f"Invalid semantic reviews: {invalid_reviews}")
    reviewer_types = {str(item["reviewer_type"]) for item in reviews}
    return {
        "status": (
            "professionally_reviewed"
            if reviewer_types == {"professional"}
            else (
                "model_reviewed_not_professional"
                if reviewer_types == {"model"}
                else "mixed_review_not_fully_professional"
            )
        ),
        "reviews": reviews,
        "passed": all(item["verdict"] == "pass" for item in reviews),
    }


def _render_benchmark_html(receipt: Mapping[str, Any]) -> str:
    extraction_rows = "".join(
        "<tr>"
        f"<td>{html.escape(str(item['case_id']))}</td>"
        f"<td>{html.escape(str(item['source_id']))}</td>"
        f"<td>{'Pass' if item['passed'] else 'Fail'}</td>"
        f"<td>{html.escape(', '.join(map(str, item['pages'])))}</td>"
        "</tr>"
        for item in receipt["extraction_cases"]
    )
    analysis_rows = "".join(
        "<tr>"
        f"<td>{html.escape(str(item['case_id']))}</td>"
        f"<td>{html.escape(str(item['outcome']))}</td>"
        f"<td>{'Pass' if item['passed'] else 'Fail'}</td>"
        f"<td>{html.escape(str(item.get('analysis_status', '—')))}</td>"
        "</tr>"
        for item in receipt["analysis_cases"]
    )
    return f"""<!doctype html><html lang="it"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>Benchmark Centrale Rischi</title><style>:root{{--navy:#002060;--blue:#006b8f;--ink:#171816;--muted:#5c6470;--rule:#d9dadd}}*{{box-sizing:border-box}}body{{margin:0;font-family:'Instrument Sans',Arial,sans-serif;color:var(--ink)}}main{{max-width:1100px;margin:auto;padding:48px 24px 72px}}header{{border-top:5px solid var(--navy);border-bottom:1px solid var(--rule);padding:28px 0}}h1{{font-size:clamp(2rem,5vw,4rem);line-height:1;margin:.3rem 0}}h2{{margin-top:42px}}.eyebrow{{color:var(--blue);font-weight:700;text-transform:uppercase;letter-spacing:.08em}}.status{{display:inline-block;border:1px solid var(--navy);padding:6px 10px}}.table-wrap{{overflow:auto;border-top:1px solid var(--navy)}}table{{width:100%;border-collapse:collapse;min-width:620px}}th,td{{padding:11px 9px;text-align:left;border-bottom:1px solid var(--rule)}}th{{color:var(--navy);font-size:.82rem}}p{{max-width:75ch}}code{{color:var(--blue)}}</style></head><body><main><header><p class="eyebrow">Vera · Evidenza riproducibile</p><h1>Benchmark Centrale Rischi</h1><p><span class="status">{html.escape(str(receipt['overall_status']))}</span></p><p>Il verdetto deterministico e il giudizio semantico sono separati. Nessun caso didattico è stato combinato con un altro salvo le sequenze esplicitamente documentate come uno stesso soggetto o report.</p></header><h2>Sintesi</h2><p>Estrazione: {receipt['summary']['extraction_passed']}/{receipt['summary']['extraction_total']} · Analisi: {receipt['summary']['analysis_passed']}/{receipt['summary']['analysis_total']} · Review semantico: {html.escape(str(receipt['semantic_review']['status']))}</p><h2>Casi di estrazione</h2><div class="table-wrap"><table><thead><tr><th>Caso</th><th>Fonte</th><th>Esito</th><th>Pagine</th></tr></thead><tbody>{extraction_rows}</tbody></table></div><h2>Casi di analisi</h2><div class="table-wrap"><table><thead><tr><th>Caso</th><th>Contratto atteso</th><th>Esito</th><th>Stato analisi</th></tr></thead><tbody>{analysis_rows}</tbody></table></div><h2>Limiti</h2><ul>{''.join(f'<li>{html.escape(str(item))}</li>' for item in receipt['limitations'])}</ul></main></body></html>"""


def run_benchmark(
    manifest_path: Path,
    source_paths: Mapping[str, Path],
    output_dir: Path,
    *,
    semantic_review_path: Path | None = None,
) -> dict[str, Any]:
    """Run all supplied gold cases and write replayable receipts."""

    manifest = _load_object(manifest_path)
    if manifest.get("schema_version") != BENCHMARK_SCHEMA:
        raise ValueError("Unsupported gold-benchmark manifest schema.")
    output_dir.mkdir(parents=True, exist_ok=True)
    configured_sources = manifest.get("sources")
    if not isinstance(configured_sources, Mapping):
        raise ValueError("Gold sources are missing.")
    verified_sources: dict[str, Path] = {}
    source_receipts: list[dict[str, Any]] = []
    for source_id, source_contract in configured_sources.items():
        if source_id not in source_paths:
            raise ValueError(f"Missing benchmark source: {source_id}")
        path = source_paths[source_id]
        if not path.is_file():
            raise ValueError(f"Benchmark source not found: {path}")
        actual_hash = sha256_file(path)
        expected_hash = str(source_contract["sha256"])
        if actual_hash != expected_hash:
            raise ValueError(f"Source hash mismatch for {source_id}.")
        with pdfplumber.open(path) as document:
            actual_page_count = len(document.pages)
        expected_page_count = int(source_contract["page_count"])
        if actual_page_count != expected_page_count:
            raise ValueError(f"Source page-count mismatch for {source_id}.")
        verified_sources[str(source_id)] = path
        source_receipts.append(
            {
                "source_id": source_id,
                "sha256": actual_hash,
                "page_count": actual_page_count,
                "role": source_contract["role"],
            }
        )

    extraction_results: list[dict[str, Any]] = []
    normalization_cache: dict[str, dict[str, Any]] = {}
    for case in manifest.get("extraction_cases", []):
        case_id = str(case["case_id"])
        source_id = str(case["source_id"])
        pages = tuple(int(value) for value in case["pages"])
        normalization = normalize_pdf(
            verified_sources[source_id],
            page_numbers=pages,
            allow_no_supported_tables=True,
        )
        normalization_cache[case_id] = normalization
        case_dir = output_dir / "extraction" / case_id
        case_dir.mkdir(parents=True, exist_ok=True)
        write_json(case_dir / "normalization.json", normalization)
        actual_counts = _row_counts(normalization)
        expected_counts = {key: 0 for key in TABLE_KEYS}
        expected_counts.update(case.get("row_counts", {}))
        row_failures: dict[str, list[dict[str, Any]]] = {}
        for key, expected_rows in case.get("expected_rows", {}).items():
            matched, missing = _check_expected_rows(
                normalization["tables"][key], expected_rows
            )
            if not matched:
                row_failures[key] = missing
        actual_issue_codes = sorted(
            issue
            for item in normalization["issues"]
            for issue in item.get("issues", [])
        )
        expected_issue_codes = sorted(case.get("issue_codes", []))
        unsupported_count = sum(
            item["review_priority"] == "unsupported_data_candidate"
            for item in normalization["unclassified_tables"]
        )
        checks = {
            "row_counts": actual_counts == expected_counts,
            "expected_rows": not row_failures,
            "issue_codes": actual_issue_codes == expected_issue_codes,
            "unsupported_data_candidates": unsupported_count
            == int(case.get("unsupported_data_candidate_count", 0)),
        }
        extraction_results.append(
            {
                "case_id": case_id,
                "source_id": source_id,
                "pages": list(pages),
                "passed": all(checks.values()),
                "checks": checks,
                "actual_row_counts": actual_counts,
                "missing_expected_rows": row_failures,
                "actual_issue_codes": actual_issue_codes,
                "unsupported_data_candidate_count": unsupported_count,
            }
        )

    negative_results: list[dict[str, Any]] = []
    for source_id in manifest.get("negative_control_sources", []):
        contract = configured_sources[source_id]
        pages = tuple(range(1, int(contract["page_count"]) + 1))
        normalization = normalize_pdf(
            verified_sources[source_id],
            page_numbers=pages,
            allow_no_supported_tables=True,
        )
        counts = _row_counts(normalization)
        unsupported_count = sum(
            item["review_priority"] == "unsupported_data_candidate"
            for item in normalization["unclassified_tables"]
        )
        negative_results.append(
            {
                "source_id": source_id,
                "passed": not any(counts.values()) and unsupported_count == 0,
                "row_counts": counts,
                "unsupported_data_candidate_count": unsupported_count,
            }
        )

    mapping_profiles = manifest.get("mapping_profiles", {})
    analysis_results: list[dict[str, Any]] = []
    analyses_by_case: dict[str, dict[str, Any]] = {}
    for case in manifest.get("analysis_cases", []):
        case_id = str(case["case_id"])
        source_id = str(case["source_id"])
        pages = tuple(int(value) for value in case["pages"])
        normalization = normalization_cache.get(str(case.get("extraction_case_id")))
        if normalization is None:
            normalization = normalize_pdf(
                verified_sources[source_id], page_numbers=pages
            )
        profile = mapping_profiles[str(case["mapping_profile"])]
        case_dir = output_dir / "analysis" / case_id
        expected = case["expected"]
        try:
            analysis, _, _ = _build_case_analysis(
                normalization, case, profile, case_dir
            )
        except (CentraleRischiContractError, ValueError) as exc:
            expected_error = str(expected.get("error_contains", ""))
            passed = (
                expected.get("outcome") == "rejected"
                and expected_error
                and expected_error in str(exc)
            )
            analysis_results.append(
                {
                    "case_id": case_id,
                    "source_id": source_id,
                    "pages": list(pages),
                    "outcome": "rejected",
                    "passed": passed,
                    "error": str(exc),
                }
            )
            continue
        if expected.get("outcome") == "rejected":
            analysis_results.append(
                {
                    "case_id": case_id,
                    "source_id": source_id,
                    "pages": list(pages),
                    "outcome": "analysis",
                    "passed": False,
                    "analysis_status": analysis["status"],
                    "error": "Analysis succeeded but the gold contract expected rejection.",
                }
            )
            continue
        analyses_by_case[case_id] = analysis
        metric_values = _metric_values(analysis)
        expected_metrics = expected.get("metrics", {})
        population_counts = {
            key: len(analysis[key])
            for key in (
                "exposures",
                "guarantees",
                "guarantees_received",
                "guarantors",
                "ceded_debtors",
                "other_risk_information",
                "summary_totals",
                "overruns",
                "inframonthly_events",
                "information_requests",
                "prejudicial_events",
            )
        }
        artifact_result = _artifact_checks(analysis, case_dir)
        context_result = _check_context_facts(
            build_model_context(analysis), case.get("required_context_facts", [])
        )
        metamorphic_result = _metamorphic_checks(
            normalization,
            case,
            profile,
            analysis,
            case_dir / "metamorphic",
        )
        checks = {
            "expected_outcome": expected.get("outcome") == "analysis",
            "analysis_status": analysis["status"] == expected["status"],
            "source_counts": all(
                analysis["source"].get(key) == value
                for key, value in expected.get("source_counts", {}).items()
            ),
            "metrics": all(
                metric_values.get(metric_id) == value
                for metric_id, value in expected_metrics.items()
            ),
            "population_counts": all(
                population_counts.get(key) == value
                for key, value in expected.get("population_counts", {}).items()
            ),
            "controls_passed": all(
                item["status"] == "passed" for item in analysis["controls"]
            ),
            "context_facts": context_result["passed"],
            "metamorphic": metamorphic_result["passed"],
            "artifacts": artifact_result["passed"],
        }
        analysis_results.append(
            {
                "case_id": case_id,
                "source_id": source_id,
                "pages": list(pages),
                "outcome": "analysis",
                "passed": all(checks.values()),
                "analysis_status": analysis["status"],
                "checks": checks,
                "actual_metrics": {
                    key: metric_values.get(key) for key in expected_metrics
                },
                "population_counts": population_counts,
                "context": context_result,
                "metamorphic": metamorphic_result,
                "artifacts": artifact_result,
            }
        )

    semantic_packet = _prepare_semantic_packet(manifest, output_dir, analyses_by_case)
    write_json(output_dir / "semantic_review_packet.json", semantic_packet)
    semantic_review = _semantic_review_summary(semantic_packet, semantic_review_path)
    deterministic_passed = all(
        item["passed"]
        for item in extraction_results + negative_results + analysis_results
    )
    semantic_passed = semantic_review.get("passed") is True
    if not deterministic_passed or semantic_review.get("passed") is False:
        overall_status = "failed"
    elif semantic_passed:
        overall_status = {
            "professionally_reviewed": "passed_professionally_reviewed",
            "model_reviewed_not_professional": (
                "passed_model_reviewed_professional_review_pending"
            ),
            "mixed_review_not_fully_professional": (
                "passed_mixed_review_professional_review_pending"
            ),
        }[semantic_review["status"]]
    else:
        overall_status = "deterministic_pass_semantic_review_pending"
    receipt = {
        "schema_version": "vera.centrale_rischi_gold_benchmark_receipt.v1",
        "workflow_id": "centrale-rischi-review",
        "overall_status": overall_status,
        "deterministic_passed": deterministic_passed,
        "sources": source_receipts,
        "extraction_cases": extraction_results,
        "negative_controls": negative_results,
        "analysis_cases": analysis_results,
        "semantic_review": semantic_review,
        "summary": {
            "extraction_total": len(extraction_results) + len(negative_results),
            "extraction_passed": sum(
                item["passed"] for item in extraction_results + negative_results
            ),
            "analysis_total": len(analysis_results),
            "analysis_passed": sum(item["passed"] for item in analysis_results),
        },
        "limitations": list(manifest.get("limitations", [])),
        "implementation_reason": (
            "Hashes, exact row facts, Decimal results, artifact contracts and "
            "metamorphic invariants are deterministic because they are mechanically "
            "reviewable. Commentary usefulness is supplied as a separate model or "
            "professional judgment and is never inferred from keywords."
        ),
    }
    write_json(output_dir / "benchmark_receipt.json", receipt)
    (output_dir / "benchmark_report.html").write_text(
        _render_benchmark_html(receipt), encoding="utf-8"
    )
    return receipt


def main(argv: list[str] | None = None) -> int:
    """Run the benchmark CLI."""

    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--manifest",
        type=Path,
        default=PLUGIN_ROOT / "evals" / "gold_cases.json",
    )
    parser.add_argument("--source", action="append", type=_source_argument, default=[])
    parser.add_argument("--output-dir", required=True, type=Path)
    parser.add_argument("--semantic-review", type=Path)
    args = parser.parse_args(argv)
    source_paths: dict[str, Path] = {}
    for source_id, path in args.source:
        if source_id in source_paths:
            parser.error(f"Duplicate source ID: {source_id}")
        source_paths[source_id] = path
    try:
        receipt = run_benchmark(
            args.manifest,
            source_paths,
            args.output_dir,
            semantic_review_path=args.semantic_review,
        )
    except (CentraleRischiContractError, OSError, ValueError) as exc:
        parser.error(str(exc))
    LOGGER.info(
        "Benchmark %s: %s/%s extraction, %s/%s analysis.",
        receipt["overall_status"],
        receipt["summary"]["extraction_passed"],
        receipt["summary"]["extraction_total"],
        receipt["summary"]["analysis_passed"],
        receipt["summary"]["analysis_total"],
    )
    semantic_failed = receipt["semantic_review"].get("passed") is False
    return 0 if receipt["deterministic_passed"] and not semantic_failed else 1


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    raise SystemExit(main())
