#!/usr/bin/env python3
"""Exact calculations and renderers for one reviewed Centrale Rischi dataset."""

from __future__ import annotations

import csv
import hashlib
import html
import json
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

__all__ = [
    "ANALYSIS_SCHEMA",
    "COMMENTARY_SCHEMA",
    "CentraleRischiContractError",
    "SourceTable",
    "build_analysis",
    "build_inspection",
    "build_model_context",
    "finalize_commentary",
    "load_json",
    "load_source_tables",
    "render_html",
    "render_markdown",
    "sha256_file",
    "write_excel",
    "write_json",
]

ANALYSIS_SCHEMA = "vera.centrale_rischi_analysis.v6"
COMMENTARY_SCHEMA = "vera.centrale_rischi_commentary.v2"
RECIPE_SCHEMA = "vera.centrale_rischi_recipe.v2"
WORKFLOW_ID = "centrale-rischi-review"
ORIGINAL_TERM_CLASSES = (
    "short",
    "medium",
    "long",
    "not_relevant",
    "unclassified",
)
RESIDUAL_TERM_CLASSES = (
    "within_one_year",
    "over_one_year",
    "not_relevant",
    "unclassified",
)
EXPOSURE_FAMILIES = ("performing", "suffering", "other")


class CentraleRischiContractError(ValueError):
    """Raised when supplied evidence does not satisfy the reviewed contract."""


@dataclass(frozen=True)
class SourceTable:
    """One source table with stable lineage metadata."""

    table_id: str
    source_path: Path
    table_label: str
    headers: tuple[str, ...]
    rows: tuple[dict[str, Any], ...]
    source_sha256: str


def sha256_file(path: Path) -> str:
    """Return the SHA-256 of one file."""

    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _json_bytes(value: Any) -> bytes:
    return (
        json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        + "\n"
    ).encode("utf-8")


def write_json(path: Path, payload: Any) -> None:
    """Write stable UTF-8 JSON."""

    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def load_json(path: Path) -> dict[str, Any]:
    """Load one JSON object."""

    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise CentraleRischiContractError(f"Expected a JSON object: {path}")
    return payload


def _cell(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if value is None:
        return ""
    return value


def _source_table(
    path: Path, label: str, headers: Sequence[Any], raw_rows: Sequence[Sequence[Any]]
) -> SourceTable:
    names = tuple(str(value).strip() for value in headers)
    if not names or any(not value for value in names) or len(names) != len(set(names)):
        raise CentraleRischiContractError(
            f"Table {label!r} has blank or duplicate headers."
        )
    source_hash = sha256_file(path)
    table_id = (
        "table_"
        + hashlib.sha256(f"{source_hash}:{label}".encode("utf-8")).hexdigest()[:16]
    )
    rows = tuple(
        {name: _cell(value) for name, value in zip(names, row)}
        for row in raw_rows
        if any(value not in (None, "") for value in row)
    )
    return SourceTable(table_id, path.resolve(), label, names, rows, source_hash)


def load_source_tables(paths: Sequence[Path]) -> list[SourceTable]:
    """Read CSV and Excel tables without assigning semantic roles."""

    tables: list[SourceTable] = []
    for path in paths:
        if not path.is_file():
            raise CentraleRischiContractError(f"Source file not found: {path}")
        suffix = path.suffix.lower()
        if suffix == ".csv":
            with path.open("r", encoding="utf-8-sig", newline="") as handle:
                reader = csv.reader(handle)
                rows = list(reader)
            if not rows:
                raise CentraleRischiContractError(f"CSV is empty: {path}")
            tables.append(_source_table(path, "CSV", rows[0], rows[1:]))
        elif suffix in {".xlsx", ".xlsm"}:
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                for sheet in workbook.worksheets:
                    rows = list(sheet.iter_rows(values_only=True))
                    if rows and any(value not in (None, "") for value in rows[0]):
                        tables.append(
                            _source_table(path, sheet.title, rows[0], rows[1:])
                        )
            finally:
                workbook.close()
        else:
            raise CentraleRischiContractError(
                f"Unsupported source format: {path.suffix}"
            )
    if not tables:
        raise CentraleRischiContractError("No readable source table was found.")
    return tables


def _type_name(value: Any) -> str:
    if value in (None, ""):
        return "empty"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float, Decimal)):
        return "number"
    return "text"


def build_inspection(
    tables: Sequence[SourceTable],
) -> tuple[dict[str, Any], dict[str, Any], dict[str, Any]]:
    """Build a bounded inspection and a deliberately unreviewed recipe."""

    public_tables: list[dict[str, Any]] = []
    private_tables: list[dict[str, Any]] = []
    for table in tables:
        observed_types = {
            header: sorted({_type_name(row.get(header)) for row in table.rows})
            for header in table.headers
        }
        public_tables.append(
            {
                "table_id": table.table_id,
                "table_label": table.table_label,
                "row_count": len(table.rows),
                "columns": list(table.headers),
                "observed_types": observed_types,
                "preview_rows": [
                    {key: str(value)[:200] for key, value in row.items()}
                    for row in table.rows[:10]
                ],
                "source_sha256": table.source_sha256,
            }
        )
        private_tables.append(
            {
                "table_id": table.table_id,
                "absolute_path": table.source_path.as_posix(),
                "table_label": table.table_label,
                "source_sha256": table.source_sha256,
            }
        )
    inventory_hash = hashlib.sha256(_json_bytes(public_tables)).hexdigest()
    inspection = {
        "schema_version": "vera.centrale_rischi_inspection.v1",
        "workflow_id": WORKFLOW_ID,
        "inventory_sha256": inventory_hash,
        "tables": public_tables,
        "semantic_roles_assigned": False,
    }
    control = {
        "schema_version": "vera.centrale_rischi_inspection_control.v1",
        "workflow_id": WORKFLOW_ID,
        "inventory_sha256": inventory_hash,
        "tables": private_tables,
    }
    recipe = {
        "schema_version": RECIPE_SCHEMA,
        "workflow_id": WORKFLOW_ID,
        "inventory_sha256": inventory_hash,
        "entity": "",
        "currency": "EUR",
        "analysis_mode": "descriptive",
        "analysis_objective": "",
        "audience": "professional",
        "source_kind": "tabular_export",
        "source_document_sha256": "",
        "table_id": "",
        "columns": {
            "reference_month": "",
            "intermediary": "",
            "risk_category": "",
            "residual_duration": "",
            "granted": "",
            "operational_granted": "",
            "used": "",
            "guarantee_type": "",
            "guaranteed_amount": "",
            "prejudicial_event": "",
            "reporting_type": "",
            "original_duration": "",
            "relationship_status": "",
            "record_status": "",
            "valid_from": "",
            "valid_to": "",
            "source_page": "",
            "source_region": "",
            "source_row_locator": "",
            "extraction_confidence": "",
        },
        "value_mappings": {
            "original_term": {},
            "residual_term": {},
            "exposure_family": {},
        },
        "control_totals": {},
        "control_tolerance": "0.01",
        "mapping_review": {"status": "pending", "reviewer": "", "reviewed_at": ""},
    }
    return inspection, control, recipe


def _decimal(value: Any, *, field: str, row_number: int) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    normalized = str(value).strip().replace(" ", "")
    if normalized.count(",") == 1 and normalized.count(".") == 0:
        normalized = normalized.replace(",", ".")
    try:
        return Decimal(normalized)
    except InvalidOperation as exc:
        raise CentraleRischiContractError(
            f"Invalid {field} at source row {row_number}: {value!r}"
        ) from exc


def _month(value: Any, *, row_number: int) -> str:
    text = str(value).strip()
    for fmt in ("%Y-%m", "%Y-%m-%d", "%d/%m/%Y", "%m/%Y"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m")
        except ValueError:
            continue
    raise CentraleRischiContractError(
        f"Invalid reference_month at source row {row_number}: {value!r}"
    )


def _text(value: Any) -> str:
    return str(value).strip() if value not in (None, "") else ""


def _mapped_text(
    row: Mapping[str, Any], columns: Mapping[str, str], field: str, default: str = ""
) -> str:
    """Return one optional mapped text value without guessing a source column."""

    column = columns.get(field, "")
    return _text(row.get(column, "")) if column else default


def _decimal_text(value: Decimal | None) -> str | None:
    if value is None:
        return None
    text = format(value, "f")
    return text.rstrip("0").rstrip(".") if "." in text else text


def _ratio(numerator: Decimal, denominator: Decimal) -> Decimal | None:
    return (
        None
        if denominator == 0
        else (numerator / denominator * Decimal("100")).quantize(Decimal("0.01"))
    )


def _require_reviewed_recipe(
    recipe: Mapping[str, Any], tables: Sequence[SourceTable]
) -> tuple[SourceTable, Mapping[str, str]]:
    if (
        recipe.get("schema_version") != RECIPE_SCHEMA
        or recipe.get("workflow_id") != WORKFLOW_ID
    ):
        raise CentraleRischiContractError(
            "Recipe schema or workflow does not match Centrale Rischi Review."
        )
    review = recipe.get("mapping_review")
    if (
        not isinstance(review, Mapping)
        or review.get("status") != "reviewed"
        or not review.get("reviewer")
        or not review.get("reviewed_at")
    ):
        raise CentraleRischiContractError(
            "A named, timestamped mapping_review.status=reviewed is required."
        )
    inspection, _, _ = build_inspection(tables)
    if recipe.get("inventory_sha256") != inspection["inventory_sha256"]:
        raise CentraleRischiContractError(
            "Recipe inventory_sha256 does not match the supplied sources."
        )
    selected = [table for table in tables if table.table_id == recipe.get("table_id")]
    if len(selected) != 1:
        raise CentraleRischiContractError(
            "Recipe must select exactly one inspected exposure table."
        )
    columns = recipe.get("columns")
    if not isinstance(columns, Mapping):
        raise CentraleRischiContractError("Recipe columns must be an object.")
    required = (
        "reference_month",
        "intermediary",
        "risk_category",
        "original_duration",
        "residual_duration",
        "granted",
        "operational_granted",
        "used",
    )
    for field in required:
        column = columns.get(field)
        if (
            not isinstance(column, str)
            or not column
            or column not in selected[0].headers
        ):
            raise CentraleRischiContractError(
                f"Missing or invalid reviewed column mapping: {field}"
            )
    for field in (
        "guarantee_type",
        "guaranteed_amount",
        "prejudicial_event",
        "reporting_type",
        "relationship_status",
        "record_status",
        "valid_from",
        "valid_to",
        "source_page",
        "source_region",
        "source_row_locator",
        "extraction_confidence",
    ):
        column = columns.get(field, "")
        if column and column not in selected[0].headers:
            raise CentraleRischiContractError(
                f"Invalid optional column mapping: {field}"
            )
    if recipe.get("analysis_mode") not in {"descriptive", "trend", "reconciled"}:
        raise CentraleRischiContractError(
            "analysis_mode must be descriptive, trend, or reconciled."
        )
    if recipe.get("analysis_mode") == "reconciled":
        raise CentraleRischiContractError(
            "Reconciled mode requires an external-evidence adapter that is not part of this initial analysis layer."
        )
    if recipe.get("source_kind") not in {
        "tabular_export",
        "native_pdf_extraction",
        "scanned_pdf_extraction",
    }:
        raise CentraleRischiContractError("Unsupported source_kind.")
    if recipe.get("source_kind") != "tabular_export":
        if not recipe.get("source_document_sha256"):
            raise CentraleRischiContractError(
                "PDF-derived rows require source_document_sha256."
            )
        for field in (
            "record_status",
            "valid_from",
            "valid_to",
            "source_page",
            "source_region",
            "source_row_locator",
            "extraction_confidence",
        ):
            if not columns.get(field):
                raise CentraleRischiContractError(
                    f"PDF-derived rows require the provenance mapping: {field}"
                )
    return selected[0], {str(key): str(value) for key, value in columns.items()}


def _metric(
    metric_id: str,
    label: str,
    value: Decimal | int | None,
    unit: str,
    availability: str = "available",
    reason: str | None = None,
) -> dict[str, Any]:
    return {
        "metric_id": metric_id,
        "label": label,
        "value": _decimal_text(value) if isinstance(value, Decimal) else value,
        "unit": unit,
        "availability": availability,
        "reason": reason,
    }


def build_analysis(
    tables: Sequence[SourceTable], recipe: Mapping[str, Any]
) -> dict[str, Any]:
    """Calculate exact CR exposure tables and source-bounded KPI metrics."""

    table, columns = _require_reviewed_recipe(recipe, tables)
    mappings = recipe.get("value_mappings")
    if not isinstance(mappings, Mapping):
        raise CentraleRischiContractError("value_mappings must be an object.")
    original_term_map = mappings.get("original_term")
    residual_term_map = mappings.get("residual_term")
    family_map = mappings.get("exposure_family")
    if (
        not isinstance(original_term_map, Mapping)
        or not isinstance(residual_term_map, Mapping)
        or not isinstance(family_map, Mapping)
    ):
        raise CentraleRischiContractError(
            "Reviewed original_term, residual_term and exposure_family mappings are required."
        )
    normalized: list[dict[str, Any]] = []
    missing_original_term: set[str] = set()
    missing_residual_term: set[str] = set()
    missing_family: set[str] = set()
    for row_number, row in enumerate(table.rows, start=2):
        original_duration_value = _text(row[columns["original_duration"]])
        residual_duration_value = _text(row[columns["residual_duration"]])
        risk_value = _text(row[columns["risk_category"]])
        original_term = original_term_map.get(original_duration_value)
        residual_term = residual_term_map.get(residual_duration_value)
        family = family_map.get(risk_value)
        if original_term not in ORIGINAL_TERM_CLASSES:
            missing_original_term.add(original_duration_value)
        if residual_term not in RESIDUAL_TERM_CLASSES:
            missing_residual_term.add(residual_duration_value)
        if family not in EXPOSURE_FAMILIES:
            missing_family.add(risk_value)
        if (
            original_term not in ORIGINAL_TERM_CLASSES
            or residual_term not in RESIDUAL_TERM_CLASSES
            or family not in EXPOSURE_FAMILIES
        ):
            continue
        record_status = _mapped_text(
            row, columns, "record_status", "current"
        ).casefold()
        if record_status not in {"current", "previous"}:
            raise CentraleRischiContractError(
                f"Invalid record_status at source row {row_number}: {record_status!r}"
            )

        def amount_value(field: str, value: Any) -> Decimal | str:
            try:
                return _decimal(value, field=field, row_number=row_number)
            except CentraleRischiContractError:
                if record_status == "previous":
                    # Previous/corrected records never enter totals. Preserve an
                    # exact non-numeric source state such as "Assenza di
                    # segnalazione" instead of coercing it to zero.
                    return _text(value)
                raise

        granted = amount_value("granted", row[columns["granted"]])
        operational = amount_value(
            "operational_granted", row[columns["operational_granted"]]
        )
        used = amount_value("used", row[columns["used"]])
        if isinstance(operational, Decimal) and isinstance(used, Decimal):
            overrun: Decimal | None = (
                Decimal("0")
                if family == "suffering"
                else max(used - operational, Decimal("0"))
            )
            available: Decimal | None = max(operational - used, Decimal("0"))
        else:
            overrun = None
            available = None
        guarantee_type = (
            _text(row.get(columns.get("guarantee_type", ""), ""))
            if columns.get("guarantee_type")
            else ""
        )
        guaranteed: Decimal | str = (
            amount_value(
                "guaranteed_amount",
                row.get(columns.get("guaranteed_amount", ""), ""),
            )
            if columns.get("guaranteed_amount")
            else Decimal("0")
        )
        prejudicial = (
            _text(row.get(columns.get("prejudicial_event", ""), ""))
            if columns.get("prejudicial_event")
            else ""
        )
        normalized.append(
            {
                "source_row": row_number,
                "reference_month": _month(
                    row[columns["reference_month"]], row_number=row_number
                ),
                "intermediary": _text(row[columns["intermediary"]])
                or "Intermediario non specificato",
                "risk_category": risk_value,
                "exposure_family": family,
                "original_duration": original_duration_value,
                "original_term": original_term,
                "residual_duration": residual_duration_value,
                "residual_term": residual_term,
                "granted": granted,
                "operational_granted": operational,
                "used": used,
                "available": available,
                "overrun": overrun,
                "guarantee_type": guarantee_type,
                "guaranteed_amount": guaranteed,
                "prejudicial_event": prejudicial,
                "reporting_type": _mapped_text(row, columns, "reporting_type"),
                "relationship_status": _mapped_text(
                    row, columns, "relationship_status"
                ),
                "record_status": record_status,
                "valid_from": _mapped_text(row, columns, "valid_from"),
                "valid_to": _mapped_text(row, columns, "valid_to"),
                "source_page": _mapped_text(row, columns, "source_page"),
                "source_region": _mapped_text(row, columns, "source_region"),
                "source_row_locator": _mapped_text(row, columns, "source_row_locator"),
                "extraction_confidence": _mapped_text(
                    row, columns, "extraction_confidence"
                ),
            }
        )
    if missing_original_term or missing_residual_term or missing_family:
        details = []
        if missing_original_term:
            details.append(
                "unmapped original_duration values: "
                + ", ".join(sorted(missing_original_term))
            )
        if missing_residual_term:
            details.append(
                "unmapped residual_duration values: "
                + ", ".join(sorted(missing_residual_term))
            )
        if missing_family:
            details.append(
                "unmapped risk_category values: " + ", ".join(sorted(missing_family))
            )
        raise CentraleRischiContractError("; ".join(details))
    if not normalized:
        raise CentraleRischiContractError(
            "The selected exposure table has no data rows."
        )

    current_rows = [row for row in normalized if row["record_status"] == "current"]
    previous_rows = [row for row in normalized if row["record_status"] == "previous"]
    if not current_rows:
        raise CentraleRischiContractError(
            "The selected exposure table has no current records."
        )
    months = sorted({str(row["reference_month"]) for row in current_rows})
    latest_month = months[-1]
    latest = [row for row in current_rows if row["reference_month"] == latest_month]
    original_term_totals: dict[str, dict[str, Decimal]] = {
        value: {
            "granted": Decimal("0"),
            "operational_granted": Decimal("0"),
            "used": Decimal("0"),
            "overrun": Decimal("0"),
        }
        for value in ORIGINAL_TERM_CLASSES
    }
    residual_term_totals: dict[str, dict[str, Decimal]] = {
        value: {
            "granted": Decimal("0"),
            "operational_granted": Decimal("0"),
            "used": Decimal("0"),
            "overrun": Decimal("0"),
        }
        for value in RESIDUAL_TERM_CLASSES
    }
    monthly: dict[str, dict[str, Decimal]] = defaultdict(lambda: defaultdict(Decimal))
    monthly_category: dict[str, dict[str, dict[str, Decimal]]] = defaultdict(
        lambda: defaultdict(lambda: defaultdict(Decimal))
    )
    intermediary: dict[str, Decimal] = defaultdict(Decimal)
    category_totals: dict[str, dict[str, Decimal]] = defaultdict(
        lambda: defaultdict(Decimal)
    )
    for row in current_rows:
        month = str(row["reference_month"])
        for field in (
            "granted",
            "operational_granted",
            "used",
            "available",
            "overrun",
            "guaranteed_amount",
        ):
            monthly[month][field] += row[field]
        for field in (
            "granted",
            "operational_granted",
            "used",
            "available",
            "overrun",
        ):
            monthly_category[month][str(row["risk_category"])][field] += row[field]
    for row in latest:
        original_bucket = original_term_totals[str(row["original_term"])]
        residual_bucket = residual_term_totals[str(row["residual_term"])]
        for field in original_bucket:
            original_bucket[field] += row[field]
            residual_bucket[field] += row[field]
        intermediary[str(row["intermediary"])] += row["used"]
        for field in (
            "granted",
            "operational_granted",
            "used",
            "available",
            "overrun",
        ):
            category_totals[str(row["risk_category"])][field] += row[field]

    total_granted = sum((row["granted"] for row in latest), Decimal("0"))
    total_operational = sum(
        (row["operational_granted"] for row in latest), Decimal("0")
    )
    total_used = sum((row["used"] for row in latest), Decimal("0"))
    total_available = sum((row["available"] for row in latest), Decimal("0"))
    total_overrun = sum((row["overrun"] for row in latest), Decimal("0"))
    total_guaranteed = sum((row["guaranteed_amount"] for row in latest), Decimal("0"))
    top_bank = max(intermediary.values()) if intermediary else Decimal("0")
    prior_used = monthly[months[-2]]["used"] if len(months) > 1 else None
    metrics = [
        _metric(
            "cr.total_granted",
            "Accordato",
            total_granted,
            str(recipe.get("currency", "EUR")),
        ),
        _metric(
            "cr.total_operational_granted",
            "Accordato operativo",
            total_operational,
            str(recipe.get("currency", "EUR")),
        ),
        _metric(
            "cr.total_used",
            "Utilizzato",
            total_used,
            str(recipe.get("currency", "EUR")),
        ),
        _metric(
            "cr.available_resources",
            "Margine CR calcolato",
            total_available,
            str(recipe.get("currency", "EUR")),
        ),
        _metric(
            "cr.overrun_amount",
            "Sconfinamento",
            total_overrun,
            str(recipe.get("currency", "EUR")),
        ),
        _metric(
            "cr.overrun_count",
            "Linee con sconfinamento",
            sum(1 for row in latest if row["overrun"] > 0),
            "count",
        ),
        _metric(
            "cr.guarantee_coverage_pct",
            "Copertura importi garantiti sulle esposizioni",
            _ratio(total_guaranteed, total_used),
            "percent",
        ),
        _metric(
            "cr.top_intermediary_share_pct",
            "Quota primo intermediario",
            _ratio(top_bank, total_used),
            "percent",
        ),
        _metric(
            "cr.original_term.short_share_pct",
            "Quota breve per durata originaria",
            _ratio(original_term_totals["short"]["used"], total_used),
            "percent",
        ),
        _metric(
            "cr.original_term.medium_share_pct",
            "Quota media per durata originaria",
            _ratio(original_term_totals["medium"]["used"], total_used),
            "percent",
        ),
        _metric(
            "cr.original_term.long_share_pct",
            "Quota lunga per durata originaria",
            _ratio(original_term_totals["long"]["used"], total_used),
            "percent",
        ),
        _metric(
            "cr.previous_record_count",
            "Segnalazioni precedenti escluse dai totali correnti",
            len(previous_rows),
            "count",
        ),
        _metric(
            "cr.used_mom_change",
            "Variazione utilizzato mese su mese",
            None if prior_used is None else total_used - prior_used,
            str(recipe.get("currency", "EUR")),
            "unavailable" if prior_used is None else "available",
            "Richiede almeno due mesi di riferimento." if prior_used is None else None,
        ),
        _metric(
            "financial.net_debt_ebitda",
            "PFN / EBITDA",
            None,
            "ratio",
            "unavailable",
            "Richiede dati di bilancio verificati esterni alla Centrale Rischi.",
        ),
        _metric(
            "financial.debt_equity",
            "Debt / Equity",
            None,
            "ratio",
            "unavailable",
            "Richiede dati patrimoniali verificati esterni alla Centrale Rischi.",
        ),
        _metric(
            "financial.dscr",
            "DSCR",
            None,
            "ratio",
            "unavailable",
            "Richiede flussi di cassa e servizio del debito verificati esterni alla Centrale Rischi.",
        ),
    ]

    category_summary: list[dict[str, Any]] = []
    for category in sorted(category_totals):
        totals = category_totals[category]
        category_id = hashlib.sha256(category.encode("utf-8")).hexdigest()[:10]
        utilization = _ratio(totals["used"], totals["operational_granted"])
        category_summary.append(
            {
                "risk_category": category,
                **{field: _decimal_text(value) for field, value in totals.items()},
                "utilization_pct": _decimal_text(utilization),
            }
        )
        category_metrics = (
            (
                "used",
                f"Utilizzato — {category}",
                totals["used"],
                str(recipe.get("currency", "EUR")),
                "available",
                None,
            ),
            (
                "available_resources",
                f"Margine CR calcolato — {category}",
                totals["available"],
                str(recipe.get("currency", "EUR")),
                "available",
                None,
            ),
            (
                "utilization_pct",
                f"Utilizzo su accordato operativo — {category}",
                utilization,
                "percent",
                "unavailable" if utilization is None else "available",
                (
                    "L'accordato operativo è zero per questa categoria."
                    if utilization is None
                    else None
                ),
            ),
        )
        for suffix, label, value, unit, availability, reason in category_metrics:
            metrics.append(
                {
                    **_metric(
                        f"cr.category.{category_id}.{suffix}",
                        label,
                        value,
                        unit,
                        availability,
                        reason,
                    ),
                    "dimension": {"risk_category": category},
                }
            )

    category_movement_summary: list[dict[str, Any]] = []
    if len(months) > 1:
        prior_month = months[-2]
        latest_categories = set(monthly_category[latest_month])
        prior_categories = set(monthly_category[prior_month])
        for category in sorted(latest_categories | prior_categories):
            category_id = hashlib.sha256(category.encode("utf-8")).hexdigest()[:10]
            prior_totals = monthly_category[prior_month][category]
            latest_totals = monthly_category[latest_month][category]
            movement = {
                "risk_category": category,
                "prior_reference_month": prior_month,
                "latest_reference_month": latest_month,
                "presence": (
                    "new_in_latest"
                    if category not in prior_categories
                    else (
                        "absent_in_latest"
                        if category not in latest_categories
                        else "continuing"
                    )
                ),
            }
            for field in (
                "granted",
                "operational_granted",
                "used",
                "available",
                "overrun",
            ):
                prior_value = prior_totals[field]
                latest_value = latest_totals[field]
                movement[f"prior_{field}"] = _decimal_text(prior_value)
                movement[f"latest_{field}"] = _decimal_text(latest_value)
                movement[f"{field}_change"] = _decimal_text(latest_value - prior_value)
            category_movement_summary.append(movement)
            for suffix, label, field in (
                (
                    "operational_granted_change",
                    f"Variazione accordato operativo — {category}",
                    "operational_granted_change",
                ),
                (
                    "used_change",
                    f"Variazione utilizzato — {category}",
                    "used_change",
                ),
                (
                    "overrun_change",
                    f"Variazione sconfinamento — {category}",
                    "overrun_change",
                ),
            ):
                metrics.append(
                    {
                        **_metric(
                            f"cr.category.{category_id}.{suffix}",
                            label,
                            Decimal(str(movement[field])),
                            str(recipe.get("currency", "EUR")),
                        ),
                        "dimension": {
                            "risk_category": category,
                            "prior_reference_month": prior_month,
                            "latest_reference_month": latest_month,
                        },
                    }
                )

    controls: list[dict[str, Any]] = []
    tolerance = _decimal(
        recipe.get("control_tolerance", "0.01"), field="control_tolerance", row_number=0
    )
    available_totals = {
        "granted": total_granted,
        "operational_granted": total_operational,
        "used": total_used,
    }
    declared_controls = recipe.get("control_totals") or {}
    if not isinstance(declared_controls, Mapping):
        raise CentraleRischiContractError("control_totals must be an object.")
    for name, expected_value in declared_controls.items():
        if name not in available_totals:
            raise CentraleRischiContractError(f"Unsupported control total: {name}")
        expected = _decimal(
            expected_value, field=f"control_totals.{name}", row_number=0
        )
        actual = available_totals[name]
        controls.append(
            {
                "control_id": f"control.latest.{name}",
                "expected": _decimal_text(expected),
                "actual": _decimal_text(actual),
                "difference": _decimal_text(actual - expected),
                "status": "passed" if abs(actual - expected) <= tolerance else "failed",
            }
        )
    failed_controls = [item for item in controls if item["status"] == "failed"]
    status = (
        "blocked"
        if failed_controls
        else (
            "partial"
            if any(
                row["original_term"] == "unclassified"
                or row["residual_term"] == "unclassified"
                for row in latest
            )
            or (recipe.get("analysis_mode") == "trend" and len(months) < 2)
            else "complete"
        )
    )

    def public_row(row: Mapping[str, Any]) -> dict[str, Any]:
        return {
            key: _decimal_text(value) if isinstance(value, Decimal) else value
            for key, value in row.items()
        }

    def auxiliary_rows(label: str) -> tuple[list[dict[str, Any]], bool]:
        matching_tables = [item for item in tables if item.table_label == label]
        rows = [public_row(row) for item in matching_tables for row in item.rows]
        # PDF normalization creates review sheets for every supported population.
        # An empty generated sheet is not evidence that the source population was
        # present and empty, so availability requires at least one extracted row.
        return rows, bool(rows)

    exposures = [public_row(row) for row in normalized]
    original_term_summary = [
        {
            "original_term": key,
            **{field: _decimal_text(value) for field, value in totals.items()},
        }
        for key, totals in original_term_totals.items()
    ]
    residual_term_summary = [
        {
            "residual_term": key,
            **{field: _decimal_text(value) for field, value in totals.items()},
        }
        for key, totals in residual_term_totals.items()
    ]
    monthly_series = [
        {
            "reference_month": month,
            **{field: _decimal_text(value) for field, value in monthly[month].items()},
        }
        for month in months
    ]
    guarantees = [public_row(row) for row in latest if row["guaranteed_amount"] > 0]
    overruns = [public_row(row) for row in latest if row["overrun"] > 0]
    prejudicial = [public_row(row) for row in latest if row["prejudicial_event"]]
    guarantees_received, guarantees_received_available = auxiliary_rows(
        "Garanzie ricevute"
    )
    guarantors, guarantors_available = auxiliary_rows("Garanti intestatario")
    ceded_debtors, ceded_debtors_available = auxiliary_rows("Debitori ceduti")
    other_risk_information, other_risk_information_available = auxiliary_rows(
        "Altre informazioni"
    )
    summary_totals, summary_totals_available = auxiliary_rows("Prospetto sintetico")
    inframonthly_events, inframonthly_events_available = auxiliary_rows(
        "Eventi inframensili"
    )
    information_requests, information_requests_available = auxiliary_rows(
        "Richieste informazioni"
    )
    limitations = [
        "L'utilizzo è calcolato per categoria di rischio confermata; non viene presentato un unico rapporto trasversale perché gli importi utilizzati hanno significati specifici per categoria.",
        "Breve, medio e lungo derivano dai valori confermati della durata originaria. La durata residua è esposta separatamente come entro un anno, oltre un anno, non rilevante o non classificata e non consente di distinguere medio da lungo termine.",
        "L'Importo garantito su un'esposizione, i Garanti dell'intestatario e le Garanzie ricevute per obbligazioni di terzi sono popolazioni diverse e non vengono unite.",
        "I Debitori ceduti e il valore nominale dei crediti ceduti sono esposti separatamente e non vengono sommati alle esposizioni creditizie dell'intestatario.",
        "Il Prospetto sintetico e le altre informazioni di rischio sono popolazioni di controllo o informative: non vengono sommate alle esposizioni analitiche e non sostituiscono la riconciliazione professionale.",
        "La Centrale Rischi da sola non consente di calcolare PFN/EBITDA, Debt/Equity o DSCR.",
        "L'output non riproduce né stima il rating proprietario di una banca.",
    ]
    if not columns.get("prejudicial_event"):
        limitations.append(
            "Le evidenze pregiudizievoli non sono disponibili perché non è stata fornita una colonna proveniente da una fonte separata e confermata."
        )
    if category_movement_summary:
        limitations.append(
            "Il confronto per categoria mostra variazioni aggregate tra i due periodi più recenti; senza una riconciliazione per singolo rapporto non prova che una specifica posizione sia stata riclassificata."
        )
    return {
        "schema_version": ANALYSIS_SCHEMA,
        "workflow_id": WORKFLOW_ID,
        "status": status,
        "review_status": "draft_pending_professional_review",
        "entity": str(recipe.get("entity", "")),
        "currency": str(recipe.get("currency", "EUR")),
        "analysis_mode": str(recipe.get("analysis_mode")),
        "analysis_objective": str(recipe.get("analysis_objective", "")),
        "audience": str(recipe.get("audience", "professional")),
        "latest_reference_month": latest_month,
        "source": {
            "table_id": table.table_id,
            "source_sha256": table.source_sha256,
            "source_kind": str(recipe.get("source_kind")),
            "source_document_sha256": str(recipe.get("source_document_sha256", "")),
            "inventory_sha256": recipe["inventory_sha256"],
            "row_count": len(normalized),
            "current_row_count": len(current_rows),
            "previous_row_count": len(previous_rows),
        },
        "metrics": metrics,
        "exposures": exposures,
        "original_term_summary": original_term_summary,
        "residual_term_summary": residual_term_summary,
        "risk_category_summary": category_summary,
        "category_movement_summary": category_movement_summary,
        "monthly_series": monthly_series,
        "guarantees": guarantees,
        "guarantees_received": guarantees_received,
        "guarantors": guarantors,
        "ceded_debtors": ceded_debtors,
        "other_risk_information": other_risk_information,
        "summary_totals": summary_totals,
        "overruns": overruns,
        "inframonthly_events": inframonthly_events,
        "information_requests": information_requests,
        "prejudicial_events": prejudicial,
        "coverage": {
            "guarantees_on_exposures": (
                "available" if columns.get("guaranteed_amount") else "unavailable"
            ),
            "guarantees_received": (
                "available" if guarantees_received_available else "unavailable"
            ),
            "guarantors": "available" if guarantors_available else "unavailable",
            "ceded_debtors": (
                "available" if ceded_debtors_available else "unavailable"
            ),
            "other_risk_information": (
                "available" if other_risk_information_available else "unavailable"
            ),
            "summary_totals": (
                "available" if summary_totals_available else "unavailable"
            ),
            "inframonthly_events": (
                "available" if inframonthly_events_available else "unavailable"
            ),
            "information_requests": (
                "available" if information_requests_available else "unavailable"
            ),
            "pregiudizievoli": (
                "available" if columns.get("prejudicial_event") else "unavailable"
            ),
            "multiple_months": "available" if len(months) > 1 else "unavailable",
            "trend_analysis": "available" if len(months) > 1 else "unavailable",
            "previous_records": "available" if previous_rows else "unavailable",
            "reconciled_analysis": "unavailable",
            "financial_statement_ratios": "unavailable",
        },
        "controls": controls,
        "assurance_levels": {
            "documentary": (
                "pdf_rows_with_reviewed_layout_provenance"
                if recipe.get("source_kind") != "tabular_export"
                else "reviewed_tabular_mapping_not_original_report_validation"
            ),
            "arithmetic": "failed" if failed_controls else "passed",
            "semantic": "reviewed_recipe",
            "professional": "pending",
        },
        "limitations": limitations,
    }


def build_model_context(analysis: Mapping[str, Any]) -> dict[str, Any]:
    """Project bounded calculated evidence for model-led interpretation."""

    def project(
        rows: Sequence[Mapping[str, Any]], fields: Sequence[str]
    ) -> list[dict[str, Any]]:
        return [
            {field: row[field] for field in fields if field in row} for row in rows[:20]
        ]

    def correction_sort_key(row: Mapping[str, Any]) -> tuple[str, date, int]:
        try:
            validity_end = datetime.strptime(
                str(row.get("valid_to", "")), "%d/%m/%Y"
            ).date()
        except ValueError:
            validity_end = date.min
        try:
            source_page = int(str(row.get("source_page", "0")))
        except ValueError:
            source_page = 0
        return str(row.get("reference_month", "")), validity_end, source_page

    return {
        "schema_version": "vera.centrale_rischi_model_context.v6",
        "workflow_id": WORKFLOW_ID,
        "status": analysis["status"],
        "entity": analysis["entity"],
        "currency": analysis["currency"],
        "latest_reference_month": analysis["latest_reference_month"],
        "metrics": analysis["metrics"],
        "original_term_summary": analysis["original_term_summary"],
        "residual_term_summary": analysis["residual_term_summary"],
        "risk_category_summary": analysis["risk_category_summary"],
        "category_movement_summary": analysis["category_movement_summary"][:50],
        "monthly_series": analysis["monthly_series"][-36:],
        "previous_records": project(
            sorted(
                (
                    row
                    for row in analysis["exposures"]
                    if row.get("record_status") == "previous"
                ),
                key=correction_sort_key,
                reverse=True,
            ),
            (
                "reference_month",
                "intermediary",
                "risk_category",
                "original_duration",
                "residual_duration",
                "granted",
                "operational_granted",
                "used",
                "guarantee_type",
                "guaranteed_amount",
                "record_status",
                "valid_from",
                "valid_to",
                "source_page",
                "source_row_locator",
                "extraction_confidence",
            ),
        ),
        "top_overruns": sorted(
            analysis["overruns"],
            key=lambda row: Decimal(str(row["overrun"])),
            reverse=True,
        )[:20],
        "top_guarantees": sorted(
            analysis["guarantees"],
            key=lambda row: Decimal(str(row["guaranteed_amount"])),
            reverse=True,
        )[:20],
        "guarantees_received": project(
            analysis["guarantees_received"],
            (
                "reference_month",
                "intermediary",
                "category",
                "location",
                "guaranteed_party",
                "relationship_status",
                "guarantee_type",
                "guarantee_value",
                "guaranteed_amount",
                "record_status",
                "valid_from",
                "valid_to",
                "source_page",
                "source_row_locator",
                "extraction_confidence",
            ),
        ),
        "guarantors": project(
            analysis["guarantors"],
            (
                "reference_month",
                "intermediary",
                "guarantor",
                "guarantee_value",
                "guaranteed_amount",
                "record_status",
                "valid_from",
                "valid_to",
                "source_page",
                "source_row_locator",
                "extraction_confidence",
            ),
        ),
        "ceded_debtors": project(
            analysis["ceded_debtors"],
            (
                "reference_month",
                "intermediary",
                "ceded_debtor",
                "nominal_value",
                "record_status",
                "valid_from",
                "valid_to",
                "source_page",
                "source_row_locator",
                "extraction_confidence",
            ),
        ),
        "other_risk_information": project(
            analysis["other_risk_information"],
            (
                "reference_month",
                "intermediary",
                "category",
                "location",
                "original_duration",
                "residual_duration",
                "currency",
                "activity_type",
                "relationship_status",
                "amount",
                "intrinsic_value",
                "record_status",
                "valid_from",
                "valid_to",
                "source_page",
                "source_row_locator",
                "extraction_confidence",
            ),
        ),
        "summary_totals": project(
            analysis["summary_totals"],
            (
                "reference_month",
                "intermediary",
                "summary_category",
                "granted",
                "operational_granted",
                "used",
                "guarantee_value",
                "guaranteed_amount",
                "intrinsic_value",
                "source_page",
                "source_row_locator",
                "extraction_confidence",
            ),
        ),
        "inframonthly_events": project(
            analysis["inframonthly_events"],
            (
                "intermediary",
                "event_date",
                "event_type",
                "event_cancelled",
                "source_page",
                "source_row_locator",
                "extraction_confidence",
            ),
        ),
        "information_requests": project(
            analysis["information_requests"],
            (
                "intermediary",
                "request_date",
                "requested_period",
                "request_type",
                "request_reason_code",
                "request_reason",
                "validity_period",
                "notes",
                "source_page",
                "source_row_locator",
                "extraction_confidence",
            ),
        ),
        "prejudicial_events": analysis["prejudicial_events"][:20],
        "coverage": analysis["coverage"],
        "controls": analysis["controls"],
        "assurance_levels": analysis["assurance_levels"],
        "limitations": analysis["limitations"],
        "excluded_by_default": [
            "current raw source population",
            "previous records beyond the bounded review projection",
            "absolute paths",
            "original filenames",
        ],
    }


def finalize_commentary(
    analysis: Mapping[str, Any], commentary: Mapping[str, Any]
) -> dict[str, Any]:
    """Validate commentary shape and evidence-reference closure."""

    if (
        analysis.get("schema_version") != ANALYSIS_SCHEMA
        or analysis.get("workflow_id") != WORKFLOW_ID
    ):
        raise CentraleRischiContractError(
            "Analysis contract does not match this workflow."
        )
    if analysis.get("status") == "blocked" or any(
        item.get("status") == "failed" for item in analysis.get("controls", [])
    ):
        raise CentraleRischiContractError("Blocked analysis cannot be finalized.")
    if (
        commentary.get("schema_version") != COMMENTARY_SCHEMA
        or commentary.get("workflow_id") != WORKFLOW_ID
    ):
        raise CentraleRischiContractError("Commentary schema or workflow is invalid.")
    model_context = build_model_context(analysis)
    valid_evidence_refs = {
        f"metric:{item['metric_id']}" for item in model_context.get("metrics", [])
    }
    valid_evidence_refs.update(
        f"control:{item['control_id']}" for item in model_context.get("controls", [])
    )

    def add_row_references(value: Any) -> None:
        if isinstance(value, Mapping):
            locator = value.get("source_row_locator")
            if locator:
                valid_evidence_refs.add(f"row:{locator}")
            for nested in value.values():
                add_row_references(nested)
        elif isinstance(value, list):
            for nested in value:
                add_row_references(nested)

    add_row_references(model_context)
    normalized: dict[str, Any] = {
        "schema_version": COMMENTARY_SCHEMA,
        "workflow_id": WORKFLOW_ID,
    }
    for section in ("observations", "hypotheses"):
        items = commentary.get(section, [])
        if not isinstance(items, list):
            raise CentraleRischiContractError(f"Commentary {section} must be a list.")
        normalized_items = []
        for item in items:
            if not isinstance(item, Mapping) or not _text(item.get("text")):
                raise CentraleRischiContractError(f"Each {section} item requires text.")
            references = item.get("evidence_refs")
            if (
                not isinstance(references, list)
                or not references
                or not set(map(str, references)) <= valid_evidence_refs
            ):
                raise CentraleRischiContractError(
                    f"Each {section} item requires only existing evidence_refs."
                )
            normalized_items.append(
                {
                    "text": _text(item["text"]),
                    "evidence_refs": [str(value) for value in references],
                }
            )
        normalized[section] = normalized_items
    for section in ("questions", "limitations"):
        items = commentary.get(section, [])
        if not isinstance(items, list) or any(
            not isinstance(item, str) or not item.strip() for item in items
        ):
            raise CentraleRischiContractError(
                f"Commentary {section} must contain non-empty strings."
            )
        normalized[section] = [item.strip() for item in items]
    normalized["status"] = "draft_pending_professional_review"
    return normalized


def _metric_rows(analysis: Mapping[str, Any]) -> str:
    return "".join(
        "<tr>"
        f"<td>{html.escape(str(item['label']))}</td>"
        f'<td class="number">{html.escape(_metric_value(item))}</td>'
        f"<td>{html.escape(_unit_label(str(item['unit'])))}</td>"
        "<td>"
        f"{html.escape(_availability_label(str(item['availability'])))}"
        + (
            f"<small>{html.escape(str(item['reason']))}</small>"
            if item.get("reason")
            else ""
        )
        + "</td></tr>"
        for item in analysis["metrics"]
        if "dimension" not in item
    )


def _availability_label(value: str) -> str:
    return {"available": "Disponibile", "unavailable": "Non disponibile"}.get(
        value, value
    )


def _unit_label(value: str) -> str:
    return {
        "count": "numero",
        "percent": "%",
        "ratio": "rapporto",
    }.get(value, value)


def _review_status_label(value: str) -> str:
    return {
        "draft_pending_professional_review": (
            "Bozza in attesa di revisione professionale"
        ),
        "pending_professional_review": "In attesa di revisione professionale",
    }.get(value, value)


def _format_number_it(value: Any) -> str:
    if value in (None, ""):
        return "—"
    try:
        number = Decimal(str(value))
    except InvalidOperation:
        return str(value)
    rendered = format(number, "f")
    integer, separator, decimal = rendered.partition(".")
    sign = "-" if integer.startswith("-") else ""
    digits = integer.lstrip("-")
    grouped = ".".join(
        reversed([digits[max(0, end - 3) : end] for end in range(len(digits), 0, -3)])
    )
    return f"{sign}{grouped}{',' + decimal if separator and decimal else ''}"


def _metric_value(item: Mapping[str, Any]) -> str:
    if item.get("value") is None:
        return "Non disponibile"
    return _format_number_it(item["value"])


_NUMERIC_FIELDS = frozenset(
    {
        "granted",
        "operational_granted",
        "used",
        "available",
        "overrun",
        "guaranteed_amount",
        "guarantee_value",
        "average_balance",
        "utilization_pct",
    }
)


def _display_cell(field: str, value: Any) -> str:
    return _format_number_it(value) if field in _NUMERIC_FIELDS else str(value or "—")


def _html_table(
    rows: Sequence[Mapping[str, Any]], columns: Sequence[tuple[str, str]]
) -> str:
    headings = "".join(f"<th>{html.escape(label)}</th>" for label, _ in columns)
    body = "".join(
        "<tr>"
        + "".join(
            (
                '<td class="number">'
                + html.escape(_display_cell(field, row.get(field)))
                + "</td>"
                if field in _NUMERIC_FIELDS
                else "<td>"
                + html.escape(_display_cell(field, row.get(field)))
                + "</td>"
            )
            for _, field in columns
        )
        + "</tr>"
        for row in rows
    )
    return (
        '<div class="table-wrap"><table><thead><tr>'
        + headings
        + "</tr></thead><tbody>"
        + body
        + "</tbody></table></div>"
    )


def _empty_population_message(
    analysis: Mapping[str, Any], coverage_key: str | None
) -> str:
    if coverage_key and analysis["coverage"].get(coverage_key) == "unavailable":
        return (
            "Non disponibile: nessuna riga è stata estratta o fornita per questa "
            "popolazione."
        )
    return "Nessuna voce rilevata nella popolazione disponibile."


def _html_population(
    analysis: Mapping[str, Any],
    *,
    title: str,
    key: str,
    columns: Sequence[tuple[str, str]],
    coverage_key: str | None = None,
) -> str:
    rows = analysis[key]
    content = (
        _html_table(rows, columns)
        if rows
        else f'<p class="empty">{html.escape(_empty_population_message(analysis, coverage_key))}</p>'
    )
    return f"<section><h2>{html.escape(title)}</h2>{content}</section>"


def _term_label(value: str) -> str:
    return {
        "short": "breve",
        "medium": "medio",
        "long": "lungo",
        "within_one_year": "entro un anno",
        "over_one_year": "oltre un anno",
        "not_relevant": "non rilevante",
        "unclassified": "non classificato",
    }.get(value, value)


def _movement_label(value: str) -> str:
    return {
        "continuing": "presente in entrambi i periodi",
        "new_in_latest": "presente solo nel periodo più recente",
        "absent_in_latest": "assente nel periodo più recente",
    }.get(value, value)


def render_markdown(
    analysis: Mapping[str, Any], commentary: Mapping[str, Any] | None = None
) -> str:
    """Render calculated facts and optional reviewed-draft commentary."""

    lines = [
        f"# Centrale Rischi — {analysis['entity']}",
        "",
        f"Periodo più recente: {analysis['latest_reference_month']}",
        "",
        f"Stato: {_review_status_label(str(analysis['review_status']))}",
        "",
        "## KPI",
        "",
        "| Metrica | Valore | Unità | Disponibilità |",
        "| --- | ---: | --- | --- |",
    ]
    for item in analysis["metrics"]:
        lines.append(
            f"| {item['label']} | {_metric_value(item)} | {_unit_label(str(item['unit']))} | {_availability_label(str(item['availability']))} |"
        )
    lines.extend(
        [
            "",
            "## Esposizioni per durata originaria",
            "",
            "| Scadenza | Accordato | Operativo | Utilizzato | Sconfinamento |",
            "| --- | ---: | ---: | ---: | ---: |",
        ]
    )
    for item in analysis["original_term_summary"]:
        lines.append(
            f"| {_term_label(str(item['original_term']))} | {item['granted']} | {item['operational_granted']} | {item['used']} | {item['overrun']} |"
        )
    lines.extend(
        [
            "",
            "## Esposizioni per durata residua",
            "",
            "| Durata residua | Accordato | Operativo | Utilizzato | Sconfinamento |",
            "| --- | ---: | ---: | ---: | ---: |",
        ]
    )
    for item in analysis["residual_term_summary"]:
        lines.append(
            f"| {_term_label(str(item['residual_term']))} | {item['granted']} | {item['operational_granted']} | {item['used']} | {item['overrun']} |"
        )
    lines.extend(
        [
            "",
            "## Indicatori per categoria di rischio",
            "",
            "| Categoria | Accordato operativo | Utilizzato | Margine | Utilizzo % |",
            "| --- | ---: | ---: | ---: | ---: |",
        ]
    )
    for item in analysis["risk_category_summary"]:
        lines.append(
            f"| {item['risk_category']} | {item['operational_granted']} | {item['used']} | {item['available']} | {item['utilization_pct'] if item['utilization_pct'] is not None else 'Non disponibile'} |"
        )
    if analysis["category_movement_summary"]:
        lines.extend(
            [
                "",
                "## Variazione tra gli ultimi due periodi per categoria",
                "",
                "| Categoria | Periodo precedente | Periodo recente | Presenza | Utilizzato precedente | Utilizzato recente | Variazione utilizzato | Sconfinamento precedente | Sconfinamento recente |",
                "| --- | --- | --- | --- | ---: | ---: | ---: | ---: | ---: |",
            ]
        )
        for item in analysis["category_movement_summary"]:
            lines.append(
                f"| {item['risk_category']} | {item['prior_reference_month']} | {item['latest_reference_month']} | {_movement_label(str(item['presence']))} | {item['prior_used']} | {item['latest_used']} | {item['used_change']} | {item['prior_overrun']} | {item['latest_overrun']} |"
            )

    def population(
        title: str,
        key: str,
        columns: Sequence[tuple[str, str]],
        coverage_key: str | None = None,
    ) -> None:
        lines.extend(["", f"## {title}", ""])
        rows = analysis[key]
        if not rows:
            lines.append(_empty_population_message(analysis, coverage_key))
            return
        lines.append("| " + " | ".join(label for label, _ in columns) + " |")
        lines.append("| " + " | ".join("---" for _ in columns) + " |")
        for row in rows:
            values = [
                _display_cell(field, row.get(field)).replace("|", "\\|")
                for _, field in columns
            ]
            lines.append("| " + " | ".join(values) + " |")

    population(
        "Garanzie collegate alle esposizioni",
        "guarantees",
        (
            ("Intermediario", "intermediary"),
            ("Categoria", "risk_category"),
            ("Tipo", "guarantee_type"),
            ("Importo garantito", "guaranteed_amount"),
        ),
        "guarantees_on_exposures",
    )
    population(
        "Garanzie ricevute per obbligazioni di terzi",
        "guarantees_received",
        (
            ("Mese", "reference_month"),
            ("Intermediario", "intermediary"),
            ("Garantito", "guaranteed_party"),
            ("Tipo", "guarantee_type"),
            ("Valore garanzia", "guarantee_value"),
            ("Importo garantito", "guaranteed_amount"),
        ),
        "guarantees_received",
    )
    population(
        "Garanti dell'intestatario",
        "guarantors",
        (
            ("Mese", "reference_month"),
            ("Intermediario", "intermediary"),
            ("Garante", "guarantor"),
            ("Valore garanzia", "guarantee_value"),
            ("Importo garantito", "guaranteed_amount"),
        ),
        "guarantors",
    )
    population(
        "Debitori ceduti",
        "ceded_debtors",
        (
            ("Mese", "reference_month"),
            ("Intermediario", "intermediary"),
            ("Debitore ceduto", "ceded_debtor"),
            ("Valore nominale", "nominal_value"),
        ),
        "ceded_debtors",
    )
    population(
        "Altre informazioni di rischio",
        "other_risk_information",
        (
            ("Mese", "reference_month"),
            ("Intermediario", "intermediary"),
            ("Categoria", "category"),
            ("Localizzazione", "location"),
            ("Importo", "amount"),
            ("Valore intrinseco", "intrinsic_value"),
        ),
        "other_risk_information",
    )
    population(
        "Prospetto sintetico",
        "summary_totals",
        (
            ("Mese", "reference_month"),
            ("Intermediario", "intermediary"),
            ("Categoria", "summary_category"),
            ("Accordato", "granted"),
            ("Operativo", "operational_granted"),
            ("Utilizzato", "used"),
            ("Valore garanzia", "guarantee_value"),
            ("Importo garantito", "guaranteed_amount"),
            ("Valore intrinseco", "intrinsic_value"),
        ),
        "summary_totals",
    )
    population(
        "Sconfinamenti",
        "overruns",
        (
            ("Intermediario", "intermediary"),
            ("Categoria", "risk_category"),
            ("Accordato operativo", "operational_granted"),
            ("Utilizzato", "used"),
            ("Sconfinamento", "overrun"),
        ),
    )
    population(
        "Eventi inframensili",
        "inframonthly_events",
        (
            ("Intermediario", "intermediary"),
            ("Data", "event_date"),
            ("Evento", "event_type"),
            ("Cancellato", "event_cancelled"),
        ),
        "inframonthly_events",
    )
    population(
        "Richieste di informazione",
        "information_requests",
        (
            ("Intermediario", "intermediary"),
            ("Data", "request_date"),
            ("Periodo richiesto", "requested_period"),
            ("Tipo", "request_type"),
            ("Causale", "request_reason"),
        ),
        "information_requests",
    )
    population(
        "Pregiudizievoli",
        "prejudicial_events",
        (
            ("Intermediario", "intermediary"),
            ("Categoria", "risk_category"),
            ("Evidenza", "prejudicial_event"),
        ),
        "pregiudizievoli",
    )
    lines.extend(["", "## Limiti", ""])
    lines.extend(f"- {item}" for item in analysis["limitations"])
    if commentary:
        lines.extend(
            ["", "## Commento — bozza in attesa di revisione professionale", ""]
        )
        for heading, key in (
            ("Osservazioni", "observations"),
            ("Ipotesi", "hypotheses"),
        ):
            lines.extend([f"### {heading}", ""])
            lines.extend(
                f"- {item['text']} ({', '.join(item['evidence_refs'])})"
                for item in commentary[key]
            )
        for heading, key in (
            ("Domande", "questions"),
            ("Limiti aggiuntivi", "limitations"),
        ):
            lines.extend(["", f"### {heading}", ""])
            lines.extend(f"- {item}" for item in commentary[key])
    return "\n".join(lines).rstrip() + "\n"


def render_html(
    analysis: Mapping[str, Any], commentary: Mapping[str, Any] | None = None
) -> str:
    """Render a self-contained, read-only review dashboard."""

    commentary_html = ""
    if commentary:
        blocks = []
        for title, key in (
            ("Osservazioni", "observations"),
            ("Ipotesi", "hypotheses"),
            ("Domande da approfondire", "questions"),
            ("Limiti aggiuntivi", "limitations"),
        ):
            if key in {"observations", "hypotheses"}:
                items = "".join(
                    f"<li><span>{html.escape(item['text'])}</span>"
                    '<details class="evidence"><summary>Evidenze</summary>'
                    f"<code>{html.escape(', '.join(item['evidence_refs']))}</code>"
                    "</details></li>"
                    for item in commentary[key]
                )
            else:
                items = "".join(
                    f"<li>{html.escape(item)}</li>" for item in commentary[key]
                )
            blocks.append(
                f"<h3>{title}</h3><ul>{items or '<li>Nessuna voce.</li>'}</ul>"
            )
        commentary_html = (
            "<section><h2>Commento professionale — bozza</h2>"
            + "".join(blocks)
            + "</section>"
        )
    original_term_rows = [
        {**item, "term": _term_label(str(item["original_term"]))}
        for item in analysis["original_term_summary"]
    ]
    residual_term_rows = [
        {**item, "term": _term_label(str(item["residual_term"]))}
        for item in analysis["residual_term_summary"]
    ]
    amount_columns = (
        ("Classe", "term"),
        ("Accordato", "granted"),
        ("Operativo", "operational_granted"),
        ("Utilizzato", "used"),
        ("Sconfinamento", "overrun"),
    )
    category_movement_rows = [
        {**item, "presence_label": _movement_label(str(item["presence"]))}
        for item in analysis["category_movement_summary"]
    ]
    category_movement_html = (
        "<section><h2>Variazione tra gli ultimi due periodi per categoria</h2>"
        + _html_table(
            category_movement_rows,
            (
                ("Categoria", "risk_category"),
                ("Periodo precedente", "prior_reference_month"),
                ("Periodo recente", "latest_reference_month"),
                ("Presenza", "presence_label"),
                ("Utilizzato precedente", "prior_used"),
                ("Utilizzato recente", "latest_used"),
                ("Variazione utilizzato", "used_change"),
                ("Sconfinamento precedente", "prior_overrun"),
                ("Sconfinamento recente", "latest_overrun"),
            ),
        )
        + "</section>"
        if category_movement_rows
        else ""
    )
    populations = "".join(
        (
            _html_population(
                analysis,
                title="Garanzie collegate alle esposizioni",
                key="guarantees",
                columns=(
                    ("Intermediario", "intermediary"),
                    ("Categoria", "risk_category"),
                    ("Tipo", "guarantee_type"),
                    ("Importo garantito", "guaranteed_amount"),
                ),
                coverage_key="guarantees_on_exposures",
            ),
            _html_population(
                analysis,
                title="Garanzie ricevute per obbligazioni di terzi",
                key="guarantees_received",
                columns=(
                    ("Mese", "reference_month"),
                    ("Intermediario", "intermediary"),
                    ("Garantito", "guaranteed_party"),
                    ("Tipo", "guarantee_type"),
                    ("Valore garanzia", "guarantee_value"),
                    ("Importo garantito", "guaranteed_amount"),
                ),
                coverage_key="guarantees_received",
            ),
            _html_population(
                analysis,
                title="Garanti dell'intestatario",
                key="guarantors",
                columns=(
                    ("Mese", "reference_month"),
                    ("Intermediario", "intermediary"),
                    ("Garante", "guarantor"),
                    ("Valore garanzia", "guarantee_value"),
                    ("Importo garantito", "guaranteed_amount"),
                ),
                coverage_key="guarantors",
            ),
            _html_population(
                analysis,
                title="Debitori ceduti",
                key="ceded_debtors",
                columns=(
                    ("Mese", "reference_month"),
                    ("Intermediario", "intermediary"),
                    ("Debitore ceduto", "ceded_debtor"),
                    ("Valore nominale", "nominal_value"),
                ),
                coverage_key="ceded_debtors",
            ),
            _html_population(
                analysis,
                title="Altre informazioni di rischio",
                key="other_risk_information",
                columns=(
                    ("Mese", "reference_month"),
                    ("Intermediario", "intermediary"),
                    ("Categoria", "category"),
                    ("Localizzazione", "location"),
                    ("Importo", "amount"),
                    ("Valore intrinseco", "intrinsic_value"),
                ),
                coverage_key="other_risk_information",
            ),
            _html_population(
                analysis,
                title="Prospetto sintetico",
                key="summary_totals",
                columns=(
                    ("Mese", "reference_month"),
                    ("Intermediario", "intermediary"),
                    ("Categoria", "summary_category"),
                    ("Accordato", "granted"),
                    ("Operativo", "operational_granted"),
                    ("Utilizzato", "used"),
                    ("Valore garanzia", "guarantee_value"),
                    ("Importo garantito", "guaranteed_amount"),
                    ("Valore intrinseco", "intrinsic_value"),
                ),
                coverage_key="summary_totals",
            ),
            _html_population(
                analysis,
                title="Sconfinamenti",
                key="overruns",
                columns=(
                    ("Intermediario", "intermediary"),
                    ("Categoria", "risk_category"),
                    ("Accordato operativo", "operational_granted"),
                    ("Utilizzato", "used"),
                    ("Sconfinamento", "overrun"),
                ),
            ),
            _html_population(
                analysis,
                title="Eventi inframensili",
                key="inframonthly_events",
                columns=(
                    ("Intermediario", "intermediary"),
                    ("Data", "event_date"),
                    ("Evento", "event_type"),
                    ("Cancellato", "event_cancelled"),
                ),
                coverage_key="inframonthly_events",
            ),
            _html_population(
                analysis,
                title="Richieste di informazione",
                key="information_requests",
                columns=(
                    ("Intermediario", "intermediary"),
                    ("Data", "request_date"),
                    ("Periodo richiesto", "requested_period"),
                    ("Tipo", "request_type"),
                    ("Causale", "request_reason"),
                ),
                coverage_key="information_requests",
            ),
            _html_population(
                analysis,
                title="Pregiudizievoli",
                key="prejudicial_events",
                columns=(
                    ("Intermediario", "intermediary"),
                    ("Categoria", "risk_category"),
                    ("Evidenza", "prejudicial_event"),
                ),
                coverage_key="pregiudizievoli",
            ),
        )
    )
    return f"""<!doctype html><html lang="it"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>Centrale Rischi — {html.escape(str(analysis['entity']))}</title><style>:root{{--navy:#002060;--blue:#006b8f;--ink:#171816;--muted:#5c6470;--rule:#d9dadd}}*{{box-sizing:border-box}}body{{font-family:'Instrument Sans',Arial,sans-serif;margin:0;color:var(--ink);background:#fff}}main{{max-width:1120px;margin:auto;padding:48px 24px 72px}}header{{border-top:5px solid var(--navy);border-bottom:1px solid #c9ccd1;padding:28px 0 24px}}.eyebrow{{color:var(--blue);font-weight:700;text-transform:uppercase;letter-spacing:.08em}}h1{{font-size:clamp(2.2rem,5vw,4.5rem);line-height:.98;margin:.4rem 0}}h2{{font-size:1.55rem;margin-bottom:14px}}h3{{font-size:1.05rem;margin-top:24px}}section{{margin-top:42px}}.table-wrap{{overflow-x:auto;border-top:1px solid var(--navy)}}table{{width:100%;border-collapse:collapse;min-width:680px}}th,td{{padding:12px 10px;border-bottom:1px solid var(--rule);text-align:left;vertical-align:top}}th{{color:var(--navy);font-size:.82rem;letter-spacing:.02em}}td.number{{font-variant-numeric:tabular-nums;text-align:right}}td small{{display:block;color:var(--muted);margin-top:4px;max-width:44ch}}.status{{display:inline-block;border:1px solid var(--navy);padding:6px 10px}}.empty{{color:var(--muted);border-top:1px solid var(--rule);padding-top:14px}}code{{font-size:.8em;color:var(--blue)}}details.evidence{{margin:.35rem 0 .8rem;color:var(--muted)}}details.evidence summary{{cursor:pointer;font-size:.82rem}}details.evidence code{{display:block;margin-top:.25rem;overflow-wrap:anywhere}}@media(max-width:700px){{main{{padding:28px 16px 52px}}table{{font-size:.82rem}}}}</style></head><body><main><header><p class="eyebrow">Vera · Centrale Rischi</p><h1>{html.escape(str(analysis['entity']))}</h1><p>Periodo più recente: {html.escape(str(analysis['latest_reference_month']))} · <span class="status">{html.escape(_review_status_label(str(analysis['review_status'])))}</span></p></header><section><h2>KPI</h2><div class="table-wrap"><table><thead><tr><th>Metrica</th><th>Valore</th><th>Unità</th><th>Copertura</th></tr></thead><tbody>{_metric_rows(analysis)}</tbody></table></div></section>{commentary_html}<section><h2>Esposizioni per durata originaria</h2>{_html_table(original_term_rows, amount_columns)}</section><section><h2>Esposizioni per durata residua</h2>{_html_table(residual_term_rows, amount_columns)}</section><section><h2>Indicatori per categoria</h2>{_html_table(analysis['risk_category_summary'], (("Categoria", "risk_category"), ("Operativo", "operational_granted"), ("Utilizzato", "used"), ("Margine CR calcolato", "available"), ("Utilizzo %", "utilization_pct")))}</section>{category_movement_html}{populations}<section><h2>Limiti</h2><ul>{''.join(f'<li>{html.escape(item)}</li>' for item in analysis['limitations'])}</ul></section></main></body></html>"""


def write_excel(path: Path, analysis: Mapping[str, Any]) -> None:
    """Write a reviewable workbook from the exact analysis payload."""

    def excel_number(value: Any) -> int | float | None:
        if value in (None, ""):
            return None
        try:
            number = Decimal(str(value))
        except InvalidOperation:
            return None
        if number == number.to_integral_value():
            return int(number)
        return float(number)

    amount_fields = {
        "granted",
        "operational_granted",
        "used",
        "available",
        "overrun",
        "guaranteed_amount",
        "guarantee_value",
        "nominal_value",
        "amount",
        "intrinsic_value",
        "average_balance",
        "expected",
        "actual",
        "difference",
        "prior_granted",
        "latest_granted",
        "granted_change",
        "prior_operational_granted",
        "latest_operational_granted",
        "operational_granted_change",
        "prior_used",
        "latest_used",
        "used_change",
        "prior_available",
        "latest_available",
        "available_change",
        "prior_overrun",
        "latest_overrun",
        "overrun_change",
    }
    percentage_fields = {"utilization_pct"}
    numeric_fields = amount_fields | percentage_fields
    hidden_audit_fields = {
        "source_row",
        "source_region",
        "source_document_sha256",
    }

    def project_rows(
        rows: Sequence[Mapping[str, Any]], fields: Sequence[str]
    ) -> list[dict[str, Any]]:
        return [
            {field: row.get(field) for field in fields if field in row} for row in rows
        ]

    workbook = Workbook()
    summary = workbook.active
    summary.title = "KPI"
    summary.append(
        ("Metric ID", "Metrica", "Valore", "Unità", "Disponibilità", "Motivo")
    )
    for item in analysis["metrics"]:
        metric_value = (
            excel_number(item["value"]) if item["availability"] == "available" else None
        )
        summary.append(
            (
                item["metric_id"],
                item["label"],
                metric_value,
                item["unit"],
                item["availability"],
                item["reason"] or "",
            )
        )
    sections = {
        "Durata originaria": analysis["original_term_summary"],
        "Durata residua": analysis["residual_term_summary"],
        "Categorie": analysis["risk_category_summary"],
        "Variazione categorie": analysis["category_movement_summary"],
        "Esposizioni": analysis["exposures"],
        "Garanzie": project_rows(
            analysis["guarantees"],
            (
                "reference_month",
                "intermediary",
                "risk_category",
                "guarantee_type",
                "guaranteed_amount",
                "record_status",
                "source_page",
                "source_row_locator",
                "extraction_confidence",
            ),
        ),
        "Garanzie ricevute": analysis["guarantees_received"],
        "Garanti intestatario": analysis["guarantors"],
        "Debitori ceduti": analysis["ceded_debtors"],
        "Altre informazioni": analysis["other_risk_information"],
        "Prospetto sintetico": analysis["summary_totals"],
        "Sconfinamenti": project_rows(
            analysis["overruns"],
            (
                "reference_month",
                "intermediary",
                "risk_category",
                "operational_granted",
                "used",
                "overrun",
                "record_status",
                "source_page",
                "source_row_locator",
                "extraction_confidence",
            ),
        ),
        "Eventi inframensili": analysis["inframonthly_events"],
        "Richieste informazioni": analysis["information_requests"],
        "Pregiudizievoli": project_rows(
            analysis["prejudicial_events"],
            (
                "reference_month",
                "intermediary",
                "risk_category",
                "prejudicial_event",
                "record_status",
                "source_page",
                "source_row_locator",
                "extraction_confidence",
            ),
        ),
        "Serie mensile": analysis["monthly_series"],
        "Controlli": analysis["controls"],
    }
    for title, rows in sections.items():
        sheet = workbook.create_sheet(title)
        if rows:
            headers = list(rows[0])
            sheet.append(headers)
            for row in rows:
                sheet.append(
                    [
                        (
                            excel_number(row.get(header))
                            if header in numeric_fields
                            else row.get(header)
                        )
                        for header in headers
                    ]
                )
        else:
            sheet.append(("Stato", "Nessun dato disponibile"))

    navy_fill = PatternFill("solid", fgColor="002060")
    alternate_fill = PatternFill("solid", fgColor="F3F6FA")
    bottom_rule = Border(bottom=Side(style="thin", color="D9DADD"))
    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "D2" if sheet.max_column > 10 else "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.row_dimensions[1].height = 30
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = navy_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        headers = {
            str(cell.value): cell.column for cell in sheet[1] if cell.value is not None
        }
        for row_number in range(2, sheet.max_row + 1):
            sheet.row_dimensions[row_number].height = 30
            for cell in sheet[row_number]:
                cell.alignment = Alignment(
                    horizontal=(
                        "right"
                        if str(sheet.cell(1, cell.column).value) in numeric_fields
                        else "left"
                    ),
                    vertical="top",
                    wrap_text=True,
                )
                cell.border = bottom_rule
                if row_number % 2 == 0:
                    cell.fill = alternate_fill
        if sheet.title == "KPI":
            for row_number in range(2, sheet.max_row + 1):
                value_cell = sheet.cell(row_number, 3)
                unit = str(sheet.cell(row_number, 4).value or "")
                value_cell.alignment = Alignment(horizontal="right", vertical="top")
                value_cell.number_format = "0.00" if unit == "percent" else "#,##0"
        else:
            for header in amount_fields:
                if header not in headers:
                    continue
                for cell in sheet.iter_rows(
                    min_row=2,
                    min_col=headers[header],
                    max_col=headers[header],
                ):
                    cell[0].number_format = "#,##0"
            for header in percentage_fields:
                if header not in headers:
                    continue
                for cell in sheet.iter_rows(
                    min_row=2,
                    min_col=headers[header],
                    max_col=headers[header],
                ):
                    cell[0].number_format = "0.00"
        for column_cells in sheet.columns:
            header = str(column_cells[0].value or "")
            width = min(
                48 if sheet.title == "KPI" and header in {"Metrica", "Motivo"} else 28,
                max(10, max(len(str(cell.value or "")) for cell in column_cells) + 2),
            )
            column = sheet.column_dimensions[column_cells[0].column_letter]
            column.width = width
            if header in hidden_audit_fields:
                column.hidden = True
    workbook.save(path)
