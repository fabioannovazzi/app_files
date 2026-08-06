#!/usr/bin/env python3
"""Deterministic case engine for Italian OIC annual accounts and XBRL drafts.

The module deliberately limits deterministic logic to mechanically verifiable
work: exact Decimal arithmetic, explicit rule-pack conditions, provenance,
revision hashes, workflow gates, and XML construction. Account meaning,
materiality, legal interpretation, and professional approval remain explicit
review decisions.
"""

from __future__ import annotations

import argparse
import base64
import binascii
import csv
import hashlib
import html
import json
import logging
import math
import re
import shutil
import sys
import zipfile
from copy import deepcopy
from dataclasses import dataclass
from datetime import UTC, date, datetime, timedelta
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from enum import StrEnum
from pathlib import Path
from typing import Any, Callable, Iterable, Mapping, Sequence

from client_history import (
    client_history_suggestions,
    remember_approved_client_history,
)
from defusedxml.ElementTree import fromstring
from disclosure_engine import (
    build_disclosure_coverage,
    disclosure_answer_complete,
    disclosure_rule_pack_hash,
    manual_disclosure_flags,
    narrative_redline,
    normalize_narrative_blocks,
    note_outline,
    prior_narrative_suggestions,
)
from external_validation import record_external_validation_result
from intelligence_contract import (
    IntelligenceTask,
    build_intelligence_packet,
    intelligence_packet_hash,
    validate_intelligence_output,
)
from lxml import etree
from mapping_memory import mapping_candidates, remember_approved_mappings
from prior_xbrl import parse_prior_xbrl
from schedule_engine import (
    SCHEDULE_TYPES,
    normalize_schedule,
    required_schedule_types,
    schedule_fact_records,
    schedule_template_fields,
    schedule_template_text_fields,
)
from schedule_taxonomy_adapter import compile_schedule_taxonomy_adapter
from statutory_presentation import (
    build_primary_presentation_inventory,
    build_statutory_presentation_coverage,
)
from validate_xbrl import validate_instance

__all__ = [
    "CaseState",
    "EvidenceStatus",
    "ParserConvention",
    "ValidationIssue",
    "approve_case",
    "archive_case",
    "attach_supporting_document",
    "activate_disclosures",
    "apply_mapping_decisions",
    "build_statements",
    "create_case",
    "determine_forms",
    "export_case",
    "generate_mapping_candidates",
    "ingest_prior_xbrl",
    "ingest_trial_balance",
    "ingest_schedule_file",
    "load_client_history",
    "migrate_regulatory_versions",
    "record_intelligence_suggestion",
    "record_issue_reviews",
    "record_adjustments",
    "record_taxonomy_facts",
    "record_taxonomy_catalogue_build",
    "record_statutory_presentation",
    "record_taxonomy_mapping_index",
    "record_taxonomy_representation",
    "record_micro_reporting",
    "load_case",
    "normalize_decimal",
    "prepare_xbrl_review",
    "record_schedule",
    "record_schedule_taxonomy_adapter",
    "record_disclosure_trigger_decisions",
    "record_comparative_reconciliation_decisions",
    "record_narrative_blocks",
    "record_external_validation",
    "record_file_security_scan",
    "record_artifact_access",
    "remember_client_history",
    "remember_mappings",
    "render_xbrl",
    "save_case",
    "validate_case",
]

LOGGER = logging.getLogger(__name__)
CASE_FILE = "case.json"
CASE_CHECKSUM_FILE = "case.json.sha256"
SCHEMA_VERSION = 1
XML_NS = "http://www.w3.org/XML/1998/namespace"
XBRLI_NS = "http://www.xbrl.org/2003/instance"
LINK_NS = "http://www.xbrl.org/2003/linkbase"
XLINK_NS = "http://www.w3.org/1999/xlink"
ISO4217_NS = "http://www.xbrl.org/2003/iso4217"
XBRLDI_NS = "http://xbrl.org/2006/xbrldi"
XSI_NS = "http://www.w3.org/2001/XMLSchema-instance"
SAFE_ID = re.compile(r"^[A-Za-z0-9][A-Za-z0-9._-]{0,127}$")
SAFE_QNAME = re.compile(r"^[A-Za-z_][A-Za-z0-9_.-]*:[A-Za-z_][A-Za-z0-9_.-]*$")
XBRL_DECIMAL_LEXICAL = re.compile(r"^[+-]?(?:\d+(?:\.\d*)?|\.\d+)$")
NORMALIZED_MONEY_LEXICAL = re.compile(
    r"^(?P<sign>[+-]?)(?P<integer>\d+)(?:\.(?P<fraction>\d+))?$"
)
MONEY_LITERAL = re.compile(
    r"(?:€|euro)\s*(\(?-?(?:\d{1,3}(?:[.\s]\d{3})+|\d+)(?:[,.]\d+)?\)?)",
    flags=re.IGNORECASE,
)
EXPORTABLE_STATUSES = {"OBSERVED", "DERIVED", "USER_CONFIRMED"}
ANNUAL_NEGATIVE_CONFIRMATION_KEYS = {
    "guarantees_and_commitments",
    "contingent_liabilities",
    "related_party_transactions",
    "off_balance_sheet_arrangements",
    "derivatives",
    "post_closing_events",
    "accounting_policy_changes",
    "prior_period_errors",
    "going_concern_uncertainties",
    "non_market_transactions",
    "double_format_events",
}
SUPPORTED_LEGAL_FORMS = {"SRL", "SPA", "SAPA"}
MAX_INPUT_BYTES = 100 * 1024 * 1024
MAX_XLSX_UNCOMPRESSED_BYTES = 250 * 1024 * 1024
MAX_TEMPLATE_ROWS = 20_000
MAX_MONEY_INTEGER_DIGITS = 18
MAX_MONEY_FRACTION_DIGITS = 6


class CaseState(StrEnum):
    """Supported workflow states from the product specification."""

    DRAFT = "DRAFT"
    INGESTING = "INGESTING"
    INPUT_REVIEW = "INPUT_REVIEW"
    MAPPING_REVIEW = "MAPPING_REVIEW"
    STATEMENT_REVIEW = "STATEMENT_REVIEW"
    DATA_GAPS = "DATA_GAPS"
    NOTE_DRAFT = "NOTE_DRAFT"
    VALIDATION_FAILED = "VALIDATION_FAILED"
    READY_FOR_REVIEW = "READY_FOR_REVIEW"
    APPROVED = "APPROVED"
    EXPORTED = "EXPORTED"
    ARCHIVED = "ARCHIVED"
    UNSUPPORTED = "UNSUPPORTED"


class EvidenceStatus(StrEnum):
    """Evidence states preserved by every canonical and XBRL fact."""

    OBSERVED = "OBSERVED"
    DERIVED = "DERIVED"
    USER_CONFIRMED = "USER_CONFIRMED"
    MODEL_SUGGESTED = "MODEL_SUGGESTED"
    ASSUMED = "ASSUMED"
    MISSING = "MISSING"
    NOT_APPLICABLE = "NOT_APPLICABLE"


class ParserConvention(StrEnum):
    """Supported meanings of debit and credit progressives."""

    TURNOVER_EXCLUDES_OPENING = "TURNOVER_EXCLUDES_OPENING"
    TURNOVER_INCLUDES_OPENING = "TURNOVER_INCLUDES_OPENING"
    TURNOVER_INCLUDES_CLOSING_ENTRIES = "TURNOVER_INCLUDES_CLOSING_ENTRIES"
    SIGNED_BALANCE_ONLY = "SIGNED_BALANCE_ONLY"
    UNKNOWN = "UNKNOWN"


@dataclass(frozen=True)
class ValidationIssue:
    """One deterministic validation result."""

    issue_id: str
    severity: str
    rule_id: str
    message: str
    affected_facts: tuple[str, ...] = ()
    source_refs: tuple[str, ...] = ()
    override_allowed: bool = False

    def as_dict(self) -> dict[str, Any]:
        """Return a JSON-compatible issue object."""

        payload = {
            "issue_id": self.issue_id,
            "severity": self.severity,
            "rule_id": self.rule_id,
            "message": self.message,
            "affected_facts": list(self.affected_facts),
            "source_refs": list(self.source_refs),
            "override_allowed": self.override_allowed,
        }
        payload["fingerprint"] = _issue_fingerprint(payload)
        return payload


def _now() -> str:
    return datetime.now(tz=UTC).replace(microsecond=0).isoformat()


def _canonical_json(value: Any) -> bytes:
    return json.dumps(
        value, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    ).encode("utf-8")


def _sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def _issue_fingerprint(issue: Mapping[str, Any]) -> str:
    """Bind a review decision to the exact issue content it considered."""

    return _sha256_bytes(
        _canonical_json(
            {
                "severity": issue["severity"],
                "rule_id": issue["rule_id"],
                "message": issue["message"],
                "affected_facts": list(issue.get("affected_facts", [])),
                "source_refs": list(issue.get("source_refs", [])),
            }
        )
    )


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _decimal_text(value: Decimal) -> str:
    _validate_monetary_decimal(value)
    text = format(value, "f")
    if value == value.to_integral():
        return f"{text.split('.', maxsplit=1)[0]}.00"
    return text


def _validate_monetary_decimal(value: Decimal) -> Decimal:
    """Reject non-finite or operationally unbounded monetary values."""

    if not value.is_finite():
        raise ValueError("Monetary values must be finite")
    if value.is_zero():
        integer_digits = 1
    else:
        integer_digits = max(value.adjusted() + 1, 1)
    fractional_digits = max(-value.as_tuple().exponent, 0)
    if integer_digits > MAX_MONEY_INTEGER_DIGITS:
        raise ValueError(
            f"Monetary values support at most {MAX_MONEY_INTEGER_DIGITS} integer digits"
        )
    if fractional_digits > MAX_MONEY_FRACTION_DIGITS:
        raise ValueError(
            f"Monetary values support at most {MAX_MONEY_FRACTION_DIGITS} fractional digits"
        )
    return value


def _decimal_from_normalized_text(text: str, raw: Any, label: str) -> Decimal:
    """Parse one bounded, non-exponent monetary lexical representation."""

    descriptor = f"{label} monetary".strip()
    match = NORMALIZED_MONEY_LEXICAL.fullmatch(text)
    if not match:
        raise ValueError(f"Invalid {descriptor} value: {raw!r}")
    integer = match.group("integer").lstrip("0") or "0"
    fraction = match.group("fraction") or ""
    if len(integer) > MAX_MONEY_INTEGER_DIGITS:
        raise ValueError(
            f"{descriptor.capitalize()} values support at most "
            f"{MAX_MONEY_INTEGER_DIGITS} integer digits"
        )
    if len(fraction) > MAX_MONEY_FRACTION_DIGITS:
        raise ValueError(
            f"{descriptor.capitalize()} values support at most "
            f"{MAX_MONEY_FRACTION_DIGITS} fractional digits"
        )
    try:
        return _validate_monetary_decimal(Decimal(text))
    except InvalidOperation as exc:
        raise ValueError(f"Invalid {descriptor} value: {raw!r}") from exc


def _reported_decimal(value: Decimal, precision: int) -> Decimal:
    """Round a presentation value without changing canonical precision."""

    quantum = Decimal(1).scaleb(-precision)
    return value.quantize(quantum, rounding=ROUND_HALF_UP)


def _previous_year_date(value: date) -> date:
    """Return the corresponding prior-year date, including a leap-day fallback."""

    try:
        return value.replace(year=value.year - 1)
    except ValueError:
        if value.month == 2 and value.day == 29:
            return value.replace(year=value.year - 1, day=28)
        raise


def _xbrl_fact_id(kind: str, stable_id: object, context_ref: str) -> str:
    """Build a stable XML NCName for one rendered fact occurrence."""

    raw = f"fact_{kind}_{stable_id}_{context_ref}"
    normalized = re.sub(r"[^A-Za-z0-9_.-]+", "_", raw).strip("_.-")
    if not normalized or not re.match(r"[A-Za-z_]", normalized):
        normalized = f"fact_{normalized}"
    return normalized


def normalize_decimal(raw: Any) -> Decimal:
    """Parse locale-neutral or Italian-formatted money as an exact Decimal."""

    if isinstance(raw, Decimal):
        return _validate_monetary_decimal(raw)
    if isinstance(raw, bool) or raw is None:
        raise ValueError("A monetary value is required")
    if isinstance(raw, int):
        return _validate_monetary_decimal(Decimal(raw))
    if isinstance(raw, float) and not math.isfinite(raw):
        raise ValueError("Monetary values must be finite")
    text = str(raw).strip().replace("\u00a0", "").replace(" ", "")
    if not text:
        raise ValueError("A monetary value is required; blank is not zero")
    has_opening_parenthesis = text.startswith("(")
    has_closing_parenthesis = text.endswith(")")
    if has_opening_parenthesis != has_closing_parenthesis:
        raise ValueError(f"Invalid monetary value: {raw!r}")
    negative = has_opening_parenthesis and has_closing_parenthesis
    if negative:
        text = text[1:-1]
        if text.startswith(("+", "-")):
            raise ValueError(
                f"Invalid monetary value with both parentheses and sign: {raw!r}"
            )
    text = re.sub(r"^(EUR|€)", "", text, flags=re.IGNORECASE)
    text = re.sub(r"(EUR|€)$", "", text, flags=re.IGNORECASE)
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(".", "").replace(",", ".")
    value = _decimal_from_normalized_text(text, raw, "")
    return -value if negative else value


def _normalize_narrative_money_literal(raw: str, language: str) -> Decimal:
    """Parse a prose monetary literal using the approved output locale."""

    text = raw.strip().replace("\u00a0", "").replace(" ", "")
    has_opening_parenthesis = text.startswith("(")
    has_closing_parenthesis = text.endswith(")")
    if has_opening_parenthesis != has_closing_parenthesis:
        raise ValueError(f"Invalid narrative monetary value: {raw!r}")
    negative = has_opening_parenthesis and has_closing_parenthesis
    if negative:
        text = text[1:-1]
        if text.startswith(("+", "-")):
            raise ValueError(
                "Invalid narrative monetary value with both parentheses and sign: "
                f"{raw!r}"
            )
    if language == "it":
        if re.fullmatch(r"-?\d{1,3}(?:\.\d{3})+", text):
            text = text.replace(".", "")
        elif "," in text:
            text = text.replace(".", "").replace(",", ".")
    elif language == "en":
        if re.fullmatch(r"-?\d{1,3}(?:,\d{3})+", text):
            text = text.replace(",", "")
        elif "." in text:
            text = text.replace(",", "")
    else:
        raise ValueError("Narrative monetary parsing supports Italian or English")
    value = _decimal_from_normalized_text(text, raw, "narrative")
    return -value if negative else value


def _case_payload_for_hash(case: Mapping[str, Any]) -> dict[str, Any]:
    payload = deepcopy(dict(case))
    payload.pop("_pending_before_hash", None)
    payload.pop("approval", None)
    payload.pop("approval_snapshots", None)
    payload.pop("audit_events", None)
    payload.pop("updated_at", None)
    return payload


def _manifest_hash(case: Mapping[str, Any]) -> str:
    documents = [
        {"document_id": item["document_id"], "sha256": item["sha256"]}
        for item in case.get("source_documents", [])
    ]
    return _sha256_bytes(_canonical_json(documents))


def _available_evidence_refs(case: Mapping[str, Any]) -> set[str]:
    """Return evidence identifiers already anchored in the current case."""

    refs = {
        str(document["document_id"]) for document in case.get("source_documents", [])
    }
    refs.update(
        str(ref)
        for document in case.get("source_documents", [])
        for ref in document.get("source_refs", [])
    )
    refs.update(
        str(anchor["source_ref"])
        for anchor in (case.get("trial_balance") or {}).get("source_anchors", [])
    )
    refs.update(
        str((fact.get("source_anchor") or {}).get("source_ref"))
        for fact in (case.get("prior_xbrl") or {}).get("facts", [])
        if (fact.get("source_anchor") or {}).get("source_ref")
    )
    refs.update(
        str(fact["fact_id"])
        for fact in [
            *case.get("canonical_facts", []),
            *case.get("taxonomy_facts", []),
            *case.get("schedule_taxonomy_facts", []),
        ]
    )
    refs.update(
        str(ref)
        for schedule in case.get("schedules", [])
        for row in schedule.get("rows", [])
        for ref in row.get("source_refs", [])
    )
    refs.update(
        str(ref)
        for answer in case.get("disclosure_answers", [])
        for ref in answer.get("source_refs", [])
    )
    return refs


def _computation_context(
    case: Mapping[str, Any],
    template_version: str,
    *,
    model_version: str | None = None,
) -> dict[str, Any]:
    """Bind a derived output to its exact inputs, decisions, and versions."""

    return {
        "case_id": case["case_id"],
        "revision_id": case["revision_id"],
        "input_manifest_hash": _manifest_hash(case),
        "mapping_version": _sha256_bytes(_canonical_json(case.get("mappings", []))),
        "rule_pack_versions": deepcopy(dict(case["rule_pack_versions"])),
        "taxonomy_checksum": case.get("taxonomy_checksum"),
        "model_version": model_version,
        "template_version": template_version,
        "computed_at": _now(),
    }


def _review_content_hash(case: Mapping[str, Any]) -> str:
    """Hash the substantive review content independently of workflow metadata."""

    payload = deepcopy(dict(case))
    payload.pop("_pending_before_hash", None)
    for key in (
        "approval",
        "approval_snapshots",
        "artifacts",
        "audit_events",
        "external_validation",
        "preview",
        "xbrl_review",
        "revision_id",
        "state",
        "updated_at",
        "validation",
    ):
        payload.pop(key, None)
    return _sha256_bytes(_canonical_json(payload))


def _reviewed_preview_bytes(preview: Mapping[str, Any]) -> bytes:
    """Decode and verify the exact HTML bytes shown during professional review."""

    encoded = preview.get("content_base64")
    if not isinstance(encoded, str) or not encoded:
        raise ValueError("The reviewed preview does not contain bound HTML bytes")
    try:
        content = base64.b64decode(encoded, validate=True)
    except (binascii.Error, ValueError) as exc:
        raise ValueError(
            "The reviewed preview contains invalid base64 content"
        ) from exc
    if len(content) != preview.get("size_bytes"):
        raise ValueError("The reviewed preview byte length does not match its receipt")
    if _sha256_bytes(content) != preview.get("sha256"):
        raise ValueError("The reviewed preview checksum does not match its receipt")
    return content


def _next_document_id(case: Mapping[str, Any]) -> str:
    numbers = [
        int(match.group(1))
        for item in case.get("source_documents", [])
        if (match := re.fullmatch(r"doc_(\d+)", str(item.get("document_id", ""))))
    ]
    return f"doc_{(max(numbers, default=0) + 1):04d}"


def _record_event(
    case: dict[str, Any],
    action: str,
    actor: str,
    details: Mapping[str, Any] | None = None,
    *,
    after_hash: str | None = None,
) -> None:
    before_hash = case.pop("_pending_before_hash", None)
    resolved_after_hash = after_hash or _sha256_bytes(
        _canonical_json(_case_payload_for_hash(case))
    )
    case.setdefault("audit_events", []).append(
        {
            "event_id": f"evt_{len(case.get('audit_events', [])) + 1:06d}",
            "action": action,
            "actor": actor,
            "at": _now(),
            "case_id": case["case_id"],
            "tenant_id": case["tenant_id"],
            "revision_id": case["revision_id"],
            "details": dict(details or {}),
            "before_hash": before_hash,
            "after_hash": resolved_after_hash,
            "originating_interface": "DIRECT_LIBRARY",
        }
    )


def _record_evidence_attached(
    case: dict[str, Any], actor: str, document: Mapping[str, Any]
) -> None:
    """Record the specification event for checksum-bound case evidence."""

    _record_event(
        case,
        "evidence_attached",
        actor,
        {
            "document_id": document["document_id"],
            "purpose": document["purpose"],
            "sha256": document["sha256"],
        },
    )


def _mapping_audit_summary(value: Mapping[str, Any] | None) -> dict[str, Any] | None:
    """Return a bounded mapping-change summary plus its exact content hash."""

    if value is None:
        return None
    allocations = list(value.get("allocations", []))
    return {
        "decision": value.get("decision"),
        "candidate_source": value.get("candidate_source"),
        "canonical_lines": sorted(
            {str(item.get("canonical_line")) for item in allocations}
        ),
        "xbrl_concepts": sorted(
            {
                str(item["xbrl_concept"])
                for item in allocations
                if item.get("xbrl_concept")
            }
        ),
        "content_sha256": _sha256_bytes(_canonical_json(value)),
    }


def _mutate(case: dict[str, Any], actor: str, action: str) -> None:
    before_hash = _sha256_bytes(_canonical_json(_case_payload_for_hash(case)))
    prior_approval = case.get("approval")
    if prior_approval:
        archived = deepcopy(prior_approval)
        archived["invalidated_at"] = _now()
        archived["invalidated_by_action"] = action
        case.setdefault("approval_snapshots", []).append(archived)
        case["approval"] = None
        _record_event(case, "snapshot_invalidated", actor, {"reason": action})
    review_metadata_actions = {
        "preview_rendered",
        "validation_run",
        "xbrl_review_prepared",
    }
    if case.get("preview") and action not in review_metadata_actions:
        case["preview"] = None
        _record_event(case, "preview_invalidated", actor, {"reason": action})
    if case.get("xbrl_review") and action not in review_metadata_actions:
        case["xbrl_review"] = None
        _record_event(case, "xbrl_review_invalidated", actor, {"reason": action})
    revision_number = int(str(case["revision_id"]).split("_")[-1]) + 1
    case["revision_id"] = f"rev_{revision_number}"
    case["updated_at"] = _now()
    case["_pending_before_hash"] = before_hash


def _clear_accounting_dependent_reviews(
    case: dict[str, Any], *, preserve_adjustments: bool = False
) -> None:
    """Invalidate reviewed outputs whose meaning depends on accounting amounts."""

    replacements: dict[str, Any] = {
        "adjustments": [],
        "comparative_reconciliation_decisions": [],
        "canonical_facts": [],
        "statements": None,
        "taxonomy_facts": [],
        "taxonomy_fact_context": None,
        "schedule_taxonomy_adapter": None,
        "schedule_taxonomy_facts": [],
        "statutory_presentation": None,
        "taxonomy_representation": None,
        "micro_reporting": None,
        "schedules": [],
        "disclosure_answers": [],
        "disclosure_trigger_flags": [],
        "disclosure_trigger_decisions": [],
        "disclosure_coverage": None,
        "questionnaire": [],
        "note_outline": [],
        "note_outline_context": None,
        "narrative_blocks": [],
        "narrative_context": None,
        "model_mapping_suggestions": [],
        "disclosure_activation_suggestions": [],
        "narrative_suggestions": [],
        "review_decisions": [],
        "validation": None,
    }
    for key, value in replacements.items():
        if key == "adjustments" and preserve_adjustments:
            continue
        case[key] = value


def _clear_narrative_reviews(case: dict[str, Any]) -> None:
    """Invalidate accepted prose after any supporting structured fact changes."""

    case["narrative_blocks"] = []
    case["narrative_context"] = None
    case["narrative_suggestions"] = []
    case["review_decisions"] = []
    case["validation"] = None


def _clear_schedule_taxonomy_adapter(case: dict[str, Any]) -> None:
    """Invalidate note-table output after any adapter input changes."""

    case["schedule_taxonomy_adapter"] = None
    case["schedule_taxonomy_facts"] = []
    case["validation"] = None


def _ensure_revision(case: Mapping[str, Any], expected_revision: str) -> None:
    if case["revision_id"] != expected_revision:
        raise ValueError(
            f"Stale revision: expected {expected_revision}, current {case['revision_id']}"
        )


def load_case(case_dir: Path) -> dict[str, Any]:
    """Load a case only after its durable JSON checksum is verified."""

    if case_dir.is_symlink():
        raise ValueError("Case directory must not be a symbolic link")
    path = case_dir.resolve() / CASE_FILE
    checksum_path = case_dir.resolve() / CASE_CHECKSUM_FILE
    if path.is_symlink() or checksum_path.is_symlink():
        raise ValueError("Case record and checksum must not be symbolic links")
    if not path.is_file():
        raise FileNotFoundError(f"Case file not found: {path}")
    if not checksum_path.is_file():
        raise ValueError("Case checksum metadata is missing")
    record = path.read_bytes()
    expected_checksum = checksum_path.read_text(encoding="ascii").strip()
    if len(expected_checksum) != 64 or any(
        character not in "0123456789abcdef" for character in expected_checksum
    ):
        raise ValueError("Case checksum metadata is invalid")
    if _sha256_bytes(record) != expected_checksum:
        raise ValueError("Case record integrity verification failed")
    payload = json.loads(record)
    if payload.get("schema_version") != SCHEMA_VERSION:
        raise ValueError("Unsupported case schema version")
    return payload


def save_case(case_dir: Path, case: Mapping[str, Any]) -> Path:
    """Atomically persist a case and checksum without following symlinks."""

    if case_dir.is_symlink():
        raise ValueError("Case directory must not be a symbolic link")
    resolved = case_dir.resolve()
    resolved.mkdir(parents=True, exist_ok=True)
    target = resolved / CASE_FILE
    checksum_target = resolved / CASE_CHECKSUM_FILE
    if target.is_symlink() or checksum_target.is_symlink():
        raise ValueError("Refusing to write a case or checksum through a symbolic link")
    temporary = resolved / f".{CASE_FILE}.tmp"
    checksum_temporary = resolved / f".{CASE_CHECKSUM_FILE}.tmp"
    if temporary.is_symlink() or checksum_temporary.is_symlink():
        raise ValueError("Refusing to write a case through a temporary symbolic link")
    record = _canonical_json(case) + b"\n"
    digest = _sha256_bytes(record)
    temporary.write_bytes(record)
    checksum_temporary.write_text(digest + "\n", encoding="ascii")
    temporary.replace(target)
    checksum_temporary.replace(checksum_target)
    if _sha256_file(target) != digest:
        raise OSError("Persisted case checksum verification failed")
    return target


def _validate_entity(entity: Mapping[str, Any]) -> list[str]:
    reasons: list[str] = []
    if str(entity.get("legal_form", "")).upper() not in SUPPORTED_LEGAL_FORMS:
        reasons.append("UNSUPPORTED_LEGAL_FORM")
    if str(entity.get("accounting_framework", "")).upper() != "OIC":
        reasons.append("UNSUPPORTED_ACCOUNTING_FRAMEWORK")
    if entity.get("listed") is not False:
        reasons.append("LISTED_OR_UNCONFIRMED")
    if entity.get("regulated_sector") is not False:
        reasons.append("REGULATED_SECTOR_OR_UNCONFIRMED")
    if entity.get("consolidated") is not False:
        reasons.append("CONSOLIDATED_OR_UNCONFIRMED")
    if entity.get("final_liquidation") is not False:
        reasons.append("FINAL_LIQUIDATION_OR_UNCONFIRMED")
    return reasons


def _validate_required_entity_profile(entity: Mapping[str, Any]) -> None:
    """Reject incomplete identity/profile data required by every MVP case."""

    missing = [
        key
        for key in ("legal_name", "tax_identifier", "registered_office")
        if not str(entity.get(key, "")).strip()
    ]
    if missing:
        raise ValueError(
            "Entity profile is missing required fields: " + ", ".join(missing)
        )
    if not isinstance(entity.get("first_financial_year"), bool):
        raise ValueError(
            "Entity profile requires an explicit first-financial-year flag"
        )
    if entity["first_financial_year"] is False and str(
        entity.get("prior_year_form", "")
    ).upper() not in {"ORDINARY", "ABBREVIATED", "MICRO"}:
        raise ValueError("A non-first-year case requires the prior-year statutory form")
    micro_exclusions = entity.get("micro_exclusion_flags")
    if not isinstance(micro_exclusions, list) or any(
        not isinstance(item, str) or not item.strip() for item in micro_exclusions
    ):
        raise ValueError(
            "Entity profile requires reviewed micro-exclusion flags, including an empty list"
        )


def _taxonomy_output_contracts(rule_pack: Mapping[str, Any]) -> dict[str, Any]:
    """Validate and copy versioned non-statement taxonomy output mappings."""

    raw_contracts = rule_pack.get("taxonomy_output_contracts")
    if not isinstance(raw_contracts, Mapping):
        raise ValueError("Statutory rule pack requires taxonomy output contracts")
    micro = raw_contracts.get("MICRO_FOOTER_TEXT")
    if not isinstance(micro, Mapping) or not SAFE_QNAME.fullmatch(
        str(micro.get("xbrl_concept", ""))
    ):
        raise ValueError("Statutory rule pack requires a micro-footer XBRL concept")
    labels = micro.get("labels")
    expected_keys = {
        "guarantees_commitments_contingencies",
        "director_auditor_compensation",
        "own_and_parent_shares",
    }
    if not isinstance(labels, Mapping) or any(
        not isinstance(labels.get(language), Mapping)
        or set(labels[language]) != expected_keys
        or any(not str(value).strip() for value in labels[language].values())
        for language in ("it", "en")
    ):
        raise ValueError("Micro-footer output labels are incomplete")
    not_applicable = micro.get("not_applicable_text")
    if not isinstance(not_applicable, Mapping) or any(
        not str(not_applicable.get(language, "")).strip() for language in ("it", "en")
    ):
        raise ValueError("Micro-footer negative-confirmation text is incomplete")
    return deepcopy(dict(raw_contracts))


def create_case(
    case_dir: Path,
    payload: Mapping[str, Any],
    rule_pack: Mapping[str, Any],
    actor: str,
) -> dict[str, Any]:
    """Create a scope-checked case locked to explicit rule and taxonomy versions."""

    entity = dict(payload.get("entity", {}))
    period = dict(payload.get("period", {}))
    case_id = str(payload.get("case_id") or "")
    if not SAFE_ID.fullmatch(case_id):
        raise ValueError("case_id must be a safe stable identifier")
    start = date.fromisoformat(str(period["start"]))
    end = date.fromisoformat(str(period["end"]))
    if start > end:
        raise ValueError("Reporting period start must not follow its end")
    prior_start_raw = str(entity.get("prior_period_start") or "")
    prior_end_raw = str(entity.get("prior_period_end") or "")
    if entity.get("first_financial_year") is True and (
        prior_start_raw or prior_end_raw
    ):
        raise ValueError("A first financial year cannot declare a comparative period")
    if bool(prior_start_raw) != bool(prior_end_raw):
        raise ValueError(
            "An explicit comparative period requires both prior start and end"
        )
    if prior_start_raw:
        prior_start_date = date.fromisoformat(prior_start_raw)
        prior_end_date = date.fromisoformat(prior_end_raw)
        if prior_start_date > prior_end_date or prior_end_date >= start:
            raise ValueError(
                "The comparative period is invalid or overlaps the current year"
            )
    if str(payload.get("currency", "EUR")) != "EUR":
        raise ValueError("MVP supports EUR only")
    taxonomy_checksum = str(payload.get("taxonomy_checksum") or "")
    if not re.fullmatch(r"[0-9a-f]{64}", taxonomy_checksum):
        raise ValueError("A lowercase SHA-256 taxonomy package checksum is required")
    unsupported_reasons = _validate_entity(entity)
    state = CaseState.UNSUPPORTED if unsupported_reasons else CaseState.DRAFT
    reporting_precision = int(payload.get("reporting_precision", 0))
    if reporting_precision not in {0, 1, 2}:
        raise ValueError("Reporting precision must be zero, one, or two decimals")
    output_language = str(payload.get("output_language", "it")).strip().lower()
    if output_language not in {"it", "en"}:
        raise ValueError("Output language must be it or en")
    _validate_required_entity_profile(entity)
    taxonomy_output_contracts = _taxonomy_output_contracts(rule_pack)
    case: dict[str, Any] = {
        "schema_version": SCHEMA_VERSION,
        "case_id": case_id,
        "tenant_id": str(payload["tenant_id"]),
        "revision_id": "rev_1",
        "state": state.value,
        "entity": entity,
        "period": {"start": start.isoformat(), "end": end.isoformat()},
        "currency": "EUR",
        "reporting_precision": reporting_precision,
        "output_language": output_language,
        "requested_form": str(payload.get("requested_form", "AUTO_RECOMMEND")),
        "rule_pack_versions": {
            "jurisdiction": "IT",
            "accounting_framework": "OIC",
            "statutory_rule_pack": rule_pack["id"],
            "oic_rule_pack": str(payload["oic_rule_pack"]),
            "taxonomy_id": str(payload.get("taxonomy_id", "PCI_2018-11-04")),
            "filing_instruction_pack": str(
                payload.get("filing_instruction_pack", "RI_2026")
            ),
            "early_adoption_flags": list(payload.get("early_adoption_flags", [])),
        },
        "rule_pack_checksum": _sha256_bytes(_canonical_json(rule_pack)),
        "taxonomy_output_contracts": taxonomy_output_contracts,
        "statutory_presentation_rule_pack_checksum": None,
        "taxonomy_checksum": taxonomy_checksum,
        "unsupported_reasons": unsupported_reasons,
        "source_documents": [],
        "file_security_scans": [],
        "prior_xbrl": None,
        "comparative_reconciliation_decisions": [],
        "trial_balance": None,
        "form_analysis": None,
        "selected_form": None,
        "mappings": [],
        "adjustments": [],
        "taxonomy_facts": [],
        "taxonomy_fact_context": None,
        "schedule_taxonomy_adapter": None,
        "schedule_taxonomy_adapter_rule_pack_checksum": None,
        "schedule_taxonomy_facts": [],
        "taxonomy_catalogue_build": None,
        "statutory_presentation_required": True,
        "statutory_presentation": None,
        "taxonomy_representation": None,
        "micro_reporting": None,
        "mapping_candidates": [],
        "taxonomy_mapping_index": None,
        "canonical_facts": [],
        "statements": None,
        "schedules": [],
        "disclosure_answers": [],
        "disclosure_trigger_flags": [],
        "disclosure_trigger_decisions": [],
        "disclosure_rule_pack": None,
        "disclosure_coverage": None,
        "questionnaire": [],
        "note_outline": [],
        "note_outline_context": None,
        "prior_narrative_suggestions": [],
        "narrative_blocks": [],
        "narrative_context": None,
        "intelligence_runs": [],
        "latest_workflow_guidance": None,
        "model_mapping_suggestions": [],
        "disclosure_activation_suggestions": [],
        "narrative_suggestions": [],
        "client_history_suggestions": None,
        "review_decisions": [],
        "regulatory_migrations": [],
        "preview": None,
        "xbrl_review": None,
        "external_validation": None,
        "validation": None,
        "approval": None,
        "approval_snapshots": [],
        "artifacts": [],
        "audit_events": [],
        "created_at": _now(),
        "updated_at": _now(),
    }
    _record_event(case, "case_created", actor)
    if unsupported_reasons:
        _record_event(
            case, "unsupported_case_detected", actor, {"reasons": unsupported_reasons}
        )
    save_case(case_dir, case)
    return case


def migrate_regulatory_versions(
    case: dict[str, Any],
    migration: Mapping[str, Any],
    actor: str,
    expected_revision: str,
) -> dict[str, Any]:
    """Explicitly migrate an open case and invalidate all regulated outputs."""

    _ensure_revision(case, expected_revision)
    if case.get("approval") or case.get("state") in {
        CaseState.APPROVED,
        CaseState.EXPORTED,
        CaseState.ARCHIVED,
    }:
        raise ValueError("Approved, exported, and archived cases cannot be migrated")
    if case.get("state") == CaseState.UNSUPPORTED:
        raise ValueError("Unsupported cases cannot be migrated")
    required = {
        "reason",
        "statutory_rule_pack",
        "oic_rule_pack",
        "taxonomy_id",
        "taxonomy_checksum",
        "filing_instruction_pack",
        "early_adoption_flags",
    }
    allowed = required | {"disclosure_rule_pack"}
    if set(migration) - allowed or not required.issubset(migration):
        raise ValueError(
            "Regulatory migration requires exactly the controlled version fields"
        )
    reason = str(migration["reason"]).strip()
    if len(reason) < 10:
        raise ValueError("Regulatory migration requires a specific reason")
    statutory_rule_pack = migration["statutory_rule_pack"]
    if not isinstance(statutory_rule_pack, Mapping):
        raise ValueError("Statutory rule pack must be an object")
    period_start = date.fromisoformat(case["period"]["start"])
    effective_from = date.fromisoformat(str(statutory_rule_pack["effective_from"]))
    effective_to = date.fromisoformat(str(statutory_rule_pack["effective_to"]))
    if not effective_from <= period_start <= effective_to:
        raise ValueError("Migrated statutory rule pack is not effective for the period")
    taxonomy_checksum = str(migration["taxonomy_checksum"])
    if not re.fullmatch(r"[0-9a-f]{64}", taxonomy_checksum):
        raise ValueError("Migrated taxonomy checksum must be a lowercase SHA-256")
    early_adoption_flags = migration["early_adoption_flags"]
    if not isinstance(early_adoption_flags, list) or not all(
        isinstance(item, str) and item.strip() for item in early_adoption_flags
    ):
        raise ValueError("Early-adoption flags must be a list of non-empty strings")
    disclosure_rule_pack = migration.get("disclosure_rule_pack")
    if disclosure_rule_pack is not None and not isinstance(
        disclosure_rule_pack, Mapping
    ):
        raise ValueError("Disclosure rule pack must be an object")

    previous_versions = deepcopy(case["rule_pack_versions"])
    target_versions = {
        "jurisdiction": "IT",
        "accounting_framework": "OIC",
        "statutory_rule_pack": str(statutory_rule_pack["id"]),
        "oic_rule_pack": str(migration["oic_rule_pack"]),
        "taxonomy_id": str(migration["taxonomy_id"]),
        "filing_instruction_pack": str(migration["filing_instruction_pack"]),
        "early_adoption_flags": list(early_adoption_flags),
    }
    if disclosure_rule_pack is not None:
        target_versions["disclosure_rule_pack"] = str(disclosure_rule_pack["id"])
    statutory_checksum = _sha256_bytes(_canonical_json(statutory_rule_pack))
    disclosure_checksum = (
        disclosure_rule_pack_hash(disclosure_rule_pack)
        if disclosure_rule_pack is not None
        else None
    )
    version_changes = {
        key: {"from": previous_versions.get(key), "to": target_versions.get(key)}
        for key in sorted(set(previous_versions) | set(target_versions))
        if previous_versions.get(key) != target_versions.get(key)
    }
    checksum_changes = {
        "statutory_rule_pack": {
            "from": case.get("rule_pack_checksum"),
            "to": statutory_checksum,
        },
        "taxonomy": {
            "from": case.get("taxonomy_checksum"),
            "to": taxonomy_checksum,
        },
        "disclosure_rule_pack": {
            "from": case.get("disclosure_rule_pack_checksum"),
            "to": disclosure_checksum,
        },
    }
    checksum_changes = {
        key: value
        for key, value in checksum_changes.items()
        if value["from"] != value["to"]
    }
    if not version_changes and not checksum_changes:
        raise ValueError("Regulatory migration target is identical to the locked case")

    reset_values: dict[str, Any] = {
        "form_analysis": None,
        "selected_form": None,
        "mappings": [],
        "adjustments": [],
        "taxonomy_facts": [],
        "taxonomy_fact_context": None,
        "schedule_taxonomy_adapter": None,
        "schedule_taxonomy_facts": [],
        "taxonomy_catalogue_build": None,
        "statutory_presentation": None,
        "taxonomy_representation": None,
        "micro_reporting": None,
        "mapping_candidates": [],
        "canonical_facts": [],
        "statements": None,
        "schedules": [],
        "disclosure_answers": [],
        "disclosure_trigger_flags": [],
        "disclosure_trigger_decisions": [],
        "disclosure_rule_pack": None,
        "disclosure_coverage": None,
        "questionnaire": [],
        "note_outline": [],
        "note_outline_context": None,
        "prior_narrative_suggestions": [],
        "narrative_blocks": [],
        "narrative_context": None,
        "intelligence_runs": [],
        "latest_workflow_guidance": None,
        "model_mapping_suggestions": [],
        "disclosure_activation_suggestions": [],
        "narrative_suggestions": [],
        "review_decisions": [],
        "preview": None,
        "xbrl_review": None,
        "external_validation": None,
        "validation": None,
        "artifacts": [],
    }
    invalidated = [
        {
            "component": key,
            "prior_sha256": _sha256_bytes(_canonical_json(case.get(key))),
        }
        for key, empty_value in reset_values.items()
        if case.get(key) not in (None, [], {}) and case.get(key) != empty_value
    ]
    _mutate(case, actor, "regulatory_versions_migrated")
    case["rule_pack_versions"] = target_versions
    case["rule_pack_checksum"] = statutory_checksum
    case["taxonomy_output_contracts"] = _taxonomy_output_contracts(statutory_rule_pack)
    case["statutory_presentation_rule_pack_checksum"] = None
    case["schedule_taxonomy_adapter_rule_pack_checksum"] = None
    case["rule_pack_versions"].pop("statutory_presentation_rule_pack", None)
    case["taxonomy_checksum"] = taxonomy_checksum
    case["disclosure_rule_pack_checksum"] = disclosure_checksum
    for key, empty_value in reset_values.items():
        case[key] = empty_value
    case["state"] = (
        CaseState.INPUT_REVIEW
        if (case.get("trial_balance") or {}).get("confirmed_convention")
        else CaseState.DRAFT
    )
    report = {
        "migration_id": f"migration_{len(case.get('regulatory_migrations', [])) + 1:04d}",
        "requested_by": actor,
        "requested_at": _now(),
        "reason": reason,
        "from_revision_id": expected_revision,
        "to_revision_id": case["revision_id"],
        "version_changes": version_changes,
        "checksum_changes": checksum_changes,
        "invalidated_components": invalidated,
        "retained_evidence": {
            "source_document_count": len(case.get("source_documents", [])),
            "trial_balance_retained": case.get("trial_balance") is not None,
            "prior_xbrl_retained": case.get("prior_xbrl") is not None,
        },
        "required_recomputation": [
            "FORM_ELIGIBILITY",
            "MAPPING",
            "STATEMENTS",
            "SCHEDULES",
            "DISCLOSURES",
            "NOTES",
            "STATUTORY_PRESENTATION",
            "LOCAL_XBRL_REVIEW",
        ],
        "revalidation_status": "REQUIRED",
        "revalidation_runs": [],
    }
    case.setdefault("regulatory_migrations", []).append(report)
    _record_event(
        case,
        "regulatory_versions_migrated",
        actor,
        {
            "migration_id": report["migration_id"],
            "changed_version_keys": sorted(version_changes),
            "changed_checksum_keys": sorted(checksum_changes),
            "invalidated_components": [item["component"] for item in invalidated],
        },
    )
    return case


def ingest_prior_xbrl(
    case: dict[str, Any], source: Path, actor: str, expected_revision: str
) -> dict[str, Any]:
    """Attach source-anchored prior XBRL after exact entity and period checks."""

    _ensure_revision(case, expected_revision)
    if case["state"] == CaseState.UNSUPPORTED:
        raise ValueError("Prior XBRL cannot be added to an unsupported case")
    if case["entity"].get("first_financial_year") is True:
        raise ValueError("A first financial year cannot have a prior filed XBRL")
    parsed = parse_prior_xbrl(source)
    allowed_identifiers = {
        str(case["entity"].get("tax_identifier", "")).strip(),
        str(case["entity"].get("prior_xbrl_identifier", "")).strip(),
    } - {""}
    if parsed["entity_identifier"] not in allowed_identifiers:
        raise ValueError("Prior-XBRL entity identifier does not match the case")
    expected_prior_end = str(case["entity"].get("prior_period_end") or "")
    if not expected_prior_end:
        current_start = date.fromisoformat(case["period"]["start"])
        expected_prior_end = (current_start - timedelta(days=1)).isoformat()
    matching_contexts = [
        context
        for context in parsed["contexts"]
        if context["period"].get("end") == expected_prior_end
    ]
    if not matching_contexts:
        raise ValueError("Prior-XBRL periods do not align with the comparative period")
    document_id = _next_document_id(case)
    for fact in parsed["facts"]:
        fact["source_anchor"]["document_id"] = document_id
        fact["source_anchor"]["source_ref"] = f"{document_id}_{fact['fact_id']}"
    document = {
        "document_id": document_id,
        "purpose": "PRIOR_XBRL",
        "file_name": parsed["file_name"],
        "media_type": "application/xbrl+xml",
        "sha256": parsed["sha256"],
        "size_bytes": source.resolve().stat().st_size,
        "parser_profile": "prior_xbrl_v4",
        "parsed_at": _now(),
    }
    _mutate(case, actor, "prior_xbrl_parsed")
    case["source_documents"] = [
        item
        for item in case.get("source_documents", [])
        if item.get("purpose") != "PRIOR_XBRL"
    ] + [document]
    case["prior_xbrl"] = {
        **parsed,
        "document_id": document_id,
        "expected_prior_end": expected_prior_end,
        "matching_context_ids": [item["context_id"] for item in matching_contexts],
        "computation_context": _computation_context(case, "prior-xbrl-parser-v4"),
    }
    case["mapping_candidates"] = []
    case["taxonomy_mapping_index"] = None
    case["comparative_reconciliation_decisions"] = []
    case["validation"] = None
    _record_event(case, "document_uploaded", actor, {"document_id": document_id})
    _record_evidence_attached(case, actor, document)
    _record_event(
        case,
        "document_parsed",
        actor,
        {"document_id": document_id, "parser_profile": "prior_xbrl_v4"},
    )
    _record_event(
        case,
        "prior_xbrl_parsed",
        actor,
        {"document_id": document_id, "fact_count": len(parsed["facts"])},
    )
    return case


def attach_supporting_document(
    case: dict[str, Any],
    source: Path,
    purpose: str,
    description: str,
    actor: str,
    expected_revision: str,
) -> dict[str, Any]:
    """Attach checksum-bound supporting evidence without interpreting its content."""

    _ensure_revision(case, expected_revision)
    if source.is_symlink() or not source.is_file():
        raise ValueError("Supporting evidence must be a regular local file")
    resolved = source.resolve()
    if resolved.stat().st_size > MAX_INPUT_BYTES:
        raise ValueError("Supporting evidence exceeds the 100 MiB size limit")
    allowed_suffixes = {
        ".csv": "text/csv",
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".pdf": "application/pdf",
        ".txt": "text/plain",
        ".json": "application/json",
        ".xml": "application/xml",
        ".html": "text/html",
        ".htm": "text/html",
    }
    suffix = resolved.suffix.lower()
    if suffix not in allowed_suffixes:
        raise ValueError("Unsupported supporting-evidence file type")
    normalized_purpose = purpose.strip().upper()
    allowed_purposes = {
        "SUPPORTING_EVIDENCE",
        "CORPORATE_RESOLUTION",
        "TAX_COMPUTATION",
        "CONTRACT",
        "QUESTIONNAIRE_EVIDENCE",
        "RESTATEMENT_WORKPAPER",
        "GOING_CONCERN_ASSESSMENT",
    }
    if normalized_purpose not in allowed_purposes:
        raise ValueError("Unsupported supporting-evidence purpose")
    normalized_description = description.strip()
    if not normalized_description:
        raise ValueError("Supporting evidence requires a description")
    document_id = _next_document_id(case)
    document = {
        "document_id": document_id,
        "purpose": normalized_purpose,
        "description": normalized_description,
        "file_name": resolved.name,
        "media_type": allowed_suffixes[suffix],
        "sha256": _sha256_file(resolved),
        "size_bytes": resolved.stat().st_size,
        "source_refs": [document_id],
        "content_treated_as_untrusted": True,
        "attached_at": _now(),
    }
    _mutate(case, actor, "supporting_evidence_attached")
    case.setdefault("source_documents", []).append(document)
    case["validation"] = None
    _record_event(case, "document_uploaded", actor, {"document_id": document_id})
    _record_evidence_attached(case, actor, document)
    return case


def _load_table(
    path: Path, sheet: str | None
) -> tuple[list[str], list[list[Any]], str]:
    if path.stat().st_size > MAX_INPUT_BYTES:
        raise ValueError("Input table exceeds the 100 MiB size limit")
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            sample = handle.read(8192)
            handle.seek(0)
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            rows = list(csv.reader(handle, dialect))
        if not rows:
            raise ValueError("The CSV file is empty")
        if len(rows) - 1 > MAX_TEMPLATE_ROWS:
            raise ValueError("Input table exceeds the 20,000-row MVP limit")
        return [str(value) for value in rows[0]], rows[1:], "csv"
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("openpyxl is required for XLSX ingestion") from exc
        try:
            with zipfile.ZipFile(path) as archive:
                total_uncompressed = 0
                for member in archive.infolist():
                    member_path = Path(member.filename)
                    if member.flag_bits & 0x1:
                        raise ValueError("Encrypted XLSX members are not supported")
                    if member_path.is_absolute() or ".." in member_path.parts:
                        raise ValueError("Unsafe XLSX member path")
                    total_uncompressed += member.file_size
                    if (
                        member.compress_size > 0
                        and member.file_size / member.compress_size > 200
                    ):
                        raise ValueError("Suspicious XLSX compression ratio")
                    lowered = member.filename.lower()
                    if "vbaproject" in lowered:
                        raise ValueError("Spreadsheet macros are not supported")
                    if lowered.startswith("xl/externallinks/"):
                        raise ValueError("XLSX external links are not supported")
                if total_uncompressed > MAX_XLSX_UNCOMPRESSED_BYTES:
                    raise ValueError("XLSX expanded content exceeds the safety limit")
                corrupt_member = archive.testzip()
                if corrupt_member:
                    raise ValueError(f"Corrupt XLSX member: {corrupt_member}")
        except zipfile.BadZipFile as exc:
            raise ValueError("XLSX input is not a valid ZIP package") from exc
        workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
        selected = workbook[sheet] if sheet else workbook[workbook.sheetnames[0]]
        selected_title = selected.title
        rows = list(selected.iter_rows(values_only=True))
        if not rows:
            raise ValueError("The XLSX worksheet is empty")
        if len(rows) - 1 > MAX_TEMPLATE_ROWS:
            raise ValueError("Input table exceeds the 20,000-row MVP limit")
        formula_workbook = load_workbook(
            path, read_only=True, data_only=False, keep_links=False
        )
        formula_sheet = formula_workbook[selected_title]
        for value_row, formula_row in zip(
            rows, formula_sheet.iter_rows(), strict=False
        ):
            for cached_value, formula_cell in zip(value_row, formula_row, strict=False):
                if formula_cell.data_type == "f" and cached_value is None:
                    raise ValueError(
                        "XLSX formula has no trusted cached value for deterministic import"
                    )
        formula_workbook.close()
        workbook.close()
        return (
            ["" if value is None else str(value) for value in rows[0]],
            [list(row) for row in rows[1:]],
            selected_title,
        )
    raise ValueError("Only CSV and XLSX trial balances are supported")


def _normalized_header(value: str) -> str:
    text = re.sub(r"[^a-z0-9]+", "_", value.strip().lower())
    aliases = {
        "codice_conto": "account_code",
        "conto": "account_code",
        "descrizione": "account_description",
        "descrizione_conto": "account_description",
        "saldo_iniziale_dare": "opening_debit",
        "saldo_iniziale_avere": "opening_credit",
        "progressivi_dare": "period_debit",
        "progressivi_avere": "period_credit",
        "saldo_finale_dare": "closing_debit",
        "saldo_finale_avere": "closing_credit",
        "saldo_precedente_dare": "prior_closing_debit",
        "saldo_precedente_avere": "prior_closing_credit",
    }
    return aliases.get(text.strip("_"), text.strip("_"))


def _column_reference(position: int) -> str:
    """Return the one-based spreadsheet column letter for a zero-based position."""

    if position < 0:
        raise ValueError("Column positions cannot be negative")
    value = position + 1
    letters = ""
    while value:
        value, remainder = divmod(value - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _normalized_header_columns(headers: Sequence[str]) -> list[dict[str, Any]]:
    """Normalize headers while preserving coordinates and rejecting collisions."""

    columns: list[dict[str, Any]] = []
    seen: dict[str, dict[str, Any]] = {}
    for position, original in enumerate(headers):
        normalized = _normalized_header(original)
        coordinate = _column_reference(position)
        if not normalized:
            raise ValueError(f"Column {coordinate} has a blank or unsupported header")
        prior = seen.get(normalized)
        if prior:
            raise ValueError(
                "Duplicate normalized column "
                f"{normalized!r}: {prior['column']} ({prior['column_header']!r}) "
                f"and {coordinate} ({original!r})"
            )
        column = {
            "column": coordinate,
            "column_index": position + 1,
            "column_header": original,
            "normalized_column": normalized,
        }
        seen[normalized] = column
        columns.append(column)
    return columns


def _row_decimal(row: Mapping[str, Any], key: str) -> Decimal:
    if key not in row:
        raise ValueError(f"Required monetary field is missing: {key}")
    return normalize_decimal(row[key])


def _calibrate(rows: Sequence[Mapping[str, Any]], tolerance: Decimal) -> dict[str, Any]:
    matches_excludes = 0
    matches_includes = 0
    tested = 0
    samples: list[dict[str, str]] = []
    for row in rows:
        opening = _row_decimal(row, "opening_signed")
        closing = _row_decimal(row, "closing_signed")
        period_net = _row_decimal(row, "period_debit") - _row_decimal(
            row, "period_credit"
        )
        excludes_diff = opening + period_net - closing
        includes_diff = period_net - closing
        if abs(excludes_diff) <= tolerance:
            matches_excludes += 1
        if abs(includes_diff) <= tolerance:
            matches_includes += 1
        tested += 1
        if len(samples) < 5:
            samples.append(
                {
                    "account_code": str(row["account_code"]),
                    "opening_signed": _decimal_text(opening),
                    "period_net": _decimal_text(period_net),
                    "closing_signed": _decimal_text(closing),
                    "excludes_opening_difference": _decimal_text(excludes_diff),
                    "includes_opening_difference": _decimal_text(includes_diff),
                }
            )
    if not rows or all(
        _row_decimal(row, "period_debit") == 0
        and _row_decimal(row, "period_credit") == 0
        for row in rows
    ):
        detected = ParserConvention.SIGNED_BALANCE_ONLY
    elif tested and matches_excludes == tested and matches_includes != tested:
        detected = ParserConvention.TURNOVER_EXCLUDES_OPENING
    elif tested and matches_includes == tested and matches_excludes != tested:
        detected = ParserConvention.TURNOVER_INCLUDES_OPENING
    else:
        detected = ParserConvention.UNKNOWN
    return {
        "detected_convention": detected.value,
        "tested_rows": tested,
        "matches_turnover_excludes_opening": matches_excludes,
        "matches_turnover_includes_opening": matches_includes,
        "unmatched_rows": tested - max(matches_excludes, matches_includes),
        "tolerance": _decimal_text(tolerance),
        "samples": samples,
        "closing_entries_assessment": {
            "appears_included": None,
            "status": "REQUIRES_PROFESSIONAL_CONFIRMATION",
            "reason": (
                "The supported numeric columns do not mechanically distinguish "
                "ordinary turnover from closing entries."
            ),
        },
    }


def ingest_trial_balance(
    case: dict[str, Any],
    source_path: Path,
    actor: str,
    expected_revision: str,
    sheet: str | None = None,
    tolerance: Decimal = Decimal("0.01"),
) -> dict[str, Any]:
    """Parse one source-anchored trial balance and produce calibration evidence."""

    _ensure_revision(case, expected_revision)
    if case["state"] == CaseState.UNSUPPORTED:
        raise ValueError("Unsupported cases cannot ingest accounting data")
    if source_path.is_symlink() or not source_path.is_file():
        raise ValueError("Trial balance must be a regular local file")
    path = source_path.resolve()
    headers, raw_rows, sheet_name = _load_table(path, sheet)
    header_columns = _normalized_header_columns(headers)
    normalized_headers = [item["normalized_column"] for item in header_columns]
    required = {"account_code", "account_description"}
    if not required.issubset(normalized_headers):
        raise ValueError("Trial balance requires account_code and account_description")
    separate_layout = {
        "opening_debit",
        "opening_credit",
        "period_debit",
        "period_credit",
        "closing_debit",
        "closing_credit",
        "prior_closing_debit",
        "prior_closing_credit",
    }.issubset(normalized_headers)
    signed_layout = {
        "opening_signed",
        "period_debit",
        "period_credit",
        "closing_signed",
        "prior_closing_signed",
    }.issubset(normalized_headers)
    if not separate_layout and not signed_layout:
        raise ValueError(
            "Trial balance does not match supported separate or signed layouts"
        )
    document_id = _next_document_id(case)
    document = {
        "document_id": document_id,
        "purpose": "TRIAL_BALANCE",
        "file_name": path.name,
        "media_type": (
            "text/csv"
            if path.suffix.lower() == ".csv"
            else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        "sha256": _sha256_file(path),
        "size_bytes": path.stat().st_size,
        "sheet": sheet_name,
        "columns": header_columns,
        "parser_profile": "generic_it_tb_v1",
        "parsed_at": _now(),
    }
    entries: list[dict[str, Any]] = []
    seen: set[str] = set()
    anchors: list[dict[str, Any]] = []
    for index, values in enumerate(raw_rows, start=2):
        row = {
            name: values[pos] if pos < len(values) else None
            for pos, name in enumerate(normalized_headers)
        }
        code = str(row.get("account_code") or "").strip()
        description = str(row.get("account_description") or "").strip()
        if not code and not description:
            continue
        if not code:
            raise ValueError(f"Missing account identifier at row {index}")
        if code in seen:
            raise ValueError(
                f"Duplicate account identifier without aggregation key: {code}"
            )
        seen.add(code)
        if separate_layout:
            opening_debit = _row_decimal(row, "opening_debit")
            opening_credit = _row_decimal(row, "opening_credit")
            closing_debit = _row_decimal(row, "closing_debit")
            closing_credit = _row_decimal(row, "closing_credit")
            prior_debit = _row_decimal(row, "prior_closing_debit")
            prior_credit = _row_decimal(row, "prior_closing_credit")
            opening_signed = opening_debit - opening_credit
            closing_signed = closing_debit - closing_credit
            prior_signed = prior_debit - prior_credit
        else:
            opening_signed = _row_decimal(row, "opening_signed")
            closing_signed = _row_decimal(row, "closing_signed")
            prior_signed = _row_decimal(row, "prior_closing_signed")
            opening_debit = max(opening_signed, Decimal("0"))
            opening_credit = max(-opening_signed, Decimal("0"))
            closing_debit = max(closing_signed, Decimal("0"))
            closing_credit = max(-closing_signed, Decimal("0"))
            prior_debit = max(prior_signed, Decimal("0"))
            prior_credit = max(-prior_signed, Decimal("0"))
        period_debit = _row_decimal(row, "period_debit")
        period_credit = _row_decimal(row, "period_credit")
        source_refs: list[str] = []
        for column in header_columns:
            position = int(column["column_index"]) - 1
            column_name = str(column["normalized_column"])
            anchor_id = f"src_{len(anchors) + 1:07d}"
            source_refs.append(anchor_id)
            raw_value = values[position] if position < len(values) else None
            anchor: dict[str, Any] = {
                "source_ref": anchor_id,
                "document_id": document_id,
                "sheet": sheet_name,
                "row": index,
                "column": column["column"],
                "column_header": column["column_header"],
                "normalized_column": column_name,
                "raw_value": "" if raw_value is None else str(raw_value),
                "parser_profile": "generic_it_tb_v1",
                "confidence": "HIGH",
            }
            if column_name not in {"account_code", "account_description"}:
                anchor["normalized_value"] = _decimal_text(
                    _row_decimal(row, column_name)
                )
            else:
                anchor["normalized_value"] = str(row[column_name] or "").strip()
            anchors.append(anchor)
        entries.append(
            {
                "account_id": f"acc_{len(entries) + 1:06d}",
                "account_code": code,
                "account_description": description,
                "opening_debit": _decimal_text(opening_debit),
                "opening_credit": _decimal_text(opening_credit),
                "opening_signed": _decimal_text(opening_signed),
                "period_debit": _decimal_text(period_debit),
                "period_credit": _decimal_text(period_credit),
                "closing_debit": _decimal_text(closing_debit),
                "closing_credit": _decimal_text(closing_credit),
                "closing_signed": _decimal_text(closing_signed),
                "prior_closing_debit": _decimal_text(prior_debit),
                "prior_closing_credit": _decimal_text(prior_credit),
                "prior_closing_signed": _decimal_text(prior_signed),
                "source_refs": source_refs,
            }
        )
    calibration = _calibrate(entries, tolerance)
    debit_total = sum(Decimal(row["closing_debit"]) for row in entries)
    credit_total = sum(Decimal(row["closing_credit"]) for row in entries)
    calibration["closing_debit_total"] = _decimal_text(debit_total)
    calibration["closing_credit_total"] = _decimal_text(credit_total)
    calibration["closing_difference"] = _decimal_text(debit_total - credit_total)
    _mutate(case, actor, "document_parsed")
    case["source_documents"] = [
        item
        for item in case.get("source_documents", [])
        if item.get("purpose") != "TRIAL_BALANCE"
    ] + [document]
    case["trial_balance"] = {
        "layout": "SEPARATE_DEBIT_CREDIT" if separate_layout else "SIGNED_BALANCES",
        "entries": entries,
        "source_anchors": anchors,
        "calibration": calibration,
        "confirmed_convention": None,
        "computation_context": _computation_context(
            case, "generic-it-trial-balance-parser-v1"
        ),
    }
    _clear_accounting_dependent_reviews(case)
    case["state"] = CaseState.INPUT_REVIEW
    case["mappings"] = []
    case["mapping_candidates"] = []
    case["taxonomy_mapping_index"] = None
    case["canonical_facts"] = []
    case["statements"] = None
    case["statutory_presentation"] = None
    case["schedules"] = []
    case["validation"] = None
    _record_event(case, "document_uploaded", actor, {"document_id": document_id})
    _record_evidence_attached(case, actor, document)
    _record_event(case, "document_parsed", actor, {"document_id": document_id})
    return case


def confirm_parser(
    case: dict[str, Any], convention: str, actor: str, expected_revision: str
) -> dict[str, Any]:
    """Record the explicit human confirmation required before computation."""

    _ensure_revision(case, expected_revision)
    if not case.get("trial_balance"):
        raise ValueError("No parsed trial balance is available")
    selected = ParserConvention(convention)
    if selected is ParserConvention.UNKNOWN:
        raise ValueError("UNKNOWN cannot be confirmed")
    closing_difference = Decimal(
        str(case["trial_balance"]["calibration"]["closing_difference"])
    )
    tolerance = Decimal(str(case["trial_balance"]["calibration"]["tolerance"]))
    if abs(closing_difference) > tolerance:
        raise ValueError(
            "Trial-balance closing debit and credit totals do not reconcile"
        )
    _mutate(case, actor, "parser_convention_confirmed")
    case["trial_balance"]["confirmed_convention"] = selected.value
    if selected is ParserConvention.TURNOVER_INCLUDES_CLOSING_ENTRIES:
        appears_included: bool | None = True
        closing_reason = "The professional explicitly confirmed that turnover includes closing entries."
    elif selected is ParserConvention.SIGNED_BALANCE_ONLY:
        appears_included = None
        closing_reason = (
            "The professional confirmed a signed-balance-only source, so closing-entry "
            "inclusion cannot be assessed from turnover columns."
        )
    else:
        appears_included = False
        closing_reason = (
            "The professional explicitly confirmed a turnover convention that excludes "
            "closing entries."
        )
    closing_review = {
        "appears_included": appears_included,
        "status": "USER_CONFIRMED",
        "reason": closing_reason,
        "confirmed_convention": selected.value,
        "confirmed_by": actor,
        "confirmed_at": _now(),
    }
    case["trial_balance"]["closing_entries_review"] = closing_review
    case["trial_balance"]["calibration"]["closing_entries_assessment"] = deepcopy(
        closing_review
    )
    case["state"] = CaseState.MAPPING_REVIEW
    case["validation"] = None
    _record_event(
        case,
        "parser_convention_confirmed",
        actor,
        {
            "convention": selected.value,
            "closing_entries_review": closing_review,
        },
    )
    return case


def _threshold_result(
    year: Mapping[str, Any], thresholds: Mapping[str, Any], actor: str
) -> dict[str, Any]:
    values = {
        "assets": normalize_decimal(year["assets"]),
        "revenue": normalize_decimal(year["revenue"]),
        "employees": normalize_decimal(year["employees"]),
    }
    breached = {key: values[key] > normalize_decimal(thresholds[key]) for key in values}
    raw_source_refs = year.get("source_refs", {})
    if raw_source_refs is not None and not isinstance(raw_source_refs, Mapping):
        raise ValueError("Form metric source_refs must be keyed by metric name")
    raw_statuses = year.get("evidence_status", {})
    if raw_statuses is not None and not isinstance(raw_statuses, Mapping):
        raise ValueError("Form metric evidence_status must be keyed by metric name")
    provenance: dict[str, Any] = {}
    for key in values:
        supplied_refs = (raw_source_refs or {}).get(key, [])
        if isinstance(supplied_refs, (str, bytes)) or not isinstance(
            supplied_refs, Sequence
        ):
            raise ValueError(f"Form metric source refs for {key} must be a list")
        source_refs = sorted(
            {str(item).strip() for item in supplied_refs if str(item).strip()}
        )
        status = str((raw_statuses or {}).get(key, "USER_CONFIRMED")).upper()
        if status not in {
            EvidenceStatus.OBSERVED,
            EvidenceStatus.DERIVED,
            EvidenceStatus.USER_CONFIRMED,
        }:
            raise ValueError(f"Form metric {key} has a non-reviewable evidence status")
        if not source_refs:
            source_refs = [f"professional_input:{actor}:{int(year['year'])}:{key}"]
            status = EvidenceStatus.USER_CONFIRMED
        provenance[key] = {
            "status": str(status),
            "source_refs": source_refs,
            "confirmed_by": actor,
        }
    return {
        "year": int(year["year"]),
        "values": {key: _decimal_text(value) for key, value in values.items()},
        "breached": breached,
        "breach_count": sum(breached.values()),
        "within_thresholds": sum(breached.values()) <= 1,
        "provenance": provenance,
    }


def _prior_xbrl_reconciliation(case: Mapping[str, Any]) -> dict[str, Any]:
    """Compare current comparative facts with an attached prior filing exactly.

    This is deterministic because both sides are reviewed monetary facts with
    explicit periods; it does not decide whether a difference is a valid
    restatement, which remains a separate professional decision.
    """

    prior_xbrl = case.get("prior_xbrl")
    if not prior_xbrl:
        return {"status": "NOT_AVAILABLE", "checks": [], "issues": []}
    matching_context_ids = set(prior_xbrl.get("matching_context_ids", []))
    contexts = {
        str(item["context_id"]): item for item in prior_xbrl.get("contexts", [])
    }
    observed: dict[str, Decimal] = {}
    observed_refs: dict[str, str] = {}
    invalid_unit_qnames: set[str] = set()
    invalid_value_qnames: set[str] = set()
    issues: list[dict[str, Any]] = []
    for fact in prior_xbrl.get("facts", []):
        context_ref = str(fact.get("context_ref", ""))
        if (
            context_ref not in matching_context_ids
            or (contexts.get(context_ref) or {}).get("has_dimensions")
            or fact.get("nil")
            or fact.get("value") in {None, ""}
        ):
            continue
        qname = str(fact["qname"])
        unit = fact.get("unit") or {}
        measure = unit.get("measure") or {}
        if not (
            unit.get("kind") == "MEASURE"
            and measure.get("namespace") == ISO4217_NS
            and measure.get("local_name") == "EUR"
        ):
            invalid_unit_qnames.add(qname)
            continue
        raw_value = str(fact["value"]).strip()
        if not XBRL_DECIMAL_LEXICAL.fullmatch(raw_value):
            invalid_value_qnames.add(qname)
            continue
        value = Decimal(raw_value)
        if qname in observed and observed[qname] != value:
            issues.append(
                {
                    "code": "PRIOR_XBRL_DUPLICATE_CONFLICT",
                    "xbrl_concept": qname,
                }
            )
            continue
        observed[qname] = value
        observed_refs[qname] = str(
            (fact.get("source_anchor") or {}).get("source_ref", "")
        )
    expected: dict[str, Decimal] = {}

    def add_expected(qname: str, value: Decimal) -> None:
        if qname in expected and expected[qname] != value:
            issues.append(
                {
                    "code": "CURRENT_COMPARATIVE_DUPLICATE_CONFLICT",
                    "xbrl_concept": qname,
                }
            )
            return
        expected[qname] = value

    precision = int(case.get("reporting_precision", 0))
    for fact in case.get("canonical_facts", []):
        qname = fact.get("xbrl_concept")
        if qname:
            add_expected(
                str(qname),
                _reported_decimal(
                    Decimal(str(fact["prior_value"]))
                    * Decimal(str(fact["xbrl_sign_multiplier"])),
                    precision,
                ),
            )
    presentation = case.get("statutory_presentation") or {}
    for fact in presentation.get("output_facts", []):
        if fact.get("prior_value") is not None:
            add_expected(
                str(fact["xbrl_concept"]),
                _reported_decimal(Decimal(str(fact["prior_value"])), precision),
            )
    for fact in case.get("taxonomy_facts", []):
        if (
            fact.get("fact_type") == "MONETARY"
            and not fact.get("dimensions")
            and str(fact.get("period", "")).startswith("prior_")
        ):
            add_expected(
                str(fact["xbrl_concept"]),
                _reported_decimal(Decimal(str(fact["value"])), precision),
            )
    inventory = presentation.get("inventory") or {}
    relevant = {
        str(item["xbrl_concept"])
        for item in [
            *inventory.get("requirements", []),
            *inventory.get("totals", []),
        ]
    } | set(expected)
    for qname in sorted(invalid_unit_qnames & relevant):
        issues.append(
            {
                "code": "PRIOR_XBRL_MONETARY_UNIT_INVALID",
                "xbrl_concept": qname,
            }
        )
    for qname in sorted(invalid_value_qnames & relevant):
        issues.append(
            {
                "code": "PRIOR_XBRL_MONETARY_VALUE_INVALID",
                "xbrl_concept": qname,
            }
        )
    checks: list[dict[str, Any]] = []
    decisions = {
        str(item["xbrl_concept"]): item
        for item in case.get("comparative_reconciliation_decisions", [])
    }
    for qname in sorted(relevant & (set(expected) | set(observed))):
        expected_value = expected.get(qname)
        observed_value = observed.get(qname)
        check = {
            "xbrl_concept": qname,
            "current_comparative": (
                None if expected_value is None else _decimal_text(expected_value)
            ),
            "prior_filed": (
                None if observed_value is None else _decimal_text(observed_value)
            ),
            "prior_source_ref": observed_refs.get(qname),
        }
        if expected_value == observed_value:
            check["status"] = "PASS"
        elif expected_value == 0 and observed_value is None:
            check["status"] = "PASS_OMITTED_ZERO"
        else:
            decision = decisions.get(qname)
            if (
                decision
                and decision.get("action") == "RESTATEMENT_CONFIRMED"
                and decision.get("current_comparative") == check["current_comparative"]
                and decision.get("prior_filed") == check["prior_filed"]
            ):
                check["status"] = "RESTATEMENT_CONFIRMED"
                check["decision_id"] = decision["decision_id"]
            else:
                check["status"] = "FAIL"
                issues.append(
                    {
                        "code": "PRIOR_XBRL_COMPARATIVE_MISMATCH",
                        **{
                            key: value
                            for key, value in check.items()
                            if key != "status"
                        },
                    }
                )
        checks.append(check)
    return {
        "status": "PASS" if not issues else "FAIL",
        "checks": checks,
        "issues": issues,
    }


def record_comparative_reconciliation_decisions(
    case: dict[str, Any],
    decisions: Sequence[Mapping[str, Any]],
    actor: str,
    expected_revision: str,
) -> dict[str, Any]:
    """Record evidence-backed professional restatements of prior filed amounts."""

    _ensure_revision(case, expected_revision)
    if not case.get("prior_xbrl"):
        raise ValueError("Comparative decisions require an attached prior XBRL")
    probe = deepcopy(case)
    probe["comparative_reconciliation_decisions"] = []
    reconciliation = _prior_xbrl_reconciliation(probe)
    failed = {
        str(item["xbrl_concept"]): item
        for item in reconciliation["checks"]
        if item["status"] == "FAIL"
    }
    if not decisions:
        raise ValueError("At least one comparative reconciliation decision is required")
    normalized: list[dict[str, Any]] = []
    seen: set[str] = set()
    for raw in decisions:
        if set(raw) != {"xbrl_concept", "action", "reason", "source_refs"}:
            raise ValueError(
                "Comparative decisions require concept, action, reason, and source_refs"
            )
        qname = str(raw["xbrl_concept"])
        action = str(raw["action"]).upper()
        reason = str(raw["reason"]).strip()
        raw_source_refs = raw["source_refs"]
        if not isinstance(raw_source_refs, list):
            raise ValueError("Comparative decision source_refs must be a list")
        source_refs = sorted(
            {str(item) for item in raw_source_refs if str(item).strip()}
        )
        if qname in seen or qname not in failed:
            raise ValueError(f"Unknown or duplicate comparative difference: {qname}")
        if action != "RESTATEMENT_CONFIRMED" or not reason or not source_refs:
            raise ValueError(
                "Comparative restatements require explicit evidence and a reason"
            )
        if not set(source_refs) <= _available_evidence_refs(case):
            raise ValueError(
                "Comparative restatement references evidence outside the case"
            )
        seen.add(qname)
        check = failed[qname]
        normalized.append(
            {
                "decision_id": f"comparative_{len(normalized) + 1:06d}",
                "xbrl_concept": qname,
                "action": action,
                "current_comparative": check["current_comparative"],
                "prior_filed": check["prior_filed"],
                "reason": reason,
                "source_refs": source_refs,
                "reviewed_by": actor,
                "reviewed_at": _now(),
            }
        )
    retained = [
        item
        for item in case.get("comparative_reconciliation_decisions", [])
        if item["xbrl_concept"] not in seen
    ]
    _mutate(case, actor, "comparative_reconciliation_decisions_recorded")
    case["comparative_reconciliation_decisions"] = [*retained, *normalized]
    case["validation"] = None
    _record_event(
        case,
        "comparative_reconciliation_decisions_recorded",
        actor,
        {"xbrl_concepts": sorted(seen)},
    )
    return case


def determine_forms(
    case: dict[str, Any],
    metrics: Sequence[Mapping[str, Any]],
    rule_pack: Mapping[str, Any],
    actor: str,
    expected_revision: str,
) -> dict[str, Any]:
    """Calculate effective-dated form eligibility without selecting the form."""

    _ensure_revision(case, expected_revision)
    if case["state"] == CaseState.UNSUPPORTED:
        raise ValueError("Unsupported cases have no eligible MVP form")
    if (
        str(rule_pack.get("id"))
        != str(case["rule_pack_versions"]["statutory_rule_pack"])
        or _sha256_bytes(_canonical_json(rule_pack)) != case["rule_pack_checksum"]
    ):
        raise ValueError(
            "Statutory rule pack differs from the locked case; use explicit migration"
        )
    period_start = date.fromisoformat(case["period"]["start"])
    effective_from = date.fromisoformat(rule_pack["effective_from"])
    effective_to = date.fromisoformat(rule_pack["effective_to"])
    if not effective_from <= period_start <= effective_to:
        raise ValueError("Rule pack is not effective for the reporting period")
    first_year = bool(case["entity"].get("first_financial_year"))
    required_years = 1 if first_year else 2
    if isinstance(metrics, (str, bytes)) or not isinstance(metrics, Sequence):
        raise ValueError("Form metrics must be a list")
    if any(not isinstance(item, Mapping) or "year" not in item for item in metrics):
        raise ValueError("Every form metric record requires a reporting year")
    metric_years = [int(item["year"]) for item in metrics]
    if len(metric_years) != len(set(metric_years)):
        raise ValueError("Form metrics require one unique record per reporting year")
    reporting_year = date.fromisoformat(case["period"]["end"]).year
    expected_years = [reporting_year]
    if not first_year:
        expected_years.append(reporting_year - 1)
    unexpected_years = sorted(set(metric_years) - set(expected_years))
    if unexpected_years:
        raise ValueError(
            "Form metric years do not align with the reporting period: "
            + ", ".join(str(year) for year in unexpected_years)
        )
    metrics_by_year = {int(item["year"]): item for item in metrics}
    missing: list[str] = []
    missing_years = [year for year in expected_years if year not in metrics_by_year]
    missing.extend(f"threshold_metrics_for_{year}" for year in missing_years)
    ordered: list[Mapping[str, Any]] = []
    for year in expected_years:
        item = metrics_by_year.get(year)
        if item is None:
            continue
        missing_metrics = [
            key
            for key in ("assets", "revenue", "employees")
            if item.get(key) is None or not str(item.get(key)).strip()
        ]
        missing.extend(f"threshold_{key}_for_{year}" for key in missing_metrics)
        if not missing_metrics:
            ordered.append(item)
    analysis: dict[str, Any] = {}
    for form in ("ABBREVIATED", "MICRO"):
        thresholds = rule_pack["forms"][form]
        years = [
            _threshold_result(item, thresholds, actor)
            for item in ordered[:required_years]
        ]
        analysis[form] = {
            "thresholds": {key: str(value) for key, value in thresholds.items()},
            "years": years,
            "eligible": len(years) == required_years
            and all(item["within_thresholds"] for item in years),
            "reasons": [],
        }
    micro_exclusions = list(case["entity"].get("micro_exclusion_flags", []))
    if micro_exclusions:
        analysis["MICRO"]["eligible"] = False
        analysis["MICRO"]["reasons"].extend(micro_exclusions)
    eligible = ["ORDINARY"]
    if analysis["ABBREVIATED"]["eligible"]:
        eligible.insert(0, "ABBREVIATED")
    if analysis["MICRO"]["eligible"]:
        eligible.insert(0, "MICRO")
    recommended = eligible[0] if not missing else None
    _mutate(case, actor, "form_eligibility_calculated")
    _clear_accounting_dependent_reviews(case)
    case["form_analysis"] = {
        "rule_pack": rule_pack["id"],
        "eligible_forms": eligible if not missing else [],
        "ineligible_forms": [
            form for form in ("MICRO", "ABBREVIATED") if form not in eligible
        ],
        "recommended_form": recommended,
        "least_burdensome_only": True,
        "prior_year_form": case["entity"].get("prior_year_form"),
        "first_year": first_year,
        "missing_fields": missing,
        "calculations": analysis,
        "computation_context": _computation_context(
            case, "statutory-form-eligibility-v1"
        ),
    }
    case["selected_form"] = None
    case["mappings"] = []
    case["mapping_candidates"] = []
    case["taxonomy_mapping_index"] = None
    case["canonical_facts"] = []
    case["statements"] = None
    case["statutory_presentation"] = None
    case["schedules"] = []
    case["validation"] = None
    _record_event(case, "form_eligibility_calculated", actor)
    return case


def select_form(
    case: dict[str, Any], selected_form: str, actor: str, expected_revision: str
) -> dict[str, Any]:
    """Record the user's form choice after deterministic eligibility analysis."""

    _ensure_revision(case, expected_revision)
    analysis = case.get("form_analysis") or {}
    selected = selected_form.upper()
    if selected not in analysis.get("eligible_forms", []):
        raise ValueError("Selected form is not eligible under the locked rule pack")
    _mutate(case, actor, "form_selected")
    _clear_accounting_dependent_reviews(case)
    case["selected_form"] = selected
    case["mappings"] = []
    case["mapping_candidates"] = []
    case["taxonomy_mapping_index"] = None
    case["canonical_facts"] = []
    case["statements"] = None
    case["statutory_presentation"] = None
    case["schedules"] = []
    case["state"] = CaseState.MAPPING_REVIEW
    case["validation"] = None
    _record_event(case, "form_selected", actor, {"form": selected})
    return case


def generate_mapping_candidates(
    case: dict[str, Any],
    memory_path: Path,
    source_system_template: str,
    actor: str,
    expected_revision: str,
) -> dict[str, Any]:
    """Attach exact tenant-isolated approved mapping candidates for review."""

    _ensure_revision(case, expected_revision)
    trial_balance = case.get("trial_balance")
    if not trial_balance or not trial_balance.get("confirmed_convention"):
        raise ValueError(
            "Parser convention must be confirmed before mapping candidates"
        )
    if not case.get("selected_form"):
        raise ValueError("A statutory form must be selected before mapping candidates")
    if case.get("statutory_presentation_required", True) and not case.get(
        "taxonomy_mapping_index"
    ):
        raise ValueError(
            "Official taxonomy mapping index must be recorded before candidates"
        )
    candidates = mapping_candidates(case, memory_path, source_system_template)
    allowed_concepts = {
        str(item["xbrl_concept"])
        for item in (case.get("taxonomy_mapping_index") or {}).get("concepts", [])
        if item.get("mapping_allowed") is True
    }
    if allowed_concepts:
        candidates = [
            candidate
            for candidate in candidates
            if all(
                not allocation.get("xbrl_concept")
                or str(allocation["xbrl_concept"]) in allowed_concepts
                for allocation in candidate.get("allocations", [])
            )
        ]
    _mutate(case, actor, "mapping_candidates_generated")
    case["mapping_candidates"] = candidates
    case["mapping_candidate_context"] = {
        "source_system_template": source_system_template,
        "memory_file_name": memory_path.name,
        "generated_at": _now(),
        "computation_context": _computation_context(
            case, "mapping-candidate-memory-v1"
        ),
    }
    case["validation"] = None
    _record_event(
        case,
        "mapping_suggested",
        actor,
        {
            "candidate_count": len(candidates),
            "account_ids_sample": [item["account_id"] for item in candidates[:50]],
            "account_ids_sha256": _sha256_bytes(
                _canonical_json([item["account_id"] for item in candidates])
            ),
            "suggestion_source": "APPROVED_MAPPING_MEMORY",
        },
    )
    _record_event(
        case,
        "mapping_candidates_generated",
        actor,
        {"candidate_count": len(candidates)},
    )
    return case


def load_client_history(
    case: dict[str, Any],
    history_path: Path,
    actor: str,
    expected_revision: str,
) -> dict[str, Any]:
    """Load prior approved client knowledge as unconfirmed suggestions only."""

    _ensure_revision(case, expected_revision)
    suggestions = client_history_suggestions(case, history_path)
    _mutate(case, actor, "client_history_loaded")
    case["client_history_suggestions"] = suggestions
    case["validation"] = None
    _record_event(
        case,
        "client_history_loaded",
        actor,
        {
            "prior_period_end": suggestions["prior_period_end"],
            "answer_suggestion_count": len(suggestions["answer_suggestions"]),
            "narrative_suggestion_count": len(suggestions["narrative_suggestions"]),
        },
    )
    return case


def apply_mapping_decisions(
    case: dict[str, Any],
    decisions: Sequence[Mapping[str, Any]],
    actor: str,
    expected_revision: str,
) -> dict[str, Any]:
    """Apply reviewed, balancing account mappings and explicit exclusions."""

    _ensure_revision(case, expected_revision)
    trial_balance = case.get("trial_balance")
    if not trial_balance or not trial_balance.get("confirmed_convention"):
        raise ValueError("Parser convention must be confirmed before mapping")
    if not case.get("selected_form"):
        raise ValueError("A statutory form must be selected before mapping")
    taxonomy_mapping_index = case.get("taxonomy_mapping_index")
    if case.get("statutory_presentation_required", True) and not taxonomy_mapping_index:
        raise ValueError(
            "Official taxonomy mapping index must be recorded before mapping"
        )
    indexed_concepts = {
        str(item["xbrl_concept"])
        for item in (taxonomy_mapping_index or {}).get("concepts", [])
        if item.get("mapping_allowed") is True
    }
    accounts = {row["account_id"]: row for row in trial_balance["entries"]}
    prior_mappings = {
        str(item["account_id"]): item for item in case.get("mappings", [])
    }
    seen: set[str] = set()
    normalized: list[dict[str, Any]] = []
    for decision in decisions:
        account_id = str(decision["account_id"])
        if account_id in seen or account_id not in accounts:
            raise ValueError(f"Unknown or duplicate mapping account: {account_id}")
        seen.add(account_id)
        account = accounts[account_id]
        status = str(decision["decision"]).upper()
        memory_scope = str(decision.get("memory_scope", "CLIENT")).upper()
        if memory_scope not in {"CLIENT", "TENANT"}:
            raise ValueError("Mapping memory scope must be CLIENT or TENANT")
        if (
            memory_scope == "TENANT"
            and decision.get("tenant_reuse_approved") is not True
        ):
            raise ValueError("Tenant-wide mapping reuse requires explicit approval")
        if status == "EXCLUDED":
            if not str(decision.get("reason", "")).strip():
                raise ValueError(f"Excluded account {account_id} requires a reason")
            allocations: list[dict[str, Any]] = []
        elif status == "ACCEPTED":
            allocations = []
            raw_allocations = list(decision.get("allocations", []))
            if not raw_allocations:
                raise ValueError(f"Accepted account {account_id} requires allocations")
            current_total = Decimal("0")
            prior_total = Decimal("0")
            for position, allocation in enumerate(raw_allocations, start=1):
                evidence_status = EvidenceStatus(
                    str(allocation["evidence_status"]).upper()
                )
                if evidence_status not in {
                    EvidenceStatus.OBSERVED,
                    EvidenceStatus.USER_CONFIRMED,
                }:
                    raise ValueError(
                        "Reviewed mapping allocations must be observed or user-confirmed"
                    )
                current = normalize_decimal(allocation["current_amount"])
                prior = normalize_decimal(allocation["prior_amount"])
                current_total += current
                prior_total += prior
                qname = allocation.get("xbrl_concept")
                if qname is not None and not SAFE_QNAME.fullmatch(str(qname)):
                    raise ValueError(f"Invalid XBRL QName: {qname}")
                if (
                    case.get("statutory_presentation_required", True)
                    and str(qname or "") not in indexed_concepts
                ):
                    raise ValueError(
                        f"Mapping concept is outside the selected-form primary network: {qname}"
                    )
                xbrl_sign_multiplier = str(allocation.get("xbrl_sign_multiplier", ""))
                if qname is not None and xbrl_sign_multiplier not in {"1", "-1"}:
                    raise ValueError(
                        "Mapped XBRL concepts require an explicit sign multiplier"
                    )
                if qname is None:
                    xbrl_sign_multiplier = "1"
                schedule_triggers = sorted(
                    {
                        str(value).upper()
                        for value in allocation.get("schedule_triggers", [])
                    }
                )
                unknown_triggers = set(schedule_triggers) - SCHEDULE_TYPES
                if unknown_triggers:
                    raise ValueError(
                        f"Unknown schedule triggers: {sorted(unknown_triggers)}"
                    )
                allocations.append(
                    {
                        "allocation_id": f"{account_id}_{position}",
                        "canonical_line": str(allocation["canonical_line"]),
                        "statement_section": str(
                            allocation["statement_section"]
                        ).upper(),
                        "xbrl_concept": qname,
                        "xbrl_sign_multiplier": xbrl_sign_multiplier,
                        "current_amount": _decimal_text(current),
                        "prior_amount": _decimal_text(prior),
                        "evidence_status": evidence_status.value,
                        "review_reason": str(allocation.get("review_reason", "")),
                        "schedule_triggers": schedule_triggers,
                        "source_refs": list(account["source_refs"]),
                    }
                )
            if current_total != Decimal(account["closing_signed"]):
                raise ValueError(f"Current split for {account_id} does not balance")
            if prior_total != Decimal(account["prior_closing_signed"]):
                raise ValueError(f"Prior split for {account_id} does not balance")
        else:
            raise ValueError(f"Unsupported mapping decision: {status}")
        normalized.append(
            {
                "account_id": account_id,
                "decision": status,
                "reason": str(decision.get("reason", "")),
                "candidate_source": str(decision.get("candidate_source", "MANUAL")),
                "memory_scope": memory_scope,
                "approved_by": actor,
                "approved_at": _now(),
                "allocations": allocations,
            }
        )
    _mutate(case, actor, "mapping_decisions_applied")
    _clear_accounting_dependent_reviews(case)
    case["mappings"] = normalized
    case["adjustments"] = []
    case["canonical_facts"] = []
    case["statements"] = None
    case["statutory_presentation"] = None
    case["schedules"] = []
    case["validation"] = None
    case["state"] = CaseState.MAPPING_REVIEW
    mapping_after_hash = _sha256_bytes(_canonical_json(_case_payload_for_hash(case)))
    for mapping in normalized:
        previous = prior_mappings.get(str(mapping["account_id"]))
        previous_signature = (
            {
                key: previous.get(key)
                for key in (
                    "decision",
                    "reason",
                    "candidate_source",
                    "memory_scope",
                    "allocations",
                )
            }
            if previous
            else None
        )
        current_signature = {
            key: mapping.get(key)
            for key in (
                "decision",
                "reason",
                "candidate_source",
                "memory_scope",
                "allocations",
            )
        }
        if previous_signature is not None and previous_signature != current_signature:
            _record_event(
                case,
                "mapping_changed",
                actor,
                {
                    "account_id": mapping["account_id"],
                    "previous": _mapping_audit_summary(previous_signature),
                    "current": _mapping_audit_summary(current_signature),
                },
                after_hash=mapping_after_hash,
            )
        _record_event(
            case,
            (
                "mapping_accepted"
                if mapping["decision"] == "ACCEPTED"
                else "mapping_excluded"
            ),
            actor,
            {"account_id": mapping["account_id"]},
            after_hash=mapping_after_hash,
        )
        if len(mapping["allocations"]) > 1:
            _record_event(
                case,
                "account_split_created",
                actor,
                {
                    "account_id": mapping["account_id"],
                    "allocation_ids": [
                        item["allocation_id"] for item in mapping["allocations"]
                    ],
                },
                after_hash=mapping_after_hash,
            )
    for removed_id in sorted(set(prior_mappings) - seen):
        previous = prior_mappings[removed_id]
        _record_event(
            case,
            "mapping_changed",
            actor,
            {
                "account_id": removed_id,
                "previous": _mapping_audit_summary(
                    {
                        key: previous.get(key)
                        for key in (
                            "decision",
                            "reason",
                            "candidate_source",
                            "memory_scope",
                            "allocations",
                        )
                    }
                ),
                "current": None,
            },
            after_hash=mapping_after_hash,
        )
    return case


def record_adjustments(
    case: dict[str, Any],
    adjustments: Sequence[Mapping[str, Any]],
    actor: str,
    expected_revision: str,
) -> dict[str, Any]:
    """Record explicit, balancing presentation reclassifications."""

    _ensure_revision(case, expected_revision)
    if not case.get("mappings"):
        raise ValueError("Reviewed mappings are required before adjustments")
    normalized: list[dict[str, Any]] = []
    seen: set[str] = set()
    for raw in adjustments:
        adjustment_id = str(raw["adjustment_id"])
        if not SAFE_ID.fullmatch(adjustment_id) or adjustment_id in seen:
            raise ValueError("Adjustment identifiers must be unique safe IDs")
        seen.add(adjustment_id)
        reason = str(raw.get("reason", "")).strip()
        if not reason:
            raise ValueError("Every presentation adjustment requires a reason")
        raw_lines = list(raw.get("lines", []))
        if len(raw_lines) < 2:
            raise ValueError("A presentation adjustment requires at least two lines")
        lines: list[dict[str, Any]] = []
        current_total = Decimal("0")
        prior_total = Decimal("0")
        for position, raw_line in enumerate(raw_lines, start=1):
            current = normalize_decimal(raw_line["current_amount"])
            prior = normalize_decimal(raw_line["prior_amount"])
            current_total += current
            prior_total += prior
            qname = raw_line.get("xbrl_concept")
            if qname is not None and not SAFE_QNAME.fullmatch(str(qname)):
                raise ValueError(f"Invalid XBRL QName: {qname}")
            xbrl_sign_multiplier = str(raw_line.get("xbrl_sign_multiplier", ""))
            if qname is not None and xbrl_sign_multiplier not in {"1", "-1"}:
                raise ValueError(
                    "Adjusted XBRL concepts require an explicit sign multiplier"
                )
            if qname is None:
                xbrl_sign_multiplier = "1"
            lines.append(
                {
                    "line_id": f"{adjustment_id}_{position}",
                    "canonical_line": str(raw_line["canonical_line"]),
                    "statement_section": str(raw_line["statement_section"]).upper(),
                    "xbrl_concept": qname,
                    "xbrl_sign_multiplier": xbrl_sign_multiplier,
                    "current_amount": _decimal_text(current),
                    "prior_amount": _decimal_text(prior),
                    "source_refs": sorted(
                        {str(item) for item in raw_line.get("source_refs", [])}
                    ),
                }
            )
            if lines[-1]["source_refs"] and not set(
                lines[-1]["source_refs"]
            ) <= _available_evidence_refs(case):
                raise ValueError("Adjustment lines reference evidence outside the case")
        if current_total or prior_total:
            raise ValueError(
                "Presentation adjustment lines must balance in both periods"
            )
        normalized.append(
            {
                "adjustment_id": adjustment_id,
                "kind": "PRESENTATION_RECLASSIFICATION",
                "reason": reason,
                "lines": lines,
                "approved_by": actor,
                "approved_at": _now(),
            }
        )
    _mutate(case, actor, "adjustments_recorded")
    _clear_accounting_dependent_reviews(case)
    case["adjustments"] = normalized
    case["canonical_facts"] = []
    case["statements"] = None
    case["statutory_presentation"] = None
    case["schedules"] = []
    case["validation"] = None
    case["state"] = CaseState.MAPPING_REVIEW
    _record_event(
        case,
        "adjustments_recorded",
        actor,
        {"adjustment_ids": [item["adjustment_id"] for item in normalized]},
    )
    return case


def record_taxonomy_facts(
    case: dict[str, Any],
    facts: Sequence[Mapping[str, Any]],
    actor: str,
    expected_revision: str,
) -> dict[str, Any]:
    """Record reviewed text, nil, or dimensional facts for the taxonomy adapter."""

    _ensure_revision(case, expected_revision)
    if not case.get("selected_form") or not case.get("statements"):
        raise ValueError("Statements and form are required before taxonomy facts")
    normalized: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    seen_keys: set[tuple[str, str, tuple[tuple[str, str], ...]]] = set()
    allowed_periods = {
        "current_instant",
        "prior_instant",
        "current_duration",
        "prior_duration",
    }
    for raw in facts:
        fact_id = str(raw["fact_id"])
        qname = str(raw["xbrl_concept"])
        period = str(raw["period"])
        fact_type = str(raw["fact_type"]).upper()
        status = EvidenceStatus(str(raw["status"]).upper())
        if not SAFE_ID.fullmatch(fact_id) or fact_id in seen_ids:
            raise ValueError("Taxonomy fact IDs must be unique safe IDs")
        seen_ids.add(fact_id)
        if not SAFE_QNAME.fullmatch(qname):
            raise ValueError(f"Invalid taxonomy fact QName: {qname}")
        if period not in allowed_periods:
            raise ValueError(f"Unsupported taxonomy fact period: {period}")
        if fact_type not in {"MONETARY", "TEXT", "NIL"}:
            raise ValueError(f"Unsupported taxonomy fact type: {fact_type}")
        if status not in {
            EvidenceStatus.OBSERVED,
            EvidenceStatus.DERIVED,
            EvidenceStatus.USER_CONFIRMED,
        }:
            raise ValueError("Taxonomy facts must be reviewed and exportable")
        dimensions = {
            str(axis): str(member)
            for axis, member in sorted(dict(raw.get("dimensions", {})).items())
        }
        if any(
            not SAFE_QNAME.fullmatch(axis) or not SAFE_QNAME.fullmatch(member)
            for axis, member in dimensions.items()
        ):
            raise ValueError("Dimension axes and members must be valid QNames")
        duplicate_key = (qname, period, tuple(dimensions.items()))
        if duplicate_key in seen_keys:
            raise ValueError("Conflicting duplicate taxonomy fact")
        seen_keys.add(duplicate_key)
        source_refs = sorted({str(item) for item in raw.get("source_refs", [])})
        derivation = raw.get("derivation")
        value: str | None
        language: str | None = None
        nil_reason: str | None = None
        if fact_type == "MONETARY":
            if str(raw.get("currency", "EUR")).upper() != "EUR":
                raise ValueError("MVP taxonomy monetary facts must use EUR")
            value = _decimal_text(normalize_decimal(raw["value"]))
        elif fact_type == "TEXT":
            value = str(raw.get("value", "")).strip()
            if not value:
                raise ValueError("Text taxonomy facts require a value")
            language = str(
                raw.get("language", case.get("output_language", "it"))
            ).lower()
            if language != case.get("output_language", "it"):
                raise ValueError(
                    "Text taxonomy facts cannot mix languages in one output"
                )
        else:
            value = None
            nil_reason = str(raw.get("nil_reason", "")).strip()
            if raw.get("nil_allowed_by_review") is not True or not nil_reason:
                raise ValueError("Nil facts require an explicit reviewed reason")
        if fact_type != "NIL" and not source_refs and not derivation:
            raise ValueError("Taxonomy facts require source refs or a derivation")
        if source_refs and not set(source_refs) <= _available_evidence_refs(case):
            raise ValueError("Taxonomy fact references evidence outside the case")
        normalized.append(
            {
                "fact_id": fact_id,
                "xbrl_concept": qname,
                "period": period,
                "fact_type": fact_type,
                "value": value,
                "currency": "EUR" if fact_type == "MONETARY" else None,
                "language": language if fact_type == "TEXT" else None,
                "status": status.value,
                "source_refs": source_refs,
                "derivation": derivation,
                "dimensions": dimensions,
                "nil_reason": nil_reason,
                "confirmed_by": actor,
                "confirmed_at": _now(),
            }
        )
    _mutate(case, actor, "taxonomy_facts_recorded")
    _clear_schedule_taxonomy_adapter(case)
    _clear_narrative_reviews(case)
    case["taxonomy_facts"] = normalized
    case["taxonomy_fact_context"] = _computation_context(
        case, "taxonomy-fact-adapter-v1"
    )
    case["statutory_presentation"] = None
    case["validation"] = None
    _refresh_disclosures(case)
    _record_event(
        case,
        "taxonomy_facts_recorded",
        actor,
        {"fact_ids": [item["fact_id"] for item in normalized]},
    )
    return case


def record_schedule_taxonomy_adapter(
    case: dict[str, Any],
    catalogue_path: Path,
    rule_pack: Mapping[str, Any],
    decisions: Sequence[Mapping[str, Any]],
    actor: str,
    expected_revision: str,
) -> dict[str, Any]:
    """Compile professionally reviewed schedules into official note-table facts."""

    _ensure_revision(case, expected_revision)
    period_start = date.fromisoformat(case["period"]["start"])
    if not (
        date.fromisoformat(str(rule_pack["effective_from"]))
        <= period_start
        <= date.fromisoformat(str(rule_pack["effective_to"]))
    ):
        raise ValueError("Schedule taxonomy rule pack is not effective for the case")
    catalogue = _taxonomy_catalogue(
        catalogue_path,
        case["rule_pack_versions"]["taxonomy_id"],
        case.get("taxonomy_checksum"),
    )
    adapter = compile_schedule_taxonomy_adapter(
        case, catalogue, rule_pack, decisions, actor
    )
    checksum = _sha256_bytes(_canonical_json(rule_pack))
    _mutate(case, actor, "schedule_taxonomy_adapter_recorded")
    _clear_narrative_reviews(case)
    recorded_at = _now()
    for fact in adapter["generated_facts"]:
        fact["confirmed_at"] = recorded_at
    adapter["recorded_at"] = recorded_at
    adapter["computation_context"] = _computation_context(case, str(rule_pack["id"]))
    case["rule_pack_versions"]["schedule_taxonomy_adapter"] = str(rule_pack["id"])
    case["schedule_taxonomy_adapter_rule_pack_checksum"] = checksum
    case["schedule_taxonomy_adapter"] = adapter
    case["schedule_taxonomy_facts"] = list(adapter["generated_facts"])
    case["validation"] = None
    _record_event(
        case,
        "schedule_taxonomy_adapter_recorded",
        actor,
        {
            "rule_pack_id": str(rule_pack["id"]),
            "inventory_sha256": adapter["inventory"]["inventory_sha256"],
            "adapter_sha256": adapter["adapter_sha256"],
            "schedule_count": len(adapter["coverage"]),
            "generated_fact_count": len(adapter["generated_facts"]),
        },
    )
    return case


def record_taxonomy_mapping_index(
    case: dict[str, Any],
    catalogue_path: Path,
    rule_pack: Mapping[str, Any],
    actor: str,
    expected_revision: str,
) -> dict[str, Any]:
    """Expose the selected form's official concepts to semantic mapping review."""

    _ensure_revision(case, expected_revision)
    if not case.get("selected_form"):
        raise ValueError("A statutory form must be selected before taxonomy indexing")
    period_start = date.fromisoformat(case["period"]["start"])
    if not (
        date.fromisoformat(str(rule_pack["effective_from"]))
        <= period_start
        <= date.fromisoformat(str(rule_pack["effective_to"]))
    ):
        raise ValueError(
            "Statutory presentation rule pack is not effective for the case period"
        )
    presentation_checksum = _sha256_bytes(_canonical_json(rule_pack))
    locked_presentation_id = case["rule_pack_versions"].get(
        "statutory_presentation_rule_pack"
    )
    locked_presentation_checksum = case.get("statutory_presentation_rule_pack_checksum")
    if locked_presentation_id is not None and (
        str(locked_presentation_id) != str(rule_pack.get("id"))
        or locked_presentation_checksum != presentation_checksum
    ):
        raise ValueError(
            "Statutory presentation rule pack differs from the locked case"
        )
    catalogue = _taxonomy_catalogue(
        catalogue_path,
        case["rule_pack_versions"]["taxonomy_id"],
        case.get("taxonomy_checksum"),
    )
    inventory = build_primary_presentation_inventory(
        catalogue, rule_pack, str(case["selected_form"])
    )
    schedule_lookup: dict[str, list[str]] = {}
    for schedule_type, concepts in inventory["schedule_trigger_concepts"].items():
        for qname in concepts:
            schedule_lookup.setdefault(str(qname), []).append(str(schedule_type))
    concepts = []
    for kind, items in (
        ("LEAF", inventory["requirements"]),
        ("TOTAL", inventory["totals"]),
    ):
        for item in items:
            qname = str(item["xbrl_concept"])
            concepts.append(
                {
                    "xbrl_concept": qname,
                    "label_it": item["label_it"],
                    "period_type": item["period_type"],
                    "balance": item["balance"],
                    "role_kinds": list(item["role_kinds"]),
                    "presentation_kind": kind,
                    "mapping_allowed": kind == "LEAF",
                    "schedule_types": sorted(schedule_lookup.get(qname, [])),
                }
            )
    _mutate(case, actor, "taxonomy_mapping_index_recorded")
    case["rule_pack_versions"]["statutory_presentation_rule_pack"] = str(
        rule_pack["id"]
    )
    case["statutory_presentation_rule_pack_checksum"] = presentation_checksum
    case["taxonomy_mapping_index"] = {
        "selected_form": case["selected_form"],
        "inventory_sha256": inventory["inventory_sha256"],
        "taxonomy_package_sha256": inventory["taxonomy_package_sha256"],
        "concepts": sorted(concepts, key=lambda item: item["xbrl_concept"]),
        "recorded_by": actor,
        "recorded_at": _now(),
        "computation_context": _computation_context(
            case, f"{rule_pack['id']}-mapping-index"
        ),
    }
    case["mapping_candidates"] = []
    case["validation"] = None
    _record_event(
        case,
        "taxonomy_mapping_index_recorded",
        actor,
        {
            "selected_form": case["selected_form"],
            "concept_count": len(concepts),
            "inventory_sha256": inventory["inventory_sha256"],
        },
    )
    return case


def record_statutory_presentation(
    case: dict[str, Any],
    catalogue_path: Path,
    rule_pack: Mapping[str, Any],
    decisions: Sequence[Mapping[str, Any]],
    actor: str,
    expected_revision: str,
) -> dict[str, Any]:
    """Record complete primary-statement coverage against official networks."""

    _ensure_revision(case, expected_revision)
    if not case.get("taxonomy_mapping_index"):
        raise ValueError(
            "The selected-form taxonomy mapping index is required before presentation review"
        )
    if case["rule_pack_versions"].get("statutory_presentation_rule_pack") != str(
        rule_pack.get("id")
    ) or case.get("statutory_presentation_rule_pack_checksum") != _sha256_bytes(
        _canonical_json(rule_pack)
    ):
        raise ValueError(
            "Statutory presentation rule pack differs from the locked mapping index"
        )
    catalogue = _taxonomy_catalogue(
        catalogue_path,
        case["rule_pack_versions"]["taxonomy_id"],
        case.get("taxonomy_checksum"),
    )
    coverage = build_statutory_presentation_coverage(
        case, catalogue, rule_pack, decisions, actor
    )
    if (
        coverage["inventory"]["inventory_sha256"]
        != case["taxonomy_mapping_index"]["inventory_sha256"]
    ):
        raise ValueError("Statutory presentation inventory differs from mapping review")
    _mutate(case, actor, "statutory_presentation_recorded")
    _clear_schedule_taxonomy_adapter(case)
    coverage["recorded_by"] = actor
    coverage["recorded_at"] = _now()
    coverage["computation_context"] = _computation_context(case, str(rule_pack["id"]))
    case["statutory_presentation"] = coverage
    case["validation"] = None
    _refresh_disclosures(case)
    _record_event(
        case,
        "statutory_presentation_recorded",
        actor,
        {
            "status": coverage["status"],
            "inventory_sha256": coverage["inventory"]["inventory_sha256"],
            **coverage["summary"],
        },
    )
    return case


def record_taxonomy_representation(
    case: dict[str, Any],
    payload: Mapping[str, Any],
    reviewer: str,
    expected_revision: str,
) -> dict[str, Any]:
    """Record the professional treatment of substantive taxonomy differences."""

    _ensure_revision(case, expected_revision)
    mismatch_present = payload.get("mismatch_present") is True
    allowed_treatments = {
        "NO_SUBSTANTIVE_MISMATCH",
        "XBRL_REPRESENTATION_ACCEPTED_WITH_DISCLOSED_DIFFERENCES",
        "ACCOUNTS_REVISED_TO_TAXONOMY",
        "DOUBLE_FORMAT_ROUTE_REFERRED_FOR_PROFESSIONAL_FILING",
    }
    treatment = str(payload.get("chosen_treatment", "")).upper()
    if treatment not in allowed_treatments:
        raise ValueError("A supported taxonomy representation treatment is required")
    if mismatch_present and treatment == "NO_SUBSTANTIVE_MISMATCH":
        raise ValueError("A substantive mismatch requires a professional treatment")
    if not mismatch_present and treatment != "NO_SUBSTANTIVE_MISMATCH":
        raise ValueError("No-mismatch cases must use NO_SUBSTANTIVE_MISMATCH")
    reason = str(payload.get("reviewer_reason", "")).strip()
    affected_sections = sorted(
        {str(item) for item in payload.get("affected_sections", []) if str(item)}
    )
    differences = [
        {
            "difference_id": str(item["difference_id"]),
            "description": str(item["description"]).strip(),
            "affected_facts": sorted(
                {str(ref) for ref in item.get("affected_facts", [])}
            ),
            "source_refs": sorted({str(ref) for ref in item.get("source_refs", [])}),
        }
        for item in payload.get("differences", [])
    ]
    if mismatch_present and (not reason or not affected_sections or not differences):
        raise ValueError(
            "A substantive taxonomy mismatch requires sections, differences, and a reviewer reason"
        )
    if any(
        not SAFE_ID.fullmatch(item["difference_id"]) or not item["description"]
        for item in differences
    ):
        raise ValueError("Taxonomy differences require safe IDs and descriptions")
    if len({item["difference_id"] for item in differences}) != len(differences):
        raise ValueError("Taxonomy difference IDs must be unique")
    if any(
        item["source_refs"]
        and not set(item["source_refs"]) <= _available_evidence_refs(case)
        for item in differences
    ):
        raise ValueError("Taxonomy difference references evidence outside the case")
    _mutate(case, reviewer, "taxonomy_representation_recorded")
    case["taxonomy_representation"] = {
        "mismatch_present": mismatch_present,
        "affected_sections": affected_sections,
        "differences": differences,
        "chosen_treatment": treatment,
        "reviewer_reason": reason,
        "reviewed_by": reviewer,
        "reviewed_at": _now(),
        "vera_did_not_select_filing_route": True,
    }
    case["validation"] = None
    _record_event(
        case,
        "taxonomy_representation_recorded",
        reviewer,
        {
            "mismatch_present": mismatch_present,
            "chosen_treatment": treatment,
        },
    )
    return case


def record_micro_reporting(
    case: dict[str, Any],
    payload: Mapping[str, Any],
    reviewer: str,
    expected_revision: str,
) -> dict[str, Any]:
    """Record whether a micro case uses reviewed footer disclosures or notes."""

    _ensure_revision(case, expected_revision)
    if case.get("selected_form") != "MICRO":
        raise ValueError("Micro reporting choices apply only to the MICRO form")
    mode = str(payload["mode"]).upper()
    if mode not in {"FOOTER_ONLY", "NOTES"}:
        raise ValueError("Micro reporting mode must be FOOTER_ONLY or NOTES")
    required_footer_keys = {
        "guarantees_commitments_contingencies",
        "director_auditor_compensation",
        "own_and_parent_shares",
    }
    footer_items: list[dict[str, Any]] = []
    if mode == "FOOTER_ONLY":
        supplied = {str(item["key"]): item for item in payload.get("footer_items", [])}
        if set(supplied) != required_footer_keys:
            raise ValueError(
                "Footer-only micro accounts require every statutory footer item"
            )
        for key in sorted(required_footer_keys):
            raw = supplied[key]
            status = str(raw["status"]).upper()
            value = raw.get("value")
            reason = str(raw.get("reason", "")).strip()
            if status not in {"PRESENT", "NOT_APPLICABLE_CONFIRMED"}:
                raise ValueError("Micro footer items require a reviewed status")
            if status == "PRESENT" and value in {None, ""}:
                raise ValueError("Present micro footer items require reviewed content")
            if status == "NOT_APPLICABLE_CONFIRMED" and not reason:
                raise ValueError("Negative micro footer items require a reason")
            footer_items.append(
                {
                    "key": key,
                    "status": status,
                    "value": value,
                    "reason": reason,
                    "source_refs": sorted(
                        {str(ref) for ref in raw.get("source_refs", [])}
                    ),
                }
            )
            if footer_items[-1]["source_refs"] and not set(
                footer_items[-1]["source_refs"]
            ) <= _available_evidence_refs(case):
                raise ValueError("Micro footer references evidence outside the case")
    output_contract = (case.get("taxonomy_output_contracts") or {}).get(
        "MICRO_FOOTER_TEXT"
    )
    if not isinstance(output_contract, Mapping):
        raise ValueError("The case has no versioned micro-footer output contract")
    output_language = str(case.get("output_language", "it"))
    labels = output_contract["labels"][output_language]
    negative_text = str(output_contract["not_applicable_text"][output_language])
    rendered_text = None
    output_source_refs: list[str] = []
    if mode == "FOOTER_ONLY":
        rendered_lines = []
        for item in footer_items:
            content = (
                str(item["value"]).strip()
                if item["status"] == "PRESENT"
                else f"{negative_text} ({item['reason']})"
            )
            rendered_lines.append(f"{labels[item['key']]}: {content}.")
            output_source_refs.extend(item["source_refs"])
        rendered_text = " ".join(rendered_lines)
    _mutate(case, reviewer, "micro_reporting_recorded")
    _clear_narrative_reviews(case)
    case["micro_reporting"] = {
        "mode": mode,
        "status": "CONFIRMED",
        "footer_items": footer_items,
        "xbrl_concept": (
            str(output_contract["xbrl_concept"]) if mode == "FOOTER_ONLY" else None
        ),
        "rendered_text": rendered_text,
        "source_refs": sorted(set(output_source_refs)),
        "reviewed_by": reviewer,
        "reviewed_at": _now(),
    }
    case["validation"] = None
    _refresh_disclosures(case)
    _record_event(
        case,
        "micro_reporting_recorded",
        reviewer,
        {"mode": mode},
    )
    return case


def build_statements(
    case: dict[str, Any], actor: str, expected_revision: str
) -> dict[str, Any]:
    """Aggregate reviewed mappings into exact current/comparative statement facts."""

    _ensure_revision(case, expected_revision)
    entries = (case.get("trial_balance") or {}).get("entries", [])
    mapped_ids = {item["account_id"] for item in case.get("mappings", [])}
    if {row["account_id"] for row in entries} != mapped_ids:
        raise ValueError("Every account must be mapped or explicitly excluded")
    grouped: dict[tuple[str, str, str | None, str], dict[str, Any]] = {}
    for mapping in case["mappings"]:
        for allocation in mapping["allocations"]:
            key = (
                allocation["statement_section"],
                allocation["canonical_line"],
                allocation["xbrl_concept"],
                allocation["xbrl_sign_multiplier"],
            )
            item = grouped.setdefault(
                key,
                {
                    "statement_section": key[0],
                    "canonical_line": key[1],
                    "xbrl_concept": key[2],
                    "xbrl_sign_multiplier": key[3],
                    "current_amount": Decimal("0"),
                    "prior_amount": Decimal("0"),
                    "source_refs": [],
                    "allocation_refs": [],
                    "adjustment_refs": [],
                },
            )
            item["current_amount"] += Decimal(allocation["current_amount"])
            item["prior_amount"] += Decimal(allocation["prior_amount"])
            item["source_refs"].extend(allocation["source_refs"])
            item["allocation_refs"].append(allocation["allocation_id"])
    for adjustment in case.get("adjustments", []):
        for line in adjustment["lines"]:
            key = (
                line["statement_section"],
                line["canonical_line"],
                line["xbrl_concept"],
                line["xbrl_sign_multiplier"],
            )
            item = grouped.setdefault(
                key,
                {
                    "statement_section": key[0],
                    "canonical_line": key[1],
                    "xbrl_concept": key[2],
                    "xbrl_sign_multiplier": key[3],
                    "current_amount": Decimal("0"),
                    "prior_amount": Decimal("0"),
                    "source_refs": [],
                    "allocation_refs": [],
                    "adjustment_refs": [],
                },
            )
            item["current_amount"] += Decimal(line["current_amount"])
            item["prior_amount"] += Decimal(line["prior_amount"])
            item["source_refs"].extend(line["source_refs"])
            item["adjustment_refs"].append(line["line_id"])
    facts: list[dict[str, Any]] = []
    for position, item in enumerate(grouped.values(), start=1):
        facts.append(
            {
                "fact_id": f"fact_{position:06d}",
                "domain": "statements",
                "key": item["canonical_line"],
                "statement_section": item["statement_section"],
                "xbrl_concept": item["xbrl_concept"],
                "xbrl_sign_multiplier": item["xbrl_sign_multiplier"],
                "period": case["period"]["end"],
                "current_value": _decimal_text(item["current_amount"]),
                "prior_value": _decimal_text(item["prior_amount"]),
                "currency": "EUR",
                "status": EvidenceStatus.DERIVED,
                "source_refs": sorted(set(item["source_refs"])),
                "derivation": {
                    "operation": "SUM_REVIEWED_ALLOCATIONS_AND_ADJUSTMENTS",
                    "allocation_refs": item["allocation_refs"],
                    "adjustment_refs": item["adjustment_refs"],
                },
                "confirmed_by": None,
                "confidence": "HIGH",
                "materiality": "UNASSESSED",
                "rule_pack": case["rule_pack_versions"]["statutory_rule_pack"],
            }
        )
    totals: dict[str, dict[str, str]] = {}
    for section in sorted({fact["statement_section"] for fact in facts}):
        selected = [fact for fact in facts if fact["statement_section"] == section]
        totals[section] = {
            "current": _decimal_text(
                sum(Decimal(fact["current_value"]) for fact in selected)
            ),
            "prior": _decimal_text(
                sum(Decimal(fact["prior_value"]) for fact in selected)
            ),
        }
    precision = int(case.get("reporting_precision", 0))
    presentation_facts = [
        {
            "fact_id": fact["fact_id"],
            "current_value": _decimal_text(
                _reported_decimal(Decimal(fact["current_value"]), precision)
            ),
            "prior_value": _decimal_text(
                _reported_decimal(Decimal(fact["prior_value"]), precision)
            ),
        }
        for fact in facts
    ]
    presentation_lookup = {item["fact_id"]: item for item in presentation_facts}
    rounding_adjustments: list[dict[str, Any]] = []
    for section, exact in totals.items():
        section_facts = [fact for fact in facts if fact["statement_section"] == section]
        for period_key, value_key in (
            ("current", "current_value"),
            ("prior", "prior_value"),
        ):
            rounded_total = _reported_decimal(Decimal(exact[period_key]), precision)
            displayed_sum = sum(
                Decimal(presentation_lookup[fact["fact_id"]][value_key])
                for fact in section_facts
            )
            difference = rounded_total - displayed_sum
            if difference:
                rounding_adjustments.append(
                    {
                        "adjustment_id": (f"rounding_{section.lower()}_{period_key}"),
                        "kind": "PRESENTATION_ROUNDING",
                        "statement_section": section,
                        "period": period_key,
                        "amount": _decimal_text(difference),
                        "precision": precision,
                        "derivation": {
                            "operation": "ROUNDED_EXACT_TOTAL_MINUS_SUM_ROUNDED_LINES",
                            "fact_refs": [fact["fact_id"] for fact in section_facts],
                        },
                        "repairs_substantive_imbalance": False,
                    }
                )
    _mutate(case, actor, "statements_computed")
    _clear_accounting_dependent_reviews(case, preserve_adjustments=True)
    case["canonical_facts"] = facts
    case["statements"] = {
        "facts": facts,
        "section_totals": totals,
        "presentation_facts": presentation_facts,
        "rounding_adjustments": rounding_adjustments,
        "reporting_precision": precision,
        "computed_at": _now(),
        "computation_context": _computation_context(case, "statement-engine-v1"),
    }
    case["statutory_presentation"] = None
    case["state"] = CaseState.STATEMENT_REVIEW
    case["schedules"] = []
    case["validation"] = None
    _record_event(case, "statements_computed", actor, {"fact_count": len(facts)})
    _record_event(
        case,
        "fact_derived",
        actor,
        {"fact_ids": [fact["fact_id"] for fact in facts]},
    )
    return case


def record_schedule(
    case: dict[str, Any],
    payload: Mapping[str, Any],
    actor: str,
    expected_revision: str,
) -> dict[str, Any]:
    """Record a source-evidenced schedule and its exact reconciliation results."""

    _ensure_revision(case, expected_revision)
    statements = case.get("statements")
    if not statements:
        raise ValueError("Statements must be computed before schedules")
    reviewed_payload = deepcopy(dict(payload))
    valid_evidence_refs = _available_evidence_refs(case)
    schedule_id = str(reviewed_payload.get("schedule_id", "schedule"))
    collections = [
        ("rows", "row_id"),
        ("items", "item_id"),
    ]
    for collection_name, identifier_key in collections:
        normalized_items = []
        for position, raw_item in enumerate(
            reviewed_payload.get(collection_name, []), start=1
        ):
            item = dict(raw_item)
            item_id = str(item.get(identifier_key) or position)
            status = str(item.get("evidence_status", "")).upper()
            refs = sorted(
                {
                    str(ref).strip()
                    for ref in item.get("source_refs", [])
                    if str(ref).strip()
                }
            )
            if status == EvidenceStatus.OBSERVED:
                if not refs or not set(refs) <= valid_evidence_refs:
                    raise ValueError(
                        "Observed schedule rows must cite evidence anchored in the case"
                    )
            elif status == EvidenceStatus.USER_CONFIRMED:
                anchored = sorted(set(refs) & valid_evidence_refs)
                item["source_refs"] = anchored or [
                    f"professional_input:{actor}:{schedule_id}:{item_id}"
                ]
            normalized_items.append(item)
        if collection_name in reviewed_payload:
            reviewed_payload[collection_name] = normalized_items
    exception = reviewed_payload.get("amortisation_reconciliation_exception")
    if isinstance(exception, Mapping):
        normalized_exception = dict(exception)
        anchored = sorted(
            set(str(ref) for ref in exception.get("source_refs", []))
            & valid_evidence_refs
        )
        normalized_exception["source_refs"] = anchored or [
            f"professional_input:{actor}:{schedule_id}:amortisation_exception"
        ]
        reviewed_payload["amortisation_reconciliation_exception"] = normalized_exception
    schedule = normalize_schedule(reviewed_payload, statements["facts"])
    if (
        schedule["schedule_type"] == "CASH_FLOW"
        and case.get("selected_form") != "ORDINARY"
    ):
        raise ValueError("A cash-flow statement belongs to ordinary accounts in MVP")
    existing = [
        item
        for item in case.get("schedules", [])
        if item["schedule_id"] != schedule["schedule_id"]
    ]
    _mutate(case, actor, "schedule_recorded")
    _clear_schedule_taxonomy_adapter(case)
    _clear_narrative_reviews(case)
    schedule["computation_context"] = _computation_context(
        case, f"schedule-{str(schedule['schedule_type']).lower()}-v1"
    )
    case["schedules"] = [*existing, schedule]
    case["validation"] = None
    _refresh_disclosures(case)
    _record_event(
        case,
        "schedule_recorded",
        actor,
        {
            "schedule_id": schedule["schedule_id"],
            "schedule_type": schedule["schedule_type"],
            "status": schedule["status"],
        },
    )
    return case


def ingest_schedule_file(
    case: dict[str, Any],
    source_path: Path,
    schedule_type: str,
    schedule_id: str,
    statement_line: str | None,
    options: Mapping[str, Any],
    actor: str,
    expected_revision: str,
    sheet: str | None = None,
) -> dict[str, Any]:
    """Import a documented CSV/XLSX schedule template with per-cell anchors."""

    _ensure_revision(case, expected_revision)
    if not case.get("statements"):
        raise ValueError("Statements must be computed before schedule ingestion")
    if source_path.is_symlink() or not source_path.is_file():
        raise ValueError("Schedule source must be a regular local file")
    normalized_type = schedule_type.upper()
    required_fields = schedule_template_fields(normalized_type)
    required_text_fields = schedule_template_text_fields(normalized_type)
    headers, raw_rows, sheet_name = _load_table(source_path.resolve(), sheet)
    header_columns = _normalized_header_columns(headers)
    normalized_headers = [item["normalized_column"] for item in header_columns]
    required_headers = {"row_id", *required_fields, *required_text_fields}
    missing_headers = sorted(required_headers - set(normalized_headers))
    if missing_headers:
        raise ValueError(
            f"Schedule template is missing columns: {', '.join(missing_headers)}"
        )
    document_id = _next_document_id(case)
    anchors: list[dict[str, Any]] = []
    rows: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row_number, values in enumerate(raw_rows, start=2):
        row = {
            name: values[position] if position < len(values) else None
            for position, name in enumerate(normalized_headers)
        }
        row_id = str(row.get("row_id") or "").strip()
        if not row_id and all(value in {None, ""} for value in row.values()):
            continue
        if not row_id or row_id in seen:
            raise ValueError("Schedule row IDs must be present and unique")
        seen.add(row_id)
        source_refs: list[str] = []
        normalized_row: dict[str, Any] = {
            "row_id": row_id,
            "label": str(row.get("label") or row_id),
            "evidence_status": EvidenceStatus.OBSERVED,
        }
        for column in header_columns:
            position = int(column["column_index"]) - 1
            column_name = str(column["normalized_column"])
            anchor_id = f"src_{document_id}_{len(anchors) + 1:07d}"
            raw_value = values[position] if position < len(values) else None
            anchor: dict[str, Any] = {
                "source_ref": anchor_id,
                "document_id": document_id,
                "sheet": sheet_name,
                "row": row_number,
                "column": column["column"],
                "column_header": column["column_header"],
                "normalized_column": column_name,
                "raw_value": "" if raw_value is None else str(raw_value),
                "parser_profile": f"{normalized_type.lower()}_template_v1",
                "confidence": "HIGH",
            }
            if column_name in required_fields:
                value = normalize_decimal(raw_value)
                normalized_row[column_name] = _decimal_text(value)
                anchor["normalized_value"] = _decimal_text(value)
            elif column_name in required_text_fields:
                text_value = str(raw_value or "").strip()
                normalized_row[column_name] = text_value or "UNKNOWN"
                anchor["normalized_value"] = text_value or "UNKNOWN"
            else:
                anchor["normalized_value"] = str(raw_value or "").strip()
            anchors.append(anchor)
            source_refs.append(anchor_id)
        normalized_row["source_refs"] = source_refs
        rows.append(normalized_row)
    payload = {
        "schedule_id": schedule_id,
        "schedule_type": normalized_type,
        "statement_line": statement_line,
        "rows": rows,
        **dict(options),
    }
    schedule = normalize_schedule(payload, case["statements"]["facts"])
    document = {
        "document_id": document_id,
        "purpose": f"{normalized_type}_SCHEDULE",
        "file_name": source_path.name,
        "media_type": (
            "text/csv"
            if source_path.suffix.lower() == ".csv"
            else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        "sha256": _sha256_file(source_path.resolve()),
        "size_bytes": source_path.resolve().stat().st_size,
        "sheet": sheet_name,
        "columns": header_columns,
        "parser_profile": f"{normalized_type.lower()}_template_v1",
        "source_anchors": anchors,
        "parsed_at": _now(),
    }
    existing = [
        item for item in case.get("schedules", []) if item["schedule_id"] != schedule_id
    ]
    _mutate(case, actor, "schedule_document_parsed")
    _clear_schedule_taxonomy_adapter(case)
    _clear_narrative_reviews(case)
    case["source_documents"].append(document)
    schedule["computation_context"] = _computation_context(
        case, f"schedule-{normalized_type.lower()}-template-v1"
    )
    case["schedules"] = [*existing, schedule]
    case["validation"] = None
    _refresh_disclosures(case)
    _record_event(case, "document_uploaded", actor, {"document_id": document_id})
    _record_evidence_attached(case, actor, document)
    _record_event(
        case,
        "document_parsed",
        actor,
        {"document_id": document_id, "schedule_type": normalized_type},
    )
    _record_event(
        case,
        "schedule_recorded",
        actor,
        {"schedule_id": schedule_id, "status": schedule["status"]},
    )
    return case


def _refresh_disclosures(case: dict[str, Any]) -> None:
    rule_pack = case.get("disclosure_rule_pack")
    if not rule_pack or not case.get("selected_form") or not case.get("statements"):
        return
    coverage = build_disclosure_coverage(case, rule_pack)
    coverage["computation_context"] = _computation_context(
        case, "disclosure-coverage-v1"
    )
    case["disclosure_coverage"] = coverage
    case["questionnaire"] = coverage["questions"]
    outline = note_outline(case)
    accepted_sections = {
        block["section_id"]
        for block in case.get("narrative_blocks", [])
        if block.get("status") == "ACCEPTED"
    }
    for section in outline:
        if section["section_id"] in accepted_sections:
            section["status"] = "ACCEPTED"
    case["note_outline"] = outline
    case["note_outline_context"] = _computation_context(case, "note-outline-v1")
    _update_prevalidation_state(case)


def record_disclosure_trigger_decisions(
    case: dict[str, Any],
    decisions: Sequence[Mapping[str, Any]],
    actor: str,
    expected_revision: str,
) -> dict[str, Any]:
    """Record professional decisions for semantically ambiguous applicability."""

    _ensure_revision(case, expected_revision)
    rule_pack = case.get("disclosure_rule_pack")
    if not rule_pack:
        raise ValueError("Disclosure rules must be activated before trigger review")
    valid_flags = manual_disclosure_flags(rule_pack)
    if not decisions:
        raise ValueError("At least one disclosure trigger decision is required")
    normalized: list[dict[str, Any]] = []
    seen: set[str] = set()
    for raw in decisions:
        if set(raw) != {"flag", "status", "reason", "source_refs"}:
            raise ValueError(
                "Disclosure trigger decisions require flag, status, reason, and source_refs"
            )
        flag = str(raw["flag"])
        status = str(raw["status"]).upper()
        reason = str(raw["reason"]).strip()
        raw_source_refs = raw["source_refs"]
        if not isinstance(raw_source_refs, list):
            raise ValueError("Disclosure trigger source_refs must be a list")
        source_refs = sorted(
            {str(item) for item in raw_source_refs if str(item).strip()}
        )
        if flag not in valid_flags or flag in seen:
            raise ValueError(f"Unknown or duplicate disclosure trigger flag: {flag}")
        if status not in {"TRIGGERED", "NOT_APPLICABLE_CONFIRMED"}:
            raise ValueError("Disclosure trigger status is not reviewable")
        if not reason or not source_refs:
            raise ValueError(
                "Disclosure trigger decisions require a reason and evidence references"
            )
        if not set(source_refs) <= _available_evidence_refs(case):
            raise ValueError(
                "Disclosure trigger decision references evidence outside the case"
            )
        seen.add(flag)
        normalized.append(
            {
                "flag": flag,
                "status": status,
                "reason": reason,
                "source_refs": source_refs,
                "reviewed_by": actor,
                "reviewed_at": _now(),
            }
        )
    retained = [
        item
        for item in case.get("disclosure_trigger_decisions", [])
        if item["flag"] not in seen
    ]
    _mutate(case, actor, "disclosure_trigger_decisions_recorded")
    _clear_narrative_reviews(case)
    case["disclosure_trigger_decisions"] = [*retained, *normalized]
    case["disclosure_trigger_flags"] = sorted(
        item["flag"]
        for item in case["disclosure_trigger_decisions"]
        if item["status"] == "TRIGGERED"
    )
    case["validation"] = None
    _refresh_disclosures(case)
    _record_event(
        case,
        "disclosure_trigger_decisions_recorded",
        actor,
        {"flags": sorted(seen)},
    )
    return case


def _update_prevalidation_state(case: dict[str, Any]) -> None:
    """Distinguish source-backed data gaps from narrative drafting work."""

    if case.get("state") in {
        CaseState.APPROVED,
        CaseState.EXPORTED,
        CaseState.ARCHIVED,
        CaseState.UNSUPPORTED,
    }:
        return
    required_schedules = required_schedule_types(case)
    schedules = case.get("schedules", [])
    present_schedules = {str(item["schedule_type"]) for item in schedules}
    schedule_gaps = bool(required_schedules - present_schedules) or any(
        item.get("status") != "COMPLETE" for item in schedules
    )
    coverage = (case.get("disclosure_coverage") or {}).get("coverage", [])
    structured_gaps = any(
        item.get("triggered")
        and any(
            not requirement.get("complete")
            and requirement.get("kind") != "NARRATIVE_SECTION"
            for requirement in item.get("requirements", [])
        )
        for item in coverage
    )
    case["state"] = (
        CaseState.DATA_GAPS
        if schedule_gaps or structured_gaps
        else CaseState.NOTE_DRAFT
    )


def activate_disclosures(
    case: dict[str, Any],
    rule_pack: Mapping[str, Any],
    actor: str,
    expected_revision: str,
) -> dict[str, Any]:
    """Lock an effective-dated disclosure pack and build relevant questions."""

    _ensure_revision(case, expected_revision)
    if not case.get("selected_form") or not case.get("statements"):
        raise ValueError("Statements and statutory form are required for disclosures")
    incoming_checksum = disclosure_rule_pack_hash(rule_pack)
    locked_id = case["rule_pack_versions"].get("disclosure_rule_pack")
    locked_checksum = case.get("disclosure_rule_pack_checksum")
    if locked_id is not None and (
        str(rule_pack.get("id")) != str(locked_id)
        or incoming_checksum != locked_checksum
    ):
        raise ValueError(
            "Disclosure rule pack differs from the locked case; use explicit migration"
        )
    _mutate(case, actor, "disclosure_rules_activated")
    case["disclosure_rule_pack"] = deepcopy(dict(rule_pack))
    case["rule_pack_versions"]["disclosure_rule_pack"] = str(rule_pack["id"])
    case["disclosure_rule_pack_checksum"] = incoming_checksum
    history_narratives = list(
        (case.get("client_history_suggestions") or {}).get("narrative_suggestions", [])
    )
    case["prior_narrative_suggestions"] = [
        *prior_narrative_suggestions(case),
        *history_narratives,
    ]
    case["validation"] = None
    _refresh_disclosures(case)
    _record_event(
        case,
        "disclosure_rules_activated",
        actor,
        {
            "rule_pack": rule_pack["id"],
            "triggered_count": case["disclosure_coverage"]["triggered_count"],
        },
    )
    return case


def record_narrative_blocks(
    case: dict[str, Any],
    blocks: Sequence[Mapping[str, Any]],
    actor: str,
    expected_revision: str,
) -> dict[str, Any]:
    """Record sentence-level provenance-bound note blocks for reviewer acceptance."""

    _ensure_revision(case, expected_revision)
    if not case.get("disclosure_rule_pack"):
        raise ValueError("Disclosure rules must be activated before note blocks")
    normalized = normalize_narrative_blocks(
        blocks, actor, str(case.get("output_language", "it"))
    )
    accepted_answers = {
        str(answer["key"])
        for answer in case.get("disclosure_answers", [])
        if disclosure_answer_complete(answer)
    }
    valid_refs = {
        str(fact["fact_id"]) for fact in case.get("canonical_facts", [])
    } | accepted_answers
    valid_refs.update(
        str(ref)
        for answer in case.get("disclosure_answers", [])
        for ref in answer.get("source_refs", [])
    )
    valid_refs.update(
        str(ref)
        for schedule in case.get("schedules", [])
        for row in schedule.get("rows", [])
        for ref in row.get("source_refs", [])
    )
    structured_facts = {
        str(fact["fact_id"]): fact
        for fact in [
            *case.get("canonical_facts", []),
            *case.get("taxonomy_facts", []),
            *case.get("schedule_taxonomy_facts", []),
            *[
                schedule_fact
                for schedule in case.get("schedules", [])
                if schedule.get("status") == "COMPLETE"
                for schedule_fact in schedule_fact_records(schedule)
            ],
        ]
    }
    valid_refs.update(structured_facts)
    accepted_qnames: set[str] = set()
    for block in normalized:
        qname = block.get("xbrl_concept")
        if qname is not None and not SAFE_QNAME.fullmatch(str(qname)):
            raise ValueError(f"Invalid narrative XBRL QName: {qname}")
        if qname and block["status"] != "ACCEPTED":
            raise ValueError("Only accepted narrative may be bound to an XBRL concept")
        if block["status"] == "ACCEPTED" and not qname:
            raise ValueError("Accepted narrative requires an XBRL concept")
        if qname in accepted_qnames:
            raise ValueError("Accepted narrative XBRL concepts must be unique")
        if qname:
            accepted_qnames.add(str(qname))
        for claim in block["claims"]:
            if (
                claim["kind"] == "FACTUAL"
                and not set(claim["source_refs"]) <= valid_refs
            ):
                raise ValueError(
                    "Narrative factual claim references unaccepted evidence"
                )
            if claim["kind"] != "FACTUAL" or block["status"] != "ACCEPTED":
                continue
            support = claim.get("semantic_support")
            if (
                not isinstance(support, Mapping)
                or str(support.get("status", "")).upper() != "SUPPORTED"
                or not str(support.get("reason", "")).strip()
            ):
                raise ValueError(
                    "Accepted factual narrative requires explicit semantic support"
                )
            assertions: list[Decimal] = []
            for assertion in claim.get("fact_assertions", []):
                if set(assertion) != {"fact_ref", "value_field", "value"}:
                    raise ValueError(
                        "Narrative fact assertions require fact_ref, value_field, and value"
                    )
                fact_ref = str(assertion["fact_ref"])
                value_field = str(assertion["value_field"])
                if (
                    fact_ref not in claim["source_refs"]
                    or fact_ref not in structured_facts
                ):
                    raise ValueError(
                        "Narrative fact assertion must reference cited structured evidence"
                    )
                fact = structured_facts[fact_ref]
                if value_field not in {"current_value", "prior_value", "value"} or (
                    value_field not in fact
                ):
                    raise ValueError(
                        "Narrative fact assertion uses an invalid value field"
                    )
                asserted_value = normalize_decimal(assertion["value"])
                if asserted_value != normalize_decimal(fact[value_field]):
                    raise ValueError(
                        "Narrative fact assertion does not match structured evidence"
                    )
                assertions.append(asserted_value)
            literal_values = [
                _normalize_narrative_money_literal(
                    match, str(case.get("output_language", "it"))
                )
                for match in MONEY_LITERAL.findall(str(claim["sentence"]))
            ]
            if sorted(literal_values) != sorted(assertions):
                raise ValueError(
                    "Narrative monetary statements must exactly match fact assertions"
                )
        prior_id = block.get("prior_suggestion_id")
        if prior_id:
            prior = next(
                (
                    item
                    for item in case.get("prior_narrative_suggestions", [])
                    if item["suggestion_id"] == prior_id
                ),
                None,
            )
            if prior is None:
                raise ValueError(
                    "Narrative block references an unknown prior suggestion"
                )
            if prior.get("language") not in {
                None,
                "",
                case.get("output_language", "it"),
            }:
                raise ValueError(
                    "Prior narrative language differs from the current output"
                )
            block["redline"] = narrative_redline(prior["text"], block["text"])
            block["prior_source_refs"] = list(prior["source_refs"])
    _mutate(case, actor, "narrative_blocks_recorded")
    case["narrative_blocks"] = normalized
    case["narrative_context"] = _computation_context(case, "reviewed-narrative-v1")
    case["validation"] = None
    _refresh_disclosures(case)
    _record_event(
        case,
        "narrative_edited",
        actor,
        {
            "block_count": len(normalized),
            "accepted_count": sum(
                block["status"] == "ACCEPTED" for block in normalized
            ),
        },
    )
    return case


def record_intelligence_suggestion(
    case: dict[str, Any],
    task: IntelligenceTask | str,
    subject_ids: Sequence[str],
    output: Mapping[str, Any],
    model_metadata: Mapping[str, Any],
    actor: str,
    expected_revision: str,
) -> dict[str, Any]:
    """Record a validated, non-authoritative semantic suggestion for review."""

    _ensure_revision(case, expected_revision)
    if case.get("state") in {
        CaseState.APPROVED,
        CaseState.EXPORTED,
        CaseState.ARCHIVED,
    }:
        raise ValueError("Intelligence suggestions cannot mutate a locked case")
    required_metadata = {"provider", "model", "prompt_template_version"}
    if set(model_metadata) != required_metadata or any(
        not str(model_metadata[key]).strip() for key in required_metadata
    ):
        raise ValueError(
            "Model metadata requires provider, model, and prompt_template_version"
        )
    packet = build_intelligence_packet(case, task, subject_ids)
    normalized = validate_intelligence_output(packet, output)
    selected_task = IntelligenceTask(str(task))
    _mutate(case, actor, "intelligence_suggestion_recorded")
    run = {
        "run_id": f"intel_{len(case.get('intelligence_runs', [])) + 1:06d}",
        "task": selected_task.value,
        "subject_ids": list(subject_ids),
        "status": EvidenceStatus.MODEL_SUGGESTED,
        "input_packet_hash": intelligence_packet_hash(packet),
        "input_revision_id": expected_revision,
        "model_metadata": dict(model_metadata),
        "output": normalized,
        "recorded_by": actor,
        "recorded_at": _now(),
        "requires_review": True,
        "computation_context": _computation_context(
            case,
            str(model_metadata["prompt_template_version"]),
            model_version=str(model_metadata["model"]),
        ),
    }
    case.setdefault("intelligence_runs", []).append(run)
    if selected_task is IntelligenceTask.WORKFLOW_GUIDANCE:
        case["latest_workflow_guidance"] = {
            "run_id": run["run_id"],
            **normalized,
        }
    elif selected_task is IntelligenceTask.ACCOUNT_MAPPING:
        case["model_mapping_suggestions"] = list(normalized["suggestions"])
    elif selected_task is IntelligenceTask.DISCLOSURE_ACTIVATION:
        case["disclosure_activation_suggestions"] = list(normalized["suggestions"])
    elif selected_task is IntelligenceTask.NARRATIVE_DRAFT:
        case["narrative_suggestions"] = list(normalized["blocks"])
    case["validation"] = None
    if selected_task is IntelligenceTask.ACCOUNT_MAPPING:
        _record_event(
            case,
            "mapping_suggested",
            actor,
            {
                "run_id": run["run_id"],
                "candidate_count": len(normalized["suggestions"]),
                "suggestion_source": "MODEL",
            },
        )
    elif selected_task is IntelligenceTask.NARRATIVE_DRAFT:
        _record_event(
            case,
            "narrative_generated",
            actor,
            {
                "run_id": run["run_id"],
                "block_count": len(normalized["blocks"]),
                "status": "MODEL_SUGGESTED",
            },
        )
    _record_event(
        case,
        "model_suggestion_recorded",
        actor,
        {"run_id": run["run_id"], "task": selected_task.value},
    )
    return case


def render_preview_html(case: Mapping[str, Any]) -> bytes:
    """Render a safe review preview from structured case data only."""

    statements = case.get("statements") or {}
    facts = statements.get("facts", [])
    schedules = case.get("schedules", [])
    questions = case.get("questionnaire", [])
    narratives = case.get("narrative_blocks", [])
    taxonomy_facts = [
        *case.get("taxonomy_facts", []),
        *case.get("schedule_taxonomy_facts", []),
    ]
    statutory_presentation = case.get("statutory_presentation") or {}
    presentation_summary = statutory_presentation.get("summary") or {}
    micro_reporting = case.get("micro_reporting") or {}
    issues = (case.get("validation") or {}).get("issues", [])
    output_language = str(case.get("output_language", "it"))

    def cell(value: Any) -> str:
        return html.escape("" if value is None else str(value))

    statement_rows = "".join(
        "<tr>"
        f"<td>{cell(fact.get('statement_section'))}</td>"
        f"<td>{cell(fact.get('key'))}</td>"
        f"<td>{cell(fact.get('current_value'))}</td>"
        f"<td>{cell(fact.get('prior_value'))}</td>"
        "</tr>"
        for fact in facts
    )
    schedule_rows = "".join(
        "<tr>"
        f"<td>{cell(item.get('schedule_type'))}</td>"
        f"<td>{cell(item.get('schedule_id'))}</td>"
        f"<td>{cell(item.get('status'))}</td>"
        f"<td>{cell(len(item.get('issues', [])))}</td>"
        "</tr>"
        for item in schedules
    )
    question_rows = "".join(
        "<tr>"
        f"<td>{cell(item.get('question_id'))}</td>"
        f"<td>{cell(item.get('title'))}</td>"
        f"<td>{cell(item.get('state'))}</td>"
        f"<td>{cell(item.get('reason'))}</td>"
        "</tr>"
        for item in questions
        if item.get("state") != "NOT_TRIGGERED"
    )
    note_blocks = "".join(
        f'<article lang="{cell(block.get("language") or output_language)}">'
        f"<h3>{cell(block.get('section_id'))}</h3>"
        f"<p>{cell(block.get('text'))}</p>"
        f"<small>{cell(block.get('status'))}</small></article>"
        for block in narratives
    )
    taxonomy_rows = "".join(
        "<tr>"
        f"<td>{cell(item.get('xbrl_concept'))}</td>"
        f"<td>{cell(item.get('fact_type'))}</td>"
        f"<td>{cell(item.get('period'))}</td>"
        f"<td>{cell(item.get('value'))}</td>"
        f"<td>{cell(item.get('dimensions'))}</td>"
        "</tr>"
        for item in taxonomy_facts
    )
    micro_footer_rows = "".join(
        "<tr>"
        f"<td>{cell(item.get('key'))}</td>"
        f"<td>{cell(item.get('status'))}</td>"
        f"<td>{cell(item.get('value'))}</td>"
        f"<td>{cell(item.get('reason'))}</td>"
        "</tr>"
        for item in micro_reporting.get("footer_items", [])
    )
    issue_rows = "".join(
        "<tr>"
        f"<td>{cell(item.get('severity'))}</td>"
        f"<td>{cell(item.get('rule_id'))}</td>"
        f"<td>{cell(item.get('message'))}</td>"
        f"<td>{cell(item.get('review_status'))}</td>"
        "</tr>"
        for item in issues
    )
    document = f"""<!doctype html>
<html lang="it"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><meta name="color-scheme" content="light">
<title>Anteprima bilancio {cell(case.get('case_id'))}</title>
<style>body{{font-family:"Instrument Sans",Arial,sans-serif;max-width:1100px;margin:2rem auto;padding:0 1rem;color:#171816;background:#fff}}.skip-link{{position:absolute;left:-9999px}}.skip-link:focus{{left:1rem;top:1rem;background:#fff;color:#002060;padding:.75rem;z-index:2}}:focus-visible{{outline:3px solid #00b0f0;outline-offset:3px}}.table-scroll{{overflow-x:auto;margin-bottom:2rem}}table{{border-collapse:collapse;width:100%}}caption{{text-align:left;font-weight:600;padding:.4rem 0}}th,td{{border-bottom:1px solid #d9d9d9;padding:.55rem;text-align:left;vertical-align:top}}h1,h2{{color:#002060}}article{{border-left:3px solid #0070c0;padding:.25rem 1rem;margin:1rem 0}}</style>
</head><body><a class="skip-link" href="#main-content">Vai al contenuto principale</a><main id="main-content" tabindex="-1" data-output-language="{cell(output_language)}"><h1>Anteprima bilancio civilistico e XBRL</h1>
<p>Caso {cell(case.get('case_id'))} · revisione {cell(case.get('revision_id'))} · forma {cell(case.get('selected_form'))}</p>
<section aria-labelledby="presentation-heading"><h2 id="presentation-heading">Copertura dei prospetti civilistici</h2><p>Stato: <strong>{cell(statutory_presentation.get('status', 'NON_REVISIONATA'))}</strong> · voci richieste {cell(presentation_summary.get('required_leaf_concepts', 0))} · decisioni esplicite {cell(presentation_summary.get('explicit_decisions', 0))} · decisioni mancanti {cell(presentation_summary.get('missing_period_decisions', 0))} · problemi aritmetici {cell(presentation_summary.get('issues', 0))}.</p></section>
<section aria-labelledby="statements-heading"><h2 id="statements-heading">Prospetti</h2><div class="table-scroll" role="region" aria-labelledby="statements-heading" tabindex="0"><table><caption>Valori correnti e comparativi</caption><thead><tr><th scope="col">Sezione</th><th scope="col">Voce</th><th scope="col">Corrente</th><th scope="col">Comparativo</th></tr></thead><tbody>{statement_rows}</tbody></table></div></section>
<section aria-labelledby="schedules-heading"><h2 id="schedules-heading">Prospetti di dettaglio</h2><div class="table-scroll" role="region" aria-labelledby="schedules-heading" tabindex="0"><table><caption>Stato delle riconciliazioni di dettaglio</caption><thead><tr><th scope="col">Tipo</th><th scope="col">ID</th><th scope="col">Stato</th><th scope="col">Problemi</th></tr></thead><tbody>{schedule_rows}</tbody></table></div></section>
<section aria-labelledby="questions-heading"><h2 id="questions-heading">Questionario</h2><div class="table-scroll" role="region" aria-labelledby="questions-heading" tabindex="0"><table><caption>Domande contestuali e motivazioni</caption><thead><tr><th scope="col">ID</th><th scope="col">Domanda</th><th scope="col">Stato</th><th scope="col">Motivo</th></tr></thead><tbody>{question_rows}</tbody></table></div></section>
<section aria-labelledby="notes-heading"><h2 id="notes-heading">Nota integrativa</h2>{note_blocks}</section>
<section aria-labelledby="micro-heading"><h2 id="micro-heading">Informazioni in calce micro-imprese</h2><div class="table-scroll" role="region" aria-labelledby="micro-heading" tabindex="0"><table><caption>Informazioni statutarie in calce</caption><thead><tr><th scope="col">Voce</th><th scope="col">Stato</th><th scope="col">Contenuto</th><th scope="col">Motivo</th></tr></thead><tbody>{micro_footer_rows}</tbody></table></div></section>
<section aria-labelledby="taxonomy-heading"><h2 id="taxonomy-heading">Fatti tassonomici aggiuntivi</h2><div class="table-scroll" role="region" aria-labelledby="taxonomy-heading" tabindex="0"><table><caption>Fatti aggiuntivi sottoposti a revisione</caption><thead><tr><th scope="col">Concetto</th><th scope="col">Tipo</th><th scope="col">Periodo</th><th scope="col">Valore</th><th scope="col">Dimensioni</th></tr></thead><tbody>{taxonomy_rows}</tbody></table></div></section>
<section aria-labelledby="issues-heading"><h2 id="issues-heading">Validazioni</h2><div class="table-scroll" role="region" aria-labelledby="issues-heading" tabindex="0"><table><caption>Problemi, gravità e stato di revisione</caption><thead><tr><th scope="col">Gravità</th><th scope="col">Regola</th><th scope="col">Messaggio</th><th scope="col">Revisione</th></tr></thead><tbody>{issue_rows}</tbody></table></div></section>
<p><strong>Confine:</strong> Vera prepara una bozza rivedibile; non approva il bilancio, non firma e non deposita.</p></main>
</body></html>"""
    return document.encode("utf-8")


def create_preview(
    case: dict[str, Any],
    output_path: Path,
    actor: str,
    expected_revision: str,
) -> dict[str, Any]:
    """Write and hash the review preview before the final validation run."""

    _ensure_revision(case, expected_revision)
    if not case.get("statements"):
        raise ValueError("Statements are required before preview rendering")
    if output_path.suffix.lower() != ".html":
        raise ValueError("Preview output must use the .html extension")
    if output_path.is_symlink():
        raise ValueError("Preview must not be written through a symbolic link")
    if output_path.parent.is_symlink():
        raise ValueError("Preview directory must not be a symbolic link")
    _mutate(case, actor, "preview_rendered")
    preview = render_preview_html(case)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_bytes(preview)
    case["preview"] = {
        "file_name": output_path.name,
        "sha256": _sha256_bytes(preview),
        "size_bytes": len(preview),
        "content_base64": base64.b64encode(preview).decode("ascii"),
        "rendered_revision_id": case["revision_id"],
        "rendered_at": _now(),
        "review_content_hash": _review_content_hash(case),
        "computation_context": _computation_context(case, "bilancio-preview-v1"),
    }
    case["validation"] = None
    _record_event(
        case,
        "preview_rendered",
        actor,
        {
            key: value
            for key, value in case["preview"].items()
            if key != "content_base64"
        },
    )
    return case


def record_external_validation(
    case: dict[str, Any],
    report_path: Path,
    result: str,
    reported_issues: Sequence[Mapping[str, Any]],
    actor: str,
    expected_revision: str,
) -> dict[str, Any]:
    """Attach a user-supplied TEBENI report without automating transmission."""

    _ensure_revision(case, expected_revision)
    if case.get("state") not in {CaseState.APPROVED, CaseState.EXPORTED}:
        raise ValueError(
            "External validation results attach only to an approved or exported snapshot"
        )
    before_hash = _sha256_bytes(_canonical_json(_case_payload_for_hash(case)))
    record = record_external_validation_result(
        case, report_path, result, reported_issues, actor
    )
    document_id = _next_document_id(case)
    document = {
        "document_id": document_id,
        "purpose": "EXTERNAL_VALIDATION_REPORT",
        **record["report"],
        "parser_profile": "user_declared_tebeni_result_v1",
        "parsed_at": _now(),
    }
    record["document_id"] = document_id
    case["source_documents"] = [
        item
        for item in case.get("source_documents", [])
        if item.get("purpose") != "EXTERNAL_VALIDATION_REPORT"
    ] + [document]
    case["external_validation"] = record
    case["updated_at"] = _now()
    case["_pending_before_hash"] = before_hash
    _record_event(case, "document_uploaded", actor, {"document_id": document_id})
    _record_evidence_attached(case, actor, document)
    _record_event(
        case,
        "external_validation_recorded",
        actor,
        {"document_id": document_id, "result": record["result"]},
    )
    return case


def record_file_security_scan(
    case: dict[str, Any], receipt: Mapping[str, Any], actor: str
) -> dict[str, Any]:
    """Bind one host scanner clean verdict to an imported document checksum."""

    if str(receipt.get("status")) != "CLEAN":
        raise ValueError("Only a clean malware-scan verdict may be recorded")
    required = {
        "sha256",
        "size_bytes",
        "engine",
        "signature_version",
        "scanned_at",
    }
    missing = sorted(key for key in required if receipt.get(key) in {None, ""})
    if missing:
        raise ValueError(
            f"Malware-scan receipt is missing fields: {', '.join(missing)}"
        )
    source_sha256 = str(receipt["sha256"])
    documents = [
        document
        for document in case.get("source_documents", [])
        if document.get("sha256") == source_sha256
        and int(document.get("size_bytes", -1)) == int(receipt["size_bytes"])
    ]
    if not documents:
        raise ValueError("Malware-scan receipt does not match an imported document")
    document = documents[-1]
    scan_id = f"scan_{len(case.get('file_security_scans', [])) + 1:06d}"
    normalized = {
        "scan_id": scan_id,
        "document_id": document["document_id"],
        "status": "CLEAN",
        "sha256": source_sha256,
        "size_bytes": int(receipt["size_bytes"]),
        "engine": str(receipt["engine"]),
        "signature_version": str(receipt["signature_version"]),
        "scanned_at": str(receipt["scanned_at"]),
    }
    document["security_scan_id"] = scan_id
    case.setdefault("file_security_scans", []).append(normalized)
    _record_event(
        case,
        "document_malware_scanned",
        actor,
        {
            "scan_id": scan_id,
            "document_id": document["document_id"],
            "engine": normalized["engine"],
            "signature_version": normalized["signature_version"],
        },
    )
    return case


def record_artifact_access(
    case: dict[str, Any], action: str, actor: str, details: Mapping[str, Any]
) -> dict[str, Any]:
    """Append an artifact grant or redemption event without changing revision."""

    if action not in {"artifact_download_grant_issued", "artifact_downloaded"}:
        raise ValueError("Unsupported artifact access event")
    required = {"grant_id", "artifact_id", "sha256"}
    missing = sorted(key for key in required if not str(details.get(key, "")))
    if missing:
        raise ValueError(f"Artifact access event is missing: {', '.join(missing)}")
    _record_event(
        case,
        action,
        actor,
        {key: str(details[key]) for key in sorted(required)},
    )
    return case


def record_taxonomy_catalogue_build(
    case: dict[str, Any],
    receipt: Mapping[str, Any],
    actor: str,
    expected_revision: str,
) -> dict[str, Any]:
    """Attach a deployment-controlled catalogue build receipt to one case."""

    _ensure_revision(case, expected_revision)
    required = {
        "file_name",
        "sha256",
        "size_bytes",
        "taxonomy_id",
        "taxonomy_package_sha256",
        "registry_sha256",
        "concept_count",
        "built_at",
    }
    missing = sorted(key for key in required if receipt.get(key) in {None, ""})
    if missing:
        raise ValueError(
            f"Taxonomy catalogue build receipt is missing: {', '.join(missing)}"
        )
    if str(receipt["taxonomy_id"]) != str(case["rule_pack_versions"]["taxonomy_id"]):
        raise ValueError("Built taxonomy catalogue identifier differs from the case")
    if str(receipt["taxonomy_package_sha256"]) != str(case.get("taxonomy_checksum")):
        raise ValueError("Built taxonomy package checksum differs from the case")
    _mutate(case, actor, "taxonomy_catalogue_built")
    case["taxonomy_catalogue_build"] = dict(receipt)
    case["statutory_presentation"] = None
    case["validation"] = None
    _record_event(
        case,
        "taxonomy_catalogue_built",
        actor,
        {
            "sha256": str(receipt["sha256"]),
            "concept_count": int(receipt["concept_count"]),
            "registry_sha256": str(receipt["registry_sha256"]),
        },
    )
    return case


def archive_case(
    case: dict[str, Any],
    actor: str,
    expected_revision: str,
    *,
    retention_days: int,
    reason: str,
) -> dict[str, Any]:
    """Archive one case under an explicit host retention policy."""

    _ensure_revision(case, expected_revision)
    if case.get("state") == CaseState.ARCHIVED:
        raise ValueError("Case is already archived")
    if isinstance(retention_days, bool) or not 1 <= retention_days <= 3650:
        raise ValueError("Retention days must be from 1 to 3650")
    normalized_reason = reason.strip()
    if not normalized_reason:
        raise ValueError("Case archive reason is required")
    archived_at = datetime.now(tz=UTC)
    before_hash = _sha256_bytes(_canonical_json(_case_payload_for_hash(case)))
    case["state"] = CaseState.ARCHIVED
    case["archive"] = {
        "archived_at": archived_at.isoformat(timespec="seconds"),
        "archived_by": actor,
        "reason": normalized_reason,
        "retention_days": retention_days,
        "retain_until": (archived_at + timedelta(days=retention_days)).isoformat(
            timespec="seconds"
        ),
    }
    case["updated_at"] = _now()
    case["_pending_before_hash"] = before_hash
    _record_event(
        case,
        "case_archived",
        actor,
        {
            "reason": normalized_reason,
            "retain_until": case["archive"]["retain_until"],
        },
    )
    return case


def remember_mappings(
    case: dict[str, Any],
    memory_path: Path,
    source_system_template: str,
    actor: str,
    expected_revision: str,
) -> dict[str, Any]:
    """Write approved classifications into an explicitly tenant-scoped store."""

    _ensure_revision(case, expected_revision)
    result = remember_approved_mappings(
        case, memory_path, source_system_template, actor
    )
    _record_event(case, "approved_mappings_remembered", actor, result)
    return case


def remember_client_history(
    case: dict[str, Any],
    history_path: Path,
    actor: str,
    expected_revision: str,
) -> dict[str, Any]:
    """Persist approved client knowledge for unconfirmed future suggestions."""

    _ensure_revision(case, expected_revision)
    result = remember_approved_client_history(case, history_path, actor)
    _record_event(case, "approved_client_history_remembered", actor, result)
    return case


def _issue(rule_id: str, severity: str, message: str, index: int) -> ValidationIssue:
    return ValidationIssue(
        issue_id=f"iss_{index:04d}",
        severity=severity,
        rule_id=rule_id,
        message=message,
        override_allowed=severity not in {"BLOCKER"},
    )


def _current_review_decision(
    case: Mapping[str, Any], issue: Mapping[str, Any]
) -> Mapping[str, Any] | None:
    fingerprint = str(issue.get("fingerprint") or _issue_fingerprint(issue))
    for decision in reversed(case.get("review_decisions", [])):
        if decision.get("issue_fingerprint") == fingerprint:
            return decision
    return None


def validate_case(case: dict[str, Any]) -> dict[str, Any]:
    """Run deterministic input, accounting, disclosure, XBRL, and review gates."""

    issues: list[ValidationIssue] = []

    def add(rule_id: str, severity: str, message: str) -> None:
        issues.append(_issue(rule_id, severity, message, len(issues) + 1))

    if case["state"] == CaseState.UNSUPPORTED or case.get("unsupported_reasons"):
        add("INPUT.UNSUPPORTED_CASE", "BLOCKER", "The entity is outside the MVP scope")
    trial_balance = case.get("trial_balance")
    if not trial_balance:
        add("INPUT.TRIAL_BALANCE_REQUIRED", "BLOCKER", "A trial balance is required")
    elif not trial_balance.get("confirmed_convention"):
        add(
            "INPUT.PARSER_CONFIRMATION_REQUIRED",
            "BLOCKER",
            "The debit/credit convention is not confirmed",
        )
    if case["entity"].get("first_financial_year") is not True and not case.get(
        "prior_xbrl"
    ):
        add(
            "INPUT.PRIOR_XBRL_RECOMMENDED",
            "MEDIUM",
            "The prior filed XBRL is not attached for comparative and opening checks",
        )
    prior_reconciliation = _prior_xbrl_reconciliation(case)
    for item in prior_reconciliation["issues"]:
        add(
            f"INPUT.{item['code']}",
            "BLOCKER",
            "Current comparative facts do not reconcile to the attached prior XBRL: "
            + str(item.get("xbrl_concept", "unknown concept")),
        )
    if any(
        item.get("status") == "RESTATEMENT_CONFIRMED"
        for item in prior_reconciliation["checks"]
    ):
        add(
            "INPUT.PRIOR_XBRL_RESTATEMENT_REVIEWED",
            "MEDIUM",
            "Comparative amounts differ from the prior filing under an evidenced restatement decision",
        )
    if not case.get("selected_form"):
        add("FORM.SELECTION_REQUIRED", "BLOCKER", "The statutory form is not confirmed")
    if case.get("selected_form") == "MICRO":
        micro_reporting = case.get("micro_reporting") or {}
        if micro_reporting.get("status") != "CONFIRMED":
            add(
                "MICRO.REPORTING_CHOICE_REQUIRED",
                "BLOCKER",
                "The micro footer-versus-notes treatment is not confirmed",
            )
        elif micro_reporting.get("mode") == "FOOTER_ONLY":
            if (
                not SAFE_QNAME.fullmatch(str(micro_reporting.get("xbrl_concept", "")))
                or not str(micro_reporting.get("rendered_text", "")).strip()
                or not str(micro_reporting.get("reviewed_by", "")).strip()
            ):
                add(
                    "MICRO.FOOTER_XBRL_FACT_REQUIRED",
                    "BLOCKER",
                    "Footer-only micro accounts require a reviewed renderable XBRL fact",
                )
            if any(
                block.get("status") == "ACCEPTED"
                for block in case.get("narrative_blocks", [])
            ):
                add(
                    "MICRO.FOOTER_ONLY_HAS_NOTES",
                    "BLOCKER",
                    "Footer-only micro accounts cannot also export a note draft",
                )
            positive_answers = []
            for answer in case.get("disclosure_answers", []):
                value = answer.get("value")
                if value is True or (
                    isinstance(value, Mapping)
                    and any(item is True for item in value.values())
                ):
                    positive_answers.append(str(answer["key"]))
            prohibited = sorted(
                set(positive_answers)
                & {
                    "accounting_policy_changes",
                    "contingent_liabilities",
                    "derivatives",
                    "double_format_events",
                    "going_concern_uncertainties",
                    "guarantees_and_commitments",
                    "non_market_transactions",
                    "off_balance_sheet_arrangements",
                    "post_closing_events",
                    "prior_period_errors",
                    "related_party_transactions",
                }
            )
            if prohibited:
                add(
                    "MICRO.FOOTER_ONLY_DISCLOSURE_CONFLICT",
                    "BLOCKER",
                    "Footer-only treatment is incompatible with positive disclosures: "
                    + ", ".join(prohibited),
                )
    entries = (trial_balance or {}).get("entries", [])
    mapping_ids = {item["account_id"] for item in case.get("mappings", [])}
    missing_mappings = [
        row["account_id"] for row in entries if row["account_id"] not in mapping_ids
    ]
    if missing_mappings:
        add(
            "MAPPING.COVERAGE",
            "BLOCKER",
            f"{len(missing_mappings)} accounts are not mapped or excluded",
        )
    for adjustment in case.get("adjustments", []):
        for amount_key in ("current_amount", "prior_amount"):
            total = sum(
                Decimal(line[amount_key]) for line in adjustment.get("lines", [])
            )
            if total:
                add(
                    "ADJUSTMENT.UNBALANCED",
                    "BLOCKER",
                    f"Adjustment {adjustment['adjustment_id']} is not presentation-neutral",
                )
    statements = case.get("statements")
    if not statements:
        add("STATEMENT.NOT_COMPUTED", "BLOCKER", "Statements have not been computed")
    else:
        totals = statements["section_totals"]
        if "ASSETS" in totals and "LIABILITIES_EQUITY" in totals:
            for period, rule_id, period_label in (
                ("current", "STATEMENT.BALANCE_SHEET", "Current"),
                (
                    "prior",
                    "STATEMENT.COMPARATIVE_BALANCE_SHEET",
                    "Comparative",
                ),
            ):
                asset_value = totals["ASSETS"].get(period)
                liability_value = totals["LIABILITIES_EQUITY"].get(period)
                if asset_value is None or liability_value is None:
                    add(
                        rule_id,
                        "BLOCKER",
                        f"{period_label} balance-sheet totals are missing",
                    )
                elif Decimal(asset_value) + Decimal(liability_value) != 0:
                    add(
                        rule_id,
                        "BLOCKER",
                        f"{period_label} assets do not equal liabilities and equity",
                    )
        result_facts = [
            fact
            for fact in statements["facts"]
            if fact["statement_section"] in {"INCOME_RESULT", "EQUITY_RESULT"}
        ]
        for value_field, rule_id, period_label in (
            ("current_value", "STATEMENT.RESULT_TIE_OUT", "Current"),
            (
                "prior_value",
                "STATEMENT.COMPARATIVE_RESULT_TIE_OUT",
                "Comparative",
            ),
        ):
            if any(fact.get(value_field) is None for fact in result_facts):
                add(
                    rule_id,
                    "BLOCKER",
                    f"{period_label} result reconciliation values are missing",
                )
                continue
            result_income = sum(
                Decimal(fact[value_field])
                for fact in result_facts
                if fact["statement_section"] == "INCOME_RESULT"
            )
            result_equity = sum(
                Decimal(fact[value_field])
                for fact in result_facts
                if fact["statement_section"] == "EQUITY_RESULT"
            )
            if (result_income or result_equity) and result_income + result_equity != 0:
                add(
                    rule_id,
                    "BLOCKER",
                    f"{period_label} income-statement result does not reconcile to equity",
                )
    if case.get("statutory_presentation_required", True):
        mapping_index = case.get("taxonomy_mapping_index")
        if not mapping_index:
            add(
                "MAPPING.TAXONOMY_INDEX_REQUIRED",
                "BLOCKER",
                "The selected-form official taxonomy mapping index is missing",
            )
        elif mapping_index.get("selected_form") != case.get(
            "selected_form"
        ) or mapping_index.get("taxonomy_package_sha256") != case.get(
            "taxonomy_checksum"
        ):
            add(
                "MAPPING.TAXONOMY_INDEX_STALE",
                "BLOCKER",
                "The mapping index does not match the selected form and taxonomy package",
            )
        presentation = case.get("statutory_presentation")
        if not presentation:
            add(
                "STATEMENT.STATUTORY_PRESENTATION_REQUIRED",
                "BLOCKER",
                "Primary statutory taxonomy presentation coverage has not been reviewed",
            )
        elif presentation.get("status") != "COMPLETE":
            summary = presentation.get("summary") or {}
            add(
                "STATEMENT.STATUTORY_PRESENTATION_INCOMPLETE",
                "BLOCKER",
                "Primary statutory presentation has "
                f"{summary.get('missing_period_decisions', 0)} missing period decisions "
                f"and {summary.get('issues', 0)} arithmetic issues",
            )
        elif (presentation.get("semantic_reconciliation") or {}).get("status") != (
            "PASS"
        ):
            add(
                "STATEMENT.XBRL_RECONCILIATION_REQUIRED",
                "BLOCKER",
                "Canonical statement totals do not reconcile to statutory XBRL roots",
            )
        elif str(
            (presentation.get("inventory") or {}).get("taxonomy_package_sha256", "")
        ) != str(case.get("taxonomy_checksum")):
            add(
                "STATEMENT.STATUTORY_PRESENTATION_TAXONOMY_MISMATCH",
                "BLOCKER",
                "Primary statutory presentation uses a different taxonomy package",
            )
        elif mapping_index and (presentation.get("inventory") or {}).get(
            "inventory_sha256"
        ) != mapping_index.get("inventory_sha256"):
            add(
                "STATEMENT.STATUTORY_PRESENTATION_MAPPING_INDEX_MISMATCH",
                "BLOCKER",
                "Statutory presentation differs from the inventory used for account mapping",
            )
        else:
            for fact in presentation.get("output_facts", []):
                if fact.get("status") not in EXPORTABLE_STATUSES:
                    add(
                        "STATEMENT.STATUTORY_PRESENTATION_NON_EXPORTABLE",
                        "BLOCKER",
                        f"Presentation fact {fact.get('fact_id')} is not exportable",
                    )
                if not fact.get("derivation") and not fact.get("confirmed_by"):
                    add(
                        "STATEMENT.STATUTORY_PRESENTATION_PROVENANCE",
                        "BLOCKER",
                        f"Presentation fact {fact.get('fact_id')} has no provenance",
                    )
    required_schedules = required_schedule_types(case)
    schedules = case.get("schedules", [])
    present_schedule_types = {item["schedule_type"] for item in schedules}
    duplicate_schedule_types = sorted(
        schedule_type
        for schedule_type in present_schedule_types
        if sum(item["schedule_type"] == schedule_type for item in schedules) > 1
    )
    if duplicate_schedule_types:
        add(
            "SCHEDULE.DUPLICATE_TYPE",
            "BLOCKER",
            "Only one reviewed schedule is allowed per type: "
            + ", ".join(duplicate_schedule_types),
        )
    missing_schedules = sorted(required_schedules - present_schedule_types)
    if missing_schedules:
        add(
            "SCHEDULE.REQUIRED",
            "BLOCKER",
            f"Required schedules are missing: {', '.join(missing_schedules)}",
        )
    incomplete_schedules = [
        item["schedule_id"] for item in schedules if item.get("status") != "COMPLETE"
    ]
    if incomplete_schedules:
        add(
            "SCHEDULE.RECONCILIATION",
            "BLOCKER",
            f"Schedules have unresolved arithmetic issues: {', '.join(incomplete_schedules)}",
        )
    non_cash_schedules = [
        item for item in schedules if item.get("schedule_type") != "CASH_FLOW"
    ]
    if non_cash_schedules and case.get("taxonomy_mapping_index") is not None:
        adapter = case.get("schedule_taxonomy_adapter")
        expected_types = sorted(
            str(item["schedule_type"]) for item in non_cash_schedules
        )
        covered_types = sorted(
            str(item.get("schedule_type"))
            for item in (adapter or {}).get("coverage", [])
            if item.get("status") == "COMPLETE"
        )
        if not adapter or adapter.get("status") != "COMPLETE":
            add(
                "SCHEDULE.TAXONOMY_ADAPTER_REQUIRED",
                "BLOCKER",
                "Every reviewed non-cash schedule requires official taxonomy-table coverage",
            )
        elif covered_types != expected_types:
            add(
                "SCHEDULE.TAXONOMY_ADAPTER_INCOMPLETE",
                "BLOCKER",
                "Schedule taxonomy coverage differs from the current schedules",
            )
        elif str(
            (adapter.get("inventory") or {}).get("taxonomy_package_sha256")
        ) != str(case.get("taxonomy_checksum")):
            add(
                "SCHEDULE.TAXONOMY_ADAPTER_TAXONOMY_MISMATCH",
                "BLOCKER",
                "Schedule taxonomy coverage uses a different taxonomy package",
            )
        elif case.get("schedule_taxonomy_facts") != adapter.get("generated_facts"):
            add(
                "SCHEDULE.TAXONOMY_ADAPTER_FACT_MISMATCH",
                "BLOCKER",
                "Schedule taxonomy facts differ from their reviewed adapter result",
            )
    if case.get("selected_form") == "ORDINARY" and case.get(
        "statutory_presentation_required", True
    ):
        cash_schedules = [
            item for item in schedules if item.get("schedule_type") == "CASH_FLOW"
        ]
        cash_flow_values = (case.get("statutory_presentation") or {}).get(
            "cash_flow_values"
        )
        if len(cash_schedules) == 1 and cash_schedules[0].get("status") == "COMPLETE":
            schedule = cash_schedules[0]
            schedule_change = Decimal(str(schedule["closing_cash"])) - Decimal(
                str(schedule["opening_cash"])
            )
            xbrl_change = (
                None
                if not isinstance(cash_flow_values, Mapping)
                or cash_flow_values.get("current_value") is None
                else Decimal(str(cash_flow_values["current_value"]))
            )
            if xbrl_change is None or xbrl_change != schedule_change:
                add(
                    "CASH_FLOW.XBRL_NET_CHANGE_RECONCILIATION",
                    "BLOCKER",
                    "The reviewed cash-flow net change does not reconcile to the statutory XBRL cash-flow root",
                )
    for fact in case.get("canonical_facts", []):
        if fact["status"] not in EXPORTABLE_STATUSES:
            add(
                "FACT.NON_EXPORTABLE_STATUS",
                "BLOCKER",
                f"Fact {fact['fact_id']} has non-exportable status {fact['status']}",
            )
        if not fact.get("source_refs") and not fact.get("derivation"):
            add(
                "FACT.PROVENANCE_REQUIRED",
                "BLOCKER",
                f"Fact {fact['fact_id']} has no provenance",
            )
        if fact.get("xbrl_concept") and str(
            fact.get("xbrl_sign_multiplier", "")
        ) not in {"1", "-1"}:
            add(
                "FACT.XBRL_SIGN_REQUIRED",
                "BLOCKER",
                f"Fact {fact['fact_id']} has no reviewed XBRL sign convention",
            )
    for fact in [
        *case.get("taxonomy_facts", []),
        *case.get("schedule_taxonomy_facts", []),
    ]:
        if fact.get("status") not in EXPORTABLE_STATUSES:
            add(
                "FACT.TAXONOMY_NON_EXPORTABLE_STATUS",
                "BLOCKER",
                f"Taxonomy fact {fact['fact_id']} is not reviewed",
            )
        if (
            fact.get("fact_type") != "NIL"
            and not fact.get("source_refs")
            and not fact.get("derivation")
        ):
            add(
                "FACT.TAXONOMY_PROVENANCE_REQUIRED",
                "BLOCKER",
                f"Taxonomy fact {fact['fact_id']} has no provenance",
            )
        if fact.get("fact_type") == "NIL" and not fact.get("nil_reason"):
            add(
                "FACT.NIL_REASON_REQUIRED",
                "BLOCKER",
                f"Nil taxonomy fact {fact['fact_id']} has no reviewed reason",
            )
    narrative_qnames: set[str] = set()
    for block in case.get("narrative_blocks", []):
        if block.get("status") != "ACCEPTED":
            continue
        qname = str(block.get("xbrl_concept") or "")
        if not qname:
            add(
                "NARRATIVE.XBRL_CONCEPT_REQUIRED",
                "BLOCKER",
                f"Accepted narrative block {block.get('block_id')} has no XBRL concept",
            )
        elif qname in narrative_qnames:
            add(
                "NARRATIVE.XBRL_CONCEPT_DUPLICATE",
                "BLOCKER",
                f"Accepted narrative concept is duplicated: {qname}",
            )
        else:
            narrative_qnames.add(qname)
    accepted_answers = {
        answer["key"]
        for answer in case.get("disclosure_answers", [])
        if disclosure_answer_complete(answer)
    }
    missing_confirmations = sorted(ANNUAL_NEGATIVE_CONFIRMATION_KEYS - accepted_answers)
    if missing_confirmations:
        add(
            "DISCLOSURE.NEGATIVE_CONFIRMATIONS",
            "BLOCKER",
            f"{len(missing_confirmations)} annual negative confirmations are missing",
        )
    double_format_answer = next(
        (
            answer
            for answer in case.get("disclosure_answers", [])
            if answer.get("key") == "double_format_events"
            and disclosure_answer_complete(answer)
        ),
        None,
    )
    double_value = (double_format_answer or {}).get("value")
    mismatch_declared = double_value is True or (
        isinstance(double_value, Mapping)
        and double_value.get("differences_present") is True
    )
    representation = case.get("taxonomy_representation")
    if mismatch_declared and not representation:
        add(
            "XBRL.SUBSTANTIVE_TAXONOMY_MISMATCH",
            "BLOCKER",
            "Substantive taxonomy differences require a reviewer-owned differences report and treatment",
        )
    elif mismatch_declared and representation.get("mismatch_present") is not True:
        add(
            "XBRL.TAXONOMY_MISMATCH_CONTRADICTION",
            "BLOCKER",
            "The disclosure answer and taxonomy representation review disagree",
        )
    elif representation and representation.get("mismatch_present") is True:
        if not mismatch_declared:
            add(
                "XBRL.TAXONOMY_MISMATCH_CONTRADICTION",
                "BLOCKER",
                "The taxonomy differences report conflicts with the annual confirmation",
            )
        else:
            add(
                "XBRL.SUBSTANTIVE_TAXONOMY_MISMATCH_REVIEWED",
                "MEDIUM",
                "Substantive taxonomy differences remain and the reviewer-selected treatment must be understood before approval",
            )
    disclosure_coverage = case.get("disclosure_coverage")
    if not case.get("disclosure_rule_pack") or not disclosure_coverage:
        add(
            "DISCLOSURE.RULE_PACK_REQUIRED",
            "BLOCKER",
            "An effective-dated disclosure rule pack has not been activated",
        )
    else:
        reviewed_manual_flags = {
            str(item["flag"]) for item in case.get("disclosure_trigger_decisions", [])
        }
        missing_manual_flags = sorted(
            manual_disclosure_flags(case["disclosure_rule_pack"])
            - reviewed_manual_flags
        )
        if missing_manual_flags:
            add(
                "DISCLOSURE.MANUAL_TRIGGER_REVIEW_REQUIRED",
                "BLOCKER",
                "Disclosure applicability requires professional decisions for: "
                + ", ".join(missing_manual_flags),
            )
        for item in disclosure_coverage["coverage"]:
            if item["triggered"] and not item["complete"]:
                missing = [
                    f"{requirement['kind']}:{requirement['key']}"
                    for requirement in item["requirements"]
                    if not requirement["complete"]
                ]
                add(
                    f"DISCLOSURE.{item['rule_id']}",
                    item["severity_if_missing"],
                    f"Triggered disclosure is incomplete: {', '.join(missing)}",
                )
    preview = case.get("preview")
    if not preview:
        add("REVIEW.PREVIEW_REQUIRED", "BLOCKER", "A review preview is required")
    elif preview.get("review_content_hash") != _review_content_hash(case):
        add(
            "REVIEW.PREVIEW_STALE",
            "BLOCKER",
            "The review preview does not match the current substantive case content",
        )
    else:
        try:
            _reviewed_preview_bytes(preview)
        except ValueError as exc:
            add(
                "REVIEW.PREVIEW_INTEGRITY",
                "BLOCKER",
                str(exc),
            )
    xbrl_review = case.get("xbrl_review")
    if xbrl_review:
        if xbrl_review.get("review_content_hash") != _review_content_hash(case):
            add(
                "XBRL.LOCAL_REVIEW_STALE",
                "BLOCKER",
                "The locally validated XBRL does not match the current case content",
            )
        elif xbrl_review.get("status") != "PASS":
            add(
                "XBRL.LOCAL_VALIDATION_FAILED",
                "BLOCKER",
                "The pre-approval XBRL instance failed local processor validation",
            )
    issue_payloads = [issue.as_dict() for issue in issues]
    unresolved_high = 0
    review_required = 0
    overrides = 0
    for issue in issue_payloads:
        decision = _current_review_decision(case, issue)
        action = str((decision or {}).get("action", ""))
        if decision:
            issue["review_decision_id"] = decision["decision_id"]
            issue["review_status"] = action
            issue["reviewed_by"] = decision["reviewed_by"]
        else:
            issue["review_status"] = "UNREVIEWED"
        if issue["severity"] == "HIGH":
            if action == "OVERRIDDEN":
                overrides += 1
            else:
                unresolved_high += 1
        elif issue["severity"] in {"MEDIUM", "LOW", "INFO"}:
            if action != "ACKNOWLEDGED":
                review_required += 1
    blockers = sum(issue.severity == "BLOCKER" for issue in issues)
    highs = sum(issue.severity == "HIGH" for issue in issues)
    status = "PASS" if blockers == 0 and unresolved_high == 0 else "FAIL"
    return {
        "status": status,
        "layers": ["INPUT", "ACCOUNTING", "DISCLOSURE", "XBRL", "REVIEW_POLICY"],
        "blockers": blockers,
        "high": highs,
        "unresolved_high": unresolved_high,
        "review_required": review_required,
        "approved_overrides": overrides,
        "issues": issue_payloads,
        "prior_xbrl_reconciliation": prior_reconciliation,
        "validated_revision_id": case["revision_id"],
        "input_manifest_hash": _manifest_hash(case),
        "validated_at": _now(),
        "computation_context": _computation_context(case, "bilancio-validation-v1"),
    }


def record_issue_reviews(
    case: dict[str, Any],
    decisions: Sequence[Mapping[str, Any]],
    reviewer: str,
    expected_revision: str,
) -> dict[str, Any]:
    """Record reviewer-owned acknowledgements and permitted HIGH overrides."""

    _ensure_revision(case, expected_revision)
    validation = case.get("validation") or {}
    if validation.get("validated_revision_id") != case["revision_id"]:
        raise ValueError("Issue reviews require validation of the current revision")
    current_issues = {
        str(issue["issue_id"]): issue for issue in validation.get("issues", [])
    }
    if not decisions:
        raise ValueError("At least one issue review decision is required")
    normalized: list[dict[str, Any]] = []
    seen: set[str] = set()
    for raw in decisions:
        issue_id = str(raw["issue_id"])
        if issue_id in seen:
            raise ValueError("Each issue may be reviewed only once per request")
        seen.add(issue_id)
        issue = current_issues.get(issue_id)
        if issue is None:
            raise ValueError(f"Unknown current validation issue: {issue_id}")
        action = str(raw["action"]).upper()
        severity = str(issue["severity"])
        reason = str(raw.get("reason", "")).strip()
        if severity == "BLOCKER" or not issue.get("override_allowed"):
            raise ValueError("Structural and blocking issues cannot be overridden")
        if action == "OVERRIDDEN":
            if severity != "HIGH":
                raise ValueError("Only HIGH issues may be professionally overridden")
            if len(reason) < 10:
                raise ValueError("A professional override requires a specific reason")
        elif action == "ACKNOWLEDGED":
            if severity not in {"MEDIUM", "LOW", "INFO"}:
                raise ValueError("Only non-blocking warnings may be acknowledged")
            if not reason:
                raise ValueError("A warning acknowledgement requires a review note")
        else:
            raise ValueError(f"Unsupported issue review action: {action}")
        normalized.append(
            {
                "issue_id": issue_id,
                "issue_fingerprint": str(
                    issue.get("fingerprint") or _issue_fingerprint(issue)
                ),
                "rule_id": issue["rule_id"],
                "severity": severity,
                "action": action,
                "reason": reason,
                "reviewed_by": reviewer,
                "reviewed_at": _now(),
            }
        )
    _mutate(case, reviewer, "issue_reviews_recorded")
    for decision in normalized:
        decision["decision_id"] = (
            f"decision_{len(case.get('review_decisions', [])) + 1:06d}"
        )
        case.setdefault("review_decisions", []).append(decision)
        current_issues[decision["issue_id"]]["review_decision_id"] = decision[
            "decision_id"
        ]
        current_issues[decision["issue_id"]]["review_status"] = decision["action"]
        current_issues[decision["issue_id"]]["reviewed_by"] = reviewer
    for decision in normalized:
        if decision["action"] == "OVERRIDDEN":
            _record_event(case, "override_created", reviewer, decision)
            _record_event(case, "override_approved", reviewer, decision)
        else:
            _record_event(case, "issue_acknowledged", reviewer, decision)
        _record_event(
            case,
            "issue_resolved",
            reviewer,
            {
                "issue_id": decision["issue_id"],
                "decision_id": decision["decision_id"],
                "resolution": decision["action"],
            },
        )
    _record_event(
        case,
        "issue_reviews_recorded",
        reviewer,
        {"decision_ids": [item["decision_id"] for item in normalized]},
    )
    return case


def record_disclosure_answers(
    case: dict[str, Any],
    answers: Sequence[Mapping[str, Any]],
    actor: str,
    expected_revision: str,
) -> dict[str, Any]:
    """Record structured, evidenced questionnaire answers without inventing zeros."""

    _ensure_revision(case, expected_revision)
    normalized: list[dict[str, Any]] = []
    seen: set[str] = set()
    allowed_statuses = {
        "OPEN",
        "ASSIGNED",
        "ANSWERED_UNREVIEWED",
        "ACCEPTED",
        "REJECTED",
        "NOT_APPLICABLE_CONFIRMED",
    }
    allowed_keys = {
        str(question["answer_key"]) for question in case.get("questionnaire", [])
    } | ANNUAL_NEGATIVE_CONFIRMATION_KEYS
    for answer in answers:
        key = str(answer["key"])
        if not key or key in seen:
            raise ValueError("Disclosure answer keys must be present and unique")
        if key not in allowed_keys:
            raise ValueError(f"Disclosure answer key is not active or annual: {key}")
        seen.add(key)
        status = str(answer["status"]).upper()
        if status not in allowed_statuses:
            raise ValueError(f"Unsupported disclosure answer status: {status}")
        if status in {"ACCEPTED", "NOT_APPLICABLE_CONFIRMED"}:
            if not actor.strip():
                raise ValueError(
                    "Accepted disclosure answers require an authenticated actor"
                )
            claimed_actor = answer.get("confirmed_by")
            if claimed_actor is not None and str(claimed_actor) != actor:
                raise ValueError(
                    "Disclosure confirmation identity must match the authenticated actor"
                )
        terminal_answer = {
            **dict(answer),
            "status": status,
            "confirmed_by": actor,
        }
        if status == "ACCEPTED" and not disclosure_answer_complete(terminal_answer):
            raise ValueError(
                "Accepted disclosure answers require a reviewed structured value"
            )
        if status == "NOT_APPLICABLE_CONFIRMED" and not disclosure_answer_complete(
            terminal_answer
        ):
            raise ValueError(
                "Not-applicable disclosure confirmations require a specific reason"
            )
        if status == "ASSIGNED" and not str(answer.get("owner", "")).strip():
            raise ValueError("Assigned disclosure answers require an owner")
        source_refs = sorted(
            {
                str(ref).strip()
                for ref in answer.get("source_refs", [])
                if str(ref).strip()
            }
        )
        if source_refs and not set(source_refs) <= _available_evidence_refs(case):
            raise ValueError("Disclosure answer references evidence outside the case")
        normalized.append(
            {
                "key": key,
                "status": status,
                "value": answer.get("value"),
                "source_refs": source_refs,
                "reason": str(answer.get("reason", "")),
                "owner": str(answer.get("owner", "")) or None,
                "confirmed_by": (
                    actor
                    if status in {"ACCEPTED", "NOT_APPLICABLE_CONFIRMED"}
                    else None
                ),
                "confirmed_at": (
                    _now()
                    if status in {"ACCEPTED", "NOT_APPLICABLE_CONFIRMED"}
                    else None
                ),
            }
        )
    existing = {
        str(answer["key"]): answer for answer in case.get("disclosure_answers", [])
    }
    for answer in normalized:
        existing[answer["key"]] = answer
    _mutate(case, actor, "questionnaire_answers_recorded")
    _clear_narrative_reviews(case)
    case["disclosure_answers"] = list(existing.values())
    case["validation"] = None
    _refresh_disclosures(case)
    _record_event(case, "question_answered", actor, {"answer_count": len(normalized)})
    return case


def run_validation(
    case: dict[str, Any], actor: str, expected_revision: str
) -> dict[str, Any]:
    """Persist a validation run and move the case to the corresponding gate."""

    _ensure_revision(case, expected_revision)
    _mutate(case, actor, "validation_run")
    result = validate_case(case)
    case["validation"] = result
    case["state"] = (
        CaseState.READY_FOR_REVIEW
        if result["status"] == "PASS"
        else CaseState.VALIDATION_FAILED
    )
    migrations = case.get("regulatory_migrations", [])
    if migrations:
        latest_migration = migrations[-1]
        latest_migration.setdefault("revalidation_runs", []).append(
            {
                "revision_id": case["revision_id"],
                "result": result["status"],
                "completed_at": _now(),
            }
        )
        latest_migration["revalidation_status"] = (
            "PASSED" if result["status"] == "PASS" else "FAILED"
        )
    _record_event(case, "validation_run", actor, {"status": result["status"]})
    return case


def prepare_xbrl_review(
    case: dict[str, Any],
    catalogue_path: Path,
    taxonomy_package: Path,
    output_dir: Path,
    actor: str,
    expected_revision: str,
    validator: (
        Callable[[Path, Path, Path | None, str | None], dict[str, object]] | None
    ) = None,
) -> dict[str, Any]:
    """Render and locally validate the current case before reviewer approval."""

    _ensure_revision(case, expected_revision)
    validation = case.get("validation") or {}
    if (
        case.get("state") != CaseState.READY_FOR_REVIEW
        or validation.get("status") != "PASS"
        or validation.get("validated_revision_id") != case["revision_id"]
    ):
        raise ValueError(
            "Pre-approval XBRL review requires current passing case validation"
        )
    if output_dir.is_symlink():
        raise ValueError("XBRL review output directory must not be a symbolic link")
    output = output_dir.resolve()
    if output in {Path("/"), Path.home().resolve()}:
        raise ValueError("Refusing a broad XBRL review output directory")
    output.mkdir(parents=True, exist_ok=True)
    if any(output.iterdir()):
        raise ValueError("XBRL review output directory must be empty")
    if catalogue_path.is_symlink() or not catalogue_path.is_file():
        raise ValueError("Taxonomy catalogue must be a regular local file")
    catalogue_bytes = catalogue_path.read_bytes()
    content_hash = _review_content_hash(case)
    snapshot = _case_payload_for_hash(case)
    snapshot_hash = _sha256_bytes(_canonical_json(snapshot))
    candidate = {
        "state": CaseState.APPROVED,
        "approval": {"snapshot": snapshot, "snapshot_hash": snapshot_hash},
    }
    xml = render_xbrl(candidate, catalogue_path, catalogue_bytes=catalogue_bytes)
    candidate_path = output / "review-candidate.xbrl"
    candidate_path.write_bytes(xml)
    candidate_sha256 = _sha256_bytes(xml)
    report_path = output / "local-xbrl-validation.json"
    selected_validator = validator or validate_instance
    result = selected_validator(
        candidate_path,
        report_path,
        taxonomy_package,
        case.get("taxonomy_checksum"),
    )
    if not report_path.is_file():
        report_path.write_text(
            json.dumps(result, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
    if _sha256_file(candidate_path) != candidate_sha256:
        raise ValueError("Local validation modified the rendered XBRL review candidate")
    status = str(result.get("status", "FAIL"))
    if status not in {"PASS", "FAIL"}:
        raise ValueError("Local XBRL validator returned an unsupported status")
    _mutate(case, actor, "xbrl_review_prepared")
    case["xbrl_review"] = {
        "review_content_hash": content_hash,
        "candidate_snapshot_hash": snapshot_hash,
        "candidate_file_name": candidate_path.name,
        "candidate_sha256": candidate_sha256,
        "validation_report_file_name": report_path.name,
        "validation_report_sha256": _sha256_file(report_path),
        "catalogue_sha256": _sha256_bytes(catalogue_bytes),
        "taxonomy_package_sha256": result.get("taxonomy_package_sha256"),
        "processor": result.get("processor", "injected-validator"),
        "status": status,
        "prepared_by": actor,
        "prepared_at": _now(),
        "computation_context": _computation_context(case, "local-xbrl-review-v1"),
    }
    current_validation = validate_case(case)
    case["validation"] = current_validation
    case["state"] = (
        CaseState.READY_FOR_REVIEW
        if status == "PASS" and current_validation["status"] == "PASS"
        else CaseState.VALIDATION_FAILED
    )
    _record_event(
        case,
        "xbrl_review_prepared",
        actor,
        {
            "status": status,
            "candidate_sha256": case["xbrl_review"]["candidate_sha256"],
            "validation_report_sha256": case["xbrl_review"]["validation_report_sha256"],
        },
    )
    return case


def approve_case(
    case: dict[str, Any],
    reviewer: str,
    expected_revision: str,
    declaration: Mapping[str, Any],
) -> dict[str, Any]:
    """Create an immutable approval snapshot only from the validated current revision."""

    _ensure_revision(case, expected_revision)
    before_hash = _sha256_bytes(_canonical_json(_case_payload_for_hash(case)))
    validation = case.get("validation") or {}
    if (
        case["state"] != CaseState.READY_FOR_REVIEW
        or validation.get("status") != "PASS"
        or validation.get("validated_revision_id") != case["revision_id"]
    ):
        raise ValueError("Only a fully validated READY_FOR_REVIEW case can be approved")
    migrations = case.get("regulatory_migrations", [])
    if migrations and migrations[-1].get("revalidation_status") != "PASSED":
        raise ValueError("Regulatory migration requires a passing full revalidation")
    xbrl_review = case.get("xbrl_review") or {}
    if xbrl_review.get("status") != "PASS" or xbrl_review.get(
        "review_content_hash"
    ) != _review_content_hash(case):
        raise ValueError(
            "Approval requires a passing local XBRL review of the current content"
        )
    required = {
        "entity_period_confirmed",
        "form_confirmed",
        "evidence_reviewed",
        "preview_reviewed",
        "filing_boundary_understood",
        "rendered_output_confirmed",
        "outstanding_warnings_understood",
    }
    if any(declaration.get(key) is not True for key in required):
        raise ValueError("The reviewer declaration is incomplete")
    if int(validation.get("review_required", 0)):
        raise ValueError("All current warnings must be reviewed before approval")
    if (
        int(validation.get("approved_overrides", 0))
        and declaration.get("overrides_reviewed") is not True
    ):
        raise ValueError("The approval must explicitly confirm professional overrides")
    snapshot_payload = _case_payload_for_hash(case)
    snapshot_hash = _sha256_bytes(_canonical_json(snapshot_payload))
    case["approval"] = {
        "snapshot_id": f"snap_{len(case.get('approval_snapshots', [])) + 1:04d}",
        "revision_id": case["revision_id"],
        "snapshot_hash": snapshot_hash,
        "input_manifest_hash": _manifest_hash(case),
        "approved_by": reviewer,
        "approved_at": _now(),
        "declaration": dict(declaration),
        "snapshot": snapshot_payload,
    }
    case["state"] = CaseState.APPROVED
    case["updated_at"] = _now()
    case["_pending_before_hash"] = before_hash
    _record_event(case, "snapshot_approved", reviewer, {"snapshot_hash": snapshot_hash})
    return case


def _taxonomy_catalogue(
    path: Path,
    expected_id: str,
    expected_checksum: str | None,
    *,
    content: bytes | None = None,
) -> dict[str, Any]:
    if content is None:
        if path.is_symlink() or not path.is_file():
            raise ValueError("Taxonomy catalogue must be a regular local file")
        content = path.read_bytes()
    try:
        catalogue = json.loads(content.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ValueError("Taxonomy catalogue must be valid UTF-8 JSON") from exc
    if catalogue.get("schema_version") != 2:
        raise ValueError("Taxonomy catalogue schema version 2 is required")
    if catalogue.get("taxonomy_id") != expected_id:
        raise ValueError("Taxonomy catalogue identifier does not match the locked case")
    package_checksum = catalogue.get("taxonomy_package_sha256")
    if not package_checksum or package_checksum == "UNVERIFIED":
        raise ValueError(
            "Taxonomy catalogue is not bound to a verified official package checksum"
        )
    if expected_checksum and package_checksum != expected_checksum:
        raise ValueError("Taxonomy package checksum does not match the locked case")
    concepts = catalogue.get("concepts")
    if not isinstance(concepts, list) or not concepts:
        raise ValueError("Taxonomy catalogue has no concepts")
    qnames: set[str] = set()
    for concept in concepts:
        if not isinstance(concept, Mapping) or not str(concept.get("qname", "")):
            raise ValueError("Taxonomy catalogue contains an invalid concept record")
        qname = str(concept["qname"])
        if qname in qnames:
            raise ValueError(f"Taxonomy catalogue repeats concept {qname}")
        qnames.add(qname)
        if not isinstance(concept.get("is_item"), bool) or not isinstance(
            concept.get("is_tuple"), bool
        ):
            raise ValueError(
                f"Taxonomy concept lacks item-versus-tuple metadata: {qname}"
            )
        if concept["is_item"] is True and concept["is_tuple"] is True:
            raise ValueError(f"Taxonomy concept cannot be both item and tuple: {qname}")
    return catalogue


def _is_reportable_item(concept: Mapping[str, Any] | None) -> bool:
    """Return whether one taxonomy concept can be emitted as an item fact."""

    return bool(
        concept
        and concept.get("abstract") is not True
        and concept.get("is_item") is True
        and concept.get("is_tuple") is False
        and concept.get("period_type") in {"instant", "duration"}
    )


def _taxonomy_presentation_order(
    catalogue: Mapping[str, Any],
    selected_form: str,
    presentation: Mapping[str, Any],
) -> dict[str, int]:
    """Return stable concept ordinals from the selected official presentation roles."""

    role_order = [
        str(item["role"])
        for item in (presentation.get("inventory") or {}).get("roles", [])
    ]
    rows = list((catalogue.get("relationships") or {}).get("presentation", []))
    order: dict[str, int] = {}
    ordinal = 0
    for role in role_order:
        role_rows = [
            row
            for row in rows
            if str(row.get("form")) == selected_form and str(row.get("role")) == role
        ]
        children: dict[str, list[Mapping[str, Any]]] = {}
        parents: set[str] = set()
        targets: set[str] = set()
        for row in role_rows:
            parent = str(row["from"])
            parents.add(parent)
            targets.add(str(row["to"]))
            children.setdefault(parent, []).append(row)
        roots = sorted(parents - targets)
        visited: set[str] = set()

        def visit(qname: str) -> None:
            nonlocal ordinal
            if qname in visited:
                return
            visited.add(qname)
            if qname not in order:
                order[qname] = ordinal
                ordinal += 1
            for row in sorted(
                children.get(qname, []),
                key=lambda item: (
                    Decimal(str(item.get("order") or "0")),
                    str(item["to"]),
                ),
            ):
                visit(str(row["to"]))

        for root_qname in roots:
            visit(root_qname)
    return order


def render_xbrl(
    case: Mapping[str, Any],
    catalogue_path: Path,
    *,
    catalogue_bytes: bytes | None = None,
) -> bytes:
    """Render an approved snapshot into deterministic XBRL XML."""

    approval = case.get("approval")
    if not approval or case.get("state") not in {
        CaseState.APPROVED,
        CaseState.EXPORTED,
    }:
        raise ValueError("XBRL rendering requires an approved snapshot")
    snapshot = approval["snapshot"]
    expected_hash = _sha256_bytes(_canonical_json(snapshot))
    if expected_hash != approval["snapshot_hash"]:
        raise ValueError("Approval snapshot hash is invalid")
    catalogue = _taxonomy_catalogue(
        catalogue_path,
        snapshot["rule_pack_versions"]["taxonomy_id"],
        snapshot.get("taxonomy_checksum"),
        content=catalogue_bytes,
    )
    concept_lookup = {item["qname"]: item for item in catalogue["concepts"]}
    selected_form = snapshot.get("selected_form")
    presentation = snapshot.get("statutory_presentation") or {}
    presentation_order = _taxonomy_presentation_order(
        catalogue, str(selected_form), presentation
    )

    def fact_sort_key(fact: Mapping[str, Any]) -> tuple[int, str, str]:
        qname = str(fact.get("xbrl_concept") or "")
        return (
            presentation_order.get(qname, len(presentation_order)),
            qname,
            str(fact.get("fact_id") or fact.get("block_id") or ""),
        )

    namespaces = {
        "xbrli": XBRLI_NS,
        "xbrldi": XBRLDI_NS,
        "xsi": XSI_NS,
        "link": LINK_NS,
        "xlink": XLINK_NS,
        "iso4217": ISO4217_NS,
        **catalogue["namespaces"],
    }
    root = etree.Element(etree.QName(XBRLI_NS, "xbrl"), nsmap=namespaces)
    schema_ref = etree.SubElement(root, etree.QName(LINK_NS, "schemaRef"))
    schema_ref.set(etree.QName(XLINK_NS, "type"), "simple")
    entry_point = catalogue.get("entry_points", {}).get(selected_form)
    if not entry_point:
        raise ValueError(f"Taxonomy catalogue has no entry point for {selected_form}")
    schema_ref.set(etree.QName(XLINK_NS, "href"), entry_point)
    entity_identifier = str(snapshot["entity"]["tax_identifier"])
    period_end = snapshot["period"]["end"]
    period_start = snapshot["period"]["start"]
    current_start = date.fromisoformat(period_start)
    prior_end = str(snapshot["entity"].get("prior_period_end") or "")
    if not prior_end:
        prior_end = (current_start - timedelta(days=1)).isoformat()
    prior_start = str(snapshot["entity"].get("prior_period_start") or "")
    if not prior_start:
        prior_start = _previous_year_date(current_start).isoformat()
    if date.fromisoformat(prior_start) > date.fromisoformat(prior_end):
        raise ValueError("Approved comparative period start is after its end")
    contexts = {
        "current_instant": (None, period_end),
        "prior_instant": (None, prior_end),
        "current_duration": (period_start, period_end),
        "prior_duration": (prior_start, prior_end),
    }
    for context_id, (start, end) in contexts.items():
        context = etree.SubElement(
            root, etree.QName(XBRLI_NS, "context"), id=context_id
        )
        entity = etree.SubElement(context, etree.QName(XBRLI_NS, "entity"))
        identifier = etree.SubElement(
            entity,
            etree.QName(XBRLI_NS, "identifier"),
            scheme="http://www.registroimprese.it",
        )
        identifier.text = entity_identifier
        period = etree.SubElement(context, etree.QName(XBRLI_NS, "period"))
        if start is None:
            etree.SubElement(period, etree.QName(XBRLI_NS, "instant")).text = end
        else:
            etree.SubElement(period, etree.QName(XBRLI_NS, "startDate")).text = start
            etree.SubElement(period, etree.QName(XBRLI_NS, "endDate")).text = end
    dimension_contexts: dict[tuple[str, tuple[tuple[str, str], ...]], str] = {}
    for fact in [
        *snapshot.get("taxonomy_facts", []),
        *snapshot.get("schedule_taxonomy_facts", []),
    ]:
        dimensions = tuple(sorted(dict(fact.get("dimensions", {})).items()))
        if not dimensions:
            continue
        period_key = str(fact["period"])
        if period_key not in contexts:
            raise ValueError(f"Unsupported dimensional fact period: {period_key}")
        key = (period_key, dimensions)
        if key in dimension_contexts:
            continue
        for axis, member in dimensions:
            if axis not in concept_lookup or member not in concept_lookup:
                raise ValueError(f"Unknown dimension axis or member: {axis}={member}")
            axis_concept = concept_lookup[axis]
            member_concept = concept_lookup[member]
            if axis_concept.get("is_dimension_item") is not True:
                raise ValueError(f"Dimension axis is not a dimension item: {axis}")
            if (
                member_concept.get("is_item") is not True
                or member_concept.get("is_tuple") is not False
                or member_concept.get("is_dimension_item") is True
                or member_concept.get("is_hypercube_item") is True
            ):
                raise ValueError(f"Dimension member is not a domain item: {member}")
            if axis.split(":", 1)[0] not in namespaces:
                raise ValueError(f"Missing namespace for dimension axis: {axis}")
            if member.split(":", 1)[0] not in namespaces:
                raise ValueError(f"Missing namespace for dimension member: {member}")
        context_id = f"ctx_dim_{len(dimension_contexts) + 1:04d}"
        dimension_contexts[key] = context_id
        start, end = contexts[period_key]
        context = etree.SubElement(
            root, etree.QName(XBRLI_NS, "context"), id=context_id
        )
        entity = etree.SubElement(context, etree.QName(XBRLI_NS, "entity"))
        identifier = etree.SubElement(
            entity,
            etree.QName(XBRLI_NS, "identifier"),
            scheme="http://www.registroimprese.it",
        )
        identifier.text = entity_identifier
        period = etree.SubElement(context, etree.QName(XBRLI_NS, "period"))
        if start is None:
            etree.SubElement(period, etree.QName(XBRLI_NS, "instant")).text = end
        else:
            etree.SubElement(period, etree.QName(XBRLI_NS, "startDate")).text = start
            etree.SubElement(period, etree.QName(XBRLI_NS, "endDate")).text = end
        scenario = etree.SubElement(context, etree.QName(XBRLI_NS, "scenario"))
        for axis, member in dimensions:
            explicit = etree.SubElement(
                scenario, etree.QName(XBRLDI_NS, "explicitMember")
            )
            explicit.set("dimension", axis)
            explicit.text = member
    unit = etree.SubElement(root, etree.QName(XBRLI_NS, "unit"), id="EUR")
    etree.SubElement(unit, etree.QName(XBRLI_NS, "measure")).text = "iso4217:EUR"
    precision = int(snapshot.get("reporting_precision", 0))
    decimals = str(precision)
    output_language = str(snapshot.get("output_language", "it"))
    if output_language not in {"it", "en"}:
        raise ValueError("Approved snapshot output language must be it or en")
    seen: set[tuple[str, str, str]] = set()
    tuple_containers: dict[tuple[tuple[str, ...], str], etree._Element] = {}

    def tuple_parent(tuple_path: list[str], instance_id: str) -> etree._Element:
        """Create or reuse one deterministic nested tuple occurrence."""

        parent = root
        for depth, tuple_qname in enumerate(tuple_path, start=1):
            path_key = tuple(tuple_path[:depth])
            key = (path_key, instance_id)
            if key in tuple_containers:
                parent = tuple_containers[key]
                continue
            tuple_concept = concept_lookup.get(tuple_qname)
            if (
                not tuple_concept
                or tuple_concept.get("is_tuple") is not True
                or tuple_concept.get("is_item") is not False
                or tuple_concept.get("abstract") is True
            ):
                raise ValueError(
                    f"Tuple path contains a non-tuple concept: {tuple_qname}"
                )
            forms = tuple_concept.get("forms")
            if forms and selected_form not in forms:
                raise ValueError(
                    f"Tuple concept {tuple_qname} is not available for {selected_form}"
                )
            prefix, local_name = tuple_qname.split(":", 1)
            namespace = namespaces.get(prefix)
            if not namespace:
                raise ValueError(f"Missing namespace for tuple concept: {tuple_qname}")
            container = etree.SubElement(parent, etree.QName(namespace, local_name))
            tuple_digest = _sha256_bytes(
                _canonical_json({"path": path_key, "instance": instance_id})
            )[:20]
            container.set("id", _xbrl_fact_id("tuple", tuple_digest, depth))
            tuple_containers[key] = container
            parent = container
        return parent

    for fact in sorted(snapshot.get("canonical_facts", []), key=fact_sort_key):
        qname = fact.get("xbrl_concept")
        if not qname:
            continue
        if fact["status"] not in EXPORTABLE_STATUSES:
            raise ValueError(f"Fact {fact['fact_id']} has a non-exportable status")
        sign_multiplier = str(fact.get("xbrl_sign_multiplier", ""))
        if sign_multiplier not in {"1", "-1"}:
            raise ValueError(
                f"Fact {fact['fact_id']} has no reviewed XBRL sign convention"
            )
        concept = concept_lookup.get(qname)
        if not _is_reportable_item(concept):
            raise ValueError(
                f"Unknown, abstract, or non-item taxonomy concept: {qname}"
            )
        if "monetaryItemType" not in str(concept.get("type", "")):
            raise ValueError(f"Statement fact concept is not monetary: {qname}")
        forms = concept.get("forms")
        if forms and selected_form not in forms:
            raise ValueError(
                f"Taxonomy concept {qname} is not available for {selected_form}"
            )
        prefix, local_name = qname.split(":", 1)
        namespace = namespaces.get(prefix)
        if not namespace:
            raise ValueError(f"Missing namespace for concept: {qname}")
        current_context = (
            "current_instant"
            if concept["period_type"] == "instant"
            else "current_duration"
        )
        prior_context = (
            "prior_instant" if concept["period_type"] == "instant" else "prior_duration"
        )
        for context_ref, value in (
            (current_context, fact["current_value"]),
            (prior_context, fact["prior_value"]),
        ):
            duplicate_key = (qname, context_ref, "")
            if duplicate_key in seen:
                raise ValueError(
                    f"Conflicting duplicate fact: {qname} in {context_ref}"
                )
            seen.add(duplicate_key)
            element = etree.SubElement(root, etree.QName(namespace, local_name))
            element.set(
                "id",
                _xbrl_fact_id("canonical", fact["fact_id"], context_ref),
            )
            element.set("contextRef", context_ref)
            element.set("unitRef", "EUR")
            element.set("decimals", decimals)
            element.text = _decimal_text(
                _reported_decimal(
                    Decimal(str(value)) * Decimal(sign_multiplier), precision
                )
            )
    if (
        snapshot.get("statutory_presentation_required", True)
        and presentation.get("status") != "COMPLETE"
    ):
        raise ValueError(
            "XBRL rendering requires complete primary statutory presentation coverage"
        )
    if (
        snapshot.get("statutory_presentation_required", True)
        and (presentation.get("semantic_reconciliation") or {}).get("status") != "PASS"
    ):
        raise ValueError(
            "XBRL rendering requires canonical statements to reconcile to statutory roots"
        )
    for presentation_index, fact in enumerate(
        sorted(presentation.get("output_facts", []), key=fact_sort_key), start=1
    ):
        if fact.get("status") not in EXPORTABLE_STATUSES:
            raise ValueError("Statutory presentation fact is not exportable")
        if not fact.get("derivation") and not fact.get("confirmed_by"):
            raise ValueError("Statutory presentation fact has no provenance")
        qname = str(fact["xbrl_concept"])
        concept = concept_lookup.get(qname)
        if not _is_reportable_item(concept):
            raise ValueError(
                f"Unknown, abstract, or non-item presentation concept: {qname}"
            )
        if "monetaryItemType" not in str(concept.get("type", "")):
            raise ValueError(f"Presentation fact concept is not monetary: {qname}")
        forms = concept.get("forms")
        if forms and selected_form not in forms:
            raise ValueError(
                f"Presentation concept {qname} is not available for {selected_form}"
            )
        prefix, local_name = qname.split(":", 1)
        namespace = namespaces.get(prefix)
        if not namespace:
            raise ValueError(f"Missing namespace for presentation concept: {qname}")
        current_context = (
            "current_instant"
            if concept["period_type"] == "instant"
            else "current_duration"
        )
        prior_context = (
            "prior_instant" if concept["period_type"] == "instant" else "prior_duration"
        )
        for context_ref, value in (
            (current_context, fact.get("current_value")),
            (prior_context, fact.get("prior_value")),
        ):
            if value is None:
                continue
            duplicate_key = (qname, context_ref, "")
            if duplicate_key in seen:
                raise ValueError(
                    f"Conflicting duplicate fact: {qname} in {context_ref}"
                )
            seen.add(duplicate_key)
            element = etree.SubElement(root, etree.QName(namespace, local_name))
            element.set(
                "id",
                _xbrl_fact_id(
                    "presentation",
                    fact.get("fact_id") or presentation_index,
                    context_ref,
                ),
            )
            element.set("contextRef", context_ref)
            element.set("unitRef", "EUR")
            element.set("decimals", decimals)
            element.text = _decimal_text(
                _reported_decimal(Decimal(str(value)), precision)
            )
    for taxonomy_index, fact in enumerate(
        sorted(
            [
                *snapshot.get("taxonomy_facts", []),
                *snapshot.get("schedule_taxonomy_facts", []),
            ],
            key=fact_sort_key,
        ),
        start=1,
    ):
        qname = str(fact["xbrl_concept"])
        concept = concept_lookup.get(qname)
        if not _is_reportable_item(concept):
            raise ValueError(
                f"Unknown, abstract, or non-item taxonomy concept: {qname}"
            )
        forms = concept.get("forms")
        if forms and selected_form not in forms:
            raise ValueError(
                f"Taxonomy concept {qname} is not available for {selected_form}"
            )
        dimensions = tuple(sorted(dict(fact.get("dimensions", {})).items()))
        period_key = str(fact["period"])
        period_type = str(concept.get("period_type", ""))
        if period_type not in {"instant", "duration"} or not period_key.endswith(
            f"_{period_type}"
        ):
            raise ValueError(
                f"Taxonomy fact period does not match concept period type: {qname}"
            )
        context_ref = (
            dimension_contexts[(period_key, dimensions)] if dimensions else period_key
        )
        tuple_path = [str(item) for item in fact.get("tuple_path", [])]
        tuple_instance_id = fact.get("tuple_instance_id")
        if tuple_path and not str(tuple_instance_id or "").strip():
            raise ValueError("Tuple facts require a stable tuple instance identifier")
        if not tuple_path and tuple_instance_id is not None:
            raise ValueError("Flat facts cannot carry a tuple instance identifier")
        occurrence = f"{'/'.join(tuple_path)}#{tuple_instance_id}" if tuple_path else ""
        duplicate_key = (qname, context_ref, occurrence)
        if duplicate_key in seen:
            raise ValueError(f"Conflicting duplicate fact: {qname} in {context_ref}")
        seen.add(duplicate_key)
        prefix, local_name = qname.split(":", 1)
        namespace = namespaces.get(prefix)
        if not namespace:
            raise ValueError(f"Missing namespace for concept: {qname}")
        parent = (
            tuple_parent(tuple_path, str(tuple_instance_id)) if tuple_path else root
        )
        element = etree.SubElement(parent, etree.QName(namespace, local_name))
        element.set(
            "id",
            _xbrl_fact_id(
                "taxonomy", fact.get("fact_id") or taxonomy_index, context_ref
            ),
        )
        element.set("contextRef", context_ref)
        fact_type = str(fact["fact_type"])
        if fact_type == "MONETARY":
            if "monetaryItemType" not in str(concept.get("type", "")):
                raise ValueError(f"Monetary fact uses a non-monetary concept: {qname}")
            element.set("unitRef", "EUR")
            element.set("decimals", decimals)
            element.text = _decimal_text(
                _reported_decimal(Decimal(str(fact["value"])), precision)
            )
        elif fact_type == "TEXT":
            if "monetaryItemType" in str(concept.get("type", "")):
                raise ValueError(f"Text fact uses a monetary concept: {qname}")
            language = str(fact.get("language") or output_language)
            if language != output_language:
                raise ValueError("Text fact language differs from approved output")
            element.set(etree.QName(XML_NS, "lang"), language)
            element.text = str(fact["value"])
        elif fact_type == "NIL":
            if concept.get("nillable") is not True:
                raise ValueError(f"Taxonomy concept does not permit nil: {qname}")
            element.set(etree.QName(XSI_NS, "nil"), "true")
            if "monetaryItemType" in str(concept.get("type", "")):
                element.set("unitRef", "EUR")
                element.set("decimals", decimals)
        else:
            raise ValueError(f"Unsupported taxonomy fact type: {fact_type}")
    for narrative_index, block in enumerate(
        sorted(snapshot.get("narrative_blocks", []), key=fact_sort_key), start=1
    ):
        if block.get("status") != "ACCEPTED":
            continue
        qname = block.get("xbrl_concept")
        if not qname:
            raise ValueError("Accepted narrative requires an XBRL concept")
        language = str(block.get("language") or output_language)
        if language != output_language:
            raise ValueError("Narrative language differs from approved output")
        concept = concept_lookup.get(qname)
        if not _is_reportable_item(concept):
            raise ValueError(
                f"Unknown, abstract, or non-item narrative concept: {qname}"
            )
        if "monetaryItemType" in str(concept.get("type", "")):
            raise ValueError(f"Narrative block uses a monetary concept: {qname}")
        forms = concept.get("forms")
        if forms and selected_form not in forms:
            raise ValueError(
                f"Narrative concept {qname} is not available for {selected_form}"
            )
        prefix, local_name = str(qname).split(":", 1)
        namespace = namespaces.get(prefix)
        if not namespace:
            raise ValueError(f"Missing namespace for narrative concept: {qname}")
        context_ref = (
            "current_instant"
            if concept["period_type"] == "instant"
            else "current_duration"
        )
        duplicate_key = (str(qname), context_ref, "")
        if duplicate_key in seen:
            raise ValueError(f"Conflicting duplicate fact: {qname} in {context_ref}")
        seen.add(duplicate_key)
        element = etree.SubElement(root, etree.QName(namespace, local_name))
        element.set(
            "id",
            _xbrl_fact_id(
                "narrative", block.get("block_id") or narrative_index, context_ref
            ),
        )
        element.set("contextRef", context_ref)
        element.set(etree.QName(XML_NS, "lang"), language)
        element.text = str(block["text"])
    micro_reporting = snapshot.get("micro_reporting") or {}
    if micro_reporting.get("mode") == "FOOTER_ONLY":
        qname = str(micro_reporting.get("xbrl_concept") or "")
        text = str(micro_reporting.get("rendered_text") or "").strip()
        if not qname or not text or not micro_reporting.get("reviewed_by"):
            raise ValueError("Micro footer is not a reviewed renderable XBRL fact")
        concept = concept_lookup.get(qname)
        if not _is_reportable_item(concept):
            raise ValueError(
                f"Unknown, abstract, or non-item micro-footer concept: {qname}"
            )
        if "monetaryItemType" in str(concept.get("type", "")):
            raise ValueError(f"Micro footer uses a monetary concept: {qname}")
        forms = concept.get("forms")
        if forms and selected_form not in forms:
            raise ValueError(
                f"Micro-footer concept {qname} is not available for {selected_form}"
            )
        prefix, local_name = qname.split(":", 1)
        namespace = namespaces.get(prefix)
        if not namespace:
            raise ValueError(f"Missing namespace for micro-footer concept: {qname}")
        context_ref = (
            "current_instant"
            if concept["period_type"] == "instant"
            else "current_duration"
        )
        duplicate_key = (qname, context_ref, "")
        if duplicate_key in seen:
            raise ValueError(f"Conflicting duplicate fact: {qname} in {context_ref}")
        seen.add(duplicate_key)
        element = etree.SubElement(root, etree.QName(namespace, local_name))
        element.set("id", _xbrl_fact_id("micro_footer", "reviewed", context_ref))
        element.set("contextRef", context_ref)
        element.set(etree.QName(XML_NS, "lang"), output_language)
        element.text = text
    xml = etree.tostring(
        root, xml_declaration=True, encoding="UTF-8", pretty_print=True
    )
    fromstring(xml)
    return xml


def _safe_slug(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")
    return slug[:64] or "societa"


def _artifact_record(path: Path) -> dict[str, Any]:
    """Return the immutable checksum record for one written export artifact."""

    return {
        "file_name": path.name,
        "sha256": _sha256_file(path),
        "size_bytes": path.stat().st_size,
    }


def _mapping_report(
    case: Mapping[str, Any], snapshot: Mapping[str, Any]
) -> dict[str, Any]:
    """Build the standalone approved account-mapping review report."""

    entries = (snapshot.get("trial_balance") or {}).get("entries", [])
    decisions = {str(item["account_id"]): item for item in snapshot.get("mappings", [])}
    candidates = {
        str(item["account_id"]): item for item in snapshot.get("mapping_candidates", [])
    }
    rows = [
        {
            "account_id": entry["account_id"],
            "account_code": entry["account_code"],
            "account_description": entry["account_description"],
            "current_balance": entry["closing_signed"],
            "prior_balance": entry["prior_closing_signed"],
            "source_refs": list(entry.get("source_refs", [])),
            "candidate": candidates.get(str(entry["account_id"])),
            "decision": decisions.get(str(entry["account_id"])),
        }
        for entry in entries
    ]
    mapped = [item for item in decisions.values() if item.get("decision") == "ACCEPTED"]
    excluded = [
        item for item in decisions.values() if item.get("decision") == "EXCLUDED"
    ]
    return {
        "schema_version": 1,
        "report_type": "ACCOUNT_MAPPING",
        "case_id": case["case_id"],
        "snapshot_id": case["approval"]["snapshot_id"],
        "snapshot_hash": case["approval"]["snapshot_hash"],
        "revision_id": case["approval"]["revision_id"],
        "selected_form": snapshot.get("selected_form"),
        "summary": {
            "total_accounts": len(entries),
            "accepted_accounts": len(mapped),
            "excluded_accounts": len(excluded),
            "split_accounts": sum(
                len(item.get("allocations", [])) > 1 for item in mapped
            ),
            "unresolved_accounts": len(entries) - len(decisions),
        },
        "rows": rows,
        "manual_splits": [
            item for item in mapped if len(item.get("allocations", [])) > 1
        ],
        "exclusions": excluded,
        "adjustments": list(snapshot.get("adjustments", [])),
        "model_suggestions": list(snapshot.get("model_mapping_suggestions", [])),
        "statutory_presentation": {
            "status": (snapshot.get("statutory_presentation") or {}).get("status"),
            "summary": (snapshot.get("statutory_presentation") or {}).get("summary"),
            "inventory_sha256": (
                (snapshot.get("statutory_presentation") or {}).get("inventory") or {}
            ).get("inventory_sha256"),
        },
        "taxonomy_mapping_index": snapshot.get("taxonomy_mapping_index"),
    }


def _issue_report(
    case: Mapping[str, Any], snapshot: Mapping[str, Any]
) -> dict[str, Any]:
    """Build the standalone approved issues, decisions, and overrides report."""

    validation = snapshot.get("validation") or {}
    decisions = list(snapshot.get("review_decisions", []))
    return {
        "schema_version": 1,
        "report_type": "CASE_ISSUES",
        "case_id": case["case_id"],
        "snapshot_id": case["approval"]["snapshot_id"],
        "snapshot_hash": case["approval"]["snapshot_hash"],
        "revision_id": case["approval"]["revision_id"],
        "validation_status": validation.get("status"),
        "summary": {
            key: validation.get(key, 0)
            for key in (
                "blockers",
                "high",
                "medium",
                "low",
                "info",
                "review_required",
                "approved_overrides",
            )
        },
        "issues": list(validation.get("issues", [])),
        "review_decisions": decisions,
        "overrides": [item for item in decisions if item.get("action") == "OVERRIDDEN"],
        "unresolved_non_blocking_warnings": [
            issue
            for issue in validation.get("issues", [])
            if issue.get("severity") in {"MEDIUM", "LOW", "INFO"}
            and issue.get("review_status") == "UNREVIEWED"
        ],
        "local_xbrl_review": snapshot.get("xbrl_review"),
        "taxonomy_representation": snapshot.get("taxonomy_representation"),
        "external_validation_addendum": case.get("external_validation"),
    }


def export_case(
    case: dict[str, Any], output_dir: Path, catalogue_path: Path, actor: str
) -> dict[str, Any]:
    """Export reproducible XBRL, workpaper, validation, and checksum artifacts."""

    if case.get("state") not in {CaseState.APPROVED, CaseState.EXPORTED}:
        raise ValueError("Only an approved case can be exported")
    approval = case.get("approval") or {}
    snapshot = approval.get("snapshot")
    if not isinstance(snapshot, Mapping):
        raise ValueError("The approved snapshot is missing")
    if _sha256_bytes(_canonical_json(snapshot)) != approval.get("snapshot_hash"):
        raise ValueError("The approved snapshot checksum does not match its receipt")
    xbrl_review = snapshot.get("xbrl_review") or {}
    if xbrl_review.get("status") != "PASS":
        raise ValueError("Export requires the approved passing local XBRL review")
    if catalogue_path.is_symlink() or not catalogue_path.is_file():
        raise ValueError("Taxonomy catalogue must be a regular local file")
    catalogue_bytes = catalogue_path.read_bytes()
    if _sha256_bytes(catalogue_bytes) != xbrl_review.get("catalogue_sha256"):
        raise ValueError(
            "The export taxonomy catalogue differs from the approved review catalogue"
        )
    xml = render_xbrl(case, catalogue_path, catalogue_bytes=catalogue_bytes)
    if _sha256_bytes(xml) != xbrl_review.get("candidate_sha256"):
        raise ValueError(
            "The final XBRL bytes differ from the approved review candidate"
        )
    preview_bytes = _reviewed_preview_bytes(snapshot.get("preview") or {})
    before_hash = _sha256_bytes(_canonical_json(_case_payload_for_hash(case)))
    if output_dir.is_symlink():
        raise ValueError("Export directory must not be a symbolic link")
    output = output_dir.resolve()
    if output in {Path("/"), Path.home().resolve()}:
        raise ValueError("Refusing a broad export directory")
    output.mkdir(parents=True, exist_ok=True)
    if any(output.iterdir()):
        raise ValueError("Export directory must be empty")
    year = date.fromisoformat(snapshot["period"]["end"]).year
    taxonomy_slug = (
        snapshot["rule_pack_versions"]["taxonomy_id"].replace("PCI_", "pci").lower()
    )
    name = f"bilancio_{_safe_slug(snapshot['entity']['legal_name'])}_{year}_{taxonomy_slug}.xbrl"
    xbrl_path = output / name
    xbrl_path.write_bytes(xml)
    additional_artifact_paths: list[Path] = []
    taxonomy_representation = snapshot.get("taxonomy_representation")
    if taxonomy_representation and taxonomy_representation.get("mismatch_present"):
        differences_path = output / "taxonomy_differences.json"
        differences_path.write_bytes(_canonical_json(taxonomy_representation) + b"\n")
        additional_artifact_paths.append(differences_path)
    mapping_report_path = output / "mapping_report.json"
    mapping_report_path.write_bytes(
        _canonical_json(_mapping_report(case, snapshot)) + b"\n"
    )
    issue_report_path = output / "issue_report.json"
    issue_report_path.write_bytes(
        _canonical_json(_issue_report(case, snapshot)) + b"\n"
    )
    validation_report = {
        "schema_version": 1,
        "report_type": "CASE_VALIDATION",
        "case_id": case["case_id"],
        "snapshot_id": approval["snapshot_id"],
        "snapshot_hash": approval["snapshot_hash"],
        "revision_id": approval["revision_id"],
        "taxonomy_id": snapshot["rule_pack_versions"]["taxonomy_id"],
        "taxonomy_package_sha256": snapshot.get("taxonomy_checksum"),
        "case_validation": snapshot["validation"],
        "local_xbrl_review": snapshot.get("xbrl_review"),
        "external_validation_addendum": case.get("external_validation"),
    }
    validation_path = output / "validation_report.json"
    validation_path.write_bytes(_canonical_json(validation_report) + b"\n")
    preview_path = output / "preview.html"
    preview_path.write_bytes(preview_bytes)
    peer_artifact_paths = [
        xbrl_path,
        mapping_report_path,
        issue_report_path,
        validation_path,
        preview_path,
        *additional_artifact_paths,
    ]
    peer_artifacts = [_artifact_record(path) for path in peer_artifact_paths]
    workpaper = {
        "case_id": case["case_id"],
        "entity": snapshot["entity"],
        "period": snapshot["period"],
        "approval": {
            key: value for key, value in approval.items() if key != "snapshot"
        },
        "rule_pack_versions": snapshot["rule_pack_versions"],
        "taxonomy_checksum": snapshot.get("taxonomy_checksum"),
        "source_documents": snapshot["source_documents"],
        "file_security_scans": snapshot.get("file_security_scans", []),
        "prior_xbrl": snapshot.get("prior_xbrl"),
        "prior_xbrl_reconciliation": _prior_xbrl_reconciliation(snapshot),
        "comparative_reconciliation_decisions": snapshot.get(
            "comparative_reconciliation_decisions", []
        ),
        "parser_calibration": (snapshot.get("trial_balance") or {}).get("calibration"),
        "mapping_candidates": snapshot.get("mapping_candidates", []),
        "taxonomy_mapping_index": snapshot.get("taxonomy_mapping_index"),
        "mappings": snapshot["mappings"],
        "adjustments": snapshot.get("adjustments", []),
        "taxonomy_facts": snapshot.get("taxonomy_facts", []),
        "schedule_taxonomy_adapter": snapshot.get("schedule_taxonomy_adapter"),
        "schedule_taxonomy_facts": snapshot.get("schedule_taxonomy_facts", []),
        "taxonomy_fact_context": snapshot.get("taxonomy_fact_context"),
        "taxonomy_representation": taxonomy_representation,
        "micro_reporting": snapshot.get("micro_reporting"),
        "form_analysis": snapshot["form_analysis"],
        "statements": snapshot["statements"],
        "statutory_presentation": snapshot.get("statutory_presentation"),
        "schedules": snapshot.get("schedules", []),
        "disclosure_coverage": snapshot.get("disclosure_coverage"),
        "disclosure_trigger_decisions": snapshot.get(
            "disclosure_trigger_decisions", []
        ),
        "disclosure_activation_suggestions": snapshot.get(
            "disclosure_activation_suggestions", []
        ),
        "questionnaire": snapshot.get("questionnaire", []),
        "disclosure_answers": snapshot["disclosure_answers"],
        "note_outline": snapshot.get("note_outline", []),
        "note_outline_context": snapshot.get("note_outline_context"),
        "narrative_blocks": snapshot.get("narrative_blocks", []),
        "narrative_context": snapshot.get("narrative_context"),
        "narrative_change_log": [
            {
                "block_id": block["block_id"],
                "prior_suggestion_id": block.get("prior_suggestion_id"),
                "redline": block.get("redline"),
                "reviewed_by": block.get("reviewed_by"),
            }
            for block in snapshot.get("narrative_blocks", [])
        ],
        "prior_narrative_suggestions": snapshot.get("prior_narrative_suggestions", []),
        "intelligence_runs": snapshot.get("intelligence_runs", []),
        "latest_workflow_guidance": snapshot.get("latest_workflow_guidance"),
        "model_mapping_suggestions": snapshot.get("model_mapping_suggestions", []),
        "narrative_suggestions": snapshot.get("narrative_suggestions", []),
        "review_decisions": snapshot.get("review_decisions", []),
        "regulatory_migrations": snapshot.get("regulatory_migrations", []),
        "overrides": [
            decision
            for decision in snapshot.get("review_decisions", [])
            if decision.get("action") == "OVERRIDDEN"
        ],
        "unresolved_non_blocking_warnings": [
            issue
            for issue in snapshot["validation"].get("issues", [])
            if issue.get("severity") in {"MEDIUM", "LOW", "INFO"}
            and issue.get("review_status") == "UNREVIEWED"
        ],
        "preview": {
            key: value
            for key, value in (snapshot.get("preview") or {}).items()
            if key != "content_base64"
        },
        "xbrl_review": snapshot.get("xbrl_review"),
        "validation": snapshot["validation"],
        "assumptions": [],
        "assumption_policy": (
            "Missing or model-suggested facts are not assumptions and are not exportable."
        ),
        "audit_events": case["audit_events"],
        "filing_boundary": "Vera does not sign, approve corporate accounts, or submit the filing.",
        "external_validation": case.get("external_validation"),
        "external_validation_documents": [
            document
            for document in case.get("source_documents", [])
            if document.get("purpose") == "EXTERNAL_VALIDATION_REPORT"
        ],
        "external_validation_boundary": "Manual user-controlled TEBENI validation remains required.",
        "artifact_manifest": {
            "file_name": "artifact_manifest.json",
            "snapshot_id": approval["snapshot_id"],
            "snapshot_hash": approval["snapshot_hash"],
            "peer_artifacts": peer_artifacts,
            "integrity_note": (
                "The workpaper and final manifest checksums are recorded in "
                "artifact_manifest.json to avoid recursive self-hashing."
            ),
        },
    }
    workpaper_path = output / "workpaper.json"
    workpaper_path.write_bytes(_canonical_json(workpaper) + b"\n")
    artifacts = [*peer_artifacts, _artifact_record(workpaper_path)]
    manifest = {
        "case_id": case["case_id"],
        "snapshot_id": approval["snapshot_id"],
        "snapshot_hash": approval["snapshot_hash"],
        "exported_by": actor,
        "exported_at": _now(),
        "artifacts": artifacts,
    }
    manifest_path = output / "artifact_manifest.json"
    manifest_path.write_bytes(_canonical_json(manifest) + b"\n")
    case["state"] = CaseState.EXPORTED
    case["updated_at"] = _now()
    case["artifacts"] = [
        *artifacts,
        {
            "file_name": manifest_path.name,
            "sha256": _sha256_file(manifest_path),
            "size_bytes": manifest_path.stat().st_size,
        },
    ]
    case["_pending_before_hash"] = before_hash
    _record_event(case, "artifact_exported", actor, {"manifest": manifest})
    return case


def _read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def _write_result(case_dir: Path, case: dict[str, Any]) -> None:
    save_case(case_dir, case)
    sys.stdout.write(
        json.dumps(
            {
                "case_id": case["case_id"],
                "revision_id": case["revision_id"],
                "state": case["state"],
            },
            ensure_ascii=False,
        )
        + "\n"
    )


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    sub = parser.add_subparsers(dest="command", required=True)
    create = sub.add_parser("create")
    create.add_argument("--case-dir", type=Path, required=True)
    create.add_argument("--payload", type=Path, required=True)
    create.add_argument("--rule-pack", type=Path, required=True)
    create.add_argument("--actor", required=True)
    ingest = sub.add_parser("ingest")
    ingest.add_argument("--case-dir", type=Path, required=True)
    ingest.add_argument("--source", type=Path, required=True)
    ingest.add_argument("--sheet")
    ingest.add_argument("--revision", required=True)
    ingest.add_argument("--actor", required=True)
    prior = sub.add_parser("ingest-prior-xbrl")
    prior.add_argument("--case-dir", type=Path, required=True)
    prior.add_argument("--source", type=Path, required=True)
    prior.add_argument("--revision", required=True)
    prior.add_argument("--actor", required=True)
    supporting = sub.add_parser("attach-supporting-document")
    supporting.add_argument("--case-dir", type=Path, required=True)
    supporting.add_argument("--source", type=Path, required=True)
    supporting.add_argument("--purpose", required=True)
    supporting.add_argument("--description", required=True)
    supporting.add_argument("--revision", required=True)
    supporting.add_argument("--actor", required=True)
    confirm = sub.add_parser("confirm-parser")
    confirm.add_argument("--case-dir", type=Path, required=True)
    confirm.add_argument(
        "--convention",
        choices=[
            item.value
            for item in ParserConvention
            if item is not ParserConvention.UNKNOWN
        ],
        required=True,
    )
    confirm.add_argument("--revision", required=True)
    confirm.add_argument("--actor", required=True)
    forms = sub.add_parser("determine-forms")
    forms.add_argument("--case-dir", type=Path, required=True)
    forms.add_argument("--metrics", type=Path, required=True)
    forms.add_argument("--rule-pack", type=Path, required=True)
    forms.add_argument("--revision", required=True)
    forms.add_argument("--actor", required=True)
    select = sub.add_parser("select-form")
    select.add_argument("--case-dir", type=Path, required=True)
    select.add_argument(
        "--form", choices=["MICRO", "ABBREVIATED", "ORDINARY"], required=True
    )
    select.add_argument("--revision", required=True)
    select.add_argument("--actor", required=True)
    taxonomy_index = sub.add_parser("taxonomy-mapping-index")
    taxonomy_index.add_argument("--case-dir", type=Path, required=True)
    taxonomy_index.add_argument("--catalogue", type=Path, required=True)
    taxonomy_index.add_argument("--rule-pack", type=Path, required=True)
    taxonomy_index.add_argument("--revision", required=True)
    taxonomy_index.add_argument("--actor", required=True)
    mapping = sub.add_parser("apply-mappings")
    mapping.add_argument("--case-dir", type=Path, required=True)
    mapping.add_argument("--decisions", type=Path, required=True)
    mapping.add_argument("--revision", required=True)
    mapping.add_argument("--actor", required=True)
    adjustments = sub.add_parser("record-adjustments")
    adjustments.add_argument("--case-dir", type=Path, required=True)
    adjustments.add_argument("--adjustments", type=Path, required=True)
    adjustments.add_argument("--revision", required=True)
    adjustments.add_argument("--actor", required=True)
    comparative = sub.add_parser("record-comparative-reconciliation")
    comparative.add_argument("--case-dir", type=Path, required=True)
    comparative.add_argument("--decisions", type=Path, required=True)
    comparative.add_argument("--revision", required=True)
    comparative.add_argument("--reviewer", required=True)
    taxonomy_facts = sub.add_parser("record-taxonomy-facts")
    taxonomy_facts.add_argument("--case-dir", type=Path, required=True)
    taxonomy_facts.add_argument("--facts", type=Path, required=True)
    taxonomy_facts.add_argument("--revision", required=True)
    taxonomy_facts.add_argument("--actor", required=True)
    statutory_presentation = sub.add_parser("record-statutory-presentation")
    statutory_presentation.add_argument("--case-dir", type=Path, required=True)
    statutory_presentation.add_argument("--catalogue", type=Path, required=True)
    statutory_presentation.add_argument("--rule-pack", type=Path, required=True)
    statutory_presentation.add_argument("--decisions", type=Path, required=True)
    statutory_presentation.add_argument("--revision", required=True)
    statutory_presentation.add_argument("--actor", required=True)
    taxonomy_representation = sub.add_parser("record-taxonomy-representation")
    taxonomy_representation.add_argument("--case-dir", type=Path, required=True)
    taxonomy_representation.add_argument("--payload", type=Path, required=True)
    taxonomy_representation.add_argument("--revision", required=True)
    taxonomy_representation.add_argument("--reviewer", required=True)
    micro_reporting = sub.add_parser("record-micro-reporting")
    micro_reporting.add_argument("--case-dir", type=Path, required=True)
    micro_reporting.add_argument("--payload", type=Path, required=True)
    micro_reporting.add_argument("--revision", required=True)
    micro_reporting.add_argument("--reviewer", required=True)
    candidates = sub.add_parser("mapping-candidates")
    candidates.add_argument("--case-dir", type=Path, required=True)
    candidates.add_argument("--memory", type=Path, required=True)
    candidates.add_argument("--source-system-template", required=True)
    candidates.add_argument("--revision", required=True)
    candidates.add_argument("--actor", required=True)
    compute = sub.add_parser("compute-statements")
    compute.add_argument("--case-dir", type=Path, required=True)
    compute.add_argument("--revision", required=True)
    compute.add_argument("--actor", required=True)
    schedule = sub.add_parser("record-schedule")
    schedule.add_argument("--case-dir", type=Path, required=True)
    schedule.add_argument("--payload", type=Path, required=True)
    schedule.add_argument("--revision", required=True)
    schedule.add_argument("--actor", required=True)
    schedule_taxonomy = sub.add_parser("record-schedule-taxonomy-adapter")
    schedule_taxonomy.add_argument("--case-dir", type=Path, required=True)
    schedule_taxonomy.add_argument("--catalogue", type=Path, required=True)
    schedule_taxonomy.add_argument("--rule-pack", type=Path, required=True)
    schedule_taxonomy.add_argument("--decisions", type=Path, required=True)
    schedule_taxonomy.add_argument("--revision", required=True)
    schedule_taxonomy.add_argument("--actor", required=True)
    schedule_file = sub.add_parser("ingest-schedule")
    schedule_file.add_argument("--case-dir", type=Path, required=True)
    schedule_file.add_argument("--source", type=Path, required=True)
    schedule_file.add_argument(
        "--schedule-type",
        choices=sorted(SCHEDULE_TYPES - {"CASH_FLOW"}),
        required=True,
    )
    schedule_file.add_argument("--schedule-id", required=True)
    schedule_file.add_argument("--statement-line")
    schedule_file.add_argument("--options", type=Path, required=True)
    schedule_file.add_argument("--sheet")
    schedule_file.add_argument("--revision", required=True)
    schedule_file.add_argument("--actor", required=True)
    disclosures = sub.add_parser("activate-disclosures")
    disclosures.add_argument("--case-dir", type=Path, required=True)
    disclosures.add_argument("--rule-pack", type=Path, required=True)
    disclosures.add_argument("--revision", required=True)
    disclosures.add_argument("--actor", required=True)
    disclosure_triggers = sub.add_parser("record-disclosure-triggers")
    disclosure_triggers.add_argument("--case-dir", type=Path, required=True)
    disclosure_triggers.add_argument("--decisions", type=Path, required=True)
    disclosure_triggers.add_argument("--revision", required=True)
    disclosure_triggers.add_argument("--reviewer", required=True)
    answers = sub.add_parser("record-answers")
    answers.add_argument("--case-dir", type=Path, required=True)
    answers.add_argument("--answers", type=Path, required=True)
    answers.add_argument("--revision", required=True)
    answers.add_argument("--actor", required=True)
    narratives = sub.add_parser("record-narratives")
    narratives.add_argument("--case-dir", type=Path, required=True)
    narratives.add_argument("--blocks", type=Path, required=True)
    narratives.add_argument("--revision", required=True)
    narratives.add_argument("--actor", required=True)
    preview = sub.add_parser("preview")
    preview.add_argument("--case-dir", type=Path, required=True)
    preview.add_argument("--output", type=Path, required=True)
    preview.add_argument("--revision", required=True)
    preview.add_argument("--actor", required=True)
    validate = sub.add_parser("validate")
    validate.add_argument("--case-dir", type=Path, required=True)
    validate.add_argument("--revision", required=True)
    validate.add_argument("--actor", required=True)
    xbrl_review = sub.add_parser("prepare-xbrl-review")
    xbrl_review.add_argument("--case-dir", type=Path, required=True)
    xbrl_review.add_argument("--catalogue", type=Path, required=True)
    xbrl_review.add_argument("--taxonomy-package", type=Path, required=True)
    xbrl_review.add_argument("--output-dir", type=Path, required=True)
    xbrl_review.add_argument("--revision", required=True)
    xbrl_review.add_argument("--actor", required=True)
    issue_reviews = sub.add_parser("record-issue-reviews")
    issue_reviews.add_argument("--case-dir", type=Path, required=True)
    issue_reviews.add_argument("--decisions", type=Path, required=True)
    issue_reviews.add_argument("--revision", required=True)
    issue_reviews.add_argument("--reviewer", required=True)
    approve = sub.add_parser("approve")
    approve.add_argument("--case-dir", type=Path, required=True)
    approve.add_argument("--declaration", type=Path, required=True)
    approve.add_argument("--revision", required=True)
    approve.add_argument("--reviewer", required=True)
    export = sub.add_parser("export")
    export.add_argument("--case-dir", type=Path, required=True)
    export.add_argument("--output-dir", type=Path, required=True)
    export.add_argument("--catalogue", type=Path, required=True)
    export.add_argument("--actor", required=True)
    external = sub.add_parser("record-external-validation")
    external.add_argument("--case-dir", type=Path, required=True)
    external.add_argument("--report", type=Path, required=True)
    external.add_argument(
        "--result", choices=["PASS", "FAIL", "WARNING"], required=True
    )
    external.add_argument("--issues", type=Path, required=True)
    external.add_argument("--revision", required=True)
    external.add_argument("--actor", required=True)
    intel_packet = sub.add_parser("intelligence-packet")
    intel_packet.add_argument("--case-dir", type=Path, required=True)
    intel_packet.add_argument(
        "--task", choices=[item.value for item in IntelligenceTask], required=True
    )
    intel_packet.add_argument("--subject-id", action="append", default=[])
    intel = sub.add_parser("record-intelligence")
    intel.add_argument("--case-dir", type=Path, required=True)
    intel.add_argument("--payload", type=Path, required=True)
    intel.add_argument("--revision", required=True)
    intel.add_argument("--actor", required=True)
    remember = sub.add_parser("remember-mappings")
    remember.add_argument("--case-dir", type=Path, required=True)
    remember.add_argument("--memory", type=Path, required=True)
    remember.add_argument("--source-system-template", required=True)
    remember.add_argument("--revision", required=True)
    remember.add_argument("--actor", required=True)
    history_load = sub.add_parser("load-client-history")
    history_load.add_argument("--case-dir", type=Path, required=True)
    history_load.add_argument("--history", type=Path, required=True)
    history_load.add_argument("--revision", required=True)
    history_load.add_argument("--actor", required=True)
    history_remember = sub.add_parser("remember-client-history")
    history_remember.add_argument("--case-dir", type=Path, required=True)
    history_remember.add_argument("--history", type=Path, required=True)
    history_remember.add_argument("--revision", required=True)
    history_remember.add_argument("--actor", required=True)
    status = sub.add_parser("status")
    status.add_argument("--case-dir", type=Path, required=True)
    return parser


def main(argv: list[str] | None = None) -> int:
    """Run the deterministic case command line interface."""

    args = _parser().parse_args(argv)
    try:
        if args.command == "create":
            case = create_case(
                args.case_dir,
                _read_json(args.payload),
                _read_json(args.rule_pack),
                args.actor,
            )
        else:
            case = load_case(args.case_dir)
            if args.command == "ingest":
                case = ingest_trial_balance(
                    case, args.source, args.actor, args.revision, args.sheet
                )
            elif args.command == "ingest-prior-xbrl":
                case = ingest_prior_xbrl(case, args.source, args.actor, args.revision)
            elif args.command == "attach-supporting-document":
                case = attach_supporting_document(
                    case,
                    args.source,
                    args.purpose,
                    args.description,
                    args.actor,
                    args.revision,
                )
            elif args.command == "confirm-parser":
                case = confirm_parser(case, args.convention, args.actor, args.revision)
            elif args.command == "determine-forms":
                case = determine_forms(
                    case,
                    _read_json(args.metrics),
                    _read_json(args.rule_pack),
                    args.actor,
                    args.revision,
                )
            elif args.command == "select-form":
                case = select_form(case, args.form, args.actor, args.revision)
            elif args.command == "taxonomy-mapping-index":
                case = record_taxonomy_mapping_index(
                    case,
                    args.catalogue,
                    _read_json(args.rule_pack),
                    args.actor,
                    args.revision,
                )
            elif args.command == "apply-mappings":
                case = apply_mapping_decisions(
                    case, _read_json(args.decisions), args.actor, args.revision
                )
            elif args.command == "record-adjustments":
                case = record_adjustments(
                    case,
                    _read_json(args.adjustments),
                    args.actor,
                    args.revision,
                )
            elif args.command == "record-comparative-reconciliation":
                case = record_comparative_reconciliation_decisions(
                    case,
                    _read_json(args.decisions),
                    args.reviewer,
                    args.revision,
                )
            elif args.command == "record-taxonomy-facts":
                case = record_taxonomy_facts(
                    case,
                    _read_json(args.facts),
                    args.actor,
                    args.revision,
                )
            elif args.command == "record-statutory-presentation":
                case = record_statutory_presentation(
                    case,
                    args.catalogue,
                    _read_json(args.rule_pack),
                    _read_json(args.decisions),
                    args.actor,
                    args.revision,
                )
            elif args.command == "record-taxonomy-representation":
                case = record_taxonomy_representation(
                    case,
                    _read_json(args.payload),
                    args.reviewer,
                    args.revision,
                )
            elif args.command == "record-micro-reporting":
                case = record_micro_reporting(
                    case,
                    _read_json(args.payload),
                    args.reviewer,
                    args.revision,
                )
            elif args.command == "mapping-candidates":
                case = generate_mapping_candidates(
                    case,
                    args.memory,
                    args.source_system_template,
                    args.actor,
                    args.revision,
                )
            elif args.command == "compute-statements":
                case = build_statements(case, args.actor, args.revision)
            elif args.command == "record-schedule":
                case = record_schedule(
                    case, _read_json(args.payload), args.actor, args.revision
                )
            elif args.command == "record-schedule-taxonomy-adapter":
                case = record_schedule_taxonomy_adapter(
                    case,
                    args.catalogue,
                    _read_json(args.rule_pack),
                    _read_json(args.decisions),
                    args.actor,
                    args.revision,
                )
            elif args.command == "ingest-schedule":
                case = ingest_schedule_file(
                    case,
                    args.source,
                    args.schedule_type,
                    args.schedule_id,
                    args.statement_line,
                    _read_json(args.options),
                    args.actor,
                    args.revision,
                    args.sheet,
                )
            elif args.command == "activate-disclosures":
                case = activate_disclosures(
                    case, _read_json(args.rule_pack), args.actor, args.revision
                )
            elif args.command == "record-disclosure-triggers":
                case = record_disclosure_trigger_decisions(
                    case,
                    _read_json(args.decisions),
                    args.reviewer,
                    args.revision,
                )
            elif args.command == "record-answers":
                case = record_disclosure_answers(
                    case, _read_json(args.answers), args.actor, args.revision
                )
            elif args.command == "record-narratives":
                case = record_narrative_blocks(
                    case, _read_json(args.blocks), args.actor, args.revision
                )
            elif args.command == "preview":
                case = create_preview(case, args.output, args.actor, args.revision)
            elif args.command == "validate":
                case = run_validation(case, args.actor, args.revision)
            elif args.command == "prepare-xbrl-review":
                case = prepare_xbrl_review(
                    case,
                    args.catalogue,
                    args.taxonomy_package,
                    args.output_dir,
                    args.actor,
                    args.revision,
                )
            elif args.command == "record-issue-reviews":
                case = record_issue_reviews(
                    case,
                    _read_json(args.decisions),
                    args.reviewer,
                    args.revision,
                )
            elif args.command == "approve":
                case = approve_case(
                    case, args.reviewer, args.revision, _read_json(args.declaration)
                )
            elif args.command == "export":
                case = export_case(case, args.output_dir, args.catalogue, args.actor)
            elif args.command == "record-external-validation":
                case = record_external_validation(
                    case,
                    args.report,
                    args.result,
                    _read_json(args.issues),
                    args.actor,
                    args.revision,
                )
            elif args.command == "intelligence-packet":
                packet = build_intelligence_packet(case, args.task, args.subject_id)
                sys.stdout.write(
                    json.dumps(packet, ensure_ascii=False, indent=2) + "\n"
                )
                return 0
            elif args.command == "record-intelligence":
                payload = _read_json(args.payload)
                case = record_intelligence_suggestion(
                    case,
                    payload["task"],
                    payload.get("subject_ids", []),
                    payload["output"],
                    payload["model_metadata"],
                    args.actor,
                    args.revision,
                )
            elif args.command == "remember-mappings":
                case = remember_mappings(
                    case,
                    args.memory,
                    args.source_system_template,
                    args.actor,
                    args.revision,
                )
            elif args.command == "load-client-history":
                case = load_client_history(
                    case, args.history, args.actor, args.revision
                )
            elif args.command == "remember-client-history":
                case = remember_client_history(
                    case, args.history, args.actor, args.revision
                )
            elif args.command == "status":
                sys.stdout.write(json.dumps(case, ensure_ascii=False, indent=2) + "\n")
                return 0
            else:
                raise AssertionError(args.command)
        _write_result(args.case_dir, case)
        return 0
    except (KeyError, ValueError, OSError, json.JSONDecodeError) as exc:
        LOGGER.error("%s", exc)
        return 2


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    raise SystemExit(main())
