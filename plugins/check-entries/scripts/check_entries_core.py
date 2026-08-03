from __future__ import annotations

import argparse
import csv
import hashlib
import json
import logging
import os
import re
import shutil
import stat

# Fixed isolated upstream assurance bridge.
import subprocess  # nosec B404
import sys
import tempfile
import unicodedata
import zipfile
from collections.abc import Mapping
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any, Sequence

import openpyxl
import polars as pl
from invoice_support import (
    InvoiceRecord,
    fatturapa_document_polarity,
    load_invoice_payloads,
    match_invoice,
)
from stable_ooxml import write_stable_xlsx

SCRIPT_DIR = Path(__file__).resolve().parent
_COMPONENT_ROOT = Path(__file__).resolve().parents[1]
_VENDOR_CANDIDATES = (
    _COMPONENT_ROOT / "vendor" / "modules",
    _COMPONENT_ROOT.parent.parent / "vendor" / "modules",
    _COMPONENT_ROOT.parent / "_shared" / "vendor" / "modules",
)
if "vera_assurance" not in sys.modules:
    for _vendor_candidate in _VENDOR_CANDIDATES:
        if (_vendor_candidate / "vera_assurance").is_dir():
            if str(_vendor_candidate) not in sys.path:
                sys.path.insert(0, str(_vendor_candidate))
            break

from vera_assurance import (  # noqa: E402
    MoneyValidationError,
    artifact_receipt,
    build_assurance_envelope,
    build_gate_register,
    build_numeric_evidence_ledger,
    build_source_qualification,
    canonical_json_sha256,
    decimal_text,
    difference_within_tolerance,
    load_client_engagement_context_file,
    parse_canonical_decimal,
    parse_localized_decimal,
    validate_artifact_receipt,
    validate_assurance_envelope,
    validate_client_engagement_context,
    validate_client_workflow_run,
    validate_gate_register,
    validate_reviewed_decision_receipt,
    validate_source_qualification,
)
from vera_assurance import write_json as write_assurance_json  # noqa: E402


def _ensure_local_review_session_import() -> None:
    """Use this plugin's review-session module in multi-plugin test runs."""

    script_dir = str(SCRIPT_DIR)
    if script_dir in sys.path:
        sys.path.remove(script_dir)
    sys.path.insert(0, script_dir)
    module = sys.modules.get("review_session")
    module_file = getattr(module, "__file__", None) if module is not None else None
    if module_file and Path(module_file).resolve().is_relative_to(SCRIPT_DIR.resolve()):
        return
    if module is not None:
        del sys.modules["review_session"]


_ensure_local_review_session_import()
from implementation_contract import (
    build_implementation_receipts,
    implementation_artifact_roots,
    validate_implementation_contract,
)
from physical_output_set import validate_initial_output_set
from review_session import write_review_session_artifacts, write_run_intake

LOGGER = logging.getLogger(__name__)

SUPPORTED_LANGUAGES = ("it", "en", "fr", "de", "es")
PDF_SUFFIXES = {".pdf"}
MAX_SUPPORT_BYTES = 100 * 1024 * 1024
MAX_ZIP_MEMBERS = 10_000
MAX_ZIP_MEMBER_BYTES = 20 * 1024 * 1024
MAX_ZIP_TOTAL_BYTES = 100 * 1024 * 1024
MAX_ZIP_COMPRESSION_RATIO = 100
RELATIONSHIP_ADAPTER_ID = "check_entries.relationship"
RELATIONSHIP_ADAPTER_VERSION = "1"
PARTY_ADAPTER_ID = "check_entries.party_perimeter"
PARTY_ADAPTER_VERSION = "1"
CURRENCY_ADAPTER_ID = "check_entries.currency"
CURRENCY_ADAPTER_VERSION = "1"
DIRECTION_ADAPTER_ID = "check_entries.direction"
DIRECTION_ADAPTER_VERSION = "1"
NORMALIZED_JOURNAL_ARTIFACT_ID = "source.normalized_journal"
NORMALIZATION_SCHEMA_VERSION = "journal_sampling.normalization.v2"
JOURNAL_HANDOFF_ARTIFACT_PATHS = frozenset(
    {
        "normalization/normalized_journal.csv",
        "normalization/normalization_diagnostics.json",
        "normalization/normalization_recipe.json",
        "normalization/suggested_recipe.json",
        "normalization/reviewed_decisions.json",
        "normalization/assurance_gates.json",
        "normalization/assurance_envelope.json",
        "normalization/qualification_review_payload.json",
        "sample/journal_sample.csv",
    }
)
JOURNAL_SAMPLING_IMPLEMENTATION_SPECS = (
    (
        "implementation",
        "scripts/check_dependencies.py",
        "implementation.journal_sampling_dependencies",
    ),
    (
        "implementation",
        "scripts/implementation_bootstrap.py",
        "implementation.journal_sampling_bootstrap",
    ),
    (
        "implementation",
        "scripts/inspect_journal.py",
        "implementation.journal_sampling_inspection_cli",
    ),
    (
        "implementation",
        "scripts/journal_sampling_core.py",
        "implementation.journal_sampling_core",
    ),
    (
        "implementation",
        "scripts/normalize_journal.py",
        "implementation.journal_sampling_normalization_cli",
    ),
    (
        "implementation",
        "scripts/replay_normalization.py",
        "implementation.journal_sampling_normalization_replay_cli",
    ),
    (
        "implementation",
        "scripts/review_session.py",
        "implementation.journal_sampling_review_session",
    ),
    (
        "implementation",
        "scripts/review_successor.py",
        "implementation.journal_sampling_review_successor",
    ),
    (
        "implementation",
        "scripts/run_sample.py",
        "implementation.journal_sampling_sample_cli",
    ),
    (
        "implementation",
        "mcp/server.cjs",
        "implementation.journal_sampling_mcp",
    ),
    (
        "implementation",
        "assets/icon.svg",
        "implementation.journal_sampling_icon",
    ),
    (
        "implementation",
        "assets/journal-sampling-review-widget.html",
        "implementation.journal_sampling_widget",
    ),
    (
        "implementation",
        "assets/review-workbench-adapter.json",
        "implementation.journal_sampling_widget_adapter",
    ),
    (
        "implementation",
        ".app.json",
        "implementation.journal_sampling_app_config",
    ),
    (
        "implementation",
        ".mcp.json",
        "implementation.journal_sampling_mcp_config",
    ),
    (
        "implementation",
        ".codex-plugin/plugin.json",
        "implementation.journal_sampling_plugin_config",
    ),
    (
        "assurance_implementation",
        "__init__.py",
        "implementation.vera_assurance_init",
    ),
    (
        "assurance_implementation",
        "contracts.py",
        "implementation.vera_assurance_contracts",
    ),
    (
        "assurance_implementation",
        "decisions.py",
        "implementation.vera_assurance_decisions",
    ),
    (
        "assurance_implementation",
        "envelope.py",
        "implementation.vera_assurance_envelope",
    ),
    (
        "assurance_implementation",
        "money.py",
        "implementation.vera_assurance_money",
    ),
    (
        "assurance_implementation",
        "relationships.py",
        "implementation.vera_assurance_relationships",
    ),
    (
        "assurance_implementation",
        "review_output_transaction.cjs",
        "implementation.vera_assurance_review_output_transaction",
    ),
    (
        "assurance_implementation",
        "serialization.py",
        "implementation.vera_assurance_serialization",
    ),
)
JOURNAL_SAMPLING_COLUMNS = [
    "entry_date",
    "movement_number",
    "line_number",
    "account",
    "account_desc",
    "line_desc",
    "debit",
    "credit",
    "amount_signed",
    "amount_abs",
    "currency",
    "unit",
    "reported_increment",
    "source_file",
    "source_sheet",
    "source_page",
    "source_row",
]
CANONICAL_ENTRY_COLUMNS = [
    "prepared_entry_id",
    "source_qualification_id",
    "movement_number",
    "line_number",
    "entry_date",
    "account",
    "account_desc",
    "description",
    "beneficiary_expected",
    "amount_signed",
    "amount_abs",
    "currency",
    "unit",
    "reported_increment",
    "source_file",
    "source_sheet",
    "source_page",
    "source_row",
]
RESULT_COLUMNS = [
    *CANONICAL_ENTRY_COLUMNS,
    "status",
    "matched_pdf",
    "checks_run",
    "mismatches",
    "review_notes",
    "amount_found",
    "date_found",
    "beneficiary_found",
    "matched_support",
    "support_type",
    "support_artifact_id",
    "support_match_status",
    "support_match_signals",
    "evidence_facts",
    "professional_conclusion",
    "assurance_gate_status",
    "support_amount_signed",
    "amount_difference_signed",
    "amount_difference_abs",
]
DATE_FORMATS = (
    "%Y-%m-%d",
    "%d/%m/%Y",
    "%d-%m-%Y",
    "%d.%m.%Y",
    "%d/%m/%y",
    "%m/%d/%Y",
    "%Y/%m/%d",
)
DATE_TOKEN_RE = re.compile(
    r"\b(\d{1,2}[./-]\d{1,2}[./-]\d{2,4}|\d{4}[./-]\d{1,2}[./-]\d{1,2})\b"
)
AMOUNT_TOKEN_RE = re.compile(
    r"(?<!\w)\(?-?\d{1,3}(?:[.,\s]\d{3})*(?:[.,]\d{2})\)?(?!\w)"
)
WORD_RE = re.compile(r"[a-z0-9]+")
ZERO = Decimal("0")
# ISO 4217 codes are a fixed mechanical vocabulary here: an explicit,
# conflicting code must never be overridden by a reviewed decision intended
# only to resolve symbol ambiguity.
EXPLICIT_CURRENCY_CODES = frozenset(
    {
        "AED",
        "AFN",
        "ALL",
        "AMD",
        "ANG",
        "AOA",
        "ARS",
        "AUD",
        "AWG",
        "AZN",
        "BAM",
        "BBD",
        "BDT",
        "BGN",
        "BHD",
        "BIF",
        "BMD",
        "BND",
        "BOB",
        "BOV",
        "BRL",
        "BSD",
        "BTN",
        "BWP",
        "BYN",
        "BZD",
        "CAD",
        "CDF",
        "CHE",
        "CHF",
        "CHW",
        "CLF",
        "CLP",
        "CNY",
        "COP",
        "COU",
        "CRC",
        "CUP",
        "CVE",
        "CZK",
        "DJF",
        "DKK",
        "DOP",
        "DZD",
        "EGP",
        "ERN",
        "ETB",
        "EUR",
        "FJD",
        "FKP",
        "GBP",
        "GEL",
        "GHS",
        "GIP",
        "GMD",
        "GNF",
        "GTQ",
        "GYD",
        "HKD",
        "HNL",
        "HTG",
        "HUF",
        "IDR",
        "ILS",
        "INR",
        "IQD",
        "IRR",
        "ISK",
        "JMD",
        "JOD",
        "JPY",
        "KES",
        "KGS",
        "KHR",
        "KMF",
        "KPW",
        "KRW",
        "KWD",
        "KYD",
        "KZT",
        "LAK",
        "LBP",
        "LKR",
        "LRD",
        "LSL",
        "LYD",
        "MAD",
        "MDL",
        "MGA",
        "MKD",
        "MMK",
        "MNT",
        "MOP",
        "MRU",
        "MUR",
        "MVR",
        "MWK",
        "MXN",
        "MXV",
        "MYR",
        "MZN",
        "NAD",
        "NGN",
        "NIO",
        "NOK",
        "NPR",
        "NZD",
        "OMR",
        "PAB",
        "PEN",
        "PGK",
        "PHP",
        "PKR",
        "PLN",
        "PYG",
        "QAR",
        "RON",
        "RSD",
        "RUB",
        "RWF",
        "SAR",
        "SBD",
        "SCR",
        "SDG",
        "SEK",
        "SGD",
        "SHP",
        "SLE",
        "SOS",
        "SRD",
        "SSP",
        "STN",
        "SVC",
        "SYP",
        "SZL",
        "THB",
        "TJS",
        "TMT",
        "TND",
        "TOP",
        "TRY",
        "TTD",
        "TWD",
        "TZS",
        "UAH",
        "UGX",
        "USD",
        "USN",
        "UYI",
        "UYU",
        "UYW",
        "UZS",
        "VES",
        "VND",
        "VUV",
        "WST",
        "XAF",
        "XAG",
        "XAU",
        "XBA",
        "XBB",
        "XBC",
        "XBD",
        "XCD",
        "XDR",
        "XOF",
        "XPD",
        "XPF",
        "XPT",
        "XSU",
        "XTS",
        "XUA",
        "XXX",
        "YER",
        "ZAR",
        "ZMW",
        "ZWG",
    }
)

__all__ = [
    "CANONICAL_ENTRY_COLUMNS",
    "RESULT_COLUMNS",
    "InspectionResult",
    "CheckRunResult",
    "add_common_args",
    "configure_logging",
    "inspect_entries",
    "load_client_engagement_context",
    "normalize_language",
    "run_entry_checks",
    "write_json",
]


@dataclass(frozen=True)
class InspectionResult:
    """Deterministic inspection output for journal entries and support files."""

    journal: dict[str, Any]
    pdfs: list[dict[str, Any]]
    invoices: list[dict[str, Any]]
    suggested_recipe: dict[str, Any]


@dataclass(frozen=True)
class CheckRunResult:
    """Check output plus reviewable audit metadata."""

    frame: pl.DataFrame
    audit: dict[str, Any]


@dataclass(frozen=True)
class SupportCapture:
    """One immutable source snapshot used for receipts and all parsing."""

    path: Path
    relative_path: str
    suffix: str
    payload: bytes
    receipt: dict[str, Any]


@dataclass(frozen=True)
class CapturedSupport:
    """Parsed facts and qualification records derived from captured bytes."""

    root: Path
    selection_path: Path
    selection_kind: str
    captures: tuple[SupportCapture, ...]
    manifest: dict[str, Any]
    pdfs: dict[str, dict[str, Any]]
    invoices: tuple[InvoiceRecord, ...]
    invoice_errors: tuple[dict[str, str], ...]
    invoice_artifact_ids: dict[str, str]
    source_qualifications: tuple[dict[str, Any], ...]


def configure_logging(verbose: bool = False) -> None:
    """Configure script logging without affecting imported use."""

    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(level=level, format="%(levelname)s: %(message)s")


def normalize_language(
    language: object | None,
    *,
    default: str = "en",
    allow_auto: bool = False,
) -> str:
    """Normalize a language tag to one supported plugin locale."""

    text = str(language or default).strip().lower().replace("_", "-")
    code = text.split("-", 1)[0]
    if allow_auto and code == "auto":
        return "auto"
    return code if code in SUPPORTED_LANGUAGES else default


def language_assumptions(
    recipe: dict[str, Any],
    *,
    language: object | None = None,
    document_language: object | None = None,
) -> dict[str, str]:
    """Resolve working and source-document language assumptions."""

    working = normalize_language(language or recipe.get("language"), default="en")
    source = normalize_language(
        document_language or recipe.get("document_language") or "auto",
        default=working,
        allow_auto=True,
    )
    return {"language": working, "document_language": source}


def write_json(path: Path, payload: Any) -> None:
    """Write JSON with stable formatting."""

    if not isinstance(payload, dict):
        raise ValueError("Check Entries JSON artifacts must be objects.")
    write_assurance_json(path, payload)


def read_json(path: Path | None) -> dict[str, Any]:
    """Return a JSON object or an empty mapping when no file is provided."""

    if path is None:
        return {}
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError(f"Recipe must be a JSON object: {path}")
    return payload


def load_client_engagement_context(path: Path) -> dict[str, Any]:
    """Load one exact client workflow context created by Studio Archive."""

    try:
        return load_client_engagement_context_file(
            path.expanduser().resolve(strict=True),
            expected_workflow_id="check-entries",
        )
    except (OSError, ValueError) as exc:
        raise ValueError(f"Client engagement context is invalid: {exc}") from exc


def _is_path_within(path: Path, parent: Path) -> bool:
    try:
        path.relative_to(parent)
    except ValueError:
        return False
    return True


def _managed_check_reference(
    path_value: Path,
    client_engagement: Mapping[str, Any] | None,
) -> str:
    """Return a run-relative reference for one managed Check Entries path."""

    if client_engagement is None:
        return path_value.as_posix()
    run_root_value = client_engagement.get("run_root")
    if not isinstance(run_root_value, str) or not run_root_value.strip():
        return path_value.as_posix()
    run_root = Path(run_root_value).expanduser().resolve()
    resolved = path_value.expanduser().resolve()
    try:
        relative = resolved.relative_to(run_root)
    except ValueError as exc:
        raise ValueError("Check Entries path is outside the run root.") from exc
    if not relative.parts or ".." in relative.parts:
        raise ValueError("Check Entries path must identify a run artifact.")
    return relative.as_posix()


def _portable_client_engagement_context(
    client_engagement: Mapping[str, Any] | None,
) -> dict[str, Any] | None:
    """Return the path-free identity persisted in managed workflow artifacts."""

    if (
        not isinstance(client_engagement, Mapping)
        or client_engagement.get("schema_version") != "vera.client_workflow_context.v2"
    ):
        return (
            dict(client_engagement) if isinstance(client_engagement, Mapping) else None
        )
    portable_fields = (
        "schema_version",
        "client_id",
        "engagement_id",
        "workflow_id",
        "workflow_version",
        "run_id",
        "label",
        "purpose",
        "created_at",
        "input_manifest",
        "input_manifest_sha256",
        "run_relative_path",
        "output_relative_path",
        "content_sha256",
    )
    return {field: client_engagement[field] for field in portable_fields}


def _bound_upstream_journal_execution(
    journal: Path,
    value: Mapping[str, Any] | None,
) -> Path:
    """Validate and return the run-local v2 normalized-journal copy.

    Both the upstream artifact and this run's execution copy are checked
    against the ledger identity.  Execution remains closed to the latter.
    """

    if value is None:
        return journal
    try:
        context = validate_client_engagement_context(value)
    except ValueError as exc:
        raise ValueError(f"Client engagement context is invalid: {exc}") from exc
    if context["schema_version"] != "vera.client_workflow_context.v2":
        return journal
    bindings = [
        item
        for item in context["input_bindings"]
        if item["kind"] == "upstream_artifact"
        and item["upstream_workflow_id"] == "journal-sampling"
        and item["upstream_artifact_id"] == "prepared.normalized_journal"
    ]
    if len(bindings) != 1:
        return journal
    binding = bindings[0]
    requested = journal.expanduser().resolve(strict=True)
    execution = Path(binding["path"]).resolve(strict=True)
    source = Path(binding["source_path"]).resolve(strict=True)
    if requested != execution:
        raise ValueError(
            "Check Entries must use the run-local normalized-journal input copy."
        )
    expected_bytes = binding["byte_count"]
    expected_sha256 = binding["sha256"]
    for candidate in (execution, source):
        payload = _stable_regular_bytes(
            candidate,
            label="Bound Journal Sampling normalized journal",
        )
        if (
            len(payload) != expected_bytes
            or hashlib.sha256(payload).hexdigest() != expected_sha256
        ):
            raise ValueError(
                "Bound Journal Sampling normalized journal no longer matches its "
                "ledger receipt."
            )
    return execution


def _validated_client_check_stage(
    value: Mapping[str, Any] | None,
    *,
    journal: Path,
    journal_diagnostics: Mapping[str, Any],
    support: Path,
    output_dir: Path,
    stage: str,
    enforce_output_path: bool,
) -> dict[str, Any] | None:
    """Close Check Entries to one upstream client engagement and support set."""

    if value is None:
        return None
    try:
        context = validate_client_engagement_context(value)
    except ValueError as exc:
        raise ValueError(f"Client engagement context is invalid: {exc}") from exc
    if context["workflow_id"] != "check-entries":
        raise ValueError("Client engagement is not for Check Entries.")
    upstream_value = journal_diagnostics.get("client_engagement")
    if not isinstance(upstream_value, Mapping):
        raise ValueError(
            "Normalized journal has no Journal Sampling client engagement."
        )
    try:
        upstream = validate_client_engagement_context(upstream_value)
    except ValueError as exc:
        raise ValueError(
            f"Normalized journal client engagement is invalid: {exc}"
        ) from exc
    if upstream["workflow_id"] != "journal-sampling":
        raise ValueError(
            "Normalized journal was not produced by a Journal Sampling engagement."
        )
    current_client_id = (
        context["client_id"]
        if context["schema_version"] == "vera.client_workflow_context.v2"
        else context["studio_client_folder"]["studio_client_id"]
    )
    upstream_client_id = (
        upstream["client_id"]
        if upstream["schema_version"] == "vera.client_workflow_context.v2"
        else upstream["studio_client_folder"]["studio_client_id"]
    )
    if (
        current_client_id != upstream_client_id
        or context["engagement_id"] != upstream["engagement_id"]
    ):
        raise ValueError(
            "Journal Sampling and Check Entries belong to different client engagements."
        )
    if context["schema_version"] == "vera.client_workflow_context.v2":
        journal_sampling_bindings = [
            item
            for item in context["input_bindings"]
            if item["kind"] == "upstream_artifact"
            and item["upstream_workflow_id"] == "journal-sampling"
        ]
        if not journal_sampling_bindings or {
            item["upstream_run_id"] for item in journal_sampling_bindings
        } != {upstream["run_id"]}:
            raise ValueError(
                "Check Entries requires Journal Sampling artifacts from one exact "
                "upstream run."
            )
        required_artifact_ids = {
            "prepared.normalized_journal",
            "internal.normalization_diagnostics",
            "prepared.journal_sample_csv",
        }
        binding_by_artifact_id = {
            item["upstream_artifact_id"]: item
            for item in journal_sampling_bindings
            if item["upstream_artifact_id"] in required_artifact_ids
        }
        if set(binding_by_artifact_id) != required_artifact_ids:
            raise ValueError(
                "Check Entries requires the exact normalized journal, normalization "
                "diagnostics, and sampled CSV from one Journal Sampling run."
            )
        upstream_output_root = (
            Path("Vera")
            / "engagements"
            / context["engagement_id"]
            / "runs"
            / upstream["run_id"]
            / "outputs"
        )
        try:
            handoff_paths = {
                Path(item["source_relative_path"])
                .relative_to(upstream_output_root)
                .as_posix()
                for item in journal_sampling_bindings
            }
        except ValueError as exc:
            raise ValueError(
                "Journal Sampling artifact paths do not match the upstream run."
            ) from exc
        if handoff_paths != JOURNAL_HANDOFF_ARTIFACT_PATHS:
            raise ValueError(
                "Check Entries requires the complete exact Journal Sampling "
                "normalization assurance handoff and sampled CSV."
            )
        journal_binding = binding_by_artifact_id["prepared.normalized_journal"]
        diagnostics_binding = binding_by_artifact_id[
            "internal.normalization_diagnostics"
        ]
        execution_journal = Path(journal_binding["path"]).resolve(strict=True)
        expected_journal = execution_journal
        diagnostics_value = journal_diagnostics.get("normalization_diagnostics")
        if not isinstance(diagnostics_value, str) or not diagnostics_value.strip():
            raise ValueError(
                "Normalized journal has no exact normalization diagnostics path."
            )
        expected_diagnostics = Path(diagnostics_binding["path"]).resolve(strict=True)
        if (
            Path(diagnostics_value).expanduser().resolve(strict=True)
            != expected_diagnostics
        ):
            raise ValueError(
                "Normalization diagnostics do not belong to the selected "
                "Journal Sampling run."
            )
        import_bindings = [
            item for item in context["input_bindings"] if item["kind"] == "import"
        ]
        if not import_bindings or any(
            item["role"] != "support" for item in import_bindings
        ):
            raise ValueError(
                "Check Entries evidence inputs must all be imported with role support."
            )
        support_path = support.expanduser().resolve(strict=True)
        selected_support_paths = {
            Path(item["path"]).resolve(strict=True) for item in import_bindings
        }
        if support_path.is_file():
            support_selection = {support_path}
        elif support_path.is_dir() and not support_path.is_symlink():
            support_selection = {
                path
                for path in selected_support_paths
                if path.is_relative_to(support_path)
            }
        else:
            raise ValueError(
                "Check Entries support must be a receipted file or closed folder."
            )
        if not support_selection or support_selection != selected_support_paths:
            raise ValueError(
                "Check Entries support selection must close over all and only this "
                "run's support receipts."
            )
        try:
            validate_client_workflow_run(
                context,
                expected_workflow_id="check-entries",
                input_paths=[execution_journal.parent, support],
                output_dir=output_dir if enforce_output_path else None,
            )
        except ValueError as exc:
            raise ValueError(f"Check Entries input binding is invalid: {exc}") from exc
    else:
        expected_journal = (
            Path(upstream["output_dir"]) / "normalization" / "normalized_journal.csv"
        ).resolve(strict=True)
    if journal.expanduser().resolve(strict=True) != expected_journal:
        raise ValueError(
            "Normalized journal does not belong to the selected client engagement."
        )
    support_path = support.expanduser().resolve(strict=True)
    if context["schema_version"] != "vera.client_workflow_context.v2":
        support_root = Path(context["input_dir"]).resolve(strict=True) / "support"
        if support_path != support_root and not _is_path_within(
            support_path, support_root
        ):
            raise ValueError(
                "Support is outside the selected client engagement support folder."
            )
    if stage not in {"inspection", "checks"}:
        raise ValueError("Unsupported Check Entries client workflow stage.")
    if enforce_output_path:
        expected_output = Path(context["output_dir"]) / stage
        if output_dir.expanduser().resolve() != expected_output.resolve():
            raise ValueError(
                "Check Entries output does not match the client engagement."
            )
    return context


def _bound_sample_entries(
    entries: pl.DataFrame,
    context: Mapping[str, Any] | None,
) -> pl.DataFrame:
    """Restrict v2 Check Entries work to the exact upstream sampled rows."""

    if (
        context is None
        or context.get("schema_version") != "vera.client_workflow_context.v2"
    ):
        return entries
    journal_sampling_bindings = [
        item
        for item in context["input_bindings"]
        if item["kind"] == "upstream_artifact"
        and item["upstream_workflow_id"] == "journal-sampling"
    ]
    run_ids = {item["upstream_run_id"] for item in journal_sampling_bindings}
    sample_bindings = [
        item
        for item in journal_sampling_bindings
        if item["upstream_artifact_id"] == "prepared.journal_sample_csv"
    ]
    if len(run_ids) != 1 or len(sample_bindings) != 1:
        raise ValueError("Check Entries has no exact Journal Sampling sample binding.")
    sample_binding = sample_bindings[0]
    locator_columns = ["source_file", "source_sheet", "source_page", "source_row"]
    try:
        normalized_locators = [
            pl.col(column).cast(pl.Utf8).fill_null("").alias(column)
            for column in locator_columns
        ]
        sampled = (
            pl.read_csv(Path(sample_binding["path"]), infer_schema=False)
            .select(locator_columns)
            .with_columns(normalized_locators)
        )
        indexed_entries = entries.with_row_index("_population_row_index")
        population_locators = indexed_entries.select(
            "_population_row_index", *locator_columns
        ).with_columns(normalized_locators)
        matched_indices = population_locators.join(
            sampled, on=locator_columns, how="semi"
        ).get_column("_population_row_index")
        filtered = indexed_entries.join(
            matched_indices.to_frame(), on="_population_row_index", how="semi"
        ).drop("_population_row_index")
        duplicate_count = sampled.height - sampled.unique().height
    except (OSError, pl.exceptions.PolarsError) as exc:
        raise ValueError(f"Bound Journal Sampling sample is invalid: {exc}") from exc
    if sampled.height == 0:
        raise ValueError("Bound Journal Sampling sample is empty.")
    if duplicate_count or filtered.height != sampled.height:
        raise ValueError(
            "Bound sample rows do not close exactly to the qualified population."
        )
    return filtered


def _captured_recipe(path: Path | None) -> tuple[dict[str, Any], bytes]:
    """Capture one exact ordinary recipe file, or the canonical empty recipe."""

    if path is None:
        return {}, b"{}\n"
    candidate = path.expanduser()
    observed_path = candidate.lstat()
    if (
        stat.S_ISLNK(observed_path.st_mode)
        or not stat.S_ISREG(observed_path.st_mode)
        or observed_path.st_nlink != 1
    ):
        raise ValueError("Check Entries recipe must be an ordinary single-link file.")
    flags = os.O_RDONLY | getattr(os, "O_NOFOLLOW", 0)
    descriptor = os.open(candidate, flags)
    try:
        before = os.fstat(descriptor)
        payload = bytearray()
        while True:
            chunk = os.read(descriptor, 1024 * 1024)
            if not chunk:
                break
            payload.extend(chunk)
        after = os.fstat(descriptor)
    finally:
        os.close(descriptor)
    identity = (
        before.st_dev,
        before.st_ino,
        before.st_mode,
        before.st_nlink,
        before.st_size,
        before.st_mtime_ns,
        before.st_ctime_ns,
    )
    if identity != (
        observed_path.st_dev,
        observed_path.st_ino,
        observed_path.st_mode,
        observed_path.st_nlink,
        observed_path.st_size,
        observed_path.st_mtime_ns,
        observed_path.st_ctime_ns,
    ) or identity != (
        after.st_dev,
        after.st_ino,
        after.st_mode,
        after.st_nlink,
        after.st_size,
        after.st_mtime_ns,
        after.st_ctime_ns,
    ):
        raise ValueError("Check Entries recipe changed while captured.")
    final_path = candidate.lstat()
    if identity != (
        final_path.st_dev,
        final_path.st_ino,
        final_path.st_mode,
        final_path.st_nlink,
        final_path.st_size,
        final_path.st_mtime_ns,
        final_path.st_ctime_ns,
    ):
        raise ValueError("Check Entries recipe path changed while captured.")
    if len(payload) != before.st_size:
        raise ValueError("Check Entries recipe size changed while captured.")
    try:
        value = json.loads(bytes(payload).decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ValueError(f"Recipe must be a UTF-8 JSON object: {candidate}") from exc
    if not isinstance(value, dict):
        raise ValueError(f"Recipe must be a JSON object: {candidate}")
    return value, bytes(payload)


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\u00a0", " ").replace("\u202f", " ").strip()


def _norm_label(value: Any) -> str:
    text = unicodedata.normalize("NFKD", _clean_text(value))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower()
    return re.sub(r"\s+", " ", text)


def _normalize_search_text(value: Any) -> str:
    text = _norm_label(value)
    return " ".join(WORD_RE.findall(text))


def _parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if not isinstance(value, str):
        return None
    text = _clean_text(value)
    if not text:
        return None
    token_match = DATE_TOKEN_RE.search(text)
    token = token_match.group(1) if token_match else text
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(token, fmt).date()
        except ValueError:
            continue
    return None


def _diagnostics_path(path: Path, recipe: dict[str, Any]) -> Path:
    journal_recipe = (
        recipe.get("journal") if isinstance(recipe.get("journal"), dict) else {}
    )
    configured = journal_recipe.get("normalization_diagnostics")
    if configured is None:
        return path.parent / "normalization_diagnostics.json"
    if not isinstance(configured, str) or not configured.strip():
        raise ValueError("journal.normalization_diagnostics must be a path string.")
    candidate = Path(configured).expanduser()
    return candidate if candidate.is_absolute() else path.parent / candidate


def _source_qualification_map(
    diagnostics: dict[str, Any],
) -> tuple[dict[str, str], list[dict[str, Any]]]:
    qualifications_raw = diagnostics.get("source_qualifications")
    if not isinstance(qualifications_raw, list) or not qualifications_raw:
        raise ValueError("Normalization diagnostics contain no source qualifications.")
    qualifications = [
        validate_source_qualification(value) for value in qualifications_raw
    ]
    if any(item["status"] != "qualified" for item in qualifications):
        raise ValueError("Every journal source must be qualified before entry checks.")

    files = diagnostics.get("files")
    if not isinstance(files, list) or not files:
        raise ValueError("Per-file source qualification diagnostics are required.")
    qualification_by_file: dict[str, str] = {}
    emitted_by_file: dict[str, int] = {}
    for index, item in enumerate(files):
        if not isinstance(item, dict):
            raise ValueError(f"File diagnostic {index + 1} must be an object.")
        source_file = item.get("source_file")
        if not isinstance(source_file, str) or not source_file.strip():
            raise ValueError(f"File diagnostic {index + 1} has no source_file.")
        if source_file in qualification_by_file:
            raise ValueError("Source file names must be unique within one population.")
        qualification = validate_source_qualification(item.get("qualification"))
        if qualification["status"] != "qualified":
            raise ValueError("Every source file must remain qualified.")
        qualification_by_file[source_file] = qualification["qualification_id"]
        emitted_by_file[source_file] = qualification["emitted_row_count"]

    top_ids = {item["qualification_id"] for item in qualifications}
    file_ids = set(qualification_by_file.values())
    if top_ids != file_ids:
        raise ValueError("Source qualification identities do not close by file.")
    diagnostics["_emitted_by_file"] = emitted_by_file
    return qualification_by_file, qualifications


def _adjacent_artifact_path(
    parent: Path,
    value: object,
    *,
    label: str,
) -> Path:
    if not isinstance(value, str) or not value.strip():
        raise ValueError(f"Normalization diagnostics are missing {label}.")
    relative = Path(value)
    if relative.is_absolute() or ".." in relative.parts:
        raise ValueError(f"{label} must stay inside the normalization directory.")
    resolved_parent = parent.resolve()
    resolved = (parent / relative).resolve()
    if not resolved.is_relative_to(resolved_parent) or not resolved.is_file():
        raise ValueError(f"{label} does not resolve to a local artifact.")
    return resolved


def _journal_sampling_component_root() -> Path:
    root = _COMPONENT_ROOT.parent / "journal-sampling"
    if not root.is_dir():
        raise ValueError(
            "Journal Sampling implementation is unavailable for assurance replay."
        )
    return root


def _validate_journal_sampling_implementation_tree(
    implementation_root: Path,
    assurance_root: Path,
) -> None:
    """Close the exact upstream implementation tree before receipt replay.

    Journal Sampling code is not executed here, but exact physical membership
    is required so a receipted file cannot hide behind a hardlink or coexist
    with an unowned executable/cache entry.
    """

    roots = {
        "implementation": implementation_root,
        "assurance_implementation": assurance_root,
    }
    expected_files = {
        (root_id, relative_path)
        for root_id, relative_path, _artifact_id in JOURNAL_SAMPLING_IMPLEMENTATION_SPECS
    }
    expected_directories: set[tuple[str, str]] = set()
    for root_id, relative_path in expected_files:
        parent = Path(relative_path).parent
        while parent != Path("."):
            expected_directories.add((root_id, parent.as_posix()))
            parent = parent.parent

    observed_files: set[tuple[str, str]] = set()
    observed_directories: set[tuple[str, str]] = set()

    def scan(root_id: str, scan_root: Path) -> None:
        root = roots[root_id]
        root_entry = scan_root.lstat()
        if stat.S_ISLNK(root_entry.st_mode) or not stat.S_ISDIR(root_entry.st_mode):
            raise ValueError("Journal Sampling implementation root is unsafe.")
        relative_root = scan_root.relative_to(root).as_posix()
        if relative_root != ".":
            observed_directories.add((root_id, relative_root))
        pending = [scan_root]
        while pending:
            current = pending.pop()
            with os.scandir(current) as iterator:
                entries = sorted(iterator, key=lambda entry: entry.name)
            for entry in entries:
                observed = entry.stat(follow_symlinks=False)
                relative = Path(entry.path).relative_to(root).as_posix()
                if stat.S_ISLNK(observed.st_mode):
                    raise ValueError(
                        "Journal Sampling implementation cannot contain symlinks."
                    )
                if stat.S_ISDIR(observed.st_mode):
                    observed_directories.add((root_id, relative))
                    pending.append(Path(entry.path))
                    continue
                if not stat.S_ISREG(observed.st_mode) or observed.st_nlink != 1:
                    raise ValueError(
                        "Journal Sampling implementation files must be ordinary "
                        "single-link files."
                    )
                observed_files.add((root_id, relative))

    for root_id, scan_root in (
        ("implementation", implementation_root / "assets"),
        ("implementation", implementation_root / "mcp"),
        ("implementation", implementation_root / "scripts"),
        ("implementation", implementation_root / ".codex-plugin"),
        ("assurance_implementation", assurance_root),
    ):
        scan(root_id, scan_root)
    for relative_path in (".app.json", ".mcp.json"):
        path = implementation_root / relative_path
        observed = path.lstat()
        if (
            stat.S_ISLNK(observed.st_mode)
            or not stat.S_ISREG(observed.st_mode)
            or observed.st_nlink != 1
        ):
            raise ValueError(
                "Journal Sampling launcher configuration must be an ordinary "
                "single-link file."
            )
        observed_files.add(("implementation", relative_path))
    if observed_files != expected_files or observed_directories != expected_directories:
        raise ValueError(
            "Journal Sampling implementation filesystem does not match its exact "
            "owned contract."
        )


def _upstream_artifact_receipt(
    parent: Path,
    path: Path,
    *,
    artifact_id: str,
) -> dict[str, Any]:
    return artifact_receipt(
        parent,
        path,
        artifact_id=artifact_id,
        root_id="normalization",
        role="source",
        media_type="application/json",
    )


_NORMALIZATION_REPLAY_FIELDS = (
    "schema_version",
    "path_reference",
    "input",
    "source_root",
    "source_receipts",
    "normalization_recipe_path",
    "normalization_recipe_receipt",
    "row_count",
    "language",
    "document_language",
    "files",
    "normalized_csv_receipt",
    "implementation_receipts",
    "reviewed_decisions",
    "assurance_gates",
    "assurance_envelope",
    "source_qualifications",
    "population_status",
    "qualification_review_payload",
)
_NORMALIZATION_REPLAY_RECEIPT_FIELDS = {
    "schema_version",
    "status",
    "normalized_csv_byte_count",
    "normalized_csv_sha256",
    "row_count",
    "normalization_content_sha256",
    "assurance_envelope_content_sha256",
    "recipe_sha256",
    "recipe_source_sha256",
    "source_receipt_set_sha256",
    "qualification_set_sha256",
    "implementation_receipt_set_sha256",
    "material_projection_sha256",
    "content_sha256",
}


def _stable_regular_bytes(
    path: Path,
    *,
    label: str,
    maximum_bytes: int | None = None,
) -> bytes:
    """Read one ordinary single-link file while proving stable identity."""

    observed = path.lstat()
    if (
        stat.S_ISLNK(observed.st_mode)
        or not stat.S_ISREG(observed.st_mode)
        or observed.st_nlink != 1
    ):
        raise ValueError(f"{label} must be an ordinary single-link file.")
    flags = os.O_RDONLY | getattr(os, "O_NOFOLLOW", 0)
    descriptor = os.open(path, flags)
    try:
        before = os.fstat(descriptor)
        payload = bytearray()
        while True:
            chunk = os.read(descriptor, 1024 * 1024)
            if not chunk:
                break
            payload.extend(chunk)
            if maximum_bytes is not None and len(payload) > maximum_bytes:
                raise ValueError(f"{label} exceeds the permitted size.")
        after = os.fstat(descriptor)
    finally:
        os.close(descriptor)
    identity = (
        before.st_dev,
        before.st_ino,
        before.st_mode,
        before.st_nlink,
        before.st_size,
        before.st_mtime_ns,
        before.st_ctime_ns,
    )
    if (
        identity
        != (
            observed.st_dev,
            observed.st_ino,
            observed.st_mode,
            observed.st_nlink,
            observed.st_size,
            observed.st_mtime_ns,
            observed.st_ctime_ns,
        )
        or identity
        != (
            after.st_dev,
            after.st_ino,
            after.st_mode,
            after.st_nlink,
            after.st_size,
            after.st_mtime_ns,
            after.st_ctime_ns,
        )
        or len(payload) != before.st_size
    ):
        raise ValueError(f"{label} changed while it was read.")
    final = path.lstat()
    if identity != (
        final.st_dev,
        final.st_ino,
        final.st_mode,
        final.st_nlink,
        final.st_size,
        final.st_mtime_ns,
        final.st_ctime_ns,
    ):
        raise ValueError(f"{label} path changed while it was read.")
    return bytes(payload)


def _validated_journal_sampling_replay_receipt(
    replay: object,
    diagnostics: dict[str, Any],
    envelope: dict[str, Any],
) -> dict[str, Any]:
    """Validate one exact replay receipt against already validated provenance."""

    if (
        not isinstance(replay, dict)
        or set(replay) != _NORMALIZATION_REPLAY_RECEIPT_FIELDS
    ):
        raise ValueError("Journal Sampling replay receipt fields are not exact.")
    recorded_digest = replay.get("content_sha256")
    replay_content = {
        key: value for key, value in replay.items() if key != "content_sha256"
    }
    if (
        replay.get("schema_version") != "journal_sampling.normalization_replay.v1"
        or replay.get("status") != "passed"
        or not isinstance(recorded_digest, str)
        or recorded_digest != canonical_json_sha256(replay_content)
    ):
        raise ValueError("Journal Sampling replay receipt is invalid or stale.")
    normalized_receipt = diagnostics["normalized_csv_receipt"]
    recipe_receipt = diagnostics["normalization_recipe_receipt"]
    recipe_source_receipt = diagnostics["normalization_recipe_source_receipt"]
    material_projection = {
        field: diagnostics.get(field) for field in _NORMALIZATION_REPLAY_FIELDS
    }
    expected = {
        "normalized_csv_byte_count": normalized_receipt["byte_count"],
        "normalized_csv_sha256": normalized_receipt["sha256"],
        "row_count": diagnostics["row_count"],
        "normalization_content_sha256": canonical_json_sha256(diagnostics),
        "assurance_envelope_content_sha256": envelope["content_sha256"],
        "recipe_sha256": recipe_receipt["sha256"],
        "recipe_source_sha256": recipe_source_receipt["sha256"],
        "source_receipt_set_sha256": canonical_json_sha256(
            {"receipts": diagnostics["source_receipts"]}
        ),
        "qualification_set_sha256": canonical_json_sha256(
            {"qualifications": diagnostics["source_qualifications"]}
        ),
        "implementation_receipt_set_sha256": canonical_json_sha256(
            {"receipts": diagnostics["implementation_receipts"]}
        ),
        "material_projection_sha256": canonical_json_sha256(material_projection),
    }
    if any(replay.get(key) != value for key, value in expected.items()):
        raise ValueError("Journal Sampling replay receipt does not close upstream.")
    return replay


def _current_upstream_journal_context(
    normalized_path: Path,
    diagnostics: Mapping[str, Any],
) -> dict[str, Any]:
    """Load the current upstream Journal Sampling context after a folder move."""

    persisted = diagnostics.get("client_engagement")
    if (
        not isinstance(persisted, Mapping)
        or persisted.get("schema_version") != "vera.client_workflow_context.v2"
    ):
        raise ValueError(
            "Journal Sampling replay requires its portable customer-run context."
        )
    engagement_id = persisted.get("engagement_id")
    run_id = persisted.get("run_id")
    if not isinstance(engagement_id, str) or not isinstance(run_id, str):
        raise ValueError("Journal Sampling portable context identity is invalid.")
    normalized_resolved = normalized_path.expanduser().resolve()
    candidates: list[Path] = []
    for candidate in (normalized_resolved, *normalized_resolved.parents):
        direct = candidate / "context.json"
        if direct.is_file() and not direct.is_symlink():
            candidates.append(direct)
        if candidate.name == "Vera":
            candidates.append(
                candidate
                / "engagements"
                / engagement_id
                / "runs"
                / run_id
                / "context.json"
            )
    for context_path in candidates:
        if not context_path.is_file() or context_path.is_symlink():
            continue
        try:
            current = load_client_engagement_context_file(
                context_path,
                expected_workflow_id="journal-sampling",
                allowed_statuses=("running", "ready_for_review", "completed"),
            )
        except (OSError, ValueError):
            continue
        if (
            current.get("engagement_id") == engagement_id
            and current.get("run_id") == run_id
            and current.get("client_id") == persisted.get("client_id")
        ):
            return current
    raise ValueError("Journal Sampling current customer-run context is unavailable.")


def _resolve_upstream_journal_reference(
    normalized_path: Path,
    diagnostics: Mapping[str, Any],
    value: object,
    *,
    label: str,
) -> Path:
    """Resolve a sealed Journal Sampling path through current context authority."""

    if not isinstance(value, str) or not value.strip() or value != value.strip():
        raise ValueError(f"{label} is unavailable.")
    reference = Path(value)
    if diagnostics.get("path_reference") != "run_root_relative":
        if not reference.is_absolute() or reference.resolve() != reference:
            raise ValueError(f"{label} is not canonical.")
        return reference
    if (
        reference.is_absolute()
        or ".." in reference.parts
        or reference.as_posix() != value
    ):
        raise ValueError(f"{label} leaves the Journal Sampling run.")
    context = _current_upstream_journal_context(normalized_path, diagnostics)
    run_root = Path(str(context["run_root"])).expanduser().resolve()
    resolved = (run_root / reference).resolve()
    if resolved == run_root or not _is_path_within(resolved, run_root):
        raise ValueError(f"{label} leaves the Journal Sampling run.")
    return resolved


def _run_journal_sampling_replay(
    normalized_path: Path,
    diagnostics: dict[str, Any],
    envelope: dict[str, Any],
) -> dict[str, Any]:
    """Run the receipted upstream replay CLI in an isolated interpreter."""

    component_root = _journal_sampling_component_root()
    replay_script = component_root / "scripts" / "replay_normalization.py"
    normalized_resolved = normalized_path.expanduser().resolve(strict=True)
    diagnostics_context = diagnostics.get("client_engagement")
    if (
        not isinstance(diagnostics_context, dict)
        or diagnostics_context.get("schema_version")
        != "vera.client_workflow_context.v2"
    ):
        raise ValueError(
            "Journal Sampling replay requires its portable customer-run context."
        )
    context_candidates: list[Path] = []
    explicit_context_path = diagnostics_context.get("context_path")
    if isinstance(explicit_context_path, str) and explicit_context_path.strip():
        context_candidates.append(Path(explicit_context_path).expanduser())
    output_root = next(
        (
            candidate
            for candidate in (normalized_resolved.parent, *normalized_resolved.parents)
            if candidate.name == "outputs"
            and (candidate.parent / "context.json").is_file()
        ),
        None,
    )
    if output_root is not None:
        context_candidates.append(output_root.parent / "context.json")
    vera_root = next(
        (
            candidate
            for candidate in normalized_resolved.parents
            if candidate.name == "Vera"
        ),
        None,
    )
    if vera_root is not None:
        context_candidates.append(
            vera_root
            / "engagements"
            / str(diagnostics_context["engagement_id"])
            / "runs"
            / str(diagnostics_context["run_id"])
            / "context.json"
        )
    client_engagement_path = next(
        (
            candidate.resolve(strict=True)
            for candidate in context_candidates
            if candidate.is_absolute() and candidate.is_file()
        ),
        None,
    )
    if client_engagement_path is None:
        raise ValueError("Journal Sampling customer-run context is unavailable.")
    upstream_output = client_engagement_path.parent / "outputs"
    replay_normalized_path = (
        upstream_output / "normalization" / "normalized_journal.csv"
    )
    replay_diagnostics_path = (
        upstream_output / "normalization" / "normalization_diagnostics.json"
    )
    with tempfile.TemporaryDirectory(
        prefix="check-entries-journal-replay-"
    ) as temporary_name:
        temporary_root = Path(temporary_name)
        receipt_path = temporary_root / "normalization_replay.json"
        try:
            # Exact receipted CLI argv; no shell or caller-controlled executable.
            completed = subprocess.run(  # nosec B603
                [
                    sys.executable,
                    "-I",
                    "-B",
                    str(replay_script),
                    str(replay_normalized_path),
                    "--diagnostics",
                    str(replay_diagnostics_path),
                    "--receipt-out",
                    str(receipt_path),
                    "--client-engagement",
                    str(client_engagement_path),
                    "--read-only-upstream",
                ],
                cwd=component_root,
                capture_output=True,
                text=True,
                check=False,
                timeout=120,
                env={
                    **os.environ,
                    "PYTHONDONTWRITEBYTECODE": "1",
                    "PYTHONPYCACHEPREFIX": str(temporary_root / "pycache"),
                },
            )
        except subprocess.TimeoutExpired as exc:
            raise ValueError(
                "Fresh Journal Sampling normalization replay timed out."
            ) from exc
        except OSError as exc:
            raise ValueError(
                "Fresh Journal Sampling normalization replay could not start."
            ) from exc
        if completed.returncode != 0 or not receipt_path.is_file():
            raise ValueError("Fresh Journal Sampling normalization replay failed.")
        payload_bytes = _stable_regular_bytes(
            receipt_path,
            label="Journal Sampling replay receipt",
            maximum_bytes=1024 * 1024,
        )
    try:
        replay = json.loads(payload_bytes.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ValueError("Journal Sampling replay receipt is not valid JSON.") from exc
    return _validated_journal_sampling_replay_receipt(
        replay,
        diagnostics,
        envelope,
    )


def _validate_upstream_assurance(
    normalized_path: Path,
    diagnostics: dict[str, Any],
    diagnostics_path: Path,
    expected_replay: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Replay the complete Journal Sampling preparation boundary."""

    normalization_root = normalized_path.parent
    source_root_value = diagnostics.get("source_root")
    if not isinstance(source_root_value, str) or not source_root_value.strip():
        raise ValueError("Normalization diagnostics are missing the source root.")
    source_root = _resolve_upstream_journal_reference(
        normalized_path,
        diagnostics,
        source_root_value,
        label="Journal Sampling source root",
    )
    envelope_path = _adjacent_artifact_path(
        normalization_root,
        diagnostics.get("assurance_envelope"),
        label="assurance_envelope",
    )
    gates_path = _adjacent_artifact_path(
        normalization_root,
        diagnostics.get("assurance_gates"),
        label="assurance_gates",
    )
    decisions_path = _adjacent_artifact_path(
        normalization_root,
        diagnostics.get("reviewed_decisions"),
        label="reviewed_decisions",
    )
    upstream_roots = {
        "source": source_root,
        "normalization": normalization_root,
        "implementation": _journal_sampling_component_root(),
        "assurance_implementation": implementation_artifact_roots()[
            "assurance_implementation"
        ],
    }
    _validate_journal_sampling_implementation_tree(
        upstream_roots["implementation"],
        upstream_roots["assurance_implementation"],
    )
    envelope = validate_assurance_envelope(
        read_json(envelope_path),
        artifact_roots=upstream_roots,
    )
    if (
        envelope["workflow_id"] != "journal-sampling-normalization"
        or envelope["workflow_version"] != "2"
    ):
        raise ValueError("Unexpected Journal Sampling assurance envelope identity.")
    recipe_path = _adjacent_artifact_path(
        normalization_root,
        diagnostics.get("normalization_recipe_path"),
        label="normalization_recipe_path",
    )
    recorded_normalization_root_value = diagnostics.get("normalization_recipe_root")
    if (
        recipe_path.name != "normalization_recipe.json"
        or not isinstance(recorded_normalization_root_value, str)
        or not recorded_normalization_root_value.strip()
    ):
        raise ValueError("Journal Sampling retained recipe path/root is not exact.")
    recorded_normalization_root = _resolve_upstream_journal_reference(
        normalized_path,
        diagnostics,
        recorded_normalization_root_value,
        label="Journal Sampling retained recipe root",
    )
    _stable_regular_bytes(
        recipe_path,
        label="Journal Sampling retained recipe",
        maximum_bytes=10 * 1024 * 1024,
    )
    recipe_receipt_raw = diagnostics.get("normalization_recipe_receipt")
    if not isinstance(recipe_receipt_raw, dict):
        raise ValueError("Journal Sampling retained recipe receipt is missing.")
    recipe_receipt = validate_artifact_receipt(
        {"normalization": normalization_root},
        recipe_receipt_raw,
    )
    if (
        recipe_receipt["artifact_id"] != "decision.normalization_recipe"
        or recipe_receipt["root_id"] != "normalization"
        or recipe_receipt["role"] != "reviewed_recipe"
        or recipe_receipt["path"] != "normalization_recipe.json"
        or recipe_receipt.get("media_type") != "application/json"
    ):
        raise ValueError("Journal Sampling retained recipe receipt is not exact.")
    envelope_recipe_receipts = [
        receipt
        for receipt in envelope["artifact_receipts"]
        if receipt["role"] == "reviewed_recipe"
    ]
    if envelope_recipe_receipts != [recipe_receipt]:
        raise ValueError(
            "Journal Sampling assurance does not bind one exact reviewed recipe."
        )
    recipe_source_value = diagnostics.get("normalization_recipe_source_path")
    recipe_source_receipt_raw = diagnostics.get("normalization_recipe_source_receipt")
    if (
        not isinstance(recipe_source_value, str)
        or not recipe_source_value.strip()
        or not isinstance(recipe_source_receipt_raw, dict)
    ):
        raise ValueError("Journal Sampling recipe source provenance is incomplete.")
    recorded_recipe_source_path = _resolve_upstream_journal_reference(
        normalized_path,
        diagnostics,
        recipe_source_value,
        label="Journal Sampling recipe source path",
    )
    try:
        recipe_source_relative = recorded_recipe_source_path.relative_to(
            recorded_normalization_root
        )
    except ValueError:
        recipe_source_path = recorded_recipe_source_path
    else:
        recipe_source_path = normalization_root / recipe_source_relative
    _stable_regular_bytes(
        recipe_source_path,
        label="Journal Sampling recipe source",
        maximum_bytes=10 * 1024 * 1024,
    )
    recipe_source_receipt = validate_artifact_receipt(
        {"normalization_recipe_source": recipe_source_path.parent},
        recipe_source_receipt_raw,
    )
    if (
        recipe_source_receipt["artifact_id"] != "decision.normalization_recipe_source"
        or recipe_source_receipt["root_id"] != "normalization_recipe_source"
        or recipe_source_receipt["role"] != "reviewed_recipe"
        or (recipe_source_path.parent / recipe_source_receipt["path"]).resolve()
        != recipe_source_path
        or recipe_source_receipt.get("media_type") != "application/json"
        or recipe_source_receipt["byte_count"] != recipe_receipt["byte_count"]
        or recipe_source_receipt["sha256"] != recipe_receipt["sha256"]
    ):
        raise ValueError(
            "Journal Sampling retained recipe does not close to its source."
        )

    gate_register = validate_gate_register(read_json(gates_path))
    if gate_register != envelope["gate_register"]:
        raise ValueError("Journal Sampling gates do not match the assurance envelope.")
    gates = gate_register["gates"]
    if (
        gates["source"]["status"] != "passed"
        or gates["preparation"]["status"] != "passed"
    ):
        raise ValueError("Journal Sampling source and preparation gates must pass.")

    decisions_payload = read_json(decisions_path)
    if (
        decisions_payload.get("schema_version")
        != "journal_sampling.reviewed_decisions.v1"
        or decisions_payload.get("decisions") != envelope["reviewed_decisions"]
    ):
        raise ValueError(
            "Journal Sampling reviewed decisions do not match the assurance envelope."
        )
    diagnostics_qualifications = diagnostics.get("source_qualifications")
    if diagnostics_qualifications != envelope["source_qualifications"]:
        raise ValueError(
            "Journal Sampling qualifications do not match the assurance envelope."
        )
    diagnostics_source_receipts = diagnostics.get("source_receipts")
    if not isinstance(diagnostics_source_receipts, list):
        raise ValueError("Normalization diagnostics are missing source receipts.")
    validated_source_receipts = [
        validate_artifact_receipt({"source": source_root}, receipt)
        for receipt in diagnostics_source_receipts
        if isinstance(receipt, dict)
    ]
    if len(validated_source_receipts) != len(diagnostics_source_receipts):
        raise ValueError("Normalization source receipts are malformed.")
    envelope_source_receipts = [
        receipt
        for receipt in envelope["artifact_receipts"]
        if receipt["role"] == "source"
    ]
    if validated_source_receipts != envelope_source_receipts:
        raise ValueError(
            "Journal Sampling source receipts do not match the assurance envelope."
        )
    normalized_receipt = diagnostics.get("normalized_csv_receipt")
    if not isinstance(normalized_receipt, dict):
        raise ValueError("Normalization diagnostics are missing the CSV receipt.")
    envelope_receipts = {
        receipt["artifact_id"]: receipt for receipt in envelope["artifact_receipts"]
    }
    if (
        envelope_receipts.get(normalized_receipt.get("artifact_id"))
        != normalized_receipt
    ):
        raise ValueError(
            "Normalized CSV receipt does not match the assurance envelope."
        )
    implementation_receipts = diagnostics.get("implementation_receipts")
    if not isinstance(implementation_receipts, list):
        raise ValueError(
            "Journal Sampling diagnostics are missing implementation receipts."
        )
    envelope_implementation_receipts = [
        receipt
        for receipt in envelope["artifact_receipts"]
        if receipt["role"] == "implementation"
    ]
    expected_implementation_metadata = [
        {
            "artifact_id": artifact_id,
            "root_id": root_id,
            "role": "implementation",
            "path": path,
            "media_type": {
                ".cjs": "text/javascript",
                ".html": "text/html",
                ".json": "application/json",
                ".py": "text/x-python",
                ".svg": "image/svg+xml",
            }[Path(path).suffix.lower()],
        }
        for root_id, path, artifact_id in JOURNAL_SAMPLING_IMPLEMENTATION_SPECS
    ]
    if (
        implementation_receipts != envelope_implementation_receipts
        or envelope["implementation_artifact_refs"]
        != [item["artifact_id"] for item in expected_implementation_metadata]
        or len(implementation_receipts) != len(expected_implementation_metadata)
        or any(
            {
                key: receipt.get(key)
                for key in ("artifact_id", "root_id", "role", "path", "media_type")
            }
            != expected
            for receipt, expected in zip(
                implementation_receipts,
                expected_implementation_metadata,
                strict=True,
            )
        )
    ):
        raise ValueError("Journal Sampling implementation receipt set is not exact.")
    normalization_replay = (
        _run_journal_sampling_replay(
            normalized_path,
            diagnostics,
            envelope,
        )
        if expected_replay is None
        else _validated_journal_sampling_replay_receipt(
            expected_replay,
            diagnostics,
            envelope,
        )
    )
    return {
        "envelope": envelope,
        "gate_register": gate_register,
        "source_receipts": validated_source_receipts,
        "normalization_recipe_receipt": recipe_receipt,
        "normalization_recipe_source_receipt": recipe_source_receipt,
        "normalization_replay": normalization_replay,
        "artifact_receipts": [
            _upstream_artifact_receipt(
                normalization_root,
                envelope_path,
                artifact_id="source.journal_sampling_assurance_envelope",
            ),
            _upstream_artifact_receipt(
                normalization_root,
                gates_path,
                artifact_id="source.journal_sampling_assurance_gates",
            ),
            _upstream_artifact_receipt(
                normalization_root,
                decisions_path,
                artifact_id="source.journal_sampling_reviewed_decisions",
            ),
        ],
    }


def _bytes_match_receipt(payload: bytes, receipt: dict[str, Any]) -> bool:
    return (
        receipt.get("byte_count") == len(payload)
        and receipt.get("sha256") == hashlib.sha256(payload).hexdigest()
    )


def _canonical_entry_frame(
    population: pl.DataFrame,
    *,
    qualification_by_file: dict[str, str],
) -> pl.DataFrame:
    records: list[dict[str, Any]] = []
    seen_entry_ids: set[str] = set()
    seen_source_locators: set[tuple[str, str, str, str]] = set()
    for row_index, row in enumerate(population.iter_rows(named=True), start=1):
        source_file = _clean_text(row.get("source_file"))
        if source_file not in qualification_by_file:
            raise ValueError(
                f"Row {row_index} does not close to a qualified source file."
            )
        if not _clean_text(row.get("account")):
            raise ValueError(f"Row {row_index} has no account.")
        parsed_date = _parse_date(row.get("entry_date"))
        if parsed_date is None:
            raise ValueError(f"Row {row_index} has no valid entry date.")
        debit = (
            ZERO
            if not _clean_text(row.get("debit"))
            else parse_canonical_decimal(row["debit"], label=f"row {row_index} debit")
        )
        credit = (
            ZERO
            if not _clean_text(row.get("credit"))
            else parse_canonical_decimal(row["credit"], label=f"row {row_index} credit")
        )
        signed = parse_canonical_decimal(
            row.get("amount_signed"), label=f"row {row_index} amount_signed"
        )
        absolute = parse_canonical_decimal(
            row.get("amount_abs"), label=f"row {row_index} amount_abs"
        )
        currency = _clean_text(row.get("currency"))
        unit = _clean_text(row.get("unit"))
        reported_increment = parse_canonical_decimal(
            row.get("reported_increment"),
            label=f"row {row_index} reported_increment",
        )
        if re.fullmatch(r"[A-Z]{3}", currency) is None or unit != "currency":
            raise ValueError(f"Row {row_index} has no valid currency/unit contract.")
        if reported_increment <= ZERO:
            raise ValueError(
                f"Row {row_index} reported increment must be strictly positive."
            )
        if signed != debit - credit or absolute != abs(signed):
            raise ValueError(f"Normalized monetary closure failed for row {row_index}.")
        if absolute == ZERO:
            raise ValueError(f"Zero-value row {row_index} cannot enter entry checks.")
        if absolute % reported_increment != ZERO:
            raise ValueError(
                f"Row {row_index} amount is not aligned to its reported increment."
            )
        source_sheet = _clean_text(row.get("source_sheet"))
        source_page = _clean_text(row.get("source_page"))
        source_row = _clean_text(row.get("source_row"))
        if not source_row.isdigit() or int(source_row) <= 0:
            raise ValueError(f"Row {row_index} has no positive physical source row.")
        if source_page and (not source_page.isdigit() or int(source_page) <= 0):
            raise ValueError(f"Row {row_index} has an invalid physical source page.")
        if (
            Path(source_file).suffix.lower() in {".xls", ".xlsx", ".xlsm"}
            and not source_sheet
        ):
            raise ValueError(f"Row {row_index} has no source worksheet locator.")
        source_locator = (source_file, source_sheet, source_page, source_row)
        if source_locator in seen_source_locators:
            raise ValueError(f"Row {row_index} duplicates a physical source locator.")
        seen_source_locators.add(source_locator)

        locator = {
            "source_qualification_id": qualification_by_file[source_file],
            "source_file": source_file,
            "source_sheet": source_sheet,
            "source_page": source_page,
            "source_row": source_row,
            "movement_number": _clean_text(row.get("movement_number")),
            "line_number": _clean_text(row.get("line_number")),
            "account": _clean_text(row.get("account")),
            "amount_signed": decimal_text(signed),
            "currency": currency,
            "unit": unit,
            "reported_increment": decimal_text(reported_increment),
        }
        prepared_entry_id = f"entry.{canonical_json_sha256(locator)}"
        if prepared_entry_id in seen_entry_ids:
            raise ValueError(
                f"Rows do not have unique stable prepared identities: {row_index}."
            )
        seen_entry_ids.add(prepared_entry_id)
        records.append(
            {
                "prepared_entry_id": prepared_entry_id,
                "source_qualification_id": qualification_by_file[source_file],
                "movement_number": _clean_text(row.get("movement_number")) or None,
                "line_number": _clean_text(row.get("line_number")) or None,
                "entry_date": parsed_date.isoformat(),
                "account": _clean_text(row.get("account")),
                "account_desc": _clean_text(row.get("account_desc")) or None,
                "description": _clean_text(row.get("line_desc")) or None,
                "beneficiary_expected": None,
                "amount_signed": decimal_text(signed),
                "amount_abs": decimal_text(absolute),
                "currency": currency,
                "unit": unit,
                "reported_increment": decimal_text(reported_increment),
                "source_file": source_file,
                "source_sheet": source_sheet or None,
                "source_page": source_page or None,
                "source_row": source_row,
            }
        )
    if not records:
        return pl.DataFrame(
            schema={column: pl.Utf8 for column in CANONICAL_ENTRY_COLUMNS}
        )
    return pl.DataFrame(records).select(CANONICAL_ENTRY_COLUMNS)


def _load_journal_entries(
    path: Path,
    recipe: dict[str, Any],
    *,
    source_bytes: bytes | None = None,
    expected_normalization_replay: dict[str, Any] | None = None,
) -> tuple[pl.DataFrame, dict[str, Any]]:
    path = path.expanduser().resolve()
    if path.suffix.lower() != ".csv":
        raise ValueError(
            "Check Entries accepts only Journal Sampling normalized_journal.csv "
            "with adjacent normalization_diagnostics.json. Normalize and qualify "
            "the source in Journal Sampling before running checks."
        )
    diagnostics_path = _diagnostics_path(path, recipe).expanduser().resolve()
    if not diagnostics_path.is_file():
        raise ValueError(
            "Check Entries requires adjacent normalization_diagnostics.json from "
            "Journal Sampling; raw journal preparation is not performed here."
        )
    captured_bytes = _stable_regular_bytes(
        path,
        label="Normalized Journal Sampling population",
    )
    if source_bytes is not None and captured_bytes != source_bytes:
        raise ValueError("Normalized journal changed after its initial capture.")
    diagnostics_bytes = _stable_regular_bytes(
        diagnostics_path,
        label="Journal Sampling normalization diagnostics",
        maximum_bytes=64 * 1024 * 1024,
    )
    try:
        raw_diagnostics = json.loads(diagnostics_bytes.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ValueError(
            "Journal Sampling normalization diagnostics are not valid JSON."
        ) from exc
    if not isinstance(raw_diagnostics, dict):
        raise ValueError(
            "Journal Sampling normalization diagnostics must be an object."
        )
    diagnostics = dict(raw_diagnostics)
    recorded_digest = diagnostics.pop("content_sha256", None)
    if not isinstance(recorded_digest, str) or recorded_digest != canonical_json_sha256(
        diagnostics
    ):
        raise ValueError("Normalization diagnostics content hash is stale.")
    if diagnostics.get("schema_version") != NORMALIZATION_SCHEMA_VERSION:
        raise ValueError("Check Entries requires journal_sampling.normalization.v2.")
    if diagnostics.get("population_status") != "complete":
        raise ValueError(
            "Entry checks are blocked because source preparation is not qualified."
        )
    receipt = diagnostics.get("normalized_csv_receipt")
    if not isinstance(receipt, dict):
        raise ValueError("Normalization diagnostics are missing the CSV receipt.")
    normalized_receipt = validate_artifact_receipt(
        {"normalization": path.parent}, receipt
    )
    if not _bytes_match_receipt(captured_bytes, normalized_receipt):
        raise ValueError(
            "Captured normalized CSV bytes do not match the preparation receipt."
        )
    upstream_assurance = _validate_upstream_assurance(
        path,
        diagnostics,
        diagnostics_path,
        expected_replay=expected_normalization_replay,
    )
    population = pl.read_csv(BytesIO(captured_bytes), infer_schema=False)
    if population.columns != JOURNAL_SAMPLING_COLUMNS:
        raise ValueError(
            "Prepared journal columns are not the ordered Journal Sampling v2 contract."
        )
    qualification_by_file, qualifications = _source_qualification_map(diagnostics)
    if diagnostics.get("row_count") != population.height or population.height <= 0:
        raise ValueError("Normalized population row count does not close.")

    actual_by_file: dict[str, int] = {}
    for value in population.get_column("source_file").to_list():
        source_file = _clean_text(value)
        actual_by_file[source_file] = actual_by_file.get(source_file, 0) + 1
    if actual_by_file != diagnostics.pop("_emitted_by_file"):
        raise ValueError(
            "Normalized rows do not close to qualified per-file populations."
        )
    frame = _canonical_entry_frame(
        population,
        qualification_by_file=qualification_by_file,
    )
    diagnostics_receipt = artifact_receipt(
        diagnostics_path.parent,
        diagnostics_path,
        artifact_id="source.normalization_diagnostics",
        root_id="normalization",
        role="source",
        media_type="application/json",
    )
    mapping = {
        "movement_number": "movement_number",
        "date": "entry_date",
        "description": "line_desc",
        "beneficiary": None,
        "amount": "amount_signed",
        "debit_amount": "debit",
        "credit_amount": "credit",
    }
    result = {
        "source_file": path.name,
        "source_preparation_status": "qualified",
        "normalization_schema_version": NORMALIZATION_SCHEMA_VERSION,
        "normalization_diagnostics": diagnostics_path.as_posix(),
        "normalization_content_sha256": recorded_digest,
        "normalized_csv_receipt": normalized_receipt,
        "normalization_diagnostics_receipt": diagnostics_receipt,
        "upstream_assurance": upstream_assurance,
        "qualification_ids": [
            qualification["qualification_id"] for qualification in qualifications
        ],
        "mapping": mapping,
        "mapping_status": "validated_canonical_contract",
        "raw_columns": list(population.columns),
        "row_count": frame.height,
        "preview": frame.head(20).to_dicts(),
        "missing_required_mapping": [],
    }
    if "client_engagement" in diagnostics:
        result["client_engagement"] = diagnostics["client_engagement"]
    return frame, result


def _missing_mapping(mapping: dict[str, Any]) -> list[str]:
    missing = []
    if not mapping.get("movement_number"):
        missing.append("movement_number")
    if not (
        mapping.get("amount")
        or (mapping.get("debit_amount") and mapping.get("credit_amount"))
    ):
        missing.append("amount or debit_amount/credit_amount")
    return missing


def supported_pdfs(pdf_path: Path) -> list[Path]:
    """Return supported PDF files from a file or folder path."""

    path = pdf_path.expanduser()
    if path.is_file():
        return [path] if path.suffix.lower() in PDF_SUFFIXES else []
    if not path.exists():
        raise FileNotFoundError(f"PDF path does not exist: {path}")
    return [
        candidate
        for candidate in sorted(path.rglob("*"))
        if candidate.is_file()
        and candidate.suffix.lower() in PDF_SUFFIXES
        and not candidate.name.startswith("~$")
    ]


def _extract_pdf_text(path: Path, payload: bytes) -> str:
    """Extract text from the already-captured PDF bytes.

    ``path`` is diagnostic context only. Parsing from ``payload`` prevents a
    live file replacement from changing the evidence after its receipt was
    computed.
    """

    import pdfplumber

    lines: list[str] = []
    with pdfplumber.open(BytesIO(payload)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            lines.extend(line.strip() for line in text.splitlines() if line.strip())
    return "\n".join(lines)


def _captured_receipt(
    *,
    root: Path,
    path: Path,
    relative_path: str,
    payload: bytes,
) -> dict[str, Any]:
    digest = hashlib.sha256(payload).hexdigest()
    artifact_id = (
        "support."
        + canonical_json_sha256(
            {
                "path": relative_path,
                "sha256": digest,
            }
        )[:24]
    )
    media_type = {
        ".pdf": "application/pdf",
        ".xml": "application/xml",
        ".p7m": "application/pkcs7-mime",
        ".zip": "application/zip",
    }.get(path.suffix.lower(), "application/octet-stream")
    receipt = {
        "schema_version": "vera.artifact_receipt.v1",
        "artifact_id": artifact_id,
        "root_id": "support",
        "role": "source",
        "path": relative_path,
        "byte_count": len(payload),
        "sha256": digest,
        "media_type": media_type,
    }
    # Replay against the live path once so canonical root/path and symlink
    # controls are shared with the cross-workflow assurance implementation.
    return validate_artifact_receipt({"support": root}, receipt)


def _capture_file(root: Path, path: Path) -> SupportCapture:
    if path.is_symlink():
        raise ValueError(f"Support artifacts cannot be symlinks: {path}")
    resolved_root = root.resolve()
    resolved = path.resolve()
    if not resolved.is_relative_to(resolved_root):
        raise ValueError("Support artifact escapes the support root.")
    relative_path = resolved.relative_to(resolved_root).as_posix()
    try:
        with resolved.open("rb") as handle:
            before = os.fstat(handle.fileno())
            if not stat.S_ISREG(before.st_mode):
                raise ValueError(f"Support artifact is not a regular file: {path}")
            if before.st_nlink != 1:
                raise ValueError(
                    f"Support artifacts cannot be hardlink aliases: {path}"
                )
            if before.st_size > MAX_SUPPORT_BYTES:
                raise ValueError(
                    f"Support artifact exceeds {MAX_SUPPORT_BYTES} bytes: {path}"
                )
            payload = handle.read(MAX_SUPPORT_BYTES + 1)
            after = os.fstat(handle.fileno())
    except OSError as exc:
        raise ValueError(f"Could not capture support artifact: {path}") from exc
    before_identity = (
        before.st_dev,
        before.st_ino,
        before.st_size,
        before.st_mtime_ns,
    )
    after_identity = (
        after.st_dev,
        after.st_ino,
        after.st_size,
        after.st_mtime_ns,
    )
    if (
        before_identity != after_identity
        or after.st_nlink != 1
        or len(payload) != after.st_size
        or len(payload) > MAX_SUPPORT_BYTES
    ):
        raise ValueError(f"Support artifact changed while captured: {path}")
    return SupportCapture(
        path=resolved,
        relative_path=relative_path,
        suffix=resolved.suffix.lower(),
        payload=payload,
        receipt=_captured_receipt(
            root=resolved_root,
            path=resolved,
            relative_path=relative_path,
            payload=payload,
        ),
    )


def _validate_unique_support_relative_paths(relative_paths: Sequence[str]) -> None:
    """Reject path aliases deterministically before payload capture."""

    canonical_paths = [
        unicodedata.normalize("NFC", relative_path).casefold()
        for relative_path in relative_paths
    ]
    if len(canonical_paths) != len(set(canonical_paths)):
        raise ValueError(
            "Support artifact paths are not unique after Unicode/casefold "
            "canonicalization."
        )


def _support_candidates(
    support_path: Path,
) -> tuple[Path, str, tuple[Path, ...]]:
    """Enumerate one canonical support selection without reading payload bytes."""

    path = support_path.expanduser()
    if not path.exists():
        raise FileNotFoundError(f"Support path does not exist: {path}")
    if path.is_symlink():
        raise ValueError("Support selection cannot be a symlink.")
    if path.is_dir():
        root = path.resolve()
        selection_kind = "directory"
        descendants = tuple(sorted(path.rglob("*")))
        if any(candidate.is_symlink() for candidate in descendants):
            raise ValueError("Support directories cannot contain symlinks.")
        if any(
            not (
                stat.S_ISDIR(candidate.lstat().st_mode)
                or stat.S_ISREG(candidate.lstat().st_mode)
            )
            for candidate in descendants
        ):
            raise ValueError(
                "Support directories cannot contain special filesystem entries."
            )
        candidates = tuple(
            candidate
            for candidate in descendants
            if stat.S_ISREG(candidate.lstat().st_mode)
        )
    else:
        root = path.parent.resolve()
        selection_kind = "file"
        candidates = (path,)
    relative_paths: list[str] = []
    for candidate in candidates:
        if candidate.is_symlink():
            raise ValueError(f"Support artifacts cannot be symlinks: {candidate}")
        candidate_stat = candidate.lstat()
        if not stat.S_ISREG(candidate_stat.st_mode):
            raise ValueError(f"Support artifact is not a regular file: {candidate}")
        if candidate_stat.st_nlink != 1:
            raise ValueError(
                f"Support artifacts cannot be hardlink aliases: {candidate}"
            )
        resolved = candidate.resolve()
        if not resolved.is_relative_to(root):
            raise ValueError("Support artifact escapes the support root.")
        relative_path = unicodedata.normalize(
            "NFC",
            resolved.relative_to(root).as_posix(),
        )
        if (
            not relative_path
            or relative_path.startswith("/")
            or ".." in Path(relative_path).parts
        ):
            raise ValueError("Support artifact has an unsafe relative path.")
        relative_paths.append(relative_path)
    _validate_unique_support_relative_paths(relative_paths)
    if relative_paths != sorted(relative_paths):
        candidates = tuple(
            candidate
            for _, candidate in sorted(
                zip(relative_paths, candidates, strict=True),
                key=lambda item: item[0],
            )
        )
    return root, selection_kind, candidates


def _support_manifest(
    *,
    selection_kind: str,
    captures: Sequence[SupportCapture],
) -> dict[str, Any]:
    """Seal canonical membership and receipts for final mechanical replay."""

    content = {
        "schema_version": "check_entries.support_manifest.v1",
        "selection_kind": selection_kind,
        "canonical_relative_paths": [capture.relative_path for capture in captures],
        "artifact_receipts": [capture.receipt for capture in captures],
    }
    return {**content, "content_sha256": canonical_json_sha256(content)}


def _capture_support(support_path: Path) -> tuple[Path, tuple[SupportCapture, ...]]:
    """Capture every relevant support artifact exactly once."""

    root, _, candidates = _support_candidates(support_path)
    captures = tuple(_capture_file(root, candidate) for candidate in candidates)
    return root, captures


def _zip_invoice_payloads(
    capture: SupportCapture,
) -> tuple[list[tuple[str, bytes]], list[dict[str, str]]]:
    payloads: list[tuple[str, bytes]] = []
    errors: list[dict[str, str]] = []
    try:
        with zipfile.ZipFile(BytesIO(capture.payload)) as archive:
            members = archive.infolist()
            if len(members) > MAX_ZIP_MEMBERS:
                raise ValueError(f"Invoice archive exceeds {MAX_ZIP_MEMBERS} members.")
            names: set[str] = set()
            total_uncompressed = 0
            for member in sorted(members, key=lambda value: value.filename):
                member_path = Path(member.filename)
                member_name = member_path.as_posix()
                canonical_name = member_name.casefold()
                if (
                    member_path.is_absolute()
                    or ".." in member_path.parts
                    or "\\" in member.filename
                    or canonical_name in names
                ):
                    raise ValueError(
                        f"Invoice archive member path is unsafe or duplicated: "
                        f"{member.filename}"
                    )
                names.add(canonical_name)
                if member.is_dir():
                    continue
                total_uncompressed += member.file_size
                if (
                    member.file_size > MAX_ZIP_MEMBER_BYTES
                    or total_uncompressed > MAX_ZIP_TOTAL_BYTES
                ):
                    raise ValueError("Invoice archive exceeds extraction budgets.")
                if member.file_size > MAX_ZIP_COMPRESSION_RATIO * max(
                    member.compress_size, 1
                ):
                    raise ValueError(
                        f"Invoice archive member compression ratio is unsafe: "
                        f"{member.filename}"
                    )
                suffix = member_path.suffix.lower()
                if suffix not in {".xml", ".p7m"}:
                    continue
                # Archive-relative names alone are lossy: separate ZIPs may
                # both contain ``invoice.xml``. The top-level captured artifact
                # path makes the locator globally unique and receipt-bound.
                source_name = f"{capture.relative_path}!/{member_name}"
                if member.flag_bits & 0x1:
                    errors.append(
                        {
                            "source_name": source_name,
                            "error": "Encrypted ZIP members are unsupported.",
                        }
                    )
                    continue
                if suffix == ".p7m":
                    errors.append(
                        {
                            "source_name": source_name,
                            "error": (
                                "P7M support is unsupported without a bounded "
                                "decoder and signature-validation policy."
                            ),
                        }
                    )
                    continue
                payloads.append((source_name, archive.read(member)))
    except (OSError, ValueError, zipfile.BadZipFile) as exc:
        errors.append(
            {
                "source_name": capture.relative_path,
                "error": str(exc),
            }
        )
    return payloads, errors


def _qualification_control(
    receipt: dict[str, Any],
    *,
    control_id: str,
    passed: bool,
    detail: str,
) -> dict[str, Any]:
    return {
        "control_id": control_id,
        "required": True,
        "status": "passed" if passed else "failed",
        "evidence_refs": [receipt["artifact_id"]],
        "detail": detail,
    }


def _support_qualification(
    capture: SupportCapture,
    *,
    source_family: str,
    adapter_id: str,
    candidate_count: int,
    emitted_count: int,
    passed: bool,
    control_id: str,
    detail: str,
    limitations: Sequence[str] = (),
) -> dict[str, Any]:
    return build_source_qualification(
        qualification_id=(
            f"qualification.{str(capture.receipt['artifact_id']).replace('.', '_')}"
        ),
        adapter_id=adapter_id,
        adapter_version="1",
        source_family=source_family,
        status="qualified" if passed else "unsupported_source_layout",
        source_artifact_refs=[capture.receipt["artifact_id"]],
        candidate_row_count=max(candidate_count, 1),
        emitted_row_count=emitted_count if passed else 0,
        reviewed_mapping_ref=None,
        controls=[
            _qualification_control(
                capture.receipt,
                control_id=control_id,
                passed=passed,
                detail=detail,
            )
        ],
        limitations=list(limitations),
    )


def _load_captured_support(
    support_path: Path,
) -> CapturedSupport:
    root, captures = _capture_support(support_path)
    selection_path = support_path.expanduser().resolve()
    selection_kind = "directory" if selection_path.is_dir() else "file"
    pdfs: dict[str, dict[str, Any]] = {}
    invoices: list[InvoiceRecord] = []
    invoice_errors: list[dict[str, str]] = []
    invoice_artifact_ids: dict[str, str] = {}
    invoice_locator_keys: set[str] = set()
    qualifications: list[dict[str, Any]] = []
    for capture in captures:
        artifact_id = str(capture.receipt["artifact_id"])
        if capture.suffix == ".pdf":
            text = ""
            error: str | None = None
            try:
                text = _extract_pdf_text(capture.path, capture.payload)
                if not text.strip():
                    error = "PDF contains no readable extracted text."
            except (OSError, ValueError, RuntimeError) as exc:
                error = str(exc)
            qualification = _support_qualification(
                capture,
                source_family="support.pdf",
                adapter_id="pdf_text",
                candidate_count=1,
                emitted_count=1 if error is None else 0,
                passed=error is None,
                control_id="readable_pdf_extraction",
                detail=(
                    "Captured PDF bytes produced non-empty readable text."
                    if error is None
                    else f"Captured PDF text extraction failed: {error}"
                ),
                limitations=[
                    "Readable extraction qualifies the PDF source layout; it does not establish entry identity."
                ],
            )
            qualifications.append(qualification)
            pdfs[capture.relative_path] = {
                "path": capture.path,
                "text": text if error is None else "",
                "text_norm": (_normalize_search_text(text) if error is None else ""),
                "error": error,
                "artifact_id": artifact_id,
                "source_qualification_id": qualification["qualification_id"],
                "capture_sha256": capture.receipt["sha256"],
            }
            continue

        if capture.suffix == ".p7m":
            error = (
                "P7M support is unsupported without a bounded decoder and "
                "signature-validation policy."
            )
            invoice_errors.append(
                {
                    "source_name": capture.relative_path,
                    "error": error,
                    "source_artifact_id": artifact_id,
                }
            )
            qualifications.append(
                _support_qualification(
                    capture,
                    source_family="support.p7m",
                    adapter_id="p7m_unsupported",
                    candidate_count=1,
                    emitted_count=0,
                    passed=False,
                    control_id="bounded_signature_decode",
                    detail=error,
                )
            )
            continue

        if capture.suffix in {".xml", ".zip"}:
            if capture.suffix == ".xml":
                payloads = [(capture.relative_path, capture.payload)]
                archive_errors: list[dict[str, str]] = []
                family = "support.fatturapa_xml"
                adapter_id = "fatturapa_xml"
            else:
                payloads, archive_errors = _zip_invoice_payloads(capture)
                family = "support.fatturapa_zip"
                adapter_id = "fatturapa_zip"
            parsed, parse_errors = load_invoice_payloads(payloads)
            errors = [*archive_errors, *parse_errors]
            for error in errors:
                invoice_errors.append(
                    {
                        **error,
                        "source_artifact_id": artifact_id,
                    }
                )
            passed = bool(payloads) and not errors and len(parsed) == len(payloads)
            qualification = _support_qualification(
                capture,
                source_family=family,
                adapter_id=adapter_id,
                candidate_count=len(payloads),
                emitted_count=len(parsed) if passed else 0,
                passed=passed,
                control_id="bounded_fatturapa_structure",
                detail=(
                    "Every captured FatturaPA XML passed bounded structural parsing."
                    if passed
                    else "One or more FatturaPA members failed bounded structural parsing."
                ),
                limitations=[
                    "Structural qualification does not establish the journal-entry relationship or party perimeter."
                ],
            )
            qualifications.append(qualification)
            if passed:
                locator_keys = [
                    unicodedata.normalize("NFC", record.source_name).casefold()
                    for record in parsed
                ]
                if len(locator_keys) != len(
                    set(locator_keys)
                ) or invoice_locator_keys.intersection(locator_keys):
                    raise ValueError(
                        "FatturaPA support locators are not globally unique."
                    )
                invoice_locator_keys.update(locator_keys)
                invoices.extend(parsed)
                for record in parsed:
                    invoice_artifact_ids[record.source_name] = artifact_id
            continue

        error = f"Unsupported support artifact type: {capture.suffix or '<none>'}"
        invoice_errors.append(
            {
                "source_name": capture.relative_path,
                "error": error,
                "source_artifact_id": artifact_id,
            }
        )
        qualifications.append(
            _support_qualification(
                capture,
                source_family="support.unsupported",
                adapter_id="unsupported_support",
                candidate_count=1,
                emitted_count=0,
                passed=False,
                control_id="supported_media_type",
                detail=error,
            )
        )
    return CapturedSupport(
        root=root,
        selection_path=selection_path,
        selection_kind=selection_kind,
        captures=captures,
        manifest=_support_manifest(
            selection_kind=selection_kind,
            captures=captures,
        ),
        pdfs=pdfs,
        invoices=tuple(invoices),
        invoice_errors=tuple(invoice_errors),
        invoice_artifact_ids=invoice_artifact_ids,
        source_qualifications=tuple(qualifications),
    )


def _validate_support_captures(captured: CapturedSupport) -> None:
    """Replay exact directory membership and every captured source receipt."""

    root, selection_kind, candidates = _support_candidates(captured.selection_path)
    if root != captured.root or selection_kind != captured.selection_kind:
        raise ValueError("Support selection identity changed during entry checks.")
    current_paths = tuple(
        unicodedata.normalize(
            "NFC",
            candidate.resolve().relative_to(root).as_posix(),
        )
        for candidate in candidates
    )
    expected_paths = tuple(captured.manifest["canonical_relative_paths"])
    if current_paths != expected_paths:
        raise ValueError("Support directory membership changed during entry checks.")
    for capture in captured.captures:
        validate_artifact_receipt({"support": captured.root}, capture.receipt)
    if (
        _support_manifest(
            selection_kind=captured.selection_kind,
            captures=captured.captures,
        )
        != captured.manifest
    ):
        raise ValueError("Captured support manifest is not reproducible.")


def _pdf_inventory_from_captured(
    captured: CapturedSupport,
) -> list[dict[str, Any]]:
    return [
        {
            "filename": filename,
            "path": str(payload["path"]),
            "support_artifact_id": payload.get("artifact_id"),
            "source_qualification_id": payload.get("source_qualification_id"),
            "capture_sha256": payload.get("capture_sha256"),
            "extractable_text": bool(payload["text"].strip()),
            "text_chars": len(str(payload["text"])),
            "error": payload["error"],
        }
        for filename, payload in captured.pdfs.items()
    ]


def _parse_tolerance(value: object) -> Decimal:
    try:
        parsed = parse_localized_decimal(value, label="amount_tolerance")
    except MoneyValidationError as exc:
        raise ValueError(str(exc)) from exc
    if parsed < ZERO:
        raise ValueError("amount_tolerance must not be negative.")
    return parsed


def _output_receipts(output_dir: Path, paths: Sequence[Path]) -> list[dict[str, Any]]:
    receipts: list[dict[str, Any]] = []
    for index, path in enumerate(paths, start=1):
        if not path.is_file():
            continue
        media_type = {
            ".csv": "text/csv",
            ".json": "application/json",
            ".md": "text/markdown",
            ".xlsx": (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        }.get(path.suffix.lower(), "application/octet-stream")
        receipts.append(
            artifact_receipt(
                output_dir,
                path,
                artifact_id=f"output.{index:04d}",
                root_id="run",
                role=(
                    "prepared"
                    if path.name
                    in {
                        "normalized_entries.csv",
                        "prepared_support_facts.csv",
                        "support_manifest.json",
                    }
                    else (
                        "workpaper"
                        if path.name
                        in {
                            "check_results.csv",
                            "check_results.xlsx",
                            "numeric_evidence_ledger.json",
                        }
                        else "output"
                    )
                ),
                media_type=media_type,
            )
        )
    return receipts


def _implementation_receipts() -> list[dict[str, Any]]:
    return build_implementation_receipts()


def _write_stable_result_workbook(frame: pl.DataFrame, path: Path) -> None:
    """Write a receiptable workbook after two-run OOXML equality."""

    def writer(candidate: Path) -> None:
        frame.write_excel(candidate)

    write_stable_xlsx(path, writer)


def _excel_column(index: int) -> str:
    letters = ""
    value = index
    while value > 0:
        value, remainder = divmod(value - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _receipt_by_path(
    receipts: Sequence[dict[str, Any]],
    path: str,
) -> dict[str, Any]:
    matches = [receipt for receipt in receipts if receipt.get("path") == path]
    if len(matches) != 1:
        raise ValueError(f"Expected one receipted artifact at {path}.")
    return matches[0]


def _support_currency_for_record(
    facts: dict[str, Any],
) -> str | None:
    """Return only a mechanically identified support currency."""

    invoice_currency = _clean_text(facts.get("invoice_currency")).upper()
    if invoice_currency:
        return invoice_currency
    explicit_codes = facts.get("currency_explicit_codes")
    if isinstance(explicit_codes, list) and len(explicit_codes) == 1:
        return _clean_text(explicit_codes[0]).upper() or None
    found = _clean_text(facts.get("currency_found")).upper()
    return found or None


def _prepared_support_facts(records: Sequence[dict[str, Any]]) -> pl.DataFrame:
    rows: list[dict[str, str]] = []
    for row in records:
        if (
            (
                row.get("support_amount_signed") in (None, "")
                and row.get("amount_found") in (None, "")
            )
            or not row.get("support_artifact_id")
            or not row.get("matched_support")
        ):
            continue
        facts = json.loads(str(row["evidence_facts"]))
        rows.append(
            {
                "prepared_entry_id": str(row["prepared_entry_id"]),
                "support_artifact_id": str(row["support_artifact_id"]),
                "support_locator": str(row["matched_support"]),
                "amount_found": str(row.get("amount_found") or ""),
                "support_amount_signed": str(row.get("support_amount_signed") or ""),
                "amount_difference_signed": str(
                    row.get("amount_difference_signed") or ""
                ),
                "amount_difference_abs": str(row.get("amount_difference_abs") or ""),
                "amount_source_locator": str(facts.get("amount_source_locator") or ""),
                "support_currency": str(_support_currency_for_record(facts) or ""),
            }
        )
    columns = [
        "prepared_entry_id",
        "support_artifact_id",
        "support_locator",
        "amount_found",
        "support_amount_signed",
        "amount_difference_signed",
        "amount_difference_abs",
        "amount_source_locator",
        "support_currency",
    ]
    if not rows:
        return pl.DataFrame(schema={column: pl.Utf8 for column in columns})
    return pl.DataFrame(rows).select(columns)


def _numeric_evidence_ledger(
    records: Sequence[dict[str, Any]],
    *,
    output_receipts: Sequence[dict[str, Any]],
) -> dict[str, Any]:
    """Bind exact journal, support, and delta values through native outputs."""

    prepared_entries = _receipt_by_path(output_receipts, "normalized_entries.csv")
    prepared_support = _receipt_by_path(output_receipts, "prepared_support_facts.csv")
    results_csv = _receipt_by_path(output_receipts, "check_results.csv")
    results_xlsx = _receipt_by_path(output_receipts, "check_results.xlsx")
    amount_signed_column = RESULT_COLUMNS.index("amount_signed") + 1
    amount_abs_column = RESULT_COLUMNS.index("amount_abs") + 1
    amount_found_column = RESULT_COLUMNS.index("amount_found") + 1
    support_amount_signed_column = RESULT_COLUMNS.index("support_amount_signed") + 1
    amount_difference_signed_column = (
        RESULT_COLUMNS.index("amount_difference_signed") + 1
    )
    amount_difference_abs_column = RESULT_COLUMNS.index("amount_difference_abs") + 1
    support_fact_row_by_entry = {
        str(row["prepared_entry_id"]): index
        for index, row in enumerate(
            (
                record
                for record in records
                if (
                    record.get("support_amount_signed") not in (None, "")
                    or record.get("amount_found") not in (None, "")
                )
                and record.get("support_artifact_id")
                and record.get("matched_support")
            ),
            start=2,
        )
    }
    entries: list[dict[str, Any]] = []
    for result_index, record in enumerate(records, start=2):
        prepared_entry_id = str(record["prepared_entry_id"])
        amount_signed = str(record["amount_signed"])
        signed_digest = canonical_json_sha256(
            {
                "prepared_entry_id": prepared_entry_id,
                "field": "amount_signed",
            }
        )
        entries.append(
            {
                "evidence_id": (f"numeric.check_entries.amount_signed.{signed_digest}"),
                "value": amount_signed,
                "unit": str(record["unit"]),
                "currency": str(record["currency"]),
                "source": {
                    "artifact_ref": NORMALIZED_JOURNAL_ARTIFACT_ID,
                    "locator": (
                        f"row={result_index};column=amount_signed;"
                        f"origin_source_file={record['source_file']};"
                        f"origin_source_row={record['source_row']}"
                    ),
                    "value": amount_signed,
                },
                "prepared": {
                    "artifact_ref": prepared_entries["artifact_id"],
                    "locator": f"row={result_index};column=amount_signed",
                    "value": amount_signed,
                },
                "outputs": [
                    {
                        "artifact_ref": results_csv["artifact_id"],
                        "locator": f"row={result_index};column=amount_signed",
                        "value": amount_signed,
                    },
                    {
                        "artifact_ref": results_xlsx["artifact_id"],
                        "locator": (
                            f"Sheet1!{_excel_column(amount_signed_column)}"
                            f"{result_index}"
                        ),
                        "value": amount_signed,
                    },
                ],
                "calculation_ref": None,
                "decision_ref": None,
                "limitations": [
                    "Exact signed transport is proven from the qualified journal row; accounting interpretation remains pending."
                ],
            }
        )
        amount_abs = str(record["amount_abs"])
        amount_digest = canonical_json_sha256(
            {
                "prepared_entry_id": prepared_entry_id,
                "field": "amount_abs",
            }
        )
        entries.append(
            {
                "evidence_id": f"numeric.check_entries.amount_abs.{amount_digest}",
                "value": amount_abs,
                "unit": str(record["unit"]),
                "currency": str(record["currency"]),
                "source": {
                    "artifact_ref": NORMALIZED_JOURNAL_ARTIFACT_ID,
                    "locator": (
                        f"row={result_index};column=amount_abs;"
                        f"origin_source_file={record['source_file']};"
                        f"origin_source_row={record['source_row']}"
                    ),
                    "value": amount_abs,
                },
                "prepared": {
                    "artifact_ref": prepared_entries["artifact_id"],
                    "locator": f"row={result_index};column=amount_abs",
                    "value": amount_abs,
                },
                "outputs": [
                    {
                        "artifact_ref": results_csv["artifact_id"],
                        "locator": f"row={result_index};column=amount_abs",
                        "value": amount_abs,
                    },
                    {
                        "artifact_ref": results_xlsx["artifact_id"],
                        "locator": (
                            f"Sheet1!{_excel_column(amount_abs_column)}{result_index}"
                        ),
                        "value": amount_abs,
                    },
                ],
                "calculation_ref": None,
                "decision_ref": None,
                "limitations": [
                    "Exact transport is proven from the qualified journal row; semantic support remains pending."
                ],
            }
        )
        amount_found = record.get("amount_found")
        if amount_found not in (None, ""):
            facts = json.loads(str(record["evidence_facts"]))
            support_row = support_fact_row_by_entry[prepared_entry_id]
            support_currency = _support_currency_for_record(facts)
            amount_found_text = str(amount_found)
            found_digest = canonical_json_sha256(
                {
                    "prepared_entry_id": prepared_entry_id,
                    "field": "amount_found",
                    "support_artifact_id": record["support_artifact_id"],
                    "support_locator": record["matched_support"],
                }
            )
            entries.append(
                {
                    "evidence_id": (
                        f"numeric.check_entries.amount_found.{found_digest}"
                    ),
                    "value": amount_found_text,
                    "unit": str(record["unit"]),
                    "currency": support_currency,
                    "source": {
                        "artifact_ref": str(record["support_artifact_id"]),
                        "locator": (
                            f"{record['matched_support']}::"
                            f"{facts['amount_source_locator']}"
                        ),
                        "value": amount_found_text,
                    },
                    "prepared": {
                        "artifact_ref": prepared_support["artifact_id"],
                        "locator": f"row={support_row};column=amount_found",
                        "value": amount_found_text,
                    },
                    "outputs": [
                        {
                            "artifact_ref": results_csv["artifact_id"],
                            "locator": (f"row={result_index};column=amount_found"),
                            "value": amount_found_text,
                        },
                        {
                            "artifact_ref": results_xlsx["artifact_id"],
                            "locator": (
                                f"Sheet1!{_excel_column(amount_found_column)}"
                                f"{result_index}"
                            ),
                            "value": amount_found_text,
                        },
                    ],
                    "calculation_ref": None,
                    "decision_ref": None,
                    "limitations": [
                        "The locator addresses deterministic extraction from captured support bytes; it does not establish party or accounting sufficiency."
                    ],
                }
            )
        support_amount_signed = record.get("support_amount_signed")
        if support_amount_signed in (None, ""):
            continue
        facts = json.loads(str(record["evidence_facts"]))
        support_row = support_fact_row_by_entry[prepared_entry_id]
        support_currency = _support_currency_for_record(facts)
        support_direction = str(facts.get("reviewed_support_direction") or "")
        direction_decision_ref = facts.get("direction_decision_ref")
        support_digest = canonical_json_sha256(
            {
                "prepared_entry_id": prepared_entry_id,
                "field": "support_amount_signed",
                "support_artifact_id": record["support_artifact_id"],
                "support_locator": record["matched_support"],
            }
        )
        entries.append(
            {
                "evidence_id": (
                    "numeric.check_entries.support_amount_signed." f"{support_digest}"
                ),
                "value": str(support_amount_signed),
                "unit": str(record["unit"]),
                "currency": support_currency,
                "source": {
                    "artifact_ref": str(record["support_artifact_id"]),
                    "locator": (
                        f"{record['matched_support']}::"
                        f"{facts['amount_source_locator']};"
                        f"reviewed_support_direction={support_direction};sign_applied"
                    ),
                    "value": str(support_amount_signed),
                },
                "prepared": {
                    "artifact_ref": prepared_support["artifact_id"],
                    "locator": f"row={support_row};column=support_amount_signed",
                    "value": str(support_amount_signed),
                },
                "outputs": [
                    {
                        "artifact_ref": results_csv["artifact_id"],
                        "locator": (f"row={result_index};column=support_amount_signed"),
                        "value": str(support_amount_signed),
                    },
                    {
                        "artifact_ref": results_xlsx["artifact_id"],
                        "locator": (
                            f"Sheet1!{_excel_column(support_amount_signed_column)}"
                            f"{result_index}"
                        ),
                        "value": str(support_amount_signed),
                    },
                ],
                "calculation_ref": ("calculation.support_reviewed_direction_sign.v1"),
                "decision_ref": direction_decision_ref,
                "limitations": [
                    "The signed value applies an exact source-bound reviewed direction to the support amount; document type alone cannot establish the journal side and professional sufficiency remains pending."
                ],
            }
        )
        for field, calculation_ref, column_index in (
            (
                "amount_difference_signed",
                "calculation.support_amount_minus_journal_amount.v1",
                amount_difference_signed_column,
            ),
            (
                "amount_difference_abs",
                "calculation.absolute_signed_amount_difference.v1",
                amount_difference_abs_column,
            ),
        ):
            raw_value = record.get(field)
            if raw_value in (None, ""):
                continue
            value = str(raw_value)
            difference_digest = canonical_json_sha256(
                {
                    "prepared_entry_id": prepared_entry_id,
                    "field": field,
                    "support_artifact_id": record["support_artifact_id"],
                }
            )
            entries.append(
                {
                    "evidence_id": (
                        f"numeric.check_entries.{field}.{difference_digest}"
                    ),
                    "value": value,
                    "unit": str(record["unit"]),
                    "currency": str(record["currency"]),
                    "source": {
                        "artifact_ref": str(record["support_artifact_id"]),
                        "locator": (
                            f"{record['matched_support']}::"
                            f"{facts['amount_source_locator']};"
                            f"reviewed_support_direction={support_direction};"
                            f"journal_amount_signed={amount_signed};field={field}"
                        ),
                        "value": value,
                    },
                    "prepared": {
                        "artifact_ref": prepared_support["artifact_id"],
                        "locator": f"row={support_row};column={field}",
                        "value": value,
                    },
                    "outputs": [
                        {
                            "artifact_ref": results_csv["artifact_id"],
                            "locator": f"row={result_index};column={field}",
                            "value": value,
                        },
                        {
                            "artifact_ref": results_xlsx["artifact_id"],
                            "locator": (
                                f"Sheet1!{_excel_column(column_index)}"
                                f"{result_index}"
                            ),
                            "value": value,
                        },
                    ],
                    "calculation_ref": calculation_ref,
                    "decision_ref": direction_decision_ref,
                    "limitations": [
                        "Exact Decimal arithmetic closes the difference; the mismatch still requires professional review."
                    ],
                }
            )
    return build_numeric_evidence_ledger(
        entries,
        ledger_id="numeric.check_entries_amounts",
    )


def _validate_numeric_artifact_addresses(
    records: Sequence[dict[str, Any]],
    *,
    normalized_path: Path,
    support_facts_path: Path,
    results_path: Path,
    xlsx_path: Path,
) -> None:
    """Re-read exact CSV/XLSX cells used in the numeric evidence ledger."""

    def rows(path: Path) -> list[dict[str, str]]:
        with path.open("r", encoding="utf-8", newline="") as handle:
            return list(csv.DictReader(handle))

    normalized_rows = rows(normalized_path)
    result_rows = rows(results_path)
    support_rows = rows(support_facts_path)
    if len(normalized_rows) != len(records) or len(result_rows) != len(records):
        raise ValueError("Numeric evidence row counts do not close.")
    support_by_entry = {row["prepared_entry_id"]: row for row in support_rows}
    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=False)
    if workbook.sheetnames != ["Sheet1"]:
        raise ValueError("Check Entries workbook sheet identity is unstable.")
    sheet = workbook["Sheet1"]
    headers = [str(cell.value or "") for cell in sheet[1]]
    if headers != RESULT_COLUMNS:
        raise ValueError("Check Entries workbook headers do not close.")
    for index, record in enumerate(records, start=2):
        row_index = index - 2
        amount_signed = str(record["amount_signed"])
        amount_abs = str(record["amount_abs"])
        if (
            normalized_rows[row_index]["amount_signed"] != amount_signed
            or result_rows[row_index]["amount_signed"] != amount_signed
            or str(sheet.cell(index, RESULT_COLUMNS.index("amount_signed") + 1).value)
            != amount_signed
        ):
            raise ValueError("amount_signed numeric evidence address does not close.")
        if (
            normalized_rows[row_index]["amount_abs"] != amount_abs
            or result_rows[row_index]["amount_abs"] != amount_abs
            or str(sheet.cell(index, RESULT_COLUMNS.index("amount_abs") + 1).value)
            != amount_abs
        ):
            raise ValueError("amount_abs numeric evidence address does not close.")
        support_amount_signed = record.get("support_amount_signed")
        if support_amount_signed not in (None, ""):
            support_amount_text = str(support_amount_signed)
            support_row = support_by_entry.get(str(record["prepared_entry_id"]))
            if support_row is None:
                raise ValueError("Prepared support numeric evidence row is missing.")
            if (
                support_row["support_amount_signed"] != support_amount_text
                or result_rows[row_index]["support_amount_signed"]
                != support_amount_text
                or str(
                    sheet.cell(
                        index,
                        RESULT_COLUMNS.index("support_amount_signed") + 1,
                    ).value
                )
                != support_amount_text
            ):
                raise ValueError(
                    "support_amount_signed numeric evidence address does not close."
                )
            raw_difference_signed = record.get("amount_difference_signed")
            raw_difference_abs = record.get("amount_difference_abs")
            if raw_difference_signed in (None, "") or raw_difference_abs in (None, ""):
                if not (
                    raw_difference_signed in (None, "")
                    and raw_difference_abs in (None, "")
                ):
                    raise ValueError("Amount difference fields must close together.")
            else:
                difference_signed = str(raw_difference_signed)
                difference_abs = str(raw_difference_abs)
                for field, value in (
                    ("amount_difference_signed", difference_signed),
                    ("amount_difference_abs", difference_abs),
                ):
                    if (
                        support_row[field] != value
                        or result_rows[row_index][field] != value
                        or str(
                            sheet.cell(
                                index,
                                RESULT_COLUMNS.index(field) + 1,
                            ).value
                        )
                        != value
                    ):
                        raise ValueError(
                            f"{field} numeric evidence address does not close."
                        )
                journal_value = parse_canonical_decimal(
                    amount_signed,
                    label="entry amount_signed",
                )
                support_value = parse_canonical_decimal(
                    support_amount_text,
                    label="support amount_signed",
                )
                signed_value = parse_canonical_decimal(
                    difference_signed,
                    label="amount difference_signed",
                )
                absolute_value = parse_canonical_decimal(
                    difference_abs,
                    label="amount difference_abs",
                )
                if (
                    support_value - journal_value != signed_value
                    or abs(signed_value) != absolute_value
                ):
                    raise ValueError("Signed support amount arithmetic does not close.")
        amount_found = record.get("amount_found")
        if amount_found in (None, ""):
            continue
        amount_found_text = str(amount_found)
        support_row = support_by_entry.get(str(record["prepared_entry_id"]))
        if (
            support_row is None
            or support_row["amount_found"] != amount_found_text
            or result_rows[row_index]["amount_found"] != amount_found_text
            or str(sheet.cell(index, RESULT_COLUMNS.index("amount_found") + 1).value)
            != amount_found_text
        ):
            raise ValueError("amount_found numeric evidence address does not close.")


def inspect_entries(
    journal: Path,
    pdf_path: Path,
    output_dir: Path,
    recipe_path: Path | None = None,
    *,
    language: object | None = None,
    document_language: object | None = None,
    client_engagement: Mapping[str, Any] | None = None,
) -> InspectionResult:
    """Inspect journal entries and PDF/XML support, then write Codex artifacts."""

    recipe, _recipe_source_bytes = _captured_recipe(recipe_path)
    languages = language_assumptions(
        recipe, language=language, document_language=document_language
    )
    bound_journal = _bound_upstream_journal_execution(journal, client_engagement)
    journal_bytes = bound_journal.read_bytes()
    frame, journal_diag = _load_journal_entries(
        bound_journal,
        recipe,
        source_bytes=journal_bytes,
    )
    normalized_client_engagement = _validated_client_check_stage(
        client_engagement,
        journal=bound_journal,
        journal_diagnostics=journal_diag,
        support=pdf_path,
        output_dir=output_dir,
        stage="inspection",
        enforce_output_path=True,
    )
    frame = _bound_sample_entries(frame, normalized_client_engagement)
    captured_support = _load_captured_support(pdf_path)
    _validate_support_captures(captured_support)
    pdfs = _pdf_inventory_from_captured(captured_support)
    invoices = list(captured_support.invoices)
    invoice_errors = list(captured_support.invoice_errors)
    suggested_recipe = {
        "version": 2,
        "description": "Qualified Check Entries evidence contract.",
        **languages,
        "journal": {
            "parser": "journal_sampling.normalization.v2",
            "normalization_diagnostics": Path(
                journal_diag["normalization_diagnostics"]
            ).name,
            "mapping": journal_diag["mapping"],
            "mapping_status": "validated_canonical_contract",
        },
        "pdf_matching": {
            "mode": "unique_labeled_movement_identifier",
            "allow_single_pdf_single_entry": False,
            "allow_filename_identity": False,
            "ambiguous_matches_require_review": True,
        },
        "acquisition_ladder": [
            "fatturapa_zip",
            "authorized_connector_export",
            "targeted_pdf_fallback",
        ],
        "xml_matching": {
            "mode": "unique_labeled_identity_plus_corroboration",
            "signals": ["invoice_number", "amount", "date"],
            "ambiguous_matches_require_review": True,
        },
        "reviewed_party_perimeters": [],
        "reviewed_support_relationships": [],
        "reviewed_currency_decisions": [],
        "reviewed_direction_decisions": [],
        "checks": {
            "amount_tolerance": "0",
            "date_window_days": 0,
            "party_perimeter": (
                "exact reviewed tax ID; structured exact name only under an "
                "explicit reviewed normalization contract"
            ),
            "beneficiary_match": "diagnostic only; free-text containment cannot promote",
            "professional_conclusion": "always_pending_review",
        },
    }
    output_dir.mkdir(parents=True, exist_ok=True)
    write_json(
        output_dir / "inspection.json",
        {
            **languages,
            **(
                {
                    "client_engagement": _portable_client_engagement_context(
                        normalized_client_engagement
                    )
                }
                if normalized_client_engagement is not None
                else {}
            ),
            "source_preparation_status": "qualified",
            "execution_eligibility": "eligible",
            "journal": journal_diag,
            "pdfs": pdfs,
            "invoices": [invoice.as_dict() for invoice in invoices],
            "invoice_errors": invoice_errors,
            "support_artifact_receipts": [
                capture.receipt for capture in captured_support.captures
            ],
            "support_manifest": captured_support.manifest,
            "support_source_qualifications": list(
                captured_support.source_qualifications
            ),
        },
    )
    write_json(output_dir / "suggested_recipe.json", suggested_recipe)
    return InspectionResult(
        journal=journal_diag,
        pdfs=pdfs,
        invoices=[invoice.as_dict() for invoice in invoices],
        suggested_recipe=suggested_recipe,
    )


def _movement_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", _norm_label(value))


def _movement_identifier_is_distinctive(value: object) -> bool:
    """Reject generic numeric/year/single-letter tokens as identity evidence."""

    parts = re.findall(r"[a-z0-9]+", _norm_label(value))
    compact = "".join(parts)
    if not compact:
        return False
    if compact.isdigit() or (len(compact) == 1 and compact.isalpha()):
        return False
    return any(character.isalpha() for character in compact)


def _contains_movement_identifier(value: object, movement_value: object) -> bool:
    if not _movement_identifier_is_distinctive(movement_value):
        return False
    movement = _movement_key(movement_value)
    if not movement:
        return False
    text = _norm_label(value)
    if movement in re.findall(r"[a-z0-9]+", text):
        return True
    parts = re.findall(r"[a-z0-9]+", _norm_label(movement_value))
    if not parts:
        return False
    pattern = (
        r"(?<![a-z0-9])"
        + r"[^a-z0-9]+".join(re.escape(part) for part in parts)
        + r"(?![a-z0-9])"
    )
    return re.search(pattern, text) is not None


def _contains_labeled_movement_identifier(
    value: object,
    movement_value: object,
) -> bool:
    """Require an accounting label around identifiers extracted from PDF text."""

    if not _movement_identifier_is_distinctive(movement_value):
        return False
    parts = re.findall(r"[a-z0-9]+", _norm_label(movement_value))
    if not parts:
        return False
    identifier = r"[^a-z0-9]+".join(re.escape(part) for part in parts)
    label = (
        r"(?:movement|movimento|registrazione|entry|posting|asiento|"
        r"ecriture|buchung)"
    )
    text = _norm_label(value)
    # A contextual label is mechanically auditable and avoids treating page,
    # row, year, or amount tokens as evidence of journal-entry identity.
    return (
        re.search(
            rf"(?<![a-z0-9]){label}(?:[^a-z0-9]+(?:n|no|nr|number))?"
            rf"[^a-z0-9]+{identifier}(?![a-z0-9])",
            text,
        )
        is not None
        or re.search(
            rf"(?<![a-z0-9]){identifier}[^a-z0-9]+{label}(?![a-z0-9])",
            text,
        )
        is not None
    )


def _reviewed_recipe_decisions(
    recipe: dict[str, Any],
    entries: pl.DataFrame,
    captured: CapturedSupport,
) -> tuple[
    dict[str, dict[str, Any]],
    dict[tuple[str, str, str], dict[str, Any]],
    dict[tuple[str, str, str], dict[str, Any]],
    dict[tuple[str, str, str], dict[str, Any]],
    list[dict[str, Any]],
]:
    """Validate source-bound reviewed party and relationship decisions."""

    entry_records = {str(row["prepared_entry_id"]): row for row in entries.to_dicts()}
    entry_ids = set(entry_records)
    support_ids = {str(capture.receipt["artifact_id"]) for capture in captured.captures}
    support_locators = {
        *captured.pdfs,
        *(invoice.source_name for invoice in captured.invoices),
    }
    locator_artifact_ids = {
        **{
            locator: str(payload["artifact_id"])
            for locator, payload in captured.pdfs.items()
        },
        **captured.invoice_artifact_ids,
    }
    party_by_entry: dict[str, dict[str, Any]] = {}
    relationship_by_key: dict[tuple[str, str, str], dict[str, Any]] = {}
    currency_by_key: dict[tuple[str, str, str], dict[str, Any]] = {}
    direction_by_key: dict[tuple[str, str, str], dict[str, Any]] = {}
    reviewed: list[dict[str, Any]] = []

    raw_party = recipe.get("reviewed_party_perimeters", [])
    if not isinstance(raw_party, list):
        raise ValueError("reviewed_party_perimeters must be a list.")
    for index, raw_decision in enumerate(raw_party):
        decision = validate_reviewed_decision_receipt(
            raw_decision,
            expected_decision_type="check_entries_party_perimeter",
            expected_source_artifact_refs=[NORMALIZED_JOURNAL_ARTIFACT_ID],
            expected_adapter_id=PARTY_ADAPTER_ID,
            expected_adapter_version=PARTY_ADAPTER_VERSION,
            require_reviewed=True,
        )
        content = decision["content"]
        expected_fields = {
            "prepared_entry_id",
            "expected_role",
            "expected_tax_ids",
            "expected_names",
            "name_normalization_contract",
        }
        if set(content) != expected_fields:
            raise ValueError(
                f"reviewed_party_perimeters[{index}] content fields are invalid."
            )
        prepared_entry_id = _clean_text(content["prepared_entry_id"])
        if prepared_entry_id not in entry_ids:
            raise ValueError("Reviewed party perimeter has a stale prepared entry.")
        if prepared_entry_id in party_by_entry:
            raise ValueError("Prepared entries may have only one party perimeter.")
        expected_role = _clean_text(content["expected_role"])
        if expected_role not in {"supplier", "customer", "either"}:
            raise ValueError("Reviewed party expected_role is unsupported.")
        raw_tax_ids = content["expected_tax_ids"]
        raw_names = content["expected_names"]
        if not isinstance(raw_tax_ids, list) or not isinstance(raw_names, list):
            raise ValueError("Reviewed party tax IDs and names must be lists.")
        tax_ids = [
            re.sub(r"[^A-Z0-9]+", "", _clean_text(value).upper())
            for value in raw_tax_ids
        ]
        names = [_clean_text(value) for value in raw_names]
        if (
            any(len(value) < 5 for value in tax_ids)
            or any(not value for value in names)
            or len(tax_ids) != len(set(tax_ids))
            or len(names) != len(set(names))
            or not (tax_ids or names)
        ):
            raise ValueError("Reviewed party perimeter values are invalid.")
        contract = content["name_normalization_contract"]
        if names:
            if contract != {"contract_id": "casefold_alnum_v1"}:
                raise ValueError(
                    "Reviewed party names require casefold_alnum_v1 normalization."
                )
        elif contract is not None:
            raise ValueError(
                "Name normalization contract requires reviewed expected names."
            )
        party_by_entry[prepared_entry_id] = {
            "decision_ref": decision["decision_id"],
            "expected_role": expected_role,
            "expected_tax_ids": tax_ids,
            "expected_names": names,
            "name_normalization_contract": contract,
        }
        reviewed.append(decision)

    raw_relationships = recipe.get("reviewed_support_relationships", [])
    if not isinstance(raw_relationships, list):
        raise ValueError("reviewed_support_relationships must be a list.")
    for index, raw_decision in enumerate(raw_relationships):
        if not isinstance(raw_decision, dict):
            raise ValueError(
                f"reviewed_support_relationships[{index}] must be an object."
            )
        raw_content = raw_decision.get("content")
        if not isinstance(raw_content, dict):
            raise ValueError("Reviewed support relationship content is missing.")
        artifact_id = _clean_text(raw_content.get("support_artifact_id"))
        decision = validate_reviewed_decision_receipt(
            raw_decision,
            expected_decision_type="check_entries_support_relationship",
            expected_source_artifact_refs=[
                NORMALIZED_JOURNAL_ARTIFACT_ID,
                artifact_id,
            ],
            expected_adapter_id=RELATIONSHIP_ADAPTER_ID,
            expected_adapter_version=RELATIONSHIP_ADAPTER_VERSION,
            require_reviewed=True,
        )
        content = decision["content"]
        expected_fields = {
            "prepared_entry_id",
            "support_artifact_id",
            "support_locator",
            "relationship_status",
            "recording_exception",
        }
        if set(content) != expected_fields:
            raise ValueError(
                f"reviewed_support_relationships[{index}] content fields are invalid."
            )
        prepared_entry_id = _clean_text(content["prepared_entry_id"])
        locator = _clean_text(content["support_locator"])
        recording_exception = _clean_text(content["recording_exception"])
        if (
            prepared_entry_id not in entry_ids
            or artifact_id not in support_ids
            or locator not in support_locators
            or locator_artifact_ids.get(locator) != artifact_id
            or content["relationship_status"] != "confirmed"
            or not recording_exception
        ):
            raise ValueError("Reviewed support relationship binding is stale.")
        key = (prepared_entry_id, artifact_id, locator)
        if key in relationship_by_key:
            raise ValueError("Reviewed support relationships must be unique.")
        if any(existing[0] == prepared_entry_id for existing in relationship_by_key):
            raise ValueError(
                "A prepared entry may have only one reviewed support relationship."
            )
        relationship_by_key[key] = {
            "decision_ref": decision["decision_id"],
            "recording_exception": recording_exception,
        }
        reviewed.append(decision)

    raw_currency = recipe.get("reviewed_currency_decisions", [])
    if not isinstance(raw_currency, list):
        raise ValueError("reviewed_currency_decisions must be a list.")
    for index, raw_decision in enumerate(raw_currency):
        if not isinstance(raw_decision, dict):
            raise ValueError(f"reviewed_currency_decisions[{index}] must be an object.")
        raw_content = raw_decision.get("content")
        if not isinstance(raw_content, dict):
            raise ValueError("Reviewed currency decision content is missing.")
        artifact_id = _clean_text(raw_content.get("support_artifact_id"))
        decision = validate_reviewed_decision_receipt(
            raw_decision,
            expected_decision_type="check_entries_currency",
            expected_source_artifact_refs=[
                NORMALIZED_JOURNAL_ARTIFACT_ID,
                artifact_id,
            ],
            expected_adapter_id=CURRENCY_ADAPTER_ID,
            expected_adapter_version=CURRENCY_ADAPTER_VERSION,
            require_reviewed=True,
        )
        content = decision["content"]
        expected_fields = {
            "prepared_entry_id",
            "support_artifact_id",
            "support_locator",
            "expected_currency",
            "currency_status",
            "recording_exception",
        }
        if set(content) != expected_fields:
            raise ValueError(
                f"reviewed_currency_decisions[{index}] content fields are invalid."
            )
        prepared_entry_id = _clean_text(content["prepared_entry_id"])
        locator = _clean_text(content["support_locator"])
        expected_currency = _clean_text(content["expected_currency"]).upper()
        recording_exception = _clean_text(content["recording_exception"])
        entry_currency = _clean_text(
            entry_records.get(prepared_entry_id, {}).get("currency")
        ).upper()
        if (
            prepared_entry_id not in entry_ids
            or locator not in captured.pdfs
            or locator_artifact_ids.get(locator) != artifact_id
            or expected_currency != entry_currency
            or re.fullmatch(r"[A-Z]{3}", expected_currency) is None
            or content["currency_status"] != "confirmed"
            or not recording_exception
        ):
            raise ValueError("Reviewed currency decision binding is stale.")
        key = (prepared_entry_id, artifact_id, locator)
        if key in currency_by_key:
            raise ValueError("Reviewed currency decisions must be unique.")
        currency_by_key[key] = {
            "decision_ref": decision["decision_id"],
            "expected_currency": expected_currency,
            "recording_exception": recording_exception,
        }
        reviewed.append(decision)

    raw_directions = recipe.get("reviewed_direction_decisions", [])
    if not isinstance(raw_directions, list):
        raise ValueError("reviewed_direction_decisions must be a list.")
    for index, raw_decision in enumerate(raw_directions):
        if not isinstance(raw_decision, dict):
            raise ValueError(
                f"reviewed_direction_decisions[{index}] must be an object."
            )
        raw_content = raw_decision.get("content")
        if not isinstance(raw_content, dict):
            raise ValueError("Reviewed direction decision content is missing.")
        artifact_id = _clean_text(raw_content.get("support_artifact_id"))
        decision = validate_reviewed_decision_receipt(
            raw_decision,
            expected_decision_type="check_entries_direction",
            expected_source_artifact_refs=[
                NORMALIZED_JOURNAL_ARTIFACT_ID,
                artifact_id,
            ],
            expected_adapter_id=DIRECTION_ADAPTER_ID,
            expected_adapter_version=DIRECTION_ADAPTER_VERSION,
            require_reviewed=True,
        )
        content = decision["content"]
        expected_fields = {
            "prepared_entry_id",
            "support_artifact_id",
            "support_locator",
            "expected_direction",
            "direction_status",
            "recording_exception",
        }
        if set(content) != expected_fields:
            raise ValueError(
                f"reviewed_direction_decisions[{index}] content fields are invalid."
            )
        prepared_entry_id = _clean_text(content["prepared_entry_id"])
        locator = _clean_text(content["support_locator"])
        expected_direction = _clean_text(content["expected_direction"])
        recording_exception = _clean_text(content["recording_exception"])
        entry_direction = _journal_direction(
            entry_records.get(prepared_entry_id, {}).get("amount_signed")
        )
        if (
            prepared_entry_id not in entry_ids
            or locator not in support_locators
            or locator_artifact_ids.get(locator) != artifact_id
            or expected_direction not in {"debit", "credit"}
            or expected_direction != entry_direction
            or content["direction_status"] != "confirmed"
            or not recording_exception
        ):
            raise ValueError("Reviewed direction decision binding is stale.")
        key = (prepared_entry_id, artifact_id, locator)
        if key in direction_by_key:
            raise ValueError("Reviewed direction decisions must be unique.")
        direction_by_key[key] = {
            "decision_ref": decision["decision_id"],
            "expected_direction": expected_direction,
            "recording_exception": recording_exception,
        }
        reviewed.append(decision)

    decision_ids = [str(decision["decision_id"]) for decision in reviewed]
    if len(decision_ids) != len(set(decision_ids)):
        raise ValueError("Reviewed Check Entries decision IDs must be unique.")
    return (
        party_by_entry,
        relationship_by_key,
        currency_by_key,
        direction_by_key,
        reviewed,
    )


def _relationship_for(
    entry: dict[str, Any],
    artifact_id: object,
    locator: object,
    relationships: dict[tuple[str, str, str], dict[str, Any]],
) -> dict[str, Any] | None:
    return relationships.get(
        (
            _clean_text(entry.get("prepared_entry_id")),
            _clean_text(artifact_id),
            _clean_text(locator),
        )
    )


def _normalized_party_name(value: object) -> str:
    return "".join(re.findall(r"[a-z0-9]+", _norm_label(value)))


def _structured_party_perimeter(
    invoice: InvoiceRecord,
    perimeter: dict[str, Any] | None,
) -> tuple[str, str | None]:
    """Compare reviewed exact party identifiers to structured XML fields."""

    if perimeter is None:
        return "missing", None
    role = str(perimeter["expected_role"])
    tax_ids = {
        re.sub(r"[^A-Z0-9]+", "", str(value).upper())
        for value in perimeter["expected_tax_ids"]
    }
    names = {_normalized_party_name(value) for value in perimeter["expected_names"]}
    candidates: list[tuple[str, str, str]] = []
    if role in {"supplier", "either"}:
        candidates.append(("supplier", invoice.supplier_tax_id, invoice.supplier_name))
    if role in {"customer", "either"}:
        candidates.append(("customer", invoice.customer_tax_id, invoice.customer_name))
    for candidate_role, tax_id, name in candidates:
        normalized_tax_id = re.sub(r"[^A-Z0-9]+", "", tax_id.upper())
        if tax_ids and normalized_tax_id in tax_ids:
            return "matched", f"{candidate_role}_tax_id"
        if names and _normalized_party_name(name) in names:
            return "matched", f"{candidate_role}_name"
    return "mismatch", None


def _pdf_party_perimeter(
    text: str,
    perimeter: dict[str, Any] | None,
) -> tuple[str, str | None]:
    """Require an exact tax ID coupled to the reviewed PDF party role.

    This is deterministic because it enforces a mechanically reviewable
    evidence contract; it does not infer a party role from document layout or
    free-text meaning.
    """

    if perimeter is None:
        return "missing", None
    tax_ids = [str(value) for value in perimeter["expected_tax_ids"]]
    normalized_text = _norm_label(text)
    tax_label = (
        r"(?:vat|vat id|tax id|partita iva|p iva|codice fiscale|"
        r"ust id|steuernummer|tva)"
    )
    role_patterns = {
        "supplier": (
            r"(?:supplier|vendor|seller|issuer|cedente(?: prestatore)?|"
            r"fornitore|fournisseur|prestataire|lieferant|anbieter|"
            r"proveedor|emisor)"
        ),
        "customer": (
            r"(?:customer|client|buyer|recipient|cessionario(?: committente)?|"
            r"committente|cliente|acquereur|destinataire|kunde|kaufer|"
            r"empfanger|comprador|receptor)"
        ),
    }
    expected_role = str(perimeter["expected_role"])
    for tax_id in tax_ids:
        token = r"[^a-z0-9]*".join(re.escape(character.lower()) for character in tax_id)
        matched_roles = {
            role
            for role, role_pattern in role_patterns.items()
            if re.search(
                (
                    rf"(?<![a-z0-9]){role_pattern}.{{0,32}}?"
                    rf"{tax_label}[^a-z0-9]+{token}(?![a-z0-9])"
                ),
                normalized_text,
            )
            or re.search(
                (
                    rf"(?<![a-z0-9]){tax_label}.{{0,32}}?"
                    rf"{role_pattern}[^a-z0-9]+{token}(?![a-z0-9])"
                ),
                normalized_text,
            )
        }
        if expected_role == "either" and matched_roles:
            role = sorted(matched_roles)[0]
            return "matched", f"{role}_labeled_tax_id"
        if expected_role in matched_roles:
            if len(matched_roles) == 1:
                return "matched", f"{expected_role}_labeled_tax_id"
            return "mismatch", "conflicting_role_labeled_tax_id"
        if matched_roles:
            return "mismatch", "opposite_role_labeled_tax_id"
        if re.search(
            rf"(?<![a-z0-9]){tax_label}[^a-z0-9]+{token}(?![a-z0-9])",
            normalized_text,
        ):
            if expected_role == "either":
                return "matched", "either_labeled_tax_id"
            return "missing", "tax_id_role_unresolved"
    # Free-text name containment is deliberately diagnostic only: text layout
    # does not prove which document party a name denotes.
    return "mismatch", None


def _match_pdf_for_entry(
    entry: dict[str, Any],
    pdfs: dict[str, dict[str, Any]],
    relationships: dict[tuple[str, str, str], dict[str, Any]],
) -> dict[str, Any]:
    movement_value = entry.get("movement_number")
    movement = _movement_key(movement_value)
    candidates: list[dict[str, Any]] = []
    for filename, payload in pdfs.items():
        signals: list[str] = []
        relationship = _relationship_for(
            entry,
            payload.get("artifact_id"),
            filename,
            relationships,
        )
        if relationship is not None:
            signals.append("reviewed_support_relationship")
        elif (
            movement
            and _movement_identifier_is_distinctive(movement_value)
            and _contains_labeled_movement_identifier(payload["text"], movement_value)
        ):
            signals.append("movement_identifier_in_text")
        if signals:
            candidates.append(
                {
                    "filename": filename,
                    "signals": signals,
                    "relationship": relationship,
                }
            )
    if len(candidates) == 1:
        return {"status": "matched", **candidates[0]}
    if len(candidates) > 1:
        return {
            "status": "ambiguous",
            "filename": None,
            "signals": [],
            "candidate_filenames": [
                str(candidate["filename"]) for candidate in candidates
            ],
        }
    return {"status": "missing", "filename": None, "signals": []}


def _amounts_in_text(text: str) -> list[Decimal]:
    values: list[Decimal] = []
    for match in AMOUNT_TOKEN_RE.finditer(text):
        try:
            values.append(
                parse_localized_decimal(
                    match.group(0),
                    label="support PDF monetary token",
                )
            )
        except MoneyValidationError:
            continue
    return values


def _dates_in_text(text: str) -> list[date]:
    values: list[date] = []
    for match in DATE_TOKEN_RE.finditer(text):
        parsed = _parse_date(match.group(1))
        if parsed is not None:
            values.append(parsed)
    return values


def _amount_found(
    expected: object,
    text: str,
    tolerance: Decimal,
) -> tuple[str | None, str | None]:
    if expected in (None, ""):
        return None, None
    expected_value = parse_canonical_decimal(expected, label="entry amount_abs")
    values = _amounts_in_text(text)
    for token_index, value in enumerate(values, start=1):
        _, within = difference_within_tolerance(
            abs(value), abs(expected_value), tolerance
        )
        if within:
            return decimal_text(value), f"extracted_text.amount_token[{token_index}]"
    return None, None


def _date_found(expected: str | None, text: str, window_days: int) -> str | None:
    parsed = _parse_date(expected)
    if parsed is None:
        return None
    for value in _dates_in_text(text):
        if abs((value - parsed).days) <= window_days:
            return value.isoformat()
    return None


def _beneficiary_found(expected: str | None, text_norm: str) -> str | None:
    expected_norm = _normalize_search_text(expected)
    if not expected_norm:
        return None
    if expected_norm in text_norm:
        return expected
    tokens = [token for token in expected_norm.split() if len(token) > 2]
    if tokens and all(token in text_norm for token in tokens):
        return expected
    return None


def _journal_direction(value: object) -> str | None:
    if value in (None, ""):
        return None
    parsed = parse_canonical_decimal(value, label="entry amount_signed")
    if parsed == ZERO:
        return None
    return "debit" if parsed > ZERO else "credit"


def _support_amount_comparison(
    journal_amount_signed: object,
    support_amount: object,
    reviewed_support_direction: str | None,
    *,
    comparable_currency: bool,
) -> dict[str, str | None]:
    """Return signed support and deltas only with a reviewed direction.

    Decimal arithmetic is mechanical, but choosing the journal side is a
    source-bound reviewer judgment. Debit is positive, credit is negative, and
    the signed delta is support minus journal once that decision exists.
    """

    empty = {
        "support_amount_signed": None,
        "amount_difference_signed": None,
        "amount_difference_abs": None,
    }
    if (
        journal_amount_signed in (None, "")
        or support_amount in (None, "")
        or reviewed_support_direction not in {"debit", "credit"}
    ):
        return empty
    journal_value = parse_canonical_decimal(
        journal_amount_signed,
        label="entry amount_signed",
    )
    support_magnitude = abs(
        parse_canonical_decimal(
            support_amount,
            label="support amount",
        )
    )
    support_signed = (
        support_magnitude
        if reviewed_support_direction == "debit"
        else -support_magnitude
    )
    if not comparable_currency:
        return {
            **empty,
            "support_amount_signed": decimal_text(support_signed),
        }
    difference_signed = support_signed - journal_value
    return {
        "support_amount_signed": decimal_text(support_signed),
        "amount_difference_signed": decimal_text(difference_signed),
        "amount_difference_abs": decimal_text(abs(difference_signed)),
    }


def _currency_found(expected: object, text: str) -> str | None:
    """Require the exact ISO code; symbols such as ``$`` are ambiguous."""

    currency = _clean_text(expected).upper()
    if re.fullmatch(r"[A-Z]{3}", currency) is None:
        return None
    if re.search(rf"(?<![A-Z]){re.escape(currency)}(?![A-Z])", text, re.IGNORECASE):
        return currency
    return None


def _explicit_currency_codes(text: str) -> tuple[str, ...]:
    """Return isolated uppercase ISO 4217 labels in stable order."""

    labels = {
        match.group(1)
        for match in re.finditer(
            r"(?<![A-Za-z])([A-Z]{3})(?![A-Za-z])",
            text,
        )
        if match.group(1) in EXPLICIT_CURRENCY_CODES
    }
    return tuple(sorted(labels))


def _explicit_currency_conflict(expected: object, text: str) -> bool:
    expected_currency = _clean_text(expected).upper()
    return bool(set(_explicit_currency_codes(text)) - {expected_currency})


def _pdf_document_polarity(text: str) -> str | None:
    """Return an explicit document polarity as diagnostic source evidence."""

    normalized = _norm_label(text)
    credit = re.search(
        r"(?<![a-z0-9])(?:credit note|nota di credito|note de credit|"
        r"avoir|gutschrift|nota de credito)(?![a-z0-9])",
        normalized,
    )
    debit = re.search(
        r"(?<![a-z0-9])(?:invoice|fattura|facture|rechnung|factura|"
        r"debit note|nota di debito)(?![a-z0-9])",
        normalized,
    )
    if bool(credit) == bool(debit):
        return None
    return "negative_document" if credit else "positive_document"


def _check_one_entry(
    entry: dict[str, Any],
    pdfs: dict[str, dict[str, Any]],
    *,
    amount_tolerance: Decimal,
    date_window_days: int,
    party_perimeter: dict[str, Any] | None,
    relationships: dict[tuple[str, str, str], dict[str, Any]],
    currency_decisions: dict[tuple[str, str, str], dict[str, Any]],
    direction_decisions: dict[tuple[str, str, str], dict[str, Any]],
    language: str = "en",
) -> dict[str, Any]:
    match = _match_pdf_for_entry(entry, pdfs, relationships)
    if match["status"] != "matched":
        ambiguous = match["status"] == "ambiguous"
        evidence_facts = {
            "support_match_status": match["status"],
            "support_match_signals": [],
            "candidate_filenames": match.get("candidate_filenames", []),
        }
        return {
            **entry,
            "status": "manual_review" if ambiguous else "missing_support",
            "matched_pdf": None,
            "checks_run": "",
            "mismatches": ("ambiguous_pdf_support" if ambiguous else "support_pdf"),
            "review_notes": (
                (
                    "Varios PDF contienen el identificador explícito del movimiento; se requiere selección de la persona revisora."
                    if ambiguous
                    else "Ningún PDF justificativo contiene el identificador explícito del movimiento."
                )
                if language == "es"
                else (
                    "Multiple PDFs contain the explicit movement identifier; reviewer selection is required."
                    if ambiguous
                    else "No supporting PDF contains the explicit movement identifier."
                )
            ),
            "amount_found": None,
            "date_found": None,
            "beneficiary_found": None,
            "matched_support": None,
            "support_type": None,
            "support_artifact_id": None,
            "support_match_status": match["status"],
            "support_match_signals": "",
            "evidence_facts": json.dumps(
                evidence_facts,
                ensure_ascii=False,
                sort_keys=True,
                separators=(",", ":"),
            ),
            "professional_conclusion": "pending_review",
            "assurance_gate_status": ("reconciliation_failed;semantic_review_withheld"),
            "support_amount_signed": None,
            "amount_difference_signed": None,
            "amount_difference_abs": None,
        }

    matched_pdf = str(match["filename"])
    pdf_payload = pdfs[matched_pdf]
    text = str(pdf_payload["text"])
    text_norm = str(pdf_payload["text_norm"])
    checks_run: list[str] = []
    mismatches: list[str] = []
    review_notes: list[str] = []

    expected_amount = entry.get("amount_abs")
    amount_value, amount_locator = _amount_found(
        expected_amount,
        text,
        amount_tolerance,
    )
    if expected_amount not in (None, ""):
        checks_run.append("amount")
        if amount_value is None:
            mismatches.append("amount")

    expected_date = entry.get("entry_date")
    date_value = _date_found(
        str(expected_date) if expected_date else None, text, date_window_days
    )
    if expected_date:
        checks_run.append("date")
        if date_value is None:
            mismatches.append("date")

    expected_beneficiary = entry.get("beneficiary_expected")
    beneficiary_value = _beneficiary_found(
        str(expected_beneficiary) if expected_beneficiary else None, text_norm
    )
    if expected_beneficiary:
        checks_run.append("beneficiary")
        if beneficiary_value is None:
            mismatches.append("beneficiary")

    expected_currency = entry.get("currency")
    currency_decision = _relationship_for(
        entry,
        pdf_payload.get("artifact_id"),
        matched_pdf,
        currency_decisions,
    )
    currency_value = _currency_found(expected_currency, text)
    explicit_currency_codes = _explicit_currency_codes(text)
    currency_conflict = _explicit_currency_conflict(expected_currency, text)
    if currency_conflict:
        currency_value = None
    if (
        currency_value is None
        and currency_decision is not None
        and not currency_conflict
    ):
        currency_value = _clean_text(expected_currency).upper()
    if expected_currency not in (None, ""):
        checks_run.append("currency")
        if currency_value is None:
            mismatches.append("currency")

    expected_direction = _journal_direction(entry.get("amount_signed"))
    document_polarity = _pdf_document_polarity(text)
    direction_decision = _relationship_for(
        entry,
        pdf_payload.get("artifact_id"),
        matched_pdf,
        direction_decisions,
    )
    reviewed_support_direction = (
        str(direction_decision["expected_direction"])
        if direction_decision is not None
        else None
    )
    amount_comparison = _support_amount_comparison(
        entry.get("amount_signed"),
        amount_value,
        reviewed_support_direction,
        comparable_currency=currency_value is not None,
    )
    direction_missing = expected_direction is None or reviewed_support_direction is None
    if direction_missing:
        mismatches.append("direction_requires_review")
        review_notes.append(
            (
                "Falta una decisión revisada y vinculada a la fuente sobre la dirección contable."
                if language == "es"
                else "A source-bound reviewed journal-direction decision is required."
            )
        )
    else:
        checks_run.append("direction")

    relationship = (
        match.get("relationship")
        if isinstance(match.get("relationship"), dict)
        else None
    )
    party_missing = False
    party_status: str
    party_signal: str | None
    party_decision_ref: str | None
    if relationship is not None:
        party_status = "reviewed_relationship_exception"
        party_signal = "reviewed_support_relationship"
        party_decision_ref = str(relationship["decision_ref"])
        checks_run.append("party_relationship")
    else:
        party_status, party_signal = _pdf_party_perimeter(text, party_perimeter)
        party_decision_ref = (
            str(party_perimeter["decision_ref"])
            if party_perimeter is not None
            else None
        )
        checks_run.append("party_perimeter")
        if party_status == "missing":
            party_missing = True
            mismatches.append("party_perimeter_requires_review")
            review_notes.append(
                (
                    "Falta un perímetro de partes revisado; el texto libre no puede establecer la parte."
                    if language == "es"
                    else "A reviewed party perimeter is missing; free-text containment cannot establish the party."
                )
            )
        elif party_status != "matched":
            mismatches.append("party_perimeter")

    if pdf_payload["error"]:
        mismatches.append("pdf_text")
        review_notes.append(
            (
                f"Error al extraer el texto del PDF: {pdf_payload['error']}"
                if language == "es"
                else f"PDF text extraction error: {pdf_payload['error']}"
            )
        )
    required_checks = {"amount", "date", "currency", "direction"}
    missing_checks = sorted(required_checks - set(checks_run))
    if missing_checks:
        status = "manual_review"
        review_notes.append(
            (
                "Faltan campos obligatorios para las comprobaciones de importe, fecha y moneda."
                if language == "es"
                else "Required amount/date/currency check fields are missing."
            )
        )
        mismatches.extend(f"missing_{check}" for check in missing_checks)
    elif party_missing or direction_missing:
        status = "manual_review"
    elif mismatches:
        status = "mismatch"
    else:
        status = "ok"

    support_signals = [str(signal) for signal in match["signals"]]
    evidence_facts = {
        "support_match_status": "matched",
        "support_match_signals": support_signals,
        "amount_found": amount_value,
        "amount_source_locator": amount_locator,
        "date_found": date_value,
        "currency_found": currency_value,
        "currency_explicit_codes": list(explicit_currency_codes),
        "currency_explicit_conflict": currency_conflict,
        "currency_decision_ref": (
            currency_decision.get("decision_ref")
            if currency_decision is not None
            else None
        ),
        "journal_direction": expected_direction,
        "document_polarity": document_polarity,
        "reviewed_support_direction": reviewed_support_direction,
        "direction_match": (
            expected_direction is not None
            and reviewed_support_direction is not None
            and expected_direction == reviewed_support_direction
        ),
        "direction_decision_ref": (
            direction_decision.get("decision_ref")
            if direction_decision is not None
            else None
        ),
        "beneficiary_found": beneficiary_value,
        "pdf_text_extraction_error": pdf_payload["error"],
        "support_source_qualification_id": pdf_payload.get("source_qualification_id"),
        "party_perimeter_status": party_status,
        "party_perimeter_signal": party_signal,
        "party_decision_ref": party_decision_ref,
        "reviewed_relationship_recording_exception": (
            relationship.get("recording_exception")
            if relationship is not None
            else None
        ),
    }
    return {
        **entry,
        "status": status,
        "matched_pdf": matched_pdf,
        "checks_run": ",".join(checks_run),
        "mismatches": ",".join(dict.fromkeys(mismatches)),
        "review_notes": " ".join(review_notes),
        "amount_found": amount_value,
        "date_found": date_value,
        "beneficiary_found": beneficiary_value,
        "matched_support": matched_pdf,
        "support_type": "pdf",
        "support_artifact_id": pdf_payload.get("artifact_id"),
        "support_match_status": "matched",
        "support_match_signals": ",".join(support_signals),
        "evidence_facts": json.dumps(
            evidence_facts, ensure_ascii=False, sort_keys=True, separators=(",", ":")
        ),
        "professional_conclusion": "pending_review",
        "assurance_gate_status": (
            "reconciliation_passed;semantic_review_withheld"
            if status == "ok"
            else "reconciliation_failed;semantic_review_withheld"
        ),
        **amount_comparison,
    }


def _check_entry_with_support_ladder(
    entry: dict[str, Any],
    invoices: list[InvoiceRecord],
    pdfs: dict[str, dict[str, Any]],
    *,
    amount_tolerance: Decimal,
    date_window_days: int,
    invoice_artifact_ids: dict[str, str],
    party_perimeter: dict[str, Any] | None,
    relationships: dict[tuple[str, str, str], dict[str, Any]],
    currency_decisions: dict[tuple[str, str, str], dict[str, Any]],
    direction_decisions: dict[tuple[str, str, str], dict[str, Any]],
    language: str = "en",
) -> dict[str, Any]:
    """Try structured XML first and fall back to a matching PDF."""

    invoice, signals, xml_issue = match_invoice(
        entry,
        invoices,
        amount_tolerance=amount_tolerance,
        date_window_days=date_window_days,
    )
    relationship: dict[str, Any] | None = None
    if invoice is None:
        reviewed_candidates = [
            (candidate, candidate_relationship)
            for candidate in invoices
            if (
                candidate_relationship := _relationship_for(
                    entry,
                    invoice_artifact_ids.get(candidate.source_name),
                    candidate.source_name,
                    relationships,
                )
            )
            is not None
        ]
        if len(reviewed_candidates) == 1:
            invoice, relationship = reviewed_candidates[0]
            signals = ["reviewed_support_relationship"]
            expected_amount = entry.get("amount_abs")
            if expected_amount not in (None, "") and invoice.total_amount is not None:
                expected_amount_value = parse_canonical_decimal(
                    expected_amount,
                    label="entry amount_abs",
                )
                invoice_amount = parse_canonical_decimal(
                    invoice.total_amount,
                    label=f"{invoice.source_name} total_amount",
                )
                expected_currency_value = _clean_text(entry.get("currency")).upper()
                if _clean_text(invoice.currency).upper() == expected_currency_value:
                    _, within = difference_within_tolerance(
                        abs(expected_amount_value),
                        abs(invoice_amount),
                        amount_tolerance,
                    )
                    if within:
                        signals.append("amount")
            expected_date_value = _parse_date(entry.get("entry_date"))
            invoice_date_value = _parse_date(invoice.invoice_date)
            if (
                expected_date_value is not None
                and invoice_date_value is not None
                and abs((expected_date_value - invoice_date_value).days)
                <= date_window_days
            ):
                signals.append("date")
            xml_issue = None
        elif len(reviewed_candidates) > 1:
            xml_issue = "multiple_invoice_candidates"
    elif relationship is None:
        relationship = _relationship_for(
            entry,
            invoice_artifact_ids.get(invoice.source_name),
            invoice.source_name,
            relationships,
        )
        if relationship is not None:
            signals = [*signals, "reviewed_support_relationship"]
    if invoice is not None:
        support_artifact_id = invoice_artifact_ids.get(invoice.source_name)
        expected_currency = _clean_text(entry.get("currency")).upper()
        invoice_currency = _clean_text(invoice.currency).upper()
        expected_direction = _journal_direction(entry.get("amount_signed"))
        document_polarity = fatturapa_document_polarity(invoice.document_type)
        direction_decision = _relationship_for(
            entry,
            support_artifact_id,
            invoice.source_name,
            direction_decisions,
        )
        reviewed_support_direction = (
            str(direction_decision["expected_direction"])
            if direction_decision is not None
            else None
        )
        amount_comparison = _support_amount_comparison(
            entry.get("amount_signed"),
            invoice.total_amount,
            reviewed_support_direction,
            comparable_currency=invoice_currency == expected_currency,
        )
        comparison_signals = list(signals)
        if expected_currency and invoice_currency == expected_currency:
            comparison_signals.append("currency")
        if expected_direction is not None and reviewed_support_direction is not None:
            comparison_signals.append("direction")
        mismatches: list[str] = []
        for field, signal in (
            ("amount_abs", "amount"),
            ("entry_date", "date"),
            ("currency", "currency"),
        ):
            if entry.get(field) not in (None, "") and signal not in comparison_signals:
                mismatches.append(signal)
        party_missing = False
        party_status: str
        party_signal: str | None
        party_decision_ref: str | None
        if relationship is not None:
            party_status = "reviewed_relationship_exception"
            party_signal = "reviewed_support_relationship"
            party_decision_ref = str(relationship["decision_ref"])
            comparison_signals.append("party_relationship")
        else:
            party_status, party_signal = _structured_party_perimeter(
                invoice,
                party_perimeter,
            )
            party_decision_ref = (
                str(party_perimeter["decision_ref"])
                if party_perimeter is not None
                else None
            )
            comparison_signals.append("party_perimeter")
            if party_status == "missing":
                party_missing = True
                mismatches.append("party_perimeter_requires_review")
            elif party_status != "matched":
                mismatches.append("party_perimeter")
        direction_missing = (
            expected_direction is None or reviewed_support_direction is None
        )
        if direction_missing:
            mismatches.append("direction_requires_review")
        required_checks = {"amount", "date", "currency", "direction"}
        missing_checks = sorted(required_checks - set(comparison_signals))
        mismatches.extend(f"missing_{check}" for check in missing_checks)
        status = (
            "manual_review"
            if party_missing or direction_missing
            else ("mismatch" if mismatches else "ok")
        )
        evidence_facts = {
            "support_match_status": "matched",
            "support_match_signals": signals,
            "invoice_number": invoice.invoice_number,
            "invoice_document_type": invoice.document_type,
            "invoice_document_polarity": document_polarity,
            "invoice_date": invoice.invoice_date,
            "invoice_total_amount": invoice.total_amount,
            "invoice_currency": invoice.currency,
            "expected_currency": expected_currency,
            "currency_match": invoice_currency == expected_currency,
            "journal_direction": expected_direction,
            "reviewed_support_direction": reviewed_support_direction,
            "direction_match": (
                expected_direction is not None
                and reviewed_support_direction is not None
                and expected_direction == reviewed_support_direction
            ),
            "direction_decision_ref": (
                direction_decision.get("decision_ref")
                if direction_decision is not None
                else None
            ),
            "supplier_name": invoice.supplier_name,
            "supplier_tax_id": invoice.supplier_tax_id,
            "customer_name": invoice.customer_name,
            "customer_tax_id": invoice.customer_tax_id,
            "amount_source_locator": (
                "FatturaElettronicaBody[1]/DatiGenerali/"
                "DatiGeneraliDocumento[1]/ImportoTotaleDocumento"
            ),
            "party_perimeter_status": party_status,
            "party_perimeter_signal": party_signal,
            "party_decision_ref": party_decision_ref,
            "reviewed_relationship_recording_exception": (
                relationship.get("recording_exception")
                if relationship is not None
                else None
            ),
        }
        return {
            **entry,
            "status": status,
            "matched_pdf": None,
            "checks_run": ",".join(comparison_signals),
            "mismatches": ",".join(mismatches),
            "review_notes": (
                (
                    "Se encontró un único XML FatturaPA mediante el número de factura y al menos un campo corroborante."
                    if language == "es"
                    else "Matched a unique FatturaPA XML using the invoice number and at least one corroborating field."
                )
                if not mismatches
                else (
                    (
                        "Se encontró el XML FatturaPA, pero se requiere una decisión revisada y vinculada a la fuente sobre la dirección contable."
                        if language == "es"
                        else "Matched the FatturaPA XML, but a source-bound reviewed journal-direction decision is required."
                    )
                    if direction_missing
                    else (
                        "Se encontró un XML FatturaPA, pero faltan o no coinciden campos obligatorios o el perímetro de partes."
                        if language == "es"
                        else "Matched a FatturaPA XML, but required fields or the reviewed party perimeter are missing or differ."
                    )
                )
            ),
            "amount_found": invoice.total_amount if "amount" in signals else None,
            "date_found": invoice.invoice_date if "date" in signals else None,
            "beneficiary_found": (
                entry.get("beneficiary_expected") if "beneficiary" in signals else None
            ),
            "matched_support": invoice.source_name,
            "support_type": "fatturapa_xml",
            "support_artifact_id": support_artifact_id,
            "support_match_status": "matched",
            "support_match_signals": ",".join(comparison_signals),
            "evidence_facts": json.dumps(
                evidence_facts,
                ensure_ascii=False,
                sort_keys=True,
                separators=(",", ":"),
            ),
            "professional_conclusion": "pending_review",
            "assurance_gate_status": (
                "reconciliation_passed;semantic_review_withheld"
                if status == "ok"
                else "reconciliation_failed;semantic_review_withheld"
            ),
            **amount_comparison,
        }
    pdf_result = _check_one_entry(
        entry,
        pdfs,
        amount_tolerance=amount_tolerance,
        date_window_days=date_window_days,
        party_perimeter=party_perimeter,
        relationships=relationships,
        currency_decisions=currency_decisions,
        direction_decisions=direction_decisions,
        language=language,
    )
    if xml_issue:
        weak_relationship = xml_issue == "invoice_relationship_requires_review"
        pdf_result["status"] = "manual_review"
        pdf_result["assurance_gate_status"] = (
            "reconciliation_failed;semantic_review_withheld"
        )
        if weak_relationship:
            pdf_result["review_notes"] = (
                "Importe y fecha coinciden, pero falta una relación explícita por número de factura o una relación revisada."
                if language == "es"
                else "Amount and date coincide, but an explicit invoice-number or reviewed relationship is missing."
            )
        else:
            pdf_result["review_notes"] = (
                "Coinciden varios XML FatturaPA candidatos; se necesita un justificante específico o la selección de la persona revisora."
                if language == "es"
                else "Multiple FatturaPA XML candidates matched; targeted support or reviewer selection is required."
            )
        existing_mismatches = [
            value
            for value in str(pdf_result.get("mismatches") or "").split(",")
            if value
        ]
        issue_code = (
            "invoice_relationship_requires_review"
            if weak_relationship
            else "ambiguous_invoice_support"
        )
        pdf_result["mismatches"] = ",".join(
            dict.fromkeys([*existing_mismatches, issue_code])
        )
        pdf_result["support_match_status"] = (
            "relationship_requires_review" if weak_relationship else "ambiguous"
        )
        facts = json.loads(str(pdf_result["evidence_facts"]))
        facts["xml_match_status"] = (
            "relationship_requires_review" if weak_relationship else "ambiguous"
        )
        pdf_result["evidence_facts"] = json.dumps(
            facts, ensure_ascii=False, sort_keys=True, separators=(",", ":")
        )
    return pdf_result


def _apply_support_reuse_review(
    records: list[dict[str, Any]],
    *,
    language: str,
) -> list[dict[str, Any]]:
    usage: dict[tuple[str, str], int] = {}
    for record in records:
        artifact_id = _clean_text(record.get("support_artifact_id"))
        locator = _clean_text(record.get("matched_support"))
        if artifact_id and locator:
            key = (artifact_id, locator)
            usage[key] = usage.get(key, 0) + 1
    for record in records:
        key = (
            _clean_text(record.get("support_artifact_id")),
            _clean_text(record.get("matched_support")),
        )
        reuse_count = usage.get(key, 0)
        if reuse_count <= 1:
            continue
        if "reviewed_support_relationship" in str(
            record.get("support_match_signals") or ""
        ).split(","):
            continue
        record["status"] = "manual_review"
        mismatches = [
            value for value in str(record.get("mismatches") or "").split(",") if value
        ]
        record["mismatches"] = ",".join(
            dict.fromkeys([*mismatches, "support_reuse_requires_review"])
        )
        record["support_match_status"] = "reused_requires_review"
        record["assurance_gate_status"] = (
            "reconciliation_failed;semantic_review_withheld"
        )
        reuse_note = (
            "El mismo justificante coincide con varios asientos; se requiere una revisión explícita de la relación."
            if language == "es"
            else "The same support evidence matches multiple entries; explicit relationship review is required."
        )
        record["review_notes"] = " ".join(
            value
            for value in (_clean_text(record.get("review_notes")), reuse_note)
            if value
        )
        facts = json.loads(str(record.get("evidence_facts") or "{}"))
        facts["support_reuse_count"] = reuse_count
        facts["support_reuse_status"] = "requires_review"
        record["evidence_facts"] = json.dumps(
            facts, ensure_ascii=False, sort_keys=True, separators=(",", ":")
        )
    return records


def _status_counts(frame: pl.DataFrame) -> dict[str, int]:
    if frame.is_empty() or "status" not in frame.columns:
        return {}
    counts = frame.group_by("status").len(name="count").to_dicts()
    return {str(item["status"]): int(item["count"]) for item in counts}


def _write_review_notes(path: Path, audit: dict[str, Any]) -> None:
    if audit.get("language") == "es":
        lines = [
            "# Notas de revisión de la comprobación de asientos",
            "",
            f"- Idioma: {audit['language']}",
            f"- Asientos del diario: {audit['journal_row_count']}",
            f"- PDF justificativos: {audit['pdf_count']}",
            f"- XML FatturaPA: {audit['invoice_count']}",
            f"- Filas de resultados: {audit['result_row_count']}",
            "",
            "## Recuento por estado",
        ]
        counts = audit.get("status_counts", {})
        if counts:
            for status, count in sorted(counts.items()):
                lines.append(f"- {status}: {count}")
        else:
            lines.append("- ninguno")
        lines.extend(
            [
                "",
                "## Política de revisión",
                "Los scripts solo comparan evidencias deterministas. Codex debe explicar los casos no resueltos, inspeccionar los justificantes cuando sea necesario y mantener explícito el juicio profesional.",
                "",
            ]
        )
        path.write_text("\n".join(lines), encoding="utf-8")
        return

    lines = [
        "# Check Entries Review Notes",
        "",
        f"- Language: {audit['language']}",
        f"- Journal rows: {audit['journal_row_count']}",
        f"- Support PDFs: {audit['pdf_count']}",
        f"- FatturaPA XMLs: {audit['invoice_count']}",
        f"- Result rows: {audit['result_row_count']}",
        "",
        "## Status Counts",
    ]
    counts = audit.get("status_counts", {})
    if counts:
        for status, count in sorted(counts.items()):
            lines.append(f"- {status}: {count}")
    else:
        lines.append("- none")
    lines.extend(
        [
            "",
            "## Review Policy",
            "The scripts only compare deterministic evidence. Codex must explain unresolved cases, inspect support where needed, and keep professional judgment explicit.",
            "",
        ]
    )
    path.write_text("\n".join(lines), encoding="utf-8")


def _build_entry_checks_run(
    journal: Path,
    pdf_path: Path,
    output_dir: Path,
    recipe_path: Path | None = None,
    *,
    amount_tolerance: Decimal | str | int = "0",
    date_window_days: int = 0,
    language: object | None = None,
    document_language: object | None = None,
    connector_name: str | None = None,
    client_engagement: Mapping[str, Any] | None = None,
    enforce_client_output_path: bool = True,
) -> CheckRunResult:
    """Build one Check Entries run in an otherwise empty output directory."""

    if (
        not isinstance(date_window_days, int)
        or isinstance(date_window_days, bool)
        or date_window_days < 0
    ):
        raise ValueError("date_window_days must be a non-negative integer.")
    tolerance = _parse_tolerance(amount_tolerance)
    tolerance_text = decimal_text(tolerance)
    recipe, recipe_source_bytes = _captured_recipe(recipe_path)
    languages = language_assumptions(
        recipe, language=language, document_language=document_language
    )
    bound_journal = _bound_upstream_journal_execution(journal, client_engagement)
    journal_bytes = bound_journal.read_bytes()
    entries, journal_diag = _load_journal_entries(
        bound_journal,
        recipe,
        source_bytes=journal_bytes,
    )
    normalized_client_engagement = _validated_client_check_stage(
        client_engagement,
        journal=bound_journal,
        journal_diagnostics=journal_diag,
        support=pdf_path,
        output_dir=output_dir,
        stage="checks",
        enforce_output_path=enforce_client_output_path,
    )
    entries = _bound_sample_entries(entries, normalized_client_engagement)
    captured_support = _load_captured_support(pdf_path)
    support_receipts = [capture.receipt for capture in captured_support.captures]
    (
        party_by_entry,
        relationships,
        currency_decisions,
        direction_decisions,
        reviewed_recipe_decisions,
    ) = _reviewed_recipe_decisions(
        recipe,
        entries,
        captured_support,
    )
    pdfs = captured_support.pdfs
    invoices = list(captured_support.invoices)
    invoice_errors = list(captured_support.invoice_errors)
    records = [
        _check_entry_with_support_ladder(
            row,
            invoices,
            pdfs,
            amount_tolerance=tolerance,
            date_window_days=date_window_days,
            invoice_artifact_ids=captured_support.invoice_artifact_ids,
            party_perimeter=party_by_entry.get(str(row["prepared_entry_id"])),
            relationships=relationships,
            currency_decisions=currency_decisions,
            direction_decisions=direction_decisions,
            language=languages["language"],
        )
        for row in entries.to_dicts()
    ]
    records = _apply_support_reuse_review(
        records,
        language=languages["language"],
    )
    _validate_support_captures(captured_support)
    result_frame = (
        pl.DataFrame(records).select(RESULT_COLUMNS)
        if records
        else pl.DataFrame(schema={column: pl.Utf8 for column in RESULT_COLUMNS})
    )
    _, replayed_journal_diag = _load_journal_entries(
        bound_journal,
        recipe,
        source_bytes=journal_bytes,
        expected_normalization_replay=journal_diag["upstream_assurance"][
            "normalization_replay"
        ],
    )
    if (
        replayed_journal_diag["upstream_assurance"]
        != journal_diag["upstream_assurance"]
    ):
        raise ValueError("Journal Sampling assurance changed during entry checks.")
    output_dir.mkdir(parents=True, exist_ok=True)
    normalized_path = output_dir / "normalized_entries.csv"
    results_path = output_dir / "check_results.csv"
    xlsx_path = output_dir / "check_results.xlsx"
    support_facts_path = output_dir / "prepared_support_facts.csv"
    inventory_path = output_dir / "pdf_inventory.json"
    invoice_inventory_path = output_dir / "invoice_inventory.json"
    support_manifest_path = output_dir / "support_manifest.json"
    execution_recipe_path = output_dir / "execution_recipe.json"
    numeric_ledger_path = output_dir / "numeric_evidence_ledger.json"
    audit_path = output_dir / "check_audit.json"
    review_notes_path = output_dir / "review_notes.md"

    entries.write_csv(normalized_path)
    result_frame.write_csv(results_path)
    support_facts = _prepared_support_facts(records)
    support_facts.write_csv(support_facts_path)
    try:
        _write_stable_result_workbook(result_frame, xlsx_path)
    except (ImportError, ModuleNotFoundError, RuntimeError, ValueError) as exc:
        raise ValueError(
            "Check Entries requires a reproducible XLSX workpaper."
        ) from exc
    _validate_numeric_artifact_addresses(
        records,
        normalized_path=normalized_path,
        support_facts_path=support_facts_path,
        results_path=results_path,
        xlsx_path=xlsx_path,
    )
    pdf_inventory = _pdf_inventory_from_captured(captured_support)
    if normalized_client_engagement is not None:
        pdf_inventory = [
            {
                **item,
                **(
                    {
                        "path": _managed_check_reference(
                            Path(str(item["path"])),
                            normalized_client_engagement,
                        )
                    }
                    if isinstance(item.get("path"), str) and str(item["path"]).strip()
                    else {}
                ),
            }
            for item in pdf_inventory
        ]
    run_intake = write_run_intake(
        output_dir,
        journal,
        pdf_path,
        normalization_diagnostics_path=Path(journal_diag["normalization_diagnostics"]),
        recipe_path=recipe_path,
        language=languages["language"],
        document_language=languages["document_language"],
        amount_tolerance=tolerance_text,
        date_window_days=date_window_days,
        mapping=journal_diag["mapping"],
        journal_row_count=entries.height,
        pdf_count=len(pdfs),
        invoice_count=len(invoices),
        connector_name=connector_name,
        client_engagement=normalized_client_engagement,
    )
    recorded_run_intake = read_json(run_intake.path)
    recorded_input_paths = recorded_run_intake.get("input_paths")
    if not isinstance(recorded_input_paths, list) or len(recorded_input_paths) != 3:
        raise ValueError("Check Entries run intake input paths are unavailable.")
    recorded_journal_path = str(recorded_input_paths[0])
    recorded_diagnostics_path = str(recorded_input_paths[1])
    recorded_support_path = str(recorded_input_paths[2])
    recorded_journal_diag = {
        **journal_diag,
        **(
            {
                "client_engagement": _portable_client_engagement_context(
                    journal_diag.get("client_engagement")
                )
            }
            if isinstance(journal_diag.get("client_engagement"), Mapping)
            else {}
        ),
        "normalization_diagnostics": recorded_diagnostics_path,
    }
    write_json(
        inventory_path,
        {
            "pdf_count": len(pdf_inventory),
            "pdfs": pdf_inventory,
            "support_artifact_receipts": support_receipts,
            "support_manifest": captured_support.manifest,
        },
    )
    write_json(
        invoice_inventory_path,
        {
            "source_kind": (
                "authorized_connector_export" if connector_name else "local_upload"
            ),
            "connector_name": connector_name,
            "invoice_count": len(invoices),
            "invoices": [invoice.as_dict() for invoice in invoices],
            "errors": invoice_errors,
            "support_artifact_receipts": support_receipts,
            "support_manifest": captured_support.manifest,
        },
    )
    write_json(support_manifest_path, captured_support.manifest)
    execution_recipe_path.write_bytes(recipe_source_bytes)
    diagnostics_receipt = journal_diag["normalization_diagnostics_receipt"]
    upstream_normalized_receipt = journal_diag["normalized_csv_receipt"]
    normalized_receipt = {
        **upstream_normalized_receipt,
        "artifact_id": "source.normalized_journal",
        "root_id": "normalization",
        "role": "source",
    }
    validate_artifact_receipt({"normalization": journal.parent}, normalized_receipt)
    upstream_assurance_receipts = journal_diag["upstream_assurance"][
        "artifact_receipts"
    ]
    upstream_assurance_refs = [
        receipt["artifact_id"] for receipt in upstream_assurance_receipts
    ]
    source_qualification = build_source_qualification(
        qualification_id="qualification.check_entries_input",
        adapter_id="journal_sampling.normalization",
        adapter_version="2",
        source_family="journal_sampling.normalized.v2",
        status="qualified",
        source_artifact_refs=[
            normalized_receipt["artifact_id"],
            diagnostics_receipt["artifact_id"],
            *upstream_assurance_refs,
        ],
        candidate_row_count=entries.height,
        emitted_row_count=entries.height,
        reviewed_mapping_ref=None,
        controls=[
            {
                "control_id": "normalization_receipt",
                "required": True,
                "status": "passed",
                "evidence_refs": [normalized_receipt["artifact_id"]],
                "detail": "Journal Sampling normalized CSV receipt replayed.",
            },
            {
                "control_id": "diagnostics_closure",
                "required": True,
                "status": "passed",
                "evidence_refs": [diagnostics_receipt["artifact_id"]],
                "detail": "Diagnostics hash, qualifications, row counts, and exact monetary closure passed.",
            },
            {
                "control_id": "upstream_assurance_replay",
                "required": True,
                "status": "passed",
                "evidence_refs": upstream_assurance_refs,
                "detail": "Journal Sampling source receipts, reviewed decisions, implementation, qualifications, and source/preparation gates replayed.",
            },
            {
                "control_id": "fresh_normalization_reperformance",
                "required": True,
                "status": "passed",
                "evidence_refs": [normalized_receipt["artifact_id"]],
                "detail": (
                    "Raw journal bytes and the exact retained reviewed recipe "
                    "freshly reproduced the normalized population; replay receipt "
                    f"{journal_diag['upstream_assurance']['normalization_replay']['content_sha256']}."
                ),
            },
        ],
        limitations=[
            "Upstream qualification IDs are preserved in row lineage; this downstream qualification does not re-judge source authority."
        ],
    )
    implementation_receipts = _implementation_receipts()
    output_receipts = _output_receipts(
        output_dir,
        [
            normalized_path,
            support_facts_path,
            results_path,
            xlsx_path,
            inventory_path,
            invoice_inventory_path,
            support_manifest_path,
        ],
    )
    execution_recipe_receipt = artifact_receipt(
        output_dir,
        execution_recipe_path,
        artifact_id="source.check_entries_execution_recipe",
        root_id="run",
        role="source",
        media_type="application/json",
    )
    output_receipts.append(execution_recipe_receipt)
    numeric_ledger = _numeric_evidence_ledger(
        records,
        output_receipts=output_receipts,
    )
    write_json(numeric_ledger_path, numeric_ledger)
    numeric_ledger_receipt = artifact_receipt(
        output_dir,
        numeric_ledger_path,
        artifact_id="output.numeric_evidence_ledger",
        root_id="run",
        role="workpaper",
        media_type="application/json",
    )
    output_receipts.append(numeric_ledger_receipt)
    prepared_entries_ref = next(
        (
            receipt["artifact_id"]
            for receipt in output_receipts
            if receipt["path"] == normalized_path.name
        ),
        "output.normalized_entries",
    )
    check_results_ref = next(
        (
            receipt["artifact_id"]
            for receipt in output_receipts
            if receipt["path"] == results_path.name
        ),
        "output.check_results",
    )
    support_manifest_ref = _receipt_by_path(
        output_receipts,
        support_manifest_path.name,
    )["artifact_id"]
    unresolved_statuses = {
        str(row.get("status"))
        for row in result_frame.to_dicts()
        if row.get("status") != "ok"
    }
    all_support_qualified = bool(captured_support.source_qualifications) and all(
        qualification["status"] == "qualified"
        for qualification in captured_support.source_qualifications
    )
    missing_support = any(
        row.get("status") == "missing_support"
        or row.get("support_match_status") == "missing"
        for row in result_frame.to_dicts()
    )
    source_status = (
        "passed"
        if all_support_qualified and not invoice_errors and not missing_support
        else "failed"
    )
    reconciliation_status = (
        "passed" if source_status == "passed" and not unresolved_statuses else "failed"
    )
    reconciliation_limitations = (
        [
            "One or more rows remain unresolved: "
            + ", ".join(sorted(unresolved_statuses))
            + "."
        ]
        if unresolved_statuses
        else [
            "Passed means mechanical checks completed without a row-level mismatch; "
            "professional evidence sufficiency is still withheld."
        ]
    )
    gate_register = build_gate_register(
        {
            "source": {
                "status": source_status,
                "evidence_refs": [
                    source_qualification["qualification_id"],
                    support_manifest_ref,
                    *[
                        qualification["qualification_id"]
                        for qualification in captured_support.source_qualifications
                    ],
                ],
                "limitations": (
                    [
                        "Every support artifact passed bounded format qualification; receipts prove bytes, not source authority."
                    ]
                    if source_status == "passed"
                    else [
                        "Source failed because support is missing or at least one support artifact did not pass bounded qualification."
                    ]
                ),
            },
            "preparation": {
                "status": "passed" if source_status == "passed" else "blocked",
                "evidence_refs": (
                    [
                        prepared_entries_ref,
                        _receipt_by_path(output_receipts, support_facts_path.name)[
                            "artifact_id"
                        ],
                        source_qualification["qualification_id"],
                    ]
                    if source_status == "passed"
                    else []
                ),
                "limitations": (
                    ["Prepared rows inherit the reviewed Journal Sampling mappings."]
                    if source_status == "passed"
                    else ["Preparation assurance is blocked by the failed source gate."]
                ),
            },
            "reconciliation": {
                "status": reconciliation_status,
                "evidence_refs": [
                    check_results_ref,
                    numeric_ledger["ledger_id"],
                ],
                "limitations": reconciliation_limitations,
            },
            "semantic_review": {
                "status": "withheld",
                "evidence_refs": [],
                "limitations": [
                    "Professional evidence sufficiency and accounting conclusion require review."
                ],
            },
            "reporting": {
                "status": "blocked",
                "evidence_refs": [],
                "limitations": ["Semantic review is not complete."],
            },
            "publication": {
                "status": "withheld",
                "evidence_refs": [],
                "limitations": ["No publication authority was requested."],
            },
        }
    )
    lineage = [
        {
            "prepared_entry_id": row["prepared_entry_id"],
            "source_qualification_id": row["source_qualification_id"],
            "prepared_artifact_id": normalized_receipt["artifact_id"],
            "support_artifact_id": row.get("support_artifact_id"),
            "support_locator": row.get("matched_support"),
            "support_type": row.get("support_type"),
            "support_match_status": row.get("support_match_status"),
            "professional_conclusion": row.get("professional_conclusion"),
        }
        for row in result_frame.to_dicts()
    ]
    assurance_path = output_dir / "assurance_envelope.json"
    _validate_support_captures(captured_support)
    assurance_envelope = build_assurance_envelope(
        run_id=run_intake.run_id,
        workflow_id="check-entries",
        workflow_version="2",
        artifact_receipts=[
            normalized_receipt,
            diagnostics_receipt,
            *upstream_assurance_receipts,
            *support_receipts,
            *output_receipts,
            *implementation_receipts,
        ],
        implementation_artifact_refs=[
            receipt["artifact_id"] for receipt in implementation_receipts
        ],
        reviewed_decisions=reviewed_recipe_decisions,
        source_qualifications=[
            source_qualification,
            *captured_support.source_qualifications,
        ],
        allocation_ledgers=[],
        numeric_evidence_ledgers=[numeric_ledger],
        gate_register=gate_register,
        limitations=[
            "The envelope proves local byte identity and mechanical closure, not source authority or professional evidence sufficiency.",
            "Semantic review, reporting, and publication remain withheld.",
        ],
        artifact_roots={
            "normalization": journal.parent,
            "support": captured_support.root,
            "run": output_dir,
            **implementation_artifact_roots(),
        },
    )
    validate_implementation_contract(assurance_envelope)
    replayed_envelope = build_assurance_envelope(
        run_id=run_intake.run_id,
        workflow_id="check-entries",
        workflow_version="2",
        artifact_receipts=assurance_envelope["artifact_receipts"],
        implementation_artifact_refs=assurance_envelope["implementation_artifact_refs"],
        reviewed_decisions=assurance_envelope["reviewed_decisions"],
        source_qualifications=assurance_envelope["source_qualifications"],
        allocation_ledgers=assurance_envelope["allocation_ledgers"],
        numeric_evidence_ledgers=assurance_envelope["numeric_evidence_ledgers"],
        gate_register=assurance_envelope["gate_register"],
        limitations=assurance_envelope["limitations"],
        artifact_roots={
            "normalization": journal.parent,
            "support": captured_support.root,
            "run": output_dir,
            **implementation_artifact_roots(),
        },
    )
    validate_implementation_contract(replayed_envelope)
    if replayed_envelope != assurance_envelope:
        raise ValueError(
            "Check Entries assurance envelope is not reproducible across two builds."
        )
    write_json(assurance_path, assurance_envelope)
    assurance_receipt = artifact_receipt(
        output_dir,
        assurance_path,
        artifact_id="output.assurance_envelope",
        root_id="run",
        role="output",
        media_type="application/json",
    )
    output_receipts.append(assurance_receipt)
    assurance_path_reference = (
        (Path(str(recorded_run_intake["output_dir"])) / assurance_path.name).as_posix()
        if recorded_run_intake.get("path_reference") == "run_root_relative"
        else assurance_path.as_posix()
    )
    audit = {
        "schema_version": "check_entries.audit.v2",
        **languages,
        **(
            {
                "client_engagement": _portable_client_engagement_context(
                    normalized_client_engagement
                )
            }
            if normalized_client_engagement is not None
            else {}
        ),
        "run_id": run_intake.run_id,
        "journal": recorded_journal_path,
        "pdf_path": recorded_support_path,
        "journal_row_count": entries.height,
        "pdf_count": len(pdfs),
        "invoice_count": len(invoices),
        "invoice_error_count": len(invoice_errors),
        "connector_name": connector_name,
        "result_row_count": result_frame.height,
        "status_counts": _status_counts(result_frame),
        "amount_tolerance": tolerance_text,
        "date_window_days": date_window_days,
        "mapping": journal_diag["mapping"],
        "source_preparation": recorded_journal_diag,
        "upstream_normalized_csv_receipt": upstream_normalized_receipt,
        "source_qualification": source_qualification,
        "support_manifest": captured_support.manifest,
        "support_source_qualifications": list(captured_support.source_qualifications),
        "reviewed_recipe_decisions": reviewed_recipe_decisions,
        "execution_recipe": {
            "path": execution_recipe_path.name,
            "artifact_receipt": execution_recipe_receipt,
        },
        "numeric_evidence_ledger": numeric_ledger,
        "reproducibility_checks": {
            "xlsx_two_run_byte_equality": "passed",
            "assurance_two_build_equality": "passed",
        },
        "input_artifact_receipts": [
            normalized_receipt,
            diagnostics_receipt,
            *upstream_assurance_receipts,
            *support_receipts,
        ],
        "output_artifact_receipts": output_receipts,
        "lineage": lineage,
        "assurance_gates": gate_register,
        "assurance_envelope": {
            "path": assurance_path_reference,
            "content_sha256": assurance_envelope["content_sha256"],
            "artifact_receipt": assurance_receipt,
        },
        "professional_conclusion_status": "pending_review",
        "outputs": {
            "normalized_entries_csv": _managed_check_reference(
                normalized_path, normalized_client_engagement
            ),
            "prepared_support_facts_csv": _managed_check_reference(
                support_facts_path, normalized_client_engagement
            ),
            "check_results_csv": _managed_check_reference(
                results_path, normalized_client_engagement
            ),
            "check_results_xlsx": (
                _managed_check_reference(xlsx_path, normalized_client_engagement)
                if xlsx_path.exists()
                else None
            ),
            "numeric_evidence_ledger_json": _managed_check_reference(
                numeric_ledger_path, normalized_client_engagement
            ),
            "pdf_inventory_json": _managed_check_reference(
                inventory_path, normalized_client_engagement
            ),
            "invoice_inventory_json": _managed_check_reference(
                invoice_inventory_path, normalized_client_engagement
            ),
            "support_manifest_json": _managed_check_reference(
                support_manifest_path, normalized_client_engagement
            ),
            "execution_recipe_json": _managed_check_reference(
                execution_recipe_path, normalized_client_engagement
            ),
            "review_notes_md": _managed_check_reference(
                review_notes_path, normalized_client_engagement
            ),
            "assurance_envelope_json": _managed_check_reference(
                assurance_path, normalized_client_engagement
            ),
        },
        "review_session": {
            "run_intake_path": "run_intake.json",
            "review_payload_path": "review_payload.json",
            "ui_decisions_path": "ui_decisions.json",
            "final_artifacts_path": "final_artifacts.json",
            "status": "pending_review",
        },
    }
    audit["content_sha256"] = canonical_json_sha256(audit)
    write_json(audit_path, audit)
    _write_review_notes(review_notes_path, audit)
    write_review_session_artifacts(
        output_dir,
        Path(recorded_journal_path),
        Path(recorded_support_path),
        run_id=run_intake.run_id,
        run_intake_path=run_intake.path,
        recipe_path=execution_recipe_path,
        language=languages["language"],
        document_language=languages["document_language"],
        amount_tolerance=tolerance_text,
        date_window_days=date_window_days,
        mapping=journal_diag["mapping"],
        result_rows=result_frame.to_dicts(),
        pdf_inventory=pdf_inventory,
        audit=audit,
        client_engagement=normalized_client_engagement,
    )
    _validate_support_captures(captured_support)
    validate_assurance_envelope(
        assurance_envelope,
        artifact_roots={
            "normalization": journal.parent,
            "support": captured_support.root,
            "run": output_dir,
            **implementation_artifact_roots(),
        },
    )
    validate_implementation_contract(assurance_envelope)
    validate_initial_output_set(output_dir)
    return CheckRunResult(frame=result_frame, audit=audit)


def _remove_transaction_path(path: Path) -> None:
    """Remove only one exact transaction path."""

    if path.is_symlink() or path.is_file():
        path.unlink()
    elif path.is_dir():
        shutil.rmtree(path)


def run_entry_checks(
    journal: Path,
    pdf_path: Path,
    output_dir: Path,
    recipe_path: Path | None = None,
    *,
    amount_tolerance: Decimal | str | int = "0",
    date_window_days: int = 0,
    language: object | None = None,
    document_language: object | None = None,
    connector_name: str | None = None,
    client_engagement: Mapping[str, Any] | None = None,
    _enforce_client_output_path: bool = True,
) -> CheckRunResult:
    """Build a fresh run and restore the exact prior directory on failure.

    Directory replacement is deterministic here because fresh-run isolation
    and byte-exact rollback are assurance requirements. No prior review state
    is carried into a successful run.
    """

    target = output_dir.expanduser()
    resolved_target = target.resolve()
    if resolved_target == Path(resolved_target.anchor):
        raise ValueError("Check Entries output directory cannot be a filesystem root.")
    if target.is_symlink():
        raise ValueError("Check Entries output directory cannot be a symlink.")
    if target.exists() and not target.is_dir():
        raise ValueError("Check Entries output path must be a directory.")
    target.parent.mkdir(parents=True, exist_ok=True)
    backup = Path(
        tempfile.mkdtemp(
            prefix=f".{target.name}.check-entries-backup-",
            dir=target.parent,
        )
    )
    backup.rmdir()
    had_prior = target.exists()
    prior_moved = False
    succeeded = False
    try:
        if had_prior:
            os.replace(target, backup)
            prior_moved = True
        target.mkdir(parents=False, exist_ok=False)
        result = _build_entry_checks_run(
            journal,
            pdf_path,
            target,
            recipe_path,
            amount_tolerance=amount_tolerance,
            date_window_days=date_window_days,
            language=language,
            document_language=document_language,
            connector_name=connector_name,
            client_engagement=client_engagement,
            enforce_client_output_path=_enforce_client_output_path,
        )
        succeeded = True
        return result
    finally:
        if succeeded:
            if backup.exists():
                _remove_transaction_path(backup)
        else:
            if target.exists() or target.is_symlink():
                _remove_transaction_path(target)
            if prior_moved:
                os.replace(backup, target)
            elif backup.exists():
                _remove_transaction_path(backup)


def add_common_args(parser: argparse.ArgumentParser) -> None:
    parser.add_argument(
        "--language",
        help="Working/output language locale: it, en, fr, de, or es. Defaults to recipe or en.",
    )
    parser.add_argument(
        "--document-language",
        help="Source-document language locale: it, en, fr, de, es, or auto. Defaults to recipe or auto.",
    )
    parser.add_argument("--verbose", action="store_true", help="Enable debug logging.")
