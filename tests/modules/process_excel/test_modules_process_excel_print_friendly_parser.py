import datetime as _dt
from io import BytesIO

import openpyxl
import polars as pl
import pytest

from modules.process_excel.print_friendly_parser import (
    _parse_date,
    build_token_map,
    derive_date_column,
    looks_like_totals,
    parse_print_friendly_journal,
)


def test_derive_date_column_carries_forward_last_valid_date():
    # Arrange
    feb_1 = _dt.date(2024, 2, 1)
    mar_5 = _dt.date(2024, 3, 5)
    rows = [
        ["x", " 01/02/2024 ", None],  # dd/mm/YYYY -> 1 Feb 2024
        ["debit", "credit", "no date"],  # no date -> carry forward
        ["2024-03-05 extra", "", ""],  # YYYY-mm-dd
        [None, " ", "n/a"],  # still carry forward
    ]

    # Act
    out = derive_date_column(rows)

    # Assert
    assert out == [feb_1, feb_1, mar_5, mar_5]


def test_derive_date_column_ignores_invalid_date_tokens():
    # Arrange
    jan_15 = _dt.date(2024, 1, 15)
    rows = [
        ["15/01/2024", "", ""],  # valid
        ["31/02/2024", "", ""],  # invalid (Feb 31) -> should not update
        [None, None, None],
    ]

    # Act
    out = derive_date_column(rows)

    # Assert
    assert out == [jan_15, jan_15, jan_15]


def test_derive_date_column_all_none_when_no_dates_present():
    # Arrange
    rows = [
        [None, "", "   "],
        ["foo", "bar", "baz"],
        [1, 2, 3],
    ]

    # Act
    out = derive_date_column(rows)

    # Assert
    assert out == [None, None, None]


def test_build_token_map_contains_expected_tokens_and_keys():
    # Act
    token_map = build_token_map()

    # Assert
    expected_keys = {
        "debit",
        "credit",
        "account",
        "account_desc",
        "line_desc",
        "rowno",
        "date_hdr",
        "totals",
    }
    assert expected_keys.issubset(set(token_map.keys()))
    assert {"dare", "debit", "dr"}.issubset(set(token_map["debit"]))
    assert {"avere", "credit", "cr"}.issubset(set(token_map["credit"]))
    assert {
        "totale",
        "total",
        "progressivo",
        "saldo",
        "riporto",
        "carry forward",
        "tot",
    }.issubset(set(token_map["totals"]))


def test_build_token_map_language_argument_is_currently_ignored():
    # Act
    default_map = build_token_map()
    en_map = build_token_map("en")

    # Assert: identical mapping regardless of argument
    assert en_map == default_map


@pytest.mark.parametrize(
    "row,debit_idx,credit_idx,expected",
    [
        (  # token-based detection (case/whitespace agnostic)
            ["foo", " Totale ", 10.0, 0.0],
            2,
            3,
            True,
        ),
        (  # both debit and credit numeric -> totals-like
            ["", "", 100.0, 50.0],
            2,
            3,
            True,
        ),
        (  # only one numeric -> not totals
            ["", "", 100.0, ""],
            2,
            3,
            False,
        ),
    ],
)
def test_looks_like_totals_various_cases(row, debit_idx, credit_idx, expected):
    tokens = build_token_map()

    # Act
    got = looks_like_totals(row, debit_idx, credit_idx, tokens)

    # Assert
    assert got is expected


def _build_print_friendly_workbook() -> bytes:
    """Return an XLSX payload with merged Dare/Avere headers and continuation rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    header_row = 7

    ws.cell(row=header_row, column=1, value="Nr. Prog")
    ws.cell(row=header_row, column=2, value="Data Reg.")
    ws.cell(row=header_row, column=3, value="Data Doc.")
    ws.cell(row=header_row, column=4, value="Nr. Doc")
    ws.cell(row=header_row, column=5, value="Descrizione")
    ws.cell(row=header_row, column=6, value="Conto")
    ws.cell(row=header_row, column=7, value="Descrizione Conto")
    ws.cell(row=header_row, column=8, value="Dare (EUR)")
    ws.cell(row=header_row, column=11, value="Avere (EUR)")
    ws.cell(row=header_row, column=14, value="Nr. Reg")
    ws.merge_cells(
        start_row=header_row, start_column=8, end_row=header_row, end_column=10
    )
    ws.merge_cells(
        start_row=header_row, start_column=11, end_row=header_row, end_column=13
    )

    row_1 = header_row + 1
    ws.cell(row=row_1, column=1, value=1)
    ws.cell(row=row_1, column=2, value="01/10/2025")
    ws.cell(row=row_1, column=3, value="30/09/2025")
    ws.cell(row=row_1, column=4, value="183")
    ws.cell(row=row_1, column=5, value="PAGAMENTO EFFETTI FORNITORE")
    ws.cell(row=row_1, column=14, value=93551)

    row_2 = header_row + 2
    ws.cell(row=row_2, column=5, value="Num. 183 del 08/09/2025 - Scad. 30/09/2025")
    ws.cell(row=row_2, column=6, value="F 21360")
    ws.cell(row=row_2, column=7, value="STUDIO PITTATORE")
    ws.cell(row=row_2, column=9, value=1857)
    ws.cell(row=row_2, column=14, value=93551)

    row_3 = header_row + 3
    ws.cell(row=row_3, column=6, value="G 514")
    ws.cell(row=row_3, column=7, value="EXAMPLE BANK SPA")
    ws.cell(row=row_3, column=12, value=1500)
    ws.cell(row=row_3, column=14, value=93551)

    row_4 = header_row + 4
    ws.cell(row=row_4, column=6, value="G 74")
    ws.cell(row=row_4, column=7, value="ERARIO C/RITENUTE D'ACCONTO")
    ws.cell(row=row_4, column=12, value=357)
    ws.cell(row=row_4, column=14, value=93551)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_parse_print_friendly_journal_handles_merged_amount_columns():
    content = _build_print_friendly_workbook()

    df = parse_print_friendly_journal(content, language="auto")

    assert df.height == 3
    assert "dare" in df.columns
    assert "avere" in df.columns
    assert df.get_column("dare").to_list() == [1857.0, 0.0, 0.0]
    assert df.get_column("avere").to_list() == [0.0, 1500.0, 357.0]
    expected_date = _dt.date(2025, 10, 1)
    assert df.get_column("data_registrazione").to_list() == [
        expected_date,
        expected_date,
        expected_date,
    ]


def _build_workbook_without_date_header() -> bytes:
    """Return an XLSX payload where row numbers could be mistaken for dates."""

    wb = openpyxl.Workbook()
    ws = wb.active
    header_row = 3

    ws.cell(row=header_row, column=1, value="Riga")
    ws.cell(row=header_row, column=2, value="Conto")
    ws.cell(row=header_row, column=3, value="Dare")
    ws.cell(row=header_row, column=4, value="Avere")

    # Section line carrying the actual registration date.
    ws.cell(row=header_row + 1, column=1, value="01/08/2025 MOVIMENTO")

    # Detail lines use row numbers in the first column.
    ws.cell(row=header_row + 2, column=1, value=1)
    ws.cell(row=header_row + 2, column=2, value="1000")
    ws.cell(row=header_row + 2, column=3, value=10)

    ws.cell(row=header_row + 3, column=1, value=2)
    ws.cell(row=header_row + 3, column=2, value="2000")
    ws.cell(row=header_row + 3, column=4, value=10)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_parse_date_ignores_small_numeric_row_numbers():
    assert _parse_date(1) is None
    assert _parse_date(99.0) is None


def test_parse_print_friendly_journal_keeps_section_date_when_row_numbers_present(
    monkeypatch,
):
    content = _build_workbook_without_date_header()
    monkeypatch.setenv("FORCE_PRINT_FRIENDLY", "1")
    monkeypatch.delenv("FORCE_RAW", raising=False)

    df = parse_print_friendly_journal(content, language="auto")

    assert df.height == 2
    expected = _dt.date(2025, 8, 1)
    assert df.get_column("data_registrazione").to_list() == [expected, expected]
