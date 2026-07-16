from __future__ import annotations

from io import BytesIO

import polars as pl
import pytest
from openpyxl import load_workbook

from modules.utils.polars_excel_writer import write_polars_excel


def _load_wb(buf: BytesIO):
    buf.seek(0)
    return load_workbook(buf)


def test_write_polars_excel_single_df_basic_roundtrip():
    # Arrange
    df = pl.DataFrame({"id": [1, 2], "name": ["a", "b"]})
    buf = BytesIO()

    # Act
    write_polars_excel(df, buf)

    # Assert
    wb = _load_wb(buf)
    assert wb.sheetnames  # at least one sheet
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("id", "name")
    assert rows[1] == (1, "a")
    assert rows[2] == (2, "b")


def test_write_polars_excel_multiple_sheets_mapping():
    # Arrange
    data = {
        "S1": pl.DataFrame({"x": [10]}),
        "S2": pl.DataFrame({"y": ["ok"]}),
    }
    buf = BytesIO()

    # Act
    write_polars_excel(data, buf)

    # Assert
    wb = _load_wb(buf)
    assert set(wb.sheetnames) == {"S1", "S2"}

    ws1 = wb["S1"]
    rows1 = list(ws1.iter_rows(values_only=True))
    assert rows1[0] == ("x",)
    assert rows1[1] == (10,)

    ws2 = wb["S2"]
    rows2 = list(ws2.iter_rows(values_only=True))
    assert rows2[0] == ("y",)
    assert rows2[1] == ("ok",)


def test_write_polars_excel_raises_when_polars_export_fails(monkeypatch):
    # Arrange: force the Polars writer to fail.
    def raise_not_implemented(*_args, **_kwargs):  # pragma: no cover - helper
        raise NotImplementedError("simulate Polars write_excel failure")

    monkeypatch.setattr(pl.DataFrame, "write_excel", raise_not_implemented)

    df = pl.DataFrame({"a": [1]})
    buf = BytesIO()

    # Act / Assert
    with pytest.raises(RuntimeError, match="Polars Excel export failed"):
        write_polars_excel(df, buf)
