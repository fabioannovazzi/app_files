from __future__ import annotations

import importlib._bootstrap_external
import importlib.util
import json
import os
import shutil
import stat
import subprocess
import sys
from pathlib import Path
from typing import Any

import openpyxl
import pytest
from polars.testing import assert_frame_equal

from scripts.validate_plugin_review_contract import validate_contract

ROOT = Path(__file__).resolve().parents[2]
SCRIPT_DIR = ROOT / "plugins" / "journal-sampling" / "scripts"
CORE_PATH = SCRIPT_DIR / "journal_sampling_core.py"
MCP_SERVER_PATH = ROOT / "plugins" / "journal-sampling" / "mcp" / "server.cjs"
TRANSITIVE_IMPLEMENTATION_ATTACKS = [
    ("plugin", "scripts/check_dependencies.py"),
    ("plugin", "scripts/implementation_bootstrap.py"),
    ("plugin", "scripts/inspect_journal.py"),
    ("plugin", "scripts/journal_sampling_core.py"),
    ("plugin", "scripts/normalize_journal.py"),
    ("plugin", "scripts/replay_normalization.py"),
    ("plugin", "scripts/review_session.py"),
    ("plugin", "scripts/review_successor.py"),
    ("plugin", "scripts/run_sample.py"),
    ("plugin", "mcp/server.cjs"),
    ("plugin", "assets/icon.svg"),
    ("plugin", "assets/journal-sampling-review-widget.html"),
    ("plugin", "assets/review-workbench-adapter.json"),
    ("plugin", ".app.json"),
    ("plugin", ".mcp.json"),
    ("plugin", ".codex-plugin/plugin.json"),
    ("assurance", "__init__.py"),
    ("assurance", "contracts.py"),
    ("assurance", "decisions.py"),
    ("assurance", "envelope.py"),
    ("assurance", "money.py"),
    ("assurance", "relationships.py"),
    ("assurance", "review_output_transaction.cjs"),
    ("assurance", "serialization.py"),
]


def load_core() -> Any:
    if str(SCRIPT_DIR) not in sys.path:
        sys.path.insert(0, str(SCRIPT_DIR))
    spec = importlib.util.spec_from_file_location("journal_sampling_core", CORE_PATH)
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def _save_workbook(path: Path, rows: list[list[Any]]) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    workbook.save(path)


def _approve_suggested_recipe(path: Path) -> None:
    core = load_core()
    payload = json.loads(path.read_text(encoding="utf-8"))
    for index, (source_name, entry) in enumerate(payload["files"].items(), start=1):
        qualification = entry["qualification"]
        if qualification["status"] == "unsupported_source_layout":
            continue
        contract = core._mapping_contract(
            parser=entry["parser"],
            source_family=entry["source_family"],
            header_rows=entry.get("header_rows", []),
            mapping=entry.get("mapping", {}),
            layout=entry.get("layout", {}),
            excluded_monetary_columns=entry.get("excluded_monetary_columns", []),
            posting_identity=entry.get("posting_identity", "source_row"),
            carry_forward_fields=entry.get("carry_forward_fields", []),
            currency=entry.get("currency", "EUR"),
            unit=entry.get("unit", "currency"),
            decimal_separator=entry.get("decimal_separator"),
            thousands_separator=entry.get("thousands_separator"),
            amount_sign_convention=entry.get("amount_sign_convention"),
        )
        adapter_id = (
            core.PRINT_ADAPTER_ID
            if entry["parser"] == "print_friendly_excel"
            else core.TABULAR_ADAPTER_ID
        )
        decision_id = f"decision.journal_mapping.{index}"
        qualification["status"] = "reviewed"
        qualification["decision_ref"] = decision_id
        qualification["decision_receipt"] = core.build_reviewed_decision_receipt(
            decision_id=decision_id,
            decision_type="source_mapping",
            status="reviewed",
            reviewer_ref="reviewer.test",
            reviewed_on="2026-07-24",
            adapter_id=adapter_id,
            adapter_version=core.ADAPTER_VERSION,
            source_artifact_refs=[core._source_artifact_ref(Path(source_name))],
            content=contract,
        )
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def _call_mcp_server(
    messages: list[dict[str, object]],
    *,
    server_path: Path = MCP_SERVER_PATH,
    env: dict[str, str] | None = None,
) -> list[dict[str, object]]:
    node = shutil.which("node")
    if node is None:
        pytest.skip("Node.js is required to exercise the Journal Sampling MCP server.")
    completed = subprocess.run(
        [node, str(server_path), "--stdio"],
        input="\n".join(json.dumps(message) for message in messages) + "\n",
        capture_output=True,
        text=True,
        check=True,
        timeout=10,
        env={**os.environ, **(env or {})},
    )
    return [json.loads(line) for line in completed.stdout.splitlines() if line.strip()]


def _run_systematic_replay(
    core: Any,
    normalized_csv: Path,
    first_output: Path,
    second_output: Path,
) -> tuple[Any, Any]:
    first = core.run_sample(
        normalized_csv,
        first_output,
        method="systematic",
        size=2,
    )
    second = core.run_sample(
        normalized_csv,
        second_output,
        method="systematic",
        size=2,
    )
    return first, second


def _prepare_assured_population(
    core: Any,
    tmp_path: Path,
) -> tuple[Path, Path]:
    journal_path = tmp_path / "journal.xlsx"
    normalization_dir = tmp_path / "normalization"
    _save_workbook(
        journal_path,
        [
            ["Date", "Account", "Description", "Debit", "Credit"],
            ["2026-01-01", "1000", "Cash", "10.00", None],
            ["2026-01-02", "2000", "Revenue", None, "20.00"],
            ["2026-01-03", "3000", "Expense", "30.00", None],
        ],
    )
    core.inspect_path(journal_path, normalization_dir)
    recipe_path = normalization_dir / "suggested_recipe.json"
    _approve_suggested_recipe(recipe_path)
    core.normalize_path(journal_path, normalization_dir, recipe_path)
    return journal_path, normalization_dir / "normalized_journal.csv"


def test_plugin_workflow_normalizes_excel_and_samples(tmp_path: Path) -> None:
    core = load_core()
    journal_path = tmp_path / "journal.xlsx"
    output_dir = tmp_path / "out"
    sample_dir = output_dir / "sample"
    _save_workbook(
        journal_path,
        [
            ["Data", "Conto", "Descrizione conto", "Descrizione", "Dare", "Avere"],
            ["2025-01-01", "1000", "Cash", "Opening", 100, None],
            ["2025-01-02", "2000", "Revenue", "Sale", None, 100],
            ["2025-01-03", "3000", "Expense", "Cost", 50, None],
        ],
    )

    inspection = core.inspect_path(
        journal_path, output_dir, language="fr", document_language="it"
    )
    _approve_suggested_recipe(output_dir / "suggested_recipe.json")
    normalized = core.normalize_path(
        journal_path, output_dir, output_dir / "suggested_recipe.json"
    )
    sample = core.run_sample(
        output_dir / "normalized_journal.csv",
        sample_dir,
        method="random",
        size=2,
    )

    assert inspection.total_rows == 0
    inspection_payload = json.loads((output_dir / "inspection.json").read_text())
    recipe_payload = json.loads((output_dir / "suggested_recipe.json").read_text())
    assert inspection_payload["language"] == "fr"
    assert inspection_payload["document_language"] == "it"
    assert recipe_payload["language"] == "fr"
    assert recipe_payload["document_language"] == "it"
    assert recipe_payload["files"]["journal.xlsx"]["qualification"]["status"] == (
        "reviewed"
    )
    assert normalized.frame.height == 3
    assert normalized.diagnostics["population_status"] == "complete"
    assert normalized.frame.get_column("source_sheet").to_list() == [
        "Sheet",
        "Sheet",
        "Sheet",
    ]
    assert normalized.frame.get_column("source_row").to_list() == [2, 3, 4]
    assert sample.frame.height == 2
    sample_rows = sample.frame.to_dicts()
    first_sample_row = sample_rows[0]
    second_sample_row = sample_rows[1]
    sampling_audit = json.loads((sample_dir / "sampling_audit.json").read_text())
    run_intake = json.loads((sample_dir / "run_intake.json").read_text())
    review_payload = json.loads((sample_dir / "review_payload.json").read_text())
    ui_decisions = json.loads((sample_dir / "ui_decisions.json").read_text())
    final_artifacts = json.loads((sample_dir / "final_artifacts.json").read_text())

    assert (output_dir / "inspection.json").exists()
    assert (output_dir / "normalized_journal.csv").exists()
    assert (output_dir / "reviewed_decisions.json").exists()
    assert (output_dir / "assurance_gates.json").exists()
    assert (output_dir / "assurance_envelope.json").exists()
    assert (sample_dir / "journal_sample.csv").exists()
    assert (sample_dir / "sampling_audit.json").exists()
    assert (sample_dir / "run_intake.json").exists()
    assert (sample_dir / "review_payload.json").exists()
    assert (sample_dir / "ui_decisions.json").exists()
    assert (sample_dir / "final_artifacts.json").exists()
    assert (sample_dir / "sample_reproducibility.json").exists()
    assert (sample_dir / "sample_material_value_ledger.json").exists()
    assert (sample_dir / "sample_assurance_gates.json").exists()
    assert (sample_dir / "sample_assurance_envelope.json").exists()
    assert (sample_dir / "sample_output_receipts.json").exists()
    assert sampling_audit["review_session"]["run_id"] == run_intake["run_id"]
    assert (
        sampling_audit["population_proof"]["assurance_gates"]["gates"]["source"][
            "status"
        ]
        == "passed"
    )
    assert (
        sampling_audit["population_proof"]["assurance_gates"]["gates"]["preparation"][
            "status"
        ]
        == "passed"
    )
    assert review_payload["plugin"] == "journal-sampling"
    assert review_payload["run_id"] == run_intake["run_id"]
    assert review_payload["review_type"] == "journal_sampling_review"
    assert review_payload["item_count"] == len(review_payload["items"])
    item_types = {item["item_type"] for item in review_payload["items"]}
    assert {"sampling_control", "sampled_entry", "sample_artifact"} <= item_types
    assert review_payload["summary"]["sample_size"] == 2
    assert ui_decisions["status"] == "pending_review"
    assert final_artifacts["status"] == "written_pending_review"
    handoff_output = next(
        output
        for output in final_artifacts["outputs"]
        if output["path"] == "review_handoff.md"
    )
    handoff_text = (sample_dir / "review_handoff.md").read_text(encoding="utf-8")
    assert handoff_output["required_text"] == [
        "Review Handoff",
        "review_payload.json",
        "ui_decisions.json",
        "applied_decisions.json",
        "final_artifacts.json",
    ]
    assert "render_journal_sampling_review" in handoff_text
    assert "apply_journal_sampling_decisions" in handoff_text
    sample_output = next(
        output
        for output in final_artifacts["outputs"]
        if output["path"] == "journal_sample.csv"
    )
    sample_xlsx_output = next(
        output
        for output in final_artifacts["outputs"]
        if output["path"] == "journal_sample.xlsx"
    )
    assert sample_output["row_count"] == sampling_audit["sample_size"]
    assert sample_output["required_columns"] == core.CANONICAL_COLUMNS
    assert {
        "entry_date",
        "account",
        "account_desc",
        "line_desc",
        "amount_abs",
        "currency",
        "unit",
        "reported_increment",
        "source_file",
        "source_row",
        first_sample_row["entry_date"],
        str(first_sample_row["account"]),
        first_sample_row["account_desc"],
        first_sample_row["line_desc"],
        first_sample_row["source_file"],
    } <= set(sample_output["required_text"])
    assert "required_text" in sample_output["qa_checks"]
    assert sample_xlsx_output["source_row_count"] == sampling_audit["sample_size"]
    assert sample_xlsx_output["required_sheets"] == ["Sheet1"]
    assert sample_xlsx_output["required_sheet_headers"] == {
        "Sheet1": core.CANONICAL_COLUMNS
    }
    required_sample_cells = sample_xlsx_output["required_cells"]["Sheet1"]
    assert {
        "A1": "entry_date",
        "B1": "movement_number",
        "C1": "line_number",
        "D1": "account",
        "E1": "account_desc",
        "F1": "line_desc",
        "G1": "debit",
        "H1": "credit",
        "I1": "amount_signed",
        "J1": "amount_abs",
        "K1": "currency",
        "L1": "unit",
        "M1": "reported_increment",
        "N1": "source_file",
        "O1": "source_sheet",
        "P1": "source_page",
        "Q1": "source_row",
    }.items() <= required_sample_cells.items()
    assert {
        "A1": "entry_date",
        "A2": first_sample_row["entry_date"],
        "A3": second_sample_row["entry_date"],
        "D1": "account",
        "D2": str(first_sample_row["account"]),
        "D3": str(second_sample_row["account"]),
        "E1": "account_desc",
        "E2": first_sample_row["account_desc"],
        "E3": second_sample_row["account_desc"],
        "F1": "line_desc",
        "F2": first_sample_row["line_desc"],
        "F3": second_sample_row["line_desc"],
        "J1": "amount_abs",
        "J2": first_sample_row["amount_abs"],
        "J3": second_sample_row["amount_abs"],
        "K1": "currency",
        "K2": first_sample_row["currency"],
        "K3": second_sample_row["currency"],
        "L1": "unit",
        "L2": first_sample_row["unit"],
        "L3": second_sample_row["unit"],
        "M1": "reported_increment",
        "M2": first_sample_row["reported_increment"],
        "M3": second_sample_row["reported_increment"],
        "N1": "source_file",
        "N2": first_sample_row["source_file"],
        "N3": second_sample_row["source_file"],
        "Q1": "source_row",
        "Q2": str(first_sample_row["source_row"]),
        "Q3": str(second_sample_row["source_row"]),
    }.items() <= required_sample_cells.items()
    assert "required_cells" in sample_xlsx_output["qa_checks"]
    contract_report = validate_contract(
        sample_dir,
        strict_data_posture=True,
        strict_execution_trace=True,
        strict_output_paths=True,
        strict_output_content=True,
    )
    assert contract_report.ok, contract_report.as_dict()


def test_sample_stage_closes_all_values_receipts_gates_and_physical_outputs(
    tmp_path: Path,
) -> None:
    core = load_core()
    _, normalized_csv = _prepare_assured_population(core, tmp_path)
    output_dir = tmp_path / "sample"

    result = core.run_sample(normalized_csv, output_dir, method="systematic", size=2)

    replay = core.validate_sample_assurance(output_dir)
    ledger = replay["material_value_ledger"]
    envelope = replay["assurance_envelope"]
    assert result.frame.height == 2
    assert ledger["row_count"] == 2
    assert len(ledger["entries"]) == 2 * len(core.CANONICAL_COLUMNS)
    assert {entry["field"] for entry in ledger["entries"]} == set(
        core.CANONICAL_COLUMNS
    )
    assert all(len(entry["outputs"]) == 2 for entry in ledger["entries"])
    assert replay["output_set"]["physical_paths"] == sorted(
        [*core.SAMPLE_OUTPUT_PAYLOAD_PATHS, core.SAMPLE_OUTPUT_SET_PATH]
    )
    assert envelope["gate_register"]["gates"]["source"]["status"] == "passed"
    assert envelope["gate_register"]["gates"]["preparation"]["status"] == "passed"
    assert (
        envelope["gate_register"]["gates"]["semantic_review"]["status"]
        == "not_assessed"
    )
    assert envelope["gate_register"]["gates"]["reporting"]["status"] == "blocked"
    assert envelope["gate_register"]["report_ready"] is False
    assert envelope["implementation_artifact_refs"] == [
        *(artifact_id for _, artifact_id in core.IMPLEMENTATION_PLUGIN_FILES),
        *(artifact_id for _, artifact_id in core.ASSURANCE_IMPLEMENTATION_FILES),
    ]
    receipt_roles = {
        receipt["artifact_id"]: receipt["role"]
        for receipt in envelope["artifact_receipts"]
    }
    assert receipt_roles["prepared.normalized_journal"] == "prepared"
    assert receipt_roles["prepared.journal_sample_csv"] == "prepared"
    assert receipt_roles["output.journal_sample_xlsx"] == "output"


def test_second_sample_row_xlsx_mutation_breaks_all_row_material_closure(
    tmp_path: Path,
) -> None:
    core = load_core()
    _, normalized_csv = _prepare_assured_population(core, tmp_path)
    output_dir = tmp_path / "sample"
    core.run_sample(normalized_csv, output_dir, method="systematic", size=2)
    workbook_path = output_dir / "journal_sample.xlsx"
    workbook = openpyxl.load_workbook(workbook_path)
    workbook["Sheet1"]["J3"] = "999"
    workbook.save(workbook_path)

    with pytest.raises(ValueError, match="material value does not close"):
        core.validate_sample_material_value_ledger(output_dir, normalized_csv)


def test_unexpected_regular_file_aborts_finalization_without_success_artifacts(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    core = load_core()
    _, normalized_csv = _prepare_assured_population(core, tmp_path)
    output_dir = tmp_path / "sample"
    original = core.write_review_session_artifacts

    def inject_unexpected_file(*args: Any, **kwargs: Any) -> Any:
        result = original(*args, **kwargs)
        Path(args[0], "unexpected.bin").write_bytes(b"unexpected")
        return result

    monkeypatch.setattr(
        core,
        "write_review_session_artifacts",
        inject_unexpected_file,
    )

    with pytest.raises(ValueError, match="unexpected=.*unexpected.bin"):
        core.run_sample(normalized_csv, output_dir, method="systematic", size=2)
    assert not output_dir.exists()
    assert not list(tmp_path.glob(".journal-sampling-stage-*"))


def test_symlink_output_target_is_rejected_without_touching_its_directory(
    tmp_path: Path,
) -> None:
    core = load_core()
    _, normalized_csv = _prepare_assured_population(core, tmp_path)
    actual_output = tmp_path / "actual-output"
    actual_output.mkdir()
    output_alias = tmp_path / "sample-alias"
    output_alias.symlink_to(actual_output, target_is_directory=True)

    with pytest.raises(ValueError, match="cannot be a symlink"):
        core.run_sample(normalized_csv, output_alias, method="systematic", size=2)
    assert output_alias.is_symlink()
    assert not list(actual_output.iterdir())


def test_missing_required_output_aborts_finalization_without_success_artifacts(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    core = load_core()
    _, normalized_csv = _prepare_assured_population(core, tmp_path)
    output_dir = tmp_path / "sample"
    original = core.write_review_session_artifacts

    def remove_required_output(*args: Any, **kwargs: Any) -> Any:
        result = original(*args, **kwargs)
        Path(args[0], "journal_sample.xlsx").unlink()
        return result

    monkeypatch.setattr(
        core,
        "write_review_session_artifacts",
        remove_required_output,
    )

    with pytest.raises(ValueError, match="missing=.*journal_sample.xlsx"):
        core.run_sample(normalized_csv, output_dir, method="systematic", size=2)
    assert not output_dir.exists()
    assert not list(tmp_path.glob(".journal-sampling-stage-*"))


def test_source_change_during_sample_finalization_blocks_delivery(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    core = load_core()
    journal_path, normalized_csv = _prepare_assured_population(core, tmp_path)
    output_dir = tmp_path / "sample"
    original = core._write_sample_output_set

    def mutate_source_after_output_seal(staging_dir: Path) -> dict[str, Any]:
        result = original(staging_dir)
        workbook = openpyxl.load_workbook(journal_path)
        workbook.active["C2"] = "Changed during sampling"
        workbook.save(journal_path)
        return result

    monkeypatch.setattr(
        core,
        "_write_sample_output_set",
        mutate_source_after_output_seal,
    )

    with pytest.raises(ValueError, match="receipt does not match current bytes"):
        core.run_sample(normalized_csv, output_dir, method="systematic", size=2)
    assert not output_dir.exists()
    assert not list(tmp_path.glob(".journal-sampling-stage-*"))


def test_normalized_population_change_during_sample_finalization_blocks_delivery(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    core = load_core()
    _, normalized_csv = _prepare_assured_population(core, tmp_path)
    output_dir = tmp_path / "sample"
    original = core._write_sample_output_set

    def mutate_population_after_output_seal(staging_dir: Path) -> dict[str, Any]:
        result = original(staging_dir)
        normalized_csv.write_text(
            normalized_csv.read_text(encoding="utf-8") + "\n",
            encoding="utf-8",
        )
        return result

    monkeypatch.setattr(
        core,
        "_write_sample_output_set",
        mutate_population_after_output_seal,
    )

    with pytest.raises(ValueError, match="receipt does not match current bytes"):
        core.run_sample(normalized_csv, output_dir, method="systematic", size=2)
    assert not output_dir.exists()
    assert not list(tmp_path.glob(".journal-sampling-stage-*"))


def test_reproducibility_surface_is_stable_and_separate_from_run_receipts(
    tmp_path: Path,
) -> None:
    core = load_core()
    _, normalized_csv = _prepare_assured_population(core, tmp_path)
    first_output = tmp_path / "first"
    second_output = tmp_path / "second"

    core.run_sample(normalized_csv, first_output, method="systematic", size=2)
    core.run_sample(normalized_csv, second_output, method="systematic", size=2)

    assert (first_output / "sample_reproducibility.json").read_bytes() == (
        second_output / "sample_reproducibility.json"
    ).read_bytes()
    assert (first_output / "journal_sample.csv").read_bytes() == (
        second_output / "journal_sample.csv"
    ).read_bytes()
    assert (first_output / "run_intake.json").read_bytes() != (
        second_output / "run_intake.json"
    ).read_bytes()


def test_spanish_run_localizes_review_artifacts_and_workbook_sheet(
    tmp_path: Path,
) -> None:
    core = load_core()
    journal_path = tmp_path / "diario.xlsx"
    normalization_dir = tmp_path / "normalizado"
    normalized_csv = normalization_dir / "normalized_journal.csv"
    output_dir = tmp_path / "muestra"
    _save_workbook(
        journal_path,
        [
            ["Date", "Account", "Account desc", "Description", "Debit", "Credit"],
            ["2026-01-15", "6000", "Servicios", "Asesoría", "125.00", None],
            ["2026-01-16", "7000", "Ingresos", "Venta", None, "210.00"],
        ],
    )
    core.inspect_path(journal_path, normalization_dir)
    _approve_suggested_recipe(normalization_dir / "suggested_recipe.json")
    core.normalize_path(
        journal_path,
        normalization_dir,
        normalization_dir / "suggested_recipe.json",
    )

    result = core.run_sample(
        normalized_csv,
        output_dir,
        method="random",
        size=1,
        language="es-ES",
    )

    run_intake = json.loads((output_dir / "run_intake.json").read_text())
    review_payload = json.loads((output_dir / "review_payload.json").read_text())
    final_artifacts = json.loads((output_dir / "final_artifacts.json").read_text())
    handoff_text = (output_dir / "review_handoff.md").read_text(encoding="utf-8")
    workbook = openpyxl.load_workbook(
        output_dir / "journal_sample.xlsx",
        read_only=True,
        data_only=True,
    )

    assert result.audit["language"] == "es"
    assert run_intake["language"] == "es"
    assert run_intake["dependency_check"]["note"].startswith("Codex debe ejecutar")
    assert run_intake["data_posture"]["notes"][0].startswith("Los scripts de muestreo")
    assert review_payload["language"] == "es"
    assert [column["label"] for column in review_payload["columns"]] == [
        "Tipo",
        "Asiento o artefacto",
        "Acción sugerida",
        "Fuente",
        "Salida",
        "Estado",
    ]
    control = next(
        item
        for item in review_payload["items"]
        if item["item_type"] == "sampling_control"
    )
    sampled_entry = next(
        item for item in review_payload["items"] if item["item_type"] == "sampled_entry"
    )
    artifact_titles = {
        item["title"]
        for item in review_payload["items"]
        if item["item_type"] in {"sample_artifact", "review_artifact"}
    }
    assert control["title"] == "Muestra aleatoria: 1 de 2"
    assert "fila" in sampled_entry["source_path"]
    assert artifact_titles == {
        "CSV de la muestra del diario",
        "Libro Excel de la muestra del diario",
        "JSON de auditoría del muestreo",
    }
    assert handoff_text.startswith("# Muestreo del diario · Entrega para revisión")
    assert "## Revisión en Codex" in handoff_text
    assert "El guardado y la aplicación persistentes" in handoff_text
    assert "# Journal Sampling Review Handoff" not in handoff_text
    assert final_artifacts["caveats"][0].startswith("La muestra determinista")
    assert final_artifacts["next_actions"][1].startswith("Revise los parámetros")
    assert workbook.sheetnames == ["Muestra del diario"]
    sheet = workbook["Muestra del diario"]
    assert [cell.value for cell in sheet[1]] == core.CANONICAL_COLUMNS

    workbook_output = next(
        output
        for output in final_artifacts["outputs"]
        if output["path"] == "journal_sample.xlsx"
    )
    handoff_output = next(
        output
        for output in final_artifacts["outputs"]
        if output["path"] == "review_handoff.md"
    )
    assert workbook_output["required_sheets"] == ["Muestra del diario"]
    assert workbook_output["required_sheet_headers"] == {
        "Muestra del diario": core.CANONICAL_COLUMNS
    }
    assert "Muestra del diario" in workbook_output["required_cells"]
    assert handoff_output["required_text"][:3] == [
        "Review Handoff",
        "Entrega para revisión",
        "Revisión en Codex",
    ]

    adapter = json.loads(
        (
            ROOT
            / "plugins"
            / "journal-sampling"
            / "assets"
            / "review-workbench-adapter.json"
        ).read_text(encoding="utf-8")
    )
    assert adapter["localized"]["es"]["title"] == "Revisión de muestreo contable"

    contract_report = validate_contract(
        output_dir,
        strict_data_posture=True,
        strict_execution_trace=True,
        strict_output_paths=True,
        strict_output_content=True,
    )
    assert contract_report.ok, contract_report.as_dict()


def test_plugin_print_friendly_excel_extracts_detail_rows(tmp_path: Path) -> None:
    core = load_core()
    journal_path = tmp_path / "print_friendly.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    header_row = 7
    headers = {
        1: "Nr. Prog",
        2: "Data Reg.",
        5: "Descrizione",
        6: "Conto",
        7: "Descrizione Conto",
        8: "Dare (EUR)",
        11: "Avere (EUR)",
        14: "Nr. Reg",
    }
    for col_idx, value in headers.items():
        sheet.cell(row=header_row, column=col_idx, value=value)
    sheet.cell(row=8, column=2, value="01/10/2025")
    sheet.cell(row=8, column=5, value="PAGAMENTO FORNITORE")
    sheet.cell(row=8, column=14, value=93551)
    sheet.cell(row=9, column=6, value="F 21360")
    sheet.cell(row=9, column=7, value="FORNITORE")
    sheet.cell(row=9, column=9, value=1857)
    sheet.cell(row=10, column=6, value="G 514")
    sheet.cell(row=10, column=7, value="BANCA")
    sheet.cell(row=10, column=12, value=1857)
    workbook.save(journal_path)

    output_dir = tmp_path / "out"
    inspection = core.inspect_path(journal_path, output_dir)

    assert inspection.total_rows == 0
    assert inspection.files[0]["qualification_status"] == "needs_review"

    _approve_suggested_recipe(output_dir / "suggested_recipe.json")
    normalized = core.normalize_path(
        journal_path,
        output_dir,
        output_dir / "suggested_recipe.json",
    )

    assert normalized.frame.height == 2
    assert normalized.frame.get_column("account").to_list() == ["F 21360", "G 514"]
    assert normalized.frame.get_column("entry_date").to_list() == [
        "2025-10-01",
        "2025-10-01",
    ]


def test_print_adapter_obeys_reviewed_empty_carry_forward_policy(
    tmp_path: Path,
) -> None:
    core = load_core()
    journal_path = tmp_path / "print-no-carry.xlsx"
    output_dir = tmp_path / "out"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    header_row = 7
    headers = {
        2: "Data Reg.",
        5: "Descrizione",
        6: "Conto",
        7: "Descrizione Conto",
        8: "Dare (EUR)",
        11: "Avere (EUR)",
        14: "Nr. Reg",
    }
    for column, value in headers.items():
        sheet.cell(row=header_row, column=column, value=value)
    sheet.cell(row=8, column=2, value="01/10/2025")
    sheet.cell(row=8, column=5, value="PAGAMENTO FORNITORE")
    sheet.cell(row=8, column=14, value=93551)
    sheet.cell(row=9, column=6, value="F 21360")
    sheet.cell(row=9, column=9, value="1857.00")
    workbook.save(journal_path)

    core.inspect_path(journal_path, output_dir)
    recipe_path = output_dir / "suggested_recipe.json"
    recipe = json.loads(recipe_path.read_text(encoding="utf-8"))
    recipe["files"][journal_path.name]["carry_forward_fields"] = []
    recipe_path.write_text(
        json.dumps(recipe, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    inspection = core.inspect_path(journal_path, output_dir, recipe_path)
    _approve_suggested_recipe(recipe_path)
    normalized = core.normalize_path(journal_path, output_dir, recipe_path)

    assert inspection.files[0]["qualification_status"] == ("unsupported_source_layout")
    assert normalized.frame.height == 0
    assert normalized.diagnostics["population_status"] == "incomplete"


def test_print_adapter_requires_disposition_of_extra_numeric_column(
    tmp_path: Path,
) -> None:
    core = load_core()
    journal_path = tmp_path / "print-extra-numeric.xlsx"
    output_dir = tmp_path / "out"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    header_row = 7
    headers = {
        2: "Data Reg.",
        5: "Descrizione",
        6: "Conto",
        7: "Descrizione Conto",
        8: "Dare (EUR)",
        11: "Avere (EUR)",
        14: "Nr. Reg",
        16: "Auxiliary Value",
    }
    for column, value in headers.items():
        sheet.cell(row=header_row, column=column, value=value)
    sheet.cell(row=8, column=2, value="01/10/2025")
    sheet.cell(row=8, column=5, value="PAGAMENTO FORNITORE")
    sheet.cell(row=8, column=14, value=93551)
    sheet.cell(row=9, column=6, value="F 21360")
    sheet.cell(row=9, column=9, value="1857.00")
    sheet.cell(row=9, column=16, value="12.00")
    workbook.save(journal_path)

    inspection = core.inspect_path(journal_path, output_dir)
    _approve_suggested_recipe(output_dir / "suggested_recipe.json")
    normalized = core.normalize_path(
        journal_path,
        output_dir,
        output_dir / "suggested_recipe.json",
    )

    assert inspection.files[0]["unresolved_monetary_columns"] == ["column_15"]
    assert normalized.frame.height == 0
    assert normalized.diagnostics["population_status"] == "incomplete"


def test_plugin_text_pdf_path_abstains_from_generic_side_reconstruction(
    tmp_path: Path,
) -> None:
    core = load_core()
    pdf_path = tmp_path / "journal.pdf"
    pdf_path.write_bytes(b"%PDF placeholder")

    normalized = core.normalize_path(pdf_path, tmp_path / "out")

    assert normalized.frame.height == 0
    assert normalized.diagnostics["population_status"] == "incomplete"
    qualification = normalized.diagnostics["source_qualifications"][0]
    assert qualification["status"] == "unsupported_source_layout"
    review_payload = json.loads(
        (tmp_path / "out" / "qualification_review_payload.json").read_text()
    )
    assert review_payload["status"] == "blocked_by_source_qualification"
    assert (
        review_payload["items"][0]["data"]["qualification_status"]
        == "unsupported_source_layout"
    )


def test_plugin_supports_french_and_german_header_labels(tmp_path: Path) -> None:
    core = load_core()
    journal_path = tmp_path / "journal_de.xlsx"
    output_dir = tmp_path / "out"
    _save_workbook(
        journal_path,
        [
            ["Datum", "Konto", "Beschreibung", "Soll", "Haben"],
            ["2025-03-01", "1000", "Start", 80, None],
            ["2025-03-02", "2000", "Umsatz", None, 80],
        ],
    )

    core.inspect_path(journal_path, output_dir, language="de", document_language="de")
    _approve_suggested_recipe(output_dir / "suggested_recipe.json")
    normalized = core.normalize_path(
        journal_path,
        output_dir,
        output_dir / "suggested_recipe.json",
        language="de",
        document_language="de",
    )

    assert normalized.frame.height == 2
    assert normalized.diagnostics["language"] == "de"
    assert normalized.diagnostics["document_language"] == "de"
    assert normalized.frame.get_column("account").to_list() == ["1000", "2000"]


def test_reviewed_french_csv_preserves_localized_exact_amounts(
    tmp_path: Path,
) -> None:
    core = load_core()
    journal_path = tmp_path / "journal_fr.csv"
    output_dir = tmp_path / "out"
    journal_path.write_text(
        "Date,Compte,Libellé,Débit,Crédit\n"
        '01/07/2025,4010,Achat,"1 234,50",\n'
        '02/07/2025,5120,Banque,,"1 234,50"\n',
        encoding="utf-8",
    )

    core.inspect_path(
        journal_path,
        output_dir,
        language="fr",
        document_language="fr",
    )
    _approve_suggested_recipe(output_dir / "suggested_recipe.json")
    normalized = core.normalize_path(
        journal_path,
        output_dir,
        output_dir / "suggested_recipe.json",
    )

    assert normalized.diagnostics["population_status"] == "complete"
    assert normalized.frame.get_column("debit").to_list() == ["1234.5", None]
    assert normalized.frame.get_column("credit").to_list() == [None, "1234.5"]
    assert normalized.frame.get_column("amount_signed").to_list() == [
        "1234.5",
        "-1234.5",
    ]


def test_reviewed_thousands_separator_cannot_be_reinterpreted_as_decimal(
    tmp_path: Path,
) -> None:
    core = load_core()
    journal_path = tmp_path / "journal.csv"
    output_dir = tmp_path / "out"
    journal_path.write_text(
        "Date,Account,Description,Debit,Credit\n" '2025-07-01,4010,Purchase,"1,23",\n',
        encoding="utf-8",
    )
    core.inspect_path(journal_path, output_dir)
    recipe_path = output_dir / "suggested_recipe.json"
    recipe = json.loads(recipe_path.read_text(encoding="utf-8"))
    file_recipe = recipe["files"][journal_path.name]
    file_recipe["decimal_separator"] = None
    file_recipe["thousands_separator"] = ","
    recipe_path.write_text(
        json.dumps(recipe, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    _approve_suggested_recipe(recipe_path)

    normalized = core.normalize_path(journal_path, output_dir, recipe_path)

    assert normalized.frame.height == 0
    assert normalized.diagnostics["population_status"] == "incomplete"
    qualification = normalized.diagnostics["source_qualifications"][0]
    assert qualification["status"] == "unsupported_source_layout"
    assert qualification["emitted_row_count"] == 0


def test_unreviewed_mapping_withholds_population_and_blocks_sampling(
    tmp_path: Path,
) -> None:
    core = load_core()
    journal_path = tmp_path / "unreviewed.xlsx"
    output_dir = tmp_path / "out"
    _save_workbook(
        journal_path,
        [
            ["Date", "Account", "Description", "Debit", "Credit"],
            ["2025-04-01", "1000", "Cash", "10.00", None],
        ],
    )

    normalized = core.normalize_path(journal_path, output_dir)

    assert normalized.frame.height == 0
    assert normalized.diagnostics["population_status"] == "incomplete"
    qualification = normalized.diagnostics["source_qualifications"][0]
    assert qualification["status"] == "needs_review"
    assert qualification["emitted_row_count"] == 0
    with pytest.raises(ValueError, match="population is incomplete"):
        core.run_sample(
            output_dir / "normalized_journal.csv",
            output_dir / "sample",
            size=1,
        )
    assert not (output_dir / "sample").exists()


def test_stale_mapping_digest_withholds_rows(tmp_path: Path) -> None:
    core = load_core()
    journal_path = tmp_path / "stale.xlsx"
    output_dir = tmp_path / "out"
    _save_workbook(
        journal_path,
        [
            ["Date", "Account", "Description", "Debit", "Credit"],
            ["2025-04-01", "1000", "Cash", "10.00", None],
        ],
    )
    core.inspect_path(journal_path, output_dir)
    recipe_path = output_dir / "suggested_recipe.json"
    _approve_suggested_recipe(recipe_path)
    recipe = json.loads(recipe_path.read_text(encoding="utf-8"))
    recipe["files"][journal_path.name]["decimal_separator"] = "."
    recipe_path.write_text(
        json.dumps(recipe, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    normalized = core.normalize_path(journal_path, output_dir, recipe_path)

    assert normalized.frame.height == 0
    qualification = normalized.diagnostics["source_qualifications"][0]
    assert qualification["status"] == "needs_review"
    reviewed_control = next(
        item
        for item in qualification["controls"]
        if item["control_id"] == "reviewed_mapping"
    )
    assert reviewed_control["status"] == "not_assessed"


def test_exact_decimal_values_remain_canonical_through_sampling(
    tmp_path: Path,
) -> None:
    core = load_core()
    journal_path = tmp_path / "exact.xlsx"
    output_dir = tmp_path / "out"
    sample_dir = tmp_path / "sample"
    _save_workbook(
        journal_path,
        [
            ["Date", "Account", "Description", "Debit", "Credit"],
            ["2025-05-01", "1000", "Small debit", "0.10", None],
            ["2025-05-02", "2000", "Small credit", None, "0.20"],
        ],
    )
    core.inspect_path(journal_path, output_dir)
    _approve_suggested_recipe(output_dir / "suggested_recipe.json")
    normalized = core.normalize_path(
        journal_path,
        output_dir,
        output_dir / "suggested_recipe.json",
    )

    rows = normalized.frame.to_dicts()
    assert rows[0]["debit"] == "0.1"
    assert rows[0]["amount_signed"] == "0.1"
    assert rows[0]["currency"] == "EUR"
    assert rows[0]["unit"] == "currency"
    assert rows[0]["reported_increment"] == "0.01"
    assert rows[1]["credit"] == "0.2"
    assert rows[1]["amount_signed"] == "-0.2"
    assert rows[1]["amount_abs"] == "0.2"
    assert "0.10000000000000001" not in (
        output_dir / "normalized_journal.csv"
    ).read_text(encoding="utf-8")

    sampled = core.run_sample(
        output_dir / "normalized_journal.csv",
        sample_dir,
        method="mus",
        size=1,
        min_abs="0.15",
    )

    assert sampled.frame.height == 1
    assert sampled.frame.get_column("amount_abs").to_list() == ["0.2"]
    assert sampled.audit["filters"]["min_abs"] == "0.15"
    assert sampled.audit["population_proof"]["population_status"] == "complete"


def test_numeric_excel_cell_preserves_displayed_reported_increment(
    tmp_path: Path,
) -> None:
    core = load_core()
    journal_path = tmp_path / "formatted.xlsx"
    output_dir = tmp_path / "out"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Date", "Account", "Debit", "Credit"])
    sheet.append(["2025-05-01", "1000", 1234.5, None])
    sheet["C2"].number_format = "0.00"
    workbook.save(journal_path)

    core.inspect_path(journal_path, output_dir)
    _approve_suggested_recipe(output_dir / "suggested_recipe.json")
    normalized = core.normalize_path(
        journal_path,
        output_dir,
        output_dir / "suggested_recipe.json",
    )

    row = normalized.frame.to_dicts()[0]
    assert row["debit"] == "1234.5"
    assert row["reported_increment"] == "0.01"


def test_tabular_iso_timestamp_date_is_normalized(tmp_path: Path) -> None:
    core = load_core()
    journal_path = tmp_path / "journal.csv"
    journal_path.write_text(
        "Date,Account,Debit,Credit\n" "2025-01-28T00:00:00.000,1000,10.00,\n",
        encoding="utf-8",
    )
    output_dir = tmp_path / "out"

    core.inspect_path(journal_path, output_dir)
    _approve_suggested_recipe(output_dir / "suggested_recipe.json")
    normalized = core.normalize_path(
        journal_path,
        output_dir,
        output_dir / "suggested_recipe.json",
    )

    assert normalized.diagnostics["population_status"] == "complete"
    assert normalized.frame.get_column("entry_date").to_list() == ["2025-01-28"]


def test_systematic_sampling_replay_is_byte_identical(tmp_path: Path) -> None:
    core = load_core()
    journal_path = tmp_path / "journal.xlsx"
    normalization_dir = tmp_path / "normalized"
    first_output = tmp_path / "sample-a"
    second_output = tmp_path / "sample-b"
    _save_workbook(
        journal_path,
        [
            ["Date", "Account", "Description", "Debit", "Credit"],
            ["2025-06-01", "1000", "One", "10.00", None],
            ["2025-06-02", "2000", "Two", None, "10.00"],
            ["2025-06-03", "3000", "Three", "20.00", None],
        ],
    )
    core.inspect_path(journal_path, normalization_dir)
    _approve_suggested_recipe(normalization_dir / "suggested_recipe.json")
    core.normalize_path(
        journal_path,
        normalization_dir,
        normalization_dir / "suggested_recipe.json",
    )

    first, second = _run_systematic_replay(
        core,
        normalization_dir / "normalized_journal.csv",
        first_output,
        second_output,
    )

    assert_frame_equal(first.frame, second.frame)
    assert (first_output / "journal_sample.csv").read_bytes() == (
        second_output / "journal_sample.csv"
    ).read_bytes()


def test_modified_normalized_csv_receipt_blocks_sampling(tmp_path: Path) -> None:
    core = load_core()
    journal_path = tmp_path / "journal.xlsx"
    output_dir = tmp_path / "out"
    _save_workbook(
        journal_path,
        [
            ["Date", "Account", "Description", "Debit", "Credit"],
            ["2025-06-01", "1000", "Cash", "10.00", None],
        ],
    )
    core.inspect_path(journal_path, output_dir)
    _approve_suggested_recipe(output_dir / "suggested_recipe.json")
    core.normalize_path(
        journal_path,
        output_dir,
        output_dir / "suggested_recipe.json",
    )
    normalized_csv = output_dir / "normalized_journal.csv"
    normalized_csv.write_text(
        normalized_csv.read_text(encoding="utf-8") + "\n",
        encoding="utf-8",
    )

    with pytest.raises(ValueError, match="receipt does not match"):
        core.run_sample(normalized_csv, tmp_path / "sample", size=1)


def test_modified_original_source_receipt_blocks_sampling(tmp_path: Path) -> None:
    core = load_core()
    journal_path = tmp_path / "journal.xlsx"
    output_dir = tmp_path / "out"
    _save_workbook(
        journal_path,
        [
            ["Date", "Account", "Description", "Debit", "Credit"],
            ["2025-06-01", "1000", "Cash", "10.00", None],
        ],
    )
    core.inspect_path(journal_path, output_dir)
    _approve_suggested_recipe(output_dir / "suggested_recipe.json")
    core.normalize_path(
        journal_path,
        output_dir,
        output_dir / "suggested_recipe.json",
    )
    workbook = openpyxl.load_workbook(journal_path)
    workbook.active["C2"] = "Changed after normalization"
    workbook.save(journal_path)

    with pytest.raises(ValueError, match="receipt does not match"):
        core.run_sample(
            output_dir / "normalized_journal.csv",
            tmp_path / "sample",
            size=1,
        )


def _write_resealed_json(core: Any, path: Path, payload: dict[str, Any]) -> None:
    content = dict(payload)
    content.pop("content_sha256", None)
    content["content_sha256"] = core.canonical_json_sha256(content)
    path.write_text(
        json.dumps(content, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def test_normalization_retains_recipe_and_freshly_replays_raw_source(
    tmp_path: Path,
) -> None:
    core = load_core()
    _, normalized_csv = _prepare_assured_population(core, tmp_path)
    diagnostics = json.loads(
        (normalized_csv.parent / "normalization_diagnostics.json").read_text(
            encoding="utf-8"
        )
    )
    recipe_source = Path(diagnostics["normalization_recipe_source_path"])
    captured_recipe = normalized_csv.parent / diagnostics["normalization_recipe_path"]

    replay = core.replay_normalization_from_provenance(normalized_csv)

    assert captured_recipe.read_bytes() == recipe_source.read_bytes()
    assert diagnostics["normalization_recipe_receipt"]["sha256"] == (
        diagnostics["normalization_recipe_source_receipt"]["sha256"]
    )
    assert replay["schema_version"] == "journal_sampling.normalization_replay.v1"
    assert replay["status"] == "passed"
    assert replay["normalized_csv_sha256"] == (
        diagnostics["normalized_csv_receipt"]["sha256"]
    )


def test_isolated_replay_cli_matches_direct_reperformance(tmp_path: Path) -> None:
    core = load_core()
    _, normalized_csv = _prepare_assured_population(core, tmp_path)
    receipt_path = tmp_path / "replay-receipt.json"

    completed = subprocess.run(
        [
            sys.executable,
            "-I",
            "-B",
            str(SCRIPT_DIR / "replay_normalization.py"),
            str(normalized_csv),
            "--receipt-out",
            str(receipt_path),
        ],
        cwd=SCRIPT_DIR.parent,
        capture_output=True,
        text=True,
        check=False,
        timeout=60,
    )

    assert completed.returncode == 0, completed.stderr
    assert json.loads(receipt_path.read_text(encoding="utf-8")) == (
        core.replay_normalization_from_provenance(normalized_csv)
    )


def test_changed_reviewed_recipe_source_blocks_fresh_replay(tmp_path: Path) -> None:
    core = load_core()
    _, normalized_csv = _prepare_assured_population(core, tmp_path)
    diagnostics = json.loads(
        (normalized_csv.parent / "normalization_diagnostics.json").read_text(
            encoding="utf-8"
        )
    )
    recipe_source = Path(diagnostics["normalization_recipe_source_path"])
    recipe_source.write_text(
        recipe_source.read_text(encoding="utf-8") + " \n",
        encoding="utf-8",
    )

    with pytest.raises(ValueError, match="receipt does not match"):
        core.replay_normalization_from_provenance(normalized_csv)


def test_resealed_qualification_review_payload_fails_fresh_reperformance(
    tmp_path: Path,
) -> None:
    core = load_core()
    _, normalized_csv = _prepare_assured_population(core, tmp_path)
    payload_path = normalized_csv.parent / "qualification_review_payload.json"
    payload = json.loads(payload_path.read_text(encoding="utf-8"))
    payload["status"] = "forged_ready"
    payload_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    with pytest.raises(
        ValueError,
        match="qualification_review_payload.json",
    ):
        core.run_sample(normalized_csv, tmp_path / "sample", size=1)


def test_self_resealed_normalized_amount_fails_fresh_reperformance(
    tmp_path: Path,
) -> None:
    core = load_core()
    _, normalized_csv = _prepare_assured_population(core, tmp_path)
    normalization_root = normalized_csv.parent
    diagnostics_path = normalization_root / "normalization_diagnostics.json"
    envelope_path = normalization_root / "assurance_envelope.json"
    frame = core.pl.read_csv(normalized_csv, infer_schema=False)
    mutated = frame.with_columns(
        core.pl.when(core.pl.int_range(core.pl.len()) == 0)
        .then(core.pl.lit("11"))
        .otherwise(core.pl.col("debit"))
        .alias("debit"),
        core.pl.when(core.pl.int_range(core.pl.len()) == 0)
        .then(core.pl.lit("11"))
        .otherwise(core.pl.col("amount_signed"))
        .alias("amount_signed"),
        core.pl.when(core.pl.int_range(core.pl.len()) == 0)
        .then(core.pl.lit("11"))
        .otherwise(core.pl.col("amount_abs"))
        .alias("amount_abs"),
    )
    mutated.write_csv(normalized_csv)
    replacement_receipt = core.artifact_receipt(
        normalization_root,
        normalized_csv,
        artifact_id="prepared.normalized_journal",
        root_id="normalization",
        role="prepared",
        media_type="text/csv",
    )
    diagnostics = json.loads(diagnostics_path.read_text(encoding="utf-8"))
    diagnostics["normalized_csv_receipt"] = replacement_receipt
    envelope = json.loads(envelope_path.read_text(encoding="utf-8"))
    envelope["artifact_receipts"] = [
        (
            replacement_receipt
            if receipt["artifact_id"] == "prepared.normalized_journal"
            else receipt
        )
        for receipt in envelope["artifact_receipts"]
    ]
    _write_resealed_json(core, envelope_path, envelope)
    _write_resealed_json(core, diagnostics_path, diagnostics)

    with pytest.raises(ValueError, match="Fresh normalization does not reproduce"):
        core.run_sample(normalized_csv, tmp_path / "sample", size=1)


def test_self_resealed_raw_source_change_fails_fresh_reperformance(
    tmp_path: Path,
) -> None:
    core = load_core()
    journal_path, normalized_csv = _prepare_assured_population(core, tmp_path)
    normalization_root = normalized_csv.parent
    diagnostics_path = normalization_root / "normalization_diagnostics.json"
    envelope_path = normalization_root / "assurance_envelope.json"
    workbook = openpyxl.load_workbook(journal_path)
    workbook.active["D2"] = "11.00"
    workbook.save(journal_path)

    diagnostics = json.loads(diagnostics_path.read_text(encoding="utf-8"))
    old_source_receipt = diagnostics["source_receipts"][0]
    byte_count, digest = core.file_snapshot(journal_path)
    replacement_receipt = {
        **old_source_receipt,
        "byte_count": byte_count,
        "sha256": digest,
    }
    diagnostics["source_receipts"][0] = replacement_receipt
    diagnostics["files"][0]["source_receipt"] = replacement_receipt
    envelope = json.loads(envelope_path.read_text(encoding="utf-8"))
    envelope["artifact_receipts"] = [
        (
            replacement_receipt
            if receipt["artifact_id"] == old_source_receipt["artifact_id"]
            else receipt
        )
        for receipt in envelope["artifact_receipts"]
    ]
    _write_resealed_json(core, envelope_path, envelope)
    _write_resealed_json(core, diagnostics_path, diagnostics)

    with pytest.raises(ValueError, match="Fresh normalization does not reproduce"):
        core.run_sample(normalized_csv, tmp_path / "sample", size=1)


def test_modified_normalization_diagnostics_hash_blocks_sampling(
    tmp_path: Path,
) -> None:
    core = load_core()
    journal_path = tmp_path / "journal.xlsx"
    output_dir = tmp_path / "out"
    _save_workbook(
        journal_path,
        [
            ["Date", "Account", "Description", "Debit", "Credit"],
            ["2025-06-01", "1000", "Cash", "10.00", None],
        ],
    )
    core.inspect_path(journal_path, output_dir)
    _approve_suggested_recipe(output_dir / "suggested_recipe.json")
    core.normalize_path(
        journal_path,
        output_dir,
        output_dir / "suggested_recipe.json",
    )
    diagnostics_path = output_dir / "normalization_diagnostics.json"
    diagnostics = json.loads(diagnostics_path.read_text(encoding="utf-8"))
    diagnostics["row_count"] = 999
    diagnostics_path.write_text(
        json.dumps(diagnostics, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    with pytest.raises(ValueError, match="content hash is stale"):
        core.run_sample(
            output_dir / "normalized_journal.csv",
            tmp_path / "sample",
            size=1,
        )


def test_modified_assurance_envelope_blocks_sampling(tmp_path: Path) -> None:
    core = load_core()
    journal_path = tmp_path / "journal.xlsx"
    output_dir = tmp_path / "out"
    _save_workbook(
        journal_path,
        [
            ["Date", "Account", "Description", "Debit", "Credit"],
            ["2025-06-01", "1000", "Cash", "10.00", None],
        ],
    )
    core.inspect_path(journal_path, output_dir)
    _approve_suggested_recipe(output_dir / "suggested_recipe.json")
    core.normalize_path(
        journal_path,
        output_dir,
        output_dir / "suggested_recipe.json",
    )
    envelope_path = output_dir / "assurance_envelope.json"
    envelope = json.loads(envelope_path.read_text(encoding="utf-8"))
    envelope["gate_register"]["gates"]["source"]["status"] = "failed"
    envelope_path.write_text(
        json.dumps(envelope, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    with pytest.raises(ValueError, match="gate_register"):
        core.run_sample(
            output_dir / "normalized_journal.csv",
            tmp_path / "sample",
            size=1,
        )


def test_print_friendly_adapter_does_not_scan_nearby_amount_cells(
    tmp_path: Path,
) -> None:
    core = load_core()
    journal_path = tmp_path / "nearby.xlsx"
    output_dir = tmp_path / "out"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for column, value in {
        2: "Data Reg.",
        5: "Descrizione",
        6: "Conto",
        7: "Descrizione Conto",
        8: "Dare (EUR)",
        11: "Avere (EUR)",
    }.items():
        sheet.cell(row=7, column=column, value=value)
    sheet.cell(row=8, column=2, value="01/10/2025")
    sheet.cell(row=9, column=6, value="F 21360")
    # The retired parser searched three cells to the right of Dare and would
    # have promoted this unrelated nearby value as the debit amount.
    sheet.cell(row=9, column=10, value="1857.00")
    workbook.save(journal_path)
    core.inspect_path(journal_path, output_dir)
    _approve_suggested_recipe(output_dir / "suggested_recipe.json")

    normalized = core.normalize_path(
        journal_path,
        output_dir,
        output_dir / "suggested_recipe.json",
    )

    assert normalized.frame.height == 0
    qualification = normalized.diagnostics["source_qualifications"][0]
    assert qualification["status"] == "unsupported_source_layout"
    file_diag = normalized.diagnostics["files"][0]
    assert file_diag["candidate_row_count"] == 0
    assert file_diag["rejected_rows"] == []
    assert file_diag["excluded_non_monetary_rows"] == [9]


def test_wide_native_tabular_journal_uses_explicit_debit_credit_columns(
    tmp_path: Path,
) -> None:
    core = load_core()
    journal_path = tmp_path / "wide.xlsx"
    output_dir = tmp_path / "out"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Prima nota"
    header_row = 8
    headers = {
        1: "Numero Registrazione",
        2: "Riga",
        3: "Data registrazione",
        8: "Conto",
        9: "Descrizione",
        20: "Importo Dare",
        21: "Importo Avere",
        30: "Importo Ritenute",
    }
    for column, value in headers.items():
        sheet.cell(row=header_row, column=column, value=value)
    sheet.cell(row=9, column=1, value=101)
    sheet.cell(row=9, column=2, value=1)
    sheet.cell(row=9, column=3, value="01/07/2025")
    sheet.cell(row=9, column=8, value="4010")
    sheet.cell(row=9, column=9, value="Supplier")
    sheet.cell(row=9, column=20, value="1234.50")
    sheet.cell(row=10, column=1, value=101)
    sheet.cell(row=10, column=2, value=2)
    sheet.cell(row=10, column=8, value="5120")
    sheet.cell(row=10, column=9, value="Bank")
    sheet.cell(row=10, column=21, value="1234.50")
    # Account-only metadata is not a monetary journal line and must not make
    # the complete population fail.
    sheet.cell(row=11, column=8, value="FTR0")
    sheet.cell(row=11, column=9, value="Non-monetary metadata")
    workbook.save(journal_path)

    inspection = core.inspect_path(journal_path, output_dir)

    file_inspection = inspection.files[0]
    assert file_inspection["parser"] == "tabular"
    assert file_inspection["qualification_status"] == "needs_review"
    assert file_inspection["mapping"]["movement_number"] == "Numero Registrazione"
    assert file_inspection["mapping"]["line_number"] == "Riga"
    assert file_inspection["mapping"]["debit"] == "Importo Dare"
    assert file_inspection["mapping"]["credit"] == "Importo Avere"
    assert file_inspection["mapping"]["amount"] is None
    assert file_inspection["candidate_row_count"] == 2
    assert file_inspection["proposed_emitted_row_count"] == 2
    assert file_inspection["unresolved_monetary_columns"] == ["Importo Ritenute"]

    recipe_path = output_dir / "suggested_recipe.json"
    recipe = json.loads(recipe_path.read_text(encoding="utf-8"))
    recipe["files"][journal_path.name]["excluded_monetary_columns"] = [
        "Importo Ritenute"
    ]
    recipe_path.write_text(
        json.dumps(recipe, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    core.inspect_path(journal_path, output_dir, recipe_path)
    _approve_suggested_recipe(recipe_path)
    normalized = core.normalize_path(
        journal_path,
        output_dir,
        recipe_path,
    )

    assert normalized.diagnostics["population_status"] == "complete"
    assert normalized.frame.height == 2
    assert normalized.frame.get_column("source_sheet").to_list() == [
        "Prima nota",
        "Prima nota",
    ]
    assert normalized.frame.get_column("source_row").to_list() == [9, 10]
    assert normalized.frame.get_column("movement_number").to_list() == ["101", "101"]
    assert normalized.frame.get_column("line_number").to_list() == ["1", "2"]
    file_diagnostics = normalized.diagnostics["files"][0]
    assert file_diagnostics["candidate_row_count"] == 2
    assert file_diagnostics["excluded_non_monetary_rows"] == [11]
    assert file_diagnostics["excluded_monetary_columns"] == ["Importo Ritenute"]


def test_unmapped_explicit_monetary_column_blocks_complete_population(
    tmp_path: Path,
) -> None:
    core = load_core()
    journal_path = tmp_path / "alternate-debit.xlsx"
    output_dir = tmp_path / "out"
    _save_workbook(
        journal_path,
        [
            ["Date", "Account", "Debit", "Credit", "Alternate Debit"],
            ["2025-06-01", "1000", "10.00", None, None],
            ["2025-06-02", "2000", None, None, "5.00"],
        ],
    )

    inspection = core.inspect_path(journal_path, output_dir)
    _approve_suggested_recipe(output_dir / "suggested_recipe.json")
    normalized = core.normalize_path(
        journal_path,
        output_dir,
        output_dir / "suggested_recipe.json",
    )

    assert inspection.files[0]["unresolved_monetary_columns"] == ["Alternate Debit"]
    assert normalized.frame.height == 0
    assert normalized.diagnostics["population_status"] == "incomplete"
    qualification = normalized.diagnostics["source_qualifications"][0]
    assert qualification["status"] == "needs_review"


def test_unmapped_neutral_numeric_column_blocks_complete_population(
    tmp_path: Path,
) -> None:
    core = load_core()
    journal_path = tmp_path / "neutral-numeric.xlsx"
    output_dir = tmp_path / "out"
    _save_workbook(
        journal_path,
        [
            ["Date", "Account", "Debit", "Credit", "Auxiliary Value"],
            ["2025-06-01", "1000", "10.00", None, None],
            ["2025-06-02", "2000", None, None, "5.00"],
        ],
    )

    inspection = core.inspect_path(journal_path, output_dir)
    _approve_suggested_recipe(output_dir / "suggested_recipe.json")
    normalized = core.normalize_path(
        journal_path,
        output_dir,
        output_dir / "suggested_recipe.json",
    )

    assert inspection.files[0]["unresolved_monetary_columns"] == ["Auxiliary Value"]
    assert normalized.frame.height == 0
    assert normalized.diagnostics["population_status"] == "incomplete"
    assert normalized.diagnostics["source_qualifications"][0]["status"] == (
        "needs_review"
    )


def test_multi_sheet_workbook_is_withheld_until_every_sheet_is_supported(
    tmp_path: Path,
) -> None:
    core = load_core()
    journal_path = tmp_path / "multi-sheet.xlsx"
    output_dir = tmp_path / "out"
    workbook = openpyxl.Workbook()
    first = workbook.active
    first.title = "January"
    first.append(["Date", "Account", "Debit", "Credit"])
    first.append(["2025-01-01", "1000", "10.00", None])
    second = workbook.create_sheet("February")
    second.append(["Date", "Account", "Debit", "Credit"])
    second.append(["2025-02-01", "2000", None, "10.00"])
    workbook.save(journal_path)

    normalized = core.normalize_path(journal_path, output_dir)

    assert normalized.frame.height == 0
    assert normalized.diagnostics["population_status"] == "incomplete"
    file_diagnostics = normalized.diagnostics["files"][0]
    assert file_diagnostics["parser"] == "multi_sheet_workbook"
    assert file_diagnostics["sheet_names"] == ["January", "February"]
    assert file_diagnostics["qualification_status"] == "unsupported_source_layout"


def test_swap_and_restore_of_original_source_cannot_change_captured_parse(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    core = load_core()
    journal_path = tmp_path / "journal.xlsx"
    output_dir = tmp_path / "out"
    _save_workbook(
        journal_path,
        [
            ["Date", "Account", "Debit", "Credit"],
            ["2025-06-01", "1000", "10.00", None],
        ],
    )
    core.inspect_path(journal_path, output_dir)
    _approve_suggested_recipe(output_dir / "suggested_recipe.json")
    original_source_bytes = journal_path.read_bytes()
    original_normalize_file = core.normalize_file

    def swap_original_while_parsing(
        path: Path,
        recipe: dict[str, Any] | None = None,
        **kwargs: Any,
    ) -> Any:
        assert path != journal_path
        workbook = openpyxl.load_workbook(journal_path)
        workbook.active["C2"] = "999.00"
        workbook.save(journal_path)
        try:
            return original_normalize_file(path, recipe, **kwargs)
        finally:
            journal_path.write_bytes(original_source_bytes)

    monkeypatch.setattr(core, "normalize_file", swap_original_while_parsing)

    normalized = core.normalize_path(
        journal_path,
        output_dir,
        output_dir / "suggested_recipe.json",
    )

    assert normalized.diagnostics["population_status"] == "complete"
    assert normalized.frame.get_column("debit").to_list() == ["10"]
    assert journal_path.read_bytes() == original_source_bytes


def test_source_change_during_normalization_withholds_parsed_rows(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    core = load_core()
    journal_path = tmp_path / "journal.xlsx"
    output_dir = tmp_path / "out"
    _save_workbook(
        journal_path,
        [
            ["Date", "Account", "Debit", "Credit"],
            ["2025-06-01", "1000", "10.00", None],
        ],
    )
    core.inspect_path(journal_path, output_dir)
    _approve_suggested_recipe(output_dir / "suggested_recipe.json")
    original_normalize_file = core.normalize_file

    def mutate_after_parse(
        path: Path,
        recipe: dict[str, Any] | None = None,
        **kwargs: Any,
    ) -> Any:
        result = original_normalize_file(path, recipe, **kwargs)
        workbook = openpyxl.load_workbook(path)
        workbook.active["C2"] = "999.00"
        workbook.save(path)
        return result

    monkeypatch.setattr(core, "normalize_file", mutate_after_parse)

    normalized = core.normalize_path(
        journal_path,
        output_dir,
        output_dir / "suggested_recipe.json",
    )

    assert normalized.frame.height == 0
    assert normalized.diagnostics["population_status"] == "incomplete"
    file_diagnostics = normalized.diagnostics["files"][0]
    assert file_diagnostics["failure_class"] == ("source_changed_during_normalization")
    assert file_diagnostics["qualification_status"] == "unsupported_source_layout"


def test_corrupt_source_in_multi_file_intake_fails_closed_without_crashing(
    tmp_path: Path,
) -> None:
    core = load_core()
    input_dir = tmp_path / "input"
    input_dir.mkdir()
    valid_path = input_dir / "valid.xlsx"
    corrupt_path = input_dir / "corrupt.xlsx"
    output_dir = tmp_path / "out"
    _save_workbook(
        valid_path,
        [
            ["Date", "Account", "Description", "Debit", "Credit"],
            ["2025-06-01", "1000", "Cash", "10.00", None],
        ],
    )
    corrupt_path.write_bytes(b"not an Excel container")

    inspection = core.inspect_path(input_dir, output_dir)
    _approve_suggested_recipe(output_dir / "suggested_recipe.json")
    normalized = core.normalize_path(
        input_dir,
        output_dir,
        output_dir / "suggested_recipe.json",
    )

    assert inspection.total_rows == 0
    assert normalized.frame.height == 1
    assert normalized.diagnostics["population_status"] == "incomplete"
    diagnostics_by_file = {
        item["source_file"]: item for item in normalized.diagnostics["files"]
    }
    assert diagnostics_by_file["corrupt.xlsx"]["qualification_status"] == (
        "unsupported_source_layout"
    )
    assert diagnostics_by_file["corrupt.xlsx"]["parser"] == "unreadable_source"
    assert diagnostics_by_file["corrupt.xlsx"]["failure_class"] == "parser_failure"
    assert diagnostics_by_file["corrupt.xlsx"]["parser_error"]["type"]
    with pytest.raises(ValueError, match="population is incomplete"):
        core.run_sample(
            output_dir / "normalized_journal.csv",
            tmp_path / "sample",
            size=1,
        )
    assert not (tmp_path / "sample").exists()


def test_mixed_qualified_and_unsupported_sources_block_partial_sampling(
    tmp_path: Path,
) -> None:
    core = load_core()
    input_dir = tmp_path / "input"
    input_dir.mkdir()
    journal_path = input_dir / "journal.xlsx"
    pdf_path = input_dir / "journal_appendix.pdf"
    output_dir = tmp_path / "out"
    _save_workbook(
        journal_path,
        [
            ["Date", "Account", "Description", "Debit", "Credit"],
            ["2025-06-01", "1000", "Cash", "10.00", None],
        ],
    )
    pdf_path.write_bytes(b"%PDF ambiguous text layout")
    core.inspect_path(input_dir, output_dir)
    _approve_suggested_recipe(output_dir / "suggested_recipe.json")

    normalized = core.normalize_path(
        input_dir,
        output_dir,
        output_dir / "suggested_recipe.json",
    )

    assert normalized.frame.height == 1
    assert normalized.diagnostics["population_status"] == "incomplete"
    assert {
        item["status"] for item in normalized.diagnostics["source_qualifications"]
    } == {
        "qualified",
        "unsupported_source_layout",
    }
    with pytest.raises(ValueError, match="population is incomplete"):
        core.run_sample(
            output_dir / "normalized_journal.csv",
            tmp_path / "sample",
            size=1,
        )


def test_journal_sampling_vendors_shared_assurance_module() -> None:
    payload = json.loads(
        (ROOT / "scripts" / "plugin_vendor_modules.json").read_text(encoding="utf-8")
    )

    assert payload["plugins"]["journal-sampling"]["module_roots"] == ["vera_assurance"]


def test_skill_tells_codex_user_does_not_run_cli_directly() -> None:
    skill_text = (
        ROOT
        / "plugins"
        / "journal-sampling"
        / "skills"
        / "journal-sampling"
        / "SKILL.md"
    ).read_text(encoding="utf-8")

    assert "The user should not interact directly with CLI scripts" in skill_text
    assert "scripts/check_dependencies.py" in skill_text
    assert "it`, `en`, `fr`, `de`, and `es`" in skill_text
    assert "missing deterministic extraction script" in skill_text
    assert "suggested next engineering action" in skill_text
    assert "Keep the improvement note local to chat or run artifacts." in skill_text
    assert "validate_journal_sampling_review" in skill_text
    assert "render_journal_sampling_review" in skill_text
    assert "mapping_sha256" in skill_text
    assert "unsupported_source_layout" in skill_text
    assert "every requested source is qualified" in skill_text


def test_static_page_exposes_four_language_switch() -> None:
    page = (ROOT / "static" / "shared" / "journal-sampling" / "index.html").read_text(
        encoding="utf-8"
    )

    for snippet in (
        'data-lang="it"',
        'data-lang="en"',
        'data-lang="fr"',
        'data-lang="de"',
        "Crea un campione riproducibile da un giornale disordinato.",
        "Create a reproducible sample from a messy journal export.",
        "Créer un échantillon reproductible depuis un journal désordonné.",
        "Eine reproduzierbare Stichprobe aus einem uneinheitlichen Journal erstellen.",
    ):
        assert snippet in page


def test_journal_sampling_mcp_server_validates_and_renders_review_payload() -> None:
    review_payload = {
        "schema_version": "1.0",
        "plugin": "journal-sampling",
        "workflow": "journal-sampling",
        "run_id": "journal-sampling-test-run",
        "review_type": "journal_sampling_review",
        "items": [
            {
                "id": "sampling-control",
                "item_type": "sampling_control",
                "title": "random sample: 2 of 3",
                "output_path": "sampling_audit.json",
                "allowed_actions": ["accept", "edit", "mark_unclear", "skip"],
                "recommended_action": "accept",
                "evidence": [{"kind": "sampling_parameters", "method": "random"}],
                "data": {"method": "random", "sample_size": 2},
                "status": "needs_review",
            },
            {
                "id": "sampled-entry-1",
                "item_type": "sampled_entry",
                "title": "2025-01-02 | 2000 | 100",
                "source_path": "journal.xlsx; row 2",
                "output_path": "journal_sample.csv",
                "allowed_actions": ["accept", "edit", "mark_unclear", "skip"],
                "recommended_action": "accept",
                "evidence": [{"kind": "sampled_entry", "account": "2000"}],
                "data": {"account": "2000", "amount_abs": 100},
                "status": "needs_review",
            },
        ],
        "item_count": 2,
        "columns": [],
        "evidence": {},
        "allowed_actions": [
            "accept",
            "reject",
            "edit",
            "mark_unclear",
            "request_more_documents",
            "skip",
        ],
        "status": "ready_for_review",
        "summary": {
            "method": "random",
            "requested_size": 2,
            "population_size_after_filters": 3,
            "sample_size": 2,
        },
    }
    messages: list[dict[str, object]] = [
        {"jsonrpc": "2.0", "id": 1, "method": "tools/list"},
        {
            "jsonrpc": "2.0",
            "id": 2,
            "method": "tools/call",
            "params": {
                "name": "validate_journal_sampling_review",
                "arguments": {"review_payload": review_payload},
            },
        },
        {
            "jsonrpc": "2.0",
            "id": 3,
            "method": "tools/call",
            "params": {
                "name": "render_journal_sampling_review",
                "arguments": {"review_payload": review_payload},
            },
        },
        {"jsonrpc": "2.0", "id": 4, "method": "resources/list"},
        {
            "jsonrpc": "2.0",
            "id": 5,
            "method": "resources/read",
            "params": {"uri": "ui://widget/journal-sampling-review.html"},
        },
    ]

    responses = {response["id"]: response for response in _call_mcp_server(messages)}

    tool_names = {tool["name"] for tool in responses[1]["result"]["tools"]}
    assert {
        "validate_journal_sampling_review",
        "render_journal_sampling_review",
    } <= tool_names
    validate_result = responses[2]["result"]["structuredContent"]
    assert validate_result["ok"] is True
    assert validate_result["item_count"] == 2
    render_result = responses[3]["result"]
    assert render_result["structuredContent"]["widget_type"] == (
        "journal_sampling_review"
    )
    assert (
        render_result["_meta"]["openai/outputTemplate"]
        == "ui://widget/journal-sampling-review.html"
    )
    resource_uris = {
        resource["uri"] for resource in responses[4]["result"]["resources"]
    }
    assert "ui://widget/journal-sampling-review.html" in resource_uris
    widget_html = responses[5]["result"]["contents"][0]["text"]
    assert "Journal Sampling Review" in widget_html


def test_journal_sampling_mcp_server_localizes_spanish_runtime_feedback(
    tmp_path: Path,
) -> None:
    output_dir = tmp_path / "mcp-es"
    review_payload = {
        "schema_version": "1.0",
        "plugin": "journal-sampling",
        "workflow": "journal-sampling",
        "run_id": "journal-sampling-es",
        "language": "es-ES",
        "review_type": "journal_sampling_review",
        "items": [
            {
                "id": "sampling-control",
                "item_type": "sampling_control",
                "title": "Muestra aleatoria: 1 de 1",
                "allowed_actions": ["accept", "mark_unclear"],
                "recommended_action": "accept",
                "status": "needs_review",
            }
        ],
        "item_count": 1,
        "status": "ready_for_review",
    }
    decisions = [{"item_id": "sampling-control", "action": "accept"}]
    run_intake = {
        "run_id": "journal-sampling-es",
        "output_dir": str(output_dir),
        "language": "es",
    }
    final_artifacts = {
        "schema_version": "1.0",
        "plugin": "journal-sampling",
        "workflow": "journal-sampling",
        "run_id": "journal-sampling-es",
        "outputs": [],
        "next_actions": [],
    }
    output_dir.mkdir()
    for name, payload in [
        ("run_intake.json", run_intake),
        ("review_payload.json", review_payload),
        ("final_artifacts.json", final_artifacts),
    ]:
        (output_dir / name).write_text(
            json.dumps(payload, indent=2) + "\n",
            encoding="utf-8",
        )
    invalid_payload = {**review_payload, "item_count": 2}
    messages: list[dict[str, object]] = [
        {
            "jsonrpc": "2.0",
            "id": 1,
            "method": "initialize",
            "params": {
                "protocolVersion": "2024-11-05",
                "_meta": {"language": "es"},
            },
        },
        {
            "jsonrpc": "2.0",
            "id": 2,
            "method": "tools/call",
            "params": {
                "name": "validate_journal_sampling_review",
                "arguments": {"review_payload": review_payload},
            },
        },
        {
            "jsonrpc": "2.0",
            "id": 3,
            "method": "tools/call",
            "params": {
                "name": "save_journal_sampling_decisions",
                "arguments": {
                    "review_payload": review_payload,
                    "decisions": decisions,
                },
            },
        },
        {
            "jsonrpc": "2.0",
            "id": 4,
            "method": "tools/call",
            "params": {
                "name": "apply_journal_sampling_decisions",
                "arguments": {
                    "review_payload": review_payload,
                    "decisions": decisions,
                    "final_artifacts": final_artifacts,
                },
            },
        },
        {
            "jsonrpc": "2.0",
            "id": 5,
            "method": "tools/call",
            "params": {
                "name": "apply_journal_sampling_decisions",
                "arguments": {
                    "run_intake": run_intake,
                    "review_payload": review_payload,
                    "decisions": decisions,
                    "final_artifacts": final_artifacts,
                },
            },
        },
        {
            "jsonrpc": "2.0",
            "id": 6,
            "method": "tools/call",
            "params": {
                "name": "validate_journal_sampling_review",
                "arguments": {"review_payload": invalid_payload},
            },
        },
    ]

    responses = {response["id"]: response for response in _call_mcp_server(messages)}

    assert (
        "antes de render_journal_sampling_review"
        in responses[1]["result"]["instructions"]
    )
    assert responses[2]["result"]["structuredContent"]["message"].startswith(
        "El payload de revisión"
    )
    assert (
        "No se proporcionó run_intake.output_dir"
        in responses[3]["result"]["structuredContent"]["message"]
    )
    no_output_apply = responses[4]["result"]["structuredContent"]
    assert "No se proporcionó run_intake.output_dir" in no_output_apply["message"]
    assert no_output_apply["final_artifacts"]["next_actions"][-1].startswith(
        "Use los artefactos solo como muestra revisada"
    )
    persisted_apply = responses[5]["result"]["structuredContent"]
    assert persisted_apply["message"].startswith("Se han aplicado 1 decisiones")
    assert persisted_apply["final_artifacts"]["next_actions"][-1].startswith(
        "Use los artefactos solo como muestra revisada"
    )
    handoff = (output_dir / "review_handoff.md").read_text(encoding="utf-8")
    assert "<!-- Review Handoff -->" in handoff
    assert "Entrega para revisión" in handoff
    assert "## Revisión en Codex" in handoff
    error_result = responses[6]["result"]
    assert error_result["isError"] is True
    assert error_result["structuredContent"]["error"].startswith(
        "No se pudo validar la solicitud:"
    )


def _journal_transaction_case(
    output_dir: Path,
) -> dict[str, dict[str, Any] | list[dict[str, str]]]:
    output_dir.mkdir(mode=0o750)
    output_dir.chmod(0o750)
    nested = output_dir / "nested"
    nested.mkdir(mode=0o711)
    nested.chmod(0o711)
    sentinel = nested / "sentinel.bin"
    sentinel.write_bytes(b"\x00journal-original\xff")
    sentinel.chmod(0o640)
    run_intake = {
        "run_id": "journal-transaction-run",
        "output_dir": str(output_dir),
        "language": "en",
        "execution_trace": [],
    }
    review_payload = {
        "schema_version": "1.0",
        "plugin": "journal-sampling",
        "workflow": "journal-sampling",
        "run_id": "journal-transaction-run",
        "review_type": "journal_sampling_review",
        "items": [
            {
                "id": "sampling-control",
                "item_type": "sampling_control",
                "title": "Sampling control",
                "allowed_actions": ["accept"],
                "recommended_action": "accept",
                "status": "needs_review",
            }
        ],
        "item_count": 1,
        "status": "ready_for_review",
    }
    final_artifacts = {
        "schema_version": "1.0",
        "plugin": "journal-sampling",
        "workflow": "journal-sampling",
        "run_id": "journal-transaction-run",
        "outputs": [],
        "next_actions": [],
    }
    for name, payload in [
        ("run_intake.json", run_intake),
        ("review_payload.json", review_payload),
        ("final_artifacts.json", final_artifacts),
    ]:
        (output_dir / name).write_text(
            json.dumps(payload, indent=2) + "\n",
            encoding="utf-8",
        )
    return {
        "run_intake": run_intake,
        "review_payload": review_payload,
        "final_artifacts": final_artifacts,
        "decisions": [{"item_id": "sampling-control", "action": "accept"}],
    }


def _journal_tree_image(root: Path) -> dict[str, tuple[Any, ...]]:
    image: dict[str, tuple[Any, ...]] = {}
    for entry in [root, *sorted(root.rglob("*"))]:
        observed = entry.lstat()
        relative = "." if entry == root else entry.relative_to(root).as_posix()
        mode = observed.st_mode & 0o7777
        if stat.S_ISREG(observed.st_mode):
            image[relative] = (
                "file",
                mode,
                observed.st_nlink,
                entry.read_bytes(),
            )
        elif stat.S_ISDIR(observed.st_mode):
            image[relative] = ("directory", mode)
        elif stat.S_ISLNK(observed.st_mode):
            image[relative] = ("symlink", mode, os.readlink(entry))
        else:
            image[relative] = ("special", stat.S_IFMT(observed.st_mode), mode)
    return image


def _journal_faulted_server(
    tmp_path: Path,
    *,
    needle: str,
    replacement: str,
) -> Path:
    source = MCP_SERVER_PATH.read_text(encoding="utf-8")
    plugin_root_line = 'const PLUGIN_ROOT = path.resolve(__dirname, "..");'
    assert source.count(plugin_root_line) == 1
    assert source.count(needle) == 1
    source = source.replace(
        plugin_root_line,
        f"const PLUGIN_ROOT = {json.dumps(str(MCP_SERVER_PATH.parents[1]))};",
        1,
    )
    source = source.replace(needle, replacement, 1)
    server_path = tmp_path / "journal-faulted-server.cjs"
    server_path.write_text(source, encoding="utf-8")
    return server_path


def _journal_transaction_call(
    tool_name: str,
    arguments: dict[str, Any],
    *,
    server_path: Path = MCP_SERVER_PATH,
    env: dict[str, str] | None = None,
) -> dict[str, Any]:
    response = _call_mcp_server(
        [
            {
                "jsonrpc": "2.0",
                "id": 1,
                "method": "tools/call",
                "params": {"name": tool_name, "arguments": arguments},
            }
        ],
        server_path=server_path,
        env=env,
    )[0]
    return response["result"]["structuredContent"]


def _real_journal_review_case(
    core: Any,
    tmp_path: Path,
    *,
    output_name: str = "sample",
) -> tuple[Path, dict[str, Any]]:
    _, normalized_csv = _prepare_assured_population(core, tmp_path)
    output_dir = tmp_path / output_name
    core.run_sample(
        normalized_csv,
        output_dir,
        method="systematic",
        size=2,
    )

    def read_json(name: str) -> dict[str, Any]:
        return json.loads((output_dir / name).read_text(encoding="utf-8"))

    review_payload = read_json("review_payload.json")
    return output_dir, {
        "run_intake": read_json("run_intake.json"),
        "review_payload": review_payload,
        "ui_decisions": read_json("ui_decisions.json"),
        "final_artifacts": read_json("final_artifacts.json"),
        "decisions": [
            {"item_id": item["id"], "action": "accept"}
            for item in review_payload["items"]
        ],
    }


def _current_real_journal_review_arguments(
    output_dir: Path,
    decisions: list[dict[str, str]],
) -> dict[str, Any]:
    def read_json(name: str) -> dict[str, Any]:
        return json.loads((output_dir / name).read_text(encoding="utf-8"))

    return {
        "run_intake": read_json("run_intake.json"),
        "review_payload": read_json("review_payload.json"),
        "ui_decisions": read_json("ui_decisions.json"),
        "final_artifacts": read_json("final_artifacts.json"),
        "decisions": decisions,
    }


def _reseal_mutated_applied_receipts(core: Any, output_dir: Path) -> None:
    envelope_path = output_dir / "sample_assurance_envelope.json"
    envelope = json.loads(envelope_path.read_text(encoding="utf-8"))
    applied_path = output_dir / "applied_decisions.json"
    applied_size, applied_sha256 = core.file_snapshot(applied_path)
    for receipt in envelope["artifact_receipts"]:
        if receipt["artifact_id"] == "workpaper.applied_decisions":
            receipt["byte_count"] = applied_size
            receipt["sha256"] = applied_sha256
    envelope_content = {
        key: value for key, value in envelope.items() if key != "content_sha256"
    }
    envelope["content_sha256"] = core.canonical_json_sha256(envelope_content)
    core.write_json(envelope_path, envelope)

    manifest_path = output_dir / core.SAMPLE_OUTPUT_SET_PATH
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    for receipt in manifest["receipts"]:
        receipt_path = output_dir / receipt["path"]
        if receipt["path"] in {
            "applied_decisions.json",
            "sample_assurance_envelope.json",
        }:
            byte_count, digest = core.file_snapshot(receipt_path)
            receipt["byte_count"] = byte_count
            receipt["sha256"] = digest
    manifest_content = {
        key: value for key, value in manifest.items() if key != "content_sha256"
    }
    manifest["content_sha256"] = core.canonical_json_sha256(manifest_content)
    core.write_json(manifest_path, manifest)
    manifest_path.chmod(core.SAMPLE_OUTPUT_SET_MODE)


@pytest.fixture(scope="module")
def assured_implementation_attack_output(
    tmp_path_factory: pytest.TempPathFactory,
) -> Path:
    core = load_core()
    fixture_root = tmp_path_factory.mktemp("journal-implementation-attacks")
    output_dir, _ = _real_journal_review_case(
        core,
        fixture_root,
        output_name="assured-sample",
    )
    return output_dir


def _copy_journal_implementation_tree(
    tmp_path: Path,
) -> tuple[Path, Path]:
    copied_plugins = tmp_path / "plugins"
    copied_plugin = copied_plugins / "journal-sampling"
    copied_assurance = (
        copied_plugins / "_shared" / "vendor" / "modules" / "vera_assurance"
    )
    shutil.copytree(
        ROOT / "plugins" / "journal-sampling",
        copied_plugin,
        ignore=shutil.ignore_patterns("__pycache__", "*.pyc", "*.pyo"),
    )
    shutil.copytree(
        ROOT / "plugins" / "_shared" / "vendor" / "modules" / "vera_assurance",
        copied_assurance,
        ignore=shutil.ignore_patterns("__pycache__", "*.pyc", "*.pyo"),
    )
    return copied_plugin, copied_assurance


def _mutate_implementation_bytes(path: Path) -> None:
    suffix = path.suffix.lower()
    if suffix == ".json":
        mutation = b" \n"
    elif suffix == ".html":
        mutation = b"\n<!-- receipt mutation -->\n"
    elif suffix == ".cjs":
        mutation = b"\n// receipt mutation\n"
    else:
        mutation = b"\n# receipt mutation\n"
    path.write_bytes(path.read_bytes() + mutation)


def _implementation_attack_arguments(output_dir: Path) -> dict[str, Any]:
    return {
        "run_intake": json.loads(
            (output_dir / "run_intake.json").read_text(encoding="utf-8")
        ),
        "review_payload": json.loads(
            (output_dir / "review_payload.json").read_text(encoding="utf-8")
        ),
        "ui_decisions": json.loads(
            (output_dir / "ui_decisions.json").read_text(encoding="utf-8")
        ),
        "final_artifacts": json.loads(
            (output_dir / "final_artifacts.json").read_text(encoding="utf-8")
        ),
    }


def test_transitive_implementation_attack_matrix_is_exact() -> None:
    core = load_core()

    assert TRANSITIVE_IMPLEMENTATION_ATTACKS == [
        *(("plugin", path) for path, _ in core.IMPLEMENTATION_PLUGIN_FILES),
        *(("assurance", path) for path, _ in core.ASSURANCE_IMPLEMENTATION_FILES),
    ]


@pytest.mark.parametrize(
    "attack_kind",
    ["empty_directory", "regular", "symlink", "hardlink", "fifo"],
)
def test_python_preimport_rejects_every_unowned_implementation_entry(
    tmp_path: Path,
    attack_kind: str,
) -> None:
    copied_plugin, _ = _copy_journal_implementation_tree(tmp_path)
    rogue = copied_plugin / "scripts" / "rogue"
    external = tmp_path / f"{attack_kind}-external"
    if attack_kind == "empty_directory":
        rogue.mkdir()
    elif attack_kind == "regular":
        rogue.write_bytes(b"unreceipted implementation")
    elif attack_kind == "symlink":
        external.write_bytes(b"external")
        rogue.symlink_to(external)
    elif attack_kind == "hardlink":
        external.write_bytes(b"external")
        rogue.hardlink_to(external)
    else:
        os.mkfifo(rogue)

    completed = subprocess.run(
        [
            sys.executable,
            "-I",
            "-B",
            str(copied_plugin / "scripts" / "replay_normalization.py"),
            "--help",
        ],
        cwd=copied_plugin,
        capture_output=True,
        text=True,
        check=False,
        timeout=30,
    )

    assert completed.returncode != 0
    assert "implementation" in completed.stderr.lower()


def test_real_python_entry_rejects_timestamp_valid_unreceipted_bytecode(
    tmp_path: Path,
) -> None:
    copied_plugin, _ = _copy_journal_implementation_tree(tmp_path)
    target = copied_plugin / "scripts" / "journal_sampling_core.py"
    source = target.read_bytes()
    source_stat = target.stat()
    marker = tmp_path / "malicious-pyc-executed.txt"
    malicious = (
        "from pathlib import Path as _AttackPath\n"
        f"_AttackPath({marker.as_posix()!r}).write_text("
        "'executed before validation\\n', encoding='utf-8')\n"
        f"exec(compile({source.decode('utf-8')!r}, {target.as_posix()!r}, "
        "'exec'), globals())\n"
    )
    code = compile(malicious, target.as_posix(), "exec")
    cache_prefix = sys.pycache_prefix
    try:
        sys.pycache_prefix = None
        cache_path = Path(importlib.util.cache_from_source(target.as_posix()))
    finally:
        sys.pycache_prefix = cache_prefix
    cache_path.parent.mkdir()
    cache_path.write_bytes(
        importlib._bootstrap_external._code_to_timestamp_pyc(
            code,
            int(source_stat.st_mtime),
            source_stat.st_size,
        )
    )

    completed = subprocess.run(
        [
            sys.executable,
            "-I",
            "-B",
            str(copied_plugin / "scripts" / "replay_normalization.py"),
            "--help",
        ],
        cwd=copied_plugin,
        capture_output=True,
        text=True,
        check=False,
        timeout=30,
    )

    assert completed.returncode != 0
    assert "implementation" in completed.stderr.lower()
    assert not marker.exists()
    assert target.read_bytes() == source
    assert target.stat().st_size == source_stat.st_size
    assert target.stat().st_mtime_ns == source_stat.st_mtime_ns


@pytest.mark.parametrize("attack_kind", ["symlink", "hardlink", "fifo"])
def test_real_python_entry_rejects_unsafe_bootstrap_before_read(
    tmp_path: Path,
    attack_kind: str,
) -> None:
    copied_plugin, _ = _copy_journal_implementation_tree(tmp_path)
    bootstrap = copied_plugin / "scripts" / "implementation_bootstrap.py"
    original = bootstrap.read_bytes()
    bootstrap.unlink()
    external = tmp_path / f"bootstrap-{attack_kind}-external.py"
    if attack_kind == "symlink":
        external.write_bytes(original)
        bootstrap.symlink_to(external)
    elif attack_kind == "hardlink":
        external.write_bytes(original)
        bootstrap.hardlink_to(external)
    else:
        os.mkfifo(bootstrap)

    completed = subprocess.run(
        [
            sys.executable,
            "-I",
            "-B",
            str(copied_plugin / "scripts" / "replay_normalization.py"),
            "--help",
        ],
        cwd=copied_plugin,
        capture_output=True,
        text=True,
        check=False,
        timeout=30,
    )

    assert completed.returncode != 0
    assert "bootstrap is not a real file" in completed.stderr


def test_mcp_rejects_unowned_implementation_path_before_stdio(
    tmp_path: Path,
) -> None:
    copied_plugin, _ = _copy_journal_implementation_tree(tmp_path)
    (copied_plugin / "scripts" / "__pycache__").mkdir()
    node = shutil.which("node")
    if node is None:
        pytest.skip("Node.js is required for the MCP implementation probe.")

    completed = subprocess.run(
        [node, str(copied_plugin / "mcp" / "server.cjs"), "--stdio"],
        input=json.dumps(
            {
                "jsonrpc": "2.0",
                "id": 1,
                "method": "tools/list",
                "params": {},
            }
        )
        + "\n",
        capture_output=True,
        text=True,
        check=False,
        timeout=30,
    )

    assert completed.returncode != 0
    assert "implementation" in completed.stderr.lower()


def test_pristine_copied_implementation_tree_replays_in_python(
    tmp_path: Path,
    assured_implementation_attack_output: Path,
) -> None:
    copied_plugin, _ = _copy_journal_implementation_tree(tmp_path)

    completed = subprocess.run(
        [
            sys.executable,
            "-I",
            "-B",
            str(copied_plugin / "scripts" / "review_successor.py"),
            "validate",
            str(assured_implementation_attack_output),
        ],
        capture_output=True,
        text=True,
        check=False,
        timeout=20,
    )

    assert completed.returncode == 0, completed.stderr
    assert json.loads(completed.stdout)["output_set"]["stage"]["kind"] == "initial"


def test_pristine_copied_implementation_tree_replays_through_mcp(
    tmp_path: Path,
    assured_implementation_attack_output: Path,
) -> None:
    copied_plugin, _ = _copy_journal_implementation_tree(tmp_path)

    result = _journal_transaction_call(
        "validate_journal_sampling_review",
        _implementation_attack_arguments(assured_implementation_attack_output),
        server_path=copied_plugin / "mcp" / "server.cjs",
        env={"JOURNAL_SAMPLING_PYTHON": sys.executable},
    )

    assert result["ok"] is True


@pytest.mark.parametrize(
    ("implementation_root", "relative_path"),
    TRANSITIVE_IMPLEMENTATION_ATTACKS,
)
def test_python_replay_rejects_each_transitive_implementation_mutation(
    tmp_path: Path,
    assured_implementation_attack_output: Path,
    implementation_root: str,
    relative_path: str,
) -> None:
    copied_plugin, copied_assurance = _copy_journal_implementation_tree(tmp_path)
    root = copied_plugin if implementation_root == "plugin" else copied_assurance
    _mutate_implementation_bytes(root / relative_path)

    completed = subprocess.run(
        [
            sys.executable,
            "-I",
            "-B",
            str(copied_plugin / "scripts" / "review_successor.py"),
            "validate",
            str(assured_implementation_attack_output),
        ],
        capture_output=True,
        text=True,
        check=False,
        timeout=20,
    )

    assert completed.returncode != 0


@pytest.mark.parametrize(
    ("implementation_root", "relative_path"),
    TRANSITIVE_IMPLEMENTATION_ATTACKS,
)
def test_mcp_replay_rejects_each_transitive_implementation_mutation(
    tmp_path: Path,
    assured_implementation_attack_output: Path,
    implementation_root: str,
    relative_path: str,
) -> None:
    copied_plugin, copied_assurance = _copy_journal_implementation_tree(tmp_path)
    root = copied_plugin if implementation_root == "plugin" else copied_assurance
    _mutate_implementation_bytes(root / relative_path)

    result = _journal_transaction_call(
        "validate_journal_sampling_review",
        _implementation_attack_arguments(assured_implementation_attack_output),
        server_path=copied_plugin / "mcp" / "server.cjs",
        env={"JOURNAL_SAMPLING_PYTHON": sys.executable},
    )

    assert result == {
        "ok": False,
        "error": "Journal Sampling assurance validate failed.",
    }


def test_real_journal_review_save_mints_replayable_limited_successor(
    tmp_path: Path,
) -> None:
    core = load_core()
    output_dir, arguments = _real_journal_review_case(core, tmp_path)
    initial = core.validate_sample_assurance(output_dir)

    result = _journal_transaction_call(
        "save_journal_sampling_decisions",
        arguments,
    )

    replay = core.validate_sample_assurance(output_dir)
    gates = replay["assurance_envelope"]["gate_register"]
    assert result["ok"] is True
    assert result["status"] == "reviewed"
    assert replay["output_set"]["stage"]["index"] == 1
    assert replay["output_set"]["stage"]["kind"] == "save"
    assert (
        replay["output_set"]["stage"]["predecessor"]["manifest_sha256"]
        == initial["output_set"]["content_sha256"]
    )
    assert replay["review_successor"]["kind"] == "save"
    assert replay["review_successor"]["applied_decisions"] is None
    assert not (output_dir / "applied_decisions.json").exists()
    assert gates["gates"]["semantic_review"]["status"] == "not_assessed"
    assert gates["gates"]["reporting"]["status"] == "blocked"
    assert gates["gates"]["publication"]["status"] == "withheld"
    assert gates["report_ready"] is False
    assert len(replay["output_set"]["physical_paths"]) == 26
    assert not list(tmp_path.glob(".generated-review-transaction-*"))


def test_real_journal_review_apply_after_save_preserves_exact_history_chain(
    tmp_path: Path,
) -> None:
    core = load_core()
    output_dir, arguments = _real_journal_review_case(core, tmp_path)
    decisions = arguments["decisions"]
    saved = _journal_transaction_call(
        "save_journal_sampling_decisions",
        arguments,
    )
    assert saved["ok"] is True
    initial_history = _journal_tree_image(
        output_dir / "assurance_history" / "000_initial"
    )
    apply_arguments = _current_real_journal_review_arguments(
        output_dir,
        decisions,
    )

    result = _journal_transaction_call(
        "apply_journal_sampling_decisions",
        apply_arguments,
    )

    replay = core.validate_sample_assurance(output_dir)
    gates = replay["assurance_envelope"]["gate_register"]
    assert result["ok"] is True
    assert result["application_status"] == "review_applied_with_assurance_limits"
    assert replay["output_set"]["stage"]["index"] == 2
    assert replay["output_set"]["stage"]["kind"] == "apply"
    assert {path.name for path in (output_dir / "assurance_history").iterdir()} == {
        "000_initial",
        "001_save",
    }
    assert (
        _journal_tree_image(output_dir / "assurance_history" / "000_initial")
        == initial_history
    )
    assert len(replay["output_set"]["physical_paths"]) == 40
    assert gates["gates"]["semantic_review"]["status"] == "not_assessed"
    assert gates["gates"]["reporting"]["status"] == "blocked"
    assert gates["gates"]["publication"]["status"] == "withheld"
    assert gates["report_ready"] is False


def test_real_journal_review_apply_regresses_supplied_successor_gap(
    tmp_path: Path,
) -> None:
    core = load_core()
    output_dir, arguments = _real_journal_review_case(core, tmp_path)
    before = core.validate_sample_assurance(output_dir)

    result = _journal_transaction_call(
        "apply_journal_sampling_decisions",
        arguments,
    )

    replay = core.validate_sample_assurance(output_dir)
    assert before["assurance_envelope"]["gate_register"]["report_ready"] is False
    assert result["ok"] is True
    assert result["application_status"] == "review_applied_with_assurance_limits"
    assert replay["output_set"]["stage"] == {
        "index": 1,
        "kind": "apply",
        "predecessor": {
            "stage_index": 0,
            "stage_kind": "initial",
            "archive_dir": "assurance_history/000_initial",
            "manifest_sha256": before["output_set"]["content_sha256"],
        },
    }
    assert replay["assurance_envelope"]["gate_register"]["report_ready"] is False


def test_real_journal_review_apply_closes_revision_file_and_directory(
    tmp_path: Path,
) -> None:
    core = load_core()
    output_dir, arguments = _real_journal_review_case(core, tmp_path)
    arguments["decisions"] = [
        {
            "item_id": item["id"],
            "action": "edit" if item["id"] == "sampled-entry-1" else "accept",
            **(
                {"edit_value": "Reviewer-authored revision"}
                if item["id"] == "sampled-entry-1"
                else {}
            ),
        }
        for item in arguments["review_payload"]["items"]
    ]

    result = _journal_transaction_call(
        "apply_journal_sampling_decisions",
        arguments,
    )

    replay = core.validate_sample_assurance(output_dir)
    applied = json.loads(
        (output_dir / "applied_decisions.json").read_text(encoding="utf-8")
    )
    assert result["ok"] is True
    assert applied["revision_count"] == 1
    assert applied["revision_paths"] == [
        "revisions/journal_sample__sampled-entry-1.txt"
    ]
    assert (output_dir / applied["revision_paths"][0]).read_text(
        encoding="utf-8"
    ) == "Reviewer-authored revision"
    assert "revisions" in replay["output_set"]["directory_paths"]
    assert applied["revision_paths"][0] in replay["output_set"]["physical_paths"]
    assert replay["review_successor"]["assurance_limits"]["report_ready"] is False


def test_real_journal_review_apply_rederives_blocked_successor(
    tmp_path: Path,
) -> None:
    core = load_core()
    output_dir, arguments = _real_journal_review_case(core, tmp_path)
    arguments["decisions"] = [
        {
            "item_id": item["id"],
            "action": (
                "mark_unclear" if item["id"] == "sampling-control" else "accept"
            ),
        }
        for item in arguments["review_payload"]["items"]
    ]

    result = _journal_transaction_call(
        "apply_journal_sampling_decisions",
        arguments,
    )

    replay = core.validate_sample_assurance(output_dir)
    final_artifacts = json.loads(
        (output_dir / "final_artifacts.json").read_text(encoding="utf-8")
    )
    assert result["ok"] is True
    assert result["application_status"] == "blocked"
    assert result["blocker_count"] == 1
    assert len(final_artifacts["blockers"]) == 1
    assert final_artifacts["status"] == "blocked"
    assert (
        replay["assurance_envelope"]["gate_register"]["gates"]["reporting"]["status"]
        == "blocked"
    )
    assert replay["assurance_envelope"]["gate_register"]["report_ready"] is False


@pytest.mark.parametrize(
    "attack_kind",
    [
        "missing_file",
        "rogue_file",
        "empty_directory",
        "nested_empty_directory",
        "symlink",
        "hardlink",
        "fifo",
        "payload_mode",
        "manifest_mode",
        "root_mode",
    ],
)
def test_sample_output_contract_rejects_every_physical_attack(
    tmp_path: Path,
    attack_kind: str,
) -> None:
    core = load_core()
    output_dir, _ = _real_journal_review_case(core, tmp_path)
    external = tmp_path / "external.bin"
    external.write_bytes(b"external")
    if attack_kind == "missing_file":
        (output_dir / "review_handoff.md").unlink()
    elif attack_kind == "rogue_file":
        (output_dir / "rogue.bin").write_bytes(b"rogue")
    elif attack_kind == "empty_directory":
        (output_dir / "empty").mkdir()
    elif attack_kind == "nested_empty_directory":
        (output_dir / "rogue" / "empty").mkdir(parents=True)
    elif attack_kind == "symlink":
        (output_dir / "rogue-link").symlink_to(external)
    elif attack_kind == "hardlink":
        os.link(output_dir / "journal_sample.csv", output_dir / "rogue-hardlink")
    elif attack_kind == "fifo":
        os.mkfifo(output_dir / "rogue-fifo")
    elif attack_kind == "payload_mode":
        target = output_dir / "journal_sample.csv"
        target.chmod((target.stat().st_mode & 0o7777) ^ 0o100)
    elif attack_kind == "manifest_mode":
        (output_dir / core.SAMPLE_OUTPUT_SET_PATH).chmod(0o644)
    else:
        output_dir.chmod((output_dir.stat().st_mode & 0o7777) ^ 0o040)

    with pytest.raises(ValueError):
        core.validate_sample_output_set(output_dir)


@pytest.mark.parametrize(
    "attack_kind",
    [
        "archived_bytes",
        "rogue_file",
        "empty_directory",
        "nested_empty_directory",
        "symlink",
        "hardlink",
        "fifo",
        "archived_file_mode",
        "archive_root_mode",
        "predecessor_binding",
        "archived_manifest_fields",
    ],
)
def test_successor_replay_rejects_every_predecessor_history_attack(
    tmp_path: Path,
    attack_kind: str,
) -> None:
    core = load_core()
    output_dir, arguments = _real_journal_review_case(core, tmp_path)
    result = _journal_transaction_call(
        "apply_journal_sampling_decisions",
        arguments,
    )
    assert result["ok"] is True
    archive = output_dir / "assurance_history" / "000_initial"
    external = tmp_path / "external.bin"
    external.write_bytes(b"external")
    if attack_kind == "archived_bytes":
        target = archive / "review_handoff.md"
        target.write_bytes(target.read_bytes() + b"forged")
    elif attack_kind == "rogue_file":
        (archive / "rogue.bin").write_bytes(b"rogue")
    elif attack_kind == "empty_directory":
        (archive / "empty").mkdir()
    elif attack_kind == "nested_empty_directory":
        (archive / "rogue" / "empty").mkdir(parents=True)
    elif attack_kind == "symlink":
        (archive / "rogue-link").symlink_to(external)
    elif attack_kind == "hardlink":
        os.link(archive / "journal_sample.csv", archive / "rogue-hardlink")
    elif attack_kind == "fifo":
        os.mkfifo(archive / "rogue-fifo")
    elif attack_kind == "archived_file_mode":
        target = archive / "journal_sample.csv"
        target.chmod((target.stat().st_mode & 0o7777) ^ 0o100)
    elif attack_kind == "archive_root_mode":
        archive.chmod((archive.stat().st_mode & 0o7777) ^ 0o040)
    elif attack_kind == "predecessor_binding":
        manifest_path = output_dir / core.SAMPLE_OUTPUT_SET_PATH
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        manifest["stage"]["predecessor"]["manifest_sha256"] = "0" * 64
        content = {
            key: value for key, value in manifest.items() if key != "content_sha256"
        }
        manifest["content_sha256"] = core.canonical_json_sha256(content)
        core.write_json(manifest_path, manifest)
        manifest_path.chmod(core.SAMPLE_OUTPUT_SET_MODE)
    else:
        manifest_path = archive / core.SAMPLE_OUTPUT_SET_PATH
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        manifest["forged_extra"] = True
        content = {
            key: value for key, value in manifest.items() if key != "content_sha256"
        }
        manifest["content_sha256"] = core.canonical_json_sha256(content)
        core.write_json(manifest_path, manifest)
        manifest_path.chmod(core.SAMPLE_OUTPUT_SET_MODE)

    with pytest.raises(ValueError):
        core.validate_sample_assurance(output_dir)


@pytest.mark.parametrize(
    "field_name",
    ["decision_count", "item_count", "blocker_count", "effects", "application_status"],
)
def test_successor_replay_rejects_stale_applied_material_fields(
    tmp_path: Path,
    field_name: str,
) -> None:
    core = load_core()
    output_dir, arguments = _real_journal_review_case(core, tmp_path)
    result = _journal_transaction_call(
        "apply_journal_sampling_decisions",
        arguments,
    )
    assert result["ok"] is True
    applied_path = output_dir / "applied_decisions.json"
    applied = json.loads(applied_path.read_text(encoding="utf-8"))
    if field_name == "effects":
        applied[field_name][0]["requires_followup"] = True
    elif field_name == "application_status":
        applied[field_name] = "final_ready"
    else:
        applied[field_name] = int(applied[field_name]) + 1
    core.write_json(applied_path, applied)
    _reseal_mutated_applied_receipts(core, output_dir)

    with pytest.raises(ValueError):
        core.validate_sample_assurance(output_dir)


def test_assured_mcp_refuses_to_archive_a_rogue_predecessor_file(
    tmp_path: Path,
) -> None:
    core = load_core()
    output_dir, arguments = _real_journal_review_case(core, tmp_path)
    rogue = output_dir / "private-unreceipted.bin"
    rogue.write_bytes(b"must-not-be-archived")
    before = _journal_tree_image(output_dir)

    result = _journal_transaction_call(
        "apply_journal_sampling_decisions",
        arguments,
    )

    assert result["ok"] is False
    assert _journal_tree_image(output_dir) == before
    assert not (output_dir / "assurance_history").exists()
    assert not list(tmp_path.glob(".generated-review-transaction-*"))


def test_assured_mcp_render_requires_fresh_whole_tree_replay(
    tmp_path: Path,
) -> None:
    core = load_core()
    output_dir, arguments = _real_journal_review_case(core, tmp_path)
    reproducibility_path = output_dir / "sample_reproducibility.json"
    reproducibility_path.write_bytes(reproducibility_path.read_bytes() + b" ")

    result = _journal_transaction_call(
        "render_journal_sampling_review",
        arguments,
    )

    assert result == {
        "ok": False,
        "error": "Journal Sampling assurance validate failed.",
    }


def test_journal_review_transaction_honest_apply_commits_without_residue(
    tmp_path: Path,
) -> None:
    output_dir = tmp_path / "output"
    arguments = _journal_transaction_case(output_dir)

    result = _journal_transaction_call(
        "apply_journal_sampling_decisions",
        arguments,
    )

    assert result["ok"] is True
    assert result["persisted"] is True
    assert result["application_status"] == "review_applied_with_assurance_limits"
    assert (output_dir / "applied_decisions.json").is_file()
    assert (output_dir / "nested" / "sentinel.bin").read_bytes() == (
        b"\x00journal-original\xff"
    )
    assert output_dir.stat().st_mode & 0o7777 == 0o750
    assert (output_dir / "nested").stat().st_mode & 0o7777 == 0o711
    assert not list(tmp_path.glob(".generated-review-transaction-*"))


def test_journal_review_transaction_rejects_forged_caller_review_target(
    tmp_path: Path,
) -> None:
    output_dir = tmp_path / "output"
    arguments = _journal_transaction_case(output_dir)
    target = output_dir / "nested" / "controlled.txt"
    target.write_bytes(b"ORIGINAL-CANONICAL-BYTES")
    target.chmod(0o640)
    before = _journal_tree_image(output_dir)
    forged_arguments = json.loads(json.dumps(arguments))
    forged_item = forged_arguments["review_payload"]["items"][0]
    forged_item.update(
        {
            "title": "Forged writable item",
            "allowed_actions": ["edit"],
            "recommended_action": "edit",
            "data": {"target_artifact": "nested/controlled.txt"},
        }
    )
    forged_arguments["decisions"] = [
        {
            "item_id": forged_item["id"],
            "action": "edit",
            "edit_value": "FORGED-CALLER-WRITE",
        }
    ]

    result = _journal_transaction_call(
        "apply_journal_sampling_decisions",
        forged_arguments,
    )

    assert result == {
        "ok": False,
        "error": (
            "Caller review payload does not match the persisted "
            "Journal Sampling review payload."
        ),
    }
    assert target.read_bytes() == b"ORIGINAL-CANONICAL-BYTES"
    assert target.stat().st_mode & 0o7777 == 0o640
    assert _journal_tree_image(output_dir) == before
    assert not list(tmp_path.glob(".generated-review-transaction-*"))


@pytest.mark.parametrize(
    ("tool_name", "context_name", "expected_error"),
    [
        (
            "save_journal_sampling_decisions",
            "run_intake",
            (
                "Caller run intake does not match the persisted "
                "Journal Sampling run intake."
            ),
        ),
        (
            "apply_journal_sampling_decisions",
            "run_intake",
            (
                "Caller run intake does not match the persisted "
                "Journal Sampling run intake."
            ),
        ),
        (
            "save_journal_sampling_decisions",
            "final_artifacts",
            (
                "Caller final artifacts do not match the persisted "
                "Journal Sampling final artifacts."
            ),
        ),
        (
            "apply_journal_sampling_decisions",
            "final_artifacts",
            (
                "Caller final artifacts do not match the persisted "
                "Journal Sampling final artifacts."
            ),
        ),
        (
            "save_journal_sampling_decisions",
            "ui_decisions",
            (
                "Caller UI decisions do not match the persisted "
                "Journal Sampling UI decisions."
            ),
        ),
        (
            "apply_journal_sampling_decisions",
            "ui_decisions",
            (
                "Caller UI decisions do not match the persisted "
                "Journal Sampling UI decisions."
            ),
        ),
    ],
)
def test_journal_review_transaction_rejects_forged_caller_context(
    tmp_path: Path,
    tool_name: str,
    context_name: str,
    expected_error: str,
) -> None:
    output_dir = tmp_path / "output"
    arguments = _journal_transaction_case(output_dir)
    if context_name == "ui_decisions":
        persisted_ui_decisions = {
            "schema_version": "1.0",
            "plugin": "journal-sampling",
            "workflow": "journal-sampling",
            "run_id": "journal-transaction-run",
            "decisions": [],
            "status": "pending_review",
        }
        (output_dir / "ui_decisions.json").write_text(
            json.dumps(persisted_ui_decisions, indent=2) + "\n",
            encoding="utf-8",
        )
        arguments["ui_decisions"] = {
            **persisted_ui_decisions,
            "status": "forged",
        }
    else:
        arguments[context_name] = {
            **arguments[context_name],
            "forged_caller_field": True,
        }
    before = _journal_tree_image(output_dir)

    result = _journal_transaction_call(tool_name, arguments)

    assert result == {"ok": False, "error": expected_error}
    assert _journal_tree_image(output_dir) == before
    assert not list(tmp_path.glob(".generated-review-transaction-*"))


@pytest.mark.parametrize(
    ("tool_name", "needle"),
    [
        (
            "save_journal_sampling_decisions",
            "      const workingResult = saveDecisionPayloadWrites(workingArgs);\n",
        ),
        (
            "apply_journal_sampling_decisions",
            "      const workingResult = applyDecisionPayloadWrites(workingArgs);\n",
        ),
    ],
)
def test_journal_review_transaction_late_failure_restores_bytes_and_modes(
    tmp_path: Path,
    tool_name: str,
    needle: str,
) -> None:
    output_dir = tmp_path / "output"
    arguments = _journal_transaction_case(output_dir)
    before = _journal_tree_image(output_dir)
    faulted = _journal_faulted_server(
        tmp_path,
        needle=needle,
        replacement=(
            needle + '      throw new Error("/private/client/journal-late-failure");\n'
        ),
    )

    result = _journal_transaction_call(
        tool_name,
        arguments,
        server_path=faulted,
    )

    assert result["ok"] is False
    assert result["error"] == (
        "Journal Sampling review "
        + ("save" if tool_name.startswith("save") else "apply")
        + " transaction failed safely."
    )
    assert "/private/client" not in result["error"]
    assert _journal_tree_image(output_dir) == before
    assert not list(tmp_path.glob(".generated-review-transaction-*"))


def test_journal_review_transaction_rejects_forged_save_response_contract(
    tmp_path: Path,
) -> None:
    output_dir = tmp_path / "output"
    arguments = _journal_transaction_case(output_dir)
    before = _journal_tree_image(output_dir)
    needle = "      const workingResult = saveDecisionPayloadWrites(workingArgs);\n"
    faulted = _journal_faulted_server(
        tmp_path,
        needle=needle,
        replacement=(needle + """
      Object.assign(workingResult, {
        validation_type: "forged_save",
        run_id: "forged-run",
        decision_count: 777,
        item_count: 888,
        status: "forged_status",
        ui_decisions_path: "/private/client/forged-ui.json",
        message: "forged message",
      });
"""),
    )

    result = _journal_transaction_call(
        "save_journal_sampling_decisions",
        arguments,
        server_path=faulted,
    )

    assert result == {
        "ok": False,
        "error": "Journal Sampling saved decisions did not close.",
    }
    assert _journal_tree_image(output_dir) == before
    assert not list(tmp_path.glob(".generated-review-transaction-*"))


def test_journal_review_transaction_rejects_forged_apply_response_contract(
    tmp_path: Path,
) -> None:
    output_dir = tmp_path / "output"
    arguments = _journal_transaction_case(output_dir)
    before = _journal_tree_image(output_dir)
    needle = "      const workingResult = applyDecisionPayloadWrites(workingArgs);\n"
    faulted = _journal_faulted_server(
        tmp_path,
        needle=needle,
        replacement=(needle + """
      Object.assign(workingResult, {
        validation_type: "forged_apply",
        run_id: "forged-run",
        decision_count: 101,
        item_count: 102,
        blocker_count: 103,
        revision_count: 104,
        target_update_count: 105,
        structured_update_count: 106,
        native_regeneration_count: 107,
        native_regenerated_count: 108,
        application_status: "final_ready",
        ui_decisions_path: "/private/client/forged-ui.json",
        applied_decisions_path: "/private/client/forged-applied.json",
        final_artifacts_path: "/private/client/forged-final.json",
        run_intake_path: "/private/client/forged-intake.json",
        message: "forged message",
      });
"""),
    )

    result = _journal_transaction_call(
        "apply_journal_sampling_decisions",
        arguments,
        server_path=faulted,
    )

    assert result == {
        "ok": False,
        "error": "Journal Sampling response did not close.",
    }
    assert _journal_tree_image(output_dir) == before
    assert not list(tmp_path.glob(".generated-review-transaction-*"))


@pytest.mark.parametrize("self_authorized_path", ["rogue.json", "ui_decisions.json"])
def test_journal_review_transaction_rejects_persisted_result_self_authorization(
    tmp_path: Path,
    self_authorized_path: str,
) -> None:
    output_dir = tmp_path / "output"
    arguments = _journal_transaction_case(output_dir)
    before = _journal_tree_image(output_dir)
    needle = "      const workingResult = applyDecisionPayloadWrites(workingArgs);\n"
    injected = """
      const selfAuthorizedPath = process.env.REVIEW_TX_SELF_AUTHORIZED_PATH;
      if (selfAuthorizedPath === "rogue.json") {
        const roguePath = path.join(workingOutputDir, selfAuthorizedPath);
        fs.writeFileSync(roguePath, "{\\"forged\\":true}\\n", "utf8");
      }
      const appliedPath = path.join(workingOutputDir, "applied_decisions.json");
      const finalPath = path.join(workingOutputDir, "final_artifacts.json");
      const applied = JSON.parse(fs.readFileSync(appliedPath, "utf8"));
      const finalArtifacts = JSON.parse(fs.readFileSync(finalPath, "utf8"));
      applied.native_regenerated_count = 1;
      applied.native_regenerated_paths = [selfAuthorizedPath];
      finalArtifacts.review_application.native_regenerated_count = 1;
      finalArtifacts.review_application.native_regenerated_paths = [selfAuthorizedPath];
      fs.writeFileSync(appliedPath, `${JSON.stringify(applied, null, 2)}\\n`, "utf8");
      fs.writeFileSync(finalPath, `${JSON.stringify(finalArtifacts, null, 2)}\\n`, "utf8");
      workingResult.native_regenerated_count = 1;
      workingResult.applied_decisions = applied;
      workingResult.final_artifacts = finalArtifacts;
"""
    faulted = _journal_faulted_server(
        tmp_path,
        needle=needle,
        replacement=needle + injected,
    )

    result = _journal_transaction_call(
        "apply_journal_sampling_decisions",
        arguments,
        server_path=faulted,
        env={"REVIEW_TX_SELF_AUTHORIZED_PATH": self_authorized_path},
    )

    assert result["ok"] is False
    assert "/" not in result["error"]
    assert "\\" not in result["error"]
    assert self_authorized_path not in result["error"]
    assert _journal_tree_image(output_dir) == before
    assert not list(tmp_path.glob(".generated-review-transaction-*"))


@pytest.mark.parametrize("attack_kind", ["symlink", "hardlink", "fifo"])
def test_journal_review_transaction_rejects_working_tree_poison(
    tmp_path: Path,
    attack_kind: str,
) -> None:
    output_dir = tmp_path / "output"
    arguments = _journal_transaction_case(output_dir)
    before = _journal_tree_image(output_dir)
    external = tmp_path / "external-target.bin"
    external.write_bytes(b"EXTERNAL-UNCHANGED")
    external.chmod(0o600)
    needle = "      const workingResult = applyDecisionPayloadWrites(workingArgs);\n"
    injected = """
      const poisonPath = path.join(workingOutputDir, "ui_decisions.json");
      fs.unlinkSync(poisonPath);
      if (process.env.REVIEW_TX_ATTACK_KIND === "symlink") {
        fs.symlinkSync(process.env.REVIEW_TX_EXTERNAL, poisonPath);
      } else if (process.env.REVIEW_TX_ATTACK_KIND === "hardlink") {
        fs.linkSync(process.env.REVIEW_TX_EXTERNAL, poisonPath);
      } else {
        require("node:child_process").spawnSync("/usr/bin/mkfifo", [poisonPath]);
      }
"""
    faulted = _journal_faulted_server(
        tmp_path,
        needle=needle,
        replacement=needle + injected,
    )

    result = _journal_transaction_call(
        "apply_journal_sampling_decisions",
        arguments,
        server_path=faulted,
        env={
            "REVIEW_TX_ATTACK_KIND": attack_kind,
            "REVIEW_TX_EXTERNAL": str(external),
        },
    )

    assert result["ok"] is False
    assert "/" not in result["error"]
    assert external.read_bytes() == b"EXTERNAL-UNCHANGED"
    assert external.stat().st_mode & 0o7777 == 0o600
    assert _journal_tree_image(output_dir) == before
    assert not list(tmp_path.glob(".generated-review-transaction-*"))


def test_journal_review_transaction_rejects_transaction_root_relocation_without_moving_canonical(
    tmp_path: Path,
) -> None:
    output_dir = tmp_path / "output"
    arguments = _journal_transaction_case(output_dir)
    before = _journal_tree_image(output_dir)
    canonical_inode = output_dir.stat().st_ino
    needle = "      const workingResult = applyDecisionPayloadWrites(workingArgs);\n"
    injected = """
      const transactionRoot = path.dirname(workingOutputDir);
      fs.renameSync(transactionRoot, `${transactionRoot}-moved`);
      throw new Error("/private/client/transaction-root-relocation");
"""
    faulted = _journal_faulted_server(
        tmp_path,
        needle=needle,
        replacement=needle + injected,
    )

    result = _journal_transaction_call(
        "apply_journal_sampling_decisions",
        arguments,
        server_path=faulted,
    )

    assert result["ok"] is False
    assert result["error"] == (
        "Journal Sampling review apply transaction failed safely."
    )
    assert output_dir.stat().st_ino == canonical_inode
    assert _journal_tree_image(output_dir) == before
    assert not list(tmp_path.glob(".generated-review-transaction-*"))
    assert not list(tmp_path.glob(".generated-review-commit-*"))
    assert not list(tmp_path.glob(".generated-review-recovery-*"))


def test_journal_review_transaction_restores_after_commit_deletion(
    tmp_path: Path,
) -> None:
    output_dir = tmp_path / "output"
    arguments = _journal_transaction_case(output_dir)
    before = _journal_tree_image(output_dir)
    needle = "    committed = true;\n    const committedImage ="
    faulted = _journal_faulted_server(
        tmp_path,
        needle=needle,
        replacement=(
            "    committed = true;\n"
            "    generatedReviewRemoveExactPath(resolvedOutputDir);\n"
            "    const committedImage ="
        ),
    )

    result = _journal_transaction_call(
        "apply_journal_sampling_decisions",
        arguments,
        server_path=faulted,
    )

    assert result["ok"] is False
    assert "/" not in result["error"]
    assert _journal_tree_image(output_dir) == before
    assert not list(tmp_path.glob(".generated-review-transaction-*"))


def test_journal_review_transaction_enforces_size_bound_before_mutation(
    tmp_path: Path,
) -> None:
    output_dir = tmp_path / "output"
    arguments = _journal_transaction_case(output_dir)
    sentinel = output_dir / "nested" / "sentinel.bin"
    oversized = output_dir / "oversized.bin"
    with oversized.open("wb") as stream:
        stream.truncate(128 * 1024 * 1024 + 1)
    root_inode = output_dir.stat().st_ino
    sentinel_inode = sentinel.stat().st_ino
    sentinel_bytes = sentinel.read_bytes()

    result = _journal_transaction_call(
        "save_journal_sampling_decisions",
        arguments,
    )

    assert result["ok"] is False
    assert output_dir.stat().st_ino == root_inode
    assert sentinel.stat().st_ino == sentinel_inode
    assert sentinel.read_bytes() == sentinel_bytes
    assert oversized.stat().st_size == 128 * 1024 * 1024 + 1
    assert not list(tmp_path.glob(".generated-review-transaction-*"))
