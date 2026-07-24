from __future__ import annotations

import copy
import hashlib
import importlib.util
import json
import sys
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace
from typing import Any, cast
from unittest.mock import patch

import jsonschema
import pytest
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[2]
CLARA_ROOT = ROOT / "plugins" / "clara"
SCRIPTS_ROOT = CLARA_ROOT / "scripts"
CASE_SCHEMA_PATH = (
    CLARA_ROOT / "contracts" / "real_general_ledger_preparation_case.v1.schema.json"
)
SEMANTIC_DECISIONS_SCHEMA_PATH = (
    CLARA_ROOT / "contracts" / "real_general_ledger_semantic_decisions.v1.schema.json"
)
VALIDATION_DATE = "2026-07-24"
PILOT_ID = "pilot-0123456789abcdef"
EXECUTION_ID = "execution-fedcba9876543210"
SOURCE_ID = "source-0123456789abcdef"
SOURCE_BYTES = b"synthetic general-journal fixture; no commercial data\n"


def _load_module(name: str, path: Path) -> Any:
    scripts_path = str(SCRIPTS_ROOT)
    inserted = scripts_path not in sys.path
    if inserted:
        sys.path.insert(0, scripts_path)
    try:
        spec = importlib.util.spec_from_file_location(name, path)
        assert spec and spec.loader
        module = importlib.util.module_from_spec(spec)
        sys.modules[spec.name] = module
        spec.loader.exec_module(module)
        return module
    finally:
        if inserted:
            sys.path.remove(scripts_path)


INTAKE = _load_module(
    "validate_real_data_pilot_intake",
    SCRIPTS_ROOT / "validate_real_data_pilot_intake.py",
)
SEMANTIC = _load_module(
    "validate_real_data_pilot_semantic_review",
    SCRIPTS_ROOT / "validate_real_data_pilot_semantic_review.py",
)
PRODUCER = _load_module(
    "clara_real_general_ledger_pilot_test",
    SCRIPTS_ROOT / "run_real_general_ledger_pilot.py",
)


@dataclass
class _Fixture:
    root: Path
    case_path: Path
    intake_contract_path: Path
    intake_receipt_path: Path
    semantic_review_path: Path
    semantic_receipt_path: Path
    semantic_decisions_path: Path
    source_path: Path
    parser_layout_path: Path
    movements: list[Any]


def _write_json(path: Path, payload: object) -> None:
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _intake_contract(source_sha256: str, byte_count: int) -> dict[str, Any]:
    return {
        "schema_version": "clara.real_data_pilot_intake.v2",
        "pilot_id": PILOT_ID,
        "purpose": "local_due_diligence_preparation_evaluation",
        "source": {
            "source_id": SOURCE_ID,
            "data_kind": "commercial_general_ledger",
            "data_classification": "consented_real",
            "media_type": "text/csv",
            "byte_count": byte_count,
            "sha256": source_sha256,
        },
        "authorization": {
            "status": "reviewed",
            "basis": "explicit_authorized_user_instruction",
            "authority_assertion": "authorizer_has_right_to_permit_this_use",
            "evidence_reference": "Synthetic unit-test declaration.",
            "authorizing_role": "unit-test fixture",
            "authorized_on": VALIDATION_DATE,
            "valid_from": VALIDATION_DATE,
            "valid_until": None,
            "purpose": "local_due_diligence_preparation_evaluation",
            "authorized_source_sha256": source_sha256,
            "permitted_actions": [
                "codex_model_processing",
                "local_deterministic_processing",
            ],
            "prohibited_actions": [
                "commit_raw_or_row_level_data",
                "package_raw_or_row_level_data",
                "publish_raw_or_row_level_data",
            ],
            "terms_summary": (
                "Synthetic unit-test declaration; no real permission is claimed."
            ),
        },
        "privacy": {
            "codex_context_acknowledged": True,
            "automatic_anonymization_claimed": False,
            "clara_external_recipient_added": False,
            "raw_and_row_level_storage": "local_run_root_only",
            "repository_recording_policy": "sanitized_summary_and_receipts_only",
        },
        "deidentification_review": {
            "status": "not_applicable",
            "basis": "Unit-test contract uses no real data.",
            "reidentification_risk_review_status": "not_applicable",
        },
        "semantic_review_plan": {
            "status": "pending",
            "required_reviews": list(INTAKE.REQUIRED_SEMANTIC_REVIEWS),
            "automatic_mapping_allowed": False,
            "unresolved_blocking_issues_block_preparation": True,
        },
        "publication_status": "withheld",
        "report_ready": False,
    }


def _evidence_registry(
    tmp_path: Path,
    *,
    semantic_decisions_path: Path,
) -> dict[str, dict[str, Any]]:
    evidence_root = tmp_path / "evidence"
    evidence_root.mkdir()
    registry: dict[str, dict[str, Any]] = {}
    for position in range(1, 9):
        evidence_id = f"evidence-{position:02d}"
        payload = f"Synthetic evidence {position}.\n".encode()
        relative_path = Path("evidence") / f"{evidence_id}.txt"
        (tmp_path / relative_path).write_bytes(payload)
        registry[evidence_id] = {
            "path": relative_path.as_posix(),
            "media_type": "text/plain",
            "byte_count": len(payload),
            "sha256": hashlib.sha256(payload).hexdigest(),
        }
    registry["evidence-semantic-decisions"] = {
        "path": semantic_decisions_path.relative_to(tmp_path).as_posix(),
        "media_type": "application/json",
        "byte_count": semantic_decisions_path.stat().st_size,
        "sha256": _sha256(semantic_decisions_path),
    }
    return registry


def _semantic_review(
    *,
    intake_receipt_sha256: str,
    evidence_registry: dict[str, dict[str, Any]],
    blocking: bool,
) -> dict[str, Any]:
    required_reviews = [
        {
            "review_id": f"topic-review-{position:02d}",
            "topic": topic,
            "status": "reviewed",
            "decision": f"Synthetic decision {position}.",
            "basis": f"Synthetic basis {position}.",
            "evidence_refs": [
                f"evidence-{position:02d}",
                "evidence-semantic-decisions",
            ],
        }
        for position, topic in enumerate(
            INTAKE.REQUIRED_SEMANTIC_REVIEWS,
            start=1,
        )
    ]
    issues = (
        {
            "issue-01": {
                "topic": "account_mapping",
                "status": "open",
                "blocking": True,
                "description": "Synthetic blocking issue.",
                "basis": "Synthetic issue basis.",
                "evidence_refs": ["evidence-08"],
                "resolution": None,
            }
        }
        if blocking
        else {}
    )
    return {
        "schema_version": "clara.real_data_pilot_semantic_review.v1",
        "pilot_id": PILOT_ID,
        "review_version": "review-fedcba9876543210",
        "review_status": "reviewed",
        "reviewed_on": VALIDATION_DATE,
        "reviewer_role": "synthetic unit-test reviewer",
        "intake_receipt_sha256": intake_receipt_sha256,
        "evidence_registry": evidence_registry,
        "required_reviews": required_reviews,
        "issues": issues,
        "mechanical_error_register_policy": "producer_owned_separate_artifact",
        "publication_status": "withheld",
        "report_ready": False,
    }


def _reviewed_decisions() -> dict[str, Any]:
    return {
        "account_identity": {
            "basis": "source_account_code",
            "normalization": "remove_whitespace_within_reviewed_account_code",
            "statement_mapping": "not_performed",
        },
        "period": {
            "basis": "posting_date",
            "calendar": "gregorian",
            "grain": "calendar_month",
            "date_assignment": "nondecreasing_carried_forward_posting_date",
        },
        "scope": {
            "dataset": "one_source_presented_export",
            "eliminations": "none_applied",
            "entity_basis": "export_level_source_scope_not_emitted",
            "consolidation_status": "not_assessed",
        },
        "currency_unit_fx": {
            "unit": "source_native_unit",
            "fx": "none",
            "currency": "not_asserted",
        },
        "sign_convention": "debit_positive_credit_negative",
        "control_basis": "exact_extracted_final_debit_and_credit_controls",
        "tolerance": "0",
        "output_grain": "source_account_x_calendar_month",
    }


def _reported_controls(
    *,
    debit: str,
    credit: str,
    expected_calendar_months: list[str],
) -> dict[str, Any]:
    return {
        "debit": debit,
        "credit": credit,
        "journal_balance_required": True,
        "monthly_balance_required": True,
        "expected_calendar_months": expected_calendar_months,
    }


def _semantic_decisions(
    *,
    debit_control: str,
    credit_control: str,
    expected_calendar_months: list[str],
) -> dict[str, Any]:
    return {
        "schema_version": "clara.real_general_ledger_semantic_decisions.v1",
        "pilot_id": PILOT_ID,
        "source_id": SOURCE_ID,
        "review_status": "reviewed",
        "reviewer_role": "synthetic unit-test reviewer",
        "reviewed_period": {"year": 2023},
        "reviewed_decisions": _reviewed_decisions(),
        "reported_controls": _reported_controls(
            debit=debit_control,
            credit=credit_control,
            expected_calendar_months=expected_calendar_months,
        ),
        "publication_status": "withheld",
        "report_ready": False,
    }


def _case_contract(
    fixture_paths: dict[str, Path],
    *,
    debit_control: str,
    credit_control: str,
    expected_calendar_months: list[str],
) -> dict[str, Any]:
    return {
        "schema_version": "clara.real_general_ledger_preparation_case.v1",
        "pilot_id": PILOT_ID,
        "execution_id": EXECUTION_ID,
        "case_status": "reviewed",
        "reviewer_role": "synthetic unit-test reviewer",
        "source": {
            "source_id": SOURCE_ID,
            "declared_data_kind": "commercial_general_ledger",
        },
        "bindings": {
            "source_sha256": _sha256(fixture_paths["source"]),
            "intake_receipt_sha256": _sha256(fixture_paths["intake_receipt"]),
            "semantic_review_receipt_sha256": _sha256(
                fixture_paths["semantic_receipt"]
            ),
            "semantic_decisions_sha256": _sha256(fixture_paths["semantic_decisions"]),
            "parser_layout_sha256": _sha256(fixture_paths["parser_layout"]),
            "parser_adapter_implementation_sha256": _sha256(
                SCRIPTS_ROOT / "run_real_general_ledger_pilot.py"
            ),
            "parser_implementation_sha256": _sha256(
                SCRIPTS_ROOT / "parse_commercial_general_journal.py"
            ),
        },
        "reviewed_period": {"year": 2023},
        "reviewed_decisions": _reviewed_decisions(),
        "reported_controls": _reported_controls(
            debit=debit_control,
            credit=credit_control,
            expected_calendar_months=expected_calendar_months,
        ),
        "publication_status": "withheld",
        "report_ready": False,
    }


def _fixture(
    tmp_path: Path,
    *,
    blocking: bool = False,
    debit_control: str = "12.5",
    credit_control: str = "12.5",
    expected_calendar_months: list[str] | None = None,
) -> _Fixture:
    expected_months = (
        ["2023-01", "2023-02"]
        if expected_calendar_months is None
        else expected_calendar_months
    )
    source_path = tmp_path / "source.csv"
    source_path.write_bytes(SOURCE_BYTES)
    intake_contract_path = tmp_path / "intake.json"
    _write_json(
        intake_contract_path,
        _intake_contract(_sha256(source_path), len(SOURCE_BYTES)),
    )
    intake_receipt = INTAKE.validate_real_data_pilot_intake_v2(
        intake_contract_path,
        source_path,
        as_of_date=VALIDATION_DATE,
        local_run_root=tmp_path,
        repository_root=ROOT,
    )
    intake_receipt_path = tmp_path / "intake_receipt.json"
    _write_json(intake_receipt_path, intake_receipt)

    parser_layout_path = tmp_path / "parser_layout.json"
    _write_json(
        parser_layout_path,
        {
            "schema_version": "synthetic.parser_layout.v1",
            "posting_date_column": 1,
            "account_code_column": 2,
            "debit_column": 3,
            "credit_column": 4,
        },
    )
    semantic_decisions_path = tmp_path / "semantic_decisions.json"
    _write_json(
        semantic_decisions_path,
        _semantic_decisions(
            debit_control=debit_control,
            credit_control=credit_control,
            expected_calendar_months=expected_months,
        ),
    )
    semantic_review_path = tmp_path / "semantic_review.json"
    _write_json(
        semantic_review_path,
        _semantic_review(
            intake_receipt_sha256=_sha256(intake_receipt_path),
            evidence_registry=_evidence_registry(
                tmp_path,
                semantic_decisions_path=semantic_decisions_path,
            ),
            blocking=blocking,
        ),
    )
    semantic_receipt = SEMANTIC.validate_real_data_pilot_semantic_review(
        semantic_review_path,
        intake_receipt_path,
        as_of_date=VALIDATION_DATE,
        local_run_root=tmp_path,
        repository_root=ROOT,
    )
    semantic_receipt_path = tmp_path / "semantic_receipt.json"
    _write_json(semantic_receipt_path, semantic_receipt)
    case_path = tmp_path / "case.json"
    _write_json(
        case_path,
        _case_contract(
            {
                "source": source_path,
                "intake_receipt": intake_receipt_path,
                "semantic_receipt": semantic_receipt_path,
                "semantic_decisions": semantic_decisions_path,
                "parser_layout": parser_layout_path,
            },
            debit_control=debit_control,
            credit_control=credit_control,
            expected_calendar_months=expected_months,
        ),
    )
    movements = [
        PRODUCER.Movement(
            "1",
            date(2023, 1, 5),
            "1000",
            Decimal("10"),
            Decimal("0"),
        ),
        PRODUCER.Movement(
            "2",
            date(2023, 1, 6),
            "2000",
            Decimal("0"),
            Decimal("10"),
        ),
        PRODUCER.Movement(
            "3",
            date(2023, 2, 7),
            "1000",
            Decimal("2.5"),
            Decimal("0"),
        ),
        PRODUCER.Movement(
            "4",
            date(2023, 2, 7),
            "1000",
            Decimal("0"),
            Decimal("2.5"),
        ),
    ]
    return _Fixture(
        root=tmp_path,
        case_path=case_path,
        intake_contract_path=intake_contract_path,
        intake_receipt_path=intake_receipt_path,
        semantic_review_path=semantic_review_path,
        semantic_receipt_path=semantic_receipt_path,
        semantic_decisions_path=semantic_decisions_path,
        source_path=source_path,
        parser_layout_path=parser_layout_path,
        movements=movements,
    )


def _refresh_semantic_review_and_case_bindings(fixture: _Fixture) -> None:
    review = json.loads(fixture.semantic_review_path.read_text(encoding="utf-8"))
    evidence = review["evidence_registry"]["evidence-semantic-decisions"]
    evidence["byte_count"] = fixture.semantic_decisions_path.stat().st_size
    evidence["sha256"] = _sha256(fixture.semantic_decisions_path)
    _write_json(fixture.semantic_review_path, review)
    receipt = SEMANTIC.validate_real_data_pilot_semantic_review(
        fixture.semantic_review_path,
        fixture.intake_receipt_path,
        as_of_date=VALIDATION_DATE,
        local_run_root=fixture.root,
        repository_root=ROOT,
    )
    _write_json(fixture.semantic_receipt_path, receipt)
    case = json.loads(fixture.case_path.read_text(encoding="utf-8"))
    case["bindings"]["semantic_decisions_sha256"] = _sha256(
        fixture.semantic_decisions_path
    )
    case["bindings"]["semantic_review_receipt_sha256"] = _sha256(
        fixture.semantic_receipt_path
    )
    _write_json(fixture.case_path, case)


def _parser(
    movements: list[Any],
    calls: list[Path] | None = None,
    *,
    source_control_debit: Decimal = Decimal("12.5"),
    source_control_credit: Decimal = Decimal("12.5"),
) -> Any:
    def parse(
        source_path: Path,
        *,
        expected_source_sha256: str,
        layout_contract: object,
    ) -> Any:
        assert expected_source_sha256 == _sha256(source_path)
        assert layout_contract is not None
        if calls is not None:
            calls.append(source_path)
        return SimpleNamespace(
            source_sha256=expected_source_sha256,
            movements=tuple(
                SimpleNamespace(
                    line_id=int(movement.movement_id),
                    posting_date=movement.posting_date,
                    account_code=movement.source_account_code,
                    debit=movement.debit,
                    credit=movement.credit,
                )
                for movement in movements
            ),
            source_control_debit_total=source_control_debit,
            source_control_credit_total=source_control_credit,
        )

    return parse


def _run(
    fixture: _Fixture,
    *,
    parser: Any | None = None,
    output_name: str = "output",
    runner_parser: Any | None = None,
    patch_parser_backend: bool = True,
) -> Any:
    controls = json.loads(fixture.case_path.read_text(encoding="utf-8"))[
        "reported_controls"
    ]
    backend = parser or _parser(
        fixture.movements,
        source_control_debit=Decimal(controls["debit"]),
        source_control_credit=Decimal(controls["credit"]),
    )

    def load_layout(
        path: Path,
        *,
        expected_contract_sha256: str,
    ) -> object:
        assert path == fixture.parser_layout_path
        assert expected_contract_sha256 == _sha256(path)
        return object()

    def execute() -> Any:
        return PRODUCER.run_real_general_ledger_pilot(
            case_path=fixture.case_path,
            intake_contract_path=fixture.intake_contract_path,
            intake_receipt_path=fixture.intake_receipt_path,
            semantic_review_path=fixture.semantic_review_path,
            semantic_receipt_path=fixture.semantic_receipt_path,
            semantic_decisions_path=fixture.semantic_decisions_path,
            source_path=fixture.source_path,
            parser_layout_path=fixture.parser_layout_path,
            parser_adapter_implementation_path=(
                SCRIPTS_ROOT / "run_real_general_ledger_pilot.py"
            ),
            parser_implementation_path=(
                SCRIPTS_ROOT / "parse_commercial_general_journal.py"
            ),
            output_directory=fixture.root / output_name,
            local_run_root=fixture.root,
            repository_root=ROOT,
            as_of_date=VALIDATION_DATE,
            parser=runner_parser or PRODUCER.parse_reviewed_commercial_general_journal,
        )

    if not patch_parser_backend:
        return execute()
    with (
        patch.object(
            PRODUCER,
            "load_general_journal_layout_contract",
            load_layout,
        ),
        patch.object(
            PRODUCER,
            "parse_commercial_general_journal",
            backend,
        ),
    ):
        return execute()


def _output_files(output_directory: Path) -> dict[str, bytes]:
    return {
        path.relative_to(output_directory).as_posix(): path.read_bytes()
        for path in sorted(output_directory.rglob("*"))
        if path.is_file()
    }


def test_case_schema_and_success_outputs_are_strict_and_source_account_only(
    tmp_path: Path,
) -> None:
    fixture = _fixture(tmp_path)
    case_schema = json.loads(CASE_SCHEMA_PATH.read_text(encoding="utf-8"))
    jsonschema.Draft202012Validator.check_schema(case_schema)
    jsonschema.Draft202012Validator(case_schema).validate(
        json.loads(fixture.case_path.read_text(encoding="utf-8"))
    )
    semantic_schema = json.loads(
        SEMANTIC_DECISIONS_SCHEMA_PATH.read_text(encoding="utf-8")
    )
    jsonschema.Draft202012Validator.check_schema(semantic_schema)
    jsonschema.Draft202012Validator(semantic_schema).validate(
        json.loads(fixture.semantic_decisions_path.read_text(encoding="utf-8"))
    )

    result = _run(fixture)

    assert result.status == "passed"
    assert result.failure_code is None
    assert result.output_directory.stat().st_mode & 0o777 == 0o700
    output_files = _output_files(result.output_directory)
    assert set(output_files) == {
        f"artifacts/{PRODUCER.ACCOUNT_MONTH_ROLE}.bin",
        f"artifacts/{PRODUCER.RECONCILIATION_ROLE}.bin",
        "mechanical_errors.json",
    }
    account_month = output_files[
        f"artifacts/{PRODUCER.ACCOUNT_MONTH_ROLE}.bin"
    ].decode()
    assert account_month.splitlines() == [
        (
            "source_account_code,calendar_month,debit_positive,"
            "credit_negative,net_movement"
        ),
        "1000,2023-01,10,0,10",
        "1000,2023-02,2.5,-2.5,0",
        "2000,2023-01,0,-10,-10",
    ]
    register = json.loads(output_files["mechanical_errors.json"])
    reconciliation = json.loads(
        output_files[f"artifacts/{PRODUCER.RECONCILIATION_ROLE}.bin"]
    )
    assert register["summary"]["overall_status"] == "passed"
    assert register["publication_status"] == "withheld"
    assert register["report_ready"] is False
    assert reconciliation["controls"]["expected_calendar_months"] == [
        "2023-01",
        "2023-02",
    ]
    assert reconciliation["controls"]["emitted_calendar_months"] == [
        "2023-01",
        "2023-02",
    ]
    assert reconciliation["controls"]["expected_month_count"] == 2
    assert reconciliation["controls"]["emitted_month_count"] == 2
    assert reconciliation["controls"]["balanced_month_count"] == 2
    assert reconciliation["controls"]["basis"] == (
        "exact_extracted_final_debit_and_credit_controls"
    )
    assert reconciliation["controls"]["source_control_debit"] == "12.5"
    assert reconciliation["controls"]["source_control_credit"] == "12.5"
    assert reconciliation["bindings"]["semantic_decisions_sha256"] == _sha256(
        fixture.semantic_decisions_path
    )
    assert register["bindings"]["producer_contract_sha256"] == (
        PRODUCER.producer_contract_sha256()
    )


@pytest.mark.parametrize(
    "expected_months",
    [
        [],
        [f"2023-{month:02d}" for month in range(1, 13)] + ["2024-01"],
        ["2023-00"],
    ],
)
def test_case_schema_rejects_invalid_expected_calendar_months(
    tmp_path: Path,
    expected_months: list[str],
) -> None:
    fixture = _fixture(tmp_path)
    case = json.loads(fixture.case_path.read_text(encoding="utf-8"))
    case["reported_controls"]["expected_calendar_months"] = expected_months
    case_schema = json.loads(CASE_SCHEMA_PATH.read_text(encoding="utf-8"))

    with pytest.raises(jsonschema.ValidationError):
        jsonschema.Draft202012Validator(case_schema).validate(case)


def test_case_schema_accepts_twelve_expected_calendar_months(
    tmp_path: Path,
) -> None:
    fixture = _fixture(tmp_path)
    case = json.loads(fixture.case_path.read_text(encoding="utf-8"))
    case["reported_controls"]["expected_calendar_months"] = [
        f"2023-{month:02d}" for month in range(1, 13)
    ]
    case_schema = json.loads(CASE_SCHEMA_PATH.read_text(encoding="utf-8"))

    jsonschema.Draft202012Validator(case_schema).validate(case)


def test_first_class_adapter_binds_layout_and_maps_journal_movements(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source_path = tmp_path / "source.xlsx"
    source_path.write_bytes(b"synthetic")
    layout_path = tmp_path / "layout.json"
    layout_path.write_text("{}\n", encoding="utf-8")
    expected_source_sha256 = _sha256(source_path)
    expected_layout_sha256 = _sha256(layout_path)
    layout_contract = object()

    def load_layout(path: Path, *, expected_contract_sha256: str) -> object:
        assert path == layout_path
        assert expected_contract_sha256 == expected_layout_sha256
        return layout_contract

    def parse_journal(
        path: Path,
        *,
        expected_source_sha256: str,
        layout_contract: object,
    ) -> Any:
        assert path == source_path
        assert expected_source_sha256 == _sha256(source_path)
        assert layout_contract is not None
        return SimpleNamespace(
            source_sha256=expected_source_sha256,
            movements=(
                SimpleNamespace(
                    line_id=1,
                    posting_date=date(2023, 1, 1),
                    account_code="1000",
                    debit=Decimal("1"),
                    credit=Decimal("0"),
                ),
            ),
            source_control_debit_total=Decimal("1"),
            source_control_credit_total=Decimal("1"),
        )

    monkeypatch.setattr(PRODUCER, "load_general_journal_layout_contract", load_layout)
    monkeypatch.setattr(PRODUCER, "parse_commercial_general_journal", parse_journal)

    result = PRODUCER.parse_reviewed_commercial_general_journal(
        source_path,
        expected_source_sha256=expected_source_sha256,
        parser_layout_path=layout_path,
        expected_parser_layout_sha256=expected_layout_sha256,
    )

    assert result == PRODUCER.ParsedMovementBatch(
        source_sha256=expected_source_sha256,
        movements=(
            PRODUCER.Movement(
                "1",
                date(2023, 1, 1),
                "1000",
                Decimal("1"),
                Decimal("0"),
            ),
        ),
        source_control_debit=Decimal("1"),
        source_control_credit=Decimal("1"),
    )


def test_first_class_adapter_integrates_with_actual_reviewed_parser(
    tmp_path: Path,
) -> None:
    source_path = tmp_path / "reviewed-journal.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Reviewed journal"
    worksheet.cell(row=2, column=1, value="Data registrazione")
    worksheet.cell(row=3, column=3, value="Riga")
    worksheet.cell(row=3, column=6, value="Conto")
    worksheet.cell(row=3, column=10, value="Dare")
    worksheet.cell(row=3, column=14, value="Avere")
    worksheet.cell(row=4, column=1, value="2023-01-31 00:00:00")
    worksheet.cell(row=5, column=3, value="1")
    worksheet.cell(row=5, column=6, value="10 / 1 / 1")
    worksheet.cell(row=5, column=10, value="100.25")
    worksheet.cell(row=6, column=3, value="2")
    worksheet.cell(row=6, column=6, value="20 / 1 / 1")
    worksheet.cell(row=6, column=14, value="100.25")
    worksheet.cell(row=7, column=1, value="Totale generale 100,25 100,25")
    workbook.save(source_path)
    layout_path = tmp_path / "reviewed-layout.json"
    _write_json(
        layout_path,
        {
            "contract_version": "clara.commercial_general_journal_layout.v2",
            "review_status": "reviewed",
            "sheet_name": "Reviewed journal",
            "date_header_label": "Data registrazione",
            "line_header_label": "Riga",
            "account_header_label": "Conto",
            "debit_header_label": "Dare",
            "credit_header_label": "Avere",
            "page_layouts": [
                {
                    "layout_id": "layout-a",
                    "date_header_column": 1,
                    "line_header_column": 3,
                    "account_header_column": 6,
                    "debit_header_column": 10,
                    "credit_header_column": 14,
                    "date_columns": [1, 2, 3, 4, 5],
                    "line_id_columns": [1, 2, 3, 4, 5],
                    "account_columns": [6],
                    "debit_amount_columns": [9, 10, 11],
                    "credit_amount_columns": [13, 14, 15],
                    "physical_first_line_columns": [6],
                }
            ],
            "date_patterns": [
                {
                    "pattern": (r"^(?P<date>[0-9]{4}-[0-9]{2}-[0-9]{2}) 00:00:00$"),
                    "strptime_format": "%Y-%m-%d",
                }
            ],
            "account_code_pattern": r"[0-9]+\s*/\s*[0-9]+\s*/\s*[0-9]+",
            "logical_candidate_pattern": (
                r"^\s*[1-9][0-9]*\s+[0-9]+\s*/\s*[0-9]+\s*/\s*[0-9]+"
            ),
            "logical_movement_patterns": [
                {
                    "layout_ids": ["layout-a"],
                    "pattern": (
                        r"^\s*(?P<line_id>[1-9][0-9]*)\s+"
                        r"(?P<account>[0-9]+\s*/\s*[0-9]+\s*/\s*[0-9]+)"
                        r"\s+(?:(?P<debit>"
                        r"(?:0|[1-9][0-9]{0,2}(?:\.[0-9]{3})*),[0-9]{2})"
                        r"\s+D|(?P<credit>"
                        r"(?:0|[1-9][0-9]{0,2}(?:\.[0-9]{3})*),[0-9]{2})"
                        r"\s+C)\s*$"
                    ),
                    "amount_format": "italian_grouped_2",
                }
            ],
            "physical_embedded_amount_patterns": [],
            "reviewed_amount_pairs": [],
            "reviewed_zero_amount_line_ids": [],
            "physical_amount_format": "canonical_dot",
            "amount_sign_policy": "nonnegative",
            "control_pattern": (
                r"^Totale generale\s+"
                r"(?P<debit>(?:0|[1-9][0-9]{0,2}(?:\.[0-9]{3})*),[0-9]{2})"
                r"\s+"
                r"(?P<credit>(?:0|[1-9][0-9]{0,2}(?:\.[0-9]{3})*),[0-9]{2})"
                r"$"
            ),
            "control_amount_format": "italian_grouped_2",
            "reviewed_final_debit_total": "100.25",
            "reviewed_final_credit_total": "100.25",
        },
    )

    result = PRODUCER.parse_reviewed_commercial_general_journal(
        source_path,
        expected_source_sha256=_sha256(source_path),
        parser_layout_path=layout_path,
        expected_parser_layout_sha256=_sha256(layout_path),
    )

    assert result.source_sha256 == _sha256(source_path)
    assert result.movements == (
        PRODUCER.Movement(
            "1",
            date(2023, 1, 31),
            "10/1/1",
            Decimal("100.25"),
            Decimal("0"),
        ),
        PRODUCER.Movement(
            "2",
            date(2023, 1, 31),
            "20/1/1",
            Decimal("0"),
            Decimal("100.25"),
        ),
    )
    assert result.source_control_debit == Decimal("100.25")
    assert result.source_control_credit == Decimal("100.25")


def test_blocking_semantic_issue_is_replayed_before_source_open(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    fixture = _fixture(tmp_path, blocking=True)
    parser_calls: list[Path] = []

    def forbidden_intake_replay(*_args: Any, **_kwargs: Any) -> Any:
        raise AssertionError("source-bound intake replay ran before semantic gate")

    monkeypatch.setattr(
        PRODUCER,
        "validate_real_data_pilot_intake_v2",
        forbidden_intake_replay,
    )

    result = _run(fixture, parser=_parser(fixture.movements, parser_calls))

    assert result.status == "failed"
    assert parser_calls == []
    assert set(_output_files(result.output_directory)) == {"mechanical_errors.json"}


def test_duplicate_movement_id_fails_without_plausible_table(
    tmp_path: Path,
) -> None:
    fixture = _fixture(tmp_path)
    duplicate = copy.copy(fixture.movements[1])
    duplicate = PRODUCER.Movement(
        fixture.movements[0].movement_id,
        duplicate.posting_date,
        duplicate.source_account_code,
        duplicate.debit,
        duplicate.credit,
    )

    result = _run(fixture, parser=_parser([fixture.movements[0], duplicate]))

    assert result.status == "failed"
    assert set(_output_files(result.output_directory)) == {"mechanical_errors.json"}


def test_posting_date_outside_reviewed_year_is_one_period_error(
    tmp_path: Path,
) -> None:
    fixture = _fixture(tmp_path)
    movement = PRODUCER.Movement(
        "1",
        date(2024, 1, 1),
        "1000",
        Decimal("12.5"),
        Decimal("5.5"),
    )

    result = _run(fixture, parser=_parser([movement]))
    register = json.loads(
        (result.output_directory / "mechanical_errors.json").read_text(encoding="utf-8")
    )

    assert result.status == "failed"
    assert register["summary"]["error_count"] == 1
    assert register["summary"]["class_counts"]["period"] == 1


@pytest.mark.parametrize(
    "movement",
    [
        PRODUCER.Movement(
            "1",
            date(2023, 1, 1),
            "1000",
            Decimal("-1"),
            Decimal("0"),
        ),
        PRODUCER.Movement(
            "1",
            date(2023, 1, 1),
            "1000",
            cast(Any, "12.5"),
            Decimal("5.5"),
        ),
        PRODUCER.Movement(
            "1",
            date(2023, 1, 1),
            "1000",
            Decimal("1"),
            Decimal("1"),
        ),
    ],
)
def test_decimal_or_source_sign_error_is_one_numeric_error(
    tmp_path: Path,
    movement: Any,
) -> None:
    fixture = _fixture(tmp_path)

    result = _run(fixture, parser=_parser([movement]))
    register = json.loads(
        (result.output_directory / "mechanical_errors.json").read_text(encoding="utf-8")
    )

    assert result.status == "failed"
    assert register["summary"]["error_count"] == 1
    assert register["summary"]["class_counts"]["numeric"] == 1


def test_exact_reported_control_mismatch_is_one_reconciliation_error(
    tmp_path: Path,
) -> None:
    fixture = _fixture(
        tmp_path,
        debit_control="12.6",
        credit_control="12.6",
    )

    result = _run(
        fixture,
        parser=_parser(
            fixture.movements,
            source_control_debit=Decimal("12.5"),
            source_control_credit=Decimal("12.5"),
        ),
    )
    register = json.loads(
        (result.output_directory / "mechanical_errors.json").read_text(encoding="utf-8")
    )

    assert result.status == "failed"
    assert result.failure_code == PRODUCER.ERROR_CODES["reported_control_mismatch"]
    assert register["summary"]["error_count"] == 1
    assert register["summary"]["class_counts"]["reconciliation"] == 1
    assert set(_output_files(result.output_directory)) == {"mechanical_errors.json"}


def test_extracted_source_control_mismatch_is_one_reconciliation_error(
    tmp_path: Path,
) -> None:
    fixture = _fixture(tmp_path)
    parser = _parser(
        fixture.movements,
        source_control_debit=Decimal("12.4"),
        source_control_credit=Decimal("12.4"),
    )

    result = _run(fixture, parser=parser)
    register = json.loads(
        (result.output_directory / "mechanical_errors.json").read_text(encoding="utf-8")
    )

    assert result.status == "failed"
    assert result.failure_code == PRODUCER.ERROR_CODES["source_control_mismatch"]
    assert register["summary"]["error_count"] == 1
    assert register["summary"]["class_counts"]["reconciliation"] == 1
    assert set(_output_files(result.output_directory)) == {"mechanical_errors.json"}


def test_parser_failure_emits_only_fixed_mechanical_register(
    tmp_path: Path,
) -> None:
    fixture = _fixture(tmp_path)

    def failed_parser(
        _source_path: Path,
        *,
        expected_source_sha256: str,
        layout_contract: object,
    ) -> Any:
        assert expected_source_sha256
        assert layout_contract
        raise PRODUCER.ParserMechanicalError

    result = _run(fixture, parser=failed_parser)
    register_text = (result.output_directory / "mechanical_errors.json").read_text(
        encoding="utf-8"
    )

    assert result.status == "failed"
    assert set(_output_files(result.output_directory)) == {"mechanical_errors.json"}
    assert "synthetic general-journal" not in register_text
    assert "line-" not in register_text


@pytest.mark.parametrize("account_code", ["=1+1", "+1000", "A\n1000", "A\0B"])
def test_unsafe_source_account_code_is_one_structure_error(
    tmp_path: Path,
    account_code: str,
) -> None:
    fixture = _fixture(tmp_path)
    movement = PRODUCER.Movement(
        "1",
        date(2023, 1, 1),
        account_code,
        Decimal("12.5"),
        Decimal("12.5"),
    )

    result = _run(fixture, parser=_parser([movement]))
    register = json.loads(
        (result.output_directory / "mechanical_errors.json").read_text(encoding="utf-8")
    )

    assert result.status == "failed"
    assert register["summary"]["error_count"] == 1
    assert register["summary"]["class_counts"]["source_structure"] == 1


def test_empty_movement_set_is_one_parser_error(tmp_path: Path) -> None:
    fixture = _fixture(tmp_path)

    result = _run(fixture, parser=_parser([]))
    register = json.loads(
        (result.output_directory / "mechanical_errors.json").read_text(encoding="utf-8")
    )

    assert result.status == "failed"
    assert register["summary"]["error_count"] == 1
    assert register["summary"]["class_counts"]["source_structure"] == 1


def test_injected_parser_cannot_claim_unbound_parser_module(
    tmp_path: Path,
) -> None:
    fixture = _fixture(tmp_path)

    result = _run(
        fixture,
        runner_parser=lambda *_args, **_kwargs: None,
    )

    assert result.status == "failed"
    assert set(_output_files(result.output_directory)) == {"mechanical_errors.json"}


def test_cross_month_mutation_is_one_monthly_reconciliation_error(
    tmp_path: Path,
) -> None:
    fixture = _fixture(tmp_path)
    shifted_credit = PRODUCER.Movement(
        fixture.movements[3].movement_id,
        date(2023, 3, 7),
        fixture.movements[3].source_account_code,
        fixture.movements[3].debit,
        fixture.movements[3].credit,
    )
    movements = [*fixture.movements[:3], shifted_credit]

    result = _run(fixture, parser=_parser(movements))
    register = json.loads(
        (result.output_directory / "mechanical_errors.json").read_text(encoding="utf-8")
    )

    assert result.status == "failed"
    assert register["summary"]["error_count"] == 1
    assert register["summary"]["class_counts"]["reconciliation"] == 1


def test_expected_calendar_month_coverage_mismatch_is_one_period_error(
    tmp_path: Path,
) -> None:
    fixture = _fixture(
        tmp_path,
        expected_calendar_months=["2023-01", "2023-03"],
    )

    result = _run(fixture)
    register = json.loads(
        (result.output_directory / "mechanical_errors.json").read_text(encoding="utf-8")
    )

    assert result.status == "failed"
    assert register["summary"]["error_count"] == 1
    assert register["summary"]["class_counts"]["period"] == 1
    assert set(_output_files(result.output_directory)) == {"mechanical_errors.json"}


@pytest.mark.parametrize(
    "expected_months",
    [
        ["2023-02", "2023-01"],
        ["2024-01"],
    ],
)
def test_runtime_rejects_noncanonical_reviewed_month_projection(
    tmp_path: Path,
    expected_months: list[str],
) -> None:
    fixture = _fixture(tmp_path, expected_calendar_months=expected_months)

    result = _run(fixture)
    register = json.loads(
        (result.output_directory / "mechanical_errors.json").read_text(encoding="utf-8")
    )

    assert result.status == "failed"
    assert register["summary"]["error_count"] == 1
    assert register["summary"]["class_counts"]["contract"] == 1
    assert set(_output_files(result.output_directory)) == {"mechanical_errors.json"}


def test_semantic_decisions_must_be_exact_case_projection(
    tmp_path: Path,
) -> None:
    fixture = _fixture(tmp_path)
    decisions = json.loads(fixture.semantic_decisions_path.read_text(encoding="utf-8"))
    decisions["reported_controls"]["expected_calendar_months"] = [
        "2023-01",
        "2023-03",
    ]
    _write_json(fixture.semantic_decisions_path, decisions)
    _refresh_semantic_review_and_case_bindings(fixture)

    result = _run(fixture)
    register = json.loads(
        (result.output_directory / "mechanical_errors.json").read_text(encoding="utf-8")
    )

    assert result.status == "failed"
    assert register["summary"]["error_count"] == 1
    assert register["summary"]["class_counts"]["contract"] == 1
    assert set(_output_files(result.output_directory)) == {"mechanical_errors.json"}


def test_semantic_decisions_evidence_must_match_exact_bytes(
    tmp_path: Path,
) -> None:
    fixture = _fixture(tmp_path)
    review = json.loads(fixture.semantic_review_path.read_text(encoding="utf-8"))
    review["evidence_registry"]["evidence-semantic-decisions"]["sha256"] = "0" * 64
    _write_json(fixture.semantic_review_path, review)
    parser_calls: list[Path] = []

    result = _run(
        fixture,
        parser=_parser(fixture.movements, parser_calls),
    )
    register = json.loads(
        (result.output_directory / "mechanical_errors.json").read_text(encoding="utf-8")
    )

    assert result.status == "failed"
    assert parser_calls == []
    assert register["summary"]["error_count"] == 1
    assert register["summary"]["class_counts"]["contract"] == 1
    assert set(_output_files(result.output_directory)) == {"mechanical_errors.json"}


def test_every_required_review_must_reference_semantic_decisions(
    tmp_path: Path,
) -> None:
    fixture = _fixture(tmp_path)
    review = json.loads(fixture.semantic_review_path.read_text(encoding="utf-8"))
    review["required_reviews"][0]["evidence_refs"].remove("evidence-semantic-decisions")
    _write_json(fixture.semantic_review_path, review)
    _refresh_semantic_review_and_case_bindings(fixture)
    parser_calls: list[Path] = []

    result = _run(
        fixture,
        parser=_parser(fixture.movements, parser_calls),
    )
    register = json.loads(
        (result.output_directory / "mechanical_errors.json").read_text(encoding="utf-8")
    )

    assert result.status == "failed"
    assert parser_calls == []
    assert register["summary"]["error_count"] == 1
    assert register["summary"]["class_counts"]["contract"] == 1
    assert set(_output_files(result.output_directory)) == {"mechanical_errors.json"}


@pytest.mark.parametrize("target", ["case", "intake_receipt"])
def test_tampered_contract_or_receipt_fails_replay_without_table(
    tmp_path: Path,
    target: str,
) -> None:
    fixture = _fixture(tmp_path)
    if target == "case":
        case = json.loads(fixture.case_path.read_text(encoding="utf-8"))
        case["reviewed_decisions"]["account_identity"][
            "statement_mapping"
        ] = "performed"
        _write_json(fixture.case_path, case)
    else:
        intake_receipt = json.loads(
            fixture.intake_receipt_path.read_text(encoding="utf-8")
        )
        intake_receipt["eligibility"]["report_ready"] = True
        _write_json(fixture.intake_receipt_path, intake_receipt)
        case = json.loads(fixture.case_path.read_text(encoding="utf-8"))
        case["bindings"]["intake_receipt_sha256"] = _sha256(fixture.intake_receipt_path)
        _write_json(fixture.case_path, case)

    result = _run(fixture)

    assert result.status == "failed"
    assert set(_output_files(result.output_directory)) == {"mechanical_errors.json"}


def test_success_output_is_byte_deterministic(
    tmp_path: Path,
) -> None:
    fixture = _fixture(tmp_path)
    parser = _parser(fixture.movements)

    first = _run(fixture, parser=parser, output_name="output-one")
    second = _run(fixture, parser=parser, output_name="output-two")

    assert first.status == second.status == "passed"
    assert first.output_receipts == second.output_receipts
    assert _output_files(first.output_directory) == _output_files(
        second.output_directory
    )
