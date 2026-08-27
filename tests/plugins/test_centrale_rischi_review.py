from __future__ import annotations

import importlib.util
import json
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Table,
    TableStyle,
)

ROOT = Path(__file__).resolve().parents[2]
PLUGIN_ROOT = ROOT / "plugins" / "centrale-rischi-review"
SCRIPTS = PLUGIN_ROOT / "scripts"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

import evaluate_pdf_corpus as evaluate_pdf_corpus_cli  # noqa: E402
import inspect_inputs  # noqa: E402
import run_analysis  # noqa: E402
import run_gold_benchmark as gold_benchmark  # noqa: E402
from centrale_rischi_core import (  # noqa: E402
    COMMENTARY_SCHEMA,
    CentraleRischiContractError,
    build_analysis,
    build_inspection,
    build_model_context,
    finalize_commentary,
    load_source_tables,
    render_html,
    write_excel,
)
from centrale_rischi_pdf import (  # noqa: E402
    evaluate_pdf_corpus,
    normalize_pdf,
    write_normalized_workbook,
)


def _write_source(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "CR"
    sheet.append(
        (
            "Mese",
            "Intermediario",
            "Categoria",
            "Durata originaria",
            "Durata residua",
            "Accordato",
            "Operativo",
            "Utilizzato",
            "Garanzia",
            "Garantito",
            "Pregiudizievole",
        )
    )
    for row in (
        (
            "2025-01",
            "Banca A",
            "Autoliquidante",
            "Fino a 1 anno",
            "Fino a 1 anno",
            1000,
            900,
            800,
            "Personale",
            300,
            "",
        ),
        (
            "2025-01",
            "Banca B",
            "A scadenza",
            "Oltre 5 anni",
            "Oltre 1 anno",
            600,
            600,
            500,
            "",
            0,
            "",
        ),
        (
            "2025-02",
            "Banca A",
            "Autoliquidante",
            "Fino a 1 anno",
            "Fino a 1 anno",
            1000,
            900,
            950,
            "Personale",
            300,
            "",
        ),
        (
            "2025-02",
            "Banca B",
            "A scadenza",
            "Oltre 5 anni",
            "Oltre 1 anno",
            600,
            600,
            450,
            "Reale",
            200,
            "Protesto segnalato",
        ),
        (
            "2025-02",
            "Banca C",
            "Sofferenze",
            "",
            "",
            100,
            0,
            100,
            "",
            0,
            "",
        ),
    ):
        sheet.append(row)
    workbook.save(path)


def _write_native_pdf_fixture(path: Path) -> None:
    styles = getSampleStyleSheet()
    document = SimpleDocTemplate(
        path.as_posix(),
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    exposure_headers = (
        "Categoria",
        "Localizzazione",
        "Durata Originaria",
        "Durata Residua",
        "Divisa",
        "Import Export",
        "Tipo Attivita",
        "Stato Rapporto",
        "Tipo Garanzia",
        "Ruolo Affidato",
        "Accordato",
        "Accordato Operativo",
        "Utilizzato",
        "Saldo Medio",
        "Importo Garantito",
    )
    current = (
        "RISCHI A SCADENZA",
        "Pavia",
        "Oltre cinque anni",
        "Oltre 1 anno",
        "Euro",
        "Operazioni diverse",
        "Prestito personale",
        "Rapporto non contestato",
        "Assenza di garanzie",
        "0",
        "45.000",
        "45.000",
        "45.000",
        "0",
        "0",
    )
    previous = (*current[:12], "60.000", *current[13:], "03/08/2026", "10/08/2026")

    def styled_table(rows: list[tuple[str, ...]]) -> Table:
        table = Table(rows, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("FONTSIZE", (0, 0), (-1, -1), 5),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        return table

    story = [
        Paragraph("DATA DI RIFERIMENTO: giugno 2026", styles["Heading2"]),
        Paragraph("Intermediario: BANCA DUE", styles["BodyText"]),
        Paragraph("Crediti per cassa - Situazione corrente", styles["BodyText"]),
        styled_table([exposure_headers, current]),
        Paragraph("Segnalazioni presenti prima delle correzioni", styles["BodyText"]),
        styled_table([(*exposure_headers, "Da", "A"), previous]),
        PageBreak(),
        Paragraph("DATA DI RIFERIMENTO: aprile 2026", styles["Heading2"]),
        Paragraph("Intermediario: BANCA TRE", styles["BodyText"]),
        Paragraph("Garanzie ricevute - Situazione corrente", styles["BodyText"]),
        styled_table(
            [
                (
                    "Categoria",
                    "Localizzazione",
                    "Garantito",
                    "Stato Rapporto",
                    "Tipo Garanzia",
                    "Valore Garanzia",
                    "Importo Garantito",
                ),
                (
                    "GARANZIE RICEVUTE",
                    "Milano",
                    "VIOLA VERDI",
                    "Rapporto contestato",
                    "Garanzia personale",
                    "130.000",
                    "80.000",
                ),
            ]
        ),
        Paragraph("Informazioni sui garanti", styles["BodyText"]),
        styled_table(
            [
                ("Garante", "Valore Garanzia", "Importo Garantito"),
                ("MARIO GARANTE", "50.000", "40.000"),
            ]
        ),
        Paragraph("Informazioni sui debitori ceduti", styles["BodyText"]),
        styled_table(
            [
                ("Ceduto", "Valore nominale del credito ceduto"),
                ("DEBITORE CEDUTO", "20.000"),
            ]
        ),
        PageBreak(),
        Paragraph("Intermediario: BANCA QUATTRO", styles["BodyText"]),
        Paragraph("Eventi inframensili", styles["BodyText"]),
        styled_table(
            [
                ("Data Evento", "Tipo Evento", "Evento Cancellato"),
                ("12/06/2026", "Regolarizzazione", "No"),
            ]
        ),
        PageBreak(),
        Paragraph("BANCA CINQUE", styles["BodyText"]),
        Paragraph("Richieste informazioni", styles["BodyText"]),
        styled_table(
            [
                (
                    "Data della richiesta di informazione",
                    "Periodo richiesto",
                    "Tipo richiesta di informazione",
                    "Causale della richiesta",
                    "Descrizione causale",
                ),
                (
                    "15/06/2026",
                    "maggio 2026",
                    "Prima informazione",
                    "01",
                    "Istruttoria affidamento",
                ),
            ]
        ),
        PageBreak(),
        Paragraph("Tabella didattica non supportata", styles["BodyText"]),
        styled_table([("Campo A", "Campo B"), ("Valore A", "Valore B")]),
    ]
    document.build(story)


def _write_fragmented_header_pdf_fixture(path: Path) -> None:
    styles = getSampleStyleSheet()
    document = SimpleDocTemplate(
        path.as_posix(),
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    rows = [
        (
            "Categor",
            "ia",
            "Localizzazione",
            "Durata Originaria",
            "Durata Residua",
            "Divisa",
            "Tipo Attivita",
            "Stato Rapporto",
            "Tipo Garanzia",
            "Ruolo Affidato",
            "Accordato",
            "Accordato Operativo",
            "Utilizzato",
            "Saldo Medio",
            "Importo Garantito",
        ),
        (
            "RISCHI A SCADENZA",
            "",
            "Milano",
            "Oltre cinque anni",
            "Oltre 1 anno",
            "Euro",
            "Prestito",
            "Rapporto non contestato",
            "Ipoteca",
            "0",
            "150.000",
            "150.000",
            "170.000",
            "0",
            "170.000",
        ),
    ]
    table = Table(rows, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 5),
            ]
        )
    )
    document.build(
        [
            Paragraph("DATA CONTABILE: settembre 2010", styles["Heading2"]),
            Paragraph("Intermediario: BANCA UNO", styles["BodyText"]),
            table,
        ]
    )


def _write_clipped_valid_to_pdf_fixture(path: Path) -> None:
    page_width, page_height = landscape(A4)
    pdf = canvas.Canvas(path.as_posix(), pagesize=(page_width, page_height))
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(24, page_height - 40, "DATA DI RIFERIMENTO: giugno 2021")
    pdf.drawString(24, page_height - 58, "Intermediario: BANCA UNO")
    headers = (
        "Categoria",
        "Localizzazione",
        "Durata Originaria",
        "Durata Residua",
        "Divisa",
        "Import Export",
        "Tipo Attivita",
        "Stato Rapporto",
        "Tipo Garanzia",
        "Ruolo Affidato",
        "Accordato",
        "Accordato Operativo",
        "Utilizzato",
        "Saldo Medio",
        "Importo Garantito",
        "Da",
    )
    row = (
        "RISCHI A SCADENZA",
        "Pavia",
        "Oltre cinque anni",
        "Oltre 1 anno",
        "Euro",
        "Altro",
        "Prestito",
        "Non contestato",
        "Nessuna",
        "0",
        "45.000",
        "45.000",
        "60.000",
        "0",
        "0",
        "02/08/2021",
    )
    column_widths = [94, 48, 48, 44, 28, 44, 44, 48, 44, 38, 42, 48, 40, 34, 42, 46]
    table = Table([headers, row], colWidths=column_widths, rowHeights=(24, 34))
    table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("FONTSIZE", (0, 0), (-1, -1), 4.5),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    table_width, table_height = table.wrapOn(pdf, page_width, page_height)
    table_x = 20
    table_y = page_height - 150
    table.drawOn(pdf, table_x, table_y)
    pdf.setFont("Helvetica", 5)
    pdf.drawString(table_x + table_width + 12, table_y + table_height - 15, "A")
    pdf.drawString(table_x + table_width + 4, table_y + 13, "10/08/2021")
    pdf.save()


def _write_repeated_header_pdf_fixture(path: Path) -> None:
    styles = getSampleStyleSheet()
    document = SimpleDocTemplate(
        path.as_posix(),
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    headers = (
        "Categoria",
        "Localizzazione",
        "Accordato",
        "Accordato Operativo",
        "Utilizzato",
    )
    table = Table(
        [
            headers,
            ("RISCHI A SCADENZA", "Milano", "100", "100", "90"),
            headers,
            ("RISCHI A REVOCA", "Roma", "200", "200", "180"),
        ]
    )
    table.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.5, colors.black)]))
    document.build(
        [
            Paragraph("DATA DI RIFERIMENTO: giugno 2021", styles["Heading2"]),
            Paragraph("Intermediario: BANCA UNO", styles["BodyText"]),
            table,
        ]
    )


def _write_merged_correction_pdf_fixture(path: Path) -> None:
    styles = getSampleStyleSheet()
    document = SimpleDocTemplate(
        path.as_posix(),
        pagesize=landscape(A4),
        leftMargin=8 * mm,
        rightMargin=8 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )
    headers = (
        "Categoria",
        "Localizzazione",
        "Durata Originaria",
        "Durata Residua",
        "Divisa",
        "Import Export",
        "Tipo Attivita",
        "Stato Rapporto",
        "Tipo Garanzia",
        "Ruolo Affidato",
        "Accordato",
        "Accordato Operativo",
        "Utilizzato",
        "Saldo Medio",
        "Importo Garantito",
        "Da",
        "A",
    )
    current = (
        "RISCHI A SCADENZA",
        "38270",
        "17",
        "18",
        "1",
        "8",
        "32",
        "90",
        "112",
        "0",
        "60.000",
        "60.000",
        "60.000",
        "0",
        "60.000",
        "",
        "",
    )
    previous_absent = (
        *current[:10],
        "Assenza di segnalazione",
        "",
        "",
        "",
        "",
        "01/12/2009",
        "10/12/2009",
    )
    previous_numeric = (
        *current[:8],
        "125",
        "0",
        "0",
        "0",
        "60.000",
        "0",
        "60.000",
        "01/12/2009",
        "10/12/2009",
    )
    table = Table([headers, current, headers, previous_absent, previous_numeric])
    table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 3.5),
            ]
        )
    )
    document.build(
        [
            Paragraph(
                "La correzione operata dalla banca sulle segnalazioni del mese di ottobre 2009 è evidenziata come segue:",
                styles["BodyText"],
            ),
            Paragraph("Situazione corrente", styles["BodyText"]),
            Paragraph(
                'Per la data contabile indicata l intermediario aveva segnalato le seguenti informazioni successivamente rettificate. Nella colonna "Da" e "A" compaiono le date.',
                styles["BodyText"],
            ),
            table,
        ]
    )


def _write_generic_request_pdf_fixture(path: Path, validity_period: str) -> None:
    styles = getSampleStyleSheet()
    document = SimpleDocTemplate(
        path.as_posix(),
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    table = Table(
        [
            (
                "Data della richiesta di informazione",
                "Periodo richiesto",
                "Tipo richiesta di informazione",
                "Causale della richiesta",
                "Descrizione causale",
                "Periodo validita",
                "Note",
            ),
            (
                "02/04/2021",
                "Marzo 2020",
                "PRIMA INFORMAZIONE",
                "01",
                "RICHIESTA PER CONCESSIONE DI FIDO",
                validity_period,
                "IN CORSO DI VALIDITA",
            ),
        ]
    )
    table.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.5, colors.black)]))
    document.build(
        [
            Paragraph("RICHIESTE DI INFORMAZIONE", styles["Heading2"]),
            Paragraph("INTERMEDIARIO", styles["BodyText"]),
            table,
        ]
    )


def _write_auxiliary_risk_pdf_fixture(path: Path) -> None:
    styles = getSampleStyleSheet()
    document = SimpleDocTemplate(
        path.as_posix(),
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    detail = Table(
        [
            ("Categoria", "Localizzazione", "Importo"),
            ("SOFFERENZE - CREDITI PASSATI A PERDITA", "Milano", "30.010"),
        ]
    )
    summary = Table(
        [
            ("", "Accordato", "Accordato Operativo", "Utilizzato"),
            ("Crediti per cassa", "308.412", "308.412", "202.890"),
        ]
    )
    for table in (detail, summary):
        table.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.5, colors.black)]))
    document.build(
        [
            Paragraph("ULTIMA DATA CONTABILE: 30/09/2010", styles["Heading2"]),
            Paragraph("Intermediario: BANCA ALFA", styles["BodyText"]),
            detail,
            Paragraph("Prospetto sintetico", styles["BodyText"]),
            summary,
        ]
    )


def _reviewed_recipe(inspection: dict[str, object]) -> dict[str, object]:
    return {
        "schema_version": "vera.centrale_rischi_recipe.v2",
        "workflow_id": "centrale-rischi-review",
        "inventory_sha256": inspection["inventory_sha256"],
        "entity": "Synthetic S.r.l.",
        "currency": "EUR",
        "analysis_mode": "trend",
        "analysis_objective": "Explain current exposure and recent movement.",
        "audience": "professional",
        "source_kind": "tabular_export",
        "source_document_sha256": "",
        "table_id": inspection["tables"][0]["table_id"],  # type: ignore[index]
        "columns": {
            "reference_month": "Mese",
            "intermediary": "Intermediario",
            "risk_category": "Categoria",
            "original_duration": "Durata originaria",
            "residual_duration": "Durata residua",
            "granted": "Accordato",
            "operational_granted": "Operativo",
            "used": "Utilizzato",
            "guarantee_type": "Garanzia",
            "guaranteed_amount": "Garantito",
            "prejudicial_event": "Pregiudizievole",
        },
        "value_mappings": {
            "original_term": {
                "Fino a 1 anno": "short",
                "Oltre 5 anni": "long",
                "": "not_relevant",
            },
            "residual_term": {
                "Fino a 1 anno": "within_one_year",
                "Oltre 1 anno": "over_one_year",
                "": "not_relevant",
            },
            "exposure_family": {
                "Autoliquidante": "performing",
                "A scadenza": "performing",
                "Sofferenze": "suffering",
            },
        },
        "control_totals": {"used": "1500"},
        "control_tolerance": "0.01",
        "mapping_review": {
            "status": "reviewed",
            "reviewer": "Test reviewer",
            "reviewed_at": "2026-08-25T12:00:00+02:00",
        },
    }


def _analysis_for(path: Path) -> tuple[dict[str, object], dict[str, object]]:
    tables = load_source_tables([path])
    inspection, _, _ = build_inspection(tables)
    return build_analysis(tables, _reviewed_recipe(inspection)), inspection


def test_inspection_does_not_assign_semantic_roles(tmp_path: Path) -> None:
    source = tmp_path / "cr.xlsx"
    _write_source(source)

    inspection, control, recipe = build_inspection(load_source_tables([source]))

    assert inspection["semantic_roles_assigned"] is False
    assert recipe["mapping_review"]["status"] == "pending"
    assert recipe["value_mappings"] == {
        "original_term": {},
        "residual_term": {},
        "exposure_family": {},
    }
    assert Path(control["tables"][0]["absolute_path"]).is_absolute()
    assert "absolute_path" not in json.dumps(inspection)


def test_analysis_calculates_maturities_overruns_guarantees_and_kpis(
    tmp_path: Path,
) -> None:
    source = tmp_path / "cr.xlsx"
    _write_source(source)

    analysis, _ = _analysis_for(source)
    metrics = {item["metric_id"]: item for item in analysis["metrics"]}
    original_term = {
        item["original_term"]: item for item in analysis["original_term_summary"]
    }
    residual_term = {
        item["residual_term"]: item for item in analysis["residual_term_summary"]
    }
    categories = {
        item["risk_category"]: item for item in analysis["risk_category_summary"]
    }
    movements = {
        item["risk_category"]: item for item in analysis["category_movement_summary"]
    }

    assert analysis["status"] == "complete"
    assert metrics["cr.total_used"]["value"] == "1500"
    assert metrics["cr.available_resources"]["value"] == "150"
    assert metrics["cr.overrun_amount"]["value"] == "50"
    assert metrics["cr.overrun_count"]["value"] == 1
    assert metrics["cr.used_mom_change"]["value"] == "200"
    assert "cr.utilization_pct" not in metrics
    assert metrics["financial.dscr"]["availability"] == "unavailable"
    assert original_term["short"]["used"] == "950"
    assert original_term["not_relevant"]["used"] == "100"
    assert original_term["long"]["used"] == "450"
    assert residual_term["within_one_year"]["used"] == "950"
    assert residual_term["over_one_year"]["used"] == "450"
    assert residual_term["not_relevant"]["used"] == "100"
    assert categories["Autoliquidante"]["utilization_pct"] == "105.56"
    assert categories["A scadenza"]["utilization_pct"] == "75"
    assert movements["Autoliquidante"]["used_change"] == "150"
    assert movements["A scadenza"]["used_change"] == "-50"
    assert movements["Sofferenze"]["used_change"] == "100"
    assert movements["Sofferenze"]["presence"] == "new_in_latest"
    assert any(
        item["label"] == "Variazione utilizzato — Sofferenze" and item["value"] == "100"
        for item in analysis["metrics"]
    )
    assert len(analysis["guarantees"]) == 2
    assert len(analysis["overruns"]) == 1
    assert len(analysis["prejudicial_events"]) == 1
    assert all(row["exposure_family"] != "suffering" for row in analysis["overruns"])


def test_analysis_rejects_unreviewed_or_incomplete_semantic_mapping(
    tmp_path: Path,
) -> None:
    source = tmp_path / "cr.xlsx"
    _write_source(source)
    tables = load_source_tables([source])
    inspection, _, _ = build_inspection(tables)
    recipe = _reviewed_recipe(inspection)
    del recipe["value_mappings"]["original_term"]["Oltre 5 anni"]

    with pytest.raises(CentraleRischiContractError, match="unmapped original_duration"):
        build_analysis(tables, recipe)


def test_pdf_derived_rows_require_layout_provenance(tmp_path: Path) -> None:
    source = tmp_path / "cr.xlsx"
    _write_source(source)
    tables = load_source_tables([source])
    inspection, _, _ = build_inspection(tables)
    recipe = _reviewed_recipe(inspection)
    recipe["source_kind"] = "native_pdf_extraction"
    recipe["source_document_sha256"] = "a" * 64

    with pytest.raises(CentraleRischiContractError, match="provenance mapping"):
        build_analysis(tables, recipe)


def test_native_pdf_normalization_feeds_analysis_without_double_counting_previous(
    tmp_path: Path,
) -> None:
    source_pdf = tmp_path / "centrale-rischi.pdf"
    normalized_workbook = tmp_path / "normalized.xlsx"
    _write_native_pdf_fixture(source_pdf)

    normalization = normalize_pdf(source_pdf)
    normalization["tables"]["exposures"][1]["granted"] = "Assenza di segnalazione"
    write_normalized_workbook(normalized_workbook, normalization)
    tables = load_source_tables([normalized_workbook])
    exposure_table = next(
        table for table in tables if table.table_label == "Esposizioni"
    )
    inspection, _, _ = build_inspection(tables)
    recipe = {
        "schema_version": "vera.centrale_rischi_recipe.v2",
        "workflow_id": "centrale-rischi-review",
        "inventory_sha256": inspection["inventory_sha256"],
        "entity": "Synthetic S.r.l.",
        "currency": "EUR",
        "analysis_mode": "descriptive",
        "analysis_objective": "Validate PDF normalization.",
        "audience": "professional",
        "source_kind": "native_pdf_extraction",
        "source_document_sha256": normalization["source"]["source_document_sha256"],
        "table_id": exposure_table.table_id,
        "columns": {
            "reference_month": "reference_month",
            "intermediary": "intermediary",
            "risk_category": "category",
            "original_duration": "original_duration",
            "residual_duration": "residual_duration",
            "granted": "granted",
            "operational_granted": "operational_granted",
            "used": "used",
            "guarantee_type": "guarantee_type",
            "guaranteed_amount": "guaranteed_amount",
            "record_status": "record_status",
            "valid_from": "valid_from",
            "valid_to": "valid_to",
            "source_page": "source_page",
            "source_region": "source_region",
            "source_row_locator": "source_row_locator",
            "extraction_confidence": "extraction_confidence",
        },
        "value_mappings": {
            "original_term": {"Oltre cinque anni": "long"},
            "residual_term": {"Oltre 1 anno": "over_one_year"},
            "exposure_family": {"RISCHI A SCADENZA": "performing"},
        },
        "control_totals": {"used": "45000"},
        "control_tolerance": "0.01",
        "mapping_review": {
            "status": "reviewed",
            "reviewer": "Test reviewer",
            "reviewed_at": "2026-08-26T12:00:00+02:00",
        },
    }

    analysis = build_analysis(tables, recipe)
    metrics = {item["metric_id"]: item for item in analysis["metrics"]}

    assert [row["record_status"] for row in normalization["tables"]["exposures"]] == [
        "current",
        "previous",
    ]
    assert len(normalization["tables"]["guarantees_received"]) == 1
    assert (
        normalization["tables"]["guarantees_received"][0]["guarantee_value"] == "130000"
    )
    assert normalization["tables"]["guarantors"][0]["guarantor"] == "MARIO GARANTE"
    assert normalization["tables"]["ceded_debtors"][0]["nominal_value"] == "20000"
    assert analysis["source"]["current_row_count"] == 1
    assert analysis["source"]["previous_row_count"] == 1
    assert analysis["exposures"][1]["granted"] == "Assenza di segnalazione"
    assert metrics["cr.total_used"]["value"] == "45000"
    assert metrics["cr.previous_record_count"]["value"] == 1
    assert analysis["guarantees"] == []
    assert len(analysis["guarantees_received"]) == 1
    assert len(analysis["guarantors"]) == 1
    assert len(analysis["ceded_debtors"]) == 1
    assert analysis["coverage"]["other_risk_information"] == "unavailable"
    assert analysis["coverage"]["summary_totals"] == "unavailable"
    assert len(analysis["inframonthly_events"]) == 1
    assert len(analysis["information_requests"]) == 1
    context = build_model_context(analysis)
    assert context["previous_records"][0]["used"] == "60000"
    assert (
        context["previous_records"][0]["source_row_locator"] == "page:1:table:2:row:2"
    )
    assert context["guarantees_received"][0]["guaranteed_party"] == "VIOLA VERDI"
    assert context["guarantors"][0]["guarantor"] == "MARIO GARANTE"
    assert context["ceded_debtors"][0]["ceded_debtor"] == "DEBITORE CEDUTO"
    assert context["inframonthly_events"][0]["event_type"] == "Regolarizzazione"
    assert context["information_requests"][0]["request_reason_code"] == "01"
    assert "source_document_sha256" not in json.dumps(context["guarantees_received"])


def test_pdf_normalization_repairs_exact_fragmented_header_and_old_month_label(
    tmp_path: Path,
) -> None:
    source_pdf = tmp_path / "fragmented-header.pdf"
    _write_fragmented_header_pdf_fixture(source_pdf)

    normalization = normalize_pdf(source_pdf)

    assert len(normalization["tables"]["exposures"]) == 1
    assert normalization["tables"]["exposures"][0]["category"] == "RISCHI A SCADENZA"
    assert normalization["tables"]["exposures"][0]["reference_month"] == "2010-09"
    assert normalization["tables"]["exposures"][0]["used"] == "170000"


def test_pdf_normalization_recovers_exact_clipped_terminal_valid_to(
    tmp_path: Path,
) -> None:
    source_pdf = tmp_path / "clipped-valid-to.pdf"
    _write_clipped_valid_to_pdf_fixture(source_pdf)

    normalization = normalize_pdf(source_pdf)
    previous = normalization["tables"]["exposures"][0]

    assert previous["record_status"] == "previous"
    assert previous["valid_from"] == "02/08/2021"
    assert previous["valid_to"] == "10/08/2021"
    assert previous["extraction_confidence"] == "high"
    assert normalization["issues"] == []


def test_pdf_normalization_quarantines_merged_tables_with_repeated_headers(
    tmp_path: Path,
) -> None:
    source_pdf = tmp_path / "repeated-header.pdf"
    _write_repeated_header_pdf_fixture(source_pdf)

    normalization = normalize_pdf(source_pdf, allow_no_supported_tables=True)

    assert normalization["tables"]["exposures"] == []
    assert len(normalization["unclassified_tables"]) == 1
    assert (
        normalization["unclassified_tables"][0]["reason"] == "repeated_header_in_body"
    )
    assert (
        normalization["unclassified_tables"][0]["review_priority"]
        == "unsupported_data_candidate"
    )


def test_pdf_normalization_recovers_exact_merged_correction_grid(
    tmp_path: Path,
) -> None:
    source_pdf = tmp_path / "merged-correction.pdf"
    _write_merged_correction_pdf_fixture(source_pdf)

    normalization = normalize_pdf(source_pdf)
    rows = normalization["tables"]["exposures"]

    assert len(rows) == 3
    assert [row["record_status"] for row in rows] == [
        "current",
        "previous",
        "previous",
    ]
    assert rows[0]["reference_month"] == "2009-10"
    assert rows[0]["used"] == "60000"
    assert rows[1]["granted"] == "Assenza di segnalazione"
    assert rows[2]["guarantee_type"] == "125"
    assert rows[2]["valid_to"] == "10/12/2009"
    assert normalization["unclassified_tables"] == []
    assert all(
        item["issues"] == ["missing_intermediary"] for item in normalization["issues"]
    )


def test_pdf_normalization_does_not_invent_generic_request_intermediary(
    tmp_path: Path,
) -> None:
    source_pdf = tmp_path / "generic-request.pdf"
    _write_generic_request_pdf_fixture(source_pdf, "Da 02/04/2021 a 31/12/9999")

    normalization = normalize_pdf(source_pdf)
    request = normalization["tables"]["information_requests"][0]

    assert request["intermediary"] == ""
    assert request["extraction_confidence"] == "review_required"
    assert normalization["issues"][0]["issues"] == ["missing_intermediary"]


def test_pdf_normalization_removes_only_exact_validity_watermark_prefix(
    tmp_path: Path,
) -> None:
    source_pdf = tmp_path / "watermarked-request.pdf"
    _write_generic_request_pdf_fixture(source_pdf, "O Da 18/06/2026 a 20/06/2026")

    normalization = normalize_pdf(source_pdf)
    request = normalization["tables"]["information_requests"][0]

    assert request["validity_period"] == "Da 18/06/2026 a 20/06/2026"


def test_pdf_normalization_separates_other_risk_and_summary_populations(
    tmp_path: Path,
) -> None:
    source_pdf = tmp_path / "auxiliary-risk.pdf"
    _write_auxiliary_risk_pdf_fixture(source_pdf)

    normalization = normalize_pdf(source_pdf)

    other = normalization["tables"]["other_risk_information"][0]
    summary = normalization["tables"]["summary_totals"][0]
    assert other["reference_month"] == "2010-09"
    assert other["category"] == "SOFFERENZE - CREDITI PASSATI A PERDITA"
    assert other["amount"] == "30010"
    assert summary["summary_category"] == "Crediti per cassa"
    assert summary["granted"] == "308412"
    assert summary["used"] == "202890"
    assert normalization["unclassified_tables"] == []


def test_pdf_corpus_evaluation_keeps_examples_separate(tmp_path: Path) -> None:
    source_pdf = tmp_path / "centrale-rischi-examples.pdf"
    _write_native_pdf_fixture(source_pdf)

    evaluation = evaluate_pdf_corpus(
        source_pdf,
        cases={
            "exposure-example": (1,),
            "guarantee-example": (2,),
            "other-information": (3, 4),
            "unsupported-example": (5,),
        },
    )

    cases = {item["case_id"]: item for item in evaluation["cases"]}
    assert evaluation["analysis_generated"] is False
    assert "analysis" not in evaluation
    assert cases["exposure-example"]["row_counts"]["exposures"] == 2
    assert cases["guarantee-example"]["row_counts"]["guarantees_received"] == 1
    assert cases["guarantee-example"]["row_counts"]["guarantors"] == 1
    assert cases["guarantee-example"]["row_counts"]["ceded_debtors"] == 1
    assert cases["other-information"]["row_counts"] == {
        "exposures": 0,
        "guarantees_received": 0,
        "guarantors": 0,
        "ceded_debtors": 0,
        "other_risk_information": 0,
        "summary_totals": 0,
        "inframonthly_events": 1,
        "information_requests": 1,
    }
    assert cases["unsupported-example"]["outcome"] == "not_recognized"
    assert cases["unsupported-example"]["unclassified_table_count"] == 1
    assert cases["unsupported-example"]["unsupported_data_candidate_count"] == 0
    assert cases["unsupported-example"]["layout_or_narrative_count"] == 1


def test_pdf_corpus_evaluation_cli_writes_coverage_only(tmp_path: Path) -> None:
    source_pdf = tmp_path / "centrale-rischi-examples.pdf"
    output = tmp_path / "coverage.json"
    _write_native_pdf_fixture(source_pdf)

    result = evaluate_pdf_corpus_cli.main(
        [
            "--input",
            source_pdf.as_posix(),
            "--output",
            output.as_posix(),
            "--case",
            "exposure-example=1",
            "--case",
            "guarantee-example=2",
        ]
    )

    assert result == 0
    payload = json.loads(output.read_text(encoding="utf-8"))
    assert payload["case_count"] == 2
    assert payload["analysis_generated"] is False


def test_gold_benchmark_runs_pdf_case_end_to_end(tmp_path: Path) -> None:
    source_pdf = tmp_path / "centrale-rischi-examples.pdf"
    manifest_path = tmp_path / "gold.json"
    output_dir = tmp_path / "benchmark"
    _write_native_pdf_fixture(source_pdf)
    manifest = {
        "schema_version": gold_benchmark.BENCHMARK_SCHEMA,
        "workflow_id": "centrale-rischi-review",
        "sources": {
            "fixture": {
                "sha256": gold_benchmark.sha256_file(source_pdf),
                "page_count": 5,
                "role": "Test fixture",
            }
        },
        "mapping_profiles": {
            "fixture": {
                "original_term": {"Oltre cinque anni": "long"},
                "residual_term": {"Oltre 1 anno": "over_one_year"},
                "exposure_family": {"RISCHI A SCADENZA": "performing"},
            }
        },
        "negative_control_sources": [],
        "extraction_cases": [
            {
                "case_id": "fixture_p1",
                "source_id": "fixture",
                "pages": [1],
                "row_counts": {"exposures": 2},
                "issue_codes": [],
                "unsupported_data_candidate_count": 0,
                "expected_rows": {
                    "exposures": [
                        {
                            "reference_month": "2026-06",
                            "used": "45000",
                            "record_status": "current",
                        },
                        {
                            "reference_month": "2026-06",
                            "used": "60000",
                            "record_status": "previous",
                        },
                    ]
                },
            }
        ],
        "analysis_cases": [
            {
                "case_id": "analysis_fixture_p1",
                "extraction_case_id": "fixture_p1",
                "source_id": "fixture",
                "pages": [1],
                "entity": "Fixture S.r.l.",
                "analysis_objective": "Test the full benchmark contract.",
                "mapping_profile": "fixture",
                "control_totals": {"used": "45000"},
                "expected": {
                    "outcome": "analysis",
                    "status": "complete",
                    "source_counts": {
                        "current_row_count": 1,
                        "previous_row_count": 1,
                    },
                    "metrics": {
                        "cr.total_used": "45000",
                        "cr.previous_record_count": 1,
                        "cr.original_term.long_share_pct": "100",
                    },
                    "population_counts": {"exposures": 2, "overruns": 0},
                },
            }
        ],
        "semantic_rubric": {"dimensions": {}},
        "semantic_review_cases": [],
        "limitations": ["Fixture only."],
    }
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")

    receipt = gold_benchmark.run_benchmark(
        manifest_path,
        {"fixture": source_pdf},
        output_dir,
    )

    assert receipt["deterministic_passed"] is True
    assert receipt["summary"] == {
        "extraction_total": 1,
        "extraction_passed": 1,
        "analysis_total": 1,
        "analysis_passed": 1,
    }
    assert receipt["overall_status"] == "deterministic_pass_semantic_review_pending"
    assert (output_dir / "benchmark_receipt.json").is_file()
    assert (output_dir / "benchmark_report.html").is_file()
    assert (
        output_dir
        / "analysis"
        / "analysis_fixture_p1"
        / "centrale_rischi_dashboard.html"
    ).is_file()


def test_gold_benchmark_rejects_semantic_review_with_missing_case(
    tmp_path: Path,
) -> None:
    source_pdf = tmp_path / "source.pdf"
    manifest_path = tmp_path / "gold.json"
    review_path = tmp_path / "semantic-review.json"
    _write_native_pdf_fixture(source_pdf)
    review_path.write_text(
        json.dumps(
            {
                "schema_version": gold_benchmark.SEMANTIC_REVIEW_SCHEMA,
                "reviews": [],
            }
        ),
        encoding="utf-8",
    )
    manifest = {
        "schema_version": gold_benchmark.BENCHMARK_SCHEMA,
        "sources": {
            "fixture": {
                "sha256": gold_benchmark.sha256_file(source_pdf),
                "page_count": 5,
                "role": "Test fixture",
            }
        },
        "extraction_cases": [],
        "negative_control_sources": [],
        "analysis_cases": [],
        "semantic_rubric": {"dimensions": {"factual_accuracy": {}}},
        "semantic_review_cases": [{"case_id": "required-case"}],
    }
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")

    with pytest.raises(ValueError, match="missing cases"):
        gold_benchmark.run_benchmark(
            manifest_path,
            {"fixture": source_pdf},
            tmp_path / "output",
            semantic_review_path=review_path,
        )


def test_pdf_inspection_cli_writes_normalized_pipeline_artifacts(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source_pdf = tmp_path / "centrale-rischi.pdf"
    output_dir = tmp_path / "inspection"
    context_path = tmp_path / "context.json"
    _write_native_pdf_fixture(source_pdf)
    context_path.write_text("{}\n", encoding="utf-8")
    monkeypatch.setattr(
        inspect_inputs,
        "load_client_engagement_context_file",
        lambda *args, **kwargs: {},
    )

    result = inspect_inputs.main(
        [
            "--input",
            source_pdf.as_posix(),
            "--output-dir",
            output_dir.as_posix(),
            "--client-engagement",
            context_path.as_posix(),
        ]
    )
    recipe = json.loads(
        (output_dir / "suggested_recipe.json").read_text(encoding="utf-8")
    )
    receipt = json.loads(
        (output_dir / "pdf_normalization_receipt.json").read_text(encoding="utf-8")
    )

    assert result == 0
    assert recipe["source_kind"] == "native_pdf_extraction"
    assert recipe["source_document_sha256"] == receipt["source_document_sha256"]
    assert receipt["normalized_row_counts"] == {
        "exposures": 2,
        "guarantees_received": 1,
        "guarantors": 1,
        "ceded_debtors": 1,
        "other_risk_information": 0,
        "summary_totals": 0,
        "information_requests": 1,
        "inframonthly_events": 1,
    }
    assert (output_dir / "centrale_rischi_normalized.xlsx").is_file()


def test_pdf_analysis_cli_verifies_and_uses_inspected_normalized_workbook(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source_pdf = tmp_path / "centrale-rischi.pdf"
    inspection_dir = tmp_path / "run" / "outputs" / "inspection"
    analysis_dir = tmp_path / "run" / "outputs" / "analysis"
    context_path = tmp_path / "run" / "context.json"
    _write_native_pdf_fixture(source_pdf)
    context_path.parent.mkdir(parents=True)
    context_path.write_text("{}\n", encoding="utf-8")
    monkeypatch.setattr(
        inspect_inputs,
        "load_client_engagement_context_file",
        lambda *args, **kwargs: {},
    )
    assert (
        inspect_inputs.main(
            [
                "--input",
                source_pdf.as_posix(),
                "--output-dir",
                inspection_dir.as_posix(),
                "--client-engagement",
                context_path.as_posix(),
            ]
        )
        == 0
    )
    inspection = json.loads(
        (inspection_dir / "inspection.json").read_text(encoding="utf-8")
    )
    recipe = json.loads(
        (inspection_dir / "suggested_recipe.json").read_text(encoding="utf-8")
    )
    recipe.update(
        {
            "entity": "Synthetic S.r.l.",
            "analysis_objective": "Validate the direct PDF handoff.",
            "table_id": inspection["tables"][0]["table_id"],
            "columns": {
                "reference_month": "reference_month",
                "intermediary": "intermediary",
                "risk_category": "category",
                "original_duration": "original_duration",
                "residual_duration": "residual_duration",
                "granted": "granted",
                "operational_granted": "operational_granted",
                "used": "used",
                "guarantee_type": "guarantee_type",
                "guaranteed_amount": "guaranteed_amount",
                "record_status": "record_status",
                "valid_from": "valid_from",
                "valid_to": "valid_to",
                "source_page": "source_page",
                "source_region": "source_region",
                "source_row_locator": "source_row_locator",
                "extraction_confidence": "extraction_confidence",
            },
            "value_mappings": {
                "original_term": {"Oltre cinque anni": "long"},
                "residual_term": {"Oltre 1 anno": "over_one_year"},
                "exposure_family": {"RISCHI A SCADENZA": "performing"},
            },
            "control_totals": {"used": "45000"},
            "mapping_review": {
                "status": "reviewed",
                "reviewer": "Test reviewer",
                "reviewed_at": "2026-08-26T12:00:00+02:00",
            },
        }
    )
    recipe_path = inspection_dir / "reviewed_recipe.json"
    recipe_path.write_text(json.dumps(recipe), encoding="utf-8")
    validated_inputs: list[list[Path]] = []

    def validate_context(*args: object, **kwargs: object) -> dict[str, object]:
        validated_inputs.append(list(kwargs["input_paths"]))
        return {}

    monkeypatch.setattr(
        run_analysis, "load_client_engagement_context_file", validate_context
    )

    result = run_analysis.main(
        [
            "--input",
            source_pdf.as_posix(),
            "--recipe",
            recipe_path.as_posix(),
            "--output-dir",
            analysis_dir.as_posix(),
            "--client-engagement",
            context_path.as_posix(),
        ]
    )
    analysis = json.loads(
        (analysis_dir / "centrale_rischi_analysis.json").read_text(encoding="utf-8")
    )
    metrics = {item["metric_id"]: item for item in analysis["metrics"]}

    assert result == 0
    assert metrics["cr.total_used"]["value"] == "45000"
    assert len(validated_inputs) == 2
    assert inspection_dir / "centrale_rischi_normalized.xlsx" in validated_inputs[1]
    assert inspection_dir / "pdf_normalization_receipt.json" in validated_inputs[1]


def test_reconciled_mode_requires_external_evidence_adapter(tmp_path: Path) -> None:
    source = tmp_path / "cr.xlsx"
    _write_source(source)
    tables = load_source_tables([source])
    inspection, _, _ = build_inspection(tables)
    recipe = _reviewed_recipe(inspection)
    recipe["analysis_mode"] = "reconciled"

    with pytest.raises(CentraleRischiContractError, match="external-evidence"):
        build_analysis(tables, recipe)


def test_failed_declared_control_blocks_analysis(tmp_path: Path) -> None:
    source = tmp_path / "cr.xlsx"
    _write_source(source)
    tables = load_source_tables([source])
    inspection, _, _ = build_inspection(tables)
    recipe = _reviewed_recipe(inspection)
    recipe["control_totals"]["used"] = "1499"

    analysis = build_analysis(tables, recipe)

    assert analysis["status"] == "blocked"
    assert analysis["controls"][0]["status"] == "failed"


def test_model_context_is_bounded_and_excludes_source_paths(tmp_path: Path) -> None:
    source = tmp_path / "cr.xlsx"
    _write_source(source)
    analysis, _ = _analysis_for(source)
    auxiliary_row = {
        "intermediary": "Banca test",
        "source_document_sha256": "a" * 64,
    }
    analysis["guarantees_received"] = [auxiliary_row] * 25
    analysis["guarantors"] = [auxiliary_row] * 25
    analysis["ceded_debtors"] = [auxiliary_row] * 25
    analysis["other_risk_information"] = [auxiliary_row] * 25
    analysis["summary_totals"] = [auxiliary_row] * 25
    analysis["inframonthly_events"] = [auxiliary_row] * 25
    analysis["information_requests"] = [auxiliary_row] * 25

    context = build_model_context(analysis)

    assert "exposures" not in context
    assert "absolute_path" not in json.dumps(context)
    assert len(context["top_overruns"]) <= 20
    assert len(context["monthly_series"]) <= 36
    assert len(context["category_movement_summary"]) <= 50
    assert len(context["previous_records"]) <= 20
    assert len(context["guarantees_received"]) == 20
    assert len(context["guarantors"]) == 20
    assert len(context["ceded_debtors"]) == 20
    assert len(context["other_risk_information"]) == 20
    assert len(context["summary_totals"]) == 20
    assert len(context["inframonthly_events"]) == 20
    assert len(context["information_requests"]) == 20
    assert "source_document_sha256" not in json.dumps(context)


def test_commentary_requires_existing_metric_references(tmp_path: Path) -> None:
    source = tmp_path / "cr.xlsx"
    _write_source(source)
    analysis, _ = _analysis_for(source)
    commentary = {
        "schema_version": COMMENTARY_SCHEMA,
        "workflow_id": "centrale-rischi-review",
        "observations": [
            {"text": "Utilizzo elevato.", "evidence_refs": ["metric:missing"]}
        ],
        "hypotheses": [],
        "questions": [],
        "limitations": [],
    }

    with pytest.raises(CentraleRischiContractError, match="existing evidence_refs"):
        finalize_commentary(analysis, commentary)


@pytest.mark.parametrize(
    "evidence_ref",
    (
        "metric:cr.total_used",
        "control:control.latest.used",
        "row:sheet:CR:row:2",
    ),
)
def test_commentary_accepts_closed_evidence_reference_types(
    tmp_path: Path, evidence_ref: str
) -> None:
    source = tmp_path / "cr.xlsx"
    _write_source(source)
    analysis, _ = _analysis_for(source)
    analysis["overruns"][0]["source_row_locator"] = "sheet:CR:row:2"
    commentary = {
        "schema_version": COMMENTARY_SCHEMA,
        "workflow_id": "centrale-rischi-review",
        "observations": [
            {"text": "Fatto collegato all'evidenza.", "evidence_refs": [evidence_ref]}
        ],
        "hypotheses": [],
        "questions": [],
        "limitations": [],
    }

    finalized = finalize_commentary(analysis, commentary)

    assert finalized["observations"][0]["evidence_refs"] == [evidence_ref]


def test_renderers_create_reviewable_html_and_excel(tmp_path: Path) -> None:
    source = tmp_path / "cr.xlsx"
    _write_source(source)
    analysis, _ = _analysis_for(source)
    output = tmp_path / "analysis.xlsx"

    write_excel(output, analysis)
    commentary = finalize_commentary(
        analysis,
        {
            "schema_version": COMMENTARY_SCHEMA,
            "workflow_id": "centrale-rischi-review",
            "observations": [
                {
                    "text": "Lo sconfinamento richiede verifica.",
                    "evidence_refs": ["metric:cr.overrun_amount"],
                }
            ],
            "hypotheses": [],
            "questions": ["Lo sconfinamento è stato regolarizzato?"],
            "limitations": ["Manca il confronto con il bilancio."],
        },
    )
    rendered = render_html(analysis, commentary)
    workbook = load_workbook(output)

    assert "draft_pending_professional_review" not in rendered
    assert "Bozza in attesa di revisione professionale" in rendered
    assert "Centrale Rischi" in rendered
    assert "1.500" in rendered
    assert "La Centrale Rischi da sola" in rendered
    assert "Requires reviewed" not in rendered
    assert "non rilevante" in rendered
    assert "Lo sconfinamento è stato regolarizzato?" in rendered
    assert "Manca il confronto con il bilancio." in rendered
    assert "Variazione tra gli ultimi due periodi per categoria" in rendered
    assert "Utilizzato — Autoliquidante" not in rendered
    assert '<details class="evidence">' in rendered
    assert "<summary>Evidenze</summary>" in rendered
    assert rendered.index("Commento professionale — bozza") < rendered.index(
        "Esposizioni per durata originaria"
    )
    assert "Garanti dell&#x27;intestatario" in rendered
    assert "Debitori ceduti" in rendered
    assert "Altre informazioni di rischio" in rendered
    assert "Prospetto sintetico" in rendered
    assert "Pregiudizievoli: 0" not in rendered
    assert set(workbook.sheetnames) == {
        "KPI",
        "Durata originaria",
        "Durata residua",
        "Categorie",
        "Variazione categorie",
        "Esposizioni",
        "Garanzie",
        "Garanzie ricevute",
        "Garanti intestatario",
        "Debitori ceduti",
        "Altre informazioni",
        "Prospetto sintetico",
        "Sconfinamenti",
        "Eventi inframensili",
        "Richieste informazioni",
        "Pregiudizievoli",
        "Serie mensile",
        "Controlli",
    }
    assert workbook["KPI"]["C2"].data_type == "n"
    assert workbook["KPI"]["C2"].number_format == "#,##0"
    assert [cell.value for cell in workbook["Garanzie"][1]] == [
        "reference_month",
        "intermediary",
        "risk_category",
        "guarantee_type",
        "guaranteed_amount",
        "record_status",
        "source_page",
        "source_row_locator",
        "extraction_confidence",
    ]
    assert workbook["Esposizioni"].freeze_panes == "D2"
    exposure_headers = {
        cell.value: cell.column_letter for cell in workbook["Esposizioni"][1]
    }
    assert (
        workbook["Esposizioni"].column_dimensions[exposure_headers["source_row"]].hidden
    )


def test_html_distinguishes_unavailable_evidence_from_zero(tmp_path: Path) -> None:
    source = tmp_path / "cr.xlsx"
    _write_source(source)
    tables = load_source_tables([source])
    inspection, _, _ = build_inspection(tables)
    recipe = _reviewed_recipe(inspection)
    recipe["columns"]["prejudicial_event"] = ""
    analysis = build_analysis(tables, recipe)

    rendered = render_html(analysis)

    assert analysis["coverage"]["pregiudizievoli"] == "unavailable"
    assert "Pregiudizievoli: 0" not in rendered
    assert (
        "Non disponibile: nessuna riga è stata estratta o fornita per questa "
        "popolazione." in rendered
    )


def test_vera_launcher_uses_the_complete_component_registry() -> None:
    scripts = ROOT / "plugins" / "vera" / "scripts"
    if str(scripts) not in sys.path:
        sys.path.insert(0, str(scripts))
    spec = importlib.util.spec_from_file_location(
        "vera_check_dependencies_for_cr_test", scripts / "check_dependencies.py"
    )
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    registry = json.loads(
        (ROOT / "plugins" / "vera" / "components.json").read_text(encoding="utf-8")
    )["plugins"]

    assert module.COMPONENTS == tuple(registry)
    assert "centrale-rischi-review" in module.COMPONENTS


def test_skill_treats_ocr_as_outside_the_centrale_rischi_workflow() -> None:
    skill = (PLUGIN_ROOT / "skills" / "centrale-rischi-review" / "SKILL.md").read_text(
        encoding="utf-8"
    )

    assert "OCR is not part of this workflow" in skill
    assert "instead of starting an OCR path" in skill
    assert "OCR is a separate adapter" not in skill


def test_public_page_and_privacy_surface_are_registered() -> None:
    page = (
        ROOT / "static" / "shared" / "centrale-rischi-review" / "index.html"
    ).read_text(encoding="utf-8")
    manifest = json.loads(
        (
            ROOT
            / "plugins"
            / "vera"
            / "privacy"
            / "workstreams"
            / "centrale-rischi-review.json"
        ).read_text(encoding="utf-8")
    )

    main = page[page.index('<main class="page-shell"') : page.index("</main>")]
    assert 'data-model-data-workflow="centrale-rischi-review"' in main
    assert 'data-model-data-status="relevant"' in main
    assert main.rstrip().endswith("</section>")
    assert "Quali dati arrivano al modello" in page
    assert "What data reaches the model" in page
    assert "garanti dell’intestatario" in page
    assert "debitori ceduti" in page
    assert "summary totals" in page
    assert "OCR" not in page
    assert (
        'href="https://centrale-rischi-synthetic.fabio3143.chatgpt.site/"'
        in page
    )
    assert 'id="example"' in page
    assert manifest["workstream"] == "centrale-rischi-review"
    assert manifest["external_boundaries"] == []
    mapping_class, review_class, commentary_class = manifest["model_context"][
        "classes"
    ]
    assert "stable table IDs" in mapping_class["content"]
    assert "each table's source SHA-256 hash" in mapping_class["content"]
    assert "inventory SHA-256 hash" in mapping_class["content"]
    assert "ceded debtors" in review_class["content"]
    assert "entity name" in review_class["content"]
    assert "up to 50 category-movement rows" in review_class["content"]
    assert "up to 20 previous-record review rows" in review_class["content"]
    assert "up to 20 rows from each separate population" in review_class["content"]
    assert commentary_class["purpose"].startswith("Prepare an evidence-linked")
    assert "metric IDs, control IDs or source-row locators" in commentary_class[
        "content"
    ]
    assert "evidence-reference closure" in commentary_class["content"]
    assert "metric-linked" not in commentary_class["purpose"]
    assert "metric-reference closure" not in commentary_class["content"]
    for phrase in (
        "identificatori stabili delle tabelle",
        "stable table identifiers",
        "identifiants stables des tables",
        "stabile Tabellenkennungen",
        "identificadores estables de las tablas",
        "hash SHA-256 della fonte di ciascuna tabella e dell’inventario",
        "source SHA-256 for each table and the inventory SHA-256",
        "SHA-256 de la source de chaque table et le SHA-256 de l’inventaire",
        "Quell-SHA-256 jeder Tabelle und den Inventar-SHA-256",
        "SHA-256 de origen de cada tabla y el SHA-256 del inventario",
    ):
        assert phrase in page
    assert "metriche, controlli o righe fonte già presenti" in page
    assert "metrics, controls or source rows already present" in page
