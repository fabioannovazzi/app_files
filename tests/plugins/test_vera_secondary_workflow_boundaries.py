from __future__ import annotations

import ast
import importlib.util
import json
import os
import shutil
import subprocess
import sys
from pathlib import Path
from types import ModuleType

import pytest
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[2]
AUDIT_SCRIPTS = ROOT / "plugins" / "open-item-reconciliation" / "scripts"
JOURNAL_SCRIPTS = ROOT / "plugins" / "journal-sampling" / "scripts"
SECONDARY_ENTRYPOINTS = (
    ("open-item-reconciliation", AUDIT_SCRIPTS / "audit_assurance.py"),
    (
        "open-item-reconciliation",
        AUDIT_SCRIPTS / "build_missing_evidence_requests.py",
    ),
    ("open-item-reconciliation", AUDIT_SCRIPTS / "build_review_sample.py"),
    ("journal-sampling", JOURNAL_SCRIPTS / "replay_normalization.py"),
    ("journal-sampling", JOURNAL_SCRIPTS / "review_successor.py"),
)


def _load_customer_ledger() -> ModuleType:
    path = ROOT / "plugins" / "studio-archive" / "scripts" / "client_ledger.py"
    module_name = "test_secondary_workflow_customer_ledger"
    existing = sys.modules.get(module_name)
    if existing is not None:
        return existing
    spec = importlib.util.spec_from_file_location(module_name, path)
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules[module_name] = module
    spec.loader.exec_module(module)
    return module


def _running_context(tmp_path: Path, workflow_id: str) -> tuple[Path, Path]:
    ledger = _load_customer_ledger()
    client_root = tmp_path / "Customer"
    client_root.mkdir()
    client_id = "client_444444444444444444444444"
    ledger.create_client_manifest(client_root, client_id)
    engagement = ledger.create_engagement(client_root, client_id, "Review helpers")
    source = tmp_path / "source.txt"
    source.write_text("exact source\n", encoding="utf-8")
    imported = ledger.import_document(
        client_root,
        client_id,
        engagement["engagement_id"],
        source,
        "source",
    )
    prepared = ledger.prepare_run(
        client_root,
        client_id,
        engagement["engagement_id"],
        workflow_id,
        "test-version",
        input_ids=[imported["receipt"]["input_id"]],
    )
    running = ledger.start_run(
        client_root,
        engagement["engagement_id"],
        prepared["run"]["run_id"],
    )
    return Path(running["context_path"]), Path(running["output_dir"])


def _write_reconciliation_workbook(path: Path) -> None:
    workbook = Workbook()
    detail = workbook.active
    detail.title = "Reconciliation detail"
    detail.append(
        [
            "record_id",
            "reconciliation_status",
            "document_no",
            "document_date",
            "amount",
            "currency",
            "expected_side",
        ]
    )
    detail.append(
        ["row-1", "open_supported", "INV-1", "2026-01-01", "100.00", "EUR", "customer"]
    )
    source_inventory = workbook.create_sheet("Source inventory")
    source_inventory.append(["source_role", "source_file"])
    source_inventory.append(["open_items", "open-items.xlsx"])
    normalized = workbook.create_sheet("Normalized records")
    normalized.append(["source_role", "source_file"])
    normalized.append(["open_items", "open-items.xlsx"])
    workbook.save(path)


@pytest.mark.parametrize(("workflow_id", "script_path"), SECONDARY_ENTRYPOINTS)
def test_secondary_entrypoint_requires_direct_customer_run_loader(
    workflow_id: str,
    script_path: Path,
) -> None:
    tree = ast.parse(script_path.read_text(encoding="utf-8"), filename=str(script_path))
    client_arguments = [
        node
        for node in ast.walk(tree)
        if isinstance(node, ast.Call)
        and isinstance(node.func, ast.Attribute)
        and node.func.attr == "add_argument"
        and any(
            isinstance(argument, ast.Constant)
            and argument.value == "--client-engagement"
            for argument in node.args
        )
    ]
    loader_calls = [
        node
        for node in ast.walk(tree)
        if isinstance(node, ast.Call)
        and isinstance(node.func, ast.Name)
        and node.func.id == "load_client_engagement_context_file"
    ]

    assert workflow_id in {"open-item-reconciliation", "journal-sampling"}
    assert len(client_arguments) == 1
    assert any(
        keyword.arg == "required"
        and isinstance(keyword.value, ast.Constant)
        and keyword.value.value is True
        for keyword in client_arguments[0].keywords
    )
    assert loader_calls


def test_audit_review_sample_rejects_unreceipted_workbook(tmp_path: Path) -> None:
    context_path, output_dir = _running_context(tmp_path, "open-item-reconciliation")
    workbook_path = tmp_path / "unreceipted.xlsx"
    _write_reconciliation_workbook(workbook_path)

    completed = subprocess.run(
        [
            sys.executable,
            "-I",
            "-B",
            str(AUDIT_SCRIPTS / "build_review_sample.py"),
            str(workbook_path),
            "--output-dir",
            str(output_dir),
            "--client-engagement",
            str(context_path),
        ],
        capture_output=True,
        text=True,
        check=False,
    )

    assert completed.returncode == 2
    assert "not one of the run's exact receipts" in completed.stderr
    assert not (output_dir / "campione_movimenti_da_controllare.xlsx").exists()


def test_audit_review_sample_continues_after_customer_folder_rename(
    tmp_path: Path,
) -> None:
    context_path, output_dir = _running_context(tmp_path, "open-item-reconciliation")
    workbook_path = output_dir / "riconciliazione_audit.xlsx"
    _write_reconciliation_workbook(workbook_path)
    original_client_root = context_path.parents[5]
    renamed_client_root = original_client_root.with_name("Renamed Customer")
    original_client_root.rename(renamed_client_root)
    renamed_context = renamed_client_root / context_path.relative_to(
        original_client_root
    )
    renamed_output = renamed_client_root / output_dir.relative_to(original_client_root)
    renamed_workbook = renamed_client_root / workbook_path.relative_to(
        original_client_root
    )

    completed = subprocess.run(
        [
            sys.executable,
            "-I",
            "-B",
            str(AUDIT_SCRIPTS / "build_review_sample.py"),
            str(renamed_workbook),
            "--client-engagement",
            str(renamed_context),
        ],
        capture_output=True,
        text=True,
        check=False,
    )

    assert completed.returncode == 0, completed.stderr
    assert (renamed_output / "campione_movimenti_da_controllare.xlsx").is_file()
    assert (renamed_output / "testo_richiesta_controllo.md").is_file()


def test_audit_missing_evidence_helper_writes_inside_running_output(
    tmp_path: Path,
) -> None:
    context_path, output_dir = _running_context(tmp_path, "open-item-reconciliation")
    workbook_path = output_dir / "riconciliazione_audit.xlsx"
    _write_reconciliation_workbook(workbook_path)

    completed = subprocess.run(
        [
            sys.executable,
            "-I",
            "-B",
            str(AUDIT_SCRIPTS / "build_missing_evidence_requests.py"),
            str(workbook_path),
            "--client-engagement",
            str(context_path),
        ],
        capture_output=True,
        text=True,
        check=False,
    )

    assert completed.returncode == 0, completed.stderr
    assert (output_dir / "richieste_mirate_evidenze.xlsx").is_file()


def test_audit_assurance_context_gate_continues_after_rename(tmp_path: Path) -> None:
    context_path, output_dir = _running_context(tmp_path, "open-item-reconciliation")
    original_client_root = context_path.parents[5]
    renamed_client_root = original_client_root.with_name("Renamed Audit Customer")
    original_client_root.rename(renamed_client_root)
    renamed_context = renamed_client_root / context_path.relative_to(
        original_client_root
    )
    renamed_output = renamed_client_root / output_dir.relative_to(original_client_root)

    completed = subprocess.run(
        [
            sys.executable,
            "-I",
            "-B",
            str(AUDIT_SCRIPTS / "audit_assurance.py"),
            "--client-engagement",
            str(renamed_context),
            "validate-context-json",
            str(renamed_output),
        ],
        capture_output=True,
        text=True,
        check=False,
    )

    assert completed.returncode == 0, completed.stderr


def test_journal_replay_classifies_private_receipt_as_read_only() -> None:
    source = (JOURNAL_SCRIPTS / "replay_normalization.py").read_text(encoding="utf-8")
    tree = ast.parse(
        source,
        filename=str(JOURNAL_SCRIPTS / "replay_normalization.py"),
    )
    loader_call = next(
        node
        for node in ast.walk(tree)
        if isinstance(node, ast.Call)
        and isinstance(node.func, ast.Name)
        and node.func.id == "load_client_engagement_context_file"
    )
    keywords = {keyword.arg for keyword in loader_call.keywords}

    assert "input_paths" in keywords
    assert "output_dir" in keywords
    assert "allowed_statuses" in keywords
    assert '"--read-only-upstream"' in source


def test_journal_replay_writer_rejects_external_receipt_output(
    tmp_path: Path,
) -> None:
    context_path, output_dir = _running_context(tmp_path, "journal-sampling")
    normalized_csv = output_dir / "normalized_journal.csv"
    normalized_csv.write_text("record_id,amount\nrow-1,1.00\n", encoding="utf-8")
    external_receipt = tmp_path / "external-replay.json"

    completed = subprocess.run(
        [
            sys.executable,
            "-I",
            "-B",
            str(JOURNAL_SCRIPTS / "replay_normalization.py"),
            str(normalized_csv),
            "--receipt-out",
            str(external_receipt),
            "--client-engagement",
            str(context_path),
        ],
        capture_output=True,
        text=True,
        check=False,
    )

    assert completed.returncode == 2
    assert "outside the client engagement run" in completed.stderr
    assert not external_receipt.exists()


def test_journal_review_successor_context_continues_after_rename(
    tmp_path: Path,
) -> None:
    context_path, output_dir = _running_context(tmp_path, "journal-sampling")
    original_client_root = context_path.parents[5]
    renamed_client_root = original_client_root.with_name("Renamed Journal Customer")
    original_client_root.rename(renamed_client_root)
    renamed_context = renamed_client_root / context_path.relative_to(
        original_client_root
    )
    renamed_output = renamed_client_root / output_dir.relative_to(original_client_root)

    completed = subprocess.run(
        [
            sys.executable,
            "-I",
            "-B",
            str(JOURNAL_SCRIPTS / "review_successor.py"),
            "context",
            str(renamed_output),
            "--client-engagement",
            str(renamed_context),
        ],
        capture_output=True,
        text=True,
        check=False,
    )

    assert completed.returncode == 0, completed.stderr


@pytest.mark.parametrize(
    ("workflow_id", "tool_name", "item_id"),
    [
        (
            "open-item-reconciliation",
            "save_open_item_reconciliation_decisions",
            "review:closed",
        ),
    ],
)
@pytest.mark.parametrize("mismatched_artifact", ["run_intake", "review_payload"])
def test_mcp_review_save_rejects_customer_run_id_mismatch_without_writes(
    tmp_path: Path,
    workflow_id: str,
    tool_name: str,
    item_id: str,
    mismatched_artifact: str,
) -> None:
    node = shutil.which("node")
    if node is None:
        pytest.skip("Node.js is required for the MCP customer-run boundary test.")
    context_path, output_dir = _running_context(tmp_path, workflow_id)
    context = json.loads(context_path.read_text(encoding="utf-8"))
    run_id = str(context["run_id"])
    wrong_run_id = "run_999999999999999999999999"
    run_intake = {
        "run_id": wrong_run_id if mismatched_artifact == "run_intake" else run_id,
        "output_dir": str(output_dir),
        "language": "en",
    }
    review_payload = {
        "schema_version": "1.0",
        "plugin": workflow_id,
        "workflow": workflow_id,
        "run_id": wrong_run_id if mismatched_artifact == "review_payload" else run_id,
        "review_type": f"{workflow_id.replace('-', '_')}_review",
        "items": [
            {
                "id": item_id,
                "item_type": "review_control",
                "title": "Review control",
                "allowed_actions": ["accept"],
                "recommended_action": "accept",
                "status": "needs_review",
            }
        ],
        "item_count": 1,
        "status": "ready_for_review",
    }
    final_artifacts = {
        "schema_version": "1.0",
        "plugin": workflow_id,
        "workflow": workflow_id,
        "run_id": run_id,
        "outputs": [],
        "next_actions": [],
    }
    for name, payload in (
        ("run_intake.json", run_intake),
        ("review_payload.json", review_payload),
        ("final_artifacts.json", final_artifacts),
    ):
        (output_dir / name).write_text(
            json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
    before = {
        path.relative_to(output_dir).as_posix(): path.read_bytes()
        for path in output_dir.rglob("*")
        if path.is_file()
    }
    server_path = ROOT / "plugins" / workflow_id / "mcp" / "server.cjs"
    request = {
        "jsonrpc": "2.0",
        "id": 1,
        "method": "tools/call",
        "params": {
            "name": tool_name,
            "arguments": {
                "run_intake": run_intake,
                "review_payload": review_payload,
                "final_artifacts": final_artifacts,
                "decisions": [{"item_id": item_id, "action": "accept"}],
            },
        },
    }

    completed = subprocess.run(
        [node, str(server_path), "--stdio"],
        input=json.dumps(request) + "\n",
        capture_output=True,
        text=True,
        check=False,
        timeout=20,
        env={**os.environ},
    )

    response = json.loads(completed.stdout.strip())
    error = response["result"]["structuredContent"]["error"].lower()
    after = {
        path.relative_to(output_dir).as_posix(): path.read_bytes()
        for path in output_dir.rglob("*")
        if path.is_file()
    }
    assert completed.returncode == 0, completed.stderr
    assert "does not match the customer-run context" in error
    assert after == before
    assert not (output_dir / "ui_decisions.json").exists()
