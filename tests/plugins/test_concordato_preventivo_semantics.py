from __future__ import annotations

import importlib
import importlib.util
import json
import sys
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl
import pytest
from docx import Document

ROOT = Path(__file__).resolve().parents[2]
SCRIPT_DIR = ROOT / "plugins" / "concordato-plan-review" / "scripts"
CORE_PATH = SCRIPT_DIR / "concordato_plan_core.py"


def _load_core() -> Any:
    if str(SCRIPT_DIR) not in sys.path:
        sys.path.insert(0, str(SCRIPT_DIR))
    spec = importlib.util.spec_from_file_location(
        "concordato_preventivo_semantic_core",
        CORE_PATH,
    )
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def _load_semantic() -> Any:
    if str(SCRIPT_DIR) not in sys.path:
        sys.path.insert(0, str(SCRIPT_DIR))
    return importlib.import_module("concordato_semantic")


def _save_workbook(path: Path, rows: list[list[Any]]) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Evidence"
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def _source_map(inventory: list[dict[str, Any]]) -> dict[str, str]:
    return {
        str(row["relative_path"]): str(row["source_artifact_ref"])
        for row in inventory
        if row.get("source_artifact_ref")
    }


def _evidence(source_ref: str, locator: str) -> list[dict[str, str]]:
    return [{"source_artifact_ref": source_ref, "locator": locator}]


def _reviewed_case_model(
    semantic: Any,
    inventory: list[dict[str, Any]],
    *,
    plan_type: str = "continuity_direct",
    missing_attestation: bool = False,
    unbalanced: bool = False,
    broken_cash_bridge: bool = False,
) -> dict[str, Any]:
    model = semantic.build_case_model_template(
        inventory,
        reference_date="2026-03-31",
        language="it",
    )
    sources = _source_map(inventory)
    plan_ref = sources["case_material_a.xlsx"]
    attestation_ref = sources["case_material_b.xlsx"]
    creditors_ref = sources["case_material_c.xlsx"]
    model["legal_framework"] = {
        "jurisdiction": "IT",
        "instrument": "concordato_preventivo",
        "framework_name": "Codice della crisi d'impresa e dell'insolvenza",
        "as_of_date": "2026-03-31",
        "authority_refs": [
            {
                "title": "D.Lgs. 12 gennaio 2019, n. 14",
                "url": (
                    "https://www.normattiva.it/eli/stato/"
                    "DECRETO_LEGISLATIVO/2019/01/12/14"
                ),
                "provisions": ["artt. 84-120 CCII"],
            }
        ],
        "judgment_basis": (
            "Framework selected by the reviewer for this Italian procedure "
            "at the stated as-of date."
        ),
    }
    model["procedure"] = {
        "identification_status": "complete",
        "debtor_name": "Northwind Restructuring S.p.A.",
        "court": "Tribunale competente",
        "procedure_reference": "CP 17/2026",
        "stage": "filed",
        "plan_type": plan_type,
        "reference_date": "2026-03-31",
        "currency": "EUR",
        "judgment_basis": "Confirmed from the reviewed filing cover and plan.",
    }
    roles_by_path = {
        "case_material_a.xlsx": {
            "roles": ["proposal", "plan", "business_plan"],
            "authoritative_for": ["proposal", "plan"],
        },
        "case_material_b.xlsx": {
            "roles": ["other_support"] if missing_attestation else ["attestation"],
            "authoritative_for": [] if missing_attestation else ["attestation"],
        },
        "case_material_c.xlsx": {
            "roles": ["creditor_schedule", "tax_social_security_schedule"],
            "authoritative_for": ["creditor_schedule"],
        },
    }
    for document in model["document_perimeter"]["documents"]:
        mapping = roles_by_path[document["relative_path"]]
        document.update(
            {
                **mapping,
                "version_date": "2026-03-31",
                "judgment_basis": (
                    "Role confirmed from document content and internal title, "
                    "not from its filename."
                ),
            }
        )
    model["document_perimeter"]["status"] = "complete"
    model["document_perimeter"]["judgment_basis"] = (
        "Every captured source was classified and the supplied folder perimeter "
        "was reviewed."
    )
    model["creditor_population"] = {
        "status": "complete",
        "cutoff_date": "2026-03-31",
        "currency": "EUR",
        "judgment_basis": "Population agreed to the reviewed creditor schedule.",
        "creditors": [
            {
                "creditor_id": "creditor-1",
                "creditor_name": "Secured Bank",
                "claim_amount": "600.00",
                "claim_status": "admitted",
                "priority": "secured",
                "class_id": "class-secured",
                "treatment_form": "cash",
                "proposed_cash_amount": "540.00",
                "proposed_non_cash_amount": "0",
                "liquidation_recovery_amount": "480.00",
                "payment_start": "2026-06-30",
                "payment_end": "2026-12-31",
                "voting_treatment": "voting",
                "evidence_refs": _evidence(creditors_ref, "Evidence!A2:N2"),
                "judgment_basis": "Reviewer confirmed claim and treatment row.",
            },
            {
                "creditor_id": "creditor-2",
                "creditor_name": "Trade Supplier",
                "claim_amount": "300.00",
                "claim_status": "asserted",
                "priority": "unsecured",
                "class_id": "class-trade",
                "treatment_form": "cash",
                "proposed_cash_amount": "120.00",
                "proposed_non_cash_amount": "0",
                "liquidation_recovery_amount": "75.00",
                "payment_start": "2027-01-31",
                "payment_end": "2027-12-31",
                "voting_treatment": "voting",
                "evidence_refs": _evidence(creditors_ref, "Evidence!A3:N3"),
                "judgment_basis": "Reviewer confirmed claim and treatment row.",
            },
            {
                "creditor_id": "creditor-3",
                "creditor_name": "Tax Authority",
                "claim_amount": "100.00",
                "claim_status": "asserted",
                "priority": "privileged",
                "class_id": "class-tax",
                "treatment_form": "mixed",
                "proposed_cash_amount": "70.00",
                "proposed_non_cash_amount": "10.00",
                "liquidation_recovery_amount": "65.00",
                "payment_start": "2026-09-30",
                "payment_end": "2028-03-31",
                "voting_treatment": "partially_voting",
                "evidence_refs": _evidence(creditors_ref, "Evidence!A4:N4"),
                "judgment_basis": (
                    "Reviewer classified priority and voting treatment; code "
                    "does not infer either."
                ),
            },
        ],
    }
    model["sources_and_uses"] = {
        "status": "complete",
        "currency": "EUR",
        "balance_tolerance": "0.01",
        "judgment_basis": "Schedule reconstructed from reviewed plan tables.",
        "items": [
            {
                "item_id": "source-operations",
                "side": "source",
                "category": "continuity cash generation",
                "description": "Cash generated during the plan",
                "amount": "500.00",
                "period": "2026-2028",
                "evidence_refs": _evidence(plan_ref, "Evidence!B2"),
                "judgment_basis": "Reviewer mapped the plan cash source.",
            },
            {
                "item_id": "source-finance",
                "side": "source",
                "category": "new finance",
                "description": "Committed new finance",
                "amount": "250.00",
                "period": "2026",
                "evidence_refs": _evidence(plan_ref, "Evidence!B3"),
                "judgment_basis": "Reviewer mapped the finance source.",
            },
            {
                "item_id": "use-distributions",
                "side": "use",
                "category": "creditor distributions",
                "description": "Cash distributions to creditors",
                "amount": "730.00" if unbalanced else "710.00",
                "period": "2026-2028",
                "evidence_refs": _evidence(plan_ref, "Evidence!B4"),
                "judgment_basis": "Reviewer mapped planned cash distributions.",
            },
            {
                "item_id": "use-costs",
                "side": "use",
                "category": "procedure costs",
                "description": "Procedure and implementation costs",
                "amount": "40.00",
                "period": "2026-2028",
                "evidence_refs": _evidence(plan_ref, "Evidence!B5"),
                "judgment_basis": "Reviewer mapped procedure costs.",
            },
        ],
    }
    model["liquidity"] = {
        "status": "complete",
        "currency": "EUR",
        "bridge_tolerance": "0.01",
        "judgment_basis": "Monthly bridge reconstructed from the reviewed plan.",
        "periods": [
            {
                "period_id": "period-1",
                "period": "2026-Q2",
                "opening_cash": "50.00",
                "operating_inflows": "200.00",
                "other_inflows": "0",
                "new_finance_inflows": "100.00",
                "operating_outflows": "150.00",
                "procedure_costs": "20.00",
                "creditor_distributions": "80.00",
                "financing_outflows": "0",
                "other_outflows": "0",
                "reported_closing_cash": "101.00" if broken_cash_bridge else "100.00",
                "evidence_refs": _evidence(plan_ref, "Evidence!C2:L2"),
                "judgment_basis": "Reviewer mapped each bridge component.",
            }
        ],
    }
    model["milestones"] = [
        {
            "milestone_id": "milestone-1",
            "date_or_period": "2026-Q3",
            "description": "First creditor distribution",
            "status": "planned",
            "evidence_refs": _evidence(plan_ref, "Evidence!M2"),
            "judgment_basis": "Reviewer mapped the stated implementation date.",
        }
    ]
    for question in model["review_questions"]:
        question.update(
            {
                "assessment": (
                    "gap"
                    if missing_attestation and question["area"] == "attestation"
                    else "addressed"
                ),
                "evidence_refs": _evidence(plan_ref, "Evidence!A1"),
                "judgment_basis": (
                    "Reviewer assessment based on the identified evidence."
                ),
                "follow_up": (
                    "Obtain the independent-professional attestation."
                    if missing_attestation and question["area"] == "attestation"
                    else ""
                ),
            }
        )
    model["assumptions"] = [
        {
            "assumption_id": "assumption-1",
            "area": "continuity_economics",
            "statement": "Operating cash generation follows the reviewed plan.",
            "status": "supported",
            "materiality": "high",
            "evidence_refs": _evidence(plan_ref, "Evidence!B2"),
            "judgment_basis": "Reviewer classified the assumption and materiality.",
        }
    ]
    model["issues"] = (
        [
            {
                "issue_id": "issue-attestation",
                "area": "attestation",
                "statement": "Independent-professional attestation was not supplied.",
                "status": "open",
                "severity": "critical",
                "evidence_refs": _evidence(plan_ref, "Evidence!A1"),
                "owner": "Engagement lead",
                "next_action": "Request the filed attestation and verify its version.",
                "judgment_basis": "Reviewer identified the missing document.",
            }
        ]
        if missing_attestation
        else []
    )
    return model


def _inspection(tmp_path: Path) -> tuple[Any, Any, Path, Any]:
    core = _load_core()
    semantic = _load_semantic()
    input_dir = tmp_path / "evidence"
    input_dir.mkdir()
    _save_workbook(
        input_dir / "case_material_a.xlsx",
        [["Plan component", "Amount"], ["Cash generation", 500]],
    )
    _save_workbook(
        input_dir / "case_material_b.xlsx",
        [["Attestation section", "Reference"], ["Feasibility", "Section 4"]],
    )
    _save_workbook(
        input_dir / "case_material_c.xlsx",
        [["Creditor", "Claim"], ["Secured Bank", 600]],
    )
    inspection = core.run_concordato_review(
        input_dir,
        tmp_path / "inspection",
        reference_date="2026-03-31",
        language="it",
        document_language="it",
        tolerance="0.01",
    )
    return core, semantic, input_dir, inspection


def _reviewed_numeric_recipe(core: Any, inspection: Any) -> dict[str, Any]:
    return core.review_source_roles(
        inspection.inventory,
        {
            "case_material_a.xlsx": "concordato_plan",
            "case_material_b.xlsx": "other_support",
            "case_material_c.xlsx": "creditor_schedule",
        },
        inspection.raw_amount_candidates,
        {
            core.candidate_id(candidate): "candidate_amount"
            for candidate in inspection.raw_amount_candidates
        },
        reviewer_ref="qualified-reviewer",
        reviewed_on="2026-07-26",
        reference_date="2026-03-31",
        tolerance="0.01",
    )


def test_filename_does_not_assign_a_semantic_source_role() -> None:
    core = _load_core()

    assert core.classify_source_role(Path("DB_31.03.2026_21052026.xlsx")) == (
        "unclassified"
    )
    assert core.classify_source_role(Path("piano CP definitive.xlsx")) == "unclassified"


def test_equal_amounts_in_unrelated_contexts_remain_candidate_evidence() -> None:
    core = _load_core()
    candidates = [
        core.AmountCandidate(
            source_file="plan.xlsx",
            source_role="concordato_plan",
            location="Plan!B2",
            amount=Decimal("100"),
            token="100",
            context="creditor distribution class A",
        ),
        core.AmountCandidate(
            source_file="support.xlsx",
            source_role="accounting_support",
            location="Ledger!F9",
            amount=Decimal("100"),
            token="100",
            context="office equipment purchase",
        ),
    ]

    matches = core.find_exact_amount_matches(candidates, tolerance="0")

    assert len(matches) == 1
    assert matches[0]["match_status"] == "candidate_amount_match"
    assert matches[0]["context_token_overlap"] == "0"


@pytest.mark.parametrize(
    "plan_type",
    ["continuity_direct", "continuity_indirect", "liquidation", "mixed"],
)
def test_reviewed_semantic_model_supports_distinct_plan_types(
    tmp_path: Path,
    plan_type: str,
) -> None:
    _, semantic, _, inspection = _inspection(tmp_path)
    model = _reviewed_case_model(
        semantic,
        inspection.inventory,
        plan_type=plan_type,
    )

    recipe = semantic.review_concordato_case_model(
        inspection.inventory,
        model,
        reviewer_ref="qualified-reviewer",
        reviewed_on="2026-07-26",
        reference_date="2026-03-31",
    )
    _, normalized = semantic.validate_semantic_recipe(
        inspection.inventory,
        recipe,
        reference_date="2026-03-31",
    )

    assert normalized is not None
    assert normalized["procedure"]["plan_type"] == plan_type


def test_semantic_run_leads_with_creditors_liquidity_and_issues(
    tmp_path: Path,
) -> None:
    core, semantic, input_dir, inspection = _inspection(tmp_path)
    model = _reviewed_case_model(semantic, inspection.inventory)
    semantic_recipe = semantic.review_concordato_case_model(
        inspection.inventory,
        model,
        reviewer_ref="qualified-reviewer",
        reviewed_on="2026-07-26",
        reference_date="2026-03-31",
    )

    output_dir = tmp_path / "reviewed"
    run = core.run_concordato_review(
        input_dir,
        output_dir,
        reference_date="2026-03-31",
        language="it",
        document_language="it",
        tolerance="0.01",
        semantic_recipe=semantic_recipe,
    )

    case_output = json.loads(
        (output_dir / "concordato_case_model.json").read_text(encoding="utf-8")
    )
    checks = json.loads(
        (output_dir / "concordato_semantic_checks.json").read_text(encoding="utf-8")
    )
    payload = json.loads(
        (output_dir / "review_payload.json").read_text(encoding="utf-8")
    )
    gates = json.loads(
        (output_dir / "assurance_gates.json").read_text(encoding="utf-8")
    )
    workbook = openpyxl.load_workbook(
        output_dir / "concordato_review_workpaper.xlsx",
        read_only=True,
        data_only=False,
    )

    assert run.semantic_status == "reviewed"
    assert run.semantic_summary["total_claim_amount"] == "1000"
    assert run.semantic_summary["total_proposed_recovery_amount"] == "740"
    assert run.semantic_summary["total_liquidation_recovery_amount"] == "620"
    assert run.semantic_summary["plan_vs_liquidation_delta"] == "120"
    assert case_output["status"] == "reviewed"
    assert checks["summary"]["surplus_shortfall"] == "0"
    assert payload["review_type"] == "concordato_preventivo_review"
    assert payload["items"][0]["item_type"] == "semantic_case_status"
    assert any(
        item["item_type"] == "creditor_class_treatment" for item in payload["items"]
    )
    assert gates["gates"]["semantic_review"]["status"] == "passed"
    assert gates["gates"]["reporting"]["status"] == "passed"
    assert gates["gates"]["reconciliation"]["status"] == "not_applicable"
    assert "Creditors" in workbook.sheetnames
    assert "Liquidity" in workbook.sheetnames
    assert "Numeric Tie-Out" in workbook.sheetnames
    assert (output_dir / "concordato_preventivo_review_summary.docx").exists()
    replay = importlib.import_module("replay_assurance").replay_assurance(output_dir)
    assert replay["ok"] is True
    assert replay["workflow_output_closure_phase"] == "initial_run_finalization"


def test_semantic_review_and_numeric_appendix_keep_independent_authority(
    tmp_path: Path,
) -> None:
    core, semantic, input_dir, inspection = _inspection(tmp_path)
    semantic_recipe = semantic.review_concordato_case_model(
        inspection.inventory,
        _reviewed_case_model(semantic, inspection.inventory),
        reviewer_ref="qualified-reviewer",
        reviewed_on="2026-07-26",
        reference_date="2026-03-31",
    )
    numeric_recipe = _reviewed_numeric_recipe(core, inspection)

    output_dir = tmp_path / "reviewed-with-appendix"
    core.run_concordato_review(
        input_dir,
        output_dir,
        reference_date="2026-03-31",
        language="it",
        document_language="it",
        tolerance="0.01",
        semantic_recipe=semantic_recipe,
        recipe=numeric_recipe,
    )

    gates = json.loads(
        (output_dir / "assurance_gates.json").read_text(encoding="utf-8")
    )
    decisions = json.loads(
        (output_dir / "reviewed_decisions.json").read_text(encoding="utf-8")
    )
    decision_types = {decision["decision_type"] for decision in decisions["decisions"]}

    assert gates["gates"]["semantic_review"]["status"] == "passed"
    assert gates["gates"]["reconciliation"]["status"] == "passed"
    assert decision_types == {
        "source_role_mapping",
        "calculation_formula_authority",
        "semantic_review",
    }


@pytest.mark.parametrize(
    ("language", "title"),
    [
        ("it-IT", "Revisione del concordato preventivo"),
        ("en-GB", "Concordato Preventivo Review"),
        ("fr-FR", "Revue du concordato preventivo"),
        ("de-DE", "Prüfung des Concordato Preventivo"),
        ("es-ES", "Revisión del concordato preventivo"),
    ],
)
def test_primary_semantic_report_contract_is_localized(
    tmp_path: Path,
    language: str,
    title: str,
) -> None:
    core, semantic, input_dir, inspection = _inspection(tmp_path)
    semantic_recipe = semantic.review_concordato_case_model(
        inspection.inventory,
        _reviewed_case_model(semantic, inspection.inventory),
        reviewer_ref="qualified-reviewer",
        reviewed_on="2026-07-26",
        reference_date="2026-03-31",
    )

    output_dir = tmp_path / f"localized-{language[:2]}"
    core.run_concordato_review(
        input_dir,
        output_dir,
        reference_date="2026-03-31",
        language=language,
        document_language=language,
        tolerance="0.01",
        semantic_recipe=semantic_recipe,
    )

    markdown = (output_dir / "concordato_semantic_review.md").read_text(
        encoding="utf-8"
    )
    document = Document(output_dir / "concordato_preventivo_review_summary.docx")
    document_text = "\n".join(paragraph.text for paragraph in document.paragraphs)
    final_artifacts = json.loads(
        (output_dir / "final_artifacts.json").read_text(encoding="utf-8")
    )
    summary_output = next(
        output
        for output in final_artifacts["outputs"]
        if output["path"] == "concordato_preventivo_review_summary.docx"
    )

    assert markdown.startswith(f"# {title}\n")
    assert title in document_text
    assert summary_output["required_text"][0] == title


def test_mechanical_checks_expose_funding_and_cash_bridge_differences(
    tmp_path: Path,
) -> None:
    _, semantic, _, inspection = _inspection(tmp_path)
    model = _reviewed_case_model(
        semantic,
        inspection.inventory,
        plan_type="liquidation",
        unbalanced=True,
        broken_cash_bridge=True,
    )
    recipe = semantic.review_concordato_case_model(
        inspection.inventory,
        model,
        reviewer_ref="qualified-reviewer",
        reviewed_on="2026-07-26",
        reference_date="2026-03-31",
    )
    _, normalized = semantic.validate_semantic_recipe(
        inspection.inventory,
        recipe,
        reference_date="2026-03-31",
    )
    assert normalized is not None

    derived = semantic.derive_case_schedules(normalized)
    checks = {row["check_id"]: row for row in derived["checks"]}

    assert derived["summary"]["surplus_shortfall"] == "-20"
    assert derived["summary"]["funding_gap"] == "20"
    assert derived["liquidity"][0]["bridge_difference"] == "1"
    assert checks["sources_and_uses_balance"]["status"] == "attention"
    assert checks["liquidity_bridge"]["status"] == "attention"
    assert "not a legal" in checks["liquidity_bridge"]["limitation"]


def test_missing_attestation_remains_a_reviewed_open_issue(
    tmp_path: Path,
) -> None:
    _, semantic, _, inspection = _inspection(tmp_path)
    model = _reviewed_case_model(
        semantic,
        inspection.inventory,
        missing_attestation=True,
    )

    recipe = semantic.review_concordato_case_model(
        inspection.inventory,
        model,
        reviewer_ref="qualified-reviewer",
        reviewed_on="2026-07-26",
        reference_date="2026-03-31",
    )
    _, normalized = semantic.validate_semantic_recipe(
        inspection.inventory,
        recipe,
        reference_date="2026-03-31",
    )

    assert normalized is not None
    attestation_question = next(
        row for row in normalized["review_questions"] if row["area"] == "attestation"
    )
    assert attestation_question["assessment"] == "gap"
    assert normalized["issues"][0]["severity"] == "critical"
    assert not any(
        "attestation" in row["authoritative_for"]
        for row in normalized["document_perimeter"]["documents"]
    )


def test_semantic_recipe_is_withheld_after_source_mutation(tmp_path: Path) -> None:
    core, semantic, input_dir, inspection = _inspection(tmp_path)
    model = _reviewed_case_model(semantic, inspection.inventory)
    semantic_recipe = semantic.review_concordato_case_model(
        inspection.inventory,
        model,
        reviewer_ref="qualified-reviewer",
        reviewed_on="2026-07-26",
        reference_date="2026-03-31",
    )
    _save_workbook(
        input_dir / "case_material_a.xlsx",
        [["Plan component", "Amount"], ["Changed cash generation", 999]],
    )

    run = core.run_concordato_review(
        input_dir,
        tmp_path / "mutated",
        reference_date="2026-03-31",
        language="it",
        document_language="it",
        semantic_recipe=semantic_recipe,
    )

    assert run.semantic_status == "invalid_semantic_recipe"
    assert "outside the capture" in str(run.audit["semantic_review_error"])
    assert run.semantic_summary == {}
