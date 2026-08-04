from __future__ import annotations

import csv
import hashlib
import importlib._bootstrap_external
import importlib.util
import json
import os
import shutil
import stat
import subprocess
import sys
import time
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl
import pytest

from scripts.validate_plugin_review_contract import validate_contract

ROOT = Path(__file__).resolve().parents[2]
SCRIPT_DIR = ROOT / "plugins" / "journal-bank-reconciliation" / "scripts"
CORE_PATH = SCRIPT_DIR / "journal_bank_core.py"
APPLY_REVIEW_EDITS_PATH = SCRIPT_DIR / "apply_review_edits.py"
SEMANTIC_REVIEW_PATH = SCRIPT_DIR / "semantic_review.py"
MCP_SERVER_PATH = (
    ROOT / "plugins" / "journal-bank-reconciliation" / "mcp" / "server.cjs"
)
EVALUATION_CONTRACT_V2_PATH = (
    ROOT
    / "docs"
    / "specs"
    / "vera_audit_assurance"
    / "journal-bank-evaluation-contract.v2.json"
)
EVALUATION_CONTRACT_V3_PATH = (
    ROOT
    / "docs"
    / "specs"
    / "vera_audit_assurance"
    / "journal-bank-evaluation-contract.v3.json"
)
EVALUATION_CONTRACT_V4_PATH = (
    ROOT
    / "docs"
    / "specs"
    / "vera_audit_assurance"
    / "journal-bank-evaluation-contract.v4.json"
)
EVALUATION_CONTRACT_PATH = (
    ROOT
    / "docs"
    / "specs"
    / "vera_audit_assurance"
    / "journal-bank-evaluation-contract.v5.json"
)
TABULAR_V7_EXTENSION_CONTRACT_PATH = (
    ROOT
    / "docs"
    / "specs"
    / "vera_audit_assurance"
    / "journal-bank-tabular-v7-extension-contract.v1.json"
)


def _load_customer_ledger() -> Any:
    ledger_path = ROOT / "plugins" / "studio-archive" / "scripts" / "client_ledger.py"
    module_name = "test_journal_bank_customer_ledger"
    ledger = sys.modules.get(module_name)
    if ledger is None:
        spec = importlib.util.spec_from_file_location(module_name, ledger_path)
        assert spec is not None and spec.loader is not None
        ledger = importlib.util.module_from_spec(spec)
        sys.modules[module_name] = ledger
        spec.loader.exec_module(ledger)
    return ledger


def _running_customer_output(tmp_path: Path) -> tuple[Path, str]:
    ledger = _load_customer_ledger()
    client_root = tmp_path / "Managed Customer"
    client_root.mkdir(parents=True)
    client_id = "client_111111111111111111111111"
    ledger.create_client_manifest(client_root, client_id)
    engagement = ledger.create_engagement(client_root, client_id, "Test engagement")
    source = tmp_path / "managed-source.txt"
    source.write_text("managed input\n", encoding="utf-8")
    imported = ledger.import_document(
        client_root,
        client_id,
        engagement["engagement_id"],
        source,
        "source",
    )
    prepared = ledger.prepare_run(
        client_root,
        client_id,
        engagement["engagement_id"],
        "journal-bank-reconciliation",
        "test-version",
        input_ids=[imported["receipt"]["input_id"]],
    )
    running = ledger.start_run(
        client_root,
        engagement["engagement_id"],
        prepared["run"]["run_id"],
    )
    return Path(running["output_dir"]), str(running["context"]["run_id"])


def _customer_context_path(output_dir: Path) -> Path:
    candidate = output_dir.resolve()
    while candidate != candidate.parent:
        context_path = candidate / "context.json"
        if context_path.is_file():
            return context_path
        candidate = candidate.parent
    raise AssertionError("customer-run context is unavailable")


def _rename_customer_output(output_dir: Path) -> tuple[Path, Path, Path]:
    context_path = _customer_context_path(output_dir)
    client_root = context_path.parents[5]
    renamed_root = client_root.with_name(f"{client_root.name} Renamed")
    relative_output = output_dir.relative_to(client_root)
    relative_context = context_path.relative_to(client_root)
    client_root.rename(renamed_root)
    return (
        renamed_root / relative_output,
        renamed_root / relative_context,
        output_dir,
    )


def load_core() -> Any:
    if str(SCRIPT_DIR) not in sys.path:
        sys.path.insert(0, str(SCRIPT_DIR))
    spec = importlib.util.spec_from_file_location("journal_bank_core", CORE_PATH)
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def load_apply_review_edits() -> Any:
    if str(SCRIPT_DIR) not in sys.path:
        sys.path.insert(0, str(SCRIPT_DIR))
    spec = importlib.util.spec_from_file_location(
        "journal_bank_apply_review_edits", APPLY_REVIEW_EDITS_PATH
    )
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def load_semantic_review() -> Any:
    if str(SCRIPT_DIR) not in sys.path:
        sys.path.insert(0, str(SCRIPT_DIR))
    spec = importlib.util.spec_from_file_location(
        "journal_bank_semantic_review", SEMANTIC_REVIEW_PATH
    )
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def _save_workbook(path: Path, rows: list[list[Any]]) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    workbook.save(path)


def _save_csv(path: Path, rows: list[list[Any]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerows(rows)


def _save_delimited_csv(
    path: Path,
    rows: list[list[Any]],
    *,
    delimiter: str,
) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, delimiter=delimiter)
        writer.writerows(rows)


def _read_csv_dicts(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def _strict_json_object(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, value in pairs:
        if key in result:
            raise ValueError(f"duplicate JSON key: {key}")
        result[key] = value
    return result


def _seal_relationship_recipe(
    core: Any,
    recipe_path: Path,
    input_receipts_path: Path,
    *,
    tolerance: str = "1",
    date_window_days: int = 7,
    policy_updates: dict[str, Any] | None = None,
) -> Path:
    recipe = json.loads(recipe_path.read_text(encoding="utf-8"))
    receipts = json.loads(input_receipts_path.read_text(encoding="utf-8"))["receipts"]
    source_refs = [
        receipt["artifact_id"]
        for receipt in receipts
        if receipt["artifact_id"].startswith(("source.bank.", "source.journal."))
    ]
    policy = {
        **recipe["relationship"]["policy"],
        "amount_tolerance": tolerance,
        "date_window_days": date_window_days,
        **(policy_updates or {}),
    }
    recipe["matching"]["amount_tolerance"] = tolerance
    recipe["matching"]["date_window_days"] = date_window_days
    recipe["relationship"] = {
        "policy": policy,
        "review_content_sha256": core.canonical_json_sha256({"policy": policy}),
        "decision": core.build_relationship_review_receipt(
            decision_id="decision.relationship",
            reviewer_ref="reviewer.test",
            reviewed_on="2026-07-24",
            source_artifact_refs=source_refs,
            policy=policy,
        ),
    }
    recipe_path.write_text(
        json.dumps(recipe, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return recipe_path


def _prepare_reviewed_recipe(
    core: Any,
    bank_path: Path,
    journal_path: Path,
    recipe_dir: Path,
    *,
    tolerance: str = "1",
    date_window_days: int = 7,
    policy_updates: dict[str, Any] | None = None,
) -> Path:
    core.inspect_inputs(bank_path, journal_path, recipe_dir)
    return _seal_relationship_recipe(
        core,
        recipe_dir / "suggested_recipe.json",
        recipe_dir / "input_receipts.json",
        tolerance=tolerance,
        date_window_days=date_window_days,
        policy_updates=policy_updates,
    )


def _prepare_reviewed_direction_recipe(
    core: Any,
    bank_path: Path,
    journal_path: Path,
    recipe_dir: Path,
    direction_value_mapping: dict[str, str],
) -> Path:
    core.inspect_inputs(bank_path, journal_path, recipe_dir)
    recipe_path = recipe_dir / "suggested_recipe.json"
    recipe = json.loads(recipe_path.read_text(encoding="utf-8"))
    receipts = json.loads(
        (recipe_dir / "input_receipts.json").read_text(encoding="utf-8")
    )["receipts"]
    journal_ref = next(
        receipt["artifact_id"]
        for receipt in receipts
        if receipt["artifact_id"].startswith("source.journal.")
    )
    journal_recipe = recipe["journal"]["files"][journal_path.name]
    journal_recipe["direction_value_mapping"] = direction_value_mapping
    journal_recipe["mapping_decision"] = core.build_mapping_review_receipt(
        decision_id="decision.mapping.journal.direction",
        reviewer_ref="reviewer.test",
        reviewed_on="2026-07-25",
        source_artifact_ref=journal_ref,
        side="journal",
        source_file=journal_path.name,
        header_rows=journal_recipe["header_rows"],
        mapping=journal_recipe["mapping"],
        potential_monetary_columns=journal_recipe["potential_monetary_columns"],
        excluded_monetary_columns=journal_recipe["excluded_monetary_columns"],
        direction_value_mapping=direction_value_mapping,
        date_convention=journal_recipe.get("date_convention"),
        csv_field_delimiter=journal_recipe["csv_field_delimiter"],
    )
    recipe_path.write_text(
        json.dumps(recipe, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return _seal_relationship_recipe(
        core,
        recipe_path,
        recipe_dir / "input_receipts.json",
        tolerance="0",
        date_window_days=0,
        policy_updates={"direction_policy": "same_sign"},
    )


def _attach_current_mapping_receipt(
    core: Any,
    recipe: dict[str, Any],
    receipts: list[dict[str, Any]],
    *,
    side: str,
    source_path: Path,
    decision_id: str,
) -> dict[str, Any]:
    source_ref = next(
        receipt["artifact_id"]
        for receipt in receipts
        if receipt["artifact_id"].startswith(f"source.{side}.")
    )
    file_recipe = recipe[side]["files"][source_path.name]
    receipt = core.build_mapping_review_receipt(
        decision_id=decision_id,
        reviewer_ref="reviewer.test",
        reviewed_on="2026-07-25",
        source_artifact_ref=source_ref,
        side=side,
        source_file=source_path.name,
        header_rows=file_recipe["header_rows"],
        mapping=file_recipe["mapping"],
        potential_monetary_columns=file_recipe["potential_monetary_columns"],
        excluded_monetary_columns=file_recipe["excluded_monetary_columns"],
        direction_value_mapping=file_recipe["direction_value_mapping"],
        date_convention=file_recipe.get("date_convention"),
        date_locale=file_recipe.get("date_locale"),
        non_movement_summary_labels=file_recipe.get("non_movement_summary_labels"),
        csv_field_delimiter=file_recipe["csv_field_delimiter"],
        decimal_separator=file_recipe.get("decimal_separator"),
        thousands_separator=file_recipe.get("thousands_separator"),
    )
    file_recipe["mapping_decision"] = receipt
    return receipt


def _prepare_reviewed_date_recipe(
    core: Any,
    bank_path: Path,
    journal_path: Path,
    recipe_dir: Path,
    *,
    date_convention: str,
) -> Path:
    core.inspect_inputs(bank_path, journal_path, recipe_dir)
    recipe_path = recipe_dir / "suggested_recipe.json"
    recipe = json.loads(recipe_path.read_text(encoding="utf-8"))
    receipts = json.loads(
        (recipe_dir / "input_receipts.json").read_text(encoding="utf-8")
    )["receipts"]
    for side, source_path in (("bank", bank_path), ("journal", journal_path)):
        file_recipe = recipe[side]["files"][source_path.name]
        file_recipe["date_convention"] = date_convention
        _attach_current_mapping_receipt(
            core,
            recipe,
            receipts,
            side=side,
            source_path=source_path,
            decision_id=f"decision.mapping.{side}.date",
        )
    recipe_path.write_text(
        json.dumps(recipe, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return _seal_relationship_recipe(
        core,
        recipe_path,
        recipe_dir / "input_receipts.json",
        tolerance="0",
        date_window_days=0,
    )


def _prepare_reviewed_date_locale_recipe(
    core: Any,
    bank_path: Path,
    journal_path: Path,
    recipe_dir: Path,
    *,
    date_locale: str,
) -> Path:
    core.inspect_inputs(bank_path, journal_path, recipe_dir)
    recipe_path = recipe_dir / "suggested_recipe.json"
    recipe = json.loads(recipe_path.read_text(encoding="utf-8"))
    receipts = json.loads(
        (recipe_dir / "input_receipts.json").read_text(encoding="utf-8")
    )["receipts"]
    for side, source_path in (("bank", bank_path), ("journal", journal_path)):
        file_recipe = recipe[side]["files"][source_path.name]
        file_recipe["date_locale"] = date_locale
        _attach_current_mapping_receipt(
            core,
            recipe,
            receipts,
            side=side,
            source_path=source_path,
            decision_id=f"decision.mapping.{side}.date_locale",
        )
    recipe_path.write_text(
        json.dumps(recipe, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return _seal_relationship_recipe(
        core,
        recipe_path,
        recipe_dir / "input_receipts.json",
        tolerance="0",
        date_window_days=0,
    )


def _prepare_reviewed_summary_recipe(
    core: Any,
    bank_path: Path,
    journal_path: Path,
    recipe_dir: Path,
    *,
    labels: list[str],
) -> Path:
    core.inspect_inputs(bank_path, journal_path, recipe_dir)
    recipe_path = recipe_dir / "suggested_recipe.json"
    recipe = json.loads(recipe_path.read_text(encoding="utf-8"))
    receipts = json.loads(
        (recipe_dir / "input_receipts.json").read_text(encoding="utf-8")
    )["receipts"]
    bank_recipe = recipe["bank"]["files"][bank_path.name]
    bank_recipe["non_movement_summary_labels"] = labels
    _attach_current_mapping_receipt(
        core,
        recipe,
        receipts,
        side="bank",
        source_path=bank_path,
        decision_id="decision.mapping.bank.summary_labels",
    )
    recipe_path.write_text(
        json.dumps(recipe, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return _seal_relationship_recipe(
        core,
        recipe_path,
        recipe_dir / "input_receipts.json",
        tolerance="0",
        date_window_days=0,
    )


def _prepare_sealed_mcp_review_run(
    tmp_path: Path,
    *,
    language: str = "en",
    portable: bool = False,
) -> tuple[
    Path,
    dict[str, Any],
    dict[str, Any],
    dict[str, Any],
    dict[str, Any],
]:
    core = load_core()
    output_dir, client_run_id = _running_customer_output(tmp_path)
    source_dir = output_dir / "sources"
    bank_path = source_dir / "bank.xlsx"
    journal_path = source_dir / "journal.xlsx"
    source_dir.mkdir(parents=True, exist_ok=True)
    _save_workbook(
        bank_path,
        [
            ["Date", "Description", "Amount", "Reference", "Beneficiary"],
            ["2025-01-02", "Payment invoice INV100 ACME", 123.45, "INV100", "ACME"],
            ["2025-01-05", "Unmatched fee", 9.99, "FEE9", "Bank"],
        ],
    )
    _save_workbook(
        journal_path,
        [
            ["Date", "Description", "Debit", "Reference", "Beneficiary"],
            ["2025-01-01", "Invoice INV100 ACME", 123.45, "INV100", "ACME"],
            ["2025-01-07", "Unmatched supplier", 77.0, "SUP77", "Supplier"],
        ],
    )
    recipe_path = _prepare_reviewed_recipe(
        core,
        bank_path,
        journal_path,
        output_dir / "recipe",
    )
    core.run_reconciliation(
        bank_path,
        journal_path,
        output_dir,
        recipe_path,
        language=language,
        document_language="en",
        client_run_id=client_run_id,
        client_run_root=(
            _customer_context_path(output_dir).parent if portable else None
        ),
    )
    return (
        output_dir,
        *(
            json.loads((output_dir / name).read_text(encoding="utf-8"))
            for name in (
                "run_intake.json",
                "review_payload.json",
                "ui_decisions.json",
                "final_artifacts.json",
            )
        ),
    )


def _prepare_closed_mcp_review_run(
    tmp_path: Path,
    *,
    portable: bool = False,
) -> tuple[
    Path,
    dict[str, Any],
    dict[str, Any],
    dict[str, Any],
    dict[str, Any],
]:
    core = load_core()
    output_dir, client_run_id = _running_customer_output(tmp_path)
    source_dir = output_dir.parent / f"{output_dir.name}-sources"
    bank_path = source_dir / "bank.csv"
    journal_path = source_dir / "journal.csv"
    source_dir.mkdir(parents=True, exist_ok=True)
    _save_csv(
        bank_path,
        [
            ["Date", "Amount", "Reference"],
            ["2025-03-10", "80.00", "INV100"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference"],
            ["2025-03-10", "80.00", "INV100"],
        ],
    )
    recipe_path = _prepare_reviewed_recipe(
        core,
        bank_path,
        journal_path,
        output_dir.parent / f"{output_dir.name}-recipe",
    )
    core.run_reconciliation(
        bank_path,
        journal_path,
        output_dir,
        recipe_path,
        client_run_id=client_run_id,
        client_run_root=(
            _customer_context_path(output_dir).parent if portable else None
        ),
    )
    return (
        output_dir,
        *(
            json.loads((output_dir / name).read_text(encoding="utf-8"))
            for name in (
                "run_intake.json",
                "review_payload.json",
                "ui_decisions.json",
                "final_artifacts.json",
            )
        ),
    )


def _prepare_two_match_run(tmp_path: Path) -> tuple[Any, Path]:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    output_dir = tmp_path / "run"
    rows = [
        ["Date", "Amount", "Reference"],
        ["2026-05-08", "80.00", "TX100"],
        ["2026-05-09", "25.50", "TX200"],
    ]
    _save_csv(bank_path, rows)
    _save_csv(journal_path, rows)
    recipe_path = _prepare_reviewed_recipe(
        core,
        bank_path,
        journal_path,
        tmp_path / "recipe",
        tolerance="0",
        date_window_days=0,
    )
    core.run_reconciliation(
        bank_path,
        journal_path,
        output_dir,
        recipe_path,
        tolerance="0",
        date_window_days=0,
    )
    return core, output_dir


def _prepare_ambiguous_semantic_run(tmp_path: Path) -> tuple[Any, Any, Path, Path]:
    core = load_core()
    semantic_review = load_semantic_review()
    case_dir = tmp_path / "case"
    bank_path = case_dir / "bank.csv"
    journal_path = case_dir / "journal.csv"
    reconciliation_dir = case_dir / "reconciliation"
    semantic_dir = case_dir / "semantic-review"
    case_dir.mkdir()
    _save_csv(
        bank_path,
        [
            ["Date", "Amount", "Description", "Beneficiary"],
            ["2026-05-08", "80.00", "Payment Alpha", "Alpha"],
            ["2026-05-08", "80.00", "Payment Beta", "Beta"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Description", "Beneficiary"],
            ["2026-05-08", "80.00", "Invoice Alpha", "Alpha"],
            ["2026-05-08", "80.00", "Invoice Beta", "Beta"],
        ],
    )
    recipe_path = _prepare_reviewed_recipe(
        core,
        bank_path,
        journal_path,
        case_dir / "recipe",
        tolerance="0",
        date_window_days=0,
    )
    result = core.run_reconciliation(
        bank_path,
        journal_path,
        reconciliation_dir,
        recipe_path,
        tolerance="0",
        date_window_days=0,
    )
    assert result.matches.is_empty()
    assert result.unmatched_bank.height == 2
    assert result.unmatched_journal.height == 2
    return core, semantic_review, reconciliation_dir, semantic_dir


def _prepare_managed_ambiguous_semantic_run(
    tmp_path: Path,
) -> tuple[Any, Any, Path, Path, Path, dict[str, Any]]:
    core = load_core()
    semantic_review = load_semantic_review()
    ledger = _load_customer_ledger()
    client_root = tmp_path / "Managed Customer"
    client_root.mkdir()
    client_id = "client_333333333333333333333333"
    ledger.create_client_manifest(client_root, client_id)
    engagement = ledger.create_engagement(
        client_root,
        client_id,
        "Managed semantic CLI",
    )
    source_dir = tmp_path / "received"
    source_dir.mkdir()
    bank_source = source_dir / "bank.csv"
    journal_source = source_dir / "journal.csv"
    _save_csv(
        bank_source,
        [
            ["Date", "Amount", "Description", "Beneficiary"],
            ["2026-05-08", "80.00", "Payment Alpha", "Alpha"],
            ["2026-05-08", "80.00", "Payment Beta", "Beta"],
        ],
    )
    _save_csv(
        journal_source,
        [
            ["Date", "Amount", "Description", "Beneficiary"],
            ["2026-05-08", "80.00", "Invoice Alpha", "Alpha"],
            ["2026-05-08", "80.00", "Invoice Beta", "Beta"],
        ],
    )
    imported_bank = ledger.import_document(
        client_root,
        client_id,
        engagement["engagement_id"],
        bank_source,
        "source",
    )
    imported_journal = ledger.import_document(
        client_root,
        client_id,
        engagement["engagement_id"],
        journal_source,
        "journal",
    )
    prepared = ledger.prepare_run(
        client_root,
        client_id,
        engagement["engagement_id"],
        "journal-bank-reconciliation",
        "test-version",
        input_ids=[
            imported_bank["receipt"]["input_id"],
            imported_journal["receipt"]["input_id"],
        ],
    )
    running = ledger.start_run(
        client_root,
        engagement["engagement_id"],
        prepared["run"]["run_id"],
    )
    input_paths = {
        binding["role"]: Path(binding["path"])
        for binding in running["context"]["input_bindings"]
    }
    bank_path = input_paths["source"]
    journal_path = input_paths["journal"]
    output_dir = Path(running["output_dir"])
    context_path = Path(running["context_path"])
    client_run_id = str(running["run"]["run_id"])
    recipe_path = _prepare_reviewed_recipe(
        core,
        bank_path,
        journal_path,
        output_dir / "recipe",
        tolerance="0",
        date_window_days=0,
    )
    reconciliation_dir = output_dir / "reconciliation"
    result = core.run_reconciliation(
        bank_path,
        journal_path,
        reconciliation_dir,
        recipe_path,
        tolerance="0",
        date_window_days=0,
        client_run_id=client_run_id,
        client_run_root=context_path.parent,
    )
    assert result.matches.is_empty()
    return (
        core,
        semantic_review,
        reconciliation_dir,
        output_dir / "semantic-review",
        context_path,
        running["context"],
    )


def _valid_semantic_response(graph: dict[str, Any]) -> dict[str, Any]:
    component = graph["selected_components"][0]
    journals_by_beneficiary = {
        record["beneficiary"]: record["transaction_id"]
        for record in component["journal_records"]
    }
    decisions = [
        {
            "bank_transaction_id": record["transaction_id"],
            "verdict": "suggest_match",
            "journal_transaction_id": journals_by_beneficiary[record["beneficiary"]],
            "evidence_fields": [
                "amount_abs",
                "transaction_date",
                "beneficiary",
                "description",
            ],
            "rationale": (
                "The amount and date are eligible, and the beneficiary plus "
                "description identify this neighboring journal row."
            ),
            "contradictions": [],
            "requested_evidence": [],
            "resolution_level": "beneficiary_match",
            "classification": None,
            "identified_counterparty": record["beneficiary"],
        }
        for record in component["bank_records"]
    ]
    return {
        "schema_version": "journal_bank.semantic_worker_response.v2",
        "candidate_graph_sha256": graph["candidate_graph_sha256"],
        "component_reviews": [
            {"component_id": component["component_id"], "decisions": decisions}
        ],
    }


def _write_semantic_worker_result(
    semantic_dir: Path,
    response: dict[str, Any],
    *,
    item_type: str = "agent_message",
    events_override: list[dict[str, Any]] | None = None,
) -> tuple[Path, Path]:
    semantic_review = sys.modules.get("journal_bank_semantic_review")
    if semantic_review is None:
        semantic_review = load_semantic_review()
    response_path = semantic_dir / "luna_response.json"
    events_path = semantic_dir / "luna_events.jsonl"
    response_path.write_text(
        json.dumps(response, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    events = events_override
    if events is None:
        events = [
            {"type": "thread.started", "thread_id": "thread_luna_test"},
            {"type": "turn.started"},
            {
                "type": "item.completed",
                "item": {
                    "id": "item_0",
                    "type": item_type,
                    "text": json.dumps(response, ensure_ascii=False),
                },
            },
            {
                "type": "turn.completed",
                "usage": {"input_tokens": 120, "output_tokens": 80},
            },
        ]
    events_bytes = "".join(
        json.dumps(event, ensure_ascii=False) + "\n" for event in events
    ).encode("utf-8")
    events_path.write_bytes(events_bytes)
    stderr_path = semantic_dir / "luna_stderr.log"
    stderr_bytes = b""
    stderr_path.write_bytes(stderr_bytes)
    response_bytes = response_path.read_bytes()
    graph_path = semantic_dir / "residual_candidate_graph.json"
    graph_bytes = graph_path.read_bytes()
    graph = json.loads(graph_bytes)
    prompt_bytes = (semantic_dir / "luna_prompt.md").read_bytes()
    schema_bytes = (semantic_dir / "luna_output_schema.json").read_bytes()
    receipt_content = {
        "schema_version": "journal_bank.semantic_launch_receipt.v1",
        "workflow_id": "journal_bank_reconciliation",
        "candidate_graph_sha256": graph["candidate_graph_sha256"],
        "packet": {
            "candidate_graph_file_sha256": hashlib.sha256(graph_bytes).hexdigest(),
            "candidate_graph_file_bytes": len(graph_bytes),
            "prompt_sha256": hashlib.sha256(prompt_bytes).hexdigest(),
            "prompt_bytes": len(prompt_bytes),
            "output_schema_sha256": hashlib.sha256(schema_bytes).hexdigest(),
            "output_schema_bytes": len(schema_bytes),
        },
        "requested_worker_configuration": graph["requested_worker_configuration"],
        "boundary": {
            "contract_id": "journal_bank.luna_seatbelt_capsule.v1",
            "platform": "Darwin",
            "darwin_build": "25F84",
            "profile_sha256": semantic_review.PINNED_SEATBELT_PROFILE_SHA256,
            "codex_path": "/Applications/ChatGPT.app/Contents/Resources/codex",
            "codex_sha256": semantic_review.PINNED_CODEX_SHA256,
            "codex_bytes": 1,
            "codex_version": "codex-cli 0.146.0-alpha.3.1",
            "sandbox_exec_path": "/usr/bin/sandbox-exec",
            "sandbox_exec_sha256": semantic_review.PINNED_SANDBOX_EXEC_SHA256,
            "canary_reader_sha256": semantic_review.PINNED_CAT_SHA256,
            "canaries": {
                "exact_schema_read_succeeded": True,
                "outside_capsule_read_denied": True,
                "codex_version_inside_boundary": "codex-cli 0.146.0-alpha.3.1",
            },
            "global_instructions_absent_or_empty": True,
            "auth_file_readable_by_codex_process": True,
            "installation_id_preexisting_and_unchanged": True,
            "outbound_network_allowed": True,
            "filesystem_scope": "capsule_plus_exact_codex_runtime_files",
            "qualification_basis": "pinned_hidden_view_image_outside_nonce_denied",
        },
        "process": {
            "return_code": 0,
            "timed_out": False,
            "duration_ms": 1,
            "redacted_argv": semantic_review._redacted_worker_argv(),
            "response_sha256": hashlib.sha256(response_bytes).hexdigest(),
            "response_bytes": len(response_bytes),
            "events_sha256": hashlib.sha256(events_bytes).hexdigest(),
            "events_bytes": len(events_bytes),
            "stderr_sha256": hashlib.sha256(stderr_bytes).hexdigest(),
            "stderr_bytes": len(stderr_bytes),
        },
        "jsonl_observation": {
            "visibility_complete": False,
            "visible_forbidden_item_count": 0,
            "tool_use_absence_observed": False,
            "thread_id": "thread_luna_test",
            "usage": {"input_tokens": 120, "output_tokens": 80},
            "completed_item_counts": {"agent_message": 1, "reasoning": 0},
        },
        "runtime_attestation": {
            "model_observed": False,
            "reasoning_effort_observed": False,
            "main_chat_model_change": False,
        },
        "advisory_only": True,
    }
    receipt = {
        **receipt_content,
        "content_sha256": semantic_review.canonical_json_sha256(receipt_content),
    }
    (semantic_dir / "luna_launch_receipt.json").write_text(
        json.dumps(receipt, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return response_path, events_path


def _mock_semantic_worker_runtime(
    monkeypatch: pytest.MonkeyPatch,
    semantic_review: Any,
    tmp_path: Path,
    process_result: dict[str, Any],
) -> dict[str, Any]:
    fake_home = tmp_path / "qualified-codex-home"
    boundary_inputs = {
        "codex_home": fake_home,
        "auth_path": fake_home / "auth.json",
        "installation_id_path": fake_home / "installation_id",
        "global_agents_path": fake_home / "AGENTS.md",
        "global_agents_override_path": fake_home / "AGENTS.override.md",
        "bindings": {
            "auth": {"sha256": "a" * 64, "byte_count": 1},
            "installation_id": {"sha256": "b" * 64, "byte_count": 36},
            "global_agents": {"exists": False, "sha256": None, "byte_count": 0},
            "global_agents_override": {
                "exists": False,
                "sha256": None,
                "byte_count": 0,
            },
        },
    }
    codex_path = Path("/qualified/codex")
    executable_bindings = {
        codex_path: {
            "sha256": semantic_review.PINNED_CODEX_SHA256,
            "byte_count": 267_702_000,
            "mode": 0o755,
        },
        semantic_review.SANDBOX_EXEC_PATH: {
            "sha256": semantic_review.PINNED_SANDBOX_EXEC_SHA256,
            "byte_count": 102_560,
            "mode": 0o755,
        },
        semantic_review.SANDBOX_CANARY_PATH: {
            "sha256": semantic_review.PINNED_CAT_SHA256,
            "byte_count": 118_992,
            "mode": 0o755,
        },
    }
    qualified = {
        "darwin_build": semantic_review.PINNED_DARWIN_BUILD,
        "codex_path": codex_path,
        "codex_binding": executable_bindings[codex_path],
        "sandbox_exec_binding": executable_bindings[semantic_review.SANDBOX_EXEC_PATH],
        "canary_binding": executable_bindings[semantic_review.SANDBOX_CANARY_PATH],
    }
    captured: dict[str, Any] = {}

    def fake_process(command: list[str], **kwargs: Any) -> dict[str, Any]:
        captured["command"] = command
        captured["kwargs"] = kwargs
        return process_result

    monkeypatch.setattr(
        semantic_review,
        "_codex_home_boundary_inputs",
        lambda: boundary_inputs,
    )
    monkeypatch.setattr(
        semantic_review,
        "_qualified_executables",
        lambda codex_bin: qualified,
    )
    monkeypatch.setattr(
        semantic_review,
        "_qualification_canaries",
        lambda **kwargs: {
            "exact_schema_read_succeeded": True,
            "outside_capsule_read_denied": True,
            "codex_version_inside_boundary": semantic_review.PINNED_CODEX_VERSION,
        },
    )
    monkeypatch.setattr(
        semantic_review,
        "_stable_executable_binding",
        lambda path, **kwargs: executable_bindings[Path(path)],
    )
    monkeypatch.setattr(semantic_review, "_run_captured_process", fake_process)
    return captured


def _reseal_semantic_source_receipt(
    reconciliation_dir: Path,
    *,
    artifact_id: str,
    artifact_path: Path,
) -> None:
    receipt_path = reconciliation_dir / "artifact_receipts.json"
    bundle = json.loads(receipt_path.read_text(encoding="utf-8"))
    receipt = next(
        item for item in bundle["output_receipts"] if item["artifact_id"] == artifact_id
    )
    payload = artifact_path.read_bytes()
    receipt["byte_count"] = len(payload)
    receipt["sha256"] = hashlib.sha256(payload).hexdigest()
    receipt_path.write_text(
        json.dumps(bundle, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def _relationship_policy(amount_tolerance: object) -> dict[str, Any]:
    return {
        "relationship_shape": "one_to_one",
        "allow_evidence_reuse": False,
        "require_same_currency": True,
        "require_same_unit": True,
        "require_same_entity": True,
        "require_same_party": False,
        "direction_policy": "absolute_amount",
        "default_currency": "EUR",
        "default_unit": "currency",
        "default_entity_ref": "entity.case",
        "default_party_ref": None,
        "amount_tolerance": amount_tolerance,
        "date_window_days": 0,
    }


def _call_mcp_server(
    messages: list[dict[str, object]],
    *,
    server_path: Path = MCP_SERVER_PATH,
    env: dict[str, str] | None = None,
) -> list[dict[str, object]]:
    node = shutil.which("node")
    if node is None:
        pytest.skip(
            "Node.js is required to exercise the Journal-Bank Reconciliation MCP server."
        )
    completed = subprocess.run(
        [node, str(server_path), "--stdio"],
        input="\n".join(json.dumps(message) for message in messages) + "\n",
        capture_output=True,
        text=True,
        check=True,
        timeout=10,
        env=env,
    )
    return [json.loads(line) for line in completed.stdout.splitlines() if line.strip()]


def _bank_transaction_call(
    tool_name: str,
    arguments: dict[str, Any],
) -> dict[str, Any]:
    response = _call_mcp_server(
        [
            {
                "jsonrpc": "2.0",
                "id": 1,
                "method": "tools/call",
                "params": {"name": tool_name, "arguments": arguments},
            }
        ]
    )[0]
    return response["result"]["structuredContent"]


def _portable_bank_transaction_case(
    tmp_path: Path,
) -> tuple[Path, dict[str, Any]]:
    output_dir, run_intake, review_payload, _, final_artifacts = (
        _prepare_sealed_mcp_review_run(tmp_path, portable=True)
    )
    decisions = [
        {
            "item_id": item["id"],
            "action": (
                "accept"
                if "accept" in item["allowed_actions"]
                else item["allowed_actions"][0]
            ),
        }
        for item in review_payload["items"]
    ]
    return output_dir, {
        "run_intake": run_intake,
        "review_payload": review_payload,
        "final_artifacts": final_artifacts,
        "decisions": decisions,
    }


def test_bank_review_save_and_apply_survive_customer_folder_rename(
    tmp_path: Path,
) -> None:
    output_dir, arguments = _portable_bank_transaction_case(tmp_path)
    old_customer_root = _customer_context_path(output_dir).parents[5].as_posix()
    old_output = output_dir.as_posix()
    assert arguments["run_intake"]["path_reference"] == "run_root_relative"
    assert arguments["run_intake"]["output_dir"] == "outputs"
    assert all(
        not Path(value).is_absolute()
        for value in arguments["run_intake"]["input_paths"]
    )

    renamed_output, current_context, stale_output = _rename_customer_output(output_dir)
    arguments["client_engagement"] = current_context.as_posix()

    saved = _bank_transaction_call(
        "save_journal_bank_decisions",
        arguments,
    )
    applied = _bank_transaction_call(
        "apply_journal_bank_decisions",
        arguments,
    )

    assert saved["ok"] is True
    assert saved["persisted"] is True
    assert applied["ok"] is True
    assert applied["persisted"] is True
    assert (renamed_output / "ui_decisions.json").is_file()
    assert (renamed_output / "applied_decisions.json").is_file()
    assert not stale_output.exists()
    assert old_output not in arguments["run_intake"]["output_dir"]
    assert all(
        old_customer_root not in artifact.read_text(encoding="utf-8")
        for artifact in renamed_output.rglob("*")
        if artifact.is_file() and artifact.suffix.lower() in {".json", ".md"}
    )


def test_bank_review_rejects_run_root_escape_without_writing(
    tmp_path: Path,
) -> None:
    output_dir, arguments = _portable_bank_transaction_case(tmp_path)
    arguments["client_engagement"] = _customer_context_path(output_dir).as_posix()
    forged = json.loads(json.dumps(arguments))
    forged["run_intake"]["output_dir"] = "../outside"
    before = _tree_snapshot(output_dir)

    result = _bank_transaction_call(
        "save_journal_bank_decisions",
        forged,
    )

    assert result["ok"] is False
    assert "leaves the customer run" in result["error"]
    assert _tree_snapshot(output_dir) == before
    assert not (output_dir.parent.parent / "outside").exists()


def _customer_run_preflight_passthrough_script_lines() -> list[str]:
    """Let stage-attack wrappers target review work after customer validation."""
    return [
        'if "--client-run-preflight-only" in sys.argv:',
        "    completed = subprocess.run([real, *sys.argv[1:]], capture_output=True)",
        "    sys.stdout.buffer.write(completed.stdout)",
        "    sys.stderr.buffer.write(completed.stderr)",
        "    raise SystemExit(completed.returncode)",
    ]


def _copy_journal_bank_implementation(tmp_path: Path) -> tuple[Path, Path]:
    copied_plugins = tmp_path / "copied" / "plugins"
    copied_plugin = copied_plugins / "journal-bank-reconciliation"
    copied_shared = copied_plugins / "_shared" / "vendor" / "modules" / "vera_assurance"
    shutil.copytree(
        ROOT / "plugins" / "journal-bank-reconciliation",
        copied_plugin,
        ignore=shutil.ignore_patterns("__pycache__", "*.pyc", "*.pyo"),
    )
    shutil.copytree(
        ROOT / "plugins" / "_shared" / "vendor" / "modules" / "vera_assurance",
        copied_shared,
        ignore=shutil.ignore_patterns("__pycache__", "*.pyc", "*.pyo"),
    )
    return copied_plugin, copied_shared


def _write_rehashed_assurance_envelope(
    path: Path,
    envelope: dict[str, Any],
) -> None:
    content = dict(envelope)
    content.pop("content_sha256", None)
    canonical = json.dumps(
        content,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    envelope["content_sha256"] = hashlib.sha256(canonical).hexdigest()
    path.write_text(
        json.dumps(envelope, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def _tree_snapshot(path: Path) -> dict[str, tuple[bytes, int]]:
    return {
        candidate.relative_to(path).as_posix(): (
            candidate.read_bytes(),
            candidate.stat().st_mode & 0o777,
        )
        for candidate in sorted(path.rglob("*"))
        if candidate.is_file()
    }


def _tree_mode_snapshot(path: Path) -> dict[str, tuple[str, int]]:
    entries = {
        ".": ("directory", path.stat().st_mode & 0o777),
    }
    for candidate in sorted(path.rglob("*")):
        relative = candidate.relative_to(path).as_posix()
        kind = "directory" if candidate.is_dir() else "file"
        entries[relative] = (kind, candidate.stat().st_mode & 0o777)
    return entries


def _mcp_review_write_message(
    tool_name: str,
    review_payload: dict[str, Any],
    run_intake: dict[str, Any],
) -> list[dict[str, object]]:
    item = next(
        entry
        for entry in review_payload["items"]
        if entry["item_type"] == "matched_pair"
    )
    return [
        {
            "jsonrpc": "2.0",
            "id": 1,
            "method": "tools/call",
            "params": {
                "name": tool_name,
                "arguments": {
                    "run_intake": run_intake,
                    "review_payload": review_payload,
                    "decisions": [
                        {
                            "item_id": item["id"],
                            "action": "edit",
                            "edit_value": "Transaction containment review.",
                        }
                    ],
                },
            },
        }
    ]


def test_evaluation_contract_v2_remains_immutable_historical_evidence() -> None:
    payload = EVALUATION_CONTRACT_V2_PATH.read_bytes()
    contract = json.loads(payload)

    assert hashlib.sha256(payload).hexdigest() == (
        "183d3bcae2a0674c637e55c26a76fca7c1647732b3b51979f160a2b1bccd6944"
    )
    assert contract["schema_version"] == "journal_bank.evaluation_contract.v2"
    assert contract["contract_id"] == "journal_bank.repository_contract.v2"
    assert contract["implementation"]["tabular_adapter"] == {
        "adapter_id": "journal_bank.tabular.v4",
        "adapter_version": "4",
    }


def test_evaluation_contract_v3_remains_immutable_pre_oracle_evidence() -> None:
    payload = EVALUATION_CONTRACT_V3_PATH.read_bytes()
    contract = json.loads(payload)

    assert hashlib.sha256(payload).hexdigest() == (
        "13b7f430805767962f7c531872cd8d91b6bb68adc74ff895acb0e6b3a2e99046"
    )
    assert contract["schema_version"] == "journal_bank.evaluation_contract.v3"
    assert contract["contract_id"] == "journal_bank.repository_contract.v3"


def test_evaluation_contract_v4_remains_immutable_historical_evidence() -> None:
    payload = EVALUATION_CONTRACT_V4_PATH.read_bytes()
    contract = json.loads(payload)

    assert hashlib.sha256(payload).hexdigest() == (
        "18c3e11da7bf263dbe392c13ad56af64b94c2093a1f2428ede52a66e85bdc97b"
    )
    assert contract["schema_version"] == "journal_bank.evaluation_contract.v4"
    assert contract["contract_id"] == "journal_bank.repository_contract.v4"
    assert contract["supersedes"]["contract_id"] == (
        "journal_bank.repository_contract.v3"
    )


def test_evaluation_contract_v5_closes_prospective_repository_contract() -> None:
    core = load_core()
    contract_bytes = EVALUATION_CONTRACT_PATH.read_bytes()
    raw_contract = contract_bytes.decode("utf-8")

    contract = json.loads(raw_contract, object_pairs_hook=_strict_json_object)

    assert hashlib.sha256(contract_bytes).hexdigest() == (
        "4824652ecdb990a844fd9b72d799a2537f46a21ddb9fefb8a664f828c0ec6657"
    )
    assert contract["schema_version"] == "journal_bank.evaluation_contract.v5"
    assert contract["contract_id"] == "journal_bank.repository_contract.v5"
    assert contract["supersedes"]["contract_id"] == (
        "journal_bank.repository_contract.v4"
    )
    assert contract["supersedes"]["contract_sha256"] == (
        "18c3e11da7bf263dbe392c13ad56af64b94c2093a1f2428ede52a66e85bdc97b"
    )
    assert contract["lifecycle"] == {
        "contract_status": "frozen_prospective",
        "promoted": False,
        "mechanical_regression_status": "GO",
        "mechanical_regression_scope": (
            "The adjudicated r3 regression is implementation and contract evidence "
            "only; it is not a fresh unseen v5 holdout."
        ),
        "m7_promotion_status": "NO_GO",
        "promotion_requires": (
            "a fresh independently authored unseen holdout bound to the exact v5 "
            "contract bytes"
        ),
    }
    assert contract["implementation"]["tabular_adapter"] == {
        "adapter_id": core.TABULAR_ADAPTER_ID,
        "adapter_version": core.TABULAR_ADAPTER_VERSION,
        "invalidates_prior_versions": True,
    }
    assert contract["implementation"]["relationship_adapter"] == {
        "adapter_id": core.RELATIONSHIP_ADAPTER_ID,
        "adapter_version": core.RELATIONSHIP_ADAPTER_VERSION,
        "invalidates_prior_versions": True,
    }
    assert contract["schemas"]["normalized_transaction_columns"] == list(
        core.TRANSACTION_COLUMNS
    )
    assert contract["schemas"]["match_columns"] == list(core.MATCH_COLUMNS)
    assert contract["schemas"]["match_material_fields"] == list(
        core.MATCH_MATERIAL_FIELDS
    )
    assert contract["schemas"]["relationship_residual_columns"] == list(
        core.RESIDUAL_COLUMNS
    )
    assert contract["schemas"]["non_movement_columns"] == list(
        core.NON_MOVEMENT_COLUMNS
    )
    assert contract["amount_semantics"] == {
        "normalized_rows": {
            "amount_signed": (
                "canonical Decimal text that preserves the normalized source sign "
                "and is not replaced by a non-negative magnitude"
            ),
            "amount_abs_formula": "abs(Decimal(amount_signed))",
            "amount_abs_non_negative": True,
        },
        "matching": {
            "eligibility_amount_field": "amount_abs",
            "tolerance_comparison_formula": (
                "abs(bank.amount_abs - journal.amount_abs) <= reviewed "
                "amount_tolerance"
            ),
            "bank_amount_formula": "matched bank endpoint amount_signed",
            "journal_amount_formula": "matched journal endpoint amount_signed",
            "amount_delta_formula": (
                "abs(abs(Decimal(bank_amount)) - abs(Decimal(journal_amount)))"
            ),
            "amount_delta_non_negative": True,
            "sign_presentation_is_distinct_from_magnitude_matching": True,
        },
        "relationship_magnitude": {
            "source_record_amount_formula": "bank endpoint amount_abs",
            "target_record_amount_formula": "journal endpoint amount_abs",
            "allocation_amount_formula": (
                "min(source record amount, target record amount) for each accepted "
                "one-to-one match"
            ),
            "allocation_amount_non_negative": True,
            "allocated_amount_formula": (
                "sum(non-negative allocation amounts for the record)"
            ),
            "residual_formula": "record_amount - allocated_amount",
            "record_allocated_and_residual_values_non_negative": True,
        },
    }
    assert contract["review_authority"]["mapping"]["fields"] == list(
        core.MAPPING_FIELDS
    )
    assert (
        "date_convention" in contract["review_authority"]["mapping"]["receipt_bindings"]
    )
    date_contract = contract["review_authority"]["date_interpretation"]
    assert date_contract["reviewed_values"] == list(core.DATE_CONVENTIONS)
    assert date_contract["list_order_guessing_allowed"] is False
    assert date_contract["classification_vocabulary"] == [
        "parsed",
        "blank",
        "ambiguous",
        "invalid",
    ]
    assert date_contract["ambiguous_without_current_receipt"] == {
        "qualification_status": "needs_review",
        "failure_kind": "mapping_review_required",
        "emitted_row_count": 0,
        "run_block_code": "mapping_review_required",
    }
    assert date_contract["invalid_populated_date"] == {
        "row_disposition": "invalid_date_value",
        "qualification_status": "unsupported_source_layout",
        "failure_kind": "candidate_row_contract_failed",
        "emitted_row_count": 0,
        "stable_reference_does_not_override": True,
    }
    delimiter_contract = contract["review_authority"]["csv_field_delimiter"]
    assert delimiter_contract["supported_one_byte_values"] == list(
        core.CSV_FIELD_DELIMITERS
    )
    assert delimiter_contract["default_value"] == core.DEFAULT_CSV_FIELD_DELIMITER
    assert delimiter_contract["profile_max_bytes"] == 128 * 1024
    assert delimiter_contract["profile_max_rows"] == 100
    assert delimiter_contract["record_terminator_transport"] == {
        "supported_byte_sequences": ["LF", "CRLF", "CR"],
        "normalization": (
            "mechanically normalize universal-newline forms to LF before the "
            "full strict CSV parse"
        ),
        "reviewed_mapping_field": False,
        "distinct_from_csv_field_delimiter": True,
        "distinct_from_numeric_separators": True,
    }
    assert delimiter_contract["full_parse"] == {
        "ignore_errors": False,
        "truncate_ragged_lines": False,
        "malformed_record_result": {
            "qualification_status": "unsupported_source_layout",
            "failure_kind": "parser_failure",
            "emitted_row_count": 0,
            "run_block_code": "parser_failure",
        },
    }
    assert delimiter_contract["parser_failure_boundary"] == (
        "parser_failure is reserved for an actual malformed, ragged, undecodable, "
        "or otherwise failed full parse; an ambiguous or unsupported delimiter is "
        "not a parser failure"
    )
    assert delimiter_contract["scored_outcomes"] == {
        "ambiguous": {
            "qualification_status": "needs_review",
            "failure_kind": "ambiguous_csv_field_delimiter",
            "emitted_row_count": 0,
            "run_block_code": "mapping_review_required",
            "gate_statuses": {
                "source": "blocked",
                "preparation": "blocked",
                "reconciliation": "blocked",
                "semantic_review": "not_assessed",
                "reporting": "blocked",
                "publication": "blocked",
            },
        },
        "unsupported": {
            "qualification_status": "unsupported_source_layout",
            "failure_kind": "unsupported_csv_field_delimiter",
            "emitted_row_count": 0,
            "run_block_code": "unsupported_source_layout",
            "gate_statuses": {
                "source": "failed",
                "preparation": "blocked",
                "reconciliation": "blocked",
                "semantic_review": "not_assessed",
                "reporting": "blocked",
                "publication": "blocked",
            },
        },
        "actual_malformed_or_failed_full_parse": {
            "qualification_status": "unsupported_source_layout",
            "failure_kind": "parser_failure",
            "emitted_row_count": 0,
            "run_block_code": "parser_failure",
            "gate_statuses": {
                "source": "failed",
                "preparation": "blocked",
                "reconciliation": "blocked",
                "semantic_review": "not_assessed",
                "reporting": "blocked",
                "publication": "blocked",
            },
        },
    }
    disposition_contract = contract["review_authority"]["monetary_column_disposition"]
    assert disposition_contract["header_tokens"] == list(core.MONETARY_HEADER_TOKENS)
    assert disposition_contract["public_reviewed_intent_fields"] == [
        "potential_monetary_columns",
        "excluded_monetary_columns",
    ]
    assert disposition_contract["derived_potential_definition"] == {
        "boolean_operator": "OR",
        "candidate_if": [
            (
                "normalized column header contains at least one exact supported "
                "monetary token"
            ),
            (
                "at least one populated value parses exactly under the reviewed "
                "numeric-separator convention"
            ),
        ],
        "mapped_non_monetary_column_is_excluded_unless_header_is_explicitly_monetary": (
            True
        ),
        "source_row_and_source_sheet_columns_are_never_candidates": True,
    }
    assert disposition_contract["declaration_invariants"] == {
        "potential_entries_are_unique_non_empty_column_names": True,
        "potential_ordered_list_equals_current_derived_evidence": True,
        "excluded_entries_are_unique_non_empty_column_names": True,
        "excluded_is_subset_of_potential": True,
        "mapped_monetary_roles": ["amount", "debit", "credit"],
        "mapped_monetary_columns_are_subset_of_potential": True,
        "mapped_and_excluded_are_disjoint": True,
        "every_potential_column_is_mapped_or_excluded": True,
    }
    assert contract["review_authority"]["direction"]["canonical_values"] == sorted(
        core.CANONICAL_DIRECTIONS
    )
    assert contract["review_authority"]["relationship_perimeter"]["fields"] == sorted(
        core.RELATIONSHIP_POLICY_FIELDS
    )
    assert contract["review_authority"]["relationship_perimeter"][
        "direction_policy"
    ] == {
        "allowed_values": [
            "absolute_amount",
            "same_sign",
            "opposite_sign",
        ],
        "semantics": {
            "absolute_amount": (
                "Do not compare bank and journal canonical directions after each "
                "source row independently passes its reviewed vocabulary and "
                "signed-amount controls."
            ),
            "same_sign": (
                "Require both rows to have canonical direction positive, negative, "
                "or zero and require exact equality."
            ),
            "opposite_sign": (
                "Require the unordered direction pair to be exactly positive and "
                "negative; zero is never eligible."
            ),
        },
        "aliases_or_free_text_allowed": False,
        "reviewed_value_must_equal_execution_value": True,
        "missing_or_unsupported_result": {
            "run_block_code": "relationship_review_required",
            "reconciliation_gate_status": "blocked",
            "reporting_gate_status": "blocked",
        },
    }
    stable_reference = contract["review_authority"]["stable_reference"]
    assert stable_reference["generic_semantic_tokens"] == sorted(
        core.REFERENCE_GENERIC_TOKENS
    )
    assert stable_reference["generic_period_prefixes"] == list(
        core.GENERIC_PERIOD_REFERENCE_PREFIXES
    )
    assert stable_reference["generic_period_words"] == sorted(core.GENERIC_PERIOD_WORDS)
    assert stable_reference["generic_period_regex"] == (
        core.GENERIC_PERIOD_REFERENCE_RE.pattern
    )
    assert stable_reference["generic_version_regex"] == (
        core.GENERIC_PERIOD_VERSION_RE.pattern
    )
    assert stable_reference["generic_period_code_regex"] == (
        core.GENERIC_PERIOD_CODE_RE.pattern
    )
    assert stable_reference["minimum_token_length"] == 5
    assert stable_reference["digit_required"] is True
    assert contract["matching"]["stage_order"] == list(core.MATCH_STAGE_ORDER)
    assert list(contract["matching"]["stages"]) == list(core.MATCH_STAGE_ORDER)
    assert contract["matching"]["row_order_may_break_collision"] is False
    condition_outcomes = {
        entry["condition"]: entry for entry in contract["condition_outcome_matrix"]
    }
    invalid_date_outcome = condition_outcomes[
        "invalid populated date on any monetary candidate row"
    ]
    assert invalid_date_outcome["qualification_status"] == ("unsupported_source_layout")
    assert invalid_date_outcome["failure_kind"] == "candidate_row_contract_failed"
    assert invalid_date_outcome["run_block_code"] == "unsupported_source_layout"
    assert invalid_date_outcome["emitted_row_count"] == 0
    ambiguous_delimiter_outcome = condition_outcomes[
        "ambiguous CSV field delimiter without a current reviewed delimiter"
    ]
    assert ambiguous_delimiter_outcome == {
        "condition": (
            "ambiguous CSV field delimiter without a current reviewed delimiter"
        ),
        **delimiter_contract["scored_outcomes"]["ambiguous"],
    }
    unsupported_delimiter_outcome = condition_outcomes[
        "unsupported CSV field delimiter"
    ]
    assert unsupported_delimiter_outcome == {
        "condition": "unsupported CSV field delimiter",
        **delimiter_contract["scored_outcomes"]["unsupported"],
    }
    parser_failure_outcome = condition_outcomes[
        "actual malformed or otherwise failed full CSV parse"
    ]
    assert parser_failure_outcome == {
        "condition": "actual malformed or otherwise failed full CSV parse",
        **delimiter_contract["scored_outcomes"][
            "actual_malformed_or_failed_full_parse"
        ],
    }
    direction_policy_outcome = condition_outcomes[
        "missing, stale, aliased, free-text, or unsupported relationship "
        "direction policy"
    ]
    assert direction_policy_outcome["run_block_code"] == (
        "relationship_review_required"
    )
    assert direction_policy_outcome["gate_statuses"]["reconciliation"] == "blocked"
    assert direction_policy_outcome["gate_statuses"]["reporting"] == "blocked"
    material_contract = contract["material_value_address_contract"]
    assert material_contract["schema_version"] == (
        core.MATERIAL_VALUE_LEDGER_SCHEMA_VERSION
    )
    assert material_contract["datasets"]["matches"]["identity_fields"] == [
        "bank_transaction_id",
        "journal_transaction_id",
    ]
    assert material_contract["datasets"]["relationship_residuals"][
        "identity_fields"
    ] == ["side", "record_ref", "transaction_id"]
    assert material_contract["coverage"]["all_prepared_rows"] is True
    assert material_contract["coverage"]["all_declared_material_fields"] is True
    assert material_contract[
        "absent_when_qualification_or_relationship_authority_blocks"
    ]
    assert material_contract["fresh_replay"][-1] == (
        "rebuild and compare the complete material-value ledger and content hash"
    )
    assert contract["outputs"]["native_files"] == list(core.NATIVE_OUTPUT_FILES)
    assert contract["outputs"]["initial_run_files"] == list(
        core.INITIAL_RUN_OUTPUT_FILES
    )
    assert contract["outputs"]["post_review_additional_files"] == list(
        core.POST_REVIEW_OUTPUT_FILES
    )
    workbook_contract = contract["outputs"]["workbook"]
    assert workbook_contract["sheet_order"] == list(core.WORKBOOK_SHEET_ORDER)
    assert list(workbook_contract["sheets"]) == list(core.WORKBOOK_SHEET_ORDER)
    assert workbook_contract["sheets"]["matches"]["columns_schema"] == "match_columns"
    assert (
        workbook_contract["sheets"]["relationship_residuals"]["columns_schema"]
        == "relationship_residual_columns"
    )
    assert (
        workbook_contract["sheets"]["unmatched_bank"]["columns_schema"]
        == "normalized_transaction_columns"
    )
    assert (
        workbook_contract["sheets"]["unmatched_journal"]["columns_schema"]
        == "normalized_transaction_columns"
    )
    assert (
        workbook_contract["sheets"]["bank_pdf_non_movements"]["columns_schema"]
        == "non_movement_columns"
    )
    assert (
        workbook_contract["sheets"]["normalized_bank"]["columns_schema"]
        == "normalized_transaction_columns"
    )
    assert (
        workbook_contract["sheets"]["normalized_journal"]["columns_schema"]
        == "normalized_transaction_columns"
    )
    assert (
        workbook_contract["closure"]["independent_oracle_ooxml_package_required"]
        is False
    )
    oracle_contract = contract["oracle_comparison_contract"]
    assert oracle_contract["candidate_ab_repeatability"]["file_count"] == 14
    assert oracle_contract["candidate_ab_repeatability"]["files"] == list(
        core.DETERMINISTIC_ARTIFACT_FILES
    )
    assert oracle_contract["hidden_expected_rows_allowed"] is False
    assert oracle_contract["example_rows_are_contract_authority"] is False
    semantic_projection = oracle_contract["independent_oracle_semantic_projection"]
    assert semantic_projection["matches"]["amount_semantics"] == {
        "bank_amount_formula": "matched bank endpoint amount_signed",
        "journal_amount_formula": "matched journal endpoint amount_signed",
        "amount_delta_formula": (
            "abs(abs(Decimal(bank_amount)) - abs(Decimal(journal_amount)))"
        ),
        "matching_eligibility_and_tolerance_field": "amount_abs",
    }
    assert semantic_projection["relationship_ledger_and_residuals"][
        "amount_semantics_reference"
    ] == ("amount_semantics.relationship_magnitude")
    assert semantic_projection["source_and_gate_outcomes"][
        "csv_delimiter_outcomes_reference"
    ] == ("review_authority.csv_field_delimiter.scored_outcomes")
    equality_contract = contract["deterministic_artifact_equality"]
    assert equality_contract["byte_identical_files"] == list(
        core.DETERMINISTIC_ARTIFACT_FILES
    )
    assert equality_contract["material_closure_files"] == list(
        core.MATERIAL_CLOSURE_FILES
    )
    assert equality_contract["excluded_run_scoped_files"] == list(
        core.RUN_SCOPED_ARTIFACT_FILES
    )
    assert set(core.NATIVE_OUTPUT_FILES) == {
        *core.DETERMINISTIC_ARTIFACT_FILES,
        *core.MATERIAL_CLOSURE_FILES,
        *core.RUN_SCOPED_ARTIFACT_FILES,
    }
    assert contract["assurance"]["gate_names"] == [
        "source",
        "preparation",
        "reconciliation",
        "semantic_review",
        "reporting",
        "publication",
    ]
    assert contract["assurance"]["gate_status_vocabulary"] == [
        "passed",
        "failed",
        "blocked",
        "not_assessed",
        "not_applicable",
        "withheld",
    ]
    assert contract["assurance"]["report_ready_required_gates"] == [
        "source",
        "preparation",
        "reconciliation",
        "semantic_review",
        "reporting",
    ]
    assert '"expected_rows"' not in raw_contract
    assert '"example_rows"' not in raw_contract


def test_tabular_v7_extension_contract_is_frozen_and_matches_production() -> None:
    # Arrange
    core = load_core()
    contract_bytes = TABULAR_V7_EXTENSION_CONTRACT_PATH.read_bytes()
    raw_contract = contract_bytes.decode("utf-8")

    # Act
    contract = json.loads(raw_contract, object_pairs_hook=_strict_json_object)

    # Assert
    assert hashlib.sha256(contract_bytes).hexdigest() == (
        "74f779325acf234cbbf126b2060d43ea63a2788f6f645d36a750cd3ec4910347"
    )
    assert contract["schema_version"] == ("journal_bank.tabular_extension_contract.v1")
    assert contract["contract_id"] == "journal_bank.tabular.v7.extension.v1"
    assert contract["base_contract"]["sha256"] == (
        "4824652ecdb990a844fd9b72d799a2537f46a21ddb9fefb8a664f828c0ec6657"
    )
    assert contract["extension_adapter"] == {
        "adapter_id": core.EXTENDED_TABULAR_ADAPTER_ID,
        "adapter_version": core.EXTENDED_TABULAR_ADAPTER_VERSION,
        "selection": {
            "v7_when": [
                "date_locale is explicitly reviewed and non-null",
                (
                    "non_movement_summary_labels contains at least one explicitly "
                    "reviewed label"
                ),
            ],
            "v6_when": ("date_locale is null and non_movement_summary_labels is empty"),
            "silent_upgrade_forbidden": True,
        },
    }
    assert contract["localized_date_contract"]["date_locale_values"] == list(
        core.DATE_LOCALES
    )
    assert contract["localized_date_contract"]["month_vocabulary"] == (
        core.ITALIAN_MONTH_NUMBERS
    )
    assert contract["reviewed_summary_contract"]["field"] == (
        "non_movement_summary_labels"
    )
    assert contract["privacy_and_evaluation"]["contains_private_source_values"] is False
    assert '"example_rows"' not in raw_contract
    assert '"expected_rows"' not in raw_contract


def test_plugin_inspects_and_runs_deterministic_journal_bank_reconciliation(
    tmp_path: Path,
) -> None:
    core = load_core()
    client_run_id = "run_" + "a" * 24
    bank_path = tmp_path / "bank.xlsx"
    journal_path = tmp_path / "journal.xlsx"
    output_dir = tmp_path / "out"
    reconciliation_dir = output_dir / "reconciliation"
    _save_workbook(
        bank_path,
        [
            ["Date", "Description", "Amount", "Reference", "Beneficiary"],
            ["2025-01-02", "Payment invoice INV100 ACME", 123.45, "INV100", "ACME"],
            ["2025-01-05", "Unmatched fee", 9.99, "FEE9", "Bank"],
        ],
    )
    _save_workbook(
        journal_path,
        [
            ["Data", "Descrizione", "Dare", "Avere", "Riferimento", "Beneficiario"],
            ["2025-01-01", "Invoice INV100 ACME", 123.45, None, "INV100", "ACME"],
            ["2025-01-07", "Unmatched supplier", 77.0, None, "SUP77", "Supplier"],
        ],
    )

    inspection = core.inspect_inputs(
        bank_path,
        journal_path,
        output_dir,
        language="it",
        document_language="it",
    )
    recipe_path = _seal_relationship_recipe(
        core,
        output_dir / "suggested_recipe.json",
        output_dir / "input_receipts.json",
    )
    result = core.run_reconciliation(
        bank_path,
        journal_path,
        reconciliation_dir,
        recipe_path,
        language="it",
        document_language="it",
        client_run_id=client_run_id,
    )

    inspection_payload = json.loads((output_dir / "inspection.json").read_text())
    recipe_payload = json.loads((output_dir / "suggested_recipe.json").read_text())
    audit_payload = json.loads(
        (reconciliation_dir / "reconciliation_audit.json").read_text()
    )
    run_intake = json.loads((reconciliation_dir / "run_intake.json").read_text())
    review_payload = json.loads(
        (reconciliation_dir / "review_payload.json").read_text()
    )
    ui_decisions = json.loads((reconciliation_dir / "ui_decisions.json").read_text())
    final_artifacts = json.loads(
        (reconciliation_dir / "final_artifacts.json").read_text()
    )
    input_receipts = json.loads(
        (reconciliation_dir / "input_receipts.json").read_text()
    )
    source_qualifications = json.loads(
        (reconciliation_dir / "source_qualifications.json").read_text()
    )
    lineage = json.loads((reconciliation_dir / "lineage.json").read_text())
    assurance_gates = json.loads(
        (reconciliation_dir / "assurance_gates.json").read_text()
    )
    match = result.matches.to_dicts()[0]
    unmatched_bank = result.unmatched_bank.to_dicts()[0]
    unmatched_journal = result.unmatched_journal.to_dicts()[0]
    relationship_residuals = _read_csv_dicts(
        reconciliation_dir / "relationship_residuals.csv"
    )

    assert inspection.bank["row_count"] == 2
    assert inspection_payload["language"] == "it"
    assert recipe_payload["bank"]["files"]["bank.xlsx"]["mapping"]["date"] == "Date"
    assert match["status"] == "matched"
    assert match["stage"] == "reference"
    assert "inv100" in match["shared_references"].split(",")
    assert result.unmatched_bank.height == 1
    assert result.unmatched_journal.height == 1
    assert audit_payload["matched_count"] == 1
    assert audit_payload["status"] == "completed_with_unresolved_reconciliation"
    assert audit_payload["source_qualification_status"] == "qualified"
    assert audit_payload["unmatched_bank_count"] == 1
    assert audit_payload["unmatched_journal_count"] == 1
    assert audit_payload["review_session"]["run_id"] == run_intake["run_id"]
    assert (reconciliation_dir / "normalized_bank.csv").exists()
    assert (reconciliation_dir / "normalized_journal.csv").exists()
    assert (reconciliation_dir / "reconciliation_matches.csv").exists()
    assert (reconciliation_dir / "relationship_residuals.csv").exists()
    assert (reconciliation_dir / "material_value_ledger.json").exists()
    assert (reconciliation_dir / "unmatched_bank.csv").exists()
    assert (reconciliation_dir / "unmatched_journal.csv").exists()
    assert (reconciliation_dir / "bank_pdf_non_movement_rows.csv").exists()
    assert (reconciliation_dir / "journal_bank_reconciliation.xlsx").exists()
    assert (reconciliation_dir / "review_notes.md").exists()
    assert (reconciliation_dir / "run_intake.json").exists()
    assert (reconciliation_dir / "review_payload.json").exists()
    assert (reconciliation_dir / "ui_decisions.json").exists()
    assert (reconciliation_dir / "final_artifacts.json").exists()
    assert (reconciliation_dir / "input_receipts.json").exists()
    assert (reconciliation_dir / "source_qualifications.json").exists()
    assert (reconciliation_dir / "lineage.json").exists()
    assert (reconciliation_dir / "assurance_gates.json").exists()
    assert (reconciliation_dir / "artifact_receipts.json").exists()
    receipt_ids = {receipt["artifact_id"] for receipt in input_receipts["receipts"]}
    assert len(receipt_ids) == 2
    assert any(value.startswith("source.bank.1.") for value in receipt_ids)
    assert any(value.startswith("source.journal.1.") for value in receipt_ids)
    assert all(
        receipt["artifact_id"].endswith(receipt["sha256"])
        for receipt in input_receipts["receipts"]
    )
    assert all(
        set(receipt)
        == {
            "artifact_id",
            "byte_count",
            "path",
            "role",
            "root_id",
            "schema_version",
            "sha256",
        }
        for receipt in input_receipts["receipts"]
    )
    assert source_qualifications["status"] == "qualified"
    assert {entry["source_artifact_ref"] for entry in lineage["entries"]} == receipt_ids
    assert assurance_gates["gates"]["source"]["status"] == "passed"
    assert assurance_gates["gates"]["reconciliation"]["status"] == "withheld"
    assert assurance_gates["report_ready"] is False
    assert run_intake["plugin"] == "journal-bank-reconciliation"
    assert run_intake["run_id"] == client_run_id
    assert review_payload["run_id"] == run_intake["run_id"]
    assert review_payload["review_type"] == "journal_bank_reconciliation_review"
    assert review_payload["item_count"] == len(review_payload["items"])
    item_types = {item["item_type"] for item in review_payload["items"]}
    assert {"matched_pair", "unmatched_bank", "unmatched_journal"} <= item_types
    matched_item = next(
        item for item in review_payload["items"] if item["item_type"] == "matched_pair"
    )
    unmatched_bank_item = next(
        item
        for item in review_payload["items"]
        if item["item_type"] == "unmatched_bank"
    )
    unmatched_journal_item = next(
        item
        for item in review_payload["items"]
        if item["item_type"] == "unmatched_journal"
    )
    assert matched_item["data"]["target_artifact"] == "reconciliation_matches.csv"
    assert matched_item["data"]["target_id_field"] == "bank_transaction_id"
    assert matched_item["data"]["target_record_id"] == match["bank_transaction_id"]
    assert matched_item["data"]["target_field"] == "review_note"
    assert unmatched_bank_item["recommended_action"] == "request_more_documents"
    assert unmatched_bank_item["data"]["requested_document"] == (
        "Journal or ledger support for bank transaction FEE9"
    )
    assert unmatched_bank_item["data"]["reason"] == (
        "Bank transaction has no deterministic journal match."
    )
    assert any(
        evidence.get("kind") == "missing_reconciliation_evidence"
        and evidence.get("requested_document")
        == "Journal or ledger support for bank transaction FEE9"
        for evidence in unmatched_bank_item["evidence"]
    )
    assert unmatched_journal_item["recommended_action"] == "request_more_documents"
    assert unmatched_journal_item["data"]["requested_document"] == (
        "Bank statement or payment evidence for journal transaction SUP77"
    )
    assert unmatched_journal_item["data"]["reason"] == (
        "Journal transaction has no deterministic bank match."
    )
    assert review_payload["summary"]["matched_count"] == 1
    assert review_payload["summary"]["unmatched_bank_count"] == 1
    assert review_payload["summary"]["unmatched_journal_count"] == 1
    assert ui_decisions["status"] == "pending_review"
    assert final_artifacts["status"] == "written_pending_review"
    outputs_by_path = {output["path"]: output for output in final_artifacts["outputs"]}
    excluded_control_paths = {
        "final_artifacts.json",
        "review_payload.json",
        "run_intake.json",
        "ui_decisions.json",
    }
    current_output_paths = {
        path.relative_to(reconciliation_dir).as_posix()
        for path in reconciliation_dir.rglob("*")
        if path.is_file() and path.name not in excluded_control_paths
    }
    assert set(outputs_by_path) == current_output_paths
    assert "assurance_envelope.json" in outputs_by_path
    for relative, output in outputs_by_path.items():
        assert output["size_bytes"] == (reconciliation_dir / relative).stat().st_size

    artifact_receipts = json.loads(
        (reconciliation_dir / "artifact_receipts.json").read_text()
    )["output_receipts"]
    receipt_paths = {receipt["path"] for receipt in artifact_receipts}
    assert "artifact_receipts.json" not in receipt_paths
    assert "assurance_envelope.json" in receipt_paths
    final_artifact_receipt = next(
        receipt
        for receipt in artifact_receipts
        if receipt["path"] == "final_artifacts.json"
    )
    assert (
        final_artifact_receipt["byte_count"]
        == (reconciliation_dir / "final_artifacts.json").stat().st_size
    )
    core.validate_artifact_receipt(reconciliation_dir, final_artifact_receipt)

    handoff_output = next(
        output
        for output in final_artifacts["outputs"]
        if output["path"] == "review_handoff.md"
    )
    handoff_text = (reconciliation_dir / "review_handoff.md").read_text(
        encoding="utf-8"
    )
    assert handoff_output["required_text"] == [
        "Review Handoff",
        "review_payload.json",
        "ui_decisions.json",
        "applied_decisions.json",
        "final_artifacts.json",
    ]
    assert handoff_output["qa_checks"] == ["nonempty_text", "required_text"]
    assert (
        handoff_output["size_bytes"]
        == (reconciliation_dir / "review_handoff.md").stat().st_size
    )
    assert "render_journal_bank_review" in handoff_text
    assert "apply_journal_bank_decisions" in handoff_text
    review_notes_output = next(
        output
        for output in final_artifacts["outputs"]
        if output["path"] == "review_notes.md"
    )
    assert review_notes_output["required_text"] == [
        "# Journal-Bank Reconciliation Review Notes",
        "## Stage Counts",
        "## Review Policy",
    ]
    workbook_output = next(
        output
        for output in final_artifacts["outputs"]
        if output["path"] == "journal_bank_reconciliation.xlsx"
    )
    assert workbook_output["required_sheets"] == [
        "matches",
        "relationship_residuals",
        "unmatched_bank",
        "unmatched_journal",
        "bank_pdf_non_movements",
        "normalized_bank",
        "normalized_journal",
    ]
    assert workbook_output["source_row_counts"] == {
        "matches": audit_payload["matched_count"],
        "relationship_residuals": audit_payload["relationship_residual_row_count"],
        "unmatched_bank": audit_payload["unmatched_bank_count"],
        "unmatched_journal": audit_payload["unmatched_journal_count"],
        "bank_pdf_non_movements": audit_payload["bank_pdf_non_movement_row_count"],
        "normalized_bank": audit_payload["bank_row_count"],
        "normalized_journal": audit_payload["journal_row_count"],
    }
    assert workbook_output["required_sheet_headers"] == {
        "matches": core.MATCH_COLUMNS,
        "relationship_residuals": core.RESIDUAL_COLUMNS,
        "unmatched_bank": core.TRANSACTION_COLUMNS,
        "unmatched_journal": core.TRANSACTION_COLUMNS,
        "bank_pdf_non_movements": core.NON_MOVEMENT_COLUMNS,
        "normalized_bank": core.TRANSACTION_COLUMNS,
        "normalized_journal": core.TRANSACTION_COLUMNS,
    }
    assert workbook_output["required_cells"] == {
        "matches": {
            "A1": "status",
            "A2": "matched",
            "B1": "stage",
            "B2": "reference",
            "C1": "bank_transaction_id",
            "C2": match["bank_transaction_id"],
            "D1": "journal_transaction_id",
            "D2": match["journal_transaction_id"],
            "M1": "shared_references",
            "M2": match["shared_references"],
        },
        "relationship_residuals": {
            "A1": "side",
            "A2": relationship_residuals[0]["side"],
            "B1": "record_ref",
            "B2": relationship_residuals[0]["record_ref"],
            "C1": "transaction_id",
            "C2": relationship_residuals[0]["transaction_id"],
            "D1": "record_amount",
            "D2": relationship_residuals[0]["record_amount"],
            "E1": "allocated_amount",
            "E2": relationship_residuals[0]["allocated_amount"],
            "F1": "residual",
            "F2": relationship_residuals[0]["residual"],
        },
        "unmatched_bank": {
            "A1": "side",
            "A2": "bank",
            "B1": "transaction_id",
            "B2": unmatched_bank["transaction_id"],
            "C1": "transaction_date",
            "C2": unmatched_bank["transaction_date"],
            "H1": "reference",
            "H2": "FEE9",
        },
        "unmatched_journal": {
            "A1": "side",
            "A2": "journal",
            "B1": "transaction_id",
            "B2": unmatched_journal["transaction_id"],
            "C1": "transaction_date",
            "C2": unmatched_journal["transaction_date"],
            "H1": "reference",
            "H2": "SUP77",
        },
        "bank_pdf_non_movements": {
            "B1": "source_file",
            "C1": "source_sheet",
            "D1": "source_row",
            "E1": "classification",
            "J1": "description",
            "I1": "amount_abs",
        },
    }
    assert "required_cells" in workbook_output["qa_checks"]
    matches_output = next(
        output
        for output in final_artifacts["outputs"]
        if output["path"] == "reconciliation_matches.csv"
    )
    assert matches_output["row_count"] == audit_payload["matched_count"]
    assert matches_output["required_columns"] == [
        "status",
        "bank_transaction_id",
        "journal_transaction_id",
        "amount_delta",
    ]
    residual_output = next(
        output
        for output in final_artifacts["outputs"]
        if output["path"] == "relationship_residuals.csv"
    )
    assert (
        residual_output["row_count"] == audit_payload["relationship_residual_row_count"]
    )
    assert residual_output["required_columns"] == core.RESIDUAL_COLUMNS
    contract_report = validate_contract(
        reconciliation_dir,
        strict_data_posture=True,
        strict_execution_trace=True,
        strict_output_paths=True,
        strict_output_content=True,
    )
    assert contract_report.ok, contract_report.as_dict()


def test_final_artifact_closure_redeclares_recreated_receipt(
    tmp_path: Path,
) -> None:
    core = load_core()
    output_dir = tmp_path / "output"
    output_dir.mkdir()
    audit_path = output_dir / "reconciliation_audit.json"
    run_intake_path = output_dir / "run_intake.json"
    final_artifacts_path = output_dir / "final_artifacts.json"
    artifact_receipts_path = output_dir / "artifact_receipts.json"
    audit_path.write_text('{"status":"completed"}\n', encoding="utf-8")
    run_intake_path.write_text(
        '{"input_paths":[],"data_posture":{"local_files_read":[]}}\n',
        encoding="utf-8",
    )
    final_artifacts_path.write_text('{"outputs":[]}\n', encoding="utf-8")
    paths = {
        "audit_json": audit_path,
        "artifact_receipts_json": artifact_receipts_path,
        "run_intake_json": run_intake_path,
        "final_artifacts_json": final_artifacts_path,
    }

    receipt_bundle = core._close_final_artifacts(
        output_dir,
        paths,
        source_receipts=[],
    )

    final_artifacts = json.loads(final_artifacts_path.read_text(encoding="utf-8"))
    declared_paths = {output["path"] for output in final_artifacts["outputs"]}
    assert declared_paths == {
        "artifact_receipts.json",
        "reconciliation_audit.json",
    }
    receipt_paths = {receipt["path"] for receipt in receipt_bundle["output_receipts"]}
    assert "artifact_receipts.json" not in receipt_paths
    final_receipt = next(
        receipt
        for receipt in receipt_bundle["output_receipts"]
        if receipt["path"] == "final_artifacts.json"
    )
    core.validate_artifact_receipt(output_dir, final_receipt)


def test_spanish_run_localizes_review_notes_and_strict_contract(tmp_path: Path) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.xlsx"
    journal_path = tmp_path / "journal.xlsx"
    output_dir = tmp_path / "output"
    _save_workbook(
        bank_path,
        [
            ["Date", "Description", "Amount", "Reference"],
            ["2026-01-15", "Pago factura ES100", 75.25, "ES100"],
            ["2026-01-16", "Comisión bancaria", -4.5, "COM-ES"],
        ],
    )
    _save_workbook(
        journal_path,
        [
            ["Date", "Description", "Debit", "Reference"],
            ["2026-01-15", "Factura ES100", 75.25, "ES100"],
        ],
    )
    recipe_path = _prepare_reviewed_recipe(
        core, bank_path, journal_path, tmp_path / "recipe"
    )

    core.run_reconciliation(
        bank_path,
        journal_path,
        output_dir,
        recipe_path,
        language="es-ES",
        document_language="en",
    )

    review_notes = (output_dir / "review_notes.md").read_text(encoding="utf-8")
    review_handoff = (output_dir / "review_handoff.md").read_text(encoding="utf-8")
    review_payload = json.loads(
        (output_dir / "review_payload.json").read_text(encoding="utf-8")
    )
    final_artifacts = json.loads(
        (output_dir / "final_artifacts.json").read_text(encoding="utf-8")
    )
    review_notes_output = next(
        output
        for output in final_artifacts["outputs"]
        if output["path"] == "review_notes.md"
    )
    handoff_output = next(
        output
        for output in final_artifacts["outputs"]
        if output["path"] == "review_handoff.md"
    )
    unmatched_item = next(
        item
        for item in review_payload["items"]
        if item["item_type"] == "unmatched_bank"
    )
    artifact_item = next(
        item
        for item in review_payload["items"]
        if item["item_type"] == "workpaper_artifact"
    )

    assert review_notes.startswith(
        "# Notas de revisión de la conciliación entre diario y banco\n"
    )
    assert "- Idioma: es" in review_notes
    assert "## Recuento por etapa" in review_notes
    assert "## Política de revisión" in review_notes
    assert review_notes_output["required_text"] == [
        "# Notas de revisión de la conciliación entre diario y banco",
        "## Recuento por etapa",
        "## Política de revisión",
    ]
    assert review_payload["language"] == "es"
    assert review_payload["columns"] == [
        {"field": "item_type", "label": "Tipo"},
        {"field": "title", "label": "Movimiento"},
        {"field": "recommended_action", "label": "Acción sugerida"},
        {"field": "source_path", "label": "Fuente"},
        {"field": "output_path", "label": "Salida"},
        {"field": "status", "label": "Estado"},
    ]
    assert unmatched_item["recommended_action"] == "request_more_documents"
    assert unmatched_item["data"]["requested_document"] == (
        "Justificante del diario o del mayor para la transacción bancaria COM-ES"
    )
    assert unmatched_item["data"]["reason"] == (
        "La transacción bancaria no tiene una correspondencia determinista en el diario."
    )
    assert artifact_item["title"] == "Libro de conciliación entre diario y banco"
    assert review_handoff.startswith(
        "# Entrega para revisión: Conciliación entre diario y banco\n"
    )
    assert "## Revisión en Codex" in review_handoff
    assert handoff_output["required_text"][0] == "Entrega para revisión"
    assert final_artifacts["caveats"][0].startswith("Las coincidencias deterministas")
    assert final_artifacts["next_actions"][1].startswith("Revise las filas bancarias")
    contract_report = validate_contract(
        output_dir,
        strict_data_posture=True,
        strict_execution_trace=True,
        strict_output_paths=True,
        strict_output_content=True,
    )
    assert contract_report.ok, contract_report.as_dict()


def test_plugin_keeps_ambiguous_rows_unmatched(tmp_path: Path) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.xlsx"
    journal_path = tmp_path / "journal.xlsx"
    output_dir = tmp_path / "out"
    _save_workbook(
        bank_path,
        [
            ["Date", "Description", "Amount"],
            ["2025-03-10", "Payment", 80],
        ],
    )
    _save_workbook(
        journal_path,
        [
            ["Date", "Description", "Amount"],
            ["2025-03-10", "Payment A", 80],
            ["2025-03-10", "Payment B", 80],
        ],
    )
    recipe_path = _prepare_reviewed_recipe(
        core, bank_path, journal_path, tmp_path / "recipe"
    )

    result = core.run_reconciliation(
        bank_path, journal_path, output_dir, recipe_path, language="en"
    )

    assert result.matches.height == 0
    assert result.unmatched_bank.height == 1
    assert result.unmatched_journal.height == 2


def test_amount_date_single_labels_only_a_later_singleton_wave(
    tmp_path: Path,
) -> None:
    # Arrange
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    recipe_dir = tmp_path / "recipe"
    output_dir = tmp_path / "out"
    _save_csv(
        bank_path,
        [
            ["Date", "Amount", "Reference", "Description"],
            ["2025-01-01", "100.00", "", "First singleton"],
            ["2025-01-02", "100.00", "", "Later singleton"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference", "Description"],
            ["2025-01-01", "100.00", "", "First target"],
            ["2025-01-03", "100.00", "", "Later target"],
        ],
    )
    recipe_path = _prepare_reviewed_recipe(
        core,
        bank_path,
        journal_path,
        recipe_dir,
        tolerance="0",
        date_window_days=1,
    )

    # Act
    result = core.run_reconciliation(
        bank_path,
        journal_path,
        output_dir,
        recipe_path,
        tolerance="0",
        date_window_days=1,
    )

    # Assert
    stages_by_description = {
        row["bank_description"]: row["stage"] for row in result.matches.to_dicts()
    }
    assert stages_by_description == {
        "First singleton": "amount_date_unique",
        "Later singleton": "amount_date_single",
    }


@pytest.mark.parametrize("shared_reference", ["", "COLLISION-100"])
@pytest.mark.parametrize("reverse_bank_rows", [False, True])
def test_competing_singletons_remain_ambiguous_in_any_bank_order(
    tmp_path: Path,
    shared_reference: str,
    reverse_bank_rows: bool,
) -> None:
    # Arrange
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    output_dir = tmp_path / "out"
    bank_rows = [
        ["2025-01-01", "100.00", shared_reference, "First bank row"],
        ["2025-01-01", "100.00", shared_reference, "Second bank row"],
    ]
    if reverse_bank_rows:
        bank_rows.reverse()
    _save_csv(
        bank_path,
        [
            ["Date", "Amount", "Reference", "Description"],
            *bank_rows,
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference", "Description"],
            ["2025-01-01", "100.00", shared_reference, "Shared journal row"],
        ],
    )
    recipe_path = _prepare_reviewed_recipe(
        core,
        bank_path,
        journal_path,
        tmp_path / "recipe",
        tolerance="0",
        date_window_days=0,
    )

    # Act
    result = core.run_reconciliation(
        bank_path,
        journal_path,
        output_dir,
        recipe_path,
        tolerance="0",
        date_window_days=0,
    )

    # Assert
    assert result.matches.is_empty()
    assert result.unmatched_bank.height == 2
    assert result.unmatched_journal.height == 1


def test_relationship_adapter_v1_receipt_is_stale_after_matching_semantics_change(
    tmp_path: Path,
) -> None:
    # Arrange
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    recipe_dir = tmp_path / "recipe"
    output_dir = tmp_path / "out"
    _save_csv(
        bank_path,
        [
            ["Date", "Amount", "Reference"],
            ["2025-01-01", "100.00", "REF-100"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference"],
            ["2025-01-01", "100.00", "REF-100"],
        ],
    )
    recipe_path = _prepare_reviewed_recipe(
        core,
        bank_path,
        journal_path,
        recipe_dir,
    )
    recipe = json.loads(recipe_path.read_text(encoding="utf-8"))
    current_decision = recipe["relationship"]["decision"]
    source_refs = current_decision["source_artifact_refs"]
    recipe["relationship"]["decision"] = core.build_reviewed_decision_receipt(
        decision_id="decision.relationship.v1",
        decision_type="journal_bank_relationship",
        status="reviewed",
        reviewer_ref="reviewer.test",
        reviewed_on="2026-07-24",
        adapter_id="journal_bank.relationship.v1",
        adapter_version="1",
        source_artifact_refs=source_refs,
        content={"policy": recipe["relationship"]["policy"]},
    )
    recipe_path.write_text(
        json.dumps(recipe, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    # Act
    with pytest.raises(core.ReconciliationBlockedError) as exc_info:
        core.run_reconciliation(
            bank_path,
            journal_path,
            output_dir,
            recipe_path,
        )

    # Assert
    audit = json.loads(
        (output_dir / "reconciliation_audit.json").read_text(encoding="utf-8")
    )
    assert exc_info.value.code == "relationship_review_required"
    assert "invalid or stale" in exc_info.value.detail
    assert audit["block_code"] == "relationship_review_required"
    assert current_decision["adapter_id"] == "journal_bank.relationship.v2"
    assert current_decision["adapter_version"] == "2"
    expected_files = set(core.INITIAL_RUN_OUTPUT_FILES) - {"material_value_ledger.json"}
    assert {
        path.name for path in output_dir.iterdir() if path.is_file()
    } == expected_files
    assert _read_csv_dicts(output_dir / "relationship_residuals.csv") == []
    assert len(_read_csv_dicts(output_dir / "unmatched_bank.csv")) == 1
    assert len(_read_csv_dicts(output_dir / "unmatched_journal.csv")) == 1
    assert not (output_dir / "material_value_ledger.json").exists()


def test_generated_workbook_is_byte_identical_across_separate_runs(
    tmp_path: Path,
) -> None:
    # Arrange
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    _save_csv(
        bank_path,
        [
            ["Date", "Amount", "Reference"],
            ["2025-01-01", "100.00", "REF-100"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference"],
            ["2025-01-01", "100.00", "REF-100"],
        ],
    )
    recipe_path = _prepare_reviewed_recipe(
        core,
        bank_path,
        journal_path,
        tmp_path / "recipe",
    )
    first_output = tmp_path / "first"
    second_output = tmp_path / "second"

    # Act
    core.run_reconciliation(bank_path, journal_path, first_output, recipe_path)
    core.run_reconciliation(bank_path, journal_path, second_output, recipe_path)

    # Assert
    first_workbook = first_output / "journal_bank_reconciliation.xlsx"
    second_workbook = second_output / "journal_bank_reconciliation.xlsx"
    assert first_workbook.read_bytes() == second_workbook.read_bytes()
    first_hashes = {
        relative: hashlib.sha256((first_output / relative).read_bytes()).hexdigest()
        for relative in core.DETERMINISTIC_ARTIFACT_FILES
    }
    second_hashes = {
        relative: hashlib.sha256((second_output / relative).read_bytes()).hexdigest()
        for relative in core.DETERMINISTIC_ARTIFACT_FILES
    }
    assert len(first_hashes) == 14
    assert first_hashes == second_hashes
    workbook = openpyxl.load_workbook(first_workbook, read_only=True)
    assert workbook.sheetnames == list(core.WORKBOOK_SHEET_ORDER)
    workbook.close()


def test_run_reconciliation_rejects_duplicate_ooxml_members(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    # Arrange
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    _save_csv(
        bank_path,
        [
            ["Date", "Amount", "Reference"],
            ["2025-01-01", "100.00", "REF-100"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference"],
            ["2025-01-01", "100.00", "REF-100"],
        ],
    )
    recipe_path = _prepare_reviewed_recipe(
        core,
        bank_path,
        journal_path,
        tmp_path / "recipe",
    )

    def save_with_duplicate_members(
        _workbook: Any,
        filename: str | Path,
    ) -> None:
        with ZipFile(filename, "w", compression=ZIP_DEFLATED) as archive:
            archive.writestr("docProps/core.xml", b"<core/>")
            archive.writestr("docProps/core.xml", b"<core/>")

    monkeypatch.setattr(core.openpyxl.Workbook, "save", save_with_duplicate_members)

    # Act / Assert
    with pytest.warns(UserWarning, match="Duplicate name"):
        with pytest.raises(
            ValueError,
            match="OOXML workbook contains duplicate member names",
        ):
            core.run_reconciliation(
                bank_path,
                journal_path,
                tmp_path / "out",
                recipe_path,
            )


def test_supplied_empty_sample_blocks_instead_of_using_full_journal(
    tmp_path: Path,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    sample_path = tmp_path / "sample.csv"
    output_dir = tmp_path / "out"
    _save_csv(
        bank_path,
        [
            ["Date", "Amount", "Reference"],
            ["2025-03-10", "80.00", "REF100"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference", "Movement"],
            ["2025-03-10", "80.00", "REF100", "M100"],
        ],
    )
    _save_csv(sample_path, [["Movement"]])

    with pytest.raises(core.ReconciliationBlockedError) as exc_info:
        core.run_reconciliation(
            bank_path,
            journal_path,
            output_dir,
            sample_path=sample_path,
        )

    audit = json.loads((output_dir / "reconciliation_audit.json").read_text())
    assert exc_info.value.code == "invalid_or_empty_sample"
    assert audit["status"] == "blocked"
    assert audit["sample_movement_count"] == 0
    assert (output_dir / "normalized_journal.csv").is_file()
    assert (output_dir / "reconciliation_matches.csv").is_file()


def test_amount_date_stages_require_actual_date_values(tmp_path: Path) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    output_dir = tmp_path / "out"
    _save_csv(
        bank_path,
        [
            ["Date", "Description", "Amount"],
            ["not-a-date", "Same description", "80.00"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Description", "Amount"],
            ["not-a-date", "Same description", "80.00"],
        ],
    )

    with pytest.raises(core.ReconciliationBlockedError) as exc_info:
        core.run_reconciliation(bank_path, journal_path, output_dir)

    audit = json.loads((output_dir / "reconciliation_audit.json").read_text())
    bank_diagnostic = audit["diagnostics"]["bank"][0]
    assert exc_info.value.code == "unsupported_source_layout"
    assert bank_diagnostic["row_disposition_counts"] == {"invalid_date_value": 1}
    assert bank_diagnostic["row_dispositions"][0]["source_row"] == 2


def test_reviewed_mapping_accepts_compact_iso_date_for_amount_date_match(
    tmp_path: Path,
) -> None:
    # Arrange
    core = load_core()
    bank_path = tmp_path / "westpac.csv"
    journal_path = tmp_path / "journal.csv"
    recipe_dir = tmp_path / "recipe"
    output_dir = tmp_path / "out"
    _save_csv(
        bank_path,
        [
            ["BankDate", "TransactionAmount", "TransactionID", "Narrative"],
            ["20260508", "-1250.00", "WB-001", "Aster Ledgerworks"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["PostingDate", "Amount", "Reference"],
            ["2026-05-08", "1250.00", ""],
        ],
    )
    core.inspect_inputs(bank_path, journal_path, recipe_dir)
    recipe_path = recipe_dir / "suggested_recipe.json"
    recipe = json.loads(recipe_path.read_text(encoding="utf-8"))
    receipts = json.loads(
        (recipe_dir / "input_receipts.json").read_text(encoding="utf-8")
    )["receipts"]
    mappings = {
        "bank": {
            "date": "BankDate",
            "amount": "TransactionAmount",
            "reference": "TransactionID",
            "description": "Narrative",
        },
        "journal": {
            "date": "PostingDate",
            "amount": "Amount",
            "reference": "Reference",
        },
    }
    for side, source_path in (("bank", bank_path), ("journal", journal_path)):
        source_ref = next(
            receipt["artifact_id"]
            for receipt in receipts
            if receipt["artifact_id"].startswith(f"source.{side}.")
        )
        file_recipe = recipe[side]["files"][source_path.name]
        file_recipe["mapping"] = mappings[side]
        file_recipe["mapping_decision"] = core.build_mapping_review_receipt(
            decision_id=f"decision.mapping.{side}",
            reviewer_ref="reviewer.test",
            reviewed_on="2026-07-25",
            source_artifact_ref=source_ref,
            side=side,
            source_file=source_path.name,
            header_rows=[1],
            mapping=mappings[side],
            potential_monetary_columns=file_recipe["potential_monetary_columns"],
            excluded_monetary_columns=file_recipe["excluded_monetary_columns"],
            csv_field_delimiter=file_recipe["csv_field_delimiter"],
        )
    recipe_path.write_text(
        json.dumps(recipe, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    _seal_relationship_recipe(
        core,
        recipe_path,
        recipe_dir / "input_receipts.json",
        tolerance="0",
        date_window_days=0,
    )

    # Act
    result = core.run_reconciliation(
        bank_path,
        journal_path,
        output_dir,
        recipe_path,
        tolerance="0",
        date_window_days=0,
    )

    # Assert
    normalized_bank = _read_csv_dicts(output_dir / "normalized_bank.csv")
    source_qualifications = json.loads(
        (output_dir / "source_qualifications.json").read_text(encoding="utf-8")
    )
    assert result.matches.height == 1
    assert result.matches.row(0, named=True)["stage"] == "amount_date_unique"
    assert normalized_bank[0]["transaction_date"] == "2026-05-08"
    assert {
        (entry["adapter_id"], entry["adapter_version"])
        for entry in source_qualifications["qualifications"]
    } == {("journal_bank.tabular.v6", "6")}


@pytest.mark.parametrize(
    ("source_date", "expected_date"),
    [
        (date(2026, 5, 8), "2026-05-08"),
        (datetime(2026, 5, 8, 14, 30), "2026-05-08"),
        (46150, "2026-05-08"),
        ("20260508", "2026-05-08"),
        ("2026/05/08", "2026-05-08"),
        ("2026.05.08", "2026-05-08"),
    ],
)
def test_native_serial_compact_and_year_first_dates_are_mechanical(
    tmp_path: Path,
    source_date: Any,
    expected_date: str,
) -> None:
    # Arrange
    core = load_core()
    bank_path = tmp_path / "bank.xlsx"
    journal_path = tmp_path / "journal.xlsx"
    for source_path in (bank_path, journal_path):
        _save_workbook(
            source_path,
            [
                ["Date", "Amount", "Reference"],
                [source_date, 80, "TX100"],
            ],
        )

    # Act
    inspection = core.inspect_inputs(
        bank_path,
        journal_path,
        tmp_path / "inspection",
    )

    # Assert
    bank_diagnostic = inspection.bank["files"][0]
    journal_diagnostic = inspection.journal["files"][0]
    assert inspection.bank["row_count"] == 1
    assert inspection.journal["row_count"] == 1
    assert bank_diagnostic["preview"][0]["transaction_date"] == expected_date
    assert journal_diagnostic["preview"][0]["transaction_date"] == expected_date
    assert bank_diagnostic["date_convention"] is None
    assert bank_diagnostic["mapping_decision"] is None


def test_ambiguous_day_month_dates_require_current_reviewed_convention(
    tmp_path: Path,
) -> None:
    # Arrange
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    for source_path in (bank_path, journal_path):
        _save_csv(
            source_path,
            [
                ["Date", "Amount", "Reference"],
                ["05/08/2026", "80.00", "TX100"],
            ],
        )

    # Act
    inspection = core.inspect_inputs(
        bank_path,
        journal_path,
        tmp_path / "inspection",
    )

    # Assert
    diagnostic = inspection.bank["files"][0]
    proposed = inspection.suggested_recipe["bank"]["files"][bank_path.name]
    assert inspection.bank["row_count"] == 0
    assert diagnostic["qualification_status"] == "needs_review"
    assert diagnostic["failure_kind"] == "mapping_review_required"
    assert diagnostic["date_convention"] is None
    assert diagnostic["date_interpretation_evidence"]["status_counts"] == {
        "parsed": 0,
        "blank": 0,
        "ambiguous": 1,
        "invalid": 0,
    }
    assert (
        "ambiguous day/month dates require a reviewed date_convention"
        in diagnostic["missing_required_mapping"]
    )
    assert proposed["date_convention"] is None
    assert proposed["mapping_decision"] is None


@pytest.mark.parametrize(
    ("date_convention", "expected_date"),
    [
        ("day_first", "2026-08-05"),
        ("month_first", "2026-05-08"),
    ],
)
def test_reviewed_day_month_conventions_produce_exact_different_dates(
    tmp_path: Path,
    date_convention: str,
    expected_date: str,
) -> None:
    # Arrange
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    output_dir = tmp_path / "out"
    for source_path in (bank_path, journal_path):
        _save_csv(
            source_path,
            [
                ["Date", "Amount", "Reference"],
                ["05/08/2026", "80.00", "TX100"],
            ],
        )
    recipe_path = _prepare_reviewed_date_recipe(
        core,
        bank_path,
        journal_path,
        tmp_path / "recipe",
        date_convention=date_convention,
    )

    # Act
    result = core.run_reconciliation(
        bank_path,
        journal_path,
        output_dir,
        recipe_path,
        tolerance="0",
        date_window_days=0,
    )

    # Assert
    normalized_bank = _read_csv_dicts(output_dir / "normalized_bank.csv")
    reviewed_decisions = json.loads(
        (output_dir / "reviewed_decisions.json").read_text(encoding="utf-8")
    )
    mapping_receipts = [
        decision
        for decision in reviewed_decisions["decisions"]
        if decision["decision_type"] == "journal_bank_mapping"
    ]
    assert result.matches.height == 1
    assert normalized_bank[0]["transaction_date"] == expected_date
    assert {receipt["content"]["date_convention"] for receipt in mapping_receipts} == {
        date_convention
    }


def test_localized_textual_month_dates_require_current_reviewed_locale(
    tmp_path: Path,
) -> None:
    # Arrange
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    for source_path in (bank_path, journal_path):
        _save_csv(
            source_path,
            [
                ["Date", "Amount", "Reference"],
                ["05 agosto 2026", "80.00", "TX100"],
            ],
        )

    # Act
    inspection = core.inspect_inputs(
        bank_path,
        journal_path,
        tmp_path / "inspection",
    )

    # Assert
    diagnostic = inspection.bank["files"][0]
    proposed = inspection.suggested_recipe["bank"]["files"][bank_path.name]
    assert inspection.bank["row_count"] == 0
    assert diagnostic["adapter_id"] == "journal_bank.tabular.v6"
    assert diagnostic["qualification_status"] == "needs_review"
    assert diagnostic["failure_kind"] == "mapping_review_required"
    assert diagnostic["date_interpretation_evidence"]["status_counts"] == {
        "parsed": 0,
        "blank": 0,
        "ambiguous": 0,
        "invalid": 0,
        "locale_required": 1,
    }
    assert diagnostic["date_interpretation_evidence"]["source_rows"][
        "locale_required"
    ] == [2]
    assert (
        "localized textual-month dates require a reviewed date_locale"
        in diagnostic["missing_required_mapping"]
    )
    assert proposed["date_locale"] is None
    assert proposed["mapping_decision"] is None


@pytest.mark.parametrize(
    "source_date",
    [
        "5 agosto 2026",
        "05 AGOSTO 2026",
        "05   agosto 2026",
        "05\u00a0agosto\u00a02026",
    ],
)
def test_reviewed_italian_textual_month_dates_use_adapter_v7(
    tmp_path: Path,
    source_date: str,
) -> None:
    # Arrange
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    output_dir = tmp_path / "out"
    for source_path in (bank_path, journal_path):
        _save_csv(
            source_path,
            [
                ["Date", "Amount", "Reference"],
                [source_date, "80.00", "TX100"],
            ],
        )
    recipe_path = _prepare_reviewed_date_locale_recipe(
        core,
        bank_path,
        journal_path,
        tmp_path / "recipe",
        date_locale="it",
    )

    # Act
    result = core.run_reconciliation(
        bank_path,
        journal_path,
        output_dir,
        recipe_path,
        tolerance="0",
        date_window_days=0,
    )

    # Assert
    normalized_bank = _read_csv_dicts(output_dir / "normalized_bank.csv")
    reviewed_decisions = json.loads(
        (output_dir / "reviewed_decisions.json").read_text(encoding="utf-8")
    )
    source_qualifications = json.loads(
        (output_dir / "source_qualifications.json").read_text(encoding="utf-8")
    )
    mapping_receipts = [
        decision
        for decision in reviewed_decisions["decisions"]
        if decision["decision_type"] == "journal_bank_mapping"
    ]
    assert result.matches.height == 1
    assert normalized_bank[0]["transaction_date"] == "2026-08-05"
    assert {
        (receipt["adapter_id"], receipt["adapter_version"])
        for receipt in mapping_receipts
    } == {("journal_bank.tabular.v7", "7")}
    assert {receipt["content"]["date_locale"] for receipt in mapping_receipts} == {"it"}
    assert {
        (qualification["adapter_id"], qualification["adapter_version"])
        for qualification in source_qualifications["qualifications"]
    } == {("journal_bank.tabular.v7", "7")}


@pytest.mark.parametrize(
    "source_date",
    [
        "31 aprile 2026",
        "05 august 2026",
        "05 ago 2026",
        "05 agosto 26",
        "05 agosto 2026 posted",
        "05\nagosto\n2026",
    ],
)
def test_reviewed_italian_locale_rejects_invalid_or_out_of_contract_dates(
    tmp_path: Path,
    source_date: str,
) -> None:
    # Arrange
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    for source_path in (bank_path, journal_path):
        _save_csv(
            source_path,
            [
                ["Date", "Amount", "Reference"],
                [source_date, "80.00", "TX100"],
            ],
        )
    recipe_path = _prepare_reviewed_date_locale_recipe(
        core,
        bank_path,
        journal_path,
        tmp_path / "recipe",
        date_locale="it",
    )

    # Act
    inspection = core.inspect_inputs(
        bank_path,
        journal_path,
        tmp_path / "reinspection",
        recipe_path,
    )

    # Assert
    diagnostic = inspection.bank["files"][0]
    assert inspection.bank["row_count"] == 0
    assert diagnostic["adapter_id"] == "journal_bank.tabular.v7"
    assert diagnostic["qualification_status"] == "unsupported_source_layout"
    assert diagnostic["failure_kind"] == "candidate_row_contract_failed"
    assert diagnostic["row_disposition_counts"] == {"invalid_date_value": 1}
    assert diagnostic["date_interpretation_evidence"]["status_counts"]["invalid"] == 1


def test_unsupported_date_locale_is_review_required_and_emits_zero_rows(
    tmp_path: Path,
) -> None:
    # Arrange
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    for source_path in (bank_path, journal_path):
        _save_csv(
            source_path,
            [
                ["Date", "Amount", "Reference"],
                ["05 agosto 2026", "80.00", "TX100"],
            ],
        )
    recipe_path = _prepare_reviewed_date_locale_recipe(
        core,
        bank_path,
        journal_path,
        tmp_path / "recipe",
        date_locale="it",
    )
    recipe = json.loads(recipe_path.read_text(encoding="utf-8"))
    recipe["bank"]["files"][bank_path.name]["date_locale"] = "en"
    recipe_path.write_text(
        json.dumps(recipe, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    # Act
    inspection = core.inspect_inputs(
        bank_path,
        journal_path,
        tmp_path / "reinspection",
        recipe_path,
    )

    # Assert
    diagnostic = inspection.bank["files"][0]
    assert inspection.bank["row_count"] == 0
    assert diagnostic["qualification_status"] == "needs_review"
    assert diagnostic["failure_kind"] == "mapping_review_required"
    assert "date_locale must be exactly it" in diagnostic["missing_required_mapping"]
    assert diagnostic["mapping_decision"] is None


def test_reviewed_summary_label_excludes_only_structural_non_movement_row(
    tmp_path: Path,
) -> None:
    # Arrange
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    output_dir = tmp_path / "out"
    _save_csv(
        bank_path,
        [
            ["Date", "Description", "Debit", "Credit", "Reference"],
            ["2026-08-05", "Customer receipt", "", "80.00", "TX100"],
            ["", "Total", "-20.00", "80.00", ""],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Description", "Amount", "Reference"],
            ["2026-08-05", "Customer receipt", "80.00", "TX100"],
        ],
    )
    recipe_path = _prepare_reviewed_summary_recipe(
        core,
        bank_path,
        journal_path,
        tmp_path / "recipe",
        labels=["total"],
    )

    # Act
    result = core.run_reconciliation(
        bank_path,
        journal_path,
        output_dir,
        recipe_path,
        tolerance="0",
        date_window_days=0,
    )

    # Assert
    audit = json.loads(
        (output_dir / "reconciliation_audit.json").read_text(encoding="utf-8")
    )
    source_qualifications = json.loads(
        (output_dir / "source_qualifications.json").read_text(encoding="utf-8")
    )
    bank_diagnostic = audit["diagnostics"]["bank"][0]
    assert result.matches.height == 1
    assert bank_diagnostic["candidate_row_count"] == 1
    assert bank_diagnostic["row_count"] == 1
    assert bank_diagnostic["row_disposition_counts"] == {
        "emitted": 1,
        "excluded_reviewed_summary": 1,
    }
    assert bank_diagnostic["non_movement_summary_labels"] == ["total"]
    assert {
        (qualification["adapter_id"], qualification["adapter_version"])
        for qualification in source_qualifications["qualifications"]
    } == {
        ("journal_bank.tabular.v6", "6"),
        ("journal_bank.tabular.v7", "7"),
    }


def test_reviewed_summary_label_does_not_override_date_or_stable_reference(
    tmp_path: Path,
) -> None:
    # Arrange
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    _save_csv(
        bank_path,
        [
            ["Date", "Description", "Amount", "Reference"],
            ["2026-08-05", "Total", "80.00", ""],
            ["", "Total", "20.00", "TOTAL-001"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Description", "Amount", "Reference"],
            ["2026-08-05", "Entry", "80.00", "TX100"],
        ],
    )
    recipe_path = _prepare_reviewed_summary_recipe(
        core,
        bank_path,
        journal_path,
        tmp_path / "recipe",
        labels=["total"],
    )

    # Act
    inspection = core.inspect_inputs(
        bank_path,
        journal_path,
        tmp_path / "reinspection",
        recipe_path,
    )

    # Assert
    diagnostic = inspection.bank["files"][0]
    assert inspection.bank["row_count"] == 2
    assert diagnostic["qualification_status"] == "qualified"
    assert diagnostic["row_disposition_counts"] == {
        "emitted": 1,
        "emitted_reference_only": 1,
    }
    assert "excluded_reviewed_summary" not in diagnostic["row_disposition_counts"]


def test_summary_label_mutation_requires_a_new_mapping_receipt(
    tmp_path: Path,
) -> None:
    # Arrange
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    _save_csv(
        bank_path,
        [
            ["Date", "Description", "Amount", "Reference"],
            ["2026-08-05", "Entry", "80.00", "TX100"],
            ["", "Total", "80.00", ""],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Description", "Amount", "Reference"],
            ["2026-08-05", "Entry", "80.00", "TX100"],
        ],
    )
    recipe_path = _prepare_reviewed_summary_recipe(
        core,
        bank_path,
        journal_path,
        tmp_path / "recipe",
        labels=["total"],
    )
    recipe = json.loads(recipe_path.read_text(encoding="utf-8"))
    recipe["bank"]["files"][bank_path.name]["non_movement_summary_labels"] = ["balance"]
    recipe_path.write_text(
        json.dumps(recipe, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    # Act
    inspection = core.inspect_inputs(
        bank_path,
        journal_path,
        tmp_path / "reinspection",
        recipe_path,
    )

    # Assert
    diagnostic = inspection.bank["files"][0]
    assert inspection.bank["row_count"] == 0
    assert diagnostic["qualification_status"] == "needs_review"
    assert diagnostic["failure_kind"] == "mapping_review_required"
    assert any(
        "Mapping receipt content does not match the current recipe." in limitation
        for limitation in diagnostic["limitations"]
    )
    assert diagnostic["mapping_decision"] is None


@pytest.mark.parametrize(
    ("mutation", "expected_error"),
    [
        (
            "date_convention",
            "Mapping receipt content does not match the current recipe.",
        ),
        ("stale_v5", "decision adapter binding is stale"),
    ],
)
def test_date_convention_mutation_or_stale_v5_receipt_emits_zero_rows(
    tmp_path: Path,
    mutation: str,
    expected_error: str,
) -> None:
    # Arrange
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    for source_path in (bank_path, journal_path):
        _save_csv(
            source_path,
            [
                ["Date", "Amount", "Reference"],
                ["05/08/2026", "80.00", "TX100"],
            ],
        )
    recipe_path = _prepare_reviewed_date_recipe(
        core,
        bank_path,
        journal_path,
        tmp_path / "recipe",
        date_convention="day_first",
    )
    recipe = json.loads(recipe_path.read_text(encoding="utf-8"))
    bank_recipe = recipe["bank"]["files"][bank_path.name]
    current = bank_recipe["mapping_decision"]
    if mutation == "date_convention":
        bank_recipe["date_convention"] = "month_first"
    else:
        bank_recipe["mapping_decision"] = core.build_reviewed_decision_receipt(
            decision_id=current["decision_id"],
            decision_type=current["decision_type"],
            status="reviewed",
            reviewer_ref=current["reviewer_ref"],
            reviewed_on=current["reviewed_on"],
            adapter_id="journal_bank.tabular.v5",
            adapter_version="5",
            source_artifact_refs=current["source_artifact_refs"],
            content=current["content"],
        )
    recipe_path.write_text(
        json.dumps(recipe, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    # Act
    inspected = core.inspect_inputs(
        bank_path,
        journal_path,
        tmp_path / "reinspection",
        recipe_path,
    )

    # Assert
    diagnostic = inspected.bank["files"][0]
    assert inspected.bank["row_count"] == 0
    assert diagnostic["qualification_status"] == "needs_review"
    assert diagnostic["failure_kind"] == "mapping_review_required"
    assert any(expected_error in value for value in diagnostic["limitations"])
    assert diagnostic["mapping_decision"] is None


def test_invalid_populated_date_with_stable_reference_fails_source(
    tmp_path: Path,
) -> None:
    # Arrange
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    _save_csv(
        bank_path,
        [
            ["Date", "Amount", "Reference"],
            ["31/02/2026", "80.00", "TX100"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference"],
            ["2026-02-28", "80.00", "TX100"],
        ],
    )

    # Act
    inspection = core.inspect_inputs(
        bank_path,
        journal_path,
        tmp_path / "inspection",
    )

    # Assert
    diagnostic = inspection.bank["files"][0]
    assert inspection.bank["row_count"] == 0
    assert diagnostic["qualification_status"] == "unsupported_source_layout"
    assert diagnostic["failure_kind"] == "candidate_row_contract_failed"
    assert diagnostic["row_disposition_counts"] == {"invalid_date_value": 1}
    assert diagnostic["date_interpretation_evidence"]["source_rows"]["invalid"] == [2]


@pytest.mark.parametrize(
    "source_date",
    [
        "posted 2026-05-08",
        "2026-05-08 posted",
        "posted\n2026-05-08",
        "2026-05-08\nposted",
        "2026-05-08 05/08/2026",
    ],
)
def test_date_parser_rejects_text_surrounding_an_embedded_date_token(
    source_date: str,
) -> None:
    # Arrange
    core = load_core()

    # Act
    result = core._date_parse_result(
        source_date,
        date_convention="month_first",
        allow_excel_serial=False,
    )

    # Assert
    assert result == ("invalid", None)


@pytest.mark.parametrize(
    ("invalid_side", "invalid_date"),
    [
        ("bank", "posted 2026-05-08"),
        ("journal", "2026-05-08 posted"),
    ],
)
def test_embedded_date_token_with_stable_reference_fails_complete_csv_source(
    tmp_path: Path,
    invalid_side: str,
    invalid_date: str,
) -> None:
    # Arrange
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    source_dates = {
        "bank": invalid_date if invalid_side == "bank" else "2026-05-08",
        "journal": invalid_date if invalid_side == "journal" else "2026-05-08",
    }
    _save_csv(
        bank_path,
        [
            ["Date", "Amount", "Reference"],
            [source_dates["bank"], "80.00", "TX100"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference"],
            [source_dates["journal"], "80.00", "TX100"],
        ],
    )
    output_dir = tmp_path / "out"

    # Act
    with pytest.raises(core.ReconciliationBlockedError) as exc_info:
        core.run_reconciliation(bank_path, journal_path, output_dir)

    # Assert
    audit = json.loads(
        (output_dir / "reconciliation_audit.json").read_text(encoding="utf-8")
    )
    diagnostic = audit["diagnostics"][invalid_side][0]
    assert exc_info.value.code == "unsupported_source_layout"
    assert diagnostic["qualification_status"] == "unsupported_source_layout"
    assert diagnostic["failure_kind"] == "candidate_row_contract_failed"
    assert diagnostic["row_count"] == 0
    assert diagnostic["row_disposition_counts"] == {"invalid_date_value": 1}
    assert diagnostic["date_interpretation_evidence"]["source_rows"]["invalid"] == [2]
    assert (output_dir / "normalized_bank.csv").is_file()
    assert (output_dir / "normalized_journal.csv").is_file()
    assert (output_dir / "reconciliation_matches.csv").is_file()


def test_mixed_valid_and_invalid_dates_withhold_the_complete_source(
    tmp_path: Path,
) -> None:
    # Arrange
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    _save_csv(
        bank_path,
        [
            ["Date", "Amount", "Reference"],
            ["2026-05-08", "80.00", "TX100"],
            ["2026-02-30", "20.00", "TX101"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference"],
            ["2026-05-08", "80.00", "TX100"],
        ],
    )

    # Act
    inspection = core.inspect_inputs(
        bank_path,
        journal_path,
        tmp_path / "inspection",
    )

    # Assert
    diagnostic = inspection.bank["files"][0]
    assert inspection.bank["row_count"] == 0
    assert diagnostic["qualification_status"] == "unsupported_source_layout"
    assert diagnostic["row_disposition_counts"] == {
        "emitted": 1,
        "invalid_date_value": 1,
    }
    assert diagnostic["invalid_candidate_rows"] == [3]


def test_noncanonical_direction_requires_reviewed_source_vocabulary(
    tmp_path: Path,
) -> None:
    # Arrange
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    inspection_dir = tmp_path / "inspection"
    _save_csv(
        bank_path,
        [
            ["Date", "Amount", "Reference"],
            ["2026-05-08", "-100.00", "TX100"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference", "Direction"],
            ["2026-05-08", "-100.00", "TX100", "Debit"],
        ],
    )

    # Act
    inspection = core.inspect_inputs(bank_path, journal_path, inspection_dir)

    # Assert
    diagnostic = inspection.journal["files"][0]
    assert inspection.journal["row_count"] == 0
    assert diagnostic["qualification_status"] == "needs_review"
    assert diagnostic["observed_direction_values"] == ["debit"]
    assert diagnostic["direction_value_mapping"] == {}
    assert diagnostic["missing_required_mapping"] == [
        "direction_value_mapping must exactly cover every observed "
        "non-canonical direction label"
    ]


def test_reviewed_direction_vocabulary_normalizes_and_matches_same_sign(
    tmp_path: Path,
) -> None:
    # Arrange
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    recipe_dir = tmp_path / "recipe"
    output_dir = tmp_path / "out"
    _save_csv(
        bank_path,
        [
            ["Date", "Amount", "Reference"],
            ["2026-05-08", "-100.00", "TX100"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference", "Direction"],
            ["2026-05-08", "-100.00", "TX100", "Debit"],
        ],
    )
    recipe_path = _prepare_reviewed_direction_recipe(
        core,
        bank_path,
        journal_path,
        recipe_dir,
        {"DEBIT": "negative"},
    )

    # Act
    result = core.run_reconciliation(
        bank_path,
        journal_path,
        output_dir,
        recipe_path,
        tolerance="0",
        date_window_days=0,
    )

    # Assert
    normalized_journal = _read_csv_dicts(output_dir / "normalized_journal.csv")
    reviewed_decisions = json.loads(
        (output_dir / "reviewed_decisions.json").read_text(encoding="utf-8")
    )
    mapping_decision = next(
        decision
        for decision in reviewed_decisions["decisions"]
        if decision["decision_type"] == "journal_bank_mapping"
    )
    assert result.matches.height == 1
    assert normalized_journal[0]["direction"] == "negative"
    assert mapping_decision["adapter_id"] == "journal_bank.tabular.v6"
    assert mapping_decision["adapter_version"] == "6"
    assert mapping_decision["content"]["direction_value_mapping"] == {
        "debit": "negative"
    }


@pytest.mark.parametrize(
    (
        "direction_policy",
        "bank_amount",
        "journal_amount",
        "expected_matches",
    ),
    [
        ("absolute_amount", "100.00", "-100.00", 1),
        ("same_sign", "-100.00", "-100.00", 1),
        ("same_sign", "100.00", "-100.00", 0),
        ("opposite_sign", "100.00", "-100.00", 1),
        ("opposite_sign", "100.00", "100.00", 0),
        ("opposite_sign", "0", "0", 0),
    ],
)
def test_relationship_direction_policy_has_exact_eligibility_semantics(
    tmp_path: Path,
    direction_policy: str,
    bank_amount: str,
    journal_amount: str,
    expected_matches: int,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    _save_csv(
        bank_path,
        [
            ["Date", "Amount", "Reference"],
            ["2026-05-08", bank_amount, "TX100"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference"],
            ["2026-05-08", journal_amount, "TX100"],
        ],
    )
    recipe_path = _prepare_reviewed_recipe(
        core,
        bank_path,
        journal_path,
        tmp_path / "recipe",
        tolerance="0",
        date_window_days=0,
        policy_updates={"direction_policy": direction_policy},
    )

    result = core.run_reconciliation(
        bank_path,
        journal_path,
        tmp_path / "run",
        recipe_path,
        tolerance="0",
        date_window_days=0,
    )

    assert result.matches.height == expected_matches


def test_unsupported_relationship_direction_policy_has_exact_block_outcome(
    tmp_path: Path,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    rows = [
        ["Date", "Amount", "Reference"],
        ["2026-05-08", "100.00", "TX100"],
    ]
    _save_csv(bank_path, rows)
    _save_csv(journal_path, rows)
    recipe_path = _prepare_reviewed_recipe(
        core,
        bank_path,
        journal_path,
        tmp_path / "recipe",
        tolerance="0",
        date_window_days=0,
    )
    recipe = json.loads(recipe_path.read_text(encoding="utf-8"))
    recipe["relationship"]["policy"]["direction_policy"] = "opposite"
    recipe_path.write_text(
        json.dumps(recipe, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    output_dir = tmp_path / "run"

    with pytest.raises(core.ReconciliationBlockedError) as exc_info:
        core.run_reconciliation(
            bank_path,
            journal_path,
            output_dir,
            recipe_path,
            tolerance="0",
            date_window_days=0,
        )

    audit = json.loads((output_dir / "reconciliation_audit.json").read_text())
    gates = json.loads((output_dir / "assurance_gates.json").read_text())
    assert exc_info.value.code == "relationship_review_required"
    assert audit["block_code"] == "relationship_review_required"
    assert gates["gates"]["reconciliation"]["status"] == "blocked"
    assert gates["gates"]["reporting"]["status"] == "blocked"


def test_reviewed_direction_vocabulary_rejects_signed_amount_conflict(
    tmp_path: Path,
) -> None:
    # Arrange
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    recipe_dir = tmp_path / "recipe"
    output_dir = tmp_path / "out"
    _save_csv(
        bank_path,
        [
            ["Date", "Amount", "Reference"],
            ["2026-05-08", "-100.00", "TX100"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference", "Direction"],
            ["2026-05-08", "-100.00", "TX100", "Credit"],
        ],
    )
    recipe_path = _prepare_reviewed_direction_recipe(
        core,
        bank_path,
        journal_path,
        recipe_dir,
        {"credit": "positive"},
    )

    # Act
    with pytest.raises(core.ReconciliationBlockedError) as exc_info:
        core.run_reconciliation(
            bank_path,
            journal_path,
            output_dir,
            recipe_path,
            tolerance="0",
            date_window_days=0,
        )

    # Assert
    audit = json.loads((output_dir / "reconciliation_audit.json").read_text())
    diagnostic = audit["diagnostics"]["journal"][0]
    assert exc_info.value.code == "unsupported_source_layout"
    assert diagnostic["qualification_status"] == "unsupported_source_layout"
    assert diagnostic["row_disposition_counts"] == {"direction_amount_mismatch": 1}
    assert diagnostic["row_count"] == 0


def test_direction_vocabulary_change_invalidates_mapping_receipt(
    tmp_path: Path,
) -> None:
    # Arrange
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    recipe_dir = tmp_path / "recipe"
    output_dir = tmp_path / "out"
    _save_csv(
        bank_path,
        [
            ["Date", "Amount", "Reference"],
            ["2026-05-08", "-100.00", "TX100"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference", "Direction"],
            ["2026-05-08", "-100.00", "TX100", "Debit"],
        ],
    )
    recipe_path = _prepare_reviewed_direction_recipe(
        core,
        bank_path,
        journal_path,
        recipe_dir,
        {"debit": "negative"},
    )
    recipe = json.loads(recipe_path.read_text(encoding="utf-8"))
    recipe["journal"]["files"][journal_path.name]["direction_value_mapping"] = {
        "debit": "positive"
    }
    recipe_path.write_text(
        json.dumps(recipe, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    # Act
    with pytest.raises(core.ReconciliationBlockedError) as exc_info:
        core.run_reconciliation(
            bank_path,
            journal_path,
            output_dir,
            recipe_path,
            tolerance="0",
            date_window_days=0,
        )

    # Assert
    audit = json.loads((output_dir / "reconciliation_audit.json").read_text())
    diagnostic = audit["diagnostics"]["journal"][0]
    assert exc_info.value.code == "mapping_review_required"
    assert diagnostic["qualification_status"] == "needs_review"
    assert "content does not match" in diagnostic["limitations"][0]


def test_invalid_compact_iso_date_blocks_without_explicit_reference(
    tmp_path: Path,
) -> None:
    # Arrange
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    output_dir = tmp_path / "out"
    _save_csv(
        bank_path,
        [
            ["Date", "Amount"],
            ["20260231", "80.00"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount"],
            ["2026-02-28", "80.00"],
        ],
    )

    # Act
    with pytest.raises(core.ReconciliationBlockedError) as exc_info:
        core.run_reconciliation(bank_path, journal_path, output_dir)

    # Assert
    audit = json.loads((output_dir / "reconciliation_audit.json").read_text())
    assert exc_info.value.code == "unsupported_source_layout"
    assert audit["diagnostics"]["bank"][0]["row_disposition_counts"] == {
        "invalid_date_value": 1
    }
    assert (output_dir / "relationship_residuals.csv").is_file()
    assert (output_dir / "review_payload.json").is_file()
    assert (output_dir / "final_artifacts.json").is_file()
    assert not (output_dir / "material_value_ledger.json").exists()


def test_reference_stage_ignores_reference_like_description_words(
    tmp_path: Path,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    output_dir = tmp_path / "out"
    _save_csv(
        bank_path,
        [
            ["Date", "Description", "Amount"],
            ["2025-03-10", "Payment INV100", "80.00"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Description", "Amount"],
            ["2025-03-10", "Invoice INV100", "80.00"],
            ["2025-03-10", "Different supplier", "80.00"],
        ],
    )
    recipe_path = _prepare_reviewed_recipe(
        core, bank_path, journal_path, tmp_path / "recipe"
    )

    result = core.run_reconciliation(bank_path, journal_path, output_dir, recipe_path)

    assert result.matches.height == 0
    assert result.unmatched_bank.height == 1
    assert result.unmatched_journal.height == 2


@pytest.mark.parametrize(
    ("journal_amount", "tolerance", "expected_matches"),
    [
        ("99.99", "0.01", 1),
        ("100.01", "0.01", 1),
        ("99.98", "0.01", 0),
        ("100.02", "0.01", 0),
        ("100.01", "0.009", 0),
    ],
)
def test_exact_decimal_tolerance_honors_cent_boundary(
    tmp_path: Path,
    journal_amount: str,
    tolerance: str,
    expected_matches: int,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    output_dir = tmp_path / "out"
    _save_csv(
        bank_path,
        [
            ["Date", "Amount"],
            ["2025-03-10", "100.00"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount"],
            ["2025-03-10", journal_amount],
        ],
    )
    recipe_path = _prepare_reviewed_recipe(
        core,
        bank_path,
        journal_path,
        tmp_path / "recipe",
        tolerance=tolerance,
    )

    result = core.run_reconciliation(
        bank_path,
        journal_path,
        output_dir,
        recipe_path,
        tolerance=tolerance,
    )

    assert result.matches.height == expected_matches
    assert result.audit["tolerance"] == tolerance
    if expected_matches:
        assert result.matches.to_dicts()[0]["amount_delta"] == "0.01"
        gates = json.loads((output_dir / "assurance_gates.json").read_text())
        assert result.audit["relationship_balanced"] is False
        assert gates["gates"]["reconciliation"]["status"] == "withheld"


def test_indexed_matching_preserves_sparse_population_matches(
    tmp_path: Path,
) -> None:
    # Arrange
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    expected_match_count = 32
    decoy_count = 512
    bank_rows = [
        [
            "2025-03-10",
            f"{10_000 + row_number}.00",
            f"MATCH-{10_000 + row_number}",
            f"Bank target {row_number}",
        ]
        for row_number in range(expected_match_count)
    ]
    target_rows = [
        [
            "2025-03-10",
            f"{10_000 + row_number}.00",
            f"MATCH-{10_000 + row_number}",
            f"Journal target {row_number}",
        ]
        for row_number in range(expected_match_count)
    ]
    decoy_rows = [
        [
            "2025-03-10",
            f"{20_000 + row_number}.00",
            f"DECOY-{20_000 + row_number}",
            f"Journal decoy {row_number}",
        ]
        for row_number in range(decoy_count)
    ]
    headers = ["Date", "Amount", "Reference", "Description"]
    _save_csv(bank_path, [headers, *bank_rows])
    _save_csv(journal_path, [headers, *decoy_rows, *target_rows])
    recipe_path = _prepare_reviewed_recipe(
        core,
        bank_path,
        journal_path,
        tmp_path / "recipe",
        tolerance="0",
        date_window_days=0,
    )

    # Act
    result = core.run_reconciliation(
        bank_path,
        journal_path,
        tmp_path / "run",
        recipe_path,
        tolerance="0",
        date_window_days=0,
    )

    # Assert
    match_rows = result.matches.to_dicts()
    assert len(match_rows) == expected_match_count
    assert {row["stage"] for row in match_rows} == {"reference"}
    assert {row["bank_description"] for row in match_rows} == {
        f"Bank target {row_number}" for row_number in range(expected_match_count)
    }
    assert {row["journal_description"] for row in match_rows} == {
        f"Journal target {row_number}" for row_number in range(expected_match_count)
    }
    assert result.unmatched_bank.is_empty()
    assert result.unmatched_journal.height == decoy_count


@pytest.mark.parametrize(
    "journal_amount",
    [
        "1.2345678901234567890123456775",
        "1.2345678901234567890123456785",
    ],
)
def test_matching_preserves_ultra_precise_inclusive_boundaries(
    tmp_path: Path,
    journal_amount: str,
) -> None:
    # Arrange
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    recipe_dir = tmp_path / "recipe"
    _save_csv(
        bank_path,
        [
            ["Date", "Amount"],
            ["2025-03-10", "1.234567890123456789012345678"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount"],
            ["2025-03-10", journal_amount],
        ],
    )
    core.inspect_inputs(bank_path, journal_path, recipe_dir)
    recipe_path = recipe_dir / "suggested_recipe.json"
    recipe = json.loads(recipe_path.read_text(encoding="utf-8"))
    receipts = json.loads(
        (recipe_dir / "input_receipts.json").read_text(encoding="utf-8")
    )["receipts"]
    for side, source_path in (("bank", bank_path), ("journal", journal_path)):
        _attach_current_mapping_receipt(
            core,
            recipe,
            receipts,
            side=side,
            source_path=source_path,
            decision_id=f"decision.mapping.{side}.precision-boundary",
        )
    recipe_path.write_text(
        json.dumps(recipe, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    recipe_path = _seal_relationship_recipe(
        core,
        recipe_path,
        recipe_dir / "input_receipts.json",
        tolerance="0.0000000000000000000000000005",
        date_window_days=0,
    )

    # Act
    result = core.run_reconciliation(
        bank_path,
        journal_path,
        tmp_path / "run",
        recipe_path,
        tolerance="0.0000000000000000000000000005",
        date_window_days=0,
    )

    # Assert
    assert result.matches.height == 1
    assert result.unmatched_bank.is_empty()
    assert result.unmatched_journal.is_empty()


@pytest.mark.parametrize(
    ("reviewed_tolerance", "expected"),
    [
        ("0.01", "0.01"),
        (Decimal("1.20"), "1.2"),
        (2, "2"),
    ],
)
def test_relationship_receipt_accepts_exact_tolerance_types(
    reviewed_tolerance: object,
    expected: str,
) -> None:
    core = load_core()

    receipt = core.build_relationship_review_receipt(
        decision_id="decision.relationship.exact-tolerance",
        reviewer_ref="reviewer.test",
        reviewed_on="2026-07-25",
        source_artifact_refs=["source.bank.1", "source.journal.1"],
        policy=_relationship_policy(reviewed_tolerance),
    )

    assert receipt["content"]["policy"]["amount_tolerance"] == expected


@pytest.mark.parametrize("invalid_tolerance", [1.0, "1,00", "1.00", True])
def test_relationship_receipt_rejects_inexact_or_noncanonical_tolerance(
    invalid_tolerance: object,
) -> None:
    core = load_core()

    with pytest.raises(core.MoneyValidationError):
        core.build_relationship_review_receipt(
            decision_id="decision.relationship.invalid-tolerance",
            reviewer_ref="reviewer.test",
            reviewed_on="2026-07-25",
            source_artifact_refs=["source.bank.1", "source.journal.1"],
            policy=_relationship_policy(invalid_tolerance),
        )


@pytest.mark.parametrize("execution_tolerance", [Decimal("0"), 0])
def test_reconciliation_api_preserves_decimal_and_integer_tolerance(
    tmp_path: Path,
    execution_tolerance: object,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    output_dir = tmp_path / "run"
    rows = [
        ["Date", "Amount", "Reference"],
        ["2026-05-08", "80.00", "TX100"],
    ]
    _save_csv(bank_path, rows)
    _save_csv(journal_path, rows)
    recipe_path = _prepare_reviewed_recipe(
        core,
        bank_path,
        journal_path,
        tmp_path / "recipe",
        tolerance="0",
        date_window_days=0,
    )

    result = core.run_reconciliation(
        bank_path,
        journal_path,
        output_dir,
        recipe_path,
        tolerance=execution_tolerance,
        date_window_days=0,
    )

    assert result.matches.height == 1
    assert result.audit["tolerance"] == "0"


def test_material_value_ledger_closes_every_match_and_residual_native_address(
    tmp_path: Path,
) -> None:
    core, output_dir = _prepare_two_match_run(tmp_path)

    ledger = core.validate_material_value_ledger(output_dir)
    residual_rows = _read_csv_dicts(output_dir / "relationship_residuals.csv")
    second_match_amount = next(
        entry
        for entry in ledger["entries"]
        if entry["evidence_id"] == "matches.2.bank_amount"
    )
    second_bank_residual = next(
        entry
        for entry in ledger["entries"]
        if entry["evidence_id"] == "relationship_residuals.2.residual"
    )

    assert ledger["entry_count"] == (
        2 * len(core.MATCH_MATERIAL_FIELDS) + 4 * len(core.RESIDUAL_MATERIAL_FIELDS)
    )
    assert ledger["datasets"][0]["row_count"] == 2
    assert ledger["datasets"][1]["row_count"] == 4
    assert residual_rows[0]["record_ref"] == "bank.1"
    assert residual_rows[1]["record_ref"] == "bank.2"
    assert residual_rows[2]["record_ref"] == "journal.1"
    assert residual_rows[3]["record_ref"] == "journal.2"
    assert residual_rows[1]["record_amount"] == "25.5"
    assert residual_rows[1]["allocated_amount"] == "25.5"
    assert residual_rows[1]["residual"] == "0"
    assert second_match_amount["prepared_locator"] == ("row=2;column=bank_amount")
    assert second_match_amount["prepared"]["locator"] == (
        second_match_amount["prepared_locator"]
    )
    assert second_match_amount["outputs"] == [
        {
            "artifact_ref": "output.reconciliation_matches_csv",
            "locator": "row=3;column=bank_amount",
            "value": "25.5",
        },
        {
            "artifact_ref": "output.workbook_xlsx",
            "locator": "matches!H3",
            "value": "25.5",
        },
    ]
    assert second_bank_residual["row_identity"] == {
        "side": "bank",
        "record_ref": "bank.2",
        "transaction_id": residual_rows[1]["transaction_id"],
    }
    assert second_bank_residual["outputs"][1]["locator"] == (
        "relationship_residuals!F3"
    )


def test_material_output_replay_streams_read_only_workbook_rows(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    # Arrange
    from openpyxl.worksheet._read_only import ReadOnlyWorksheet

    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    output_dir = tmp_path / "run"
    rows = [
        ["Date", "Amount", "Reference"],
        ["2026-05-08", "80.00", "TX100"],
        ["2026-05-09", "25.50", "TX200"],
    ]
    _save_csv(bank_path, rows)
    _save_csv(journal_path, rows)
    recipe_path = _prepare_reviewed_recipe(
        core,
        bank_path,
        journal_path,
        tmp_path / "recipe",
        tolerance="0",
        date_window_days=0,
    )

    def reject_random_cell_access(
        self: ReadOnlyWorksheet,
        row: int,
        column: int,
    ) -> None:
        raise AssertionError(
            f"read-only workbook replay used random access at {row}:{column}"
        )

    monkeypatch.setattr(ReadOnlyWorksheet, "cell", reject_random_cell_access)

    # Act
    result = core.run_reconciliation(
        bank_path,
        journal_path,
        output_dir,
        recipe_path,
        tolerance="0",
        date_window_days=0,
    )

    # Assert
    replay = core.validate_material_value_ledger(output_dir)
    assert result.matches.height == 2
    assert replay["entry_count"] > 0


def test_relationship_ledger_emits_shared_v1_amount_field(
    tmp_path: Path,
) -> None:
    core, output_dir = _prepare_two_match_run(tmp_path)

    ledger = json.loads(
        (output_dir / "relationship_ledger.json").read_text(encoding="utf-8")
    )
    validated = core.validate_allocation_ledger(ledger)

    assert validated == ledger
    assert ledger["allocations"]
    assert all(
        set(allocation)
        == {
            "allocation_id",
            "source_record_ref",
            "target_record_ref",
            "amount",
            "currency",
            "unit",
            "evidence_refs",
        }
        for allocation in ledger["allocations"]
    )
    assert [allocation["amount"] for allocation in ledger["allocations"]] == [
        "80",
        "25.5",
    ]


def test_material_value_replay_rejects_second_match_csv_row_mutation(
    tmp_path: Path,
) -> None:
    core, output_dir = _prepare_two_match_run(tmp_path)
    matches_path = output_dir / "reconciliation_matches.csv"
    rows = _read_csv_dicts(matches_path)
    rows[1]["bank_amount"] = "999"
    _save_csv(matches_path, [list(rows[0]), *[list(row.values()) for row in rows]])

    with pytest.raises(
        ValueError,
        match="match material values do not replay exactly",
    ):
        core.validate_material_value_ledger(output_dir)


def test_material_value_replay_rejects_second_residual_workbook_cell_mutation(
    tmp_path: Path,
) -> None:
    core, output_dir = _prepare_two_match_run(tmp_path)
    workbook_path = output_dir / "journal_bank_reconciliation.xlsx"
    workbook = openpyxl.load_workbook(workbook_path)
    workbook["relationship_residuals"]["F3"] = "999"
    workbook.save(workbook_path)
    workbook.close()

    with pytest.raises(
        ValueError,
        match=r"relationship_residuals!F3",
    ):
        core.validate_material_value_ledger(output_dir)


def test_material_value_replay_rejects_duplicate_ooxml_member(
    tmp_path: Path,
) -> None:
    core, output_dir = _prepare_two_match_run(tmp_path)
    workbook_path = output_dir / "journal_bank_reconciliation.xlsx"
    member_name = "xl/worksheets/sheet1.xml"
    with ZipFile(workbook_path, "a") as archive:
        member_bytes = archive.read(member_name)
        with pytest.warns(UserWarning, match="Duplicate name"):
            archive.writestr(member_name, member_bytes)

    with pytest.raises(
        ValueError,
        match="OOXML workbook contains duplicate member names",
    ):
        core.validate_material_value_ledger(output_dir)


def test_separator_ambiguous_money_blocks_source_qualification(
    tmp_path: Path,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    output_dir = tmp_path / "out"
    _save_csv(
        bank_path,
        [
            ["Date", "Amount"],
            ["2025-03-10", "1.000"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount"],
            ["2025-03-10", "1000"],
        ],
    )

    with pytest.raises(core.ReconciliationBlockedError) as exc_info:
        core.run_reconciliation(bank_path, journal_path, output_dir)

    qualifications = json.loads((output_dir / "source_qualifications.json").read_text())
    audit = json.loads((output_dir / "reconciliation_audit.json").read_text())
    assert exc_info.value.code == "mapping_review_required"
    assert qualifications["status"] == "needs_review"
    bank_qualification = next(
        entry
        for entry in qualifications["qualifications"]
        if entry["qualification_id"] == "qualification.bank.1"
    )
    bank_diagnostic = audit["diagnostics"]["bank"][0]
    assert bank_qualification["adapter_id"] == "journal_bank.tabular.v6"
    assert bank_qualification["adapter_version"] == "6"
    assert bank_qualification["status"] == "needs_review"
    assert bank_qualification["emitted_row_count"] == 0
    assert bank_diagnostic["failure_kind"] == "mapping_review_required"
    assert bank_diagnostic["ambiguous_numeric_rows"] == [2]
    assert (
        "ambiguous numeric separators require reviewed decimal_separator "
        "and thousands_separator values" in bank_diagnostic["missing_required_mapping"]
    )


def test_reviewed_thousands_separator_cannot_be_reinterpreted_as_decimal(
    tmp_path: Path,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    recipe_dir = tmp_path / "recipe"
    output_dir = tmp_path / "out"
    _save_csv(
        bank_path,
        [
            ["Date", "Amount", "Reference"],
            ["2025-03-10", "1,23", "TX100"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference"],
            ["2025-03-10", "123", "TX100"],
        ],
    )
    inspection = core.inspect_inputs(bank_path, journal_path, recipe_dir)
    recipe = inspection.suggested_recipe
    receipts = json.loads(
        (recipe_dir / "input_receipts.json").read_text(encoding="utf-8")
    )["receipts"]
    bank_recipe = recipe["bank"]["files"][bank_path.name]
    bank_recipe["decimal_separator"] = None
    bank_recipe["thousands_separator"] = ","
    _attach_current_mapping_receipt(
        core,
        recipe,
        receipts,
        side="bank",
        source_path=bank_path,
        decision_id="decision.mapping.bank.numeric_locale",
    )
    recipe_path = recipe_dir / "reviewed_recipe.json"
    recipe_path.write_text(
        json.dumps(recipe, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    _seal_relationship_recipe(
        core,
        recipe_path,
        recipe_dir / "input_receipts.json",
    )

    with pytest.raises(core.ReconciliationBlockedError) as exc_info:
        core.run_reconciliation(bank_path, journal_path, output_dir, recipe_path)

    qualifications = json.loads(
        (output_dir / "source_qualifications.json").read_text(encoding="utf-8")
    )
    bank_qualification = next(
        entry
        for entry in qualifications["qualifications"]
        if entry["qualification_id"] == "qualification.bank.1"
    )
    assert exc_info.value.code == "unsupported_source_layout"
    assert bank_qualification["status"] == "unsupported_source_layout"
    assert bank_qualification["emitted_row_count"] == 0


@pytest.mark.parametrize("delimiter", [";", "\t", "|"])
def test_supported_nondefault_csv_delimiter_is_a_zero_row_review_proposal(
    tmp_path: Path,
    delimiter: str,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    inspection_dir = tmp_path / "inspection"
    _save_delimited_csv(
        bank_path,
        [
            ["Date", "Amount", "Reference"],
            ["2026-05-08", "80.00", "TX100"],
            ["2026-05-09", "-25.00", "TX101"],
        ],
        delimiter=delimiter,
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference"],
            ["2026-05-08", "80.00", "TX100"],
        ],
    )

    inspection = core.inspect_inputs(bank_path, journal_path, inspection_dir)

    diagnostic = inspection.bank["files"][0]
    proposed = inspection.suggested_recipe["bank"]["files"]["bank.csv"]
    assert inspection.bank["row_count"] == 0
    assert diagnostic["qualification_status"] == "needs_review"
    assert diagnostic["failure_kind"] == "mapping_review_required"
    assert diagnostic["csv_field_delimiter"] == delimiter
    assert diagnostic["csv_field_delimiter_origin"] == "profiled_proposal"
    assert diagnostic["raw_columns"] == ["Date", "Amount", "Reference"]
    assert diagnostic["candidate_row_count"] == 2
    assert diagnostic["preview"] == []
    assert proposed["csv_field_delimiter"] == delimiter
    assert proposed["potential_monetary_columns"] == ["Amount"]
    assert proposed["excluded_monetary_columns"] == []
    assert proposed["mapping_decision"] is None


def test_complete_reviewed_semicolon_v6_mapping_qualifies_exact_rows(
    tmp_path: Path,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    first_dir = tmp_path / "first"
    reviewed_dir = tmp_path / "reviewed"
    _save_delimited_csv(
        bank_path,
        [
            ["Date", "Amount", "Reference"],
            ["2026-05-08", "80.00", "TX100"],
            ["2026-05-09", "-25.00", "TX101"],
        ],
        delimiter=";",
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference"],
            ["2026-05-08", "80.00", "TX100"],
        ],
    )
    first = core.inspect_inputs(bank_path, journal_path, first_dir)
    recipe = first.suggested_recipe
    receipts = json.loads(
        (first_dir / "input_receipts.json").read_text(encoding="utf-8")
    )["receipts"]
    receipt = _attach_current_mapping_receipt(
        core,
        recipe,
        receipts,
        side="bank",
        source_path=bank_path,
        decision_id="decision.mapping.bank.semicolon",
    )
    recipe_path = tmp_path / "reviewed-recipe.json"
    recipe_path.write_text(json.dumps(recipe, indent=2) + "\n", encoding="utf-8")

    reviewed = core.inspect_inputs(
        bank_path,
        journal_path,
        reviewed_dir,
        recipe_path,
    )

    diagnostic = reviewed.bank["files"][0]
    assert reviewed.bank["row_count"] == 2
    assert diagnostic["qualification_status"] == "qualified"
    assert diagnostic["csv_field_delimiter"] == ";"
    assert diagnostic["reviewed_mapping_ref"] == receipt["decision_id"]
    assert [row["source_row"] for row in diagnostic["preview"]] == [2, 3]
    assert [row["amount_signed"] for row in diagnostic["preview"]] == ["80", "-25"]
    assert receipt["adapter_id"] == "journal_bank.tabular.v6"
    assert receipt["adapter_version"] == "6"
    assert receipt["content"]["csv_field_delimiter"] == ";"
    assert receipt["content"]["potential_monetary_columns"] == ["Amount"]
    assert receipt["content"]["excluded_monetary_columns"] == []


@pytest.mark.parametrize(
    ("mutation", "expected_error"),
    [
        ("delimiter", "Mapping receipt content does not match the current recipe."),
        ("stale_v5", "decision adapter binding is stale"),
    ],
)
def test_delimiter_mutation_or_stale_v5_mapping_receipt_emits_zero_rows(
    tmp_path: Path,
    mutation: str,
    expected_error: str,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    first_dir = tmp_path / "first"
    _save_delimited_csv(
        bank_path,
        [
            ["Date", "Amount", "Reference"],
            ["2026-05-08", "80.00", "TX100"],
        ],
        delimiter=";",
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference"],
            ["2026-05-08", "80.00", "TX100"],
        ],
    )
    first = core.inspect_inputs(bank_path, journal_path, first_dir)
    recipe = first.suggested_recipe
    receipts = json.loads(
        (first_dir / "input_receipts.json").read_text(encoding="utf-8")
    )["receipts"]
    current = _attach_current_mapping_receipt(
        core,
        recipe,
        receipts,
        side="bank",
        source_path=bank_path,
        decision_id="decision.mapping.bank.semicolon",
    )
    bank_recipe = recipe["bank"]["files"]["bank.csv"]
    if mutation == "delimiter":
        bank_recipe["csv_field_delimiter"] = "|"
    else:
        bank_recipe["mapping_decision"] = core.build_reviewed_decision_receipt(
            decision_id=current["decision_id"],
            decision_type=current["decision_type"],
            status="reviewed",
            reviewer_ref=current["reviewer_ref"],
            reviewed_on=current["reviewed_on"],
            adapter_id="journal_bank.tabular.v5",
            adapter_version="5",
            source_artifact_refs=current["source_artifact_refs"],
            content=current["content"],
        )
    recipe_path = tmp_path / f"{mutation}-recipe.json"
    recipe_path.write_text(json.dumps(recipe, indent=2) + "\n", encoding="utf-8")

    inspected = core.inspect_inputs(
        bank_path,
        journal_path,
        tmp_path / f"{mutation}-inspection",
        recipe_path,
    )

    diagnostic = inspected.bank["files"][0]
    assert inspected.bank["row_count"] == 0
    assert diagnostic["qualification_status"] == "needs_review"
    expected_candidate_count = 0 if mutation == "delimiter" else 1
    assert diagnostic["candidate_row_count"] == expected_candidate_count
    assert diagnostic["preview"] == []
    assert any(expected_error in value for value in diagnostic["limitations"])
    assert diagnostic["mapping_decision"] is None


@pytest.mark.parametrize(
    ("layout", "expected_status", "expected_failure"),
    [
        (
            "ambiguous",
            "needs_review",
            "ambiguous_csv_field_delimiter",
        ),
        (
            "unsupported",
            "unsupported_source_layout",
            "unsupported_csv_field_delimiter",
        ),
    ],
)
def test_ambiguous_or_unsupported_csv_delimiter_has_no_plausible_rows(
    tmp_path: Path,
    layout: str,
    expected_status: str,
    expected_failure: str,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    recipe_path: Path | None = None
    if layout == "ambiguous":
        bank_path.write_text(
            "Date,Posting;Amount,Reference\n" "2026-05-08,memo;80.00,TX100\n",
            encoding="utf-8",
        )
    else:
        bank_path.write_text(
            "Date:Amount:Reference\n2026-05-08:80.00:TX100\n",
            encoding="utf-8",
        )
        recipe_path = tmp_path / "unsupported-recipe.json"
        recipe_path.write_text(
            json.dumps(
                {
                    "bank": {
                        "files": {
                            "bank.csv": {
                                "csv_field_delimiter": ":",
                            }
                        }
                    }
                }
            )
            + "\n",
            encoding="utf-8",
        )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference"],
            ["2026-05-08", "80.00", "TX100"],
        ],
    )

    inspection = core.inspect_inputs(
        bank_path,
        journal_path,
        tmp_path / f"{layout}-inspection",
        recipe_path,
    )

    diagnostic = inspection.bank["files"][0]
    assert inspection.bank["row_count"] == 0
    assert diagnostic["qualification_status"] == expected_status
    assert diagnostic["failure_kind"] == expected_failure
    assert diagnostic["candidate_row_count"] == 0
    assert diagnostic["row_count"] == 0
    assert diagnostic["preview"] == []
    assert diagnostic["row_dispositions"] == []


@pytest.mark.parametrize(
    ("bank_text", "expected_status", "expected_failure"),
    [
        (
            "date,amount;description,reference\n"
            "2026-07-24,-10.00;Ambiguous delimiter,CTRL84197\n",
            "needs_review",
            "ambiguous_csv_field_delimiter",
        ),
        (
            "date:amount:description:reference\n"
            "2026-07-24:-10.00:Unsupported delimiter:CTRL84197\n",
            "unsupported_source_layout",
            "unsupported_csv_field_delimiter",
        ),
    ],
    ids=["ambiguous-profile", "unsupported-profile"],
)
def test_unreviewed_explicit_comma_preserves_delimiter_profile_taxonomy(
    tmp_path: Path,
    bank_text: str,
    expected_status: str,
    expected_failure: str,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    bank_path.write_text(bank_text, encoding="utf-8")
    _save_csv(
        journal_path,
        [
            ["date", "amount", "reference"],
            ["2026-07-24", "10.00", "CTRL84197"],
        ],
    )
    recipe_path = tmp_path / "recipe.json"
    recipe_path.write_text(
        json.dumps(
            {
                "bank": {
                    "files": {
                        bank_path.name: {
                            "csv_field_delimiter": ",",
                            "mapping_decision": None,
                        }
                    }
                }
            },
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )

    inspection = core.inspect_inputs(
        bank_path,
        journal_path,
        tmp_path / "inspection",
        recipe_path,
    )

    diagnostic = inspection.bank["files"][0]
    assert inspection.bank["row_count"] == 0
    assert diagnostic["qualification_status"] == expected_status
    assert diagnostic["failure_kind"] == expected_failure
    assert diagnostic["requested_csv_field_delimiter"] == ","
    assert diagnostic["candidate_row_count"] == 0


def test_explicit_null_csv_delimiter_preserves_ambiguous_review_boundary(
    tmp_path: Path,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    bank_path.write_text(
        "Date,Posting;Amount,Reference\n" "2026-05-08,memo;80.00,TX100\n",
        encoding="utf-8",
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference"],
            ["2026-05-08", "80.00", "TX100"],
        ],
    )
    recipe_path = tmp_path / "null-delimiter-recipe.json"
    recipe_path.write_text(
        json.dumps(
            {
                "bank": {
                    "files": {
                        "bank.csv": {
                            "csv_field_delimiter": None,
                        }
                    }
                }
            }
        )
        + "\n",
        encoding="utf-8",
    )

    inspection = core.inspect_inputs(
        bank_path,
        journal_path,
        tmp_path / "inspection",
        recipe_path,
    )

    diagnostic = inspection.bank["files"][0]
    assert inspection.bank["row_count"] == 0
    assert diagnostic["qualification_status"] == "needs_review"
    assert diagnostic["failure_kind"] == "ambiguous_csv_field_delimiter"
    assert diagnostic["candidate_row_count"] == 0
    assert diagnostic["csv_field_delimiter"] is None
    assert diagnostic["requested_csv_field_delimiter"] is None
    assert diagnostic["row_dispositions"] == []


def test_unique_comma_exact_contract_remains_qualified_without_receipt(
    tmp_path: Path,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    for source_path in (bank_path, journal_path):
        _save_csv(
            source_path,
            [
                ["Date", "Amount", "Reference"],
                ["2026-05-08", "80.00", "TX100"],
            ],
        )

    inspection = core.inspect_inputs(
        bank_path,
        journal_path,
        tmp_path / "inspection",
    )

    for side in (inspection.bank, inspection.journal):
        diagnostic = side["files"][0]
        assert side["row_count"] == 1
        assert diagnostic["qualification_status"] == "qualified"
        assert diagnostic["csv_field_delimiter"] == ","
        assert diagnostic["csv_field_delimiter_origin"] == "default"
        assert diagnostic["mapping_origin"] == "bounded_exact_headers"
        assert diagnostic["mapping_decision"] is None


def test_canonical_snake_case_mapping_runs_amount_date_cascade_and_native_closure(
    tmp_path: Path,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    inspection_dir = tmp_path / "inspection"
    output_dir = tmp_path / "run"
    headers = [
        "date",
        "amount",
        "description",
        "beneficiary",
        "reference",
        "movement_number",
        "account",
        "currency",
        "unit",
        "entity_ref",
        "party_ref",
        "direction",
    ]
    _save_csv(
        bank_path,
        [
            headers,
            [
                "2026-07-10",
                "-100.00",
                "Bank leg early",
                "Asterix Services",
                "",
                "",
                "MAIN",
                "EUR",
                "OPS",
                "ENT-A",
                "PARTY-X",
                "negative",
            ],
            [
                "2026-07-11",
                "-100.00",
                "Bank leg later",
                "Asterix Services",
                "",
                "",
                "MAIN",
                "EUR",
                "OPS",
                "ENT-A",
                "PARTY-X",
                "negative",
            ],
        ],
    )
    _save_csv(
        journal_path,
        [
            headers,
            [
                "2026-07-11",
                "100.00",
                "Journal leg middle",
                "Asterix Services",
                "",
                "",
                "MAIN",
                "EUR",
                "OPS",
                "ENT-A",
                "PARTY-X",
                "positive",
            ],
            [
                "2026-07-12",
                "100.00",
                "Journal leg latest",
                "Asterix Services",
                "",
                "",
                "MAIN",
                "EUR",
                "OPS",
                "ENT-A",
                "PARTY-X",
                "positive",
            ],
        ],
    )

    inspection = core.inspect_inputs(
        bank_path,
        journal_path,
        inspection_dir,
    )
    recipe_path = _seal_relationship_recipe(
        core,
        inspection_dir / "suggested_recipe.json",
        inspection_dir / "input_receipts.json",
        tolerance="0",
        date_window_days=1,
    )
    result = core.run_reconciliation(
        bank_path,
        journal_path,
        output_dir,
        recipe_path,
        tolerance="0",
        date_window_days=1,
    )

    expected_mapping = {
        **{field: field for field in headers},
        "credit": None,
        "debit": None,
    }
    for side in (inspection.bank, inspection.journal):
        diagnostic = side["files"][0]
        assert side["row_count"] == 2
        assert diagnostic["qualification_status"] == "qualified"
        assert diagnostic["mapping"] == expected_mapping
        assert diagnostic["mapping_origin"] == "bounded_exact_headers"
        assert diagnostic["mapping_decision"] is None
    assert result.matches.height == 2
    assert set(result.matches["stage"]) == {
        "amount_date_unique",
        "amount_date_single",
    }
    assert all(
        (output_dir / relative_path).is_file()
        for relative_path in core.INITIAL_RUN_OUTPUT_FILES
    )
    assert core.validate_material_value_ledger(output_dir)["entry_count"] > 0
    envelope = json.loads(
        (output_dir / "assurance_envelope.json").read_text(encoding="utf-8")
    )
    assert (
        len(
            [
                receipt
                for receipt in envelope["artifact_receipts"]
                if receipt["role"] == "implementation"
            ]
        )
        == 24
    )


@pytest.mark.parametrize(
    ("headers", "mapping"),
    [
        (
            [
                "date",
                "amount",
                "movement_no",
                "entity_reference",
                "party_reference",
            ],
            {
                "date": "date",
                "amount": "amount",
                "movement_number": "movement_no",
                "entity_ref": "entity_reference",
                "party_ref": "party_reference",
            },
        ),
        (
            ["date", "transaction date", "amount"],
            {
                "date": "date",
                "amount": "amount",
            },
        ),
    ],
    ids=["near-match-optional-fields", "ambiguous-exact-date-fields"],
)
def test_automatic_exact_mapping_abstains_on_near_match_or_ambiguous_layout(
    tmp_path: Path,
    headers: list[str],
    mapping: dict[str, str],
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    bank_row = {
        "date": "2026-07-10",
        "transaction date": "2026-07-10",
        "amount": "100.00",
        "movement_no": "M-100",
        "entity_reference": "ENT-A",
        "party_reference": "PARTY-X",
    }
    _save_csv(bank_path, [headers, [bank_row[header] for header in headers]])
    _save_csv(
        journal_path,
        [
            ["date", "amount", "reference"],
            ["2026-07-10", "100.00", "TX100"],
        ],
    )
    recipe_path = tmp_path / "recipe.json"
    recipe_path.write_text(
        json.dumps(
            {
                "bank": {
                    "files": {
                        bank_path.name: {
                            "header_rows": [1],
                            "mapping": mapping,
                            "potential_monetary_columns": ["amount"],
                            "excluded_monetary_columns": [],
                            "csv_field_delimiter": ",",
                            "date_convention": None,
                            "decimal_separator": None,
                            "thousands_separator": None,
                            "direction_value_mapping": {},
                            "mapping_decision": None,
                        }
                    }
                }
            },
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )

    inspection = core.inspect_inputs(
        bank_path,
        journal_path,
        tmp_path / "inspection",
        recipe_path,
    )

    diagnostic = inspection.bank["files"][0]
    assert inspection.bank["row_count"] == 0
    assert diagnostic["qualification_status"] == "needs_review"
    assert diagnostic["failure_kind"] == "mapping_review_required"
    assert diagnostic["mapping_origin"] == "reviewed_recipe"
    assert diagnostic["mapping_decision"] is None
    assert (
        "A reviewed mapping receipt is required for this layout."
        in diagnostic["limitations"]
    )


@pytest.mark.parametrize(
    ("separator_field", "separator_value", "reviewed_value"),
    [
        ("decimal_separator", ".", "."),
        ("thousands_separator", ",", ","),
        ("decimal_separator", "", None),
        ("thousands_separator", "", None),
    ],
)
def test_explicit_numeric_separator_requires_current_mapping_receipt(
    tmp_path: Path,
    separator_field: str,
    separator_value: str,
    reviewed_value: str | None,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    for source_path in (bank_path, journal_path):
        _save_csv(
            source_path,
            [
                ["Date", "Amount", "Reference"],
                ["2026-05-08", "80.00", "TX100"],
            ],
        )
    first_dir = tmp_path / "first"
    first = core.inspect_inputs(bank_path, journal_path, first_dir)
    recipe = first.suggested_recipe
    receipts = json.loads(
        (first_dir / "input_receipts.json").read_text(encoding="utf-8")
    )["receipts"]
    bank_recipe = recipe["bank"]["files"][bank_path.name]
    bank_recipe[separator_field] = separator_value
    recipe_path = tmp_path / "unreviewed-recipe.json"
    recipe_path.write_text(json.dumps(recipe, indent=2) + "\n", encoding="utf-8")

    unreviewed = core.inspect_inputs(
        bank_path,
        journal_path,
        tmp_path / "unreviewed",
        recipe_path,
    )

    unreviewed_diagnostic = unreviewed.bank["files"][0]
    assert unreviewed.bank["row_count"] == 0
    assert unreviewed_diagnostic["qualification_status"] == "needs_review"
    assert unreviewed_diagnostic["failure_kind"] == "mapping_review_required"
    assert unreviewed_diagnostic["mapping_decision"] is None
    assert (
        unreviewed_diagnostic["mapping_review_content"][separator_field]
        == reviewed_value
    )
    assert (
        "reviewed mapping receipt is required"
        in " ".join(unreviewed_diagnostic["limitations"]).lower()
    )

    _attach_current_mapping_receipt(
        core,
        recipe,
        receipts,
        side="bank",
        source_path=bank_path,
        decision_id=f"decision.mapping.bank.{separator_field}",
    )
    recipe_path.write_text(json.dumps(recipe, indent=2) + "\n", encoding="utf-8")
    reviewed = core.inspect_inputs(
        bank_path,
        journal_path,
        tmp_path / "reviewed",
        recipe_path,
    )

    reviewed_diagnostic = reviewed.bank["files"][0]
    assert reviewed.bank["row_count"] == 1
    assert reviewed_diagnostic["qualification_status"] == "qualified"
    assert reviewed_diagnostic["mapping_origin"] == "reviewed_recipe"
    assert (
        reviewed_diagnostic["mapping_decision"]["content"][separator_field]
        == reviewed_value
    )


def test_exact_automatic_contract_rejects_stale_monetary_declaration(
    tmp_path: Path,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    for source_path in (bank_path, journal_path):
        _save_csv(
            source_path,
            [
                ["Date", "Amount", "Reference"],
                ["2026-05-08", "80.00", "TX100"],
            ],
        )
    first = core.inspect_inputs(
        bank_path,
        journal_path,
        tmp_path / "first",
    )
    recipe = first.suggested_recipe
    recipe["bank"]["files"][bank_path.name]["potential_monetary_columns"] = []
    recipe_path = tmp_path / "stale-recipe.json"
    recipe_path.write_text(json.dumps(recipe, indent=2) + "\n", encoding="utf-8")

    inspected = core.inspect_inputs(
        bank_path,
        journal_path,
        tmp_path / "stale",
        recipe_path,
    )

    diagnostic = inspected.bank["files"][0]
    assert inspected.bank["row_count"] == 0
    assert diagnostic["qualification_status"] == "needs_review"
    assert diagnostic["failure_kind"] == "mapping_review_required"
    assert diagnostic["potential_monetary_columns"] == ["Amount"]
    assert diagnostic["mapping_decision"] is None
    assert (
        "potential_monetary_columns does not match current source evidence"
        in diagnostic["limitations"]
    )


def test_exact_automatic_contract_does_not_ignore_supplied_stale_receipt(
    tmp_path: Path,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    for source_path in (bank_path, journal_path):
        _save_csv(
            source_path,
            [
                ["Date", "Amount", "Reference"],
                ["2026-05-08", "80.00", "TX100"],
            ],
        )
    first_dir = tmp_path / "first"
    first = core.inspect_inputs(bank_path, journal_path, first_dir)
    recipe = first.suggested_recipe
    receipts = json.loads(
        (first_dir / "input_receipts.json").read_text(encoding="utf-8")
    )["receipts"]
    current = _attach_current_mapping_receipt(
        core,
        recipe,
        receipts,
        side="bank",
        source_path=bank_path,
        decision_id="decision.mapping.bank.exact",
    )
    recipe["bank"]["files"][bank_path.name]["mapping_decision"] = (
        core.build_reviewed_decision_receipt(
            decision_id=current["decision_id"],
            decision_type=current["decision_type"],
            status="reviewed",
            reviewer_ref=current["reviewer_ref"],
            reviewed_on=current["reviewed_on"],
            adapter_id="journal_bank.tabular.v4",
            adapter_version="4",
            source_artifact_refs=current["source_artifact_refs"],
            content=current["content"],
        )
    )
    recipe_path = tmp_path / "stale-receipt-recipe.json"
    recipe_path.write_text(json.dumps(recipe, indent=2) + "\n", encoding="utf-8")

    inspected = core.inspect_inputs(
        bank_path,
        journal_path,
        tmp_path / "stale-receipt",
        recipe_path,
    )

    diagnostic = inspected.bank["files"][0]
    assert inspected.bank["row_count"] == 0
    assert diagnostic["qualification_status"] == "needs_review"
    assert diagnostic["failure_kind"] == "mapping_review_required"
    assert diagnostic["mapping_decision"] is None
    assert any(
        "decision adapter binding is stale" in limitation
        for limitation in diagnostic["limitations"]
    )


@pytest.mark.parametrize(
    ("field", "invalid_value", "expected_limitation"),
    [
        (
            "header_rows",
            "1",
            "header_rows must be a non-empty list of unique positive integers",
        ),
        (
            "mapping",
            ["Date", "Amount", "Reference"],
            (
                "mapping must be an object containing only supported fields "
                "with string or null column names"
            ),
        ),
    ],
)
def test_exact_automatic_contract_rejects_invalid_explicit_recipe_shape(
    tmp_path: Path,
    field: str,
    invalid_value: object,
    expected_limitation: str,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    for source_path in (bank_path, journal_path):
        _save_csv(
            source_path,
            [
                ["Date", "Amount", "Reference"],
                ["2026-05-08", "80.00", "TX100"],
            ],
        )
    first = core.inspect_inputs(
        bank_path,
        journal_path,
        tmp_path / "first",
    )
    recipe = first.suggested_recipe
    recipe["bank"]["files"][bank_path.name][field] = invalid_value
    recipe_path = tmp_path / f"invalid-{field}.json"
    recipe_path.write_text(json.dumps(recipe, indent=2) + "\n", encoding="utf-8")

    inspected = core.inspect_inputs(
        bank_path,
        journal_path,
        tmp_path / f"invalid-{field}",
        recipe_path,
    )

    diagnostic = inspected.bank["files"][0]
    assert inspected.bank["row_count"] == 0
    assert diagnostic["qualification_status"] == "needs_review"
    assert diagnostic["failure_kind"] == "mapping_review_required"
    assert diagnostic["mapping_decision"] is None
    assert expected_limitation in diagnostic["limitations"]


@pytest.mark.parametrize(
    "invalid_bank_recipe",
    [
        [],
        {"files": []},
        {"files": {"bank.csv": []}},
        {"files": {"bank.csv": None}},
    ],
    ids=[
        "side-list",
        "files-list",
        "file-list",
        "file-null",
    ],
)
def test_exact_automatic_contract_rejects_malformed_recipe_container(
    tmp_path: Path,
    invalid_bank_recipe: object,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    for source_path in (bank_path, journal_path):
        _save_csv(
            source_path,
            [
                ["Date", "Amount", "Reference"],
                ["2026-05-08", "80.00", "TX100"],
            ],
        )
    recipe_path = tmp_path / "invalid-container.json"
    recipe_path.write_text(
        json.dumps({"bank": invalid_bank_recipe}, indent=2) + "\n",
        encoding="utf-8",
    )

    inspected = core.inspect_inputs(
        bank_path,
        journal_path,
        tmp_path / "invalid-container",
        recipe_path,
    )

    diagnostic = inspected.bank["files"][0]
    assert inspected.bank["row_count"] == 0
    assert diagnostic["qualification_status"] == "needs_review"
    assert diagnostic["failure_kind"] == "mapping_review_required"
    assert diagnostic["mapping_decision"] is None
    assert (
        "recipe container must use object-valued side, files, and file entries"
        in diagnostic["limitations"]
    )


@pytest.mark.parametrize(
    "record_terminator",
    [b"\n", b"\r\n", b"\r"],
    ids=["lf", "crlf", "cr"],
)
def test_csv_record_terminators_are_equivalent_transport_syntax(
    tmp_path: Path,
    record_terminator: bytes,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    rows = [
        b"Date,Amount,Reference",
        b"2026-05-08,80.00,TX100",
    ]
    bank_path.write_bytes(record_terminator.join(rows) + record_terminator)
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference"],
            ["2026-05-08", "80.00", "TX100"],
        ],
    )

    inspection = core.inspect_inputs(
        bank_path,
        journal_path,
        tmp_path / "inspection",
    )

    diagnostic = inspection.bank["files"][0]
    assert inspection.bank["row_count"] == 1
    assert diagnostic["qualification_status"] == "qualified"
    assert diagnostic["candidate_row_count"] == 1
    assert diagnostic["csv_field_delimiter"] == ","
    assert diagnostic["mapping_origin"] == "bounded_exact_headers"


@pytest.mark.parametrize(
    "boundary_terminator",
    [b"\r\n", b"\r"],
    ids=["split-crlf", "trailing-cr"],
)
def test_csv_record_terminator_at_transport_chunk_boundary_is_preserved(
    tmp_path: Path,
    boundary_terminator: bytes,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    row_prefix = (
        b"Date,Amount,Reference,Description"
        + boundary_terminator
        + b"2026-05-08,80.00,TX100,"
    )
    filler = b"x" * (core.CSV_TRANSPORT_CHUNK_BYTES - len(row_prefix) - 1)
    bank_path.write_bytes(row_prefix + filler + boundary_terminator)
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference"],
            ["2026-05-08", "80.00", "TX100"],
        ],
    )

    inspection = core.inspect_inputs(
        bank_path,
        journal_path,
        tmp_path / "inspection",
    )

    diagnostic = inspection.bank["files"][0]
    assert inspection.bank["row_count"] == 1
    assert diagnostic["qualification_status"] == "qualified"
    assert diagnostic["candidate_row_count"] == 1
    assert diagnostic["csv_field_delimiter"] == ","


@pytest.mark.parametrize(
    "malformed_row",
    [
        ["2026-05-08", "80.00", "TX100", "unexpected-field"],
        ["2026-05-08", "ragged-only"],
    ],
    ids=["extra-field", "short-row"],
)
def test_malformed_csv_row_beyond_delimiter_profile_fails_closed(
    tmp_path: Path,
    malformed_row: list[str],
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    output_dir = tmp_path / "run"
    valid_data_row = ["2026-05-08", "80.00", "TX100"]
    _save_csv(
        bank_path,
        [["Date", "Amount", "Reference"]] + [valid_data_row] * 100,
    )
    with bank_path.open("a", encoding="utf-8", newline="") as handle:
        csv.writer(handle).writerow(malformed_row)
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference"],
            valid_data_row,
        ],
    )

    with pytest.raises(core.ReconciliationBlockedError) as exc_info:
        core.run_reconciliation(bank_path, journal_path, output_dir)

    audit = json.loads((output_dir / "reconciliation_audit.json").read_text())
    qualifications = json.loads((output_dir / "source_qualifications.json").read_text())
    bank_qualification = next(
        item
        for item in qualifications["qualifications"]
        if item["qualification_id"] == "qualification.bank.1"
    )
    assert exc_info.value.code == "parser_failure"
    assert audit["bank_row_count"] == 0
    assert audit["diagnostics"]["bank"][0]["failure_kind"] == "parser_failure"
    assert audit["diagnostics"]["bank"][0]["candidate_row_count"] == 0
    assert bank_qualification["status"] == "unsupported_source_layout"
    assert bank_qualification["emitted_row_count"] == 0
    bank_outcome = next(
        item for item in qualifications["source_outcomes"] if item["side"] == "bank"
    )
    assert bank_outcome == {
        "side": "bank",
        "qualification_status": "unsupported_source_layout",
        "failure_kind": "parser_failure",
        "emitted_row_count": 0,
    }
    assert (output_dir / "artifact_receipts.json").is_file()
    assert (output_dir / "assurance_gates.json").is_file()


def test_semicolon_sample_csv_is_parsed_by_the_bounded_profile(
    tmp_path: Path,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    sample_path = tmp_path / "sample.csv"
    _save_csv(
        bank_path,
        [
            ["Date", "Amount", "Reference"],
            ["2026-05-08", "80.00", "TX100"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference", "Movement"],
            ["2026-05-08", "80.00", "TX100", "M100"],
        ],
    )
    _save_delimited_csv(
        sample_path,
        [
            ["Movement", "Selection"],
            ["M100", "reviewed"],
        ],
        delimiter=";",
    )

    inspection = core.inspect_inputs(
        bank_path,
        journal_path,
        tmp_path / "inspection",
        sample_path=sample_path,
    )

    inspection_json = json.loads(
        (tmp_path / "inspection" / "inspection.json").read_text(encoding="utf-8")
    )
    sample_diagnostic = inspection_json["sample"]["diagnostics"][0]
    assert inspection.sample == {"movement_count": 1, "status": "qualified"}
    assert inspection_json["sample"]["movements"] == ["M100"]
    assert sample_diagnostic["status"] == "qualified"
    assert sample_diagnostic["csv_field_delimiter"] == ";"
    assert sample_diagnostic["movement_count"] == 1


def test_profile_and_positional_mapping_remain_review_proposals(
    tmp_path: Path,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    inspection_dir = tmp_path / "inspection"
    _save_csv(
        bank_path,
        [
            ["Posting", "Value", "Memo"],
            ["2025-03-10", "80.00", "ABC123"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference"],
            ["2025-03-10", "80.00", "ABC123"],
        ],
    )

    inspection = core.inspect_inputs(bank_path, journal_path, inspection_dir)

    diagnostic = inspection.bank["files"][0]
    qualifications = json.loads(
        (inspection_dir / "source_qualifications.json").read_text()
    )
    assert inspection.bank["row_count"] == 0
    assert diagnostic["qualification_status"] == "needs_review"
    assert diagnostic["mapping_origin"] == "proposal"
    assert diagnostic["mapping"]["date"] == "Posting"
    assert diagnostic["mapping"]["amount"] is None
    assert qualifications["status"] == "needs_review"
    assert qualifications["qualifications"][0]["emitted_row_count"] == 0


def test_reviewed_mapping_is_source_bound_and_stale_after_source_change(
    tmp_path: Path,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    recipe_dir = tmp_path / "recipe"
    _save_csv(
        bank_path,
        [
            ["Posting", "Value", "Doc ID"],
            ["2025-03-10", "80.00", "ABC123"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference"],
            ["2025-03-10", "80.00", "ABC123"],
        ],
    )
    core.inspect_inputs(bank_path, journal_path, recipe_dir)
    recipe_path = recipe_dir / "suggested_recipe.json"
    recipe = json.loads(recipe_path.read_text())
    receipts = json.loads((recipe_dir / "input_receipts.json").read_text())["receipts"]
    bank_ref = next(
        receipt["artifact_id"]
        for receipt in receipts
        if receipt["artifact_id"].startswith("source.bank.")
    )
    mapping = {
        "date": "Posting",
        "amount": "Value",
        "reference": "Doc ID",
    }
    bank_recipe = recipe["bank"]["files"]["bank.csv"]
    bank_recipe["mapping"] = mapping
    bank_recipe["mapping_decision"] = core.build_mapping_review_receipt(
        decision_id="decision.mapping.bank",
        reviewer_ref="reviewer.test",
        reviewed_on="2026-07-24",
        source_artifact_ref=bank_ref,
        side="bank",
        source_file="bank.csv",
        header_rows=[1],
        mapping=mapping,
        potential_monetary_columns=bank_recipe["potential_monetary_columns"],
        excluded_monetary_columns=bank_recipe["excluded_monetary_columns"],
        csv_field_delimiter=bank_recipe["csv_field_delimiter"],
    )
    recipe_path.write_text(json.dumps(recipe, indent=2) + "\n", encoding="utf-8")
    _seal_relationship_recipe(
        core,
        recipe_path,
        recipe_dir / "input_receipts.json",
    )

    result = core.run_reconciliation(
        bank_path,
        journal_path,
        tmp_path / "first-run",
        recipe_path,
    )
    _save_csv(
        bank_path,
        [
            ["Posting", "Value", "Doc ID"],
            ["2025-03-10", "80.00", "ABC123"],
            [],
        ],
    )
    stale_output = tmp_path / "stale-run"
    with pytest.raises(core.ReconciliationBlockedError) as exc_info:
        core.run_reconciliation(
            bank_path,
            journal_path,
            stale_output,
            recipe_path,
        )

    stale_audit = json.loads((stale_output / "reconciliation_audit.json").read_text())
    assert result.matches.height == 1
    assert exc_info.value.code == "mapping_review_required"
    assert "invalid or stale" in stale_audit["diagnostics"]["bank"][0]["limitations"][0]


def test_lineage_preserves_physical_sheet_and_row_after_preamble_and_blanks(
    tmp_path: Path,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.xlsx"
    journal_path = tmp_path / "journal.xlsx"
    _save_workbook(
        bank_path,
        [
            ["Bank export"],
            ["Generated", "2025-03-31"],
            ["Date", "Amount", "Reference"],
            [],
            ["2025-03-10", 80, "ABC123"],
        ],
    )
    _save_workbook(
        journal_path,
        [
            ["Ledger export"],
            [],
            ["Date", "Amount", "Reference"],
            [],
            ["2025-03-10", 80, "ABC123"],
        ],
    )
    recipe_path = _prepare_reviewed_recipe(
        core, bank_path, journal_path, tmp_path / "recipe"
    )

    result = core.run_reconciliation(
        bank_path,
        journal_path,
        tmp_path / "run",
        recipe_path,
    )
    lineage = json.loads((tmp_path / "run" / "lineage.json").read_text())

    assert result.matches.height == 1
    assert result.matches.to_dicts()[0]["bank_transaction_id"].endswith(":Sheet:5")
    assert {entry["source_locator"] for entry in lineage["entries"]} == {
        "sheet:Sheet;row:5"
    }


@pytest.mark.parametrize(
    "reference",
    [
        "INV100",
        "FQ-2026-071-HO",
        "DOC-2025-9981",
        "Budget-2025-INV100",
        "INV-JAN25-100",
    ],
)
def test_missing_date_with_distinctive_identifier_is_reference_only(
    tmp_path: Path,
    reference: str,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    _save_csv(
        bank_path,
        [
            ["Date", "Amount", "Reference"],
            ["", "80.00", reference],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference"],
            ["", "80.00", reference],
        ],
    )
    recipe_path = _prepare_reviewed_recipe(
        core, bank_path, journal_path, tmp_path / "recipe"
    )

    result = core.run_reconciliation(
        bank_path,
        journal_path,
        tmp_path / "run",
        recipe_path,
    )

    assert result.matches.to_dicts()[0]["stage"] == "reference"
    assert result.audit["diagnostics"]["bank"][0]["row_disposition_counts"] == {
        "emitted_reference_only": 1
    }


def test_generic_period_fragment_does_not_join_distinctive_references(
    tmp_path: Path,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    _save_csv(
        bank_path,
        [
            ["Date", "Amount", "Reference"],
            ["", "80.00", "INV-JAN25-100"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference"],
            ["", "80.00", "DOC-JAN25-200"],
        ],
    )
    recipe_path = _prepare_reviewed_recipe(
        core,
        bank_path,
        journal_path,
        tmp_path / "recipe",
    )

    result = core.run_reconciliation(
        bank_path,
        journal_path,
        tmp_path / "run",
        recipe_path,
    )

    assert result.matches.height == 0
    assert result.unmatched_bank.height == 1
    assert result.unmatched_journal.height == 1


def test_relationship_policy_prevents_cross_currency_and_missing_perimeter(
    tmp_path: Path,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    _save_csv(
        bank_path,
        [
            ["Date", "Amount", "Reference", "Currency"],
            ["2025-03-10", "80.00", "ABC123", "EUR"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference", "Currency"],
            ["2025-03-10", "80.00", "ABC123", "USD"],
        ],
    )
    recipe_path = _prepare_reviewed_recipe(
        core, bank_path, journal_path, tmp_path / "recipe"
    )

    result = core.run_reconciliation(
        bank_path,
        journal_path,
        tmp_path / "mismatch",
        recipe_path,
    )

    assert result.matches.height == 0
    assert result.audit["status"] == "completed_with_unresolved_reconciliation"
    gates = json.loads((tmp_path / "mismatch" / "assurance_gates.json").read_text())
    assert gates["gates"]["reconciliation"]["status"] == "withheld"

    missing_perimeter_recipe = _prepare_reviewed_recipe(
        core,
        bank_path,
        journal_path,
        tmp_path / "missing-perimeter-recipe",
        policy_updates={"default_unit": None},
    )
    with pytest.raises(core.ReconciliationBlockedError) as exc_info:
        core.run_reconciliation(
            bank_path,
            journal_path,
            tmp_path / "missing-perimeter",
            missing_perimeter_recipe,
        )
    assert exc_info.value.code == "relationship_perimeter_incomplete"


def test_sample_failure_preserves_independent_source_gate(
    tmp_path: Path,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    sample_path = tmp_path / "sample.csv"
    output_dir = tmp_path / "run"
    _save_csv(
        bank_path,
        [
            ["Date", "Amount", "Reference"],
            ["2025-03-10", "80.00", "ABC123"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference", "Movement"],
            ["2025-03-10", "80.00", "ABC123", "M100"],
        ],
    )
    _save_csv(sample_path, [["Movement"], ["M999"]])
    recipe_path = _prepare_reviewed_recipe(
        core, bank_path, journal_path, tmp_path / "recipe"
    )

    with pytest.raises(core.ReconciliationBlockedError) as exc_info:
        core.run_reconciliation(
            bank_path,
            journal_path,
            output_dir,
            recipe_path,
            sample_path=sample_path,
        )

    gates = json.loads((output_dir / "assurance_gates.json").read_text())
    assert exc_info.value.code == "sample_not_found_in_journal"
    assert gates["gates"]["source"]["status"] == "passed"
    assert gates["gates"]["source"]["limitations"] == []
    assert gates["gates"]["preparation"]["status"] == "failed"
    assert gates["gates"]["reconciliation"]["status"] == "blocked"


def test_corrupt_workbook_writes_parser_failure_assurance_artifacts(
    tmp_path: Path,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.xlsx"
    journal_path = tmp_path / "journal.csv"
    output_dir = tmp_path / "run"
    bank_path.write_bytes(b"not-an-xlsx")
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference"],
            ["2025-03-10", "80.00", "ABC123"],
        ],
    )

    with pytest.raises(core.ReconciliationBlockedError) as exc_info:
        core.run_reconciliation(bank_path, journal_path, output_dir)

    audit = json.loads((output_dir / "reconciliation_audit.json").read_text())
    gates = json.loads((output_dir / "assurance_gates.json").read_text())
    assert exc_info.value.code == "parser_failure"
    assert audit["diagnostics"]["bank"][0]["failure_kind"] == "parser_failure"
    assert audit["bank_row_count"] == 0
    assert gates["gates"]["source"]["status"] == "failed"
    assert (output_dir / "artifact_receipts.json").is_file()
    assert (output_dir / "lineage.json").is_file()


def test_generic_explicit_reference_words_do_not_create_a_match(
    tmp_path: Path,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    _save_csv(
        bank_path,
        [
            ["Date", "Amount", "Reference"],
            ["2025-03-10", "80.00", "invoice"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference"],
            ["2025-03-10", "80.00", "invoice"],
            ["2025-03-10", "80.00", "payment"],
        ],
    )
    recipe_path = _prepare_reviewed_recipe(
        core, bank_path, journal_path, tmp_path / "recipe"
    )

    result = core.run_reconciliation(
        bank_path,
        journal_path,
        tmp_path / "run",
        recipe_path,
    )

    assert result.matches.height == 0
    assert result.unmatched_bank.height == 1
    assert result.unmatched_journal.height == 2


def test_run_reconciliation_sanitizes_illegal_excel_characters_before_workbook_export(
    tmp_path: Path,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    output_dir = tmp_path / "out"
    raw_bank_description = "Saldo iniziale al 31.03.2025 +133\x19 318, 47 EUR"
    excel_bank_description = "Saldo iniziale al 31.03.2025 +133 318, 47 EUR"
    _save_csv(
        bank_path,
        [
            ["Date", "Description", "Amount", "Reference"],
            ["2025-03-31", raw_bank_description, "47.00", "CTRL19"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Description", "Amount", "Reference"],
            ["2025-03-31", "Journal movement CTRL19", "47.00", "CTRL19"],
        ],
    )
    recipe_path = _prepare_reviewed_recipe(
        core, bank_path, journal_path, tmp_path / "recipe"
    )

    result = core.run_reconciliation(
        bank_path, journal_path, output_dir, recipe_path, language="en"
    )

    match = result.matches.to_dicts()[0]
    workbook = openpyxl.load_workbook(output_dir / "journal_bank_reconciliation.xlsx")
    assert match["bank_description"] == raw_bank_description
    assert workbook["matches"]["K2"].value == excel_bank_description
    assert workbook["normalized_bank"]["F2"].value == excel_bank_description
    assert "\x19" not in workbook["matches"]["K2"].value


def test_bank_pdf_non_movement_rows_are_excluded_with_multilingual_rules(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.pdf"
    journal_path = tmp_path / "journal.csv"
    output_dir = tmp_path / "out"
    bank_path.write_text("stub pdf content", encoding="utf-8")
    bank_pdf_text = "\n".join(
        [
            "Saldo iniziale al 31.03.2025 +133 318,47 EUR",
            "Totale accrediti 1.000,00 EUR",
            "Riassunto scalare interessi 12,34 EUR",
            "Condizioni economiche canone 5,00 EUR",
            "Opening balance at 31/03/2025 133,318.47 EUR",
            "Total fees 12.00 EUR",
            "Account conditions 5.00 EUR",
            "Solde initial au 31/03/2025 133 318,47 EUR",
            "Total des credits 1 000,00 EUR",
            "Conditions economiques 5,00 EUR",
            "Anfangssaldo zum 31.03.2025 133.318,47 EUR",
            "Summe der Gutschriften 1.000,00 EUR",
            "Kontokonditionen 5,00 EUR",
            "31/03/2025 Bonifico cliente ACME INV100 1.000,00 EUR",
            "01/04/2025 Commissione bonifico FEEIT 3,00 EUR",
            "02/04/2025 Bank transfer fee FEEEN 4.00 EUR",
            "03/04/2025 Frais de virement FEEFR 5,00 EUR",
            "04/04/2025 Ueberweisungsgebuehr FEEDE 6,00 EUR",
        ]
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Description", "Amount", "Reference"],
            ["2025-03-31", "Invoice INV100 ACME", "1000.00", "INV100"],
            ["2025-04-01", "Commissione bonifico FEEIT", "3.00", "FEEIT"],
            ["2025-04-02", "Bank transfer fee FEEEN", "4.00", "FEEEN"],
            ["2025-04-03", "Frais de virement FEEFR", "5.00", "FEEFR"],
            ["2025-04-04", "Ueberweisungsgebuehr FEEDE", "6.00", "FEEDE"],
        ],
    )

    def fake_extract_pdf_text(path: Path) -> str:
        return bank_pdf_text if path == bank_path else ""

    monkeypatch.setattr(core, "_extract_pdf_text", fake_extract_pdf_text)

    inspection = core.inspect_inputs(
        bank_path,
        journal_path,
        tmp_path / "inspection",
        language="en",
        document_language="auto",
    )
    with pytest.raises(core.ReconciliationBlockedError) as exc_info:
        core.run_reconciliation(bank_path, journal_path, output_dir, language="en")

    non_movement_rows = _read_csv_dicts(output_dir / "bank_pdf_non_movement_rows.csv")
    audit = json.loads((output_dir / "reconciliation_audit.json").read_text())
    qualifications = json.loads((output_dir / "source_qualifications.json").read_text())
    classifications = [row["classification"] for row in non_movement_rows]

    assert inspection.bank["row_count"] == 0
    assert (
        json.loads((tmp_path / "inspection" / "inspection.json").read_text())[
            "qualification_status"
        ]
        == "unsupported_source_layout"
    )
    assert exc_info.value.code == "unsupported_source_layout"
    assert audit["status"] == "blocked"
    assert audit["source_qualification_status"] == "unsupported_source_layout"
    assert audit["bank_row_count"] == 0
    assert audit["bank_pdf_non_movement_row_count"] == 13
    assert audit["bank_pdf_non_movement_classifications"] == {
        "balance": 4,
        "conditions": 4,
        "scalare": 1,
        "total": 4,
    }
    assert qualifications["status"] == "unsupported_source_layout"
    assert qualifications["qualifications"][0]["emitted_row_count"] == 0
    assert len(non_movement_rows) == 13
    assert {"balance", "conditions", "scalare", "total"} <= set(classifications)
    assert any("Saldo iniziale" in row["description"] for row in non_movement_rows)
    assert any("Opening balance" in row["description"] for row in non_movement_rows)
    assert any("Solde initial" in row["description"] for row in non_movement_rows)
    assert any("Anfangssaldo" in row["description"] for row in non_movement_rows)
    assert (output_dir / "normalized_bank.csv").is_file()
    assert (output_dir / "reconciliation_matches.csv").is_file()


def test_duplicate_basenames_keep_distinct_source_identity_and_lineage(
    tmp_path: Path,
) -> None:
    core = load_core()
    bank_root = tmp_path / "bank"
    journal_path = tmp_path / "journal.csv"
    (bank_root / "north").mkdir(parents=True)
    (bank_root / "south").mkdir(parents=True)
    for folder, reference, amount in (
        ("north", "NORTH100", "10.00"),
        ("south", "SOUTH200", "20.00"),
    ):
        _save_csv(
            bank_root / folder / "export.csv",
            [
                ["Date", "Amount", "Reference"],
                ["2025-03-10", amount, reference],
            ],
        )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference"],
            ["2025-03-10", "10.00", "NORTH100"],
            ["2025-03-10", "20.00", "SOUTH200"],
        ],
    )

    result = core.inspect_inputs(bank_root, journal_path, tmp_path / "inspection")
    receipts = json.loads(
        (tmp_path / "inspection" / "input_receipts.json").read_text()
    )["receipts"]
    lineage = json.loads((tmp_path / "inspection" / "lineage.json").read_text())
    bank_receipts = [
        receipt for receipt in receipts if receipt["root_id"] == "source_bank"
    ]
    bank_entries = [entry for entry in lineage["entries"] if entry["side"] == "bank"]

    assert result.bank["row_count"] == 2
    assert {receipt["path"] for receipt in bank_receipts} == {
        "north/export.csv",
        "south/export.csv",
    }
    assert len({entry["prepared_id"] for entry in bank_entries}) == 2
    assert len({entry["source_artifact_ref"] for entry in bank_entries}) == 2


def test_invalid_utf8_csv_blocks_with_replayable_parser_failure_artifacts(
    tmp_path: Path,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    output_dir = tmp_path / "run"
    bank_path.write_bytes(b"Date,Amount,Reference\n2025-03-10,80.00,ABC\xff123\n")
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference"],
            ["2025-03-10", "80.00", "ABC123"],
        ],
    )

    with pytest.raises(core.ReconciliationBlockedError) as exc_info:
        core.run_reconciliation(bank_path, journal_path, output_dir)

    audit = json.loads((output_dir / "reconciliation_audit.json").read_text())
    envelope = json.loads((output_dir / "assurance_envelope.json").read_text())
    assert exc_info.value.code == "parser_failure"
    assert audit["diagnostics"]["bank"][0]["failure_kind"] == "parser_failure"
    assert (output_dir / "artifact_receipts.json").is_file()
    assert (output_dir / "assurance_gates.json").is_file()
    core.validate_assurance_envelope(
        envelope,
        artifact_roots={
            "source_bank": bank_path.parent,
            "source_journal": journal_path.parent,
            "run": output_dir,
            "implementation": ROOT / "plugins" / "journal-bank-reconciliation",
            "shared_implementation": (
                ROOT / "plugins" / "_shared" / "vendor" / "modules" / "vera_assurance"
            ),
        },
    )


def test_initial_assurance_envelope_binds_exact_transitive_implementation_set(
    tmp_path: Path,
) -> None:
    core = load_core()
    output_dir = tmp_path / "run"
    output_dir, *_ = _prepare_closed_mcp_review_run(output_dir)
    envelope = json.loads(
        (output_dir / "assurance_envelope.json").read_text(encoding="utf-8")
    )
    implementation_receipts = [
        receipt
        for receipt in envelope["artifact_receipts"]
        if receipt["role"] == "implementation"
    ]

    validated = core.validate_exact_implementation_receipts(envelope)

    assert validated == implementation_receipts
    assert envelope["implementation_artifact_refs"] == [
        artifact_id for artifact_id, _, _ in core.IMPLEMENTATION_ARTIFACT_SPECS
    ]
    assert [
        (receipt["root_id"], receipt["path"]) for receipt in implementation_receipts
    ] == [
        (root_id, relative_path)
        for _, root_id, relative_path in core.IMPLEMENTATION_ARTIFACT_SPECS
    ]
    assert len(implementation_receipts) == 24


@pytest.mark.parametrize(
    "attack_kind",
    ["empty_directory", "regular", "symlink", "hardlink", "fifo"],
)
def test_python_preimport_rejects_every_unowned_implementation_entry(
    tmp_path: Path,
    attack_kind: str,
) -> None:
    output_dir = tmp_path / "run"
    output_dir, *_ = _prepare_closed_mcp_review_run(output_dir)
    before = _tree_snapshot(output_dir)
    copied_plugin, _ = _copy_journal_bank_implementation(tmp_path)
    rogue = copied_plugin / "scripts" / "rogue"
    external = tmp_path / f"{attack_kind}-external"
    if attack_kind == "empty_directory":
        rogue.mkdir()
    elif attack_kind == "regular":
        rogue.write_bytes(b"unreceipted implementation")
    elif attack_kind == "symlink":
        external.write_bytes(b"external")
        rogue.symlink_to(external)
    elif attack_kind == "hardlink":
        external.write_bytes(b"external")
        rogue.hardlink_to(external)
    else:
        os.mkfifo(rogue)

    completed = subprocess.run(
        [
            sys.executable,
            "-I",
            "-B",
            str(copied_plugin / "scripts" / "apply_review_edits.py"),
            "--output-dir",
            str(output_dir),
            "--preflight-only",
        ],
        capture_output=True,
        text=True,
        check=False,
        timeout=10,
    )

    assert completed.returncode != 0
    assert "implementation" in completed.stderr.lower()
    assert _tree_snapshot(output_dir) == before


def test_real_python_entry_rejects_timestamp_valid_unreceipted_bytecode(
    tmp_path: Path,
) -> None:
    copied_plugin, _ = _copy_journal_bank_implementation(tmp_path)
    target = copied_plugin / "scripts" / "journal_bank_core.py"
    source = target.read_bytes()
    source_stat = target.stat()
    marker = tmp_path / "malicious-pyc-executed.txt"
    malicious = (
        "from pathlib import Path as _AttackPath\n"
        f"_AttackPath({marker.as_posix()!r}).write_text("
        "'executed before validation\\n', encoding='utf-8')\n"
        f"exec(compile({source.decode('utf-8')!r}, {target.as_posix()!r}, "
        "'exec'), globals())\n"
    )
    code = compile(malicious, target.as_posix(), "exec")
    cache_prefix = sys.pycache_prefix
    try:
        sys.pycache_prefix = None
        cache_path = Path(importlib.util.cache_from_source(target.as_posix()))
    finally:
        sys.pycache_prefix = cache_prefix
    cache_path.parent.mkdir()
    cache_path.write_bytes(
        importlib._bootstrap_external._code_to_timestamp_pyc(
            code,
            int(source_stat.st_mtime),
            source_stat.st_size,
        )
    )

    completed = subprocess.run(
        [
            sys.executable,
            "-I",
            "-B",
            str(copied_plugin / "scripts" / "run_reconciliation.py"),
            "--help",
        ],
        cwd=copied_plugin,
        capture_output=True,
        text=True,
        check=False,
        timeout=30,
    )

    assert completed.returncode != 0
    assert "implementation" in completed.stderr.lower()
    assert not marker.exists()
    assert target.read_bytes() == source
    assert target.stat().st_size == source_stat.st_size
    assert target.stat().st_mtime_ns == source_stat.st_mtime_ns


@pytest.mark.parametrize("attack_kind", ["symlink", "hardlink", "fifo"])
def test_real_python_entry_rejects_unsafe_bootstrap_before_read(
    tmp_path: Path,
    attack_kind: str,
) -> None:
    copied_plugin, _ = _copy_journal_bank_implementation(tmp_path)
    bootstrap = copied_plugin / "scripts" / "implementation_bootstrap.py"
    original = bootstrap.read_bytes()
    bootstrap.unlink()
    external = tmp_path / f"bootstrap-{attack_kind}-external.py"
    if attack_kind == "symlink":
        external.write_bytes(original)
        bootstrap.symlink_to(external)
    elif attack_kind == "hardlink":
        external.write_bytes(original)
        bootstrap.hardlink_to(external)
    else:
        os.mkfifo(bootstrap)

    completed = subprocess.run(
        [
            sys.executable,
            "-I",
            "-B",
            str(copied_plugin / "scripts" / "run_reconciliation.py"),
            "--help",
        ],
        cwd=copied_plugin,
        capture_output=True,
        text=True,
        check=False,
        timeout=30,
    )

    assert completed.returncode != 0
    assert "bootstrap is not a real file" in completed.stderr


def test_mcp_rejects_unowned_implementation_path_before_stdio(
    tmp_path: Path,
) -> None:
    copied_plugin, _ = _copy_journal_bank_implementation(tmp_path)
    (copied_plugin / "scripts" / "__pycache__").mkdir()
    node = shutil.which("node")
    if node is None:
        pytest.skip("Node.js is required for the MCP implementation probe.")

    completed = subprocess.run(
        [node, str(copied_plugin / "mcp" / "server.cjs"), "--stdio"],
        input=json.dumps(
            {
                "jsonrpc": "2.0",
                "id": 1,
                "method": "tools/list",
                "params": {},
            }
        )
        + "\n",
        capture_output=True,
        text=True,
        check=False,
        timeout=30,
    )

    assert completed.returncode != 0
    assert "implementation" in completed.stderr.lower()


@pytest.mark.parametrize(
    ("implementation_root", "relative_path"),
    [
        ("plugin", "scripts/journal_bank_core.py"),
        ("plugin", "mcp/server.cjs"),
        ("plugin", "assets/journal-bank-review-widget.html"),
        ("plugin", "assets/review-workbench-adapter.json"),
        ("shared", "serialization.py"),
        ("shared", "money.py"),
        ("shared", "review_output_transaction.cjs"),
    ],
)
def test_python_preflight_rejects_copied_implementation_byte_mutation(
    tmp_path: Path,
    implementation_root: str,
    relative_path: str,
) -> None:
    output_dir = tmp_path / "run"
    output_dir, *_ = _prepare_closed_mcp_review_run(output_dir)
    before = _tree_snapshot(output_dir)
    copied_plugin, copied_shared = _copy_journal_bank_implementation(tmp_path)
    root = copied_plugin if implementation_root == "plugin" else copied_shared
    target = root / relative_path
    target.write_bytes(target.read_bytes() + b"\n")

    completed = subprocess.run(
        [
            sys.executable,
            str(copied_plugin / "scripts" / "apply_review_edits.py"),
            "--output-dir",
            str(output_dir),
            "--preflight-only",
        ],
        capture_output=True,
        text=True,
        check=False,
        timeout=10,
    )

    assert completed.returncode != 0
    assert "implementation" in completed.stderr.lower()
    assert _tree_snapshot(output_dir) == before


@pytest.mark.parametrize(
    ("implementation_root", "relative_path"),
    [
        ("plugin", "scripts/journal_bank_core.py"),
        ("plugin", "mcp/server.cjs"),
        ("plugin", "assets/journal-bank-review-widget.html"),
        ("plugin", "assets/review-workbench-adapter.json"),
        ("shared", "serialization.py"),
        ("shared", "money.py"),
        ("shared", "review_output_transaction.cjs"),
    ],
)
def test_mcp_replay_rejects_copied_implementation_byte_mutation(
    tmp_path: Path,
    implementation_root: str,
    relative_path: str,
) -> None:
    output_dir = tmp_path / "run"
    output_dir, run_intake, review_payload, _, _ = _prepare_closed_mcp_review_run(
        output_dir
    )
    before = _tree_snapshot(output_dir)
    copied_plugin, copied_shared = _copy_journal_bank_implementation(tmp_path)
    root = copied_plugin if implementation_root == "plugin" else copied_shared
    target = root / relative_path
    target.write_bytes(target.read_bytes() + b"\n")
    env = {**os.environ, "PYTHON": sys.executable}

    response = _call_mcp_server(
        _mcp_review_write_message(
            "apply_journal_bank_decisions",
            review_payload,
            run_intake,
        ),
        server_path=copied_plugin / "mcp" / "server.cjs",
        env=env,
    )[0]["result"]

    assert response["isError"] is True
    assert "receipt" in response["structuredContent"]["error"].lower()
    assert _tree_snapshot(output_dir) == before


@pytest.mark.parametrize("replay_surface", ["python", "mcp"])
@pytest.mark.parametrize("link_kind", ["symlink", "hardlink"])
def test_copied_implementation_links_are_never_authoritative(
    tmp_path: Path,
    replay_surface: str,
    link_kind: str,
) -> None:
    output_dir = tmp_path / "run"
    output_dir, run_intake, review_payload, _, _ = _prepare_closed_mcp_review_run(
        output_dir
    )
    before = _tree_snapshot(output_dir)
    copied_plugin, copied_shared = _copy_journal_bank_implementation(tmp_path)
    target = copied_shared / "review_output_transaction.cjs"
    outside = tmp_path / "outside-review-output-transaction.cjs"
    outside.write_bytes(target.read_bytes())
    if link_kind == "symlink":
        target.unlink()
        target.symlink_to(outside)
    else:
        outside.unlink()
        os.link(target, outside)

    if replay_surface == "python":
        completed = subprocess.run(
            [
                sys.executable,
                str(copied_plugin / "scripts" / "apply_review_edits.py"),
                "--output-dir",
                str(output_dir),
                "--preflight-only",
            ],
            capture_output=True,
            text=True,
            check=False,
            timeout=10,
        )
        rejected = completed.returncode != 0
    else:
        try:
            response = _call_mcp_server(
                _mcp_review_write_message(
                    "apply_journal_bank_decisions",
                    review_payload,
                    run_intake,
                ),
                server_path=copied_plugin / "mcp" / "server.cjs",
                env={**os.environ, "PYTHON": sys.executable},
            )[0]["result"]
            rejected = response["isError"] is True
        except subprocess.CalledProcessError:
            rejected = True

    assert rejected is True
    assert _tree_snapshot(output_dir) == before
    assert outside.read_bytes() == target.read_bytes()


@pytest.mark.parametrize("replay_surface", ["python", "mcp"])
def test_copied_implementation_set_cannot_self_expand(
    tmp_path: Path,
    replay_surface: str,
) -> None:
    output_dir = tmp_path / "run"
    output_dir, run_intake, review_payload, _, _ = _prepare_closed_mcp_review_run(
        output_dir
    )
    before = _tree_snapshot(output_dir)
    copied_plugin, _ = _copy_journal_bank_implementation(tmp_path)
    (copied_plugin / "scripts" / "untrusted_extension.py").write_text(
        "from __future__ import annotations\n",
        encoding="utf-8",
    )

    if replay_surface == "python":
        completed = subprocess.run(
            [
                sys.executable,
                str(copied_plugin / "scripts" / "apply_review_edits.py"),
                "--output-dir",
                str(output_dir),
                "--preflight-only",
            ],
            capture_output=True,
            text=True,
            check=False,
            timeout=10,
        )
        rejected = completed.returncode != 0
    else:
        try:
            response = _call_mcp_server(
                _mcp_review_write_message(
                    "apply_journal_bank_decisions",
                    review_payload,
                    run_intake,
                ),
                server_path=copied_plugin / "mcp" / "server.cjs",
                env={**os.environ, "PYTHON": sys.executable},
            )[0]["result"]
            rejected = response["isError"] is True
        except subprocess.CalledProcessError:
            rejected = True

    assert rejected is True
    assert _tree_snapshot(output_dir) == before


@pytest.mark.parametrize("replay_surface", ["python", "mcp"])
@pytest.mark.parametrize("set_attack", ["missing", "expanded"])
def test_implementation_receipt_set_must_remain_exact(
    tmp_path: Path,
    replay_surface: str,
    set_attack: str,
) -> None:
    output_dir = tmp_path / "run"
    output_dir, run_intake, review_payload, _, _ = _prepare_closed_mcp_review_run(
        output_dir
    )
    envelope_path = output_dir / "assurance_envelope.json"
    envelope = json.loads(envelope_path.read_text(encoding="utf-8"))
    if set_attack == "missing":
        removed_ref = envelope["implementation_artifact_refs"].pop()
        envelope["artifact_receipts"] = [
            receipt
            for receipt in envelope["artifact_receipts"]
            if receipt["artifact_id"] != removed_ref
        ]
    else:
        readme_path = ROOT / "plugins" / "journal-bank-reconciliation" / "README.md"
        readme_bytes = readme_path.read_bytes()
        expanded_receipt = {
            "schema_version": "vera.artifact_receipt.v1",
            "artifact_id": "implementation.plugin.readme_md",
            "root_id": "implementation",
            "role": "implementation",
            "path": "README.md",
            "byte_count": len(readme_bytes),
            "sha256": hashlib.sha256(readme_bytes).hexdigest(),
        }
        envelope["implementation_artifact_refs"].append(expanded_receipt["artifact_id"])
        envelope["artifact_receipts"].append(expanded_receipt)
    _write_rehashed_assurance_envelope(envelope_path, envelope)
    before = _tree_snapshot(output_dir)

    if replay_surface == "python":
        completed = subprocess.run(
            [
                sys.executable,
                str(APPLY_REVIEW_EDITS_PATH),
                "--output-dir",
                str(output_dir),
                "--preflight-only",
            ],
            capture_output=True,
            text=True,
            check=False,
            timeout=10,
        )
        rejected = completed.returncode != 0
    else:
        response = _call_mcp_server(
            _mcp_review_write_message(
                "apply_journal_bank_decisions",
                review_payload,
                run_intake,
            )
        )[0]["result"]
        rejected = response["isError"] is True

    assert rejected is True
    assert _tree_snapshot(output_dir) == before


def test_multi_sheet_workbook_is_not_silently_reduced_to_first_sheet(
    tmp_path: Path,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.xlsx"
    journal_path = tmp_path / "journal.csv"
    output_dir = tmp_path / "run"
    workbook = openpyxl.Workbook()
    first = workbook.active
    first.title = "January"
    first.append(["Date", "Amount", "Reference"])
    first.append(["2025-01-10", "80.00", "JAN100"])
    second = workbook.create_sheet("February")
    second.append(["Date", "Amount", "Reference"])
    second.append(["2025-02-10", "90.00", "FEB200"])
    workbook.save(bank_path)
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference"],
            ["2025-01-10", "80.00", "JAN100"],
            ["2025-02-10", "90.00", "FEB200"],
        ],
    )

    with pytest.raises(core.ReconciliationBlockedError) as exc_info:
        core.run_reconciliation(bank_path, journal_path, output_dir)

    audit = json.loads((output_dir / "reconciliation_audit.json").read_text())
    diagnostic = audit["diagnostics"]["bank"][0]
    assert exc_info.value.code == "unsupported_source_layout"
    assert diagnostic["failure_kind"] == "multiple_sheets_unsupported"
    assert diagnostic["workbook_sheets"] == ["January", "February"]
    assert audit["bank_row_count"] == 0
    assert (output_dir / "normalized_bank.csv").is_file()


def test_source_change_during_normalization_discards_prepared_rows(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    output_dir = tmp_path / "run"
    _save_csv(
        bank_path,
        [
            ["Date", "Amount", "Reference"],
            ["2025-03-10", "80.00", "ABC123"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference"],
            ["2025-03-10", "80.00", "ABC123"],
        ],
    )
    original_normalize_files = core._normalize_files
    changed = False

    def normalize_then_change(*args: Any, **kwargs: Any) -> Any:
        nonlocal changed
        result = original_normalize_files(*args, **kwargs)
        if args[1] == "bank" and not changed:
            changed = True
            _save_csv(
                bank_path,
                [
                    ["Date", "Amount", "Reference"],
                    ["2025-03-10", "90.00", "ABC123"],
                ],
            )
        return result

    monkeypatch.setattr(core, "_normalize_files", normalize_then_change)

    with pytest.raises(core.ReconciliationBlockedError) as exc_info:
        core.run_reconciliation(bank_path, journal_path, output_dir)

    audit = json.loads((output_dir / "reconciliation_audit.json").read_text())
    qualifications = json.loads((output_dir / "source_qualifications.json").read_text())
    bank_outcome = next(
        outcome
        for outcome in qualifications["source_outcomes"]
        if outcome["side"] == "bank"
    )
    assert exc_info.value.code == "source_changed_during_run"
    assert audit["source_snapshot_changed"] is True
    assert audit["bank_row_count"] == 0
    assert audit["diagnostics"]["bank"][0]["failure_kind"] == (
        "source_changed_during_run"
    )
    assert bank_outcome == {
        "side": "bank",
        "qualification_status": "unsupported_source_layout",
        "failure_kind": "source_changed_during_run",
        "emitted_row_count": 0,
    }
    assert (output_dir / "normalized_bank.csv").is_file()


def test_source_membership_change_discards_entire_affected_root(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank"
    bank_path.mkdir()
    journal_path = tmp_path / "journal.csv"
    output_dir = tmp_path / "run"
    _save_csv(
        bank_path / "bank.csv",
        [
            ["Date", "Amount", "Reference"],
            ["2025-03-10", "80.00", "ABC123"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference"],
            ["2025-03-10", "80.00", "ABC123"],
        ],
    )
    original_normalize_files = core._normalize_files
    changed = False

    def normalize_then_add_source(*args: Any, **kwargs: Any) -> Any:
        nonlocal changed
        result = original_normalize_files(*args, **kwargs)
        if args[1] == "bank" and not changed:
            changed = True
            _save_csv(
                bank_path / "late.csv",
                [
                    ["Date", "Amount", "Reference"],
                    ["2025-03-11", "20.00", "LATE456"],
                ],
            )
        return result

    monkeypatch.setattr(core, "_normalize_files", normalize_then_add_source)

    with pytest.raises(core.ReconciliationBlockedError) as exc_info:
        core.run_reconciliation(bank_path, journal_path, output_dir)

    audit = json.loads((output_dir / "reconciliation_audit.json").read_text())
    qualifications = json.loads((output_dir / "source_qualifications.json").read_text())
    bank_outcome = next(
        outcome
        for outcome in qualifications["source_outcomes"]
        if outcome["side"] == "bank"
    )

    assert exc_info.value.code == "source_changed_during_run"
    assert audit["changed_sources"] == [{"root_id": "source_bank", "path": "late.csv"}]
    assert audit["bank_row_count"] == 0
    assert bank_outcome == {
        "side": "bank",
        "qualification_status": "unsupported_source_layout",
        "failure_kind": "source_changed_during_run",
        "emitted_row_count": 0,
    }
    assert all(
        item["status"] == "unsupported_source_layout" and item["emitted_row_count"] == 0
        for item in qualifications["qualifications"]
        if item["qualification_id"].startswith("qualification.bank.")
    )
    assert len((output_dir / "normalized_bank.csv").read_text().splitlines()) == 1


def test_populated_fee_column_requires_explicit_reviewed_disposition(
    tmp_path: Path,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    first_inspection_dir = tmp_path / "inspection-first"
    reviewed_inspection_dir = tmp_path / "inspection-reviewed"
    _save_csv(
        bank_path,
        [
            ["Date", "Amount", "Fee", "Reference"],
            ["2025-03-10", "80.00", "2.00", "ABC123"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference"],
            ["2025-03-10", "80.00", "ABC123"],
        ],
    )

    first = core.inspect_inputs(bank_path, journal_path, first_inspection_dir)
    recipe = first.suggested_recipe
    bank_recipe = recipe["bank"]["files"]["bank.csv"]
    bank_receipt = next(
        receipt
        for receipt in json.loads(
            (first_inspection_dir / "input_receipts.json").read_text()
        )["receipts"]
        if receipt["root_id"] == "source_bank"
    )
    bank_recipe["excluded_monetary_columns"] = ["Fee"]
    bank_recipe["mapping_decision"] = core.build_mapping_review_receipt(
        decision_id="decision.mapping.bank.fee",
        reviewer_ref="reviewer.test",
        reviewed_on="2026-07-24",
        source_artifact_ref=bank_receipt["artifact_id"],
        side="bank",
        source_file="bank.csv",
        header_rows=bank_recipe["header_rows"],
        mapping=bank_recipe["mapping"],
        potential_monetary_columns=bank_recipe["potential_monetary_columns"],
        excluded_monetary_columns=["Fee"],
        csv_field_delimiter=bank_recipe["csv_field_delimiter"],
    )
    recipe_path = tmp_path / "reviewed-recipe.json"
    recipe_path.write_text(json.dumps(recipe, indent=2) + "\n", encoding="utf-8")

    reviewed = core.inspect_inputs(
        bank_path,
        journal_path,
        reviewed_inspection_dir,
        recipe_path,
    )

    first_diag = first.bank["files"][0]
    reviewed_diag = reviewed.bank["files"][0]
    assert first_diag["qualification_status"] == "needs_review"
    assert first_diag["unresolved_monetary_columns"] == ["Fee"]
    assert reviewed_diag["qualification_status"] == "qualified"
    assert reviewed_diag["excluded_monetary_columns"] == ["Fee"]
    assert reviewed.bank["row_count"] == 1


def test_mapping_receipt_rejects_mapped_and_excluded_monetary_overlap() -> None:
    core = load_core()

    with pytest.raises(
        ValueError,
        match="mapped monetary column cannot also be explicitly excluded",
    ):
        core.build_mapping_review_receipt(
            decision_id="decision.mapping.bank.overlap",
            reviewer_ref="reviewer.test",
            reviewed_on="2026-07-25",
            source_artifact_ref="source.bank.fixture.sha256",
            side="bank",
            source_file="bank.csv",
            header_rows=[1],
            mapping={"date": "Date", "amount": "Amount"},
            potential_monetary_columns=["Amount"],
            excluded_monetary_columns=["Amount"],
            csv_field_delimiter=",",
        )


def test_monetary_header_cannot_be_hidden_by_nonmonetary_mapping(
    tmp_path: Path,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    first_dir = tmp_path / "first"
    _save_csv(
        bank_path,
        [
            ["Date", "Amount", "Fee", "Reference"],
            ["2025-03-10", "80.00", "2.00", "ABC123"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference"],
            ["2025-03-10", "80.00", "ABC123"],
        ],
    )
    first = core.inspect_inputs(bank_path, journal_path, first_dir)
    recipe = first.suggested_recipe
    receipts = json.loads(
        (first_dir / "input_receipts.json").read_text(encoding="utf-8")
    )["receipts"]
    bank_recipe = recipe["bank"]["files"]["bank.csv"]
    bank_recipe["mapping"]["description"] = "Fee"
    bank_recipe["potential_monetary_columns"] = ["Amount"]
    bank_recipe["excluded_monetary_columns"] = []
    bank_recipe["mapping_decision"] = core.build_mapping_review_receipt(
        decision_id="decision.mapping.bank.hidden-fee",
        reviewer_ref="reviewer.test",
        reviewed_on="2026-07-25",
        source_artifact_ref=next(
            receipt["artifact_id"]
            for receipt in receipts
            if receipt["root_id"] == "source_bank"
        ),
        side="bank",
        source_file="bank.csv",
        header_rows=bank_recipe["header_rows"],
        mapping=bank_recipe["mapping"],
        potential_monetary_columns=["Amount"],
        excluded_monetary_columns=[],
        csv_field_delimiter=",",
    )
    recipe_path = tmp_path / "hidden-fee-recipe.json"
    recipe_path.write_text(json.dumps(recipe, indent=2) + "\n", encoding="utf-8")

    inspected = core.inspect_inputs(
        bank_path,
        journal_path,
        tmp_path / "inspected",
        recipe_path,
    )

    diagnostic = inspected.bank["files"][0]
    assert inspected.bank["row_count"] == 0
    assert diagnostic["qualification_status"] == "needs_review"
    assert diagnostic["potential_monetary_columns"] == ["Amount", "Fee"]
    assert diagnostic["unresolved_monetary_columns"] == ["Fee"]
    assert any(
        "potential_monetary_columns does not match current source evidence"
        in limitation
        for limitation in diagnostic["limitations"]
    )


@pytest.mark.parametrize(
    "reference",
    [
        "invoice 2025",
        "FY2025",
        "fy-2025",
        "FY 2025/26",
        "FY2025 period",
        "FY2025 actual",
        "FY2025 budget",
        "FY2025 forecast",
        "Budget 2025",
        "Forecast 2025",
        "Actual 2025",
        "Budget FY2025",
        "Forecast FY2025",
        "Actual FY2025",
        "Budget 2025 payment",
        "Forecast 2025 final",
        "Actual 2025 EUR",
        "FY2025 revised",
        "FY 2025 Budget v1",
        "Budget 2025 version 1",
        "2025 Budget revised",
        "Fiscal 2025",
        "Fiscal 2025 report",
        "January 2025",
        "Month 03 2025",
        "Week 10 2025",
        "W10 2025",
        "H1 2025",
        "Semester 1 2025",
        "2025 M03",
        "March FY2025",
        "FY2025 Month 03",
        "January 25",
        "Jan25",
        "Budget 25",
        "Forecast 25",
        "FY25 actual",
        "Month 03 FY25",
        "W10 FY25",
        "Fiscal Year 2025",
        "fiscal-year-2025",
        "fiscal period 2025",
        "Financial 2025",
        "financial_year_2025",
        "period2025",
        "period-2025",
        "period FY2025",
        "accounting period 2025",
        "Q1 2025",
        "Q1 2025/26",
        "2025Q1",
        "2025-2026 Q1",
        "2025/26",
        "2025-2026",
    ],
)
def test_missing_date_with_generic_period_reference_is_not_stable_evidence(
    tmp_path: Path,
    reference: str,
) -> None:
    core = load_core()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    output_dir = tmp_path / "run"
    for path in (bank_path, journal_path):
        _save_csv(
            path,
            [
                ["Date", "Amount", "Reference"],
                ["", "80.00", reference],
            ],
        )

    with pytest.raises(core.ReconciliationBlockedError) as exc_info:
        core.run_reconciliation(bank_path, journal_path, output_dir)

    audit = json.loads((output_dir / "reconciliation_audit.json").read_text())
    assert exc_info.value.code == "unsupported_source_layout"
    assert audit["diagnostics"]["bank"][0]["row_disposition_counts"] == {
        "missing_date_without_stable_reference": 1
    }
    assert (output_dir / "reconciliation_matches.csv").is_file()


def test_accept_decisions_cannot_authorize_quantitative_match_tamper(
    tmp_path: Path,
) -> None:
    core = load_core()
    apply_review_edits_module = load_apply_review_edits()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    output_dir = tmp_path / "run"
    for path in (bank_path, journal_path):
        _save_csv(
            path,
            [
                ["Date", "Amount", "Reference"],
                ["2025-03-10", "80.00", "ABC123"],
            ],
        )
    recipe_path = _prepare_reviewed_recipe(
        core, bank_path, journal_path, tmp_path / "recipe"
    )
    core.run_reconciliation(bank_path, journal_path, output_dir, recipe_path)
    apply_review_edits_module.preflight_review_application(output_dir)
    original_receipt = next(
        receipt
        for receipt in json.loads((output_dir / "artifact_receipts.json").read_text())[
            "output_receipts"
        ]
        if receipt["path"] == "reconciliation_matches.csv"
    )
    rows = _read_csv_dicts(output_dir / "reconciliation_matches.csv")
    rows[0]["bank_amount"] = "999"
    _save_csv(
        output_dir / "reconciliation_matches.csv",
        [list(rows[0]), list(rows[0].values())],
    )
    review_payload = json.loads((output_dir / "review_payload.json").read_text())
    applied_path = output_dir / "applied_decisions.json"
    applied_path.write_text(
        json.dumps(
            {
                "effects": [
                    {"action": "accept", "item_id": item["id"]}
                    for item in review_payload["items"]
                ],
                "blocker_count": 0,
                "decision_count": review_payload["item_count"],
                "item_count": review_payload["item_count"],
            }
        )
        + "\n",
        encoding="utf-8",
    )

    result = apply_review_edits_module.apply_review_edits(
        output_dir,
        applied_path,
        output_dir / "final_artifacts.json",
    )

    assert result["application_status"] == "blocked"
    assert result["assurance_report_ready"] is False
    assert any(
        "reconciliation_matches" in limitation
        for limitation in result["assurance_limitations"]
    )
    with pytest.raises(ValueError):
        core.validate_artifact_receipt(output_dir, original_receipt)


def test_apply_review_edits_sanitizes_illegal_excel_characters_during_regeneration(
    tmp_path: Path,
) -> None:
    core = load_core()
    apply_review_edits_module = load_apply_review_edits()
    raw_review_note = "Reviewer accepted\x19 reference match."
    excel_review_note = "Reviewer accepted reference match."
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    output_dir = tmp_path / "run"
    _save_csv(
        bank_path,
        [
            ["Date", "Amount", "Reference"],
            ["2025-03-10", "80.00", "ABC123"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference"],
            ["2025-03-10", "80.00", "ABC123"],
        ],
    )
    recipe_path = _prepare_reviewed_recipe(
        core, bank_path, journal_path, tmp_path / "recipe"
    )
    core.run_reconciliation(bank_path, journal_path, output_dir, recipe_path)
    apply_review_edits_module.preflight_review_application(output_dir)
    review_payload = json.loads((output_dir / "review_payload.json").read_text())
    matched_item = next(
        item for item in review_payload["items"] if item["item_type"] == "matched_pair"
    )
    backup_relative = (
        Path("revisions")
        / "originals"
        / f"reconciliation_matches__{matched_item['id']}.csv"
    )
    backup_path = output_dir / backup_relative
    backup_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(output_dir / "reconciliation_matches.csv", backup_path)
    match_rows = _read_csv_dicts(output_dir / "reconciliation_matches.csv")
    match_rows[0]["review_note"] = raw_review_note
    _save_csv(
        output_dir / "reconciliation_matches.csv",
        [list(match_rows[0]), list(match_rows[0].values())],
    )
    applied_decisions = {
        "effects": [
            {
                "action": "edit",
                "artifact_update": "structured_artifact_updated",
                "target_artifact": "reconciliation_matches.csv",
                "derived_native_regeneration_paths": [
                    "journal_bank_reconciliation.xlsx"
                ],
                "edit_value": raw_review_note,
                "item_id": matched_item["id"],
                "target_id_field": "bank_transaction_id",
                "target_record_id": matched_item["data"]["target_record_id"],
                "target_field": "review_note",
                "original_artifact_backup": backup_relative.as_posix(),
                "structured_update": {
                    "id_field": "bank_transaction_id",
                    "record_id": matched_item["data"]["target_record_id"],
                    "target_field": "review_note",
                },
            },
            *[
                {"action": "accept", "item_id": item["id"]}
                for item in review_payload["items"]
                if item["id"] != matched_item["id"]
            ],
        ],
        "blocker_count": 0,
        "decision_count": review_payload["item_count"],
        "item_count": review_payload["item_count"],
    }
    applied_path = output_dir / "applied_decisions.json"
    final_artifacts_path = output_dir / "final_artifacts.json"
    applied_path.write_text(json.dumps(applied_decisions) + "\n", encoding="utf-8")

    result = apply_review_edits_module.apply_review_edits(
        output_dir,
        applied_path,
        final_artifacts_path,
    )

    workbook = openpyxl.load_workbook(output_dir / "journal_bank_reconciliation.xlsx")
    written_final_artifacts = json.loads(final_artifacts_path.read_text())
    workbook_output = next(
        output
        for output in written_final_artifacts["outputs"]
        if output["path"] == "journal_bank_reconciliation.xlsx"
    )
    assert result["ok"] is True
    match_sheet = workbook["matches"]
    review_note_column = next(
        cell.column for cell in match_sheet[1] if cell.value == "review_note"
    )
    assert match_sheet.cell(row=2, column=review_note_column).value == excel_review_note
    assert list(workbook_output["required_cells"]["matches"].values()) == [
        excel_review_note
    ]


def test_completed_review_promotes_only_closed_assurance_and_reseals_outputs(
    tmp_path: Path,
) -> None:
    core = load_core()
    apply_review_edits_module = load_apply_review_edits()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    output_dir = tmp_path / "run"
    _save_csv(
        bank_path,
        [
            ["Date", "Amount", "Reference"],
            ["2025-03-10", "80.00", "ABC123"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference"],
            ["2025-03-10", "80.00", "ABC123"],
        ],
    )
    recipe_path = _prepare_reviewed_recipe(
        core, bank_path, journal_path, tmp_path / "recipe"
    )
    core.run_reconciliation(bank_path, journal_path, output_dir, recipe_path)
    apply_review_edits_module.preflight_review_application(output_dir)
    old_receipts = json.loads((output_dir / "artifact_receipts.json").read_text())[
        "output_receipts"
    ]
    old_match_sha = next(
        receipt["sha256"]
        for receipt in old_receipts
        if receipt["path"] == "reconciliation_matches.csv"
    )
    match_rows = _read_csv_dicts(output_dir / "reconciliation_matches.csv")
    backup_relative = (
        Path("revisions") / "originals" / "reconciliation_matches__reviewed-match.csv"
    )
    backup_path = output_dir / backup_relative
    backup_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(output_dir / "reconciliation_matches.csv", backup_path)
    match_rows[0]["review_note"] = "Reviewer accepted exact source-bound match."
    _save_csv(
        output_dir / "reconciliation_matches.csv",
        [list(match_rows[0]), list(match_rows[0].values())],
    )
    review_payload = json.loads((output_dir / "review_payload.json").read_text())
    matched_item = next(
        item for item in review_payload["items"] if item["item_type"] == "matched_pair"
    )
    effects = [
        {
            "action": "edit",
            "artifact_update": "structured_artifact_updated",
            "target_artifact": "reconciliation_matches.csv",
            "derived_native_regeneration_paths": ["journal_bank_reconciliation.xlsx"],
            "requires_native_regeneration": True,
            "edit_value": "Reviewer accepted exact source-bound match.",
            "item_id": matched_item["id"],
            "target_id_field": "bank_transaction_id",
            "target_record_id": matched_item["data"]["target_record_id"],
            "target_field": "review_note",
            "original_artifact_backup": backup_relative.as_posix(),
            "structured_update": {
                "id_field": "bank_transaction_id",
                "record_id": matched_item["data"]["target_record_id"],
                "target_field": "review_note",
            },
        },
        *[
            {"action": "accept", "item_id": item["id"]}
            for item in review_payload["items"]
            if item["id"] != matched_item["id"]
        ],
    ]
    applied_path = output_dir / "applied_decisions.json"
    applied_path.write_text(
        json.dumps(
            {
                "effects": effects,
                "blocker_count": 0,
                "decision_count": review_payload["item_count"],
                "item_count": review_payload["item_count"],
            }
        )
        + "\n",
        encoding="utf-8",
    )

    result = apply_review_edits_module.apply_review_edits(
        output_dir,
        applied_path,
        output_dir / "final_artifacts.json",
    )

    gates = json.loads((output_dir / "assurance_gates.json").read_text())
    receipts = json.loads((output_dir / "artifact_receipts.json").read_text())[
        "output_receipts"
    ]
    new_match_sha = next(
        receipt["sha256"]
        for receipt in receipts
        if receipt["path"] == "reconciliation_matches.csv"
    )
    assert result["application_status"] == "final_ready"
    assert result["assurance_report_ready"] is True
    assert gates["gates"]["semantic_review"]["status"] == "passed"
    assert gates["gates"]["reporting"]["status"] == "passed"
    assert gates["report_ready"] is True
    assert old_match_sha != new_match_sha
    assert any(receipt["path"] == "applied_decisions.json" for receipt in receipts)
    for receipt in receipts:
        core.validate_artifact_receipt(output_dir, receipt)


def test_accepting_unresolved_rows_cannot_make_report_final_ready(
    tmp_path: Path,
) -> None:
    core = load_core()
    apply_review_edits_module = load_apply_review_edits()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    output_dir = tmp_path / "run"
    _save_csv(
        bank_path,
        [
            ["Date", "Amount", "Reference"],
            ["2025-03-10", "80.00", "ABC123"],
            ["2025-03-11", "10.00", "FEE999"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference"],
            ["2025-03-10", "80.00", "ABC123"],
        ],
    )
    recipe_path = _prepare_reviewed_recipe(
        core, bank_path, journal_path, tmp_path / "recipe"
    )
    core.run_reconciliation(bank_path, journal_path, output_dir, recipe_path)
    apply_review_edits_module.preflight_review_application(output_dir)
    review_payload = json.loads((output_dir / "review_payload.json").read_text())
    applied_path = output_dir / "applied_decisions.json"
    applied_path.write_text(
        json.dumps(
            {
                "effects": [
                    {"action": "accept", "item_id": item["id"]}
                    for item in review_payload["items"]
                ],
                "blocker_count": 0,
                "decision_count": review_payload["item_count"],
                "item_count": review_payload["item_count"],
            }
        )
        + "\n",
        encoding="utf-8",
    )

    result = apply_review_edits_module.apply_review_edits(
        output_dir,
        applied_path,
        output_dir / "final_artifacts.json",
    )

    gates = json.loads((output_dir / "assurance_gates.json").read_text())
    assert result["application_status"] == "blocked"
    assert result["assurance_report_ready"] is False
    assert gates["gates"]["reconciliation"]["status"] == "withheld"
    assert gates["gates"]["semantic_review"]["status"] == "passed"
    assert gates["gates"]["reporting"]["status"] == "blocked"
    assert gates["report_ready"] is False


def test_unexpected_output_tamper_blocks_review_and_is_not_resealed(
    tmp_path: Path,
) -> None:
    core = load_core()
    apply_review_edits_module = load_apply_review_edits()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    output_dir = tmp_path / "run"
    _save_csv(
        bank_path,
        [
            ["Date", "Amount", "Reference"],
            ["2025-03-10", "80.00", "ABC123"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference"],
            ["2025-03-10", "80.00", "ABC123"],
        ],
    )
    recipe_path = _prepare_reviewed_recipe(
        core, bank_path, journal_path, tmp_path / "recipe"
    )
    core.run_reconciliation(bank_path, journal_path, output_dir, recipe_path)
    apply_review_edits_module.preflight_review_application(output_dir)
    review_payload = json.loads((output_dir / "review_payload.json").read_text())
    normalized_bank_path = output_dir / "normalized_bank.csv"
    normalized_bank_path.write_text(
        normalized_bank_path.read_text() + "tampered\n",
        encoding="utf-8",
    )
    applied_path = output_dir / "applied_decisions.json"
    applied_path.write_text(
        json.dumps(
            {
                "effects": [
                    {"action": "accept", "item_id": item["id"]}
                    for item in review_payload["items"]
                ],
                "blocker_count": 0,
                "decision_count": review_payload["item_count"],
                "item_count": review_payload["item_count"],
            }
        )
        + "\n",
        encoding="utf-8",
    )

    result = apply_review_edits_module.apply_review_edits(
        output_dir,
        applied_path,
        output_dir / "final_artifacts.json",
    )

    receipt = next(
        receipt
        for receipt in json.loads((output_dir / "artifact_receipts.json").read_text())[
            "output_receipts"
        ]
        if receipt["path"] == "normalized_bank.csv"
    )
    assert result["application_status"] == "blocked"
    assert result["assurance_report_ready"] is False
    assert any(
        "output.normalized_bank_csv" in limitation
        for limitation in result["assurance_limitations"]
    )
    with pytest.raises(ValueError):
        core.validate_artifact_receipt(output_dir, receipt)


def test_semantic_prepare_cli_replays_renamed_managed_run_and_closes_outputs(
    tmp_path: Path,
) -> None:
    ledger = _load_customer_ledger()
    core = load_core()
    client_root = tmp_path / "Managed Customer"
    client_root.mkdir()
    client_id = "client_222222222222222222222222"
    ledger.create_client_manifest(client_root, client_id)
    engagement = ledger.create_engagement(
        client_root,
        client_id,
        "Portable semantic review",
    )
    source_dir = tmp_path / "received"
    source_dir.mkdir()
    bank_source = source_dir / "bank.csv"
    journal_source = source_dir / "journal.csv"
    _save_csv(
        bank_source,
        [
            ["Date", "Amount", "Description", "Beneficiary"],
            ["2026-05-08", "80.00", "Payment Alpha", "Alpha"],
            ["2026-05-08", "80.00", "Payment Beta", "Beta"],
        ],
    )
    _save_csv(
        journal_source,
        [
            ["Date", "Amount", "Description", "Beneficiary"],
            ["2026-05-08", "80.00", "Invoice Alpha", "Alpha"],
            ["2026-05-08", "80.00", "Invoice Beta", "Beta"],
        ],
    )
    imported_bank = ledger.import_document(
        client_root,
        client_id,
        engagement["engagement_id"],
        bank_source,
        "source",
    )
    imported_journal = ledger.import_document(
        client_root,
        client_id,
        engagement["engagement_id"],
        journal_source,
        "journal",
    )
    prepared = ledger.prepare_run(
        client_root,
        client_id,
        engagement["engagement_id"],
        "journal-bank-reconciliation",
        "test-version",
        input_ids=[
            imported_bank["receipt"]["input_id"],
            imported_journal["receipt"]["input_id"],
        ],
    )
    running = ledger.start_run(
        client_root,
        engagement["engagement_id"],
        prepared["run"]["run_id"],
    )
    input_paths = {
        binding["role"]: Path(binding["path"])
        for binding in running["context"]["input_bindings"]
    }
    output_dir = Path(running["output_dir"])
    context_path = Path(running["context_path"])
    recipe_path = _prepare_reviewed_recipe(
        core,
        input_paths["source"],
        input_paths["journal"],
        output_dir / "recipe",
        tolerance="0",
        date_window_days=0,
    )
    reconciliation_dir = output_dir / "reconciliation"
    reconciliation = subprocess.run(
        [
            sys.executable,
            "-B",
            str(SCRIPT_DIR / "run_reconciliation.py"),
            str(input_paths["source"]),
            str(input_paths["journal"]),
            "--output-dir",
            str(reconciliation_dir),
            "--recipe",
            str(recipe_path),
            "--tolerance",
            "0",
            "--date-window-days",
            "0",
            "--client-engagement",
            str(context_path),
        ],
        capture_output=True,
        check=False,
        text=True,
        timeout=120,
    )
    assert reconciliation.returncode == 0, reconciliation.stderr
    run_intake = json.loads(
        (reconciliation_dir / "run_intake.json").read_text(encoding="utf-8")
    )
    assert run_intake["path_reference"] == "run_root_relative"
    assert run_intake["output_dir"] == "outputs/reconciliation"
    assert all(not Path(value).is_absolute() for value in run_intake["input_paths"])

    renamed_output, renamed_context, stale_output = _rename_customer_output(output_dir)
    renamed_reconciliation = renamed_output / "reconciliation"
    semantic_dir = renamed_output / "semantic-review"
    portable_intake_path = renamed_reconciliation / "run_intake.json"
    portable_intake_bytes = portable_intake_path.read_bytes()
    forged_intake = json.loads(portable_intake_bytes.decode("utf-8"))
    forged_intake["assumptions"]["bank_path"] = "outputs/recipe/input_receipts.json"
    portable_intake_path.write_text(
        json.dumps(forged_intake, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    blocked = subprocess.run(
        [
            sys.executable,
            "-B",
            str(SEMANTIC_REVIEW_PATH),
            "prepare",
            str(renamed_reconciliation),
            "--output-dir",
            str(semantic_dir),
            "--required-level",
            "identifier_match",
            "--client-engagement",
            str(renamed_context),
        ],
        capture_output=True,
        check=False,
        text=True,
        timeout=120,
    )
    assert blocked.returncode == 2
    assert "exact receipts" in blocked.stderr
    assert not any(semantic_dir.iterdir())
    portable_intake_path.write_bytes(portable_intake_bytes)

    semantic = subprocess.run(
        [
            sys.executable,
            "-B",
            str(SEMANTIC_REVIEW_PATH),
            "prepare",
            str(renamed_reconciliation),
            "--output-dir",
            str(semantic_dir),
            "--required-level",
            "identifier_match",
            "--client-engagement",
            str(renamed_context),
        ],
        capture_output=True,
        check=False,
        text=True,
        timeout=120,
    )
    assert semantic.returncode == 0, semantic.stderr
    assert not stale_output.exists()
    assert {
        "residual_candidate_graph.json",
        "luna_output_schema.json",
        "luna_prompt.md",
        "semantic_review_status.json",
        "semantic_resolution_application.json",
        "resolution_funnel.json",
        "human_review_queue.json",
    }.issubset({path.name for path in semantic_dir.iterdir()})
    graph = json.loads(
        (semantic_dir / "residual_candidate_graph.json").read_text(encoding="utf-8")
    )
    assert graph["resolution_policy"]["required_level"] == "identifier_match"

    physical_outputs = sorted(
        path for path in renamed_output.rglob("*") if path.is_file()
    )
    media_types = {
        ".csv": "text/csv",
        ".json": "application/json",
        ".md": "text/markdown",
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    declarations = [
        {
            "artifact_id": f"journal_bank.output.{index:03d}",
            "path": path.relative_to(renamed_output).as_posix(),
            "purpose": (
                "Preserve this Journal–Bank run output for professional review: "
                f"{path.relative_to(renamed_output).as_posix()}."
            ),
            "audience": "review",
            "media_type": media_types.get(path.suffix.lower(), "text/plain"),
        }
        for index, path in enumerate(physical_outputs, start=1)
    ]
    renamed_client_root = renamed_context.parents[5]
    finalized = ledger.finalize_run(
        renamed_client_root,
        engagement["engagement_id"],
        running["run"]["run_id"],
        declarations,
    )
    artifacts = finalized["artifact_manifest"]["artifacts"]
    assert {artifact["path"] for artifact in artifacts} == {
        path.relative_to(renamed_output).as_posix() for path in physical_outputs
    }
    assert all(artifact["purpose"] for artifact in artifacts)
    completed = ledger.complete_run(
        renamed_client_root,
        engagement["engagement_id"],
        running["run"]["run_id"],
    )
    assert completed["run"]["status"] == "completed"


def test_semantic_prepare_builds_bounded_hash_bound_advisory_graph(
    tmp_path: Path,
) -> None:
    _, semantic_review, reconciliation_dir, semantic_dir = (
        _prepare_ambiguous_semantic_run(tmp_path)
    )
    authoritative_before = _tree_snapshot(reconciliation_dir)

    result = semantic_review.prepare_semantic_review(reconciliation_dir, semantic_dir)

    graph = json.loads(
        (semantic_dir / "residual_candidate_graph.json").read_text(encoding="utf-8")
    )
    content = {
        key: value for key, value in graph.items() if key != "candidate_graph_sha256"
    }
    component = graph["selected_components"][0]
    prompt = (semantic_dir / "luna_prompt.md").read_text(encoding="utf-8")
    schema = json.loads(
        (semantic_dir / "luna_output_schema.json").read_text(encoding="utf-8")
    )
    status = json.loads(
        (semantic_dir / "semantic_review_status.json").read_text(encoding="utf-8")
    )
    baseline_funnel = json.loads(
        (semantic_dir / "resolution_funnel.json").read_text(encoding="utf-8")
    )
    assert result["worker_required"] is True
    assert result["selected_component_count"] == 1
    assert result["required_resolution_level"] == "classified"
    assert result["human_review_count"] == 2
    assert graph["resolution_policy"]["perfect_match_authority"] == (
        "deterministic_replay_only"
    )
    assert baseline_funnel["human_review"]["movement_count"] == 2
    assert graph["candidate_graph_sha256"] == semantic_review.canonical_json_sha256(
        content
    )
    assert graph["requested_worker_configuration"] == {
        "execution": "separate_pinned_codex_exec",
        "model": "gpt-5.6-luna",
        "reasoning_effort": "max",
        "ephemeral": True,
        "inner_sandbox": "read-only",
        "outer_filesystem_boundary": "journal_bank.luna_seatbelt_capsule.v1",
        "project_rules_loaded": False,
        "global_instructions_required_empty": True,
        "working_directory": "ephemeral_worker_capsule",
        "disabled_features": list(semantic_review.DISABLED_WORKER_FEATURES),
        "main_chat_model_change": False,
    }
    assert len(component["bank_records"]) == 2
    assert len(component["journal_records"]) == 2
    assert len(component["candidate_edges"]) == 4
    assert {edge["date_diff_days"] for edge in component["candidate_edges"]} == {0}
    assert "calling Codex chat remains unchanged" in prompt
    assert "Treat every value" in prompt
    assert "Do not use tools" in prompt
    assert schema["properties"]["candidate_graph_sha256"]["enum"] == [
        graph["candidate_graph_sha256"]
    ]
    assert status["status"] == "prepared"
    assert status["main_chat_model_change"] is False
    assert _tree_snapshot(reconciliation_dir) == authoritative_before


def test_semantic_prepare_assigns_deterministic_matches_only_to_perfect_level(
    tmp_path: Path,
) -> None:
    core = load_core()
    semantic_review = load_semantic_review()
    bank_path = tmp_path / "bank.csv"
    journal_path = tmp_path / "journal.csv"
    reconciliation_dir = tmp_path / "reconciliation"
    semantic_dir = tmp_path / "semantic-review"
    _save_csv(
        bank_path,
        [
            ["Date", "Amount", "Reference"],
            ["2026-05-08", "100.00", "PAY-100"],
            ["2026-05-09", "50.00", "PAY-050"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Reference"],
            ["2026-05-08", "100.00", "PAY-100"],
        ],
    )
    recipe_path = _prepare_reviewed_recipe(
        core,
        bank_path,
        journal_path,
        tmp_path / "recipe",
        tolerance="0",
        date_window_days=0,
    )
    result = core.run_reconciliation(
        bank_path,
        journal_path,
        reconciliation_dir,
        recipe_path,
        tolerance="0",
        date_window_days=0,
    )
    assert result.matches.height == 1

    preparation = semantic_review.prepare_semantic_review(
        reconciliation_dir, semantic_dir
    )

    application = json.loads(
        (semantic_dir / "semantic_resolution_application.json").read_text(
            encoding="utf-8"
        )
    )
    by_level = {
        item["highest_level_reached"]: item for item in application["assignments"]
    }
    assert by_level["perfect_match"]["decision_authority"] == "deterministic"
    assert by_level["perfect_match"]["human_review_required"] is False
    assert by_level["unresolved"]["human_review_required"] is True
    graph_path = semantic_dir / "residual_candidate_graph.json"
    graph = json.loads(graph_path.read_text(encoding="utf-8"))
    component = graph["selected_components"][0]
    assert preparation["worker_required"] is True
    assert component["journal_records"] == []
    assert component["candidate_edges"] == []
    response = {
        "schema_version": "journal_bank.semantic_worker_response.v2",
        "candidate_graph_sha256": graph["candidate_graph_sha256"],
        "component_reviews": [
            {
                "component_id": component["component_id"],
                "decisions": [
                    {
                        "bank_transaction_id": component["bank_records"][0][
                            "transaction_id"
                        ],
                        "verdict": "no_match",
                        "journal_transaction_id": None,
                        "evidence_fields": ["reference"],
                        "rationale": "The bank reference supports an expense classification.",
                        "contradictions": [],
                        "requested_evidence": [],
                        "resolution_level": "classified",
                        "classification": "other expense",
                        "identified_counterparty": None,
                    }
                ],
            }
        ],
    }
    response_path, events_path = _write_semantic_worker_result(semantic_dir, response)

    semantic_review.validate_semantic_review(
        reconciliation_dir,
        semantic_dir,
        graph_path,
        response_path,
        events_path,
    )

    applied = json.loads(
        (semantic_dir / "semantic_resolution_application.json").read_text(
            encoding="utf-8"
        )
    )
    assert applied["summary"]["human_review_count"] == 0
    assert {item["highest_level_reached"] for item in applied["assignments"]} == {
        "classified",
        "perfect_match",
    }


def test_semantic_prepare_rejects_authoritative_and_semantic_same_directory(
    tmp_path: Path,
) -> None:
    _, semantic_review, reconciliation_dir, _ = _prepare_ambiguous_semantic_run(
        tmp_path
    )
    same_dir = reconciliation_dir.parent / "semantic-review"
    reconciliation_dir.rename(same_dir)

    with pytest.raises(ValueError, match="must be distinct"):
        semantic_review.prepare_semantic_review(same_dir, same_dir)

    assert not (same_dir / "residual_candidate_graph.json").exists()


def test_semantic_prepare_rejects_mutable_receipt_not_closed_to_envelope(
    tmp_path: Path,
) -> None:
    _, semantic_review, reconciliation_dir, semantic_dir = (
        _prepare_ambiguous_semantic_run(tmp_path)
    )
    unmatched_path = reconciliation_dir / "unmatched_bank.csv"
    with unmatched_path.open("r", encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle))
    rows[0]["description"] = "FORGED SEMANTIC CONTEXT"
    with unmatched_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)
    _reseal_semantic_source_receipt(
        reconciliation_dir,
        artifact_id="output.unmatched_bank_csv",
        artifact_path=unmatched_path,
    )

    with pytest.raises(ValueError, match="artifact receipt"):
        semantic_review.prepare_semantic_review(reconciliation_dir, semantic_dir)

    assert not (semantic_dir / "residual_candidate_graph.json").exists()


def test_semantic_prepare_rejects_source_change_after_receipt_validation(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    _, semantic_review, reconciliation_dir, semantic_dir = (
        _prepare_ambiguous_semantic_run(tmp_path)
    )
    intake_path = reconciliation_dir / "run_intake.json"
    real_validate = semantic_review.validate_artifact_receipt
    changed = False

    def mutate_after_validation(root: Path, receipt: dict[str, Any]) -> dict[str, Any]:
        nonlocal changed
        validated = real_validate(root, receipt)
        if receipt.get("artifact_id") == "output.run_intake_json" and not changed:
            intake = json.loads(intake_path.read_text(encoding="utf-8"))
            intake["run_id"] = "raced_run_id"
            intake_path.write_text(
                json.dumps(intake, ensure_ascii=False, indent=2) + "\n",
                encoding="utf-8",
            )
            changed = True
        return validated

    monkeypatch.setattr(
        semantic_review, "validate_artifact_receipt", mutate_after_validation
    )

    with pytest.raises(ValueError, match="changed during validation"):
        semantic_review.prepare_semantic_review(reconciliation_dir, semantic_dir)


def test_semantic_prepare_rejects_hardlinked_graph_input(tmp_path: Path) -> None:
    _, semantic_review, reconciliation_dir, semantic_dir = (
        _prepare_ambiguous_semantic_run(tmp_path)
    )
    source = reconciliation_dir / "unmatched_bank.csv"
    os.link(source, tmp_path / "unmatched_bank_alias.csv")

    with pytest.raises(ValueError, match="ordinary single-link file"):
        semantic_review.prepare_semantic_review(reconciliation_dir, semantic_dir)


def test_semantic_prepare_defers_candidate_discovery_at_hard_edge_cap(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    _, semantic_review, reconciliation_dir, semantic_dir = (
        _prepare_ambiguous_semantic_run(tmp_path)
    )
    monkeypatch.setattr(semantic_review, "MAX_DISCOVERED_EDGES", 3)

    result = semantic_review.prepare_semantic_review(reconciliation_dir, semantic_dir)

    graph = json.loads(
        (semantic_dir / "residual_candidate_graph.json").read_text(encoding="utf-8")
    )
    assert result["worker_required"] is False
    assert graph["counts"]["candidate_discovery_complete"] is False
    assert graph["counts"]["eligible_component_count"] is None
    assert graph["selected_components"] == []
    assert graph["deferred_components"] == [
        {
            "component_id": graph["deferred_components"][0]["component_id"],
            "bank_count": 2,
            "journal_count": 2,
            "observed_edge_count": 4,
            "observed_candidate_comparison_count": 4,
            "reason": "candidate_discovery_edge_cap_exceeded",
        }
    ]


def test_semantic_prepare_defers_candidate_discovery_at_raw_comparison_cap(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    core = load_core()
    semantic_review = load_semantic_review()
    case_dir = tmp_path / "case"
    bank_path = case_dir / "bank.csv"
    journal_path = case_dir / "journal.csv"
    reconciliation_dir = case_dir / "reconciliation"
    semantic_dir = case_dir / "semantic-review"
    case_dir.mkdir()
    _save_csv(
        bank_path,
        [
            ["Date", "Amount", "Entity", "Description"],
            ["2026-05-08", "80.00", "bank-a", "Payment Alpha"],
            ["2026-05-08", "80.00", "bank-b", "Payment Beta"],
        ],
    )
    _save_csv(
        journal_path,
        [
            ["Date", "Amount", "Entity", "Description"],
            ["2026-05-08", "80.00", "journal-a", "Invoice Alpha"],
            ["2026-05-08", "80.00", "journal-b", "Invoice Beta"],
        ],
    )
    recipe_path = _prepare_reviewed_recipe(
        core,
        bank_path,
        journal_path,
        case_dir / "recipe",
        tolerance="0",
        date_window_days=0,
    )
    run_result = core.run_reconciliation(
        bank_path,
        journal_path,
        reconciliation_dir,
        recipe_path,
        tolerance="0",
        date_window_days=0,
    )
    assert run_result.matches.is_empty()
    monkeypatch.setattr(semantic_review, "MAX_DISCOVERED_CANDIDATE_COMPARISONS", 3)

    result = semantic_review.prepare_semantic_review(reconciliation_dir, semantic_dir)

    graph = json.loads(
        (semantic_dir / "residual_candidate_graph.json").read_text(encoding="utf-8")
    )
    assert result["worker_required"] is False
    assert graph["counts"]["candidate_discovery_complete"] is False
    assert graph["counts"]["eligible_component_count"] is None
    assert graph["selected_components"] == []
    assert graph["deferred_components"] == [
        {
            "component_id": graph["deferred_components"][0]["component_id"],
            "bank_count": 2,
            "journal_count": 2,
            "observed_edge_count": 0,
            "observed_candidate_comparison_count": 4,
            "reason": "candidate_discovery_comparison_cap_exceeded",
        }
    ]


def test_semantic_deferred_summary_keeps_graph_replay_within_byte_cap(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    _, semantic_review, reconciliation_dir, semantic_dir = (
        _prepare_ambiguous_semantic_run(tmp_path)
    )
    components = [
        {
            "component_id": f"component.synthetic.{index}",
            "bank_records": [{"transaction_id": f"bank:{index}"}],
            "journal_records": [{"transaction_id": f"journal:{index}"}],
            "candidate_edges": [
                {
                    "bank_transaction_id": f"bank:{index}",
                    "journal_transaction_id": f"journal:{index}",
                    "amount_delta": "0",
                    "date_diff_days": 0,
                    "shared_references": [],
                }
            ],
        }
        for index in range(5)
    ]
    monkeypatch.setattr(semantic_review, "MAX_DEFERRED_SUMMARIES", 2)
    monkeypatch.setattr(
        semantic_review,
        "_candidate_components",
        lambda *args, **kwargs: (components, None),
    )

    semantic_review.prepare_semantic_review(reconciliation_dir, semantic_dir)

    graph_path = semantic_dir / "residual_candidate_graph.json"
    graph = json.loads(graph_path.read_text(encoding="utf-8"))
    response = {
        "schema_version": "journal_bank.semantic_worker_response.v2",
        "candidate_graph_sha256": graph["candidate_graph_sha256"],
        "component_reviews": [],
    }
    response_path, events_path = _write_semantic_worker_result(semantic_dir, response)
    result = semantic_review.validate_semantic_review(
        reconciliation_dir,
        semantic_dir,
        graph_path,
        response_path,
        events_path,
    )
    assert graph_path.stat().st_size <= semantic_review.MAX_GRAPH_BYTES
    assert len(graph["deferred_components"]) == 2
    assert graph["deferred_component_summary"] == {
        "omitted_component_count": 3,
        "omitted_bank_count": 3,
        "omitted_journal_count": 3,
        "known_observed_edge_count": 3,
        "unknown_observed_edge_component_count": 0,
        "known_observed_candidate_comparison_count": 0,
        "unknown_observed_candidate_comparison_component_count": 3,
        "reason_counts": {"unexpected_deterministic_singleton": 3},
        "omitted_components_sha256": graph["deferred_component_summary"][
            "omitted_components_sha256"
        ],
    }
    assert result["summary"]["decision_count"] == 0


def test_semantic_validate_applies_certainty_funnel_without_changing_strict_ledger(
    tmp_path: Path,
) -> None:
    _, semantic_review, reconciliation_dir, semantic_dir = (
        _prepare_ambiguous_semantic_run(tmp_path)
    )
    semantic_review.prepare_semantic_review(reconciliation_dir, semantic_dir)
    graph = json.loads(
        (semantic_dir / "residual_candidate_graph.json").read_text(encoding="utf-8")
    )
    response = _valid_semantic_response(graph)
    response_path, events_path = _write_semantic_worker_result(semantic_dir, response)
    authoritative_before = _tree_snapshot(reconciliation_dir)

    result = semantic_review.validate_semantic_review(
        reconciliation_dir,
        semantic_dir,
        semantic_dir / "residual_candidate_graph.json",
        response_path,
        events_path,
    )

    validated = json.loads(
        (semantic_dir / "semantic_suggestions_validated.json").read_text(
            encoding="utf-8"
        )
    )
    worker_run = json.loads(
        (semantic_dir / "semantic_worker_run.json").read_text(encoding="utf-8")
    )
    status = json.loads(
        (semantic_dir / "semantic_review_status.json").read_text(encoding="utf-8")
    )
    application = json.loads(
        (semantic_dir / "semantic_resolution_application.json").read_text(
            encoding="utf-8"
        )
    )
    funnel = json.loads(
        (semantic_dir / "resolution_funnel.json").read_text(encoding="utf-8")
    )
    review_queue = json.loads(
        (semantic_dir / "human_review_queue.json").read_text(encoding="utf-8")
    )
    assert result["summary"] == {
        "component_count": 1,
        "decision_count": 2,
        "suggest_match_count": 2,
        "abstention_count": 0,
        "no_match_count": 0,
        "meets_threshold_count": 2,
        "human_review_count": 0,
    }
    assert validated["worker_output_advisory_until_validated"] is True
    assert validated["application_status"] == "applied_to_resolution_funnel"
    assert validated["strict_reconciliation_unchanged"] is True
    assert validated["main_codex_review_required"] is False
    assert application["summary"] == {
        "movement_count": 2,
        "meets_threshold_count": 2,
        "human_review_count": 0,
    }
    assert {item["highest_level_reached"] for item in application["assignments"]} == {
        "beneficiary_match"
    }
    assert {item["level"]: item["movement_count"] for item in funnel["at_least"]} == {
        "classified": 2,
        "candidate_match": 2,
        "beneficiary_match": 2,
        "identifier_match": 0,
        "perfect_match": 0,
    }
    assert review_queue["items"] == []
    assert worker_run["requested_worker_configuration"]["model"] == "gpt-5.6-luna"
    assert worker_run["requested_worker_configuration"]["reasoning_effort"] == "max"
    assert worker_run["runtime_attestation"] == {
        "separate_thread_and_usage_observed": True,
        "model_observed": False,
        "reasoning_effort_observed": False,
        "filesystem_boundary_receipt_validated": True,
        "jsonl_visibility_complete": False,
        "tool_use_absence_observed": False,
        "trust_boundary": "journal_bank.luna_seatbelt_capsule.v1",
    }
    assert worker_run["jsonl_observation"] == {
        "visibility_complete": False,
        "visible_forbidden_item_count": 0,
        "tool_use_absence_observed": False,
    }
    assert "tool_item_count" not in worker_run
    assert worker_run["main_chat_model_change"] is False
    assert status["status"] == "completed_validated"
    assert _tree_snapshot(reconciliation_dir) == authoritative_before


def test_semantic_classification_clears_review_at_classified_threshold(
    tmp_path: Path,
) -> None:
    _, semantic_review, reconciliation_dir, semantic_dir = (
        _prepare_ambiguous_semantic_run(tmp_path)
    )
    semantic_review.prepare_semantic_review(
        reconciliation_dir,
        semantic_dir,
        required_resolution_level="classified",
    )
    graph_path = semantic_dir / "residual_candidate_graph.json"
    graph = json.loads(graph_path.read_text(encoding="utf-8"))
    response = _valid_semantic_response(graph)
    for decision in response["component_reviews"][0]["decisions"]:
        decision.update(
            {
                "verdict": "no_match",
                "journal_transaction_id": None,
                "evidence_fields": ["description"],
                "resolution_level": "classified",
                "classification": "payroll",
                "identified_counterparty": None,
                "rationale": "The bank narrative is sufficient to classify payroll.",
            }
        )
    response_path, events_path = _write_semantic_worker_result(semantic_dir, response)

    semantic_review.validate_semantic_review(
        reconciliation_dir,
        semantic_dir,
        graph_path,
        response_path,
        events_path,
    )

    application = json.loads(
        (semantic_dir / "semantic_resolution_application.json").read_text(
            encoding="utf-8"
        )
    )
    assert application["summary"]["human_review_count"] == 0
    assert {
        (item["highest_level_reached"], item["classification"])
        for item in application["assignments"]
    } == {("classified", "payroll")}


def test_semantic_beneficiary_resolution_stays_in_review_at_identifier_threshold(
    tmp_path: Path,
) -> None:
    _, semantic_review, reconciliation_dir, semantic_dir = (
        _prepare_ambiguous_semantic_run(tmp_path)
    )
    semantic_review.prepare_semantic_review(
        reconciliation_dir,
        semantic_dir,
        required_resolution_level="identifier_match",
    )
    graph_path = semantic_dir / "residual_candidate_graph.json"
    graph = json.loads(graph_path.read_text(encoding="utf-8"))
    response_path, events_path = _write_semantic_worker_result(
        semantic_dir,
        _valid_semantic_response(graph),
    )

    semantic_review.validate_semantic_review(
        reconciliation_dir,
        semantic_dir,
        graph_path,
        response_path,
        events_path,
    )

    review_queue = json.loads(
        (semantic_dir / "human_review_queue.json").read_text(encoding="utf-8")
    )
    assert review_queue["required_resolution_level"] == "identifier_match"
    assert len(review_queue["items"]) == 2
    assert {item["highest_level_reached"] for item in review_queue["items"]} == {
        "beneficiary_match"
    }


def test_semantic_validate_is_idempotent_for_exact_worker_generation(
    tmp_path: Path,
) -> None:
    _, semantic_review, reconciliation_dir, semantic_dir = (
        _prepare_ambiguous_semantic_run(tmp_path)
    )
    semantic_review.prepare_semantic_review(reconciliation_dir, semantic_dir)
    graph_path = semantic_dir / "residual_candidate_graph.json"
    graph = json.loads(graph_path.read_text(encoding="utf-8"))
    response_path, events_path = _write_semantic_worker_result(
        semantic_dir, _valid_semantic_response(graph)
    )
    semantic_review.validate_semantic_review(
        reconciliation_dir, semantic_dir, graph_path, response_path, events_path
    )
    first_pair = {
        name: (semantic_dir / name).read_bytes()
        for name in (
            "semantic_suggestions_validated.json",
            "semantic_worker_run.json",
        )
    }

    result = semantic_review.validate_semantic_review(
        reconciliation_dir, semantic_dir, graph_path, response_path, events_path
    )

    status = json.loads(
        (semantic_dir / "semantic_review_status.json").read_text(encoding="utf-8")
    )
    assert result["summary"]["decision_count"] == 2
    assert status["status"] == "completed_validated"
    assert {
        name: (semantic_dir / name).read_bytes() for name in first_pair
    } == first_pair


def test_semantic_reprepare_archives_prior_worker_generation(
    tmp_path: Path,
) -> None:
    _, semantic_review, reconciliation_dir, semantic_dir = (
        _prepare_ambiguous_semantic_run(tmp_path)
    )
    semantic_review.prepare_semantic_review(reconciliation_dir, semantic_dir)
    graph = json.loads(
        (semantic_dir / "residual_candidate_graph.json").read_text(encoding="utf-8")
    )
    response_path, events_path = _write_semantic_worker_result(
        semantic_dir, _valid_semantic_response(graph)
    )
    semantic_review.validate_semantic_review(
        reconciliation_dir,
        semantic_dir,
        semantic_dir / "residual_candidate_graph.json",
        response_path,
        events_path,
    )

    prepared = semantic_review.prepare_semantic_review(reconciliation_dir, semantic_dir)

    archive_dir = prepared["archived_generation"]
    assert isinstance(archive_dir, Path)
    assert (archive_dir / "semantic_suggestions_validated.json").is_file()
    assert (archive_dir / "semantic_worker_run.json").is_file()
    assert (archive_dir / "luna_response.json").is_file()
    assert (archive_dir / "luna_events.jsonl").is_file()
    assert not (semantic_dir / "semantic_suggestions_validated.json").exists()
    assert not (semantic_dir / "semantic_worker_run.json").exists()
    status = json.loads(
        (semantic_dir / "semantic_review_status.json").read_text(encoding="utf-8")
    )
    assert status["status"] == "prepared"


def test_semantic_reprepare_failure_removes_completed_marker_first(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    _, semantic_review, reconciliation_dir, semantic_dir = (
        _prepare_ambiguous_semantic_run(tmp_path)
    )
    semantic_review.prepare_semantic_review(reconciliation_dir, semantic_dir)
    graph_path = semantic_dir / "residual_candidate_graph.json"
    graph = json.loads(graph_path.read_text(encoding="utf-8"))
    response_path, events_path = _write_semantic_worker_result(
        semantic_dir, _valid_semantic_response(graph)
    )
    semantic_review.validate_semantic_review(
        reconciliation_dir, semantic_dir, graph_path, response_path, events_path
    )
    real_replace = Path.replace

    def fail_after_status(path: Path, target: Path) -> Path:
        if path.name == "semantic_suggestions_validated.json":
            raise OSError("synthetic archive failure")
        return real_replace(path, target)

    monkeypatch.setattr(Path, "replace", fail_after_status)

    with pytest.raises(OSError, match="synthetic archive failure"):
        semantic_review.prepare_semantic_review(reconciliation_dir, semantic_dir)

    assert not (semantic_dir / "semantic_review_status.json").exists()
    assert (semantic_dir / "semantic_suggestions_validated.json").is_file()


def test_semantic_validate_preflights_both_advisory_outputs(
    tmp_path: Path,
) -> None:
    _, semantic_review, reconciliation_dir, semantic_dir = (
        _prepare_ambiguous_semantic_run(tmp_path)
    )
    semantic_review.prepare_semantic_review(reconciliation_dir, semantic_dir)
    graph = json.loads(
        (semantic_dir / "residual_candidate_graph.json").read_text(encoding="utf-8")
    )
    response_path, events_path = _write_semantic_worker_result(
        semantic_dir, _valid_semantic_response(graph)
    )
    (semantic_dir / "semantic_worker_run.json").mkdir()

    with pytest.raises(ValueError, match="already exists"):
        semantic_review.validate_semantic_review(
            reconciliation_dir,
            semantic_dir,
            semantic_dir / "residual_candidate_graph.json",
            response_path,
            events_path,
        )

    assert not (semantic_dir / "semantic_suggestions_validated.json").exists()


def test_semantic_validate_rejects_file_identity_change_during_snapshot(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    _, semantic_review, reconciliation_dir, semantic_dir = (
        _prepare_ambiguous_semantic_run(tmp_path)
    )
    semantic_review.prepare_semantic_review(reconciliation_dir, semantic_dir)
    graph = json.loads(
        (semantic_dir / "residual_candidate_graph.json").read_text(encoding="utf-8")
    )
    response_path, events_path = _write_semantic_worker_result(
        semantic_dir, _valid_semantic_response(graph)
    )
    real_fstat = semantic_review.os.fstat
    call_count = 0

    def drifting_fstat(file_descriptor: int) -> Any:
        nonlocal call_count
        current = real_fstat(file_descriptor)
        call_count += 1
        if call_count != 2:
            return current
        values = {
            name: getattr(current, name)
            for name in (
                "st_dev",
                "st_ino",
                "st_size",
                "st_mtime_ns",
                "st_nlink",
                "st_mode",
            )
        }
        values["st_mtime_ns"] += 1
        return SimpleNamespace(**values)

    monkeypatch.setattr(semantic_review.os, "fstat", drifting_fstat)

    with pytest.raises(ValueError, match="changed while it was read"):
        semantic_review.validate_semantic_review(
            reconciliation_dir,
            semantic_dir,
            semantic_dir / "residual_candidate_graph.json",
            response_path,
            events_path,
        )


@pytest.mark.parametrize(
    "invalid_case",
    [
        "stale_hash",
        "invented_edge",
        "journal_reuse",
        "missing_bank",
        "non_match_with_journal",
        "unexpected_field",
        "overlong_rationale",
    ],
)
def test_semantic_validate_rejects_invalid_worker_decisions(
    tmp_path: Path,
    invalid_case: str,
) -> None:
    _, semantic_review, reconciliation_dir, semantic_dir = (
        _prepare_ambiguous_semantic_run(tmp_path)
    )
    semantic_review.prepare_semantic_review(reconciliation_dir, semantic_dir)
    graph = json.loads(
        (semantic_dir / "residual_candidate_graph.json").read_text(encoding="utf-8")
    )
    response = _valid_semantic_response(graph)
    decisions = response["component_reviews"][0]["decisions"]
    if invalid_case == "stale_hash":
        response["candidate_graph_sha256"] = "0" * 64
    elif invalid_case == "invented_edge":
        decisions[0]["journal_transaction_id"] = "journal:invented"
    elif invalid_case == "journal_reuse":
        decisions[1]["journal_transaction_id"] = decisions[0]["journal_transaction_id"]
    elif invalid_case == "missing_bank":
        decisions.pop()
    elif invalid_case == "non_match_with_journal":
        decisions[0]["verdict"] = "ambiguous"
    elif invalid_case == "unexpected_field":
        decisions[0]["confidence"] = "high"
    else:
        decisions[0]["rationale"] = "x" * 601
    response_path, events_path = _write_semantic_worker_result(semantic_dir, response)
    authoritative_before = _tree_snapshot(reconciliation_dir)

    with pytest.raises(ValueError):
        semantic_review.validate_semantic_review(
            reconciliation_dir,
            semantic_dir,
            semantic_dir / "residual_candidate_graph.json",
            response_path,
            events_path,
        )

    assert not (semantic_dir / "semantic_suggestions_validated.json").exists()
    assert not (semantic_dir / "semantic_worker_run.json").exists()
    assert _tree_snapshot(reconciliation_dir) == authoritative_before


def test_semantic_validate_rejects_duplicate_json_keys(tmp_path: Path) -> None:
    _, semantic_review, reconciliation_dir, semantic_dir = (
        _prepare_ambiguous_semantic_run(tmp_path)
    )
    semantic_review.prepare_semantic_review(reconciliation_dir, semantic_dir)
    graph = json.loads(
        (semantic_dir / "residual_candidate_graph.json").read_text(encoding="utf-8")
    )
    response = _valid_semantic_response(graph)
    response_path, events_path = _write_semantic_worker_result(semantic_dir, response)
    original = response_path.read_text(encoding="utf-8").strip()
    duplicate = original[:-1] + ',"schema_version":"duplicate"}'
    response_path.write_text(duplicate + "\n", encoding="utf-8")

    with pytest.raises(ValueError, match="duplicate JSON key"):
        semantic_review.validate_semantic_review(
            reconciliation_dir,
            semantic_dir,
            semantic_dir / "residual_candidate_graph.json",
            response_path,
            events_path,
        )


def test_semantic_validate_rejects_worker_tool_events(tmp_path: Path) -> None:
    _, semantic_review, reconciliation_dir, semantic_dir = (
        _prepare_ambiguous_semantic_run(tmp_path)
    )
    semantic_review.prepare_semantic_review(reconciliation_dir, semantic_dir)
    graph = json.loads(
        (semantic_dir / "residual_candidate_graph.json").read_text(encoding="utf-8")
    )
    response = _valid_semantic_response(graph)
    response_path, events_path = _write_semantic_worker_result(
        semantic_dir, response, item_type="command_execution"
    )

    with pytest.raises(ValueError, match="forbidden item type"):
        semantic_review.validate_semantic_review(
            reconciliation_dir,
            semantic_dir,
            semantic_dir / "residual_candidate_graph.json",
            response_path,
            events_path,
        )


@pytest.mark.parametrize(
    "events_builder",
    [
        pytest.param(
            lambda message: [
                {"type": "thread.started", "thread_id": "thread_luna_test"},
                {
                    "type": "item.completed",
                    "item": {
                        "id": "item_0",
                        "type": "agent_message",
                        "text": message,
                    },
                },
                {
                    "type": "turn.completed",
                    "usage": {"input_tokens": 120, "output_tokens": 80},
                },
            ],
            id="missing-turn-start",
        ),
        pytest.param(
            lambda message: [
                {"type": "thread.started", "thread_id": "thread_luna_test"},
                {"type": "turn.started"},
                {"type": "turn.started"},
                {
                    "type": "item.completed",
                    "item": {
                        "id": "item_0",
                        "type": "agent_message",
                        "text": message,
                    },
                },
                {
                    "type": "turn.completed",
                    "usage": {"input_tokens": 120, "output_tokens": 80},
                },
            ],
            id="duplicate-turn-start",
        ),
        pytest.param(
            lambda message: [
                {"type": "turn.started"},
                {"type": "thread.started", "thread_id": "thread_luna_test"},
                {
                    "type": "item.completed",
                    "item": {
                        "id": "item_0",
                        "type": "agent_message",
                        "text": message,
                    },
                },
                {
                    "type": "turn.completed",
                    "usage": {"input_tokens": 120, "output_tokens": 80},
                },
            ],
            id="thread-not-first",
        ),
        pytest.param(
            lambda message: [
                {"type": "thread.started", "thread_id": "thread_luna_test"},
                {"type": "turn.started"},
                {
                    "type": "item.completed",
                    "item": {
                        "id": "item_0",
                        "type": "agent_message",
                        "text": message,
                    },
                },
                {
                    "type": "turn.completed",
                    "usage": {"input_tokens": 120, "output_tokens": 80},
                },
                {
                    "type": "item.completed",
                    "item": {"id": "item_1", "type": "reasoning"},
                },
            ],
            id="item-after-turn-completion",
        ),
        pytest.param(
            lambda message: [
                {"type": "thread.started", "thread_id": "thread_luna_test"},
                {"type": "turn.started"},
                {
                    "type": "item.completed",
                    "item": {
                        "id": "item_0",
                        "type": "agent_message",
                        "text": message,
                    },
                },
                {
                    "type": "item.completed",
                    "item": {
                        "id": "item_1",
                        "type": "agent_message",
                        "text": message,
                    },
                },
                {
                    "type": "turn.completed",
                    "usage": {"input_tokens": 120, "output_tokens": 80},
                },
            ],
            id="multiple-agent-messages",
        ),
        pytest.param(
            lambda message: [
                {"type": "thread.started", "thread_id": "thread_luna_test"},
                {"type": "turn.started"},
                {
                    "type": "item.completed",
                    "item": {"id": "item_1", "type": "reasoning"},
                },
                {
                    "type": "item.completed",
                    "item": {"id": "item_1", "type": "reasoning"},
                },
                {
                    "type": "item.completed",
                    "item": {
                        "id": "item_0",
                        "type": "agent_message",
                        "text": message,
                    },
                },
                {
                    "type": "turn.completed",
                    "usage": {"input_tokens": 120, "output_tokens": 80},
                },
            ],
            id="duplicate-item-completion",
        ),
    ],
)
def test_semantic_validate_rejects_impossible_worker_event_lifecycle(
    tmp_path: Path,
    events_builder: Any,
) -> None:
    _, semantic_review, reconciliation_dir, semantic_dir = (
        _prepare_ambiguous_semantic_run(tmp_path)
    )
    semantic_review.prepare_semantic_review(reconciliation_dir, semantic_dir)
    graph = json.loads(
        (semantic_dir / "residual_candidate_graph.json").read_text(encoding="utf-8")
    )
    response = _valid_semantic_response(graph)
    events = events_builder(json.dumps(response, ensure_ascii=False))
    response_path, events_path = _write_semantic_worker_result(
        semantic_dir,
        response,
        events_override=events,
    )

    with pytest.raises(ValueError, match="Worker"):
        semantic_review.validate_semantic_review(
            reconciliation_dir,
            semantic_dir,
            semantic_dir / "residual_candidate_graph.json",
            response_path,
            events_path,
        )


def test_semantic_validate_requires_pinned_launch_receipt(tmp_path: Path) -> None:
    _, semantic_review, reconciliation_dir, semantic_dir = (
        _prepare_ambiguous_semantic_run(tmp_path)
    )
    semantic_review.prepare_semantic_review(reconciliation_dir, semantic_dir)
    graph = json.loads(
        (semantic_dir / "residual_candidate_graph.json").read_text(encoding="utf-8")
    )
    response_path, events_path = _write_semantic_worker_result(
        semantic_dir,
        _valid_semantic_response(graph),
    )
    (semantic_dir / "luna_launch_receipt.json").unlink()

    with pytest.raises(ValueError, match="launch_receipt.json does not exist"):
        semantic_review.validate_semantic_review(
            reconciliation_dir,
            semantic_dir,
            semantic_dir / "residual_candidate_graph.json",
            response_path,
            events_path,
        )


def test_semantic_validate_rejects_rehashed_forged_launch_boundary(
    tmp_path: Path,
) -> None:
    _, semantic_review, reconciliation_dir, semantic_dir = (
        _prepare_ambiguous_semantic_run(tmp_path)
    )
    semantic_review.prepare_semantic_review(reconciliation_dir, semantic_dir)
    graph = json.loads(
        (semantic_dir / "residual_candidate_graph.json").read_text(encoding="utf-8")
    )
    response_path, events_path = _write_semantic_worker_result(
        semantic_dir,
        _valid_semantic_response(graph),
    )
    receipt_path = semantic_dir / "luna_launch_receipt.json"
    receipt = json.loads(receipt_path.read_text(encoding="utf-8"))
    receipt["boundary"]["codex_sha256"] = "0" * 64
    receipt_content = dict(receipt)
    receipt_content.pop("content_sha256")
    receipt["content_sha256"] = semantic_review.canonical_json_sha256(receipt_content)
    receipt_path.write_text(
        json.dumps(receipt, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    with pytest.raises(ValueError, match="codex_sha256"):
        semantic_review.validate_semantic_review(
            reconciliation_dir,
            semantic_dir,
            semantic_dir / "residual_candidate_graph.json",
            response_path,
            events_path,
        )


def test_semantic_run_worker_launches_separate_pinned_luna_capsule(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    _, semantic_review, reconciliation_dir, semantic_dir = (
        _prepare_ambiguous_semantic_run(tmp_path)
    )
    semantic_review.prepare_semantic_review(reconciliation_dir, semantic_dir)
    graph = json.loads(
        (semantic_dir / "residual_candidate_graph.json").read_text(encoding="utf-8")
    )
    response = _valid_semantic_response(graph)
    events = [
        {"type": "thread.started", "thread_id": "thread_luna_mock"},
        {"type": "turn.started"},
        {
            "type": "item.completed",
            "item": {
                "id": "item_0",
                "type": "agent_message",
                "text": json.dumps(response, ensure_ascii=False),
            },
        },
        {
            "type": "turn.completed",
            "usage": {"input_tokens": 120, "output_tokens": 80},
        },
    ]
    events_bytes = "".join(
        json.dumps(event, ensure_ascii=False) + "\n" for event in events
    ).encode("utf-8")
    captured = _mock_semantic_worker_runtime(
        monkeypatch,
        semantic_review,
        tmp_path,
        {
            "return_code": 0,
            "stdout": events_bytes,
            "stderr": b"bounded worker warning\n",
            "duration_ms": 42,
        },
    )
    authoritative_before = _tree_snapshot(reconciliation_dir)

    result = semantic_review.run_semantic_worker(
        reconciliation_dir,
        semantic_dir,
        semantic_dir / "residual_candidate_graph.json",
    )

    command = captured["command"]
    receipt = json.loads(
        (semantic_dir / "luna_launch_receipt.json").read_text(encoding="utf-8")
    )
    status = json.loads(
        (semantic_dir / "semantic_review_status.json").read_text(encoding="utf-8")
    )
    assert command[0] == "/usr/bin/sandbox-exec"
    assert command[command.index("--model") + 1] == "gpt-5.6-luna"
    assert 'model_reasoning_effort="max"' in command
    assert "--disable" in command
    assert "shell_tool" in command
    assert "unified_exec" in command
    assert captured["kwargs"]["stdin_path"].name == "prompt.stdin"
    assert result["main_chat_model_change"] is False
    assert receipt["runtime_attestation"]["main_chat_model_change"] is False
    assert receipt["jsonl_observation"]["visibility_complete"] is False
    assert receipt["jsonl_observation"]["tool_use_absence_observed"] is False
    assert status["status"] == "worker_completed_pending_validation"
    assert (semantic_dir / "luna_response.json").is_file()
    assert (semantic_dir / "luna_events.jsonl").read_bytes() == events_bytes
    assert not list(semantic_dir.glob(".luna-worker-capsule.*"))
    assert _tree_snapshot(reconciliation_dir) == authoritative_before


def test_semantic_run_worker_fails_closed_on_unqualified_platform(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    semantic_review = load_semantic_review()
    monkeypatch.setattr(semantic_review.platform, "system", lambda: "Linux")

    with pytest.raises(ValueError, match="only on macOS"):
        semantic_review._darwin_build_version()


def test_semantic_run_worker_does_not_launch_after_failed_boundary_canary(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    _, semantic_review, reconciliation_dir, semantic_dir = (
        _prepare_ambiguous_semantic_run(tmp_path)
    )
    semantic_review.prepare_semantic_review(reconciliation_dir, semantic_dir)
    captured = _mock_semantic_worker_runtime(
        monkeypatch,
        semantic_review,
        tmp_path,
        {
            "return_code": 0,
            "stdout": b"",
            "stderr": b"",
            "duration_ms": 1,
        },
    )

    def reject_canary(**kwargs: Any) -> dict[str, Any]:
        raise ValueError("outside-read canary failed")

    monkeypatch.setattr(semantic_review, "_qualification_canaries", reject_canary)

    with pytest.raises(ValueError, match="outside-read canary failed"):
        semantic_review.run_semantic_worker(
            reconciliation_dir,
            semantic_dir,
            semantic_dir / "residual_candidate_graph.json",
        )

    assert captured == {}
    assert not (semantic_dir / "luna_launch_receipt.json").exists()
    assert not list(semantic_dir.glob(".luna-worker-capsule.*"))


def test_semantic_run_worker_rejects_nonzero_child_without_partial_outputs(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    _, semantic_review, reconciliation_dir, semantic_dir = (
        _prepare_ambiguous_semantic_run(tmp_path)
    )
    semantic_review.prepare_semantic_review(reconciliation_dir, semantic_dir)
    _mock_semantic_worker_runtime(
        monkeypatch,
        semantic_review,
        tmp_path,
        {
            "return_code": 1,
            "stdout": b"",
            "stderr": b"worker failed",
            "duration_ms": 3,
        },
    )

    with pytest.raises(ValueError, match="nonzero"):
        semantic_review.run_semantic_worker(
            reconciliation_dir,
            semantic_dir,
            semantic_dir / "residual_candidate_graph.json",
        )

    assert not (semantic_dir / "luna_response.json").exists()
    assert not (semantic_dir / "luna_events.jsonl").exists()
    assert not (semantic_dir / "luna_stderr.log").exists()
    assert not (semantic_dir / "luna_launch_receipt.json").exists()


def test_semantic_run_worker_rolls_back_partial_atomic_publication(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    _, semantic_review, reconciliation_dir, semantic_dir = (
        _prepare_ambiguous_semantic_run(tmp_path)
    )
    semantic_review.prepare_semantic_review(reconciliation_dir, semantic_dir)
    graph = json.loads(
        (semantic_dir / "residual_candidate_graph.json").read_text(encoding="utf-8")
    )
    response = _valid_semantic_response(graph)
    events = [
        {"type": "thread.started", "thread_id": "thread_luna_mock"},
        {"type": "turn.started"},
        {
            "type": "item.completed",
            "item": {
                "id": "item_0",
                "type": "agent_message",
                "text": json.dumps(response, ensure_ascii=False),
            },
        },
        {
            "type": "turn.completed",
            "usage": {"input_tokens": 120, "output_tokens": 80},
        },
    ]
    events_bytes = "".join(
        json.dumps(event, ensure_ascii=False) + "\n" for event in events
    ).encode("utf-8")
    _mock_semantic_worker_runtime(
        monkeypatch,
        semantic_review,
        tmp_path,
        {
            "return_code": 0,
            "stdout": events_bytes,
            "stderr": b"",
            "duration_ms": 4,
        },
    )
    real_replace = Path.replace

    def fail_receipt_publish(path: Path, target: Path) -> Path:
        if path.name == "luna_launch_receipt.json":
            raise OSError("synthetic receipt publication failure")
        return real_replace(path, target)

    monkeypatch.setattr(Path, "replace", fail_receipt_publish)

    with pytest.raises(OSError, match="synthetic receipt publication failure"):
        semantic_review.run_semantic_worker(
            reconciliation_dir,
            semantic_dir,
            semantic_dir / "residual_candidate_graph.json",
        )

    assert not (semantic_dir / "luna_response.json").exists()
    assert not (semantic_dir / "luna_events.jsonl").exists()
    assert not (semantic_dir / "luna_stderr.log").exists()
    assert not (semantic_dir / "luna_launch_receipt.json").exists()


def test_semantic_validate_cli_records_worker_failure_status(tmp_path: Path) -> None:
    _, semantic_review, reconciliation_dir, semantic_dir, context_path, context = (
        _prepare_managed_ambiguous_semantic_run(tmp_path)
    )
    semantic_review.prepare_semantic_review(
        reconciliation_dir,
        semantic_dir,
        client_engagement=context,
    )

    completed = subprocess.run(
        [
            sys.executable,
            str(SEMANTIC_REVIEW_PATH),
            "validate",
            str(reconciliation_dir),
            "--candidate-graph",
            str(semantic_dir / "residual_candidate_graph.json"),
            "--output-dir",
            str(semantic_dir),
            "--response",
            str(semantic_dir / "luna_response.json"),
            "--events",
            str(semantic_dir / "luna_events.jsonl"),
            "--client-engagement",
            str(context_path),
        ],
        capture_output=True,
        check=False,
        text=True,
    )

    status = json.loads(
        (semantic_dir / "semantic_review_status.json").read_text(encoding="utf-8")
    )
    assert completed.returncode == 2
    assert "SEMANTIC_WORKER_VALIDATION_FAILED" in completed.stderr
    assert status["status"] == "worker_failed"
    assert status["failure_reason"] == "worker_command_or_validation_failed"


def test_semantic_completed_status_survives_invalid_validation_retry(
    tmp_path: Path,
) -> None:
    _, semantic_review, reconciliation_dir, semantic_dir, context_path, context = (
        _prepare_managed_ambiguous_semantic_run(tmp_path)
    )
    semantic_review.prepare_semantic_review(
        reconciliation_dir,
        semantic_dir,
        client_engagement=context,
    )
    graph_path = semantic_dir / "residual_candidate_graph.json"
    graph = json.loads(graph_path.read_text(encoding="utf-8"))
    response_path, events_path = _write_semantic_worker_result(
        semantic_dir,
        _valid_semantic_response(graph),
    )
    semantic_review.validate_semantic_review(
        reconciliation_dir,
        semantic_dir,
        graph_path,
        response_path,
        events_path,
        client_engagement=context,
    )
    response_path.write_text("{}\n", encoding="utf-8")

    completed = subprocess.run(
        [
            sys.executable,
            str(SEMANTIC_REVIEW_PATH),
            "validate",
            str(reconciliation_dir),
            "--candidate-graph",
            str(graph_path),
            "--output-dir",
            str(semantic_dir),
            "--response",
            str(response_path),
            "--events",
            str(events_path),
            "--client-engagement",
            str(context_path),
        ],
        capture_output=True,
        check=False,
        text=True,
    )

    status = json.loads(
        (semantic_dir / "semantic_review_status.json").read_text(encoding="utf-8")
    )
    assert completed.returncode == 2
    assert status["status"] == "completed_validated"
    assert (semantic_dir / "semantic_suggestions_validated.json").is_file()
    assert (semantic_dir / "semantic_worker_run.json").is_file()


def test_semantic_source_tamper_does_not_forge_bound_worker_failure_status(
    tmp_path: Path,
) -> None:
    _, semantic_review, reconciliation_dir, semantic_dir, context_path, context = (
        _prepare_managed_ambiguous_semantic_run(tmp_path)
    )
    semantic_review.prepare_semantic_review(
        reconciliation_dir,
        semantic_dir,
        client_engagement=context,
    )
    unmatched_path = reconciliation_dir / "unmatched_bank.csv"
    unmatched_path.write_bytes(unmatched_path.read_bytes() + b"\n")

    completed = subprocess.run(
        [
            sys.executable,
            str(SEMANTIC_REVIEW_PATH),
            "validate",
            str(reconciliation_dir),
            "--candidate-graph",
            str(semantic_dir / "residual_candidate_graph.json"),
            "--output-dir",
            str(semantic_dir),
            "--response",
            str(semantic_dir / "luna_response.json"),
            "--events",
            str(semantic_dir / "luna_events.jsonl"),
            "--client-engagement",
            str(context_path),
        ],
        capture_output=True,
        check=False,
        text=True,
    )

    status = json.loads(
        (semantic_dir / "semantic_review_status.json").read_text(encoding="utf-8")
    )
    assert completed.returncode == 2
    assert status["status"] == "prepared"
    assert status["failure_reason"] is None


def test_skill_and_scripts_keep_codex_as_the_review_layer() -> None:
    skill_text = (
        ROOT
        / "plugins"
        / "journal-bank-reconciliation"
        / "skills"
        / "journal-bank-reconciliation"
        / "SKILL.md"
    ).read_text(encoding="utf-8")
    script_text = "\n".join(
        path.read_text(encoding="utf-8") for path in SCRIPT_DIR.glob("*.py")
    )

    assert "The user should not interact directly with CLI scripts" in skill_text
    assert "must not make direct OpenAI API calls" in skill_text
    assert "scripts/check_dependencies.py" in skill_text
    assert "it`, `en`, `fr`, `de`, and `es`" in skill_text
    assert "missing deterministic extraction script" in skill_text
    assert "Keep the improvement note local to chat or run artifacts." in skill_text
    assert "validate_journal_bank_review" in skill_text
    assert "render_journal_bank_review" in skill_text
    assert "Codex-Only Luna Max Residual Resolution Funnel" in skill_text
    assert "semantic_review.py run-worker" in skill_text
    assert "journal_bank.luna_seatbelt_capsule.v1" in skill_text
    assert "current chat unchanged" in skill_text
    assert "Codex JSONL visibility is incomplete" in skill_text
    assert "luna_launch_receipt.json" in skill_text
    assert "copy or reconstruct its underlying `codex exec` command" in skill_text
    assert "codex exec \\" not in skill_text
    assert "Raw worker output is advisory until the validator accepts it" in skill_text
    assert "human_review_queue.json" in skill_text
    assert "modules.llm" not in script_text
    assert "model_router" not in script_text
    assert "openai" not in script_text.lower()


def test_static_page_exposes_five_language_switch() -> None:
    page = (
        ROOT / "static" / "shared" / "journal-bank-reconciliation" / "index.html"
    ).read_text(encoding="utf-8")

    for snippet in (
        'data-lang="it"',
        'data-lang="en"',
        'data-lang="fr"',
        'data-lang="de"',
        'data-lang="es"',
        "Porta banca e contabilità in una riconciliazione con eccezioni visibili.",
        "Bring bank and accounting into one reconciliation with visible exceptions.",
        "Réunissez banque et comptabilité dans un rapprochement aux exceptions visibles.",
        "Reúna banco y contabilidad en una conciliación con excepciones visibles.",
        "Führen Sie Bank und Buchhaltung in einer Abstimmung mit sichtbaren Ausnahmen zusammen.",
    ):
        assert snippet in page


def test_journal_bank_mcp_server_validates_renders_and_applies_review_payload(
    tmp_path: Path,
) -> None:
    output_dir, run_intake, review_payload, ui_decisions, final_artifacts = (
        _prepare_sealed_mcp_review_run(tmp_path)
    )
    matches_path = output_dir / "reconciliation_matches.csv"
    workbook_path = output_dir / "journal_bank_reconciliation.xlsx"
    matched_item = next(
        item for item in review_payload["items"] if item["item_type"] == "matched_pair"
    )
    unmatched_bank_item = next(
        item
        for item in review_payload["items"]
        if item["item_type"] == "unmatched_bank"
    )
    decisions = [
        {
            "item_id": matched_item["id"],
            "action": "edit",
            "edit_value": "Reviewer accepted reference match.",
        },
        {
            "item_id": unmatched_bank_item["id"],
            "action": "request_more_documents",
            "reviewer_note": "Need accounting support for FEE9.",
            "requested_documents": ["ledger_support_FEE9.pdf"],
        },
        *[
            {"item_id": item["id"], "action": "accept"}
            for item in review_payload["items"]
            if item["id"] not in {matched_item["id"], unmatched_bank_item["id"]}
        ],
    ]
    messages: list[dict[str, object]] = [
        {"jsonrpc": "2.0", "id": 1, "method": "tools/list"},
        {
            "jsonrpc": "2.0",
            "id": 2,
            "method": "tools/call",
            "params": {
                "name": "validate_journal_bank_review",
                "arguments": {"review_payload": review_payload},
            },
        },
        {
            "jsonrpc": "2.0",
            "id": 3,
            "method": "tools/call",
            "params": {
                "name": "render_journal_bank_review",
                "arguments": {
                    "run_intake": run_intake,
                    "review_payload": review_payload,
                    "ui_decisions": ui_decisions,
                },
            },
        },
        {"jsonrpc": "2.0", "id": 4, "method": "resources/list"},
        {
            "jsonrpc": "2.0",
            "id": 5,
            "method": "resources/read",
            "params": {"uri": "ui://widget/journal-bank-review.html"},
        },
        {
            "jsonrpc": "2.0",
            "id": 6,
            "method": "tools/call",
            "params": {
                "name": "save_journal_bank_decisions",
                "arguments": {
                    "run_intake": run_intake,
                    "review_payload": review_payload,
                    "ui_decisions": ui_decisions,
                    "decisions": decisions,
                },
            },
        },
        {
            "jsonrpc": "2.0",
            "id": 7,
            "method": "tools/call",
            "params": {
                "name": "apply_journal_bank_decisions",
                "arguments": {
                    "run_intake": run_intake,
                    "review_payload": review_payload,
                    "ui_decisions": ui_decisions,
                    "final_artifacts": final_artifacts,
                    "decisions": decisions,
                },
            },
        },
    ]

    responses = {response["id"]: response for response in _call_mcp_server(messages)}

    tool_names = {tool["name"] for tool in responses[1]["result"]["tools"]}
    assert {
        "validate_journal_bank_review",
        "render_journal_bank_review",
        "save_journal_bank_decisions",
        "apply_journal_bank_decisions",
    } <= tool_names
    validate_result = responses[2]["result"]["structuredContent"]
    assert validate_result["ok"] is True
    assert validate_result["item_count"] == review_payload["item_count"]
    render_result = responses[3]["result"]
    assert render_result["structuredContent"]["widget_type"] == "journal_bank_review"
    assert (
        render_result["_meta"]["openai/outputTemplate"]
        == "ui://widget/journal-bank-review.html"
    )
    resource_uris = {
        resource["uri"] for resource in responses[4]["result"]["resources"]
    }
    assert "ui://widget/journal-bank-review.html" in resource_uris
    widget_html = responses[5]["result"]["contents"][0]["text"]
    assert "Journal-Bank Review" in widget_html
    save_result = responses[6]["result"]["structuredContent"]
    assert save_result["ok"] is True
    assert save_result["persisted"] is True
    assert save_result["decision_count"] == len(decisions)
    written_decisions = json.loads((output_dir / "ui_decisions.json").read_text())
    assert written_decisions["decisions"][0]["edit_value"] == (
        "Reviewer accepted reference match."
    )
    assert written_decisions["decisions"][1]["requested_documents"] == [
        "ledger_support_FEE9.pdf"
    ]
    apply_result = responses[7]["result"]["structuredContent"]
    assert apply_result["ok"] is True
    assert apply_result["persisted"] is True
    assert apply_result["run_intake_path"] == str(output_dir / "run_intake.json")
    assert apply_result["decision_count"] == len(decisions)
    assert apply_result["blocker_count"] == 1
    assert apply_result["structured_update_count"] == 1
    assert apply_result["native_regeneration_count"] == 0
    assert apply_result["native_regenerated_count"] == 1
    assert apply_result["application_status"] == "blocked"
    assert "Reviewer accepted reference match." in matches_path.read_text(
        encoding="utf-8"
    )
    applied = json.loads((output_dir / "applied_decisions.json").read_text())
    assert applied["effects"][0]["structured_update"] == {
        "id_field": "bank_transaction_id",
        "record_id": matched_item["data"]["target_record_id"],
        "target_field": "review_note",
        "records_key": None,
        "updated_rows": 1,
    }
    assert applied["effects"][0]["derived_native_regeneration_paths"] == [
        "journal_bank_reconciliation.xlsx"
    ]
    assert applied["effects"][0]["requires_native_regeneration"] is False
    assert applied["effects"][0]["native_regeneration_status"] == "regenerated"
    assert applied["structured_update_paths"] == ["reconciliation_matches.csv"]
    assert applied["native_regeneration_paths"] == []
    assert applied["native_regenerated_paths"] == ["journal_bank_reconciliation.xlsx"]
    workbook = openpyxl.load_workbook(workbook_path)
    match_sheet = workbook["matches"]
    review_note_column = next(
        cell.column for cell in match_sheet[1] if cell.value == "review_note"
    )
    review_note_cell = match_sheet.cell(row=2, column=review_note_column)
    assert review_note_cell.value == "Reviewer accepted reference match."
    final_artifacts = json.loads((output_dir / "final_artifacts.json").read_text())
    assert final_artifacts["review_application"]["structured_update_count"] == 1
    assert final_artifacts["review_application"]["structured_update_paths"] == [
        "reconciliation_matches.csv"
    ]
    assert final_artifacts["review_application"]["native_regeneration_paths"] == []
    assert final_artifacts["review_application"]["native_regenerated_paths"] == [
        "journal_bank_reconciliation.xlsx"
    ]
    outputs_by_path = {output["path"]: output for output in final_artifacts["outputs"]}
    assert outputs_by_path["journal_bank_reconciliation.xlsx"]["status"] == (
        "updated_from_review"
    )
    assert (
        outputs_by_path["journal_bank_reconciliation.xlsx"]["native_regenerated"]
        is True
    )
    assert outputs_by_path["journal_bank_reconciliation.xlsx"]["source_artifact"] == (
        "reconciliation_matches.csv"
    )
    assert outputs_by_path["journal_bank_reconciliation.xlsx"]["source_row_count"] == 1
    assert outputs_by_path["journal_bank_reconciliation.xlsx"][
        "required_sheet_headers"
    ]["matches"] == [cell.value for cell in match_sheet[1]]
    assert outputs_by_path["journal_bank_reconciliation.xlsx"]["required_cells"][
        "matches"
    ] == {review_note_cell.coordinate: "Reviewer accepted reference match."}
    assert {
        "reconciliation_matches.csv",
        f"revisions/originals/reconciliation_matches__{matched_item['id']}.csv",
        (
            "revisions/originals/"
            f"journal_bank_reconciliation__{matched_item['id']}.xlsx"
        ),
    } <= {output["path"] for output in final_artifacts["outputs"]}
    run_intake = json.loads((output_dir / "run_intake.json").read_text())
    review_apply_steps = [
        step
        for step in run_intake["execution_trace"]
        if step["kind"] == "deterministic_review_apply"
    ]
    assert len(review_apply_steps) == 1
    assert {
        "applied_decisions.json",
        "final_artifacts.json",
        "journal_bank_reconciliation.xlsx",
        "reconciliation_matches.csv",
        (
            "revisions/originals/"
            f"journal_bank_reconciliation__{matched_item['id']}.xlsx"
        ),
        f"revisions/originals/reconciliation_matches__{matched_item['id']}.csv",
        "ui_decisions.json",
    } <= set(review_apply_steps[0]["outputs"])
    contract = validate_contract(
        output_dir,
        strict_data_posture=True,
        strict_execution_trace=True,
        strict_output_paths=True,
        strict_output_content=True,
    )
    assert contract.ok is True, contract.errors


def test_journal_bank_mcp_closed_review_uses_validated_canonical_readiness(
    tmp_path: Path,
) -> None:
    output_dir = tmp_path / "run"
    output_dir, run_intake, review_payload, _, final_artifacts = (
        _prepare_closed_mcp_review_run(output_dir, portable=True)
    )
    decisions = [
        {"item_id": item["id"], "action": "accept"} for item in review_payload["items"]
    ]

    response = _call_mcp_server(
        [
            {
                "jsonrpc": "2.0",
                "id": 1,
                "method": "tools/call",
                "params": {
                    "name": "apply_journal_bank_decisions",
                    "arguments": {
                        "client_engagement": _customer_context_path(
                            output_dir
                        ).as_posix(),
                        "run_intake": run_intake,
                        "review_payload": review_payload,
                        "final_artifacts": final_artifacts,
                        "decisions": decisions,
                    },
                },
            }
        ]
    )[0]["result"]

    assert response["isError"] is False, response
    result = response["structuredContent"]
    assert result["application_status"] == "final_ready"
    assert result["assurance_report_ready"] is True
    assert (
        json.loads((output_dir / "assurance_gates.json").read_text())["report_ready"]
        is True
    )
    assert (output_dir / "assurance_envelope.reviewed.json").is_file()
    initial_envelope = json.loads(
        (output_dir / "assurance_envelope.json").read_text(encoding="utf-8")
    )
    reviewed_envelope = json.loads(
        (output_dir / "assurance_envelope.reviewed.json").read_text(encoding="utf-8")
    )
    expected_refs = [
        artifact_id for artifact_id, _, _ in load_core().IMPLEMENTATION_ARTIFACT_SPECS
    ]
    assert initial_envelope["implementation_artifact_refs"] == expected_refs
    assert reviewed_envelope["implementation_artifact_refs"] == expected_refs
    assert [
        receipt
        for receipt in reviewed_envelope["artifact_receipts"]
        if receipt["role"] == "implementation"
    ] == [
        receipt
        for receipt in initial_envelope["artifact_receipts"]
        if receipt["role"] == "implementation"
    ]


def test_journal_bank_mcp_delayed_native_regeneration_allows_modified_timestamp(
    tmp_path: Path,
) -> None:
    output_dir = tmp_path / "run"
    output_dir, run_intake, review_payload, _, final_artifacts = (
        _prepare_closed_mcp_review_run(output_dir)
    )
    matched_item = next(
        item for item in review_payload["items"] if item["item_type"] == "matched_pair"
    )
    decisions = [
        (
            {
                "item_id": item["id"],
                "action": "edit",
                "edit_value": "Reviewed closed reconciliation.",
            }
            if item["id"] == matched_item["id"]
            else {"item_id": item["id"], "action": "accept"}
        )
        for item in review_payload["items"]
    ]
    time.sleep(2.2)

    response = _call_mcp_server(
        [
            {
                "jsonrpc": "2.0",
                "id": 1,
                "method": "tools/call",
                "params": {
                    "name": "apply_journal_bank_decisions",
                    "arguments": {
                        "run_intake": run_intake,
                        "review_payload": review_payload,
                        "final_artifacts": final_artifacts,
                        "decisions": decisions,
                    },
                },
            }
        ]
    )[0]["result"]

    assert response["isError"] is False, response
    result = response["structuredContent"]
    assert result["application_status"] == "final_ready"
    assert result["native_regenerated_count"] == 1


def test_journal_bank_mcp_complete_unresolved_review_remains_gate_bound(
    tmp_path: Path,
) -> None:
    output_dir = tmp_path / "run"
    output_dir, run_intake, review_payload, _, final_artifacts = (
        _prepare_sealed_mcp_review_run(output_dir)
    )
    decisions = [
        {"item_id": item["id"], "action": "accept"} for item in review_payload["items"]
    ]

    response = _call_mcp_server(
        [
            {
                "jsonrpc": "2.0",
                "id": 1,
                "method": "tools/call",
                "params": {
                    "name": "apply_journal_bank_decisions",
                    "arguments": {
                        "run_intake": run_intake,
                        "review_payload": review_payload,
                        "final_artifacts": final_artifacts,
                        "decisions": decisions,
                    },
                },
            }
        ]
    )[0]["result"]

    assert response["isError"] is False, response
    result = response["structuredContent"]
    assert result["application_status"] == "blocked"
    assert result["assurance_report_ready"] is False
    gates = json.loads((output_dir / "assurance_gates.json").read_text())
    assert gates["gates"]["semantic_review"]["status"] == "passed"
    assert gates["gates"]["reporting"]["status"] == "blocked"
    assert gates["report_ready"] is False
    assert not (output_dir / "assurance_envelope.reviewed.json").exists()


@pytest.mark.parametrize(
    "forgery_mode",
    [
        "reviewed_envelope_limitations",
        "reviewed_envelope_implementation",
        "final_output_status",
        "nonready_reviewed_envelope",
    ],
)
def test_journal_bank_mcp_child_cannot_author_assurance_envelope_fields(
    monkeypatch: Any,
    tmp_path: Path,
    forgery_mode: str,
) -> None:
    output_dir = tmp_path / "run"
    if forgery_mode == "nonready_reviewed_envelope":
        output_dir, run_intake, review_payload, _, final_artifacts = (
            _prepare_sealed_mcp_review_run(output_dir)
        )
        item = next(
            entry
            for entry in review_payload["items"]
            if entry["item_type"] == "matched_pair"
        )
        decisions = [{"item_id": item["id"], "action": "accept"}]
    else:
        output_dir, run_intake, review_payload, _, final_artifacts = (
            _prepare_closed_mcp_review_run(output_dir)
        )
        decisions = [
            {"item_id": item["id"], "action": "accept"}
            for item in review_payload["items"]
        ]
    before = _tree_snapshot(output_dir)
    wrapper = tmp_path / "python-journal-bank-envelope-forgery"
    wrapper.write_text(
        "\n".join(
            [
                "#!/usr/bin/env python3",
                "import hashlib",
                "import json",
                "import os",
                "import subprocess",
                "import sys",
                "from pathlib import Path",
                f"real = {str(Path(sys.executable))!r}",
                *_customer_run_preflight_passthrough_script_lines(),
                "completed = subprocess.run([real, *sys.argv[1:]], capture_output=True)",
                "if completed.returncode != 0:",
                "    sys.stdout.buffer.write(completed.stdout)",
                "    sys.stderr.buffer.write(completed.stderr)",
                "    raise SystemExit(completed.returncode)",
                'if "--preflight-only" in sys.argv:',
                "    sys.stdout.buffer.write(completed.stdout)",
                "    raise SystemExit(0)",
                'mode = os.environ["JB_TEST_ENVELOPE_FORGERY"]',
                'output_dir = Path(sys.argv[sys.argv.index("--output-dir") + 1])',
                'envelope_path = output_dir / "assurance_envelope.reviewed.json"',
                'if mode == "nonready_reviewed_envelope":',
                '    envelope_path.write_text(\'{"forged": "Traceback /Users/private/nonready-envelope.csv"}\\n\')',
                "else:",
                "    envelope = json.loads(envelope_path.read_text())",
                '    if mode == "reviewed_envelope_limitations":',
                '        envelope["limitations"] = ["Traceback /Users/private/envelope-secret.csv"]',
                '    if mode == "reviewed_envelope_implementation":',
                "        implementation = next(",
                '            value for value in envelope["artifact_receipts"]',
                '            if value["artifact_id"] == "implementation.plugin.scripts.apply_review_edits_py"',
                "        )",
                "        script_path = Path(next(",
                "            value for value in sys.argv[1:]",
                '            if value.endswith(".py")',
                "        ))",
                '        false_path = script_path.resolve().parents[1] / "README.md"',
                "        false_bytes = false_path.read_bytes()",
                '        implementation["path"] = "README.md"',
                '        implementation["byte_count"] = len(false_bytes)',
                '        implementation["sha256"] = hashlib.sha256(false_bytes).hexdigest()',
                '    if mode == "final_output_status":',
                '        final_path = output_dir / "final_artifacts.json"',
                "        final = json.loads(final_path.read_text())",
                "        applied_output = next(",
                '            value for value in final["outputs"]',
                '            if value["path"] == "applied_decisions.json"',
                "        )",
                '        applied_output["status"] = "Traceback /Users/private/output-secret.csv"',
                "        final_path.write_text(",
                '            json.dumps(final, indent=2, sort_keys=True) + "\\n"',
                "        )",
                '        for receipt in envelope["artifact_receipts"]:',
                '            if receipt["root_id"] != "run":',
                "                continue",
                '            artifact = output_dir / receipt["path"]',
                "            content = artifact.read_bytes()",
                '            receipt["byte_count"] = len(content)',
                '            receipt["sha256"] = hashlib.sha256(content).hexdigest()',
                "    content = dict(envelope)",
                '    content.pop("content_sha256")',
                "    canonical = json.dumps(",
                "        content, ensure_ascii=False, sort_keys=True, separators=(',', ':')",
                "    ).encode()",
                '    envelope["content_sha256"] = hashlib.sha256(canonical).hexdigest()',
                "    envelope_path.write_text(",
                '        json.dumps(envelope, indent=2, sort_keys=True) + "\\n"',
                "    )",
                'receipts_path = output_dir / "artifact_receipts.json"',
                "receipts = json.loads(receipts_path.read_text())",
                "if not any(",
                '    value["path"] == "assurance_envelope.reviewed.json"',
                '    for value in receipts["output_receipts"]',
                "):",
                '    receipts["output_receipts"].append({',
                '        "schema_version": "vera.artifact_receipt.v1",',
                '        "artifact_id": "output.assurance_envelope_reviewed_json",',
                '        "root_id": "run",',
                '        "role": "journal-bank reconciliation assurance_envelope.reviewed",',
                '        "path": "assurance_envelope.reviewed.json",',
                '        "byte_count": 0,',
                '        "sha256": "0" * 64,',
                "    })",
                'receipts["output_receipts"].sort(key=lambda value: value["path"])',
                'for receipt in receipts["output_receipts"]:',
                '    artifact = output_dir / receipt["path"]',
                "    if artifact.is_file():",
                "        content = artifact.read_bytes()",
                '        receipt["byte_count"] = len(content)',
                '        receipt["sha256"] = hashlib.sha256(content).hexdigest()',
                "receipts_path.write_text(",
                '    json.dumps(receipts, indent=2, sort_keys=True) + "\\n"',
                ")",
                "sys.stdout.buffer.write(completed.stdout)",
                "raise SystemExit(0)",
                "",
            ]
        ),
        encoding="utf-8",
    )
    wrapper.chmod(0o700)
    monkeypatch.setenv("PYTHON", wrapper.as_posix())
    monkeypatch.setenv("JB_TEST_ENVELOPE_FORGERY", forgery_mode)

    response = _call_mcp_server(
        [
            {
                "jsonrpc": "2.0",
                "id": 1,
                "method": "tools/call",
                "params": {
                    "name": "apply_journal_bank_decisions",
                    "arguments": {
                        "run_intake": run_intake,
                        "review_payload": review_payload,
                        "final_artifacts": final_artifacts,
                        "decisions": decisions,
                    },
                },
            }
        ]
    )[0]["result"]

    assert response["isError"] is True
    assert response["structuredContent"]["error"] == (
        "Journal-Bank review application returned an invalid result."
    )
    assert _tree_snapshot(output_dir) == before
    assert list(output_dir.parent.glob(".journal-bank-apply-*")) == []


def test_journal_bank_mcp_apply_rejects_unsealed_output_before_mutation(
    tmp_path: Path,
) -> None:
    matches_path = tmp_path / "reconciliation_matches.csv"
    original_matches = (
        "status,stage,bank_transaction_id,journal_transaction_id,review_note\n"
        "matched,reference,bank:1,journal:1,\n"
    )
    matches_path.write_text(original_matches, encoding="utf-8")
    review_payload = {
        "schema_version": "1.0",
        "plugin": "journal-bank-reconciliation",
        "workflow": "journal-bank-reconciliation",
        "run_id": "journal-bank-unsealed-run",
        "review_type": "journal_bank_reconciliation_review",
        "items": [
            {
                "id": "matched-pair-1",
                "item_type": "matched_pair",
                "title": "Unsealed match",
                "output_path": "reconciliation_matches.csv",
                "allowed_actions": ["accept", "edit", "mark_unclear", "skip"],
                "recommended_action": "accept",
                "data": {
                    "target_artifact": "reconciliation_matches.csv",
                    "target_id_field": "bank_transaction_id",
                    "target_record_id": "bank:1",
                    "target_field": "review_note",
                },
            }
        ],
        "item_count": 1,
        "status": "ready_for_review",
    }
    messages = [
        {
            "jsonrpc": "2.0",
            "id": 1,
            "method": "tools/call",
            "params": {
                "name": "apply_journal_bank_decisions",
                "arguments": {
                    "run_intake": {
                        "run_id": review_payload["run_id"],
                        "language": "en",
                        "output_dir": tmp_path.as_posix(),
                    },
                    "review_payload": review_payload,
                    "decisions": [
                        {
                            "item_id": "matched-pair-1",
                            "action": "edit",
                            "edit_value": "This must not be written.",
                        }
                    ],
                },
            },
        }
    ]

    response = _call_mcp_server(messages)[0]
    result = response["result"]["structuredContent"]

    assert response["result"]["isError"] is True
    assert result["ok"] is False
    assert isinstance(result["error"], str)
    assert result["error"]
    assert matches_path.read_text(encoding="utf-8") == original_matches
    for relative_path in (
        "review_baseline_replay.json",
        "ui_decisions.json",
        "applied_decisions.json",
        "final_artifacts.json",
        "review_handoff.md",
    ):
        assert not (tmp_path / relative_path).exists()


def test_spanish_mcp_runtime_feedback_handoff_and_errors(tmp_path: Path) -> None:
    output_dir, run_intake, review_payload, _, _ = _prepare_sealed_mcp_review_run(
        tmp_path,
        language="es-ES",
    )
    decisions = [
        {"item_id": item["id"], "action": "accept"} for item in review_payload["items"]
    ]
    messages = [
        {
            "jsonrpc": "2.0",
            "id": 1,
            "method": "initialize",
            "params": {"_meta": {"locale": "es"}},
        },
        {
            "jsonrpc": "2.0",
            "id": 2,
            "method": "tools/call",
            "params": {
                "name": "validate_journal_bank_review",
                "arguments": {"review_payload": review_payload},
            },
        },
        {
            "jsonrpc": "2.0",
            "id": 3,
            "method": "tools/call",
            "params": {
                "name": "save_journal_bank_decisions",
                "arguments": {
                    "review_payload": review_payload,
                    "decisions": decisions,
                },
            },
        },
        {
            "jsonrpc": "2.0",
            "id": 4,
            "method": "tools/call",
            "params": {
                "name": "apply_journal_bank_decisions",
                "arguments": {
                    "run_intake": run_intake,
                    "review_payload": review_payload,
                    "decisions": decisions,
                },
            },
        },
        {
            "jsonrpc": "2.0",
            "id": 5,
            "method": "tools/call",
            "params": {
                "name": "validate_journal_bank_review",
                "arguments": {"review_payload": {**review_payload, "items": "invalid"}},
            },
        },
    ]

    responses = {response["id"]: response for response in _call_mcp_server(messages)}
    validation = responses[2]["result"]["structuredContent"]
    saved = responses[3]["result"]["structuredContent"]
    applied = responses[4]["result"]["structuredContent"]
    invalid = responses[5]["result"]["structuredContent"]
    handoff = (output_dir / "review_handoff.md").read_text(encoding="utf-8")

    assert (
        "Use validate_journal_bank_review antes"
        in responses[1]["result"]["instructions"]
    )
    assert validation["message"].startswith("Los datos de revisión")
    assert saved["message"].startswith("Las decisiones son válidas")
    assert applied["message"].startswith(f"Se aplicaron {len(decisions)} decisiones")
    assert applied["application_status"] == "blocked"
    assert applied["assurance_report_ready"] is False
    assert applied["final_artifacts"]["next_actions"][-1] == (
        "Resolve withheld or failed assurance gates before final handoff."
    )
    assert handoff.startswith(
        "# Entrega para revisión: Conciliación entre diario y banco\n"
    )
    assert "<!-- Review Handoff -->" in handoff
    handoff_output = next(
        output
        for output in applied["final_artifacts"]["outputs"]
        if output["path"] == "review_handoff.md"
    )
    assert handoff_output["required_text"][:2] == [
        "Entrega para revisión",
        "Review Handoff",
    ]
    assert invalid["error"] == "review_payload.items debe ser una matriz"


@pytest.mark.parametrize(
    "tool_name",
    ["save_journal_bank_decisions", "apply_journal_bank_decisions"],
)
@pytest.mark.parametrize(
    ("entry_kind", "expected_error"),
    [
        ("symlink", "symbolic links"),
        ("dangling_symlink", "symbolic links"),
        ("hardlink", "hardlink aliases"),
        ("fifo", "special filesystem entries"),
    ],
)
def test_journal_bank_mcp_review_transactions_reject_unsafe_internal_entries(
    tmp_path: Path,
    tool_name: str,
    entry_kind: str,
    expected_error: str,
) -> None:
    output_dir = tmp_path / "run"
    output_dir, run_intake, review_payload, _, _ = _prepare_sealed_mcp_review_run(
        output_dir
    )
    target = output_dir / "ui_decisions.json"
    outside = tmp_path / f"outside-{tool_name}-{entry_kind}.json"
    outside_before: bytes | None = None
    if entry_kind in {"symlink", "hardlink"}:
        outside.write_bytes(target.read_bytes())
        outside_before = outside.read_bytes()
        target.unlink()
        if entry_kind == "symlink":
            target.symlink_to(outside)
        else:
            os.link(outside, target)
    elif entry_kind == "dangling_symlink":
        target.unlink()
        target.symlink_to(tmp_path / "missing-ui-decisions.json")
    else:
        target = output_dir / "unexpected.fifo"
        os.mkfifo(target)

    response = _call_mcp_server(
        _mcp_review_write_message(tool_name, review_payload, run_intake)
    )[0]["result"]

    assert response["isError"] is True
    assert expected_error in response["structuredContent"]["error"]
    if outside_before is not None:
        assert outside.read_bytes() == outside_before
    if entry_kind in {"symlink", "dangling_symlink"}:
        assert target.is_symlink()
    elif entry_kind == "hardlink":
        assert target.stat().st_nlink == 2
    else:
        assert stat.S_ISFIFO(target.lstat().st_mode)
    assert list(output_dir.parent.glob(".journal-bank-apply-*")) == []


@pytest.mark.parametrize("link_kind", ["symlink", "hardlink"])
def test_journal_bank_mcp_post_preflight_link_swap_rolls_back_exact_tree(
    monkeypatch: Any,
    tmp_path: Path,
    link_kind: str,
) -> None:
    output_dir = tmp_path / "run"
    output_dir, run_intake, review_payload, _, _ = _prepare_sealed_mcp_review_run(
        output_dir
    )
    before = _tree_snapshot(output_dir)
    outside = tmp_path / "outside-reconciliation-matches.csv"
    wrapper = tmp_path / "python-preflight-journal-bank-swap"
    wrapper.write_text(
        "\n".join(
            [
                "#!/usr/bin/env python3",
                "import os",
                "from pathlib import Path",
                "import subprocess",
                "import sys",
                f"real = {str(Path(sys.executable))!r}",
                *_customer_run_preflight_passthrough_script_lines(),
                "completed = subprocess.run([real, *sys.argv[1:]], capture_output=True)",
                "sys.stdout.buffer.write(completed.stdout)",
                "sys.stderr.buffer.write(completed.stderr)",
                'if completed.returncode == 0 and "--preflight-only" in sys.argv:',
                '    out = Path(sys.argv[sys.argv.index("--output-dir") + 1])',
                '    outside = Path(os.environ["JB_TEST_OUTSIDE"])',
                '    link_kind = os.environ["JB_TEST_LINK_KIND"]',
                '    target = out / "reconciliation_matches.csv"',
                "    outside.write_bytes(target.read_bytes())",
                "    target.unlink()",
                "    if link_kind == 'symlink':",
                "        target.symlink_to(outside)",
                "    else:",
                "        os.link(outside, target)",
                "raise SystemExit(completed.returncode)",
                "",
            ]
        ),
        encoding="utf-8",
    )
    wrapper.chmod(0o700)
    monkeypatch.setenv("PYTHON", wrapper.as_posix())
    monkeypatch.setenv("JB_TEST_OUTSIDE", outside.as_posix())
    monkeypatch.setenv("JB_TEST_LINK_KIND", link_kind)

    response = _call_mcp_server(
        _mcp_review_write_message(
            "apply_journal_bank_decisions",
            review_payload,
            run_intake,
        )
    )[0]["result"]

    assert response["isError"] is True
    expected_error = "symbolic links" if link_kind == "symlink" else "hardlink aliases"
    assert expected_error in response["structuredContent"]["error"]
    assert _tree_snapshot(output_dir) == before
    assert (output_dir / "reconciliation_matches.csv").stat().st_nlink == 1
    assert b"Transaction containment review." not in outside.read_bytes()
    assert list(output_dir.parent.glob(".journal-bank-apply-*")) == []


@pytest.mark.parametrize("link_kind", ["symlink", "hardlink"])
def test_journal_bank_mcp_python_apply_link_swap_cannot_mutate_external_file(
    monkeypatch: Any,
    tmp_path: Path,
    link_kind: str,
) -> None:
    output_dir = tmp_path / "run"
    output_dir, run_intake, review_payload, _, _ = _prepare_sealed_mcp_review_run(
        output_dir
    )
    before = _tree_snapshot(output_dir)
    outside = tmp_path / f"outside-applied-{link_kind}.json"
    wrapper = tmp_path / f"python-apply-journal-bank-{link_kind}-swap"
    wrapper.write_text(
        "\n".join(
            [
                "#!/usr/bin/env python3",
                "import os",
                "from pathlib import Path",
                "import subprocess",
                "import sys",
                f"real = {str(Path(sys.executable))!r}",
                *_customer_run_preflight_passthrough_script_lines(),
                'if "--preflight-only" in sys.argv:',
                "    completed = subprocess.run([real, *sys.argv[1:]])",
                "    raise SystemExit(completed.returncode)",
                'out = Path(sys.argv[sys.argv.index("--output-dir") + 1])',
                'outside = Path(os.environ["JB_TEST_OUTSIDE"])',
                'link_kind = os.environ["JB_TEST_LINK_KIND"]',
                'target = out / "applied_decisions.json"',
                "outside.write_bytes(target.read_bytes())",
                "target.unlink()",
                "if link_kind == 'symlink':",
                "    target.symlink_to(outside)",
                "else:",
                "    os.link(outside, target)",
                "completed = subprocess.run([real, *sys.argv[1:]])",
                "raise SystemExit(completed.returncode)",
                "",
            ]
        ),
        encoding="utf-8",
    )
    wrapper.chmod(0o700)
    monkeypatch.setenv("PYTHON", wrapper.as_posix())
    monkeypatch.setenv("JB_TEST_OUTSIDE", outside.as_posix())
    monkeypatch.setenv("JB_TEST_LINK_KIND", link_kind)

    response = _call_mcp_server(
        _mcp_review_write_message(
            "apply_journal_bank_decisions",
            review_payload,
            run_intake,
        )
    )[0]["result"]

    assert response["isError"] is True
    assert (
        response["structuredContent"]["error"]
        == "Journal-Bank review application failed."
    )
    assert _tree_snapshot(output_dir) == before
    outside_payload = json.loads(outside.read_text(encoding="utf-8"))
    assert "native_regenerated_count" not in outside_payload
    assert list(output_dir.parent.glob(".journal-bank-apply-*")) == []


def test_journal_bank_mcp_dangling_canonical_swap_restores_exact_tree(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    output_dir = tmp_path / "run"
    output_dir, run_intake, review_payload, _, _ = _prepare_sealed_mcp_review_run(
        output_dir
    )
    before = _tree_snapshot(output_dir)
    wrapper = tmp_path / "python-journal-bank-dangling-output"
    wrapper.write_text(
        "\n".join(
            [
                "#!/usr/bin/env python3",
                "from pathlib import Path",
                "import shutil",
                "import subprocess",
                "import sys",
                f"real = {str(Path(sys.executable))!r}",
                *_customer_run_preflight_passthrough_script_lines(),
                'if "--preflight-only" in sys.argv:',
                "    completed = subprocess.run([real, *sys.argv[1:]])",
                "    raise SystemExit(completed.returncode)",
                'canonical = Path(sys.argv[sys.argv.index("--canonical-output-dir") + 1])',
                "shutil.rmtree(canonical)",
                "canonical.symlink_to(canonical.parent / 'missing-output-target')",
                "completed = subprocess.run([real, *sys.argv[1:]])",
                "raise SystemExit(completed.returncode)",
                "",
            ]
        ),
        encoding="utf-8",
    )
    wrapper.chmod(0o700)
    monkeypatch.setenv("PYTHON", wrapper.as_posix())

    response = _call_mcp_server(
        _mcp_review_write_message(
            "apply_journal_bank_decisions",
            review_payload,
            run_intake,
        )
    )[0]["result"]

    assert response["isError"] is True
    assert (
        response["structuredContent"]["error"]
        == "Journal-Bank review application failed."
    )
    assert not output_dir.is_symlink()
    assert _tree_snapshot(output_dir) == before
    assert list(output_dir.parent.glob(".journal-bank-apply-*")) == []


def test_journal_bank_mcp_rejects_intermediate_output_symlink(
    tmp_path: Path,
) -> None:
    output_dir = tmp_path / "real" / "run"
    output_dir, run_intake, review_payload, _, _ = _prepare_sealed_mcp_review_run(
        output_dir
    )
    before = _tree_snapshot(output_dir)
    alias = tmp_path / "alias"
    alias.symlink_to(output_dir.parent, target_is_directory=True)
    aliased_intake = {
        **run_intake,
        "output_dir": (alias / output_dir.name).as_posix(),
    }

    response = _call_mcp_server(
        _mcp_review_write_message(
            "save_journal_bank_decisions",
            review_payload,
            aliased_intake,
        )
    )[0]["result"]

    assert response["isError"] is True
    assert "parent must be a real directory" in response["structuredContent"]["error"]
    assert _tree_snapshot(output_dir) == before
    assert alias.is_symlink()
    assert list(output_dir.parent.glob(".journal-bank-apply-*")) == []


def test_journal_bank_python_preflight_rejects_replay_symlink(
    tmp_path: Path,
) -> None:
    output_dir = tmp_path / "run"
    output_dir, *_ = _prepare_sealed_mcp_review_run(output_dir)
    outside = tmp_path / "outside-replay.json"
    outside.write_bytes(b'{"outside":"unchanged"}\n')
    outside_before = outside.read_bytes()
    replay_path = output_dir / "review_baseline_replay.json"
    replay_path.symlink_to(outside)

    with pytest.raises(ValueError, match="symbolic links"):
        load_apply_review_edits().preflight_review_application(output_dir)

    assert outside.read_bytes() == outside_before
    assert replay_path.is_symlink()


def test_journal_bank_mcp_preflight_persisted_replay_must_match_stdout(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    output_dir = tmp_path / "run"
    output_dir, run_intake, review_payload, _, _ = _prepare_sealed_mcp_review_run(
        output_dir
    )
    before = _tree_snapshot(output_dir)
    wrapper = tmp_path / "python-journal-bank-preflight-divergence"
    wrapper.write_text(
        "\n".join(
            [
                "#!/usr/bin/env python3",
                "import hashlib",
                "import json",
                "import subprocess",
                "import sys",
                "from pathlib import Path",
                f"real = {str(Path(sys.executable))!r}",
                *_customer_run_preflight_passthrough_script_lines(),
                "completed = subprocess.run([real, *sys.argv[1:]], capture_output=True)",
                "if completed.returncode != 0:",
                "    sys.stdout.buffer.write(completed.stdout)",
                "    sys.stderr.buffer.write(completed.stderr)",
                "    raise SystemExit(completed.returncode)",
                'if "--preflight-only" in sys.argv:',
                '    output_dir = Path(sys.argv[sys.argv.index("--output-dir") + 1])',
                '    replay_path = output_dir / "review_baseline_replay.json"',
                "    replay = json.loads(replay_path.read_text())",
                '    replay["replayed_on"] = "1900-01-01"',
                "    content = dict(replay)",
                '    content.pop("content_sha256")',
                "    canonical = json.dumps(",
                "        content, ensure_ascii=False, sort_keys=True, separators=(',', ':')",
                "    ).encode()",
                '    replay["content_sha256"] = hashlib.sha256(canonical).hexdigest()',
                "    replay_path.write_text(",
                '        json.dumps(replay, indent=2, sort_keys=True) + "\\n"',
                "    )",
                "sys.stdout.buffer.write(completed.stdout)",
                "raise SystemExit(0)",
                "",
            ]
        ),
        encoding="utf-8",
    )
    wrapper.chmod(0o700)
    monkeypatch.setenv("PYTHON", wrapper.as_posix())

    response = _call_mcp_server(
        _mcp_review_write_message(
            "apply_journal_bank_decisions",
            review_payload,
            run_intake,
        )
    )[0]["result"]

    assert response["isError"] is True
    assert response["structuredContent"]["error"] == (
        "Journal-Bank assurance preflight returned an invalid result."
    )
    assert _tree_snapshot(output_dir) == before
    assert list(output_dir.parent.glob(".journal-bank-apply-*")) == []


@pytest.mark.parametrize(
    ("phase", "expected_error"),
    [
        ("preflight", "Journal-Bank assurance preflight failed."),
        ("apply", "Journal-Bank review application failed."),
    ],
)
def test_journal_bank_mcp_child_failure_is_path_free_and_rolls_back(
    monkeypatch: Any,
    tmp_path: Path,
    phase: str,
    expected_error: str,
) -> None:
    output_dir = tmp_path / "run"
    output_dir, run_intake, review_payload, _, _ = _prepare_sealed_mcp_review_run(
        output_dir
    )
    before = _tree_snapshot(output_dir)
    wrapper = tmp_path / f"python-journal-bank-malicious-{phase}"
    wrapper.write_text(
        "\n".join(
            [
                "#!/usr/bin/env python3",
                "import os",
                "import subprocess",
                "import sys",
                f"real = {str(Path(sys.executable))!r}",
                *_customer_run_preflight_passthrough_script_lines(),
                'phase = os.environ["JB_TEST_FAILURE_PHASE"]',
                'if phase == "apply" and "--preflight-only" in sys.argv:',
                "    completed = subprocess.run([real, *sys.argv[1:]], capture_output=True)",
                "    sys.stdout.buffer.write(completed.stdout)",
                "    sys.stderr.buffer.write(completed.stderr)",
                "    raise SystemExit(completed.returncode)",
                'sys.stdout.write("untrusted /Users/private/client/run.csv\\n")',
                'sys.stderr.write("Traceback (most recent call last):\\n")',
                'sys.stderr.write("ValueError: C:\\\\private\\\\client\\\\run.csv\\n")',
                "raise SystemExit(23)",
                "",
            ]
        ),
        encoding="utf-8",
    )
    wrapper.chmod(0o700)
    monkeypatch.setenv("PYTHON", wrapper.as_posix())
    monkeypatch.setenv("JB_TEST_FAILURE_PHASE", phase)

    response = _call_mcp_server(
        _mcp_review_write_message(
            "apply_journal_bank_decisions",
            review_payload,
            run_intake,
        )
    )[0]["result"]

    assert response["isError"] is True
    error = response["structuredContent"]["error"]
    assert error == expected_error
    assert "\n" not in error
    assert "Traceback" not in error
    assert "/Users/private" not in error
    assert "\\private\\client" not in error
    assert tmp_path.as_posix() not in error
    assert _tree_snapshot(output_dir) == before
    assert list(output_dir.parent.glob(".journal-bank-apply-*")) == []


def test_journal_bank_mcp_rejects_transaction_root_relocation_without_moving_canonical(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    output_dir = tmp_path / "run"
    output_dir, run_intake, review_payload, _, _ = _prepare_sealed_mcp_review_run(
        output_dir
    )
    before = _tree_snapshot(output_dir)
    before_modes = _tree_mode_snapshot(output_dir)
    canonical_inode = output_dir.stat().st_ino
    wrapper = tmp_path / "python-journal-bank-root-relocation"
    wrapper.write_text(
        "\n".join(
            [
                "#!/usr/bin/env python3",
                "from pathlib import Path",
                "import subprocess",
                "import sys",
                f"real = {str(Path(sys.executable))!r}",
                *_customer_run_preflight_passthrough_script_lines(),
                "completed = subprocess.run([real, *sys.argv[1:]], capture_output=True)",
                "if completed.returncode == 0 and '--preflight-only' not in sys.argv:",
                "    working = Path(sys.argv[sys.argv.index('--output-dir') + 1])",
                "    transaction_root = working.parent",
                "    transaction_root.rename(transaction_root.with_name(transaction_root.name + '-moved'))",
                "sys.stdout.buffer.write(completed.stdout)",
                "sys.stderr.buffer.write(completed.stderr)",
                "raise SystemExit(completed.returncode)",
                "",
            ]
        ),
        encoding="utf-8",
    )
    wrapper.chmod(0o700)
    monkeypatch.setenv("PYTHON", wrapper.as_posix())

    response = _call_mcp_server(
        _mcp_review_write_message(
            "apply_journal_bank_decisions",
            review_payload,
            run_intake,
        )
    )[0]["result"]

    assert response["isError"] is True
    assert output_dir.stat().st_ino == canonical_inode
    assert _tree_snapshot(output_dir) == before
    assert _tree_mode_snapshot(output_dir) == before_modes
    assert list(output_dir.parent.glob(".journal-bank-apply-*")) == []
    assert list(output_dir.parent.glob(".journal-bank-commit-*")) == []
    assert list(output_dir.parent.glob(".journal-bank-recovery-*")) == []


@pytest.mark.parametrize(
    "attack",
    [
        "directory",
        "symlink",
        "fifo",
    ],
)
def test_journal_bank_mcp_child_created_snapshot_is_non_authoritative(
    monkeypatch: Any,
    tmp_path: Path,
    attack: str,
) -> None:
    output_dir = tmp_path / "run"
    output_dir, run_intake, review_payload, _, _ = _prepare_sealed_mcp_review_run(
        output_dir
    )
    outside = tmp_path / "outside-snapshot-target"
    outside.mkdir()
    marker = outside / "marker.txt"
    marker.write_text("outside unchanged\n", encoding="utf-8")
    wrapper = tmp_path / "python-journal-bank-snapshot-attack"
    wrapper.write_text(
        "\n".join(
            [
                "#!/usr/bin/env python3",
                "import json",
                "import os",
                "import subprocess",
                "import sys",
                "from pathlib import Path",
                f"real = {str(Path(sys.executable))!r}",
                *_customer_run_preflight_passthrough_script_lines(),
                "completed = subprocess.run([real, *sys.argv[1:]], capture_output=True)",
                "if completed.returncode != 0:",
                "    sys.stdout.buffer.write(completed.stdout)",
                "    sys.stderr.buffer.write(completed.stderr)",
                "    raise SystemExit(completed.returncode)",
                'if "--preflight-only" in sys.argv:',
                '    working = Path(sys.argv[sys.argv.index("--output-dir") + 1])',
                '    snapshot = working.parent / "snapshot"',
                '    attack = os.environ["JB_TEST_SNAPSHOT_ATTACK"]',
                '    if attack == "directory":',
                "        snapshot.mkdir()",
                '        (snapshot / "final_artifacts.json").write_text(',
                '            json.dumps({"poisoned_by_child": "/Users/private/client-secret.csv"}) + "\\n"',
                "        )",
                '    elif attack == "symlink":',
                '        snapshot.symlink_to(Path(os.environ["JB_TEST_SNAPSHOT_OUTSIDE"]), target_is_directory=True)',
                '    elif attack == "fifo":',
                "        os.mkfifo(snapshot)",
                "sys.stdout.buffer.write(completed.stdout)",
                "sys.stderr.buffer.write(completed.stderr)",
                "raise SystemExit(completed.returncode)",
                "",
            ]
        ),
        encoding="utf-8",
    )
    wrapper.chmod(0o700)
    monkeypatch.setenv("PYTHON", wrapper.as_posix())
    monkeypatch.setenv("JB_TEST_SNAPSHOT_ATTACK", attack)
    monkeypatch.setenv("JB_TEST_SNAPSHOT_OUTSIDE", outside.as_posix())

    response = _call_mcp_server(
        _mcp_review_write_message(
            "apply_journal_bank_decisions",
            review_payload,
            run_intake,
        )
    )[0]["result"]

    assert response["isError"] is False, response
    final_artifacts = json.loads(
        (output_dir / "final_artifacts.json").read_text(encoding="utf-8")
    )
    assert "poisoned_by_child" not in final_artifacts
    assert marker.read_text(encoding="utf-8") == "outside unchanged\n"
    assert list(output_dir.parent.glob(".journal-bank-apply-*")) == []
    assert list(output_dir.parent.glob(".journal-bank-commit-*")) == []
    assert list(output_dir.parent.glob(".journal-bank-recovery-*")) == []


@pytest.mark.parametrize(
    ("phase", "target", "expected_error"),
    [
        (
            "preflight",
            "root",
            "Journal-Bank assurance preflight returned an invalid result.",
        ),
        (
            "apply",
            "root",
            "Journal-Bank review application returned an invalid result.",
        ),
        (
            "apply",
            "file",
            "Journal-Bank review application returned an invalid result.",
        ),
    ],
)
def test_journal_bank_mcp_child_cannot_change_output_tree_modes(
    monkeypatch: Any,
    tmp_path: Path,
    phase: str,
    target: str,
    expected_error: str,
) -> None:
    output_dir = tmp_path / "run"
    output_dir, run_intake, review_payload, _, _ = _prepare_sealed_mcp_review_run(
        output_dir
    )
    before = _tree_snapshot(output_dir)
    before_modes = _tree_mode_snapshot(output_dir)
    wrapper = tmp_path / "python-journal-bank-mode-attack"
    wrapper.write_text(
        "\n".join(
            [
                "#!/usr/bin/env python3",
                "import os",
                "import subprocess",
                "import sys",
                "from pathlib import Path",
                f"real = {str(Path(sys.executable))!r}",
                *_customer_run_preflight_passthrough_script_lines(),
                "completed = subprocess.run([real, *sys.argv[1:]], capture_output=True)",
                "if completed.returncode != 0:",
                "    sys.stdout.buffer.write(completed.stdout)",
                "    sys.stderr.buffer.write(completed.stderr)",
                "    raise SystemExit(completed.returncode)",
                'is_preflight = "--preflight-only" in sys.argv',
                'phase = os.environ["JB_TEST_MODE_PHASE"]',
                "if (phase == 'preflight') == is_preflight:",
                '    output_dir = Path(sys.argv[sys.argv.index("--output-dir") + 1])',
                '    target = output_dir if os.environ["JB_TEST_MODE_TARGET"] == "root" else output_dir / "final_artifacts.json"',
                "    target.chmod(0o777)",
                "sys.stdout.buffer.write(completed.stdout)",
                "raise SystemExit(0)",
                "",
            ]
        ),
        encoding="utf-8",
    )
    wrapper.chmod(0o700)
    monkeypatch.setenv("PYTHON", wrapper.as_posix())
    monkeypatch.setenv("JB_TEST_MODE_PHASE", phase)
    monkeypatch.setenv("JB_TEST_MODE_TARGET", target)

    response = _call_mcp_server(
        _mcp_review_write_message(
            "apply_journal_bank_decisions",
            review_payload,
            run_intake,
        )
    )[0]["result"]

    assert response["isError"] is True
    assert response["structuredContent"]["error"] == expected_error
    assert _tree_snapshot(output_dir) == before
    assert _tree_mode_snapshot(output_dir) == before_modes
    assert list(output_dir.parent.glob(".journal-bank-apply-*")) == []
    assert list(output_dir.parent.glob(".journal-bank-recovery-*")) == []


@pytest.mark.parametrize(
    ("phase", "expected_error"),
    [
        (
            "preflight",
            "Journal-Bank assurance preflight returned an invalid result.",
        ),
        (
            "apply",
            "Journal-Bank review application returned an invalid result.",
        ),
    ],
)
@pytest.mark.parametrize("child_output", ["{not-json", "{}"])
def test_journal_bank_mcp_invalid_child_output_fails_closed(
    monkeypatch: Any,
    tmp_path: Path,
    phase: str,
    expected_error: str,
    child_output: str,
) -> None:
    output_dir = tmp_path / "run"
    output_dir, run_intake, review_payload, _, _ = _prepare_sealed_mcp_review_run(
        output_dir
    )
    before = _tree_snapshot(output_dir)
    wrapper = tmp_path / f"python-journal-bank-invalid-{phase}"
    wrapper.write_text(
        "\n".join(
            [
                "#!/usr/bin/env python3",
                "import os",
                "import subprocess",
                "import sys",
                f"real = {str(Path(sys.executable))!r}",
                *_customer_run_preflight_passthrough_script_lines(),
                'phase = os.environ["JB_TEST_INVALID_PHASE"]',
                'if phase == "apply" and "--preflight-only" in sys.argv:',
                "    completed = subprocess.run([real, *sys.argv[1:]], capture_output=True)",
                "    sys.stdout.buffer.write(completed.stdout)",
                "    sys.stderr.buffer.write(completed.stderr)",
                "    raise SystemExit(completed.returncode)",
                'sys.stdout.write(os.environ["JB_TEST_INVALID_OUTPUT"] + "\\n")',
                "raise SystemExit(0)",
                "",
            ]
        ),
        encoding="utf-8",
    )
    wrapper.chmod(0o700)
    monkeypatch.setenv("PYTHON", wrapper.as_posix())
    monkeypatch.setenv("JB_TEST_INVALID_PHASE", phase)
    monkeypatch.setenv("JB_TEST_INVALID_OUTPUT", child_output)

    response = _call_mcp_server(
        _mcp_review_write_message(
            "apply_journal_bank_decisions",
            review_payload,
            run_intake,
        )
    )[0]["result"]

    assert response["isError"] is True
    assert response["structuredContent"]["error"] == expected_error
    assert _tree_snapshot(output_dir) == before
    assert list(output_dir.parent.glob(".journal-bank-apply-*")) == []


@pytest.mark.parametrize(
    "forgery_mode",
    [
        "ack_only",
        "staged_final",
        "staged_final_with_receipts",
        "staged_counts_with_receipts",
        "forged_gate_with_receipts",
        "forged_limitations_with_receipts",
        "forged_native_counts_with_receipts",
        "forged_trace_with_receipts",
    ],
)
def test_journal_bank_mcp_child_cannot_self_author_final_readiness(
    monkeypatch: Any,
    tmp_path: Path,
    forgery_mode: str,
) -> None:
    output_dir = tmp_path / "run"
    output_dir, run_intake, review_payload, _, _ = _prepare_sealed_mcp_review_run(
        output_dir
    )
    before = _tree_snapshot(output_dir)
    wrapper = tmp_path / "python-journal-bank-readiness-forgery"
    wrapper.write_text(
        "\n".join(
            [
                "#!/usr/bin/env python3",
                "import hashlib",
                "import json",
                "import os",
                "import subprocess",
                "import sys",
                "from pathlib import Path",
                f"real = {str(Path(sys.executable))!r}",
                *_customer_run_preflight_passthrough_script_lines(),
                "completed = subprocess.run([real, *sys.argv[1:]], capture_output=True)",
                "if completed.returncode != 0:",
                "    sys.stdout.buffer.write(completed.stdout)",
                "    sys.stderr.buffer.write(completed.stderr)",
                "    raise SystemExit(completed.returncode)",
                'if "--preflight-only" in sys.argv:',
                "    sys.stdout.buffer.write(completed.stdout)",
                "    raise SystemExit(0)",
                "payload = json.loads(completed.stdout.decode().strip().splitlines()[-1])",
                'mode = os.environ["JB_TEST_FORGERY_MODE"]',
                'output_dir = Path(sys.argv[sys.argv.index("--output-dir") + 1])',
                'if mode in {"ack_only", "staged_final", "staged_final_with_receipts", "forged_gate_with_receipts"}:',
                '    payload["application_status"] = "final_ready"',
                '    payload["assurance_report_ready"] = True',
                '    payload["applied_decisions"]["application_status"] = "final_ready"',
                '    payload["applied_decisions"]["assurance_report_ready"] = True',
                '    payload["final_artifacts"]["status"] = "final_ready"',
                '    payload["final_artifacts"]["review_status"] = "final_ready"',
                '    payload["final_artifacts"]["review_application"]["application_status"] = "final_ready"',
                '    payload["final_artifacts"]["review_application"]["assurance_report_ready"] = True',
                'if mode == "staged_counts_with_receipts":',
                '    payload["applied_decisions"]["decision_count"] = 999',
                '    payload["applied_decisions"]["item_count"] = 999',
                '    payload["applied_decisions"]["blocker_count"] = 0',
                '    payload["final_artifacts"]["review_application"]["decision_count"] = 999',
                '    payload["final_artifacts"]["review_application"]["item_count"] = 999',
                '    payload["final_artifacts"]["review_application"]["blocker_count"] = 0',
                'if mode == "forged_limitations_with_receipts":',
                '    payload["applied_decisions"]["assurance_limitations"] = ["Traceback /Users/private/client-secret.csv"]',
                '    payload["final_artifacts"]["review_application"]["assurance_limitations"] = ["Traceback /Users/private/client-secret.csv"]',
                'if mode == "forged_native_counts_with_receipts":',
                '    payload["applied_decisions"]["native_regeneration_count"] = 999',
                '    payload["final_artifacts"]["review_application"]["native_regeneration_count"] = 999',
                '    payload["final_artifacts"]["next_actions"] = ["Traceback /Users/private/count-secret.csv"]',
                'if mode != "ack_only":',
                '    applied_path = output_dir / "applied_decisions.json"',
                '    final_path = output_dir / "final_artifacts.json"',
                '    applied_path.write_text(json.dumps(payload["applied_decisions"], indent=2, sort_keys=True) + "\\n")',
                '    final_path.write_text(json.dumps(payload["final_artifacts"], indent=2, sort_keys=True) + "\\n")',
                'if mode == "forged_gate_with_receipts":',
                '    gates_path = output_dir / "assurance_gates.json"',
                "    gates = json.loads(gates_path.read_text())",
                '    gates["gates"]["reconciliation"] = {"status": "passed", "evidence_refs": ["output.relationship_ledger_json"], "limitations": []}',
                '    gates["gates"]["semantic_review"] = {"status": "passed", "evidence_refs": ["decision.review_application.forged"], "limitations": []}',
                '    gates["gates"]["reporting"] = {"status": "passed", "evidence_refs": ["output.final_artifacts_json"], "limitations": []}',
                '    gates["report_ready"] = True',
                '    gates_path.write_text(json.dumps(gates, indent=2, sort_keys=True) + "\\n")',
                'if mode == "forged_trace_with_receipts":',
                '    run_intake_path = output_dir / "run_intake.json"',
                "    run_intake = json.loads(run_intake_path.read_text())",
                '    run_intake["execution_trace"].append({"step_id": "forged_final_ready", "kind": "deterministic_review_apply", "status": "passed", "execution_location": "local_codex_workspace", "command": ["forged"], "inputs": [], "outputs": ["/Users/private/client-secret.csv"]})',
                '    run_intake_path.write_text(json.dumps(run_intake, indent=2, sort_keys=True) + "\\n")',
                'if mode in {"staged_final_with_receipts", "staged_counts_with_receipts", "forged_gate_with_receipts", "forged_limitations_with_receipts", "forged_native_counts_with_receipts", "forged_trace_with_receipts"}:',
                '    receipts_path = output_dir / "artifact_receipts.json"',
                "    receipts = json.loads(receipts_path.read_text())",
                '    for receipt in receipts["output_receipts"]:',
                '        artifact = output_dir / receipt["path"]',
                "        if artifact.is_file():",
                "            content = artifact.read_bytes()",
                '            receipt["byte_count"] = len(content)',
                '            receipt["sha256"] = hashlib.sha256(content).hexdigest()',
                '    receipts_path.write_text(json.dumps(receipts, indent=2, sort_keys=True) + "\\n")',
                'sys.stdout.write(json.dumps(payload) + "\\n")',
                "raise SystemExit(0)",
                "",
            ]
        ),
        encoding="utf-8",
    )
    wrapper.chmod(0o700)
    monkeypatch.setenv("PYTHON", wrapper.as_posix())
    monkeypatch.setenv("JB_TEST_FORGERY_MODE", forgery_mode)

    response = _call_mcp_server(
        _mcp_review_write_message(
            "apply_journal_bank_decisions",
            review_payload,
            run_intake,
        )
    )[0]["result"]

    if forgery_mode == "ack_only":
        assert response["isError"] is False
        result = response["structuredContent"]
        assert result["application_status"] != "final_ready"
        assert result["assurance_report_ready"] is False
        assert (
            json.loads((output_dir / "assurance_gates.json").read_text())[
                "report_ready"
            ]
            is False
        )
    else:
        assert response["isError"] is True
        assert response["structuredContent"]["error"] == (
            "Journal-Bank review application returned an invalid result."
        )
        assert _tree_snapshot(output_dir) == before
        assert list(output_dir.parent.glob(".journal-bank-apply-*")) == []


def test_journal_bank_mcp_no_native_child_cannot_extend_trace_outputs(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    output_dir = tmp_path / "run"
    output_dir, run_intake, review_payload, _, _ = _prepare_sealed_mcp_review_run(
        output_dir
    )
    before = _tree_snapshot(output_dir)
    wrapper = tmp_path / "python-journal-bank-trace-forgery"
    wrapper.write_text(
        "\n".join(
            [
                "#!/usr/bin/env python3",
                "import hashlib",
                "import json",
                "import subprocess",
                "import sys",
                "from pathlib import Path",
                f"real = {str(Path(sys.executable))!r}",
                *_customer_run_preflight_passthrough_script_lines(),
                "completed = subprocess.run([real, *sys.argv[1:]], capture_output=True)",
                "if completed.returncode != 0:",
                "    sys.stdout.buffer.write(completed.stdout)",
                "    sys.stderr.buffer.write(completed.stderr)",
                "    raise SystemExit(completed.returncode)",
                'if "--preflight-only" in sys.argv:',
                "    sys.stdout.buffer.write(completed.stdout)",
                "    raise SystemExit(0)",
                'output_dir = Path(sys.argv[sys.argv.index("--output-dir") + 1])',
                'run_intake_path = output_dir / "run_intake.json"',
                "run_intake = json.loads(run_intake_path.read_text())",
                'run_intake["execution_trace"][-1]["outputs"].append(',
                '    "journal_bank_reconciliation.xlsx"',
                ")",
                "run_intake_path.write_text(",
                '    json.dumps(run_intake, indent=2, sort_keys=True) + "\\n"',
                ")",
                'receipts_path = output_dir / "artifact_receipts.json"',
                "receipts = json.loads(receipts_path.read_text())",
                'for receipt in receipts["output_receipts"]:',
                '    artifact = output_dir / receipt["path"]',
                "    if artifact.is_file():",
                "        content = artifact.read_bytes()",
                '        receipt["byte_count"] = len(content)',
                '        receipt["sha256"] = hashlib.sha256(content).hexdigest()',
                "receipts_path.write_text(",
                '    json.dumps(receipts, indent=2, sort_keys=True) + "\\n"',
                ")",
                "sys.stdout.write('{\"ok\": true}\\n')",
                "raise SystemExit(0)",
                "",
            ]
        ),
        encoding="utf-8",
    )
    wrapper.chmod(0o700)
    monkeypatch.setenv("PYTHON", wrapper.as_posix())
    item = next(
        entry
        for entry in review_payload["items"]
        if entry["item_type"] == "matched_pair"
    )

    response = _call_mcp_server(
        [
            {
                "jsonrpc": "2.0",
                "id": 1,
                "method": "tools/call",
                "params": {
                    "name": "apply_journal_bank_decisions",
                    "arguments": {
                        "run_intake": run_intake,
                        "review_payload": review_payload,
                        "decisions": [{"item_id": item["id"], "action": "accept"}],
                    },
                },
            }
        ]
    )[0]["result"]

    assert response["isError"] is True
    assert response["structuredContent"]["error"] == (
        "Journal-Bank review application returned an invalid result."
    )
    assert _tree_snapshot(output_dir) == before
    assert list(output_dir.parent.glob(".journal-bank-apply-*")) == []


@pytest.mark.parametrize(
    "forgery_mode",
    [
        "garbage_bytes",
        "missing_sheet",
        "changed_header",
        "changed_key_cell",
        "extra_sheet",
        "formula_cell",
        "hidden_sheet",
        "hidden_row",
        "hidden_column",
        "hyperlink_cell",
        "styled_cell",
        "default_style_concealment",
        "extra_package_entry",
        "prefixed_formula",
        "wrong_namespace_cell",
        "core_creator",
        "core_timestamp_mismatch",
        "core_invalid_timestamp",
        "invalid_numeric_entity",
        "bad_crc",
        "malformed_xml",
        "xml_declaration_after_root",
        "invalid_xml_declaration",
        "unbound_namespace",
        "zip_traversal",
        "duplicate_zip_entry",
    ],
)
def test_journal_bank_mcp_rejects_forged_regenerated_workbook_bytes(
    monkeypatch: Any,
    tmp_path: Path,
    forgery_mode: str,
) -> None:
    output_dir = tmp_path / "run"
    output_dir, run_intake, review_payload, _, final_artifacts = (
        _prepare_closed_mcp_review_run(output_dir, portable=True)
    )
    decisions = []
    edited = False
    for item in review_payload["items"]:
        if item["item_type"] == "matched_pair" and not edited:
            decisions.append(
                {
                    "item_id": item["id"],
                    "action": "edit",
                    "edit_value": "Reviewed closed reconciliation.",
                }
            )
            edited = True
        else:
            decisions.append({"item_id": item["id"], "action": "accept"})
    assert edited is True
    before = _tree_snapshot(output_dir)
    wrapper = tmp_path / "python-journal-bank-workbook-forgery"
    wrapper.write_text(
        "\n".join(
            [
                f"#!{sys.executable}",
                "import hashlib",
                "import json",
                "import os",
                "import re",
                "import subprocess",
                "import sys",
                "import zipfile",
                "from pathlib import Path",
                "import openpyxl",
                f"real = {str(Path(sys.executable))!r}",
                *_customer_run_preflight_passthrough_script_lines(),
                "completed = subprocess.run([real, *sys.argv[1:]], capture_output=True)",
                "if completed.returncode != 0:",
                "    sys.stdout.buffer.write(completed.stdout)",
                "    sys.stderr.buffer.write(completed.stderr)",
                "    raise SystemExit(completed.returncode)",
                'if "--preflight-only" in sys.argv:',
                "    sys.stdout.buffer.write(completed.stdout)",
                "    raise SystemExit(0)",
                'mode = os.environ["JB_TEST_WORKBOOK_FORGERY"]',
                'output_dir = Path(sys.argv[sys.argv.index("--output-dir") + 1])',
                'workbook_path = output_dir / "journal_bank_reconciliation.xlsx"',
                'if mode == "garbage_bytes":',
                '    workbook_path.write_bytes(b"forged workbook bytes")',
                'elif mode in {"missing_sheet", "changed_header", "changed_key_cell", "extra_sheet", "formula_cell", "hidden_sheet", "hidden_row", "hidden_column", "hyperlink_cell", "styled_cell"}:',
                "    workbook = openpyxl.load_workbook(workbook_path)",
                '    if mode == "missing_sheet":',
                '        workbook.remove(workbook["normalized_journal"])',
                '    elif mode == "changed_header":',
                '        workbook["matches"]["A1"] = "forged header"',
                '    elif mode == "changed_key_cell":',
                '        workbook["matches"]["A2"] = "forged value"',
                '    elif mode == "extra_sheet":',
                '        workbook.create_sheet("forged")',
                '    elif mode == "formula_cell":',
                '        workbook["matches"]["A1"] = "=1+1"',
                '    elif mode == "hidden_sheet":',
                '        workbook["matches"].sheet_state = "hidden"',
                '    elif mode == "hidden_row":',
                '        workbook["matches"].row_dimensions[2].hidden = True',
                '    elif mode == "hidden_column":',
                '        workbook["matches"].column_dimensions["A"].hidden = True',
                '    elif mode == "hyperlink_cell":',
                '        workbook["matches"]["A1"].hyperlink = "https://example.invalid/"',
                "    else:",
                "        from openpyxl.styles import Font",
                '        workbook["matches"]["A1"].font = Font(color="FFFFFF")',
                "    workbook.save(workbook_path)",
                'elif mode == "bad_crc":',
                "    archive = bytearray(workbook_path.read_bytes())",
                '    target = b"xl/workbook.xml"',
                "    found = False",
                "    for offset in range(len(archive) - 46):",
                '        if archive[offset:offset + 4] != b"PK\\x01\\x02":',
                "            continue",
                "        name_length = int.from_bytes(archive[offset + 28:offset + 30], 'little')",
                "        name = bytes(archive[offset + 46:offset + 46 + name_length])",
                "        if name != target:",
                "            continue",
                "        local_offset = int.from_bytes(",
                "            archive[offset + 42:offset + 46], 'little'",
                "        )",
                "        archive[offset + 16:offset + 20] = bytes(4)",
                "        archive[local_offset + 14:local_offset + 18] = bytes(4)",
                "        found = True",
                "        break",
                "    if not found:",
                '        raise RuntimeError("workbook.xml ZIP record missing")',
                "    workbook_path.write_bytes(archive)",
                'elif mode in {"malformed_xml", "xml_declaration_after_root", "invalid_xml_declaration", "unbound_namespace", "default_style_concealment", "prefixed_formula", "wrong_namespace_cell", "core_creator", "core_timestamp_mismatch", "core_invalid_timestamp", "invalid_numeric_entity"}:',
                '    rebuilt_path = workbook_path.with_suffix(".rebuilt.xlsx")',
                '    with zipfile.ZipFile(workbook_path, "r") as source:',
                '        with zipfile.ZipFile(rebuilt_path, "w") as target:',
                "            for info in source.infolist():",
                "                content = source.read(info.filename)",
                '                if info.filename == "xl/workbook.xml":',
                '                    if mode == "malformed_xml":',
                '                        content += b"<broken"',
                '                    elif mode == "xml_declaration_after_root":',
                "                        content += b'<?xml version=\"1.0\"?>'",
                '                    elif mode == "invalid_xml_declaration":',
                '                        content = b"<?xml?>" + content',
                '                    elif mode == "unbound_namespace":',
                '                        content = content.replace(b"</workbook>", b"<x:evil/></workbook>")',
                '                    elif mode == "invalid_numeric_entity":',
                '                        content = content.replace(b"<workbook ", b\'<workbook bogus="&#0;" \', 1)',
                '                if mode == "default_style_concealment" and info.filename == "xl/styles.xml":',
                "                    changed = content.replace(b'<color theme=\"1\"/>', b'<color rgb=\"00FFFFFF\"/>', 1)",
                "                    if changed == content:",
                '                        raise RuntimeError("default style color not found")',
                "                    content = changed",
                '                if info.filename == "docProps/core.xml":',
                '                    if mode == "core_creator":',
                '                        changed = content.replace(b"openpyxl", b"forged", 1)',
                '                    elif mode == "core_timestamp_mismatch":',
                "                        changed = re.sub(",
                '                            rb"(<dcterms:created\\b[^>]*>)[^<]*(</dcterms:created>)",',
                '                            rb"\\g<1>2000-01-01T00:00:00Z\\g<2>",',
                "                            content,",
                "                            count=1,",
                "                        )",
                '                    elif mode == "core_invalid_timestamp":',
                "                        changed = re.sub(",
                '                            rb"(<dcterms:modified\\b[^>]*>)[^<]*(</dcterms:modified>)",',
                '                            rb"\\g<1>2026-02-31T00:00:00Z\\g<2>",',
                "                            content,",
                "                            count=1,",
                "                        )",
                "                    else:",
                "                        changed = content",
                '                    if mode.startswith("core_") and changed == content:',
                '                        raise RuntimeError("core property target not found")',
                "                    content = changed",
                '                if mode == "wrong_namespace_cell" and info.filename == "xl/worksheets/sheet1.xml":',
                "                    changed = content.replace(",
                "                        b'<c r=\"A2\"',",
                '                        b\'<c xmlns="urn:evil" r="A2"\',',
                "                        1,",
                "                    )",
                "                    if changed == content:",
                '                        raise RuntimeError("A2 worksheet cell not found")',
                "                    content = changed",
                '                if mode == "prefixed_formula" and info.filename == "xl/worksheets/sheet1.xml":',
                "                    changed = re.sub(",
                "                        rb'<c r=\"A2\"[^>]*>.*?</c>',",
                '                        b\'<c r="A2" t="str"><x:f xmlns:x="http://schemas.openxmlformats.org/spreadsheetml/2006/main">1+1</x:f><v>matched</v></c>\',',
                "                        content,",
                "                        count=1,",
                "                    )",
                "                    if changed == content:",
                '                        raise RuntimeError("A2 worksheet cell not found")',
                "                    content = changed",
                "                target.writestr(info, content)",
                "    rebuilt_path.replace(workbook_path)",
                'elif mode == "extra_package_entry":',
                '    with zipfile.ZipFile(workbook_path, "a") as archive:',
                '        archive.writestr("xl/media/secret.txt", "/Users/private/client-secret.csv")',
                'elif mode == "zip_traversal":',
                '    with zipfile.ZipFile(workbook_path, "a") as archive:',
                '        archive.writestr("../forged.xml", "<forged/>")',
                "else:",
                '    with zipfile.ZipFile(workbook_path, "a") as archive:',
                '        workbook_xml = archive.read("xl/workbook.xml")',
                '        archive.writestr("xl/workbook.xml", workbook_xml)',
                'final_path = output_dir / "final_artifacts.json"',
                "final = json.loads(final_path.read_text())",
                "workbook_output = next(",
                '    value for value in final["outputs"]',
                '    if value["path"] == "journal_bank_reconciliation.xlsx"',
                ")",
                'workbook_output["size_bytes"] = workbook_path.stat().st_size',
                "final_path.write_text(",
                '    json.dumps(final, indent=2, sort_keys=True) + "\\n"',
                ")",
                'envelope_path = output_dir / "assurance_envelope.reviewed.json"',
                "envelope = json.loads(envelope_path.read_text())",
                'for receipt in envelope["artifact_receipts"]:',
                '    if receipt["root_id"] != "run":',
                "        continue",
                '    artifact = output_dir / receipt["path"]',
                "    content = artifact.read_bytes()",
                '    receipt["byte_count"] = len(content)',
                '    receipt["sha256"] = hashlib.sha256(content).hexdigest()',
                "envelope_content = dict(envelope)",
                'envelope_content.pop("content_sha256")',
                "canonical = json.dumps(",
                "    envelope_content, ensure_ascii=False, sort_keys=True, separators=(',', ':')",
                " ).encode()",
                'envelope["content_sha256"] = hashlib.sha256(canonical).hexdigest()',
                "envelope_path.write_text(",
                '    json.dumps(envelope, indent=2, sort_keys=True) + "\\n"',
                ")",
                'receipts_path = output_dir / "artifact_receipts.json"',
                "receipts = json.loads(receipts_path.read_text())",
                'for receipt in receipts["output_receipts"]:',
                '    artifact = output_dir / receipt["path"]',
                "    if artifact.is_file():",
                "        content = artifact.read_bytes()",
                '        receipt["byte_count"] = len(content)',
                '        receipt["sha256"] = hashlib.sha256(content).hexdigest()',
                "receipts_path.write_text(",
                '    json.dumps(receipts, indent=2, sort_keys=True) + "\\n"',
                ")",
                "sys.stdout.write('{\"ok\": true}\\n')",
                "raise SystemExit(0)",
                "",
            ]
        ),
        encoding="utf-8",
    )
    wrapper.chmod(0o700)
    monkeypatch.setenv("PYTHON", wrapper.as_posix())
    monkeypatch.setenv("JB_TEST_WORKBOOK_FORGERY", forgery_mode)

    response = _call_mcp_server(
        [
            {
                "jsonrpc": "2.0",
                "id": 1,
                "method": "tools/call",
                "params": {
                    "name": "apply_journal_bank_decisions",
                    "arguments": {
                        "client_engagement": _customer_context_path(
                            output_dir
                        ).as_posix(),
                        "run_intake": run_intake,
                        "review_payload": review_payload,
                        "final_artifacts": final_artifacts,
                        "decisions": decisions,
                    },
                },
            }
        ]
    )[0]["result"]

    assert response["isError"] is True
    assert response["structuredContent"]["error"] == (
        "Journal-Bank review application returned an invalid result."
    )
    assert _tree_snapshot(output_dir) == before
    assert list(output_dir.parent.glob(".journal-bank-apply-*")) == []


def test_journal_bank_mcp_preflight_start_failure_is_fixed(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    output_dir = tmp_path / "run"
    output_dir, run_intake, review_payload, _, _ = _prepare_sealed_mcp_review_run(
        output_dir
    )
    before = _tree_snapshot(output_dir)
    blocked = tmp_path / "python-journal-bank-not-executable"
    blocked.write_text(
        "\n".join(
            [
                "#!/usr/bin/env python3",
                "from pathlib import Path",
                "import subprocess",
                "import sys",
                f"real = {str(Path(sys.executable))!r}",
                "completed = subprocess.run([real, *sys.argv[1:]], capture_output=True)",
                "sys.stdout.buffer.write(completed.stdout)",
                "sys.stderr.buffer.write(completed.stderr)",
                'if completed.returncode == 0 and "--client-run-preflight-only" in sys.argv:',
                "    Path(__file__).chmod(0o600)",
                "raise SystemExit(completed.returncode)",
                "",
            ]
        ),
        encoding="utf-8",
    )
    blocked.chmod(0o700)
    monkeypatch.setenv("PYTHON", blocked.as_posix())

    response = _call_mcp_server(
        _mcp_review_write_message(
            "apply_journal_bank_decisions",
            review_payload,
            run_intake,
        )
    )[0]["result"]

    assert response["isError"] is True
    assert response["structuredContent"]["error"] == (
        "Journal-Bank assurance preflight could not start."
    )
    assert blocked.as_posix() not in response["structuredContent"]["error"]
    assert _tree_snapshot(output_dir) == before
    assert list(output_dir.parent.glob(".journal-bank-apply-*")) == []


def test_journal_bank_mcp_apply_start_failure_is_fixed(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    output_dir = tmp_path / "run"
    output_dir, run_intake, review_payload, _, _ = _prepare_sealed_mcp_review_run(
        output_dir
    )
    before = _tree_snapshot(output_dir)
    wrapper = tmp_path / "python-journal-bank-disable-after-preflight"
    wrapper.write_text(
        "\n".join(
            [
                "#!/usr/bin/env python3",
                "from pathlib import Path",
                "import subprocess",
                "import sys",
                f"real = {str(Path(sys.executable))!r}",
                *_customer_run_preflight_passthrough_script_lines(),
                "completed = subprocess.run([real, *sys.argv[1:]], capture_output=True)",
                "sys.stdout.buffer.write(completed.stdout)",
                "sys.stderr.buffer.write(completed.stderr)",
                'if completed.returncode == 0 and "--preflight-only" in sys.argv:',
                "    Path(__file__).chmod(0o600)",
                "raise SystemExit(completed.returncode)",
                "",
            ]
        ),
        encoding="utf-8",
    )
    wrapper.chmod(0o700)
    monkeypatch.setenv("PYTHON", wrapper.as_posix())

    response = _call_mcp_server(
        _mcp_review_write_message(
            "apply_journal_bank_decisions",
            review_payload,
            run_intake,
        )
    )[0]["result"]

    assert response["isError"] is True
    assert response["structuredContent"]["error"] == (
        "Journal-Bank review application could not start."
    )
    assert wrapper.as_posix() not in response["structuredContent"]["error"]
    assert _tree_snapshot(output_dir) == before
    assert list(output_dir.parent.glob(".journal-bank-apply-*")) == []


def test_journal_bank_mcp_save_commits_only_decision_file(tmp_path: Path) -> None:
    output_dir = tmp_path / "run"
    output_dir, run_intake, review_payload, _, _ = _prepare_sealed_mcp_review_run(
        output_dir
    )
    before = _tree_snapshot(output_dir)

    response = _call_mcp_server(
        _mcp_review_write_message(
            "save_journal_bank_decisions",
            review_payload,
            run_intake,
        )
    )[0]["result"]

    assert response["isError"] is False
    assert (
        response["structuredContent"]["ui_decisions_path"]
        == (output_dir / "ui_decisions.json").as_posix()
    )
    after = _tree_snapshot(output_dir)
    changed_paths = {
        relative
        for relative in set(before) | set(after)
        if before.get(relative) != after.get(relative)
    }
    assert changed_paths == {"ui_decisions.json"}
    assert list(output_dir.parent.glob(".journal-bank-apply-*")) == []
