from __future__ import annotations

import csv
import hashlib
import importlib._bootstrap_external
import importlib.util
import json
import os
import shutil
import stat
import subprocess
import sys
import zipfile
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl
import pytest
from reportlab.pdfgen import canvas

from scripts.validate_plugin_review_contract import validate_contract

ROOT = Path(__file__).resolve().parents[2]
SCRIPT_DIR = ROOT / "plugins" / "check-entries" / "scripts"
CORE_PATH = SCRIPT_DIR / "check_entries_core.py"
APPLY_REVIEW_EDITS_PATH = SCRIPT_DIR / "apply_review_edits.py"
MCP_SERVER_PATH = ROOT / "plugins" / "check-entries" / "mcp" / "server.cjs"
JOURNAL_SAMPLING_SCRIPT_DIR = ROOT / "plugins" / "journal-sampling" / "scripts"
JOURNAL_SAMPLING_CORE_PATH = JOURNAL_SAMPLING_SCRIPT_DIR / "journal_sampling_core.py"
STUDIO_ARCHIVE_CORE_PATH = (
    ROOT / "plugins" / "studio-archive" / "scripts" / "archive_core.py"
)


def _load_client_ledger() -> Any:
    ledger_path = ROOT / "plugins" / "studio-archive" / "scripts" / "client_ledger.py"
    module_name = "test_check_entries_customer_ledger"
    ledger = sys.modules.get(module_name)
    if ledger is None:
        spec = importlib.util.spec_from_file_location(module_name, ledger_path)
        assert spec is not None and spec.loader is not None
        ledger = importlib.util.module_from_spec(spec)
        sys.modules[module_name] = ledger
        spec.loader.exec_module(ledger)
    return ledger


def _running_customer_output(tmp_path: Path) -> tuple[Path, str]:
    ledger = _load_client_ledger()
    client_root = tmp_path / "Managed Customer"
    client_root.mkdir(parents=True)
    client_id = "client_111111111111111111111111"
    ledger.create_client_manifest(client_root, client_id)
    engagement = ledger.create_engagement(client_root, client_id, "Test engagement")
    source = tmp_path / "managed-source.txt"
    source.write_text("managed input\n", encoding="utf-8")
    imported = ledger.import_document(
        client_root,
        client_id,
        engagement["engagement_id"],
        source,
        "source",
    )
    prepared = ledger.prepare_run(
        client_root,
        client_id,
        engagement["engagement_id"],
        "check-entries",
        "test-version",
        input_ids=[imported["receipt"]["input_id"]],
    )
    running = ledger.start_run(
        client_root,
        engagement["engagement_id"],
        prepared["run"]["run_id"],
    )
    return Path(running["output_dir"]), str(running["context"]["run_id"])


def _customer_context_path(output_dir: Path) -> Path:
    candidate = output_dir.resolve()
    while candidate != candidate.parent:
        context_path = candidate / "context.json"
        if context_path.is_file():
            return context_path
        candidate = candidate.parent
    raise AssertionError("customer-run context is unavailable")


def _rename_customer_output(output_dir: Path) -> tuple[Path, Path, Path]:
    context_path = _customer_context_path(output_dir)
    client_root = context_path.parents[5]
    renamed_root = client_root.with_name(f"{client_root.name} Renamed")
    relative_output = output_dir.relative_to(client_root)
    relative_context = context_path.relative_to(client_root)
    client_root.rename(renamed_root)
    return (
        renamed_root / relative_output,
        renamed_root / relative_context,
        output_dir,
    )


def _tree_snapshot(path: Path) -> dict[str, tuple[bytes, int]]:
    return {
        candidate.relative_to(path).as_posix(): (
            candidate.read_bytes(),
            candidate.stat().st_mode & 0o777,
        )
        for candidate in sorted(path.rglob("*"))
        if candidate.is_file()
    }


def _transaction_tree_state(
    root: Path,
) -> dict[str, tuple[str, bytes | str | None, int]]:
    state = {
        ".": (
            "directory",
            None,
            stat.S_IMODE(root.lstat().st_mode),
        )
    }
    for candidate in sorted(root.rglob("*")):
        entry = candidate.lstat()
        relative = candidate.relative_to(root).as_posix()
        mode = stat.S_IMODE(entry.st_mode)
        if stat.S_ISREG(entry.st_mode):
            state[relative] = ("file", candidate.read_bytes(), mode)
        elif stat.S_ISDIR(entry.st_mode):
            state[relative] = ("directory", None, mode)
        elif stat.S_ISLNK(entry.st_mode):
            state[relative] = ("symlink", os.readlink(candidate), mode)
        elif stat.S_ISFIFO(entry.st_mode):
            state[relative] = ("fifo", None, mode)
        else:
            state[relative] = ("special", None, mode)
    return state


def _install_review_transaction_commit_fault(
    monkeypatch: Any,
    tmp_path: Path,
    output_dir: Path,
    scenario: str,
) -> tuple[Path, Path]:
    preload = tmp_path / f"review-transaction-{scenario}.cjs"
    preload.write_text(
        """
"use strict";
const childProcess = require("node:child_process");
const fs = require("node:fs");
const path = require("node:path");
const originalRenameSync = fs.renameSync;
const canonical = path.resolve(process.env.REVIEW_TX_CANONICAL);
const external = path.resolve(process.env.REVIEW_TX_EXTERNAL);
const marker = path.resolve(process.env.REVIEW_TX_MARKER);
const scenario = process.env.REVIEW_TX_SCENARIO;
let triggered = false;
fs.renameSync = function reviewTransactionCommitFault(source, target) {
  const commitRoot = path.dirname(source);
  const isCommit =
    path.basename(source) === "candidate" &&
    path.basename(commitRoot).startsWith(
      ".generated-review-commit-",
    ) &&
    path.resolve(target) === canonical;
  if (!triggered && isCommit) {
    triggered = true;
    fs.writeFileSync(marker, "commit fault triggered\\n");
    const trustedBackup = path.join(commitRoot, "trusted-backup");
    if (scenario === "poison_file") {
      fs.writeFileSync(
        path.join(trustedBackup, "run_intake.json"),
        "POISONED SNAPSHOT\\n",
      );
    } else if (scenario === "delete_snapshot") {
      fs.rmSync(trustedBackup, { recursive: true, force: true });
    } else if (scenario === "symlink_snapshot") {
      fs.rmSync(trustedBackup, { recursive: true, force: true });
      fs.symlinkSync(external, trustedBackup, "dir");
    } else if (scenario === "fifo_snapshot") {
      fs.rmSync(trustedBackup, { recursive: true, force: true });
      fs.mkdirSync(trustedBackup, { mode: 0o700 });
      const fifo = path.join(trustedBackup, "poison.fifo");
      const created = childProcess.spawnSync("mkfifo", [fifo]);
      if (created.status !== 0) {
        throw new Error("FIFO probe setup failed");
      }
    } else if (scenario === "canonical_delete") {
      originalRenameSync.call(fs, source, target);
      fs.rmSync(target, { recursive: true, force: true });
      throw new Error(
        "/private/client/run injected late transaction failure",
      );
    } else {
      throw new Error("Unknown transaction fault scenario");
    }
    throw new Error(
      "/private/client/run injected late transaction failure",
    );
  }
  return originalRenameSync.call(fs, source, target);
};
""".lstrip(),
        encoding="utf-8",
    )
    external = tmp_path / f"external-{scenario}"
    external.mkdir()
    (external / "sentinel.bin").write_bytes(b"external unchanged")
    marker = tmp_path / f"commit-fault-{scenario}.marker"
    monkeypatch.setenv("NODE_OPTIONS", f"--require={preload}")
    monkeypatch.setenv("REVIEW_TX_CANONICAL", output_dir.as_posix())
    monkeypatch.setenv("REVIEW_TX_EXTERNAL", external.as_posix())
    monkeypatch.setenv("REVIEW_TX_MARKER", marker.as_posix())
    monkeypatch.setenv("REVIEW_TX_SCENARIO", scenario)
    return marker, external


def _install_review_transaction_root_relocation(
    monkeypatch: Any,
    tmp_path: Path,
) -> Path:
    preload = tmp_path / "review-transaction-root-relocation.cjs"
    preload.write_text(
        """
"use strict";
const fs = require("node:fs");
const Module = require("node:module");
const path = require("node:path");
const originalLoad = Module._load;
const marker = path.resolve(process.env.REVIEW_TX_MARKER);
let triggered = false;
Module._load = function reviewTransactionRootRelocation(
  request,
  parent,
  isMain,
) {
  const loaded = originalLoad.call(this, request, parent, isMain);
  if (
    typeof request !== "string" ||
    !request.endsWith("review_output_transaction.cjs") ||
    typeof loaded.withGeneratedReviewOutputTransaction !== "function"
  ) {
    return loaded;
  }
  return {
    ...loaded,
    withGeneratedReviewOutputTransaction(outputDir, operation, options) {
      return loaded.withGeneratedReviewOutputTransaction(
        outputDir,
        (context) => {
          const envelope = operation(context);
          if (!triggered && context.workingOutputDir) {
            triggered = true;
            const transactionRoot = path.dirname(context.workingOutputDir);
            fs.renameSync(
              transactionRoot,
              `${transactionRoot}-moved`,
            );
            fs.writeFileSync(marker, "transaction root relocated\\n");
          }
          return envelope;
        },
        options,
      );
    },
  };
};
""".lstrip(),
        encoding="utf-8",
    )
    marker = tmp_path / "transaction-root-relocation.marker"
    monkeypatch.setenv("NODE_OPTIONS", f"--require={preload}")
    monkeypatch.setenv("REVIEW_TX_MARKER", marker.as_posix())
    return marker


def _seal_review_payload(payload: dict[str, Any]) -> dict[str, Any]:
    content = dict(payload)
    content.pop("content_sha256", None)
    encoded = json.dumps(
        content,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    ).encode("utf-8")
    payload["content_sha256"] = hashlib.sha256(encoded).hexdigest()
    return payload


def load_core() -> Any:
    if str(SCRIPT_DIR) not in sys.path:
        sys.path.insert(0, str(SCRIPT_DIR))
    spec = importlib.util.spec_from_file_location("check_entries_core", CORE_PATH)
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def load_apply_review_edits() -> Any:
    spec = importlib.util.spec_from_file_location(
        "check_entries_apply_review_edits",
        APPLY_REVIEW_EDITS_PATH,
    )
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def load_studio_archive_core() -> Any:
    """Load the local Studio Archive implementation for integration tests."""

    module_name = "studio_archive_core_for_check_entries"
    spec = importlib.util.spec_from_file_location(module_name, STUDIO_ARCHIVE_CORE_PATH)
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    sys.modules[module_name] = module
    spec.loader.exec_module(module)
    return module


def _save_workbook(path: Path, rows: list[list[Any]]) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    workbook.save(path)


def _text_pdf_bytes(lines: list[str]) -> bytes:
    """Return a small real text PDF that production extraction can replay."""

    buffer = BytesIO()
    document = canvas.Canvas(buffer)
    vertical = 800
    for line in lines:
        document.drawString(72, vertical, line)
        vertical -= 18
    document.save()
    return buffer.getvalue()


def _load_journal_sampling_core() -> Any:
    previous_bootstrap = sys.modules.pop("implementation_bootstrap", None)
    script_dir = str(JOURNAL_SAMPLING_SCRIPT_DIR)
    inserted = script_dir not in sys.path
    if inserted:
        sys.path.insert(0, script_dir)
    spec = importlib.util.spec_from_file_location(
        "journal_sampling_core_for_check_entries",
        JOURNAL_SAMPLING_CORE_PATH,
    )
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    try:
        spec.loader.exec_module(module)
    finally:
        sys.modules.pop("implementation_bootstrap", None)
        if previous_bootstrap is not None:
            sys.modules["implementation_bootstrap"] = previous_bootstrap
        if inserted:
            sys.path.remove(script_dir)
    return module


def _qualified_journal(
    root: Path,
    rows: list[dict[str, object]],
    *,
    source_path: Path | None = None,
    normalization_name: str = "journal_normalization",
    client_engagement: dict[str, Any] | None = None,
    write_source: bool = True,
) -> Path:
    root.mkdir(parents=True, exist_ok=True)
    managed_fixture: tuple[Any, Path, str, str] | None = None
    source_path = source_path or root / "entries.xlsx"
    source_path.parent.mkdir(parents=True, exist_ok=True)
    normalization_dir = root / normalization_name
    if write_source:
        _save_workbook(
            source_path,
            [
                [
                    "Data",
                    "Nr. Reg",
                    "Conto",
                    "Descrizione conto",
                    "Descrizione",
                    "Dare",
                    "Avere",
                ],
                *[
                    [
                        row["date"],
                        row.get("movement"),
                        row.get("account", "4000"),
                        row.get("account_desc", "Trade payable"),
                        row.get("description"),
                        row.get("debit"),
                        row.get("credit"),
                    ]
                    for row in rows
                ],
            ],
        )
    if client_engagement is None:
        ledger = _load_client_ledger()
        fixture_root = root / ".managed-journal-fixtures"
        fixture_number = 1
        client_root = fixture_root / f"Client-{fixture_number}"
        while client_root.exists():
            fixture_number += 1
            client_root = fixture_root / f"Client-{fixture_number}"
        client_root.mkdir(parents=True)
        client_id = "client_" + "1" * 24
        ledger.create_client_manifest(client_root, client_id)
        engagement = ledger.create_engagement(
            client_root,
            client_id,
            "Check Entries upstream fixture",
        )
        imported = ledger.import_document(
            client_root,
            client_id,
            engagement["engagement_id"],
            source_path,
            "journal",
        )
        prepared = ledger.prepare_run(
            client_root,
            client_id,
            engagement["engagement_id"],
            "journal-sampling",
            "test-version",
            input_ids=[imported["receipt"]["input_id"]],
        )
        running = ledger.start_run(
            client_root,
            engagement["engagement_id"],
            prepared["run"]["run_id"],
        )
        client_engagement = running["context"]
        source_path = Path(client_engagement["input_bindings"][0]["path"])
        normalization_dir = Path(client_engagement["output_dir"]) / "normalization"
        managed_fixture = (
            ledger,
            client_root,
            engagement["engagement_id"],
            prepared["run"]["run_id"],
        )
    sampling_core = _load_journal_sampling_core()
    sampling_core.inspect_path(
        source_path,
        normalization_dir,
        client_engagement=client_engagement,
    )
    recipe_path = normalization_dir / "suggested_recipe.json"
    recipe = json.loads(recipe_path.read_text(encoding="utf-8"))
    for index, (source_name, entry) in enumerate(
        recipe["files"].items(),
        start=1,
    ):
        requested_currencies = {
            str(row["currency"]).upper()
            for row in rows
            if row.get("currency") not in (None, "")
        }
        if len(requested_currencies) == 1:
            entry["currency"] = next(iter(requested_currencies))
        qualification = entry["qualification"]
        contract = sampling_core._mapping_contract(
            parser=entry["parser"],
            source_family=entry["source_family"],
            header_rows=entry.get("header_rows", []),
            mapping=entry.get("mapping", {}),
            layout=entry.get("layout", {}),
            excluded_monetary_columns=entry.get(
                "excluded_monetary_columns",
                [],
            ),
            posting_identity=entry.get("posting_identity", "source_row"),
            carry_forward_fields=entry.get("carry_forward_fields", []),
            currency=entry.get("currency", "EUR"),
            unit=entry.get("unit", "currency"),
            decimal_separator=entry.get("decimal_separator"),
            thousands_separator=entry.get("thousands_separator"),
            amount_sign_convention=entry.get("amount_sign_convention"),
        )
        adapter_id = (
            sampling_core.PRINT_ADAPTER_ID
            if entry["parser"] == "print_friendly_excel"
            else sampling_core.TABULAR_ADAPTER_ID
        )
        decision_id = f"decision.journal_mapping.{index}"
        qualification["status"] = "reviewed"
        qualification["mapping_sha256"] = sampling_core.canonical_json_sha256(contract)
        qualification["decision_ref"] = decision_id
        qualification["decision_receipt"] = (
            sampling_core.build_reviewed_decision_receipt(
                decision_id=decision_id,
                decision_type="source_mapping",
                status="reviewed",
                reviewer_ref="reviewer.check_entries_test",
                reviewed_on="2026-07-24",
                adapter_id=adapter_id,
                adapter_version=sampling_core.ADAPTER_VERSION,
                source_artifact_refs=[
                    sampling_core._source_artifact_ref(Path(source_name))
                ],
                content=contract,
            )
        )
    recipe_path.write_text(
        json.dumps(recipe, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    normalized = sampling_core.normalize_path(
        source_path,
        normalization_dir,
        recipe_path,
        client_engagement=client_engagement,
    )
    assert normalized.diagnostics["population_status"] == "complete"
    if managed_fixture is not None:
        ledger, client_root, engagement_id, run_id = managed_fixture
        declarations = [
            {
                "artifact_id": f"internal.journal_fixture.{index:03d}",
                "path": artifact.relative_to(normalization_dir.parent).as_posix(),
                "purpose": "Preserve a Journal Sampling assurance fixture.",
                "audience": "internal",
                "media_type": (
                    "application/json" if artifact.suffix == ".json" else "text/csv"
                ),
            }
            for index, artifact in enumerate(
                sorted(
                    path
                    for path in normalization_dir.parent.rglob("*")
                    if path.is_file()
                ),
                start=1,
            )
        ]
        ledger.finalize_run(client_root, engagement_id, run_id, declarations)
    return normalization_dir / "normalized_journal.csv"


def _qualified_journal_source(normalized: Path) -> Path:
    """Resolve the managed source bound to one qualified Journal run."""

    diagnostics = json.loads(
        (normalized.parent / "normalization_diagnostics.json").read_text(
            encoding="utf-8"
        )
    )
    reference = Path(diagnostics["input"])
    assert diagnostics["path_reference"] == "run_root_relative"
    assert not reference.is_absolute() and ".." not in reference.parts
    return (_customer_context_path(normalized).parent / reference).resolve(strict=True)


def _journal_tamper_source(normalized: Path) -> Path:
    """Move source-receipt testing off the immutable v2 input binding."""

    original = _qualified_journal_source(normalized)
    run_root = _customer_context_path(normalized).parent
    source_root = run_root / "outputs" / "receipt-test-source"
    source_root.mkdir(parents=True)
    source = source_root / original.name
    shutil.copy2(original, source)
    diagnostics_path = normalized.parent / "normalization_diagnostics.json"
    diagnostics = json.loads(diagnostics_path.read_text(encoding="utf-8"))
    diagnostics["input"] = source.relative_to(run_root).as_posix()
    diagnostics["source_root"] = source_root.relative_to(run_root).as_posix()
    _seal_review_payload(diagnostics)
    diagnostics_path.write_text(
        json.dumps(diagnostics, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return source


def _portable_context_projection(context: dict[str, Any]) -> dict[str, Any]:
    """Return the stable v2 context identity persisted by managed workflows."""

    runtime_fields = {
        "studio_client_folder",
        "input_bindings",
        "input_dir",
        "workspace_root",
        "output_dir",
        "run_root",
        "run_manifest_path",
        "input_manifest_path",
        "context_path",
    }
    return {key: value for key, value in context.items() if key not in runtime_fields}


def _client_bound_check_inputs(
    tmp_path: Path,
    rows: list[dict[str, object]],
) -> tuple[Path, Path, dict[str, Any], dict[str, Any]]:
    """Create one valid Journal Sampling -> Check Entries client boundary."""

    from vera_assurance import (
        build_client_engagement_context,
        build_studio_client_folder_binding,
    )

    archive_root = tmp_path / "Studio"
    client_root = archive_root / "Zecca SPA"
    input_dir = client_root / "Vera engagements" / ("eng_" + "1" * 24) / "inputs"
    journal_source = input_dir / "journal" / "entries.xlsx"
    support = input_dir / "support"
    journal_source.parent.mkdir(parents=True)
    support.mkdir(parents=True)
    workspace_root = tmp_path / "vera-client-work"
    workspace_root.mkdir()
    relative_dir = "Zecca SPA"
    scope_id = (
        "scope_"
        + hashlib.sha256(relative_dir.casefold().encode("utf-8")).hexdigest()[:24]
    )
    folder = build_studio_client_folder_binding(
        studio_client_id="client_" + "a" * 24,
        scope_id=scope_id,
        archive_root=archive_root,
        scope_relative_dir=relative_dir,
        client_root=client_root,
        display_name=relative_dir,
    )
    journal_context = build_client_engagement_context(
        studio_client_folder=folder,
        engagement_id="eng_" + "1" * 24,
        workflow_id="journal-sampling",
        run_id="run_" + "2" * 24,
        input_dir=input_dir,
        workspace_root=workspace_root,
    )
    normalized = _qualified_journal(
        Path(journal_context["output_dir"]),
        rows,
        source_path=journal_source,
        normalization_name="normalization",
        client_engagement=journal_context,
    )
    check_context = build_client_engagement_context(
        studio_client_folder=folder,
        engagement_id="eng_" + "1" * 24,
        workflow_id="check-entries",
        run_id="run_" + "3" * 24,
        input_dir=input_dir,
        workspace_root=workspace_root,
    )
    return normalized, support, journal_context, check_context


def _v2_client_bound_check_inputs(
    tmp_path: Path,
    rows: list[dict[str, object]],
) -> tuple[Path, Path, dict[str, Any], dict[str, Any]]:
    """Create one running v2 Journal Sampling -> Check Entries boundary."""

    check_core = load_core()
    archive_core = load_studio_archive_core()
    archive_root = tmp_path / "Studio"
    client_root = archive_root / "Zecca SPA"
    client_root.mkdir(parents=True)
    state_dir = tmp_path / "private-state"
    configured = archive_core.configure_archive(archive_root, state_dir=state_dir)
    scope_id = next(
        item["scope_id"]
        for item in configured["scopes"]
        if item["display_name"] == "Zecca SPA"
    )
    client_id = archive_core.set_studio_client_identity(
        scope_id,
        legal_names=["Zecca SPA"],
        state_dir=state_dir,
    )["client"]["client_id"]
    engagement_id = archive_core.create_studio_client_engagement(
        client_id,
        "Journal sample checks",
        state_dir=state_dir,
    )["engagement"]["engagement_id"]

    received = tmp_path / "received"
    received.mkdir()
    journal_source = received / "journal.xlsx"
    _save_workbook(
        journal_source,
        [
            [
                "Data",
                "Nr. Reg",
                "Conto",
                "Descrizione conto",
                "Descrizione",
                "Dare",
                "Avere",
            ],
            *[
                [
                    row["date"],
                    row.get("movement"),
                    row.get("account", "4000"),
                    row.get("account_desc", "Trade payable"),
                    row.get("description"),
                    row.get("debit"),
                    row.get("credit"),
                ]
                for row in rows
            ],
        ],
    )
    journal_import = archive_core.import_studio_client_document(
        client_id,
        journal_source,
        "journal",
        engagement_id=engagement_id,
        state_dir=state_dir,
    )
    journal_run = archive_core.prepare_studio_client_workflow(
        engagement_id,
        "journal-sampling",
        input_ids=[journal_import["input_id"]],
        state_dir=state_dir,
    )
    journal_run_id = journal_run["run"]["run_id"]
    archive_core.start_studio_client_workflow(
        client_id,
        engagement_id,
        journal_run_id,
        state_dir=state_dir,
    )
    journal_context = journal_run["client_engagement"]
    journal_output = Path(journal_context["output_dir"])
    journal_execution = next(
        Path(item["path"])
        for item in journal_context["input_bindings"]
        if item["binding_id"] == journal_import["input_id"]
    )
    normalized = _qualified_journal(
        journal_output,
        rows,
        source_path=journal_execution,
        normalization_name="normalization",
        client_engagement=journal_context,
        write_source=False,
    )
    sampling_core = _load_journal_sampling_core()
    sampling_core.run_sample(
        normalized,
        journal_output / "sample",
        method="systematic",
        size=1,
        client_engagement=journal_context,
    )

    special_artifact_ids = {
        "normalization/normalized_journal.csv": "prepared.normalized_journal",
        "normalization/normalization_diagnostics.json": (
            "internal.normalization_diagnostics"
        ),
        "sample/journal_sample.csv": "prepared.journal_sample_csv",
    }
    declarations = []
    for index, artifact_path in enumerate(
        sorted(path for path in journal_output.rglob("*") if path.is_file()),
        start=1,
    ):
        relative_path = artifact_path.relative_to(journal_output).as_posix()
        declarations.append(
            {
                "artifact_id": special_artifact_ids.get(
                    relative_path, f"internal.journal_sampling.{index:03d}"
                ),
                "path": relative_path,
                "purpose": f"Preserve Journal Sampling artifact {relative_path}.",
                "audience": (
                    "review" if relative_path.startswith("sample/") else "internal"
                ),
                "media_type": (
                    "application/json"
                    if artifact_path.suffix == ".json"
                    else (
                        "text/csv"
                        if artifact_path.suffix == ".csv"
                        else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                ),
            }
        )
    archive_core.finalize_studio_client_workflow(
        client_id,
        engagement_id,
        journal_run_id,
        declarations,
        state_dir=state_dir,
    )
    archive_core.complete_studio_client_workflow(
        client_id,
        engagement_id,
        journal_run_id,
        state_dir=state_dir,
    )
    upstream_artifacts = [
        {
            "run_id": journal_run_id,
            "artifact_id": declaration["artifact_id"],
            "role": (
                "journal_sample"
                if declaration["artifact_id"] == "prepared.journal_sample_csv"
                else "journal_normalization"
            ),
        }
        for declaration in declarations
        if declaration["path"] in check_core.JOURNAL_HANDOFF_ARTIFACT_PATHS
    ]

    support_source = received / "invoice.pdf"
    support_source.write_bytes(
        _text_pdf_bytes(
            [
                "Invoice",
                "Movement M-1001",
                "Supplier VAT: 01234567890",
                "02/01/2025 EUR 123.45",
            ]
        )
    )
    support_import = archive_core.import_studio_client_document(
        client_id,
        support_source,
        "support",
        engagement_id=engagement_id,
        state_dir=state_dir,
    )
    check_run = archive_core.prepare_studio_client_workflow(
        engagement_id,
        "check-entries",
        input_ids=[support_import["input_id"]],
        upstream_artifacts=upstream_artifacts,
        state_dir=state_dir,
    )
    archive_core.start_studio_client_workflow(
        client_id,
        engagement_id,
        check_run["run"]["run_id"],
        state_dir=state_dir,
    )
    check_context = check_run["client_engagement"]
    check_journal = next(
        Path(item["path"])
        for item in check_context["input_bindings"]
        if item.get("upstream_artifact_id") == "prepared.normalized_journal"
    )
    support = next(
        Path(item["path"])
        for item in check_context["input_bindings"]
        if item["binding_id"] == support_import["input_id"]
    )
    return check_journal, support, journal_context, check_context


def _fatturapa_xml(
    *,
    document_type: str = "TD01",
    number: str = "INV-42",
    invoice_date: str = "2025-01-02",
    amount: str = "123.45",
    supplier: str = "ACME SPA",
    currency: str = "EUR",
) -> bytes:
    """Return the smallest representative FatturaPA invoice fixture."""

    return f"""<?xml version="1.0" encoding="UTF-8"?>
<FatturaElettronica>
  <FatturaElettronicaHeader>
    <CedentePrestatore><DatiAnagrafici><IdFiscaleIVA><IdPaese>IT</IdPaese><IdCodice>01234567890</IdCodice></IdFiscaleIVA><Anagrafica><Denominazione>{supplier}</Denominazione></Anagrafica></DatiAnagrafici></CedentePrestatore>
    <CessionarioCommittente><DatiAnagrafici><IdFiscaleIVA><IdPaese>IT</IdPaese><IdCodice>09876543210</IdCodice></IdFiscaleIVA><Anagrafica><Denominazione>CLIENTE SRL</Denominazione></Anagrafica></DatiAnagrafici></CessionarioCommittente>
  </FatturaElettronicaHeader>
  <FatturaElettronicaBody><DatiGenerali><DatiGeneraliDocumento><TipoDocumento>{document_type}</TipoDocumento><Divisa>{currency}</Divisa><Data>{invoice_date}</Data><Numero>{number}</Numero><ImportoTotaleDocumento>{amount}</ImportoTotaleDocumento></DatiGeneraliDocumento></DatiGenerali></FatturaElettronicaBody>
</FatturaElettronica>
""".encode()


def _reviewed_party_recipe(
    core: Any,
    normalized: Path,
    path: Path,
    *,
    tax_id: str = "01234567890",
    expected_role: str = "supplier",
) -> Path:
    from vera_assurance import build_reviewed_decision_receipt

    frame, _ = core._load_journal_entries(normalized, {})
    decisions = []
    for index, row in enumerate(frame.to_dicts(), start=1):
        decisions.append(
            build_reviewed_decision_receipt(
                decision_id=f"decision.check_entries_party.{index}",
                decision_type="check_entries_party_perimeter",
                status="reviewed",
                reviewer_ref="reviewer.check_entries_test",
                reviewed_on="2026-07-24",
                adapter_id="check_entries.party_perimeter",
                adapter_version="1",
                source_artifact_refs=["source.normalized_journal"],
                content={
                    "prepared_entry_id": row["prepared_entry_id"],
                    "expected_role": expected_role,
                    "expected_tax_ids": [tax_id],
                    "expected_names": [],
                    "name_normalization_contract": None,
                },
            )
        )
    recipe = (
        json.loads(path.read_text(encoding="utf-8"))
        if path.is_file()
        else {"version": 2}
    )
    recipe["reviewed_party_perimeters"] = decisions
    recipe.setdefault("reviewed_support_relationships", [])
    path.write_text(
        json.dumps(recipe, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return path


def _reviewed_name_party_recipe(
    core: Any,
    normalized: Path,
    path: Path,
    *,
    name: str = "ACME SPA",
    expected_role: str = "supplier",
) -> Path:
    from vera_assurance import build_reviewed_decision_receipt

    frame, _ = core._load_journal_entries(normalized, {})
    row = frame.to_dicts()[0]
    decision = build_reviewed_decision_receipt(
        decision_id="decision.check_entries_party.name",
        decision_type="check_entries_party_perimeter",
        status="reviewed",
        reviewer_ref="reviewer.check_entries_test",
        reviewed_on="2026-07-24",
        adapter_id="check_entries.party_perimeter",
        adapter_version="1",
        source_artifact_refs=["source.normalized_journal"],
        content={
            "prepared_entry_id": row["prepared_entry_id"],
            "expected_role": expected_role,
            "expected_tax_ids": [],
            "expected_names": [name],
            "name_normalization_contract": {"contract_id": "casefold_alnum_v1"},
        },
    )
    path.write_text(
        json.dumps(
            {
                "version": 2,
                "reviewed_party_perimeters": [decision],
                "reviewed_support_relationships": [],
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )
    return path


def _reviewed_relationship_recipe(
    core: Any,
    normalized: Path,
    support_path: Path,
    path: Path,
    *,
    support_locator: str,
) -> Path:
    from vera_assurance import build_reviewed_decision_receipt

    frame, _ = core._load_journal_entries(normalized, {})
    _, captures = core._capture_support(support_path)
    assert len(captures) == 1
    row = frame.to_dicts()[0]
    artifact_id = captures[0].receipt["artifact_id"]
    decision = build_reviewed_decision_receipt(
        decision_id="decision.check_entries_relationship.1",
        decision_type="check_entries_support_relationship",
        status="reviewed",
        reviewer_ref="reviewer.check_entries_test",
        reviewed_on="2026-07-24",
        adapter_id="check_entries.relationship",
        adapter_version="1",
        source_artifact_refs=["source.normalized_journal", artifact_id],
        content={
            "prepared_entry_id": row["prepared_entry_id"],
            "support_artifact_id": artifact_id,
            "support_locator": support_locator,
            "relationship_status": "confirmed",
            "recording_exception": (
                "The journal export omits the document party identifier; "
                "the reviewer confirmed the exact support relationship."
            ),
        },
    )
    path.write_text(
        json.dumps(
            {
                "version": 2,
                "reviewed_party_perimeters": [],
                "reviewed_support_relationships": [decision],
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )
    return path


def _reviewed_pdf_assertion_recipe(
    core: Any,
    normalized: Path,
    support_path: Path,
    path: Path,
    *,
    support_locator: str,
    currency: bool = False,
    direction: bool = False,
    entry_index: int = 0,
    decision_suffix: str = "1",
) -> Path:
    from vera_assurance import build_reviewed_decision_receipt

    frame, _ = core._load_journal_entries(normalized, {})
    root, captures = core._capture_support(support_path)
    del root
    captured_locator = support_locator.split("!/", 1)[0]
    capture = next(item for item in captures if item.relative_path == captured_locator)
    row = frame.to_dicts()[entry_index]
    recipe = (
        json.loads(path.read_text(encoding="utf-8"))
        if path.is_file()
        else {"version": 2}
    )
    recipe.setdefault("reviewed_party_perimeters", [])
    recipe.setdefault("reviewed_support_relationships", [])
    recipe.setdefault("reviewed_currency_decisions", [])
    recipe.setdefault("reviewed_direction_decisions", [])
    source_refs = ["source.normalized_journal", capture.receipt["artifact_id"]]
    if currency:
        recipe["reviewed_currency_decisions"].append(
            build_reviewed_decision_receipt(
                decision_id=f"decision.check_entries_currency.{decision_suffix}",
                decision_type="check_entries_currency",
                status="reviewed",
                reviewer_ref="reviewer.check_entries_test",
                reviewed_on="2026-07-24",
                adapter_id="check_entries.currency",
                adapter_version="1",
                source_artifact_refs=source_refs,
                content={
                    "prepared_entry_id": row["prepared_entry_id"],
                    "support_artifact_id": capture.receipt["artifact_id"],
                    "support_locator": support_locator,
                    "expected_currency": row["currency"],
                    "currency_status": "confirmed",
                    "recording_exception": (
                        "The reviewer confirmed the currency where the PDF "
                        "uses an otherwise ambiguous symbol."
                    ),
                },
            )
        )
    if direction:
        expected_direction = (
            "debit" if Decimal(str(row["amount_signed"])) > 0 else "credit"
        )
        recipe["reviewed_direction_decisions"].append(
            build_reviewed_decision_receipt(
                decision_id=f"decision.check_entries_direction.{decision_suffix}",
                decision_type="check_entries_direction",
                status="reviewed",
                reviewer_ref="reviewer.check_entries_test",
                reviewed_on="2026-07-24",
                adapter_id="check_entries.direction",
                adapter_version="1",
                source_artifact_refs=source_refs,
                content={
                    "prepared_entry_id": row["prepared_entry_id"],
                    "support_artifact_id": capture.receipt["artifact_id"],
                    "support_locator": support_locator,
                    "expected_direction": expected_direction,
                    "direction_status": "confirmed",
                    "recording_exception": (
                        "The reviewer confirmed which journal line the exact "
                        "support artifact and locator substantiate."
                    ),
                },
            )
        )
    path.write_text(
        json.dumps(recipe, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return path


def _call_mcp_server(
    messages: list[dict[str, object]],
    *,
    server_path: Path = MCP_SERVER_PATH,
) -> list[dict[str, object]]:
    node = shutil.which("node")
    if node is None:
        pytest.skip("Node.js is required to exercise the Check Entries MCP server.")
    completed = subprocess.run(
        [node, str(server_path), "--stdio"],
        input="\n".join(json.dumps(message) for message in messages) + "\n",
        capture_output=True,
        text=True,
        check=True,
        timeout=10,
    )
    return [json.loads(line) for line in completed.stdout.splitlines() if line.strip()]


def _check_transaction_call(
    tool_name: str,
    arguments: dict[str, Any],
) -> dict[str, Any]:
    response = _call_mcp_server(
        [
            {
                "jsonrpc": "2.0",
                "id": 1,
                "method": "tools/call",
                "params": {"name": tool_name, "arguments": arguments},
            }
        ]
    )[0]
    return response["result"]["structuredContent"]


def _portable_check_transaction_case(
    tmp_path: Path,
) -> tuple[Path, dict[str, Any]]:
    output_dir, run_id = _running_customer_output(tmp_path)
    context_path = _customer_context_path(output_dir)
    core = load_core()
    context = core.load_client_engagement_context(context_path)
    source_path = Path(context["input_bindings"][0]["path"])
    core.write_run_intake(
        output_dir,
        source_path,
        source_path,
        normalization_diagnostics_path=source_path,
        recipe_path=None,
        language="en",
        document_language="en",
        amount_tolerance="0.01",
        date_window_days=0,
        mapping={},
        journal_row_count=1,
        pdf_count=1,
        client_engagement=context,
    )
    run_intake = json.loads(
        (output_dir / "run_intake.json").read_text(encoding="utf-8")
    )
    review_payload = {
        "schema_version": "2.0",
        "plugin": "check-entries",
        "workflow": "check-entries",
        "run_id": run_id,
        "source_paths": list(run_intake["input_paths"]),
        "review_type": "journal_entry_support_review",
        "items": [
            {
                "id": "entry-1",
                "item_type": "supported_entry",
                "title": "Managed entry",
                "source_path": run_intake["input_paths"][0],
                "output_path": "final_artifacts.json",
                "allowed_actions": ["accept", "mark_unclear", "skip"],
                "recommended_action": "accept",
                "evidence": [],
                "data": {"status": "ok"},
                "status": "needs_review",
            }
        ],
        "item_count": 1,
        "columns": [],
        "source_artifacts": {"run_intake": "run_intake.json"},
        "evidence": {},
        "allowed_actions": ["accept", "mark_unclear", "skip"],
        "status": "ready_for_review",
        "summary": {"result_row_count": 1, "ok_count": 1},
    }
    _seal_review_payload(review_payload)
    final_artifacts = {
        "schema_version": "2.0",
        "plugin": "check-entries",
        "workflow": "check-entries",
        "run_id": run_id,
        "outputs": [],
        "status": "written_pending_review",
    }
    for name, payload in (
        ("review_payload.json", review_payload),
        ("final_artifacts.json", final_artifacts),
    ):
        (output_dir / name).write_text(
            json.dumps(payload, indent=2) + "\n",
            encoding="utf-8",
        )
    return output_dir, {
        "run_intake": run_intake,
        "review_payload": review_payload,
        "final_artifacts": final_artifacts,
        "decisions": [{"item_id": "entry-1", "action": "accept"}],
    }


def test_check_review_save_and_apply_survive_customer_folder_rename(
    tmp_path: Path,
) -> None:
    output_dir, arguments = _portable_check_transaction_case(tmp_path)
    old_customer_root = _customer_context_path(output_dir).parents[5].as_posix()
    old_output = output_dir.as_posix()
    assert arguments["run_intake"]["path_reference"] == "run_root_relative"
    assert arguments["run_intake"]["output_dir"] == "outputs"
    assert all(
        not Path(value).is_absolute()
        for value in arguments["run_intake"]["input_paths"]
    )

    renamed_output, current_context, stale_output = _rename_customer_output(output_dir)
    arguments["client_engagement"] = current_context.as_posix()

    saved = _check_transaction_call(
        "save_check_entries_decisions",
        arguments,
    )
    applied = _check_transaction_call(
        "apply_check_entries_decisions",
        arguments,
    )

    assert saved["ok"] is True, saved
    assert saved["persisted"] is True
    assert applied["ok"] is True, applied
    assert applied["persisted"] is True
    assert (renamed_output / "ui_decisions.json").is_file()
    assert (renamed_output / "applied_decisions.json").is_file()
    assert not stale_output.exists()
    assert old_output not in arguments["run_intake"]["output_dir"]
    assert all(
        old_customer_root not in artifact.read_text(encoding="utf-8")
        for artifact in renamed_output.rglob("*")
        if artifact.is_file() and artifact.suffix.lower() in {".json", ".md"}
    )


def test_check_review_rejects_run_root_escape_without_writing(
    tmp_path: Path,
) -> None:
    output_dir, arguments = _portable_check_transaction_case(tmp_path)
    arguments["client_engagement"] = _customer_context_path(output_dir).as_posix()
    forged = json.loads(json.dumps(arguments))
    forged["run_intake"]["output_dir"] = "../outside"
    before = _transaction_tree_state(output_dir)

    result = _check_transaction_call(
        "save_check_entries_decisions",
        forged,
    )

    assert result["ok"] is False
    assert "leaves the customer run" in result["error"]
    assert _transaction_tree_state(output_dir) == before
    assert not (output_dir.parent.parent / "outside").exists()


def _supported_assurance_run(
    monkeypatch: Any,
    root: Path,
) -> tuple[Any, Path, dict[str, Any], dict[str, Any]]:
    del monkeypatch
    core = load_core()
    normalized, support_path, _, check_context = _v2_client_bound_check_inputs(
        root,
        [
            {
                "date": "2025-01-02",
                "movement": "M-1001",
                "description": "Invoice payment",
                "debit": "123.45",
            }
        ],
    )
    output_dir = Path(check_context["output_dir"]) / "checks"
    recipe_path = _reviewed_party_recipe(
        core,
        normalized,
        root / "check_entries_recipe.json",
    )
    _reviewed_pdf_assertion_recipe(
        core,
        normalized,
        support_path,
        recipe_path,
        support_locator=support_path.name,
        direction=True,
    )
    core.run_entry_checks(
        normalized,
        support_path,
        output_dir,
        recipe_path,
        client_engagement=check_context,
    )
    return (
        core,
        output_dir,
        json.loads((output_dir / "review_payload.json").read_text()),
        json.loads((output_dir / "run_intake.json").read_text()),
    )


def test_managed_journal_to_check_review_survives_customer_folder_rename(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    _, output_dir, review_payload, run_intake = _supported_assurance_run(
        monkeypatch,
        tmp_path,
    )
    final_artifacts = json.loads(
        (output_dir / "final_artifacts.json").read_text(encoding="utf-8")
    )
    decisions = [
        {
            "item_id": item["id"],
            "action": (
                "accept"
                if "accept" in item["allowed_actions"]
                else item["allowed_actions"][0]
            ),
        }
        for item in review_payload["items"]
    ]
    old_customer_root = _customer_context_path(output_dir).parents[5]
    old_customer_path = old_customer_root.as_posix()
    renamed_output, current_context, stale_output = _rename_customer_output(output_dir)
    arguments = {
        "client_engagement": current_context.as_posix(),
        "run_intake": run_intake,
        "review_payload": review_payload,
        "final_artifacts": final_artifacts,
        "decisions": decisions,
    }

    saved = _check_transaction_call(
        "save_check_entries_decisions",
        arguments,
    )
    arguments.update(
        {
            "run_intake": json.loads(
                (renamed_output / "run_intake.json").read_text(encoding="utf-8")
            ),
            "review_payload": json.loads(
                (renamed_output / "review_payload.json").read_text(encoding="utf-8")
            ),
            "final_artifacts": json.loads(
                (renamed_output / "final_artifacts.json").read_text(encoding="utf-8")
            ),
        }
    )
    applied = _check_transaction_call(
        "apply_check_entries_decisions",
        arguments,
    )

    assert saved["ok"] is True, saved
    assert saved["persisted"] is True
    assert applied["ok"] is True, applied
    assert applied["persisted"] is True
    assert (
        applied["applied_decisions"]["assurance_preflight"]["assurance_replayed"]
        is True
    )
    assert (renamed_output / "ui_decisions.json").is_file()
    assert (renamed_output / "applied_decisions.json").is_file()
    assert not stale_output.exists()
    assert run_intake["path_reference"] == "run_root_relative"
    assert old_customer_path not in run_intake["output_dir"]
    renamed_client_root = current_context.parents[5]
    managed_output_files = [
        artifact
        for output_root in renamed_client_root.glob("Vera/engagements/*/runs/*/outputs")
        for artifact in output_root.rglob("*")
        if artifact.is_file() and artifact.suffix.lower() in {".json", ".md"}
    ]
    assert managed_output_files
    assert all(
        old_customer_path not in artifact.read_text(encoding="utf-8")
        for artifact in managed_output_files
    )


def _replace_receipt(
    receipts: list[dict[str, Any]],
    replacement: dict[str, Any],
) -> list[dict[str, Any]]:
    return [
        (
            replacement
            if receipt.get("artifact_id") == replacement["artifact_id"]
            else receipt
        )
        for receipt in receipts
    ]


def _managed_run_context_path(output_dir: Path) -> Path:
    """Return the current context path owning a managed Check output."""

    return output_dir.parent.parent / "context.json"


def _managed_check_mcp_arguments(
    output_dir: Path,
    arguments: dict[str, Any],
) -> dict[str, Any]:
    """Bind a durable MCP write to its current managed-run context."""

    return {
        "client_engagement": _customer_context_path(output_dir).as_posix(),
        **arguments,
    }


def _resolve_run_reference(
    output_dir: Path,
    run_intake: dict[str, Any],
    value: object,
) -> Path:
    """Resolve one persisted path against the current managed run root."""

    reference = Path(str(value))
    if reference.is_absolute():
        return reference
    if run_intake.get("path_reference") != "run_root_relative":
        raise AssertionError("test fixture has a nonportable relative path")
    run_root = output_dir.parent.parent
    resolved = (run_root / reference).resolve(strict=True)
    if not resolved.is_relative_to(run_root.resolve(strict=True)):
        raise AssertionError("test fixture path escapes the managed run")
    return resolved


def _support_reference_root(
    output_dir: Path,
    run_intake: dict[str, Any],
    value: object,
) -> Path:
    """Return the artifact root for a support file or directory reference."""

    resolved = _resolve_run_reference(output_dir, run_intake, value)
    return resolved if resolved.is_dir() else resolved.parent


def _forge_self_resealed_material_state(
    core: Any,
    output_dir: Path,
    mutation: str,
) -> tuple[dict[str, Any], dict[str, Any], dict[str, Any]]:
    """Reissue every local seal after one material/gate forgery."""

    envelope_path = output_dir / "assurance_envelope.json"
    audit_path = output_dir / "check_audit.json"
    payload_path = output_dir / "review_payload.json"
    final_path = output_dir / "final_artifacts.json"
    ui_path = output_dir / "ui_decisions.json"
    applied_path = output_dir / "applied_decisions.json"
    envelope = json.loads(envelope_path.read_text(encoding="utf-8"))
    audit = json.loads(audit_path.read_text(encoding="utf-8"))
    payload = json.loads(payload_path.read_text(encoding="utf-8"))
    run_intake = json.loads(
        (output_dir / "run_intake.json").read_text(encoding="utf-8")
    )
    final_artifacts = json.loads(final_path.read_text(encoding="utf-8"))
    ui_decisions = json.loads(ui_path.read_text(encoding="utf-8"))
    applied = json.loads(applied_path.read_text(encoding="utf-8"))
    replacements: list[dict[str, Any]] = []

    if mutation == "csv_amount":
        results_path = output_dir / "check_results.csv"
        with results_path.open("r", encoding="utf-8", newline="") as handle:
            reader = csv.DictReader(handle)
            fieldnames = list(reader.fieldnames or [])
            rows = list(reader)
        rows[0]["amount_signed"] = "999999.99"
        with results_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(
                handle,
                fieldnames=fieldnames,
                lineterminator="\n",
            )
            writer.writeheader()
            writer.writerows(rows)
        prior = next(
            receipt
            for receipt in envelope["artifact_receipts"]
            if receipt.get("path") == "check_results.csv"
        )
        replacements.append(
            core.artifact_receipt(
                output_dir,
                results_path,
                artifact_id=prior["artifact_id"],
                root_id="run",
                role=prior["role"],
                media_type=prior["media_type"],
            )
        )
    elif mutation == "xlsx_amount":
        workbook_path = output_dir / "check_results.xlsx"
        workbook = openpyxl.load_workbook(workbook_path)
        worksheet = workbook["Sheet1"]
        worksheet.cell(
            row=2,
            column=core.RESULT_COLUMNS.index("amount_signed") + 1,
            value="999999.99",
        )
        workbook.save(workbook_path)
        prior = next(
            receipt
            for receipt in envelope["artifact_receipts"]
            if receipt.get("path") == "check_results.xlsx"
        )
        replacements.append(
            core.artifact_receipt(
                output_dir,
                workbook_path,
                artifact_id=prior["artifact_id"],
                root_id="run",
                role=prior["role"],
                media_type=prior["media_type"],
            )
        )
    elif mutation == "numeric_ledger":
        ledger_path = output_dir / "numeric_evidence_ledger.json"
        ledger = json.loads(ledger_path.read_text(encoding="utf-8"))
        entry = ledger["entries"][0]
        entry["value"] = "999999.99"
        entry["source"]["value"] = "999999.99"
        entry["prepared"]["value"] = "999999.99"
        for target in entry["outputs"]:
            target["value"] = "999999.99"
        ledger.pop("content_sha256", None)
        ledger["content_sha256"] = core.canonical_json_sha256(ledger)
        ledger_path.write_text(
            json.dumps(ledger, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        audit["numeric_evidence_ledger"] = ledger
        envelope["numeric_evidence_ledgers"] = [ledger]
        prior = next(
            receipt
            for receipt in envelope["artifact_receipts"]
            if receipt.get("path") == "numeric_evidence_ledger.json"
        )
        replacements.append(
            core.artifact_receipt(
                output_dir,
                ledger_path,
                artifact_id=prior["artifact_id"],
                root_id="run",
                role=prior["role"],
                media_type=prior["media_type"],
            )
        )
    elif mutation != "gate_status":
        raise AssertionError(f"unknown material forgery: {mutation}")

    review_decision = next(
        decision
        for decision in envelope["reviewed_decisions"]
        if decision["decision_type"] == "check_entries_review_actions"
    )
    gates = json.loads(json.dumps(envelope["gate_register"]))
    gates["gates"]["semantic_review"] = {
        "status": "passed",
        "evidence_refs": [review_decision["decision_id"]],
        "limitations": ["Self-resealed promotion is not external authority."],
    }
    gates["gates"]["reporting"] = {
        "status": "passed",
        "evidence_refs": ["numeric.check_entries_amounts"],
        "limitations": ["Self-resealed promotion is not external authority."],
    }
    gates["report_ready"] = True
    gates = core.validate_gate_register(gates)

    payload["summary"]["assurance_gates"] = gates
    payload["summary"]["professional_conclusion_status"] = "reviewed"
    payload.pop("content_sha256", None)
    payload["content_sha256"] = core.canonical_json_sha256(payload)
    payload_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    payload_receipt = core.artifact_receipt(
        output_dir,
        payload_path,
        artifact_id="source.review_payload",
        root_id="run",
        role="source",
        media_type="application/json",
    )
    replacements.append(payload_receipt)

    receipts = list(envelope["artifact_receipts"])
    for replacement in replacements:
        receipts = _replace_receipt(receipts, replacement)
    roots = {
        "normalization": _resolve_run_reference(
            output_dir,
            run_intake,
            audit["journal"],
        ).parent,
        "support": _support_reference_root(
            output_dir,
            run_intake,
            audit["pdf_path"],
        ),
        "run": output_dir,
        **core.implementation_artifact_roots(),
    }
    forged_envelope = core.build_assurance_envelope(
        run_id=envelope["run_id"],
        workflow_id=envelope["workflow_id"],
        workflow_version=envelope["workflow_version"],
        artifact_receipts=receipts,
        implementation_artifact_refs=envelope["implementation_artifact_refs"],
        reviewed_decisions=envelope["reviewed_decisions"],
        source_qualifications=envelope["source_qualifications"],
        allocation_ledgers=envelope["allocation_ledgers"],
        numeric_evidence_ledgers=envelope["numeric_evidence_ledgers"],
        gate_register=gates,
        limitations=envelope["limitations"],
        artifact_roots=roots,
    )
    envelope_path.write_text(
        json.dumps(forged_envelope, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    envelope_receipt = core.artifact_receipt(
        output_dir,
        envelope_path,
        artifact_id="output.assurance_envelope",
        root_id="run",
        role="output",
        media_type="application/json",
    )
    envelope_binding = {
        "path": envelope_path.as_posix(),
        "content_sha256": forged_envelope["content_sha256"],
        "artifact_receipt": envelope_receipt,
    }

    audit["assurance_gates"] = gates
    audit["professional_conclusion_status"] = "reviewed"
    audit["review_payload_binding"] = {
        "content_sha256": payload["content_sha256"],
        "artifact_receipt": payload_receipt,
    }
    audit["assurance_envelope"] = envelope_binding
    for replacement in [*replacements, envelope_receipt]:
        audit["output_artifact_receipts"] = _replace_receipt(
            audit["output_artifact_receipts"],
            replacement,
        )
    audit.pop("content_sha256", None)
    audit["content_sha256"] = core.canonical_json_sha256(audit)
    audit_path.write_text(
        json.dumps(audit, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    final_artifacts["assurance_gates"] = gates
    final_artifacts["professional_conclusion_status"] = "reviewed"
    final_artifacts["status"] = "final_ready"
    final_artifacts["review_status"] = "final_ready"
    final_artifacts["review_payload_content_sha256"] = payload["content_sha256"]
    final_artifacts["assurance_envelope"] = envelope_binding
    for output in final_artifacts["outputs"]:
        for replacement in [*replacements, envelope_receipt]:
            if output.get("path") == replacement.get("path"):
                output["artifact_receipt"] = replacement
                output["size_bytes"] = replacement["byte_count"]
    final_path.write_text(
        json.dumps(final_artifacts, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    ui_decisions["review_payload_content_sha256"] = payload["content_sha256"]
    ui_path.write_text(
        json.dumps(ui_decisions, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    applied["review_payload"]["content_sha256"] = payload["content_sha256"]
    applied["assurance_envelope_content_sha256"] = forged_envelope["content_sha256"]
    applied_path.write_text(
        json.dumps(applied, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    run_intake = json.loads(
        (output_dir / "run_intake.json").read_text(encoding="utf-8")
    )
    return payload, final_artifacts, run_intake


def _copied_implementation_roots(
    core: Any,
    root: Path,
) -> tuple[dict[str, Path], dict[str, Any]]:
    source_roots = core.implementation_artifact_roots()
    copied_roots = {
        "implementation": root / "check-entries",
        "assurance_implementation": root / "vera_assurance",
    }
    receipts = core.build_implementation_receipts()
    for receipt in receipts:
        source = source_roots[receipt["root_id"]] / receipt["path"]
        destination = copied_roots[receipt["root_id"]] / receipt["path"]
        destination.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source, destination)
    envelope = {
        "artifact_receipts": receipts,
        "implementation_artifact_refs": [
            receipt["artifact_id"] for receipt in receipts
        ],
    }
    return copied_roots, envelope


def _copy_check_entries_runtime(root: Path) -> tuple[Path, Path]:
    plugin = root / "check-entries"
    shared = root / "_shared" / "vendor" / "modules" / "vera_assurance"
    shutil.copytree(ROOT / "plugins" / "check-entries", plugin)
    shutil.copytree(
        ROOT / "plugins" / "_shared" / "vendor" / "modules" / "vera_assurance",
        shared,
    )
    for cache in sorted(root.rglob("__pycache__"), reverse=True):
        shutil.rmtree(cache)
    return plugin, shared


def test_plugin_inspects_entries_and_runs_deterministic_checks(
    monkeypatch: Any, tmp_path: Path
) -> None:
    core = load_core()
    journal_path = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-01-02",
                "movement": "M-1001",
                "description": "Invoice payment",
                "debit": "123.45",
            }
        ],
    )
    pdf_dir = tmp_path / "pdfs"
    output_dir = tmp_path / "out"
    checks_dir = output_dir / "checks"
    pdf_dir.mkdir()
    support_pdf = pdf_dir / "support_M-1001.pdf"
    support_pdf.write_bytes(b"%PDF placeholder")

    def fake_extract_text(path: Path, payload: bytes) -> str:
        assert path == support_pdf
        assert payload == b"%PDF placeholder"
        return (
            "Movimento M-1001\nPartita IVA fornitore: 01234567890\n"
            "Pagamento fattura ACME Spa 02/01/2025 EUR 123,45"
        )

    monkeypatch.setattr(core, "_extract_pdf_text", fake_extract_text)

    inspection = core.inspect_entries(
        journal_path, pdf_dir, output_dir, language="it", document_language="it"
    )
    _reviewed_party_recipe(
        core,
        journal_path,
        output_dir / "suggested_recipe.json",
    )
    _reviewed_pdf_assertion_recipe(
        core,
        journal_path,
        pdf_dir,
        output_dir / "suggested_recipe.json",
        support_locator="support_M-1001.pdf",
        direction=True,
    )
    result = core.run_entry_checks(
        journal_path,
        pdf_dir,
        checks_dir,
        output_dir / "suggested_recipe.json",
        language="it",
        document_language="it",
    )

    inspection_payload = json.loads((output_dir / "inspection.json").read_text())
    recipe_payload = json.loads((output_dir / "suggested_recipe.json").read_text())
    audit_payload = json.loads((checks_dir / "check_audit.json").read_text())
    run_intake = json.loads((checks_dir / "run_intake.json").read_text())
    review_payload = json.loads((checks_dir / "review_payload.json").read_text())
    ui_decisions = json.loads((checks_dir / "ui_decisions.json").read_text())
    final_artifacts = json.loads((checks_dir / "final_artifacts.json").read_text())
    assurance_envelope = json.loads(
        (checks_dir / "assurance_envelope.json").read_text()
    )
    row = result.frame.to_dicts()[0]

    assert inspection.journal["row_count"] == 1
    assert inspection_payload["language"] == "it"
    assert recipe_payload["journal"]["mapping"]["movement_number"] == (
        "movement_number"
    )
    assert row["status"] == "ok"
    assert row["matched_pdf"] == "support_M-1001.pdf"
    assert row["checks_run"] == "amount,date,currency,direction,party_perimeter"
    assert row["professional_conclusion"] == "pending_review"
    assert row["source_qualification_id"].startswith("qualification.")
    assert audit_payload["status_counts"] == {"ok": 1}
    assert audit_payload["source_preparation"]["source_preparation_status"] == (
        "qualified"
    )
    normalization_replay = audit_payload["source_preparation"]["upstream_assurance"][
        "normalization_replay"
    ]
    assert normalization_replay["status"] == "passed"
    assert normalization_replay["schema_version"] == (
        "journal_sampling.normalization_replay.v1"
    )
    assert audit_payload["assurance_gates"]["gates"]["semantic_review"]["status"] == (
        "withheld"
    )
    assert audit_payload["assurance_gates"]["report_ready"] is False
    assert audit_payload["input_artifact_receipts"]
    assert audit_payload["lineage"][0]["prepared_entry_id"] == row["prepared_entry_id"]
    assert (checks_dir / "normalized_entries.csv").exists()
    support_manifest = json.loads(
        (checks_dir / "support_manifest.json").read_text(encoding="utf-8")
    )
    assert support_manifest["canonical_relative_paths"] == ["support_M-1001.pdf"]
    assert support_manifest["artifact_receipts"][0]["sha256"] == (
        hashlib.sha256(b"%PDF placeholder").hexdigest()
    )
    assert (checks_dir / "pdf_inventory.json").exists()
    assert (checks_dir / "check_results.csv").exists()
    assert (checks_dir / "review_notes.md").exists()
    assert run_intake["plugin"] == "check-entries"
    assert run_intake["workflow"] == "check-entries"
    assert run_intake["dependency_check"]["status"] == "not_run"
    assert journal_path.as_posix() in run_intake["data_posture"]["local_files_read"]
    assert pdf_dir.as_posix() in run_intake["data_posture"]["local_files_read"]
    assert run_intake["data_posture"]["external_connectors_used"] == []
    assert run_intake["data_posture"]["upload_paths_used"] == []
    assert review_payload["run_id"] == run_intake["run_id"]
    assert review_payload["review_type"] == "journal_entry_support_review"
    assert review_payload["item_count"] == len(review_payload["items"])
    review_payload_content = dict(review_payload)
    review_payload_digest = review_payload_content.pop("content_sha256")
    assert review_payload_digest == core.canonical_json_sha256(review_payload_content)
    item_types = {item["item_type"] for item in review_payload["items"]}
    assert {"supported_entry", "pdf_inventory", "review_artifact"} <= item_types
    supported_item = next(
        item
        for item in review_payload["items"]
        if item["item_type"] == "supported_entry"
    )
    assert supported_item["data"]["target_artifact"] == "check_results.csv"
    assert supported_item["data"]["target_id_field"] == "prepared_entry_id"
    assert supported_item["data"]["target_record_id"] == row["prepared_entry_id"]
    assert supported_item["data"]["target_field"] == "review_notes"
    assert review_payload["summary"]["ok_count"] == 1
    assert ui_decisions["status"] == "pending_review"
    assert ui_decisions["decision_source"] == "not_collected"
    assert ui_decisions["review_payload_content_sha256"] == review_payload_digest
    assert final_artifacts["run_id"] == run_intake["run_id"]
    assert final_artifacts["status"] == "written_pending_review"
    assert final_artifacts["assurance_envelope"]["content_sha256"] == (
        assurance_envelope["content_sha256"]
    )
    from vera_assurance import validate_assurance_envelope

    assert (
        validate_assurance_envelope(
            assurance_envelope,
            artifact_roots={
                "normalization": journal_path.parent,
                "support": pdf_dir,
                "run": checks_dir,
                **core.implementation_artifact_roots(),
            },
        )["content_sha256"]
        == assurance_envelope["content_sha256"]
    )
    handoff_output = next(
        output
        for output in final_artifacts["outputs"]
        if output["path"] == "review_handoff.md"
    )
    handoff_text = (checks_dir / "review_handoff.md").read_text(encoding="utf-8")
    assert handoff_output["required_text"] == [
        "Review Handoff",
        "review_payload.json",
        "ui_decisions.json",
        "applied_decisions.json",
        "final_artifacts.json",
    ]
    assert "render_check_entries_review" in handoff_text
    assert "apply_check_entries_decisions" in handoff_text
    review_notes_output = next(
        output
        for output in final_artifacts["outputs"]
        if output["path"] == "review_notes.md"
    )
    assert review_notes_output["required_text"] == [
        "# Check Entries Review Notes",
        "## Status Counts",
        "## Review Policy",
    ]
    check_results_output = next(
        output
        for output in final_artifacts["outputs"]
        if output["path"] == "check_results.csv"
    )
    check_results_xlsx_output = next(
        output
        for output in final_artifacts["outputs"]
        if output["path"] == "check_results.xlsx"
    )
    assert check_results_output["row_count"] == audit_payload["result_row_count"]
    assert check_results_output["required_columns"] == [
        "prepared_entry_id",
        "source_qualification_id",
        "movement_number",
        "source_file",
        "source_sheet",
        "source_page",
        "source_row",
        "status",
        "matched_pdf",
        "currency",
        "unit",
        "reported_increment",
        "professional_conclusion",
        "assurance_gate_status",
    ]
    assert (
        check_results_xlsx_output["source_row_count"]
        == audit_payload["result_row_count"]
    )
    assert check_results_xlsx_output["required_sheets"] == ["Sheet1"]
    assert check_results_xlsx_output["required_sheet_headers"] == {
        "Sheet1": [
            "prepared_entry_id",
            "movement_number",
            "source_row",
            "status",
            "matched_pdf",
            "checks_run",
        ]
    }
    assert check_results_xlsx_output["required_cells"] == {
        "Sheet1": {
            "A1": "prepared_entry_id",
            "A2": row["prepared_entry_id"],
            "C1": "movement_number",
            "C2": "M-1001",
            "R1": "source_row",
            "R2": str(row["source_row"]),
            "S1": "status",
            "S2": "ok",
            "T1": "matched_pdf",
            "T2": "support_M-1001.pdf",
            "U1": "checks_run",
            "U2": "amount,date,currency,direction,party_perimeter",
        }
    }
    assert "required_cells" in check_results_xlsx_output["qa_checks"]
    contract_report = validate_contract(
        checks_dir,
        strict_data_posture=True,
        strict_execution_trace=True,
        strict_output_paths=True,
        strict_output_content=True,
    )
    assert contract_report.ok, contract_report.as_dict()


@pytest.mark.parametrize("mutation", ["missing", "reordered", "expanded"])
def test_check_entries_rejects_resealed_upstream_implementation_set_mutation(
    tmp_path: Path,
    mutation: str,
) -> None:
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-01-02",
                "movement": "M-1001",
                "description": "Invoice payment",
                "debit": "123.45",
            }
        ],
    )
    normalization_dir = normalized.parent
    envelope_path = normalization_dir / "assurance_envelope.json"
    diagnostics_path = normalization_dir / "normalization_diagnostics.json"
    envelope = json.loads(envelope_path.read_text(encoding="utf-8"))
    diagnostics = json.loads(diagnostics_path.read_text(encoding="utf-8"))
    implementation_receipts = [
        receipt
        for receipt in envelope["artifact_receipts"]
        if receipt["role"] == "implementation"
    ]
    non_implementation_receipts = [
        receipt
        for receipt in envelope["artifact_receipts"]
        if receipt["role"] != "implementation"
    ]
    core = load_core()
    if mutation == "missing":
        tampered_receipts = implementation_receipts[:-1]
    elif mutation == "reordered":
        tampered_receipts = list(reversed(implementation_receipts))
    else:
        journal_sampling_root = core._journal_sampling_component_root()
        tampered_receipts = [
            *implementation_receipts,
            core.artifact_receipt(
                journal_sampling_root,
                journal_sampling_root / "README.md",
                artifact_id="implementation.journal_sampling_rogue",
                root_id="implementation",
                role="implementation",
                media_type="text/markdown",
            ),
        ]
    envelope["artifact_receipts"] = [
        *non_implementation_receipts,
        *tampered_receipts,
    ]
    envelope["implementation_artifact_refs"] = [
        receipt["artifact_id"] for receipt in tampered_receipts
    ]
    _seal_review_payload(envelope)
    envelope_path.write_text(
        json.dumps(envelope, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    diagnostics["implementation_receipts"] = tampered_receipts
    _seal_review_payload(diagnostics)
    diagnostics_path.write_text(
        json.dumps(diagnostics, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    with pytest.raises(
        ValueError,
        match="Journal Sampling implementation receipt set is not exact",
    ):
        core._load_journal_entries(normalized, {})


@pytest.mark.parametrize(
    "mutation",
    ["unowned_script", "implementation_fifo", "receipted_hardlink"],
)
def test_check_entries_rejects_unsafe_upstream_implementation_physical_tree(
    monkeypatch: Any,
    tmp_path: Path,
    mutation: str,
) -> None:
    if mutation == "implementation_fifo" and sys.platform == "win32":
        pytest.skip("FIFO implementation probe requires a POSIX host.")
    normalized = _qualified_journal(
        tmp_path / "source",
        [
            {
                "date": "2025-01-02",
                "movement": "M-1001",
                "description": "Invoice payment",
                "debit": "123.45",
            }
        ],
    )
    copied_plugin = tmp_path / "runtime" / "journal-sampling"
    copied_shared = (
        tmp_path / "runtime" / "_shared" / "vendor" / "modules" / "vera_assurance"
    )
    shutil.copytree(ROOT / "plugins" / "journal-sampling", copied_plugin)
    shutil.copytree(
        ROOT / "plugins" / "_shared" / "vendor" / "modules" / "vera_assurance",
        copied_shared,
    )
    for cache in sorted((tmp_path / "runtime").rglob("__pycache__"), reverse=True):
        shutil.rmtree(cache)
    if mutation == "unowned_script":
        (copied_plugin / "scripts" / "rogue.py").write_text(
            "VALUE = 'unowned'\n",
            encoding="utf-8",
        )
    elif mutation == "implementation_fifo":
        os.mkfifo(copied_plugin / "scripts" / "rogue.pyc")
    else:
        target = copied_plugin / "scripts" / "journal_sampling_core.py"
        external = tmp_path / "journal_sampling_core.py"
        shutil.copy2(target, external)
        target.unlink()
        os.link(external, target)
    core = load_core()
    monkeypatch.setattr(
        core,
        "_journal_sampling_component_root",
        lambda: copied_plugin,
    )
    monkeypatch.setattr(
        core,
        "implementation_artifact_roots",
        lambda: {
            "implementation": ROOT / "plugins" / "check-entries",
            "assurance_implementation": copied_shared,
        },
    )
    support_dir = tmp_path / "support"
    support_dir.mkdir()

    with pytest.raises(ValueError, match="Journal Sampling implementation"):
        core.run_entry_checks(
            normalized,
            support_dir,
            tmp_path / "out",
        )


def test_review_edit_reseals_assurance_envelope(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    core = load_core()
    apply_review_edits = load_apply_review_edits()
    journal_path = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-01-02",
                "movement": "M-1001",
                "description": "Invoice payment",
                "debit": "123.45",
            }
        ],
    )
    support_dir = tmp_path / "support"
    output_dir = tmp_path / "checks"
    support_dir.mkdir()
    support_pdf = support_dir / "support_M-1001.pdf"
    support_pdf.write_bytes(b"%PDF placeholder")
    monkeypatch.setattr(
        core,
        "_extract_pdf_text",
        lambda _path, _payload: (
            "Movimento M-1001\nPartita IVA fornitore: 01234567890\n"
            "Pagamento fattura ACME Spa 02/01/2025 EUR 123,45"
        ),
    )

    recipe_path = _reviewed_party_recipe(
        core,
        journal_path,
        tmp_path / "check_entries_recipe.json",
    )
    _reviewed_pdf_assertion_recipe(
        core,
        journal_path,
        support_dir,
        recipe_path,
        support_locator="support_M-1001.pdf",
        direction=True,
    )
    core.run_entry_checks(
        journal_path,
        support_dir,
        output_dir,
        recipe_path,
        language="it",
        document_language="it",
    )
    review_payload = json.loads((output_dir / "review_payload.json").read_text())
    final_artifacts_path = output_dir / "final_artifacts.json"
    applied_decisions_path = output_dir / "applied_decisions.json"
    prior_envelope = json.loads((output_dir / "assurance_envelope.json").read_text())
    entry_item = next(
        item
        for item in review_payload["items"]
        if item["item_type"] == "supported_entry"
    )

    check_results_path = output_dir / "check_results.csv"
    original_results_path = (
        output_dir
        / "revisions"
        / "originals"
        / f"check_results__{entry_item['id']}.csv"
    )
    original_results_path.parent.mkdir(parents=True)
    shutil.copy2(check_results_path, original_results_path)
    with check_results_path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        fieldnames = list(reader.fieldnames or [])
        rows = list(reader)
    rows[0]["review_notes"] = "Reviewer confirmed the support."
    with check_results_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    decisions: list[dict[str, Any]] = []
    effects: list[dict[str, Any]] = []
    for item in review_payload["items"]:
        if item["id"] == entry_item["id"]:
            decisions.append(
                {
                    "item_id": item["id"],
                    "action": "edit",
                    "edit_value": "Reviewer confirmed the support.",
                }
            )
            effects.append(
                {
                    "item_id": item["id"],
                    "item_type": item["item_type"],
                    "title": item["title"],
                    "action": "edit",
                    "artifact_update": "structured_artifact_updated",
                    "target_artifact": "check_results.csv",
                    "target_id_field": "prepared_entry_id",
                    "target_record_id": item["data"]["target_record_id"],
                    "target_field": "review_notes",
                    "edit_value": "Reviewer confirmed the support.",
                    "original_artifact_backup": (
                        original_results_path.relative_to(output_dir).as_posix()
                    ),
                    "structured_update": {
                        "id_field": "prepared_entry_id",
                        "record_id": item["data"]["target_record_id"],
                        "target_field": "review_notes",
                        "records_key": None,
                        "updated_rows": 1,
                    },
                    "derived_native_regeneration_paths": ["check_results.xlsx"],
                    "requires_native_regeneration": True,
                    "requires_followup": False,
                }
            )
        else:
            decisions.append({"item_id": item["id"], "action": "accept"})
            effects.append(
                {
                    "item_id": item["id"],
                    "item_type": item["item_type"],
                    "title": item["title"],
                    "action": "accept",
                    "artifact_update": "none",
                    "requires_native_regeneration": False,
                    "requires_followup": False,
                }
            )
    applied = {
        "schema_version": "2.0",
        "plugin": "check-entries",
        "workflow": "check-entries",
        "run_id": review_payload["run_id"],
        "applied_at": "2026-07-24T10:00:00+00:00",
        "review_payload": {
            "path": "review_payload.json",
            "content_sha256": review_payload["content_sha256"],
            "item_count": review_payload["item_count"],
        },
        "decisions": decisions,
        "effects": effects,
        "decision_count": len(decisions),
        "item_count": review_payload["item_count"],
        "blocker_count": 0,
        "native_regeneration_count": 1,
        "application_status": "partial_review_applied",
    }
    applied_decisions_path.write_text(
        json.dumps(applied, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    rows[0]["amount_abs"] = "999"
    with check_results_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    with pytest.raises(ValueError, match="outside authorized review notes"):
        apply_review_edits.apply_review_edits(
            output_dir,
            applied_decisions_path,
            final_artifacts_path,
        )

    rows[0]["amount_abs"] = "123.45"
    with check_results_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    result = apply_review_edits.apply_review_edits(
        output_dir,
        applied_decisions_path,
        final_artifacts_path,
    )

    resealed = json.loads((output_dir / "assurance_envelope.json").read_text())
    audit = json.loads((output_dir / "check_audit.json").read_text())
    final_artifacts = json.loads(final_artifacts_path.read_text())
    assert result["application_status"] == "blocked"
    assert resealed["content_sha256"] != prior_envelope["content_sha256"]
    assert resealed["gate_register"]["report_ready"] is False
    review_decision = next(
        decision
        for decision in resealed["reviewed_decisions"]
        if decision["decision_type"] == "check_entries_review_actions"
    )
    assert review_decision["status"] == "reviewed"
    assert review_decision["source_artifact_refs"] == ["source.review_payload"]
    receipt_by_path = {
        receipt["path"]: receipt for receipt in resealed["artifact_receipts"]
    }
    assert receipt_by_path["review_payload.json"]["role"] == "source"
    core.validate_artifact_receipt(
        output_dir,
        receipt_by_path["check_results.csv"],
    )
    core.validate_artifact_receipt(
        output_dir,
        receipt_by_path["check_results.xlsx"],
    )
    audit_content = dict(audit)
    audit_digest = audit_content.pop("content_sha256")
    assert audit_digest == core.canonical_json_sha256(audit_content)
    assert final_artifacts["status"] == "blocked"
    assert final_artifacts["assurance_envelope"]["content_sha256"] == (
        resealed["content_sha256"]
    )


def test_accept_only_review_is_replayed_and_resealed(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    core, output_dir, review_payload, run_intake = _supported_assurance_run(
        monkeypatch,
        tmp_path,
    )
    prior_envelope = json.loads((output_dir / "assurance_envelope.json").read_text())
    decisions = [
        {
            "item_id": item["id"],
            "action": ("accept" if "accept" in item["allowed_actions"] else "skip"),
        }
        for item in review_payload["items"]
    ]
    responses = _call_mcp_server(
        [
            {
                "jsonrpc": "2.0",
                "id": 1,
                "method": "tools/call",
                "params": {
                    "name": "apply_check_entries_decisions",
                    "arguments": {
                        "client_engagement": _managed_run_context_path(
                            output_dir
                        ).as_posix(),
                        "run_intake": run_intake,
                        "review_payload": review_payload,
                        "decisions": decisions,
                    },
                },
            }
        ]
    )

    result = responses[0]["result"]["structuredContent"]
    applied = json.loads((output_dir / "applied_decisions.json").read_text())
    resealed = json.loads((output_dir / "assurance_envelope.json").read_text())

    assert result["ok"] is True
    assert result["application_status"] == "blocked"
    assert applied["assurance_preflight"]["assurance_replayed"] is True
    assert applied["assurance_replayed"] is True
    assert resealed["content_sha256"] != prior_envelope["content_sha256"]
    assert resealed["reviewed_decisions"][-1]["status"] == "reviewed"
    assert resealed["gate_register"]["gates"]["semantic_review"]["status"] == (
        "withheld"
    )
    assert resealed["gate_register"]["report_ready"] is False
    core.validate_assurance_envelope(
        resealed,
        artifact_roots={
            "normalization": _resolve_run_reference(
                output_dir,
                run_intake,
                json.loads((output_dir / "check_audit.json").read_text())["journal"],
            ).parent,
            "support": _support_reference_root(
                output_dir,
                run_intake,
                run_intake["input_paths"][2],
            ),
            "run": output_dir,
            **core.implementation_artifact_roots(),
        },
    )


@pytest.mark.parametrize(
    "mutation",
    [
        "csv_amount",
        "xlsx_amount",
        "numeric_ledger",
        "gate_status",
    ],
)
def test_fresh_rederivation_rejects_fully_self_resealed_material_and_gate_forgery(
    monkeypatch: Any,
    tmp_path: Path,
    mutation: str,
) -> None:
    # Arrange
    core, output_dir, review_payload, run_intake = _supported_assurance_run(
        monkeypatch,
        tmp_path,
    )
    complete_decisions = [
        {
            "item_id": item["id"],
            "action": "accept" if "accept" in item["allowed_actions"] else "skip",
        }
        for item in review_payload["items"]
    ]
    honest = _call_mcp_server(
        [
            {
                "jsonrpc": "2.0",
                "id": 1,
                "method": "tools/call",
                "params": {
                    "name": "apply_check_entries_decisions",
                    "arguments": {
                        "client_engagement": _managed_run_context_path(
                            output_dir
                        ).as_posix(),
                        "run_intake": run_intake,
                        "review_payload": review_payload,
                        "decisions": complete_decisions,
                    },
                },
            }
        ]
    )[0]["result"]
    assert honest["isError"] is False
    assert honest["structuredContent"]["application_status"] == "blocked"

    forged_payload, forged_final, persisted_intake = (
        _forge_self_resealed_material_state(
            core,
            output_dir,
            mutation,
        )
    )
    before = _transaction_tree_state(output_dir)
    decision_item = forged_payload["items"][0]
    decision = {
        "item_id": decision_item["id"],
        "action": (
            "accept" if "accept" in decision_item["allowed_actions"] else "skip"
        ),
    }

    # Act / Assert: the Python boundary rederives facts rather than trusting seals.
    with pytest.raises(ValueError):
        load_apply_review_edits().preflight_assurance(output_dir)

    responses = _call_mcp_server(
        [
            {
                "jsonrpc": "2.0",
                "id": index,
                "method": "tools/call",
                "params": {
                    "name": tool_name,
                    "arguments": {
                        "client_engagement": _managed_run_context_path(
                            output_dir
                        ).as_posix(),
                        "run_intake": persisted_intake,
                        "review_payload": forged_payload,
                        **(
                            {"final_artifacts": forged_final}
                            if tool_name == "render_check_entries_review"
                            else {}
                        ),
                        **(
                            {"decisions": [decision]}
                            if tool_name
                            in {
                                "save_check_entries_decisions",
                                "apply_check_entries_decisions",
                            }
                            else {}
                        ),
                    },
                },
            }
            for index, tool_name in enumerate(
                (
                    "validate_check_entries_review",
                    "render_check_entries_review",
                    "save_check_entries_decisions",
                    "apply_check_entries_decisions",
                ),
                start=1,
            )
        ]
    )
    assert all(response["result"]["isError"] is True for response in responses)
    assert _transaction_tree_state(output_dir) == before


@pytest.mark.parametrize(
    "tool_name",
    [
        "validate_check_entries_review",
        "render_check_entries_review",
        "save_check_entries_decisions",
        "apply_check_entries_decisions",
    ],
)
def test_assured_review_tools_reject_foreign_physical_output_path(
    monkeypatch: Any,
    tmp_path: Path,
    tool_name: str,
) -> None:
    # Arrange
    _, output_dir, review_payload, run_intake = _supported_assurance_run(
        monkeypatch,
        tmp_path,
    )
    rogue = output_dir / "rogue-unowned.bin"
    rogue.write_bytes(b"not workflow owned")
    before = _transaction_tree_state(output_dir)
    item = review_payload["items"][0]
    action = "accept" if "accept" in item["allowed_actions"] else "skip"
    arguments: dict[str, Any] = _managed_check_mcp_arguments(
        output_dir,
        {
            "run_intake": run_intake,
            "review_payload": review_payload,
        },
    )
    if tool_name in {
        "save_check_entries_decisions",
        "apply_check_entries_decisions",
    }:
        arguments["decisions"] = [{"item_id": item["id"], "action": action}]

    # Act
    response = _call_mcp_server(
        [
            {
                "jsonrpc": "2.0",
                "id": 1,
                "method": "tools/call",
                "params": {
                    "name": tool_name,
                    "arguments": arguments,
                },
            }
        ]
    )[0]["result"]

    # Assert
    assert response["isError"] is True
    assert _transaction_tree_state(output_dir) == before
    assert rogue.read_bytes() == b"not workflow owned"
    assert not list(tmp_path.glob(".generated-review-transaction-*"))


def test_fresh_run_rejects_late_foreign_physical_output_path(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    # Arrange
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-01-02",
                "movement": "M-1001",
                "description": "Invoice payment",
                "debit": "123.45",
            }
        ],
    )
    support_dir = tmp_path / "support"
    support_dir.mkdir()
    (support_dir / "support_M-1001.pdf").write_bytes(b"%PDF placeholder")
    monkeypatch.setattr(
        core,
        "_extract_pdf_text",
        lambda _path, _payload: (
            "Invoice\nMovement M-1001\nSupplier VAT: 01234567890\n"
            "02/01/2025 EUR 123,45"
        ),
    )
    recipe_path = _reviewed_party_recipe(
        core,
        normalized,
        tmp_path / "check_entries_recipe.json",
    )
    _reviewed_pdf_assertion_recipe(
        core,
        normalized,
        support_dir,
        recipe_path,
        support_locator="support_M-1001.pdf",
        direction=True,
    )
    output_dir = tmp_path / "checks"
    output_dir.mkdir()
    (output_dir / "prior.bin").write_bytes(b"prior")
    prior = _transaction_tree_state(output_dir)
    original_write = core.write_review_session_artifacts

    def inject_foreign_path(*args: Any, **kwargs: Any) -> Any:
        result = original_write(*args, **kwargs)
        (Path(args[0]) / "rogue-unowned.bin").write_bytes(b"foreign")
        return result

    monkeypatch.setattr(
        core,
        "write_review_session_artifacts",
        inject_foreign_path,
    )

    # Act / Assert
    with pytest.raises(ValueError, match="physical output set"):
        core.run_entry_checks(
            normalized,
            support_dir,
            output_dir,
            recipe_path,
        )
    assert _transaction_tree_state(output_dir) == prior


@pytest.mark.parametrize(
    ("root_id", "relative_path"),
    [
        ("implementation", "scripts/check_entries_core.py"),
        ("implementation", "mcp/server.cjs"),
        ("implementation", "assets/check-entries-review-widget.html"),
        ("assurance_implementation", "serialization.py"),
        (
            "assurance_implementation",
            "review_output_transaction.cjs",
        ),
    ],
)
def test_transitive_implementation_contract_rejects_changed_contributor(
    tmp_path: Path,
    root_id: str,
    relative_path: str,
) -> None:
    # Arrange
    core = load_core()
    roots, envelope = _copied_implementation_roots(
        core,
        tmp_path / "implementation-copy",
    )
    target = roots[root_id] / relative_path
    target.write_bytes(target.read_bytes() + b"\n# changed contributor\n")

    # Act / Assert
    with pytest.raises(ValueError, match="artifact receipt"):
        core.validate_implementation_contract(
            envelope,
            artifact_roots=roots,
        )


def test_transitive_implementation_contract_rejects_hardlink_alias(
    tmp_path: Path,
) -> None:
    # Arrange
    core = load_core()
    roots, envelope = _copied_implementation_roots(
        core,
        tmp_path / "implementation-copy",
    )
    target = roots["implementation"] / "mcp" / "server.cjs"
    alias_source = tmp_path / "server-alias.cjs"
    shutil.copy2(target, alias_source)
    target.unlink()
    os.link(alias_source, target)

    # Act / Assert
    with pytest.raises(ValueError, match="ordinary single-link"):
        core.validate_implementation_contract(
            envelope,
            artifact_roots=roots,
        )


def test_real_python_product_entry_rejects_timestamp_valid_malicious_bytecode(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    # Arrange
    _, output_dir, _, _ = _supported_assurance_run(monkeypatch, tmp_path / "run")
    plugin, _shared = _copy_check_entries_runtime(tmp_path / "runtime")
    target = plugin / "scripts" / "implementation_contract.py"
    source_bytes = target.read_bytes()
    source_stat = target.stat()
    marker = tmp_path / "MALICIOUS_PYC_EXECUTED.txt"
    malicious_source = (
        "from pathlib import Path as _AttackPath\n"
        f"_AttackPath({marker.as_posix()!r}).write_text("
        "'executed before implementation validation\\n', encoding='utf-8')\n"
        f"exec(compile({source_bytes.decode('utf-8')!r}, "
        f"{target.as_posix()!r}, 'exec'), globals())\n"
    )
    code = compile(malicious_source, target.as_posix(), "exec")
    cache_payload = importlib._bootstrap_external._code_to_timestamp_pyc(
        code,
        int(source_stat.st_mtime),
        source_stat.st_size,
    )
    monkeypatch.setattr(sys, "pycache_prefix", None)
    cache_path = Path(importlib.util.cache_from_source(target.as_posix()))
    cache_path.parent.mkdir()
    cache_path.write_bytes(cache_payload)
    before = _transaction_tree_state(output_dir)

    # Act
    completed = subprocess.run(
        [
            sys.executable,
            "-I",
            "-B",
            (plugin / "scripts" / "apply_review_edits.py").as_posix(),
            "--output-dir",
            output_dir.as_posix(),
            "--preflight-only",
        ],
        cwd=plugin,
        capture_output=True,
        text=True,
        check=False,
        timeout=30,
    )

    # Assert
    assert completed.returncode != 0
    assert marker.exists() is False
    assert target.read_bytes() == source_bytes
    assert _transaction_tree_state(output_dir) == before


def test_real_python_product_entry_does_not_expose_vendor_parent_modules(
    tmp_path: Path,
) -> None:
    # Arrange
    plugin, shared = _copy_check_entries_runtime(tmp_path / "runtime")
    marker = tmp_path / "UNOWNED_VENDOR_MODULE_EXECUTED.txt"
    (shared.parent / "argparse.py").write_text(
        "from pathlib import Path\n"
        f"Path({marker.as_posix()!r}).write_text('executed\\n', encoding='utf-8')\n",
        encoding="utf-8",
    )

    # Act
    completed = subprocess.run(
        [
            sys.executable,
            "-I",
            "-B",
            (plugin / "scripts" / "run_checks.py").as_posix(),
            "--help",
        ],
        cwd=plugin,
        capture_output=True,
        text=True,
        check=False,
        timeout=30,
    )

    # Assert
    assert completed.returncode == 0, completed.stderr
    assert marker.exists() is False
    assert "usage:" in completed.stdout


@pytest.mark.parametrize("entry_kind", ["symlink", "hardlink", "fifo"])
def test_real_python_product_entry_rejects_unsafe_bootstrap_entry(
    tmp_path: Path,
    entry_kind: str,
) -> None:
    if entry_kind == "fifo" and sys.platform == "win32":
        pytest.skip("FIFO bootstrap probe requires a POSIX host.")
    plugin, _shared = _copy_check_entries_runtime(tmp_path / "runtime")
    bootstrap = plugin / "scripts" / "implementation_bootstrap.py"
    external = tmp_path / f"bootstrap-{entry_kind}.py"
    external.write_bytes(bootstrap.read_bytes())
    bootstrap.unlink()
    if entry_kind == "symlink":
        bootstrap.symlink_to(external)
    elif entry_kind == "hardlink":
        os.link(external, bootstrap)
    else:
        os.mkfifo(bootstrap)

    completed = subprocess.run(
        [
            sys.executable,
            "-I",
            "-B",
            (plugin / "scripts" / "apply_review_edits.py").as_posix(),
            "--help",
        ],
        cwd=plugin,
        capture_output=True,
        text=True,
        check=False,
        timeout=30,
    )

    assert completed.returncode != 0
    assert "ordinary single-link regular file" in completed.stderr


def test_mcp_launches_python_with_isolated_no_bytecode_flags(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    _, output_dir, review_payload, run_intake = _supported_assurance_run(
        monkeypatch,
        tmp_path,
    )
    wrapper = tmp_path / "isolated-python-wrapper"
    wrapper.write_text(
        "\n".join(
            [
                "#!/usr/bin/env python3",
                "import subprocess",
                "import sys",
                f"real = {sys.executable!r}",
                'if sys.argv[1:3] != ["-I", "-B"]:',
                "    raise SystemExit(91)",
                "completed = subprocess.run([real, *sys.argv[1:]])",
                "raise SystemExit(completed.returncode)",
                "",
            ]
        ),
        encoding="utf-8",
    )
    wrapper.chmod(0o700)
    monkeypatch.setenv("PYTHON", wrapper.as_posix())

    response = _call_mcp_server(
        [
            {
                "jsonrpc": "2.0",
                "id": 1,
                "method": "tools/call",
                "params": {
                    "name": "validate_check_entries_review",
                    "arguments": _managed_check_mcp_arguments(
                        output_dir,
                        {
                            "run_intake": run_intake,
                            "review_payload": review_payload,
                        },
                    ),
                },
            }
        ]
    )[0]["result"]

    assert response["isError"] is False


@pytest.mark.parametrize(
    "relative_path",
    [
        "mcp/server.cjs",
        "assets/check-entries-review-widget.html",
        "../_shared/vendor/modules/vera_assurance/serialization.py",
        ("../_shared/vendor/modules/vera_assurance/" "review_output_transaction.cjs"),
    ],
)
def test_mcp_replay_rejects_changed_transitive_implementation(
    monkeypatch: Any,
    tmp_path: Path,
    relative_path: str,
) -> None:
    # Arrange
    _, output_dir, review_payload, run_intake = _supported_assurance_run(
        monkeypatch,
        tmp_path,
    )
    copied_root = tmp_path / "copied-runtime"
    copied_plugin = copied_root / "check-entries"
    copied_shared = copied_root / "_shared" / "vendor" / "modules" / "vera_assurance"
    shutil.copytree(ROOT / "plugins" / "check-entries", copied_plugin)
    shutil.copytree(
        ROOT / "plugins" / "_shared" / "vendor" / "modules" / "vera_assurance",
        copied_shared,
    )
    target = copied_plugin / relative_path
    target.write_bytes(target.read_bytes() + b"\n/* changed contributor */\n")

    # Act
    response = _call_mcp_server(
        [
            {
                "jsonrpc": "2.0",
                "id": 1,
                "method": "tools/call",
                "params": {
                    "name": "validate_check_entries_review",
                    "arguments": _managed_check_mcp_arguments(
                        output_dir,
                        {
                            "run_intake": run_intake,
                            "review_payload": review_payload,
                        },
                    ),
                },
            }
        ],
        server_path=copied_plugin / "mcp" / "server.cjs",
    )[0]["result"]

    # Assert
    assert response["isError"] is True
    assert (
        response["structuredContent"]["error"]
        == "Check Entries persisted review authorization failed."
    )


def test_transitive_implementation_contract_rejects_self_expanded_set(
    tmp_path: Path,
) -> None:
    # Arrange
    core = load_core()
    roots, envelope = _copied_implementation_roots(
        core,
        tmp_path / "implementation-copy",
    )
    rogue = roots["implementation"] / "scripts" / "rogue.py"
    rogue.write_text("VALUE = 1\n", encoding="utf-8")
    rogue_receipt = core.artifact_receipt(
        roots["implementation"],
        rogue,
        artifact_id="implementation.check_entries.scripts.rogue.py",
        root_id="implementation",
        role="implementation",
        media_type="text/x-python",
    )
    envelope["artifact_receipts"].append(rogue_receipt)
    envelope["implementation_artifact_refs"].append(rogue_receipt["artifact_id"])

    # Act / Assert
    with pytest.raises(ValueError, match="reference set is not exact"):
        core.validate_implementation_contract(
            envelope,
            artifact_roots=roots,
        )


def test_review_preflight_rejects_unrelated_check_result_mutation(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    _, output_dir, review_payload, run_intake = _supported_assurance_run(
        monkeypatch,
        tmp_path,
    )
    entry_item = next(
        item
        for item in review_payload["items"]
        if item["item_type"] == "supported_entry"
    )
    results_path = output_dir / "check_results.csv"
    before_final = (output_dir / "final_artifacts.json").read_bytes()
    with results_path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        fieldnames = list(reader.fieldnames or [])
        rows = list(reader)
    rows[0]["amount_abs"] = "999"
    original_note = rows[0]["review_notes"]
    with results_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    responses = _call_mcp_server(
        [
            {
                "jsonrpc": "2.0",
                "id": 1,
                "method": "tools/call",
                "params": {
                    "name": "apply_check_entries_decisions",
                    "arguments": _managed_check_mcp_arguments(
                        output_dir,
                        {
                            "run_intake": run_intake,
                            "review_payload": review_payload,
                            "decisions": [
                                {
                                    "item_id": entry_item["id"],
                                    "action": "edit",
                                    "edit_value": "Reviewer-confirmed note",
                                }
                            ],
                        },
                    ),
                },
            }
        ]
    )

    result = responses[0]["result"]
    with results_path.open("r", encoding="utf-8", newline="") as handle:
        row = next(csv.DictReader(handle))
    assert result["isError"] is True
    assert (
        result["structuredContent"]["error"]
        == "Check Entries persisted review authorization failed."
    )
    assert row["amount_abs"] == "999"
    assert row["review_notes"] == original_note
    assert (output_dir / "final_artifacts.json").read_bytes() == before_final
    assert not (output_dir / "applied_decisions.json").exists()


def test_spanish_run_localizes_review_notes_and_strict_contract(tmp_path: Path) -> None:
    core = load_core()
    journal_path = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2026-01-15",
                "movement": "ES-1",
                "description": "Pago",
                "debit": "75.25",
            }
        ],
    )
    support_dir = tmp_path / "support"
    output_dir = tmp_path / "output"
    support_dir.mkdir()
    result = core.run_entry_checks(
        journal_path,
        support_dir,
        output_dir,
        language="es-ES",
        document_language="en",
    )

    review_notes = (output_dir / "review_notes.md").read_text(encoding="utf-8")
    review_handoff = (output_dir / "review_handoff.md").read_text(encoding="utf-8")
    review_payload = json.loads(
        (output_dir / "review_payload.json").read_text(encoding="utf-8")
    )
    final_artifacts = json.loads(
        (output_dir / "final_artifacts.json").read_text(encoding="utf-8")
    )
    review_notes_output = next(
        output
        for output in final_artifacts["outputs"]
        if output["path"] == "review_notes.md"
    )
    handoff_output = next(
        output
        for output in final_artifacts["outputs"]
        if output["path"] == "review_handoff.md"
    )
    missing_item = next(
        item
        for item in review_payload["items"]
        if item["item_type"] == "missing_support"
    )
    artifact_item = next(
        item for item in review_payload["items"] if item["id"] == "check-results-xlsx"
    )

    assert review_notes.startswith(
        "# Notas de revisión de la comprobación de asientos\n"
    )
    assert "- Idioma: es" in review_notes
    assert "## Recuento por estado" in review_notes
    assert "## Política de revisión" in review_notes
    assert review_notes_output["required_text"] == [
        "# Notas de revisión de la comprobación de asientos",
        "## Recuento por estado",
        "## Política de revisión",
    ]
    assert result.frame.to_dicts()[0]["status"] == "missing_support"
    assert result.frame.to_dicts()[0]["review_notes"] == (
        "Ningún PDF justificativo contiene el identificador explícito del movimiento."
    )
    assert review_payload["language"] == "es"
    assert review_payload["columns"] == [
        {"field": "item_type", "label": "Tipo"},
        {"field": "title", "label": "Asiento"},
        {"field": "recommended_action", "label": "Acción sugerida"},
        {"field": "source_path", "label": "Fuente"},
        {"field": "output_path", "label": "Salida"},
        {"field": "status", "label": "Estado"},
    ]
    assert missing_item["recommended_action"] == "request_more_documents"
    assert missing_item["data"]["requested_document"] == (
        "PDF justificativo del movimiento ES-1"
    )
    assert missing_item["data"]["reason"] == (
        "Ningún PDF justificativo contiene el identificador explícito del movimiento."
    )
    assert artifact_item["title"] == "Libro de resultados de la comprobación"
    assert review_handoff.startswith(
        "# Entrega para revisión: Comprobación de asientos\n"
    )
    assert "## Revisión en Codex" in review_handoff
    assert handoff_output["required_text"][0] == "Entrega para revisión"
    assert final_artifacts["caveats"][0].startswith(
        "Los scripts solo comparan evidencias deterministas"
    )
    assert final_artifacts["next_actions"][0].startswith("Revise las filas")
    contract_report = validate_contract(
        output_dir,
        strict_data_posture=True,
        strict_execution_trace=True,
        strict_output_paths=True,
        strict_output_content=True,
    )
    assert contract_report.ok, contract_report.as_dict()


def test_plugin_marks_missing_support_without_model_calls(tmp_path: Path) -> None:
    core = load_core()
    journal_path = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-03-10",
                "movement": "2001",
                "description": "Payment",
                "debit": "80",
            }
        ],
    )
    pdf_dir = tmp_path / "pdfs"
    output_dir = tmp_path / "out"
    pdf_dir.mkdir()
    result = core.run_entry_checks(journal_path, pdf_dir, output_dir, language="en")
    row = result.frame.to_dicts()[0]
    review_payload = json.loads((output_dir / "review_payload.json").read_text())
    missing_item = next(
        item
        for item in review_payload["items"]
        if item["item_type"] == "missing_support"
    )

    assert row["status"] == "missing_support"
    assert row["mismatches"] == "support_pdf"
    assert missing_item["recommended_action"] == "request_more_documents"
    assert missing_item["data"]["requested_document"] == (
        "Supporting PDF for movement 2001"
    )
    assert missing_item["data"]["reason"] == (
        "No supporting PDF contains the explicit movement identifier."
    )
    assert any(
        evidence.get("kind") == "missing_document_request"
        and evidence.get("requested_document") == "Supporting PDF for movement 2001"
        for evidence in missing_item["evidence"]
    )


def test_plugin_matches_sampled_entry_from_fatturapa_zip(tmp_path: Path) -> None:
    core = load_core()
    journal_path = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-01-02",
                "movement": "3001",
                "description": "Invoice INV-42",
                "debit": "123.45",
            }
        ],
    )
    invoice_zip = tmp_path / "all_invoices.zip"
    output_dir = tmp_path / "out"
    with zipfile.ZipFile(invoice_zip, "w") as archive:
        archive.writestr("IT01234567890_INV42.xml", _fatturapa_xml())

    inspection_dir = tmp_path / "inspect"
    inspection = core.inspect_entries(journal_path, invoice_zip, inspection_dir)
    recipe_path = _reviewed_party_recipe(
        core,
        journal_path,
        inspection_dir / "suggested_recipe.json",
    )
    _reviewed_pdf_assertion_recipe(
        core,
        journal_path,
        invoice_zip,
        recipe_path,
        support_locator="all_invoices.zip!/IT01234567890_INV42.xml",
        direction=True,
    )
    result = core.run_entry_checks(
        journal_path,
        invoice_zip,
        output_dir,
        recipe_path,
    )

    row = result.frame.to_dicts()[0]
    inventory = json.loads((output_dir / "invoice_inventory.json").read_text())
    assert len(inspection.invoices) == 1
    assert inspection.suggested_recipe["acquisition_ladder"] == [
        "fatturapa_zip",
        "authorized_connector_export",
        "targeted_pdf_fallback",
    ]
    assert row["status"] == "ok"
    assert row["support_type"] == "fatturapa_xml"
    assert row["matched_support"] == ("all_invoices.zip!/IT01234567890_INV42.xml")
    assert row["matched_pdf"] is None
    assert set(row["checks_run"].split(",")) == {
        "invoice_number",
        "amount",
        "date",
        "currency",
        "direction",
        "party_perimeter",
    }
    assert row["support_artifact_id"].startswith("support.")
    assert inventory["invoice_count"] == 1
    assert inventory["errors"] == []


def test_plugin_records_authorized_connector_and_requests_targeted_fallback(
    tmp_path: Path,
) -> None:
    core = load_core()
    journal_path = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-01-02",
                "movement": "4001",
                "description": "Supplier invoice",
                "debit": "123.45",
            }
        ],
    )
    connector_export = tmp_path / "connector_export"
    output_dir = tmp_path / "out"
    connector_export.mkdir()
    (connector_export / "invoice_a.xml").write_bytes(_fatturapa_xml(number="A"))
    (connector_export / "invoice_b.xml").write_bytes(_fatturapa_xml(number="B"))

    result = core.run_entry_checks(
        journal_path,
        connector_export,
        output_dir,
        connector_name="authorized-accounting-system",
    )

    row = result.frame.to_dicts()[0]
    run_intake = json.loads((output_dir / "run_intake.json").read_text())
    assert row["status"] == "manual_review"
    assert "invoice_relationship_requires_review" in row["mismatches"]
    assert row["support_match_status"] == "relationship_requires_review"
    assert "explicit invoice-number" in row["review_notes"]
    assert run_intake["data_posture"]["external_connectors_used"] == [
        "authorized-accounting-system"
    ]
    assert run_intake["data_posture"]["external_routes_used"] == [
        {
            "route": "authorized-accounting-system",
            "destination_or_origin": "authorized-accounting-system",
            "payload_category": (
                "accounting_system_export_materialized_as_local_support"
            ),
            "network_used": True,
            "access_basis": None,
        }
    ]
    assert run_intake["assumptions"]["invoice_count"] == 2


def test_amount_and_date_cannot_establish_invoice_identity(tmp_path: Path) -> None:
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-01-02",
                "movement": "4002",
                "description": "Payment to a different supplier",
                "debit": "123.45",
            }
        ],
    )
    invoice = tmp_path / "unrelated_invoice.xml"
    invoice.write_bytes(
        _fatturapa_xml(
            number="UNRELATED-99",
            supplier="UNRELATED SUPPLIER SPA",
        )
    )

    recipe_path = _reviewed_party_recipe(
        core,
        normalized,
        tmp_path / "check_entries_recipe.json",
    )
    result = core.run_entry_checks(
        normalized,
        invoice,
        tmp_path / "out",
        recipe_path,
    )
    row = result.frame.to_dicts()[0]

    assert row["status"] == "manual_review"
    assert row["matched_support"] is None
    assert row["support_match_status"] == "relationship_requires_review"
    assert "invoice_relationship_requires_review" in row["mismatches"]


def test_raw_journal_and_binary_float_tolerance_are_rejected(tmp_path: Path) -> None:
    core = load_core()
    raw_journal = tmp_path / "raw.xlsx"
    support_dir = tmp_path / "support"
    support_dir.mkdir()
    _save_workbook(
        raw_journal,
        [["Date", "Amount"], ["2025-01-01", 10.0]],
    )

    with pytest.raises(ValueError, match="accepts only Journal Sampling"):
        core.run_entry_checks(raw_journal, support_dir, tmp_path / "raw-out")

    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-01-01",
                "movement": "F-1",
                "description": "Float rejection",
                "debit": "10",
            }
        ],
    )
    with pytest.raises(ValueError, match="binary float"):
        core.run_entry_checks(
            normalized,
            support_dir,
            tmp_path / "float-out",
            amount_tolerance=0.0,
        )


def test_modified_population_or_diagnostics_blocks_checks(tmp_path: Path) -> None:
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-02-01",
                "movement": "T-1",
                "description": "Tamper test",
                "debit": "10",
            }
        ],
    )
    support_dir = tmp_path / "support"
    support_dir.mkdir()
    normalized.write_text(
        normalized.read_text(encoding="utf-8") + "\n",
        encoding="utf-8",
    )

    with pytest.raises(ValueError, match="artifact receipt"):
        core.run_entry_checks(normalized, support_dir, tmp_path / "csv-tamper-out")

    normalized = _qualified_journal(
        tmp_path / "fresh",
        [
            {
                "date": "2025-02-01",
                "movement": "T-2",
                "description": "Diagnostics tamper",
                "debit": "10",
            }
        ],
    )
    diagnostics_path = normalized.parent / "normalization_diagnostics.json"
    diagnostics = json.loads(diagnostics_path.read_text(encoding="utf-8"))
    diagnostics["population_status"] = "incomplete"
    diagnostics_path.write_text(
        json.dumps(diagnostics, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    with pytest.raises(ValueError, match="content hash is stale"):
        core.run_entry_checks(
            normalized,
            support_dir,
            tmp_path / "diagnostics-tamper-out",
        )


def test_deleted_upstream_assurance_blocks_checks(tmp_path: Path) -> None:
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-02-01",
                "movement": "T-3",
                "description": "Missing upstream envelope",
                "debit": "10",
            }
        ],
    )
    (normalized.parent / "assurance_envelope.json").unlink()
    support_dir = tmp_path / "support"
    support_dir.mkdir()

    with pytest.raises(ValueError, match="assurance_envelope"):
        core.run_entry_checks(normalized, support_dir, tmp_path / "out")


def test_changed_original_journal_blocks_upstream_replay(tmp_path: Path) -> None:
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-02-01",
                "movement": "T-4",
                "description": "Changed source workbook",
                "debit": "10",
            }
        ],
    )
    source_path = _journal_tamper_source(normalized)
    source_path.write_bytes(source_path.read_bytes() + b"changed-after-normalization")
    support_dir = tmp_path / "support"
    support_dir.mkdir()

    with pytest.raises(ValueError, match="artifact receipt"):
        core.run_entry_checks(normalized, support_dir, tmp_path / "out")


def test_managed_journal_input_tamper_fails_before_check_output(
    tmp_path: Path,
) -> None:
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-02-01",
                "movement": "T-4-MANAGED",
                "description": "Changed managed source workbook",
                "debit": "10",
            }
        ],
    )
    source_path = _qualified_journal_source(normalized)
    source_path.write_bytes(source_path.read_bytes() + b"managed-input-tamper")
    support_dir = tmp_path / "support"
    support_dir.mkdir()
    output_dir = tmp_path / "out"

    with pytest.raises(
        ValueError,
        match="Journal Sampling current customer-run context is unavailable",
    ):
        core.run_entry_checks(normalized, support_dir, output_dir)

    assert not output_dir.exists()


def test_fresh_upstream_replay_runs_before_polars_population_parse(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-02-01",
                "movement": "T-4A",
                "description": "Replay before parse",
                "debit": "10",
            }
        ],
    )

    def fail_replay(*_args: Any, **_kwargs: Any) -> dict[str, Any]:
        raise ValueError("fresh replay sentinel")

    def forbid_parse(*_args: Any, **_kwargs: Any) -> Any:
        raise AssertionError("Polars parsed before fresh replay")

    monkeypatch.setattr(core, "_run_journal_sampling_replay", fail_replay)
    monkeypatch.setattr(core.pl, "read_csv", forbid_parse)

    with pytest.raises(ValueError, match="fresh replay sentinel"):
        core._load_journal_entries(normalized, {})


def test_self_resealed_raw_journal_forgery_fails_fresh_reperformance(
    tmp_path: Path,
) -> None:
    core = load_core()
    sampling_core = _load_journal_sampling_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-02-01",
                "movement": "T-4B",
                "description": "Self-resealed source",
                "debit": "10",
            }
        ],
    )
    source_path = _journal_tamper_source(normalized)
    workbook = openpyxl.load_workbook(source_path)
    workbook.active["F2"] = "11"
    workbook.save(source_path)
    diagnostics_path = normalized.parent / "normalization_diagnostics.json"
    envelope_path = normalized.parent / "assurance_envelope.json"
    diagnostics = json.loads(diagnostics_path.read_text(encoding="utf-8"))
    envelope = json.loads(envelope_path.read_text(encoding="utf-8"))
    old_receipt = diagnostics["source_receipts"][0]
    byte_count, digest = sampling_core.file_snapshot(source_path)
    replacement = {
        **old_receipt,
        "byte_count": byte_count,
        "sha256": digest,
    }
    diagnostics["source_receipts"][0] = replacement
    diagnostics["files"][0]["source_receipt"] = replacement
    envelope["artifact_receipts"] = [
        replacement if receipt["artifact_id"] == old_receipt["artifact_id"] else receipt
        for receipt in envelope["artifact_receipts"]
    ]
    _seal_review_payload(envelope)
    _seal_review_payload(diagnostics)
    envelope_path.write_text(
        json.dumps(envelope, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    diagnostics_path.write_text(
        json.dumps(diagnostics, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    support_dir = tmp_path / "support"
    support_dir.mkdir()
    output_dir = tmp_path / "out"

    with pytest.raises(ValueError, match="normalization replay failed"):
        core.run_entry_checks(normalized, support_dir, output_dir)

    assert not output_dir.exists()


def test_self_resealed_reviewed_recipe_forgery_fails_fresh_reperformance(
    tmp_path: Path,
) -> None:
    core = load_core()
    sampling_core = _load_journal_sampling_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-02-01",
                "movement": "T-4C",
                "description": "Self-resealed recipe",
                "debit": "10",
            }
        ],
    )
    normalization_root = normalized.parent
    source_path = _qualified_journal_source(normalized)
    recipe_source = normalization_root / "suggested_recipe.json"
    recipe = json.loads(recipe_source.read_text(encoding="utf-8"))
    for source_name, entry in recipe["files"].items():
        entry["currency"] = "USD"
        qualification = entry["qualification"]
        old_decision = qualification["decision_receipt"]
        contract = sampling_core._mapping_contract(
            parser=entry["parser"],
            source_family=entry["source_family"],
            header_rows=entry.get("header_rows", []),
            mapping=entry.get("mapping", {}),
            layout=entry.get("layout", {}),
            excluded_monetary_columns=entry.get("excluded_monetary_columns", []),
            posting_identity=entry.get("posting_identity", "source_row"),
            carry_forward_fields=entry.get("carry_forward_fields", []),
            currency="USD",
            unit=entry.get("unit", "currency"),
            decimal_separator=entry.get("decimal_separator"),
            thousands_separator=entry.get("thousands_separator"),
            amount_sign_convention=entry.get("amount_sign_convention"),
        )
        qualification["mapping_sha256"] = sampling_core.canonical_json_sha256(contract)
        qualification["decision_receipt"] = (
            sampling_core.build_reviewed_decision_receipt(
                decision_id=old_decision["decision_id"],
                decision_type=old_decision["decision_type"],
                status=old_decision["status"],
                reviewer_ref=old_decision["reviewer_ref"],
                reviewed_on=old_decision["reviewed_on"],
                adapter_id=old_decision["adapter_id"],
                adapter_version=old_decision["adapter_version"],
                source_artifact_refs=[
                    sampling_core._source_artifact_ref(Path(source_name))
                ],
                content=contract,
            )
        )
    recipe_source.write_text(
        json.dumps(recipe, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    forged_root = tmp_path / "forged-normalization"
    sampling_core.normalize_path(source_path, forged_root, recipe_source)
    original_diagnostics_path = normalization_root / "normalization_diagnostics.json"
    original_diagnostics = json.loads(
        original_diagnostics_path.read_text(encoding="utf-8")
    )
    old_normalized_receipt = original_diagnostics["normalized_csv_receipt"]
    forged_diagnostics = json.loads(
        (forged_root / "normalization_diagnostics.json").read_text(encoding="utf-8")
    )
    forged_envelope = json.loads(
        (forged_root / "assurance_envelope.json").read_text(encoding="utf-8")
    )
    for field in (
        "client_engagement",
        "path_reference",
        "input",
        "source_root",
        "normalization_recipe_root",
        "normalization_recipe_source_path",
        "output_csv",
    ):
        forged_diagnostics[field] = original_diagnostics[field]
    forged_diagnostics["normalized_csv_receipt"] = old_normalized_receipt
    forged_envelope["artifact_receipts"] = [
        (
            old_normalized_receipt
            if receipt["artifact_id"] == "prepared.normalized_journal"
            else receipt
        )
        for receipt in forged_envelope["artifact_receipts"]
    ]
    _seal_review_payload(forged_envelope)
    _seal_review_payload(forged_diagnostics)
    for artifact_name in (
        "normalization_recipe.json",
        "reviewed_decisions.json",
        "assurance_gates.json",
        "qualification_review_payload.json",
    ):
        shutil.copy2(forged_root / artifact_name, normalization_root / artifact_name)
    (normalization_root / "assurance_envelope.json").write_text(
        json.dumps(forged_envelope, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    original_diagnostics_path.write_text(
        json.dumps(forged_diagnostics, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    support_dir = tmp_path / "support"
    support_dir.mkdir()
    output_dir = tmp_path / "out"

    with pytest.raises(ValueError, match="normalization replay failed"):
        core.run_entry_checks(normalized, support_dir, output_dir)

    assert not output_dir.exists()


def test_normalized_journal_change_during_support_read_blocks_run(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-02-01",
                "movement": "T-5",
                "description": "Mid-run mutation",
                "debit": "10",
            }
        ],
    )
    support_dir = tmp_path / "support"
    support_dir.mkdir()
    (support_dir / "support_T-5.pdf").write_bytes(b"%PDF placeholder")

    def mutate_normalized(_: Path, _payload: bytes) -> str:
        normalized.write_text(
            normalized.read_text(encoding="utf-8") + "\n",
            encoding="utf-8",
        )
        return "Movement T-5, 01/02/2025, EUR 10.00"

    monkeypatch.setattr(core, "_extract_pdf_text", mutate_normalized)

    with pytest.raises(ValueError, match="changed after its initial capture"):
        core.run_entry_checks(normalized, support_dir, tmp_path / "out")


def test_noncanonical_money_is_rejected_even_when_receipts_are_resealed(
    tmp_path: Path,
) -> None:
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-02-02",
                "movement": "M-1",
                "description": "Canonical money",
                "debit": "123.45",
            }
        ],
    )
    text = normalized.read_text(encoding="utf-8")
    normalized.write_text(text.replace("123.45", "123.450"), encoding="utf-8")
    diagnostics_path = normalized.parent / "normalization_diagnostics.json"
    diagnostics = json.loads(diagnostics_path.read_text(encoding="utf-8"))
    diagnostics.pop("content_sha256")
    diagnostics["normalized_csv_receipt"] = core.artifact_receipt(
        normalized.parent,
        normalized,
        artifact_id="prepared.normalized_journal",
        root_id="normalization",
        role="canonical journal rows; population status is recorded separately",
        media_type="text/csv",
    )
    diagnostics["content_sha256"] = core.canonical_json_sha256(diagnostics)
    diagnostics_path.write_text(
        json.dumps(diagnostics, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    support_dir = tmp_path / "support"
    support_dir.mkdir()

    with pytest.raises(ValueError, match="artifact receipt"):
        core.run_entry_checks(normalized, support_dir, tmp_path / "out")


def test_single_unrelated_pdf_never_auto_matches(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-03-01",
                "movement": "5001",
                "description": "Single PDF",
                "debit": "25",
            }
        ],
    )
    support_dir = tmp_path / "support"
    support_dir.mkdir()
    pdf = support_dir / "unrelated.pdf"
    pdf.write_bytes(b"%PDF placeholder")
    monkeypatch.setattr(
        core,
        "_extract_pdf_text",
        lambda _path, _payload: ("01/03/2025 EUR 25.00 without a movement reference"),
    )

    result = core.run_entry_checks(normalized, support_dir, tmp_path / "out")
    row = result.frame.to_dicts()[0]

    assert row["status"] == "missing_support"
    assert row["matched_pdf"] is None
    assert result.audit["assurance_gates"]["gates"]["reconciliation"]["status"] == (
        "failed"
    )


def test_page_number_is_not_a_movement_identity(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-03-01",
                "movement": "1",
                "description": "Short movement identifier",
                "debit": "25",
            }
        ],
    )
    support_dir = tmp_path / "support"
    support_dir.mkdir()
    (support_dir / "unrelated.pdf").write_bytes(b"%PDF placeholder")
    monkeypatch.setattr(
        core,
        "_extract_pdf_text",
        lambda _path, _payload: "Page 1\n01/03/2025\nEUR 25.00",
    )

    result = core.run_entry_checks(normalized, support_dir, tmp_path / "out")
    row = result.frame.to_dicts()[0]

    assert row["status"] == "missing_support"
    assert row["matched_pdf"] is None


def test_multiple_explicit_pdf_candidates_remain_review(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-03-02",
                "movement": "M-5002",
                "description": "Ambiguous PDFs",
                "debit": "25",
            }
        ],
    )
    support_dir = tmp_path / "support"
    support_dir.mkdir()
    for name in ("support_5002_a.pdf", "support_5002_b.pdf"):
        (support_dir / name).write_bytes(b"%PDF placeholder")
    monkeypatch.setattr(
        core,
        "_extract_pdf_text",
        lambda _path, _payload: "Movement M-5002, 02/03/2025, EUR 25.00",
    )

    result = core.run_entry_checks(normalized, support_dir, tmp_path / "out")
    row = result.frame.to_dicts()[0]

    assert row["status"] == "manual_review"
    assert row["support_match_status"] == "ambiguous"
    assert row["matched_pdf"] is None
    assert "ambiguous_pdf_support" in row["mismatches"]


def test_prepared_identity_is_stable_when_unrelated_later_row_is_added(
    tmp_path: Path,
) -> None:
    core = load_core()
    first_root = tmp_path / "first"
    second_root = tmp_path / "second"
    first_root.mkdir()
    second_root.mkdir()
    common = {
        "date": "2025-04-01",
        "movement": "S-1",
        "description": "Stable row",
        "debit": "19.99",
    }
    first_path = _qualified_journal(first_root, [common])
    second_path = _qualified_journal(
        second_root,
        [
            common,
            {
                "date": "2025-04-02",
                "movement": "S-2",
                "description": "Unrelated later row",
                "debit": "5",
            },
        ],
    )

    first_frame, _ = core._load_journal_entries(first_path, {})
    second_frame, _ = core._load_journal_entries(second_path, {})

    assert first_frame.to_dicts()[0]["prepared_entry_id"] == (
        second_frame.to_dicts()[0]["prepared_entry_id"]
    )


def test_high_precision_xml_amount_remains_exact(tmp_path: Path) -> None:
    core = load_core()
    amount = "12345678901234567890.123456"
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-05-01",
                "movement": "HP-1",
                "description": "Invoice HP-42",
                "debit": amount,
            }
        ],
    )
    invoice = tmp_path / "invoice.xml"
    invoice.write_bytes(
        _fatturapa_xml(
            number="HP-42",
            invoice_date="2025-05-01",
            amount=amount,
        )
    )

    recipe_path = _reviewed_party_recipe(
        core,
        normalized,
        tmp_path / "check_entries_recipe.json",
    )
    _reviewed_pdf_assertion_recipe(
        core,
        normalized,
        invoice,
        recipe_path,
        support_locator="invoice.xml",
        direction=True,
    )
    result = core.run_entry_checks(
        normalized,
        invoice,
        tmp_path / "out",
        recipe_path,
    )
    row = result.frame.to_dicts()[0]

    assert row["status"] == "ok"
    assert row["amount_abs"] == amount
    assert row["amount_found"] == amount
    inventory = json.loads(
        (tmp_path / "out" / "invoice_inventory.json").read_text(encoding="utf-8")
    )
    assert inventory["invoices"][0]["total_amount"] == amount


def test_xml_currency_mismatch_cannot_emit_ok(tmp_path: Path) -> None:
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-05-02",
                "movement": "FX-1",
                "description": "Invoice FX-42",
                "debit": "100",
            }
        ],
    )
    invoice = tmp_path / "invoice.xml"
    invoice.write_bytes(
        _fatturapa_xml(
            number="FX-42",
            invoice_date="2025-05-02",
            amount="100",
            currency="USD",
        )
    )

    recipe_path = _reviewed_party_recipe(
        core,
        normalized,
        tmp_path / "check_entries_recipe.json",
    )
    _reviewed_pdf_assertion_recipe(
        core,
        normalized,
        invoice,
        recipe_path,
        support_locator="invoice.xml",
        direction=True,
    )
    result = core.run_entry_checks(
        normalized,
        invoice,
        tmp_path / "out",
        recipe_path,
    )
    row = result.frame.to_dicts()[0]

    assert row["status"] == "mismatch"
    assert "amount" in row["mismatches"]
    assert json.loads(row["evidence_facts"])["invoice_currency"] == "USD"


def test_same_support_reused_for_two_entries_requires_review(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-06-01",
                "movement": "R-1",
                "description": "First use",
                "debit": "50",
            },
            {
                "date": "2025-06-01",
                "movement": "R-1",
                "description": "Second use",
                "debit": "50",
            },
        ],
    )
    support_dir = tmp_path / "support"
    support_dir.mkdir()
    (support_dir / "support_R-1.pdf").write_bytes(b"%PDF placeholder")
    monkeypatch.setattr(
        core,
        "_extract_pdf_text",
        lambda _path, _payload: "Movement R-1, 01/06/2025, EUR 50.00",
    )

    result = core.run_entry_checks(normalized, support_dir, tmp_path / "out")

    assert result.frame.get_column("status").to_list() == [
        "manual_review",
        "manual_review",
    ]
    assert all(
        "support_reuse_requires_review" in value
        for value in result.frame.get_column("mismatches").to_list()
    )
    assert result.audit["assurance_gates"]["gates"]["reconciliation"]["status"] == (
        "failed"
    )


def test_audit_hash_closes_and_professional_gate_stays_withheld(
    tmp_path: Path,
) -> None:
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-07-01",
                "movement": "A-1",
                "description": "Audit closure",
                "debit": "10",
            }
        ],
    )
    support_dir = tmp_path / "support"
    support_dir.mkdir()

    core.run_entry_checks(normalized, support_dir, tmp_path / "out")
    audit = json.loads(
        (tmp_path / "out" / "check_audit.json").read_text(encoding="utf-8")
    )
    recorded_hash = audit.pop("content_sha256")

    assert recorded_hash == core.canonical_json_sha256(audit)
    assert audit["professional_conclusion_status"] == "pending_review"
    assert audit["assurance_gates"]["gates"]["semantic_review"]["status"] == (
        "withheld"
    )
    assert audit["assurance_gates"]["report_ready"] is False


def test_check_entries_run_is_bound_to_journal_sampling_client_engagement(
    tmp_path: Path,
) -> None:
    # Arrange
    core = load_core()
    journal, support, journal_context, check_context = _v2_client_bound_check_inputs(
        tmp_path,
        [
            {
                "date": "2025-01-02",
                "movement": "M-1001",
                "description": "Invoice payment",
                "debit": "123.45",
            }
        ],
    )
    output_dir = Path(check_context["output_dir"])

    # Act
    inspection = core.inspect_entries(
        journal,
        support,
        output_dir / "inspection",
        client_engagement=check_context,
    )
    result = core.run_entry_checks(
        journal,
        support,
        output_dir / "checks",
        client_engagement=check_context,
    )

    # Assert
    expected_journal_context = _portable_context_projection(journal_context)
    expected_check_context = _portable_context_projection(check_context)
    assert inspection.journal["client_engagement"] == expected_journal_context
    assert result.audit["client_engagement"] == expected_check_context
    for filename in (
        "run_intake.json",
        "review_payload.json",
        "final_artifacts.json",
    ):
        payload = json.loads((output_dir / "checks" / filename).read_text())
        assert payload["client_engagement"] == expected_check_context
        assert not {
            "studio_client_folder",
            "input_bindings",
            "input_dir",
            "workspace_root",
            "output_dir",
            "run_root",
            "run_manifest_path",
            "input_manifest_path",
            "context_path",
        }.intersection(payload["client_engagement"])
        assert payload["run_id"] == check_context["run_id"]


def test_customer_folder_handoff_separates_support_batches_and_uses_sample_only(
    tmp_path: Path,
) -> None:
    # Arrange: one explicit client engagement and one sealed Journal Sampling run.
    check_core = load_core()
    archive_core = load_studio_archive_core()
    archive_root = tmp_path / "Studio"
    client_root = archive_root / "Zecca SPA"
    client_root.mkdir(parents=True)
    state_dir = tmp_path / "private-state"
    configured = archive_core.configure_archive(archive_root, state_dir=state_dir)
    scope_id = next(
        item["scope_id"]
        for item in configured["scopes"]
        if item["display_name"] == "Zecca SPA"
    )
    client_id = archive_core.set_studio_client_identity(
        scope_id,
        legal_names=["Zecca SPA"],
        state_dir=state_dir,
    )["client"]["client_id"]
    engagement_id = archive_core.create_studio_client_engagement(
        client_id,
        "2025 journal sample checks",
        state_dir=state_dir,
    )["engagement"]["engagement_id"]
    received = tmp_path / "received"
    received.mkdir()
    original_journal = received / "zecca-journal.xlsx"
    journal_rows = [
        {
            "date": "2025-01-02",
            "movement": "M-1001",
            "description": "Invoice one",
            "debit": "100.00",
        },
        {
            "date": "2025-01-03",
            "movement": "M-1002",
            "description": "Invoice two",
            "debit": "200.00",
        },
        {
            "date": "2025-01-04",
            "movement": "M-1003",
            "description": "Invoice three",
            "debit": "300.00",
        },
    ]
    _save_workbook(
        original_journal,
        [
            [
                "Data",
                "Nr. Reg",
                "Conto",
                "Descrizione conto",
                "Descrizione",
                "Dare",
                "Avere",
            ],
            *[
                [
                    row["date"],
                    row["movement"],
                    "4000",
                    "Trade payable",
                    row["description"],
                    row["debit"],
                    None,
                ]
                for row in journal_rows
            ],
        ],
    )
    journal_import = archive_core.import_studio_client_document(
        client_id,
        original_journal,
        "journal",
        engagement_id=engagement_id,
        state_dir=state_dir,
    )
    journal_run = archive_core.prepare_studio_client_workflow(
        engagement_id,
        "journal-sampling",
        input_ids=[journal_import["input_id"]],
        idempotency_key="journal-sample-2025",
        state_dir=state_dir,
    )
    journal_run_id = journal_run["run"]["run_id"]
    archive_core.start_studio_client_workflow(
        client_id,
        engagement_id,
        journal_run_id,
        state_dir=state_dir,
    )
    journal_context = journal_run["client_engagement"]
    journal_output = Path(journal_context["output_dir"])
    sample_dir = journal_output / "sample"
    journal_execution = next(
        Path(item["path"])
        for item in journal_context["input_bindings"]
        if item["binding_id"] == journal_import["input_id"]
    )
    normalized_path = _qualified_journal(
        journal_output,
        journal_rows,
        source_path=journal_execution,
        normalization_name="normalization",
        client_engagement=journal_context,
        write_source=False,
    )
    sampling_core = _load_journal_sampling_core()
    sampling_core.run_sample(
        normalized_path,
        sample_dir,
        method="systematic",
        size=1,
        client_engagement=journal_context,
    )
    population_rows = check_core.pl.read_csv(
        normalized_path, infer_schema=False
    ).to_dicts()
    sample_path = sample_dir / "journal_sample.csv"
    sampled_rows = check_core.pl.read_csv(sample_path, infer_schema=False).to_dicts()
    special_artifact_ids = {
        "normalization/normalized_journal.csv": "prepared.normalized_journal",
        "normalization/normalization_diagnostics.json": (
            "internal.normalization_diagnostics"
        ),
        "sample/journal_sample.csv": "prepared.journal_sample_csv",
    }
    declarations = []
    for index, artifact_path in enumerate(
        sorted(path for path in journal_output.rglob("*") if path.is_file()),
        start=1,
    ):
        relative_path = artifact_path.relative_to(journal_output).as_posix()
        declarations.append(
            {
                "artifact_id": special_artifact_ids.get(
                    relative_path, f"internal.journal_sampling.{index:03d}"
                ),
                "path": relative_path,
                "purpose": f"Preserve Journal Sampling artifact {relative_path}.",
                "audience": (
                    "review" if relative_path.startswith("sample/") else "internal"
                ),
                "media_type": (
                    "application/json"
                    if artifact_path.suffix == ".json"
                    else (
                        "text/csv"
                        if artifact_path.suffix == ".csv"
                        else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                ),
            }
        )
    archive_core.finalize_studio_client_workflow(
        client_id,
        engagement_id,
        journal_run_id,
        declarations,
        state_dir=state_dir,
    )
    archive_core.complete_studio_client_workflow(
        client_id,
        engagement_id,
        journal_run_id,
        state_dir=state_dir,
    )
    wrong_role_source = received / "not-support.pdf"
    wrong_role_source.write_bytes(_text_pdf_bytes(["Not a support receipt"]))
    wrong_role_input = archive_core.import_studio_client_document(
        client_id,
        wrong_role_source,
        "source",
        engagement_id=engagement_id,
        state_dir=state_dir,
    )
    with pytest.raises(archive_core.ArchiveError, match="support role"):
        archive_core.start_check_entries_from_sample(
            client_id,
            engagement_id,
            journal_run_id,
            support_input_ids=[wrong_role_input["input_id"]],
            idempotency_key="wrong-support-role",
            state_dir=state_dir,
        )

    def prepare_support_batch(name: str, payload: bytes) -> tuple[dict[str, Any], str]:
        support_source = received / f"support-{name}.pdf"
        support_source.write_bytes(payload)
        imported = archive_core.import_studio_client_document(
            client_id,
            support_source,
            "support",
            engagement_id=engagement_id,
            state_dir=state_dir,
        )
        prepared = archive_core.start_check_entries_from_sample(
            client_id,
            engagement_id,
            journal_run_id,
            support_input_ids=[imported["input_id"]],
            idempotency_key=f"check-support-{name}",
            state_dir=state_dir,
        )
        return prepared, imported["input_id"]

    # Act: batch A runs first; batch B then gets a separate immutable run.
    batch_a, support_a_id = prepare_support_batch(
        "a",
        _text_pdf_bytes(
            [
                "Evidence batch A",
                "Movement M-1001 EUR 100.00 02/01/2025",
                "Movement M-1002 EUR 200.00 03/01/2025",
                "Movement M-1003 EUR 300.00 04/01/2025",
            ]
        ),
    )
    replayed_batch_a = archive_core.start_check_entries_from_sample(
        client_id,
        engagement_id,
        journal_run_id,
        support_input_ids=[support_a_id],
        idempotency_key="check-support-a",
        state_dir=state_dir,
    )
    manifest_a_path = Path(batch_a["client_engagement"]["input_manifest_path"])
    support_a_binding = next(
        item
        for item in batch_a["client_engagement"]["input_bindings"]
        if item["binding_id"] == support_a_id
    )
    normalized_a_binding = next(
        item
        for item in batch_a["client_engagement"]["input_bindings"]
        if item.get("upstream_artifact_id") == "prepared.normalized_journal"
    )
    with pytest.raises(ValueError, match="run-local normalized-journal"):
        check_core._bound_upstream_journal_execution(
            Path(normalized_a_binding["source_path"]),
            batch_a["client_engagement"],
        )

    def validate_batch(
        prepared: dict[str, Any], support_id: str
    ) -> tuple[Any, Path, Path]:
        context = prepared["client_engagement"]
        bindings = context["input_bindings"]
        normalized_binding = next(
            item
            for item in bindings
            if item.get("upstream_artifact_id") == "prepared.normalized_journal"
        )
        diagnostics_binding = next(
            item
            for item in bindings
            if item.get("upstream_artifact_id") == "internal.normalization_diagnostics"
        )
        support_binding = next(
            item for item in bindings if item["binding_id"] == support_id
        )
        check_core._validated_client_check_stage(
            context,
            journal=Path(normalized_binding["path"]),
            journal_diagnostics={
                "client_engagement": journal_context,
                "normalization_diagnostics": diagnostics_binding["path"],
            },
            support=Path(support_binding["path"]),
            output_dir=Path(context["output_dir"]) / "checks",
            stage="checks",
            enforce_output_path=True,
        )
        return (
            check_core._bound_sample_entries(
                check_core.pl.DataFrame(population_rows),
                context,
            ),
            Path(normalized_binding["path"]),
            Path(support_binding["path"]),
        )

    filtered_a, batch_a_journal, batch_a_support = validate_batch(batch_a, support_a_id)
    result_a = check_core.run_entry_checks(
        batch_a_journal,
        batch_a_support,
        Path(batch_a["client_engagement"]["output_dir"]) / "checks",
        client_engagement=batch_a["client_engagement"],
    )
    manifest_a_before = manifest_a_path.read_bytes()
    support_a_before = Path(support_a_binding["path"]).read_bytes()
    outputs_a_before = _tree_snapshot(Path(batch_a["client_engagement"]["output_dir"]))
    batch_b, support_b_id = prepare_support_batch(
        "b",
        _text_pdf_bytes(
            [
                "Evidence batch B",
                "Movement M-1001 EUR 100.00 02/01/2025",
                "Movement M-1002 EUR 200.00 03/01/2025",
                "Movement M-1003 EUR 300.00 04/01/2025",
            ]
        ),
    )
    filtered_b, _, _ = validate_batch(batch_b, support_b_id)

    # Assert: both checks use only the sample, while batch B cannot alter batch A.
    locator_columns = ["source_file", "source_sheet", "source_page", "source_row"]
    assert filtered_a.select(locator_columns).to_dicts() == [
        {column: sampled_rows[0][column] for column in locator_columns}
    ]
    assert filtered_b.select(locator_columns).to_dicts() == [
        {column: sampled_rows[0][column] for column in locator_columns}
    ]
    assert result_a.audit["result_row_count"] == 1
    assert result_a.frame.get_column("movement_number").to_list() == [
        sampled_rows[0]["movement_number"]
    ]
    assert batch_a["run"]["run_id"] != batch_b["run"]["run_id"]
    assert batch_a["status"] == "running"
    assert replayed_batch_a["run"]["run_id"] == batch_a["run"]["run_id"]
    assert (
        sum(
            item["kind"] == "upstream_artifact"
            for item in batch_a["input_manifest"]["inputs"]
        )
        == 9
    )
    assert manifest_a_path.read_bytes() == manifest_a_before
    assert Path(support_a_binding["path"]).read_bytes() == support_a_before
    assert _tree_snapshot(Path(batch_a["client_engagement"]["output_dir"])) == (
        outputs_a_before
    )
    batch_a_binding_ids = {
        item["binding_id"] for item in batch_a["input_manifest"]["inputs"]
    }
    batch_b_binding_ids = {
        item["binding_id"] for item in batch_b["input_manifest"]["inputs"]
    }
    assert support_a_id in batch_a_binding_ids
    assert support_b_id not in batch_a_binding_ids
    assert support_b_id in batch_b_binding_ids
    assert support_a_id not in batch_b_binding_ids


def test_check_entries_v2_rejects_missing_normalization_diagnostics_binding(
    tmp_path: Path,
) -> None:
    check_core = load_core()
    archive_core = load_studio_archive_core()
    archive_root = tmp_path / "Studio"
    client_root = archive_root / "Client"
    client_root.mkdir(parents=True)
    state_dir = tmp_path / "private-state"
    configured = archive_core.configure_archive(archive_root, state_dir=state_dir)
    scope_id = next(
        item["scope_id"]
        for item in configured["scopes"]
        if item["display_name"] == "Client"
    )
    client_id = archive_core.set_studio_client_identity(
        scope_id,
        legal_names=["Client"],
        state_dir=state_dir,
    )["client"]["client_id"]
    engagement_id = archive_core.create_studio_client_engagement(
        client_id,
        "Incomplete handoff",
        state_dir=state_dir,
    )["engagement"]["engagement_id"]
    received = tmp_path / "received"
    received.mkdir()
    journal_source = received / "journal.csv"
    journal_source.write_text("journal\n", encoding="utf-8")
    journal_input = archive_core.import_studio_client_document(
        client_id,
        journal_source,
        "journal",
        engagement_id=engagement_id,
        state_dir=state_dir,
    )
    journal_run = archive_core.prepare_studio_client_workflow(
        engagement_id,
        "journal-sampling",
        input_ids=[journal_input["input_id"]],
        state_dir=state_dir,
    )
    journal_run_id = journal_run["run"]["run_id"]
    archive_core.start_studio_client_workflow(
        client_id,
        engagement_id,
        journal_run_id,
        state_dir=state_dir,
    )
    output = Path(journal_run["client_engagement"]["output_dir"])
    (output / "normalization").mkdir()
    (output / "sample").mkdir()
    (output / "normalization" / "normalized_journal.csv").write_text(
        "source_file,source_sheet,source_page,source_row\njournal.csv,,,2\n",
        encoding="utf-8",
    )
    (output / "normalization" / "normalization_diagnostics.json").write_text(
        "{}\n", encoding="utf-8"
    )
    (output / "sample" / "journal_sample.csv").write_text(
        "source_file,source_sheet,source_page,source_row\njournal.csv,,,2\n",
        encoding="utf-8",
    )
    archive_core.finalize_studio_client_workflow(
        client_id,
        engagement_id,
        journal_run_id,
        [
            {
                "artifact_id": "prepared.normalized_journal",
                "path": "normalization/normalized_journal.csv",
                "purpose": "Provide the qualified journal population.",
                "audience": "internal",
                "media_type": "text/csv",
            },
            {
                "artifact_id": "internal.normalization_diagnostics",
                "path": "normalization/normalization_diagnostics.json",
                "purpose": "Preserve normalization diagnostics.",
                "audience": "internal",
                "media_type": "application/json",
            },
            {
                "artifact_id": "prepared.journal_sample_csv",
                "path": "sample/journal_sample.csv",
                "purpose": "Identify the exact selected journal rows.",
                "audience": "review",
                "media_type": "text/csv",
            },
        ],
        state_dir=state_dir,
    )
    support_source = received / "support.pdf"
    support_source.write_bytes(b"support\n")
    support_input = archive_core.import_studio_client_document(
        client_id,
        support_source,
        "support",
        engagement_id=engagement_id,
        state_dir=state_dir,
    )
    check_run = archive_core.prepare_studio_client_workflow(
        engagement_id,
        "check-entries",
        input_ids=[support_input["input_id"]],
        upstream_artifacts=[
            {
                "run_id": journal_run_id,
                "artifact_id": "prepared.normalized_journal",
                "role": "normalized_journal",
            },
            {
                "run_id": journal_run_id,
                "artifact_id": "prepared.journal_sample_csv",
                "role": "journal_sample",
            },
        ],
        state_dir=state_dir,
    )
    context = check_run["client_engagement"]
    journal_binding = next(
        item
        for item in context["input_bindings"]
        if item.get("upstream_artifact_id") == "prepared.normalized_journal"
    )
    support_binding = next(
        item
        for item in context["input_bindings"]
        if item["binding_id"] == support_input["input_id"]
    )

    with pytest.raises(ValueError, match="exact normalized journal"):
        check_core._validated_client_check_stage(
            context,
            journal=Path(journal_binding["path"]),
            journal_diagnostics={
                "client_engagement": journal_run["client_engagement"],
                "normalization_diagnostics": str(
                    output / "normalization" / "normalization_diagnostics.json"
                ),
            },
            support=Path(support_binding["path"]),
            output_dir=Path(context["output_dir"]) / "checks",
            stage="checks",
            enforce_output_path=True,
        )

    three_artifact_run = archive_core.prepare_studio_client_workflow(
        engagement_id,
        "check-entries",
        input_ids=[support_input["input_id"]],
        upstream_artifacts=[
            {
                "run_id": journal_run_id,
                "artifact_id": artifact_id,
                "role": "journal_handoff",
            }
            for artifact_id in (
                "prepared.normalized_journal",
                "internal.normalization_diagnostics",
                "prepared.journal_sample_csv",
            )
        ],
        idempotency_key="three-artifact-only-handoff",
        state_dir=state_dir,
    )
    three_context = three_artifact_run["client_engagement"]
    three_journal = next(
        item
        for item in three_context["input_bindings"]
        if item.get("upstream_artifact_id") == "prepared.normalized_journal"
    )
    three_diagnostics = next(
        item
        for item in three_context["input_bindings"]
        if item.get("upstream_artifact_id") == "internal.normalization_diagnostics"
    )
    three_support = next(
        item
        for item in three_context["input_bindings"]
        if item["binding_id"] == support_input["input_id"]
    )
    with pytest.raises(ValueError, match="complete exact Journal Sampling"):
        check_core._validated_client_check_stage(
            three_context,
            journal=Path(three_journal["path"]),
            journal_diagnostics={
                "client_engagement": journal_run["client_engagement"],
                "normalization_diagnostics": three_diagnostics["path"],
            },
            support=Path(three_support["path"]),
            output_dir=Path(three_context["output_dir"]) / "checks",
            stage="checks",
            enforce_output_path=True,
        )


def test_check_entries_rejects_cross_client_context(tmp_path: Path) -> None:
    # Arrange
    core = load_core()
    _, _, journal_context, _ = _v2_client_bound_check_inputs(
        tmp_path / "selected",
        [
            {
                "date": "2025-01-02",
                "movement": "M-1001",
                "description": "Invoice payment",
                "debit": "123.45",
            }
        ],
    )
    other_journal, other_support, _, other_context = _v2_client_bound_check_inputs(
        tmp_path / "other",
        [
            {
                "date": "2025-01-02",
                "movement": "OTHER-1",
                "description": "Other client invoice",
                "debit": "50.00",
            }
        ],
    )

    # Act / Assert
    with pytest.raises(ValueError, match="different client engagements"):
        core._validated_client_check_stage(
            other_context,
            journal=other_journal,
            journal_diagnostics={
                "client_engagement": journal_context,
                "normalization_diagnostics": "unreached.json",
            },
            support=other_support,
            output_dir=Path(other_context["output_dir"]) / "inspection",
            stage="inspection",
            enforce_output_path=True,
        )


def test_check_entries_rejects_support_outside_selected_engagement(
    tmp_path: Path,
) -> None:
    # Arrange
    core = load_core()
    journal, _, _, check_context = _v2_client_bound_check_inputs(
        tmp_path,
        [
            {
                "date": "2025-01-02",
                "movement": "M-1001",
                "description": "Invoice payment",
                "debit": "123.45",
            }
        ],
    )
    foreign_support = tmp_path / "received elsewhere"
    foreign_support.mkdir()

    # Act / Assert
    with pytest.raises(ValueError, match="support selection must close"):
        core.inspect_entries(
            journal,
            foreign_support,
            Path(check_context["output_dir"]) / "inspection",
            client_engagement=check_context,
        )


def test_support_swap_attack_uses_captured_bytes_not_live_reread(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-08-01",
                "movement": "M-999",
                "description": "Captured support",
                "debit": "999",
            }
        ],
    )
    support_dir = tmp_path / "support"
    support_dir.mkdir()
    support = support_dir / "support.pdf"
    support.write_bytes(b"999.00")

    def swap_live_file(path: Path, payload: bytes) -> str:
        assert payload == b"999.00"
        path.write_bytes(b"100.00")
        assert path.read_bytes() == b"100.00"
        path.write_bytes(b"999.00")
        return (
            "Invoice\nMovement M-999\nSupplier VAT: 01234567890\n"
            f"01/08/2025 EUR {payload.decode('ascii')}"
        )

    monkeypatch.setattr(core, "_extract_pdf_text", swap_live_file)
    recipe_path = _reviewed_party_recipe(
        core,
        normalized,
        tmp_path / "recipe.json",
    )
    _reviewed_pdf_assertion_recipe(
        core,
        normalized,
        support_dir,
        recipe_path,
        support_locator="support.pdf",
        direction=True,
    )
    result = core.run_entry_checks(
        normalized,
        support_dir,
        tmp_path / "out",
        recipe_path,
    )

    row = result.frame.to_dicts()[0]
    assert row["status"] == "ok"
    assert row["amount_found"] == "999"
    assert row["amount_found"] != "100"
    receipt = next(
        receipt
        for receipt in result.audit["input_artifact_receipts"]
        if receipt["path"] == "support.pdf"
    )
    assert receipt["sha256"] == hashlib.sha256(b"999.00").hexdigest()


def test_support_change_that_is_not_restored_blocks_before_assurance_write(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-08-02",
                "movement": "M-100",
                "description": "Changed support",
                "debit": "100",
            }
        ],
    )
    support_dir = tmp_path / "support"
    support_dir.mkdir()
    support = support_dir / "support.pdf"
    support.write_bytes(b"100.00")

    def mutate_live_file(path: Path, payload: bytes) -> str:
        assert payload == b"100.00"
        path.write_bytes(b"999.00")
        return "Movement M-100\nSupplier VAT: 01234567890\n" "02/08/2025 EUR 100.00"

    monkeypatch.setattr(core, "_extract_pdf_text", mutate_live_file)
    recipe_path = _reviewed_party_recipe(
        core,
        normalized,
        tmp_path / "recipe.json",
    )
    with pytest.raises(ValueError, match="artifact receipt"):
        core.run_entry_checks(
            normalized,
            support_dir,
            tmp_path / "out",
            recipe_path,
        )
    assert not (tmp_path / "out" / "assurance_envelope.json").exists()


@pytest.mark.parametrize("movement", ["2025", "A", "12345"])
def test_generic_labeled_tokens_do_not_establish_pdf_identity(
    monkeypatch: Any,
    tmp_path: Path,
    movement: str,
) -> None:
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-08-03",
                "movement": movement,
                "description": "Generic token",
                "debit": "10",
            }
        ],
    )
    support_dir = tmp_path / "support"
    support_dir.mkdir()
    (support_dir / f"movement_{movement}.pdf").write_bytes(b"%PDF")
    monkeypatch.setattr(
        core,
        "_extract_pdf_text",
        lambda _path, _payload: (
            f"Movement {movement}\nSupplier VAT: 01234567890\n" "03/08/2025 EUR 10.00"
        ),
    )

    result = core.run_entry_checks(normalized, support_dir, tmp_path / "out")

    assert result.frame.to_dicts()[0]["status"] == "missing_support"
    assert result.audit["assurance_gates"]["gates"]["source"]["status"] == "failed"


def test_filename_occurrence_alone_does_not_establish_pdf_identity(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-08-04",
                "movement": "M-77",
                "description": "Filename is not evidence",
                "debit": "10",
            }
        ],
    )
    support_dir = tmp_path / "support"
    support_dir.mkdir()
    (support_dir / "movement_M-77.pdf").write_bytes(b"%PDF")
    monkeypatch.setattr(
        core,
        "_extract_pdf_text",
        lambda _path, _payload: "04/08/2025 EUR 10.00",
    )

    result = core.run_entry_checks(normalized, support_dir, tmp_path / "out")

    assert result.frame.to_dicts()[0]["status"] == "missing_support"


def test_labeled_pdf_identity_and_exact_reviewed_tax_id_can_emit_ok(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-08-05",
                "movement": "M-78",
                "description": "Explicit support",
                "debit": "10",
            }
        ],
    )
    support_dir = tmp_path / "support"
    support_dir.mkdir()
    (support_dir / "unrelated-name.pdf").write_bytes(b"%PDF")
    monkeypatch.setattr(
        core,
        "_extract_pdf_text",
        lambda _path, _payload: (
            "Invoice\nMovement M-78\nSupplier VAT: 01234567890\n" "05/08/2025 EUR 10.00"
        ),
    )
    recipe_path = _reviewed_party_recipe(
        core,
        normalized,
        tmp_path / "recipe.json",
    )
    _reviewed_pdf_assertion_recipe(
        core,
        normalized,
        support_dir,
        recipe_path,
        support_locator="unrelated-name.pdf",
        direction=True,
    )

    result = core.run_entry_checks(
        normalized,
        support_dir,
        tmp_path / "out",
        recipe_path,
    )

    assert result.frame.to_dicts()[0]["status"] == "ok"
    assert (
        result.audit["assurance_gates"]["gates"]["reconciliation"]["status"] == "passed"
    )


def test_numeric_invoice_number_requires_reviewed_relationship(
    tmp_path: Path,
) -> None:
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-08-06",
                "movement": "M-79",
                "description": "Invoice 42",
                "debit": "10",
            }
        ],
    )
    invoice = tmp_path / "invoice.xml"
    invoice.write_bytes(
        _fatturapa_xml(number="42", invoice_date="2025-08-06", amount="10")
    )
    party_recipe = _reviewed_party_recipe(
        core,
        normalized,
        tmp_path / "party_recipe.json",
    )

    without_relationship = core.run_entry_checks(
        normalized,
        invoice,
        tmp_path / "without-relationship",
        party_recipe,
    )
    relationship_recipe = _reviewed_relationship_recipe(
        core,
        normalized,
        invoice,
        tmp_path / "relationship_recipe.json",
        support_locator="invoice.xml",
    )
    _reviewed_pdf_assertion_recipe(
        core,
        normalized,
        invoice,
        relationship_recipe,
        support_locator="invoice.xml",
        direction=True,
    )
    with_relationship = core.run_entry_checks(
        normalized,
        invoice,
        tmp_path / "with-relationship",
        relationship_recipe,
    )

    first = without_relationship.frame.to_dicts()[0]
    second = with_relationship.frame.to_dicts()[0]
    assert first["status"] == "manual_review"
    assert first["support_match_status"] == "relationship_requires_review"
    assert second["status"] == "ok"
    assert "reviewed_support_relationship" in second["support_match_signals"]
    assert json.loads(second["evidence_facts"])[
        "reviewed_relationship_recording_exception"
    ]


def test_missing_party_perimeter_never_emits_ok_for_structured_invoice(
    tmp_path: Path,
) -> None:
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-08-07",
                "movement": "M-80",
                "description": "Invoice INV-80",
                "debit": "10",
            }
        ],
    )
    invoice = tmp_path / "invoice.xml"
    invoice.write_bytes(
        _fatturapa_xml(number="INV-80", invoice_date="2025-08-07", amount="10")
    )

    result = core.run_entry_checks(normalized, invoice, tmp_path / "out")

    row = result.frame.to_dicts()[0]
    assert row["status"] == "manual_review"
    assert "party_perimeter_requires_review" in row["mismatches"]
    assert result.audit["assurance_gates"]["gates"]["reconciliation"]["status"] == (
        "failed"
    )


@pytest.mark.parametrize(
    ("expected_role", "wrong_tax_id"),
    [
        ("supplier", "11111111111"),
        ("customer", "22222222222"),
    ],
)
def test_wrong_supplier_or_customer_perimeter_is_mismatch(
    tmp_path: Path,
    expected_role: str,
    wrong_tax_id: str,
) -> None:
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-08-08",
                "movement": "M-81",
                "description": "Invoice INV-81",
                "debit": "10",
            }
        ],
    )
    invoice = tmp_path / "invoice.xml"
    invoice.write_bytes(
        _fatturapa_xml(number="INV-81", invoice_date="2025-08-08", amount="10")
    )
    recipe_path = _reviewed_party_recipe(
        core,
        normalized,
        tmp_path / "recipe.json",
        tax_id=wrong_tax_id,
        expected_role=expected_role,
    )
    _reviewed_pdf_assertion_recipe(
        core,
        normalized,
        invoice,
        recipe_path,
        support_locator="invoice.xml",
        direction=True,
    )

    result = core.run_entry_checks(
        normalized,
        invoice,
        tmp_path / "out",
        recipe_path,
    )

    row = result.frame.to_dicts()[0]
    assert row["status"] == "mismatch"
    assert "party_perimeter" in row["mismatches"]


def test_free_text_party_name_cannot_promote_pdf_match(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-08-09",
                "movement": "M-82",
                "description": "Name containment",
                "debit": "10",
            }
        ],
    )
    support_dir = tmp_path / "support"
    support_dir.mkdir()
    (support_dir / "support.pdf").write_bytes(b"%PDF")
    monkeypatch.setattr(
        core,
        "_extract_pdf_text",
        lambda _path, _payload: (
            "Invoice\nMovement M-82\nSupplier: ACME SPA\n" "09/08/2025 EUR 10.00"
        ),
    )
    recipe_path = _reviewed_name_party_recipe(
        core,
        normalized,
        tmp_path / "recipe.json",
    )
    _reviewed_pdf_assertion_recipe(
        core,
        normalized,
        support_dir,
        recipe_path,
        support_locator="support.pdf",
        direction=True,
    )

    result = core.run_entry_checks(
        normalized,
        support_dir,
        tmp_path / "out",
        recipe_path,
    )

    row = result.frame.to_dicts()[0]
    assert row["status"] == "mismatch"
    assert "party_perimeter" in row["mismatches"]
    assert row["beneficiary_found"] is None


@pytest.mark.parametrize(
    ("suffix", "payload", "expected_error"),
    [
        (".xml", b"<not-an-invoice/>", "FatturaPA"),
        (".p7m", _fatturapa_xml(), "P7M"),
    ],
)
def test_unqualified_support_artifact_forces_failed_source_gate(
    tmp_path: Path,
    suffix: str,
    payload: bytes,
    expected_error: str,
) -> None:
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-08-10",
                "movement": "M-83",
                "description": "Unsupported source",
                "debit": "10",
            }
        ],
    )
    support = tmp_path / f"support{suffix}"
    support.write_bytes(payload)

    result = core.run_entry_checks(normalized, support, tmp_path / "out")

    qualification = result.audit["support_source_qualifications"][0]
    assert qualification["status"] == "unsupported_source_layout"
    assert result.audit["assurance_gates"]["gates"]["source"]["status"] == "failed"
    inventory = json.loads(
        (tmp_path / "out" / "invoice_inventory.json").read_text(encoding="utf-8")
    )
    assert expected_error in inventory["errors"][0]["error"]


def test_pdf_readability_qualification_is_separate_from_identity(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-08-11",
                "movement": "M-84",
                "description": "Readable but unrelated",
                "debit": "10",
            }
        ],
    )
    support_dir = tmp_path / "support"
    support_dir.mkdir()
    (support_dir / "other.pdf").write_bytes(b"%PDF")
    monkeypatch.setattr(
        core,
        "_extract_pdf_text",
        lambda _path, _payload: "Readable unrelated document",
    )

    result = core.run_entry_checks(normalized, support_dir, tmp_path / "out")

    qualification = result.audit["support_source_qualifications"][0]
    assert qualification["status"] == "qualified"
    assert result.frame.to_dicts()[0]["status"] == "missing_support"
    assert result.audit["assurance_gates"]["gates"]["source"]["status"] == "failed"


def test_xlsx_and_assurance_are_equal_across_repeated_runs(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-08-12",
                "movement": "M-85",
                "description": "Replay equality",
                "debit": "10",
            }
        ],
    )
    support_dir = tmp_path / "support"
    support_dir.mkdir()
    (support_dir / "support.pdf").write_bytes(b"%PDF")
    monkeypatch.setattr(
        core,
        "_extract_pdf_text",
        lambda _path, _payload: (
            "Movement M-85\nSupplier VAT: 01234567890\n" "12/08/2025 EUR 10.00"
        ),
    )
    review_module = sys.modules["review_session"]
    monkeypatch.setattr(
        review_module,
        "_utc_now",
        lambda: "2026-07-24T12:00:00+00:00",
    )
    recipe_path = _reviewed_party_recipe(
        core,
        normalized,
        tmp_path / "recipe.json",
    )

    first = tmp_path / "first"
    second = tmp_path / "second"
    first_result = core.run_entry_checks(
        normalized,
        support_dir,
        first,
        recipe_path,
    )
    core.run_entry_checks(
        normalized,
        support_dir,
        second,
        recipe_path,
    )

    assert (first / "check_results.xlsx").read_bytes() == (
        second / "check_results.xlsx"
    ).read_bytes()
    assert (first / "assurance_envelope.json").read_bytes() == (
        second / "assurance_envelope.json"
    ).read_bytes()
    ledger = first_result.audit["numeric_evidence_ledger"]
    evidence_ids = {entry["evidence_id"] for entry in ledger["entries"]}
    assert any(".amount_signed." in evidence_id for evidence_id in evidence_ids)
    assert any(".amount_abs." in evidence_id for evidence_id in evidence_ids)
    assert any(".amount_found." in evidence_id for evidence_id in evidence_ids)
    assert all(
        len(entry["outputs"]) == 2
        and entry["outputs"][0]["locator"].startswith("row=2;column=")
        and entry["outputs"][1]["locator"].startswith("Sheet1!")
        for entry in ledger["entries"]
    )
    assert first_result.audit["reproducibility_checks"] == {
        "xlsx_two_run_byte_equality": "passed",
        "assurance_two_build_equality": "passed",
    }


def test_skill_and_scripts_keep_codex_as_the_review_layer() -> None:
    skill_text = (
        ROOT / "plugins" / "check-entries" / "skills" / "check-entries" / "SKILL.md"
    ).read_text(encoding="utf-8")
    script_text = "\n".join(
        path.read_text(encoding="utf-8") for path in SCRIPT_DIR.glob("*.py")
    )

    assert "The user should not interact directly with CLI scripts" in skill_text
    assert "must not make direct OpenAI API calls" in skill_text
    assert "scripts/check_dependencies.py" in skill_text
    assert "`it`, `en`, `fr`, `de`, and `es`" in skill_text
    assert "missing deterministic extraction script" in skill_text
    assert "Keep the improvement note local to chat or run artifacts." in skill_text
    assert "validate_check_entries_review" in skill_text
    assert "render_check_entries_review" in skill_text
    assert "save_check_entries_decisions" in skill_text
    assert "apply_check_entries_decisions" in skill_text
    assert "modules.llm" not in script_text
    assert "model_router" not in script_text


def test_static_page_exposes_five_language_switch() -> None:
    page = (ROOT / "static" / "shared" / "check-entries" / "index.html").read_text(
        encoding="utf-8"
    )

    for snippet in (
        'data-lang="it"',
        'data-lang="en"',
        'data-lang="fr"',
        'data-lang="de"',
        'data-lang="es"',
        "Dalla scrittura campionata al supporto che la spiega.",
        "From a sampled entry to the document that explains it.",
        "De l'écriture échantillonnée au document qui l'explique.",
        "Von der ausgewählten Buchung zum erklärenden Beleg.",
        "Del asiento muestreado al documento que lo explica.",
        "authorized_connector_export",
        "invoice_inventory.json",
        "up to the first 20 rows of the normalized population",
        "for every parsed FatturaPA record",
        "1,500 results and 500 PDFs",
        "2,500 items or 2,000,000 bytes",
        "at most 25 per call",
        "500,000 bytes",
        "only an opaque handle, type, status, proposed action",
        "One index and the same limits in every mode",
        "without rereading all complete review data",
    ):
        assert snippet in page


def test_check_entries_mcp_server_validates_renders_and_saves_review_payload(
    tmp_path: Path,
) -> None:
    output_dir, client_run_id = _running_customer_output(tmp_path)
    check_results_path = output_dir / "check_results.csv"
    check_results_xlsx_path = output_dir / "check_results.xlsx"
    check_results_path.write_text(
        "\n".join(
            [
                "source_row,review_notes,status,matched_pdf",
                "1,Original deterministic note,ok,support_1001.pdf",
                "2,Missing support,missing_support,",
            ]
        )
        + "\n",
        encoding="utf-8",
    )
    _save_workbook(
        check_results_xlsx_path,
        [
            ["source_row", "review_notes", "status", "matched_pdf"],
            [1, "Original deterministic note", "ok", "support_1001.pdf"],
            [2, "Missing support", "missing_support", None],
        ],
    )
    review_payload = {
        "schema_version": "2.0",
        "plugin": "check-entries",
        "workflow": "check-entries",
        "run_id": client_run_id,
        "source_paths": ["entries.xlsx", "support_1001.pdf"],
        "review_type": "journal_entry_support_review",
        "items": [
            {
                "id": "entry-1",
                "item_type": "supported_entry",
                "title": "1001 | 123.45 | 2025-01-02",
                "source_path": "entries.xlsx",
                "output_path": "check_results.csv",
                "allowed_actions": ["accept", "edit", "mark_unclear", "skip"],
                "recommended_action": "accept",
                "evidence": [{"kind": "deterministic_checks", "checks_run": "amount"}],
                "data": {
                    "status": "ok",
                    "matched_pdf": "support_1001.pdf",
                    "target_artifact": "check_results.csv",
                    "target_id_field": "source_row",
                    "target_record_id": "1",
                    "target_field": "review_notes",
                },
                "status": "needs_review",
            },
            {
                "id": "entry-2",
                "item_type": "missing_support",
                "title": "1002 | 88.0",
                "source_path": "entries.xlsx",
                "output_path": "check_results.csv",
                "allowed_actions": [
                    "accept",
                    "edit",
                    "mark_unclear",
                    "request_more_documents",
                    "skip",
                ],
                "recommended_action": "request_more_documents",
                "evidence": [
                    {
                        "kind": "deterministic_checks",
                        "mismatches": "support_pdf",
                        "requested_document": "support_1002.pdf",
                    }
                ],
                "data": {
                    "status": "missing_support",
                    "requested_document": "support_1002.pdf",
                },
                "status": "needs_review",
            },
        ],
        "item_count": 2,
        "columns": [],
        "source_artifacts": {
            "run_intake": "run_intake.json",
            "check_results_csv": "check_results.csv",
            "check_results_xlsx": "check_results.xlsx",
        },
        "evidence": {},
        "allowed_actions": [
            "accept",
            "edit",
            "mark_unclear",
            "request_more_documents",
            "skip",
        ],
        "status": "ready_for_review",
        "summary": {
            "result_row_count": 2,
            "ok_count": 1,
            "missing_support_count": 1,
            "pdf_count": 1,
        },
    }
    _seal_review_payload(review_payload)
    run_intake = {
        "schema_version": "2.0",
        "plugin": "check-entries",
        "workflow": "check-entries",
        "run_id": client_run_id,
        "created_at": "2026-01-01T00:00:00Z",
        "language": "en",
        "document_language": "en",
        "input_paths": ["entries.xlsx", "support_1001.pdf"],
        "output_dir": output_dir.as_posix(),
        "inferred_task": "journal_entry_support_check",
        "assumptions": {},
        "unresolved_questions": [],
        "dependency_check": {"status": "not_run"},
        "data_posture": {
            "local_files_read": ["entries.xlsx", "support_1001.pdf"],
            "external_connectors_used": [],
            "upload_paths_used": [],
            "remote_sql_execution_used": False,
            "hosted_notebook_execution_used": False,
        },
        "execution_trace": [
            {
                "step_id": "check_entries_run",
                "kind": "deterministic_review_session",
                "status": "passed",
                "execution_location": "local_codex_workspace",
                "command": [
                    "python",
                    "plugins/check-entries/scripts/run_check_entries.py",
                ],
                "inputs": ["entries.xlsx", "support_1001.pdf"],
                "outputs": [
                    "review_payload.json",
                    "check_results.xlsx",
                    "final_artifacts.json",
                ],
            }
        ],
    }
    ui_decisions = {
        "schema_version": "2.0",
        "plugin": "check-entries",
        "workflow": "check-entries",
        "run_id": client_run_id,
        "review_payload_path": "review_payload.json",
        "review_payload_content_sha256": review_payload["content_sha256"],
        "decisions": [],
        "decision_count": 0,
        "status": "pending_review",
    }
    (output_dir / "run_intake.json").write_text(
        json.dumps(run_intake, indent=2) + "\n",
        encoding="utf-8",
    )
    (output_dir / "review_payload.json").write_text(
        json.dumps(review_payload, indent=2) + "\n",
        encoding="utf-8",
    )
    messages: list[dict[str, object]] = [
        {"jsonrpc": "2.0", "id": 1, "method": "tools/list"},
        {
            "jsonrpc": "2.0",
            "id": 2,
            "method": "tools/call",
            "params": {
                "name": "validate_check_entries_review",
                "arguments": {
                    "run_intake_path": str(output_dir / "run_intake.json"),
                    "review_payload_path": str(output_dir / "review_payload.json"),
                },
            },
        },
        {
            "jsonrpc": "2.0",
            "id": 3,
            "method": "tools/call",
            "params": {
                "name": "render_check_entries_review",
                "arguments": {
                    "run_intake_path": str(output_dir / "run_intake.json"),
                    "review_payload_path": str(output_dir / "review_payload.json"),
                    "ui_decisions": ui_decisions,
                },
            },
        },
        {"jsonrpc": "2.0", "id": 4, "method": "resources/list"},
        {
            "jsonrpc": "2.0",
            "id": 5,
            "method": "resources/read",
            "params": {"uri": "ui://widget/check-entries-review.html"},
        },
        {
            "jsonrpc": "2.0",
            "id": 6,
            "method": "tools/call",
            "params": {
                "name": "save_check_entries_decisions",
                "arguments": {
                    "client_engagement": _customer_context_path(output_dir).as_posix(),
                    "run_intake": run_intake,
                    "review_payload": review_payload,
                    "ui_decisions": ui_decisions,
                    "decisions": [
                        {
                            "item_id": "entry-1",
                            "action": "edit",
                            "edit_value": "Reviewer confirmed support evidence.",
                        },
                        {
                            "item_id": "entry-2",
                            "action": "request_more_documents",
                            "reviewer_note": "Support file is still missing.",
                        },
                    ],
                },
            },
        },
        {
            "jsonrpc": "2.0",
            "id": 7,
            "method": "tools/call",
            "params": {
                "name": "apply_check_entries_decisions",
                "arguments": {
                    "client_engagement": _customer_context_path(output_dir).as_posix(),
                    "run_intake": run_intake,
                    "review_payload": review_payload,
                    "final_artifacts": {
                        "schema_version": "2.0",
                        "plugin": "check-entries",
                        "workflow": "check-entries",
                        "run_id": client_run_id,
                        "outputs": [
                            {
                                "path": "review_payload.json",
                                "kind": "json",
                                "status": "written",
                            },
                            {
                                "path": "check_results.xlsx",
                                "kind": "xlsx",
                                "status": "written",
                            },
                        ],
                        "status": "written_pending_review",
                    },
                    "decisions": [
                        {
                            "item_id": "entry-1",
                            "action": "edit",
                            "edit_value": "Reviewer confirmed support evidence.",
                        },
                        {
                            "item_id": "entry-2",
                            "action": "request_more_documents",
                            "reviewer_note": "Support file is still missing.",
                        },
                    ],
                },
            },
        },
    ]

    responses = {response["id"]: response for response in _call_mcp_server(messages)}

    tool_names = {tool["name"] for tool in responses[1]["result"]["tools"]}
    assert {
        "validate_check_entries_review",
        "render_check_entries_review",
        "get_check_entries_case_context",
        "save_check_entries_decisions",
        "apply_check_entries_decisions",
    } <= tool_names
    validate_result = responses[2]["result"]["structuredContent"]
    assert validate_result["ok"] is True
    assert validate_result["item_count"] == 2
    assert "review_payload" not in validate_result
    assert "model_context_index" in validate_result
    render_result = responses[3]["result"]
    assert render_result["structuredContent"]["widget_type"] == "check_entries_review"
    assert "review_payload" not in render_result["structuredContent"]
    private_payload = render_result["_meta"]["private_review_payload"]
    assert private_payload["review_payload"] == review_payload
    assert private_payload["decision_policy"]["can_persist"] is True
    assert (
        render_result["_meta"]["openai/outputTemplate"]
        == "ui://widget/check-entries-review.html"
    )
    resource_uris = {
        resource["uri"] for resource in responses[4]["result"]["resources"]
    }
    assert "ui://widget/check-entries-review.html" in resource_uris
    widget_html = responses[5]["result"]["contents"][0]["text"]
    assert "Check Entries Review" in widget_html
    assert "Save decisions" in widget_html
    assert "Apply decisions" in widget_html
    assert "Applica decisioni" in widget_html
    assert "Preview sample review" in widget_html
    assert "Final outputs" in widget_html
    assert "toolResponseMetadata" in widget_html
    assert "private_review_payload" in widget_html
    save_result = responses[6]["result"]["structuredContent"]
    assert save_result["ok"] is True
    assert save_result["persisted"] is True
    assert save_result["decision_count"] == 2
    assert save_result["status"] == "reviewed"
    written_decisions = json.loads((output_dir / "ui_decisions.json").read_text())
    assert written_decisions["decision_source"] == "mcp_widget"
    assert written_decisions["status"] == "reviewed"
    assert written_decisions["decision_count"] == 2
    assert written_decisions["decisions"][0]["edit_value"] == (
        "Reviewer confirmed support evidence."
    )
    assert written_decisions["decisions"][1]["requested_documents"] == [
        "support_1002.pdf"
    ]
    assert written_decisions["decisions"][1]["followup_context"] == {
        "reason": "support_pdf"
    }
    apply_result = responses[7]["result"]["structuredContent"]
    assert apply_result["ok"] is True
    assert apply_result["persisted"] is True
    assert apply_result["run_intake_path"] == str(output_dir / "run_intake.json")
    assert apply_result["decision_count"] == 2
    assert apply_result["blocker_count"] == 1
    assert apply_result["structured_update_count"] == 1
    assert apply_result["native_regeneration_count"] == 0
    assert apply_result["native_regenerated_count"] == 1
    assert apply_result["application_status"] == "blocked"
    applied = json.loads((output_dir / "applied_decisions.json").read_text())
    assert applied["plugin"] == "check-entries"
    assert applied["effects"][0]["structured_update"] == {
        "id_field": "source_row",
        "record_id": "1",
        "target_field": "review_notes",
        "records_key": None,
        "updated_rows": 1,
    }
    assert applied["effects"][0]["derived_native_regeneration_paths"] == [
        "check_results.xlsx"
    ]
    assert applied["effects"][0]["requires_native_regeneration"] is False
    assert applied["effects"][0]["native_regeneration_status"] == "regenerated"
    assert applied["native_regeneration_paths"] == []
    assert applied["native_regenerated_paths"] == ["check_results.xlsx"]
    assert applied["effects"][1]["requires_followup"] is True
    assert applied["effects"][1]["followup_context"] == {"reason": "support_pdf"}
    assert "Reviewer confirmed support evidence." in check_results_path.read_text(
        encoding="utf-8"
    )
    workbook = openpyxl.load_workbook(check_results_xlsx_path)
    assert workbook.active["B2"].value == "Reviewer confirmed support evidence."
    final_artifacts = json.loads((output_dir / "final_artifacts.json").read_text())
    assert final_artifacts["status"] == "blocked"
    assert final_artifacts["review_application"]["structured_update_count"] == 1
    assert final_artifacts["review_application"]["structured_update_paths"] == [
        "check_results.csv"
    ]
    assert final_artifacts["review_application"]["native_regeneration_paths"] == []
    assert final_artifacts["review_application"]["native_regenerated_paths"] == [
        "check_results.xlsx"
    ]
    outputs_by_path = {output["path"]: output for output in final_artifacts["outputs"]}
    assert outputs_by_path["check_results.xlsx"]["status"] == "updated_from_review"
    assert outputs_by_path["check_results.xlsx"]["native_regenerated"] is True
    assert outputs_by_path["check_results.xlsx"]["source_artifact"] == (
        "check_results.csv"
    )
    assert outputs_by_path["check_results.xlsx"]["source_row_count"] == 2
    assert outputs_by_path["check_results.xlsx"]["required_sheets"] == ["Sheet1"]
    assert outputs_by_path["check_results.xlsx"]["required_cells"] == {
        "Sheet1": {"B2": "Reviewer confirmed support evidence."}
    }
    assert {"ui_decisions.json", "applied_decisions.json"} <= {
        output["path"] for output in final_artifacts["outputs"]
    }
    assert {
        "check_results.csv",
        "revisions/originals/check_results__entry-1.csv",
        "revisions/originals/check_results__entry-1.xlsx",
    } <= {output["path"] for output in final_artifacts["outputs"]}
    run_intake = json.loads((output_dir / "run_intake.json").read_text())
    review_apply_steps = [
        step
        for step in run_intake["execution_trace"]
        if step["kind"] == "deterministic_review_apply"
    ]
    assert len(review_apply_steps) == 1
    assert {
        "applied_decisions.json",
        "check_results.csv",
        "check_results.xlsx",
        "final_artifacts.json",
        "revisions/originals/check_results__entry-1.csv",
        "revisions/originals/check_results__entry-1.xlsx",
        "ui_decisions.json",
    } <= set(review_apply_steps[0]["outputs"])
    contract = validate_contract(
        output_dir,
        strict_data_posture=True,
        strict_execution_trace=True,
        strict_output_paths=True,
        strict_output_content=True,
    )
    assert contract.ok is True, contract.errors


@pytest.mark.parametrize(
    "decisions",
    [
        [{"item_id": "missing-item", "action": "accept"}],
        [{"item_id": "entry-1", "action": "edit"}],
        [{"item_id": "entry-1", "action": "request_more_documents"}],
    ],
)
def test_check_entries_mcp_server_rejects_invalid_review_decisions(
    tmp_path: Path,
    decisions: list[dict[str, object]],
) -> None:
    output_dir, managed_arguments = _portable_check_transaction_case(tmp_path)
    review_payload = managed_arguments["review_payload"]
    run_intake = managed_arguments["run_intake"]
    messages: list[dict[str, object]] = [
        {
            "jsonrpc": "2.0",
            "id": 1,
            "method": "tools/call",
            "params": {
                "name": "save_check_entries_decisions",
                "arguments": _managed_check_mcp_arguments(
                    output_dir,
                    {
                        "run_intake": run_intake,
                        "review_payload": review_payload,
                        "decisions": decisions,
                    },
                ),
            },
        },
    ]

    responses = {response["id"]: response for response in _call_mcp_server(messages)}

    result = responses[1]["result"]
    assert result["isError"] is True
    assert (
        result["structuredContent"]["error"]
        == "Check Entries review transaction failed safely."
    )
    assert not (output_dir / "ui_decisions.json").exists()


def test_forged_review_summary_cannot_grant_final_ready(tmp_path: Path) -> None:
    output_dir, managed_arguments = _portable_check_transaction_case(tmp_path)
    review_payload = managed_arguments["review_payload"]
    review_payload["summary"] = {
        "assurance_gates": {"report_ready": True},
        "professional_conclusion_status": "reviewed",
    }
    _seal_review_payload(review_payload)
    (output_dir / "review_payload.json").write_text(
        json.dumps(review_payload, indent=2) + "\n",
        encoding="utf-8",
    )
    run_intake = managed_arguments["run_intake"]
    forged_final = managed_arguments["final_artifacts"]
    forged_final["assurance_gates"] = {"report_ready": True}
    forged_final["professional_conclusion_status"] = "reviewed"
    (output_dir / "final_artifacts.json").write_text(
        json.dumps(forged_final, indent=2) + "\n",
        encoding="utf-8",
    )

    responses = _call_mcp_server(
        [
            {
                "jsonrpc": "2.0",
                "id": 1,
                "method": "tools/call",
                "params": {
                    "name": "apply_check_entries_decisions",
                    "arguments": _managed_check_mcp_arguments(
                        output_dir,
                        {
                            "run_intake": run_intake,
                            "review_payload": review_payload,
                            "final_artifacts": forged_final,
                            "decisions": [{"item_id": "entry-1", "action": "accept"}],
                        },
                    ),
                },
            }
        ]
    )

    result = responses[0]["result"]["structuredContent"]
    assert result["application_status"] == "blocked"
    assert (
        result["applied_decisions"]["assurance_preflight"]["assurance_replayed"]
        is False
    )
    assert result["final_artifacts"]["status"] == "blocked"


def test_check_entries_mcp_rejects_stale_review_payload_bindings(
    tmp_path: Path,
) -> None:
    review_payload = _seal_review_payload(
        {
            "schema_version": "2.0",
            "plugin": "check-entries",
            "workflow": "check-entries",
            "run_id": "check-entries-stale-review",
            "items": [
                {
                    "id": "entry-1",
                    "item_type": "supported_entry",
                    "title": "Original title",
                    "allowed_actions": ["accept", "skip"],
                    "recommended_action": "accept",
                }
            ],
            "item_count": 1,
            "status": "ready_for_review",
        }
    )
    tampered_payload = {
        **review_payload,
        "items": [{**review_payload["items"][0], "title": "Tampered title"}],
    }
    stale_ui_decisions = {
        "schema_version": "2.0",
        "plugin": "check-entries",
        "workflow": "check-entries",
        "run_id": review_payload["run_id"],
        "review_payload_path": "review_payload.json",
        "review_payload_content_sha256": "0" * 64,
        "decisions": [],
        "decision_count": 0,
        "status": "pending_review",
    }
    messages: list[dict[str, object]] = [
        {
            "jsonrpc": "2.0",
            "id": 1,
            "method": "tools/call",
            "params": {
                "name": "validate_check_entries_review",
                "arguments": {"review_payload": tampered_payload},
            },
        },
        {
            "jsonrpc": "2.0",
            "id": 2,
            "method": "tools/call",
            "params": {
                "name": "save_check_entries_decisions",
                "arguments": {
                    "review_payload": review_payload,
                    "ui_decisions": stale_ui_decisions,
                    "decisions": [{"item_id": "entry-1", "action": "accept"}],
                },
            },
        },
    ]

    responses = {response["id"]: response for response in _call_mcp_server(messages)}

    assert responses[1]["result"]["isError"] is True
    assert (
        responses[1]["result"]["structuredContent"]["error"]
        == "review_payload.content_sha256 is stale"
    )
    assert responses[2]["result"]["isError"] is True
    assert (
        responses[2]["result"]["structuredContent"]["error"]
        == "ui_decisions is bound to a different review_payload"
    )
    assert not (tmp_path / "ui_decisions.json").exists()


@pytest.mark.parametrize(
    "tool_name",
    ["save_check_entries_decisions", "apply_check_entries_decisions"],
)
@pytest.mark.parametrize(
    "intake_state",
    ["tampered", "missing"],
)
def test_check_entries_mcp_rejects_invalid_run_intake_before_any_write(
    monkeypatch: Any,
    tmp_path: Path,
    tool_name: str,
    intake_state: str,
) -> None:
    _, output_dir, review_payload, run_intake = _supported_assurance_run(
        monkeypatch,
        tmp_path,
    )
    submitted_intake = json.loads(json.dumps(run_intake))
    if intake_state == "tampered":
        submitted_intake["assumptions"]["amount_tolerance"] = "999"
    else:
        (output_dir / "run_intake.json").unlink()
    item = review_payload["items"][0]
    action = "accept" if "accept" in item["allowed_actions"] else "skip"
    ui_path = output_dir / "ui_decisions.json"
    final_path = output_dir / "final_artifacts.json"
    before_ui = ui_path.read_bytes()
    before_final = final_path.read_bytes()

    response = _call_mcp_server(
        [
            {
                "jsonrpc": "2.0",
                "id": 1,
                "method": "tools/call",
                "params": {
                    "name": tool_name,
                    "arguments": _managed_check_mcp_arguments(
                        output_dir,
                        {
                            "run_intake": submitted_intake,
                            "review_payload": review_payload,
                            "decisions": [
                                {
                                    "item_id": item["id"],
                                    "action": action,
                                }
                            ],
                        },
                    ),
                },
            }
        ]
    )[0]["result"]

    assert response["isError"] is True
    assert (
        response["structuredContent"]["error"]
        == "Check Entries persisted review authorization failed."
    )
    assert ui_path.read_bytes() == before_ui
    assert final_path.read_bytes() == before_final
    assert not (output_dir / "applied_decisions.json").exists()


@pytest.mark.parametrize(
    "mutation",
    ["add_nested", "add_unsupported", "delete"],
)
def test_support_directory_membership_change_restores_exact_prior_run(
    monkeypatch: Any,
    tmp_path: Path,
    mutation: str,
) -> None:
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-08-13",
                "movement": "M-90",
                "description": "Invoice support",
                "debit": "10",
            }
        ],
    )
    support_dir = tmp_path / "support"
    support_dir.mkdir()
    support = support_dir / "support.pdf"
    support.write_bytes(b"%PDF")
    output_dir = tmp_path / "out"
    (output_dir / "revisions").mkdir(parents=True)
    (output_dir / "prior.json").write_bytes(b'{"prior":true}\n')
    (output_dir / "revisions" / "prior.txt").write_bytes(b"prior revision")
    prior = _tree_snapshot(output_dir)
    monkeypatch.setattr(
        core,
        "_extract_pdf_text",
        lambda _path, _payload: (
            "Invoice\nMovement M-90\nSupplier VAT: 01234567890\n" "13/08/2025 EUR 10.00"
        ),
    )
    recipe_path = _reviewed_party_recipe(
        core,
        normalized,
        tmp_path / "recipe.json",
    )
    original_write_review = core.write_review_session_artifacts

    def mutate_membership(*args: Any, **kwargs: Any) -> Any:
        result = original_write_review(*args, **kwargs)
        if mutation == "add_nested":
            nested = support_dir / "late" / "added.xml"
            nested.parent.mkdir()
            nested.write_bytes(_fatturapa_xml())
        elif mutation == "add_unsupported":
            (support_dir / "late.txt").write_text("late", encoding="utf-8")
        else:
            support.unlink()
        return result

    monkeypatch.setattr(
        core,
        "write_review_session_artifacts",
        mutate_membership,
    )

    with pytest.raises((FileNotFoundError, ValueError), match="membership|receipt"):
        core.run_entry_checks(
            normalized,
            support_dir,
            output_dir,
            recipe_path,
        )

    assert _tree_snapshot(output_dir) == prior


def test_late_run_failure_restores_exact_prior_tree(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-08-14",
                "movement": "M-91",
                "description": "Invoice support",
                "debit": "10",
            }
        ],
    )
    support_dir = tmp_path / "support"
    support_dir.mkdir()
    (support_dir / "support.pdf").write_bytes(b"%PDF")
    output_dir = tmp_path / "out"
    (output_dir / "revisions").mkdir(parents=True)
    (output_dir / "prior.bin").write_bytes(b"\x00prior\xff")
    (output_dir / "revisions" / "prior.txt").write_text(
        "prior",
        encoding="utf-8",
    )
    prior = _tree_snapshot(output_dir)
    monkeypatch.setattr(
        core,
        "_extract_pdf_text",
        lambda _path, _payload: (
            "Invoice\nMovement M-91\nSupplier VAT: 01234567890\n" "14/08/2025 EUR 10.00"
        ),
    )
    recipe_path = _reviewed_party_recipe(
        core,
        normalized,
        tmp_path / "recipe.json",
    )

    def fail_after_envelope(path: Path, audit: dict[str, Any]) -> None:
        del audit
        path.write_text("partial", encoding="utf-8")
        raise RuntimeError("injected late run failure")

    monkeypatch.setattr(core, "_write_review_notes", fail_after_envelope)

    with pytest.raises(RuntimeError, match="injected late run failure"):
        core.run_entry_checks(
            normalized,
            support_dir,
            output_dir,
            recipe_path,
        )

    assert _tree_snapshot(output_dir) == prior


def test_successful_rerun_removes_prior_review_state(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-08-15",
                "movement": "M-92",
                "description": "Invoice support",
                "debit": "10",
            }
        ],
    )
    support_dir = tmp_path / "support"
    support_dir.mkdir()
    (support_dir / "support.pdf").write_bytes(b"%PDF")
    output_dir = tmp_path / "out"
    (output_dir / "revisions").mkdir(parents=True)
    (output_dir / "applied_decisions.json").write_text(
        '{"stale":true}\n',
        encoding="utf-8",
    )
    (output_dir / "revisions" / "stale.txt").write_text(
        "stale",
        encoding="utf-8",
    )
    monkeypatch.setattr(
        core,
        "_extract_pdf_text",
        lambda _path, _payload: (
            "Invoice\nMovement M-92\nSupplier VAT: 01234567890\n" "15/08/2025 EUR 10.00"
        ),
    )
    recipe_path = _reviewed_party_recipe(
        core,
        normalized,
        tmp_path / "recipe.json",
    )
    _reviewed_pdf_assertion_recipe(
        core,
        normalized,
        support_dir,
        recipe_path,
        support_locator="support.pdf",
        direction=True,
    )

    result = core.run_entry_checks(
        normalized,
        support_dir,
        output_dir,
        recipe_path,
    )

    assert result.frame.to_dicts()[0]["status"] == "ok"
    assert not (output_dir / "applied_decisions.json").exists()
    assert not (output_dir / "revisions").exists()


def test_two_zip_member_locators_bind_the_actual_archive_in_numeric_ledger(
    tmp_path: Path,
) -> None:
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-08-16",
                "movement": "M-93",
                "description": "Invoice INV-A",
                "debit": "10",
            },
            {
                "date": "2025-08-16",
                "movement": "M-94",
                "description": "Invoice INV-B",
                "debit": "20",
            },
        ],
    )
    support_dir = tmp_path / "support"
    support_dir.mkdir()
    with zipfile.ZipFile(support_dir / "a.zip", "w") as archive:
        archive.writestr(
            "invoice.xml",
            _fatturapa_xml(number="INV-A", invoice_date="2025-08-16", amount="10"),
        )
    with zipfile.ZipFile(support_dir / "b.zip", "w") as archive:
        archive.writestr(
            "invoice.xml",
            _fatturapa_xml(number="INV-B", invoice_date="2025-08-16", amount="20"),
        )
    recipe_path = _reviewed_party_recipe(
        core,
        normalized,
        tmp_path / "recipe.json",
    )
    _reviewed_pdf_assertion_recipe(
        core,
        normalized,
        support_dir,
        recipe_path,
        support_locator="a.zip!/invoice.xml",
        direction=True,
        entry_index=0,
        decision_suffix="1",
    )
    _reviewed_pdf_assertion_recipe(
        core,
        normalized,
        support_dir,
        recipe_path,
        support_locator="b.zip!/invoice.xml",
        direction=True,
        entry_index=1,
        decision_suffix="2",
    )

    result = core.run_entry_checks(
        normalized,
        support_dir,
        tmp_path / "out",
        recipe_path,
    )

    rows = result.frame.to_dicts()
    assert [row["status"] for row in rows] == ["ok", "ok"]
    assert [row["matched_support"] for row in rows] == [
        "a.zip!/invoice.xml",
        "b.zip!/invoice.xml",
    ]
    assert rows[0]["support_artifact_id"] != rows[1]["support_artifact_id"]
    ledger = json.loads((tmp_path / "out" / "numeric_evidence_ledger.json").read_text())
    support_entries = [
        entry for entry in ledger["entries"] if ".amount_found." in entry["evidence_id"]
    ]
    source_by_locator = {
        entry["source"]["locator"].split("::", 1)[0]: entry["source"]["artifact_ref"]
        for entry in support_entries
    }
    assert source_by_locator == {
        row["matched_support"]: row["support_artifact_id"] for row in rows
    }


def test_zip_and_top_level_xml_have_distinct_global_support_locators(
    tmp_path: Path,
) -> None:
    core = load_core()
    support_dir = tmp_path / "support"
    support_dir.mkdir()
    (support_dir / "invoice.xml").write_bytes(
        _fatturapa_xml(number="INV-X", amount="10")
    )
    with zipfile.ZipFile(support_dir / "archive.zip", "w") as archive:
        archive.writestr(
            "invoice.xml",
            _fatturapa_xml(number="INV-Z", amount="20"),
        )

    captured = core._load_captured_support(support_dir)

    locators = {invoice.source_name for invoice in captured.invoices}
    assert locators == {"invoice.xml", "archive.zip!/invoice.xml"}
    assert captured.invoice_artifact_ids["invoice.xml"] != (
        captured.invoice_artifact_ids["archive.zip!/invoice.xml"]
    )


def test_support_membership_rejects_casefold_duplicate_paths() -> None:
    core = load_core()

    with pytest.raises(ValueError, match="Unicode/casefold"):
        core._validate_unique_support_relative_paths(
            ["nested/Invoice.pdf", "nested/invoice.PDF"]
        )


def test_pdf_cad_label_and_dollar_symbol_cannot_pass_usd(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-08-17",
                "movement": "M-95",
                "description": "Invoice support",
                "debit": "10",
                "currency": "USD",
            }
        ],
    )
    support_dir = tmp_path / "support"
    support_dir.mkdir()
    (support_dir / "support.pdf").write_bytes(b"%PDF")
    monkeypatch.setattr(
        core,
        "_extract_pdf_text",
        lambda _path, _payload: (
            "Invoice\nMovement M-95\nSupplier VAT: 01234567890\n"
            "17/08/2025 CAD $10.00"
        ),
    )
    recipe_path = _reviewed_party_recipe(
        core,
        normalized,
        tmp_path / "recipe.json",
    )
    _reviewed_pdf_assertion_recipe(
        core,
        normalized,
        support_dir,
        recipe_path,
        support_locator="support.pdf",
        direction=True,
    )

    result = core.run_entry_checks(
        normalized,
        support_dir,
        tmp_path / "out",
        recipe_path,
    )

    row = result.frame.to_dicts()[0]
    assert row["status"] == "mismatch"
    assert "currency" in row["mismatches"].split(",")
    assert json.loads(row["evidence_facts"])["currency_explicit_conflict"] is True


def test_reviewed_currency_decision_can_close_ambiguous_pdf_symbol(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-08-18",
                "movement": "M-96",
                "description": "Invoice support",
                "debit": "10",
                "currency": "USD",
            }
        ],
    )
    support_dir = tmp_path / "support"
    support_dir.mkdir()
    (support_dir / "support.pdf").write_bytes(b"%PDF")
    monkeypatch.setattr(
        core,
        "_extract_pdf_text",
        lambda _path, _payload: (
            "Invoice\nMovement M-96\nSupplier VAT: 01234567890\n" "18/08/2025 $10.00"
        ),
    )
    recipe_path = _reviewed_party_recipe(
        core,
        normalized,
        tmp_path / "recipe.json",
    )
    _reviewed_pdf_assertion_recipe(
        core,
        normalized,
        support_dir,
        recipe_path,
        support_locator="support.pdf",
        currency=True,
        direction=True,
    )

    result = core.run_entry_checks(
        normalized,
        support_dir,
        tmp_path / "out",
        recipe_path,
    )

    row = result.frame.to_dicts()[0]
    assert row["status"] == "ok"
    facts = json.loads(row["evidence_facts"])
    assert facts["currency_found"] == "USD"
    assert facts["currency_decision_ref"] == "decision.check_entries_currency.1"


@pytest.mark.parametrize(
    ("document_type", "posting_field", "expected_polarity"),
    [
        ("TD01", "debit", "positive_document"),
        ("TD01", "credit", "positive_document"),
        ("TD04", "debit", "negative_document"),
        ("TD04", "credit", "negative_document"),
    ],
)
def test_fatturapa_type_does_not_override_reviewed_journal_direction(
    tmp_path: Path,
    document_type: str,
    posting_field: str,
    expected_polarity: str,
) -> None:
    core = load_core()
    row = {
        "date": "2025-08-19",
        "movement": "M-97",
        "description": "Invoice DOC-1",
        posting_field: "10",
    }
    normalized = _qualified_journal(tmp_path, [row])
    support = tmp_path / "document.xml"
    support.write_bytes(
        _fatturapa_xml(
            document_type=document_type,
            number="DOC-1",
            invoice_date="2025-08-19",
            amount="10",
        )
    )
    recipe_path = _reviewed_party_recipe(
        core,
        normalized,
        tmp_path / "recipe.json",
    )
    _reviewed_pdf_assertion_recipe(
        core,
        normalized,
        support,
        recipe_path,
        support_locator="document.xml",
        direction=True,
    )

    result = core.run_entry_checks(
        normalized,
        support,
        tmp_path / "out",
        recipe_path,
    )

    checked = result.frame.to_dicts()[0]
    facts = json.loads(checked["evidence_facts"])
    expected_direction = "debit" if posting_field == "debit" else "credit"
    expected_signed = "10" if posting_field == "debit" else "-10"
    assert checked["status"] == "ok"
    assert checked["support_amount_signed"] == expected_signed
    assert checked["amount_difference_signed"] == "0"
    assert facts["invoice_document_type"] == document_type
    assert facts["invoice_document_polarity"] == expected_polarity
    assert facts["reviewed_support_direction"] == expected_direction
    assert facts["direction_decision_ref"] == "decision.check_entries_direction.1"
    assert "direction" not in facts["support_match_signals"]


@pytest.mark.parametrize("document_type", ["TD01", "TD04"])
def test_fatturapa_without_reviewed_direction_withholds_signed_closure(
    tmp_path: Path,
    document_type: str,
) -> None:
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-08-19",
                "movement": "M-97-NO-DIRECTION",
                "description": "Invoice DOC-NO-DIRECTION",
                "credit": "10",
            }
        ],
    )
    support = tmp_path / "document.xml"
    support.write_bytes(
        _fatturapa_xml(
            document_type=document_type,
            number="DOC-NO-DIRECTION",
            invoice_date="2025-08-19",
            amount="10",
        )
    )
    recipe_path = _reviewed_party_recipe(
        core,
        normalized,
        tmp_path / "recipe.json",
    )

    result = core.run_entry_checks(
        normalized,
        support,
        tmp_path / "out",
        recipe_path,
    )

    checked = result.frame.to_dicts()[0]
    facts = json.loads(checked["evidence_facts"])
    assert checked["status"] == "manual_review"
    assert "direction_requires_review" in checked["mismatches"].split(",")
    assert checked["support_amount_signed"] is None
    assert checked["amount_difference_signed"] is None
    assert checked["amount_difference_abs"] is None
    assert facts["invoice_document_type"] == document_type
    assert facts["reviewed_support_direction"] is None
    assert facts["direction_decision_ref"] is None


@pytest.mark.parametrize("mutation", ["opposite_direction", "stale_entry"])
def test_fatturapa_rejects_mismatched_or_stale_direction_decision(
    tmp_path: Path,
    mutation: str,
) -> None:
    core = load_core()
    from vera_assurance import build_reviewed_decision_receipt

    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-08-19",
                "movement": "M-97-STALE",
                "description": "Invoice DOC-STALE",
                "credit": "10",
            }
        ],
    )
    support = tmp_path / "document.xml"
    support.write_bytes(
        _fatturapa_xml(
            document_type="TD01",
            number="DOC-STALE",
            invoice_date="2025-08-19",
            amount="10",
        )
    )
    recipe_path = _reviewed_party_recipe(
        core,
        normalized,
        tmp_path / "recipe.json",
    )
    frame, _ = core._load_journal_entries(normalized, {})
    _, captures = core._capture_support(support)
    row = frame.to_dicts()[0]
    artifact_id = captures[0].receipt["artifact_id"]
    content = {
        "prepared_entry_id": row["prepared_entry_id"],
        "support_artifact_id": artifact_id,
        "support_locator": "document.xml",
        "expected_direction": "credit",
        "direction_status": "confirmed",
        "recording_exception": (
            "The reviewer confirmed which journal line the document supports."
        ),
    }
    if mutation == "opposite_direction":
        content["expected_direction"] = "debit"
    else:
        content["prepared_entry_id"] = "prepared.check_entries.stale"
    direction_decision = build_reviewed_decision_receipt(
        decision_id=f"decision.check_entries_direction.{mutation}",
        decision_type="check_entries_direction",
        status="reviewed",
        reviewer_ref="reviewer.check_entries_test",
        reviewed_on="2026-07-24",
        adapter_id="check_entries.direction",
        adapter_version="1",
        source_artifact_refs=["source.normalized_journal", artifact_id],
        content=content,
    )
    recipe = json.loads(recipe_path.read_text(encoding="utf-8"))
    recipe["reviewed_direction_decisions"] = [direction_decision]
    recipe_path.write_text(
        json.dumps(recipe, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    with pytest.raises(ValueError, match="direction decision binding is stale"):
        core.run_entry_checks(
            normalized,
            support,
            tmp_path / "out",
            recipe_path,
        )


@pytest.mark.parametrize(
    ("document_label", "posting_field", "expected_polarity"),
    [
        ("Invoice", "debit", "positive_document"),
        ("Credit note", "credit", "negative_document"),
    ],
)
def test_pdf_document_label_does_not_replace_reviewed_direction(
    monkeypatch: Any,
    tmp_path: Path,
    document_label: str,
    posting_field: str,
    expected_polarity: str,
) -> None:
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-08-20",
                "movement": "M-98-LABEL",
                "description": "Labeled document support",
                posting_field: "10",
            }
        ],
    )
    support_dir = tmp_path / "support"
    support_dir.mkdir()
    (support_dir / "support.pdf").write_bytes(b"%PDF")
    monkeypatch.setattr(
        core,
        "_extract_pdf_text",
        lambda _path, _payload: (
            f"{document_label}\nMovement M-98-LABEL\n"
            "Supplier VAT: 01234567890\n20/08/2025 EUR 10.00"
        ),
    )
    recipe_path = _reviewed_party_recipe(
        core,
        normalized,
        tmp_path / "recipe.json",
    )

    result = core.run_entry_checks(
        normalized,
        support_dir,
        tmp_path / "out",
        recipe_path,
    )

    row = result.frame.to_dicts()[0]
    facts = json.loads(row["evidence_facts"])
    assert row["status"] == "manual_review"
    assert "direction_requires_review" in row["mismatches"].split(",")
    assert row["support_amount_signed"] is None
    assert row["amount_difference_signed"] is None
    assert facts["document_polarity"] == expected_polarity
    assert facts["reviewed_support_direction"] is None


def test_ambiguous_pdf_direction_fails_closed_without_review(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-08-20",
                "movement": "M-98",
                "description": "Support",
                "debit": "10",
            }
        ],
    )
    support_dir = tmp_path / "support"
    support_dir.mkdir()
    (support_dir / "support.pdf").write_bytes(b"%PDF")
    monkeypatch.setattr(
        core,
        "_extract_pdf_text",
        lambda _path, _payload: (
            "Movement M-98\nSupplier VAT: 01234567890\n" "20/08/2025 EUR 10.00"
        ),
    )
    recipe_path = _reviewed_party_recipe(
        core,
        normalized,
        tmp_path / "recipe.json",
    )

    result = core.run_entry_checks(
        normalized,
        support_dir,
        tmp_path / "out",
        recipe_path,
    )

    row = result.frame.to_dicts()[0]
    assert row["status"] == "manual_review"
    assert "direction_requires_review" in row["mismatches"].split(",")


def test_reviewed_direction_decision_can_close_ambiguous_pdf_direction(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-08-21",
                "movement": "M-99",
                "description": "Support",
                "debit": "10",
            }
        ],
    )
    support_dir = tmp_path / "support"
    support_dir.mkdir()
    (support_dir / "support.pdf").write_bytes(b"%PDF")
    monkeypatch.setattr(
        core,
        "_extract_pdf_text",
        lambda _path, _payload: (
            "Movement M-99\nSupplier VAT: 01234567890\n" "21/08/2025 EUR 10.00"
        ),
    )
    recipe_path = _reviewed_party_recipe(
        core,
        normalized,
        tmp_path / "recipe.json",
    )
    _reviewed_pdf_assertion_recipe(
        core,
        normalized,
        support_dir,
        recipe_path,
        support_locator="support.pdf",
        direction=True,
    )

    result = core.run_entry_checks(
        normalized,
        support_dir,
        tmp_path / "out",
        recipe_path,
    )

    row = result.frame.to_dicts()[0]
    assert row["status"] == "ok"
    assert json.loads(row["evidence_facts"])["direction_decision_ref"] == (
        "decision.check_entries_direction.1"
    )


def test_mcp_apply_python_failure_rolls_back_exact_output_tree(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    _, output_dir, review_payload, run_intake = _supported_assurance_run(
        monkeypatch,
        tmp_path,
    )
    entry_item = next(
        item
        for item in review_payload["items"]
        if item["item_type"] == "supported_entry"
    )
    prior = _tree_snapshot(output_dir)
    real_python = Path(sys.executable)
    failing_python = tmp_path / "failing-python"
    failing_python.write_text(
        "\n".join(
            [
                "#!/usr/bin/env python3",
                "import subprocess",
                "import sys",
                f"real = {str(real_python)!r}",
                'if "--client-run-preflight-only" in sys.argv:',
                "    completed = subprocess.run([real, *sys.argv[1:]])",
                "    raise SystemExit(completed.returncode)",
                "raise SystemExit(1)",
                "",
            ]
        ),
        encoding="utf-8",
    )
    failing_python.chmod(0o700)
    monkeypatch.setenv("PYTHON", str(failing_python))

    response = _call_mcp_server(
        [
            {
                "jsonrpc": "2.0",
                "id": 1,
                "method": "tools/call",
                "params": {
                    "name": "apply_check_entries_decisions",
                    "arguments": _managed_check_mcp_arguments(
                        output_dir,
                        {
                            "run_intake": run_intake,
                            "review_payload": review_payload,
                            "decisions": [
                                {
                                    "item_id": entry_item["id"],
                                    "action": "edit",
                                    "edit_value": "Injected rollback test.",
                                }
                            ],
                        },
                    ),
                },
            }
        ]
    )[0]["result"]

    assert response["isError"] is True
    assert (
        response["structuredContent"]["error"]
        == "Check Entries assurance preflight failed."
    )
    assert _tree_snapshot(output_dir) == prior


def test_pdf_tax_id_with_opposite_party_role_cannot_emit_ok(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    # Arrange
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-09-01",
                "movement": "M-ROLE",
                "description": "Invoice support",
                "debit": "10",
            }
        ],
    )
    support_dir = tmp_path / "support"
    support_dir.mkdir()
    (support_dir / "support.pdf").write_bytes(b"%PDF")
    monkeypatch.setattr(
        core,
        "_extract_pdf_text",
        lambda _path, _payload: (
            "Invoice\nMovement M-ROLE\nCustomer VAT: 01234567890\n"
            "01/09/2025 EUR 10.00"
        ),
    )
    recipe_path = _reviewed_party_recipe(
        core,
        normalized,
        tmp_path / "recipe.json",
        tax_id="01234567890",
        expected_role="supplier",
    )
    _reviewed_pdf_assertion_recipe(
        core,
        normalized,
        support_dir,
        recipe_path,
        support_locator="support.pdf",
        direction=True,
    )

    # Act
    result = core.run_entry_checks(
        normalized,
        support_dir,
        tmp_path / "out",
        recipe_path,
    )

    # Assert
    row = result.frame.to_dicts()[0]
    facts = json.loads(row["evidence_facts"])
    assert row["status"] == "mismatch"
    assert "party_perimeter" in row["mismatches"].split(",")
    assert facts["party_perimeter_status"] == "mismatch"
    assert facts["party_perimeter_signal"] == "opposite_role_labeled_tax_id"
    assert (
        result.audit["assurance_gates"]["gates"]["reconciliation"]["status"] == "failed"
    )


def test_reviewed_currency_cannot_override_explicit_mxn_conflict(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    # Arrange
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-09-05",
                "movement": "M-MXN",
                "description": "Invoice support",
                "debit": "10",
                "currency": "USD",
            }
        ],
    )
    support_dir = tmp_path / "support"
    support_dir.mkdir()
    (support_dir / "support.pdf").write_bytes(b"%PDF")
    monkeypatch.setattr(
        core,
        "_extract_pdf_text",
        lambda _path, _payload: (
            "Invoice\nMovement M-MXN\nSupplier VAT: 01234567890\n"
            "05/09/2025 MXN $10.00"
        ),
    )
    recipe_path = _reviewed_party_recipe(
        core,
        normalized,
        tmp_path / "recipe.json",
    )
    _reviewed_pdf_assertion_recipe(
        core,
        normalized,
        support_dir,
        recipe_path,
        support_locator="support.pdf",
        currency=True,
        direction=True,
    )

    # Act
    result = core.run_entry_checks(
        normalized,
        support_dir,
        tmp_path / "out",
        recipe_path,
    )

    # Assert
    row = result.frame.to_dicts()[0]
    facts = json.loads(row["evidence_facts"])
    assert row["status"] == "mismatch"
    assert "currency" in row["mismatches"].split(",")
    assert facts["currency_found"] is None
    assert facts["currency_explicit_conflict"] is True


def test_lowercase_word_that_is_iso_code_is_not_an_explicit_currency_conflict() -> None:
    # Arrange
    core = load_core()

    # Act
    ordinary_word_conflict = core._explicit_currency_conflict(
        "USD",
        "Invoice for all services USD 10.00",
    )
    uppercase_code_conflict = core._explicit_currency_conflict(
        "USD",
        "Invoice total MXN 10.00",
    )

    # Assert
    assert ordinary_word_conflict is False
    assert uppercase_code_conflict is True


def test_support_hardlink_aliases_are_rejected(tmp_path: Path) -> None:
    # Arrange
    core = load_core()
    support_dir = tmp_path / "support"
    support_dir.mkdir()
    original = support_dir / "a.xml"
    alias = support_dir / "b.xml"
    original.write_bytes(_fatturapa_xml(number="HARDLINK-1"))
    alias.hardlink_to(original)

    # Act / Assert
    with pytest.raises(ValueError, match="hardlink"):
        core._load_captured_support(support_dir)


def test_temporary_prefixed_file_is_sealed_and_fails_qualification(
    tmp_path: Path,
) -> None:
    # Arrange
    core = load_core()
    support_dir = tmp_path / "support"
    support_dir.mkdir()
    (support_dir / "invoice.xml").write_bytes(_fatturapa_xml(number="TEMP-1"))
    (support_dir / "~$unqualified.txt").write_text(
        "not support",
        encoding="utf-8",
    )

    # Act
    captured = core._load_captured_support(support_dir)

    # Assert
    assert captured.manifest["canonical_relative_paths"] == [
        "invoice.xml",
        "~$unqualified.txt",
    ]
    status_by_source = {
        qualification["source_artifact_refs"][0]: qualification["status"]
        for qualification in captured.source_qualifications
    }
    receipt_by_path = {
        receipt["path"]: receipt["artifact_id"]
        for receipt in captured.manifest["artifact_receipts"]
    }
    assert status_by_source[receipt_by_path["invoice.xml"]] == "qualified"
    assert (
        status_by_source[receipt_by_path["~$unqualified.txt"]]
        == "unsupported_source_layout"
    )


def test_mismatched_credit_note_closes_signed_support_and_difference_ledger(
    tmp_path: Path,
) -> None:
    # Arrange
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-09-02",
                "movement": "M-DIFF",
                "description": "Invoice CR-DIFF",
                "credit": "10",
            }
        ],
    )
    invoice = tmp_path / "credit.xml"
    invoice.write_bytes(
        _fatturapa_xml(
            document_type="TD04",
            number="CR-DIFF",
            invoice_date="2025-09-02",
            amount="11",
        )
    )
    recipe_path = _reviewed_party_recipe(
        core,
        normalized,
        tmp_path / "recipe.json",
    )
    _reviewed_pdf_assertion_recipe(
        core,
        normalized,
        invoice,
        recipe_path,
        support_locator="credit.xml",
        direction=True,
    )
    output_dir = tmp_path / "out"

    # Act
    result = core.run_entry_checks(
        normalized,
        invoice,
        output_dir,
        recipe_path,
    )

    # Assert
    row = result.frame.to_dicts()[0]
    assert row["status"] == "mismatch"
    assert row["amount_signed"] == "-10"
    assert row["support_amount_signed"] == "-11"
    assert row["amount_difference_signed"] == "-1"
    assert row["amount_difference_abs"] == "1"
    assert row["amount_found"] is None
    ledger = result.audit["numeric_evidence_ledger"]
    value_by_field = {
        field: next(
            entry["value"]
            for entry in ledger["entries"]
            if f".{field}." in entry["evidence_id"]
        )
        for field in (
            "amount_signed",
            "support_amount_signed",
            "amount_difference_signed",
            "amount_difference_abs",
        )
    }
    assert value_by_field == {
        "amount_signed": "-10",
        "support_amount_signed": "-11",
        "amount_difference_signed": "-1",
        "amount_difference_abs": "1",
    }
    with (output_dir / "prepared_support_facts.csv").open(
        "r",
        encoding="utf-8",
        newline="",
    ) as handle:
        prepared_rows = list(csv.DictReader(handle))
    assert prepared_rows[0]["support_amount_signed"] == "-11"
    assert prepared_rows[0]["amount_difference_signed"] == "-1"
    assert prepared_rows[0]["amount_difference_abs"] == "1"


def test_setup_failure_after_prior_move_restores_prior_without_backup(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    # Arrange
    core = load_core()
    normalized = _qualified_journal(
        tmp_path,
        [
            {
                "date": "2025-09-03",
                "movement": "M-SETUP",
                "description": "Setup rollback",
                "debit": "10",
            }
        ],
    )
    support_dir = tmp_path / "support"
    support_dir.mkdir()
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    (output_dir / "prior.bin").write_bytes(b"prior")
    original_mkdir = Path.mkdir

    def fail_target_recreation(
        self: Path,
        mode: int = 0o777,
        parents: bool = False,
        exist_ok: bool = False,
    ) -> None:
        if self == output_dir and not self.exists():
            raise OSError("injected target recreation failure")
        original_mkdir(
            self,
            mode=mode,
            parents=parents,
            exist_ok=exist_ok,
        )

    monkeypatch.setattr(Path, "mkdir", fail_target_recreation)

    # Act / Assert
    with pytest.raises(OSError, match="injected target recreation failure"):
        core.run_entry_checks(
            normalized,
            support_dir,
            output_dir,
        )
    assert (output_dir / "prior.bin").read_bytes() == b"prior"
    assert list(tmp_path.glob(".out.check-entries-backup-*")) == []


def _mcp_review_write_message(
    tool_name: str,
    review_payload: dict[str, Any],
    run_intake: dict[str, Any],
    *,
    output_dir: Path,
) -> list[dict[str, object]]:
    item = next(
        entry
        for entry in review_payload["items"]
        if entry["item_type"] == "supported_entry"
    )
    return [
        {
            "jsonrpc": "2.0",
            "id": 1,
            "method": "tools/call",
            "params": {
                "name": tool_name,
                "arguments": _managed_check_mcp_arguments(
                    output_dir,
                    {
                        "run_intake": run_intake,
                        "review_payload": review_payload,
                        "decisions": [
                            {
                                "item_id": item["id"],
                                "action": "edit",
                                "edit_value": "Transaction containment review.",
                            }
                        ],
                    },
                ),
            },
        }
    ]


@pytest.mark.parametrize(
    "tool_name",
    ["save_check_entries_decisions", "apply_check_entries_decisions"],
)
@pytest.mark.parametrize(
    "scenario",
    [
        "poison_file",
        "delete_snapshot",
        "symlink_snapshot",
        "fifo_snapshot",
        "canonical_delete",
    ],
)
def test_mcp_review_transaction_restores_trusted_bytes_and_modes_after_commit_fault(
    monkeypatch: Any,
    tmp_path: Path,
    tool_name: str,
    scenario: str,
) -> None:
    # Arrange
    if scenario == "fifo_snapshot" and sys.platform == "win32":
        pytest.skip("FIFO transaction probe requires a POSIX host.")
    _, output_dir, review_payload, run_intake = _supported_assurance_run(
        monkeypatch,
        tmp_path,
    )
    untouched_file = output_dir / "review_notes.md"
    untouched_file.chmod(0o640)
    output_dir.chmod(0o750)
    before = _transaction_tree_state(output_dir)
    marker, external = _install_review_transaction_commit_fault(
        monkeypatch,
        tmp_path,
        output_dir,
        scenario,
    )
    external_before = _transaction_tree_state(external)

    # Act
    response = _call_mcp_server(
        _mcp_review_write_message(
            tool_name,
            review_payload,
            run_intake,
            output_dir=output_dir,
        )
    )[0]["result"]

    # Assert
    assert marker.read_text(encoding="utf-8") == "commit fault triggered\n"
    assert response["isError"] is True
    error = response["structuredContent"]["error"]
    assert error == "Check Entries review transaction failed safely."
    assert len(error) <= 240
    assert "\n" not in error
    assert "/" not in error
    assert "\\" not in error
    assert "Traceback" not in error
    assert _transaction_tree_state(output_dir) == before
    assert _transaction_tree_state(external) == external_before
    assert not (output_dir / "poison.fifo").exists()
    assert not list(tmp_path.glob(".generated-review-transaction-*"))
    assert not list(tmp_path.glob(".generated-review-commit-*"))
    assert not list(tmp_path.glob(".generated-review-recovery-*"))


def test_mcp_review_transaction_rejects_root_relocation_without_moving_canonical(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    _, output_dir, review_payload, run_intake = _supported_assurance_run(
        monkeypatch,
        tmp_path,
    )
    before = _transaction_tree_state(output_dir)
    canonical_inode = output_dir.stat().st_ino
    marker = _install_review_transaction_root_relocation(
        monkeypatch,
        tmp_path,
    )

    response = _call_mcp_server(
        _mcp_review_write_message(
            "apply_check_entries_decisions",
            review_payload,
            run_intake,
            output_dir=output_dir,
        )
    )[0]["result"]

    assert marker.read_text(encoding="utf-8") == ("transaction root relocated\n")
    assert response["isError"] is True
    assert output_dir.stat().st_ino == canonical_inode
    assert _transaction_tree_state(output_dir) == before
    assert not list(tmp_path.glob(".generated-review-transaction-*"))
    assert not list(tmp_path.glob(".generated-review-commit-*"))
    assert not list(tmp_path.glob(".generated-review-recovery-*"))


@pytest.mark.parametrize(
    "tool_name",
    ["save_check_entries_decisions", "apply_check_entries_decisions"],
)
def test_mcp_review_transaction_bounds_fail_before_canonical_mutation(
    monkeypatch: Any,
    tmp_path: Path,
    tool_name: str,
) -> None:
    # Arrange
    _, output_dir, review_payload, run_intake = _supported_assurance_run(
        monkeypatch,
        tmp_path,
    )
    sentinel = output_dir / "run_intake.json"
    sentinel_before = (
        sentinel.read_bytes(),
        sentinel.lstat().st_ino,
        stat.S_IMODE(sentinel.lstat().st_mode),
    )
    oversized = output_dir / "oversized.bin"
    with oversized.open("wb") as handle:
        handle.truncate(128 * 1024 * 1024 + 1)
    oversized.chmod(0o640)
    oversized_before = oversized.lstat()
    marker, external = _install_review_transaction_commit_fault(
        monkeypatch,
        tmp_path,
        output_dir,
        "poison_file",
    )
    external_before = _transaction_tree_state(external)

    # Act
    response = _call_mcp_server(
        _mcp_review_write_message(
            tool_name,
            review_payload,
            run_intake,
            output_dir=output_dir,
        )
    )[0]["result"]

    # Assert
    assert response["isError"] is True
    assert (
        response["structuredContent"]["error"]
        == "Check Entries review transaction failed safely."
    )
    assert not marker.exists()
    assert (
        sentinel.read_bytes(),
        sentinel.lstat().st_ino,
        stat.S_IMODE(sentinel.lstat().st_mode),
    ) == sentinel_before
    oversized_after = oversized.lstat()
    assert oversized_after.st_ino == oversized_before.st_ino
    assert oversized_after.st_size == oversized_before.st_size
    assert stat.S_IMODE(oversized_after.st_mode) == 0o640
    assert _transaction_tree_state(external) == external_before
    assert not list(tmp_path.glob(".generated-review-transaction-*"))


@pytest.mark.parametrize(
    "tool_name",
    ["save_check_entries_decisions", "apply_check_entries_decisions"],
)
def test_mcp_review_transaction_honest_commit_preserves_unwritten_nested_modes(
    monkeypatch: Any,
    tmp_path: Path,
    tool_name: str,
) -> None:
    # Arrange
    _, output_dir, review_payload, run_intake = _supported_assurance_run(
        monkeypatch,
        tmp_path,
    )
    untouched_file = output_dir / "review_notes.md"
    untouched_file.chmod(0o640)
    output_dir.chmod(0o750)
    before = _transaction_tree_state(output_dir)

    # Act
    response = _call_mcp_server(
        _mcp_review_write_message(
            tool_name,
            review_payload,
            run_intake,
            output_dir=output_dir,
        )
    )[0]["result"]

    # Assert
    assert response["isError"] is False
    after = _transaction_tree_state(output_dir)
    assert after["."] == before["."]
    assert after["review_notes.md"] == before["review_notes.md"]
    expected_artifact = (
        "ui_decisions.json"
        if tool_name == "save_check_entries_decisions"
        else "applied_decisions.json"
    )
    assert (output_dir / expected_artifact).is_file()
    assert not list(tmp_path.glob(".generated-review-transaction-*"))


@pytest.mark.parametrize(
    "tool_name",
    ["save_check_entries_decisions", "apply_check_entries_decisions"],
)
@pytest.mark.parametrize(
    ("entry_kind", "expected_error"),
    [
        ("symlink", "symbolic links"),
        ("hardlink", "hardlink aliases"),
        ("fifo", "special filesystem entries"),
    ],
)
def test_mcp_review_transactions_reject_unsafe_internal_entries_before_writes(
    monkeypatch: Any,
    tmp_path: Path,
    tool_name: str,
    entry_kind: str,
    expected_error: str,
) -> None:
    # Arrange
    _, output_dir, review_payload, run_intake = _supported_assurance_run(
        monkeypatch,
        tmp_path,
    )
    target = output_dir / "run_intake.json"
    outside = tmp_path / f"outside-{tool_name}-{entry_kind}.json"
    outside_before: bytes | None = None
    if entry_kind in {"symlink", "hardlink"}:
        outside.write_bytes(target.read_bytes())
        outside_before = outside.read_bytes()
        target.unlink()
        if entry_kind == "symlink":
            target.symlink_to(outside)
        else:
            os.link(outside, target)
    else:
        target = output_dir / "unexpected.fifo"
        os.mkfifo(target)

    # Act
    response = _call_mcp_server(
        _mcp_review_write_message(
            tool_name,
            review_payload,
            run_intake,
            output_dir=output_dir,
        )
    )[0]["result"]

    # Assert
    assert response["isError"] is True
    assert expected_error in response["structuredContent"]["error"]
    if outside_before is not None:
        assert outside.read_bytes() == outside_before
    if entry_kind == "symlink":
        assert target.is_symlink()
    elif entry_kind == "hardlink":
        assert target.stat().st_nlink == 2
    else:
        assert stat.S_ISFIFO(target.lstat().st_mode)
    assert list(output_dir.parent.glob(".check-entries-apply-*")) == []


def test_mcp_rejects_preexisting_internal_symlink_without_external_write(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    # Arrange
    _, output_dir, review_payload, run_intake = _supported_assurance_run(
        monkeypatch,
        tmp_path,
    )
    intake_path = output_dir / "run_intake.json"
    outside = tmp_path / "outside-run-intake.json"
    outside.write_bytes(intake_path.read_bytes())
    before_outside = outside.read_bytes()
    intake_path.unlink()
    intake_path.symlink_to(outside)
    item = next(
        entry
        for entry in review_payload["items"]
        if entry["item_type"] == "supported_entry"
    )

    # Act
    response = _call_mcp_server(
        [
            {
                "jsonrpc": "2.0",
                "id": 1,
                "method": "tools/call",
                "params": {
                    "name": "apply_check_entries_decisions",
                    "arguments": _managed_check_mcp_arguments(
                        output_dir,
                        {
                            "run_intake": run_intake,
                            "review_payload": review_payload,
                            "decisions": [
                                {
                                    "item_id": item["id"],
                                    "action": "edit",
                                    "edit_value": "Must stay inside the run.",
                                }
                            ],
                        },
                    ),
                },
            }
        ]
    )[0]["result"]

    # Assert
    assert response["isError"] is True
    assert "symbolic links" in response["structuredContent"]["error"]
    assert outside.read_bytes() == before_outside
    assert intake_path.is_symlink()


def test_mcp_post_preflight_symlink_swap_rolls_back_without_external_edit(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    # Arrange
    _, output_dir, review_payload, run_intake = _supported_assurance_run(
        monkeypatch,
        tmp_path,
    )
    before = _tree_snapshot(output_dir)
    outside = tmp_path / "outside-check-results.csv"
    real_python = Path(sys.executable)
    wrapper = tmp_path / "python-preflight-symlink-swap"
    wrapper.write_text(
        "\n".join(
            [
                "#!/usr/bin/env python3",
                "import os",
                "from pathlib import Path",
                "import subprocess",
                "import sys",
                f"real = {str(real_python)!r}",
                "completed = subprocess.run([real, *sys.argv[1:]], capture_output=True)",
                "sys.stdout.buffer.write(completed.stdout)",
                "sys.stderr.buffer.write(completed.stderr)",
                'if completed.returncode == 0 and "--preflight-only" in sys.argv:',
                '    out = Path(sys.argv[sys.argv.index("--output-dir") + 1])',
                '    outside = Path(os.environ["CE_TEST_OUTSIDE"])',
                '    target = out / "check_results.csv"',
                "    outside.write_bytes(target.read_bytes())",
                "    target.unlink()",
                "    target.symlink_to(outside)",
                "raise SystemExit(completed.returncode)",
                "",
            ]
        ),
        encoding="utf-8",
    )
    wrapper.chmod(0o700)
    monkeypatch.setenv("PYTHON", wrapper.as_posix())
    monkeypatch.setenv("CE_TEST_OUTSIDE", outside.as_posix())
    item = next(
        entry
        for entry in review_payload["items"]
        if entry["item_type"] == "supported_entry"
    )

    # Act
    response = _call_mcp_server(
        [
            {
                "jsonrpc": "2.0",
                "id": 1,
                "method": "tools/call",
                "params": {
                    "name": "apply_check_entries_decisions",
                    "arguments": _managed_check_mcp_arguments(
                        output_dir,
                        {
                            "run_intake": run_intake,
                            "review_payload": review_payload,
                            "decisions": [
                                {
                                    "item_id": item["id"],
                                    "action": "edit",
                                    "edit_value": "Independent transaction probe.",
                                }
                            ],
                        },
                    ),
                },
            }
        ]
    )[0]["result"]

    # Assert
    assert response["isError"] is True
    assert "symbolic links" in response["structuredContent"]["error"]
    assert _tree_snapshot(output_dir) == before
    assert not (output_dir / "check_results.csv").is_symlink()
    assert outside.exists()
    assert b"Independent transaction probe." not in outside.read_bytes()


@pytest.mark.parametrize("link_kind", ["symlink", "hardlink"])
def test_mcp_apply_phase_link_swap_is_rejected_before_external_mutation(
    monkeypatch: Any,
    tmp_path: Path,
    link_kind: str,
) -> None:
    # Arrange
    _, output_dir, review_payload, run_intake = _supported_assurance_run(
        monkeypatch,
        tmp_path,
    )
    before = _tree_snapshot(output_dir)
    outside = tmp_path / f"outside-applied-{link_kind}.json"
    real_python = Path(sys.executable)
    wrapper = tmp_path / f"python-apply-{link_kind}-swap"
    wrapper.write_text(
        "\n".join(
            [
                "#!/usr/bin/env python3",
                "import os",
                "from pathlib import Path",
                "import subprocess",
                "import sys",
                f"real = {str(real_python)!r}",
                'if "--preflight-only" in sys.argv or "--client-run-preflight-only" in sys.argv:',
                "    completed = subprocess.run([real, *sys.argv[1:]])",
                "    raise SystemExit(completed.returncode)",
                'out = Path(sys.argv[sys.argv.index("--output-dir") + 1])',
                'outside = Path(os.environ["CE_TEST_OUTSIDE"])',
                'link_kind = os.environ["CE_TEST_LINK_KIND"]',
                'target = out / "applied_decisions.json"',
                "outside.write_bytes(target.read_bytes())",
                "target.unlink()",
                "if link_kind == 'symlink':",
                "    target.symlink_to(outside)",
                "else:",
                "    os.link(outside, target)",
                "completed = subprocess.run([real, *sys.argv[1:]])",
                "raise SystemExit(completed.returncode)",
                "",
            ]
        ),
        encoding="utf-8",
    )
    wrapper.chmod(0o700)
    monkeypatch.setenv("PYTHON", wrapper.as_posix())
    monkeypatch.setenv("CE_TEST_OUTSIDE", outside.as_posix())
    monkeypatch.setenv("CE_TEST_LINK_KIND", link_kind)

    # Act
    response = _call_mcp_server(
        _mcp_review_write_message(
            "apply_check_entries_decisions",
            review_payload,
            run_intake,
            output_dir=output_dir,
        )
    )[0]["result"]

    # Assert
    assert response["isError"] is True
    expected_error = "symbolic links" if link_kind == "symlink" else "hardlink aliases"
    assert expected_error in response["structuredContent"]["error"]
    assert _tree_snapshot(output_dir) == before
    outside_payload = json.loads(outside.read_text(encoding="utf-8"))
    assert "native_regenerated_count" not in outside_payload
    assert all(
        effect.get("native_regeneration_status") != "regenerated"
        for effect in outside_payload["effects"]
    )
    assert list(output_dir.parent.glob(".check-entries-apply-*")) == []


def test_mcp_dangling_canonical_output_swap_restores_exact_prior_tree(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    # Arrange
    _, output_dir, review_payload, run_intake = _supported_assurance_run(
        monkeypatch,
        tmp_path,
    )
    before = _tree_snapshot(output_dir)
    real_python = Path(sys.executable)
    wrapper = tmp_path / "python-apply-dangling-output-swap"
    wrapper.write_text(
        "\n".join(
            [
                "#!/usr/bin/env python3",
                "from pathlib import Path",
                "import shutil",
                "import subprocess",
                "import sys",
                f"real = {str(real_python)!r}",
                'if "--preflight-only" in sys.argv:',
                "    completed = subprocess.run([real, *sys.argv[1:]])",
                "    raise SystemExit(completed.returncode)",
                'canonical = Path(sys.argv[sys.argv.index("--canonical-output-dir") + 1])',
                "shutil.rmtree(canonical)",
                "canonical.symlink_to(canonical.parent / 'missing-output-target')",
                "completed = subprocess.run([real, *sys.argv[1:]])",
                "raise SystemExit(completed.returncode)",
                "",
            ]
        ),
        encoding="utf-8",
    )
    wrapper.chmod(0o700)
    monkeypatch.setenv("PYTHON", wrapper.as_posix())

    # Act
    response = _call_mcp_server(
        _mcp_review_write_message(
            "apply_check_entries_decisions",
            review_payload,
            run_intake,
            output_dir=output_dir,
        )
    )[0]["result"]

    # Assert
    assert response["isError"] is True
    assert response["structuredContent"]["error"] == (
        "Check Entries review application failed."
    )
    assert "rollback was incomplete" not in response["structuredContent"]["error"]
    assert not output_dir.is_symlink()
    assert _tree_snapshot(output_dir) == before
    assert not list(output_dir.parent.glob(".generated-review-transaction-*"))
    assert not list(output_dir.parent.glob(".generated-review-commit-*"))
    assert not list(output_dir.parent.glob(".generated-review-recovery-*"))


def test_mcp_save_decisions_commits_only_canonical_decision_artifact(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    # Arrange
    _, output_dir, review_payload, run_intake = _supported_assurance_run(
        monkeypatch,
        tmp_path,
    )
    before = _tree_snapshot(output_dir)

    # Act
    response = _call_mcp_server(
        _mcp_review_write_message(
            "save_check_entries_decisions",
            review_payload,
            run_intake,
            output_dir=output_dir,
        )
    )[0]["result"]

    # Assert
    assert response["isError"] is False
    assert (
        response["structuredContent"]["ui_decisions_path"]
        == (output_dir / "ui_decisions.json").as_posix()
    )
    after = _tree_snapshot(output_dir)
    changed_paths = {
        relative
        for relative in set(before) | set(after)
        if before.get(relative) != after.get(relative)
    }
    assert changed_paths == {"ui_decisions.json"}
    assert list(output_dir.parent.glob(".check-entries-apply-*")) == []


def test_mcp_save_rejects_rehashed_caller_review_payload_without_mutation(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    # Arrange
    _, output_dir, review_payload, run_intake = _supported_assurance_run(
        monkeypatch,
        tmp_path,
    )
    forged_review = json.loads(json.dumps(review_payload))
    forged_review["items"][0]["title"] = "FORGED CALLER REVIEW TITLE"
    _seal_review_payload(forged_review)
    before = _transaction_tree_state(output_dir)

    # Act
    response = _call_mcp_server(
        [
            {
                "jsonrpc": "2.0",
                "id": 1,
                "method": "tools/call",
                "params": {
                    "name": "save_check_entries_decisions",
                    "arguments": _managed_check_mcp_arguments(
                        output_dir,
                        {
                            "run_intake": run_intake,
                            "review_payload": forged_review,
                            "decisions": [
                                {
                                    "item_id": forged_review["items"][0]["id"],
                                    "action": "accept",
                                }
                            ],
                        },
                    ),
                },
            }
        ]
    )[0]["result"]

    # Assert
    assert response["isError"] is True
    assert response["structuredContent"]["error"] == (
        "review_payload does not match the persisted assured review"
    )
    assert _transaction_tree_state(output_dir) == before
    assert not list(output_dir.parent.glob(".generated-review-transaction-*"))
    assert not list(output_dir.parent.glob(".generated-review-commit-*"))
    assert not list(output_dir.parent.glob(".generated-review-recovery-*"))


def test_mcp_apply_rejects_fake_ready_child_without_mutation(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    # Arrange
    _, output_dir, review_payload, run_intake = _supported_assurance_run(
        monkeypatch,
        tmp_path,
    )
    before = _transaction_tree_state(output_dir)
    real_python = Path(sys.executable)
    fake_python = tmp_path / "fake-check-entries-ready-python"
    fake_python.write_text(
        "\n".join(
            [
                "#!/usr/bin/env python3",
                "from __future__ import annotations",
                "import json",
                "import subprocess",
                "import sys",
                "from pathlib import Path",
                f"real = {str(real_python)!r}",
                'if "--client-run-preflight-only" in sys.argv:',
                "    completed = subprocess.run([real, *sys.argv[1:]])",
                "    raise SystemExit(completed.returncode)",
                'output_dir = Path(sys.argv[sys.argv.index("--output-dir") + 1])',
                "envelope = json.loads(",
                '    (output_dir / "assurance_envelope.json").read_text(encoding="utf-8")',
                ")",
                'if "--preflight-only" in sys.argv:',
                "    payload = {",
                '        "ok": True,',
                '        "assurance_replayed": True,',
                '        "report_ready": True,',
                '        "professional_conclusion_status": "reviewed",',
                '        "envelope_content_sha256": envelope["content_sha256"],',
                "    }",
                "else:",
                "    applied_path = Path(",
                '        sys.argv[sys.argv.index("--applied-decisions") + 1]',
                "    )",
                "    final_path = Path(",
                '        sys.argv[sys.argv.index("--final-artifacts") + 1]',
                "    )",
                '    applied = json.loads(applied_path.read_text(encoding="utf-8"))',
                '    final = json.loads(final_path.read_text(encoding="utf-8"))',
                "    payload = {",
                '        "ok": True,',
                '        "updated_effect_count": 0,',
                '        "assurance_replayed": True,',
                '        "application_status": applied["application_status"],',
                '        "native_regenerated_paths": [],',
                '        "backup_paths": [],',
                '        "applied_decisions": applied,',
                '        "final_artifacts": final,',
                "    }",
                'sys.stdout.write(json.dumps(payload) + "\\n")',
                "",
            ]
        ),
        encoding="utf-8",
    )
    fake_python.chmod(0o700)
    monkeypatch.setenv("PYTHON", fake_python.as_posix())
    decisions = [
        {"item_id": item["id"], "action": "accept"} for item in review_payload["items"]
    ]

    # Act
    response = _call_mcp_server(
        [
            {
                "jsonrpc": "2.0",
                "id": 1,
                "method": "tools/call",
                "params": {
                    "name": "apply_check_entries_decisions",
                    "arguments": _managed_check_mcp_arguments(
                        output_dir,
                        {
                            "run_intake": run_intake,
                            "review_payload": review_payload,
                            "decisions": decisions,
                        },
                    ),
                },
            }
        ]
    )[0]["result"]

    # Assert
    assert response["isError"] is True
    assert response["structuredContent"]["error"] == (
        "Check Entries assurance preflight returned an invalid result."
    )
    assert _transaction_tree_state(output_dir) == before
    assert not list(output_dir.parent.glob(".generated-review-transaction-*"))
    assert not list(output_dir.parent.glob(".generated-review-commit-*"))
    assert not list(output_dir.parent.glob(".generated-review-recovery-*"))


def test_mcp_preflight_child_failure_is_bounded_and_discloses_no_paths(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    # Arrange
    _, output_dir, review_payload, run_intake = _supported_assurance_run(
        monkeypatch,
        tmp_path,
    )
    before = _tree_snapshot(output_dir)
    outside = tmp_path / "outside-preflight.txt"
    outside.write_bytes(b"outside unchanged")
    outside_before = outside.read_bytes()
    posix_path = "/Users/private/repository/client/run/secret.csv"
    windows_path = "C:\\Users\\private\\repository\\client\\run\\secret.csv"
    real_python = Path(sys.executable)
    wrapper = tmp_path / "python-malicious-preflight"
    wrapper.write_text(
        "\n".join(
            [
                "#!/usr/bin/env python3",
                "import os",
                "import subprocess",
                "import sys",
                f"real = {str(real_python)!r}",
                'if "--client-run-preflight-only" in sys.argv:',
                "    completed = subprocess.run([real, *sys.argv[1:]])",
                "    raise SystemExit(completed.returncode)",
                'posix_path = os.environ["CE_TEST_POSIX_PATH"]',
                'windows_path = os.environ["CE_TEST_WINDOWS_PATH"]',
                'sys.stdout.write(f"arbitrary child output {posix_path} {windows_path}\\n")',
                'sys.stderr.write("Traceback (most recent call last):\\n")',
                "sys.stderr.write(f'  File \"{posix_path}\", line 9, in child\\n')",
                'sys.stderr.write(f"RuntimeError: terminal failure at {posix_path} and {windows_path}\\n")',
                "raise SystemExit(17)",
                "",
            ]
        ),
        encoding="utf-8",
    )
    wrapper.chmod(0o700)
    monkeypatch.setenv("PYTHON", wrapper.as_posix())
    monkeypatch.setenv("CE_TEST_POSIX_PATH", posix_path)
    monkeypatch.setenv("CE_TEST_WINDOWS_PATH", windows_path)

    # Act
    response = _call_mcp_server(
        _mcp_review_write_message(
            "apply_check_entries_decisions",
            review_payload,
            run_intake,
            output_dir=output_dir,
        )
    )[0]["result"]

    # Assert
    assert response["isError"] is True
    error = response["structuredContent"]["error"]
    assert error == "Check Entries assurance preflight failed."
    assert len(error) <= 240
    assert error.count("failed") == 1
    assert "\n" not in error
    assert "Traceback" not in error
    assert "arbitrary child output" not in error
    assert posix_path not in error
    assert windows_path not in error
    assert tmp_path.as_posix() not in error
    assert outside.read_bytes() == outside_before
    assert _tree_snapshot(output_dir) == before
    assert list(output_dir.parent.glob(".check-entries-apply-*")) == []


def test_mcp_apply_child_failure_is_bounded_and_rolls_back_link_swap(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    # Arrange
    _, output_dir, review_payload, run_intake = _supported_assurance_run(
        monkeypatch,
        tmp_path,
    )
    before = _tree_snapshot(output_dir)
    outside = tmp_path / "outside-apply.json"
    outside.write_bytes(b'{"outside":"unchanged"}\n')
    outside_before = outside.read_bytes()
    posix_path = "/Users/private/repository/client/run/applied_decisions.json"
    windows_path = "D:\\Clients\\private\\run\\applied_decisions.json"
    real_python = Path(sys.executable)
    wrapper = tmp_path / "python-malicious-apply"
    wrapper.write_text(
        "\n".join(
            [
                "#!/usr/bin/env python3",
                "import os",
                "from pathlib import Path",
                "import subprocess",
                "import sys",
                f"real = {str(real_python)!r}",
                'if "--preflight-only" in sys.argv:',
                "    completed = subprocess.run([real, *sys.argv[1:]], capture_output=True)",
                "    sys.stdout.buffer.write(completed.stdout)",
                "    sys.stderr.buffer.write(completed.stderr)",
                "    raise SystemExit(completed.returncode)",
                'out = Path(sys.argv[sys.argv.index("--output-dir") + 1])',
                'outside = Path(os.environ["CE_TEST_OUTSIDE"])',
                'target = out / "applied_decisions.json"',
                "target.unlink()",
                "target.symlink_to(outside)",
                'posix_path = os.environ["CE_TEST_POSIX_PATH"]',
                'windows_path = os.environ["CE_TEST_WINDOWS_PATH"]',
                'sys.stdout.write(f"untrusted stdout {posix_path}\\n")',
                'sys.stderr.write("Traceback (most recent call last):\\n")',
                "sys.stderr.write(f'  File \"{posix_path}\", line 12, in apply\\n')",
                'sys.stderr.write(f"ValueError: failed for {posix_path} and {windows_path}\\n")',
                "raise SystemExit(23)",
                "",
            ]
        ),
        encoding="utf-8",
    )
    wrapper.chmod(0o700)
    monkeypatch.setenv("PYTHON", wrapper.as_posix())
    monkeypatch.setenv("CE_TEST_OUTSIDE", outside.as_posix())
    monkeypatch.setenv("CE_TEST_POSIX_PATH", posix_path)
    monkeypatch.setenv("CE_TEST_WINDOWS_PATH", windows_path)

    # Act
    response = _call_mcp_server(
        _mcp_review_write_message(
            "apply_check_entries_decisions",
            review_payload,
            run_intake,
            output_dir=output_dir,
        )
    )[0]["result"]

    # Assert
    assert response["isError"] is True
    error = response["structuredContent"]["error"]
    assert error == "Check Entries review application failed."
    assert len(error) <= 240
    assert error.count("failed") == 1
    assert "\n" not in error
    assert "Traceback" not in error
    assert "untrusted stdout" not in error
    assert posix_path not in error
    assert windows_path not in error
    assert tmp_path.as_posix() not in error
    assert outside.read_bytes() == outside_before
    assert _tree_snapshot(output_dir) == before
    assert list(output_dir.parent.glob(".check-entries-apply-*")) == []


def test_mcp_preflight_child_start_failure_is_fixed_and_rolls_back(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    # Arrange
    _, output_dir, review_payload, run_intake = _supported_assurance_run(
        monkeypatch,
        tmp_path,
    )
    before = _tree_snapshot(output_dir)
    real_python = Path(sys.executable)
    blocked_executable = tmp_path / "python-not-executable"
    blocked_executable.write_text(
        "\n".join(
            [
                "#!/usr/bin/env python3",
                "from pathlib import Path",
                "import subprocess",
                "import sys",
                f"real = {str(real_python)!r}",
                'if "--client-run-preflight-only" in sys.argv:',
                "    completed = subprocess.run([real, *sys.argv[1:]], capture_output=True)",
                "    sys.stdout.buffer.write(completed.stdout)",
                "    sys.stderr.buffer.write(completed.stderr)",
                "    if completed.returncode == 0:",
                "        Path(__file__).chmod(0o600)",
                "    raise SystemExit(completed.returncode)",
                "raise SystemExit(0)",
                "",
            ]
        ),
        encoding="utf-8",
    )
    blocked_executable.chmod(0o700)
    monkeypatch.setenv("PYTHON", blocked_executable.as_posix())

    # Act
    response = _call_mcp_server(
        _mcp_review_write_message(
            "apply_check_entries_decisions",
            review_payload,
            run_intake,
            output_dir=output_dir,
        )
    )[0]["result"]

    # Assert
    assert response["isError"] is True
    assert (
        response["structuredContent"]["error"]
        == "Check Entries assurance preflight could not start."
    )
    assert blocked_executable.as_posix() not in response["structuredContent"]["error"]
    assert _tree_snapshot(output_dir) == before
    assert list(output_dir.parent.glob(".check-entries-apply-*")) == []


def test_mcp_apply_child_start_failure_is_fixed_and_rolls_back(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    # Arrange
    _, output_dir, review_payload, run_intake = _supported_assurance_run(
        monkeypatch,
        tmp_path,
    )
    before = _tree_snapshot(output_dir)
    real_python = Path(sys.executable)
    wrapper = tmp_path / "python-disable-after-preflight"
    wrapper.write_text(
        "\n".join(
            [
                "#!/usr/bin/env python3",
                "from pathlib import Path",
                "import subprocess",
                "import sys",
                f"real = {str(real_python)!r}",
                "completed = subprocess.run([real, *sys.argv[1:]], capture_output=True)",
                "sys.stdout.buffer.write(completed.stdout)",
                "sys.stderr.buffer.write(completed.stderr)",
                'if completed.returncode == 0 and "--preflight-only" in sys.argv:',
                "    Path(__file__).chmod(0o600)",
                "raise SystemExit(completed.returncode)",
                "",
            ]
        ),
        encoding="utf-8",
    )
    wrapper.chmod(0o700)
    monkeypatch.setenv("PYTHON", wrapper.as_posix())

    # Act
    response = _call_mcp_server(
        _mcp_review_write_message(
            "apply_check_entries_decisions",
            review_payload,
            run_intake,
            output_dir=output_dir,
        )
    )[0]["result"]

    # Assert
    assert response["isError"] is True
    assert (
        response["structuredContent"]["error"]
        == "Check Entries review application could not start."
    )
    assert wrapper.as_posix() not in response["structuredContent"]["error"]
    assert _tree_snapshot(output_dir) == before
    assert list(output_dir.parent.glob(".check-entries-apply-*")) == []


@pytest.mark.parametrize(
    ("phase", "expected_error"),
    [
        (
            "preflight",
            "Check Entries assurance preflight returned an invalid result.",
        ),
        (
            "apply",
            "Check Entries review application returned an invalid result.",
        ),
    ],
)
@pytest.mark.parametrize(
    ("result_kind", "child_output"),
    [
        ("empty", ""),
        ("malformed", "{not-json"),
        ("array", "[]"),
        ("null", "null"),
        (
            "string",
            '"arbitrary child /Users/private/run.json C:\\\\Users\\\\private\\\\run.json"',
        ),
        ("number", "17"),
        ("empty_object", "{}"),
    ],
)
def test_mcp_status_zero_invalid_child_result_fails_closed_and_rolls_back(
    monkeypatch: Any,
    tmp_path: Path,
    phase: str,
    expected_error: str,
    result_kind: str,
    child_output: str,
) -> None:
    # Arrange
    _, output_dir, review_payload, run_intake = _supported_assurance_run(
        monkeypatch,
        tmp_path,
    )
    before = _tree_snapshot(output_dir)
    outside = tmp_path / f"outside-{phase}-{result_kind}.txt"
    outside.write_bytes(b"outside remains unchanged")
    outside_before = outside.read_bytes()
    real_python = Path(sys.executable)
    wrapper = tmp_path / f"python-invalid-{phase}-{result_kind}"
    wrapper.write_text(
        "\n".join(
            [
                "#!/usr/bin/env python3",
                "import os",
                "import subprocess",
                "import sys",
                f"real = {str(real_python)!r}",
                'phase = os.environ["CE_TEST_INVALID_PHASE"]',
                'if "--client-run-preflight-only" in sys.argv:',
                "    completed = subprocess.run([real, *sys.argv[1:]], capture_output=True)",
                "    sys.stdout.buffer.write(completed.stdout)",
                "    sys.stderr.buffer.write(completed.stderr)",
                "    raise SystemExit(completed.returncode)",
                'if phase == "apply" and "--preflight-only" in sys.argv:',
                "    completed = subprocess.run([real, *sys.argv[1:]], capture_output=True)",
                "    sys.stdout.buffer.write(completed.stdout)",
                "    sys.stderr.buffer.write(completed.stderr)",
                "    raise SystemExit(completed.returncode)",
                'payload = os.environ["CE_TEST_INVALID_OUTPUT"]',
                "sys.stdout.write(payload)",
                'if payload and not payload.endswith("\\n"):',
                '    sys.stdout.write("\\n")',
                "raise SystemExit(0)",
                "",
            ]
        ),
        encoding="utf-8",
    )
    wrapper.chmod(0o700)
    monkeypatch.setenv("PYTHON", wrapper.as_posix())
    monkeypatch.setenv("CE_TEST_INVALID_PHASE", phase)
    monkeypatch.setenv("CE_TEST_INVALID_OUTPUT", child_output)

    # Act
    response = _call_mcp_server(
        _mcp_review_write_message(
            "apply_check_entries_decisions",
            review_payload,
            run_intake,
            output_dir=output_dir,
        )
    )[0]["result"]

    # Assert
    assert response["isError"] is True
    error = response["structuredContent"]["error"]
    assert error == expected_error
    assert len(error) <= 240
    assert "\n" not in error
    assert "arbitrary child" not in error
    assert "/Users/private/run.json" not in error
    assert "C:\\Users\\private\\run.json" not in error
    assert outside.read_bytes() == outside_before
    assert _tree_snapshot(output_dir) == before
    assert list(output_dir.parent.glob(".check-entries-apply-*")) == []


def test_mcp_valid_preflight_and_apply_child_objects_remain_accepted(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    # Arrange
    _, output_dir, review_payload, run_intake = _supported_assurance_run(
        monkeypatch,
        tmp_path,
    )
    real_python = Path(sys.executable)
    wrapper = tmp_path / "python-valid-child-proxy"
    wrapper.write_text(
        "\n".join(
            [
                "#!/usr/bin/env python3",
                "import subprocess",
                "import sys",
                f"real = {str(real_python)!r}",
                "completed = subprocess.run([real, *sys.argv[1:]], capture_output=True)",
                "sys.stdout.buffer.write(completed.stdout)",
                "sys.stderr.buffer.write(completed.stderr)",
                "raise SystemExit(completed.returncode)",
                "",
            ]
        ),
        encoding="utf-8",
    )
    wrapper.chmod(0o700)
    monkeypatch.setenv("PYTHON", wrapper.as_posix())

    # Act
    response = _call_mcp_server(
        _mcp_review_write_message(
            "apply_check_entries_decisions",
            review_payload,
            run_intake,
            output_dir=output_dir,
        )
    )[0]["result"]

    # Assert
    assert response["isError"] is False
    applied = response["structuredContent"]["applied_decisions"]
    assert applied["assurance_preflight"]["ok"] is True
    assert applied["assurance_preflight"]["assurance_replayed"] is True
    assert response["structuredContent"]["native_regenerated_count"] == 1
    assert (output_dir / "applied_decisions.json").is_file()
    assert list(output_dir.parent.glob(".check-entries-apply-*")) == []


def test_mcp_forged_shaped_apply_result_cannot_commit_or_pollute_trace(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    # Arrange
    _, output_dir, review_payload, run_intake = _supported_assurance_run(
        monkeypatch,
        tmp_path,
    )
    before = _tree_snapshot(output_dir)
    outside = tmp_path / "outside-forged-result.xlsx"
    outside.write_bytes(b"outside remains unchanged")
    outside_before = outside.read_bytes()
    real_python = Path(sys.executable)
    wrapper = tmp_path / "python-forged-shaped-apply"
    wrapper.write_text(
        "\n".join(
            [
                "#!/usr/bin/env python3",
                "import json",
                "from pathlib import Path",
                "import subprocess",
                "import sys",
                f"real = {str(real_python)!r}",
                'if "--preflight-only" in sys.argv or "--client-run-preflight-only" in sys.argv:',
                "    completed = subprocess.run([real, *sys.argv[1:]], capture_output=True)",
                "    sys.stdout.buffer.write(completed.stdout)",
                "    sys.stderr.buffer.write(completed.stderr)",
                "    raise SystemExit(completed.returncode)",
                'out = Path(sys.argv[sys.argv.index("--output-dir") + 1])',
                'applied = json.loads((out / "applied_decisions.json").read_text())',
                'final = json.loads((out / "final_artifacts.json").read_text())',
                'applied["run_id"] = "forged-run"',
                'applied["native_regeneration_count"] = 0',
                'applied["native_regeneration_paths"] = []',
                'applied["native_regenerated_count"] = 777',
                'applied["native_regenerated_paths"] = ["../../outside.xlsx"]',
                'applied["application_status"] = "final_ready"',
                'final["status"] = "final_ready"',
                'final["review_status"] = "final_ready"',
                'final.setdefault("review_application", {})["original_backup_paths"] = [',
                '    "/private/client-secret/original.xlsx"',
                "]",
                "payload = {",
                '    "ok": True,',
                '    "updated_effect_count": 777,',
                '    "assurance_replayed": True,',
                '    "application_status": "final_ready",',
                '    "native_regenerated_paths": ["../../outside.xlsx"],',
                '    "backup_paths": ["/private/client-secret/original.xlsx"],',
                '    "applied_decisions": applied,',
                '    "final_artifacts": final,',
                "}",
                "sys.stdout.write(json.dumps(payload) + '\\n')",
                "raise SystemExit(0)",
                "",
            ]
        ),
        encoding="utf-8",
    )
    wrapper.chmod(0o700)
    monkeypatch.setenv("PYTHON", wrapper.as_posix())

    # Act
    response = _call_mcp_server(
        _mcp_review_write_message(
            "apply_check_entries_decisions",
            review_payload,
            run_intake,
            output_dir=output_dir,
        )
    )[0]["result"]

    # Assert
    assert response["isError"] is True
    assert (
        response["structuredContent"]["error"]
        == "Check Entries review application returned an invalid result."
    )
    assert _tree_snapshot(output_dir) == before
    assert outside.read_bytes() == outside_before
    serialized = json.dumps(
        json.loads((output_dir / "run_intake.json").read_text(encoding="utf-8"))
    )
    assert "forged-run" not in serialized
    assert "../../outside.xlsx" not in serialized
    assert "/private/client-secret/original.xlsx" not in serialized
    assert list(output_dir.parent.glob(".check-entries-apply-*")) == []


@pytest.mark.parametrize(
    ("field_name", "forged_value"),
    [
        ("backup_paths", [{"bad": "object"}]),
        ("backup_paths", ["/private/client-secret/original.xlsx"]),
        (
            "native_regenerated_paths",
            ["check_results.xlsx", {"bad": "object"}],
        ),
        (
            "native_regenerated_paths",
            [
                "check_results.xlsx",
                "../../outside.xlsx",
                "/private/client-secret/x.xlsx",
            ],
        ),
    ],
)
def test_mcp_ignores_forged_top_level_child_path_metadata(
    monkeypatch: Any,
    tmp_path: Path,
    field_name: str,
    forged_value: list[object],
) -> None:
    # Arrange
    _, output_dir, review_payload, run_intake = _supported_assurance_run(
        monkeypatch,
        tmp_path,
    )
    entry_item = next(
        item
        for item in review_payload["items"]
        if item["item_type"] == "supported_entry"
    )
    item_id = entry_item["id"]
    outside = tmp_path / "outside-child-metadata.xlsx"
    outside.write_bytes(b"outside remains unchanged")
    outside_before = outside.read_bytes()
    real_python = Path(sys.executable)
    wrapper = tmp_path / f"python-forged-{field_name}-{len(forged_value)}"
    wrapper.write_text(
        "\n".join(
            [
                "#!/usr/bin/env python3",
                "import json",
                "import os",
                "import subprocess",
                "import sys",
                f"real = {str(real_python)!r}",
                "completed = subprocess.run([real, *sys.argv[1:]], capture_output=True)",
                "if completed.returncode != 0:",
                "    sys.stdout.buffer.write(completed.stdout)",
                "    sys.stderr.buffer.write(completed.stderr)",
                "    raise SystemExit(completed.returncode)",
                'if "--preflight-only" in sys.argv:',
                "    sys.stdout.buffer.write(completed.stdout)",
                "    raise SystemExit(0)",
                "lines = [line for line in completed.stdout.decode().splitlines() if line]",
                "payload = json.loads(lines[-1])",
                'payload[os.environ["CE_TEST_FORGED_FIELD"]] = json.loads(',
                '    os.environ["CE_TEST_FORGED_VALUE"]',
                ")",
                "sys.stdout.write(json.dumps(payload) + '\\n')",
                "raise SystemExit(0)",
                "",
            ]
        ),
        encoding="utf-8",
    )
    wrapper.chmod(0o700)
    monkeypatch.setenv("PYTHON", wrapper.as_posix())
    monkeypatch.setenv("CE_TEST_FORGED_FIELD", field_name)
    monkeypatch.setenv(
        "CE_TEST_FORGED_VALUE",
        json.dumps([*forged_value, outside.as_posix()]),
    )

    # Act
    response = _call_mcp_server(
        _mcp_review_write_message(
            "apply_check_entries_decisions",
            review_payload,
            run_intake,
            output_dir=output_dir,
        )
    )[0]["result"]

    # Assert
    assert response.get("isError") is not True
    assert response["structuredContent"]["ok"] is True
    serialized = json.dumps(response["structuredContent"])
    assert "../../outside.xlsx" not in serialized
    assert "/private/client-secret" not in serialized
    assert outside.as_posix() not in serialized
    assert field_name not in response["structuredContent"]
    applied = json.loads(
        (output_dir / "applied_decisions.json").read_text(encoding="utf-8")
    )
    expected_revision = f"revisions/check_results__{item_id}.txt"
    expected_csv_backup = f"revisions/originals/check_results__{item_id}.csv"
    expected_xlsx_backup = f"revisions/originals/check_results__{item_id}.xlsx"
    assert applied["native_regeneration_paths"] == []
    assert applied["native_regenerated_paths"] == ["check_results.xlsx"]
    assert applied["original_backup_paths"] == [
        expected_csv_backup,
        expected_xlsx_backup,
    ]
    assert len(applied["effects"]) == 1
    effect = applied["effects"][0]
    assert effect["item_id"] == item_id
    assert effect["target_artifact"] == "check_results.csv"
    assert effect["target_id_field"] == "prepared_entry_id"
    assert effect["target_record_id"] == entry_item["data"]["target_record_id"]
    assert effect["target_field"] == "review_notes"
    assert effect["revision_artifact"] == expected_revision
    assert effect["original_artifact_backup"] == expected_csv_backup
    assert effect["derived_native_regeneration_paths"] == ["check_results.xlsx"]
    assert effect["native_regenerated_paths"] == ["check_results.xlsx"]
    assert effect["requires_native_regeneration"] is False
    assert effect["native_regeneration_status"] == "regenerated"
    final_artifacts = json.loads(
        (output_dir / "final_artifacts.json").read_text(encoding="utf-8")
    )
    review_application = final_artifacts["review_application"]
    assert review_application["revision_paths"] == [expected_revision]
    assert review_application["target_update_paths"] == ["check_results.csv"]
    assert review_application["structured_update_paths"] == ["check_results.csv"]
    assert review_application["native_regeneration_paths"] == []
    assert review_application["native_regenerated_paths"] == ["check_results.xlsx"]
    assert review_application["original_backup_paths"] == [
        expected_csv_backup,
        expected_xlsx_backup,
    ]
    assert outside.read_bytes() == outside_before
    assert not list(output_dir.parent.glob(".check-entries-apply-*"))
    assert not list(output_dir.parent.glob(".generated-review-transaction-*"))


@pytest.mark.parametrize(
    "mutation_kind",
    [
        "native_count",
        "windows_final_output",
        "effect_escape_path",
        "effect_reset_to_pending",
        "assurance_replayed_false",
        "padded_final_output",
        "control_final_output",
        "effect_top_path_mismatch",
        "effect_original_backup",
        "effect_derived_paths",
        "blocker_count",
        "extra_missing_final_output",
        "extra_existing_final_output",
        "final_native_flag_false",
        "applied_hidden_backup",
        "applied_review_path",
        "application_status_final_ready",
        "applied_decision_mutation",
    ],
)
def test_mcp_contradictory_persisted_child_state_fails_closed(
    monkeypatch: Any,
    tmp_path: Path,
    mutation_kind: str,
) -> None:
    # Arrange
    _, output_dir, review_payload, run_intake = _supported_assurance_run(
        monkeypatch,
        tmp_path,
    )
    before = _tree_snapshot(output_dir)
    real_python = Path(sys.executable)
    wrapper = tmp_path / f"python-contradictory-{mutation_kind}"
    wrapper.write_text(
        "\n".join(
            [
                "#!/usr/bin/env python3",
                "import json",
                "import os",
                "from pathlib import Path",
                "import subprocess",
                "import sys",
                f"real = {str(real_python)!r}",
                "completed = subprocess.run([real, *sys.argv[1:]], capture_output=True)",
                "if completed.returncode != 0:",
                "    sys.stdout.buffer.write(completed.stdout)",
                "    sys.stderr.buffer.write(completed.stderr)",
                "    raise SystemExit(completed.returncode)",
                'if "--preflight-only" in sys.argv or "--client-run-preflight-only" in sys.argv:',
                "    sys.stdout.buffer.write(completed.stdout)",
                "    raise SystemExit(0)",
                'out = Path(sys.argv[sys.argv.index("--output-dir") + 1])',
                "payload = json.loads(completed.stdout.decode().splitlines()[-1])",
                'applied_path = out / "applied_decisions.json"',
                'final_path = out / "final_artifacts.json"',
                "applied = json.loads(applied_path.read_text())",
                "final = json.loads(final_path.read_text())",
                'kind = os.environ["CE_TEST_CONTRADICTION"]',
                'if kind == "native_count":',
                '    applied["native_regenerated_count"] = 777',
                '    payload["updated_effect_count"] = 777',
                'elif kind == "windows_final_output":',
                '    final.setdefault("outputs", []).append({"path": "C:/Users/private/x.xlsx"})',
                'elif kind == "effect_escape_path":',
                '    applied["effects"][0].setdefault("native_regenerated_paths", []).append(',
                '        "../../outside.xlsx"',
                "    )",
                'elif kind == "effect_reset_to_pending":',
                '    applied["effects"][0]["native_regeneration_status"] = "pending"',
                '    applied["effects"][0]["requires_native_regeneration"] = True',
                'elif kind == "assurance_replayed_false":',
                '    applied["assurance_replayed"] = False',
                'elif kind == "padded_final_output":',
                '    final.setdefault("outputs", []).append({"path": " check_results.xlsx "})',
                'elif kind == "control_final_output":',
                '    final.setdefault("outputs", []).append({"path": "bad\\\\npath.xlsx"})',
                'elif kind == "effect_top_path_mismatch":',
                '    applied["effects"][0]["native_regenerated_paths"] = ["other.xlsx"]',
                'elif kind == "effect_original_backup":',
                '    applied["effects"][0]["original_artifact_backup"] = (',
                '        "/private/client-secret/original.csv"',
                "    )",
                'elif kind == "effect_derived_paths":',
                '    applied["effects"][0]["derived_native_regeneration_paths"] = [',
                '        "../../outside.xlsx"',
                "    ]",
                'elif kind == "blocker_count":',
                '    applied["blocker_count"] = 99',
                '    final["review_application"]["blocker_count"] = 99',
                'elif kind == "extra_missing_final_output":',
                '    final.setdefault("outputs", []).append(',
                '        {"path": "forged-nonexistent.txt", "status": "written"}',
                "    )",
                'elif kind == "extra_existing_final_output":',
                '    (out / "forged-existing.txt").write_text("forged")',
                '    final.setdefault("outputs", []).append(',
                '        {"path": "forged-existing.txt", "status": "written"}',
                "    )",
                'elif kind == "final_native_flag_false":',
                '    output = next(item for item in final["outputs"] if item.get("path") == "check_results.xlsx")',
                '    output["native_regenerated"] = False',
                'elif kind == "applied_hidden_backup":',
                '    applied.setdefault("original_backup_paths", []).append(',
                '        "/private/client-secret/hidden.xlsx"',
                "    )",
                'elif kind == "applied_review_path":',
                '    applied["review_payload"]["path"] = "/private/client-secret/review_payload.json"',
                'elif kind == "application_status_final_ready":',
                '    applied["application_status"] = "final_ready"',
                '    final["status"] = "final_ready"',
                '    final["review_status"] = "final_ready"',
                '    payload["application_status"] = "final_ready"',
                "else:",
                '    applied["decisions"][0]["action"] = "accept"',
                "applied_path.write_text(json.dumps(applied, indent=2) + '\\n')",
                "final_path.write_text(json.dumps(final, indent=2) + '\\n')",
                'payload["applied_decisions"] = applied',
                'payload["final_artifacts"] = final',
                "sys.stdout.write(json.dumps(payload) + '\\n')",
                "raise SystemExit(0)",
                "",
            ]
        ),
        encoding="utf-8",
    )
    wrapper.chmod(0o700)
    monkeypatch.setenv("PYTHON", wrapper.as_posix())
    monkeypatch.setenv("CE_TEST_CONTRADICTION", mutation_kind)

    # Act
    response = _call_mcp_server(
        _mcp_review_write_message(
            "apply_check_entries_decisions",
            review_payload,
            run_intake,
            output_dir=output_dir,
        )
    )[0]["result"]

    # Assert
    assert response["isError"] is True
    assert (
        response["structuredContent"]["error"]
        == "Check Entries review application returned an invalid result."
    )
    assert _tree_snapshot(output_dir) == before
    serialized = json.dumps(
        json.loads((output_dir / "run_intake.json").read_text(encoding="utf-8"))
    )
    assert "C:/Users/private" not in serialized
    assert "../../outside.xlsx" not in serialized
    assert "/private/client-secret" not in serialized
    assert list(output_dir.parent.glob(".check-entries-apply-*")) == []


def test_mcp_assured_apply_commits_without_staging_paths(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    # Arrange
    _, output_dir, review_payload, run_intake = _supported_assurance_run(
        monkeypatch,
        tmp_path,
    )
    item = next(
        entry
        for entry in review_payload["items"]
        if entry["item_type"] == "supported_entry"
    )

    # Act
    response = _call_mcp_server(
        [
            {
                "jsonrpc": "2.0",
                "id": 1,
                "method": "tools/call",
                "params": {
                    "name": "apply_check_entries_decisions",
                    "arguments": _managed_check_mcp_arguments(
                        output_dir,
                        {
                            "run_intake": run_intake,
                            "review_payload": review_payload,
                            "decisions": [
                                {
                                    "item_id": item["id"],
                                    "action": "edit",
                                    "edit_value": "Reviewer confirmed exact support.",
                                }
                            ],
                        },
                    ),
                },
            }
        ]
    )[0]["result"]

    # Assert
    assert response["isError"] is False
    audit = json.loads((output_dir / "check_audit.json").read_text())
    assert (
        audit["assurance_envelope"]["path"]
        == (output_dir / "assurance_envelope.json")
        .relative_to(_customer_context_path(output_dir).parent)
        .as_posix()
    )
    assert ".check-entries-apply-" not in json.dumps(audit)
    assert list(output_dir.parent.glob(".check-entries-apply-*")) == []
    preflight = load_apply_review_edits().preflight_assurance(output_dir)
    assert preflight["ok"] is True
    assert preflight["assurance_replayed"] is True


def test_spanish_mcp_runtime_feedback_handoff_and_errors(tmp_path: Path) -> None:
    output_dir, managed_arguments = _portable_check_transaction_case(tmp_path)
    review_payload = managed_arguments["review_payload"]
    review_payload["language"] = "es-ES"
    _seal_review_payload(review_payload)
    (output_dir / "review_payload.json").write_text(
        json.dumps(review_payload, indent=2) + "\n",
        encoding="utf-8",
    )
    run_intake = managed_arguments["run_intake"]
    run_intake["language"] = "es"
    (output_dir / "run_intake.json").write_text(
        json.dumps(run_intake, indent=2) + "\n",
        encoding="utf-8",
    )
    decision = {"item_id": review_payload["items"][0]["id"], "action": "accept"}
    invalid_payload = _seal_review_payload({**review_payload, "items": "invalid"})
    messages = [
        {
            "jsonrpc": "2.0",
            "id": 1,
            "method": "initialize",
            "params": {"_meta": {"locale": "es-ES"}},
        },
        {
            "jsonrpc": "2.0",
            "id": 2,
            "method": "tools/call",
            "params": {
                "name": "validate_check_entries_review",
                "arguments": {"review_payload": review_payload},
            },
        },
        {
            "jsonrpc": "2.0",
            "id": 3,
            "method": "tools/call",
            "params": {
                "name": "save_check_entries_decisions",
                "arguments": {
                    "review_payload": review_payload,
                    "decisions": [decision],
                },
            },
        },
        {
            "jsonrpc": "2.0",
            "id": 4,
            "method": "tools/call",
            "params": {
                "name": "apply_check_entries_decisions",
                "arguments": _managed_check_mcp_arguments(
                    output_dir,
                    {
                        "run_intake": run_intake,
                        "review_payload": review_payload,
                        "decisions": [decision],
                    },
                ),
            },
        },
        {
            "jsonrpc": "2.0",
            "id": 5,
            "method": "tools/call",
            "params": {
                "name": "validate_check_entries_review",
                "arguments": {"review_payload": invalid_payload},
            },
        },
    ]

    responses = {response["id"]: response for response in _call_mcp_server(messages)}
    validation = responses[2]["result"]["structuredContent"]
    saved = responses[3]["result"]["structuredContent"]
    applied = responses[4]["result"]["structuredContent"]
    invalid = responses[5]["result"]["structuredContent"]
    handoff = (output_dir / "review_handoff.md").read_text(encoding="utf-8")

    assert (
        "Use review_payload_path con validate_check_entries_review"
        in responses[1]["result"]["instructions"]
    )
    assert validation["message"].startswith("Los datos de revisión")
    assert saved["message"].startswith("Las decisiones son válidas")
    assert applied["message"].startswith("Se aplicaron 1 decisiones")
    assert applied["application_status"] == "blocked"
    assert not any(
        action.startswith("Use final_artifacts.json como galería")
        for action in applied["final_artifacts"]["next_actions"]
    )
    handoff_output = next(
        output
        for output in applied["final_artifacts"]["outputs"]
        if output["path"] == "review_handoff.md"
    )
    assert handoff.startswith("# Entrega para revisión: Comprobación de asientos\n")
    assert "## Revisión en Codex" in handoff
    assert "<!-- Review Handoff -->" in handoff
    assert handoff_output["required_text"][:2] == [
        "Entrega para revisión",
        "Review Handoff",
    ]
    assert invalid["error"] == "review_payload.items debe ser una matriz"


@pytest.mark.parametrize(
    "artifact", ["__pycache__/ambient.pyc", "ambient.pyc", "ambient.pyo"]
)
def test_mcp_initializes_with_incidental_bytecode(
    tmp_path: Path, artifact: str
) -> None:
    plugin, shared = _copy_check_entries_runtime(tmp_path / "runtime")
    cache = shared / artifact
    cache.parent.mkdir(parents=True, exist_ok=True)
    cache.write_bytes(b"inert generated cache")

    responses = _call_mcp_server(
        [
            {"jsonrpc": "2.0", "id": 1, "method": "initialize", "params": {}},
            {"jsonrpc": "2.0", "id": 2, "method": "tools/list", "params": {}},
        ],
        server_path=plugin / "mcp/server.cjs",
    )

    assert responses[0]["result"]["serverInfo"]["name"] == "check-entries-widgets"
    assert responses[1]["result"]["tools"]
    assert cache.read_bytes() == b"inert generated cache"


@pytest.mark.parametrize(
    "artifact", ["__pycache__/ambient.pyc", "ambient.pyc", "ambient.pyo"]
)
def test_check_entries_accepts_upstream_incidental_bytecode(
    monkeypatch: Any,
    tmp_path: Path,
    artifact: str,
) -> None:
    normalized = _qualified_journal(
        tmp_path / "source",
        [
            {
                "date": "2025-01-02",
                "movement": "M-1001",
                "description": "Invoice payment",
                "debit": "123.45",
            }
        ],
    )
    plugin = tmp_path / "runtime/journal-sampling"
    shared = tmp_path / "runtime/_shared/vendor/modules/vera_assurance"
    ignore = shutil.ignore_patterns("__pycache__", "*.pyc", "*.pyo")
    shutil.copytree(ROOT / "plugins/journal-sampling", plugin, ignore=ignore)
    shutil.copytree(
        ROOT / "plugins/_shared/vendor/modules/vera_assurance", shared, ignore=ignore
    )
    cache = plugin / "scripts" / artifact
    cache.parent.mkdir(parents=True, exist_ok=True)
    cache.write_bytes(b"inert generated cache")
    core = load_core()
    monkeypatch.setattr(core, "_journal_sampling_component_root", lambda: plugin)
    monkeypatch.setattr(
        core,
        "implementation_artifact_roots",
        lambda: {
            "implementation": ROOT / "plugins/check-entries",
            "assurance_implementation": shared,
        },
    )
    support = tmp_path / "support"
    support.mkdir()
    output = tmp_path / "checks"

    core.run_entry_checks(normalized, support, output)

    assert (output / "check_audit.json").is_file()
    assert cache.read_bytes() == b"inert generated cache"
