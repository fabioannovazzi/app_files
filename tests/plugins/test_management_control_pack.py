from __future__ import annotations

import importlib.util
import json
import re
import subprocess
import sys
from pathlib import Path
from types import ModuleType

import pytest
from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[2]
PLUGIN_ROOT = ROOT / "plugins" / "management-control-pack"
SCRIPTS = PLUGIN_ROOT / "scripts"
SHARED_MODULES = ROOT / "plugins" / "_shared" / "vendor" / "modules"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))
if str(SHARED_MODULES) not in sys.path:
    sys.path.insert(0, str(SHARED_MODULES))

from management_control_core import (  # noqa: E402
    COMMENTARY_SCHEMA,
    PackContractError,
    build_inspection,
    build_management_pack,
    build_model_context,
    finalize_commentary,
    load_source_tables,
    write_excel,
)


def _write_workbook(path: Path, *, full: bool = True) -> None:
    workbook = Workbook()
    ledger = workbook.active
    ledger.title = "GL"
    ledger.append(("Date", "Account", "Category", "Amount"))
    ledger_rows = (
        ("2025-01-31", "4000", "Revenue", 1000),
        ("2025-01-31", "5000", "COGS", -400),
        ("2025-01-31", "6000", "Opex", -200),
        ("2025-02-28", "4000", "Revenue", 1400),
        ("2025-02-28", "5000", "COGS", -500),
        ("2025-02-28", "6000", "Opex", -300),
    )
    for row in ledger_rows:
        ledger.append(row)
    if not full:
        workbook.save(path)
        return

    budget = workbook.create_sheet("Budget")
    budget.append(("Date", "Category", "Amount"))
    for row in (
        ("2025-01-31", "Revenue", 900),
        ("2025-01-31", "COGS", -350),
        ("2025-01-31", "Opex", -220),
        ("2025-02-28", "Revenue", 1300),
        ("2025-02-28", "COGS", -480),
        ("2025-02-28", "Opex", -280),
    ):
        budget.append(row)

    receivables = workbook.create_sheet("AR")
    receivables.append(("Customer", "Due", "Outstanding"))
    receivables.append(("Customer A", "2025-01-15", 400))
    receivables.append(("Customer B", "2025-03-15", 600))

    payables = workbook.create_sheet("AP")
    payables.append(("Supplier", "Due", "Outstanding"))
    payables.append(("Supplier X", "2025-01-01", 300))

    bank = workbook.create_sheet("Bank")
    bank.append(("Date", "Account", "Amount", "Balance"))
    bank.append(("2025-01-10", "Main", 2000, 2000))
    bank.append(("2025-01-20", "Main", -800, 1200))
    bank.append(("2025-02-10", "Main", 1000, 2200))
    bank.append(("2025-02-20", "Main", -500, 1700))

    sales = workbook.create_sheet("Sales")
    sales.append(("Date", "Customer", "Service", "Revenue", "Direct cost"))
    sales.append(("2025-01-10", "Customer A", "Consulting", 800, 200))
    sales.append(("2025-01-20", "Customer B", "Product", 200, 120))
    sales.append(("2025-02-10", "Customer A", "Consulting", 1000, 250))
    sales.append(("2025-02-20", "Customer C", "Product", 400, 200))
    workbook.save(path)


def _reviewed_recipe(
    inspection: dict[str, object], *, full: bool = True
) -> dict[str, object]:
    table_ids = {
        str(item["table_label"]): str(item["table_id"])
        for item in inspection["tables"]  # type: ignore[index]
    }
    tables: dict[str, object] = {
        "general_ledger": {
            "table_id": table_ids["GL"],
            "columns": {
                "date": "Date",
                "account_code": "Account",
                "category": "Category",
                "amount": "Amount",
            },
        }
    }
    controls: dict[str, str] = {"general_ledger": "1000"}
    if full:
        tables.update(
            {
                "budget": {
                    "table_id": table_ids["Budget"],
                    "columns": {
                        "date": "Date",
                        "category": "Category",
                        "amount": "Amount",
                    },
                },
                "receivables": {
                    "table_id": table_ids["AR"],
                    "columns": {
                        "customer_name": "Customer",
                        "due_date": "Due",
                        "outstanding_amount": "Outstanding",
                    },
                },
                "payables": {
                    "table_id": table_ids["AP"],
                    "columns": {
                        "supplier_name": "Supplier",
                        "due_date": "Due",
                        "outstanding_amount": "Outstanding",
                    },
                },
                "bank": {
                    "table_id": table_ids["Bank"],
                    "columns": {
                        "date": "Date",
                        "account": "Account",
                        "amount": "Amount",
                        "balance": "Balance",
                    },
                },
                "sales_lines": {
                    "table_id": table_ids["Sales"],
                    "columns": {
                        "date": "Date",
                        "customer_name": "Customer",
                        "service": "Service",
                        "revenue": "Revenue",
                        "direct_cost": "Direct cost",
                    },
                },
            }
        )
        controls.update(
            {
                "budget": "870",
                "receivables": "1000",
                "payables": "300",
                "bank": "1700",
                "sales_lines": "2400",
            }
        )
    return {
        "schema_version": "vera.management_control_recipe.v1",
        "workflow_id": "management-control-pack",
        "inventory_sha256": inspection["inventory_sha256"],
        "entity": "Synthetic S.r.l.",
        "reporting_period": {
            "start": "2025-01-01",
            "end": "2025-02-28",
            "cutoff": "2025-02-28",
        },
        "currency": "EUR",
        "fiscal_year_start_month": 1,
        "number_format": "dot_decimal",
        "date_format": "%Y-%m-%d",
        "tables": tables,
        "category_roles": {
            "Revenue": "revenue",
            "COGS": "cogs",
            "Opex": "operating_expense",
        },
        "category_multipliers": {},
        "aging_buckets": [30, 60, 90],
        "top_customers": 10,
        "control_totals": controls,
        "control_tolerance": "0.01",
        "mapping_review": {
            "status": "reviewed",
            "reviewer": "Test reviewer",
            "reviewed_at": "2026-08-25T12:00:00+02:00",
        },
    }


def _inspection_for(path: Path) -> tuple[list[object], dict[str, object]]:
    tables = load_source_tables([path])
    inspection, _, _ = build_inspection(tables)
    return tables, inspection


def _load_client_ledger() -> ModuleType:
    path = ROOT / "plugins" / "studio-archive" / "scripts" / "client_ledger.py"
    spec = importlib.util.spec_from_file_location(
        "test_management_pack_client_ledger", path
    )
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def test_inspection_never_assigns_semantic_table_roles(tmp_path: Path) -> None:
    source = tmp_path / "management.xlsx"
    _write_workbook(source)

    _, inspection = _inspection_for(source)

    assert inspection["mapping_policy"] == "semantic_roles_require_review"
    assert "suggested_role" not in json.dumps(inspection)
    assert inspection["table_count"] == 6
    assert all(len(item["preview_rows"]) <= 10 for item in inspection["tables"])


def test_full_pack_calculates_exact_supported_sections(tmp_path: Path) -> None:
    source = tmp_path / "management.xlsx"
    _write_workbook(source)
    tables, inspection = _inspection_for(source)

    pack = build_management_pack(tables, _reviewed_recipe(inspection))

    assert pack["status"] == "ready_for_review"
    assert pack["metrics"]["pnl.total.revenue"]["value"] == "2400"
    assert pack["metrics"]["pnl.total.ebitda"]["value"] == "1000"
    assert pack["metrics"]["budget.total.ebitda_variance"]["value"] == "130"
    assert pack["metrics"]["ar.total.overdue"]["value"] == "400"
    assert pack["metrics"]["cash.latest.reported_balance"]["value"] == "1700"
    assert pack["metrics"]["customers.top1.share"]["value"] == "0.75"
    assert pack["metrics"]["services.total.margin"]["value"] == "1630"
    assert pack["sections"]["customer_concentration"]["rows"][1]["share"] == (
        "0.166667"
    )
    assert (
        pack["sections"]["service_profitability"]["rows"][1]["margin_rate"]
        == "0.466667"
    )
    assert {control["status"] for control in pack["controls"]} == {"passed"}


def test_cash_uses_reporting_window_and_balance_cutoff(tmp_path: Path) -> None:
    source = tmp_path / "management.xlsx"
    _write_workbook(source)
    workbook = load_workbook(source)
    bank = workbook["Bank"]
    bank.append(("2024-12-31", "Main", 9000, 9000))
    bank.append(("2025-03-31", "Main", 7000, 16000))
    workbook.save(source)
    workbook.close()
    tables, inspection = _inspection_for(source)

    pack = build_management_pack(tables, _reviewed_recipe(inspection))

    assert pack["status"] == "ready_for_review"
    assert [row["period"] for row in pack["sections"]["cash_movement"]["rows"]] == [
        "2025-01",
        "2025-02",
    ]
    assert pack["metrics"]["cash.total.net_movement"]["value"] == "1700"
    assert pack["metrics"]["cash.latest.reported_balance"]["value"] == "1700"


def test_budget_without_reporting_period_rows_is_unavailable(tmp_path: Path) -> None:
    source = tmp_path / "management.xlsx"
    _write_workbook(source)
    workbook = load_workbook(source)
    budget = workbook["Budget"]
    budget["A2"] = "2024-01-31"
    budget["A3"] = "2024-01-31"
    budget["A4"] = "2024-01-31"
    budget["A5"] = "2024-02-29"
    budget["A6"] = "2024-02-29"
    budget["A7"] = "2024-02-29"
    workbook.save(source)
    workbook.close()
    tables, inspection = _inspection_for(source)

    pack = build_management_pack(tables, _reviewed_recipe(inspection))

    assert pack["status"] == "partial"
    assert pack["sections"]["budget_variance"] == {
        "status": "unavailable",
        "reason": "Budget has no rows inside the reporting period.",
    }
    assert "budget.total.ebitda_variance" not in pack["metrics"]


def test_reviewed_sign_multipliers_normalize_bank_and_sales(tmp_path: Path) -> None:
    source = tmp_path / "management.xlsx"
    _write_workbook(source)
    workbook = load_workbook(source)
    bank = workbook["Bank"]
    bank["C2"], bank["D2"] = -2000, -2000
    bank["C3"], bank["D3"] = 800, -1200
    bank["C4"], bank["D4"] = -1000, -2200
    bank["C5"], bank["D5"] = 500, -1700
    sales = workbook["Sales"]
    sales["D2"], sales["E2"] = -800, -200
    sales["D3"], sales["E3"] = -200, -120
    sales["D4"], sales["E4"] = -1000, -250
    sales["D5"], sales["E5"] = -400, -200
    workbook.save(source)
    workbook.close()
    tables, inspection = _inspection_for(source)
    recipe = _reviewed_recipe(inspection)
    recipe["tables"]["bank"]["amount_multiplier"] = "-1"  # type: ignore[index]
    recipe["tables"]["bank"]["balance_multiplier"] = "-1"  # type: ignore[index]
    recipe["tables"]["sales_lines"]["revenue_multiplier"] = "-1"  # type: ignore[index]
    recipe["tables"]["sales_lines"]["direct_cost_multiplier"] = "-1"  # type: ignore[index]

    pack = build_management_pack(tables, recipe)

    assert pack["status"] == "ready_for_review"
    assert pack["metrics"]["cash.total.net_movement"]["value"] == "1700"
    assert pack["metrics"]["cash.latest.reported_balance"]["value"] == "1700"
    assert pack["metrics"]["services.total.margin"]["value"] == "1630"
    assert {control["status"] for control in pack["controls"]} == {"passed"}


def test_sales_without_reporting_period_rows_is_unavailable(tmp_path: Path) -> None:
    source = tmp_path / "management.xlsx"
    _write_workbook(source)
    workbook = load_workbook(source)
    sales = workbook["Sales"]
    sales["A2"] = "2024-01-10"
    sales["A3"] = "2024-01-20"
    sales["A4"] = "2024-02-10"
    sales["A5"] = "2024-02-20"
    workbook.save(source)
    workbook.close()
    tables, inspection = _inspection_for(source)

    pack = build_management_pack(tables, _reviewed_recipe(inspection))

    assert pack["status"] == "partial"
    assert pack["sections"]["customer_concentration"]["status"] == "unavailable"
    assert pack["sections"]["service_profitability"]["status"] == "unavailable"
    assert "customers.top1.share" not in pack["metrics"]


def test_missing_optional_exports_remain_visible_and_partial(tmp_path: Path) -> None:
    source = tmp_path / "ledger.xlsx"
    _write_workbook(source, full=False)
    tables, inspection = _inspection_for(source)

    pack = build_management_pack(tables, _reviewed_recipe(inspection, full=False))

    assert pack["status"] == "partial"
    unavailable = {
        item["section"] for item in pack["coverage"] if item["status"] == "unavailable"
    }
    assert unavailable == {
        "budget_variance",
        "receivables_aging",
        "payables_aging",
        "cash_movement",
        "customer_concentration",
        "service_profitability",
    }


def test_failed_declared_control_blocks_the_pack(tmp_path: Path) -> None:
    source = tmp_path / "management.xlsx"
    _write_workbook(source)
    tables, inspection = _inspection_for(source)
    recipe = _reviewed_recipe(inspection)
    recipe["control_totals"]["general_ledger"] = "999"  # type: ignore[index]

    pack = build_management_pack(tables, recipe)

    assert pack["status"] == "blocked"
    assert any(
        control["role"] == "general_ledger" and control["status"] == "failed"
        for control in pack["controls"]
    )


def test_post_calculation_context_excludes_raw_files_and_absolute_paths(
    tmp_path: Path,
) -> None:
    source = tmp_path / "confidential-management.xlsx"
    _write_workbook(source)
    tables, inspection = _inspection_for(source)
    pack = build_management_pack(tables, _reviewed_recipe(inspection))

    serialized = json.dumps(build_model_context(pack), ensure_ascii=False)

    assert source.name not in serialized
    assert str(tmp_path) not in serialized
    assert '"source_id": "source_001"' in serialized
    assert '"sha256"' in serialized


def test_pack_rejects_unreviewed_mapping(tmp_path: Path) -> None:
    source = tmp_path / "management.xlsx"
    _write_workbook(source)
    tables, inspection = _inspection_for(source)
    recipe = _reviewed_recipe(inspection)
    recipe["mapping_review"]["status"] = "not_reviewed"  # type: ignore[index]

    with pytest.raises(PackContractError, match="explicitly reviewed"):
        build_management_pack(tables, recipe)


def test_commentary_closes_only_existing_metric_references(tmp_path: Path) -> None:
    source = tmp_path / "management.xlsx"
    _write_workbook(source)
    tables, inspection = _inspection_for(source)
    pack = build_management_pack(tables, _reviewed_recipe(inspection))
    pack_bytes = (
        json.dumps(pack, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        + "\n"
    ).encode()
    commentary = {
        "schema_version": COMMENTARY_SCHEMA,
        "workflow_id": "management-control-pack",
        "pack_sha256": __import__("hashlib").sha256(pack_bytes).hexdigest(),
        "observations": [
            {
                "text": "Revenue for the reviewed period is 2,400 EUR.",
                "metric_ids": ["pnl.total.revenue"],
            }
        ],
        "hypotheses": [
            {
                "text": "Customer concentration may warrant a contract review.",
                "metric_ids": ["customers.top1.share"],
            }
        ],
        "questions": [{"text": "Is the concentration intentional?", "metric_ids": []}],
        "limitations": [],
    }

    reviewed = finalize_commentary(pack, commentary)
    assert reviewed["status"] == "draft_pending_professional_review"
    commentary["observations"][0]["metric_ids"] = ["missing.metric"]
    with pytest.raises(PackContractError, match="unknown metric"):
        finalize_commentary(pack, commentary)


def test_commentary_finalizer_rejects_blocked_pack(tmp_path: Path) -> None:
    source = tmp_path / "management.xlsx"
    _write_workbook(source)
    tables, inspection = _inspection_for(source)
    recipe = _reviewed_recipe(inspection)
    recipe["control_totals"]["general_ledger"] = "999"  # type: ignore[index]
    pack = build_management_pack(tables, recipe)
    pack_bytes = (
        json.dumps(pack, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        + "\n"
    ).encode()
    commentary = {
        "schema_version": COMMENTARY_SCHEMA,
        "workflow_id": "management-control-pack",
        "pack_sha256": __import__("hashlib").sha256(pack_bytes).hexdigest(),
        "observations": [],
        "hypotheses": [],
        "questions": [],
        "limitations": [],
    }

    with pytest.raises(PackContractError, match="cannot be finalized"):
        finalize_commentary(pack, commentary)


def test_excel_writes_untrusted_labels_as_literal_text(tmp_path: Path) -> None:
    source = tmp_path / "management.xlsx"
    _write_workbook(source)
    tables, inspection = _inspection_for(source)
    pack = build_management_pack(tables, _reviewed_recipe(inspection))
    pack["entity"] = "=1+1"
    pack["sections"]["customer_concentration"]["rows"][0]["customer"] = "@SUM(1,1)"
    pack["sections"]["service_profitability"]["rows"][0]["service"] = "+cmd"
    output = tmp_path / "safe.xlsx"

    write_excel(output, pack)

    workbook = load_workbook(output, data_only=False)
    try:
        assert workbook["Summary"]["B1"].data_type == "s"
        assert workbook["Summary"]["B1"].value == "'=1+1"
        assert workbook["Customers"]["A2"].data_type == "s"
        assert workbook["Customers"]["A2"].value == "'@SUM(1,1)"
        assert workbook["Services"]["A2"].data_type == "s"
        assert workbook["Services"]["A2"].value == "'+cmd"
    finally:
        workbook.close()


def test_connectorless_cli_runs_end_to_end_inside_studio_archive(
    tmp_path: Path,
) -> None:
    source = tmp_path / "management.xlsx"
    _write_workbook(source)
    ledger = _load_client_ledger()
    client_root = tmp_path / "Client"
    client_root.mkdir()
    client_id = "client_111111111111111111111111"
    ledger.create_client_manifest(client_root, client_id)
    engagement = ledger.create_engagement(
        client_root, client_id, "Management reporting"
    )
    imported = ledger.import_document(
        client_root,
        client_id,
        engagement["engagement_id"],
        source,
        "source",
    )
    prepared = ledger.prepare_run(
        client_root,
        client_id,
        engagement["engagement_id"],
        "management-control-pack",
        "0.1.0",
        input_ids=[imported["receipt"]["input_id"]],
    )
    running = ledger.start_run(
        client_root,
        engagement["engagement_id"],
        prepared["run"]["run_id"],
    )
    context_path = Path(running["context_path"])
    bound_input = Path(running["context"]["input_bindings"][0]["path"])
    output_dir = Path(running["output_dir"])
    inspection_dir = output_dir / "inspection"
    pack_dir = output_dir / "pack"

    inspected = subprocess.run(
        [
            sys.executable,
            str(SCRIPTS / "inspect_inputs.py"),
            "--input",
            str(bound_input),
            "--client-engagement",
            str(context_path),
            "--output-dir",
            str(inspection_dir),
        ],
        capture_output=True,
        text=True,
        check=False,
    )
    assert inspected.returncode == 0, inspected.stderr
    inspection = json.loads((inspection_dir / "inspection.json").read_text())
    recipe = _reviewed_recipe(inspection)
    recipe_path = inspection_dir / "reviewed_recipe.json"
    recipe_path.write_text(json.dumps(recipe), encoding="utf-8")

    calculated = subprocess.run(
        [
            sys.executable,
            str(SCRIPTS / "run_pack.py"),
            "--input",
            str(bound_input),
            "--recipe",
            str(recipe_path),
            "--client-engagement",
            str(context_path),
            "--output-dir",
            str(pack_dir),
        ],
        capture_output=True,
        text=True,
        check=False,
    )
    assert calculated.returncode == 0, calculated.stderr
    for name in (
        "management_control_pack.json",
        "management_control_pack.xlsx",
        "management_control_facts.md",
        "management_control_dashboard.html",
        "model_context.json",
        "execution_receipt.json",
    ):
        assert (pack_dir / name).is_file(), name
    workbook = load_workbook(pack_dir / "management_control_pack.xlsx")
    try:
        assert {"Summary", "Monthly P&L", "Budget variance", "AR aging"}.issubset(
            workbook.sheetnames
        )
        assert workbook["Summary"]["C10"].data_type == "n"
        assert workbook["Summary"].page_setup.fitToWidth == 1
        assert workbook["Customers"]["C2"].number_format == "0.00%"
    finally:
        workbook.close()

    commentary = json.loads((pack_dir / "commentary_template.json").read_text())
    commentary["observations"] = [
        {
            "text": "The reviewed period reports EBITDA of 1,000 EUR.",
            "metric_ids": ["pnl.total.ebitda"],
        }
    ]
    commentary_path = pack_dir / "management_commentary.json"
    commentary_path.write_text(json.dumps(commentary), encoding="utf-8")
    final_dir = pack_dir / "final"
    finalized = subprocess.run(
        [
            sys.executable,
            str(SCRIPTS / "finalize_pack.py"),
            "--pack",
            str(pack_dir / "management_control_pack.json"),
            "--commentary",
            str(commentary_path),
            "--client-engagement",
            str(context_path),
            "--output-dir",
            str(final_dir),
        ],
        capture_output=True,
        text=True,
        check=False,
    )
    assert finalized.returncode == 0, finalized.stderr
    assert (final_dir / "management_control_report.md").is_file()
    assert (final_dir / "management_control_dashboard_reviewed.html").is_file()
    assert (final_dir / "commentary_receipt.json").is_file()
    final_html = (final_dir / "management_control_dashboard_reviewed.html").read_text(
        encoding="utf-8"
    )
    final_markdown = (final_dir / "management_control_report.md").read_text(
        encoding="utf-8"
    )
    for heading in ("Payables aging", "Cash movement", "Top customers", "Services"):
        assert heading in final_html
    for heading in (
        "Payables aging",
        "Cash movement",
        "Customer concentration",
        "Service profitability",
    ):
        assert heading in final_markdown


def test_public_page_states_connector_and_model_data_boundaries() -> None:
    page = (
        ROOT / "static" / "shared" / "management-control-pack" / "index.html"
    ).read_text(encoding="utf-8")
    main = page[page.index('<main class="page-shell"') : page.index("</main>")]

    for snippet in (
        "Pacchetto controllo di gestione | Vera",
        "Management control pack | Vera",
        "Prepare recurring management reports from accounting exports.",
        "Non serve un connettore al gestionale",
        "No accounting-system connector is required",
        "Quali dati arrivano al modello",
        "What data reaches the model",
        "Quelles données parviennent au modèle",
        "Welche Daten das Modell erhält",
        "Qué datos recibe el modelo",
        "fino a dieci righe di anteprima per tabella",
        "non riceve di default le popolazioni complete",
        "non monitora i sistemi in background",
        'data-model-data-workflow="management-control-pack"',
        'data-model-data-status="relevant"',
    ):
        assert snippet in page
    assert "files already available" not in page
    visible_keys = set(re.findall(r'data-i18n(?:-aria-label)?="([^"]+)"', page))
    assert all(page.count(f'"{key}":') == 5 for key in visible_keys)
    assert main.rstrip().endswith("</section>")
    assert main.rindex('class="function-model-data"') > main.rindex('id="prompt"')
