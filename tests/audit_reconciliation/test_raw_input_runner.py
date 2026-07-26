from __future__ import annotations

import builtins
import importlib.util
import json
import sys
import zipfile
from pathlib import Path

from openpyxl import Workbook

SCRIPTS = (
    Path(__file__).resolve().parents[2] / "plugins" / "audit-reconciliation" / "scripts"
)
RUNNER = SCRIPTS / "raw_input_runner.py"


def load_runner():
    spec = importlib.util.spec_from_file_location(
        "audit_reconciliation_raw_runner", RUNNER
    )
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def reviewed_source_input(
    *,
    role: str,
    adapter_family: str,
    direction_policy: str = "not_applicable",
    decimal_separator: str | None = ".",
    thousands_separator: str | None = ",",
) -> dict[str, object]:
    return {
        "role": role,
        "adapter_family": adapter_family,
        "reviewer_ref": "reviewer.test",
        "reviewed_on": "2026-07-25",
        "perimeter": {
            "entity_ref": "entity.test",
            "party_ref": "party.test",
            "currency": "EUR",
            "unit": "currency_amount",
            "direction_policy": direction_policy,
            "allocation_policy": "one_to_one",
        },
        "money": {
            "decimal_separator": decimal_separator,
            "thousands_separator": thousands_separator,
            "reported_unit": "EUR",
            "reported_increment": "0.01",
        },
        "date": {"order": "day_first"},
    }


def direct_reviewed_assumptions(
    source_file: str,
    *,
    role: str,
    adapter_family: str,
    direction_policy: str = "not_applicable",
    decimal_separator: str | None = ".",
    thousands_separator: str | None = ",",
) -> dict[str, object]:
    reviewed = reviewed_source_input(
        role=role,
        adapter_family=adapter_family,
        direction_policy=direction_policy,
        decimal_separator=decimal_separator,
        thousands_separator=thousands_separator,
    )
    return {
        "currency": "EUR",
        "_reviewed_source_decision_receipts": {
            source_file: {
                "decision_id": "source_mapping.test",
                "content": {
                    "source_path": source_file,
                    "role": reviewed["role"],
                    "adapter_family": reviewed["adapter_family"],
                    "perimeter": reviewed["perimeter"],
                    "money": reviewed["money"],
                    "date": reviewed["date"],
                },
            }
        },
    }


def test_open_item_document_key_matches_invoice_style_document_key():
    runner = load_runner()

    assert (
        runner.normalize_open_item_document("23FE01/000120", "2023-01-31")
        == "120FE|2023"
    )


def test_two_digit_bank_dates_are_normalized():
    runner = load_runner()

    assert (
        runner.iso_date(
            "03/07/23",
            convention={"order": "day_first"},
        )
        == "2023-07-03"
    )


def test_source_role_suggestions_are_generic_and_not_reviewed_mappings():
    runner = load_runner()

    assert (
        runner.infer_source_role("Open items 2023.pdf", language="en") == "open_items"
    )
    assert (
        runner.infer_source_role("Estratto_Conto_corrente_30_09_2023.pdf")
        == "bank_statement"
    )
    assert runner.infer_source_role("GIORNALE 2023.xlsx") == "journal"
    assert runner.infer_source_role("DistintaPagamento.doc") == "payment_order"
    assert runner.infer_source_role("Unrecognized annex.pdf") == "unknown"

    resolution = runner.resolve_source_role(
        "Open items 2023.pdf",
        assumptions={},
        language="en",
    )

    assert resolution["source_role"] == "unknown"
    assert resolution["suggested_source_role"] == "open_items"
    assert resolution["status"] == "needs_review"


def test_source_role_inference_supports_configured_languages():
    runner = load_runner()

    assert (
        runner.infer_source_role("Customer open items 2023.pdf", language="en_US")
        == "open_items"
    )
    assert (
        runner.infer_source_role("Releve bancaire 2023.pdf", language="fr_FR")
        == "bank_statement"
    )
    assert (
        runner.infer_source_role("Hauptjournal 2023.xlsx", language="de_DE")
        == "journal"
    )
    assert (
        runner.infer_source_role("Zahlungsauftrag 42.doc", language="de_DE")
        == "payment_order"
    )


def test_ambiguous_source_role_suggestions_require_reviewed_input():
    runner = load_runner()

    text = "MASTRINO DI SOTTOCONTO Banca c/anticipi DISPONIBILITA LIQUIDE"

    assert (
        runner.infer_source_role("Banca anticipi.pdf", sample_text=text, language="it")
        == "unknown"
    )
    assert set(
        runner.infer_source_role_candidates(
            "Banca anticipi.pdf",
            sample_text=text,
            language="it",
        )
    ) >= {"bank_statement", "ledger", "factoring_statement"}


def test_generic_statement_filename_abstains_when_roles_are_ambiguous():
    runner = load_runner()

    assert (
        runner.infer_source_role("Customer statement 2023.pdf", language="en_US")
        == "unknown"
    )
    assert set(
        runner.infer_source_role_candidates(
            "Customer statement 2023.pdf",
            language="en_US",
        )
    ) == {"bank_statement", "open_items"}


def test_reviewed_source_role_is_authoritative_over_advisory_filename():
    runner = load_runner()

    resolution = runner.resolve_source_role(
        "ambiguous bank journal.pdf",
        assumptions=direct_reviewed_assumptions(
            "ambiguous bank journal.pdf",
            role="bank_statement",
            adapter_family="bank_statement_text_v1",
        ),
        language="en",
    )

    assert resolution["source_role"] == "bank_statement"
    assert resolution["status"] == "reviewed"


def test_base_requirements_declare_pymupdf():
    requirements = (SCRIPTS.parent / "requirements.txt").read_text(encoding="utf-8")

    assert "PyMuPDF" in requirements


def test_configure_ocr_environment_defaults_to_paddle_bos(tmp_path, monkeypatch):
    runner = load_runner()
    cache_dir = tmp_path / "cache"

    monkeypatch.delenv("PADDLE_PDX_MODEL_SOURCE", raising=False)
    monkeypatch.delenv("PADDLE_PDX_CACHE_HOME", raising=False)

    runner.configure_ocr_environment(cache_dir)

    assert (cache_dir / "paddlex").exists()
    assert runner.os.environ["PADDLE_PDX_MODEL_SOURCE"] == "bos"
    assert runner.os.environ["PADDLE_PDX_CACHE_HOME"] == str(cache_dir / "paddlex")


def test_configure_ocr_environment_preserves_explicit_model_source(
    tmp_path, monkeypatch
):
    runner = load_runner()

    monkeypatch.setenv("PADDLE_PDX_MODEL_SOURCE", "huggingface")

    runner.configure_ocr_environment(tmp_path / "cache")

    assert runner.os.environ["PADDLE_PDX_MODEL_SOURCE"] == "huggingface"


def test_optional_pdf_imports_are_independent(monkeypatch):
    real_import = builtins.__import__

    def fake_import(name, globals=None, locals=None, fromlist=(), level=0):
        if name == "fitz":
            raise ModuleNotFoundError("No module named 'fitz'")
        return real_import(name, globals, locals, fromlist, level)

    monkeypatch.setattr(builtins, "__import__", fake_import)

    runner = load_runner()

    assert runner.fitz is None
    assert runner.load_workbook is not None
    assert runner.pdfplumber is not None


def test_ocr_page_text_uses_local_paddle_fallback(tmp_path, monkeypatch):
    runner = load_runner()
    captured: dict[str, object] = {}

    class FakePixmap:
        def tobytes(self, image_format):
            assert image_format == "png"
            return b"image-bytes"

    class FakePage:
        def get_pixmap(self, matrix, alpha=False):
            assert matrix is not None
            assert alpha is False
            return FakePixmap()

    class FakeDoc:
        def __getitem__(self, index):
            assert index == 0
            return FakePage()

    class FakeFitz:
        @staticmethod
        def open(path):
            assert path.name == "scan.pdf"
            return FakeDoc()

        @staticmethod
        def Matrix(*_args):
            return object()

    def fake_local_ocr(image_bytes, *, lang, text_recognition_model_name=None):
        captured["image_bytes"] = image_bytes
        captured["lang"] = lang
        captured["text_recognition_model_name"] = text_recognition_model_name
        return "testo OCR"

    runner.fitz = FakeFitz
    monkeypatch.setattr(
        runner, "_shared_ocr_text_from_image_bytes", lambda *args, **kwargs: None
    )
    monkeypatch.setattr(
        runner, "_local_paddle_ocr_text_from_image_bytes", fake_local_ocr
    )

    text = runner._ocr_page_text(
        tmp_path / "scan.pdf",
        0,
        tmp_path / "cache",
        language="it_IT",
    )

    assert text == "testo OCR"
    assert captured == {
        "image_bytes": b"image-bytes",
        "lang": "it",
        "text_recognition_model_name": "PP-OCRv5_server_rec",
    }


def test_raw_ocr_text_supports_paddle_v3_rec_texts():
    runner = load_runner()

    assert (
        runner._raw_ocr_text([{"rec_texts": ["Riga 1", "Riga 2"]}]) == "Riga 1\nRiga 2"
    )


def test_extract_normalized_records_defaults_cache_inside_output_dir(tmp_path):
    runner = load_runner()
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()

    result = runner.extract_normalized_records(input_dir, {}, output_dir=output_dir)

    assert Path(result["cache_dir"]) == output_dir / ".audit_reconciliation_cache"
    assert (output_dir / ".audit_reconciliation_cache").exists()


def test_unreviewed_source_role_abstains_and_surfaces_needs_review(tmp_path):
    runner = load_runner()
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    (input_dir / "bank journal.pdf").write_bytes(b"not parsed")

    result = runner.extract_normalized_records(
        input_dir,
        {},
        output_dir=output_dir,
    )

    assert result["normalized_records"] == []
    assert result["source_qualifications"][0]["status"] == "needs_review"
    assert result["source_inventory"][0]["source_role"] == "unknown"
    assert result["extraction_errors"][0]["status"] == "needs_review"


def test_plain_legacy_role_mapping_never_grants_parser_authority(tmp_path):
    runner = load_runner()
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    source_name = "customer ledger.pdf"
    (input_dir / source_name).write_bytes(b"not parsed")

    result = runner.extract_normalized_records(
        input_dir,
        {"reviewed_source_roles": {source_name: "ledger"}},
        output_dir=output_dir,
    )

    qualification = result["source_qualifications"][0]
    assert result["normalized_records"] == []
    assert qualification["status"] == "needs_review"
    assert qualification["emitted_row_count"] == 0
    assert result["extraction_errors"][0]["status"] == "needs_review"


def test_unknown_reviewed_adapter_does_not_unlock_source_parser(tmp_path):
    runner = load_runner()
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    source_name = "customer open items.pdf"
    (input_dir / source_name).write_bytes(b"not parsed")

    result = runner.extract_normalized_records(
        input_dir,
        {
            "reviewed_source_decisions": {
                source_name: reviewed_source_input(
                    role="open_items",
                    adapter_family="custom_unknown_v1",
                    direction_policy="customer",
                )
            },
        },
        output_dir=output_dir,
    )

    qualification = result["source_qualifications"][0]
    assert result["normalized_records"] == []
    assert qualification["adapter_id"] == "custom_unknown_v1"
    assert qualification["status"] == "unsupported_source_layout"
    assert result["extraction_errors"][0]["status"] == "unsupported_source_layout"


def test_journal_amount_ownership_requires_exact_columns_or_legacy_adapter():
    runner = load_runner()
    row = ("1", "Account", "Description", "40.00", "999.00", "60.00")
    layout = {"operation_col": 3, "debit_col": 4, "credit_col": 6}

    exact = runner.journal_amount_sides(row, layout)
    legacy = runner.journal_amount_sides(
        row,
        layout,
        adapter_family=runner.LEGACY_ADAPTER_FAMILY,
    )

    assert exact == (runner.Decimal("40.00"), runner.Decimal("60.00"))
    assert legacy == (runner.Decimal("1039.00"), runner.Decimal("60.00"))


def test_authoritative_money_abstains_on_ambiguous_separator_and_float():
    runner = load_runner()

    assert runner.parse_money("1.000") is None
    assert runner.parse_money(1000.25) is None
    assert runner.parse_money(
        "1.000",
        convention={
            "decimal_separator": ",",
            "thousands_separator": ".",
            "reported_increment": "0.01",
        },
    ) == runner.Decimal("1000")
    assert (
        runner.parse_money(
            "10.005",
            convention={
                "decimal_separator": ".",
                "thousands_separator": ",",
                "reported_increment": "0.01",
            },
        )
        is None
    )


def test_authoritative_money_honors_explicit_thousands_separator_role():
    runner = load_runner()
    convention = {
        "decimal_separator": None,
        "thousands_separator": ",",
        "reported_increment": "0.01",
    }

    result = runner.parse_money("1,23", convention=convention)

    assert result is None


def test_journal_float_cell_never_becomes_authoritative_money():
    runner = load_runner()

    assert runner.journal_money_cell(1000.25) is None
    assert runner.journal_money_cell("1000.25") == runner.Decimal("1000.25")


def test_bank_row_without_document_reference_is_unallocated_external():
    runner = load_runner()
    page = runner.SourcePage(
        source_file="bank.pdf",
        source_role="bank_statement",
        source_page=1,
        extraction_method="pdf_text",
        text_length=100,
        line_count=1,
        text="",
    )

    rows = runner._bank_row_from_text(
        page,
        "03/07/23 03/07/23 27.400,77 BONIFICO o/c: CUSTOMER S.DO DIST.PG.152",
        1,
        {
            **direct_reviewed_assumptions(
                page.source_file,
                role="bank_statement",
                adapter_family="bank_statement_text_v1",
                decimal_separator=",",
                thousands_separator=".",
            ),
            "counterparty_keywords": ["customer"],
        },
        ["customer"],
    )

    assert rows[0]["evidence_type"] == "unallocated_external_bank"
    assert rows[0]["amount"] == "27400.77"
    assert rows[0]["posting_date"] == "2023-07-03"


def test_bank_distinta_range_is_batch_not_invoice_reference():
    runner = load_runner()
    page = runner.SourcePage(
        source_file="bank.pdf",
        source_role="bank_statement",
        source_page=1,
        extraction_method="pdf_text",
        text_length=100,
        line_count=1,
        text="",
    )

    rows = runner._bank_row_from_text(
        page,
        "11/07/23 11/07/23 133.633,12 BONIFICO o/c: CUSTOMER S.DO DIST.PG.1-7 CUSTOMER",
        1,
        {
            **direct_reviewed_assumptions(
                page.source_file,
                role="bank_statement",
                adapter_family="bank_statement_text_v1",
                decimal_separator=",",
                thousands_separator=".",
            ),
            "counterparty_keywords": ["customer"],
        },
        ["customer"],
    )

    assert rows[0]["evidence_type"] == "unallocated_external_bank"
    assert rows[0]["document_key"] == ""
    assert (
        rows[0]["batch_ids"]
        == "distinta:1;distinta:2;distinta:3;distinta:4;distinta:5;distinta:6;distinta:7"
    )


def test_bank_fatt_number_reference_is_invoice_key():
    runner = load_runner()
    page = runner.SourcePage(
        source_file="bank.pdf",
        source_role="bank_statement",
        source_page=1,
        extraction_method="pdf_text",
        text_length=100,
        line_count=1,
        text="",
    )

    rows = runner._bank_row_from_text(
        page,
        "13/01/23 13/01/23 2.752,96 RIENTRO ANTICIPO/FINANZIAMENTO FATT. 587 CUSTOMER",
        1,
        {
            **direct_reviewed_assumptions(
                page.source_file,
                role="bank_statement",
                adapter_family="bank_statement_text_v1",
                decimal_separator=",",
                thousands_separator=".",
            ),
            "counterparty_keywords": ["customer"],
        },
        ["customer"],
    )

    assert rows[0]["evidence_type"] == "external_bank"
    assert rows[0]["document_key"] == "587|2023"


def test_bank_ft_number_reference_is_invoice_key():
    runner = load_runner()
    page = runner.SourcePage(
        source_file="bank.pdf",
        source_role="bank_statement",
        source_page=1,
        extraction_method="pdf_text",
        text_length=100,
        line_count=1,
        text="",
    )

    rows = runner._bank_row_from_text(
        page,
        "19/05/23 19/05/23 26.067,14 BONIFICI ESTERI o/c: PAYMENTCO SAS FACTORCO-22651 ACCONTO ID 12411 FT. 293 - CUSTOMER",
        1,
        {
            **direct_reviewed_assumptions(
                page.source_file,
                role="bank_statement",
                adapter_family="bank_statement_text_v1",
                decimal_separator=",",
                thousands_separator=".",
            ),
            "counterparty_keywords": ["customer"],
            "factoring_operator_keywords": ["factorco", "paymentco"],
        },
        ["customer"],
        ["factorco", "paymentco"],
    )

    assert rows[0]["evidence_type"] == "external_bank"
    assert rows[0]["document_no"] == "293"
    assert rows[0]["document_key"] == "293|2023"


def test_bank_invoice_reference_supports_english_keyword_and_decimal_point():
    runner = load_runner()
    page = runner.SourcePage(
        source_file="bank.pdf",
        source_role="bank_statement",
        source_page=1,
        extraction_method="pdf_text",
        text_length=100,
        line_count=1,
        text="",
    )

    rows = runner._bank_row_from_text(
        page,
        "19/05/23 19/05/23 26,067.14 WIRE TRANSFER CUSTOMER INVOICE 293",
        1,
        {
            **direct_reviewed_assumptions(
                page.source_file,
                role="bank_statement",
                adapter_family="bank_statement_text_v1",
                decimal_separator=".",
                thousands_separator=",",
            ),
            "counterparty_keywords": ["customer"],
            "document_language": "en",
        },
        ["customer"],
    )

    assert rows[0]["evidence_type"] == "external_bank"
    assert rows[0]["amount"] == "26067.14"
    assert rows[0]["document_key"] == "293|2023"


def test_parse_open_items_from_ocr_like_lines():
    runner = load_runner()
    page = runner.SourcePage(
        source_file="Open-items-customers-2023.pdf",
        source_role="open_items",
        source_page=1,
        extraction_method="paddle_ocr",
        text_length=120,
        line_count=5,
        text="\n".join(
            [
                "RICONCILIAZIONE SCHEDA CLIENTE",
                "23FE01/000120",
                "31/01/23",
                "77.032,64",
                "77.032,64",
            ]
        ),
    )

    rows = runner.parse_open_items(
        [page],
        direct_reviewed_assumptions(
            page.source_file,
            role="open_items",
            adapter_family="open_items_text_v1",
            direction_policy="customer",
            decimal_separator=",",
            thousands_separator=".",
        ),
    )

    assert len(rows) == 1
    assert rows[0]["document_key"] == "120FE|2023"
    assert rows[0]["amount"] == "77032.64"
    assert rows[0]["expected_side"] == "customer"


def test_parse_open_items_keywords_cannot_choose_accounting_direction():
    runner = load_runner()
    page = runner.SourcePage(
        source_file="Supplier statement 2023.pdf",
        source_role="open_items",
        source_page=1,
        extraction_method="pdf_text",
        text_length=120,
        line_count=5,
        text="\n".join(
            [
                "SUPPLIER STATEMENT",
                "23FF01/000120",
                "31/01/23",
                "1,000.00",
                "1,000.00",
            ]
        ),
    )

    rows = runner.parse_open_items([page], {"currency": "EUR"})

    assert rows == []


def test_extract_pdf_pages_uses_page_cache(tmp_path):
    runner = load_runner()
    pdf_path = tmp_path / "sample.pdf"
    pdf_path.write_bytes(b"%PDF cached test")
    cache_dir = tmp_path / "cache"
    calls = {"open": 0}

    class FakePage:
        def extract_text(self):
            return "Estratto conto corrente con testo gia leggibile e sufficiente per evitare OCR"

    class FakePdf:
        pages = [FakePage()]

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            return False

    class FakePdfPlumber:
        def open(self, path):
            calls["open"] += 1
            return FakePdf()

    class FailingPdfPlumber:
        def open(self, path):
            raise AssertionError("pdfplumber should not be called on cache hit")

    runner.pdfplumber = FakePdfPlumber()
    first = runner.extract_pdf_pages(pdf_path, cache_dir)
    runner.pdfplumber = FailingPdfPlumber()
    second = runner.extract_pdf_pages(pdf_path, cache_dir)

    assert calls["open"] == 1
    assert first[0].text == second[0].text
    assert second[0].source_file == "sample.pdf"


def test_extract_pdf_pages_emits_ocr_progress_events(tmp_path, monkeypatch):
    runner = load_runner()
    pdf_path = tmp_path / "scan.pdf"
    pdf_path.write_bytes(b"%PDF scanned test")
    cache_dir = tmp_path / "cache"
    events: list[dict[str, object]] = []

    class FakePage:
        def extract_text(self):
            return ""

    class FakePdf:
        pages = [FakePage()]

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            return False

    class FakePdfPlumber:
        def open(self, path):
            assert path == pdf_path
            return FakePdf()

    runner.pdfplumber = FakePdfPlumber()
    monkeypatch.setattr(runner, "_ocr_page_text", lambda *_args, **_kwargs: "testo OCR")

    pages = runner.extract_pdf_pages(
        pdf_path,
        cache_dir,
        progress_callback=events.append,
    )

    assert pages[0].extraction_method == "paddle_ocr"
    assert [event["event"] for event in events] == [
        "pdf_file_start",
        "ocr_page_start",
        "ocr_page_done",
        "pdf_page_done",
        "pdf_file_done",
    ]
    assert events[0]["source_file"] == "scan.pdf"
    assert events[0]["page_count"] == 1
    assert events[1]["source_page"] == 1
    assert events[1]["page_count"] == 1
    assert events[2]["text_length"] == len("testo OCR")
    assert events[3]["extraction_method"] == "paddle_ocr"
    assert events[4]["ocr_page_count"] == 1


def test_raw_run_writes_extracted_source_pages_to_output(tmp_path, monkeypatch):
    runner = load_runner()
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "out"
    input_dir.mkdir()

    extracted_pages = [
        {
            "source_file": "scan.pdf",
            "source_role": "open_items",
            "source_page": 1,
            "extraction_method": "paddle_ocr",
            "text_length": 18,
            "line_count": 1,
            "text": "OCR extracted text",
        }
    ]

    def fake_extract_normalized_records(input_dir, assumptions, *, output_dir=None):
        return {
            "source_inventory": [
                {"source_file": "scan.pdf", "source_role": "open_items"}
            ],
            "source_pages": extracted_pages,
            "open_items": [],
            "evidence_rows": [],
            "ledger_balance_rows": [],
            "journal_rollforward_rows": [],
            "journal_rollforward_summary": [],
            "normalized_records": [],
            "extraction_errors": [],
            "cache_dir": str(tmp_path / ".audit_reconciliation_cache"),
        }

    monkeypatch.setattr(
        runner, "extract_normalized_records", fake_extract_normalized_records
    )

    result = runner.run_raw_input_reconciliation(
        input_dir=input_dir,
        output_dir=output_dir,
        assumptions={"scope_year": "2023", "cutoff_date": "2023-12-31"},
    )

    source_pages_path = output_dir / "source_pages.json"
    manifest = json.loads(
        (output_dir / "run_manifest.json").read_text(encoding="utf-8")
    )
    source_pages = json.loads(source_pages_path.read_text(encoding="utf-8"))

    assert result["manifest"]["source_pages_path"] == str(source_pages_path)
    assert Path(result["manifest"]["accountant_report_path"]).exists()
    assert manifest["source_pages_path"] == str(source_pages_path)
    assert Path(manifest["accountant_report_path"]).exists()
    assert manifest["counts"]["source_pages"] == 1
    assert source_pages == extracted_pages


def test_payment_order_zip_extracts_invoice_rows_and_batch_total(tmp_path):
    runner = load_runner()
    zip_path = tmp_path / "orders.zip"
    html = """
    Distinta di Pagamento Distinta 0000401 Del 21/12/2023 Valuta 29/12/2023
    FA Tipo Data FA FP Data FP %QC Importo Importo QC
    10179-23 Fattura 13/10/2023 1515-23 26/10/2023 1,50 27.400,77 411,01
    Totale Distinta 26.989,76
    """
    with zipfile.ZipFile(zip_path, "w") as zf:
        zf.writestr("DistintaPagamento.doc", html)

    rows = runner.parse_payment_order_zip(
        zip_path,
        direct_reviewed_assumptions(
            zip_path.name,
            role="payment_order",
            adapter_family="payment_order_html_zip_v1",
            decimal_separator=",",
            thousands_separator=".",
        ),
    )

    assert len(rows) == 1
    assert rows[0]["document_key"] == "1515|2023"
    assert rows[0]["batch_id"] == "distinta:401"
    assert rows[0]["batch_total"] == "26989.76"
    assert rows[0]["value_date"] == "2023-12-29"


def test_parse_pdf_journal_page_when_no_spreadsheet_journal():
    runner = load_runner()
    page = runner.SourcePage(
        source_file="giornale.pdf",
        source_role="journal",
        source_page=7,
        extraction_method="pdf_text",
        text_length=200,
        line_count=2,
        text="\n".join(
            [
                "07/04/2023 INCASSATA FATTURA",
                "11746 15 / 5 / 1003 Example Bank S.p.A. N.3389-23 del 07042023 EXAMPLE SUPPLIER 244.324,25",
            ]
        ),
    )

    rows = runner.parse_journal_pages(
        [page],
        {
            **direct_reviewed_assumptions(
                page.source_file,
                role="journal",
                adapter_family=runner.LEGACY_ADAPTER_FAMILY,
                decimal_separator=",",
                thousands_separator=".",
            ),
        },
    )

    assert len(rows) == 1
    assert rows[0]["source_role"] == "journal"
    assert rows[0]["evidence_type"] == "internal_closure"
    assert rows[0]["document_key"] == "3389|2023"
    assert rows[0]["amount"] == "244324.25"


def test_parse_pdf_journal_page_abstains_without_reviewed_legacy_adapter():
    runner = load_runner()
    page = runner.SourcePage(
        source_file="giornale.pdf",
        source_role="journal",
        source_page=7,
        extraction_method="pdf_text",
        text_length=200,
        line_count=2,
        text="\n".join(
            [
                "07/04/2023 INCASSATA FATTURA",
                "11746 15 / 5 / 1003 Example Bank S.p.A. "
                "N.3389-23 del 07042023 EXAMPLE SUPPLIER 244.324,25",
            ]
        ),
    )

    rows = runner.parse_journal_pages([page], {"currency": "EUR"})

    assert rows == []


def test_parse_journal_rollforward_xlsx_extracts_counterparty_opening_and_movements(
    tmp_path,
):
    runner = load_runner()
    path = tmp_path / "GIORNALE 2023.xlsx"
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Table 1"
    ws.cell(row=4, column=1).value = "Data registrazione"
    ws.cell(row=4, column=7).value = "Causale"
    ws.cell(row=5, column=1).value = "Riga"
    ws.cell(row=5, column=7).value = "Conto"
    ws.cell(row=5, column=12).value = "Descrizione conto"
    ws.cell(row=5, column=19).value = "Descrizione dell'operazione"
    ws.cell(row=5, column=25).value = "Dare"
    ws.cell(row=5, column=43).value = "Avere"
    ws.cell(row=6, column=1).value = "01/01/2023 APERTURA ESERCIZIO"
    ws.cell(row=7, column=1).value = 126
    ws.cell(row=7, column=7).value = "9 / 5 / 3"
    ws.cell(row=7, column=12).value = "Customer Srl"
    ws.cell(row=7, column=19).value = "Apertura esercizio in data 01/01/2023"
    ws.cell(row=7, column=25).value = "1000.00"
    ws.cell(row=8, column=1).value = 149
    ws.cell(row=8, column=7).value = "22 / 5 / 15"
    ws.cell(row=8, column=12).value = "Customer Srl"
    ws.cell(row=8, column=19).value = "Apertura esercizio in data 01/01/2023"
    ws.cell(row=8, column=43).value = "300.00"
    ws.cell(row=9, column=1).value = "15/01/2023 INCASSATA FATTURA"
    ws.cell(row=10, column=1).value = 300
    ws.cell(row=10, column=7).value = "9 / 5 / 3"
    ws.cell(row=10, column=12).value = "Customer Srl"
    ws.cell(row=10, column=19).value = "Incasso fattura"
    ws.cell(row=10, column=43).value = "200.00"
    ws.cell(row=11, column=1).value = 301
    ws.cell(row=11, column=7).value = "22 / 5 / 15"
    ws.cell(row=11, column=12).value = "Customer Srl"
    ws.cell(row=11, column=19).value = "Pagamento fattura"
    ws.cell(row=11, column=25).value = "50.00"
    workbook.save(path)

    rows = runner.parse_journal_rollforward_xlsx(
        path,
        {
            **direct_reviewed_assumptions(
                path.name,
                role="journal",
                adapter_family=runner.LEGACY_ADAPTER_FAMILY,
            ),
            "counterparty_keywords": ["customer"],
        },
    )
    summary = runner.summarize_journal_rollforward(rows)
    total = summary[0]

    assert len(rows) == 4
    assert rows[0]["movement_type"] == "opening"
    assert rows[0]["debit_amount"] == "1000"
    assert rows[1]["credit_amount"] == "300"
    assert total["account"] == "TOTAL"
    assert total["opening_net_debit_minus_credit"] == "700.00"
    assert total["period_net_debit_minus_credit"] == "-150.00"
    assert total["closing_net_debit_minus_credit"] == "550.00"


def test_rollforward_keywords_are_inferred_from_non_bank_ledgers():
    runner = load_runner()

    keywords = runner.rollforward_counterparty_keywords(
        [
            {"account": "TOTAL", "account_name": "All matched counterparty ledgers"},
            {
                "account": "15 / 5 / 1003",
                "account_name": "Banco Example S.p.A.",
                "source_file": "bank.pdf",
            },
            {
                "account": "9 / 5 / 3",
                "account_name": "Customer Srl",
                "source_file": "customer-ledger.pdf",
            },
        ],
        {},
    )

    assert keywords == ["customer srl"]


def test_build_account_rollforward_check_compares_journal_to_ledger():
    runner = load_runner()

    check = runner.build_account_rollforward_check(
        [
            {
                "account": "9 / 5 / 3",
                "account_name": "Customer Srl",
                "source_file": "customer-ledger.pdf",
                "source_pages": "1-2",
                "opening_balance_signed_debit_minus_credit": "1000.00",
                "closing_balance_signed_debit_minus_credit": "550.00",
            }
        ],
        [
            {
                "account": "TOTAL",
                "account_name": "All matched counterparty journal accounts",
                "rows": 2,
                "opening_net_debit_minus_credit": "1000.00",
                "period_net_debit_minus_credit": "-450.00",
                "closing_net_debit_minus_credit": "550.00",
            },
            {
                "account": "9 / 5 / 3",
                "account_name": "Customer Srl",
                "rows": 2,
                "opening_net_debit_minus_credit": "1000.00",
                "period_net_debit_minus_credit": "-450.00",
                "closing_net_debit_minus_credit": "550.00",
            },
        ],
        {"amount_tolerance": "0.01"},
    )

    assert check[0]["account"] == "TOTAL"
    assert check[0]["status"] == "PASS"
    assert check[1]["closing_difference_journal_minus_ledger"] == "0.00"


def test_parse_ledger_balance_pages_extracts_pre_closing_balance():
    runner = load_runner()
    page = runner.SourcePage(
        source_file="ledger.pdf",
        source_role="ledger",
        source_page=1,
        extraction_method="pdf_text",
        text_length=500,
        line_count=8,
        text="\n".join(
            [
                "Conto: 9 / 5 / 3 Customer Srl",
                "15808 01/01/2023 A 1 752 APERTURA ESERCIZIO",
                "Apertura esercizio in data 01/01/2023 2023 1.000,00 1.000,00 +",
                "12480 29/12/2023 1939 A 1 1 FATTURA DI VENDITA",
                "29/12/2023 1895-FE Del 29122023 n1895-FE Pr1939 2023 200,00 1.200,00 +",
                "17119 31/12/2023 A 1 751 CHIUSURA ESERCIZIO",
                "Chiusura Clienti in data 31/12/2023 2023 1.200,00 0,00 +",
            ]
        ),
    )

    rows = runner.parse_ledger_balance_pages(
        [page],
        {
            "counterparty_keywords": ["customer"],
            **direct_reviewed_assumptions(
                page.source_file,
                role="ledger",
                adapter_family=runner.LEGACY_ADAPTER_FAMILY,
                decimal_separator=",",
                thousands_separator=".",
            ),
        },
    )

    assert rows[0]["account"] == "TOTAL"
    assert rows[0]["opening_balance_signed_debit_minus_credit"] == "1000.00"
    assert rows[0]["closing_balance_signed_debit_minus_credit"] == "1200.00"
    assert rows[1]["account"] == "9 / 5 / 3"
