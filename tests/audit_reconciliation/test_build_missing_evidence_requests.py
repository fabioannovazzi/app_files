from __future__ import annotations

import importlib.util
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

SCRIPTS = (
    Path(__file__).resolve().parents[2] / "plugins" / "audit-reconciliation" / "scripts"
)
SCRIPT = SCRIPTS / "build_missing_evidence_requests.py"


def load_missing_requests():
    sys.path.insert(0, str(SCRIPTS))
    spec = importlib.util.spec_from_file_location(
        "audit_reconciliation_missing_requests", SCRIPT
    )
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def write_reconciliation_workbook(path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    inventory = workbook.create_sheet("Source inventory")
    inventory.append(["source_file", "source_role"])
    inventory.append(["All.A.pdf", "open_items"])
    inventory.append(["bank.pdf", "bank_statement"])
    inventory.append(["giornale.xlsx", "journal"])

    normalized = workbook.create_sheet("Normalized records")
    normalized.append(["record_id", "source_file", "source_role"])
    normalized.append(["open-1", "All.A.pdf", "open_items"])
    normalized.append(["bank-1", "bank.pdf", "bank_statement"])
    normalized.append(["journal-1", "giornale.xlsx", "journal"])

    detail = workbook.create_sheet("Reconciliation detail")
    detail.append(
        [
            "record_id",
            "reconciliation_status",
            "rule_applied",
            "amount",
            "document_no",
            "document_date",
            "expected_side",
            "matched_evidence_reference",
            "missing_evidence",
            "probable_bank_reference",
            "probable_bank_description",
            "description",
        ]
    )
    detail.append(
        [
            "probable-1",
            "probable_payment",
            "probable_bank_payment_candidate",
            "120.00",
            "INV-PROB",
            "2023-03-01",
            "customer",
            "file=bank.pdf; row=9",
            "",
            "file=bank.pdf; row=9",
            "BONIFICO DISTINTA 7",
            "",
        ]
    )
    detail.append(
        [
            "internal-1",
            "needs_evidence",
            "internal_closure_without_external",
            "80.00",
            "INV-INT",
            "2023-04-01",
            "supplier",
            "file=giornale.xlsx; row=40",
            "Serve evidenza esterna.",
            "",
            "",
            "Chiusura interna",
        ]
    )
    detail.append(
        [
            "open-1",
            "open_supported",
            "internal_booking_open_support",
            "300.00",
            "INV-OPEN",
            "2023-05-01",
            "supplier",
            "file=mastro.pdf; page=4",
            "Serve evidenza esterna specifica.",
            "",
            "",
            "Saldo aperto",
        ]
    )
    detail.append(
        [
            "unresolved-1",
            "unresolved",
            "unresolved",
            "10.00",
            "INV-UNR",
            "2023-06-01",
            "customer",
            "",
            "Acquisire evidenza specifica.",
            "",
            "",
            "Senza match",
        ]
    )
    detail.append(
        [
            "closed-1",
            "closed",
            "external_bank_match",
            "50.00",
            "INV-CLOSED",
            "2023-07-01",
            "customer",
            "file=bank.pdf; row=10",
            "",
            "",
            "",
            "",
        ]
    )
    workbook.save(path)


def test_missing_evidence_pack_separates_available_from_missing(tmp_path):
    missing_requests = load_missing_requests()
    workbook_path = tmp_path / "riconciliazione.xlsx"
    write_reconciliation_workbook(workbook_path)

    context = missing_requests.load_reconciliation_context(workbook_path)
    pack = missing_requests.build_missing_evidence_request_pack(
        context["reconciliation_rows"],
        source_inventory=context["source_inventory"],
        normalized_records=context["normalized_records"],
        entity_name="ExampleCo",
        counterparty_name="Example Supplier",
        cutoff_date="2023-12-31",
    )

    assert len(pack.request_sections["reconciled_strong"]) == 1
    assert len(pack.request_sections["probable_payment"]) == 1
    assert len(pack.request_sections["accounting_support_needed"]) == 1
    assert len(pack.request_sections["open_balance_confirmation"]) == 1
    assert len(pack.request_sections["unresolved"]) == 1
    assert all(row["section"] != "closed" for row in pack.summary)

    probable = pack.request_sections["probable_payment"][0]
    assert probable["operational_owner"] == "ExampleCo"
    assert "gia presenti" in probable["available_evidence"]
    assert "allocazione fattura-per-fattura" in probable["targeted_missing_item"]

    open_row = pack.request_sections["open_balance_confirmation"][0]
    assert open_row["operational_owner"] == (
        "ExampleCo; eventuale conferma saldo Example Supplier"
    )
    assert "2023-12-31" in open_row["targeted_missing_item"]


def test_missing_evidence_workbook_has_operational_tabs(tmp_path):
    missing_requests = load_missing_requests()
    workbook_path = tmp_path / "riconciliazione.xlsx"
    output_path = tmp_path / "richieste.xlsx"
    write_reconciliation_workbook(workbook_path)
    context = missing_requests.load_reconciliation_context(workbook_path)
    pack = missing_requests.build_missing_evidence_request_pack(
        context["reconciliation_rows"],
        source_inventory=context["source_inventory"],
        normalized_records=context["normalized_records"],
        entity_name="ExampleCo",
        counterparty_name="Example Supplier",
        cutoff_date="2023-12-31",
    )

    missing_requests.write_missing_evidence_workbook(output_path, pack)

    workbook = load_workbook(output_path)
    assert workbook.sheetnames == [
        "istruzioni",
        "sintesi",
        "evidenze_disponibili",
        "riconciliate_forti",
        "pagamenti_probabili",
        "scritture_da_supportare",
        "evidenze_da_integrare",
        "saldi_aperti",
        "non_risolte",
    ]
    headers = [
        cell.value
        for cell in next(workbook["pagamenti_probabili"].iter_rows(max_row=1))
    ]
    assert "evidenza_gia_disponibile" in headers
    assert "dato_mancante_mirato" in headers
    assert "stato_attuale" not in headers
    assert "regola" not in headers


@pytest.mark.parametrize(
    ("language", "sheet_name", "header", "category"),
    [
        (
            "fr",
            "paiements_probables",
            "preuve_deja_disponible",
            "Paiements probables a affecter",
        ),
        (
            "de",
            "wahrscheinliche_zahlungen",
            "bereits_verfuegbare_evidenz",
            "Wahrscheinliche Zahlungen zuzuordnen",
        ),
        (
            "en",
            "probable_payments",
            "available_evidence",
            "Probable payments to allocate",
        ),
    ],
)
def test_missing_evidence_workbook_localizes_operational_language(
    tmp_path, language, sheet_name, header, category
):
    missing_requests = load_missing_requests()
    workbook_path = tmp_path / "riconciliazione.xlsx"
    output_path = tmp_path / f"requests-{language}.xlsx"
    write_reconciliation_workbook(workbook_path)
    context = missing_requests.load_reconciliation_context(workbook_path)
    pack = missing_requests.build_missing_evidence_request_pack(
        context["reconciliation_rows"],
        source_inventory=context["source_inventory"],
        normalized_records=context["normalized_records"],
        entity_name="ExampleCo",
        counterparty_name="Example Supplier",
        cutoff_date="2023-12-31",
        language=language,
    )

    missing_requests.write_missing_evidence_workbook(output_path, pack)

    workbook = load_workbook(output_path)
    assert sheet_name in workbook.sheetnames
    rows = list(workbook[sheet_name].iter_rows(values_only=True))
    assert header in rows[0]
    flat_text = " ".join(str(value) for row in rows for value in row if value)
    assert category in flat_text
    assert "probable_bank_payment_candidate" not in flat_text
    assert "open_supported" not in flat_text
