from __future__ import annotations

import importlib.util
import sys
from pathlib import Path

from openpyxl import load_workbook

SCRIPTS = (
    Path(__file__).resolve().parents[2] / "plugins" / "audit-reconciliation" / "scripts"
)
ACCOUNTANT_REPORT = SCRIPTS / "accountant_report.py"


def load_accountant_report():
    sys.path.insert(0, str(SCRIPTS))
    spec = importlib.util.spec_from_file_location(
        "audit_accountant_report", ACCOUNTANT_REPORT
    )
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def test_build_accountant_report_rows_explains_probable_bank_candidate():
    report = load_accountant_report()

    main_rows, detail_rows, legend_rows = report.build_accountant_report_rows(
        [
            {
                "record_id": "open-1",
                "document_no": "23FE01/000001",
                "document_key": "1FE|2023",
                "document_date": "2023-01-31",
                "amount": "100.00",
                "balance": "100.00",
                "reconciliation_status": "probable_payment",
                "evidence_level": "bridge_only",
                "rule_applied": "candidate_bank_matches_nonclosed_open_total",
                "missing_evidence": "Confermare allocazione banca-fattura.",
            }
        ],
        bank_allocation_candidates=[
            {
                "candidate_open_record_ids": "open-1",
                "candidate_confidence": "medium",
                "candidate_amount_match": "YES",
                "bank_date": "2023-02-10",
                "bank_amount": "100.00",
                "bank_source_file": "bank.pdf",
                "bank_source_page": "2",
                "bank_source_row": "7",
                "bank_description": "Bonifico fattura 1",
            }
        ],
    )

    assert main_rows[0]["partita"] == "23FE01/000001"
    assert main_rows[0]["data pagamento"] == "2023-02-10"
    assert main_rows[0]["stato riscontro"] == "Probabile, da verificare"
    assert main_rows[0]["confidenza"] == "Media"
    assert main_rows[0]["non capita"] == "SI"
    assert detail_rows[0]["tipo evidenza"] == "Banca candidata"
    assert legend_rows[0]["campo"] == "Scopo"


def test_write_accountant_report_workbook_creates_operational_tabs(tmp_path):
    report = load_accountant_report()
    path = tmp_path / "scheda_operativa_commercialista.xlsx"

    report.write_accountant_report_workbook(
        path,
        [
            {
                "record_id": "open-1",
                "document_no": "23FE01/000001",
                "document_key": "1FE|2023",
                "amount": "100.00",
                "reconciliation_status": "closed",
                "evidence_level": "strong_external",
                "matched_evidence_type": "external_bank",
                "matched_evidence_id": "bank-1",
            }
        ],
        normalized_records=[
            {
                "record_id": "bank-1",
                "document_key": "1FE|2023",
                "evidence_type": "external_bank",
                "source_role": "bank_statement",
                "source_file": "bank.pdf",
                "posting_date": "2023-02-10",
                "amount": "100.00",
                "description": "Bonifico fattura 1",
            }
        ],
    )

    workbook = load_workbook(path)
    assert workbook.sheetnames == [
        "Legenda",
        "Scheda operativa",
        "Dettaglio riscontri",
    ]
    main_headers = [cell.value for cell in workbook["Scheda operativa"][1]]
    main_values = [cell.value for cell in workbook["Scheda operativa"][2]]
    assert "data pagamento" in main_headers
    assert "azione richiesta" in main_headers
    assert "Riscontro forte" in main_values
    assert workbook["Scheda operativa"].freeze_panes == "A2"
