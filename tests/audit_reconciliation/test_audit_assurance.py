from __future__ import annotations

import copy
import hashlib
import importlib._bootstrap_external
import importlib.util
import json
import os
import shutil
import stat
import subprocess
import sys
from pathlib import Path

import pytest
from docx import Document
from openpyxl import Workbook, load_workbook

sys.dont_write_bytecode = True
sys.pycache_prefix = "/dev/null/audit-reconciliation-tests"

ROOT = Path(__file__).resolve().parents[2]
SCRIPTS = ROOT / "plugins" / "audit-reconciliation" / "scripts"
ASSURANCE_PATH = SCRIPTS / "audit_assurance.py"
WORKFLOW_PATH = SCRIPTS / "reconciliation_workflow.py"
REVIEW_SERVER_PATH = SCRIPTS / "review_server.py"
PLUGIN_ROOT = SCRIPTS.parent
SHARED_ASSURANCE_ROOT = (
    PLUGIN_ROOT.parent / "_shared" / "vendor" / "modules" / "vera_assurance"
)


def running_audit_context(tmp_path: Path) -> tuple[Path, dict[str, object]]:
    ledger = load_script_module(
        f"audit_reconciliation_customer_ledger_{tmp_path.name}",
        ROOT / "plugins" / "studio-archive" / "scripts" / "client_ledger.py",
    )
    client_root = tmp_path / "Audit Customer"
    client_root.mkdir()
    client_id = "client_333333333333333333333333"
    ledger.create_client_manifest(client_root, client_id)
    engagement = ledger.create_engagement(client_root, client_id, "Audit review")
    source = tmp_path / "audit-source.txt"
    source.write_text("audit source\n", encoding="utf-8")
    imported = ledger.import_document(
        client_root,
        client_id,
        engagement["engagement_id"],
        source,
        "source",
    )
    prepared = ledger.prepare_run(
        client_root,
        client_id,
        engagement["engagement_id"],
        "audit-reconciliation",
        "test-version",
        input_ids=[imported["receipt"]["input_id"]],
    )
    running = ledger.start_run(
        client_root,
        engagement["engagement_id"],
        prepared["run"]["run_id"],
    )
    context_path = Path(running["run_root"]) / "context.json"
    return context_path, running["context"]


def load_assurance():
    spec = importlib.util.spec_from_file_location(
        "audit_reconciliation_assurance",
        ASSURANCE_PATH,
    )
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def load_script_module(name: str, path: Path):
    if str(SCRIPTS) not in sys.path:
        sys.path.insert(0, str(SCRIPTS))
    spec = importlib.util.spec_from_file_location(name, path)
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def load_retained_module(module_name: str):
    load_script_module(
        f"audit_reconciliation_retained_loader_{module_name}",
        WORKFLOW_PATH,
    )
    return sys.modules[module_name]


def reviewed_source_decision() -> dict[str, object]:
    return {
        "role": "open_items",
        "adapter_family": "open_items_text_v1",
        "reviewer_ref": "reviewer.test",
        "reviewed_on": "2026-07-25",
        "perimeter": {
            "entity_ref": "entity.test",
            "party_ref": "party.test",
            "currency": "EUR",
            "unit": "currency_amount",
            "direction_policy": "customer",
            "allocation_policy": "one_to_one",
        },
        "money": {
            "decimal_separator": ".",
            "thousands_separator": ",",
            "reported_unit": "EUR",
            "reported_increment": "0.01",
        },
        "date": {"order": "day_first"},
    }


def source_rows() -> list[dict[str, object]]:
    return [
        {
            "record_id": "open-1",
            "source_file": "source.csv",
            "source_row": 2,
            "source_role": "open_items",
            "document_key": "INV-1|2026",
            "amount": "100.00",
            "currency": "EUR",
            "unit": "currency_amount",
            "entity_ref": "entity.test",
            "party_ref": "party.test",
            "direction_policy": "customer",
            "allocation_policy": "one_to_one",
            "reported_unit": "EUR",
            "reported_increment": "0.01",
        },
        {
            "record_id": "open-2",
            "source_file": "source.csv",
            "source_row": 3,
            "source_role": "open_items",
            "document_key": "INV-2|2026",
            "amount": "25.50",
            "currency": "EUR",
            "unit": "currency_amount",
            "entity_ref": "entity.test",
            "party_ref": "party.test",
            "direction_policy": "customer",
            "allocation_policy": "one_to_one",
            "reported_unit": "EUR",
            "reported_increment": "0.01",
        },
    ]


def workbook_bytes(tmp_path: Path) -> bytes:
    path = tmp_path / "template.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reconciliation detail"
    sheet.append(["record_id", "amount"])
    sheet.append(["open-1", "100.00"])
    sheet.append(["open-2", "25.50"])
    workbook.save(path)
    return path.read_bytes()


def stable_write(path: Path, value: object) -> None:
    path.write_text(
        json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def exact_tree_image(root: Path) -> dict[str, tuple[object, ...]]:
    image: dict[str, tuple[object, ...]] = {}
    for path in [root, *sorted(root.rglob("*"))]:
        current = path.lstat()
        relative = "." if path == root else path.relative_to(root).as_posix()
        mode = current.st_mode & 0o7777
        if stat.S_ISDIR(current.st_mode):
            image[relative] = ("directory", mode)
        elif stat.S_ISREG(current.st_mode):
            image[relative] = ("file", mode, current.st_nlink, path.read_bytes())
        elif stat.S_ISLNK(current.st_mode):
            image[relative] = ("symlink", mode, os.readlink(path))
        else:
            image[relative] = ("special", stat.S_IFMT(current.st_mode), mode)
    return image


def copied_implementation_tree(tmp_path: Path) -> Path:
    plugin_copy = tmp_path / "audit-reconciliation"
    for directory in ("assets", "mcp", "scripts"):
        shutil.copytree(
            PLUGIN_ROOT / directory,
            plugin_copy / directory,
            ignore=shutil.ignore_patterns("__pycache__", "*.pyc", "*.pyo"),
        )
    shutil.copytree(
        SHARED_ASSURANCE_ROOT,
        plugin_copy / "vendor" / "modules" / "vera_assurance",
        ignore=shutil.ignore_patterns("__pycache__", "*.pyc", "*.pyo"),
    )
    return plugin_copy


def reseal_outer(assurance: object, payload: dict[str, object]) -> None:
    content = {key: payload[key] for key in payload if key != "content_sha256"}
    payload["content_sha256"] = assurance.canonical_json_sha256(content)


def reseal_numeric(assurance: object, payload: dict[str, object]) -> None:
    content = {
        "schema_version": payload["schema_version"],
        "ledger_id": payload["ledger_id"],
        "entries": payload["entries"],
    }
    payload["content_sha256"] = assurance.canonical_json_sha256(content)


def prepared_case(
    tmp_path: Path,
) -> tuple[object, Path, Path, list[dict[str, object]], dict[str, object]]:
    assurance = load_assurance()
    source_dir = tmp_path / "source"
    output_dir = tmp_path / "output"
    source_dir.mkdir(parents=True)
    output_dir.mkdir(parents=True)
    source_path = source_dir / "source.csv"
    source_path.write_text("document,amount\nINV-1,100.00\nINV-2,25.50\n")
    receipts = assurance.build_source_receipts(source_dir, [source_path])
    decisions, errors = assurance.build_reviewed_source_decisions(
        input_root=source_dir,
        source_receipts=receipts,
        adapter_families={"source.csv": "open_items_text_v1"},
        assumptions={
            "reviewed_source_decisions": {"source.csv": reviewed_source_decision()}
        },
    )
    assert errors == {}
    rows = source_rows()
    source_qualifications = qualified_sources(
        assurance,
        {
            "source_receipts": receipts,
            "reviewed_source_decisions": list(decisions.values()),
        },
    )
    context = assurance.prepare_assurance_run(
        output_dir=output_dir,
        open_items=rows,
        evidence_rows=[],
        assumptions={"currency": "EUR", "amount_tolerance": "0"},
        source_root=source_dir,
        source_receipts=receipts,
        reviewed_source_decisions=list(decisions.values()),
        source_qualifications=source_qualifications,
    )
    return assurance, source_path, output_dir, rows, context


def qualified_sources(
    assurance: object,
    context: dict[str, object],
) -> list[dict[str, object]]:
    source_receipt = context["source_receipts"][0]
    decision = context["reviewed_source_decisions"][0]
    return [
        assurance.validate_source_qualification(
            {
                "schema_version": "vera.source_qualification.v1",
                "qualification_id": "qualification.test",
                "adapter_id": "open_items_text_v1",
                "adapter_version": "2",
                "source_family": "csv.open_items",
                "status": "qualified",
                "source_artifact_refs": [source_receipt["artifact_id"]],
                "controls": [
                    {
                        "control_id": "reviewed_mapping",
                        "required": True,
                        "status": "passed",
                        "evidence_refs": [source_receipt["artifact_id"]],
                        "detail": "Reviewed mapping and test adapter passed.",
                    }
                ],
                "candidate_row_count": 2,
                "emitted_row_count": 2,
                "reviewed_mapping_ref": decision["decision_id"],
                "limitations": [],
            }
        )
    ]


def finalize_case(
    tmp_path: Path,
    *,
    review_status: str = "PASS",
    allocation_ledgers: list[dict[str, object]] | None = None,
    workbook_content: bytes | None = None,
    source_processing: dict[str, object] | None = None,
    analyses: dict[str, object] | None = None,
) -> tuple[object, Path, dict[str, object]]:
    assurance, _, output_dir, rows, context = prepared_case(tmp_path)
    workbook_path = output_dir / "audit.xlsx"
    workbook_path.write_bytes(workbook_content or workbook_bytes(tmp_path))
    report_path = output_dir / "report.txt"
    report_path.write_text("Audit reconciliation report\n")
    payload = assurance.finalize_assurance_run(
        output_dir=output_dir,
        context=context,
        reconciliation_rows=rows,
        allocation_ledgers=allocation_ledgers or [],
        checks=[{"check": "row_count", "status": "PASS"}],
        review_rows=[
            {
                "record_id": row["record_id"],
                "review_status": review_status,
                **(
                    {
                        "reviewer_ref": "reviewer.test",
                        "reviewed_on": "2026-07-25",
                    }
                    if review_status == "PASS"
                    else {}
                ),
            }
            for row in rows
        ],
        source_qualifications=qualified_sources(assurance, context),
        declared_outputs=[workbook_path, report_path],
        workbook_name=workbook_path.name,
        source_processing=source_processing,
        analyses=analyses,
    )
    return assurance, output_dir, payload


def test_canonical_result_consolidates_schedules_and_artifact_contract(tmp_path):
    assurance = load_assurance()
    source_processing = {field: [] for field in assurance.SOURCE_PROCESSING_FIELDS}
    source_processing["extraction_errors"] = [
        {
            "source_file": "unsupported.csv",
            "status": "unsupported_source_layout",
        }
    ]
    analyses = {field: [] for field in assurance.ANALYSIS_FIELDS}
    analyses["aging_summary"] = [{"aging_bucket": "0-30", "amount_total": "100.00"}]

    _, output_dir, _ = finalize_case(
        tmp_path,
        source_processing=source_processing,
        analyses=analyses,
    )

    canonical = json.loads(
        (
            output_dir / "assurance_final_outputs" / "reconciliation_results.json"
        ).read_text(encoding="utf-8")
    )
    assert (
        canonical["schema_version"] == assurance.RECONCILIATION_RESULTS_SCHEMA_VERSION
    )
    assert canonical["source_processing"] == source_processing
    assert canonical["analyses"] == analyses
    assert set(assurance.RUN_ROOT_ARTIFACT_CONTRACT) == set(
        assurance.RUN_ROOT_FIXED_FILES
    )
    assert all(
        record["audience"] and record["purpose"]
        for record in assurance.RUN_ROOT_ARTIFACT_CONTRACT.values()
    )
    removed_sidecars = {
        "aging_summary.json",
        "bank_allocation_candidates.json",
        "cutoff_window_movements.json",
        "document_source_map.json",
        "evidence_concentration.json",
        "external_evidence_detail.json",
        "external_evidence_summary.json",
        "extraction_errors.json",
        "journal_rollforward_rows.json",
        "journal_rollforward_summary.json",
        "ledger_balance_rows.json",
        "post_cutoff_candidates.json",
        "relationship_allocation_ledgers.json",
        "reversal_candidates.json",
        "review_signals.json",
        "source_qualifications.json",
    }
    assert removed_sidecars.isdisjoint(assurance.RUN_ROOT_FIXED_FILES)


def assured_browser_case(tmp_path: Path) -> tuple[object, Path, dict[str, object]]:
    workflow = load_script_module(
        f"audit_reconciliation_workflow_{tmp_path.name}",
        WORKFLOW_PATH,
    )
    review_server = load_script_module(
        f"audit_reconciliation_review_server_{tmp_path.name}",
        REVIEW_SERVER_PATH,
    )
    output_dir = tmp_path / "output"
    workflow.build_reconciliation_artifacts(
        output_dir=output_dir,
        open_items=[
            {
                "record_id": "open-1",
                "document_key": "INV-1|2026",
                "document_no": "INV-1",
                "document_date": "2026-01-01",
                "amount": "100.00",
                "currency": "EUR",
            }
        ],
        evidence_rows=[],
        assumptions={"scope_year": "2026", "amount_tolerance": "0"},
        require_completed_review=False,
    )
    review_payload = json.loads(
        (output_dir / "review_payload.json").read_text(encoding="utf-8")
    )
    expected_predecessor_checkpoint = json.loads(
        (output_dir / "assurance_receipts.json").read_text(encoding="utf-8")
    )["content_sha256"]
    decisions = [
        {"item_id": item["id"], "action": "accept"} for item in review_payload["items"]
    ]
    return (
        review_server,
        output_dir,
        {
            "decisions": decisions,
            "expected_predecessor_checkpoint": expected_predecessor_checkpoint,
        },
    )


def successor_lifecycle_case(
    tmp_path: Path,
) -> tuple[
    object,
    object,
    Path,
    dict[str, object],
    dict[str, bytes],
    dict[str, object],
]:
    assurance = load_assurance()
    workflow = load_script_module(
        f"audit_reconciliation_successor_workflow_{tmp_path.name}",
        WORKFLOW_PATH,
    )
    review_server = load_script_module(
        f"audit_reconciliation_successor_review_server_{tmp_path.name}",
        REVIEW_SERVER_PATH,
    )
    output_dir = tmp_path / "output"
    open_items = [
        {
            "record_id": "open-1",
            "document_key": "INV-1|2026",
            "document_no": "INV-1",
            "document_date": "2026-01-01",
            "amount": "100.00",
            "currency": "EUR",
        }
    ]
    workflow_args: dict[str, object] = {
        "output_dir": output_dir,
        "open_items": open_items,
        "evidence_rows": [],
        "assumptions": {
            "scope_year": "2026",
            "amount_tolerance": "0",
            "assurance_run_date": "2026-07-25",
        },
        "review_rows": [
            {
                "record_id": "open-1",
                "review_status": "PENDING",
                "reviewer_ref": "",
                "reviewed_on": "",
            }
        ],
        "require_completed_review": False,
        "fail_on_check_errors": False,
        "language": "en",
    }
    workflow.build_reconciliation_artifacts(**workflow_args)
    predecessor_bytes = {
        "assurance": (output_dir / "assurance_receipts.json").read_bytes(),
        "professional": (output_dir / "professional_review.json").read_bytes(),
        "reconciliation": (
            output_dir / "assurance_final_outputs" / "reconciliation_results.json"
        ).read_bytes(),
        "review_payload": (output_dir / "review_payload.json").read_bytes(),
    }
    review_payload = json.loads(predecessor_bytes["review_payload"])
    expected_predecessor_checkpoint = json.loads(predecessor_bytes["assurance"])[
        "content_sha256"
    ]
    decisions = {
        "decisions": [
            {"item_id": item["id"], "action": "accept"}
            for item in review_payload["items"]
        ],
        "expected_predecessor_checkpoint": expected_predecessor_checkpoint,
    }
    first_apply = review_server.apply_decisions(output_dir, decisions)
    authority = json.loads(
        (output_dir / "professional_review.json").read_text(encoding="utf-8")
    )
    workflow_args["review_rows"] = authority["records"]
    workflow_args["expected_predecessor_checkpoint"] = expected_predecessor_checkpoint
    workflow.build_reconciliation_artifacts(**workflow_args)
    return (
        assurance,
        review_server,
        output_dir,
        decisions,
        predecessor_bytes,
        first_apply,
    )


def refreshed_artifact_receipt(
    assurance: object,
    root: Path,
    path: Path,
    receipt: dict[str, object],
) -> dict[str, object]:
    return assurance.artifact_receipt(
        root,
        path,
        artifact_id=receipt["artifact_id"],
        role=receipt["role"],
        root_id=receipt["root_id"],
        **({"media_type": receipt["media_type"]} if "media_type" in receipt else {}),
    )


def fully_resealed_contradictory_predecessor(
    tmp_path: Path,
) -> tuple[object, Path, str]:
    """Build the former complete-reseal bypass with synthetic run data."""

    assurance, _, output_dir, _, _, _ = successor_lifecycle_case(tmp_path)
    seal_path = output_dir / "assurance_receipts.json"
    seal = json.loads(seal_path.read_text(encoding="utf-8"))
    old_digest = seal["professional_review_authority"]["predecessor_assurance_sha256"]
    history_root = output_dir / "assurance_transition_history"
    old_history = history_root / old_digest
    predecessor_path = old_history / "predecessor_assurance_receipts.json"
    snapshot_assurance_path = (
        old_history / "predecessor_run" / "assurance_receipts.json"
    )
    predecessor = json.loads(predecessor_path.read_text(encoding="utf-8"))
    predecessor["run_date"] = "2026-07-24"
    reseal_outer(assurance, predecessor)
    new_digest = predecessor["content_sha256"]
    stable_write(predecessor_path, predecessor)
    stable_write(snapshot_assurance_path, predecessor)

    predecessor_professional = json.loads(
        (old_history / "predecessor_professional_review.json").read_text(
            encoding="utf-8"
        )
    )
    predecessor_reconciliation = json.loads(
        (old_history / "predecessor_reconciliation_results.json").read_text(
            encoding="utf-8"
        )
    )
    mapping = json.loads(
        (old_history / "review_payload_mapping.json").read_text(encoding="utf-8")
    )
    applied = json.loads(
        (old_history / "applied_decisions.json").read_text(encoding="utf-8")
    )
    successor = assurance.build_applied_review_authority(
        predecessor_assurance=predecessor,
        predecessor_professional_review=predecessor_professional,
        predecessor_reconciliation=predecessor_reconciliation,
        review_payload_mapping=mapping,
        effects=applied["effects"],
        reviewer_ref=applied.get("reviewer"),
    )
    stable_write(old_history / "successor_professional_review.json", successor)
    professional_path = output_dir / "professional_review.json"
    stable_write(professional_path, successor)

    result_path = output_dir / "assurance_final_outputs" / "reconciliation_results.json"
    result = json.loads(result_path.read_text(encoding="utf-8"))
    result["review_rows"] = successor["records"]
    stable_write(result_path, result)

    transition_path = old_history / "transition_receipt.json"
    transition = json.loads(transition_path.read_text(encoding="utf-8"))
    refreshed_transition_receipts = [
        refreshed_artifact_receipt(
            assurance,
            old_history,
            old_history / receipt["path"],
            receipt,
        )
        for receipt in transition["artifact_receipts"]
    ]
    transition_content = {
        "schema_version": transition["schema_version"],
        "transition_id": f"review_transition.{new_digest}",
        "predecessor_assurance_sha256": new_digest,
        "decision_fingerprint": successor["decision_fingerprint"],
        "successor_professional_review_sha256": successor["content_sha256"],
        "artifact_receipts": refreshed_transition_receipts,
    }
    transition = {
        **transition_content,
        "content_sha256": assurance.canonical_json_sha256(transition_content),
    }
    stable_write(transition_path, transition)
    new_history = history_root / new_digest
    old_history.rename(new_history)

    seal["professional_review_authority"] = successor
    seal["professional_review_receipt"] = refreshed_artifact_receipt(
        assurance,
        output_dir,
        professional_path,
        seal["professional_review_receipt"],
    )
    seal["review_transition_receipts"] = [transition]

    final_root = output_dir / "assurance_final_outputs"
    inventory = seal["final_output_inventory"]
    inventory_content = {
        "schema_version": inventory["schema_version"],
        "boundary_root": inventory["boundary_root"],
        "declared_paths": inventory["declared_paths"],
        "artifact_receipts": [
            refreshed_artifact_receipt(
                assurance,
                final_root,
                final_root / receipt["path"],
                receipt,
            )
            for receipt in inventory["artifact_receipts"]
        ],
    }
    inventory = {
        **inventory_content,
        "content_sha256": assurance.canonical_json_sha256(inventory_content),
    }
    seal["final_output_inventory"] = inventory
    stable_write(output_dir / "final_output_inventory.json", inventory)

    for entry in seal["run_tree_contract"]["entries"]:
        entry["path"] = entry["path"].replace(old_digest, new_digest)
    seal["run_tree_contract"]["entries"].sort(key=lambda entry: entry["path"])
    tree_content = {
        "schema_version": seal["run_tree_contract"]["schema_version"],
        "entries": seal["run_tree_contract"]["entries"],
    }
    seal["run_tree_contract"]["content_sha256"] = assurance.canonical_json_sha256(
        tree_content
    )
    reseal_outer(assurance, seal)
    stable_write(seal_path, seal)
    return assurance, output_dir, old_digest


def test_implementation_receipts_match_exact_ordered_25_file_contract() -> None:
    assurance = load_assurance()

    receipts = assurance.build_implementation_receipts()

    contract = [(item["root_id"], item["path"]) for item in receipts]
    assert contract == list(assurance.IMPLEMENTATION_CONTRACT)
    assert len(contract) == 25
    assert (
        hashlib.sha256(
            "\n".join(f"{root_id}:{path}" for root_id, path in contract).encode()
        ).hexdigest()
        == "11e5226f586e22f0222ac36eba82c3937daf4185d0c53d66d3ea744e2c968c41"
    )


def test_professional_authority_marks_reviewer_ref_unsigned_and_untrusted() -> None:
    assurance = load_assurance()

    authority = assurance.build_professional_review_authority(
        [
            {
                "record_id": "open-1",
                "review_status": "PASS",
                "reviewer_ref": "reviewer.test",
                "reviewed_on": "2026-07-25",
            }
        ]
    )

    assert authority["reviewer_ref_trust"] == "unsigned_untrusted_label"
    assert assurance.validate_professional_review_authority(authority) == authority


@pytest.mark.parametrize("rogue_kind", ["file", "directory"])
def test_implementation_contract_rejects_expanded_tree(
    tmp_path: Path,
    rogue_kind: str,
) -> None:
    assurance = load_assurance()
    plugin_copy = tmp_path / "audit-reconciliation"
    for directory in ("assets", "mcp", "scripts"):
        shutil.copytree(
            PLUGIN_ROOT / directory,
            plugin_copy / directory,
            ignore=shutil.ignore_patterns("__pycache__", "*.pyc", "*.pyo"),
        )
    rogue = plugin_copy / "scripts" / "rogue"
    if rogue_kind == "file":
        rogue.write_text("rogue", encoding="utf-8")
    else:
        rogue.mkdir()

    with pytest.raises(ValueError, match="exact.*contract"):
        assurance.build_implementation_receipts(plugin_copy)


@pytest.mark.parametrize(
    "substitution",
    ["parent_symlink", "file_symlink", "file_hardlink", "file_fifo"],
)
def test_implementation_contract_rejects_link_or_special_substitution(
    tmp_path: Path,
    substitution: str,
) -> None:
    assurance = load_assurance()
    plugin_copy = tmp_path / "audit-reconciliation"
    for directory in ("assets", "mcp", "scripts"):
        shutil.copytree(
            PLUGIN_ROOT / directory,
            plugin_copy / directory,
            ignore=shutil.ignore_patterns("__pycache__", "*.pyc", "*.pyo"),
        )
    target = plugin_copy / "assets" / "icon.svg"
    if substitution == "parent_symlink":
        assets_real = plugin_copy / "assets-real"
        (plugin_copy / "assets").rename(assets_real)
        (plugin_copy / "assets").symlink_to(assets_real, target_is_directory=True)
    else:
        original = tmp_path / "original.svg"
        shutil.copyfile(target, original)
        target.unlink()
        if substitution == "file_symlink":
            target.symlink_to(original)
        elif substitution == "file_hardlink":
            os.link(original, target)
        else:
            os.mkfifo(target)

    with pytest.raises(ValueError, match="ordinary|real directory"):
        assurance.build_implementation_receipts(plugin_copy)


@pytest.mark.parametrize(
    "rogue_kind",
    ["regular", "empty_directory", "symlink", "hardlink", "fifo"],
)
@pytest.mark.parametrize("cache_root", ["scripts", "shared_assurance"])
def test_implementation_contract_rejects_every_cache_namespace_entry(
    tmp_path: Path,
    rogue_kind: str,
    cache_root: str,
) -> None:
    assurance = load_assurance()
    plugin_copy = copied_implementation_tree(tmp_path)
    selected_root = (
        plugin_copy / "scripts"
        if cache_root == "scripts"
        else plugin_copy / "vendor" / "modules" / "vera_assurance"
    )
    cache_dir = selected_root / "__pycache__"
    cache_dir.mkdir()
    rogue = cache_dir / "rogue"
    link_target = tmp_path / "link-target"
    link_target.write_bytes(b"target")
    if rogue_kind == "regular":
        rogue.write_bytes(b"rogue")
    elif rogue_kind == "empty_directory":
        rogue.mkdir()
    elif rogue_kind == "symlink":
        rogue.symlink_to(link_target)
    elif rogue_kind == "hardlink":
        os.link(link_target, rogue)
    elif rogue_kind == "fifo":
        os.mkfifo(rogue)
    if cache_root == "scripts":
        with pytest.raises(ValueError, match="implementation"):
            assurance.build_implementation_receipts(plugin_copy)
    else:
        completed = subprocess.run(
            [
                sys.executable,
                (plugin_copy / "scripts" / "audit_assurance.py").as_posix(),
                "--help",
            ],
            check=False,
            capture_output=True,
            text=True,
        )
        assert completed.returncode != 0
        assert "implementation" in completed.stderr


@pytest.mark.parametrize(
    "module_name",
    ["audit_assurance", "reconciliation_helpers"],
)
@pytest.mark.parametrize(
    "entrypoint_name",
    ["audit_assurance.py", "review_server.py"],
)
def test_public_python_entrypoints_reject_timestamp_valid_local_bytecode_before_import(
    tmp_path: Path,
    module_name: str,
    entrypoint_name: str,
) -> None:
    plugin_copy = copied_implementation_tree(tmp_path)
    target_source = plugin_copy / "scripts" / f"{module_name}.py"
    metadata_source = (
        target_source
        if target_source.exists()
        else plugin_copy / "scripts" / "retained_sources" / f"{module_name}.source"
    )
    cache_dir = target_source.parent / "__pycache__"
    cache_dir.mkdir()
    marker = tmp_path / f"{module_name}-bytecode-executed"
    source_stat = metadata_source.stat()
    code = compile(
        (
            "from pathlib import Path\n"
            f"Path({marker.as_posix()!r}).write_text('executed', encoding='utf-8')\n"
        ),
        target_source.as_posix(),
        "exec",
    )
    pyc = importlib._bootstrap_external._code_to_timestamp_pyc(
        code,
        int(source_stat.st_mtime),
        source_stat.st_size,
    )
    cache_path = cache_dir / f"{module_name}.{sys.implementation.cache_tag}.pyc"
    cache_path.write_bytes(pyc)
    environment = dict(os.environ)
    environment.pop("PYTHONDONTWRITEBYTECODE", None)
    environment.pop("PYTHONPYCACHEPREFIX", None)

    completed = subprocess.run(
        [
            sys.executable,
            (plugin_copy / "scripts" / entrypoint_name).as_posix(),
            "--help",
        ],
        check=False,
        capture_output=True,
        text=True,
        env=environment,
    )

    assert completed.returncode != 0
    assert not marker.exists()
    assert "exact 25-file contract" in completed.stderr


@pytest.mark.parametrize(
    "module_name",
    [
        "accountant_report",
        "locale_support",
        "reconciliation_helpers",
        "review_session",
        "workpaper_outputs",
    ],
)
def test_ordinary_import_cannot_resolve_retained_source_or_execute_local_bytecode(
    tmp_path: Path,
    module_name: str,
) -> None:
    plugin_copy = copied_implementation_tree(tmp_path)
    target_source = plugin_copy / "scripts" / f"{module_name}.py"
    retained_source = (
        plugin_copy / "scripts" / "retained_sources" / f"{module_name}.source"
    )
    assert not target_source.exists()
    assert retained_source.is_file()
    cache_dir = target_source.parent / "__pycache__"
    cache_dir.mkdir()
    marker = tmp_path / f"{module_name}-direct-bytecode-executed"
    source_stat = retained_source.stat()
    code = compile(
        (
            "from pathlib import Path\n"
            f"Path({marker.as_posix()!r}).write_text('executed', encoding='utf-8')\n"
        ),
        target_source.as_posix(),
        "exec",
    )
    pyc = importlib._bootstrap_external._code_to_timestamp_pyc(
        code,
        int(source_stat.st_mtime),
        source_stat.st_size,
    )
    (cache_dir / f"{module_name}.{sys.implementation.cache_tag}.pyc").write_bytes(pyc)
    environment = dict(os.environ)
    environment.pop("PYTHONDONTWRITEBYTECODE", None)
    environment.pop("PYTHONPYCACHEPREFIX", None)
    environment["PYTHONPATH"] = target_source.parent.as_posix()

    completed = subprocess.run(
        [sys.executable, "-c", f"import {module_name}"],
        cwd=target_source.parent,
        check=False,
        capture_output=True,
        text=True,
        env=environment,
    )

    assert completed.returncode != 0
    assert not marker.exists()
    assert "ModuleNotFoundError" in completed.stderr


@pytest.mark.parametrize("substitution", ["symlink", "hardlink"])
def test_public_entrypoint_rejects_unsafe_bootstrap_before_reading_it(
    tmp_path: Path,
    substitution: str,
) -> None:
    plugin_copy = copied_implementation_tree(tmp_path)
    bootstrap = plugin_copy / "scripts" / "implementation_bootstrap.py"
    marker = tmp_path / "bootstrap-executed"
    malicious = tmp_path / "malicious-bootstrap.py"
    malicious.write_text(
        (
            "from pathlib import Path\n"
            f"Path({marker.as_posix()!r}).write_text('executed', encoding='utf-8')\n"
        ),
        encoding="utf-8",
    )
    bootstrap.unlink()
    if substitution == "symlink":
        bootstrap.symlink_to(malicious)
    else:
        os.link(malicious, bootstrap)

    completed = subprocess.run(
        [
            sys.executable,
            (plugin_copy / "scripts" / "review_server.py").as_posix(),
            "--help",
        ],
        check=False,
        capture_output=True,
        text=True,
    )

    assert completed.returncode != 0
    assert not marker.exists()
    assert "ordinary single-link regular file" in completed.stderr


@pytest.mark.parametrize("attack", ["missing", "expanded", "reordered"])
def test_assurance_replay_rejects_nonexact_implementation_receipt_list(
    tmp_path: Path,
    attack: str,
) -> None:
    assurance, output_dir, _ = finalize_case(tmp_path)
    seal_path = output_dir / "assurance_receipts.json"
    seal = json.loads(seal_path.read_text(encoding="utf-8"))
    receipts = seal["implementation_receipts"]
    if attack == "missing":
        receipts.pop()
    elif attack == "reordered":
        seal["implementation_receipts"] = list(reversed(receipts))
    else:
        receipts.append(
            assurance.artifact_receipt(
                PLUGIN_ROOT,
                PLUGIN_ROOT / "README.md",
                artifact_id="implementation.expanded_readme",
                role="implementation",
                root_id="plugin",
                media_type="text/markdown",
            )
        )
    reseal_outer(assurance, seal)
    stable_write(seal_path, seal)

    with pytest.raises(ValueError, match="implementation receipts"):
        assurance.validate_assurance_run(output_dir)


@pytest.mark.parametrize("location", ["root", "final_boundary"])
def test_assurance_replay_rejects_rogue_empty_directory(
    tmp_path: Path,
    location: str,
) -> None:
    assurance, output_dir, _ = finalize_case(tmp_path)
    parent = (
        output_dir if location == "root" else output_dir / "assurance_final_outputs"
    )
    (parent / "rogue-empty").mkdir()

    with pytest.raises(ValueError, match="closure|unsafe entry"):
        assurance.validate_assurance_run(output_dir)


@pytest.mark.parametrize(
    "entry_kind",
    ["regular", "symlink", "hardlink", "fifo"],
)
def test_assurance_replay_rejects_unrelated_root_entry(
    tmp_path: Path,
    entry_kind: str,
) -> None:
    assurance, output_dir, _ = finalize_case(tmp_path)
    rogue = output_dir / "rogue-entry"
    external = tmp_path / "external-entry"
    external.write_text("external", encoding="utf-8")
    if entry_kind == "regular":
        rogue.write_text("rogue", encoding="utf-8")
    elif entry_kind == "symlink":
        rogue.symlink_to(external)
    elif entry_kind == "hardlink":
        os.link(external, rogue)
    else:
        os.mkfifo(rogue)

    with pytest.raises(ValueError):
        assurance.validate_assurance_run(output_dir)


def test_assurance_replay_rejects_self_expanded_root_tree_contract(
    tmp_path: Path,
) -> None:
    assurance, output_dir, _ = finalize_case(tmp_path)
    (output_dir / "rogue.json").write_text("{}\n", encoding="utf-8")
    seal_path = output_dir / "assurance_receipts.json"
    seal = json.loads(seal_path.read_text(encoding="utf-8"))
    tree = seal["run_tree_contract"]
    tree["entries"].append({"path": "rogue.json", "kind": "file"})
    tree["entries"].sort(key=lambda item: item["path"])
    tree_content = {
        "schema_version": tree["schema_version"],
        "entries": tree["entries"],
    }
    tree["content_sha256"] = assurance.canonical_json_sha256(tree_content)
    reseal_outer(assurance, seal)
    stable_write(seal_path, seal)

    with pytest.raises(ValueError, match="unrelated entry"):
        assurance.validate_assurance_run(output_dir)


def test_browser_save_rejects_rogue_root_directory_without_writes(
    tmp_path: Path,
) -> None:
    review_server, output_dir, decisions = assured_browser_case(tmp_path)
    (output_dir / "rogue-empty").mkdir()
    before = exact_tree_image(output_dir)

    with pytest.raises(ValueError, match="closure"):
        review_server.save_decisions(output_dir, decisions)

    assert exact_tree_image(output_dir) == before


def test_browser_save_failure_restores_exact_whole_tree(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    review_server, output_dir, decisions = assured_browser_case(tmp_path)
    before = exact_tree_image(output_dir)
    root_inode = output_dir.stat().st_ino
    decisions_inode = (output_dir / "ui_decisions.json").stat().st_ino
    original_write = review_server._write_json

    def fail_after_write(path: Path, payload: dict[str, object]) -> None:
        original_write(path, payload)
        raise OSError("injected save failure")

    monkeypatch.setattr(review_server, "_write_json", fail_after_write)

    with pytest.raises(OSError, match="injected save failure"):
        review_server.save_decisions(output_dir, decisions)

    assert exact_tree_image(output_dir) == before
    assert output_dir.stat().st_ino == root_inode
    assert (output_dir / "ui_decisions.json").stat().st_ino == decisions_inode


@pytest.mark.parametrize("checkpoint_case", ["missing", "wrong"])
def test_browser_apply_rejects_untrusted_predecessor_checkpoint_without_writes(
    tmp_path: Path,
    checkpoint_case: str,
) -> None:
    review_server, output_dir, decisions = assured_browser_case(tmp_path)
    supplied = copy.deepcopy(decisions)
    if checkpoint_case == "missing":
        supplied.pop("expected_predecessor_checkpoint")
        expected_message = "external expected predecessor checkpoint is required"
    else:
        current = str(supplied["expected_predecessor_checkpoint"])
        supplied["expected_predecessor_checkpoint"] = (
            "f" * 64 if current != "f" * 64 else "e" * 64
        )
        expected_message = "external expected predecessor checkpoint does not match"
    before = exact_tree_image(output_dir)

    with pytest.raises(ValueError, match=expected_message):
        review_server.apply_decisions(output_dir, supplied)

    assert exact_tree_image(output_dir) == before
    assert not list(tmp_path.glob(".audit-review-transaction-*"))


def test_browser_apply_failure_restores_exact_whole_tree(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    review_server, output_dir, decisions = assured_browser_case(tmp_path)
    before = exact_tree_image(output_dir)
    root_inode = output_dir.stat().st_ino
    decisions_inode = (output_dir / "ui_decisions.json").stat().st_ino
    original_write = review_server._write_json
    write_count = 0

    def fail_after_second_write(path: Path, payload: dict[str, object]) -> None:
        nonlocal write_count
        original_write(path, payload)
        write_count += 1
        if write_count == 2:
            raise OSError("injected apply failure")

    monkeypatch.setattr(review_server, "_write_json", fail_after_second_write)

    with pytest.raises(OSError, match="injected apply failure"):
        review_server.apply_decisions(output_dir, decisions)

    assert exact_tree_image(output_dir) == before
    assert output_dir.stat().st_ino == root_inode
    assert (output_dir / "ui_decisions.json").stat().st_ino == decisions_inode


def test_browser_transition_retention_failure_restores_exact_whole_tree(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    review_server, output_dir, decisions = assured_browser_case(tmp_path)
    before = exact_tree_image(output_dir)
    root_inode = output_dir.stat().st_ino
    original_retain = review_server.retain_review_transition

    def fail_after_transition(
        output: Path,
        capture: Path,
        **kwargs: object,
    ) -> None:
        original_retain(output, capture, **kwargs)
        raise OSError("injected transition retention failure")

    monkeypatch.setattr(
        review_server,
        "retain_review_transition",
        fail_after_transition,
    )

    with pytest.raises(OSError, match="transition retention failure"):
        review_server.apply_decisions(output_dir, decisions)

    assert exact_tree_image(output_dir) == before
    assert output_dir.stat().st_ino == root_inode
    assert not list(tmp_path.glob(".audit-review-transaction-*"))


def test_stale_source_receipt_blocks_before_prepared_rows(tmp_path: Path) -> None:
    assurance = load_assurance()
    source_dir = tmp_path / "source"
    output_dir = tmp_path / "output"
    source_dir.mkdir()
    output_dir.mkdir()
    source_path = source_dir / "source.csv"
    source_path.write_text("amount\n100.00\n")
    receipts = assurance.build_source_receipts(source_dir, [source_path])
    source_path.write_text("amount\n999.00\n")

    with pytest.raises(ValueError, match="does not match current bytes"):
        assurance.prepare_assurance_run(
            output_dir=output_dir,
            open_items=source_rows(),
            evidence_rows=[],
            assumptions={},
            source_root=source_dir,
            source_receipts=receipts,
        )

    assert not (output_dir / "prepared_records.json").exists()


def test_stale_source_decision_blocks_before_prepared_rows(tmp_path: Path) -> None:
    assurance = load_assurance()
    source_dir = tmp_path / "source"
    output_dir = tmp_path / "output"
    source_dir.mkdir()
    output_dir.mkdir()
    source_path = source_dir / "source.csv"
    source_path.write_text("amount\n100.00\n")
    old_receipts = assurance.build_source_receipts(source_dir, [source_path])
    decisions, errors = assurance.build_reviewed_source_decisions(
        input_root=source_dir,
        source_receipts=old_receipts,
        adapter_families={"source.csv": "open_items_text_v1"},
        assumptions={
            "reviewed_source_decisions": {"source.csv": reviewed_source_decision()}
        },
    )
    assert errors == {}
    source_path.write_text("amount\n999.00\n")
    current_receipts = assurance.build_source_receipts(source_dir, [source_path])

    with pytest.raises(ValueError, match="stale source identity"):
        assurance.prepare_assurance_run(
            output_dir=output_dir,
            open_items=source_rows(),
            evidence_rows=[],
            assumptions={},
            source_root=source_dir,
            source_receipts=current_receipts,
            reviewed_source_decisions=list(decisions.values()),
        )

    assert not (output_dir / "prepared_records.json").exists()


def test_stale_implementation_receipt_blocks_final_promotion(tmp_path: Path) -> None:
    assurance, _, output_dir, rows, context = prepared_case(tmp_path)
    stale = copy.deepcopy(context)
    stale["implementation_receipts"][0]["sha256"] = "0" * 64
    workbook_path = output_dir / "audit.xlsx"
    workbook_path.write_bytes(workbook_bytes(tmp_path))

    with pytest.raises(ValueError, match="does not match current bytes"):
        assurance.finalize_assurance_run(
            output_dir=output_dir,
            context=stale,
            reconciliation_rows=rows,
            allocation_ledgers=[],
            checks=[{"check": "row_count", "status": "PASS"}],
            review_rows=[],
            source_qualifications=[{"status": "qualified"}],
            declared_outputs=[workbook_path],
            workbook_name=workbook_path.name,
        )

    assert not (output_dir / "assurance_final_outputs").exists()


def test_final_output_inventory_rejects_missing_and_unexpected_files(
    tmp_path: Path,
) -> None:
    assurance, output_dir, payload = finalize_case(tmp_path)
    boundary = output_dir / "assurance_final_outputs"
    inventory = payload["final_output_inventory"]
    contract = payload["workflow_output_contract"]
    assert contract["declared_paths"] == [
        "audit.xlsx",
        "reconciliation_results.json",
        "report.txt",
    ]
    (boundary / "unexpected.txt").write_text("unexpected")

    with pytest.raises(ValueError, match="file sets do not match"):
        assurance.validate_final_output_inventory(output_dir, inventory)
    with pytest.raises(ValueError, match="file sets do not match"):
        assurance.validate_assurance_run(output_dir)

    (boundary / "unexpected.txt").unlink()
    (boundary / "report.txt").unlink()
    with pytest.raises(ValueError, match="file sets do not match"):
        assurance.validate_final_output_inventory(output_dir, inventory)


def test_all_native_amount_rows_are_covered_by_output_receipt(
    tmp_path: Path,
) -> None:
    assurance, output_dir, payload = finalize_case(tmp_path)
    boundary_workbook = output_dir / "assurance_final_outputs" / "audit.xlsx"
    workbook = load_workbook(boundary_workbook)
    sheet = workbook["Reconciliation detail"]
    sheet["B2"] = "999.00"
    sheet["B3"] = "888.00"
    workbook.save(boundary_workbook)

    with pytest.raises(ValueError, match="does not match current bytes"):
        assurance.validate_final_output_inventory(
            output_dir,
            payload["final_output_inventory"],
        )


def test_exact_residual_is_preserved_and_blocks_reconciliation_gate(
    tmp_path: Path,
) -> None:
    assurance = load_assurance()
    allocation = assurance.build_allocation_ledger(
        ledger_id="allocation.test",
        policy={
            "relationship_shape": "one_to_one",
            "require_same_currency": True,
            "require_same_unit": True,
            "require_same_entity": True,
            "require_same_party": True,
            "allow_evidence_reuse": False,
            "tolerance": "0",
        },
        source_records=[
            {
                "record_id": "source.record",
                "amount": "100",
                "currency": "EUR",
                "unit": "currency_amount",
                "entity_ref": "entity.test",
                "party_ref": "party.test",
            }
        ],
        target_records=[
            {
                "record_id": "target.record",
                "amount": "70",
                "currency": "EUR",
                "unit": "currency_amount",
                "entity_ref": "entity.test",
                "party_ref": "party.test",
            }
        ],
        allocations=[
            {
                "allocation_id": "allocation.link",
                "source_record_ref": "source.record",
                "target_record_ref": "target.record",
                "amount": "70",
                "currency": "EUR",
                "unit": "currency_amount",
                "evidence_refs": ["evidence.link"],
            }
        ],
    )

    _, _, payload = finalize_case(tmp_path, allocation_ledgers=[allocation])

    residuals = [
        item
        for item in payload["allocation_value_addresses"]
        if item["kind"] == "source_residual"
    ]
    assert residuals[0]["value"] == "30"
    assert residuals[0]["artifact_ref"].startswith("final_output.")
    assert "prepared_locator" not in residuals[0]
    assert payload["gate_register"]["gates"]["reconciliation"]["status"] == "failed"
    assert payload["gate_register"]["gates"]["reporting"]["status"] == "blocked"


def test_independent_gates_distinguish_pending_review_from_valid_report(
    tmp_path: Path,
) -> None:
    _, _, pending = finalize_case(tmp_path / "pending", review_status="PENDING")
    _, _, failed = finalize_case(tmp_path / "failed", review_status="FAIL")
    _, _, reviewed = finalize_case(tmp_path / "reviewed", review_status="PASS")

    pending_gates = pending["gate_register"]["gates"]
    reviewed_gates = reviewed["gate_register"]["gates"]
    assert pending_gates["source"]["status"] == "passed"
    assert pending_gates["preparation"]["status"] == "passed"
    assert pending_gates["reconciliation"]["status"] == "passed"
    assert pending_gates["semantic_review"]["status"] == "withheld"
    assert pending_gates["reporting"]["status"] == "blocked"
    assert failed["gate_register"]["gates"]["semantic_review"]["status"] == "failed"
    assert failed["gate_register"]["gates"]["reporting"]["status"] == "blocked"
    assert reviewed_gates["semantic_review"]["status"] == "passed"
    assert reviewed_gates["reporting"]["status"] == "passed"
    assert reviewed_gates["publication"]["status"] == "withheld"
    assert reviewed["gate_register"]["report_ready"] is True


def test_pass_without_reviewer_identity_is_withheld(tmp_path: Path) -> None:
    assurance, _, output_dir, rows, context = prepared_case(tmp_path)
    workbook_path = output_dir / "audit.xlsx"
    workbook_path.write_bytes(workbook_bytes(tmp_path))

    payload = assurance.finalize_assurance_run(
        output_dir=output_dir,
        context=context,
        reconciliation_rows=rows,
        allocation_ledgers=[],
        checks=[{"check": "row_count", "status": "PASS"}],
        review_rows=[
            {"record_id": row["record_id"], "review_status": "PASS"} for row in rows
        ],
        source_qualifications=qualified_sources(assurance, context),
        declared_outputs=[workbook_path],
        workbook_name=workbook_path.name,
    )

    gates = payload["gate_register"]["gates"]
    assert gates["semantic_review"]["status"] == "withheld"
    assert gates["reporting"]["status"] == "blocked"


def test_missing_required_review_is_withheld(tmp_path: Path) -> None:
    assurance, _, output_dir, rows, context = prepared_case(tmp_path)
    workbook_path = output_dir / "audit.xlsx"
    workbook_path.write_bytes(workbook_bytes(tmp_path))

    payload = assurance.finalize_assurance_run(
        output_dir=output_dir,
        context=context,
        reconciliation_rows=rows,
        allocation_ledgers=[],
        checks=[{"check": "row_count", "status": "PASS"}],
        review_rows=[],
        source_qualifications=qualified_sources(assurance, context),
        declared_outputs=[workbook_path],
        workbook_name=workbook_path.name,
    )

    gates = payload["gate_register"]["gates"]
    assert gates["semantic_review"]["status"] == "withheld"
    assert gates["reporting"]["status"] == "blocked"


def test_assurance_receipts_are_repeatable_for_same_bytes(tmp_path: Path) -> None:
    content = workbook_bytes(tmp_path)
    _, _, first = finalize_case(tmp_path / "first", workbook_content=content)
    _, _, second = finalize_case(tmp_path / "second", workbook_content=content)

    assert first["final_output_inventory"] == second["final_output_inventory"]
    assert first["numeric_evidence_ledger"] == second["numeric_evidence_ledger"]
    assert first["gate_register"] == second["gate_register"]


def test_assurance_replay_rejects_gate_file_that_outruns_seal(
    tmp_path: Path,
) -> None:
    assurance, output_dir, _ = finalize_case(
        tmp_path,
        review_status="PENDING",
    )
    gates_path = output_dir / "assurance_gates.json"
    gates = json.loads(gates_path.read_text(encoding="utf-8"))
    gates["gates"]["semantic_review"]["status"] = "passed"
    gates["gates"]["semantic_review"]["evidence_refs"] = ["forged.review"]
    gates_path.write_text(json.dumps(gates, indent=2) + "\n", encoding="utf-8")

    with pytest.raises(ValueError, match="gate file is stale"):
        assurance.validate_assurance_run(output_dir)


def test_review_application_cannot_outrun_latest_assurance_replay(
    tmp_path: Path,
) -> None:
    workflow = load_script_module(
        "audit_reconciliation_assured_workflow",
        WORKFLOW_PATH,
    )
    review_server = load_script_module(
        "audit_reconciliation_assured_review_server",
        REVIEW_SERVER_PATH,
    )
    output_dir = tmp_path / "output"
    workflow.build_reconciliation_artifacts(
        output_dir=output_dir,
        open_items=[
            {
                "record_id": "open-1",
                "document_key": "INV-1|2026",
                "document_no": "INV-1",
                "document_date": "2026-01-01",
                "amount": "100.00",
                "currency": "EUR",
            }
        ],
        evidence_rows=[
            {
                "record_id": "bank-1",
                "source_role": "bank_statement",
                "evidence_type": "external_bank",
                "document_key": "INV-1|2026",
                "document_no": "INV-1",
                "document_date": "2026-01-01",
                "posting_date": "2026-01-02",
                "amount": "100.00",
                "currency": "EUR",
                "source_file": "bank.csv",
                "source_row": 2,
            }
        ],
        assumptions={"scope_year": "2026", "amount_tolerance": "0"},
        require_completed_review=False,
    )
    review_payload = json.loads(
        (output_dir / "review_payload.json").read_text(encoding="utf-8")
    )
    expected_predecessor_checkpoint = json.loads(
        (output_dir / "assurance_receipts.json").read_text(encoding="utf-8")
    )["content_sha256"]

    result = review_server.apply_decisions(
        output_dir,
        {
            "decisions": [
                {"item_id": item["id"], "action": "accept"}
                for item in review_payload["items"]
            ],
            "expected_predecessor_checkpoint": expected_predecessor_checkpoint,
        },
    )

    assert result["application_status"] == "blocked"
    assert any(
        blocker["kind"] == "assurance_replay_required"
        for blocker in result["applied_decisions"]["completion_blockers"]
    )


def test_honest_successor_retains_and_replays_exact_transition(
    tmp_path: Path,
) -> None:
    (
        assurance,
        review_server,
        output_dir,
        decisions,
        predecessor_bytes,
        first_apply,
    ) = successor_lifecycle_case(tmp_path)
    predecessor = json.loads(predecessor_bytes["assurance"])
    predecessor_digest = predecessor["content_sha256"]
    history_dir = output_dir / "assurance_transition_history" / predecessor_digest

    assert first_apply["application_status"] == "blocked"
    assert {path.name for path in history_dir.iterdir()} == {
        "predecessor_assurance_receipts.json",
        "predecessor_professional_review.json",
        "predecessor_reconciliation_results.json",
        "predecessor_review_payload.json",
        "review_payload_mapping.json",
        "applied_decisions.json",
        "successor_professional_review.json",
        "transition_receipt.json",
        "predecessor_run",
    }
    assert (
        history_dir / "predecessor_assurance_receipts.json"
    ).read_bytes() == predecessor_bytes["assurance"]
    assert (
        history_dir / "predecessor_professional_review.json"
    ).read_bytes() == predecessor_bytes["professional"]
    assert (
        history_dir / "predecessor_reconciliation_results.json"
    ).read_bytes() == predecessor_bytes["reconciliation"]
    assert (
        history_dir / "predecessor_review_payload.json"
    ).read_bytes() == predecessor_bytes["review_payload"]
    predecessor_snapshot = history_dir / "predecessor_run"
    assert (
        predecessor_snapshot / "assurance_receipts.json"
    ).read_bytes() == predecessor_bytes["assurance"]
    assert (
        predecessor_snapshot / "professional_review.json"
    ).read_bytes() == predecessor_bytes["professional"]
    assert (
        predecessor_snapshot / "assurance_final_outputs" / "reconciliation_results.json"
    ).read_bytes() == predecessor_bytes["reconciliation"]
    assert (
        predecessor_snapshot / "review_payload.json"
    ).read_bytes() == predecessor_bytes["review_payload"]
    assert (
        assurance.validate_assurance_run(predecessor_snapshot)["content_sha256"]
        == predecessor_digest
    )
    successor = assurance.validate_assurance_run(
        output_dir,
        expected_predecessor_checkpoint=predecessor_digest,
    )
    assert (
        successor["professional_review_authority"]["predecessor_assurance_sha256"]
        == predecessor_digest
    )
    assert len(successor["review_transition_receipts"]) == 1
    assert (
        successor["review_transition_receipts"][0]["predecessor_assurance_sha256"]
        == predecessor_digest
    )
    assert successor["gate_register"]["gates"]["publication"]["status"] == "withheld"

    review_payload = json.loads(
        (output_dir / "review_payload.json").read_text(encoding="utf-8")
    )
    result = review_server.apply_decisions(
        output_dir,
        {
            "decisions": [
                {"item_id": item["id"], "action": "accept"}
                for item in review_payload["items"]
            ],
            "expected_predecessor_checkpoint": predecessor_digest,
        },
    )

    assert result["application_status"] == "final_ready"
    assert (
        result["applied_decisions"]["professional_review"][
            "successor_assurance_replayed"
        ]
        is True
    )
    assert (
        assurance.validate_assurance_run(
            output_dir,
            expected_predecessor_checkpoint=predecessor_digest,
        )["gate_register"]["gates"]["publication"]["status"]
        == "withheld"
    )


def test_successor_validation_requires_external_checkpoint_and_rejects_wrong_one(
    tmp_path: Path,
) -> None:
    assurance, _, output_dir, decisions, _, _ = successor_lifecycle_case(tmp_path)
    expected = str(decisions["expected_predecessor_checkpoint"])
    wrong = "f" * 64 if expected != "f" * 64 else "e" * 64

    with pytest.raises(
        ValueError,
        match="external expected predecessor checkpoint is required",
    ):
        assurance.validate_assurance_run(output_dir)
    with pytest.raises(
        ValueError,
        match="external expected predecessor checkpoint does not match",
    ):
        assurance.validate_assurance_run(
            output_dir,
            expected_predecessor_checkpoint=wrong,
        )

    replay = assurance.validate_assurance_run(
        output_dir,
        expected_predecessor_checkpoint=expected,
    )
    assert (
        replay["professional_review_authority"]["predecessor_assurance_sha256"]
        == expected
    )


def test_isolated_successor_cli_requires_matching_external_checkpoint(
    tmp_path: Path,
) -> None:
    _, _, output_dir, decisions, _, _ = successor_lifecycle_case(tmp_path)
    managed_root = tmp_path / "managed"
    managed_root.mkdir()
    context_path, client_engagement = running_audit_context(managed_root)
    managed_output = Path(str(client_engagement["output_dir"]))
    shutil.copytree(output_dir, managed_output, dirs_exist_ok=True)
    output_dir = managed_output
    expected = str(decisions["expected_predecessor_checkpoint"])
    wrong = "f" * 64 if expected != "f" * 64 else "e" * 64
    environment = {
        **os.environ,
        "PYTHONDONTWRITEBYTECODE": "1",
        "PYTHONPYCACHEPREFIX": (tmp_path / "pycache").as_posix(),
    }

    missing = subprocess.run(
        [
            sys.executable,
            "-I",
            "-B",
            ASSURANCE_PATH.as_posix(),
            "--client-engagement",
            context_path.as_posix(),
            "validate-run-json",
            output_dir.as_posix(),
        ],
        check=False,
        capture_output=True,
        text=True,
        env=environment,
    )
    wrong_result = subprocess.run(
        [
            sys.executable,
            "-I",
            "-B",
            ASSURANCE_PATH.as_posix(),
            "--client-engagement",
            context_path.as_posix(),
            "validate-run-json",
            output_dir.as_posix(),
            "--expected-predecessor-checkpoint",
            wrong,
        ],
        check=False,
        capture_output=True,
        text=True,
        env=environment,
    )
    correct = subprocess.run(
        [
            sys.executable,
            "-I",
            "-B",
            ASSURANCE_PATH.as_posix(),
            "--client-engagement",
            context_path.as_posix(),
            "validate-run-json",
            output_dir.as_posix(),
            "--expected-predecessor-checkpoint",
            expected,
        ],
        check=False,
        capture_output=True,
        text=True,
        env=environment,
    )

    assert missing.returncode == 1
    assert wrong_result.returncode == 1
    assert json.loads(missing.stdout) == {
        "ok": False,
        "error": "complete assurance replay failed",
    }
    assert json.loads(wrong_result.stdout) == {
        "ok": False,
        "error": "complete assurance replay failed",
    }
    assert correct.returncode == 0
    assert json.loads(correct.stdout)["ok"] is True
    assert missing.stderr == wrong_result.stderr == correct.stderr == ""


def test_run_id_is_digest_bound_and_self_resealed_mismatch_is_rejected(
    tmp_path: Path,
) -> None:
    assurance, _, output_dir, decisions, _, _ = successor_lifecycle_case(tmp_path)
    seal_path = output_dir / "assurance_receipts.json"
    seal = json.loads(seal_path.read_text(encoding="utf-8"))
    review_payload = json.loads(
        (output_dir / "review_payload.json").read_text(encoding="utf-8")
    )
    assert seal["run_id"] == review_payload["run_id"]
    original_digest = seal["content_sha256"]
    forged_run_id = f"{seal['run_id']}-forged"
    seal["run_id"] = forged_run_id
    reseal_outer(assurance, seal)
    assert seal["content_sha256"] != original_digest
    stable_write(seal_path, seal)

    with pytest.raises(ValueError, match="sealed review run identity is stale"):
        assurance.validate_assurance_run(
            output_dir,
            expected_predecessor_checkpoint=decisions[
                "expected_predecessor_checkpoint"
            ],
        )


def test_self_resealed_successor_predecessor_replacement_is_rejected(
    tmp_path: Path,
) -> None:
    assurance, review_server, output_dir, decisions, _, _ = successor_lifecycle_case(
        tmp_path
    )
    professional_path = output_dir / "professional_review.json"
    professional = json.loads(professional_path.read_text(encoding="utf-8"))
    professional["predecessor_assurance_sha256"] = "f" * 64
    professional_content = {
        key: value for key, value in professional.items() if key != "content_sha256"
    }
    professional["content_sha256"] = assurance.canonical_json_sha256(
        professional_content
    )
    stable_write(professional_path, professional)
    seal_path = output_dir / "assurance_receipts.json"
    seal = json.loads(seal_path.read_text(encoding="utf-8"))
    seal["professional_review_authority"] = professional
    professional_bytes = professional_path.read_bytes()
    seal["professional_review_receipt"]["byte_count"] = len(professional_bytes)
    seal["professional_review_receipt"]["sha256"] = hashlib.sha256(
        professional_bytes
    ).hexdigest()
    reseal_outer(assurance, seal)
    stable_write(seal_path, seal)
    before = exact_tree_image(output_dir)
    review_payload = json.loads(
        (output_dir / "review_payload.json").read_text(encoding="utf-8")
    )

    with pytest.raises(
        ValueError,
        match="external expected predecessor checkpoint|retained transition",
    ):
        assurance.validate_assurance_run(
            output_dir,
            expected_predecessor_checkpoint=decisions[
                "expected_predecessor_checkpoint"
            ],
        )
    with pytest.raises(
        ValueError,
        match="external expected predecessor checkpoint|retained transition",
    ):
        review_server.apply_decisions(
            output_dir,
            {
                "decisions": [
                    {"item_id": item["id"], "action": "accept"}
                    for item in review_payload["items"]
                ],
                "expected_predecessor_checkpoint": decisions[
                    "expected_predecessor_checkpoint"
                ],
            },
        )
    assert exact_tree_image(output_dir) == before


@pytest.mark.parametrize(
    "mutation_family",
    [
        "material_amount",
        "currency",
        "cutoff_date",
        "run_date",
        "run_id",
        "scope_year",
        "amount_tolerance",
    ],
)
def test_external_checkpoint_rejects_every_resealed_predecessor_family(
    tmp_path: Path,
    mutation_family: str,
) -> None:
    assurance, _, output_dir, decisions, predecessor_bytes, _ = (
        successor_lifecycle_case(tmp_path)
    )
    expected = str(decisions["expected_predecessor_checkpoint"])
    forged_predecessor = json.loads(predecessor_bytes["assurance"])
    if mutation_family == "run_date":
        forged_predecessor["run_date"] = "2026-07-24"
    elif mutation_family == "run_id":
        forged_predecessor["run_id"] = (
            f"{forged_predecessor['run_id'] or 'predecessor'}-forged"
        )
    else:
        forged_predecessor["prepared_receipt"]["sha256"] = hashlib.sha256(
            f"fully-resealed-{mutation_family}".encode()
        ).hexdigest()
    reseal_outer(assurance, forged_predecessor)
    forged_digest = forged_predecessor["content_sha256"]
    assert forged_digest != expected

    professional_path = output_dir / "professional_review.json"
    professional = json.loads(professional_path.read_text(encoding="utf-8"))
    professional["predecessor_assurance_sha256"] = forged_digest
    professional_content = {
        key: value for key, value in professional.items() if key != "content_sha256"
    }
    professional["content_sha256"] = assurance.canonical_json_sha256(
        professional_content
    )
    stable_write(professional_path, professional)
    seal_path = output_dir / "assurance_receipts.json"
    seal = json.loads(seal_path.read_text(encoding="utf-8"))
    seal["professional_review_authority"] = professional
    professional_bytes = professional_path.read_bytes()
    seal["professional_review_receipt"]["byte_count"] = len(professional_bytes)
    seal["professional_review_receipt"]["sha256"] = hashlib.sha256(
        professional_bytes
    ).hexdigest()
    reseal_outer(assurance, seal)
    stable_write(seal_path, seal)

    with pytest.raises(
        ValueError,
        match="external expected predecessor checkpoint does not match",
    ):
        assurance.validate_assurance_run(
            output_dir,
            expected_predecessor_checkpoint=expected,
        )


def test_fully_resealed_contradictory_predecessor_is_rejected_in_process(
    tmp_path: Path,
) -> None:
    assurance, output_dir, expected = fully_resealed_contradictory_predecessor(tmp_path)

    with pytest.raises(
        ValueError,
        match="external expected predecessor checkpoint|prepared run date|complete replay",
    ):
        assurance.validate_assurance_run(
            output_dir,
            expected_predecessor_checkpoint=expected,
        )


def test_fully_resealed_contradictory_predecessor_is_rejected_by_isolated_cli(
    tmp_path: Path,
) -> None:
    _, output_dir, expected = fully_resealed_contradictory_predecessor(tmp_path)
    managed_root = tmp_path / "managed"
    managed_root.mkdir()
    context_path, client_engagement = running_audit_context(managed_root)
    managed_output = Path(str(client_engagement["output_dir"]))
    shutil.copytree(output_dir, managed_output, dirs_exist_ok=True)
    output_dir = managed_output

    completed = subprocess.run(
        [
            sys.executable,
            "-I",
            "-B",
            ASSURANCE_PATH.as_posix(),
            "--client-engagement",
            context_path.as_posix(),
            "validate-run-json",
            output_dir.as_posix(),
            "--expected-predecessor-checkpoint",
            expected,
        ],
        check=False,
        capture_output=True,
        text=True,
        env={
            **os.environ,
            "PYTHONDONTWRITEBYTECODE": "1",
            "PYTHONPYCACHEPREFIX": (tmp_path / "pycache").as_posix(),
        },
    )

    assert completed.returncode == 1
    assert json.loads(completed.stdout) == {
        "ok": False,
        "error": "complete assurance replay failed",
    }
    assert completed.stderr == ""


@pytest.mark.parametrize(
    "attack",
    ["missing", "changed", "expanded", "reordered", "history_path"],
)
def test_successor_rejects_transition_history_forgery(
    tmp_path: Path,
    attack: str,
) -> None:
    assurance, _, output_dir, decisions, _, _ = successor_lifecycle_case(tmp_path)
    seal_path = output_dir / "assurance_receipts.json"
    seal = json.loads(seal_path.read_text(encoding="utf-8"))
    predecessor_digest = seal["professional_review_authority"][
        "predecessor_assurance_sha256"
    ]
    history_dir = output_dir / "assurance_transition_history" / predecessor_digest
    if attack == "missing":
        (history_dir / "predecessor_review_payload.json").unlink()
    elif attack == "changed":
        (history_dir / "predecessor_professional_review.json").write_bytes(b"{}\n")
    elif attack == "expanded":
        (history_dir / "rogue.json").write_text("{}\n", encoding="utf-8")
    elif attack == "history_path":
        history_dir.rename(history_dir.with_name("f" * 64))
    else:
        transition_path = history_dir / "transition_receipt.json"
        transition = json.loads(transition_path.read_text(encoding="utf-8"))
        transition["artifact_receipts"] = list(
            reversed(transition["artifact_receipts"])
        )
        transition_content = {
            key: value for key, value in transition.items() if key != "content_sha256"
        }
        transition["content_sha256"] = assurance.canonical_json_sha256(
            transition_content
        )
        stable_write(transition_path, transition)
        seal["review_transition_receipts"] = [transition]
        reseal_outer(assurance, seal)
        stable_write(seal_path, seal)

    with pytest.raises(ValueError, match="transition|run tree|artifact receipt"):
        assurance.validate_assurance_run(
            output_dir,
            expected_predecessor_checkpoint=decisions[
                "expected_predecessor_checkpoint"
            ],
        )


def test_assurance_contract_is_documented_in_skill_and_workflow_reference() -> None:
    skill = (PLUGIN_ROOT / "skills" / "audit-reconciliation" / "SKILL.md").read_text(
        encoding="utf-8"
    )
    workflow = (PLUGIN_ROOT / "references" / "workflow-reference.md").read_text(
        encoding="utf-8"
    )
    normalized_workflow = " ".join(workflow.split())

    assert "reviewed_source_decisions" in skill
    assert "assurance_final_outputs/" in skill
    assert "publication remains withheld" in skill
    assert "full-byte" in workflow
    assert "`semantic_review`" in workflow
    assert "professional" in workflow.lower()
    assert "unsigned, unauthenticated, untrusted label" in workflow
    assert "unsigned, unauthenticated, untrusted label" in skill
    assert "ordered 25-file implementation contract" in skill
    assert "ordered 25-file contract" in workflow
    assert "pre-import" in workflow
    assert "bootstrap" in workflow
    assert "`__pycache__`" in workflow
    assert "assurance_transition_history" in workflow
    assert "predecessor-seal-content-sha256" in workflow
    assert "predecessor_run/" in skill
    assert "predecessor_run/" in workflow
    assert "complete physical snapshot" in workflow
    assert "outside this in-process 25-file boundary" in workflow
    assert "`expected_predecessor_checkpoint`" in skill
    assert "`expected_predecessor_checkpoint`" in workflow
    assert "separate review channel" in skill
    assert "never inferred from the candidate output tree" in workflow
    assert "`scripts/retained_sources/`" in skill
    assert "`scripts/retained_sources/`" in workflow
    assert "Ordinary Python import cannot resolve them" in normalized_workflow
    assert "same operating-system user" in workflow
    assert "`run_id` is an explicit assurance-seal field" in normalized_workflow


@pytest.mark.parametrize(
    ("field", "value"),
    [
        ("reviewer_ref", None),
        ("reviewer_ref", 123),
        ("reviewer_ref", " reviewer.test"),
        ("reviewed_on", None),
        ("reviewed_on", "not-a-date"),
        ("reviewed_on", "2999-01-01"),
    ],
)
def test_source_decision_rejects_noncanonical_or_future_review_metadata(
    tmp_path: Path,
    field: str,
    value: object,
) -> None:
    assurance = load_assurance()
    source = tmp_path / "source.csv"
    source.write_text("document,amount\nINV-1,100.00\n", encoding="utf-8")
    receipts = assurance.build_source_receipts(tmp_path, [source])
    supplied = reviewed_source_decision()
    supplied[field] = value

    decisions, errors = assurance.build_reviewed_source_decisions(
        input_root=tmp_path,
        source_receipts=receipts,
        adapter_families={"source.csv": "open_items_text_v1"},
        assumptions={
            "assurance_run_date": "2026-07-25",
            "reviewed_source_decisions": {"source.csv": supplied},
        },
    )

    assert decisions == {}
    assert "source.csv" in errors


def test_source_decision_v2_rejects_stale_v1_receipt(tmp_path: Path) -> None:
    assurance = load_assurance()
    source = tmp_path / "source.csv"
    source.write_text("document,amount\nINV-1,100.00\n", encoding="utf-8")
    receipts = assurance.build_source_receipts(tmp_path, [source])
    decisions, errors = assurance.build_reviewed_source_decisions(
        input_root=tmp_path,
        source_receipts=receipts,
        adapter_families={"source.csv": "open_items_text_v1"},
        assumptions={
            "assurance_run_date": "2026-07-25",
            "reviewed_source_decisions": {"source.csv": reviewed_source_decision()},
        },
    )
    assert errors == {}
    stale = copy.deepcopy(decisions["source.csv"])
    stale["adapter_version"] = "1"

    replayed, replay_errors = assurance.build_reviewed_source_decisions(
        input_root=tmp_path,
        source_receipts=receipts,
        adapter_families={"source.csv": "open_items_text_v1"},
        assumptions={
            "assurance_run_date": "2026-07-25",
            "reviewed_source_decisions": {"source.csv": stale},
        },
    )

    assert replayed == {}
    assert "adapter version is stale" in replay_errors["source.csv"]


def test_date_authority_is_source_bound_and_invalid_bank_date_emits_no_row(
    tmp_path: Path,
) -> None:
    assurance = load_assurance()
    runner = load_script_module(
        "audit_reconciliation_blackbox_dates",
        SCRIPTS / "raw_input_runner.py",
    )
    assert runner.iso_date("01/02/2026") == ""
    assert (
        runner.iso_date(
            "01/02/2026",
            convention={"order": "day_first"},
        )
        == "2026-02-01"
    )
    assert (
        runner.iso_date(
            "01/02/2026",
            convention={"order": "month_first"},
        )
        == "2026-01-02"
    )
    source = tmp_path / "bank.pdf"
    source.write_bytes(b"placeholder")
    receipts = assurance.build_source_receipts(tmp_path, [source])
    supplied = reviewed_source_decision()
    supplied["role"] = "bank_statement"
    supplied["adapter_family"] = "bank_statement_text_v1"
    supplied["perimeter"]["direction_policy"] = "not_applicable"
    supplied["money"]["decimal_separator"] = ","
    supplied["money"]["thousands_separator"] = "."
    decisions, errors = assurance.build_reviewed_source_decisions(
        input_root=tmp_path,
        source_receipts=receipts,
        adapter_families={"bank.pdf": "bank_statement_text_v1"},
        assumptions={
            "assurance_run_date": "2026-07-25",
            "reviewed_source_decisions": {"bank.pdf": supplied},
        },
    )
    assert errors == {}
    page = runner.SourcePage(
        source_file="bank.pdf",
        source_role="bank_statement",
        source_page=1,
        extraction_method="text",
        text_length=0,
        line_count=1,
        text="",
    )
    active = {"_reviewed_source_decision_receipts": {"bank.pdf": decisions["bank.pdf"]}}

    assert (
        runner._bank_row_from_text(
            page,
            "31/02/26 31/02/26 100,00 123-FE",
            1,
            active,
            [],
        )
        == []
    )


def test_invalid_populated_critical_date_blocks_before_prepared_rows(
    tmp_path: Path,
) -> None:
    assurance = load_assurance()
    source_dir = tmp_path / "source"
    output_dir = tmp_path / "output"
    source_dir.mkdir()
    output_dir.mkdir()
    source = source_dir / "source.csv"
    source.write_text(
        "document,amount\nINV-1,100.00\nINV-2,25.50\n",
        encoding="utf-8",
    )
    receipts = assurance.build_source_receipts(source_dir, [source])
    decisions, errors = assurance.build_reviewed_source_decisions(
        input_root=source_dir,
        source_receipts=receipts,
        adapter_families={"source.csv": "open_items_text_v1"},
        assumptions={
            "reviewed_source_decisions": {"source.csv": reviewed_source_decision()}
        },
    )
    assert errors == {}
    qualifications = qualified_sources(
        assurance,
        {
            "source_receipts": receipts,
            "reviewed_source_decisions": list(decisions.values()),
        },
    )
    rows = source_rows()
    rows[0]["document_date"] = "31/02/2026"

    with pytest.raises(ValueError, match="invalid populated critical date"):
        assurance.prepare_assurance_run(
            output_dir=output_dir,
            open_items=rows,
            evidence_rows=[],
            assumptions={"amount_tolerance": "0"},
            source_root=source_dir,
            source_receipts=receipts,
            reviewed_source_decisions=list(decisions.values()),
            source_qualifications=qualifications,
        )

    assert not (output_dir / "prepared_records.json").exists()


def test_non_cent_increment_is_rejected_without_downstream_rounding(
    tmp_path: Path,
) -> None:
    assurance = load_assurance()
    runner = load_script_module(
        "audit_reconciliation_blackbox_increment_runner",
        SCRIPTS / "raw_input_runner.py",
    )
    workpapers = load_retained_module("workpaper_outputs")
    accountant = load_retained_module("accountant_report")
    source = tmp_path / "source.csv"
    source.write_text("document,amount\nINV-1,1.234\n", encoding="utf-8")
    receipts = assurance.build_source_receipts(tmp_path, [source])
    supplied = reviewed_source_decision()
    supplied["money"]["reported_increment"] = "0.001"

    decisions, errors = assurance.build_reviewed_source_decisions(
        input_root=tmp_path,
        source_receipts=receipts,
        adapter_families={"source.csv": "open_items_text_v1"},
        assumptions={"reviewed_source_decisions": {"source.csv": supplied}},
    )

    assert decisions == {}
    assert "exactly 0.01" in errors["source.csv"]
    assert (
        runner.parse_money(
            "1.234",
            convention={
                "decimal_separator": ".",
                "thousands_separator": ",",
                "reported_increment": "0.001",
            },
        )
        is None
    )
    assert (
        workpapers.summary_from_reconciliation(
            [
                {
                    "reconciliation_status": "unresolved",
                    "rule_applied": "unresolved",
                    "amount": "1.234",
                }
            ]
        )[0]["amount"]
        == "1.234"
    )
    assert accountant.parse_decimal("1.234") is None


@pytest.mark.parametrize("invalid", [True, 0.0, -1, "-0.01"])
def test_amount_tolerance_rejects_bool_float_and_negative(invalid: object) -> None:
    helpers = load_retained_module("reconciliation_helpers")

    with pytest.raises(ValueError, match="non-negative canonical Decimal"):
        helpers.amounts_equal("100", "100", invalid)


def test_zero_tolerance_is_not_promoted_to_one_cent() -> None:
    helpers = load_retained_module("reconciliation_helpers")

    rows = helpers.reconcile_open_items(
        [
            {
                "record_id": "open-zero",
                "document_key": "INV-ZERO|2026",
                "amount": "100",
                "currency": "EUR",
                "unit": "currency_amount",
                "entity_ref": "entity.test",
                "party_ref": "party.test",
            }
        ],
        [
            {
                "record_id": "bank-zero",
                "source_role": "bank_statement",
                "evidence_type": "external_bank",
                "document_key": "INV-ZERO|2026",
                "amount": "100.005",
                "currency": "EUR",
                "unit": "currency_amount",
                "entity_ref": "entity.test",
                "party_ref": "party.test",
            }
        ],
        {"amount_tolerance": "0"},
    )

    assert rows[0]["reconciliation_status"] != "closed"


@pytest.mark.parametrize(
    "review_rows",
    [
        [
            {
                "record_id": "open-1",
                "review_status": "PASS",
                "reviewer_ref": "reviewer.test",
                "reviewed_on": "2026-07-25",
            }
        ],
        [
            {
                "record_id": "open-1",
                "review_status": "PASS",
                "reviewer_ref": "reviewer.test",
                "reviewed_on": "2026-07-25",
            },
            {
                "record_id": "open-1",
                "review_status": "PASS",
                "reviewer_ref": "reviewer.test",
                "reviewed_on": "2026-07-25",
            },
        ],
        [
            {
                "record_id": "unknown",
                "review_status": "PASS",
                "reviewer_ref": "reviewer.test",
                "reviewed_on": "2026-07-25",
            }
        ],
        [
            {
                "review_status": "PASS",
                "reviewer_ref": "reviewer.test",
                "reviewed_on": "2026-07-25",
            }
        ],
        [
            {
                "record_id": "open-1",
                "review_status": "PASS",
                "reviewer_ref": "reviewer.test",
                "reviewed_on": "not-a-date",
            },
            {
                "record_id": "open-2",
                "review_status": "PASS",
                "reviewer_ref": "reviewer.test",
                "reviewed_on": "not-a-date",
            },
        ],
        [
            {
                "record_id": "open-1",
                "review_status": "PASS",
                "reviewer_ref": "reviewer.test",
                "reviewed_on": "2999-01-01",
            },
            {
                "record_id": "open-2",
                "review_status": "PASS",
                "reviewer_ref": "reviewer.test",
                "reviewed_on": "2999-01-01",
            },
        ],
        [
            {
                "record_id": "open-1",
                "review_status": "PASS",
                "reviewer_ref": 123,
                "reviewed_on": "2026-07-25",
            },
            {
                "record_id": "open-2",
                "review_status": "PASS",
                "reviewer_ref": 123,
                "reviewed_on": "2026-07-25",
            },
        ],
    ],
)
def test_review_gate_requires_exact_record_set_and_bounded_identity(
    tmp_path: Path,
    review_rows: list[dict[str, object]],
) -> None:
    assurance, _, output_dir, rows, context = prepared_case(tmp_path)
    workbook_path = output_dir / "audit.xlsx"
    workbook_path.write_bytes(workbook_bytes(tmp_path))

    payload = assurance.finalize_assurance_run(
        output_dir=output_dir,
        context=context,
        reconciliation_rows=rows,
        allocation_ledgers=[],
        checks=[{"check": "row_count", "status": "PASS"}],
        review_rows=review_rows,
        source_qualifications=qualified_sources(assurance, context),
        declared_outputs=[workbook_path],
        workbook_name=workbook_path.name,
    )

    assert payload["gate_register"]["report_ready"] is False
    assert payload["gate_register"]["gates"]["reporting"]["status"] == "blocked"


@pytest.mark.parametrize(
    "checks",
    [
        [],
        [{"check": "", "status": "PASS"}],
        [{"check": "row_count", "status": ""}],
    ],
)
def test_rows_require_nonblank_mechanical_check_coverage(
    tmp_path: Path,
    checks: list[dict[str, str]],
) -> None:
    assurance, _, output_dir, rows, context = prepared_case(tmp_path)
    workbook_path = output_dir / "audit.xlsx"
    workbook_path.write_bytes(workbook_bytes(tmp_path))

    payload = assurance.finalize_assurance_run(
        output_dir=output_dir,
        context=context,
        reconciliation_rows=rows,
        allocation_ledgers=[],
        checks=checks,
        review_rows=[
            {
                "record_id": row["record_id"],
                "review_status": "PASS",
                "reviewer_ref": "reviewer.test",
                "reviewed_on": "2026-07-25",
            }
            for row in rows
        ],
        source_qualifications=qualified_sources(assurance, context),
        declared_outputs=[workbook_path],
        workbook_name=workbook_path.name,
    )

    assert payload["gate_register"]["gates"]["reconciliation"]["status"] == "failed"
    assert payload["gate_register"]["report_ready"] is False


def test_missing_decision_or_mapping_blocks_before_prepared_rows(
    tmp_path: Path,
) -> None:
    assurance = load_assurance()
    source_dir = tmp_path / "source"
    output_dir = tmp_path / "output"
    source_dir.mkdir()
    output_dir.mkdir()
    source = source_dir / "source.csv"
    source.write_text(
        "document,amount\nINV-1,100.00\nINV-2,25.50\n",
        encoding="utf-8",
    )
    receipts = assurance.build_source_receipts(source_dir, [source])

    with pytest.raises(ValueError, match="exactly one current reviewed"):
        assurance.prepare_assurance_run(
            output_dir=output_dir,
            open_items=source_rows(),
            evidence_rows=[],
            assumptions={"amount_tolerance": "0"},
            source_root=source_dir,
            source_receipts=receipts,
        )

    assert not (output_dir / "prepared_records.json").exists()


@pytest.mark.parametrize("status", ["needs_review", "unsupported_source_layout"])
def test_unqualified_source_cannot_emit_plausible_prepared_rows(
    tmp_path: Path,
    status: str,
) -> None:
    assurance = load_assurance()
    source_dir = tmp_path / "source"
    output_dir = tmp_path / "output"
    source_dir.mkdir()
    output_dir.mkdir()
    source = source_dir / "source.csv"
    source.write_text(
        "document,amount\nINV-1,100.00\nINV-2,25.50\n",
        encoding="utf-8",
    )
    receipts = assurance.build_source_receipts(source_dir, [source])
    decisions, errors = assurance.build_reviewed_source_decisions(
        input_root=source_dir,
        source_receipts=receipts,
        adapter_families={"source.csv": "open_items_text_v1"},
        assumptions={
            "reviewed_source_decisions": {"source.csv": reviewed_source_decision()}
        },
    )
    assert errors == {}
    qualification = assurance.validate_source_qualification(
        {
            "schema_version": "vera.source_qualification.v1",
            "qualification_id": f"qualification.{status}",
            "adapter_id": "open_items_text_v1",
            "adapter_version": "2",
            "source_family": "csv.open_items",
            "status": status,
            "source_artifact_refs": [receipts[0]["artifact_id"]],
            "reviewed_mapping_ref": decisions["source.csv"]["decision_id"],
            "candidate_row_count": 2,
            "emitted_row_count": 0,
            "controls": [
                {
                    "control_id": "reviewed_mapping",
                    "required": True,
                    "status": (
                        "not_assessed" if status == "needs_review" else "failed"
                    ),
                    "evidence_refs": [receipts[0]["artifact_id"]],
                    "detail": "black-box hold",
                }
            ],
            "limitations": ["black-box hold"],
        }
    )

    with pytest.raises(ValueError, match="cannot emit plausible prepared rows"):
        assurance.prepare_assurance_run(
            output_dir=output_dir,
            open_items=source_rows(),
            evidence_rows=[],
            assumptions={"amount_tolerance": "0"},
            source_root=source_dir,
            source_receipts=receipts,
            reviewed_source_decisions=[decisions["source.csv"]],
            source_qualifications=[qualification],
        )

    assert not (output_dir / "prepared_records.json").exists()


def test_source_qualification_requires_reviewed_mapping_reference(
    tmp_path: Path,
) -> None:
    assurance = load_assurance()
    source_dir = tmp_path / "source"
    output_dir = tmp_path / "output"
    source_dir.mkdir()
    output_dir.mkdir()
    source = source_dir / "source.csv"
    source.write_text(
        "document,amount\nINV-1,100.00\nINV-2,25.50\n",
        encoding="utf-8",
    )
    receipts = assurance.build_source_receipts(source_dir, [source])
    decisions, errors = assurance.build_reviewed_source_decisions(
        input_root=source_dir,
        source_receipts=receipts,
        adapter_families={"source.csv": "open_items_text_v1"},
        assumptions={
            "reviewed_source_decisions": {"source.csv": reviewed_source_decision()}
        },
    )
    assert errors == {}
    qualification = qualified_sources(
        assurance,
        {
            "source_receipts": receipts,
            "reviewed_source_decisions": list(decisions.values()),
        },
    )[0]
    qualification["reviewed_mapping_ref"] = None

    with pytest.raises(ValueError, match="reviewed mapping reference"):
        assurance.prepare_assurance_run(
            output_dir=output_dir,
            open_items=source_rows(),
            evidence_rows=[],
            assumptions={"amount_tolerance": "0"},
            source_root=source_dir,
            source_receipts=receipts,
            reviewed_source_decisions=list(decisions.values()),
            source_qualifications=[qualification],
        )

    assert not (output_dir / "prepared_records.json").exists()


def test_wrong_workbook_cannot_fall_back_to_json_and_leaves_no_final_tree(
    tmp_path: Path,
) -> None:
    assurance, _, output_dir, rows, context = prepared_case(tmp_path)
    workbook_path = output_dir / "audit.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Other"
    sheet.append(["record_id", "note"])
    for row in rows:
        sheet.append([row["record_id"], "not addressed"])
    workbook.save(workbook_path)

    with pytest.raises(ValueError, match="exact material record-id set"):
        assurance.finalize_assurance_run(
            output_dir=output_dir,
            context=context,
            reconciliation_rows=rows,
            allocation_ledgers=[],
            checks=[{"check": "row_count", "status": "PASS"}],
            review_rows=[
                {
                    "record_id": row["record_id"],
                    "review_status": "PASS",
                    "reviewer_ref": "reviewer.test",
                    "reviewed_on": "2026-07-25",
                }
                for row in rows
            ],
            source_qualifications=qualified_sources(assurance, context),
            declared_outputs=[workbook_path],
            workbook_name=workbook_path.name,
        )

    assert not (output_dir / "assurance_final_outputs").exists()
    assert not (output_dir / "assurance_receipts.json").exists()
    assert not (output_dir / "reconciliation_results.json").exists()


def test_workbook_with_unaddressed_material_figure_is_withheld(
    tmp_path: Path,
) -> None:
    assurance, _, output_dir, rows, context = prepared_case(tmp_path)
    workbook_path = output_dir / "audit.xlsx"
    workbook_path.write_bytes(workbook_bytes(tmp_path))
    workbook = load_workbook(workbook_path)
    summary = workbook.create_sheet("Summary")
    summary.append(["total", "125.50"])
    workbook.save(workbook_path)

    with pytest.raises(ValueError, match="material figure without a record address"):
        assurance.finalize_assurance_run(
            output_dir=output_dir,
            context=context,
            reconciliation_rows=rows,
            allocation_ledgers=[],
            checks=[{"check": "row_count", "status": "PASS"}],
            review_rows=[
                {
                    "record_id": row["record_id"],
                    "review_status": "PASS",
                    "reviewer_ref": "reviewer.test",
                    "reviewed_on": "2026-07-25",
                }
                for row in rows
            ],
            source_qualifications=qualified_sources(assurance, context),
            declared_outputs=[workbook_path],
            workbook_name=workbook_path.name,
        )

    assert not (output_dir / "assurance_final_outputs").exists()


def test_word_report_with_unaddressed_material_figure_is_withheld(
    tmp_path: Path,
) -> None:
    assurance, _, output_dir, rows, context = prepared_case(tmp_path)
    workbook_path = output_dir / "audit.xlsx"
    workbook_path.write_bytes(workbook_bytes(tmp_path))
    report_path = output_dir / "report.docx"
    document = Document()
    document.add_paragraph("Unaddressed total: 125.50 EUR")
    document.save(report_path)

    with pytest.raises(ValueError, match="material figures without record addresses"):
        assurance.finalize_assurance_run(
            output_dir=output_dir,
            context=context,
            reconciliation_rows=rows,
            allocation_ledgers=[],
            checks=[{"check": "row_count", "status": "PASS"}],
            review_rows=[
                {
                    "record_id": row["record_id"],
                    "review_status": "PASS",
                    "reviewer_ref": "reviewer.test",
                    "reviewed_on": "2026-07-25",
                }
                for row in rows
            ],
            source_qualifications=qualified_sources(assurance, context),
            declared_outputs=[workbook_path, report_path],
            workbook_name=workbook_path.name,
        )

    assert not (output_dir / "assurance_final_outputs").exists()


def test_native_json_with_unaddressed_material_figure_is_withheld(
    tmp_path: Path,
) -> None:
    assurance, _, output_dir, rows, context = prepared_case(tmp_path)
    workbook_path = output_dir / "audit.xlsx"
    workbook_path.write_bytes(workbook_bytes(tmp_path))
    native_json_path = output_dir / "summary.json"
    stable_write(native_json_path, {"total_amount": "125.50"})

    with pytest.raises(ValueError, match="material figure without a record address"):
        assurance.finalize_assurance_run(
            output_dir=output_dir,
            context=context,
            reconciliation_rows=rows,
            allocation_ledgers=[],
            checks=[{"check": "row_count", "status": "PASS"}],
            review_rows=[
                {
                    "record_id": row["record_id"],
                    "review_status": "PASS",
                    "reviewer_ref": "reviewer.test",
                    "reviewed_on": "2026-07-25",
                }
                for row in rows
            ],
            source_qualifications=qualified_sources(assurance, context),
            declared_outputs=[workbook_path, native_json_path],
            workbook_name=workbook_path.name,
        )

    assert not (output_dir / "assurance_final_outputs").exists()


def test_control_json_with_unaddressed_check_amount_is_withheld(
    tmp_path: Path,
) -> None:
    assurance, _, output_dir, rows, context = prepared_case(tmp_path)
    workbook_path = output_dir / "audit.xlsx"
    workbook_path.write_bytes(workbook_bytes(tmp_path))

    with pytest.raises(ValueError, match="material figure without a record address"):
        assurance.finalize_assurance_run(
            output_dir=output_dir,
            context=context,
            reconciliation_rows=rows,
            allocation_ledgers=[],
            checks=[
                {
                    "check": "row_count",
                    "status": "PASS",
                    "difference": "1.00",
                }
            ],
            review_rows=[
                {
                    "record_id": row["record_id"],
                    "review_status": "PASS",
                    "reviewer_ref": "reviewer.test",
                    "reviewed_on": "2026-07-25",
                }
                for row in rows
            ],
            source_qualifications=qualified_sources(assurance, context),
            declared_outputs=[workbook_path],
            workbook_name=workbook_path.name,
        )

    assert not (output_dir / "assurance_final_outputs").exists()


def test_expected_final_output_hardlink_is_rejected(tmp_path: Path) -> None:
    assurance, output_dir, _ = finalize_case(tmp_path)
    boundary = output_dir / "assurance_final_outputs"
    target = boundary / "report.txt"
    external = tmp_path / "external-report.txt"
    shutil.copyfile(target, external)
    target.unlink()
    os.link(external, target)

    with pytest.raises(ValueError, match="hardlinked"):
        assurance.validate_assurance_run(output_dir)


def test_rehashed_changed_source_is_rejected_by_bounded_locator(
    tmp_path: Path,
) -> None:
    assurance, output_dir, _ = finalize_case(tmp_path)
    source = tmp_path / "source" / "source.csv"
    source.write_text(
        "document,amount\nINV-1,777.00\nINV-2,888.00\n",
        encoding="utf-8",
    )
    seal_path = output_dir / "assurance_receipts.json"
    seal = json.loads(seal_path.read_text(encoding="utf-8"))
    raw = source.read_bytes()
    seal["source_receipts"][0]["byte_count"] = len(raw)
    seal["source_receipts"][0]["sha256"] = hashlib.sha256(raw).hexdigest()
    reseal_outer(assurance, seal)
    stable_write(seal_path, seal)

    with pytest.raises(ValueError, match="source receipt identity is stale"):
        assurance.validate_assurance_run(output_dir)


def test_rehashed_invalid_decision_is_semantically_rejected(tmp_path: Path) -> None:
    assurance, output_dir, _ = finalize_case(tmp_path)
    seal_path = output_dir / "assurance_receipts.json"
    seal = json.loads(seal_path.read_text(encoding="utf-8"))
    decision = seal["reviewed_source_decisions"][0]
    decision["content"]["role"] = "forged-role"
    decision["content_sha256"] = assurance.canonical_json_sha256(decision["content"])
    reseal_outer(assurance, seal)
    stable_write(seal_path, seal)

    with pytest.raises(ValueError, match="source role"):
        assurance.validate_assurance_run(output_dir)


@pytest.mark.parametrize(
    ("mutation", "expected"),
    [
        ("decision_id", "decision identity is stale"),
        ("source_path", "content is stale"),
        ("adapter", "adapter binding is stale"),
        ("future_review", "cannot be after the sealed run date"),
    ],
)
def test_rehashed_decision_identity_adapter_path_and_date_are_replayed(
    tmp_path: Path,
    mutation: str,
    expected: str,
) -> None:
    assurance, output_dir, _ = finalize_case(tmp_path)
    seal_path = output_dir / "assurance_receipts.json"
    seal = json.loads(seal_path.read_text(encoding="utf-8"))
    decision = seal["reviewed_source_decisions"][0]
    if mutation == "decision_id":
        decision["decision_id"] = "source_mapping.forged"
    elif mutation == "source_path":
        decision["content"]["source_path"] = "other.csv"
        decision["content_sha256"] = assurance.canonical_json_sha256(
            decision["content"]
        )
    elif mutation == "adapter":
        decision["adapter_id"] = "forged.adapter"
        decision["content"]["adapter_family"] = "forged.adapter"
        decision["content_sha256"] = assurance.canonical_json_sha256(
            decision["content"]
        )
    else:
        decision["reviewed_on"] = "2999-01-01"
    reseal_outer(assurance, seal)
    stable_write(seal_path, seal)

    with pytest.raises(ValueError, match=expected):
        assurance.validate_assurance_run(output_dir)


def test_rehashed_forged_source_locator_is_rejected(tmp_path: Path) -> None:
    assurance, output_dir, _ = finalize_case(tmp_path)
    seal_path = output_dir / "assurance_receipts.json"
    numeric_path = output_dir / "numeric_evidence_ledger.json"
    seal = json.loads(seal_path.read_text(encoding="utf-8"))
    numeric = seal["numeric_evidence_ledger"]
    numeric["entries"][1]["source"]["locator"] = (
        '{"kind":"csv_record_value","record_id":"open-2",' '"row":999,"value_column":1}'
    )
    reseal_numeric(assurance, numeric)
    stable_write(numeric_path, numeric)
    reseal_outer(assurance, seal)
    stable_write(seal_path, seal)

    with pytest.raises(
        ValueError,
        match="record/source/output closure|output value locator is stale",
    ):
        assurance.validate_assurance_run(output_dir)


def test_rehashed_equal_amount_prepared_identity_swap_is_rejected(
    tmp_path: Path,
) -> None:
    assurance, source, output_dir, rows, context = prepared_case(tmp_path)
    source.write_text(
        "document,amount\nINV-1,50.00\nINV-2,50.00\n",
        encoding="utf-8",
    )
    receipts = assurance.build_source_receipts(source.parent, [source])
    decisions, errors = assurance.build_reviewed_source_decisions(
        input_root=source.parent,
        source_receipts=receipts,
        adapter_families={"source.csv": "open_items_text_v1"},
        assumptions={
            "reviewed_source_decisions": {"source.csv": reviewed_source_decision()}
        },
    )
    assert errors == {}
    equal_rows = copy.deepcopy(rows)
    for row in equal_rows:
        row["amount"] = "50.00"
    pre_context = {
        "source_receipts": receipts,
        "reviewed_source_decisions": list(decisions.values()),
    }
    qualifications = qualified_sources(assurance, pre_context)
    context = assurance.prepare_assurance_run(
        output_dir=output_dir,
        open_items=equal_rows,
        evidence_rows=[],
        assumptions={"amount_tolerance": "0"},
        source_root=source.parent,
        source_receipts=receipts,
        reviewed_source_decisions=list(decisions.values()),
        source_qualifications=qualifications,
    )
    workbook_path = output_dir / "audit.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reconciliation detail"
    sheet.append(["record_id", "amount"])
    for row in equal_rows:
        sheet.append([row["record_id"], row["amount"]])
    workbook.save(workbook_path)
    assurance.finalize_assurance_run(
        output_dir=output_dir,
        context=context,
        reconciliation_rows=equal_rows,
        allocation_ledgers=[],
        checks=[{"check": "row_count", "status": "PASS"}],
        review_rows=[
            {
                "record_id": row["record_id"],
                "review_status": "PASS",
                "reviewer_ref": "reviewer.test",
                "reviewed_on": "2026-07-25",
            }
            for row in equal_rows
        ],
        source_qualifications=qualifications,
        declared_outputs=[workbook_path],
        workbook_name=workbook_path.name,
    )
    prepared_path = output_dir / "prepared_records.json"
    prepared = json.loads(prepared_path.read_text(encoding="utf-8"))
    prepared["open_items"][1]["record_id"] = "forged-second-record"
    prepared["open_items"][1]["source_row"] = 999
    stable_write(prepared_path, prepared)
    seal_path = output_dir / "assurance_receipts.json"
    seal = json.loads(seal_path.read_text(encoding="utf-8"))
    raw = prepared_path.read_bytes()
    seal["prepared_receipt"]["byte_count"] = len(raw)
    seal["prepared_receipt"]["sha256"] = hashlib.sha256(raw).hexdigest()
    reseal_outer(assurance, seal)
    stable_write(seal_path, seal)

    with pytest.raises(ValueError, match="absent from prepared records"):
        assurance.validate_assurance_run(output_dir)


def test_rehashed_second_workbook_row_is_rejected_by_record_value_replay(
    tmp_path: Path,
) -> None:
    assurance, output_dir, _ = finalize_case(tmp_path)
    workbook_path = output_dir / "assurance_final_outputs" / "audit.xlsx"
    workbook = load_workbook(workbook_path)
    workbook["Reconciliation detail"]["B3"] = "999.00"
    workbook.save(workbook_path)
    seal_path = output_dir / "assurance_receipts.json"
    inventory_path = output_dir / "final_output_inventory.json"
    seal = json.loads(seal_path.read_text(encoding="utf-8"))
    inventory = seal["final_output_inventory"]
    receipt = next(
        item for item in inventory["artifact_receipts"] if item["path"] == "audit.xlsx"
    )
    raw = workbook_path.read_bytes()
    receipt["byte_count"] = len(raw)
    receipt["sha256"] = hashlib.sha256(raw).hexdigest()
    inventory_content = {
        "schema_version": inventory["schema_version"],
        "boundary_root": inventory["boundary_root"],
        "declared_paths": inventory["declared_paths"],
        "artifact_receipts": inventory["artifact_receipts"],
    }
    inventory["content_sha256"] = assurance.canonical_json_sha256(inventory_content)
    stable_write(inventory_path, inventory)
    reseal_outer(assurance, seal)
    stable_write(seal_path, seal)

    with pytest.raises(
        ValueError,
        match="record/source/output closure|output value locator is stale",
    ):
        assurance.validate_assurance_run(output_dir)


def test_late_failure_preserves_exact_prior_assurance_tree(tmp_path: Path) -> None:
    assurance, output_dir, _ = finalize_case(tmp_path)

    def image() -> dict[str, bytes]:
        return {
            path.relative_to(output_dir).as_posix(): path.read_bytes()
            for path in sorted(output_dir.rglob("*"))
            if path.is_file()
            and (
                path.name.startswith("assurance_")
                or "assurance_final_outputs" in path.parts
                or path.name
                in {
                    "final_output_inventory.json",
                    "numeric_evidence_ledger.json",
                    "reconciliation_results.json",
                }
            )
        }

    before = image()
    workbook_path = output_dir / "wrong.xlsx"
    workbook = Workbook()
    workbook.active.title = "Other"
    workbook.active.append(["record_id", "note"])
    workbook.save(workbook_path)
    context = {
        key: value
        for key, value in json.loads(
            (output_dir / "assurance_receipts.json").read_text(encoding="utf-8")
        ).items()
        if key
        in {
            "run_date",
            "source_root",
            "source_receipts",
            "reviewed_source_decisions",
            "source_qualifications",
            "implementation_receipts",
            "prepared_receipt",
        }
    }
    prepared = json.loads(
        (output_dir / "prepared_records.json").read_text(encoding="utf-8")
    )

    with pytest.raises(ValueError, match="exact material record-id set"):
        assurance.finalize_assurance_run(
            output_dir=output_dir,
            context=context,
            reconciliation_rows=prepared["open_items"],
            allocation_ledgers=[],
            checks=[{"check": "row_count", "status": "PASS"}],
            review_rows=[
                {
                    "record_id": row["record_id"],
                    "review_status": "PASS",
                    "reviewer_ref": "reviewer.test",
                    "reviewed_on": "2026-07-25",
                }
                for row in prepared["open_items"]
            ],
            source_qualifications=context["source_qualifications"],
            declared_outputs=[workbook_path],
            workbook_name=workbook_path.name,
        )

    assert image() == before
