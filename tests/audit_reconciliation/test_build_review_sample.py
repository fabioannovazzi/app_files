from __future__ import annotations

import importlib.util
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook


SCRIPTS = (
    Path(__file__).resolve().parents[2]
    / "plugins"
    / "audit-reconciliation"
    / "scripts"
)
SCRIPT = SCRIPTS / "build_review_sample.py"


def load_review_sample():
    sys.path.insert(0, str(SCRIPTS))
    spec = importlib.util.spec_from_file_location("audit_reconciliation_review_sample", SCRIPT)
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def write_reconciliation_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reconciliation detail"
    headers = [
        "record_id",
        "reconciliation_status",
        "rule_applied",
        "amount",
        "balance",
        "document_no",
        "document_date",
        "document_key",
        "expected_side",
        "source_file",
        "source_page",
        "source_row",
        "matched_evidence_id",
        "matched_evidence_amounts",
        "matched_evidence_reference",
    ]
    sheet.append(headers)
    sheet.append(
        [
            "open-high",
            "open_supported",
            "internal_booking_open_support",
            "250000.00",
            "250000.00",
            "INV-HIGH",
            "2023-12-19",
            "HIGH|2023",
            "supplier",
            "open-items.pdf",
            "7",
            "61",
            "ledger-high",
            "amount=250000.00",
            "file=ledger.pdf; page=83; row=50",
        ]
    )
    sheet.append(
        [
            "open-group-a",
            "open_supported",
            "grouped_open_amount_internal_booking_support",
            "100.00",
            "100.00",
            "INV-GROUP",
            "2023-03-31",
            "GROUP|2023",
            "customer",
            "open-items.pdf",
            "1",
            "45",
            "ledger-group",
            "amount=300.00; group_open_amount_total=300.00; group_rows=3",
            "file=ledger.pdf; page=4; row=64",
        ]
    )
    sheet.append(
        [
            "open-group-b",
            "open_supported",
            "grouped_open_amount_internal_booking_support",
            "100.00",
            "100.00",
            "INV-GROUP",
            "2023-03-31",
            "GROUP|2023",
            "customer",
            "open-items.pdf",
            "1",
            "49",
            "ledger-group",
            "amount=300.00; group_open_amount_total=300.00; group_rows=3",
            "file=ledger.pdf; page=4; row=64",
        ]
    )
    sheet.append(
        [
            "open-group-c",
            "open_supported",
            "grouped_open_amount_internal_booking_support",
            "100.00",
            "100.00",
            "INV-GROUP",
            "2023-03-31",
            "GROUP|2023",
            "customer",
            "open-items.pdf",
            "1",
            "53",
            "ledger-group",
            "amount=300.00; group_open_amount_total=300.00; group_rows=3",
            "file=ledger.pdf; page=4; row=64",
        ]
    )
    sheet.append(
        [
            "closed-row",
            "closed",
            "direct_external_or_documented",
            "1000.00",
            "1000.00",
            "INV-CLOSED",
            "2023-02-01",
            "CLOSED|2023",
            "customer",
            "open-items.pdf",
            "2",
            "1",
            "bank-1",
            "amount=1000.00",
            "file=bank.pdf; page=1; row=1",
        ]
    )
    workbook.save(path)


def test_build_review_sample_selects_material_and_grouped_rows(tmp_path):
    review_sample = load_review_sample()
    workbook_path = tmp_path / "riconciliazione.xlsx"
    write_reconciliation_workbook(workbook_path)

    rows = review_sample.load_reconciliation_rows(workbook_path)
    sample = review_sample.build_review_sample(rows, count=2)

    documents = [row["documento"] for row in sample.selected_rows]
    assert documents == ["INV-HIGH", "INV-GROUP"]
    assert len(sample.related_rows) == 3
    assert any("sommando più righe" in row["lettura_operativa"] for row in sample.selected_rows)


def test_review_sample_outputs_italian_operational_text(tmp_path):
    review_sample = load_review_sample()
    workbook_path = tmp_path / "riconciliazione.xlsx"
    write_reconciliation_workbook(workbook_path)
    rows = review_sample.load_reconciliation_rows(workbook_path)
    sample = review_sample.build_review_sample(rows, count=2)

    output_path = review_sample.write_review_sample_workbook(tmp_path / "campione.xlsx", sample)
    request_path = review_sample.write_review_request(tmp_path / "richiesta.md", sample)

    workbook = load_workbook(output_path)
    assert workbook.sheetnames == [
        "movimenti_da_controllare",
        "righe_collegate",
        "domande_per_revisione",
        "criteri_di_scelta",
    ]
    sheet_text = "\n".join(
        str(cell.value)
        for row in workbook["movimenti_da_controllare"].iter_rows()
        for cell in row
        if cell.value is not None
    )
    request_text = request_path.read_text(encoding="utf-8")

    assert "open_supported" not in sheet_text
    assert "grouped_open_amount_internal_booking_support" not in sheet_text
    assert "page=" not in sheet_text
    assert "row=" not in sheet_text
    assert "open_supported" not in request_text
    assert "pagina" in request_text
    assert "Risposta utile" in request_text
