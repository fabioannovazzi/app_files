"""Parse "print-friendly" Excel journal exports into tabular data."""

from __future__ import annotations

import logging
import math
import os
import re
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple
from zipfile import BadZipFile

import openpyxl
import polars as pl

from modules.utilities.utils import get_schema_and_column_names

LOGGER = logging.getLogger(__name__)


@dataclass
class SectionInfo:
    header_row: int
    dare_col: int
    avere_col: int
    debit_offset: int
    credit_offset: int
    confidence: float
    length: int


_DATE_TOKENS = re.compile(
    r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})|(\d{1,2}\.\d{1,2}\.\d{2,4})|(\d{4}-\d{1,2}-\d{1,2})"
)
_DATE_FORMATS = [
    "%d/%m/%Y",
    "%d-%m-%Y",
    "%d.%m.%Y",
    "%Y-%m-%d",
    "%d/%m/%y",
    "%m/%d/%Y",
    "%Y/%m/%d",
]


def _extract_date_token(text: str) -> Optional[str]:
    """Return the first date-like token in *text*, or ``None``."""

    m = _DATE_TOKENS.search(text)
    return m.group(0) if m else None


def _parse_token(token: str) -> Optional[date]:
    """Try parsing a date token with known formats."""

    for fmt in _DATE_FORMATS:
        try:
            parsed = datetime.strptime(token, fmt).date()
        except ValueError:
            continue
        except Exception as exc:  # pragma: no cover - unexpected parsing error
            LOGGER.debug("Failed to parse date token %r with %s: %s", token, fmt, exc)
            continue
        if parsed.year < 1:
            continue
        return parsed
    return None


def derive_date_column(rows: Sequence[Sequence[Any]]) -> List[Optional[date]]:
    """Return per-row parsed dates by scanning each row for a date token."""

    result: List[Optional[date]] = []
    current: Optional[date] = None
    for row in rows:
        token = None
        for val in row[:3]:
            if isinstance(val, str) and val.strip():
                token = _extract_date_token(val)
                if token:
                    break
        if token:
            parsed = _parse_token(token)
            if parsed:
                current = parsed
        result.append(current)
    return result


def _extract_leading_context_date(
    row: Sequence[Any],
    *,
    allow_generic_scan: bool,
) -> date | None:
    """Parse a section date from leading cells while ignoring page banners."""

    leading_text = " ".join(
        _nonempty_text(value).lower() for value in row[:12] if _nonempty_text(value)
    )
    if "libro giornale" in leading_text:
        return None

    first_value = row[0] if row else None
    second_value = row[1] if len(row) > 1 else None
    for pos, value in enumerate((first_value, second_value)):
        if not isinstance(value, str):
            continue
        text = value.strip()
        if not text:
            continue
        token = _extract_date_token(text)
        if not token:
            continue
        lower = text.lower()
        parsed = _parse_token(token)
        if parsed is None:
            continue
        if re.search(r"^\d{1,8}\s+\d{1,2}[/-]\d{1,2}[/-]\d{2,4}", text):
            return parsed
        if any(
            marker in lower
            for marker in ("movimento", "registrazione", "reg.", "n. reg", "nr. reg")
        ):
            return parsed
        if pos == 1 and _is_number_like(first_value):
            return parsed

    if not allow_generic_scan:
        return None
    if any(marker in leading_text for marker in ("num.", "scad.", "fattura")):
        return None
    for value in row[:3]:
        parsed = _parse_date_with_context(value, allow_numeric=False)
        if parsed is not None:
            return parsed
    return None


def _derive_date_column_from_index(
    rows: Sequence[Sequence[Any]],
    date_idx: int | None,
    *,
    debit_idx: int | None = None,
    credit_idx: int | None = None,
) -> List[Optional[date]]:
    """Carry forward dates from an explicit date column when available."""

    if date_idx is None or date_idx < 0:
        return derive_date_column(rows)

    result: List[Optional[date]] = []
    current: Optional[date] = None
    for row in rows:
        parsed: Optional[date] = None
        has_amount = False
        if debit_idx is not None and 0 <= debit_idx < len(row):
            has_amount = _to_float(row[debit_idx]) is not None
        if not has_amount and credit_idx is not None and 0 <= credit_idx < len(row):
            has_amount = _to_float(row[credit_idx]) is not None
        if date_idx < len(row):
            parsed = _parse_date_with_context(row[date_idx], allow_numeric=True)
        if parsed is None:
            parsed = _extract_leading_context_date(
                row,
                allow_generic_scan=not has_amount,
            )
        if parsed is not None:
            current = parsed
        result.append(current)
    return result


def _parse_date(value: Any) -> date | None:
    """Return a ``datetime.date`` parsed from common journal formats."""

    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, (int, float)):
        # Ignore small integers commonly used as movement/row numbers.
        try:
            serial = float(value)
        except (TypeError, ValueError):  # pragma: no cover - defensive cast guard
            return None
        if not (20000 <= serial <= 60000):
            return None
        try:
            from openpyxl.utils.datetime import from_excel

            parsed = from_excel(serial)
            if isinstance(parsed, datetime):
                return parsed.date()
            if isinstance(parsed, date):
                return parsed
        except (
            TypeError,
            ValueError,
            OverflowError,
        ):  # pragma: no cover - non-date numerics
            return None
    if isinstance(value, str):
        s = value.strip()
        token = _extract_date_token(s)
        if token:
            return _parse_token(token)
    return None


def _parse_date_with_context(value: Any, *, allow_numeric: bool) -> date | None:
    """Return a date parsed from *value* with optional numeric serial support."""

    if isinstance(value, (int, float)) and not allow_numeric:
        return None
    return _parse_date(value)


def build_token_map(language: str = "auto") -> Dict[str, Sequence[str]]:
    """Return a dictionary of header tokens for multiple languages.

    Extend these lists to support additional languages. For example::

        token_map["debit"].append("soll")  # German
        token_map["credit"].append("haben")

    Parameters
    ----------
    language:
        Currently ignored; retained for future customisation.
    """
    return {
        "debit": ["dare", "debit", "debitore", "dr", "charge"],
        "credit": ["avere", "credit", "credito", "cr", "creditore"],
        "account": ["conto", "account"],
        "account_desc": ["descrizione conto", "account description"],
        "line_desc": [
            "descrizione dell'operazione",
            "descrizione operazione",
            "operation description",
            "description",
            "memo",
            "narration",
        ],
        "rowno": [
            "riga",
            "row",
            "line",
            "entry",
            "number",
            "n.reg",
            "nr reg",
            "nr. reg",
            "n. reg",
            "numero registrazione",
            "nr. prog",
            "nr prog",
        ],
        "date_hdr": ["data registrazione", "data", "date"],
        "totals": [
            "totale",
            "progressivo",
            "saldo",
            "riporto",
            "tot",
            "total",
            "carry forward",
        ],
    }


def _split_words(text: str) -> List[str]:
    return [w for w in re.split(r"[^a-z0-9]+", text) if w]


def _match_token(cell: str, token: str) -> bool:
    if cell == token:
        return True
    if not token or len(token) < 3:
        return False
    if " " in token:
        return token in cell
    return token in _split_words(cell)


def _row_has(tokens: Dict[str, Sequence[str]], row: Sequence[Any], key: str) -> bool:
    lower = [str(x).strip().lower() for x in row if x is not None and str(x).strip()]
    for cell in lower:
        for token in tokens.get(key, []):
            if _match_token(cell, token):
                return True
    return False


def _find_header_rows(
    rows: Sequence[Sequence[Any]], tokens: Dict[str, Sequence[str]]
) -> List[int]:
    hdr: List[int] = []
    for i, row in enumerate(rows):
        if _row_has(tokens, row, "debit") and _row_has(tokens, row, "credit"):
            hdr.append(i)
    return hdr


def _find_token_index(row: Sequence[Any], options: Sequence[str]) -> Optional[int]:
    for idx, cell in enumerate(row):
        norm = _norm(cell)
        if not norm:
            continue
        for opt in options:
            if _match_token(norm, opt):
                return idx
    return None


def _find_token_indexes(row: Sequence[Any], options: Sequence[str]) -> List[int]:
    """Return every header position that matches one of *options*."""

    matches: List[int] = []
    for idx, cell in enumerate(row):
        norm = _norm(cell)
        if not norm:
            continue
        if any(_match_token(norm, opt) for opt in options):
            matches.append(idx)
    return matches


def _header_text_signature(row: Sequence[Any]) -> str:
    cleaned = [re.sub(r"[^a-z0-9]+", "", str(x).lower()) for x in row if str(x).strip()]
    return "|".join(cleaned)


def _infer_offsets_for_section(
    rows: Sequence[Sequence[Any]],
    header_idx: int,
    dare_idx: int,
    avere_idx: int,
    lookahead: int = 30,
) -> Tuple[int, int, float]:
    votes_debit: Dict[int, int] = {}
    votes_credit: Dict[int, int] = {}

    def _bump(d: Dict[int, int], k: int) -> None:
        d[k] = d.get(k, 0) + 1

    for r in range(header_idx + 1, min(header_idx + 1 + lookahead, len(rows))):
        row = rows[r]
        if (
            0 <= dare_idx < len(row)
            and 0 <= avere_idx < len(row)
            and re.fullmatch(r"[-+]?\d+([.,]\d+)?", str(row[dare_idx]).replace(" ", ""))
            and re.fullmatch(
                r"[-+]?\d+([.,]\d+)?", str(row[avere_idx]).replace(" ", "")
            )
        ):
            continue
        for off in range(-2, 4):
            ci = dare_idx + off
            if (
                0 <= ci < len(row)
                and ci != avere_idx
                and re.fullmatch(r"[-+]?\d+([.,]\d+)?", str(row[ci]).replace(" ", ""))
            ):
                _bump(votes_debit, off)
            cj = avere_idx + off
            if (
                0 <= cj < len(row)
                and cj != dare_idx
                and re.fullmatch(r"[-+]?\d+([.,]\d+)?", str(row[cj]).replace(" ", ""))
            ):
                _bump(votes_credit, off)

    if votes_debit:
        debit_off = max(votes_debit, key=votes_debit.get)
        debit_conf = votes_debit[debit_off] / sum(votes_debit.values())
    else:
        debit_off, debit_conf = 0, 0.0

    if votes_credit:
        credit_off = max(votes_credit, key=votes_credit.get)
        credit_conf = votes_credit[credit_off] / sum(votes_credit.values())
    else:
        credit_off, credit_conf = 0, 0.0

    conf = min(1.0, (debit_conf + credit_conf) / 2.0)
    return debit_off, credit_off, conf


def _should_use_print_friendly(
    rows: Sequence[Sequence[Any]],
    tokens: Dict[str, Sequence[str]],
    header_rows: List[int],
) -> Tuple[bool, str, List[SectionInfo]]:
    """Decide whether *rows* resemble a paginated print-friendly export."""

    if not header_rows:
        return False, "no debit/credit header found", []

    # filter out headers that repeat too closely (<10 rows apart)
    cleaned: List[int] = []
    prev = -999
    for h in header_rows:
        if h - prev >= 10:
            cleaned.append(h)
            prev = h
    header_rows = cleaned

    sections: List[SectionInfo] = []
    for hi in header_rows:
        row = rows[hi]
        dare_idx = _find_token_index(row, tokens.get("debit", []))
        avere_idx = _find_token_index(row, tokens.get("credit", []))
        if dare_idx is None or avere_idx is None:
            continue
        d_off, c_off, conf = _infer_offsets_for_section(
            rows, hi, dare_idx, avere_idx, lookahead=40
        )
        next_hi = next((h for h in header_rows if h > hi), len(rows))
        sections.append(
            SectionInfo(
                header_row=hi,
                dare_col=dare_idx,
                avere_col=avere_idx,
                debit_offset=d_off,
                credit_offset=c_off,
                confidence=conf,
                length=next_hi - hi - 1,
            )
        )

    if not sections:
        return False, "no valid sections", []

    if len(sections) >= 2:
        return True, f"{len(sections)} header rows", sections

    sec = sections[0]

    if (sec.debit_offset != 0 or sec.credit_offset != 0) and sec.confidence >= 0.5:
        return (
            True,
            f"non-zero offsets (d={sec.debit_offset}, c={sec.credit_offset})",
            sections,
        )

    sig = _header_text_signature(rows[sec.header_row])
    repeats = 0
    for hi in header_rows[1:]:
        if _header_text_signature(rows[hi]) == sig:
            dist = hi - sections[0].header_row
            if 20 <= dist <= 120:
                repeats += 1
    if repeats >= 1:
        return True, "header pattern repeats at page-like distance", sections

    totals_tokens = [
        "totale",
        "progressivo",
        "saldo",
        "carry forward",
        "riporto",
        "tot",
    ]
    totals_hits = 0
    for row in rows:
        lower = " ".join([str(x).lower() for x in row if x is not None])
        if any(t in lower for t in totals_tokens):
            totals_hits += 1
    if totals_hits >= 3 and len(rows) > 200:
        return True, f"many totals/progressivo lines ({totals_hits})", sections

    # corridor blankness
    corridor_blank_signal = 0.0
    try:
        hi = sections[0].header_row
        dcol = sections[0].dare_col
        acol = sections[0].avere_col
        end = min(len(rows), hi + 1 + 40)
        blanks = total = 0
        for r in range(hi + 1, end):
            row = rows[r]
            for c in range(max(0, dcol - 1), min(len(row), dcol + 2)):
                total += 1
                blanks += 1 if str(row[c]).strip() in ("", "None") else 0
            for c in range(max(0, acol - 1), min(len(row), acol + 3)):
                total += 1
                blanks += 1 if str(row[c]).strip() in ("", "None") else 0
        if total > 0:
            corridor_blank_signal = blanks / total
    except Exception as e:
        logging.exception(e)
        corridor_blank_signal = 0.0

    if corridor_blank_signal >= 0.7 and (
        repeats >= 1 or sec.debit_offset != 0 or sec.credit_offset != 0
    ):
        return True, f"spacer corridor ~{corridor_blank_signal:.0%}", sections

    return False, "single header & zero offsets", sections


def _norm(val: Any) -> str:
    return str(val).strip().lower() if val is not None else ""


def _is_number_like(val: Any) -> bool:
    if isinstance(val, (int, float)):
        return not (isinstance(val, float) and math.isnan(val))
    if isinstance(val, str):
        txt = val.strip()
        if not txt:
            return False
        txt = txt.replace(".", "").replace(",", ".")
        return bool(re.fullmatch(r"[-+]?[0-9]*\.?[0-9]+", txt))
    return False


def _to_float(val: Any) -> Optional[float]:
    if isinstance(val, (int, float)):
        if isinstance(val, float) and math.isnan(val):
            return None
        return float(val)
    if isinstance(val, str):
        txt = val.strip()
        if not txt:
            return None
        txt = txt.replace(".", "").replace(",", ".")
        try:
            return float(txt)
        except ValueError:
            return None
    return None


def _find_header_cols(
    row: Sequence[Any], tokens: Dict[str, Sequence[str]]
) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    normed = [_norm(cell) for cell in row]
    for idx, cell in enumerate(normed):
        if not cell:
            continue
        for key, opts in tokens.items():
            if key == "totals" or key in mapping:
                continue
            if cell in opts:
                mapping[key] = idx
    for idx, cell in enumerate(normed):
        if not cell:
            continue
        for key, opts in tokens.items():
            if key == "totals" or key in mapping:
                continue
            if any(_match_token(cell, opt) for opt in opts):
                mapping[key] = idx
    return mapping


def _header_label_positions(row: Sequence[Any]) -> List[int]:
    positions: List[int] = []
    for idx, cell in enumerate(row):
        norm = _norm(cell)
        if not norm:
            continue
        if any(ch.isalpha() for ch in norm):
            positions.append(idx)
    return positions


def _header_span(
    label_positions: Sequence[int], idx: int, width: int
) -> Optional[Tuple[int, int]]:
    if idx < 0 or width <= 0:
        return None
    next_idx = next((pos for pos in label_positions if pos > idx), width)
    end = max(idx, min(width - 1, next_idx - 1))
    return (idx, end)


def _merged_ranges(
    ws: openpyxl.worksheet.worksheet.Worksheet,
) -> List[Tuple[int, int, int, int]]:
    ranges: List[Tuple[int, int, int, int]] = []
    merged = getattr(ws, "merged_cells", None)
    if merged is None:
        return ranges
    for cell_range in getattr(merged, "ranges", []):
        try:
            ranges.append(
                (
                    cell_range.min_row - 1,
                    cell_range.max_row - 1,
                    cell_range.min_col - 1,
                    cell_range.max_col - 1,
                )
            )
        except Exception:  # pragma: no cover - defensive guard
            continue
    return ranges


def _merged_span_for_cell(
    merged_ranges: Sequence[Tuple[int, int, int, int]],
    row_idx: int,
    col_idx: int,
) -> Optional[Tuple[int, int]]:
    for rmin, rmax, cmin, cmax in merged_ranges:
        if rmin <= row_idx <= rmax and cmin <= col_idx <= cmax:
            return (cmin, cmax)
    return None


def _candidate_columns(
    base_idx: int,
    span: Optional[Tuple[int, int]],
    width: int,
    *,
    right_window: int = 6,
) -> List[int]:
    if width <= 0:
        return []
    if span is not None:
        start, end = span
    else:
        start = max(0, base_idx - 2)
        end = min(width - 1, base_idx + right_window)
    return list(range(start, min(end, width - 1) + 1))


def _candidate_text_columns(
    base_idx: int,
    span: Optional[Tuple[int, int]],
    width: int,
    *,
    left_window: int = 2,
    right_window: int = 8,
) -> List[int]:
    """Return candidate columns for text-like fields near *base_idx*."""

    if width <= 0 or base_idx < 0:
        return []
    if span is not None:
        start, end = span
    else:
        start = base_idx
        end = base_idx
    start = max(0, start - left_window)
    end = min(width - 1, end + right_window)
    return list(range(start, end + 1))


def _nonempty_text(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _looks_like_account_code(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return False
        if float(value).is_integer():
            digits = len(str(int(abs(float(value)))))
            return 2 <= digits <= 12
        return False
    text = _nonempty_text(value)
    if not text:
        return False
    normalized = text.replace(" ", "")
    if len(normalized) <= 2 and normalized.isalpha():
        return False
    if re.fullmatch(r"[A-Z]?[0-9]+([./-][0-9A-Z]+)*", normalized, re.I):
        return True
    if any(ch.isdigit() for ch in text) and len(text) <= 32:
        return True
    return False


def _looks_like_integer_identifier(value: Any) -> bool:
    """Return whether *value* resembles an integer movement identifier."""

    if isinstance(value, int):
        return True
    if isinstance(value, float):
        return not math.isnan(value) and float(value).is_integer()
    text = _nonempty_text(value)
    if not text:
        return False
    compact = text.replace(" ", "")
    return bool(re.fullmatch(r"[A-Za-z]?\d{1,12}", compact))


def _looks_like_decimal_amount(value: Any) -> bool:
    """Return whether *value* looks like an amount with decimals."""

    if isinstance(value, float):
        return not math.isnan(value) and not float(value).is_integer()
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return False
        if "." in text or "," in text:
            return _is_number_like(text)
    return False


def _split_account_marker(value: Any) -> tuple[str, str]:
    """Return ``(marker, code)`` for account fragments such as ``G`` + ``361``."""

    text = _nonempty_text(value)
    if not text:
        return "", ""
    marker = ""
    code = text
    match = re.match(r"^([A-Za-z]{1,3})\s+(.+)$", text)
    if match:
        marker = match.group(1).strip()
        code = match.group(2).strip()
    return marker, code


def _score_account_column(rows: Sequence[Sequence[Any]], idx: int) -> float:
    if not rows:
        return -1.0
    non_empty = 0
    account_like = 0
    short_markers = 0
    long_text = 0
    seen: set[str] = set()
    for row in rows:
        if idx >= len(row):
            continue
        value = row[idx]
        text = _nonempty_text(value)
        if not text:
            continue
        non_empty += 1
        seen.add(text)
        if _looks_like_account_code(value):
            account_like += 1
        if len(text) <= 2 and text.isalpha():
            short_markers += 1
        if len(text) > 32:
            long_text += 1
    if non_empty == 0:
        return -1.0
    unique_ratio = len(seen) / non_empty
    return (
        account_like * 2.5
        + unique_ratio * 3.0
        + non_empty * 0.05
        - short_markers * 2.0
        - long_text * 1.5
    )


def _score_text_column(rows: Sequence[Sequence[Any]], idx: int) -> float:
    if not rows:
        return -1.0
    non_empty = 0
    alpha = 0
    long_text = 0
    numeric_like = 0
    seen: set[str] = set()
    for row in rows:
        if idx >= len(row):
            continue
        value = row[idx]
        text = _nonempty_text(value)
        if not text:
            continue
        non_empty += 1
        seen.add(text)
        if any(ch.isalpha() for ch in text):
            alpha += 1
        if len(text) >= 8:
            long_text += 1
        if _is_number_like(value):
            numeric_like += 1
    if non_empty == 0:
        return -1.0
    unique_ratio = len(seen) / non_empty
    return (
        alpha * 1.5
        + long_text * 1.5
        + unique_ratio * 2.0
        + non_empty * 0.05
        - numeric_like * 2.0
    )


def _score_rowno_column(rows: Sequence[Sequence[Any]], idx: int) -> float:
    if not rows:
        return -1.0
    non_empty = 0
    integer_like = 0
    decimal_like = 0
    alpha = 0
    seen: set[str] = set()
    for row in rows:
        if idx >= len(row):
            continue
        value = row[idx]
        text = _nonempty_text(value)
        if not text:
            continue
        non_empty += 1
        seen.add(text)
        if _looks_like_integer_identifier(value):
            integer_like += 1
        elif _looks_like_decimal_amount(value):
            decimal_like += 1
        if any(ch.isalpha() for ch in text):
            alpha += 1
    if non_empty == 0:
        return -1.0
    unique_ratio = len(seen) / non_empty
    return (
        integer_like * 3.0
        + min(unique_ratio, 0.4) * 1.0
        + non_empty * 0.05
        - decimal_like * 3.0
        - alpha * 1.0
    )


def _score_account_marker_column(rows: Sequence[Sequence[Any]], idx: int) -> float:
    if not rows:
        return -1.0
    non_empty = 0
    short_alpha = 0
    seen: set[str] = set()
    for row in rows:
        if idx >= len(row):
            continue
        text = _nonempty_text(row[idx])
        if not text:
            continue
        non_empty += 1
        seen.add(text)
        if len(text) <= 3 and text.isalpha():
            short_alpha += 1
    if non_empty == 0 or short_alpha == 0:
        return -1.0
    if len(seen) > 8:
        return -1.0
    return short_alpha * 2.0 + non_empty * 0.05


def _best_column(
    rows: Sequence[Sequence[Any]],
    candidates: Sequence[int],
    scorer,
) -> Optional[int]:
    best_idx: Optional[int] = None
    best_score = -1.0
    for idx in candidates:
        score = scorer(rows, idx)
        if score > best_score:
            best_idx = idx
            best_score = score
    return best_idx if best_score >= 0 else None


def _rank_columns(
    rows: Sequence[Sequence[Any]],
    candidates: Sequence[int],
    scorer,
) -> List[int]:
    """Return candidate columns ordered by descending scorer output."""

    scored: List[tuple[float, int]] = []
    for idx in sorted(set(candidates)):
        score = scorer(rows, idx)
        if score >= 0:
            scored.append((score, idx))
    scored.sort(key=lambda item: item[0], reverse=True)
    return [idx for _score, idx in scored]


def _merge_account_parts(marker_value: Any, account_value: Any) -> str | None:
    marker = _nonempty_text(marker_value)
    account_text = _nonempty_text(account_value)
    if not account_text:
        return None
    if not marker or len(marker) > 3 or not marker.isalpha():
        return account_text
    existing_marker, _existing_code = _split_account_marker(account_text)
    if existing_marker:
        return account_text
    return f"{marker} {account_text}"


def _normalise_row_identifier(value: Any) -> str | None:
    """Return a stable row or movement identifier from mixed cell values."""

    if value in (None, ""):
        return None
    if _is_number_like(value):
        try:
            numeric = float(value)
        except (TypeError, ValueError):
            return _nonempty_text(value) or None
        if not numeric.is_integer():
            return None
        return str(int(numeric))
    text = _nonempty_text(value)
    if not text:
        return None
    leading = re.match(r"^(\d{1,8})\b", text)
    if leading:
        return leading.group(1)
    if re.fullmatch(r"[A-Za-z]\s+\d{1,12}", text):
        return None
    compact = text.replace(" ", "")
    if re.fullmatch(r"[A-Za-z]?\d{1,12}", compact):
        return compact
    return text


def _clean_ascii_fragment(value: Any) -> str | None:
    text = _nonempty_text(value)
    if not text:
        return None
    cleaned = text.replace("\n", " ")
    cleaned = cleaned.replace("!", " ")
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" +-")
    return cleaned or None


def _extract_ascii_account(row: Sequence[Any], *, prefer_credit: bool) -> str | None:
    positions = [2, 1, 0, 3] if prefer_credit else [0, 1, 2, 3]
    for idx in positions:
        if idx >= len(row):
            continue
        cleaned = _clean_ascii_fragment(row[idx])
        if not cleaned:
            continue
        if cleaned.startswith("****"):
            continue
        if _looks_like_account_code(cleaned):
            return cleaned
    return None


def _extract_ascii_description(row: Sequence[Any]) -> str | None:
    for idx in range(6, min(len(row), 18)):
        cleaned = _clean_ascii_fragment(row[idx])
        if not cleaned:
            continue
        if cleaned.startswith("****"):
            continue
        return cleaned
    return None


def _extract_ascii_narrative(row: Sequence[Any]) -> str | None:
    parts: List[str] = []
    for idx in list(range(6, min(len(row), 18))) + list(range(9, min(len(row), 24))):
        cleaned = _clean_ascii_fragment(row[idx])
        if not cleaned:
            continue
        if "****" in cleaned or any(ch.isalpha() for ch in cleaned):
            parts.append(cleaned.replace("****", "").strip())
    if not parts:
        return None
    narrative = re.sub(r"\s+", " ", " ".join(parts)).strip()
    return narrative or None


def _extract_ascii_amount(row: Sequence[Any], *, debit: bool) -> float | None:
    ranges = (
        range(28, min(len(row), 34)),
        range(34, min(len(row), 41)),
    )
    indices = ranges[0] if debit else ranges[1]
    for idx in indices:
        value = _to_float(row[idx])
        if value is not None:
            return value
    return None


def _extract_ascii_separator_context(text: str) -> tuple[date | None, str | None]:
    date_token = _extract_date_token(text)
    parsed_date = _parse_token(date_token) if date_token else None
    movement: str | None = None
    movement_match = re.search(
        r"(\d{1,6})\s*-+\s*\d{2}/\d{2}/\d{4}",
        text,
    )
    if movement_match:
        movement = movement_match.group(1)
    return parsed_date, movement


def _looks_like_ascii_print_journal(rows: Sequence[Sequence[Any]]) -> bool:
    for row in rows[: min(len(rows), 20)]:
        text = " ".join(_nonempty_text(val) for val in row if _nonempty_text(val))
        normalized = text.lower()
        if "conto dare" in normalized and "conto avere" in normalized:
            return True
    return False


def _parse_ascii_print_journal(rows: Sequence[Sequence[Any]]) -> pl.DataFrame:
    """Parse ASCII-art journal exports that use pipe-style visual separators."""

    records: List[Dict[str, Any]] = []
    current_date: date | None = None
    current_movement: str | None = None
    current_narrative: str | None = None

    for row_idx, row in enumerate(rows, start=1):
        combined_text = " ".join(
            _nonempty_text(val) for val in row if _nonempty_text(val)
        )
        normalized = combined_text.lower()
        if not combined_text:
            continue

        if "conto dare" in normalized and "conto avere" in normalized:
            current_narrative = None
            parsed_date, movement = _extract_ascii_separator_context(combined_text)
            if parsed_date is not None:
                current_date = parsed_date
            if movement is not None:
                current_movement = movement
            continue

        if combined_text.startswith("+---------------+"):
            current_narrative = None
            parsed_date, movement = _extract_ascii_separator_context(combined_text)
            if parsed_date is not None:
                current_date = parsed_date
            if movement is not None:
                current_movement = movement
            continue

        embedded_date, embedded_movement = _extract_ascii_separator_context(
            combined_text
        )
        if embedded_date is not None and (
            "seguito registrazione" in normalized
            or "------------------" in combined_text
        ):
            current_date = embedded_date
            if embedded_movement is not None:
                current_movement = embedded_movement
            current_narrative = None
            continue

        debit = _extract_ascii_amount(row, debit=True)
        credit = _extract_ascii_amount(row, debit=False)
        description = _extract_ascii_description(row)
        narrative = _extract_ascii_narrative(row)

        if debit is None and credit is None:
            if narrative:
                current_narrative = narrative
            continue

        prefer_credit = credit is not None and debit is None
        account = _extract_ascii_account(row, prefer_credit=prefer_credit)
        if account is None:
            continue

        if narrative and current_narrative is None:
            current_narrative = narrative

        records.append(
            {
                "date": current_date,
                "movement_number": current_movement,
                "account": account,
                "account_desc": description,
                "line_desc": current_narrative,
                "debit_amount": debit,
                "credit_amount": credit,
                "amount": (debit or 0.0) - (credit or 0.0),
                "page_index": 0,
                "source_row": row_idx,
            }
        )

    if not records:
        return _empty_df()

    return _records_to_frame(records)


def _records_to_frame(records: Sequence[Dict[str, Any]]) -> pl.DataFrame:
    """Return canonical journal frame built from parsed record dictionaries."""

    if not records:
        return _empty_df()

    schema = {
        "date": pl.Date,
        "movement_number": pl.Utf8,
        "account": pl.Utf8,
        "account_desc": pl.Utf8,
        "line_desc": pl.Utf8,
        "debit_amount": pl.Float64,
        "credit_amount": pl.Float64,
        "amount": pl.Float64,
        "page_index": pl.Int64,
        "source_row": pl.Int64,
    }
    df = pl.DataFrame(records, schema=schema, strict=False)
    for col in ("debit_amount", "credit_amount"):
        columns, _ = get_schema_and_column_names(df)
        if col not in columns:
            df = df.with_columns(pl.lit(0.0).alias(col))

    df = df.with_columns(
        [
            pl.col("date").cast(pl.Date, strict=False),
            pl.col("movement_number").cast(pl.Utf8),
            pl.col("account").cast(pl.Utf8),
            pl.col("account_desc").cast(pl.Utf8),
            pl.col("line_desc").cast(pl.Utf8),
            pl.col("debit_amount").cast(pl.Float64, strict=False).fill_null(0.0),
            pl.col("credit_amount").cast(pl.Float64, strict=False).fill_null(0.0),
        ]
    )

    df = df.with_columns(
        (pl.col("debit_amount") - pl.col("credit_amount")).alias("amount")
    )

    df = df.rename(
        {
            "date": "data_registrazione",
            "movement_number": "riga",
            "account": "conto",
            "account_desc": "descrizione_conto",
            "line_desc": "descrizione_operazione",
            "debit_amount": "dare",
            "credit_amount": "avere",
        }
    )

    columns, _ = get_schema_and_column_names(df)
    drop_cols = [
        col for col in ("page_index", "amount", "source_row") if col in columns
    ]
    if drop_cols:
        df = df.drop(drop_cols)
    return df


def _select_amount_columns(
    rows: Sequence[Sequence[Any]],
    debit_candidates: Sequence[int],
    credit_candidates: Sequence[int],
) -> Tuple[Optional[int], Optional[int], float]:
    if not rows or not debit_candidates or not credit_candidates:
        return None, None, 0.0
    row_count = len(rows)
    candidates = sorted(set(debit_candidates) | set(credit_candidates))
    flags: Dict[int, List[bool]] = {}
    for idx in candidates:
        flags[idx] = [
            _is_number_like(row[idx]) if idx < len(row) else False for row in rows
        ]

    best_score = -1.0
    best_pair: Tuple[Optional[int], Optional[int]] = (None, None)
    for d_idx in debit_candidates:
        d_flags = flags.get(d_idx, [])
        if not d_flags:
            continue
        d_hits = sum(d_flags)
        for c_idx in credit_candidates:
            if c_idx == d_idx:
                continue
            c_flags = flags.get(c_idx, [])
            if not c_flags:
                continue
            any_hits = both_hits = 0
            for d_hit, c_hit in zip(d_flags, c_flags):
                if d_hit or c_hit:
                    any_hits += 1
                    if d_hit and c_hit:
                        both_hits += 1
            if any_hits == 0:
                continue
            exclusive_hits = any_hits - both_hits
            if exclusive_hits == 0:
                continue
            c_hits = sum(c_flags)
            coverage = (d_hits + c_hits) / row_count
            exclusivity = exclusive_hits / any_hits
            score = coverage * exclusivity
            for count in (d_hits, c_hits):
                if count / row_count > 0.9:
                    score *= 0.7
            if score > best_score:
                best_score = score
                best_pair = (d_idx, c_idx)

    if best_score < 0.05:
        return None, None, best_score
    return best_pair[0], best_pair[1], best_score


def _first_numeric(
    row: Sequence[Any], start: int, stop: Optional[int] = None
) -> Optional[Tuple[int, float]]:
    limit = stop if stop is not None else len(row)
    for i in range(start, min(limit, len(row))):
        if _is_number_like(row[i]):
            num = _to_float(row[i])
            if num is not None:
                return i, num
    return None


def _first_text(
    row: Sequence[Any], start: int = 0, stop: Optional[int] = None
) -> Optional[Tuple[int, str]]:
    limit = stop if stop is not None else len(row)
    for i in range(start, min(limit, len(row))):
        val = row[i]
        if isinstance(val, str) and val.strip():
            return i, val
    return None


def _infer_offsets(
    rows: Sequence[Sequence[Any]], header_map: Dict[str, int]
) -> Tuple[int, int]:
    d_idx = header_map.get("debit", -1)
    c_idx = header_map.get("credit", -1)
    d_off = c_off = 0
    found_d = found_c = False
    for row in rows:
        if d_idx >= 0 and not found_d:
            d = _first_numeric(row, d_idx)
            if d is not None:
                d_off = d[0] - d_idx
                found_d = True
        if c_idx >= 0 and not found_c:
            c = _first_numeric(row, c_idx)
            if c is not None:
                c_off = c[0] - c_idx
                found_c = True
        if found_d and found_c:
            break
    return d_off, c_off


def looks_like_totals(
    row: Sequence[Any],
    debit_idx: int,
    credit_idx: int,
    tokens: Dict[str, Sequence[str]],
) -> bool:
    normed = [_norm(c) for c in row]
    if any(tok in normed for tok in tokens.get("totals", [])):
        return True
    has_both = False
    if 0 <= debit_idx < len(row) and 0 <= credit_idx < len(row):
        has_both = _is_number_like(row[debit_idx]) and _is_number_like(row[credit_idx])
    return has_both


def _empty_df() -> pl.DataFrame:
    """Return an empty DataFrame used to signal fallback."""

    return pl.DataFrame()


def parse_print_friendly_journal(
    content: bytes, language: str = "auto"
) -> pl.DataFrame:
    """Parse a "print-friendly" Excel journal into a tidy DataFrame.

    Parameters
    ----------
    content:
        Raw binary Excel content.
    language:
        Optional hint for language-specific tokens. ``"auto"`` inspects all
        supported languages.
    """

    if not content.startswith(b"PK\x03\x04"):
        logging.info(
            "Skipping print-friendly parser: uploaded file is not an XLSX archive."
        )
        return _empty_df()

    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True, read_only=True)
    except BadZipFile:
        logging.info(
            "Skipping print-friendly parser: uploaded file is not an Excel workbook."
        )
        return _empty_df()
    except ValueError:
        logging.info(
            "Skipping print-friendly parser: workbook XML is invalid or malformed."
        )
        return _empty_df()
    except Exception as e:  # pragma: no cover - corrupt or non-Excel file
        LOGGER.debug("Skipping print-friendly parser due to unexpected error: %s", e)
        return _empty_df()

    ws = wb.active
    rows: List[List[Any]] = [list(r) for r in ws.iter_rows(values_only=True)]
    merged_ranges = _merged_ranges(ws)
    tokens = build_token_map(language)
    header_rows = _find_header_rows(rows, tokens)
    use_pf, _reason, sections = _should_use_print_friendly(rows, tokens, header_rows)

    force_pf = os.getenv("FORCE_PRINT_FRIENDLY") == "1"
    force_raw = os.getenv("FORCE_RAW") == "1"
    if force_pf:
        use_pf = True
    if force_raw:
        use_pf = False

    if not use_pf:
        return _empty_df()

    if _looks_like_ascii_print_journal(rows):
        return _parse_ascii_print_journal(rows)

    records: List[Dict[str, Any]] = []
    previous_date: date | None = None
    previous_context: str | None = None
    for idx, sec in enumerate(sections):
        header_row = rows[sec.header_row]
        mapping = _find_header_cols(header_row, tokens)
        if "debit" not in mapping or "credit" not in mapping:
            continue
        start = sec.header_row + 1
        end = start + sec.length
        section_rows = rows[start:end]
        row_width = max((len(r) for r in section_rows), default=len(header_row))
        label_positions = _header_label_positions(header_row)
        header_d_span = _header_span(label_positions, mapping["debit"], row_width)
        header_c_span = _header_span(label_positions, mapping["credit"], row_width)
        d_span = _merged_span_for_cell(merged_ranges, sec.header_row, mapping["debit"])
        c_span = _merged_span_for_cell(merged_ranges, sec.header_row, mapping["credit"])
        if header_d_span:
            d_span = (
                header_d_span
                if d_span is None
                else (
                    max(d_span[0], header_d_span[0]),
                    min(d_span[1], header_d_span[1]),
                )
            )
        if header_c_span:
            c_span = (
                header_c_span
                if c_span is None
                else (
                    max(c_span[0], header_c_span[0]),
                    min(c_span[1], header_c_span[1]),
                )
            )
        debit_candidates = _candidate_columns(mapping["debit"], d_span, row_width)
        credit_candidates = _candidate_columns(mapping["credit"], c_span, row_width)
        d_col, c_col, _score = _select_amount_columns(
            section_rows, debit_candidates, credit_candidates
        )
        if d_col is None or c_col is None:
            d_col = mapping["debit"] + sec.debit_offset
            c_col = mapping["credit"] + sec.credit_offset
        if d_col >= row_width:
            d_col = mapping["debit"]
        if c_col >= row_width:
            c_col = mapping["credit"]

        detail_rows = [
            row
            for row in section_rows
            if not looks_like_totals(row, d_col, c_col, tokens)
            and (
                (d_col < len(row) and _to_float(row[d_col]) is not None)
                or (c_col < len(row) and _to_float(row[c_col]) is not None)
            )
        ]

        def _resolve_text_candidates(
            field_key: str,
            *,
            left_window: int = 2,
            right_window: int = 8,
        ) -> List[int]:
            base_idx = mapping.get(field_key)
            if base_idx is None:
                return []
            header_span = _header_span(label_positions, base_idx, row_width)
            merged_span = _merged_span_for_cell(merged_ranges, sec.header_row, base_idx)
            if header_span:
                merged_span = (
                    header_span
                    if merged_span is None
                    else (
                        max(merged_span[0], header_span[0]),
                        min(merged_span[1], header_span[1]),
                    )
                )
            return _candidate_text_columns(
                base_idx,
                merged_span,
                row_width,
                left_window=left_window,
                right_window=right_window,
            )

        account_candidates = _resolve_text_candidates(
            "account",
            left_window=2,
            right_window=4,
        )
        account_col = _best_column(
            detail_rows, account_candidates, _score_account_column
        )
        if account_col is None:
            account_col = mapping.get("account")

        account_marker_candidates = [
            candidate for candidate in account_candidates if candidate != account_col
        ]
        account_marker_col = _best_column(
            detail_rows,
            account_marker_candidates,
            _score_account_marker_column,
        )

        account_desc_candidates = _resolve_text_candidates(
            "account_desc",
            left_window=1,
            right_window=0,
        )
        account_desc_col = _best_column(
            detail_rows,
            account_desc_candidates,
            _score_text_column,
        )
        if account_desc_col is None:
            account_desc_col = mapping.get("account_desc")

        line_desc_candidates = _resolve_text_candidates(
            "line_desc",
            left_window=1,
            right_window=8,
        )
        line_desc_col = _best_column(
            detail_rows,
            line_desc_candidates,
            _score_text_column,
        )
        if line_desc_col is None:
            line_desc_col = mapping.get("line_desc")

        rowno_header_indexes = _find_token_indexes(header_row, tokens.get("rowno", []))
        if not rowno_header_indexes and mapping.get("rowno") is not None:
            rowno_header_indexes = [mapping["rowno"]]
        rowno_candidates: List[int] = []
        for rowno_idx in rowno_header_indexes:
            header_span = _header_span(label_positions, rowno_idx, row_width)
            merged_span = _merged_span_for_cell(
                merged_ranges, sec.header_row, rowno_idx
            )
            if header_span:
                merged_span = (
                    header_span
                    if merged_span is None
                    else (
                        max(merged_span[0], header_span[0]),
                        min(merged_span[1], header_span[1]),
                    )
                )
            rowno_candidates.extend(
                _candidate_text_columns(
                    rowno_idx,
                    merged_span,
                    row_width,
                    left_window=8,
                    right_window=2,
                )
            )
        rowno_candidates = sorted(set(rowno_candidates))
        if not rowno_candidates:
            rowno_candidates = [
                candidate for candidate in (0, 1, 2) if candidate < row_width
            ]
        ranked_rowno_cols = _rank_columns(
            detail_rows, rowno_candidates, _score_rowno_column
        )
        rowno_col = ranked_rowno_cols[0] if ranked_rowno_cols else None
        if rowno_col is None:
            rowno_col = mapping.get("rowno")
            ranked_rowno_cols = [rowno_col] if rowno_col is not None else []

        current_date: date | None = previous_date
        current_context: str | None = previous_context
        current_movement: str | None = None
        section_records: List[Dict[str, Any]] = []
        for rel, row in enumerate(section_rows):
            debit = _to_float(row[d_col]) if d_col < len(row) else None
            credit = _to_float(row[c_col]) if c_col < len(row) else None
            has_amount = debit is not None or credit is not None
            date_idx = mapping.get("date_hdr")
            parsed_date: date | None = None
            if date_idx is not None and 0 <= date_idx < len(row):
                parsed_date = _parse_date_with_context(
                    row[date_idx], allow_numeric=True
                )
            if parsed_date is None:
                parsed_date = _extract_leading_context_date(
                    row,
                    allow_generic_scan=not has_amount,
                )
            if parsed_date is not None:
                current_date = parsed_date
            if looks_like_totals(row, d_col, c_col, tokens):
                continue
            if debit is None and credit is None:
                if line_desc_col is not None and line_desc_col < len(row):
                    context_value = _nonempty_text(row[line_desc_col])
                    if context_value:
                        current_context = context_value
                else:
                    context_text = _first_text(row, start=0, stop=min(len(row), 4))
                    if context_text is not None and any(
                        ch.isalpha() for ch in context_text[1]
                    ):
                        current_context = context_text[1].strip()
                continue

            account_raw = (
                row[account_col]
                if account_col is not None and account_col < len(row)
                else None
            )
            account = (
                _merge_account_parts(
                    (
                        row[account_marker_col]
                        if account_marker_col is not None
                        and account_marker_col < len(row)
                        else None
                    ),
                    account_raw,
                )
                or account_raw
            )
            account_desc = (
                row[account_desc_col]
                if account_desc_col is not None and account_desc_col < len(row)
                else None
            )
            line_desc = (
                row[line_desc_col]
                if line_desc_col is not None and line_desc_col < len(row)
                else None
            )
            if line_desc in (None, "") and current_context:
                line_desc = current_context

            movement_str: str | None = None
            candidate: Any = None
            for candidate_idx in ranked_rowno_cols:
                if candidate_idx is None or candidate_idx >= len(row):
                    continue
                candidate = row[candidate_idx]
                movement_str = _normalise_row_identifier(candidate)
                if movement_str is not None:
                    break
            if movement_str is None and not _is_number_like(candidate):
                for c in (0, 1):
                    if c < len(row) and _is_number_like(row[c]):
                        candidate = row[c]
                        movement_str = _normalise_row_identifier(candidate)
                        break
            if movement_str is None and current_movement is not None:
                movement_str = current_movement
            elif movement_str is not None:
                current_movement = movement_str
            amount = (debit or 0.0) - (credit or 0.0)
            section_records.append(
                {
                    "date": current_date,
                    "movement_number": movement_str,
                    "account": account,
                    "account_desc": account_desc,
                    "line_desc": line_desc,
                    "debit_amount": debit,
                    "credit_amount": credit,
                    "amount": amount,
                    "page_index": idx,
                    "source_row": start + rel + 1,
                }
            )

        if any(r["date"] is None for r in section_records):
            derived_dates = _derive_date_column_from_index(
                section_rows,
                date_idx,
                debit_idx=d_col,
                credit_idx=c_col,
            )
            for rec in section_records:
                rel_idx = rec["source_row"] - start - 1
                if 0 <= rel_idx < len(derived_dates) and rec["date"] is None:
                    rec["date"] = derived_dates[rel_idx]

        if current_date is not None:
            previous_date = current_date
        if current_context:
            previous_context = current_context
        records.extend(section_records)

    if not records:
        return _empty_df()

    return _records_to_frame(records)
