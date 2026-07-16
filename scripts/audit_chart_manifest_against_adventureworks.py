from __future__ import annotations

import json
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

__all__ = ["audit_manifest_against_adventureworks", "main"]

REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_MANIFEST = (
    REPO_ROOT / "runs" / "chart_selection_manifest_rebuild" / "selection_manifest.json"
)
DEFAULT_DATASET = REPO_ROOT / "test_data" / "adventureworks.xlsx"
DEFAULT_OUTPUT = (
    REPO_ROOT
    / "runs"
    / "chart_selection_manifest_rebuild"
    / "adventureworks_orchestrator_audit.md"
)


ROLE_COVERAGE: dict[str, tuple[str, str]] = {
    "comparison_metric": ("direct", "Salesamount, Units, Cogs, or Discount"),
    "variance_metric": ("direct", "Salesamount, Units, Cogs, or Discount"),
    "distribution_metric": (
        "direct",
        "Salesamount, Units, Cogs, Discount, or derived rates",
    ),
    "primary_metric": ("direct", "Salesamount, Units, Cogs, Discount"),
    "primary_additive_metric": ("direct", "Salesamount, Units, Cogs, Discount"),
    "related_marker_metric": ("direct", "Discount or derived margin/rate"),
    "x_metric": ("direct", "Any numeric metric, e.g. Salesamount"),
    "y_metric": ("direct", "Any second numeric metric, e.g. Cogs"),
    "size_metric": ("direct", "Any third numeric metric, e.g. Units"),
    "current_period_metric": ("derived", "Salesamount aggregated where Scenario = AC"),
    "baseline_period_metric": ("derived", "Salesamount aggregated where Scenario = PL"),
    "current_metric": ("derived", "Metric aggregated where Scenario = AC"),
    "baseline_metric": ("derived", "Metric aggregated where Scenario = PL"),
    "delta_metric": ("derived", "AC minus PL after aggregation"),
    "percent_delta_metric": ("derived", "Delta divided by PL after aggregation"),
    "value_metric": ("direct", "Salesamount"),
    "volume_metric": ("direct", "Units"),
    "price_or_rate_metric": ("derived", "Salesamount divided by Units"),
    "width_metric": ("direct", "Salesamount or Units"),
    "height_metric": ("direct", "Cogs, Discount, or derived rate"),
    "area_metric": ("derived", "Width metric times height metric"),
    "stage_start_count": ("missing", "No ordered funnel stage columns"),
    "stage_pass_count": ("missing", "No ordered funnel stage columns"),
    "dropoff_count": ("missing", "No ordered funnel stage columns"),
    "conversion_rate": ("missing", "No ordered funnel stage columns"),
    "statement_value": ("missing", "No financial statement line-item table"),
    "focus_share": ("missing", "No attribute-bundle evidence table"),
    "baseline_share": ("missing", "No attribute-bundle evidence table"),
    "index_metric": ("missing", "No attribute-bundle evidence table"),
    "current_signal_metric": ("missing", "No current/emerging signal table"),
    "emerging_signal_metric": ("missing", "No current/emerging signal table"),
    "alignment_metric": ("missing", "No current/emerging signal table"),
    "gross_weight": ("missing", "No rank-weighted visibility table"),
    "incremental_weight": ("missing", "No rank-weighted visibility table"),
    "cumulative_weight": ("missing", "No rank-weighted visibility table"),
    "robustness_metric": ("missing", "No rank-weighted visibility table"),
    "product_signal_score": ("missing", "No product-signal evidence table"),
    "validation_metric": ("missing", "No product-signal evidence table"),
    "category": (
        "direct",
        "Category, Subcategory, Productline, Color, Region, Country",
    ),
    "component_category": ("direct", "Color, Subcategory, Productline, Country"),
    "component_dimension": ("direct", "Color, Subcategory, Productline, Country"),
    "nested_category": ("direct", "Category -> Subcategory -> Productname"),
    "width_category": ("direct", "Category, Subcategory, Country, Region"),
    "stack_category": ("direct", "Color, Subcategory, Productline, Country"),
    "height_category": ("direct", "Category, Subcategory, Country, Region"),
    "point_dimension": ("direct", "Customer, Productname, Category, Region"),
    "comparison_series": ("direct", "Scenario"),
    "comparison_item": ("direct", "Category, Subcategory, Region, Country"),
    "comparison_window": ("derived", "Date windows from Orderdate"),
    "bridge_component_period": (
        "derived",
        "Monthly or scenario deltas from Orderdate/Scenario",
    ),
    "stable_population_flag": ("derived", "Customers present in both scenarios"),
    "first_active_cohort": ("derived", "First Orderdate by Customer"),
    "lost_or_last_active_cohort": ("derived", "Last Orderdate by Customer"),
    "variance_driver": ("direct", "Category, Subcategory, Region, Country, Color"),
    "dimension_member": (
        "direct",
        "Members of Category, Subcategory, Region, Country, Color",
    ),
    "parent_driver": ("direct", "Category or Country"),
    "child_driver": ("direct", "Subcategory, Productname, Region, or Color"),
    "root_cause_driver": ("derived", "Candidate drivers ranked by scenario delta"),
    "component": ("direct", "Category, Region, or Productline"),
    "component_driver": ("direct", "Subcategory, Productname, Country, or Color"),
    "scenario_or_period_pair": ("direct", "Scenario AC/PL or Orderdate windows"),
    "set_membership_fields": (
        "derived",
        "Customer-to-category or product-to-region memberships",
    ),
    "panel_or_segment": ("direct", "Country, Region, Category, or Group"),
    "two_or_three_set_membership_fields": (
        "derived",
        "Customer-to-category or product-to-region memberships",
    ),
    "ordered_stage": ("missing", "No funnel stage column"),
    "statement_line_item": ("missing", "No statement line-item column"),
    "attribute_bundle": ("missing", "No attribute bundle column"),
    "signal_bundle": ("missing", "No signal bundle column"),
    "cohort_layer": ("missing", "No current/emerging cohort layer"),
    "rank_or_lane": ("missing", "No rank/lane visibility structure"),
    "product": ("direct", "Productname"),
}


def _dataset_profile(dataset_path: Path) -> dict[str, Any]:
    workbook = load_workbook(dataset_path, read_only=True, data_only=True)
    worksheet = workbook.active
    headers = list(next(worksheet.iter_rows(max_row=1, values_only=True)))
    header_index = {header: index for index, header in enumerate(headers)}
    counters: dict[str, Counter[Any]] = {
        column: Counter()
        for column in (
            "Scenario",
            "Category",
            "Subcategory",
            "Color",
            "Region",
            "Country",
            "Group",
        )
        if column in header_index
    }
    rows = 0
    min_date = None
    max_date = None
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        rows += 1
        date_value = row[header_index["Orderdate"]]
        if date_value is not None:
            min_date = (
                date_value if min_date is None or date_value < min_date else min_date
            )
            max_date = (
                date_value if max_date is None or date_value > max_date else max_date
            )
        for column, counter in counters.items():
            value = row[header_index[column]]
            if value is not None:
                counter[value] += 1
    return {
        "path": str(dataset_path),
        "rows": rows,
        "columns": headers,
        "date_range": [str(min_date), str(max_date)],
        "sample_distinct_values": {
            column: [value for value, _count in counter.most_common(8)]
            for column, counter in counters.items()
        },
    }


def _role_coverage(role: str) -> tuple[str, str]:
    return ROLE_COVERAGE.get(role, ("unknown", "No mapping in this dataset audit."))


def _capability_result(capability: dict[str, Any]) -> dict[str, Any]:
    contract = capability["selection_contract"]
    requirements = contract["dataset_requirements"]
    metric_requirements = requirements["metrics"]
    source_metric_roles = [
        role["role"]
        for role in metric_requirements["source_metric_roles"]
        if role.get("required", True)
    ]
    derived_metric_roles = [
        role["role"] for role in metric_requirements["derived_metric_roles"]
    ]
    roles = source_metric_roles + list(requirements["dimensions"]["required_roles"])
    covered = [_role_coverage(role) for role in roles]
    states = {state for state, _detail in covered}
    if "unknown" in states:
        status = "manifest_role_unmapped_in_audit"
    elif "missing" in states:
        status = "understand_reject_for_this_dataset"
    elif "derived" in states:
        status = "understand_use_after_aggregation_or_derivation"
    else:
        status = "understand_use_directly_after_role_mapping"
    return {
        "capability_id": capability["capability_id"],
        "selection_emphasis": capability["selection_emphasis"],
        "status": status,
        "required_roles": roles,
        "source_metric_roles": source_metric_roles,
        "derived_metric_roles": derived_metric_roles,
        "role_coverage": [
            {"role": role, "status": state, "evidence": detail}
            for role, (state, detail) in zip(roles, covered)
        ],
        "period": requirements["period"],
        "accept_when": contract["accept_when"],
        "reject_when": contract["reject_when"],
    }


def audit_manifest_against_adventureworks(
    *,
    manifest_path: Path = DEFAULT_MANIFEST,
    dataset_path: Path = DEFAULT_DATASET,
) -> dict[str, Any]:
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    results = [
        _capability_result(capability)
        for capability in manifest["capabilities"].values()
    ]
    results_by_capability = {result["capability_id"]: result for result in results}
    artifact_results = []
    for artifact in manifest["artifacts"]:
        capability_id = artifact["capability_id"]
        capability_result = results_by_capability[capability_id]
        artifact_results.append(
            {
                "label": artifact["label"],
                "capability_id": capability_id,
                "status": capability_result["status"],
                "selection_emphasis": capability_result["selection_emphasis"],
                "output": artifact["output"],
                "role_coverage": capability_result["role_coverage"],
                "period": capability_result["period"],
            }
        )
    return {
        "dataset": _dataset_profile(dataset_path),
        "summary": dict(Counter(result["status"] for result in results)),
        "artifact_summary": dict(
            Counter(result["status"] for result in artifact_results)
        ),
        "results": sorted(results, key=lambda item: item["capability_id"]),
        "artifact_results": sorted(
            artifact_results,
            key=lambda item: (item["status"], item["label"]),
        ),
    }


def _markdown(audit: dict[str, Any]) -> str:
    dataset = audit["dataset"]
    lines = [
        "# AdventureWorks Orchestrator Manifest Audit",
        "",
        "## Dataset",
        "",
        f"- Path: `{dataset['path']}`",
        f"- Rows: `{dataset['rows']}`",
        f"- Date range: `{dataset['date_range'][0]}` to `{dataset['date_range'][1]}`",
        f"- Columns: `{len(dataset['columns'])}`",
        "",
        "## Summary",
        "",
    ]
    for status, count in sorted(audit["summary"].items()):
        lines.append(f"- `{status}`: `{count}`")
    lines.extend(["", "## Artifact Summary", ""])
    for status, count in sorted(audit["artifact_summary"].items()):
        lines.append(f"- `{status}`: `{count}`")
    lines.extend(["", "## Capability Results", ""])
    for result in audit["results"]:
        lines.append(
            f"- `{result['capability_id']}` -> `{result['status']}` "
            f"({result['selection_emphasis']})"
        )
        lines.append(f"  - Period: `{result['period']['role']}`")
        if result["required_roles"]:
            role_text = ", ".join(
                f"`{entry['role']}`={entry['status']}"
                for entry in result["role_coverage"]
            )
            lines.append(f"  - Required roles: {role_text}")
        else:
            lines.append("  - Required roles: none")
        lines.append(f"  - Use: {result['accept_when']}")
        lines.append(f"  - Reject: {result['reject_when']}")
    lines.extend(["", "## Artifact Results", ""])
    for result in audit["artifact_results"]:
        lines.append(
            f"- `{result['label']}` -> `{result['status']}` "
            f"via `{result['capability_id']}` ({result['selection_emphasis']})"
        )
        lines.append(f"  - Output: `{result['output']}`")
        lines.append(f"  - Period: `{result['period']['role']}`")
        if result["role_coverage"]:
            role_text = ", ".join(
                f"`{entry['role']}`={entry['status']}"
                for entry in result["role_coverage"]
            )
            lines.append(f"  - Required roles: {role_text}")
        else:
            lines.append("  - Required roles: none")
    return "\n".join(lines) + "\n"


def main() -> int:
    audit = audit_manifest_against_adventureworks()
    DEFAULT_OUTPUT.write_text(_markdown(audit), encoding="utf-8")
    print(DEFAULT_OUTPUT)
    print(json.dumps(audit["summary"], sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
