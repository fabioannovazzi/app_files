"""Deterministic accounting core for the shared Business Planning workflow.

Fixed code is limited to mechanically verifiable contract validation, exact
Decimal arithmetic, statement roll-forwards, reconciliation, rendering, and
artifact preparation. Business meaning, evidence relevance, assumptions,
scenario design, interpretation, and professional approval remain model-led or
professional judgments.
"""

from __future__ import annotations

import csv
import html
import json
import re
from collections import Counter
from collections.abc import Mapping, Sequence
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from business_planning_handoff import HANDOFF_SCHEMA, counterpart_handoff_status
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from vera_assurance import (
    MoneyValidationError,
    decimal_text,
    parse_canonical_decimal,
)

__all__ = [
    "CASE_SCHEMA",
    "COMMENTARY_SCHEMA",
    "PLAN_SCHEMA",
    "BusinessPlanningContractError",
    "build_counterpart_handoff",
    "build_business_plan",
    "build_model_context",
    "load_json",
    "render_html",
    "render_markdown",
    "validate_case",
    "write_assumption_ledger",
    "write_excel",
]

WORKFLOW_ID = "business-planning"
CASE_SCHEMA = "mparanza.business_planning_financial_case.v1"
PLAN_SCHEMA = "mparanza.business_planning_financial_plan.v2"
COMMENTARY_SCHEMA = "mparanza.business_planning_financial_commentary.v1"
MAX_PERIODS = 60
MAX_SCENARIOS = 8
_IDENTIFIER_RE = re.compile(r"^[a-z][a-z0-9_.-]{0,119}$")
_CURRENCY_RE = re.compile(r"^[A-Z]{3}$")
_EVIDENCE_KINDS = {
    "historical_actual",
    "opening_fact",
    "management_assumption",
    "external_evidence",
    "model_hypothesis",
}
_EVIDENCE_STATUSES = {"reviewed", "confirmed", "unverified"}
_OPENING_VALUE_FIELDS = (
    "cash",
    "accounts_receivable",
    "inventory",
    "other_current_assets",
    "net_fixed_assets",
    "other_non_current_assets",
    "accounts_payable",
    "debt",
    "other_liabilities",
    "equity",
)
_SCHEDULE_MONEY_FIELDS = (
    "revenue",
    "cogs",
    "operating_expenses",
    "depreciation_amortization",
    "interest_expense",
    "tax_expense",
    "capital_expenditure",
    "ending_accounts_receivable",
    "ending_inventory",
    "ending_other_current_assets",
    "ending_accounts_payable",
    "ending_other_liabilities",
    "debt_draws",
    "debt_repayments",
    "equity_contributions",
    "dividends",
)
_TOP_LEVEL_FIELDS = {
    "schema_version",
    "case_id",
    "entity_name",
    "company_stage",
    "planning_objective",
    "professional_lens",
    "audience",
    "reporting_currency",
    "periods",
    "reconciliation_tolerance",
    "review",
    "evidence_register",
    "opening_balance",
    "assumptions",
    "scenarios",
}


class BusinessPlanningContractError(ValueError):
    """Raised when a reviewed planning case is mechanically inconsistent."""


def load_json(path: Path) -> dict[str, Any]:
    """Load one JSON object from a regular file."""

    if path.is_symlink() or not path.is_file():
        raise BusinessPlanningContractError(
            f"JSON input must be a regular file: {path}"
        )
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise BusinessPlanningContractError(
            f"Invalid JSON in {path.name}: {exc}"
        ) from exc
    if not isinstance(payload, dict):
        raise BusinessPlanningContractError(f"JSON must contain an object: {path.name}")
    return payload


def _mapping(value: object, *, label: str) -> Mapping[str, Any]:
    if not isinstance(value, Mapping):
        raise BusinessPlanningContractError(f"{label} must be an object")
    return value


def _list(value: object, *, label: str) -> list[Any]:
    if not isinstance(value, Sequence) or isinstance(value, (str, bytes, bytearray)):
        raise BusinessPlanningContractError(f"{label} must be a list")
    return list(value)


def _exact_keys(value: Mapping[str, Any], *, expected: set[str], label: str) -> None:
    observed = set(value)
    if observed != expected:
        missing = sorted(expected - observed)
        extra = sorted(observed - expected)
        raise BusinessPlanningContractError(
            f"{label} fields do not match the contract; missing={missing}, extra={extra}"
        )


def _text(value: object, *, label: str, maximum: int = 400) -> str:
    if not isinstance(value, str) or value != value.strip() or not value:
        raise BusinessPlanningContractError(f"{label} must be non-empty trimmed text")
    if len(value) > maximum:
        raise BusinessPlanningContractError(
            f"{label} must contain at most {maximum} characters"
        )
    return value


def _identifier(value: object, *, label: str) -> str:
    text = _text(value, label=label, maximum=120)
    if _IDENTIFIER_RE.fullmatch(text) is None:
        raise BusinessPlanningContractError(f"{label} must be a lowercase identifier")
    return text


def _money(value: object, *, label: str, non_negative: bool = True) -> Decimal:
    try:
        amount = parse_canonical_decimal(value, label=label)
    except MoneyValidationError as exc:
        raise BusinessPlanningContractError(str(exc)) from exc
    if non_negative and amount < 0:
        raise BusinessPlanningContractError(f"{label} must be non-negative")
    return amount


def _unique_ids(values: object, *, label: str, allow_empty: bool = False) -> list[str]:
    items = [
        _identifier(item, label=f"{label} item") for item in _list(values, label=label)
    ]
    if not items and not allow_empty:
        raise BusinessPlanningContractError(f"{label} must not be empty")
    if len(items) != len(set(items)):
        raise BusinessPlanningContractError(f"{label} must contain unique identifiers")
    return items


def _review_timestamp(value: object, *, label: str) -> str:
    timestamp = _text(value, label=label, maximum=80)
    try:
        parsed = datetime.fromisoformat(timestamp.replace("Z", "+00:00"))
    except ValueError as exc:
        raise BusinessPlanningContractError(
            f"{label} must be an ISO 8601 timestamp"
        ) from exc
    if parsed.tzinfo is None or parsed.utcoffset() is None:
        raise BusinessPlanningContractError(f"{label} must include a timezone")
    return timestamp


def validate_case(case: Mapping[str, Any]) -> None:
    """Validate one reviewed case without making semantic planning judgments."""

    _exact_keys(case, expected=_TOP_LEVEL_FIELDS, label="case")
    if case["schema_version"] != CASE_SCHEMA:
        raise BusinessPlanningContractError(
            f"case.schema_version must be {CASE_SCHEMA}"
        )
    _identifier(case["case_id"], label="case.case_id")
    _text(case["entity_name"], label="case.entity_name", maximum=200)
    _text(case["company_stage"], label="case.company_stage", maximum=300)
    _text(case["planning_objective"], label="case.planning_objective", maximum=500)
    if case["professional_lens"] != "accounting_financial":
        raise BusinessPlanningContractError(
            "case.professional_lens must be accounting_financial for this runner"
        )
    _text(case["audience"], label="case.audience", maximum=200)
    currency = _text(
        case["reporting_currency"], label="case.reporting_currency", maximum=3
    )
    if _CURRENCY_RE.fullmatch(currency) is None:
        raise BusinessPlanningContractError(
            "case.reporting_currency must be a three-letter uppercase code"
        )
    tolerance = _money(
        case["reconciliation_tolerance"],
        label="case.reconciliation_tolerance",
    )
    if tolerance > Decimal("1000"):
        raise BusinessPlanningContractError(
            "case.reconciliation_tolerance exceeds the supported maximum"
        )

    periods = [
        _text(item, label="case.periods item", maximum=40)
        for item in _list(case["periods"], label="case.periods")
    ]
    if not 1 <= len(periods) <= MAX_PERIODS:
        raise BusinessPlanningContractError(
            f"case.periods must contain between 1 and {MAX_PERIODS} labels"
        )
    if len(periods) != len(set(periods)):
        raise BusinessPlanningContractError("case.periods must be unique")

    review = _mapping(case["review"], label="case.review")
    _exact_keys(
        review,
        expected={"status", "reviewer", "reviewed_at"},
        label="case.review",
    )
    if review["status"] != "reviewed":
        raise BusinessPlanningContractError(
            "case.review.status must be reviewed before calculation"
        )
    _text(review["reviewer"], label="case.review.reviewer", maximum=200)
    _review_timestamp(review["reviewed_at"], label="case.review.reviewed_at")

    evidence_items = _list(case["evidence_register"], label="case.evidence_register")
    if not evidence_items:
        raise BusinessPlanningContractError("case.evidence_register must not be empty")
    evidence_ids: set[str] = set()
    evidence_statuses: dict[str, str] = {}
    for index, raw in enumerate(evidence_items):
        item = _mapping(raw, label=f"case.evidence_register[{index}]")
        _exact_keys(
            item,
            expected={"id", "kind", "description", "source_ref", "status"},
            label=f"case.evidence_register[{index}]",
        )
        evidence_id = _identifier(item["id"], label=f"evidence[{index}].id")
        if evidence_id in evidence_ids:
            raise BusinessPlanningContractError(f"Duplicate evidence ID: {evidence_id}")
        evidence_ids.add(evidence_id)
        kind = _text(item["kind"], label=f"evidence[{index}].kind", maximum=80)
        if kind not in _EVIDENCE_KINDS:
            raise BusinessPlanningContractError(
                f"evidence[{index}].kind is not supported: {kind}"
            )
        _text(
            item["description"],
            label=f"evidence[{index}].description",
            maximum=500,
        )
        _text(item["source_ref"], label=f"evidence[{index}].source_ref", maximum=300)
        status = _text(item["status"], label=f"evidence[{index}].status", maximum=40)
        if status not in _EVIDENCE_STATUSES:
            raise BusinessPlanningContractError(
                f"evidence[{index}].status is not supported: {status}"
            )
        evidence_statuses[evidence_id] = status

    opening = _mapping(case["opening_balance"], label="case.opening_balance")
    _exact_keys(
        opening,
        expected={"values", "evidence_ids"},
        label="case.opening_balance",
    )
    opening_values = _mapping(opening["values"], label="case.opening_balance.values")
    _exact_keys(
        opening_values,
        expected=set(_OPENING_VALUE_FIELDS),
        label="case.opening_balance.values",
    )
    for field in _OPENING_VALUE_FIELDS:
        _money(
            opening_values[field],
            label=f"case.opening_balance.values.{field}",
            non_negative=field != "equity",
        )
    opening_evidence_ids = _unique_ids(
        opening["evidence_ids"], label="case.opening_balance.evidence_ids"
    )
    unknown_opening_evidence = sorted(set(opening_evidence_ids) - evidence_ids)
    if unknown_opening_evidence:
        raise BusinessPlanningContractError(
            f"Opening balance references unknown evidence: {unknown_opening_evidence}"
        )

    assumption_items = _list(case["assumptions"], label="case.assumptions")
    if not assumption_items:
        raise BusinessPlanningContractError("case.assumptions must not be empty")
    assumption_ids: set[str] = set()
    assumption_periods: dict[str, set[str]] = {}
    for index, raw in enumerate(assumption_items):
        item = _mapping(raw, label=f"case.assumptions[{index}]")
        _exact_keys(
            item,
            expected={
                "id",
                "category",
                "description",
                "evidence_ids",
                "effective_periods",
                "rationale",
                "status",
            },
            label=f"case.assumptions[{index}]",
        )
        assumption_id = _identifier(item["id"], label=f"assumption[{index}].id")
        if assumption_id in assumption_ids:
            raise BusinessPlanningContractError(
                f"Duplicate assumption ID: {assumption_id}"
            )
        assumption_ids.add(assumption_id)
        _text(item["category"], label=f"assumption[{index}].category", maximum=120)
        _text(
            item["description"],
            label=f"assumption[{index}].description",
            maximum=500,
        )
        linked_evidence = _unique_ids(
            item["evidence_ids"], label=f"assumption[{index}].evidence_ids"
        )
        unknown_evidence = sorted(set(linked_evidence) - evidence_ids)
        if unknown_evidence:
            raise BusinessPlanningContractError(
                f"Assumption {assumption_id} references unknown evidence: {unknown_evidence}"
            )
        effective_periods = [
            _text(
                value, label=f"assumption[{index}].effective_periods item", maximum=40
            )
            for value in _list(
                item["effective_periods"],
                label=f"assumption[{index}].effective_periods",
            )
        ]
        if not effective_periods:
            raise BusinessPlanningContractError(
                f"Assumption {assumption_id} has no effective period"
            )
        if len(effective_periods) != len(set(effective_periods)):
            raise BusinessPlanningContractError(
                f"Assumption {assumption_id} has duplicate effective periods"
            )
        unknown_periods = sorted(set(effective_periods) - set(periods))
        if unknown_periods:
            raise BusinessPlanningContractError(
                f"Assumption {assumption_id} references unknown periods: {unknown_periods}"
            )
        assumption_periods[assumption_id] = set(effective_periods)
        _text(item["rationale"], label=f"assumption[{index}].rationale", maximum=800)
        if item["status"] != "confirmed":
            raise BusinessPlanningContractError(
                f"Assumption {assumption_id} must have status=confirmed"
            )

    scenario_items = _list(case["scenarios"], label="case.scenarios")
    if not 1 <= len(scenario_items) <= MAX_SCENARIOS:
        raise BusinessPlanningContractError(
            f"case.scenarios must contain between 1 and {MAX_SCENARIOS} scenarios"
        )
    scenario_ids: set[str] = set()
    referenced_assumptions: set[str] = set()
    schedule_fields = {"period", "assumption_ids", *_SCHEDULE_MONEY_FIELDS}
    for scenario_index, raw in enumerate(scenario_items):
        scenario = _mapping(raw, label=f"case.scenarios[{scenario_index}]")
        _exact_keys(
            scenario,
            expected={"id", "label", "schedule"},
            label=f"case.scenarios[{scenario_index}]",
        )
        scenario_id = _identifier(
            scenario["id"], label=f"case.scenarios[{scenario_index}].id"
        )
        if scenario_id in scenario_ids:
            raise BusinessPlanningContractError(f"Duplicate scenario ID: {scenario_id}")
        scenario_ids.add(scenario_id)
        _text(
            scenario["label"],
            label=f"case.scenarios[{scenario_index}].label",
            maximum=160,
        )
        schedule = _list(
            scenario["schedule"], label=f"case.scenarios[{scenario_index}].schedule"
        )
        observed_periods: list[str] = []
        if len(schedule) != len(periods):
            raise BusinessPlanningContractError(
                f"Scenario {scenario_id} must contain exactly {len(periods)} rows"
            )
        for row_index, raw_row in enumerate(schedule):
            row = _mapping(
                raw_row,
                label=f"case.scenarios[{scenario_index}].schedule[{row_index}]",
            )
            _exact_keys(
                row,
                expected=schedule_fields,
                label=f"scenario {scenario_id} row {row_index}",
            )
            period = _text(
                row["period"],
                label=f"scenario {scenario_id} row {row_index}.period",
                maximum=40,
            )
            observed_periods.append(period)
            linked_assumptions = _unique_ids(
                row["assumption_ids"],
                label=f"scenario {scenario_id} row {row_index}.assumption_ids",
            )
            referenced_assumptions.update(linked_assumptions)
            unknown_assumptions = sorted(set(linked_assumptions) - assumption_ids)
            if unknown_assumptions:
                raise BusinessPlanningContractError(
                    f"Scenario {scenario_id} references unknown assumptions: {unknown_assumptions}"
                )
            inactive = sorted(
                assumption_id
                for assumption_id in linked_assumptions
                if period not in assumption_periods[assumption_id]
            )
            if inactive:
                raise BusinessPlanningContractError(
                    f"Scenario {scenario_id} period {period} references inactive assumptions: {inactive}"
                )
            for field in _SCHEDULE_MONEY_FIELDS:
                _money(
                    row[field],
                    label=f"scenario {scenario_id} period {period}.{field}",
                )
        if observed_periods != periods:
            raise BusinessPlanningContractError(
                f"Scenario {scenario_id} period order does not match case.periods"
            )

    unused_assumptions = sorted(assumption_ids - referenced_assumptions)
    if unused_assumptions:
        raise BusinessPlanningContractError(
            f"Confirmed assumptions are not applied by any scenario: {unused_assumptions}"
        )

    referenced_evidence = {
        evidence_id
        for raw in assumption_items
        for evidence_id in _list(
            _mapping(raw, label="assumption")["evidence_ids"],
            label="assumption.evidence_ids",
        )
    } | set(opening_evidence_ids)
    if not referenced_evidence:
        raise BusinessPlanningContractError("The case references no evidence")


def _decimal_mapping(
    values: Mapping[str, Any],
    fields: Sequence[str],
    *,
    prefix: str,
    allow_negative_fields: frozenset[str] = frozenset(),
) -> dict[str, Decimal]:
    return {
        field: _money(
            values[field],
            label=f"{prefix}.{field}",
            non_negative=field not in allow_negative_fields,
        )
        for field in fields
    }


def _money_strings(values: Mapping[str, Decimal]) -> dict[str, str]:
    return {key: decimal_text(value) for key, value in values.items()}


def _opening_reconciliation(opening: Mapping[str, Decimal]) -> dict[str, Decimal]:
    total_assets = (
        opening["cash"]
        + opening["accounts_receivable"]
        + opening["inventory"]
        + opening["other_current_assets"]
        + opening["net_fixed_assets"]
        + opening["other_non_current_assets"]
    )
    total_liabilities_equity = (
        opening["accounts_payable"]
        + opening["debt"]
        + opening["other_liabilities"]
        + opening["equity"]
    )
    return {
        "total_assets": total_assets,
        "total_liabilities_and_equity": total_liabilities_equity,
        "difference": total_assets - total_liabilities_equity,
    }


def build_business_plan(case: Mapping[str, Any]) -> dict[str, Any]:
    """Build linked statements from one reviewed planning case."""

    validate_case(case)
    tolerance = _money(
        case["reconciliation_tolerance"], label="case.reconciliation_tolerance"
    )
    opening_record = _mapping(case["opening_balance"], label="case.opening_balance")
    opening_values = _decimal_mapping(
        _mapping(opening_record["values"], label="opening values"),
        _OPENING_VALUE_FIELDS,
        prefix="opening_balance",
        allow_negative_fields=frozenset({"equity"}),
    )
    opening_reconciliation = _opening_reconciliation(opening_values)
    opening_passed = abs(opening_reconciliation["difference"]) <= tolerance

    evidence_items = [
        dict(_mapping(item, label="evidence"))
        for item in _list(case["evidence_register"], label="case.evidence_register")
    ]
    evidence_status = {str(item["id"]): str(item["status"]) for item in evidence_items}
    referenced_evidence = set(
        _list(opening_record["evidence_ids"], label="opening evidence IDs")
    )
    for raw_assumption in _list(case["assumptions"], label="case.assumptions"):
        assumption = _mapping(raw_assumption, label="assumption")
        referenced_evidence.update(
            _list(assumption["evidence_ids"], label="assumption evidence IDs")
        )
    unverified_evidence = sorted(
        evidence_id
        for evidence_id in referenced_evidence
        if evidence_status[evidence_id] == "unverified"
    )

    scenarios: list[dict[str, Any]] = []
    reconciliation_rows: list[dict[str, Any]] = []
    any_failed = not opening_passed
    for raw_scenario in _list(case["scenarios"], label="case.scenarios"):
        scenario = _mapping(raw_scenario, label="scenario")
        previous = dict(opening_values)
        scenario_periods: list[dict[str, Any]] = []
        ending_cash_values: list[Decimal] = []
        total_revenue = Decimal("0")
        total_net_income = Decimal("0")
        total_debt_draws = Decimal("0")
        total_equity_contributions = Decimal("0")
        break_even_period: str | None = None
        for raw_row in _list(scenario["schedule"], label="scenario.schedule"):
            row = _mapping(raw_row, label="schedule row")
            inputs = _decimal_mapping(
                row,
                _SCHEDULE_MONEY_FIELDS,
                prefix=f"scenario {scenario['id']} period {row['period']}",
            )
            revenue = inputs["revenue"]
            cogs = inputs["cogs"]
            operating_expenses = inputs["operating_expenses"]
            depreciation = inputs["depreciation_amortization"]
            interest = inputs["interest_expense"]
            tax = inputs["tax_expense"]
            gross_profit = revenue - cogs
            ebitda = gross_profit - operating_expenses
            ebit = ebitda - depreciation
            profit_before_tax = ebit - interest
            net_income = profit_before_tax - tax

            change_ar = (
                inputs["ending_accounts_receivable"] - previous["accounts_receivable"]
            )
            change_inventory = inputs["ending_inventory"] - previous["inventory"]
            change_other_current_assets = (
                inputs["ending_other_current_assets"] - previous["other_current_assets"]
            )
            change_ap = inputs["ending_accounts_payable"] - previous["accounts_payable"]
            change_other_liabilities = (
                inputs["ending_other_liabilities"] - previous["other_liabilities"]
            )
            working_capital_investment = (
                change_ar
                + change_inventory
                + change_other_current_assets
                - change_ap
                - change_other_liabilities
            )
            operating_cash_flow = net_income + depreciation - working_capital_investment
            investing_cash_flow = -inputs["capital_expenditure"]
            financing_cash_flow = (
                inputs["debt_draws"]
                - inputs["debt_repayments"]
                + inputs["equity_contributions"]
                - inputs["dividends"]
            )
            net_cash_change = (
                operating_cash_flow + investing_cash_flow + financing_cash_flow
            )
            ending_cash = previous["cash"] + net_cash_change
            ending_fixed_assets = (
                previous["net_fixed_assets"]
                + inputs["capital_expenditure"]
                - depreciation
            )
            ending_debt = (
                previous["debt"] + inputs["debt_draws"] - inputs["debt_repayments"]
            )
            ending_equity = (
                previous["equity"]
                + net_income
                + inputs["equity_contributions"]
                - inputs["dividends"]
            )
            if ending_fixed_assets < 0:
                raise BusinessPlanningContractError(
                    f"Scenario {scenario['id']} period {row['period']} makes net fixed assets negative"
                )
            if ending_debt < 0:
                raise BusinessPlanningContractError(
                    f"Scenario {scenario['id']} period {row['period']} makes debt negative"
                )

            balance_sheet = {
                "cash": ending_cash,
                "accounts_receivable": inputs["ending_accounts_receivable"],
                "inventory": inputs["ending_inventory"],
                "other_current_assets": inputs["ending_other_current_assets"],
                "net_fixed_assets": ending_fixed_assets,
                "other_non_current_assets": previous["other_non_current_assets"],
                "accounts_payable": inputs["ending_accounts_payable"],
                "debt": ending_debt,
                "other_liabilities": inputs["ending_other_liabilities"],
                "equity": ending_equity,
            }
            reconciliation = _opening_reconciliation(balance_sheet)
            passed = abs(reconciliation["difference"]) <= tolerance
            any_failed = any_failed or not passed
            reconciliation_rows.append(
                {
                    "scenario_id": scenario["id"],
                    "period": row["period"],
                    **_money_strings(reconciliation),
                    "passed": passed,
                }
            )
            profit_and_loss = {
                "revenue": revenue,
                "cogs": cogs,
                "gross_profit": gross_profit,
                "operating_expenses": operating_expenses,
                "ebitda": ebitda,
                "depreciation_amortization": depreciation,
                "ebit": ebit,
                "interest_expense": interest,
                "profit_before_tax": profit_before_tax,
                "tax_expense": tax,
                "net_income": net_income,
            }
            cash_flow = {
                "net_income": net_income,
                "depreciation_amortization": depreciation,
                "working_capital_investment": working_capital_investment,
                "operating_cash_flow": operating_cash_flow,
                "capital_expenditure": -inputs["capital_expenditure"],
                "investing_cash_flow": investing_cash_flow,
                "debt_draws": inputs["debt_draws"],
                "debt_repayments": -inputs["debt_repayments"],
                "equity_contributions": inputs["equity_contributions"],
                "dividends": -inputs["dividends"],
                "financing_cash_flow": financing_cash_flow,
                "net_cash_change": net_cash_change,
                "ending_cash": ending_cash,
            }
            scenario_periods.append(
                {
                    "period": row["period"],
                    "assumption_ids": list(row["assumption_ids"]),
                    "profit_and_loss": _money_strings(profit_and_loss),
                    "cash_flow": _money_strings(cash_flow),
                    "balance_sheet": _money_strings(balance_sheet),
                    "reconciliation": {
                        **_money_strings(reconciliation),
                        "passed": passed,
                    },
                }
            )
            ending_cash_values.append(ending_cash)
            total_revenue += revenue
            total_net_income += net_income
            total_debt_draws += inputs["debt_draws"]
            total_equity_contributions += inputs["equity_contributions"]
            if break_even_period is None and ebitda >= 0:
                break_even_period = str(row["period"])
            previous = balance_sheet

        minimum_cash = min(ending_cash_values)
        funding_requirement = max(Decimal("0"), -minimum_cash)
        scenarios.append(
            {
                "id": scenario["id"],
                "label": scenario["label"],
                "periods": scenario_periods,
                "summary": {
                    "total_revenue": decimal_text(total_revenue),
                    "total_net_income": decimal_text(total_net_income),
                    "ending_cash": decimal_text(ending_cash_values[-1]),
                    "minimum_cash": decimal_text(minimum_cash),
                    "funding_requirement": decimal_text(funding_requirement),
                    "total_debt_draws": decimal_text(total_debt_draws),
                    "total_equity_contributions": decimal_text(
                        total_equity_contributions
                    ),
                    "break_even_period": break_even_period,
                },
            }
        )

    status = (
        "blocked"
        if any_failed
        else "partial" if unverified_evidence else "ready_for_professional_review"
    )
    evidence_counts = Counter(str(item["status"]) for item in evidence_items)
    return {
        "schema_version": PLAN_SCHEMA,
        "workflow_id": WORKFLOW_ID,
        "case_id": case["case_id"],
        "entity_name": case["entity_name"],
        "company_stage": case["company_stage"],
        "planning_objective": case["planning_objective"],
        "professional_lens": case["professional_lens"],
        "audience": case["audience"],
        "reporting_currency": case["reporting_currency"],
        "periods": list(case["periods"]),
        "status": status,
        "review_status": "draft_pending_professional_review",
        "reconciliation_tolerance": case["reconciliation_tolerance"],
        "evidence_coverage": {
            "referenced_evidence_ids": sorted(referenced_evidence),
            "status_counts": dict(sorted(evidence_counts.items())),
            "unverified_evidence_ids": unverified_evidence,
        },
        "evidence_register": [
            {
                "id": item["id"],
                "kind": item["kind"],
                "description": item["description"],
                "status": item["status"],
            }
            for item in evidence_items
            if item["id"] in referenced_evidence
        ],
        "opening_balance": {
            "values": _money_strings(opening_values),
            "evidence_ids": list(opening_record["evidence_ids"]),
            "reconciliation": {
                **_money_strings(opening_reconciliation),
                "passed": opening_passed,
            },
        },
        "assumptions": [
            dict(_mapping(item, label="assumption")) for item in case["assumptions"]
        ],
        "scenarios": scenarios,
        "reconciliation": {
            "opening": {
                **_money_strings(opening_reconciliation),
                "passed": opening_passed,
            },
            "periods": reconciliation_rows,
            "all_passed": not any_failed,
        },
        "limitations": [
            "Interest and tax expense are treated as paid in the same period.",
            "Deferred tax, tax payable, non-cash interest, leases, and asset disposals are outside the v1 contract.",
            "Break-even period means the first period with non-negative EBITDA; it is not a unit or revenue threshold calculation.",
            *(
                ["One or more referenced evidence items remain unverified."]
                if unverified_evidence
                else []
            ),
        ],
    }


def build_model_context(plan: Mapping[str, Any]) -> dict[str, Any]:
    """Project calculated facts without source references or raw populations."""

    assumptions = [
        {
            "id": item["id"],
            "category": item["category"],
            "description": item["description"],
            "effective_periods": item["effective_periods"],
            "rationale": item["rationale"],
            "status": item["status"],
        }
        for item in plan["assumptions"]
    ]
    return {
        "schema_version": "mparanza.business_planning_financial_model_context.v2",
        "workflow_id": WORKFLOW_ID,
        "case_id": plan["case_id"],
        "entity_name": plan["entity_name"],
        "company_stage": plan["company_stage"],
        "planning_objective": plan["planning_objective"],
        "professional_lens": plan["professional_lens"],
        "audience": plan["audience"],
        "reporting_currency": plan["reporting_currency"],
        "status": plan["status"],
        "review_status": plan["review_status"],
        "evidence_coverage": plan["evidence_coverage"],
        "evidence_register": plan["evidence_register"],
        "assumptions": assumptions,
        "scenario_summaries": [
            {
                "id": scenario["id"],
                "label": scenario["label"],
                "summary": scenario["summary"],
            }
            for scenario in plan["scenarios"]
        ],
        "scenario_periods": [
            {
                "scenario_id": scenario["id"],
                "periods": scenario["periods"],
            }
            for scenario in plan["scenarios"]
        ],
        "reconciliation": plan["reconciliation"],
        "limitations": plan["limitations"],
        "excluded_by_default": [
            "raw source populations",
            "absolute source paths",
            "original filenames",
            "evidence source_ref values",
        ],
    }


def build_counterpart_handoff(plan: Mapping[str, Any]) -> dict[str, Any]:
    """Build the reviewed Vera-to-Clara bridge without semantic merging."""

    return {
        "schema_version": HANDOFF_SCHEMA,
        "workflow_id": WORKFLOW_ID,
        "case_id": plan["case_id"],
        "entity_name": plan["entity_name"],
        "company_stage": plan["company_stage"],
        "planning_objective": plan["planning_objective"],
        "audience": plan["audience"],
        "from_product": "Vera",
        "from_lens": "accounting_financial",
        "to_product": "Clara",
        "to_lens": "strategic_commercial",
        "status": counterpart_handoff_status(plan["status"]),
        "source_plan_status": plan["status"],
        "source_review_status": plan["review_status"],
        "assumptions": [
            {
                "id": item["id"],
                "category": item["category"],
                "description": item["description"],
                "effective_periods": item["effective_periods"],
                "rationale": item["rationale"],
                "status": item["status"],
            }
            for item in plan["assumptions"]
        ],
        "financial_scenario_summaries": [
            {
                "id": scenario["id"],
                "label": scenario["label"],
                "summary": scenario["summary"],
            }
            for scenario in plan["scenarios"]
        ],
        "reconciliation": plan["reconciliation"],
        "limitations": plan["limitations"],
        "handoff_boundary": (
            "Clara may use this reviewed bridge as evidence for strategic work, but "
            "must not alter accounting assumptions, figures, or reconciliation status "
            "silently. Any divergence must be stated and returned for professional review."
        ),
    }


def write_assumption_ledger(path: Path, plan: Mapping[str, Any]) -> None:
    """Write the reviewed assumption register as CSV."""

    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=(
                "id",
                "category",
                "description",
                "evidence_ids",
                "effective_periods",
                "rationale",
                "status",
            ),
        )
        writer.writeheader()
        for item in plan["assumptions"]:
            writer.writerow(
                {
                    **item,
                    "evidence_ids": "|".join(item["evidence_ids"]),
                    "effective_periods": "|".join(item["effective_periods"]),
                }
            )


def _style_sheet(worksheet: Any) -> None:
    worksheet.freeze_panes = "A2"
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="002060")
        cell.alignment = Alignment(horizontal="center")
    for column_cells in worksheet.columns:
        width = min(
            42,
            max(12, max(len(str(cell.value or "")) for cell in column_cells) + 2),
        )
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = (
            width
        )


def _append_statement_rows(
    worksheet: Any,
    plan: Mapping[str, Any],
    statement: str,
) -> None:
    worksheet.append(["Scenario", "Period", "Line", "Amount", "Currency"])
    for scenario in plan["scenarios"]:
        for period in scenario["periods"]:
            for line, amount in period[statement].items():
                if line == "passed":
                    continue
                worksheet.append(
                    [
                        scenario["label"],
                        period["period"],
                        line,
                        float(Decimal(amount)),
                        plan["reporting_currency"],
                    ]
                )
    for row in worksheet.iter_rows(min_row=2, min_col=4, max_col=4):
        row[0].number_format = "#,##0.00;[Red]-#,##0.00"
    _style_sheet(worksheet)


def write_excel(path: Path, plan: Mapping[str, Any]) -> None:
    """Write a review workbook; exact monetary authority remains the JSON."""

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    review = workbook.active
    review.title = "Review"
    review.append(["Field", "Value"])
    review_rows = (
        ("Entity", plan["entity_name"]),
        ("Company stage", plan["company_stage"]),
        ("Planning objective", plan["planning_objective"]),
        ("Professional lens", plan["professional_lens"]),
        ("Audience", plan["audience"]),
        ("Currency", plan["reporting_currency"]),
        ("Status", plan["status"]),
        ("Review status", plan["review_status"]),
        (
            "Reconciliation",
            "passed" if plan["reconciliation"]["all_passed"] else "failed",
        ),
    )
    for row in review_rows:
        review.append(row)
    _style_sheet(review)

    assumptions = workbook.create_sheet("Assumptions")
    assumptions.append(
        [
            "ID",
            "Category",
            "Description",
            "Evidence IDs",
            "Effective periods",
            "Rationale",
            "Status",
        ]
    )
    for item in plan["assumptions"]:
        assumptions.append(
            [
                item["id"],
                item["category"],
                item["description"],
                " | ".join(item["evidence_ids"]),
                " | ".join(item["effective_periods"]),
                item["rationale"],
                item["status"],
            ]
        )
    _style_sheet(assumptions)

    _append_statement_rows(
        workbook.create_sheet("Profit and loss"), plan, "profit_and_loss"
    )
    _append_statement_rows(workbook.create_sheet("Cash flow"), plan, "cash_flow")
    _append_statement_rows(
        workbook.create_sheet("Balance sheet"), plan, "balance_sheet"
    )

    reconciliation = workbook.create_sheet("Reconciliation")
    reconciliation.append(
        [
            "Scenario",
            "Period",
            "Total assets",
            "Liabilities and equity",
            "Difference",
            "Passed",
        ]
    )
    for item in plan["reconciliation"]["periods"]:
        reconciliation.append(
            [
                item["scenario_id"],
                item["period"],
                float(Decimal(item["total_assets"])),
                float(Decimal(item["total_liabilities_and_equity"])),
                float(Decimal(item["difference"])),
                item["passed"],
            ]
        )
    for row in reconciliation.iter_rows(min_row=2, min_col=3, max_col=5):
        for cell in row:
            cell.number_format = "#,##0.00;[Red]-#,##0.00"
    _style_sheet(reconciliation)
    workbook.save(path)


def render_markdown(plan: Mapping[str, Any]) -> str:
    """Render a factual planning summary without semantic conclusions."""

    lines = [
        f"# Business plan facts — {plan['entity_name']}",
        "",
        f"- Status: `{plan['status']}`",
        f"- Review: `{plan['review_status']}`",
        f"- Company stage: {plan['company_stage']}",
        f"- Planning objective: {plan['planning_objective']}",
        f"- Professional lens: `{plan['professional_lens']}`",
        f"- Audience: {plan['audience']}",
        f"- Currency: {plan['reporting_currency']}",
        f"- Reconciliation: {'passed' if plan['reconciliation']['all_passed'] else 'failed'}",
        "",
        "## Scenario summaries",
        "",
        "| Scenario | Revenue | Net income | Ending cash | Minimum cash | Funding requirement | Break-even |",
        "| --- | ---: | ---: | ---: | ---: | ---: | --- |",
    ]
    for scenario in plan["scenarios"]:
        summary = scenario["summary"]
        lines.append(
            "| {label} | {revenue} | {income} | {ending} | {minimum} | {funding} | {break_even} |".format(
                label=scenario["label"],
                revenue=summary["total_revenue"],
                income=summary["total_net_income"],
                ending=summary["ending_cash"],
                minimum=summary["minimum_cash"],
                funding=summary["funding_requirement"],
                break_even=summary["break_even_period"] or "not reached",
            )
        )
    lines.extend(["", "## Limitations", ""])
    lines.extend(f"- {item}" for item in plan["limitations"])
    lines.extend(
        [
            "",
            "Exact statement closure does not validate the assumptions or approve the plan.",
        ]
    )
    return "\n".join(lines) + "\n"


def render_html(plan: Mapping[str, Any]) -> str:
    """Render a self-contained factual page with the complete review surface."""

    def render_table(headers: Sequence[str], rows: Sequence[Sequence[object]]) -> str:
        head = "".join(f"<th>{html.escape(header)}</th>" for header in headers)
        body = "".join(
            "<tr>"
            + "".join(f"<td>{html.escape(str(value))}</td>" for value in row)
            + "</tr>"
            for row in rows
        )
        return (
            '<div class="table-wrap"><table><thead><tr>'
            f"{head}</tr></thead><tbody>{body}</tbody></table></div>"
        )

    scenario_summary = render_table(
        (
            "Scenario",
            "Revenue",
            "Net income",
            "Ending cash",
            "Minimum cash",
            "Funding requirement",
            "Break-even",
        ),
        [
            (
                scenario["label"],
                scenario["summary"]["total_revenue"],
                scenario["summary"]["total_net_income"],
                scenario["summary"]["ending_cash"],
                scenario["summary"]["minimum_cash"],
                scenario["summary"]["funding_requirement"],
                scenario["summary"]["break_even_period"] or "not reached",
            )
            for scenario in plan["scenarios"]
        ],
    )
    assumption_table = render_table(
        (
            "ID",
            "Category",
            "Description",
            "Evidence",
            "Effective periods",
            "Rationale",
            "Status",
        ),
        [
            (
                item["id"],
                item["category"],
                item["description"],
                " · ".join(item["evidence_ids"]),
                " · ".join(item["effective_periods"]),
                item["rationale"],
                item["status"],
            )
            for item in plan["assumptions"]
        ],
    )
    evidence_table = render_table(
        ("ID", "Kind", "Description", "Status"),
        [
            (
                item["id"],
                item["kind"],
                item["description"],
                item["status"],
            )
            for item in plan["evidence_register"]
        ],
    )
    opening = plan["opening_balance"]
    opening_table = render_table(
        ("Line", "Amount"),
        [(line, value) for line, value in opening["values"].items()],
    )
    reconciliation_table = render_table(
        (
            "Scenario",
            "Period",
            "Total assets",
            "Liabilities and equity",
            "Difference",
            "Passed",
        ),
        [
            (
                item["scenario_id"],
                item["period"],
                item["total_assets"],
                item["total_liabilities_and_equity"],
                item["difference"],
                item["passed"],
            )
            for item in plan["reconciliation"]["periods"]
        ],
    )
    scenario_details: list[str] = []
    for scenario in plan["scenarios"]:
        scenario_details.append(
            f"<section><h3>{html.escape(str(scenario['label']))}</h3>"
        )
        for statement, title in (
            ("profit_and_loss", "Profit and loss"),
            ("cash_flow", "Cash flow"),
            ("balance_sheet", "Balance sheet"),
        ):
            line_names = list(scenario["periods"][0][statement])
            scenario_details.append(f"<h4>{title}</h4>")
            scenario_details.append(
                render_table(
                    ("Period", *line_names),
                    [
                        (
                            period["period"],
                            *(period[statement][line] for line in line_names),
                        )
                        for period in scenario["periods"]
                    ],
                )
            )
        scenario_details.append("</section>")
    limitations = "".join(
        f"<li>{html.escape(str(item))}</li>" for item in plan["limitations"]
    )
    evidence = plan["evidence_coverage"]
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Business plan facts — {html.escape(str(plan['entity_name']))}</title>
  <style>
    :root {{ color-scheme: light; font-family: "Instrument Sans", sans-serif; color: #17213a; background: #f6f8fc; }}
    body {{ margin: 0; }} main {{ max-width: 1180px; margin: 0 auto; padding: 48px 24px 72px; }}
    h1 {{ color: #002060; font-size: clamp(2rem, 5vw, 4rem); margin: 0 0 12px; }}
    h2 {{ margin-top: 48px; }} h3 {{ margin-top: 32px; }} h4 {{ margin: 28px 0 8px; color: #31527e; }}
    .meta {{ display: flex; flex-wrap: wrap; gap: 12px 24px; border-top: 1px solid #bfd1ed; border-bottom: 1px solid #bfd1ed; padding: 18px 0; }}
    .table-wrap {{ overflow-x: auto; }} table {{ width: 100%; border-collapse: collapse; margin-top: 12px; background: white; }}
    th, td {{ padding: 12px; border-bottom: 1px solid #dbe4f2; text-align: right; }}
    th:first-child, td:first-child {{ text-align: left; }} th {{ color: #002060; white-space: nowrap; }} td {{ vertical-align: top; }}
    .warning {{ margin-top: 28px; padding: 18px; border-left: 3px solid #00a7d8; background: #edf8fc; }}
  </style>
</head>
<body><main>
  <p>Vera · Business planning · accounting and financial lens</p>
  <h1>{html.escape(str(plan['entity_name']))}</h1>
  <div class="meta"><span>Company stage: {html.escape(str(plan['company_stage']))}</span><span>Status: {html.escape(str(plan['status']))}</span><span>Review: {html.escape(str(plan['review_status']))}</span><span>Currency: {html.escape(str(plan['reporting_currency']))}</span></div>
  <p><strong>Objective:</strong> {html.escape(str(plan['planning_objective']))}</p>
  <p><strong>Audience:</strong> {html.escape(str(plan['audience']))}</p>
  <h2>Evidence coverage</h2>{evidence_table}
  <p>Unverified evidence: {html.escape(' · '.join(evidence['unverified_evidence_ids']) or 'none')}.</p>
  <h2>Opening balance</h2>{opening_table}
  <p>Evidence: {html.escape(' · '.join(opening['evidence_ids']))}. Reconciliation difference: {html.escape(str(opening['reconciliation']['difference']))}; passed: {html.escape(str(opening['reconciliation']['passed']))}.</p>
  <h2>Confirmed assumptions</h2>{assumption_table}
  <h2>Scenario summaries</h2>
  {scenario_summary}
  <h2>Integrated statements</h2>{''.join(scenario_details)}
  <h2>Reconciliation</h2>{reconciliation_table}
  <h2>Limitations</h2><ul>{limitations}</ul>
  <p class="warning">Exact statement closure does not validate the assumptions or approve the plan.</p>
</main></body></html>
"""
