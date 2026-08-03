from __future__ import annotations

import argparse
import hashlib
import json
import logging
import math
import os
import re
import shutil
import stat
import sys
import tempfile
from collections.abc import Mapping
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any, Sequence
from zipfile import BadZipFile

import fastexcel
import openpyxl
import polars as pl
from implementation_bootstrap import (
    IMPLEMENTATION_CONTRACT,
    validate_implementation_tree,
)
from openpyxl.utils.exceptions import InvalidFileException

_COMPONENT_ROOT = Path(__file__).resolve().parents[1]
_VENDOR_CANDIDATES = (
    _COMPONENT_ROOT / "vendor" / "modules",
    _COMPONENT_ROOT.parent.parent / "vendor" / "modules",
    _COMPONENT_ROOT.parent / "_shared" / "vendor" / "modules",
)
for _vendor_candidate in _VENDOR_CANDIDATES:
    if (_vendor_candidate / "vera_assurance").is_dir():
        if str(_vendor_candidate) not in sys.path:
            sys.path.insert(0, str(_vendor_candidate))
        break

import vera_assurance as _vera_assurance  # noqa: E402
from vera_assurance import (  # noqa: E402
    DecisionReceiptError,
    MoneyValidationError,
    artifact_receipt,
    build_assurance_envelope,
    build_gate_register,
    build_reviewed_decision_receipt,
    build_source_qualification,
    canonical_json_sha256,
    decimal_text,
    file_snapshot,
    load_client_engagement_context_file,
    parse_canonical_decimal,
    parse_localized_decimal,
    validate_artifact_receipt,
    validate_assurance_envelope,
    validate_client_engagement_context,
    validate_reviewed_decision_receipt,
    validate_source_qualification,
)
from vera_assurance import write_json as write_assurance_json  # noqa: E402

try:
    from .review_session import (
        workbook_sheet_name,
        write_review_session_artifacts,
        write_run_intake,
    )
except ImportError:  # pragma: no cover - supports direct script imports
    import importlib.util

    _review_session_path = Path(__file__).resolve().parent / "review_session.py"
    _review_session_spec = importlib.util.spec_from_file_location(
        "mparanza_journal_sampling_review_session",
        _review_session_path,
    )
    if _review_session_spec is None or _review_session_spec.loader is None:
        raise ImportError("Could not load Journal Sampling review-session helpers.")
    _review_session = importlib.util.module_from_spec(_review_session_spec)
    sys.modules[_review_session_spec.name] = _review_session
    _review_session_spec.loader.exec_module(_review_session)
    workbook_sheet_name = _review_session.workbook_sheet_name
    write_review_session_artifacts = _review_session.write_review_session_artifacts
    write_run_intake = _review_session.write_run_intake

LOGGER = logging.getLogger(__name__)

SUPPORTED_SUFFIXES = {".csv", ".xls", ".xlsx", ".xlsm", ".pdf"}
SUPPORTED_LANGUAGES = ("it", "en", "fr", "de", "es")
CANONICAL_COLUMNS = [
    "entry_date",
    "movement_number",
    "line_number",
    "account",
    "account_desc",
    "line_desc",
    "debit",
    "credit",
    "amount_signed",
    "amount_abs",
    "currency",
    "unit",
    "reported_increment",
    "source_file",
    "source_sheet",
    "source_page",
    "source_row",
]
CANONICAL_SCHEMA: dict[str, pl.DataType] = {
    "entry_date": pl.Utf8,
    "movement_number": pl.Utf8,
    "line_number": pl.Utf8,
    "account": pl.Utf8,
    "account_desc": pl.Utf8,
    "line_desc": pl.Utf8,
    "debit": pl.Utf8,
    "credit": pl.Utf8,
    "amount_signed": pl.Utf8,
    "amount_abs": pl.Utf8,
    "currency": pl.Utf8,
    "unit": pl.Utf8,
    "reported_increment": pl.Utf8,
    "source_file": pl.Utf8,
    "source_sheet": pl.Utf8,
    "source_page": pl.Int64,
    "source_row": pl.Int64,
}
TABULAR_ADAPTER_ID = "journal.tabular.v2"
TABULAR_SOURCE_FAMILY = "tabular.explicit_mapping.v1"
PRINT_ADAPTER_ID = "journal.print_friendly_excel.v2"
PRINT_SOURCE_FAMILY = "print_friendly.debit_credit_columns.v1"
DISABLED_PDF_ADAPTER_ID = "journal.text_pdf.disabled.v2"
ADAPTER_VERSION = "2"
NORMALIZATION_SCHEMA_VERSION = "journal_sampling.normalization.v2"
QUALIFICATION_REVIEW_SCHEMA_VERSION = "journal_sampling.qualification_review.v1"
SAMPLE_ASSURANCE_VERSION = "1"
SAMPLE_REPRODUCIBILITY_SCHEMA_VERSION = "journal_sampling.sample_reproducibility.v1"
SAMPLE_MATERIAL_LEDGER_SCHEMA_VERSION = (
    "journal_sampling.sample_material_value_ledger.v1"
)
SAMPLE_OUTPUT_SET_SCHEMA_VERSION = "journal_sampling.sample_output_set.v2"
SAMPLE_REVIEW_SUCCESSOR_SCHEMA_VERSION = "journal_sampling.review_successor.v1"
SAMPLE_OUTPUT_SET_MODE = 0o600
SAMPLE_MATERIAL_FIELDS = tuple(CANONICAL_COLUMNS)
SAMPLE_DECIMAL_FIELDS = {
    "debit",
    "credit",
    "amount_signed",
    "amount_abs",
    "reported_increment",
}
SAMPLE_INTEGER_FIELDS = {"source_page", "source_row"}
SAMPLE_OUTPUT_PAYLOAD_PATHS = (
    "journal_sample.csv",
    "journal_sample.xlsx",
    "sampling_audit.json",
    "run_intake.json",
    "review_payload.json",
    "ui_decisions.json",
    "review_handoff.md",
    "final_artifacts.json",
    "sample_reproducibility.json",
    "sample_material_value_ledger.json",
    "sample_assurance_gates.json",
    "sample_assurance_envelope.json",
)
SAMPLE_OUTPUT_SET_PATH = "sample_output_receipts.json"
SAMPLE_ASSURANCE_HISTORY_DIR = "assurance_history"
SAMPLE_APPLIED_DECISIONS_PATH = "applied_decisions.json"
IMPLEMENTATION_PLUGIN_FILES = (
    ("scripts/check_dependencies.py", "implementation.journal_sampling_dependencies"),
    (
        "scripts/implementation_bootstrap.py",
        "implementation.journal_sampling_bootstrap",
    ),
    ("scripts/inspect_journal.py", "implementation.journal_sampling_inspection_cli"),
    ("scripts/journal_sampling_core.py", "implementation.journal_sampling_core"),
    (
        "scripts/normalize_journal.py",
        "implementation.journal_sampling_normalization_cli",
    ),
    (
        "scripts/replay_normalization.py",
        "implementation.journal_sampling_normalization_replay_cli",
    ),
    ("scripts/review_session.py", "implementation.journal_sampling_review_session"),
    ("scripts/review_successor.py", "implementation.journal_sampling_review_successor"),
    ("scripts/run_sample.py", "implementation.journal_sampling_sample_cli"),
    ("mcp/server.cjs", "implementation.journal_sampling_mcp"),
    ("assets/icon.svg", "implementation.journal_sampling_icon"),
    (
        "assets/journal-sampling-review-widget.html",
        "implementation.journal_sampling_widget",
    ),
    (
        "assets/review-workbench-adapter.json",
        "implementation.journal_sampling_widget_adapter",
    ),
    (".app.json", "implementation.journal_sampling_app_config"),
    (".mcp.json", "implementation.journal_sampling_mcp_config"),
    (".codex-plugin/plugin.json", "implementation.journal_sampling_plugin_config"),
)
ASSURANCE_IMPLEMENTATION_FILES = (
    ("__init__.py", "implementation.vera_assurance_init"),
    ("contracts.py", "implementation.vera_assurance_contracts"),
    ("decisions.py", "implementation.vera_assurance_decisions"),
    ("envelope.py", "implementation.vera_assurance_envelope"),
    ("money.py", "implementation.vera_assurance_money"),
    ("relationships.py", "implementation.vera_assurance_relationships"),
    (
        "review_output_transaction.cjs",
        "implementation.vera_assurance_review_output_transaction",
    ),
    ("serialization.py", "implementation.vera_assurance_serialization"),
)
ZERO = Decimal("0")
DATE_FORMATS = (
    "%Y-%m-%d",
    "%d/%m/%Y",
    "%d-%m-%Y",
    "%d.%m.%Y",
    "%d/%m/%y",
    "%m/%d/%Y",
    "%Y/%m/%d",
)
DATE_TOKEN_RE = re.compile(
    r"\b(\d{1,2}[./-]\d{1,2}[./-]\d{2,4}|\d{4}[./-]\d{1,2}[./-]\d{1,2})\b"
)
ACCOUNT_RE = re.compile(r"^[A-Za-z]?\s*\d+(?:\s*[./-]\s*\d+){0,4}[A-Za-z]?$")


__all__ = [
    "CANONICAL_COLUMNS",
    "CANONICAL_SCHEMA",
    "InspectionResult",
    "NormalizationResult",
    "SampleResult",
    "language_assumptions",
    "inspect_path",
    "load_client_engagement_context",
    "normalize_language",
    "normalize_path",
    "prepare_sample_review_successor",
    "replay_normalization_from_provenance",
    "finalize_sample_review_successor",
    "run_sample",
    "validate_sample_assurance",
    "validate_sample_material_value_ledger",
    "validate_sample_output_set",
    "write_json",
]


def normalize_language(
    language: object | None,
    *,
    default: str = "en",
    allow_auto: bool = False,
) -> str:
    """Normalize a language tag to one of the supported plugin locales."""

    text = str(language or default).strip().lower().replace("_", "-")
    code = text.split("-", 1)[0]
    if allow_auto and code == "auto":
        return "auto"
    return code if code in SUPPORTED_LANGUAGES else default


def language_assumptions(
    recipe: dict[str, Any],
    *,
    language: object | None = None,
    document_language: object | None = None,
) -> dict[str, str]:
    """Resolve working and source-document language assumptions."""

    working = normalize_language(language or recipe.get("language"), default="en")
    source = normalize_language(
        document_language or recipe.get("document_language") or "auto",
        default=working,
        allow_auto=True,
    )
    return {"language": working, "document_language": source}


@dataclass(frozen=True)
class InspectionResult:
    """Deterministic inspection output for one or more journal files."""

    files: list[dict[str, Any]]
    total_rows: int
    suggested_recipe: dict[str, Any]


@dataclass(frozen=True)
class NormalizationResult:
    """Normalized journal rows plus parser diagnostics."""

    frame: pl.DataFrame
    diagnostics: dict[str, Any]


@dataclass(frozen=True)
class SampleResult:
    """Sample rows plus reproducibility metadata."""

    frame: pl.DataFrame
    audit: dict[str, Any]


@dataclass(frozen=True)
class QualificationPlan:
    """Pre-normalization decision for one bounded source adapter."""

    parser: str
    adapter_id: str
    source_family: str
    status: str
    candidate_row_count: int
    controls: list[dict[str, Any]]
    limitations: list[str]
    reviewed_mapping_ref: str | None
    reviewed_decision: dict[str, Any] | None
    header_rows: list[int]
    mapping: dict[str, Any]
    layout: dict[str, Any]
    excluded_monetary_columns: list[str]
    unresolved_monetary_columns: list[str]
    posting_identity: str
    carry_forward_fields: list[str]
    currency: str
    unit: str
    decimal_separator: str | None
    thousands_separator: str | None
    amount_sign_convention: str | None


def configure_logging(verbose: bool = False) -> None:
    """Configure script logging without affecting imported use."""

    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(level=level, format="%(levelname)s: %(message)s")


def write_json(path: Path, payload: Any) -> None:
    """Write JSON using the shared no-binary-float assurance contract."""

    if not isinstance(payload, dict):
        raise ValueError("Journal Sampling JSON artifacts must be objects.")
    write_assurance_json(path, payload)


def read_json(path: Path | None) -> dict[str, Any]:
    """Return a JSON object or an empty mapping when no file is provided."""

    if path is None:
        return {}
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError(f"Recipe must be a JSON object: {path}")
    return payload


def load_client_engagement_context(path: Path) -> dict[str, Any]:
    """Load one exact client workflow context created by Studio Archive."""

    try:
        return load_client_engagement_context_file(
            path.expanduser().resolve(strict=True),
            expected_workflow_id="journal-sampling",
        )
    except (OSError, ValueError) as exc:
        raise ValueError(f"Client engagement context is invalid: {exc}") from exc


def _is_path_within(path: Path, parent: Path) -> bool:
    try:
        path.relative_to(parent)
    except ValueError:
        return False
    return True


def _managed_run_reference(
    path_value: Path,
    client_engagement: Mapping[str, Any] | None,
) -> str:
    """Persist a managed artifact beneath its run root without folder identity."""

    if client_engagement is None:
        return path_value.as_posix()
    run_root_value = client_engagement.get("run_root")
    if not isinstance(run_root_value, str) or not run_root_value.strip():
        return path_value.as_posix()
    run_root = Path(run_root_value).expanduser().resolve()
    resolved = path_value.expanduser().resolve()
    try:
        relative = resolved.relative_to(run_root)
    except ValueError as exc:
        raise ValueError("Journal Sampling path is outside the run root.") from exc
    if not relative.parts or ".." in relative.parts:
        raise ValueError("Journal Sampling path must identify a run artifact.")
    return relative.as_posix()


def _portable_client_engagement_context(
    client_engagement: Mapping[str, Any] | None,
) -> dict[str, Any] | None:
    """Return the path-free identity persisted in managed workflow artifacts."""

    if (
        not isinstance(client_engagement, Mapping)
        or client_engagement.get("schema_version") != "vera.client_workflow_context.v2"
    ):
        return (
            dict(client_engagement) if isinstance(client_engagement, Mapping) else None
        )
    portable_fields = (
        "schema_version",
        "client_id",
        "engagement_id",
        "workflow_id",
        "workflow_version",
        "run_id",
        "label",
        "purpose",
        "created_at",
        "input_manifest",
        "input_manifest_sha256",
        "run_relative_path",
        "output_relative_path",
        "content_sha256",
    )
    return {field: client_engagement[field] for field in portable_fields}


def _current_journal_context(
    anchor: Path,
    persisted_context: Mapping[str, Any],
) -> dict[str, Any]:
    """Load the current Journal Sampling run context from the renamed tree."""

    if persisted_context.get("schema_version") != "vera.client_workflow_context.v2":
        raise ValueError("Journal Sampling portable context is unavailable.")
    engagement_id = persisted_context.get("engagement_id")
    run_id = persisted_context.get("run_id")
    if not isinstance(engagement_id, str) or not isinstance(run_id, str):
        raise ValueError("Journal Sampling portable context identity is invalid.")
    resolved_anchor = anchor.expanduser().resolve()
    candidates: list[Path] = []
    for candidate in (resolved_anchor, *resolved_anchor.parents):
        direct = candidate / "context.json"
        if direct.is_file() and not direct.is_symlink():
            candidates.append(direct)
        if candidate.name == "Vera":
            candidates.append(
                candidate
                / "engagements"
                / engagement_id
                / "runs"
                / run_id
                / "context.json"
            )
    current: dict[str, Any] | None = None
    for context_path in candidates:
        if not context_path.is_file() or context_path.is_symlink():
            continue
        try:
            candidate_context = load_client_engagement_context_file(
                context_path,
                expected_workflow_id="journal-sampling",
                allowed_statuses=("running", "ready_for_review", "completed"),
            )
        except (OSError, ValueError):
            continue
        if (
            candidate_context.get("engagement_id") == engagement_id
            and candidate_context.get("run_id") == run_id
            and candidate_context.get("client_id") == persisted_context.get("client_id")
        ):
            current = candidate_context
            break
    if current is None:
        raise ValueError(
            "Journal Sampling current customer-run context is unavailable."
        )
    return current


def _resolve_normalization_reference(
    anchor: Path,
    diagnostics: Mapping[str, Any],
    value: object,
    *,
    label: str,
) -> Path:
    """Resolve one sealed normalization reference through current run authority."""

    if not isinstance(value, str) or not value.strip() or value != value.strip():
        raise ValueError(f"{label} is unavailable.")
    reference = Path(value)
    if diagnostics.get("path_reference") != "run_root_relative":
        if not reference.is_absolute() or reference.resolve() != reference:
            raise ValueError(f"{label} is not canonical.")
        return reference
    if (
        reference.is_absolute()
        or ".." in reference.parts
        or reference.as_posix() != value
    ):
        raise ValueError(f"{label} leaves the Journal Sampling run.")
    persisted_context = diagnostics.get("client_engagement")
    if not isinstance(persisted_context, Mapping):
        raise ValueError("Journal Sampling portable context is unavailable.")
    current_context = _current_journal_context(anchor, persisted_context)
    run_root = Path(str(current_context["run_root"])).expanduser().resolve()
    resolved = (run_root / reference).resolve()
    if not _is_path_within(resolved, run_root) or resolved == run_root:
        raise ValueError(f"{label} leaves the Journal Sampling run.")
    return resolved


def _validated_client_normalization_stage(
    value: Mapping[str, Any] | None,
    *,
    input_path: Path,
    output_dir: Path,
) -> dict[str, Any] | None:
    if value is None:
        return None
    try:
        context = validate_client_engagement_context(value)
    except ValueError as exc:
        raise ValueError(f"Client engagement context is invalid: {exc}") from exc
    if context["workflow_id"] != "journal-sampling":
        raise ValueError("Client engagement is not for Journal Sampling.")
    resolved_input = input_path.expanduser().resolve(strict=True)
    if context["schema_version"] == "vera.client_workflow_context.v2":
        journal_inputs = {
            Path(item["path"]).resolve(strict=True)
            for item in context["input_bindings"]
            if item["kind"] == "import" and item["role"] == "journal"
        }
        if journal_inputs != {resolved_input}:
            raise ValueError(
                "Journal input is not the run's one exact journal receipt."
            )
    else:
        input_root = Path(context["input_dir"]).resolve(strict=True)
        if not _is_path_within(resolved_input, input_root):
            raise ValueError("Journal input is outside the selected client engagement.")
    expected_output = Path(context["output_dir"]) / "normalization"
    if output_dir.expanduser().resolve() != expected_output.resolve():
        raise ValueError(
            "Journal normalization output does not match the client engagement."
        )
    return context


def _validated_client_sample_stage(
    value: Mapping[str, Any] | None,
    *,
    normalized_csv: Path,
    output_dir: Path,
) -> dict[str, Any] | None:
    if value is None:
        return None
    try:
        context = validate_client_engagement_context(value)
    except ValueError as exc:
        raise ValueError(f"Client engagement context is invalid: {exc}") from exc
    if context["workflow_id"] != "journal-sampling":
        raise ValueError("Client engagement is not for Journal Sampling.")
    expected_csv = (
        Path(context["output_dir"]) / "normalization" / "normalized_journal.csv"
    )
    expected_output = Path(context["output_dir"]) / "sample"
    if normalized_csv.expanduser().resolve(strict=True) != expected_csv.resolve(
        strict=True
    ):
        raise ValueError(
            "Normalized journal does not belong to the selected client engagement."
        )
    if output_dir.expanduser().resolve() != expected_output.resolve():
        raise ValueError("Sample output does not match the client engagement.")
    diagnostics = read_json(expected_csv.parent / "normalization_diagnostics.json")
    diagnostics_context = diagnostics.get("client_engagement")
    if not isinstance(diagnostics_context, Mapping):
        raise ValueError("Normalized journal client engagement is missing or stale.")
    try:
        normalized_diagnostics_context = validate_client_engagement_context(
            diagnostics_context
        )
    except ValueError as exc:
        raise ValueError(
            "Normalized journal client engagement is missing or stale."
        ) from exc
    stable_keys = ("engagement_id", "workflow_id", "run_id", "content_sha256")
    if any(
        normalized_diagnostics_context.get(key) != context.get(key)
        for key in stable_keys
    ):
        raise ValueError("Normalized journal client engagement is missing or stale.")
    return context


def _read_recipe_with_receipt(
    path: Path | None,
) -> tuple[
    dict[str, Any],
    Path | None,
    dict[str, Any] | None,
    bytes | None,
]:
    """Read one stable reviewed recipe and bind its original bytes."""

    if path is None:
        return {}, None, None, None
    recipe_path = path.expanduser().resolve()
    _require_ordinary_single_link(recipe_path, label="Normalization recipe")
    recipe_root = recipe_path.parent
    receipt_before = artifact_receipt(
        recipe_root,
        recipe_path,
        artifact_id="decision.normalization_recipe_source",
        root_id="normalization_recipe_source",
        role="reviewed_recipe",
        media_type="application/json",
    )
    with recipe_path.open("rb") as handle:
        before = os.fstat(handle.fileno())
        recipe_bytes = handle.read()
        after = os.fstat(handle.fileno())
    before_identity = (
        before.st_dev,
        before.st_ino,
        before.st_size,
        before.st_mtime_ns,
    )
    after_identity = (
        after.st_dev,
        after.st_ino,
        after.st_size,
        after.st_mtime_ns,
    )
    if (
        not stat.S_ISREG(before.st_mode)
        or before.st_nlink != 1
        or before_identity != after_identity
        or len(recipe_bytes) != after.st_size
    ):
        raise ValueError("Normalization recipe changed while it was read.")
    try:
        payload = json.loads(recipe_bytes.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ValueError(f"Recipe must be valid UTF-8 JSON: {recipe_path}") from exc
    if not isinstance(payload, dict):
        raise ValueError(f"Recipe must be a JSON object: {recipe_path}")
    receipt_after = artifact_receipt(
        recipe_root,
        recipe_path,
        artifact_id="decision.normalization_recipe_source",
        root_id="normalization_recipe_source",
        role="reviewed_recipe",
        media_type="application/json",
    )
    if receipt_before != receipt_after:
        raise ValueError("Normalization recipe changed while it was receipted.")
    return payload, recipe_path, receipt_after, recipe_bytes


def _assurance_package_root() -> Path:
    package_file = getattr(_vera_assurance, "__file__", None)
    if not isinstance(package_file, str) or not package_file:
        raise ValueError("The Vera assurance implementation root is unavailable.")
    return Path(package_file).resolve().parent


def _implementation_artifact_roots() -> dict[str, Path]:
    return {
        "implementation": _COMPONENT_ROOT,
        "assurance_implementation": _assurance_package_root(),
    }


def _require_ordinary_single_link(path: Path, *, label: str) -> None:
    """Require one receipt input to be an ordinary single-link file.

    This deterministic check is justified because an assurance receipt must
    identify one mechanically stable file, never an alias or special entry.
    """

    try:
        observed = path.lstat()
    except FileNotFoundError as exc:
        raise ValueError(f"{label} is missing.") from exc
    if (
        stat.S_ISLNK(observed.st_mode)
        or not stat.S_ISREG(observed.st_mode)
        or observed.st_nlink != 1
    ):
        raise ValueError(f"{label} must be an ordinary single-link file.")


def _implementation_media_type(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".py":
        return "text/x-python"
    if suffix == ".cjs":
        return "text/javascript"
    if suffix == ".html":
        return "text/html"
    if suffix == ".svg":
        return "image/svg+xml"
    return "application/json"


def _validate_implementation_configuration() -> None:
    """Require the receipted discovery and review configuration to agree."""

    plugin_manifest = read_json(_COMPONENT_ROOT / ".codex-plugin" / "plugin.json")
    app_manifest = read_json(_COMPONENT_ROOT / ".app.json")
    mcp_manifest = read_json(_COMPONENT_ROOT / ".mcp.json")
    adapter = read_json(_COMPONENT_ROOT / "assets" / "review-workbench-adapter.json")
    if (
        plugin_manifest.get("name") != "journal-sampling"
        or plugin_manifest.get("skills") != "./skills/"
        or plugin_manifest.get("apps") != "./.app.json"
        or plugin_manifest.get("mcpServers") != "./.mcp.json"
        or app_manifest != {"apps": {}}
    ):
        raise ValueError("Journal Sampling plugin discovery configuration is stale.")
    servers = mcp_manifest.get("mcpServers")
    if not isinstance(servers, dict) or set(servers) != {"journalSamplingWidgets"}:
        raise ValueError("Journal Sampling MCP configuration is not exact.")
    server = servers["journalSamplingWidgets"]
    if (
        not isinstance(server, dict)
        or server.get("cwd") != "."
        or server.get("command") != "node"
        or server.get("args") != ["./mcp/server.cjs", "--stdio"]
    ):
        raise ValueError("Journal Sampling MCP launch contract is stale.")
    if (
        adapter.get("plugin") != "journal-sampling"
        or adapter.get("saveTool") != "save_journal_sampling_decisions"
        or adapter.get("applyTool") != "apply_journal_sampling_decisions"
        or adapter.get("widgetType") != "journal_sampling_review"
    ):
        raise ValueError("Journal Sampling review adapter contract is stale.")
    embedded_adapter = (
        "const CONFIG = "
        + json.dumps(adapter, ensure_ascii=True, separators=(",", ":"))
        + ";"
    )
    widget = (
        _COMPONENT_ROOT / "assets" / "journal-sampling-review-widget.html"
    ).read_text(encoding="utf-8")
    if widget.count(embedded_adapter) != 1:
        raise ValueError(
            "Journal Sampling widget does not embed the exact review adapter."
        )


def _implementation_receipts() -> list[dict[str, Any]]:
    """Return the static ordered transitive implementation receipt set."""

    assurance_root = _assurance_package_root()
    declared_contract = (
        *(
            ("plugin", relative_path)
            for relative_path, _ in IMPLEMENTATION_PLUGIN_FILES
        ),
        *(
            ("shared_assurance", relative_path)
            for relative_path, _ in ASSURANCE_IMPLEMENTATION_FILES
        ),
    )
    if declared_contract != IMPLEMENTATION_CONTRACT:
        raise ValueError(
            "Journal Sampling receipt and execution-boundary contracts diverged."
        )
    validate_implementation_tree(
        str(_COMPONENT_ROOT),
        shared_assurance_root=str(assurance_root),
    )

    receipts: list[dict[str, Any]] = []
    for relative_path, artifact_id in IMPLEMENTATION_PLUGIN_FILES:
        path = _COMPONENT_ROOT / relative_path
        _require_ordinary_single_link(path, label=f"Implementation {relative_path}")
        receipts.append(
            artifact_receipt(
                _COMPONENT_ROOT,
                path,
                artifact_id=artifact_id,
                root_id="implementation",
                role="implementation",
                media_type=_implementation_media_type(path),
            )
        )
    for relative_path, artifact_id in ASSURANCE_IMPLEMENTATION_FILES:
        path = assurance_root / relative_path
        _require_ordinary_single_link(
            path,
            label=f"Assurance implementation {relative_path}",
        )
        receipts.append(
            artifact_receipt(
                assurance_root,
                path,
                artifact_id=artifact_id,
                root_id="assurance_implementation",
                role="implementation",
                media_type=_implementation_media_type(path),
            )
        )
    _validate_implementation_configuration()
    return receipts


def _validate_exact_implementation_receipts(
    envelope: dict[str, Any],
) -> list[dict[str, Any]]:
    expected = _implementation_receipts()
    actual = [
        receipt
        for receipt in envelope.get("artifact_receipts", [])
        if isinstance(receipt, dict) and receipt.get("role") == "implementation"
    ]
    expected_refs = [receipt["artifact_id"] for receipt in expected]
    if (
        actual != expected
        or envelope.get("implementation_artifact_refs") != expected_refs
    ):
        raise ValueError("Transitive implementation receipts are not exact or ordered.")
    return expected


def supported_files(input_path: Path) -> list[Path]:
    """Return supported journal files from a file or folder path."""

    path = input_path.expanduser()
    if path.is_file():
        return [path] if path.suffix.lower() in SUPPORTED_SUFFIXES else []
    if not path.exists():
        raise FileNotFoundError(f"Input path does not exist: {path}")
    files = [
        candidate
        for candidate in sorted(path.rglob("*"))
        if candidate.is_file()
        and candidate.suffix.lower() in SUPPORTED_SUFFIXES
        and not candidate.name.startswith("~$")
    ]
    return files


def _excel_column_name(index: int) -> str:
    idx = index + 1
    letters: list[str] = []
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        letters.append(chr(65 + rem))
    return "".join(reversed(letters))


def _unique_names(values: Sequence[Any]) -> list[str]:
    names: list[str] = []
    seen: dict[str, int] = {}
    for idx, value in enumerate(values):
        text = _clean_text(value)
        base = (
            text
            if text and text.lower() not in {"none", "nan"}
            else _excel_column_name(idx)
        )
        count = seen.get(base, 0)
        seen[base] = count + 1
        names.append(base if count == 0 else f"{base}_{count + 1}")
    return names


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\u00a0", " ").strip()


def _norm_label(value: Any) -> str:
    return re.sub(r"\s+", " ", _clean_text(value).lower())


def _nonempty_count(row: Sequence[Any]) -> int:
    return sum(1 for value in row if _clean_text(value))


def _excel_number_format_decimal_places(number_format: object) -> int | None:
    """Return a fixed-point scale declared by a simple Excel number format."""

    if not isinstance(number_format, str) or not number_format.strip():
        return None
    section = number_format.split(";", 1)[0]
    if "%" in section or re.search(r"[eE][+-]0", section):
        return None
    section = re.sub(r'"[^"]*"|\[[^\]]*\]|\\.|_.|\*.', "", section)
    if not any(character in section for character in "0#?"):
        return None
    if "." not in section:
        return 0
    fractional = section.split(".", 1)[1]
    placeholders = []
    for character in fractional:
        if character in "0#?":
            placeholders.append(character)
        elif placeholders:
            break
    return len(placeholders)


def _raw_excel_text(
    value: Any,
    *,
    number_format: object | None = None,
) -> str | None:
    """Represent heterogeneous spreadsheet cells without type inference loss."""

    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (int, float, Decimal)):
        if isinstance(value, float) and not math.isfinite(value):
            return str(value)
        text = str(value)
        decimal_places = _excel_number_format_decimal_places(number_format)
        if decimal_places is not None:
            decimal_value = Decimal(text)
            quantum = Decimal(1).scaleb(-decimal_places)
            try:
                if decimal_value.quantize(quantum) == decimal_value:
                    return f"{decimal_value:.{decimal_places}f}"
            except InvalidOperation:
                pass
        return text
    return str(value)


def _read_excel_raw(
    path: Path,
    source_bytes: bytes | None = None,
) -> pl.DataFrame:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        source: Path | BytesIO = path if source_bytes is None else BytesIO(source_bytes)
        workbook = openpyxl.load_workbook(source, data_only=True, read_only=True)
        sheet = workbook.worksheets[0]
        rows = [list(row) for row in sheet.iter_rows()]
        workbook.close()
        max_width = max((len(row) for row in rows), default=0)
        if max_width == 0:
            return pl.DataFrame()
        data = {
            f"column_{idx}": [
                (
                    _raw_excel_text(
                        row[idx].value,
                        number_format=row[idx].number_format,
                    )
                    if idx < len(row)
                    else None
                )
                for row in rows
            ]
            for idx in range(max_width)
        }
        return pl.DataFrame(data)

    try:
        return pl.read_excel(
            path if source_bytes is None else source_bytes,
            has_header=False,
            drop_empty_rows=False,
            drop_empty_cols=False,
        )
    except (fastexcel.CalamineError, ValueError, RuntimeError, OSError) as exc:
        LOGGER.info(
            "Polars Excel read failed for %s; trying openpyxl: %s", path.name, exc
        )

    source = path if source_bytes is None else BytesIO(source_bytes)
    workbook = openpyxl.load_workbook(source, data_only=True, read_only=False)
    sheet = workbook.worksheets[0]
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    max_width = max((len(row) for row in rows), default=0)
    if max_width == 0:
        return pl.DataFrame()
    data = {
        f"column_{idx}": [
            _raw_excel_text(row[idx]) if idx < len(row) else None for row in rows
        ]
        for idx in range(max_width)
    }
    return pl.DataFrame(data)


def _read_csv_raw(
    path: Path,
    source_bytes: bytes | None = None,
) -> pl.DataFrame:
    source: Path | BytesIO = path if source_bytes is None else BytesIO(source_bytes)
    return pl.read_csv(source, has_header=False, infer_schema=False, ignore_errors=True)


def _raw_table(
    path: Path,
    source_bytes: bytes | None = None,
) -> pl.DataFrame:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv_raw(path, source_bytes)
    if suffix in {".xls", ".xlsx", ".xlsm"}:
        return _read_excel_raw(path, source_bytes)
    raise ValueError(f"Unsupported tabular file: {path}")


def _source_sheet_names(
    path: Path,
    source_bytes: bytes | None = None,
) -> list[str]:
    """Return every worksheet name so workbook scope can be qualified."""

    if path.suffix.lower() not in {".xls", ".xlsx", ".xlsm"}:
        return []
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        source: Path | BytesIO = path if source_bytes is None else BytesIO(source_bytes)
        workbook = openpyxl.load_workbook(source, read_only=True)
        titles = [sheet.title for sheet in workbook.worksheets]
        workbook.close()
        return titles
    reader = fastexcel.read_excel(path if source_bytes is None else source_bytes)
    return [str(value) for value in reader.sheet_names]


def _source_sheet_name(
    path: Path,
    source_bytes: bytes | None = None,
) -> str | None:
    """Return the exact worksheet used by the bounded single-sheet adapters."""

    sheet_names = _source_sheet_names(path, source_bytes)
    return sheet_names[0] if sheet_names else None


def _row_values(df: pl.DataFrame, idx: int) -> list[Any]:
    return list(df.row(idx))


def _suggest_header_rows(df: pl.DataFrame) -> list[int]:
    if df.is_empty():
        return [1]
    best_idx = 0
    best_score = -1
    limit = min(df.height, 20)
    header_tokens = (
        "data",
        "date",
        "datum",
        "conto",
        "compte",
        "konto",
        "account",
        "descrizione",
        "description",
        "beschreibung",
        "libelle",
        "libellé",
        "dare",
        "soll",
        "avere",
        "haben",
        "debit",
        "débit",
        "credit",
        "crédit",
        "amount",
        "montant",
        "betrag",
        "importo",
    )
    for idx in range(limit):
        row = _row_values(df, idx)
        score = 0
        for value in row:
            text = _norm_label(value)
            if not text:
                continue
            score += 1
            if any(token in text for token in header_tokens):
                score += 3
            elif any(ch.isalpha() for ch in text):
                score += 1
        if score > best_score:
            best_idx = idx
            best_score = score
    if best_idx > 0:
        previous = _row_values(df, best_idx - 1)
        current = _row_values(df, best_idx)
        fillable = 0
        for top, base in zip(previous, current):
            if _clean_text(top) and not _clean_text(base):
                fillable += 1
        if fillable >= 2:
            return [best_idx, best_idx + 1]
    return [best_idx + 1]


def _merge_header_rows(rows: Sequence[Sequence[Any]]) -> list[str]:
    if not rows:
        return []
    width = max((len(row) for row in rows), default=0)
    labels: list[str] = []
    for idx in range(width):
        parts = []
        for row in rows:
            value = _clean_text(row[idx] if idx < len(row) else "")
            if value and value.lower() not in {"none", "nan"}:
                parts.append(value)
        labels.append(" ".join(parts))
    return _unique_names(labels)


def _apply_header(df: pl.DataFrame, rows_1_indexed: Sequence[int]) -> pl.DataFrame:
    if df.is_empty():
        return df
    row_indexes = sorted({int(row) - 1 for row in rows_1_indexed})
    if not row_indexes or min(row_indexes) < 0:
        raise ValueError("Header rows must be 1-indexed positive integers.")
    if max(row_indexes) >= df.height:
        raise ValueError("Header row exceeds available rows.")
    labels = _merge_header_rows([_row_values(df, idx) for idx in row_indexes])
    body = df.slice(max(row_indexes) + 1)
    if body.width != len(labels):
        labels = _unique_names(labels[: body.width])
    body.columns = labels
    # Preserve mapped columns even when this particular extract has no values
    # on one side (for example, a debit-only holdout population). Keep empty
    # physical rows too: removing them would make source_row cease to be an
    # exact workbook locator.
    return body


def _parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        serial = float(value)
        if 20000 <= serial <= 60000:
            return date(1899, 12, 30) + timedelta(days=int(serial))
        return None
    text = _clean_text(value)
    if not text:
        return None
    iso_datetime_match = re.match(r"^(\d{4}-\d{2}-\d{2})[T\s]", text)
    if iso_datetime_match:
        try:
            return datetime.strptime(iso_datetime_match.group(1), "%Y-%m-%d").date()
        except ValueError:
            return None
    token_match = DATE_TOKEN_RE.search(text)
    token = token_match.group(1) if token_match else text
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(token, fmt).date()
        except ValueError:
            continue
    return None


def _parse_number(
    value: Any,
    *,
    decimal_separator: str | None = None,
    thousands_separator: str | None = None,
) -> Decimal | None:
    """Parse one exact amount.

    Exact arithmetic and canonical text are mechanically verifiable. Spreadsheet
    readers may surface OOXML numeric cells as Python floats, so this bounded
    adapter explicitly converts their shortest decimal rendering rather than
    performing binary-float arithmetic.
    """

    if value is None:
        return None
    text = _clean_text(value)
    if not text:
        return None
    try:
        return parse_localized_decimal(
            value,
            decimal_separator=decimal_separator,
            thousands_separator=thousands_separator,
            allow_float=True,
        )
    except MoneyValidationError:
        return None


def _reported_increment(
    value: Any,
    *,
    decimal_separator: str | None = None,
) -> str:
    """Return the exact increment implied by the source representation."""

    text = _clean_text(value)
    if not text:
        return "1"
    compact = re.sub(r"\s+", "", text)
    separator = decimal_separator
    if separator is None:
        if "," in compact and "." in compact:
            separator = "," if compact.rfind(",") > compact.rfind(".") else "."
        elif "," in compact:
            separator = ","
        elif "." in compact:
            separator = "."
    decimal_places = 0
    if separator and separator in compact:
        fractional = re.sub(r"\D", "", compact.rsplit(separator, 1)[1])
        decimal_places = len(fractional)
    increment = Decimal(1).scaleb(-decimal_places)
    return decimal_text(increment)


def _date_ratio(series: pl.Series) -> float:
    values = [_parse_date(value) for value in series.drop_nulls().head(100).to_list()]
    if not values:
        return 0.0
    return sum(value is not None for value in values) / len(values)


def _account_ratio(series: pl.Series) -> float:
    values = [_clean_text(value) for value in series.drop_nulls().head(100).to_list()]
    values = [value for value in values if value]
    if not values:
        return 0.0
    matches = sum(bool(ACCOUNT_RE.fullmatch(value)) for value in values)
    return matches / len(values)


def _amount_ratio(series: pl.Series) -> float:
    values = series.drop_nulls().head(100).to_list()
    if not values:
        return 0.0
    matches = sum(_parse_number(value) is not None for value in values)
    return matches / len(values)


def infer_mapping(df: pl.DataFrame) -> dict[str, str | None]:
    """Infer canonical field mapping from headers and simple column profiles."""

    mapping: dict[str, str | None] = {
        "date": None,
        "movement_number": None,
        "line_number": None,
        "account": None,
        "account_desc": None,
        "line_desc": None,
        "debit": None,
        "credit": None,
        "amount": None,
    }
    for col in df.columns:
        label = _norm_label(col)
        if mapping["date"] is None and any(
            token in label for token in ("data", "date", "datum")
        ):
            mapping["date"] = col
        if mapping["movement_number"] is None and any(
            token in label
            for token in (
                "nr. reg",
                "n. reg",
                "numero registrazione",
                "movimento",
                "movement",
                "mouvement",
                "bewegung",
            )
        ):
            mapping["movement_number"] = col
        if (
            mapping["line_number"] is None
            and col != mapping["movement_number"]
            and any(
                token in label
                for token in (
                    "riga",
                    "ligne",
                    "zeile",
                )
            )
        ):
            mapping["line_number"] = col
        if mapping["account"] is None and (
            label in {"conto", "account", "compte", "konto"}
            or (
                any(token in label for token in ("conto", "account", "compte", "konto"))
                and "desc" not in label
                and "descr" not in label
                and "beschreibung" not in label
            )
        ):
            mapping["account"] = col
        if mapping["account_desc"] is None and (
            "descrizione conto" in label
            or "account desc" in label
            or "description compte" in label
            or "kontobeschreibung" in label
            or "konto beschreibung" in label
        ):
            mapping["account_desc"] = col
        if mapping["line_desc"] is None and (
            "descrizione oper" in label
            or "description" in label
            or "libelle" in label
            or "libellé" in label
            or "beschreibung" in label
            or "causale" in label
            or label == "descrizione"
        ):
            mapping["line_desc"] = col
        if mapping["debit"] is None and any(
            token in label for token in ("dare", "debit", "débit", "addebit", "soll")
        ):
            mapping["debit"] = col
        if mapping["credit"] is None and any(
            token in label
            for token in ("avere", "credit", "crédit", "accredit", "haben")
        ):
            mapping["credit"] = col
        if mapping["amount"] is None and any(
            token in label
            for token in ("amount", "importo", "saldo", "montant", "betrag")
        ):
            mapping["amount"] = col

    if mapping["debit"] is not None and mapping["credit"] is not None:
        # Explicit source-owned debit and credit columns take precedence over
        # other fields whose labels merely contain "amount" or "importo".
        mapping["amount"] = None

    if mapping["date"] is None:
        candidates = [(col, _date_ratio(df.get_column(col))) for col in df.columns]
        mapping["date"] = (
            max(candidates, key=lambda item: item[1])[0]
            if candidates and max(score for _, score in candidates) >= 0.5
            else None
        )
    if mapping["account"] is None:
        candidates = [(col, _account_ratio(df.get_column(col))) for col in df.columns]
        mapping["account"] = (
            max(candidates, key=lambda item: item[1])[0]
            if candidates and max(score for _, score in candidates) >= 0.4
            else None
        )
    if (
        mapping["debit"] is None
        and mapping["credit"] is None
        and mapping["amount"] is None
    ):
        candidates = [(col, _amount_ratio(df.get_column(col))) for col in df.columns]
        amount_cols = [col for col, score in candidates if score >= 0.3]
        if len(amount_cols) >= 2:
            mapping["debit"] = amount_cols[-2]
            mapping["credit"] = amount_cols[-1]
        elif amount_cols:
            mapping["amount"] = amount_cols[-1]
    return mapping


def _recipe_for_file(recipe: dict[str, Any], path: Path) -> dict[str, Any]:
    files = recipe.get("files")
    if isinstance(files, dict):
        item = files.get(path.name) or files.get(path.as_posix())
        if isinstance(item, dict):
            merged = {key: value for key, value in recipe.items() if key != "files"}
            merged.update(item)
            return merged
    return recipe


def _field(mapping: dict[str, Any], name: str) -> str | None:
    value = mapping.get(name)
    return str(value) if value else None


def _identifier_fragment(value: str) -> str:
    text = re.sub(r"[^A-Za-z0-9._-]+", "-", value).strip("-._")
    return text or "source"


def _source_artifact_ref(path: Path) -> str:
    return f"source.{_identifier_fragment(path.name)}"


def _control(
    control_id: str,
    status: str,
    detail: str,
    *,
    required: bool = True,
    evidence_refs: Sequence[str] = (),
) -> dict[str, Any]:
    return {
        "control_id": control_id,
        "required": required,
        "status": status,
        "evidence_refs": list(evidence_refs),
        "detail": detail,
    }


def _mapping_contract(
    *,
    parser: str,
    source_family: str,
    header_rows: Sequence[int] = (),
    mapping: dict[str, Any] | None = None,
    layout: dict[str, Any] | None = None,
    excluded_monetary_columns: Sequence[str] = (),
    posting_identity: str = "source_row",
    carry_forward_fields: Sequence[str] = (),
    currency: str = "EUR",
    unit: str = "currency",
    decimal_separator: str | None = None,
    thousands_separator: str | None = None,
    amount_sign_convention: str | None = None,
) -> dict[str, Any]:
    return {
        "parser": parser,
        "source_family": source_family,
        "header_rows": list(header_rows),
        "mapping": dict(mapping or {}),
        "layout": dict(layout or {}),
        "excluded_monetary_columns": list(excluded_monetary_columns),
        "posting_identity": posting_identity,
        "carry_forward_fields": list(carry_forward_fields),
        "currency": currency,
        "unit": unit,
        "decimal_separator": decimal_separator,
        "thousands_separator": thousands_separator,
        "amount_sign_convention": amount_sign_convention,
    }


def _reviewed_mapping(
    file_recipe: dict[str, Any],
    contract: dict[str, Any],
    *,
    path: Path,
    adapter_id: str,
) -> tuple[bool, str | None, dict[str, Any] | None]:
    qualification = file_recipe.get("qualification")
    if not isinstance(qualification, dict):
        return False, None, None
    receipt = qualification.get("decision_receipt")
    if qualification.get("status") != "reviewed" or not isinstance(receipt, dict):
        return False, None, None
    try:
        validated = validate_reviewed_decision_receipt(
            receipt,
            expected_source_artifact_refs=[_source_artifact_ref(path)],
            expected_adapter_id=adapter_id,
            expected_adapter_version=ADAPTER_VERSION,
            require_reviewed=True,
        )
    except DecisionReceiptError:
        return False, None, None
    if validated["decision_type"] != "source_mapping":
        return False, None, None
    if validated["content"] != contract:
        return False, None, None
    if qualification.get("mapping_sha256") != canonical_json_sha256(contract):
        return False, None, None
    return True, str(validated["decision_id"]), validated


def _tabular_candidate_count(
    table: pl.DataFrame,
    mapping: dict[str, Any],
) -> int:
    money_cols = [
        column
        for column in (
            _field(mapping, "debit"),
            _field(mapping, "credit"),
            _field(mapping, "amount"),
        )
        if column
    ]
    count = 0
    for row in table.iter_rows(named=True):
        has_money = any(_clean_text(row.get(column)) for column in money_cols)
        if has_money:
            count += 1
    return count


def _potential_monetary_columns(
    table: pl.DataFrame,
    mapping: dict[str, Any],
    *,
    decimal_separator: str | None,
    thousands_separator: str | None,
) -> list[str]:
    """Return every non-semantic column that may contain monetary values.

    Headers are useful evidence but cannot establish completeness by
    themselves. An otherwise-unmapped column with parseable numeric cells must
    also be mapped or explicitly excluded in the reviewed contract.
    """

    monetary_tokens = (
        "amount",
        "importo",
        "dare",
        "avere",
        "debit",
        "débit",
        "credit",
        "crédit",
        "soll",
        "haben",
        "montant",
        "betrag",
        "saldo",
        "balance",
        "total",
    )
    mapped_non_monetary = {
        column
        for field_name, column in mapping.items()
        if field_name not in {"debit", "credit", "amount"} and column
    }
    candidates: list[str] = []
    for column in table.columns:
        if column in mapped_non_monetary:
            continue
        explicit_header = any(token in _norm_label(column) for token in monetary_tokens)
        numeric_values = any(
            _parse_number(
                value,
                decimal_separator=decimal_separator,
                thousands_separator=thousands_separator,
            )
            is not None
            for value in table.get_column(column).drop_nulls().to_list()
            if _clean_text(value)
        )
        if explicit_header or numeric_values:
            candidates.append(column)
    return candidates


def _excluded_monetary_columns(
    file_recipe: dict[str, Any],
) -> tuple[list[str], bool]:
    raw = file_recipe.get("excluded_monetary_columns", [])
    if not isinstance(raw, list) or any(
        not isinstance(value, str) or not value.strip() for value in raw
    ):
        return [], False
    normalized = [value.strip() for value in raw]
    return normalized, len(normalized) == len(set(normalized))


def _posting_identity(
    file_recipe: dict[str, Any],
    mapping: dict[str, Any],
) -> tuple[str, bool]:
    proposed = (
        "movement_number_and_line_number"
        if _field(mapping, "movement_number") and _field(mapping, "line_number")
        else (
            "movement_number_and_source_row"
            if _field(mapping, "movement_number")
            else "source_row"
        )
    )
    value = _clean_text(file_recipe.get("posting_identity")) or proposed
    valid = value in {
        "source_row",
        "movement_number_and_source_row",
        "movement_number_and_line_number",
    }
    if value.startswith("movement_number") and not _field(mapping, "movement_number"):
        valid = False
    if value == "movement_number_and_line_number" and not _field(
        mapping, "line_number"
    ):
        valid = False
    return value, valid


def _carry_forward_fields(
    file_recipe: dict[str, Any],
    mapping: dict[str, Any],
    table: pl.DataFrame,
) -> tuple[list[str], bool]:
    if "carry_forward_fields" not in file_recipe:
        proposed: list[str] = []
        money_columns = [
            column
            for column in (
                _field(mapping, "debit"),
                _field(mapping, "credit"),
                _field(mapping, "amount"),
            )
            if column
        ]
        seen_date = False
        seen_movement = False
        for row in table.iter_rows(named=True):
            if not any(_clean_text(row.get(column)) for column in money_columns):
                continue
            date_value = row.get(_field(mapping, "date") or "")
            movement_value = row.get(_field(mapping, "movement_number") or "")
            if _parse_date(date_value) is not None:
                seen_date = True
            elif seen_date and "date" not in proposed:
                proposed.append("date")
            if _clean_text(movement_value):
                seen_movement = True
            elif (
                _field(mapping, "movement_number")
                and seen_movement
                and "movement_number" not in proposed
            ):
                proposed.append("movement_number")
        raw: object = proposed
    else:
        raw = file_recipe.get("carry_forward_fields")
    if not isinstance(raw, list) or any(
        not isinstance(value, str) or not value.strip() for value in raw
    ):
        return [], False
    fields = [value.strip() for value in raw]
    valid = (
        len(fields) == len(set(fields))
        and set(fields) <= {"date", "movement_number"}
        and all(_field(mapping, field) for field in fields)
    )
    return fields, valid


def _tabular_plan(
    raw_df: pl.DataFrame,
    file_recipe: dict[str, Any],
    path: Path,
) -> tuple[QualificationPlan, pl.DataFrame]:
    header_rows_raw = file_recipe.get("header_rows")
    header_rows = (
        [int(value) for value in header_rows_raw]
        if isinstance(header_rows_raw, list)
        and header_rows_raw
        and all(
            isinstance(value, int) and not isinstance(value, bool)
            for value in header_rows_raw
        )
        else _suggest_header_rows(raw_df)
    )
    try:
        table = _apply_header(raw_df, header_rows)
    except ValueError:
        table = pl.DataFrame()
    explicit_mapping = file_recipe.get("mapping")
    mapping = (
        dict(explicit_mapping)
        if isinstance(explicit_mapping, dict)
        else infer_mapping(table)
    )
    decimal_separator = file_recipe.get("decimal_separator")
    thousands_separator = file_recipe.get("thousands_separator")
    amount_sign_convention = file_recipe.get("amount_sign_convention")
    excluded_monetary_columns, exclusions_valid = _excluded_monetary_columns(
        file_recipe
    )
    posting_identity, posting_identity_valid = _posting_identity(file_recipe, mapping)
    carry_forward_fields, carry_forward_valid = _carry_forward_fields(
        file_recipe, mapping, table
    )
    currency = (_clean_text(file_recipe.get("currency")) or "EUR").upper()
    unit = _clean_text(file_recipe.get("unit")) or "currency"
    unit_contract_valid = (
        bool(re.fullmatch(r"[A-Z]{3}", currency)) and unit == "currency"
    )
    mapped_columns = [str(value) for value in mapping.values() if value]
    columns_exist = bool(table.width) and all(
        column in table.columns for column in mapped_columns
    )
    unique_columns = len(mapped_columns) == len(set(mapped_columns))
    date_mapped = bool(_field(mapping, "date"))
    account_mapped = bool(_field(mapping, "account"))
    has_amount = bool(_field(mapping, "amount"))
    has_debit_credit = bool(_field(mapping, "debit") and _field(mapping, "credit"))
    potential_monetary_columns = _potential_monetary_columns(
        table,
        mapping,
        decimal_separator=(
            str(decimal_separator) if decimal_separator is not None else None
        ),
        thousands_separator=(
            str(thousands_separator) if thousands_separator is not None else None
        ),
    )
    mapped_monetary_columns = {
        column
        for column in (
            _field(mapping, "debit"),
            _field(mapping, "credit"),
            _field(mapping, "amount"),
        )
        if column
    }
    excluded_set = set(excluded_monetary_columns)
    exclusions_exist = excluded_set <= set(potential_monetary_columns)
    unresolved_monetary_columns = sorted(
        set(potential_monetary_columns) - mapped_monetary_columns - excluded_set
    )
    if has_amount and amount_sign_convention is None:
        amount_sign_convention = "debit_positive"
    contract = _mapping_contract(
        parser="tabular",
        source_family=TABULAR_SOURCE_FAMILY,
        header_rows=header_rows,
        mapping=mapping,
        excluded_monetary_columns=excluded_monetary_columns,
        posting_identity=posting_identity,
        carry_forward_fields=carry_forward_fields,
        currency=currency,
        unit=unit,
        decimal_separator=decimal_separator,
        thousands_separator=thousands_separator,
        amount_sign_convention=amount_sign_convention,
    )
    reviewed, decision_ref, reviewed_decision = _reviewed_mapping(
        file_recipe,
        contract,
        path=path,
        adapter_id=TABULAR_ADAPTER_ID,
    )
    monetary_shape = has_amount != has_debit_credit
    sign_convention_ok = (
        amount_sign_convention in {"debit_positive", "credit_positive"}
        if has_amount
        else amount_sign_convention in {None, ""}
    )
    source_family_explicit = (
        file_recipe.get("parser") == "tabular"
        and file_recipe.get("source_family") == TABULAR_SOURCE_FAMILY
    )
    candidate_row_count = _tabular_candidate_count(table, mapping)
    structure_ok = (
        bool(table.width)
        and columns_exist
        and unique_columns
        and date_mapped
        and account_mapped
        and monetary_shape
        and sign_convention_ok
        and candidate_row_count > 0
    )
    controls = [
        _control(
            "bounded_adapter",
            "passed" if source_family_explicit else "not_assessed",
            (
                "Explicit tabular source-family adapter selected."
                if source_family_explicit
                else "Select the tabular source family explicitly in the reviewed recipe."
            ),
        ),
        _control(
            "schema_mapping",
            "passed" if structure_ok else "failed",
            (
                "Mapped date, account, and one unambiguous monetary representation exist."
                if structure_ok
                else "Required mapped columns are missing, duplicated, or monetarily ambiguous."
            ),
        ),
        _control(
            "monetary_field_disposition",
            (
                "failed"
                if not exclusions_valid or not exclusions_exist
                else ("not_assessed" if unresolved_monetary_columns else "passed")
            ),
            (
                "Every monetary-labelled or otherwise numeric source column is mapped or explicitly excluded."
                if (
                    exclusions_valid
                    and exclusions_exist
                    and not unresolved_monetary_columns
                )
                else (
                    "Excluded monetary columns are malformed or do not exist."
                    if not exclusions_valid or not exclusions_exist
                    else "Review and map or explicitly exclude every unresolved monetary column."
                )
            ),
        ),
        _control(
            "posting_identity_and_carry_forward",
            (
                "passed"
                if posting_identity_valid
                and carry_forward_valid
                and unit_contract_valid
                else "failed"
            ),
            (
                "Posting identity, carry-forward policy, currency, and unit are explicit."
                if posting_identity_valid
                and carry_forward_valid
                and unit_contract_valid
                else "Posting identity, carry-forward policy, currency, or unit is invalid."
            ),
        ),
        _control(
            "reviewed_mapping",
            "passed" if reviewed else "not_assessed",
            (
                "Review binding matches the exact mapping contract."
                if reviewed
                else "The exact mapping contract has not been reviewed or its digest is stale."
            ),
        ),
    ]
    failed = any(item["status"] == "failed" for item in controls)
    unassessed = any(item["status"] == "not_assessed" for item in controls)
    status = (
        "unsupported_source_layout"
        if failed
        else "needs_review" if unassessed else "qualified"
    )
    limitations = []
    if status != "qualified":
        limitations.append("No rows from this file may enter the sampling population.")
    if unresolved_monetary_columns:
        limitations.append(
            "Potential monetary columns remain unmapped and not reviewed for exclusion."
        )
    return (
        QualificationPlan(
            parser="tabular",
            adapter_id=TABULAR_ADAPTER_ID,
            source_family=TABULAR_SOURCE_FAMILY,
            status=status,
            candidate_row_count=candidate_row_count,
            controls=controls,
            limitations=limitations,
            reviewed_mapping_ref=decision_ref,
            reviewed_decision=reviewed_decision,
            header_rows=header_rows,
            mapping=mapping,
            layout={},
            excluded_monetary_columns=excluded_monetary_columns,
            unresolved_monetary_columns=unresolved_monetary_columns,
            posting_identity=posting_identity,
            carry_forward_fields=carry_forward_fields,
            currency=currency,
            unit=unit,
            decimal_separator=(
                str(decimal_separator) if decimal_separator is not None else None
            ),
            thousands_separator=(
                str(thousands_separator) if thousands_separator is not None else None
            ),
            amount_sign_convention=(
                str(amount_sign_convention)
                if amount_sign_convention is not None
                else None
            ),
        ),
        table,
    )


def _suggest_print_layout(raw_df: pl.DataFrame) -> dict[str, Any]:
    header_idx = _find_header_index(raw_df, ("dare", "avere"))
    if header_idx is None:
        return {}
    header = _row_values(raw_df, header_idx)
    date_col = _find_header_col(header, "data reg", "data", "date", "datum")
    movement_col = _find_header_col(
        header,
        "nr. reg",
        "n. reg",
        "numero registrazione",
        "movement",
        "mouvement",
        "bewegung",
    )
    line_desc_col = _find_header_col(
        header, "descrizione", "description", "libelle", "libellé", "beschreibung"
    )
    account_col = _find_header_col(header, "conto", "account", "compte", "konto")
    debit_col = _find_header_col(header, "dare", "debit", "débit", "soll")
    credit_col = _find_header_col(header, "avere", "credit", "crédit", "haben")
    if None in {date_col, account_col, debit_col, credit_col}:
        return {}
    columns = {
        "date": int(date_col) + 1,
        "movement_number": int(movement_col) + 1 if movement_col is not None else None,
        "line_desc": int(line_desc_col) + 1 if line_desc_col is not None else None,
        "account": int(account_col) + 1,
        "account_desc": int(account_col) + 2,
        # This source family has merged debit/credit headers with values in the
        # immediately following physical column. The explicit reviewed layout
        # prevents the old three-cell search from promoting unrelated numbers.
        "debit": int(debit_col) + 2,
        "credit": int(credit_col) + 2,
    }
    return {"header_row": header_idx + 1, "columns": columns}


def _print_candidate_count(raw_df: pl.DataFrame, layout: dict[str, Any]) -> int:
    header_row = layout.get("header_row")
    if not isinstance(header_row, int) or isinstance(header_row, bool):
        return 0
    count = 0
    for row_idx in range(header_row, raw_df.height):
        row = _row_values(raw_df, row_idx)
        if any(
            _clean_text(_layout_cell(row, layout, field))
            for field in ("debit", "credit")
        ):
            count += 1
    return count


def _print_layout_column_name(
    raw_df: pl.DataFrame,
    columns: dict[str, Any],
    field: str,
) -> str | None:
    value = columns.get(field)
    if (
        not isinstance(value, int)
        or isinstance(value, bool)
        or value <= 0
        or value > raw_df.width
    ):
        return None
    return raw_df.columns[value - 1]


def _potential_print_monetary_columns(
    raw_df: pl.DataFrame,
    layout: dict[str, Any],
    *,
    decimal_separator: str | None,
    thousands_separator: str | None,
) -> list[str]:
    """Return physical print-layout columns requiring monetary disposition."""

    columns = layout.get("columns") if isinstance(layout.get("columns"), dict) else {}
    mapped_non_monetary = {
        column
        for field in (
            "date",
            "movement_number",
            "line_number",
            "line_desc",
            "account",
            "account_desc",
        )
        if (column := _print_layout_column_name(raw_df, columns, field)) is not None
    }
    header_row = layout.get("header_row")
    candidates: list[str] = []
    for column in raw_df.columns:
        if column in mapped_non_monetary:
            continue
        values = (
            raw_df.get_column(column).slice(header_row).drop_nulls().to_list()
            if isinstance(header_row, int)
            and not isinstance(header_row, bool)
            and header_row >= 0
            else raw_df.get_column(column).drop_nulls().to_list()
        )
        numeric_values = any(
            _parse_number(
                value,
                decimal_separator=decimal_separator,
                thousands_separator=thousands_separator,
            )
            is not None
            for value in values
            if _clean_text(value)
        )
        if numeric_values:
            candidates.append(column)
    return candidates


def _print_plan(
    raw_df: pl.DataFrame,
    file_recipe: dict[str, Any],
    path: Path,
) -> QualificationPlan:
    recipe_layout = file_recipe.get("layout")
    layout = (
        dict(recipe_layout)
        if isinstance(recipe_layout, dict)
        else _suggest_print_layout(raw_df)
    )
    decimal_separator = file_recipe.get("decimal_separator")
    thousands_separator = file_recipe.get("thousands_separator")
    columns = layout.get("columns") if isinstance(layout.get("columns"), dict) else {}
    excluded_monetary_columns, exclusions_valid = _excluded_monetary_columns(
        file_recipe
    )
    potential_monetary_columns = _potential_print_monetary_columns(
        raw_df,
        layout,
        decimal_separator=(
            str(decimal_separator) if decimal_separator is not None else None
        ),
        thousands_separator=(
            str(thousands_separator) if thousands_separator is not None else None
        ),
    )
    mapped_monetary_columns = {
        column
        for field in ("debit", "credit")
        if (column := _print_layout_column_name(raw_df, columns, field)) is not None
    }
    excluded_set = set(excluded_monetary_columns)
    exclusions_exist = excluded_set <= set(potential_monetary_columns)
    unresolved_monetary_columns = sorted(
        set(potential_monetary_columns) - mapped_monetary_columns - excluded_set
    )
    proposed_identity = (
        "movement_number_and_source_row"
        if columns.get("movement_number")
        else "source_row"
    )
    posting_identity = (
        _clean_text(file_recipe.get("posting_identity")) or proposed_identity
    )
    carry_forward_fields_raw = file_recipe.get(
        "carry_forward_fields",
        [
            "date",
            *(["movement_number"] if columns.get("movement_number") else []),
            *(["line_desc"] if columns.get("line_desc") else []),
        ],
    )
    carry_forward_fields = (
        [str(value).strip() for value in carry_forward_fields_raw]
        if isinstance(carry_forward_fields_raw, list)
        else []
    )
    currency = (_clean_text(file_recipe.get("currency")) or "EUR").upper()
    unit = _clean_text(file_recipe.get("unit")) or "currency"
    contract = _mapping_contract(
        parser="print_friendly_excel",
        source_family=PRINT_SOURCE_FAMILY,
        layout=layout,
        excluded_monetary_columns=excluded_monetary_columns,
        posting_identity=posting_identity,
        carry_forward_fields=carry_forward_fields,
        currency=currency,
        unit=unit,
        decimal_separator=decimal_separator,
        thousands_separator=thousands_separator,
    )
    reviewed, decision_ref, reviewed_decision = _reviewed_mapping(
        file_recipe,
        contract,
        path=path,
        adapter_id=PRINT_ADAPTER_ID,
    )
    required_fields = {"date", "account", "debit", "credit"}
    layout_ok = (
        isinstance(layout.get("header_row"), int)
        and required_fields <= set(columns)
        and all(
            isinstance(columns[field], int)
            and not isinstance(columns[field], bool)
            and columns[field] > 0
            for field in required_fields
        )
    )
    source_family_explicit = (
        file_recipe.get("parser") == "print_friendly_excel"
        and file_recipe.get("source_family") == PRINT_SOURCE_FAMILY
    )
    posting_contract_ok = (
        posting_identity in {"source_row", "movement_number_and_source_row"}
        and (posting_identity == "source_row" or bool(columns.get("movement_number")))
        and len(carry_forward_fields) == len(set(carry_forward_fields))
        and set(carry_forward_fields) <= {"date", "movement_number", "line_desc"}
        and (
            "movement_number" not in carry_forward_fields
            or bool(columns.get("movement_number"))
        )
        and ("line_desc" not in carry_forward_fields or bool(columns.get("line_desc")))
        and bool(re.fullmatch(r"[A-Z]{3}", currency))
        and unit == "currency"
    )
    controls = [
        _control(
            "bounded_adapter",
            "passed" if source_family_explicit else "not_assessed",
            (
                "Explicit print-friendly source-family adapter selected."
                if source_family_explicit
                else "Select the bounded print-friendly source family explicitly."
            ),
        ),
        _control(
            "layout_contract",
            "passed" if layout_ok else "failed",
            (
                "Header row and exact field columns satisfy the bounded adapter."
                if layout_ok
                else "The bounded print-friendly layout fields are incomplete."
            ),
        ),
        _control(
            "monetary_field_disposition",
            (
                "failed"
                if not exclusions_valid or not exclusions_exist
                else ("not_assessed" if unresolved_monetary_columns else "passed")
            ),
            (
                "Every monetary-labelled or otherwise numeric physical column is mapped or explicitly excluded."
                if (
                    exclusions_valid
                    and exclusions_exist
                    and not unresolved_monetary_columns
                )
                else (
                    "Excluded monetary columns are malformed or do not exist."
                    if not exclusions_valid or not exclusions_exist
                    else "Review and map or explicitly exclude every unresolved physical numeric column."
                )
            ),
        ),
        _control(
            "posting_identity_and_carry_forward",
            "passed" if posting_contract_ok else "failed",
            (
                "Posting identity, carry-forward policy, currency, and unit are explicit."
                if posting_contract_ok
                else "Posting identity, carry-forward policy, currency, or unit is invalid."
            ),
        ),
        _control(
            "reviewed_mapping",
            "passed" if reviewed else "not_assessed",
            (
                "Review binding matches the exact print layout."
                if reviewed
                else "The exact print layout has not been reviewed or its digest is stale."
            ),
        ),
    ]
    failed = any(item["status"] == "failed" for item in controls)
    unassessed = any(item["status"] == "not_assessed" for item in controls)
    status = (
        "unsupported_source_layout"
        if failed
        else "needs_review" if unassessed else "qualified"
    )
    limitations = (
        ["No rows from this file may enter the sampling population."]
        if status != "qualified"
        else []
    )
    if unresolved_monetary_columns:
        limitations.append(
            "Potential monetary columns remain unmapped and not reviewed for exclusion."
        )
    return QualificationPlan(
        parser="print_friendly_excel",
        adapter_id=PRINT_ADAPTER_ID,
        source_family=PRINT_SOURCE_FAMILY,
        status=status,
        candidate_row_count=_print_candidate_count(raw_df, layout),
        controls=controls,
        limitations=limitations,
        reviewed_mapping_ref=decision_ref,
        reviewed_decision=reviewed_decision,
        header_rows=[],
        mapping={},
        layout=layout,
        excluded_monetary_columns=excluded_monetary_columns,
        unresolved_monetary_columns=unresolved_monetary_columns,
        posting_identity=posting_identity,
        carry_forward_fields=carry_forward_fields,
        currency=currency,
        unit=unit,
        decimal_separator=(
            str(decimal_separator) if decimal_separator is not None else None
        ),
        thousands_separator=(
            str(thousands_separator) if thousands_separator is not None else None
        ),
        amount_sign_convention=None,
    )


def _pdf_plan() -> QualificationPlan:
    return QualificationPlan(
        parser="text_pdf",
        adapter_id=DISABLED_PDF_ADAPTER_ID,
        source_family="text_pdf.unqualified.v1",
        status="unsupported_source_layout",
        candidate_row_count=0,
        controls=[
            _control(
                "bounded_adapter",
                "failed",
                "No tested source-family PDF adapter establishes amount side and row boundaries.",
            )
        ],
        limitations=[
            "Generic text-PDF amount/side reconstruction is disabled.",
            "Implement and test a source-family-specific adapter before emitting rows.",
        ],
        reviewed_mapping_ref=None,
        reviewed_decision=None,
        header_rows=[],
        mapping={},
        layout={},
        excluded_monetary_columns=[],
        unresolved_monetary_columns=[],
        posting_identity="source_row",
        carry_forward_fields=[],
        currency="EUR",
        unit="currency",
        decimal_separator=None,
        thousands_separator=None,
        amount_sign_convention=None,
    )


def _qualification_payload(
    path: Path,
    plan: QualificationPlan,
    *,
    emitted_row_count: int,
    controls: Sequence[dict[str, Any]] | None = None,
    status: str | None = None,
    limitations: Sequence[str] | None = None,
) -> dict[str, Any]:
    source_id = _identifier_fragment(path.name)
    return build_source_qualification(
        qualification_id=f"qualification.{source_id}",
        adapter_id=plan.adapter_id,
        adapter_version=ADAPTER_VERSION,
        source_family=plan.source_family,
        status=status or plan.status,
        source_artifact_refs=[f"source.{source_id}"],
        reviewed_mapping_ref=(
            plan.reviewed_mapping_ref
            if (status or plan.status) == "qualified"
            else None
        ),
        candidate_row_count=plan.candidate_row_count,
        emitted_row_count=emitted_row_count,
        controls=list(controls or plan.controls),
        limitations=list(limitations if limitations is not None else plan.limitations),
    )


def _suggested_recipe_from_plan(plan: QualificationPlan) -> dict[str, Any]:
    contract = _mapping_contract(
        parser=plan.parser,
        source_family=plan.source_family,
        header_rows=plan.header_rows,
        mapping=plan.mapping,
        layout=plan.layout,
        excluded_monetary_columns=plan.excluded_monetary_columns,
        posting_identity=plan.posting_identity,
        carry_forward_fields=plan.carry_forward_fields,
        currency=plan.currency,
        unit=plan.unit,
        decimal_separator=plan.decimal_separator,
        thousands_separator=plan.thousands_separator,
        amount_sign_convention=(
            plan.amount_sign_convention if _field(plan.mapping, "amount") else None
        ),
    )
    entry: dict[str, Any] = {
        "parser": plan.parser,
        "source_family": plan.source_family,
        "qualification": {
            "status": (
                "unsupported_source_layout"
                if plan.status == "unsupported_source_layout"
                else "needs_review"
            ),
            "decision_ref": None,
            "mapping_sha256": canonical_json_sha256(contract),
        },
    }
    if plan.header_rows:
        entry["header_rows"] = plan.header_rows
    if plan.mapping:
        entry["mapping"] = plan.mapping
        if _field(plan.mapping, "amount"):
            entry["amount_sign_convention"] = plan.amount_sign_convention
    if plan.layout:
        entry["layout"] = plan.layout
    if plan.excluded_monetary_columns:
        entry["excluded_monetary_columns"] = plan.excluded_monetary_columns
    entry["posting_identity"] = plan.posting_identity
    entry["carry_forward_fields"] = plan.carry_forward_fields
    entry["currency"] = plan.currency
    entry["unit"] = plan.unit
    if plan.decimal_separator is not None:
        entry["decimal_separator"] = plan.decimal_separator
    if plan.thousands_separator is not None:
        entry["thousands_separator"] = plan.thousands_separator
    return entry


def _recorded_recipe_from_plan(
    plan: QualificationPlan,
    qualification_status: str,
) -> dict[str, Any]:
    entry = _suggested_recipe_from_plan(plan)
    if qualification_status == "qualified":
        entry["qualification"]["status"] = "reviewed"
        entry["qualification"]["decision_ref"] = plan.reviewed_mapping_ref
        entry["qualification"]["decision_receipt"] = plan.reviewed_decision
    return entry


def _normalize_record(
    source: dict[str, Any],
    *,
    source_file: str,
    source_sheet: str | None,
    source_page: int | None,
    source_row: int | None,
) -> dict[str, Any] | None:
    decimal_separator = source.get("__decimal_separator")
    thousands_separator = source.get("__thousands_separator")
    debit = _parse_number(
        source.get("debit"),
        decimal_separator=decimal_separator,
        thousands_separator=thousands_separator,
    )
    credit = _parse_number(
        source.get("credit"),
        decimal_separator=decimal_separator,
        thousands_separator=thousands_separator,
    )
    amount = _parse_number(
        source.get("amount"),
        decimal_separator=decimal_separator,
        thousands_separator=thousands_separator,
    )
    if amount is not None and debit is None and credit is None:
        amount_signed = amount
        debit = amount if amount >= ZERO else None
        credit = abs(amount) if amount < 0 else None
    else:
        amount_signed = (debit or ZERO) - (credit or ZERO)
    account = _clean_text(source.get("account"))
    line_desc = _clean_text(source.get("line_desc"))
    account_desc = _clean_text(source.get("account_desc"))
    if not account and not line_desc and amount_signed == ZERO:
        return None
    entry_date = _parse_date(source.get("entry_date"))
    return {
        "entry_date": entry_date.isoformat() if entry_date else None,
        "movement_number": _clean_text(source.get("movement_number")) or None,
        "line_number": _clean_text(source.get("line_number")) or None,
        "account": account or None,
        "account_desc": account_desc or None,
        "line_desc": line_desc or None,
        "debit": decimal_text(debit) if debit is not None else None,
        "credit": decimal_text(credit) if credit is not None else None,
        "amount_signed": decimal_text(amount_signed),
        "amount_abs": decimal_text(abs(amount_signed)),
        "currency": _clean_text(source.get("currency")) or None,
        "unit": _clean_text(source.get("unit")) or None,
        "reported_increment": (_clean_text(source.get("reported_increment")) or None),
        "source_file": source_file,
        "source_sheet": source_sheet,
        "source_page": source_page,
        "source_row": source_row,
    }


def _normalize_tabular(
    df: pl.DataFrame,
    mapping: dict[str, Any],
    path: Path,
    *,
    decimal_separator: str | None,
    thousands_separator: str | None,
    amount_sign_convention: str | None,
    source_sheet: str | None,
    source_row_offset: int,
    carry_forward_fields: Sequence[str],
    currency: str,
    unit: str,
) -> tuple[pl.DataFrame, dict[str, Any]]:
    records: list[dict[str, Any]] = []
    rejected_rows: list[dict[str, Any]] = []
    excluded_non_monetary_rows: list[int] = []
    candidate_row_count = 0
    current_date: Any = None
    current_movement: Any = None
    for row_idx, row in enumerate(df.iter_rows(named=True), start=1):
        source_row = source_row_offset + row_idx
        date_col = _field(mapping, "date")
        movement_col = _field(mapping, "movement_number")
        row_date = row.get(date_col) if date_col else None
        row_movement = row.get(movement_col) if movement_col else None
        if "date" in carry_forward_fields and _parse_date(row_date) is not None:
            current_date = row_date
        if "movement_number" in carry_forward_fields and _clean_text(row_movement):
            current_movement = row_movement
        debit_value = row.get(_field(mapping, "debit") or "")
        credit_value = row.get(_field(mapping, "credit") or "")
        amount_value = row.get(_field(mapping, "amount") or "")
        account_value = row.get(_field(mapping, "account") or "")
        has_account = bool(_clean_text(account_value))
        has_money = any(
            _clean_text(value) for value in (debit_value, credit_value, amount_value)
        )
        if not has_money:
            if has_account:
                excluded_non_monetary_rows.append(source_row)
            continue
        candidate_row_count += 1
        debit = _parse_number(
            debit_value,
            decimal_separator=decimal_separator,
            thousands_separator=thousands_separator,
        )
        credit = _parse_number(
            credit_value,
            decimal_separator=decimal_separator,
            thousands_separator=thousands_separator,
        )
        amount = _parse_number(
            amount_value,
            decimal_separator=decimal_separator,
            thousands_separator=thousands_separator,
        )
        reason: str | None = None
        if not has_account:
            reason = "missing_account"
        effective_date = current_date if "date" in carry_forward_fields else row_date
        effective_movement = (
            current_movement
            if "movement_number" in carry_forward_fields
            else row_movement
        )
        elif_date = _parse_date(effective_date)
        if reason is None and elif_date is None:
            reason = "missing_entry_date"
        elif reason is None and _field(mapping, "amount"):
            if amount is None:
                reason = "invalid_or_missing_amount"
            elif amount == ZERO:
                reason = "zero_amount"
        elif reason is None and debit is None and credit is None:
            reason = "invalid_or_missing_debit_credit"
        elif reason is None and (debit or ZERO) != ZERO and (credit or ZERO) != ZERO:
            reason = "both_debit_and_credit_nonzero"
        elif reason is None and (debit or ZERO) == ZERO and (credit or ZERO) == ZERO:
            reason = "zero_amount"
        if reason is not None:
            rejected_rows.append({"source_row": source_row, "reason": reason})
            continue
        if amount is not None and amount_sign_convention == "credit_positive":
            amount = -amount
        source = {
            "entry_date": effective_date,
            "movement_number": effective_movement,
            "line_number": row.get(_field(mapping, "line_number") or ""),
            "account": account_value,
            "account_desc": row.get(_field(mapping, "account_desc") or ""),
            "line_desc": row.get(_field(mapping, "line_desc") or ""),
            "debit": debit,
            "credit": credit,
            "amount": amount,
            "currency": currency,
            "unit": unit,
            "reported_increment": _reported_increment(
                (
                    amount_value
                    if _field(mapping, "amount")
                    else (
                        debit_value
                        if debit is not None and debit != ZERO
                        else credit_value
                    )
                ),
                decimal_separator=decimal_separator,
            ),
            "__decimal_separator": decimal_separator,
            "__thousands_separator": thousands_separator,
        }
        record = _normalize_record(
            source,
            source_file=path.name,
            source_sheet=source_sheet,
            source_page=None,
            source_row=source_row,
        )
        if record:
            records.append(record)
    frame = (
        pl.DataFrame(records, schema=CANONICAL_SCHEMA, strict=False)
        if records
        else pl.DataFrame(schema=CANONICAL_SCHEMA)
    )
    return frame, {
        "candidate_row_count": candidate_row_count,
        "emitted_row_count": frame.height,
        "rejected_row_count": len(rejected_rows),
        "rejected_rows": rejected_rows[:50],
        "excluded_non_monetary_row_count": len(excluded_non_monetary_rows),
        "excluded_non_monetary_rows": excluded_non_monetary_rows[:50],
    }


def _find_header_index(raw_df: pl.DataFrame, tokens: Sequence[str]) -> int | None:
    for idx in range(min(raw_df.height, 40)):
        labels = [_norm_label(value) for value in _row_values(raw_df, idx)]
        joined = " ".join(labels)
        if all(token in joined for token in tokens):
            return idx
    return None


def _find_header_col(header: Sequence[Any], *tokens: str) -> int | None:
    for idx, value in enumerate(header):
        label = _norm_label(value)
        if any(token in label for token in tokens):
            return idx
    return None


def _layout_cell(row: Sequence[Any], layout: dict[str, Any], field: str) -> Any:
    columns = layout.get("columns")
    if not isinstance(columns, dict):
        return None
    value = columns.get(field)
    if not isinstance(value, int) or isinstance(value, bool) or value <= 0:
        return None
    index = value - 1
    return row[index] if index < len(row) else None


def _parse_print_friendly_excel(
    path: Path,
    raw_df: pl.DataFrame,
    *,
    layout: dict[str, Any],
    decimal_separator: str | None,
    thousands_separator: str | None,
    source_sheet: str | None,
    carry_forward_fields: Sequence[str],
    currency: str,
    unit: str,
) -> tuple[pl.DataFrame, dict[str, Any]]:
    header_row = layout.get("header_row")
    if not isinstance(header_row, int) or isinstance(header_row, bool):
        return pl.DataFrame(schema=CANONICAL_SCHEMA), {
            "parser": "print_friendly_excel",
            "accepted": False,
            "candidate_row_count": 0,
            "emitted_row_count": 0,
            "rejected_row_count": 0,
            "rejected_rows": [],
        }
    header_idx = header_row - 1
    records: list[dict[str, Any]] = []
    rejected_rows: list[dict[str, Any]] = []
    excluded_non_monetary_rows: list[int] = []
    candidate_row_count = 0
    current_date: Any = None
    current_movement: Any = None
    current_desc = ""

    for row_idx in range(header_idx + 1, raw_df.height):
        row = _row_values(raw_df, row_idx)
        parsed_date = _parse_date(_layout_cell(row, layout, "date"))
        if "date" in carry_forward_fields and parsed_date is not None:
            current_date = parsed_date
        effective_date = current_date if "date" in carry_forward_fields else parsed_date
        row_movement = _clean_text(_layout_cell(row, layout, "movement_number"))
        if "movement_number" in carry_forward_fields and row_movement:
            current_movement = row_movement
        effective_movement = (
            current_movement
            if "movement_number" in carry_forward_fields
            else row_movement
        )
        row_desc = _clean_text(_layout_cell(row, layout, "line_desc"))
        if (
            "line_desc" in carry_forward_fields
            and row_desc
            and not ACCOUNT_RE.fullmatch(row_desc)
        ):
            current_desc = row_desc
        effective_desc = (
            current_desc if "line_desc" in carry_forward_fields else row_desc
        )
        account = _clean_text(_layout_cell(row, layout, "account"))
        account_desc = _clean_text(_layout_cell(row, layout, "account_desc"))
        debit_value = _layout_cell(row, layout, "debit")
        credit_value = _layout_cell(row, layout, "credit")
        has_money = bool(_clean_text(debit_value) or _clean_text(credit_value))
        if not has_money:
            if account:
                excluded_non_monetary_rows.append(row_idx + 1)
            continue
        candidate_row_count += 1
        debit = _parse_number(
            debit_value,
            decimal_separator=decimal_separator,
            thousands_separator=thousands_separator,
        )
        credit = _parse_number(
            credit_value,
            decimal_separator=decimal_separator,
            thousands_separator=thousands_separator,
        )
        reason: str | None = None
        if not account:
            reason = "missing_account"
        elif effective_date is None:
            reason = "missing_entry_date"
        elif debit is None and credit is None:
            reason = "invalid_or_missing_debit_credit"
        elif (debit or ZERO) != ZERO and (credit or ZERO) != ZERO:
            reason = "both_debit_and_credit_nonzero"
        elif (debit or ZERO) == ZERO and (credit or ZERO) == ZERO:
            reason = "zero_amount"
        if reason is not None:
            rejected_rows.append({"source_row": row_idx + 1, "reason": reason})
            continue
        record = _normalize_record(
            {
                "entry_date": effective_date,
                "movement_number": effective_movement,
                "line_number": _layout_cell(row, layout, "line_number"),
                "account": account,
                "account_desc": account_desc,
                "line_desc": effective_desc,
                "debit": debit,
                "credit": credit,
                "currency": currency,
                "unit": unit,
                "reported_increment": _reported_increment(
                    (
                        debit_value
                        if debit is not None and debit != ZERO
                        else credit_value
                    ),
                    decimal_separator=decimal_separator,
                ),
            },
            source_file=path.name,
            source_sheet=source_sheet,
            source_page=None,
            source_row=row_idx + 1,
        )
        if record:
            records.append(record)

    accepted = bool(records) and not rejected_rows
    diagnostics = {
        "parser": "print_friendly_excel",
        "accepted": accepted,
        "header_row": header_idx + 1,
        "candidate_row_count": candidate_row_count,
        "emitted_row_count": len(records),
        "rejected_row_count": len(rejected_rows),
        "rejected_rows": rejected_rows[:50],
        "excluded_non_monetary_row_count": len(excluded_non_monetary_rows),
        "excluded_non_monetary_rows": excluded_non_monetary_rows[:50],
    }
    frame = (
        pl.DataFrame(records, schema=CANONICAL_SCHEMA, strict=False)
        if records
        else pl.DataFrame(schema=CANONICAL_SCHEMA)
    )
    return frame, diagnostics


def _parse_text_pdf(_path: Path) -> tuple[pl.DataFrame, dict[str, Any]]:
    """Abstain from generic PDF side reconstruction.

    Text position alone cannot mechanically establish whether the last amount
    is debit, credit, balance, or a line total. A source-family-specific adapter
    must be implemented and tested before PDF rows may enter the population.
    """

    return pl.DataFrame(schema=CANONICAL_SCHEMA), {
        "parser": "text_pdf",
        "accepted": False,
        "candidate_row_count": 0,
        "emitted_row_count": 0,
        "rejected_row_count": 0,
        "rejected_rows": [],
        "status": "unsupported_source_layout",
        "reason": "generic_text_pdf_amount_side_reconstruction_disabled",
    }


def _execute_plan(
    path: Path,
    raw_df: pl.DataFrame,
    table: pl.DataFrame,
    plan: QualificationPlan,
    *,
    source_sheet: str | None,
) -> tuple[pl.DataFrame, dict[str, Any]]:
    """Execute one already-selected bounded adapter without promoting it."""

    if plan.parser == "print_friendly_excel":
        return _parse_print_friendly_excel(
            path,
            raw_df,
            layout=plan.layout,
            decimal_separator=plan.decimal_separator,
            thousands_separator=plan.thousands_separator,
            source_sheet=source_sheet,
            carry_forward_fields=plan.carry_forward_fields,
            currency=plan.currency,
            unit=plan.unit,
        )
    frame, diagnostics = _normalize_tabular(
        table,
        plan.mapping,
        path,
        decimal_separator=plan.decimal_separator,
        thousands_separator=plan.thousands_separator,
        amount_sign_convention=plan.amount_sign_convention,
        source_sheet=source_sheet,
        source_row_offset=max(plan.header_rows, default=0),
        carry_forward_fields=plan.carry_forward_fields,
        currency=plan.currency,
        unit=plan.unit,
    )
    diagnostics["parser"] = "tabular"
    diagnostics["accepted"] = frame.height > 0
    return frame, diagnostics


def _parser_result_is_complete(
    plan: QualificationPlan,
    parser_diagnostics: dict[str, Any],
) -> bool:
    """Return whether every source-owned monetary candidate closed exactly."""

    return bool(
        parser_diagnostics["candidate_row_count"] > 0
        and parser_diagnostics["rejected_row_count"] == 0
        and parser_diagnostics["emitted_row_count"]
        == parser_diagnostics["candidate_row_count"]
        and parser_diagnostics["candidate_row_count"] == plan.candidate_row_count
    )


def _unreadable_source_result(path: Path, error: BaseException) -> NormalizationResult:
    """Represent an unreadable source without crashing a multi-file intake."""

    plan = QualificationPlan(
        parser="unreadable_source",
        adapter_id="journal.unreadable_source.v2",
        source_family="unreadable.source_container.v1",
        status="unsupported_source_layout",
        candidate_row_count=0,
        controls=[
            _control(
                "source_container_readable",
                "failed",
                f"The source container could not be read ({type(error).__name__}).",
            )
        ],
        limitations=[
            "No rows from this unreadable source may enter the sampling population."
        ],
        reviewed_mapping_ref=None,
        reviewed_decision=None,
        header_rows=[],
        mapping={},
        layout={},
        excluded_monetary_columns=[],
        unresolved_monetary_columns=[],
        posting_identity="source_row",
        carry_forward_fields=[],
        currency="EUR",
        unit="currency",
        decimal_separator=None,
        thousands_separator=None,
        amount_sign_convention=None,
    )
    frame = pl.DataFrame(schema=CANONICAL_SCHEMA)
    qualification = _qualification_payload(path, plan, emitted_row_count=0)
    return NormalizationResult(
        frame=frame,
        diagnostics={
            "source_file": path.name,
            "parser": plan.parser,
            "accepted": False,
            "qualification": qualification,
            "qualification_status": qualification["status"],
            "review_required": False,
            "suggested_recipe": _recorded_recipe_from_plan(
                plan, qualification["status"]
            ),
            "candidate_row_count": 0,
            "emitted_row_count": 0,
            "rejected_row_count": 0,
            "rejected_rows": [],
            "row_count": 0,
            "failure_class": "parser_failure",
            "missing_fields": _missing_fields(frame),
            "parser_error": {
                "type": type(error).__name__,
                "message": str(error),
            },
        },
    )


def _unsupported_workbook_scope_result(
    path: Path,
    sheet_names: Sequence[str],
) -> NormalizationResult:
    """Withhold multi-sheet workbooks until every sheet has a bounded adapter."""

    plan = QualificationPlan(
        parser="multi_sheet_workbook",
        adapter_id="journal.multi_sheet_workbook.disabled.v2",
        source_family="workbook.multi_sheet.unqualified.v1",
        status="unsupported_source_layout",
        candidate_row_count=0,
        controls=[
            _control(
                "workbook_sheet_scope",
                "failed",
                "Every supplied worksheet must be qualified; this adapter is single-sheet only.",
            )
        ],
        limitations=[
            "No rows from a multi-sheet workbook may enter the population until every sheet is explicitly supported."
        ],
        reviewed_mapping_ref=None,
        reviewed_decision=None,
        header_rows=[],
        mapping={},
        layout={},
        excluded_monetary_columns=[],
        unresolved_monetary_columns=[],
        posting_identity="source_row",
        carry_forward_fields=[],
        currency="EUR",
        unit="currency",
        decimal_separator=None,
        thousands_separator=None,
        amount_sign_convention=None,
    )
    frame = pl.DataFrame(schema=CANONICAL_SCHEMA)
    qualification = _qualification_payload(path, plan, emitted_row_count=0)
    return NormalizationResult(
        frame=frame,
        diagnostics={
            "source_file": path.name,
            "parser": plan.parser,
            "accepted": False,
            "qualification": qualification,
            "qualification_status": qualification["status"],
            "review_required": False,
            "suggested_recipe": _recorded_recipe_from_plan(
                plan, qualification["status"]
            ),
            "candidate_row_count": 0,
            "emitted_row_count": 0,
            "rejected_row_count": 0,
            "rejected_rows": [],
            "row_count": 0,
            "sheet_names": list(sheet_names),
            "failure_class": "unsupported_workbook_scope",
            "missing_fields": _missing_fields(frame),
        },
    )


def _source_changed_result(path: Path) -> NormalizationResult:
    """Withhold rows when source bytes change across the parsing boundary."""

    error = RuntimeError("source bytes changed during normalization")
    result = _unreadable_source_result(path, error)
    diagnostics = dict(result.diagnostics)
    diagnostics["parser"] = "source_changed_during_normalization"
    diagnostics["failure_class"] = "source_changed_during_normalization"
    diagnostics["parser_error"] = {
        "type": type(error).__name__,
        "message": str(error),
    }
    return NormalizationResult(frame=result.frame, diagnostics=diagnostics)


def normalize_file(
    path: Path,
    recipe: dict[str, Any] | None = None,
    *,
    source_bytes: bytes | None = None,
) -> NormalizationResult:
    """Qualify one source, then normalize only a complete qualified population."""

    recipe = recipe or {}
    file_recipe = _recipe_for_file(recipe, path)
    suffix = path.suffix.lower()
    diagnostics: dict[str, Any] = {"source_file": path.name}
    if suffix == ".pdf":
        plan = _pdf_plan()
        frame, parser_diag = _parse_text_pdf(path)
        qualification = _qualification_payload(path, plan, emitted_row_count=0)
        diagnostics.update(parser_diag)
        diagnostics.update(
            {
                "qualification": qualification,
                "qualification_status": qualification["status"],
                "review_required": True,
                "suggested_recipe": _recorded_recipe_from_plan(
                    plan, qualification["status"]
                ),
                "missing_fields": _missing_fields(frame),
                "row_count": 0,
            }
        )
        return NormalizationResult(frame=frame, diagnostics=diagnostics)

    try:
        sheet_names = _source_sheet_names(path, source_bytes)
        if len(sheet_names) > 1:
            return _unsupported_workbook_scope_result(path, sheet_names)
        raw_df = _raw_table(path, source_bytes)
        source_sheet = sheet_names[0] if sheet_names else None
    except (
        fastexcel.CalamineError,
        BadZipFile,
        InvalidFileException,
        OSError,
        pl.exceptions.PolarsError,
        ValueError,
        RuntimeError,
    ) as exc:
        return _unreadable_source_result(path, exc)

    requested_parser = file_recipe.get("parser")
    table = pl.DataFrame()
    if requested_parser == "print_friendly_excel":
        plan = _print_plan(raw_df, file_recipe, path)
    elif requested_parser == "tabular":
        plan, table = _tabular_plan(raw_df, file_recipe, path)
    elif requested_parser is None:
        tabular_plan, tabular_table = _tabular_plan(raw_df, file_recipe, path)
        print_plan = _print_plan(raw_df, file_recipe, path)
        if tabular_plan.status != "unsupported_source_layout":
            plan, table = tabular_plan, tabular_table
        elif print_plan.status != "unsupported_source_layout":
            plan = print_plan
        elif print_plan.candidate_row_count > tabular_plan.candidate_row_count:
            plan = print_plan
        else:
            plan, table = tabular_plan, tabular_table
    else:
        parser_name = _identifier_fragment(str(requested_parser))
        plan = QualificationPlan(
            parser=str(requested_parser),
            adapter_id="journal.unsupported_adapter.v2",
            source_family=f"unsupported.{parser_name}",
            status="unsupported_source_layout",
            candidate_row_count=0,
            controls=[
                _control(
                    "bounded_adapter",
                    "failed",
                    f"Unsupported parser adapter: {requested_parser}",
                )
            ],
            limitations=["No rows from this file may enter the sampling population."],
            reviewed_mapping_ref=None,
            reviewed_decision=None,
            header_rows=[],
            mapping={},
            layout={},
            excluded_monetary_columns=[],
            unresolved_monetary_columns=[],
            posting_identity="source_row",
            carry_forward_fields=[],
            currency="EUR",
            unit="currency",
            decimal_separator=None,
            thousands_separator=None,
            amount_sign_convention=None,
        )

    if plan.status != "qualified":
        frame = pl.DataFrame(schema=CANONICAL_SCHEMA)
        proposal_diagnostics: dict[str, Any] = {}
        qualification_controls = list(plan.controls)
        qualification_status = plan.status
        qualification_limitations = list(plan.limitations)
        if plan.status == "needs_review":
            proposed_frame, proposal_diagnostics = _execute_plan(
                path,
                raw_df,
                table,
                plan,
                source_sheet=source_sheet,
            )
            proposal_complete = _parser_result_is_complete(plan, proposal_diagnostics)
            proposal_diagnostics = {
                **proposal_diagnostics,
                "proposed_emitted_row_count": proposed_frame.height,
            }
            if proposal_complete:
                qualification_controls.append(
                    _control(
                        "suggested_mapping_dry_run",
                        "passed",
                        "The proposed mapping closes every monetary candidate.",
                    )
                )
            else:
                qualification_status = "unsupported_source_layout"
                qualification_controls.append(
                    _control(
                        "suggested_mapping_dry_run",
                        "failed",
                        "The proposed mapping does not close every monetary candidate.",
                    )
                )
                qualification_limitations.append(
                    "The proposed mapping failed mechanical dry-run closure."
                )
        qualification = _qualification_payload(
            path,
            plan,
            emitted_row_count=0,
            controls=qualification_controls,
            status=qualification_status,
            limitations=qualification_limitations,
        )
        diagnostics.update(
            {
                **proposal_diagnostics,
                "parser": plan.parser,
                "accepted": False,
                "qualification": qualification,
                "qualification_status": qualification["status"],
                "review_required": qualification["status"] == "needs_review",
                "suggested_recipe": _recorded_recipe_from_plan(
                    plan, qualification["status"]
                ),
                "reviewed_decision": plan.reviewed_decision,
                "header_rows": plan.header_rows,
                "mapping": plan.mapping,
                "layout": plan.layout,
                "excluded_monetary_columns": plan.excluded_monetary_columns,
                "unresolved_monetary_columns": plan.unresolved_monetary_columns,
                "raw_columns": table.columns if plan.parser == "tabular" else [],
                "row_count": 0,
                "candidate_row_count": plan.candidate_row_count,
                "emitted_row_count": 0,
                "rejected_row_count": proposal_diagnostics.get("rejected_row_count", 0),
                "rejected_rows": proposal_diagnostics.get("rejected_rows", []),
                "missing_fields": _missing_fields(frame),
            }
        )
        return NormalizationResult(frame=frame, diagnostics=diagnostics)

    frame, parser_diag = _execute_plan(
        path,
        raw_df,
        table,
        plan,
        source_sheet=source_sheet,
    )
    complete = _parser_result_is_complete(plan, parser_diag)
    if not complete:
        completeness_control = _control(
            "population_completeness",
            "failed",
            "Every adapter candidate row must emit exactly one canonical row.",
        )
        qualification = _qualification_payload(
            path,
            plan,
            emitted_row_count=0,
            controls=[*plan.controls, completeness_control],
            status="unsupported_source_layout",
            limitations=[
                *plan.limitations,
                "Candidate and emitted rows do not close exactly; all rows were withheld.",
            ],
        )
        frame = pl.DataFrame(schema=CANONICAL_SCHEMA)
    else:
        completeness_control = _control(
            "population_completeness",
            "passed",
            "Every adapter candidate row emitted exactly one canonical row.",
        )
        qualification = _qualification_payload(
            path,
            plan,
            emitted_row_count=frame.height,
            controls=[*plan.controls, completeness_control],
        )
    diagnostics.update(
        {
            **parser_diag,
            "accepted": qualification["status"] == "qualified",
            "qualification": qualification,
            "qualification_status": qualification["status"],
            "review_required": qualification["status"] != "qualified",
            "suggested_recipe": _recorded_recipe_from_plan(
                plan, qualification["status"]
            ),
            "reviewed_decision": plan.reviewed_decision,
            "header_rows": plan.header_rows,
            "mapping": plan.mapping,
            "layout": plan.layout,
            "excluded_monetary_columns": plan.excluded_monetary_columns,
            "unresolved_monetary_columns": plan.unresolved_monetary_columns,
            "raw_columns": table.columns if plan.parser == "tabular" else [],
            "row_count": frame.height,
            "missing_fields": _missing_fields(frame),
        }
    )
    return NormalizationResult(frame=frame, diagnostics=diagnostics)


def _missing_fields(frame: pl.DataFrame) -> list[str]:
    missing: list[str] = []
    for field in ("entry_date", "account", "amount_abs"):
        if field not in frame.columns or frame.is_empty():
            missing.append(field)
            continue
        series = frame.get_column(field)
        if series.null_count() == series.len():
            missing.append(field)
    return missing


def _confidence(diagnostics: dict[str, Any]) -> int:
    """Expose a binary contract result instead of an arbitrary parser score."""

    return 1 if diagnostics.get("qualification_status") == "qualified" else 0


def _recipe_entry(diag: dict[str, Any]) -> dict[str, Any]:
    suggested = diag.get("suggested_recipe")
    return dict(suggested) if isinstance(suggested, dict) else {}


def _qualification_review_payload(
    file_diagnostics: Sequence[dict[str, Any]],
) -> dict[str, Any]:
    items: list[dict[str, Any]] = []
    for index, diagnostic in enumerate(file_diagnostics, start=1):
        qualification = validate_source_qualification(diagnostic["qualification"])
        status = qualification["status"]
        allowed_actions = [
            "edit",
            "mark_unclear",
            "request_more_documents",
            "skip",
        ]
        if status == "qualified":
            allowed_actions.insert(0, "accept")
        items.append(
            {
                "id": f"source-qualification-{index}",
                "item_type": "source_qualification",
                "title": f"{diagnostic['source_file']}: {status}",
                "source_path": diagnostic["source_file"],
                "allowed_actions": allowed_actions,
                "recommended_action": (
                    "accept"
                    if status == "qualified"
                    else (
                        "request_more_documents"
                        if status == "unsupported_source_layout"
                        else "mark_unclear"
                    )
                ),
                "status": "needs_review",
                "evidence": [
                    {
                        "kind": "source_qualification",
                        "qualification": qualification,
                    }
                ],
                "data": {
                    "qualification_status": status,
                    "candidate_row_count": qualification["candidate_row_count"],
                    "emitted_row_count": qualification["emitted_row_count"],
                    "limitations": qualification["limitations"],
                    "suggested_recipe": diagnostic.get("suggested_recipe", {}),
                },
            }
        )
    blocked = any(item["data"]["qualification_status"] != "qualified" for item in items)
    return {
        "schema_version": QUALIFICATION_REVIEW_SCHEMA_VERSION,
        "plugin": "journal-sampling",
        "workflow": "journal-sampling",
        "review_type": "journal_source_qualification_review",
        "items": items,
        "item_count": len(items),
        "status": "blocked_by_source_qualification" if blocked else "ready_for_review",
        "summary": {
            "source_count": len(items),
            "qualified_source_count": sum(
                item["data"]["qualification_status"] == "qualified" for item in items
            ),
            "blocked_source_count": sum(
                item["data"]["qualification_status"] != "qualified" for item in items
            ),
        },
    }


def inspect_path(
    input_path: Path,
    output_dir: Path | None = None,
    recipe_path: Path | None = None,
    *,
    language: object | None = None,
    document_language: object | None = None,
    client_engagement: Mapping[str, Any] | None = None,
) -> InspectionResult:
    """Inspect supported files and optionally write inspection artifacts."""

    normalized_client_engagement = (
        _validated_client_normalization_stage(
            client_engagement,
            input_path=input_path,
            output_dir=output_dir,
        )
        if output_dir is not None
        else None
    )
    recipe, _, _, _ = _read_recipe_with_receipt(recipe_path)
    languages = language_assumptions(
        recipe, language=language, document_language=document_language
    )
    files = supported_files(input_path)
    inspections: list[dict[str, Any]] = []
    recipe_files: dict[str, Any] = {}
    total_rows = 0
    for file_path in files:
        result = normalize_file(file_path, recipe)
        preview = result.frame.head(20).to_dicts()
        diag = dict(result.diagnostics)
        diag.update(languages)
        diag["confidence"] = _confidence(diag)
        diag["preview"] = preview
        inspections.append(diag)
        recipe_files[file_path.name] = _recipe_entry(diag)
        total_rows += result.frame.height

    suggested_recipe = {
        "version": 2,
        "description": (
            "Source-family recipe suggested for review. Rows remain withheld until "
            "qualification.status is reviewed and bound to this mapping digest."
        ),
        **languages,
        "files": recipe_files,
    }
    result = InspectionResult(
        files=inspections, total_rows=total_rows, suggested_recipe=suggested_recipe
    )
    if output_dir is not None:
        output_dir.mkdir(parents=True, exist_ok=True)
        review_payload = _qualification_review_payload(inspections)
        write_json(
            output_dir / "inspection.json",
            {
                "files": inspections,
                "total_rows": total_rows,
                "client_engagement": _portable_client_engagement_context(
                    normalized_client_engagement
                ),
                **languages,
                "qualification_review_payload": "qualification_review_payload.json",
            },
        )
        write_json(output_dir / "suggested_recipe.json", suggested_recipe)
        write_json(
            output_dir / "qualification_review_payload.json",
            review_payload,
        )
    return result


def normalize_path(
    input_path: Path,
    output_dir: Path,
    recipe_path: Path | None = None,
    *,
    language: object | None = None,
    document_language: object | None = None,
    client_engagement: Mapping[str, Any] | None = None,
) -> NormalizationResult:
    """Normalize all supported files under a path and write canonical outputs."""

    normalized_client_engagement = _validated_client_normalization_stage(
        client_engagement,
        input_path=input_path,
        output_dir=output_dir,
    )
    resolved_input = input_path.expanduser().resolve()
    output_dir = output_dir.expanduser().resolve()
    (
        recipe,
        recipe_source_path,
        recipe_source_receipt,
        recipe_source_bytes,
    ) = _read_recipe_with_receipt(recipe_path)
    languages = language_assumptions(
        recipe, language=language, document_language=document_language
    )
    files = supported_files(resolved_input)
    source_root = resolved_input if resolved_input.is_dir() else resolved_input.parent
    frames: list[pl.DataFrame] = []
    diagnostics: list[dict[str, Any]] = []
    source_receipts: list[dict[str, Any]] = []
    for file_path in files:
        artifact_id = f"source.{_identifier_fragment(file_path.name)}"
        source_receipt_before = artifact_receipt(
            source_root,
            file_path,
            artifact_id=artifact_id,
            root_id="source",
            role="source",
        )
        with tempfile.TemporaryDirectory(
            prefix="vera-journal-source-snapshot-"
        ) as snapshot_dir_name:
            snapshot_root = Path(snapshot_dir_name)
            snapshot_path = snapshot_root / file_path.name
            try:
                shutil.copyfile(file_path, snapshot_path)
                snapshot_receipt_before = artifact_receipt(
                    snapshot_root,
                    snapshot_path,
                    artifact_id=artifact_id,
                    root_id="source",
                    role="source",
                )
                if snapshot_receipt_before != source_receipt_before:
                    result = _source_changed_result(file_path)
                else:
                    captured_source_bytes = snapshot_path.read_bytes()
                    result = normalize_file(
                        snapshot_path,
                        recipe,
                        source_bytes=captured_source_bytes,
                    )
                snapshot_receipt_after = artifact_receipt(
                    snapshot_root,
                    snapshot_path,
                    artifact_id=artifact_id,
                    root_id="source",
                    role="source",
                )
            except OSError as exc:
                result = _unreadable_source_result(file_path, exc)
                snapshot_receipt_before = None
                snapshot_receipt_after = None
        source_receipt = artifact_receipt(
            source_root,
            file_path,
            artifact_id=artifact_id,
            root_id="source",
            role="source",
        )
        if (
            source_receipt != source_receipt_before
            or snapshot_receipt_before != source_receipt_before
            or snapshot_receipt_after != source_receipt_before
        ):
            result = _source_changed_result(file_path)
        frames.append(result.frame)
        file_diag = dict(result.diagnostics)
        file_diag.update(languages)
        file_diag["source_receipt"] = source_receipt
        source_receipts.append(source_receipt)
        diagnostics.append(file_diag)
    frame = (
        pl.concat(frames, how="vertical")
        if frames
        else pl.DataFrame(schema=CANONICAL_SCHEMA)
    )
    frame = frame.select([col for col in CANONICAL_COLUMNS if col in frame.columns])
    output_dir.mkdir(parents=True, exist_ok=True)
    captured_recipe_path: Path | None = None
    recipe_receipt: dict[str, Any] | None = None
    if recipe_source_bytes is not None:
        captured_recipe_path = output_dir / "normalization_recipe.json"
        captured_recipe_path.write_bytes(recipe_source_bytes)
        recipe_receipt = artifact_receipt(
            output_dir,
            captured_recipe_path,
            artifact_id="decision.normalization_recipe",
            root_id="normalization",
            role="reviewed_recipe",
            media_type="application/json",
        )
        if (
            recipe_source_receipt is None
            or recipe_source_receipt["byte_count"] != recipe_receipt["byte_count"]
            or recipe_source_receipt["sha256"] != recipe_receipt["sha256"]
        ):
            raise ValueError(
                "Captured normalization recipe does not match its stable source bytes."
            )
    csv_path = output_dir / "normalized_journal.csv"
    frame.write_csv(csv_path)
    qualifications = [
        validate_source_qualification(item["qualification"]) for item in diagnostics
    ]
    population_complete = bool(qualifications) and all(
        qualification["status"] == "qualified" for qualification in qualifications
    )
    emitted_total = sum(
        qualification["emitted_row_count"] for qualification in qualifications
    )
    population_complete = (
        population_complete
        and emitted_total == frame.height
        and frame.height > 0
        and recipe_receipt is not None
    )
    csv_receipt = artifact_receipt(
        output_dir,
        csv_path,
        artifact_id="prepared.normalized_journal",
        root_id="normalization",
        role="prepared",
        media_type="text/csv",
    )
    implementation_receipts = _implementation_receipts()
    reviewed_decisions = [
        item["reviewed_decision"]
        for item in diagnostics
        if isinstance(item.get("reviewed_decision"), dict)
    ]
    source_status = (
        "passed"
        if qualifications
        and all(item["status"] == "qualified" for item in qualifications)
        else (
            "failed"
            if any(
                item["status"] == "unsupported_source_layout" for item in qualifications
            )
            or not qualifications
            else "withheld"
        )
    )
    gate_register = build_gate_register(
        {
            "source": {
                "status": source_status,
                "evidence_refs": [item["qualification_id"] for item in qualifications],
                "limitations": (
                    []
                    if source_status == "passed"
                    else ["At least one source is not qualified."]
                ),
            },
            "preparation": {
                "status": "passed" if population_complete else "blocked",
                "evidence_refs": (
                    [csv_receipt["artifact_id"]] if population_complete else []
                ),
                "limitations": (
                    []
                    if population_complete
                    else ["Canonical population preparation is incomplete."]
                ),
            },
            "reconciliation": {
                "status": "not_applicable",
                "evidence_refs": [],
                "limitations": [],
            },
            "semantic_review": {
                "status": "not_assessed",
                "evidence_refs": [],
                "limitations": ["Professional sampling-basis review remains pending."],
            },
            "reporting": {
                "status": "not_applicable",
                "evidence_refs": [],
                "limitations": [],
            },
            "publication": {
                "status": "not_applicable",
                "evidence_refs": [],
                "limitations": [],
            },
        }
    )
    assurance_envelope = build_assurance_envelope(
        run_id=f"journal-sampling-normalization-{csv_receipt['sha256'][:16]}",
        workflow_id="journal-sampling-normalization",
        workflow_version=ADAPTER_VERSION,
        artifact_receipts=[
            *source_receipts,
            *([recipe_receipt] if recipe_receipt is not None else []),
            csv_receipt,
            *implementation_receipts,
        ],
        implementation_artifact_refs=[
            receipt["artifact_id"] for receipt in implementation_receipts
        ],
        reviewed_decisions=reviewed_decisions,
        source_qualifications=qualifications,
        allocation_ledgers=[],
        numeric_evidence_ledgers=[],
        gate_register=gate_register,
        limitations=(
            []
            if population_complete
            else ["Sampling remains blocked until every source is qualified."]
        ),
        artifact_roots={
            "source": source_root,
            "normalization": output_dir,
            **_implementation_artifact_roots(),
        },
    )
    write_json(
        output_dir / "reviewed_decisions.json",
        {
            "schema_version": "journal_sampling.reviewed_decisions.v1",
            "decisions": reviewed_decisions,
        },
    )
    write_json(output_dir / "assurance_gates.json", gate_register)
    write_json(output_dir / "assurance_envelope.json", assurance_envelope)
    review_payload = _qualification_review_payload(diagnostics)
    write_json(
        output_dir / "qualification_review_payload.json",
        review_payload,
    )
    managed_run = (
        isinstance(normalized_client_engagement, Mapping)
        and isinstance(normalized_client_engagement.get("run_root"), str)
        and bool(str(normalized_client_engagement["run_root"]).strip())
    )
    input_reference = _managed_run_reference(
        resolved_input,
        normalized_client_engagement,
    )
    source_root_reference = _managed_run_reference(
        source_root,
        normalized_client_engagement,
    )
    recipe_root_reference = (
        _managed_run_reference(output_dir, normalized_client_engagement)
        if captured_recipe_path is not None
        else ""
    )
    recipe_source_reference = (
        _managed_run_reference(recipe_source_path, normalized_client_engagement)
        if recipe_source_path is not None
        else ""
    )
    output_csv_reference = _managed_run_reference(
        csv_path,
        normalized_client_engagement,
    )
    diagnostics_payload = {
        "schema_version": NORMALIZATION_SCHEMA_VERSION,
        "client_engagement": _portable_client_engagement_context(
            normalized_client_engagement
        ),
        **({"path_reference": "run_root_relative"} if managed_run else {}),
        "input": input_reference,
        "source_root": source_root_reference,
        "source_receipts": source_receipts,
        "normalization_recipe_path": (
            captured_recipe_path.relative_to(output_dir).as_posix()
            if captured_recipe_path is not None
            else ""
        ),
        "normalization_recipe_root": (
            recipe_root_reference if captured_recipe_path is not None else ""
        ),
        "normalization_recipe_receipt": recipe_receipt,
        "normalization_recipe_source_path": (
            recipe_source_reference if recipe_source_path is not None else ""
        ),
        "normalization_recipe_source_receipt": recipe_source_receipt,
        "row_count": frame.height,
        **languages,
        "files": diagnostics,
        "output_csv": output_csv_reference,
        "normalized_csv_receipt": csv_receipt,
        "implementation_receipts": implementation_receipts,
        "reviewed_decisions": "reviewed_decisions.json",
        "assurance_gates": "assurance_gates.json",
        "assurance_envelope": "assurance_envelope.json",
        "source_qualifications": qualifications,
        "population_status": "complete" if population_complete else "incomplete",
        "qualification_review_payload": "qualification_review_payload.json",
    }
    write_json(
        output_dir / "normalization_diagnostics.json",
        {
            **diagnostics_payload,
            "content_sha256": canonical_json_sha256(diagnostics_payload),
        },
    )
    return NormalizationResult(
        frame=frame,
        diagnostics={
            "schema_version": NORMALIZATION_SCHEMA_VERSION,
            "client_engagement": _portable_client_engagement_context(
                normalized_client_engagement
            ),
            **({"path_reference": "run_root_relative"} if managed_run else {}),
            "files": diagnostics,
            "source_qualifications": qualifications,
            "source_root": source_root_reference,
            "source_receipts": source_receipts,
            "normalization_recipe_path": (
                captured_recipe_path.relative_to(output_dir).as_posix()
                if captured_recipe_path is not None
                else ""
            ),
            "normalization_recipe_root": (
                recipe_root_reference if captured_recipe_path is not None else ""
            ),
            "normalization_recipe_receipt": recipe_receipt,
            "normalization_recipe_source_path": (
                recipe_source_reference if recipe_source_path is not None else ""
            ),
            "normalization_recipe_source_receipt": recipe_source_receipt,
            "population_status": "complete" if population_complete else "incomplete",
            "row_count": frame.height,
            **languages,
        },
    )


def _apply_filters(
    frame: pl.DataFrame,
    *,
    include_accounts: Sequence[str] = (),
    exclude_accounts: Sequence[str] = (),
    date_start: str | None = None,
    date_end: str | None = None,
    min_abs: Decimal | str | int | None = None,
    keyword: str | None = None,
) -> pl.DataFrame:
    result = frame
    if include_accounts:
        result = result.filter(
            pl.col("account").cast(pl.Utf8).is_in(list(include_accounts))
        )
    if exclude_accounts:
        result = result.filter(
            ~pl.col("account").cast(pl.Utf8).is_in(list(exclude_accounts))
        )
    if date_start:
        result = result.filter(
            pl.col("entry_date").cast(pl.Date, strict=False)
            >= pl.lit(date_start).cast(pl.Date)
        )
    if date_end:
        result = result.filter(
            pl.col("entry_date").cast(pl.Date, strict=False)
            <= pl.lit(date_end).cast(pl.Date)
        )
    if min_abs is not None:
        threshold = (
            min_abs
            if isinstance(min_abs, Decimal)
            else parse_localized_decimal(min_abs, label="min_abs")
        )
        mask = [
            parse_canonical_decimal(value, label="amount_abs") >= threshold
            for value in result.get_column("amount_abs").to_list()
        ]
        result = result.filter(pl.Series("__min_abs_mask", mask))
    if keyword:
        lowered = keyword.lower()
        result = result.filter(
            pl.col("line_desc")
            .cast(pl.Utf8, strict=False)
            .fill_null("")
            .str.to_lowercase()
            .str.contains(lowered, literal=True)
        )
    return result


def _systematic_sample(frame: pl.DataFrame, size: int) -> pl.DataFrame:
    if frame.height == 0 or size <= 0:
        return frame.head(0)
    if size >= frame.height:
        return frame
    step = frame.height / size
    indexes = sorted({min(int(idx * step), frame.height - 1) for idx in range(size)})
    return (
        frame.with_row_index("__idx")
        .filter(pl.col("__idx").is_in(indexes))
        .drop("__idx")
    )


def _stratified_sample(
    frame: pl.DataFrame, size: int, group_column: str
) -> pl.DataFrame:
    if frame.height == 0 or size <= 0:
        return frame.head(0)
    if group_column not in frame.columns:
        raise ValueError(f"Stratified sampling group column not found: {group_column}")
    groups = frame.partition_by(group_column, maintain_order=True)
    per_group = max(1, math.ceil(size / len(groups)))
    parts = [
        group.sample(n=min(per_group, group.height), seed=42)
        for group in groups
        if group.height > 0
    ]
    return pl.concat(parts).head(size) if parts else frame.head(0)


def _mus_sample(frame: pl.DataFrame, size: int) -> pl.DataFrame:
    if frame.height == 0 or size <= 0:
        return frame.head(0)
    indexed_rows = frame.with_row_index("__idx").to_dicts()
    indexed_rows.sort(
        key=lambda row: parse_canonical_decimal(row["amount_abs"], label="amount_abs"),
        reverse=True,
    )
    total = sum(
        (
            parse_canonical_decimal(row["amount_abs"], label="amount_abs")
            for row in indexed_rows
        ),
        ZERO,
    )
    if total <= ZERO:
        indexes = [int(row["__idx"]) for row in indexed_rows[: min(size, frame.height)]]
        return (
            frame.with_row_index("__idx")
            .filter(pl.col("__idx").is_in(indexes))
            .drop("__idx")
        )
    selection_count = min(size, frame.height)
    interval = total / Decimal(selection_count)
    thresholds = [
        (Decimal(index) + Decimal("0.5")) * interval for index in range(selection_count)
    ]
    running = ZERO
    picked: list[int] = []
    threshold_idx = 0
    for row in indexed_rows:
        running += parse_canonical_decimal(row["amount_abs"], label="amount_abs")
        while threshold_idx < len(thresholds) and running >= thresholds[threshold_idx]:
            picked.append(int(row["__idx"]))
            threshold_idx += 1
    unique = sorted(dict.fromkeys(picked))
    return (
        frame.with_row_index("__idx")
        .filter(pl.col("__idx").is_in(unique))
        .drop("__idx")
        .head(size)
    )


def _validated_normalization_recipe(
    normalized_csv: Path,
    diagnostics: dict[str, Any],
) -> tuple[Path, dict[str, Any], dict[str, Any]]:
    """Validate the retained reviewed recipe and its original source receipt."""

    normalization_root = normalized_csv.parent
    recipe_path_value = diagnostics.get("normalization_recipe_path")
    recipe_root_value = diagnostics.get("normalization_recipe_root")
    recipe_receipt = diagnostics.get("normalization_recipe_receipt")
    if (
        recipe_path_value != "normalization_recipe.json"
        or not isinstance(recipe_root_value, str)
        or not recipe_root_value.strip()
        or not isinstance(recipe_receipt, dict)
    ):
        raise ValueError(
            "Normalization diagnostics do not retain the exact reviewed recipe."
        )
    recorded_root = _resolve_normalization_reference(
        normalized_csv,
        diagnostics,
        recipe_root_value,
        label="Normalization recipe root",
    )
    recipe_path = normalization_root / recipe_path_value
    _require_ordinary_single_link(recipe_path, label="Captured normalization recipe")
    validated_recipe_receipt = validate_artifact_receipt(
        {"normalization": normalization_root},
        recipe_receipt,
    )
    if (
        validated_recipe_receipt["artifact_id"] != "decision.normalization_recipe"
        or validated_recipe_receipt["root_id"] != "normalization"
        or validated_recipe_receipt["role"] != "reviewed_recipe"
        or validated_recipe_receipt["path"] != recipe_path_value
        or validated_recipe_receipt.get("media_type") != "application/json"
    ):
        raise ValueError("Captured normalization recipe receipt is not exact.")

    source_path_value = diagnostics.get("normalization_recipe_source_path")
    source_receipt = diagnostics.get("normalization_recipe_source_receipt")
    if (
        not isinstance(source_path_value, str)
        or not source_path_value.strip()
        or not isinstance(source_receipt, dict)
    ):
        raise ValueError("Normalization recipe source provenance is incomplete.")
    recorded_source_path = _resolve_normalization_reference(
        normalized_csv,
        diagnostics,
        source_path_value,
        label="Normalization recipe source path",
    )
    try:
        source_relative = recorded_source_path.relative_to(recorded_root)
    except ValueError:
        source_path = recorded_source_path
    else:
        source_path = normalization_root / source_relative
    _require_ordinary_single_link(source_path, label="Normalization recipe source")
    validated_source_receipt = validate_artifact_receipt(
        {"normalization_recipe_source": source_path.parent},
        source_receipt,
    )
    if (
        validated_source_receipt["artifact_id"]
        != "decision.normalization_recipe_source"
        or validated_source_receipt["root_id"] != "normalization_recipe_source"
        or validated_source_receipt["role"] != "reviewed_recipe"
        or (source_path.parent / validated_source_receipt["path"]).resolve()
        != source_path
        or validated_source_receipt.get("media_type") != "application/json"
        or validated_source_receipt["byte_count"]
        != validated_recipe_receipt["byte_count"]
        or validated_source_receipt["sha256"] != validated_recipe_receipt["sha256"]
    ):
        raise ValueError(
            "Retained normalization recipe does not close to its reviewed source."
        )
    try:
        recipe_payload = json.loads(recipe_path.read_text(encoding="utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ValueError(
            "Captured normalization recipe is not valid UTF-8 JSON."
        ) from exc
    if not isinstance(recipe_payload, dict):
        raise ValueError("Captured normalization recipe must be a JSON object.")
    return recipe_path, validated_recipe_receipt, validated_source_receipt


_NORMALIZATION_REPLAY_FIELDS = (
    "schema_version",
    "path_reference",
    "input",
    "source_root",
    "source_receipts",
    "normalization_recipe_path",
    "normalization_recipe_receipt",
    "row_count",
    "language",
    "document_language",
    "files",
    "normalized_csv_receipt",
    "implementation_receipts",
    "reviewed_decisions",
    "assurance_gates",
    "assurance_envelope",
    "source_qualifications",
    "population_status",
    "qualification_review_payload",
)


def _fresh_normalization_replay(
    normalized_csv: Path,
    diagnostics: dict[str, Any],
    *,
    recipe_path: Path,
) -> dict[str, Any]:
    """Re-run raw normalization and compare every material preparation claim."""

    input_value = diagnostics.get("input")
    source_root_value = diagnostics.get("source_root")
    if (
        not isinstance(input_value, str)
        or not input_value.strip()
        or not isinstance(source_root_value, str)
        or not source_root_value.strip()
    ):
        raise ValueError("Normalization diagnostics do not retain the raw input path.")
    source_input = _resolve_normalization_reference(
        normalized_csv,
        diagnostics,
        input_value,
        label="Normalization raw input",
    )
    source_root = _resolve_normalization_reference(
        normalized_csv,
        diagnostics,
        source_root_value,
        label="Normalization source root",
    )
    if source_input != source_root and source_input.parent != source_root:
        raise ValueError("Normalization raw input provenance is not canonical.")

    original_bytes = normalized_csv.read_bytes()
    original_receipt = validate_artifact_receipt(
        {"normalization": normalized_csv.parent},
        diagnostics["normalized_csv_receipt"],
    )
    if (
        len(original_bytes) != original_receipt["byte_count"]
        or hashlib.sha256(original_bytes).hexdigest() != original_receipt["sha256"]
    ):
        raise ValueError("Normalized population changed during fresh replay.")

    with tempfile.TemporaryDirectory(
        prefix="vera-journal-normalization-replay-"
    ) as replay_name:
        replay_root = Path(replay_name)
        normalize_path(
            source_input,
            replay_root,
            recipe_path,
            language=diagnostics.get("language"),
            document_language=diagnostics.get("document_language"),
        )
        replay_csv = replay_root / "normalized_journal.csv"
        replay_bytes = replay_csv.read_bytes()
        replay_diagnostics_payload = read_json(
            replay_root / "normalization_diagnostics.json"
        )
        replay_digest = replay_diagnostics_payload.pop("content_sha256", None)
        if not isinstance(replay_digest, str) or replay_digest != canonical_json_sha256(
            replay_diagnostics_payload
        ):
            raise ValueError("Fresh normalization diagnostics did not seal.")
        original_projection = {
            field: diagnostics.get(field) for field in _NORMALIZATION_REPLAY_FIELDS
        }
        replay_projection = {
            field: replay_diagnostics_payload.get(field)
            for field in _NORMALIZATION_REPLAY_FIELDS
        }
        if diagnostics.get("path_reference") == "run_root_relative":
            for field in ("path_reference", "input", "source_root"):
                replay_projection[field] = original_projection[field]
        if replay_bytes != original_bytes:
            raise ValueError(
                "Fresh normalization does not reproduce normalized_journal.csv."
            )
        if replay_projection != original_projection:
            raise ValueError(
                "Fresh normalization does not reproduce the preparation contract."
            )
        for artifact_name in (
            "reviewed_decisions.json",
            "assurance_gates.json",
            "assurance_envelope.json",
            "qualification_review_payload.json",
        ):
            if read_json(replay_root / artifact_name) != read_json(
                normalized_csv.parent / artifact_name
            ):
                raise ValueError(
                    f"Fresh normalization does not reproduce {artifact_name}."
                )

    replay_receipt = {
        "schema_version": "journal_sampling.normalization_replay.v1",
        "status": "passed",
        "normalized_csv_byte_count": original_receipt["byte_count"],
        "normalized_csv_sha256": original_receipt["sha256"],
        "row_count": diagnostics["row_count"],
        "normalization_content_sha256": canonical_json_sha256(diagnostics),
        "assurance_envelope_content_sha256": read_json(
            normalized_csv.parent / "assurance_envelope.json"
        )["content_sha256"],
        "recipe_sha256": diagnostics["normalization_recipe_receipt"]["sha256"],
        "recipe_source_sha256": diagnostics["normalization_recipe_source_receipt"][
            "sha256"
        ],
        "source_receipt_set_sha256": canonical_json_sha256(
            {"receipts": diagnostics["source_receipts"]}
        ),
        "qualification_set_sha256": canonical_json_sha256(
            {"qualifications": diagnostics["source_qualifications"]}
        ),
        "implementation_receipt_set_sha256": canonical_json_sha256(
            {"receipts": diagnostics["implementation_receipts"]}
        ),
        "material_projection_sha256": canonical_json_sha256(original_projection),
    }
    return {
        **replay_receipt,
        "content_sha256": canonical_json_sha256(replay_receipt),
    }


def _validate_population_proof(
    normalized_csv: Path,
    frame: pl.DataFrame,
    diagnostics_path: Path,
) -> dict[str, Any]:
    diagnostics = read_json(diagnostics_path)
    recorded_digest = diagnostics.pop("content_sha256", None)
    if not isinstance(recorded_digest, str) or recorded_digest != canonical_json_sha256(
        diagnostics
    ):
        raise ValueError("Normalization diagnostics content hash is stale.")
    if diagnostics.get("schema_version") != NORMALIZATION_SCHEMA_VERSION:
        raise ValueError(
            "Sampling requires journal_sampling.normalization.v2 diagnostics."
        )
    if diagnostics.get("population_status") != "complete":
        raise ValueError(
            "Sampling is blocked because the normalized population is incomplete."
        )
    if frame.columns != CANONICAL_COLUMNS:
        raise ValueError("Normalized population columns are not canonical or ordered.")
    receipt = diagnostics.get("normalized_csv_receipt")
    if not isinstance(receipt, dict):
        raise ValueError("Normalization diagnostics are missing the CSV receipt.")
    validate_artifact_receipt({"normalization": normalized_csv.parent}, receipt)
    qualifications_raw = diagnostics.get("source_qualifications")
    if not isinstance(qualifications_raw, list) or not qualifications_raw:
        raise ValueError("Normalization diagnostics contain no source qualifications.")
    qualifications = [
        validate_source_qualification(value) for value in qualifications_raw
    ]
    if any(item["status"] != "qualified" for item in qualifications):
        raise ValueError("Every source must be qualified before sampling.")
    source_root_reference = diagnostics.get("source_root")
    source_receipts_raw = diagnostics.get("source_receipts")
    if not isinstance(source_root_reference, str) or not source_root_reference.strip():
        raise ValueError("Normalization diagnostics are missing the source root.")
    if not isinstance(source_receipts_raw, list) or not source_receipts_raw:
        raise ValueError("Normalization diagnostics are missing source receipts.")
    source_root = _resolve_normalization_reference(
        normalized_csv,
        diagnostics,
        source_root_reference,
        label="Normalization source root",
    )
    source_receipts = [
        validate_artifact_receipt({"source": source_root}, receipt)
        for receipt in source_receipts_raw
        if isinstance(receipt, dict)
    ]
    if len(source_receipts) != len(source_receipts_raw):
        raise ValueError("Normalization source receipts are malformed.")
    (
        recipe_path,
        recipe_receipt,
        recipe_source_receipt,
    ) = _validated_normalization_recipe(normalized_csv, diagnostics)
    receipt_ids = [str(receipt["artifact_id"]) for receipt in source_receipts]
    expected_source_refs = [
        source_ref
        for qualification in qualifications
        for source_ref in qualification["source_artifact_refs"]
    ]
    if len(receipt_ids) != len(set(receipt_ids)) or sorted(receipt_ids) != sorted(
        expected_source_refs
    ):
        raise ValueError(
            "Source receipts do not close to source qualification references."
        )
    envelope_name = diagnostics.get("assurance_envelope")
    if not isinstance(envelope_name, str) or not envelope_name.strip():
        raise ValueError(
            "Normalization diagnostics are missing the assurance envelope."
        )
    envelope = validate_assurance_envelope(
        read_json(normalized_csv.parent / envelope_name),
        artifact_roots={
            "source": source_root,
            "normalization": normalized_csv.parent,
            **_implementation_artifact_roots(),
        },
    )
    _validate_exact_implementation_receipts(envelope)
    envelope_receipts = {
        receipt["artifact_id"]: receipt for receipt in envelope["artifact_receipts"]
    }
    if envelope_receipts.get(recipe_receipt["artifact_id"]) != recipe_receipt:
        raise ValueError(
            "Assurance envelope does not bind the retained normalization recipe."
        )
    if diagnostics.get("implementation_receipts") != [
        receipt
        for receipt in envelope["artifact_receipts"]
        if receipt["role"] == "implementation"
    ]:
        raise ValueError(
            "Normalization diagnostics implementation receipts are not exact."
        )
    if {item["qualification_id"] for item in envelope["source_qualifications"]} != {
        item["qualification_id"] for item in qualifications
    }:
        raise ValueError("Assurance envelope does not close to source qualifications.")
    gates = envelope["gate_register"]["gates"]
    if (
        gates["source"]["status"] != "passed"
        or gates["preparation"]["status"] != "passed"
    ):
        raise ValueError("Assurance gates do not permit sampling.")
    if diagnostics.get("row_count") != frame.height or frame.height <= 0:
        raise ValueError("Normalized population row count does not close.")
    file_diagnostics = diagnostics.get("files")
    if not isinstance(file_diagnostics, list) or not file_diagnostics:
        raise ValueError("Per-file qualification diagnostics are required.")
    expected_by_file: dict[str, int] = {}
    for item in file_diagnostics:
        if not isinstance(item, dict) or not isinstance(item.get("source_file"), str):
            raise ValueError("Per-file qualification diagnostics are malformed.")
        qualification = validate_source_qualification(item.get("qualification"))
        if qualification["status"] != "qualified":
            raise ValueError("Every source file must remain qualified.")
        source_file = item["source_file"]
        if source_file in expected_by_file:
            raise ValueError("Source file names must be unique within one population.")
        expected_by_file[source_file] = qualification["emitted_row_count"]
    actual_by_file: dict[str, int] = {}
    for value in frame.get_column("source_file").to_list():
        source_file = str(value)
        actual_by_file[source_file] = actual_by_file.get(source_file, 0) + 1
    if actual_by_file != expected_by_file:
        raise ValueError(
            "Normalized rows do not close to the qualified per-file populations."
        )
    locators: set[tuple[str, str, str, str]] = set()
    for row_index, row in enumerate(frame.iter_rows(named=True), start=1):
        if not _clean_text(row.get("account")):
            raise ValueError(f"Row {row_index} has no account.")
        if _parse_date(row.get("entry_date")) is None:
            raise ValueError(f"Row {row_index} has no valid entry date.")
        currency = _clean_text(row.get("currency"))
        unit = _clean_text(row.get("unit"))
        if re.fullmatch(r"[A-Z]{3}", currency) is None or unit != "currency":
            raise ValueError(f"Row {row_index} has no valid currency/unit contract.")
        reported_increment = parse_canonical_decimal(
            row["reported_increment"],
            label=f"row {row_index} reported_increment",
        )
        if reported_increment <= ZERO:
            raise ValueError(
                f"Row {row_index} reported increment must be strictly positive."
            )
        locator = (
            _clean_text(row.get("source_file")),
            _clean_text(row.get("source_sheet")),
            _clean_text(row.get("source_page")),
            _clean_text(row.get("source_row")),
        )
        if not locator[0] or (not locator[2] and not locator[3]):
            raise ValueError(f"Row {row_index} has no complete source locator.")
        if locator in locators:
            raise ValueError(f"Row {row_index} duplicates a source locator.")
        locators.add(locator)
        debit = (
            ZERO
            if not _clean_text(row.get("debit"))
            else parse_canonical_decimal(row["debit"], label=f"row {row_index} debit")
        )
        credit = (
            ZERO
            if not _clean_text(row.get("credit"))
            else parse_canonical_decimal(row["credit"], label=f"row {row_index} credit")
        )
        signed = parse_canonical_decimal(
            row["amount_signed"], label=f"row {row_index} amount_signed"
        )
        absolute = parse_canonical_decimal(
            row["amount_abs"], label=f"row {row_index} amount_abs"
        )
        if signed != debit - credit or absolute != abs(signed):
            raise ValueError(f"Normalized monetary closure failed for row {row_index}.")
        if absolute == ZERO:
            raise ValueError(f"Zero-value row {row_index} cannot enter the population.")
    normalization_replay = _fresh_normalization_replay(
        normalized_csv,
        diagnostics,
        recipe_path=recipe_path,
    )
    managed_paths = diagnostics.get("path_reference") == "run_root_relative"
    output_csv_reference = diagnostics.get("output_csv")
    if managed_paths:
        if not isinstance(output_csv_reference, str) or not output_csv_reference:
            raise ValueError("Normalization diagnostics output reference is missing.")
        output_reference = Path(output_csv_reference)
        diagnostics_reference = output_reference.with_name(
            diagnostics_path.name
        ).as_posix()
        envelope_reference = output_reference.with_name(envelope_name).as_posix()
    else:
        diagnostics_reference = diagnostics_path.as_posix()
        envelope_reference = (normalized_csv.parent / envelope_name).as_posix()
    return {
        **({"path_reference": "run_root_relative"} if managed_paths else {}),
        "diagnostics_path": diagnostics_reference,
        "source_root": source_root_reference,
        "normalization_content_sha256": recorded_digest,
        "population_status": "complete",
        "normalized_csv_receipt": receipt,
        "source_receipts": source_receipts,
        "normalization_recipe_receipt": recipe_receipt,
        "normalization_recipe_source_receipt": recipe_source_receipt,
        "normalization_replay": normalization_replay,
        "assurance_envelope_path": envelope_reference,
        "assurance_envelope_content_sha256": envelope["content_sha256"],
        "assurance_gates": envelope["gate_register"],
        "qualification_ids": [
            qualification["qualification_id"] for qualification in qualifications
        ],
    }


def replay_normalization_from_provenance(
    normalized_csv: Path,
    diagnostics_path: Path | None = None,
) -> dict[str, Any]:
    """Freshly re-perform raw normalization under the retained reviewed recipe."""

    resolved_csv = normalized_csv.expanduser().resolve()
    resolved_diagnostics = (
        diagnostics_path.expanduser().resolve()
        if diagnostics_path is not None
        else resolved_csv.parent / "normalization_diagnostics.json"
    )
    frame = pl.read_csv(resolved_csv, infer_schema=False)
    proof = _validate_population_proof(
        resolved_csv,
        frame,
        resolved_diagnostics,
    )
    return dict(proof["normalization_replay"])


def _select_sample(
    population: pl.DataFrame,
    *,
    method: str,
    size: int,
    group_column: str,
) -> pl.DataFrame:
    """Select rows under the reviewed deterministic sampling contract."""

    if size <= 0:
        raise ValueError("Requested sample size must be strictly positive.")
    if population.height <= 0:
        raise ValueError("Sampling is blocked because filters produced no rows.")
    if method == "random":
        return population.sample(n=min(size, population.height), seed=42)
    if method == "systematic":
        return _systematic_sample(population, size)
    if method == "stratified":
        return _stratified_sample(population, size, group_column)
    if method == "mus":
        return _mus_sample(population, size)
    raise ValueError(f"Unsupported sampling method: {method}")


def _material_value(value: object, field: str) -> dict[str, str]:
    """Return an exact typed text value for a material sample field.

    This deterministic conversion is justified because CSV/XLSX cell equality,
    canonical decimal syntax, and integer locators are mechanically verifiable.
    It does not decide whether a sampled item is sufficient audit evidence.
    """

    if value is None or value == "":
        return {"kind": "empty", "value": ""}
    text = str(value)
    if field in SAMPLE_DECIMAL_FIELDS:
        parsed = parse_canonical_decimal(text, label=field)
        if decimal_text(parsed) != text:
            raise ValueError(f"{field} is not canonical Decimal text.")
        return {"kind": "decimal", "value": text}
    if field in SAMPLE_INTEGER_FIELDS:
        try:
            parsed_integer = int(text)
        except ValueError as exc:
            raise ValueError(f"{field} is not canonical integer text.") from exc
        if str(parsed_integer) != text:
            raise ValueError(f"{field} is not canonical integer text.")
        return {"kind": "integer", "value": text}
    return {"kind": "text", "value": text}


def _excel_cell(field: str, row_number: int) -> str:
    column_number = CANONICAL_COLUMNS.index(field) + 1
    letters = ""
    while column_number:
        column_number, remainder = divmod(column_number - 1, 26)
        letters = chr(65 + remainder) + letters
    return f"{letters}{row_number}"


def _canonical_sample_rows(frame: pl.DataFrame) -> list[dict[str, str]]:
    return [
        {
            field: _material_value(row.get(field), field)["value"]
            for field in CANONICAL_COLUMNS
        }
        for row in frame.iter_rows(named=True)
    ]


def _verify_native_sample_outputs(
    output_dir: Path,
    selected: pl.DataFrame,
    *,
    worksheet: str,
) -> None:
    """Verify every material value in both native sample outputs."""

    expected = _canonical_sample_rows(selected.select(CANONICAL_COLUMNS))
    csv_path = output_dir / "journal_sample.csv"
    csv_frame = pl.read_csv(csv_path, infer_schema=False)
    if csv_frame.columns != CANONICAL_COLUMNS:
        raise ValueError("Sample CSV columns do not close to the canonical contract.")
    if _canonical_sample_rows(csv_frame) != expected:
        raise ValueError("Sample CSV values do not close to the prepared selection.")

    workbook = openpyxl.load_workbook(
        output_dir / "journal_sample.xlsx",
        read_only=True,
        data_only=False,
    )
    try:
        if workbook.sheetnames != [worksheet]:
            raise ValueError("Sample XLSX worksheet set does not close.")
        sheet = workbook[worksheet]
        headers = [
            (
                ""
                if sheet.cell(row=1, column=index).value is None
                else str(sheet.cell(row=1, column=index).value)
            )
            for index in range(1, len(CANONICAL_COLUMNS) + 1)
        ]
        if headers != CANONICAL_COLUMNS:
            raise ValueError("Sample XLSX headers do not close.")
        if sheet.max_row != len(expected) + 1:
            raise ValueError("Sample XLSX row count does not close.")
        for output_index, expected_row in enumerate(expected, start=2):
            for column_index, field in enumerate(CANONICAL_COLUMNS, start=1):
                actual = _material_value(
                    sheet.cell(row=output_index, column=column_index).value,
                    field,
                )["value"]
                if actual != expected_row[field]:
                    raise ValueError(
                        "Sample XLSX material value does not close at "
                        f"{worksheet}!{_excel_cell(field, output_index)}."
                    )
    finally:
        workbook.close()


def _build_sample_material_value_ledger(
    output_dir: Path,
    selected: pl.DataFrame,
    *,
    worksheet: str,
) -> dict[str, Any]:
    """Build all-row prepared-to-CSV/XLSX material-value address closure."""

    _verify_native_sample_outputs(output_dir, selected, worksheet=worksheet)
    entries: list[dict[str, Any]] = []
    for sample_index, row in enumerate(selected.iter_rows(named=True), start=1):
        prepared_row_number = int(row["__prepared_row_number"]) + 1
        output_row_number = sample_index + 1
        identity = {
            field: _material_value(row.get(field), field)["value"]
            for field in (
                "source_file",
                "source_sheet",
                "source_page",
                "source_row",
            )
        }
        monetary_context = {
            "currency": _material_value(row.get("currency"), "currency")["value"],
            "unit": _material_value(row.get("unit"), "unit")["value"],
            "reported_increment": _material_value(
                row.get("reported_increment"),
                "reported_increment",
            )["value"],
        }
        if (
            re.fullmatch(r"[A-Z]{3}", monetary_context["currency"]) is None
            or monetary_context["unit"] != "currency"
            or parse_canonical_decimal(
                monetary_context["reported_increment"],
                label=f"sample row {sample_index} reported_increment",
            )
            <= ZERO
        ):
            raise ValueError(
                f"Sample row {sample_index} has no closed monetary context."
            )
        for field in SAMPLE_MATERIAL_FIELDS:
            canonical = _material_value(row.get(field), field)
            entries.append(
                {
                    "evidence_id": f"sample.{sample_index}.{field}",
                    "sample_row_number": sample_index,
                    "field": field,
                    "value_kind": canonical["kind"],
                    "canonical_value": canonical["value"],
                    "monetary_context": monetary_context,
                    "prepared": {
                        "artifact_ref": "prepared.normalized_journal",
                        "locator": (f"row={prepared_row_number};column={field}"),
                        "prepared_row_number": prepared_row_number,
                        "identity": identity,
                        "value": canonical["value"],
                    },
                    "outputs": [
                        {
                            "artifact_ref": "prepared.journal_sample_csv",
                            "locator": (f"row={output_row_number};column={field}"),
                            "value": canonical["value"],
                        },
                        {
                            "artifact_ref": "output.journal_sample_xlsx",
                            "locator": (
                                f"{worksheet}!{_excel_cell(field, output_row_number)}"
                            ),
                            "value": canonical["value"],
                        },
                    ],
                }
            )
    content = {
        "schema_version": SAMPLE_MATERIAL_LEDGER_SCHEMA_VERSION,
        "ledger_id": "journal_sampling.sample_material_values",
        "row_count": selected.height,
        "material_fields": list(SAMPLE_MATERIAL_FIELDS),
        "entries": entries,
    }
    return {**content, "content_sha256": canonical_json_sha256(content)}


def validate_sample_material_value_ledger(
    output_dir: Path,
    normalized_csv: Path | None = None,
) -> dict[str, Any]:
    """Replay every prepared, CSV, and XLSX address in the sample ledger."""

    unresolved_output_dir = output_dir.expanduser()
    if unresolved_output_dir.is_symlink():
        raise ValueError("Sample output directory cannot be a symlink.")
    output_dir = unresolved_output_dir.resolve()
    payload = read_json(output_dir / "sample_material_value_ledger.json")
    recorded_digest = payload.pop("content_sha256", None)
    if not isinstance(recorded_digest, str) or recorded_digest != canonical_json_sha256(
        payload
    ):
        raise ValueError("Sample material-value ledger content hash is stale.")
    if payload.get("schema_version") != SAMPLE_MATERIAL_LEDGER_SCHEMA_VERSION:
        raise ValueError("Unsupported sample material-value ledger schema.")
    if payload.get("material_fields") != list(SAMPLE_MATERIAL_FIELDS):
        raise ValueError("Sample material fields do not close.")
    entries = payload.get("entries")
    row_count = payload.get("row_count")
    if (
        not isinstance(entries, list)
        or not isinstance(row_count, int)
        or isinstance(row_count, bool)
        or row_count <= 0
        or len(entries) != row_count * len(SAMPLE_MATERIAL_FIELDS)
    ):
        raise ValueError("Sample material-value ledger cardinality does not close.")
    if normalized_csv is None:
        audit = read_json(output_dir / "sampling_audit.json")
        run_intake = read_json(output_dir / "run_intake.json")
        normalized_value = audit.get("normalized_csv")
        if not isinstance(normalized_value, str) or not normalized_value:
            raise ValueError("Sampling audit is missing the normalized CSV path.")
        normalized_csv = _resolve_run_intake_reference(
            output_dir,
            run_intake,
            normalized_value,
        )
    normalized_frame = pl.read_csv(normalized_csv, infer_schema=False).with_row_index(
        "__prepared_row_number",
        offset=1,
    )
    prepared_indexes: list[int] = []
    for sample_index in range(row_count):
        entry_index = sample_index * len(SAMPLE_MATERIAL_FIELDS)
        entry = entries[entry_index]
        if not isinstance(entry, dict):
            raise ValueError("Sample material-value ledger entry is malformed.")
        prepared = entry.get("prepared")
        if not isinstance(prepared, dict):
            raise ValueError("Sample material-value prepared locator is malformed.")
        prepared_row_number = prepared.get("prepared_row_number")
        if (
            not isinstance(prepared_row_number, int)
            or isinstance(prepared_row_number, bool)
            or prepared_row_number <= 1
        ):
            raise ValueError("Sample material prepared row is invalid.")
        prepared_indexes.append(prepared_row_number - 1)
    if len(prepared_indexes) != len(set(prepared_indexes)):
        raise ValueError("Sample material prepared rows must be unique.")
    by_index = {
        int(row["__prepared_row_number"]): row
        for row in normalized_frame.iter_rows(named=True)
    }
    if any(index not in by_index for index in prepared_indexes):
        raise ValueError("Sample material ledger references an unknown prepared row.")
    selected = pl.DataFrame(
        [by_index[index] for index in prepared_indexes],
        schema=normalized_frame.schema,
        strict=False,
    )
    expected = _build_sample_material_value_ledger(
        output_dir,
        selected,
        worksheet=workbook_sheet_name(
            read_json(output_dir / "sampling_audit.json").get("language")
        ),
    )
    actual = {**payload, "content_sha256": recorded_digest}
    if actual != expected:
        raise ValueError("Sample material-value ledger does not replay exactly.")
    return actual


def _build_sample_reproducibility(
    *,
    normalized_receipt: dict[str, Any],
    method: str,
    size: int,
    group_column: str,
    include_accounts: Sequence[str],
    exclude_accounts: Sequence[str],
    date_start: str | None,
    date_end: str | None,
    min_abs: str | None,
    keyword: str | None,
    population_size_before: int,
    population_size_after: int,
    selected: pl.DataFrame,
    sample_csv: Path,
) -> dict[str, Any]:
    """Build the run-independent deterministic replay surface."""

    _, sample_csv_sha256 = file_snapshot(sample_csv)
    _, implementation_sha256 = file_snapshot(Path(__file__))
    rows = _canonical_sample_rows(selected.select(CANONICAL_COLUMNS))
    content = {
        "schema_version": SAMPLE_REPRODUCIBILITY_SCHEMA_VERSION,
        "workflow_version": SAMPLE_ASSURANCE_VERSION,
        "normalization_receipt": {
            "sha256": normalized_receipt["sha256"],
            "byte_count": normalized_receipt["byte_count"],
        },
        "implementation_sha256": implementation_sha256,
        "sampling_contract": {
            "method": method,
            "seed": 42 if method == "random" else None,
            "requested_size": size,
            "group_column": group_column,
            "filters": {
                "include_accounts": list(include_accounts),
                "exclude_accounts": list(exclude_accounts),
                "date_start": date_start,
                "date_end": date_end,
                "min_abs": min_abs,
                "keyword": keyword,
            },
        },
        "population_size_before_filters": population_size_before,
        "population_size_after_filters": population_size_after,
        "selected_prepared_row_numbers": [
            int(value)
            for value in selected.get_column("__prepared_row_number").to_list()
        ],
        "sample_rows_sha256": canonical_json_sha256(
            {"columns": CANONICAL_COLUMNS, "rows": rows}
        ),
        "sample_csv_sha256": sample_csv_sha256,
    }
    return {**content, "content_sha256": canonical_json_sha256(content)}


def _validate_sample_reproducibility(
    output_dir: Path,
    normalized_csv: Path,
) -> dict[str, Any]:
    payload = read_json(output_dir / "sample_reproducibility.json")
    recorded_digest = payload.pop("content_sha256", None)
    if not isinstance(recorded_digest, str) or recorded_digest != canonical_json_sha256(
        payload
    ):
        raise ValueError("Sample reproducibility content hash is stale.")
    if payload.get("schema_version") != SAMPLE_REPRODUCIBILITY_SCHEMA_VERSION:
        raise ValueError("Unsupported sample reproducibility schema.")
    contract = payload.get("sampling_contract")
    if not isinstance(contract, dict):
        raise ValueError("Sample reproducibility contract is malformed.")
    filters = contract.get("filters")
    if not isinstance(filters, dict):
        raise ValueError("Sample reproducibility filters are malformed.")
    normalized_receipt = artifact_receipt(
        normalized_csv.parent,
        normalized_csv,
        artifact_id="prepared.normalized_journal",
        root_id="normalization",
        role="prepared",
        media_type="text/csv",
    )
    frame = pl.read_csv(normalized_csv, infer_schema=False).with_row_index(
        "__prepared_row_number",
        offset=1,
    )
    min_abs = filters.get("min_abs")
    population = _apply_filters(
        frame,
        include_accounts=filters.get("include_accounts") or (),
        exclude_accounts=filters.get("exclude_accounts") or (),
        date_start=filters.get("date_start"),
        date_end=filters.get("date_end"),
        min_abs=min_abs,
        keyword=filters.get("keyword"),
    )
    selected = _select_sample(
        population,
        method=str(contract.get("method") or ""),
        size=int(contract.get("requested_size") or 0),
        group_column=str(contract.get("group_column") or ""),
    )
    expected = _build_sample_reproducibility(
        normalized_receipt=normalized_receipt,
        method=str(contract["method"]),
        size=int(contract["requested_size"]),
        group_column=str(contract["group_column"]),
        include_accounts=filters.get("include_accounts") or (),
        exclude_accounts=filters.get("exclude_accounts") or (),
        date_start=filters.get("date_start"),
        date_end=filters.get("date_end"),
        min_abs=min_abs,
        keyword=filters.get("keyword"),
        population_size_before=frame.height,
        population_size_after=population.height,
        selected=selected,
        sample_csv=output_dir / "journal_sample.csv",
    )
    actual = {**payload, "content_sha256": recorded_digest}
    if actual != expected:
        raise ValueError("Sample reproducibility replay does not close.")
    return actual


def _validated_normalization_envelope(
    normalized_csv: Path,
    diagnostics_path: Path,
) -> tuple[dict[str, Any], dict[str, Any]]:
    diagnostics = read_json(diagnostics_path)
    source_root_reference = diagnostics.get("source_root")
    envelope_name = diagnostics.get("assurance_envelope")
    if (
        not isinstance(source_root_reference, str)
        or not source_root_reference
        or not isinstance(envelope_name, str)
        or not envelope_name
    ):
        raise ValueError("Normalization assurance context is incomplete.")
    _, recipe_receipt, _ = _validated_normalization_recipe(
        normalized_csv,
        diagnostics,
    )
    source_root = _resolve_normalization_reference(
        normalized_csv,
        diagnostics,
        source_root_reference,
        label="Normalization source root",
    )
    envelope = validate_assurance_envelope(
        read_json(normalized_csv.parent / envelope_name),
        artifact_roots={
            "source": source_root,
            "normalization": normalized_csv.parent,
            **_implementation_artifact_roots(),
        },
    )
    _validate_exact_implementation_receipts(envelope)
    envelope_receipts = {
        receipt["artifact_id"]: receipt for receipt in envelope["artifact_receipts"]
    }
    if envelope_receipts.get(recipe_receipt["artifact_id"]) != recipe_receipt:
        raise ValueError(
            "Normalization assurance does not bind the retained reviewed recipe."
        )
    return diagnostics, envelope


def _build_sample_assurance(
    output_dir: Path,
    *,
    normalized_csv: Path,
    diagnostics_path: Path,
    run_id: str,
) -> tuple[dict[str, Any], dict[str, Any]]:
    """Build a replayable sample-stage envelope from current artifact bytes."""

    diagnostics, upstream = _validated_normalization_envelope(
        normalized_csv,
        diagnostics_path,
    )
    normalized_receipt = validate_artifact_receipt(
        {"normalization": normalized_csv.parent},
        diagnostics["normalized_csv_receipt"],
    )
    implementation_receipts = _implementation_receipts()
    sample_receipts = [
        artifact_receipt(
            output_dir,
            output_dir / "journal_sample.csv",
            artifact_id="prepared.journal_sample_csv",
            root_id="sample",
            role="prepared",
            media_type="text/csv",
        ),
        artifact_receipt(
            output_dir,
            output_dir / "journal_sample.xlsx",
            artifact_id="output.journal_sample_xlsx",
            root_id="sample",
            role="output",
            media_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        ),
        artifact_receipt(
            output_dir,
            output_dir / "sampling_audit.json",
            artifact_id="workpaper.sampling_audit",
            root_id="sample",
            role="workpaper",
            media_type="application/json",
        ),
        artifact_receipt(
            output_dir,
            output_dir / "sample_reproducibility.json",
            artifact_id="workpaper.sample_reproducibility",
            root_id="sample",
            role="workpaper",
            media_type="application/json",
        ),
        artifact_receipt(
            output_dir,
            output_dir / "sample_material_value_ledger.json",
            artifact_id="workpaper.sample_material_value_ledger",
            root_id="sample",
            role="workpaper",
            media_type="application/json",
        ),
    ]
    source_evidence_refs = [
        qualification["qualification_id"]
        for qualification in upstream["source_qualifications"]
    ]
    preparation_evidence_refs = [
        normalized_receipt["artifact_id"],
        "prepared.journal_sample_csv",
        "output.journal_sample_xlsx",
        "workpaper.sample_reproducibility",
        "workpaper.sample_material_value_ledger",
    ]
    gate_register = build_gate_register(
        {
            "source": {
                "status": "passed",
                "evidence_refs": source_evidence_refs,
                "limitations": [],
            },
            "preparation": {
                "status": "passed",
                "evidence_refs": preparation_evidence_refs,
                "limitations": [],
            },
            "reconciliation": {
                "status": "not_applicable",
                "evidence_refs": [],
                "limitations": [],
            },
            "semantic_review": {
                "status": "not_assessed",
                "evidence_refs": [],
                "limitations": [
                    "Professional review of sampling basis and sufficiency remains pending."
                ],
            },
            "reporting": {
                "status": "blocked",
                "evidence_refs": [],
                "limitations": [
                    "Reporting remains blocked until professional sample review."
                ],
            },
            "publication": {
                "status": "withheld",
                "evidence_refs": [],
                "limitations": [
                    "Publication authority is outside deterministic sample preparation."
                ],
            },
        }
    )
    write_json(output_dir / "sample_assurance_gates.json", gate_register)
    gate_receipt = artifact_receipt(
        output_dir,
        output_dir / "sample_assurance_gates.json",
        artifact_id="workpaper.sample_assurance_gates",
        root_id="sample",
        role="workpaper",
        media_type="application/json",
    )
    source_receipts = [
        receipt
        for receipt in upstream["artifact_receipts"]
        if receipt["role"] == "source"
    ]
    reviewed_recipe_receipts = [
        receipt
        for receipt in upstream["artifact_receipts"]
        if receipt["role"] == "reviewed_recipe"
    ]
    if reviewed_recipe_receipts != [diagnostics["normalization_recipe_receipt"]]:
        raise ValueError("Upstream reviewed recipe receipt is not exact.")
    envelope = build_assurance_envelope(
        run_id=run_id,
        workflow_id="journal-sampling-sample",
        workflow_version=SAMPLE_ASSURANCE_VERSION,
        artifact_receipts=[
            *source_receipts,
            *reviewed_recipe_receipts,
            normalized_receipt,
            *implementation_receipts,
            *sample_receipts,
            gate_receipt,
        ],
        implementation_artifact_refs=[
            receipt["artifact_id"] for receipt in implementation_receipts
        ],
        reviewed_decisions=upstream["reviewed_decisions"],
        source_qualifications=upstream["source_qualifications"],
        allocation_ledgers=[],
        numeric_evidence_ledgers=[],
        gate_register=gate_register,
        limitations=[
            "The deterministic sample and native-value closure do not establish audit sufficiency or a professional conclusion."
        ],
        artifact_roots={
            "source": _resolve_normalization_reference(
                normalized_csv,
                diagnostics,
                diagnostics["source_root"],
                label="Normalization source root",
            ),
            "normalization": normalized_csv.parent,
            "sample": output_dir,
            **_implementation_artifact_roots(),
        },
    )
    write_json(output_dir / "sample_assurance_envelope.json", envelope)
    return gate_register, envelope


def _output_set_receipt(
    output_dir: Path,
    path: str,
    index: int,
) -> dict[str, Any]:
    return artifact_receipt(
        output_dir,
        output_dir / path,
        artifact_id=f"sample.output_set.{index:02d}",
        root_id="sample_output",
        role="output_set_member",
    )


def _canonical_output_relative_path(value: object, *, label: str) -> str:
    if not isinstance(value, str) or not value or value != value.strip():
        raise ValueError(f"{label} must be a non-empty canonical path.")
    if "\\" in value:
        raise ValueError(f"{label} must use canonical separators.")
    candidate = Path(value)
    if (
        candidate.is_absolute()
        or ".." in candidate.parts
        or candidate.as_posix() != value
        or value in {".", SAMPLE_ASSURANCE_HISTORY_DIR}
    ):
        raise ValueError(f"{label} must stay inside the sample output.")
    return value


def _scan_sample_tree(output_dir: Path) -> tuple[list[str], list[str], dict[str, str]]:
    """Return exact regular files, directories, and modes without following links."""

    files: list[str] = []
    directories: list[str] = []
    modes: dict[str, str] = {}
    pending = [output_dir]
    while pending:
        current = pending.pop()
        with os.scandir(current) as entries:
            for entry in sorted(entries, key=lambda item: item.name):
                path = Path(entry.path)
                relative = path.relative_to(output_dir).as_posix()
                observed = path.lstat()
                if stat.S_ISLNK(observed.st_mode):
                    raise ValueError("Sample output set cannot contain symlinks.")
                if stat.S_ISDIR(observed.st_mode):
                    directories.append(relative)
                    modes[relative] = f"{stat.S_IMODE(observed.st_mode):04o}"
                    pending.append(path)
                    continue
                if not stat.S_ISREG(observed.st_mode):
                    raise ValueError("Sample output set cannot contain special files.")
                if observed.st_nlink != 1:
                    raise ValueError(
                        "Sample output set cannot contain hard-linked files."
                    )
                files.append(relative)
                modes[relative] = f"{stat.S_IMODE(observed.st_mode):04o}"
    return sorted(files), sorted(directories), modes


def _manifest_content(path: Path) -> tuple[dict[str, Any], str]:
    try:
        manifest = read_json(path)
    except PermissionError as exc:
        raise ValueError("Sample output-set manifest is not safely readable.") from exc
    recorded_digest = manifest.pop("content_sha256", None)
    if not isinstance(recorded_digest, str) or recorded_digest != canonical_json_sha256(
        manifest
    ):
        raise ValueError("Sample output-set manifest content hash is stale.")
    if manifest.get("schema_version") != SAMPLE_OUTPUT_SET_SCHEMA_VERSION:
        raise ValueError("Unsupported sample output-set schema.")
    if manifest.get("bootstrap_path") != SAMPLE_OUTPUT_SET_PATH:
        raise ValueError("Sample output-set bootstrap path is invalid.")
    if set(manifest) != {
        "schema_version",
        "boundary",
        "bootstrap_path",
        "bootstrap_mode",
        "root_mode",
        "stage",
        "payload_paths",
        "physical_paths",
        "directory_paths",
        "directory_modes",
        "modes",
        "receipts",
    }:
        raise ValueError("Sample output-set manifest fields are not exact.")
    return manifest, recorded_digest


def _validated_stage(manifest: dict[str, Any]) -> dict[str, Any]:
    stage = manifest.get("stage")
    if not isinstance(stage, dict) or set(stage) != {
        "index",
        "kind",
        "predecessor",
    }:
        raise ValueError("Sample output-set stage declaration is not exact.")
    index = stage.get("index")
    kind = stage.get("kind")
    if (
        not isinstance(index, int)
        or isinstance(index, bool)
        or index < 0
        or kind not in {"initial", "save", "apply"}
        or (index == 0) != (kind == "initial")
    ):
        raise ValueError("Sample output-set stage identity is invalid.")
    expected_boundary = (
        "sample_stage_finalization_pre_review"
        if kind == "initial"
        else f"sample_review_{kind}_successor"
    )
    if "boundary" in manifest and manifest.get("boundary") != expected_boundary:
        raise ValueError("Sample output-set boundary is stale.")
    predecessor = stage.get("predecessor")
    if index == 0:
        if predecessor is not None:
            raise ValueError("Initial sample stage cannot declare a predecessor.")
    else:
        if not isinstance(predecessor, dict) or set(predecessor) != {
            "stage_index",
            "stage_kind",
            "archive_dir",
            "manifest_sha256",
        }:
            raise ValueError("Sample successor predecessor declaration is not exact.")
        if (
            predecessor.get("stage_index") != index - 1
            or predecessor.get("stage_kind") not in {"initial", "save", "apply"}
            or not isinstance(predecessor.get("manifest_sha256"), str)
            or re.fullmatch(r"[0-9a-f]{64}", predecessor["manifest_sha256"]) is None
        ):
            raise ValueError("Sample successor predecessor identity is invalid.")
        expected_archive = (
            f"{SAMPLE_ASSURANCE_HISTORY_DIR}/{index - 1:03d}_"
            f"{predecessor['stage_kind']}"
        )
        if predecessor.get("archive_dir") != expected_archive:
            raise ValueError("Sample successor archive directory is not canonical.")
    return stage


def _declared_review_extra_paths(output_dir: Path) -> list[str]:
    applied_path = output_dir / SAMPLE_APPLIED_DECISIONS_PATH
    if not applied_path.exists():
        return []
    applied = read_json(applied_path)
    values: list[object] = []
    for field in ("revision_paths", "original_backup_paths"):
        field_value = applied.get(field)
        if not isinstance(field_value, list):
            raise ValueError(f"Applied decisions {field} must be a list.")
        values.extend(field_value)
    paths = [
        _canonical_output_relative_path(value, label="review artifact path")
        for value in values
    ]
    if len(paths) != len(set(paths)) or any(
        not path.startswith("revisions/") for path in paths
    ):
        raise ValueError("Applied review artifact paths are not exact.")
    return sorted(paths)


def _current_operational_payload_paths(output_dir: Path) -> list[str]:
    paths = list(SAMPLE_OUTPUT_PAYLOAD_PATHS)
    if (output_dir / SAMPLE_APPLIED_DECISIONS_PATH).exists():
        paths.append(SAMPLE_APPLIED_DECISIONS_PATH)
    paths.extend(_declared_review_extra_paths(output_dir))
    if len(paths) != len(set(paths)):
        raise ValueError("Current sample output paths are not unique.")
    return paths


def _parent_directories(paths: Sequence[str]) -> set[str]:
    directories: set[str] = set()
    for relative in paths:
        parent = Path(relative).parent
        while parent.as_posix() not in {"", "."}:
            directories.add(parent.as_posix())
            parent = parent.parent
    return directories


def _validate_archived_manifest_receipts(
    output_dir: Path,
    archive_dir: str,
    archived_manifest: dict[str, Any],
) -> None:
    payload_paths = archived_manifest.get("payload_paths")
    receipts = archived_manifest.get("receipts")
    if not isinstance(payload_paths, list) or not isinstance(receipts, list):
        raise ValueError("Archived sample output receipts are malformed.")
    if [receipt.get("path") for receipt in receipts if isinstance(receipt, dict)] != (
        payload_paths
    ):
        raise ValueError("Archived sample output receipt ordering is stale.")
    archive_root = output_dir / archive_dir
    for receipt in receipts:
        if not isinstance(receipt, dict):
            raise ValueError("Archived sample output receipt is malformed.")
        receipt_path = str(receipt.get("path") or "")
        root = (
            output_dir
            if receipt_path.startswith(f"{SAMPLE_ASSURANCE_HISTORY_DIR}/")
            else archive_root
        )
        validate_artifact_receipt({"sample_output": root}, receipt)


def _predecessor_archive_contract(
    output_dir: Path,
    manifest: dict[str, Any],
) -> tuple[set[str], set[str]]:
    expected_files: set[str] = set()
    expected_directories: set[str] = set()
    current = manifest
    current_stage = _validated_stage(current)
    seen_indexes: set[int] = set()
    while current_stage["index"] > 0:
        predecessor = current_stage["predecessor"]
        if not isinstance(predecessor, dict):
            raise ValueError("Sample predecessor chain is incomplete.")
        predecessor_index = int(predecessor["stage_index"])
        if predecessor_index in seen_indexes:
            raise ValueError("Sample predecessor chain contains a cycle.")
        seen_indexes.add(predecessor_index)
        archive_dir = str(predecessor["archive_dir"])
        archive_root = output_dir / archive_dir
        archived_path = archive_root / SAMPLE_OUTPUT_SET_PATH
        archived_manifest, archived_digest = _manifest_content(archived_path)
        archived_stage = _validated_stage(archived_manifest)
        if (
            archived_digest != predecessor["manifest_sha256"]
            or archived_stage["index"] != predecessor_index
            or archived_stage["kind"] != predecessor["stage_kind"]
        ):
            raise ValueError("Sample predecessor manifest binding is stale.")
        archived_physical = archived_manifest.get("physical_paths")
        if not isinstance(archived_physical, list):
            raise ValueError("Archived sample physical paths are malformed.")
        operational = sorted(
            _canonical_output_relative_path(path, label="archived physical path")
            for path in archived_physical
            if isinstance(path, str)
            and not path.startswith(f"{SAMPLE_ASSURANCE_HISTORY_DIR}/")
        )
        archive_files, archive_directories, archive_modes = _scan_sample_tree(
            archive_root
        )
        if archive_files != operational:
            raise ValueError("Archived predecessor file set does not close.")
        expected_archive_directories = sorted(_parent_directories(operational))
        if archive_directories != expected_archive_directories:
            raise ValueError("Archived predecessor directory set does not close.")
        archived_modes = archived_manifest.get("modes")
        archived_directory_modes = archived_manifest.get("directory_modes")
        if not isinstance(archived_modes, dict) or not isinstance(
            archived_directory_modes, dict
        ):
            raise ValueError("Archived predecessor modes are missing.")
        for relative in operational:
            if relative == SAMPLE_OUTPUT_SET_PATH:
                continue
            if archive_modes.get(relative) != archived_modes.get(relative):
                raise ValueError("Archived predecessor file mode is stale.")
        for relative in expected_archive_directories:
            if archive_modes.get(relative) != archived_directory_modes.get(relative):
                raise ValueError("Archived predecessor directory mode is stale.")
        if archive_modes.get(SAMPLE_OUTPUT_SET_PATH) != archived_manifest.get(
            "bootstrap_mode"
        ):
            raise ValueError("Archived predecessor bootstrap mode is stale.")
        if (
            f"{stat.S_IMODE(archive_root.lstat().st_mode):04o}"
            != archived_manifest.get("root_mode")
        ):
            raise ValueError("Archived predecessor root mode is stale.")
        _validate_archived_manifest_receipts(
            output_dir,
            archive_dir,
            archived_manifest,
        )
        expected_files.update(f"{archive_dir}/{path}" for path in operational)
        expected_directories.add(SAMPLE_ASSURANCE_HISTORY_DIR)
        expected_directories.add(archive_dir)
        expected_directories.update(
            f"{archive_dir}/{path}" for path in expected_archive_directories
        )
        current = archived_manifest
        current_stage = archived_stage
    if len(seen_indexes) != int(_validated_stage(manifest)["index"]):
        raise ValueError("Sample predecessor chain cardinality does not close.")
    return expected_files, expected_directories


def _ordered_sample_payload_paths(
    output_dir: Path,
    *,
    history_files: set[str],
) -> list[str]:
    return [
        *_current_operational_payload_paths(output_dir),
        *sorted(history_files),
    ]


def _write_sample_output_set(
    output_dir: Path,
    *,
    stage: dict[str, Any] | None = None,
) -> dict[str, Any]:
    stage_value = stage or {"index": 0, "kind": "initial", "predecessor": None}
    _validated_stage({"stage": stage_value})
    history_files: set[str] = set()
    history_directories: set[str] = set()
    if stage_value["index"] > 0:
        history_files, history_directories = _predecessor_archive_contract(
            output_dir,
            {"stage": stage_value},
        )
    payload_paths = _ordered_sample_payload_paths(
        output_dir,
        history_files=history_files,
    )
    expected_files = sorted([*payload_paths, SAMPLE_OUTPUT_SET_PATH])
    expected_directories = sorted(
        {
            *_parent_directories(_current_operational_payload_paths(output_dir)),
            *history_directories,
        }
    )
    actual_files, actual_directories, modes = _scan_sample_tree(output_dir)
    actual_without_bootstrap = [
        path for path in actual_files if path != SAMPLE_OUTPUT_SET_PATH
    ]
    if actual_without_bootstrap != sorted(payload_paths):
        missing = sorted(set(payload_paths) - set(actual_without_bootstrap))
        unexpected = sorted(set(actual_without_bootstrap) - set(payload_paths))
        raise ValueError(
            "Sample output set cannot be sealed; "
            f"missing={missing}, unexpected={unexpected}."
        )
    if actual_directories != expected_directories:
        raise ValueError(
            "Sample output directory set cannot be sealed; "
            f"missing={sorted(set(expected_directories) - set(actual_directories))}, "
            f"unexpected={sorted(set(actual_directories) - set(expected_directories))}."
        )
    receipts = [
        _output_set_receipt(output_dir, path, index)
        for index, path in enumerate(payload_paths, start=1)
    ]
    content = {
        "schema_version": SAMPLE_OUTPUT_SET_SCHEMA_VERSION,
        "boundary": (
            "sample_stage_finalization_pre_review"
            if stage_value["kind"] == "initial"
            else f"sample_review_{stage_value['kind']}_successor"
        ),
        "bootstrap_path": SAMPLE_OUTPUT_SET_PATH,
        "bootstrap_mode": f"{SAMPLE_OUTPUT_SET_MODE:04o}",
        "root_mode": f"{stat.S_IMODE(output_dir.lstat().st_mode):04o}",
        "stage": stage_value,
        "payload_paths": payload_paths,
        "physical_paths": expected_files,
        "directory_paths": expected_directories,
        "directory_modes": {path: modes[path] for path in expected_directories},
        "modes": {path: modes[path] for path in payload_paths},
        "receipts": receipts,
    }
    manifest = {**content, "content_sha256": canonical_json_sha256(content)}
    write_json(output_dir / SAMPLE_OUTPUT_SET_PATH, manifest)
    (output_dir / SAMPLE_OUTPUT_SET_PATH).chmod(SAMPLE_OUTPUT_SET_MODE)
    return validate_sample_output_set(output_dir)


def validate_sample_output_set(output_dir: Path) -> dict[str, Any]:
    """Reject any non-canonical file, directory, link, mode, or stale receipt."""

    unresolved_output_dir = output_dir.expanduser()
    if unresolved_output_dir.is_symlink():
        raise ValueError("Sample output directory cannot be a symlink.")
    output_dir = unresolved_output_dir.resolve()
    manifest, recorded_digest = _manifest_content(output_dir / SAMPLE_OUTPUT_SET_PATH)
    stage = _validated_stage(manifest)
    history_files, history_directories = _predecessor_archive_contract(
        output_dir,
        manifest,
    )
    expected_payload = _ordered_sample_payload_paths(
        output_dir,
        history_files=history_files,
    )
    expected_physical = sorted([*expected_payload, SAMPLE_OUTPUT_SET_PATH])
    expected_directories = sorted(
        {
            *_parent_directories(_current_operational_payload_paths(output_dir)),
            *history_directories,
        }
    )
    payload_paths = manifest.get("payload_paths")
    physical_paths = manifest.get("physical_paths")
    directory_paths = manifest.get("directory_paths")
    receipts = manifest.get("receipts")
    modes = manifest.get("modes")
    directory_modes = manifest.get("directory_modes")
    if (
        payload_paths != expected_payload
        or physical_paths != expected_physical
        or directory_paths != expected_directories
        or not isinstance(receipts, list)
        or len(receipts) != len(expected_payload)
        or not isinstance(modes, dict)
        or set(modes) != set(expected_payload)
        or not isinstance(directory_modes, dict)
        or set(directory_modes) != set(expected_directories)
    ):
        raise ValueError("Sample output-set declaration is not exact.")
    actual_files, actual_directories, actual_modes = _scan_sample_tree(output_dir)
    if actual_files != expected_physical:
        missing = sorted(set(expected_physical) - set(actual_files))
        unexpected = sorted(set(actual_files) - set(expected_physical))
        raise ValueError(
            "Sample physical output set does not close; "
            f"missing={missing}, unexpected={unexpected}."
        )
    if actual_directories != expected_directories:
        raise ValueError(
            "Sample physical directory set does not close; "
            f"missing={sorted(set(expected_directories) - set(actual_directories))}, "
            f"unexpected={sorted(set(actual_directories) - set(expected_directories))}."
        )
    if any(actual_modes.get(path) != modes.get(path) for path in expected_payload):
        raise ValueError("Sample output modes do not close.")
    if any(
        actual_modes.get(path) != directory_modes.get(path)
        for path in expected_directories
    ):
        raise ValueError("Sample output directory modes do not close.")
    if actual_modes.get(SAMPLE_OUTPUT_SET_PATH) != manifest.get("bootstrap_mode"):
        raise ValueError("Sample output-set bootstrap mode does not close.")
    if f"{stat.S_IMODE(output_dir.lstat().st_mode):04o}" != manifest.get("root_mode"):
        raise ValueError("Sample output root mode does not close.")
    validated_receipts = [
        validate_artifact_receipt({"sample_output": output_dir}, receipt)
        for receipt in receipts
        if isinstance(receipt, dict)
    ]
    if (
        len(validated_receipts) != len(receipts)
        or [receipt["path"] for receipt in validated_receipts] != expected_payload
    ):
        raise ValueError("Sample output receipts do not close to declared paths.")
    return {
        **manifest,
        "stage": stage,
        "content_sha256": recorded_digest,
    }


def _archive_current_sample_stage(
    output_dir: Path,
    manifest: dict[str, Any],
) -> str:
    stage = _validated_stage(manifest)
    archive_dir = f"{SAMPLE_ASSURANCE_HISTORY_DIR}/{stage['index']:03d}_{stage['kind']}"
    archive_root = output_dir / archive_dir
    if archive_root.exists() or archive_root.is_symlink():
        raise ValueError("Sample predecessor archive already exists.")
    operational_files = [
        path
        for path in manifest["physical_paths"]
        if not path.startswith(f"{SAMPLE_ASSURANCE_HISTORY_DIR}/")
    ]
    operational_directories = sorted(
        _parent_directories(operational_files),
        key=lambda value: (len(Path(value).parts), value),
    )
    history_root = output_dir / SAMPLE_ASSURANCE_HISTORY_DIR
    history_root.mkdir(mode=0o700, exist_ok=True)
    history_root.chmod(0o700)
    predecessor_root_mode = int(str(manifest["root_mode"]), 8)
    archive_root.mkdir(mode=predecessor_root_mode)
    archive_root.chmod(predecessor_root_mode)
    for relative in operational_directories:
        source = output_dir / relative
        target = archive_root / relative
        observed = source.lstat()
        if not stat.S_ISDIR(observed.st_mode) or stat.S_ISLNK(observed.st_mode):
            raise ValueError("Sample predecessor directory is unsafe.")
        target.mkdir(parents=True, exist_ok=True, mode=0o700)
        target.chmod(stat.S_IMODE(observed.st_mode))
    for relative in operational_files:
        source = output_dir / relative
        _require_ordinary_single_link(
            source,
            label=f"Sample predecessor artifact {relative}",
        )
        target = archive_root / relative
        target.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source, target, follow_symlinks=False)
        _require_ordinary_single_link(
            target,
            label=f"Archived predecessor artifact {relative}",
        )
    return archive_dir


def prepare_sample_review_successor(
    output_dir: Path,
    *,
    kind: str,
) -> dict[str, Any]:
    """Replay and archive one trusted predecessor before a review write."""

    if kind not in {"save", "apply"}:
        raise ValueError("Review successor kind must be save or apply.")
    replay = validate_sample_assurance(output_dir)
    resolved = output_dir.expanduser().resolve()
    manifest = replay["output_set"]
    archive_dir = _archive_current_sample_stage(resolved, manifest)
    return {
        "ok": True,
        "kind": kind,
        "predecessor_stage": manifest["stage"],
        "predecessor_manifest_sha256": manifest["content_sha256"],
        "archive_dir": archive_dir,
    }


def _resolve_run_intake_reference(
    output_dir: Path,
    run_intake: Mapping[str, Any],
    value: object,
) -> Path:
    """Resolve one persisted run-root-relative reference from the current tree."""

    if not isinstance(value, str) or not value.strip():
        raise ValueError("Sample run intake path reference is unavailable.")
    reference = Path(value)
    if reference.is_absolute():
        return reference.expanduser().resolve()
    if (
        run_intake.get("path_reference") != "run_root_relative"
        or ".." in reference.parts
    ):
        raise ValueError("Sample run intake path reference is invalid.")
    candidate = output_dir.expanduser().resolve()
    while True:
        context_path = candidate / "context.json"
        if context_path.is_file() and not context_path.is_symlink():
            return (candidate / reference).resolve()
        if candidate == candidate.parent:
            raise ValueError("Sample customer-run context is unavailable.")
        candidate = candidate.parent


def _review_item_contract(
    output_dir: Path,
    audit: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any], list[dict[str, str]]]:
    """Freshly rederive the bounded review item identities and row payloads."""

    review = read_json(output_dir / "review_payload.json")
    run_intake = read_json(output_dir / "run_intake.json")
    sample_frame = pl.read_csv(output_dir / "journal_sample.csv", infer_schema=False)
    sample_rows = _canonical_sample_rows(sample_frame)
    expected_ids = [
        "sampling-control",
        *(f"sampled-entry-{index}" for index in range(1, len(sample_rows) + 1)),
        "artifact-1",
        "artifact-2",
        "artifact-3",
    ]
    expected_types = [
        "sampling_control",
        *("sampled_entry" for _ in sample_rows),
        "sample_artifact",
        "sample_artifact",
        "review_artifact",
    ]
    expected_outputs = [
        "sampling_audit.json",
        *("journal_sample.csv" for _ in sample_rows),
        "journal_sample.csv",
        "journal_sample.xlsx",
        "sampling_audit.json",
    ]
    items = review.get("items")
    if (
        not isinstance(items, list)
        or review.get("item_count") != len(items)
        or len(items) != len(expected_ids)
        or [item.get("id") for item in items if isinstance(item, dict)] != expected_ids
        or [item.get("item_type") for item in items if isinstance(item, dict)]
        != expected_types
        or [item.get("output_path") for item in items if isinstance(item, dict)]
        != expected_outputs
    ):
        raise ValueError("Sample review item identity or cardinality is stale.")
    for index, row in enumerate(sample_rows, start=1):
        item = items[index]
        data = item.get("data")
        if not isinstance(data, dict):
            raise ValueError("Sampled review item data is malformed.")
        actual_row = {
            field: _material_value(data.get(field), field)["value"]
            for field in CANONICAL_COLUMNS
        }
        if actual_row != row:
            raise ValueError("Sampled review item does not close to the sample row.")
    summary = review.get("summary")
    if not isinstance(summary, dict) or {
        "method": summary.get("method"),
        "requested_size": summary.get("requested_size"),
        "population_size_before_filters": summary.get("population_size_before_filters"),
        "population_size_after_filters": summary.get("population_size_after_filters"),
        "sample_size": summary.get("sample_size"),
        "filters": summary.get("filters"),
        "population_proof": summary.get("population_proof"),
    } != {
        "method": audit.get("method"),
        "requested_size": audit.get("requested_size"),
        "population_size_before_filters": audit.get("population_size_before_filters"),
        "population_size_after_filters": audit.get("population_size_after_filters"),
        "sample_size": audit.get("sample_size"),
        "filters": audit.get("filters"),
        "population_proof": audit.get("population_proof"),
    }:
        raise ValueError("Sample review summary is stale.")
    if (
        review.get("plugin") != "journal-sampling"
        or review.get("workflow") != "journal-sampling"
        or review.get("run_id") != run_intake.get("run_id")
        or review.get("run_id") != audit.get("review_session", {}).get("run_id")
    ):
        raise ValueError("Sample review run identity is stale.")
    assumptions = run_intake.get("assumptions")
    normalized_csv_reference = (
        _resolve_run_intake_reference(
            output_dir,
            run_intake,
            assumptions.get("normalized_csv"),
        ).as_posix()
        if isinstance(assumptions, dict)
        else None
    )
    audit_normalized_csv_reference = _resolve_run_intake_reference(
        output_dir,
        run_intake,
        audit.get("normalized_csv"),
    ).as_posix()
    if not isinstance(assumptions, dict) or {
        "normalized_csv": normalized_csv_reference,
        "method": assumptions.get("method"),
        "requested_size": assumptions.get("requested_size"),
        "group_column": assumptions.get("group_column"),
        "include_accounts": assumptions.get("include_accounts"),
        "exclude_accounts": assumptions.get("exclude_accounts"),
        "date_start": assumptions.get("date_start"),
        "date_end": assumptions.get("date_end"),
        "min_abs": assumptions.get("min_abs"),
        "keyword": assumptions.get("keyword"),
    } != {
        "normalized_csv": audit_normalized_csv_reference,
        "method": audit.get("method"),
        "requested_size": audit.get("requested_size"),
        "group_column": read_json(output_dir / "sample_reproducibility.json")
        .get("sampling_contract", {})
        .get("group_column"),
        "include_accounts": audit.get("filters", {}).get("include_accounts"),
        "exclude_accounts": audit.get("filters", {}).get("exclude_accounts"),
        "date_start": audit.get("filters", {}).get("date_start"),
        "date_end": audit.get("filters", {}).get("date_end"),
        "min_abs": audit.get("filters", {}).get("min_abs"),
        "keyword": audit.get("filters", {}).get("keyword"),
    }:
        raise ValueError("Sample run intake no longer closes to the sampling contract.")
    return run_intake, review, sample_rows


def _iso_datetime(value: object, *, label: str) -> str:
    if not isinstance(value, str) or not value or value != value.strip():
        raise ValueError(f"{label} must be an ISO timestamp.")
    try:
        datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError as exc:
        raise ValueError(f"{label} must be an ISO timestamp.") from exc
    return value


def _compact_review_context_value(value: object) -> str:
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int) and not isinstance(value, bool):
        return str(value)
    return ""


def _expected_followup_context(
    item: dict[str, Any],
    *,
    action: str,
) -> dict[str, str]:
    if action not in {"reject", "mark_unclear", "request_more_documents"}:
        return {}
    data = item.get("data") if isinstance(item.get("data"), dict) else {}
    evidence = [value for value in item.get("evidence", []) if isinstance(value, dict)]
    records = [data, *evidence]
    fields = (
        (
            "owner",
            ("owner", "responsible_party", "assignee", "contact", "client_contact"),
        ),
        ("source_system", ("source_system", "system", "source_system_name")),
        (
            "source_file",
            ("source_file", "filename", "file_name", "source_workbook"),
        ),
        ("source_table", ("source_table", "sheet", "worksheet", "table")),
        ("due_date", ("due_date", "deadline", "response_due_date")),
        ("period", ("period", "tax_period", "fiscal_year", "year")),
        (
            "entity",
            (
                "entity",
                "client",
                "company",
                "account",
                "counterparty",
                "beneficiary",
            ),
        ),
        (
            "record_id",
            (
                "record_id",
                "source_row",
                "movement_number",
                "bank_transaction_id",
                "journal_entry_id",
                "claim_index",
            ),
        ),
        ("amount", ("amount", "amount_abs", "amount_value")),
        ("reason", ("reason", "missing_reason", "blocking_reason", "mismatches")),
        ("priority", ("priority", "severity")),
    )
    context: dict[str, str] = {}
    for target, source_fields in fields:
        for record in records:
            value = next(
                (
                    compact
                    for source in source_fields
                    if (compact := _compact_review_context_value(record.get(source)))
                ),
                "",
            )
            if value:
                context[target] = value
                break
    return context


def _validate_ui_decisions(
    output_dir: Path,
    review: dict[str, Any],
) -> dict[str, Any]:
    ui = read_json(output_dir / "ui_decisions.json")
    decisions = ui.get("decisions")
    items = review["items"]
    item_by_id = {item["id"]: item for item in items}
    if not isinstance(decisions, list) or len(decisions) > len(items):
        raise ValueError("UI decision cardinality is invalid.")
    seen: set[str] = set()
    action_status = {
        "accept": "accepted",
        "reject": "rejected",
        "edit": "edited",
        "mark_unclear": "needs_evidence",
        "request_more_documents": "needs_evidence",
        "skip": "skipped",
    }
    for decision in decisions:
        if not isinstance(decision, dict):
            raise ValueError("UI decision is malformed.")
        expected_fields = {
            "item_id",
            "item_type",
            "title",
            "action",
            "status",
            "decided_at",
        }
        expected_fields.update(
            field
            for field in (
                "reviewer_note",
                "edit_value",
                "requested_documents",
                "followup_context",
            )
            if field in decision
        )
        if set(decision) != expected_fields:
            raise ValueError("UI decision fields are not exact.")
        item_id = decision.get("item_id")
        item = item_by_id.get(item_id)
        action = decision.get("action")
        if (
            not isinstance(item_id, str)
            or item_id in seen
            or not isinstance(item, dict)
            or action not in action_status
            or action not in item.get("allowed_actions", [])
            or decision.get("item_type") != item.get("item_type")
            or decision.get("title") != item.get("title")
            or decision.get("status") != action_status[action]
        ):
            raise ValueError("UI decision does not close to its review item.")
        seen.add(item_id)
        _iso_datetime(decision.get("decided_at"), label="decision.decided_at")
        if action == "edit" and not str(decision.get("edit_value") or "").strip():
            raise ValueError("Edited UI decision has no edit value.")
        if "reviewer_note" in decision and (
            not isinstance(decision["reviewer_note"], str)
            or not decision["reviewer_note"].strip()
        ):
            raise ValueError("UI reviewer note is malformed.")
        if "requested_documents" in decision and (
            not isinstance(decision["requested_documents"], list)
            or not decision["requested_documents"]
            or any(
                not isinstance(value, str) or not value.strip()
                for value in decision["requested_documents"]
            )
        ):
            raise ValueError("UI requested documents are malformed.")
        if "followup_context" in decision and not isinstance(
            decision["followup_context"], dict
        ):
            raise ValueError("UI follow-up context is malformed.")
        expected_followup = _expected_followup_context(item, action=action)
        if decision.get("followup_context", {}) != expected_followup:
            raise ValueError("UI follow-up context does not freshly rederive.")
    expected_status = (
        "pending_review"
        if not decisions
        else "reviewed" if len(decisions) == len(items) else "partial_review"
    )
    expected_ui_fields = {
        "schema_version",
        "plugin",
        "workflow",
        "run_id",
        "decided_at",
        "decision_source",
        "review_payload_path",
        "decisions",
        "decision_count",
        "item_count",
        "status",
    }
    if "reviewer" in ui:
        expected_ui_fields.add("reviewer")
    if (
        set(ui) != expected_ui_fields
        or ui.get("schema_version") != review.get("schema_version")
        or ui.get("plugin") != review.get("plugin")
        or ui.get("workflow") != review.get("workflow")
        or ui.get("run_id") != review.get("run_id")
        or ui.get("review_payload_path") != "review_payload.json"
        or ui.get("decision_count") != len(decisions)
        or ui.get("item_count") != len(items)
        or ui.get("status") != expected_status
        or not isinstance(ui.get("decision_source"), str)
        or not ui["decision_source"].strip()
        or (
            "reviewer" in ui
            and (not isinstance(ui["reviewer"], str) or not ui["reviewer"].strip())
        )
    ):
        raise ValueError("UI decision receipt is stale.")
    if decisions:
        decided_at = _iso_datetime(
            ui.get("decided_at"), label="ui_decisions.decided_at"
        )
        if any(decision["decided_at"] != decided_at for decision in decisions):
            raise ValueError("UI decision timestamps do not close.")
    elif ui.get("decided_at") is not None:
        raise ValueError("Empty UI decisions cannot carry a decision timestamp.")
    return ui


def _safe_review_segment(value: object, fallback: str) -> str:
    cleaned = re.sub(r"[^a-zA-Z0-9._-]+", "-", str(value or "").strip())
    cleaned = cleaned.strip("-._")[:80]
    return cleaned or fallback


def _review_revision_path(item: dict[str, Any], decision: dict[str, Any]) -> str:
    target = str(item.get("output_path") or item.get("data", {}).get("path") or "")
    target_suffix = Path(target).suffix
    extension = (
        target_suffix.lower()
        if target_suffix.lower()
        in {".htm", ".html", ".md", ".sql", ".txt", ".xml", ".yaml", ".yml"}
        else ".txt"
    )
    source_base = Path(target or "review-item").stem
    return (
        f"revisions/{_safe_review_segment(source_base, 'review-item')}__"
        f"{_safe_review_segment(decision['item_id'], 'item')}{extension}"
    )


def _expected_application_effect(
    decision: dict[str, Any],
    item: dict[str, Any],
    *,
    applied_at: str,
) -> dict[str, Any]:
    data = item.get("data") if isinstance(item.get("data"), dict) else {}
    target_artifact = (
        str(data.get("target_artifact") or "").strip()
        or str(item.get("output_path") or "").strip()
        or str(data.get("path") or "").strip()
        or None
    )
    target_path = (
        str(data.get("target_path") or "").strip()
        or str(data.get("field_path") or "").strip()
        or str(data.get("field") or "").strip()
        or None
    )
    requires_followup = decision["action"] in {
        "reject",
        "mark_unclear",
        "request_more_documents",
    }
    effect: dict[str, Any] = {
        "item_id": decision["item_id"],
        "item_type": decision["item_type"],
        "title": decision["title"],
        "action": decision["action"],
        "status": decision["status"],
        "applied_at": applied_at,
        "applied": True,
        "requires_followup": requires_followup,
        "target_artifact": target_artifact,
        "target_path": target_path,
        "target_id_field": None,
        "target_record_id": None,
        "target_field": None,
        "target_records_key": None,
        "source_path": str(item.get("source_path") or "").strip() or None,
        "artifact_update": (
            "revision_artifact_pending"
            if decision["action"] == "edit"
            else "decision_manifest_only" if target_artifact else "review_record_only"
        ),
    }
    for optional in (
        "reviewer_note",
        "edit_value",
        "requested_documents",
        "followup_context",
    ):
        if optional in decision:
            effect[optional] = decision[optional]
    if decision["action"] == "edit":
        revision_path = _review_revision_path(item, decision)
        effect["revision_artifact"] = revision_path
        effect["artifact_update"] = "revision_artifact_written"
        if Path(target_artifact or "").suffix.lower() in {
            ".docx",
            ".pdf",
            ".pptx",
            ".xls",
            ".xlsm",
            ".xlsx",
        }:
            effect["requires_native_regeneration"] = True
            effect["native_regeneration_status"] = "pending"
            effect["artifact_update"] = "native_regeneration_pending"
    return effect


def _safe_application_status(
    effects: Sequence[dict[str, Any]],
    *,
    decision_count: int,
    item_count: int,
) -> str:
    if decision_count == 0:
        return "pending_review"
    if any(effect["requires_followup"] for effect in effects):
        return "blocked"
    if decision_count < item_count or any(
        effect.get("requires_native_regeneration") for effect in effects
    ):
        return "partial_review_applied"
    return "review_applied_with_assurance_limits"


def _validate_and_refresh_applied_decisions(
    output_dir: Path,
    review: dict[str, Any],
    ui: dict[str, Any],
    *,
    write_status: bool,
) -> dict[str, Any]:
    applied = read_json(output_dir / SAMPLE_APPLIED_DECISIONS_PATH)
    applied_at = _iso_datetime(applied.get("applied_at"), label="applied_at")
    item_by_id = {item["id"]: item for item in review["items"]}
    effects = [
        _expected_application_effect(
            decision,
            item_by_id[decision["item_id"]],
            applied_at=applied_at,
        )
        for decision in ui["decisions"]
    ]
    actual_effects = applied.get("effects")
    if actual_effects != effects:
        raise ValueError("Applied review effects do not freshly rederive.")
    revision_paths = [
        effect["revision_artifact"]
        for effect in effects
        if isinstance(effect.get("revision_artifact"), str)
    ]
    native_paths = sorted(
        {
            str(effect["target_artifact"])
            for effect in effects
            if effect.get("requires_native_regeneration")
        }
    )
    for effect in effects:
        revision_path = effect.get("revision_artifact")
        if not isinstance(revision_path, str):
            continue
        revision = output_dir / revision_path
        _require_ordinary_single_link(
            revision,
            label=f"Review revision {revision_path}",
        )
        if revision.read_text(encoding="utf-8") != effect["edit_value"]:
            raise ValueError("Review revision bytes do not close to the edit decision.")
    safe_status = _safe_application_status(
        effects,
        decision_count=len(ui["decisions"]),
        item_count=len(review["items"]),
    )
    expected = {
        "schema_version": review["schema_version"],
        "plugin": review["plugin"],
        "workflow": review["workflow"],
        "run_id": review["run_id"],
        "applied_at": applied_at,
        "decision_source": ui.get("decision_source", "mcp_widget"),
        "review_payload": {
            "path": "review_payload.json",
            "item_count": len(review["items"]),
            "review_type": review.get("review_type"),
        },
        "decisions": ui["decisions"],
        "effects": effects,
        "decision_count": len(ui["decisions"]),
        "item_count": len(review["items"]),
        "blocker_count": sum(1 for effect in effects if effect["requires_followup"]),
        "revision_count": len(revision_paths),
        "revision_paths": revision_paths,
        "target_update_count": 0,
        "target_update_paths": [],
        "structured_update_count": 0,
        "structured_update_paths": [],
        "native_regeneration_count": len(native_paths),
        "native_regeneration_paths": native_paths,
        "original_backup_paths": [],
        "application_status": safe_status,
    }
    if "reviewer" in ui:
        expected["reviewer"] = ui["reviewer"]
    exact_fields = {
        "schema_version",
        "plugin",
        "workflow",
        "run_id",
        "applied_at",
        "decision_source",
        "review_payload",
        "decisions",
        "effects",
        "decision_count",
        "item_count",
        "blocker_count",
        "revision_count",
        "revision_paths",
        "target_update_count",
        "target_update_paths",
        "structured_update_count",
        "structured_update_paths",
        "native_regeneration_count",
        "native_regeneration_paths",
        "original_backup_paths",
        "application_status",
    }
    if "reviewer" in ui:
        exact_fields.add("reviewer")
    if set(applied) != exact_fields:
        raise ValueError("Applied decision fields are not exact.")
    for field in exact_fields - {"application_status"}:
        if applied.get(field) != expected.get(field):
            raise ValueError(f"Applied decision field {field} does not close.")
    if write_status:
        write_json(output_dir / SAMPLE_APPLIED_DECISIONS_PATH, expected)
        return expected
    if applied.get("application_status") != safe_status:
        raise ValueError("Applied decision status is not successor-authorized.")
    return applied


def _successor_context(
    output_dir: Path,
    *,
    kind: str,
) -> tuple[dict[str, Any], str, dict[str, Any]]:
    predecessor_manifest, predecessor_digest = _manifest_content(
        output_dir / SAMPLE_OUTPUT_SET_PATH
    )
    predecessor_stage = _validated_stage(predecessor_manifest)
    archive_dir = (
        f"{SAMPLE_ASSURANCE_HISTORY_DIR}/{predecessor_stage['index']:03d}_"
        f"{predecessor_stage['kind']}"
    )
    archived_manifest, archived_digest = _manifest_content(
        output_dir / archive_dir / SAMPLE_OUTPUT_SET_PATH
    )
    if (
        archived_digest != predecessor_digest
        or archived_manifest != predecessor_manifest
    ):
        raise ValueError(
            "Prepared predecessor archive does not match the trusted stage."
        )
    stage = {
        "index": predecessor_stage["index"] + 1,
        "kind": kind,
        "predecessor": {
            "stage_index": predecessor_stage["index"],
            "stage_kind": predecessor_stage["kind"],
            "archive_dir": archive_dir,
            "manifest_sha256": predecessor_digest,
        },
    }
    _validated_stage({"stage": stage})
    return predecessor_manifest, archive_dir, stage


def _successor_binding(
    output_dir: Path,
    *,
    stage: dict[str, Any],
    ui: dict[str, Any],
    applied: dict[str, Any] | None,
) -> dict[str, Any]:
    ui_bytes, ui_sha256 = file_snapshot(output_dir / "ui_decisions.json")
    applied_receipt = None
    if applied is not None:
        applied_bytes, applied_sha256 = file_snapshot(
            output_dir / SAMPLE_APPLIED_DECISIONS_PATH
        )
        applied_receipt = {
            "path": SAMPLE_APPLIED_DECISIONS_PATH,
            "byte_count": applied_bytes,
            "sha256": applied_sha256,
            "decision_count": applied["decision_count"],
            "application_status": applied["application_status"],
        }
    predecessor = stage["predecessor"]
    return {
        "schema_version": SAMPLE_REVIEW_SUCCESSOR_SCHEMA_VERSION,
        "stage_index": stage["index"],
        "kind": stage["kind"],
        "predecessor_manifest_sha256": predecessor["manifest_sha256"],
        "ui_decisions": {
            "path": "ui_decisions.json",
            "byte_count": ui_bytes,
            "sha256": ui_sha256,
            "decision_count": ui["decision_count"],
            "status": ui["status"],
        },
        "applied_decisions": applied_receipt,
        "assurance_limits": {
            "semantic_review": "not_assessed",
            "reporting": "blocked",
            "publication": "withheld",
            "report_ready": False,
        },
    }


def _successor_next_actions(
    *,
    language: str,
    kind: str,
    ui: dict[str, Any],
    applied: dict[str, Any] | None,
) -> list[str]:
    spanish = language == "es"
    boundary = (
        "La revisión profesional de la suficiencia de la muestra, los informes y la publicación siguen pendientes."
        if spanish
        else "Professional sample-sufficiency review, reporting, and publication remain withheld."
    )
    if kind == "save":
        action = (
            "Aplique las decisiones guardadas antes de cualquier uso posterior."
            if spanish
            else "Apply the saved decisions before any downstream use."
        )
    elif applied is not None and applied["application_status"] == "blocked":
        action = (
            "Resuelva las decisiones bloqueadas y vuelva a aplicar la revisión."
            if spanish
            else "Resolve blocked decisions and apply the review again."
        )
    elif (
        applied is not None
        and applied["application_status"] == "partial_review_applied"
    ):
        action = (
            "Complete las decisiones o regeneraciones pendientes antes de la entrega."
            if spanish
            else "Complete pending decisions or native regeneration before handoff."
        )
    else:
        action = (
            "Use los artefactos solo como muestra preparada y revisada, no como conclusión de auditoría."
            if spanish
            else "Use the artifacts only as a prepared, reviewed sample—not as an audit conclusion."
        )
    if ui["decision_count"] == 0:
        action = (
            "Registre decisiones de revisión antes de continuar."
            if spanish
            else "Record review decisions before continuing."
        )
    return [action, boundary]


def _successor_output_records(
    output_dir: Path,
    *,
    predecessor_final: dict[str, Any],
    kind: str,
    ui: dict[str, Any],
    applied: dict[str, Any] | None,
) -> list[dict[str, Any]]:
    trusted_by_path = {
        output["path"]: output
        for output in predecessor_final.get("outputs", [])
        if isinstance(output, dict) and isinstance(output.get("path"), str)
    }
    paths = [
        path
        for path in _current_operational_payload_paths(output_dir)
        if path != "final_artifacts.json"
    ]
    paths.append(SAMPLE_OUTPUT_SET_PATH)
    records: list[dict[str, Any]] = []
    for relative in paths:
        trusted = trusted_by_path.get(relative, {})
        record: dict[str, Any] = {
            "path": relative,
            "kind": Path(relative).suffix.lower().lstrip(".") or "file",
            "status": "written_assured",
        }
        for field in (
            "row_count",
            "source_row_count",
            "required_columns",
            "required_text",
            "required_sheets",
            "required_sheet_headers",
            "required_cells",
            "qa_checks",
        ):
            if field in trusted:
                record[field] = trusted[field]
        if relative == "ui_decisions.json":
            record["status"] = ui["status"]
        elif relative == SAMPLE_APPLIED_DECISIONS_PATH and applied is not None:
            record["status"] = applied["application_status"]
        elif relative.startswith("revisions/"):
            record["status"] = "written_revision"
        elif relative == SAMPLE_OUTPUT_SET_PATH:
            record["status"] = f"{kind}_successor_sealed"
        records.append(record)
    return records


def _successor_run_and_final_payloads(
    output_dir: Path,
    *,
    archive_dir: str,
    stage: dict[str, Any],
    kind: str,
    ui: dict[str, Any],
    applied: dict[str, Any] | None,
    event_at: str,
) -> tuple[dict[str, Any], dict[str, Any], dict[str, Any]]:
    predecessor_run = read_json(output_dir / archive_dir / "run_intake.json")
    predecessor_final = read_json(output_dir / archive_dir / "final_artifacts.json")
    traces = [
        trace
        for trace in predecessor_run.get("execution_trace", [])
        if isinstance(trace, dict)
    ]
    traces.append(
        {
            "step_id": f"journal_sampling_review_{kind}_{stage['index']:03d}",
            "kind": f"deterministic_review_{kind}",
            "status": "passed",
            "execution_location": "cowork_connected_folder",
            "command": [
                "journal-sampling-widgets",
                (
                    "save_journal_sampling_decisions"
                    if kind == "save"
                    else "apply_journal_sampling_decisions"
                ),
            ],
            "inputs": [
                "review_payload.json",
                "ui_decisions.json",
                *([SAMPLE_APPLIED_DECISIONS_PATH] if applied is not None else []),
            ],
            "outputs": [
                "sampling_audit.json",
                "sample_assurance_gates.json",
                "sample_assurance_envelope.json",
                SAMPLE_OUTPUT_SET_PATH,
                "final_artifacts.json",
                "run_intake.json",
            ],
            "successor_stage": stage["index"],
            "completed_at": event_at,
        }
    )
    language = normalize_language(predecessor_run.get("language"), default="en")
    successor_binding = _successor_binding(
        output_dir,
        stage=stage,
        ui=ui,
        applied=applied,
    )
    run_intake = {
        **predecessor_run,
        "execution_trace": traces,
        "status": (
            "review_decisions_saved"
            if kind == "save"
            else str(applied["application_status"])
        ),
        "review_successor": successor_binding,
    }
    blockers = (
        [
            {
                "item_id": effect["item_id"],
                "item_type": effect["item_type"],
                "title": effect["title"],
                "action": effect["action"],
                "status": effect["status"],
            }
            for effect in applied["effects"]
            if effect["requires_followup"]
        ]
        if applied is not None
        else []
    )
    final_status = (
        "review_saved_pending_application"
        if kind == "save"
        else str(applied["application_status"])
    )
    final_artifacts: dict[str, Any] = {
        "schema_version": predecessor_final.get("schema_version", "1.0"),
        "plugin": "journal-sampling",
        "workflow": "journal-sampling",
        "run_id": run_intake["run_id"],
        "completed_at": event_at,
        "outputs": _successor_output_records(
            output_dir,
            predecessor_final=predecessor_final,
            kind=kind,
            ui=ui,
            applied=applied,
        ),
        "caveats": [
            "Deterministic preparation and review persistence do not establish sample sufficiency or an audit conclusion.",
            "Reporting and publication remain withheld without separate reviewed authority.",
        ],
        "blockers": blockers,
        "next_actions": _successor_next_actions(
            language=language,
            kind=kind,
            ui=ui,
            applied=applied,
        ),
        "status": final_status,
        "review_status": ui["status"],
        "review_successor": successor_binding,
    }
    if applied is not None:
        final_artifacts["review_application"] = {
            "applied_at": applied["applied_at"],
            "application_status": applied["application_status"],
            "decision_count": applied["decision_count"],
            "item_count": applied["item_count"],
            "blocker_count": applied["blocker_count"],
            "revision_count": applied["revision_count"],
            "revision_paths": applied["revision_paths"],
            "target_update_count": 0,
            "target_update_paths": [],
            "structured_update_count": 0,
            "structured_update_paths": [],
            "native_regeneration_count": applied["native_regeneration_count"],
            "native_regeneration_paths": applied["native_regeneration_paths"],
            "original_backup_paths": [],
            "applied_decisions_path": SAMPLE_APPLIED_DECISIONS_PATH,
        }
    return run_intake, final_artifacts, successor_binding


def _write_successor_run_and_final(
    output_dir: Path,
    *,
    archive_dir: str,
    stage: dict[str, Any],
    kind: str,
    ui: dict[str, Any],
    applied: dict[str, Any] | None,
) -> tuple[dict[str, Any], dict[str, Any], dict[str, Any]]:
    event_at = (
        applied["applied_at"]
        if applied is not None
        else ui.get("decided_at") or datetime.now().astimezone().isoformat()
    )
    run_intake, final_artifacts, successor_binding = _successor_run_and_final_payloads(
        output_dir,
        archive_dir=archive_dir,
        stage=stage,
        kind=kind,
        ui=ui,
        applied=applied,
        event_at=event_at,
    )
    write_json(output_dir / "run_intake.json", run_intake)
    write_json(output_dir / "final_artifacts.json", final_artifacts)
    return run_intake, final_artifacts, successor_binding


def _successor_gate_register(
    upstream: dict[str, Any],
    *,
    applied: bool,
) -> dict[str, Any]:
    source_refs = [
        qualification["qualification_id"]
        for qualification in upstream["source_qualifications"]
    ]
    semantic_refs = ["workpaper.ui_decisions"]
    if applied:
        semantic_refs.append("workpaper.applied_decisions")
    return build_gate_register(
        {
            "source": {
                "status": "passed",
                "evidence_refs": source_refs,
                "limitations": [],
            },
            "preparation": {
                "status": "passed",
                "evidence_refs": [
                    "prepared.normalized_journal",
                    "prepared.journal_sample_csv",
                    "output.journal_sample_xlsx",
                    "workpaper.sample_reproducibility",
                    "workpaper.sample_material_value_ledger",
                ],
                "limitations": [],
            },
            "reconciliation": {
                "status": "not_applicable",
                "evidence_refs": [],
                "limitations": [],
            },
            "semantic_review": {
                "status": "not_assessed",
                "evidence_refs": semantic_refs,
                "limitations": [
                    "Recorded item decisions are not a professional sample-sufficiency or audit-conclusion receipt."
                ],
            },
            "reporting": {
                "status": "blocked",
                "evidence_refs": [],
                "limitations": [
                    "Reporting remains blocked without separate reviewed reporting authority."
                ],
            },
            "publication": {
                "status": "withheld",
                "evidence_refs": [],
                "limitations": [
                    "Publication remains withheld without explicit publication authority."
                ],
            },
        }
    )


def _successor_receipt_specifications(
    *,
    applied: bool,
) -> list[tuple[str, str, str, str]]:
    specifications = [
        (
            "journal_sample.csv",
            "prepared.journal_sample_csv",
            "prepared",
            "text/csv",
        ),
        (
            "journal_sample.xlsx",
            "output.journal_sample_xlsx",
            "output",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
        (
            "sample_reproducibility.json",
            "workpaper.sample_reproducibility",
            "workpaper",
            "application/json",
        ),
        (
            "sample_material_value_ledger.json",
            "workpaper.sample_material_value_ledger",
            "workpaper",
            "application/json",
        ),
        (
            "sampling_audit.json",
            "workpaper.sampling_audit",
            "workpaper",
            "application/json",
        ),
        (
            "review_payload.json",
            "workpaper.review_payload",
            "workpaper",
            "application/json",
        ),
        (
            "ui_decisions.json",
            "workpaper.ui_decisions",
            "workpaper",
            "application/json",
        ),
        (
            "final_artifacts.json",
            "workpaper.final_artifacts",
            "workpaper",
            "application/json",
        ),
        (
            "run_intake.json",
            "workpaper.run_intake",
            "workpaper",
            "application/json",
        ),
    ]
    if applied:
        specifications.append(
            (
                SAMPLE_APPLIED_DECISIONS_PATH,
                "workpaper.applied_decisions",
                "workpaper",
                "application/json",
            )
        )
    return specifications


def _expected_successor_envelope(
    output_dir: Path,
    *,
    normalized_csv: Path,
    diagnostics: dict[str, Any],
    upstream: dict[str, Any],
    run_id: str,
    stage: dict[str, Any],
    applied: bool,
    gates: dict[str, Any],
) -> dict[str, Any]:
    normalized_receipt = validate_artifact_receipt(
        {"normalization": normalized_csv.parent},
        diagnostics["normalized_csv_receipt"],
    )
    implementation_receipts = _implementation_receipts()
    sample_receipts = [
        artifact_receipt(
            output_dir,
            output_dir / relative,
            artifact_id=artifact_id,
            root_id="sample",
            role=role,
            media_type=media_type,
        )
        for relative, artifact_id, role, media_type in (
            _successor_receipt_specifications(applied=applied)
        )
    ]
    gate_receipt = artifact_receipt(
        output_dir,
        output_dir / "sample_assurance_gates.json",
        artifact_id="workpaper.sample_assurance_gates",
        root_id="sample",
        role="workpaper",
        media_type="application/json",
    )
    source_receipts = [
        receipt
        for receipt in upstream["artifact_receipts"]
        if receipt["role"] == "source"
    ]
    return build_assurance_envelope(
        run_id=run_id,
        workflow_id="journal-sampling-sample",
        workflow_version=f"{SAMPLE_ASSURANCE_VERSION}-review-{stage['index']}",
        artifact_receipts=[
            *source_receipts,
            normalized_receipt,
            *implementation_receipts,
            *sample_receipts,
            gate_receipt,
        ],
        implementation_artifact_refs=[
            receipt["artifact_id"] for receipt in implementation_receipts
        ],
        reviewed_decisions=upstream["reviewed_decisions"],
        source_qualifications=upstream["source_qualifications"],
        allocation_ledgers=[],
        numeric_evidence_ledgers=[],
        gate_register=gates,
        limitations=[
            "The successor proves deterministic preparation and review persistence only; professional sufficiency, reporting, and publication remain withheld."
        ],
        artifact_roots={
            "source": _resolve_normalization_reference(
                normalized_csv,
                diagnostics,
                diagnostics["source_root"],
                label="Normalization source root",
            ),
            "normalization": normalized_csv.parent,
            "sample": output_dir,
            **_implementation_artifact_roots(),
        },
    )


def _build_successor_assurance(
    output_dir: Path,
    *,
    normalized_csv: Path,
    diagnostics_path: Path,
    run_id: str,
    stage: dict[str, Any],
    applied: bool,
) -> tuple[dict[str, Any], dict[str, Any]]:
    diagnostics, upstream = _validated_normalization_envelope(
        normalized_csv,
        diagnostics_path,
    )
    gates = _successor_gate_register(upstream, applied=applied)
    write_json(output_dir / "sample_assurance_gates.json", gates)
    envelope = _expected_successor_envelope(
        output_dir,
        normalized_csv=normalized_csv,
        diagnostics=diagnostics,
        upstream=upstream,
        run_id=run_id,
        stage=stage,
        applied=applied,
        gates=gates,
    )
    write_json(output_dir / "sample_assurance_envelope.json", envelope)
    return gates, envelope


def finalize_sample_review_successor(
    output_dir: Path,
    *,
    kind: str,
) -> dict[str, Any]:
    """Mint, replay, and return one real post-review successor closure."""

    if kind not in {"save", "apply"}:
        raise ValueError("Review successor kind must be save or apply.")
    unresolved = output_dir.expanduser()
    if unresolved.is_symlink():
        raise ValueError("Sample output directory cannot be a symlink.")
    output_dir = unresolved.resolve()
    _, archive_dir, stage = _successor_context(output_dir, kind=kind)
    audit = read_json(output_dir / "sampling_audit.json")
    run_intake, review, _ = _review_item_contract(output_dir, audit)
    ui = _validate_ui_decisions(output_dir, review)
    applied = None
    if kind == "apply":
        applied = _validate_and_refresh_applied_decisions(
            output_dir,
            review,
            ui,
            write_status=True,
        )
    elif (output_dir / SAMPLE_APPLIED_DECISIONS_PATH).exists():
        raise ValueError("A save successor cannot introduce applied decisions.")
    _, _, successor_binding = _write_successor_run_and_final(
        output_dir,
        archive_dir=archive_dir,
        stage=stage,
        kind=kind,
        ui=ui,
        applied=applied,
    )
    audit["review_successor"] = successor_binding
    audit["review_session"] = {
        **audit["review_session"],
        "stage_index": stage["index"],
        "stage_kind": stage["kind"],
        "ui_status": ui["status"],
        "application_status": (
            applied["application_status"] if applied is not None else None
        ),
    }
    write_json(output_dir / "sampling_audit.json", audit)
    population_proof = audit.get("population_proof")
    if not isinstance(population_proof, dict):
        raise ValueError("Sampling audit lost its population proof.")
    normalized_csv = _resolve_run_intake_reference(
        output_dir,
        run_intake,
        audit["normalized_csv"],
    )
    diagnostics_path = _resolve_run_intake_reference(
        output_dir,
        run_intake,
        population_proof["diagnostics_path"],
    )
    _build_successor_assurance(
        output_dir,
        normalized_csv=normalized_csv,
        diagnostics_path=diagnostics_path,
        run_id=str(run_intake["run_id"]),
        stage=stage,
        applied=applied is not None,
    )
    _write_sample_output_set(output_dir, stage=stage)
    replay = validate_sample_assurance(output_dir)
    return {
        "ok": True,
        "kind": kind,
        "stage": replay["output_set"]["stage"],
        "manifest_sha256": replay["output_set"]["content_sha256"],
        "application_status": (
            applied["application_status"] if applied is not None else None
        ),
        "physical_paths": replay["output_set"]["physical_paths"],
        "directory_paths": replay["output_set"]["directory_paths"],
    }


def _validate_review_successor_state(
    output_dir: Path,
    *,
    output_set: dict[str, Any],
    audit: dict[str, Any],
    envelope: dict[str, Any],
    normalized_csv: Path,
    diagnostics_path: Path,
) -> dict[str, Any] | None:
    stage = output_set["stage"]
    run_intake, review, _ = _review_item_contract(output_dir, audit)
    if stage["kind"] == "initial":
        ui = read_json(output_dir / "ui_decisions.json")
        if (
            ui.get("decisions") != []
            or ui.get("decision_count") != 0
            or ui.get("status") != "pending_review"
            or (output_dir / SAMPLE_APPLIED_DECISIONS_PATH).exists()
            or "review_successor" in audit
            or "review_successor" in run_intake
        ):
            raise ValueError("Initial sample stage contains post-review state.")
        return None

    ui = _validate_ui_decisions(output_dir, review)
    applied = None
    if stage["kind"] == "apply":
        applied = _validate_and_refresh_applied_decisions(
            output_dir,
            review,
            ui,
            write_status=False,
        )
    elif (output_dir / SAMPLE_APPLIED_DECISIONS_PATH).exists():
        raise ValueError("Save successor contains unbound applied decisions.")
    expected_binding = _successor_binding(
        output_dir,
        stage=stage,
        ui=ui,
        applied=applied,
    )
    final_artifacts = read_json(output_dir / "final_artifacts.json")
    if (
        audit.get("review_successor") != expected_binding
        or run_intake.get("review_successor") != expected_binding
        or final_artifacts.get("review_successor") != expected_binding
    ):
        raise ValueError("Review successor binding is stale.")
    predecessor = stage["predecessor"]
    predecessor_audit = read_json(
        output_dir / predecessor["archive_dir"] / "sampling_audit.json"
    )
    predecessor_review_session = predecessor_audit.get("review_session")
    if not isinstance(predecessor_review_session, dict):
        raise ValueError("Sample predecessor review session is malformed.")
    expected_audit = {
        **predecessor_audit,
        "review_successor": expected_binding,
        "review_session": {
            **predecessor_review_session,
            "stage_index": stage["index"],
            "stage_kind": stage["kind"],
            "ui_status": ui["status"],
            "application_status": (
                applied["application_status"] if applied is not None else None
            ),
        },
    }
    if audit != expected_audit:
        raise ValueError("Review successor audit state is stale.")
    event_at = (
        applied["applied_at"]
        if applied is not None
        else ui.get("decided_at")
        or _iso_datetime(
            final_artifacts.get("completed_at"),
            label="final_artifacts.completed_at",
        )
    )
    expected_run, expected_final, payload_binding = _successor_run_and_final_payloads(
        output_dir,
        archive_dir=predecessor["archive_dir"],
        stage=stage,
        kind=stage["kind"],
        ui=ui,
        applied=applied,
        event_at=event_at,
    )
    if payload_binding != expected_binding:
        raise ValueError("Review successor payload binding is stale.")
    if run_intake != expected_run or final_artifacts != expected_final:
        raise ValueError("Review successor final or run state is stale.")
    successor_diagnostics, upstream = _validated_normalization_envelope(
        normalized_csv,
        diagnostics_path,
    )
    expected_gates = _successor_gate_register(
        upstream,
        applied=applied is not None,
    )
    if envelope["gate_register"] != expected_gates:
        raise ValueError("Review successor gates do not freshly rederive.")
    expected_envelope = _expected_successor_envelope(
        output_dir,
        normalized_csv=normalized_csv,
        diagnostics=successor_diagnostics,
        upstream=upstream,
        run_id=str(run_intake["run_id"]),
        stage=stage,
        applied=applied is not None,
        gates=expected_gates,
    )
    if envelope != expected_envelope:
        raise ValueError("Review successor assurance envelope is stale.")
    expected_implementation = _implementation_receipts()
    expected_ids = [
        *(
            receipt["artifact_id"]
            for receipt in upstream["artifact_receipts"]
            if receipt["role"] == "source"
        ),
        "prepared.normalized_journal",
        *(receipt["artifact_id"] for receipt in expected_implementation),
        "prepared.journal_sample_csv",
        "output.journal_sample_xlsx",
        "workpaper.sample_reproducibility",
        "workpaper.sample_material_value_ledger",
        "workpaper.sampling_audit",
        "workpaper.review_payload",
        "workpaper.ui_decisions",
        "workpaper.final_artifacts",
        "workpaper.run_intake",
        *(["workpaper.applied_decisions"] if applied is not None else []),
        "workpaper.sample_assurance_gates",
    ]
    if [
        receipt["artifact_id"] for receipt in envelope["artifact_receipts"]
    ] != expected_ids:
        raise ValueError("Review successor assurance receipt set is not exact.")
    return expected_binding


def validate_sample_assurance(output_dir: Path) -> dict[str, Any]:
    """Freshly replay upstream proof, selection, native values, and output set."""

    unresolved_output_dir = output_dir.expanduser()
    if unresolved_output_dir.is_symlink():
        raise ValueError("Sample output directory cannot be a symlink.")
    output_dir = unresolved_output_dir.resolve()
    output_set = validate_sample_output_set(output_dir)
    audit = read_json(output_dir / "sampling_audit.json")
    run_intake = read_json(output_dir / "run_intake.json")
    normalized_value = audit.get("normalized_csv")
    population_proof = audit.get("population_proof")
    if (
        not isinstance(normalized_value, str)
        or not normalized_value
        or not isinstance(population_proof, dict)
    ):
        raise ValueError("Sampling audit does not contain replayable population proof.")
    normalized_csv = _resolve_run_intake_reference(
        output_dir,
        run_intake,
        normalized_value,
    )
    diagnostics_value = population_proof.get("diagnostics_path")
    if not isinstance(diagnostics_value, str) or not diagnostics_value:
        raise ValueError("Sampling audit is missing normalization diagnostics.")
    diagnostics_path = _resolve_run_intake_reference(
        output_dir,
        run_intake,
        diagnostics_value,
    )
    frame = pl.read_csv(normalized_csv, infer_schema=False)
    fresh_proof = _validate_population_proof(
        normalized_csv,
        frame,
        diagnostics_path,
    )
    if fresh_proof != population_proof:
        raise ValueError("Sampling population proof is stale.")
    reproducibility = _validate_sample_reproducibility(output_dir, normalized_csv)
    ledger = validate_sample_material_value_ledger(output_dir, normalized_csv)
    diagnostics, _ = _validated_normalization_envelope(
        normalized_csv,
        diagnostics_path,
    )
    gates = read_json(output_dir / "sample_assurance_gates.json")
    envelope = validate_assurance_envelope(
        read_json(output_dir / "sample_assurance_envelope.json"),
        artifact_roots={
            "source": _resolve_normalization_reference(
                normalized_csv,
                diagnostics,
                diagnostics["source_root"],
                label="Normalization source root",
            ),
            "normalization": normalized_csv.parent,
            "sample": output_dir,
            **_implementation_artifact_roots(),
        },
    )
    _validate_exact_implementation_receipts(envelope)
    if envelope["gate_register"] != gates:
        raise ValueError("Sample assurance gates do not close to the envelope.")
    sample_gates = envelope["gate_register"]["gates"]
    if (
        sample_gates["source"]["status"] != "passed"
        or sample_gates["preparation"]["status"] != "passed"
        or sample_gates["semantic_review"]["status"] != "not_assessed"
        or sample_gates["reporting"]["status"] != "blocked"
        or sample_gates["publication"]["status"] != "withheld"
        or envelope["gate_register"]["report_ready"]
    ):
        raise ValueError("Sample assurance gate state is success-shaped or stale.")
    if (
        audit.get("population_size_before_filters")
        != reproducibility["population_size_before_filters"]
        or audit.get("population_size_after_filters")
        != reproducibility["population_size_after_filters"]
        or audit.get("sample_size")
        != len(reproducibility["selected_prepared_row_numbers"])
        or audit.get("sample_size") != ledger["row_count"]
        or audit.get("method") != reproducibility["sampling_contract"]["method"]
        or audit.get("requested_size")
        != reproducibility["sampling_contract"]["requested_size"]
        or audit.get("filters") != reproducibility["sampling_contract"]["filters"]
    ):
        raise ValueError("Sampling audit counts or contract are stale.")
    successor = _validate_review_successor_state(
        output_dir,
        output_set=output_set,
        audit=audit,
        envelope=envelope,
        normalized_csv=normalized_csv,
        diagnostics_path=diagnostics_path,
    )
    return {
        "output_set": output_set,
        "population_proof": fresh_proof,
        "reproducibility": reproducibility,
        "material_value_ledger": ledger,
        "assurance_envelope": envelope,
        "review_successor": successor,
    }


def _write_sample_xlsx(
    sample: pl.DataFrame,
    path: Path,
    *,
    language: str,
) -> None:
    """Write the required native workbook or fail the whole sample stage."""

    try:
        sample.write_excel(path, worksheet=workbook_sheet_name(language))
    except (ImportError, ModuleNotFoundError, RuntimeError, ValueError) as exc:
        raise RuntimeError("Required sample XLSX generation failed.") from exc


def _require_clean_output_target(output_dir: Path) -> bool:
    if output_dir.is_symlink():
        raise ValueError("Sample output directory cannot be a symlink.")
    if not output_dir.exists():
        return False
    if not output_dir.is_dir():
        raise ValueError("Sample output path must be a directory.")
    if any(output_dir.iterdir()):
        raise ValueError(
            "Sample output directory must be absent or empty for exact closure."
        )
    return True


def run_sample(
    normalized_csv: Path,
    output_dir: Path,
    *,
    method: str = "random",
    size: int = 25,
    group_column: str = "account",
    include_accounts: Sequence[str] = (),
    exclude_accounts: Sequence[str] = (),
    date_start: str | None = None,
    date_end: str | None = None,
    min_abs: Decimal | str | int | None = None,
    keyword: str | None = None,
    language: object | None = None,
    normalization_diagnostics: Path | None = None,
    client_engagement: Mapping[str, Any] | None = None,
) -> SampleResult:
    """Run and atomically finalize a replayable, exactly closed sample stage."""

    normalized_csv = normalized_csv.expanduser().resolve()
    unresolved_output_dir = output_dir.expanduser()
    if unresolved_output_dir.is_symlink():
        raise ValueError("Sample output directory cannot be a symlink.")
    output_dir = unresolved_output_dir.resolve()
    normalized_client_engagement = _validated_client_sample_stage(
        client_engagement,
        normalized_csv=normalized_csv,
        output_dir=output_dir,
    )
    method_key = method.strip().lower()
    if method_key in {"monetary unit", "monetary unit sampling"}:
        method_key = "mus"
    language_code = normalize_language(language, default="en")
    diagnostics_path = (
        normalization_diagnostics.expanduser().resolve()
        if normalization_diagnostics is not None
        else normalized_csv.parent / "normalization_diagnostics.json"
    )
    frame = pl.read_csv(normalized_csv, infer_schema=False)
    population_proof = _validate_population_proof(
        normalized_csv,
        frame,
        diagnostics_path,
    )
    min_abs_value = (
        None
        if min_abs is None
        else (
            min_abs
            if isinstance(min_abs, Decimal)
            else parse_localized_decimal(min_abs, label="min_abs")
        )
    )
    min_abs_text = decimal_text(min_abs_value) if min_abs_value is not None else None
    target_existed = _require_clean_output_target(output_dir)
    indexed_frame = frame.with_row_index("__prepared_row_number", offset=1)
    population = _apply_filters(
        indexed_frame,
        include_accounts=include_accounts,
        exclude_accounts=exclude_accounts,
        date_start=date_start,
        date_end=date_end,
        min_abs=min_abs_value,
        keyword=keyword,
    )
    selected = _select_sample(
        population,
        method=method_key,
        size=size,
        group_column=group_column,
    )
    sample = selected.select(CANONICAL_COLUMNS)

    output_dir.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(
        prefix=".journal-sampling-stage-",
        dir=output_dir.parent,
    ) as staging_name:
        staging_dir = Path(staging_name)
        run_intake = write_run_intake(
            staging_dir,
            normalized_csv=normalized_csv,
            method=method_key,
            size=size,
            group_column=group_column,
            include_accounts=include_accounts,
            exclude_accounts=exclude_accounts,
            date_start=date_start,
            date_end=date_end,
            min_abs=min_abs_text,
            keyword=keyword,
            language=language_code,
            declared_output_dir=output_dir,
            client_engagement=normalized_client_engagement,
            run_id=(
                str(normalized_client_engagement["run_id"])
                if normalized_client_engagement is not None
                else None
            ),
        )
        run_intake_payload = read_json(run_intake.path)
        normalized_csv_reference = run_intake_payload.get("assumptions", {}).get(
            "normalized_csv"
        )
        if not isinstance(normalized_csv_reference, str):
            raise ValueError("Sample run intake has no normalized CSV reference.")
        sample_csv = staging_dir / "journal_sample.csv"
        sample_xlsx = staging_dir / "journal_sample.xlsx"
        sample.write_csv(sample_csv)
        _write_sample_xlsx(sample, sample_xlsx, language=language_code)
        reproducibility = _build_sample_reproducibility(
            normalized_receipt=population_proof["normalized_csv_receipt"],
            method=method_key,
            size=size,
            group_column=group_column,
            include_accounts=include_accounts,
            exclude_accounts=exclude_accounts,
            date_start=date_start,
            date_end=date_end,
            min_abs=min_abs_text,
            keyword=keyword,
            population_size_before=frame.height,
            population_size_after=population.height,
            selected=selected,
            sample_csv=sample_csv,
        )
        write_json(staging_dir / "sample_reproducibility.json", reproducibility)
        material_ledger = _build_sample_material_value_ledger(
            staging_dir,
            selected,
            worksheet=workbook_sheet_name(language_code),
        )
        write_json(
            staging_dir / "sample_material_value_ledger.json",
            material_ledger,
        )
        review_paths = {
            "run_id": run_intake.run_id,
            "run_intake_path": _managed_run_reference(
                output_dir / "run_intake.json",
                normalized_client_engagement,
            ),
            "review_payload_path": _managed_run_reference(
                output_dir / "review_payload.json",
                normalized_client_engagement,
            ),
            "ui_decisions_path": _managed_run_reference(
                output_dir / "ui_decisions.json",
                normalized_client_engagement,
            ),
            "final_artifacts_path": _managed_run_reference(
                output_dir / "final_artifacts.json",
                normalized_client_engagement,
            ),
            "review_item_count": sample.height + 4,
        }
        audit = {
            "schema_version": "journal_sampling.sample_audit.v1",
            "client_engagement": _portable_client_engagement_context(
                normalized_client_engagement
            ),
            **(
                {"path_reference": "run_root_relative"}
                if run_intake_payload.get("path_reference") == "run_root_relative"
                else {}
            ),
            "normalized_csv": normalized_csv_reference,
            "language": language_code,
            "method": method_key,
            "seed": 42 if method_key == "random" else None,
            "requested_size": size,
            "population_size_before_filters": frame.height,
            "population_size_after_filters": population.height,
            "sample_size": sample.height,
            "filters": {
                "include_accounts": list(include_accounts),
                "exclude_accounts": list(exclude_accounts),
                "date_start": date_start,
                "date_end": date_end,
                "min_abs": min_abs_text,
                "keyword": keyword,
            },
            "population_proof": population_proof,
            "reproducibility": "sample_reproducibility.json",
            "material_value_ledger": "sample_material_value_ledger.json",
            "sample_assurance_gates": "sample_assurance_gates.json",
            "sample_assurance_envelope": "sample_assurance_envelope.json",
            "sample_output_receipts": SAMPLE_OUTPUT_SET_PATH,
            "outputs": {
                "csv": "journal_sample.csv",
                "xlsx": "journal_sample.xlsx",
            },
            "review_session": review_paths,
        }
        write_json(staging_dir / "sampling_audit.json", audit)
        _build_sample_assurance(
            staging_dir,
            normalized_csv=normalized_csv,
            diagnostics_path=diagnostics_path,
            run_id=run_intake.run_id,
        )
        review_session = write_review_session_artifacts(
            staging_dir,
            run_id=run_intake.run_id,
            run_intake_path=run_intake.path,
            sample=sample,
            audit=audit,
            client_engagement=normalized_client_engagement,
        )
        if review_session.review_item_count != review_paths["review_item_count"]:
            raise ValueError("Sample review-item cardinality does not close.")
        _write_sample_output_set(staging_dir)
        validate_sample_assurance(staging_dir)
        final_frame = pl.read_csv(normalized_csv, infer_schema=False)
        final_proof = _validate_population_proof(
            normalized_csv,
            final_frame,
            diagnostics_path,
        )
        if final_proof != population_proof:
            raise ValueError("Sampling population changed before finalization.")

        if target_existed:
            output_dir.rmdir()
        try:
            staging_dir.rename(output_dir)
        except OSError:
            if target_existed and not output_dir.exists():
                output_dir.mkdir()
            raise
    return SampleResult(frame=sample, audit=audit)


def comma_list(value: str | None) -> list[str]:
    """Parse comma-separated CLI values."""

    if not value:
        return []
    return [item.strip() for item in value.split(",") if item.strip()]


def temp_output_dir(prefix: str) -> Path:
    """Create a temporary output directory for Claude scratch runs."""

    return Path(tempfile.mkdtemp(prefix=prefix))


def add_common_args(parser: argparse.ArgumentParser) -> None:
    parser.add_argument(
        "--language",
        help="Working/output language locale: it, en, fr, de, or es. Defaults to recipe or en.",
    )
    parser.add_argument(
        "--document-language",
        help="Source-document language locale: it, en, fr, de, es, or auto. Defaults to recipe or auto.",
    )
    parser.add_argument("--verbose", action="store_true", help="Enable debug logging.")
