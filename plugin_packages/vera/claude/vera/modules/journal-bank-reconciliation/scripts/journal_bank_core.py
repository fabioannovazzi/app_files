from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import logging
import os
import re
import stat
import sys
import tempfile
import unicodedata
from bisect import bisect_left
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Any, Sequence
from zipfile import ZIP_DEFLATED, BadZipFile, ZipFile, ZipInfo

import fastexcel
import openpyxl
import polars as pl
from excel_sanitization import excel_safe_value
from implementation_bootstrap import (
    IMPLEMENTATION_CONTRACT,
    validate_implementation_tree,
)
from openpyxl.utils.exceptions import InvalidFileException

_COMPONENT_ROOT = Path(__file__).resolve().parents[1]
_VENDOR_CANDIDATES = (
    _COMPONENT_ROOT / "vendor" / "modules",
    _COMPONENT_ROOT.parent.parent / "vendor" / "modules",
    _COMPONENT_ROOT.parent / "_shared" / "vendor" / "modules",
)
_VERA_ASSURANCE_ROOT: Path | None = None
for _vendor_candidate in _VENDOR_CANDIDATES:
    if (_vendor_candidate / "vera_assurance").is_dir():
        _VERA_ASSURANCE_ROOT = _vendor_candidate / "vera_assurance"
        if str(_vendor_candidate) not in sys.path:
            sys.path.insert(0, str(_vendor_candidate))
        break
if _VERA_ASSURANCE_ROOT is None:
    raise ImportError("Cannot locate the vendored vera_assurance implementation.")

from vera_assurance import (
    MoneyValidationError,
    artifact_receipt,
    build_allocation_ledger,
    build_assurance_envelope,
    build_gate_register,
    build_reviewed_decision_receipt,
    build_source_qualification,
    canonical_json_sha256,
    decimal_text,
    difference_within_tolerance,
    file_snapshot,
    parse_canonical_decimal,
    parse_localized_decimal,
    validate_allocation_ledger,
    validate_artifact_receipt,
    validate_assurance_envelope,
    validate_gate_register,
    validate_reviewed_decision_receipt,
    validate_source_qualification,
)
from vera_assurance import write_json as write_assurance_json  # noqa: E402

try:
    from .review_session import (
        FINAL_ARTIFACT_EXCLUDED_NAMES,
        refresh_final_artifacts,
        refresh_review_execution_trace,
        write_review_session_artifacts,
        write_run_intake,
    )
except ImportError:  # pragma: no cover - supports direct script imports
    import importlib.util

    _review_session_path = Path(__file__).resolve().parent / "review_session.py"
    _review_session_spec = importlib.util.spec_from_file_location(
        "mparanza_journal_bank_reconciliation_review_session",
        _review_session_path,
    )
    if _review_session_spec is None or _review_session_spec.loader is None:
        raise ImportError(f"Cannot load review session helper: {_review_session_path}")
    _review_session = importlib.util.module_from_spec(_review_session_spec)
    sys.modules[_review_session_spec.name] = _review_session
    _review_session_spec.loader.exec_module(_review_session)
    FINAL_ARTIFACT_EXCLUDED_NAMES = _review_session.FINAL_ARTIFACT_EXCLUDED_NAMES
    refresh_final_artifacts = _review_session.refresh_final_artifacts
    refresh_review_execution_trace = _review_session.refresh_review_execution_trace
    write_review_session_artifacts = _review_session.write_review_session_artifacts
    write_run_intake = _review_session.write_run_intake

LOGGER = logging.getLogger(__name__)

SUPPORTED_LANGUAGES = ("it", "en", "fr", "de", "es")
SUPPORTED_INPUT_SUFFIXES = {".csv", ".xls", ".xlsx", ".xlsm", ".pdf"}
TABULAR_ADAPTER_VERSION = "6"
TABULAR_ADAPTER_ID = "journal_bank.tabular.v6"
EXTENDED_TABULAR_ADAPTER_VERSION = "7"
EXTENDED_TABULAR_ADAPTER_ID = "journal_bank.tabular.v7"
TEXT_PDF_ADAPTER_VERSION = "2"
TEXT_PDF_ADAPTER_ID = "journal_bank.text_pdf.disabled.v2"
RELATIONSHIP_ADAPTER_ID = "journal_bank.relationship.v3"
RELATIONSHIP_ADAPTER_VERSION = "3"
NORMALIZATION_SCHEMA_VERSION = "journal_bank.normalization.v2"
LINEAGE_SCHEMA_VERSION = "journal_bank.lineage.v1"
ASSURANCE_WORKFLOW_ID = "journal_bank_reconciliation"
ASSURANCE_WORKFLOW_VERSION = "2"
FINAL_ARTIFACT_CLOSURE_MAX_PASSES = 8
MATCH_STAGE_ORDER = (
    "reference",
    "reference_group",
    "amount_date_unique",
    "amount_date_single",
)
WORKBOOK_SHEET_ORDER = (
    "matches",
    "relationship_residuals",
    "unmatched_bank",
    "unmatched_journal",
    "bank_pdf_non_movements",
    "normalized_bank",
    "normalized_journal",
)
NATIVE_OUTPUT_FILES = (
    "normalized_bank.csv",
    "normalized_journal.csv",
    "reconciliation_matches.csv",
    "relationship_residuals.csv",
    "unmatched_bank.csv",
    "unmatched_journal.csv",
    "bank_pdf_non_movement_rows.csv",
    "journal_bank_reconciliation.xlsx",
    "reconciliation_audit.json",
    "review_notes.md",
    "input_receipts.json",
    "source_qualifications.json",
    "reviewed_decisions.json",
    "lineage.json",
    "relationship_ledger.json",
    "material_value_ledger.json",
    "assurance_gates.json",
    "artifact_receipts.json",
    "assurance_envelope.json",
    "run_intake.json",
    "review_payload.json",
    "ui_decisions.json",
    "final_artifacts.json",
    "review_handoff.md",
    "applied_decisions.json",
)
INITIAL_RUN_OUTPUT_FILES = NATIVE_OUTPUT_FILES[:-1]
POST_REVIEW_OUTPUT_FILES = ("applied_decisions.json",)
DETERMINISTIC_ARTIFACT_FILES = (
    "normalized_bank.csv",
    "normalized_journal.csv",
    "reconciliation_matches.csv",
    "unmatched_bank.csv",
    "unmatched_journal.csv",
    "bank_pdf_non_movement_rows.csv",
    "journal_bank_reconciliation.xlsx",
    "review_notes.md",
    "input_receipts.json",
    "source_qualifications.json",
    "reviewed_decisions.json",
    "lineage.json",
    "relationship_ledger.json",
    "assurance_gates.json",
)
MATERIAL_CLOSURE_FILES = (
    "relationship_residuals.csv",
    "material_value_ledger.json",
)
IMPLEMENTATION_ARTIFACT_SPECS = (
    (
        "implementation.plugin.codex_plugin.plugin_json",
        "implementation",
        ".codex-plugin/plugin.json",
    ),
    (
        "implementation.plugin.app_json",
        "implementation",
        ".app.json",
    ),
    (
        "implementation.plugin.mcp_json",
        "implementation",
        ".mcp.json",
    ),
    (
        "implementation.plugin.assets.icon_svg",
        "implementation",
        "assets/icon.svg",
    ),
    (
        "implementation.plugin.assets.journal_bank_review_widget_html",
        "implementation",
        "assets/journal-bank-review-widget.html",
    ),
    (
        "implementation.plugin.assets.review_workbench_adapter_json",
        "implementation",
        "assets/review-workbench-adapter.json",
    ),
    (
        "implementation.plugin.mcp.server_cjs",
        "implementation",
        "mcp/server.cjs",
    ),
    (
        "implementation.plugin.scripts.apply_review_edits_py",
        "implementation",
        "scripts/apply_review_edits.py",
    ),
    (
        "implementation.plugin.scripts.check_dependencies_py",
        "implementation",
        "scripts/check_dependencies.py",
    ),
    (
        "implementation.plugin.scripts.excel_sanitization_py",
        "implementation",
        "scripts/excel_sanitization.py",
    ),
    (
        "implementation.plugin.scripts.implementation_bootstrap_py",
        "implementation",
        "scripts/implementation_bootstrap.py",
    ),
    (
        "implementation.plugin.scripts.inspect_inputs_py",
        "implementation",
        "scripts/inspect_inputs.py",
    ),
    (
        "implementation.plugin.scripts.journal_bank_core_py",
        "implementation",
        "scripts/journal_bank_core.py",
    ),
    (
        "implementation.plugin.scripts.review_session_py",
        "implementation",
        "scripts/review_session.py",
    ),
    (
        "implementation.plugin.scripts.run_reconciliation_py",
        "implementation",
        "scripts/run_reconciliation.py",
    ),
    (
        "implementation.plugin.scripts.semantic_review_py",
        "implementation",
        "scripts/semantic_review.py",
    ),
    (
        "implementation.shared.vera_assurance.init_py",
        "shared_implementation",
        "__init__.py",
    ),
    (
        "implementation.shared.vera_assurance.contracts_py",
        "shared_implementation",
        "contracts.py",
    ),
    (
        "implementation.shared.vera_assurance.decisions_py",
        "shared_implementation",
        "decisions.py",
    ),
    (
        "implementation.shared.vera_assurance.envelope_py",
        "shared_implementation",
        "envelope.py",
    ),
    (
        "implementation.shared.vera_assurance.money_py",
        "shared_implementation",
        "money.py",
    ),
    (
        "implementation.shared.vera_assurance.relationships_py",
        "shared_implementation",
        "relationships.py",
    ),
    (
        "implementation.shared.vera_assurance.review_output_transaction_cjs",
        "shared_implementation",
        "review_output_transaction.cjs",
    ),
    (
        "implementation.shared.vera_assurance.serialization_py",
        "shared_implementation",
        "serialization.py",
    ),
)
RUN_SCOPED_ARTIFACT_FILES = (
    "reconciliation_audit.json",
    "artifact_receipts.json",
    "assurance_envelope.json",
    "run_intake.json",
    "review_payload.json",
    "ui_decisions.json",
    "final_artifacts.json",
    "review_handoff.md",
    "applied_decisions.json",
)
OOXML_TIMESTAMP = b"2000-01-01T00:00:00Z"
OOXML_ZIP_TIMESTAMP = (1980, 1, 1, 0, 0, 0)
OOXML_CORE_TIMESTAMP_RE = re.compile(
    rb"(<dcterms:(created|modified)\b[^>]*>).*?(</dcterms:\2>)",
    flags=re.DOTALL,
)
ZERO = Decimal("0")
CANONICAL_DIRECTIONS = frozenset({"positive", "negative", "zero"})
CSV_FIELD_DELIMITERS = (",", ";", "\t", "|")
DEFAULT_CSV_FIELD_DELIMITER = ","
DATE_CONVENTIONS = ("day_first", "month_first")
DATE_LOCALES = ("it",)
ITALIAN_MONTH_NUMBERS = {
    "gennaio": 1,
    "febbraio": 2,
    "marzo": 3,
    "aprile": 4,
    "maggio": 5,
    "giugno": 6,
    "luglio": 7,
    "agosto": 8,
    "settembre": 9,
    "ottobre": 10,
    "novembre": 11,
    "dicembre": 12,
}
MATERIAL_VALUE_LEDGER_SCHEMA_VERSION = "journal_bank.material_value_ledger.v1"
CSV_DELIMITER_PROFILE_MAX_BYTES = 128 * 1024
CSV_DELIMITER_PROFILE_MAX_ROWS = 100
CSV_TRANSPORT_CHUNK_BYTES = 64 * 1024
_RECIPE_CONTAINER_ERROR_FIELD = "_recipe_container_error"
SOURCE_ROW_COLUMN = "__source_row__"
SOURCE_SHEET_COLUMN = "__source_sheet__"
TRANSACTION_COLUMNS = [
    "side",
    "transaction_id",
    "transaction_date",
    "amount_signed",
    "amount_abs",
    "description",
    "beneficiary",
    "reference",
    "movement_number",
    "account",
    "currency",
    "unit",
    "entity_ref",
    "party_ref",
    "direction",
    "source_file",
    "source_sheet",
    "source_row",
]
NON_MOVEMENT_COLUMNS = [
    "side",
    "source_file",
    "source_sheet",
    "source_row",
    "classification",
    "reason",
    "transaction_date",
    "amount_signed",
    "amount_abs",
    "description",
]
EXACT_HEADER_ALIASES: dict[str, frozenset[str]] = {
    "date": frozenset(
        {
            "date",
            "data",
            "datum",
            "booking date",
            "transaction date",
            "data registrazione",
        }
    ),
    "amount": frozenset(
        {
            "amount",
            "importo",
            "montant",
            "betrag",
            "signed amount",
            "importo movimento",
        }
    ),
    "debit": frozenset({"debit", "dare", "addebito", "soll", "importo dare"}),
    "credit": frozenset(
        {"credit", "avere", "credito", "haben", "accredito", "importo avere"}
    ),
    "description": frozenset(
        {
            "description",
            "descrizione",
            "causale",
            "libelle",
            "libellé",
            "beschreibung",
            "narrative",
            "details",
        }
    ),
    "beneficiary": frozenset(
        {
            "beneficiary",
            "beneficiario",
            "payee",
            "payer",
            "counterparty",
            "contrepartie",
            "beguenstigter",
            "begünstigter",
        }
    ),
    "reference": frozenset(
        {
            "reference",
            "riferimento",
            "document reference",
            "documento",
            "cro",
            "trn",
        }
    ),
    "movement_number": frozenset(
        {
            "movement",
            "movement number",
            "movement_number",
            "movimento",
            "numero movimento",
            "numero registrazione",
            "n. registrazione",
            "nr. reg",
            "n. reg",
            "beleg",
        }
    ),
    "account": frozenset({"account", "conto", "konto"}),
    "currency": frozenset({"currency", "valuta", "devise", "wahrung", "währung"}),
    "unit": frozenset({"unit", "unita", "unità", "einheit"}),
    "entity_ref": frozenset(
        {
            "entity",
            "entity ref",
            "entity_ref",
            "societa",
            "società",
            "company",
            "legal entity",
        }
    ),
    "party_ref": frozenset(
        {
            "party",
            "party ref",
            "party_ref",
            "counterparty id",
            "controparte",
            "soggetto",
        }
    ),
    "direction": frozenset(
        {"direction", "flow direction", "segno", "verso", "richtung"}
    ),
}
REFERENCE_GENERIC_TOKENS = frozenset(
    {
        "invoice",
        "fattura",
        "payment",
        "pagamento",
        "transfer",
        "bonifico",
        "reference",
        "riferimento",
        "document",
        "documento",
        "transaction",
        "movimento",
        "credit",
        "debit",
        "bank",
        "banca",
        "statement",
        "estratto",
        "year",
        "anno",
    }
)
GENERIC_PERIOD_REFERENCE_PREFIXES = (
    "fy",
    "fq",
    "fiscal",
    "financial",
    "fiscalyear",
    "financialyear",
    "fiscalperiod",
    "financialperiod",
    "accountingperiod",
    "reportingperiod",
    "period",
    "esercizio",
    "exercice",
    "quarter",
)
_GENERIC_PERIOD_MARKER = "(?:" + "|".join(GENERIC_PERIOD_REFERENCE_PREFIXES) + ")"
_GENERIC_PERIOD_QUALIFIER = (
    r"(?:actuals?|budget(?:ed)?|forecast(?:ed)?|quarter(?:ly)?|q[1-4])"
)
_GENERIC_PERIOD_YEAR = r"(?:19|20)\d{2}(?:(?:19|20)\d{2}|\d{2})?"
GENERIC_PERIOD_REFERENCE_RE = re.compile(
    rf"(?:"
    rf"{_GENERIC_PERIOD_MARKER}(?:{_GENERIC_PERIOD_QUALIFIER})*"
    rf"(?:fy)?{_GENERIC_PERIOD_YEAR}(?:{_GENERIC_PERIOD_QUALIFIER})*"
    rf"(?:{_GENERIC_PERIOD_MARKER})?"
    rf"|q[1-4](?:fy)?{_GENERIC_PERIOD_YEAR}"
    rf"(?:{_GENERIC_PERIOD_QUALIFIER})*"
    rf"|{_GENERIC_PERIOD_YEAR}(?:{_GENERIC_PERIOD_QUALIFIER})+"
    rf"|(?:{_GENERIC_PERIOD_QUALIFIER})+(?:fy)?{_GENERIC_PERIOD_YEAR}"
    rf"(?:{_GENERIC_PERIOD_MARKER})?"
    rf")"
)
GENERIC_PERIOD_WORDS = frozenset(
    {
        *GENERIC_PERIOD_REFERENCE_PREFIXES,
        "actual",
        "actuals",
        "budget",
        "budgeted",
        "forecast",
        "forecasted",
        "q1",
        "q2",
        "q3",
        "q4",
        "quarterly",
        "month",
        "monthly",
        "week",
        "weekly",
        "half",
        "semester",
        "sem",
        "jan",
        "january",
        "feb",
        "february",
        "mar",
        "march",
        "apr",
        "april",
        "may",
        "jun",
        "june",
        "jul",
        "july",
        "aug",
        "august",
        "sep",
        "sept",
        "september",
        "oct",
        "october",
        "nov",
        "november",
        "dec",
        "december",
    }
)
GENERIC_PERIOD_VERSION_RE = re.compile(r"(?:version|revision|draft|final|ver|rev|v)\d*")
GENERIC_PERIOD_CODE_RE = re.compile(
    r"(?:"
    r"(?:month|m)(?:0?[1-9]|1[0-2])"
    r"|(?:week|wk|w)(?:0?[1-9]|[1-4]\d|5[0-3])"
    r"|(?:half|h|semester|sem)[12]"
    r"|q[1-4]"
    r")"
)
MONETARY_HEADER_TOKENS = (
    "amount",
    "importo",
    "dare",
    "avere",
    "debit",
    "débit",
    "credit",
    "crédit",
    "soll",
    "haben",
    "montant",
    "betrag",
    "saldo",
    "balance",
    "fee",
    "commission",
    "commissione",
    "tax",
    "imposta",
    "ritenuta",
    "total",
    "totale",
)
MATCH_COLUMNS = [
    "status",
    "stage",
    "bank_transaction_id",
    "journal_transaction_id",
    "bank_date",
    "journal_date",
    "date_diff_days",
    "bank_amount",
    "journal_amount",
    "amount_delta",
    "bank_description",
    "journal_description",
    "shared_references",
    "review_note",
]
MATCH_MATERIAL_FIELDS = (
    "status",
    "stage",
    "bank_transaction_id",
    "journal_transaction_id",
    "bank_date",
    "journal_date",
    "date_diff_days",
    "bank_amount",
    "journal_amount",
    "amount_delta",
    "shared_references",
)
RESIDUAL_COLUMNS = [
    "side",
    "record_ref",
    "transaction_id",
    "record_amount",
    "allocated_amount",
    "residual",
    "currency",
    "unit",
    "entity_ref",
    "party_ref",
]
RESIDUAL_MATERIAL_FIELDS = tuple(RESIDUAL_COLUMNS)
MATERIAL_DECIMAL_FIELDS = frozenset(
    {
        "bank_amount",
        "journal_amount",
        "amount_delta",
        "record_amount",
        "allocated_amount",
        "residual",
    }
)
MATERIAL_INTEGER_FIELDS = frozenset({"date_diff_days"})
COMPACT_ISO_DATE_RE = re.compile(r"^(?P<year>\d{4})(?P<month>\d{2})(?P<day>\d{2})$")
YEAR_FIRST_DATE_RE = re.compile(
    r"^(?P<year>\d{4})(?P<separator>[./-])"
    r"(?P<month>\d{1,2})(?P=separator)(?P<day>\d{1,2})$"
)
DAY_MONTH_DATE_RE = re.compile(
    r"^(?P<first>\d{1,2})(?P<separator>[./-])"
    r"(?P<second>\d{1,2})(?P=separator)(?P<year>\d{2}|\d{4})$"
)
TEXTUAL_MONTH_DATE_RE = re.compile(
    r"^(?P<day>\d{1,2})[ \t]+(?P<month>[^\W\d_]+)" r"[ \t]+(?P<year>\d{4})$",
    flags=re.UNICODE,
)
AMOUNT_TOKEN_RE = re.compile(
    r"(?<!\w)\(?-?\d{1,3}(?:[.,\s]\d{3})*(?:[.,]\d{2})\)?(?!\w)"
)
BANK_PDF_NON_MOVEMENT_PATTERNS = (
    (
        "balance",
        "Bank statement balance line, not an ordinary bank movement.",
        re.compile(
            r"\b(?:"
            r"saldo iniziale|saldo finale|saldo precedente|saldo contabile|"
            r"saldo disponibile|saldo al|saldo a inizio|saldo a fine|"
            r"saldo per|"
            r"opening balance|closing balance|initial balance|final balance|"
            r"previous balance|balance brought forward|balance carried forward|"
            r"solde initial|solde final|solde au|"
            r"anfangssaldo|endsaldo|anfangsbestand|schlussbestand|kontostand"
            r")\b"
        ),
    ),
    (
        "total",
        "Bank statement total line, not an ordinary bank movement.",
        re.compile(
            r"^(?:totale|totali|total|totaux|summe|gesamtbetrag)\s+"
            r"(?:(?:de|del|della|dei|degli|des|du|der|die|of)\s+)?"
            r"(?:entrate|uscite|accrediti|addebiti|versamenti|prelievi|"
            r"movimenti|competenze|spese|commissioni|dare|avere|a credito|"
            r"a debito|debits?|credits?|incoming|outgoing|transactions?|"
            r"operations?|charges?|fees?|interests?|interets?|"
            r"gutschriften|belastungen|gebuhren|gebuehren|soll|haben|zinsen)\b"
        ),
    ),
    (
        "scalare",
        "Bank statement scalare/interest summary line, not an ordinary bank movement.",
        re.compile(
            r"\b(?:"
            r"scalare|riassunto scalare|riepilogo scalare|numeri creditori|"
            r"numeri debitori|calcolo competenze|riepilogo competenze|"
            r"competenze scalari|interest scale|interest summary|"
            r"echelle d interets|echelle interets|zinsstaffel|zinsenstaffel"
            r")\b"
        ),
    ),
    (
        "conditions",
        "Bank statement conditions line, not an ordinary bank movement.",
        re.compile(
            r"\b(?:"
            r"condizioni economiche|condizioni applicate|riepilogo condizioni|"
            r"condizioni del conto|conditions applied|account conditions|"
            r"conditions economiques|conditions appliquees|"
            r"kontokonditionen|konditionen|wirtschaftliche bedingungen"
            r")\b"
        ),
    ),
)

__all__ = [
    "CANONICAL_DIRECTIONS",
    "CSV_FIELD_DELIMITERS",
    "CSV_TRANSPORT_CHUNK_BYTES",
    "DATE_CONVENTIONS",
    "DATE_LOCALES",
    "DEFAULT_CSV_FIELD_DELIMITER",
    "DETERMINISTIC_ARTIFACT_FILES",
    "IMPLEMENTATION_ARTIFACT_SPECS",
    "INITIAL_RUN_OUTPUT_FILES",
    "InspectionResult",
    "ITALIAN_MONTH_NUMBERS",
    "MAPPING_FIELDS",
    "MATERIAL_CLOSURE_FILES",
    "MATCH_COLUMNS",
    "MATCH_MATERIAL_FIELDS",
    "MATCH_STAGE_ORDER",
    "NATIVE_OUTPUT_FILES",
    "NON_MOVEMENT_COLUMNS",
    "EXTENDED_TABULAR_ADAPTER_ID",
    "EXTENDED_TABULAR_ADAPTER_VERSION",
    "POST_REVIEW_OUTPUT_FILES",
    "ReconciliationBlockedError",
    "ReconciliationRunResult",
    "RELATIONSHIP_ADAPTER_ID",
    "RELATIONSHIP_ADAPTER_VERSION",
    "RELATIONSHIP_POLICY_FIELDS",
    "RUN_SCOPED_ARTIFACT_FILES",
    "RESIDUAL_COLUMNS",
    "RESIDUAL_MATERIAL_FIELDS",
    "TABULAR_ADAPTER_ID",
    "TABULAR_ADAPTER_VERSION",
    "TRANSACTION_COLUMNS",
    "WORKBOOK_SHEET_ORDER",
    "add_common_args",
    "build_mapping_review_receipt",
    "build_implementation_artifact_receipts",
    "build_relationship_review_receipt",
    "configure_logging",
    "excel_safe_value",
    "inspect_inputs",
    "implementation_artifact_roots",
    "normalize_language",
    "run_reconciliation",
    "validate_exact_implementation_receipts",
    "validate_material_value_ledger",
    "write_json",
]

MAPPING_FIELDS = (
    "date",
    "amount",
    "debit",
    "credit",
    "description",
    "beneficiary",
    "reference",
    "movement_number",
    "account",
    "currency",
    "unit",
    "entity_ref",
    "party_ref",
    "direction",
)
RELATIONSHIP_POLICY_FIELDS = {
    "relationship_shape",
    "allow_evidence_reuse",
    "require_same_currency",
    "require_same_unit",
    "require_same_entity",
    "require_same_party",
    "direction_policy",
    "default_currency",
    "default_unit",
    "default_entity_ref",
    "default_party_ref",
    "amount_tolerance",
    "date_window_days",
}


class ReconciliationBlockedError(ValueError):
    """Raised after evidence artifacts are written for a fail-closed run."""

    def __init__(self, code: str, detail: str) -> None:
        super().__init__(detail)
        self.code = code
        self.detail = detail


@dataclass(frozen=True)
class InspectionResult:
    """Deterministic inspection output for bank and journal inputs."""

    bank: dict[str, Any]
    journal: dict[str, Any]
    sample: dict[str, Any]
    suggested_recipe: dict[str, Any]


@dataclass(frozen=True)
class ReconciliationRunResult:
    """Reconciliation output plus audit metadata."""

    matches: pl.DataFrame
    unmatched_bank: pl.DataFrame
    unmatched_journal: pl.DataFrame
    audit: dict[str, Any]


def configure_logging(verbose: bool = False) -> None:
    """Configure script logging without affecting imported use."""

    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(level=level, format="%(levelname)s: %(message)s")


def normalize_language(
    language: object | None,
    *,
    default: str = "en",
    allow_auto: bool = False,
) -> str:
    """Normalize a language tag to one supported plugin locale."""

    text = str(language or default).strip().lower().replace("_", "-")
    code = text.split("-", 1)[0]
    if allow_auto and code == "auto":
        return "auto"
    return code if code in SUPPORTED_LANGUAGES else default


def language_assumptions(
    recipe: dict[str, Any],
    *,
    language: object | None = None,
    document_language: object | None = None,
) -> dict[str, str]:
    """Resolve working and source-document language assumptions."""

    working = normalize_language(language or recipe.get("language"), default="en")
    source = normalize_language(
        document_language or recipe.get("document_language") or "auto",
        default=working,
        allow_auto=True,
    )
    return {"language": working, "document_language": source}


def write_json(path: Path, payload: Any) -> None:
    """Write JSON with stable formatting."""

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )


def read_json(path: Path | None) -> dict[str, Any]:
    """Return a JSON object or an empty mapping when no file is provided."""

    if path is None:
        return {}
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError(f"Recipe must be a JSON object: {path}")
    return payload


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\u00a0", " ").replace("\u202f", " ").strip()


def _norm_label(value: Any) -> str:
    text = unicodedata.normalize("NFKD", _clean_text(value))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower()
    return re.sub(r"\s+", " ", text)


def _identifier_fragment(value: object) -> str:
    text = re.sub(r"[^A-Za-z0-9._-]+", "_", _clean_text(value)).strip("._-")
    return text or "source"


def _excel_column_name(index: int) -> str:
    idx = index + 1
    letters: list[str] = []
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        letters.append(chr(65 + rem))
    return "".join(reversed(letters))


def _unique_names(values: Sequence[Any]) -> list[str]:
    names: list[str] = []
    seen: dict[str, int] = {}
    for idx, value in enumerate(values):
        text = _clean_text(value)
        base = (
            text
            if text and text.lower() not in {"none", "nan"}
            else _excel_column_name(idx)
        )
        count = seen.get(base, 0)
        seen[base] = count + 1
        names.append(base if count == 0 else f"{base}_{count + 1}")
    return names


def supported_files(input_path: Path) -> list[Path]:
    """Return supported files from a file or folder path."""

    path = input_path.expanduser()
    if path.is_file():
        return [path] if path.suffix.lower() in SUPPORTED_INPUT_SUFFIXES else []
    if not path.exists():
        raise FileNotFoundError(f"Input path does not exist: {path}")
    return [
        candidate
        for candidate in sorted(path.rglob("*"))
        if candidate.is_file()
        and candidate.suffix.lower() in SUPPORTED_INPUT_SUFFIXES
        and not candidate.name.startswith("~$")
    ]


def _input_root(input_path: Path) -> Path:
    path = input_path.expanduser()
    return path if path.is_dir() else path.parent


def _source_identity(input_path: Path, source: Path) -> str:
    """Return a collision-safe, input-root-relative source identity."""

    root = _input_root(input_path).resolve()
    resolved = source.resolve()
    if not resolved.is_relative_to(root):
        raise ValueError(f"Source escapes its declared input root: {source}")
    return resolved.relative_to(root).as_posix()


def _source_identity_fragment(source_identity: str) -> str:
    path_text = Path(source_identity).with_suffix("").as_posix()
    readable = _identifier_fragment(path_text)
    digest = hashlib.sha256(source_identity.encode("utf-8")).hexdigest()[:12]
    return f"{readable}.{digest}"


def _read_excel_raw(path: Path) -> pl.DataFrame:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        workbook = openpyxl.load_workbook(path, data_only=True, read_only=False)
        try:
            sheet = workbook.worksheets[0]
            rows = [
                [
                    (
                        value.date().isoformat()
                        if isinstance(value, datetime)
                        else value.isoformat() if isinstance(value, date) else value
                    )
                    for value in row
                ]
                for row in sheet.iter_rows(values_only=True)
            ]
        finally:
            workbook.close()
        max_width = max((len(row) for row in rows), default=0)
        if max_width == 0:
            return pl.DataFrame()
        return pl.DataFrame(
            {
                f"column_{idx}": [row[idx] if idx < len(row) else None for row in rows]
                for idx in range(max_width)
            },
            strict=False,
        )
    try:
        return pl.read_excel(
            path,
            has_header=False,
            drop_empty_rows=False,
            drop_empty_cols=False,
        )
    except (fastexcel.CalamineError, ValueError, RuntimeError, OSError) as exc:
        raise ValueError(f"Unable to read legacy Excel source: {path}") from exc


def _bounded_csv_profile_text(path: Path) -> str:
    """Read a bounded UTF-8 prefix for mechanically reproducible delimiter checks."""

    with path.open("rb") as handle:
        payload = handle.read(CSV_DELIMITER_PROFILE_MAX_BYTES + 1)
    if len(payload) > CSV_DELIMITER_PROFILE_MAX_BYTES:
        payload = payload[:CSV_DELIMITER_PROFILE_MAX_BYTES]
        boundary = max(payload.rfind(b"\n"), payload.rfind(b"\r"))
        if boundary < 0:
            return ""
        payload = payload[: boundary + 1]
    return payload.decode("utf-8-sig")


def _profile_csv_field_delimiter(path: Path) -> dict[str, Any]:
    """Profile one unique consistent delimiter without assigning review authority.

    This rule is deterministic because it checks only the fixed supported byte
    set, strict CSV record structure, and bounded row/byte limits.
    """

    text = _bounded_csv_profile_text(path)
    candidates: list[str] = []
    parse_error_delimiters: list[str] = []
    sampled_row_count = 0
    for delimiter in CSV_FIELD_DELIMITERS:
        try:
            rows: list[list[str]] = []
            reader = csv.reader(
                io.StringIO(text, newline=""),
                delimiter=delimiter,
                strict=True,
            )
            for row in reader:
                if not any(_clean_text(value) for value in row):
                    continue
                rows.append(row)
                if len(rows) >= CSV_DELIMITER_PROFILE_MAX_ROWS:
                    break
        except csv.Error:
            parse_error_delimiters.append(delimiter)
            continue
        sampled_row_count = max(sampled_row_count, len(rows))
        widths = [len(row) for row in rows]
        if len(widths) >= 2 and widths[0] > 1 and len(set(widths)) == 1:
            candidates.append(delimiter)
    status = (
        "unique"
        if len(candidates) == 1
        else "ambiguous" if len(candidates) > 1 else "unsupported"
    )
    return {
        "status": status,
        "candidate_delimiters": candidates,
        "parse_error_delimiters": parse_error_delimiters,
        "sampled_row_count": sampled_row_count,
        "max_bytes": CSV_DELIMITER_PROFILE_MAX_BYTES,
        "max_rows": CSV_DELIMITER_PROFILE_MAX_ROWS,
    }


def _validated_csv_field_delimiter(value: object) -> str:
    if (
        not isinstance(value, str)
        or value not in CSV_FIELD_DELIMITERS
        or len(value.encode("utf-8")) != 1
    ):
        raise ValueError(
            "csv_field_delimiter must be one reviewed byte from comma, "
            "semicolon, tab, or pipe"
        )
    return value


def _mapping_csv_field_delimiter(
    source_file: str,
    value: object | None,
) -> str | None:
    is_csv = Path(source_file).suffix.lower() == ".csv"
    if not is_csv:
        if value is not None:
            raise ValueError("csv_field_delimiter is only valid for CSV sources")
        return None
    if value is None:
        return DEFAULT_CSV_FIELD_DELIMITER
    return _validated_csv_field_delimiter(value)


def _resolve_csv_field_delimiter(
    path: Path,
    file_recipe: dict[str, Any],
) -> dict[str, Any]:
    if path.suffix.lower() != ".csv":
        return {
            "status": "not_applicable",
            "delimiter": None,
            "origin": "not_applicable",
            "review_required": False,
            "profile": None,
            "requested_delimiter": None,
        }
    profile = _profile_csv_field_delimiter(path)
    if (
        "csv_field_delimiter" in file_recipe
        and file_recipe.get("csv_field_delimiter") is not None
    ):
        requested = file_recipe.get("csv_field_delimiter")
        try:
            delimiter = _validated_csv_field_delimiter(requested)
        except ValueError:
            return {
                "status": "unsupported",
                "delimiter": None,
                "origin": "reviewed_recipe",
                "review_required": True,
                "profile": profile,
                "requested_delimiter": requested,
            }
        profile_matches = profile["status"] == "unique" and profile[
            "candidate_delimiters"
        ] == [delimiter]
        requested_parse_failed = delimiter in profile["parse_error_delimiters"]
        if (
            profile["status"] in {"ambiguous", "unsupported"}
            and not requested_parse_failed
            and not isinstance(file_recipe.get("mapping_decision"), dict)
        ):
            return {
                "status": str(profile["status"]),
                "delimiter": None,
                "origin": "reviewed_recipe",
                "review_required": True,
                "profile": profile,
                "requested_delimiter": requested,
            }
        return {
            "status": "resolved",
            "delimiter": delimiter,
            "origin": "reviewed_recipe",
            "review_required": (
                delimiter != DEFAULT_CSV_FIELD_DELIMITER or not profile_matches
            ),
            "profile": profile,
            "requested_delimiter": requested,
        }
    if profile["status"] == "unique":
        delimiter = str(profile["candidate_delimiters"][0])
        return {
            "status": "resolved",
            "delimiter": delimiter,
            "origin": (
                "default"
                if delimiter == DEFAULT_CSV_FIELD_DELIMITER
                else "profiled_proposal"
            ),
            "review_required": delimiter != DEFAULT_CSV_FIELD_DELIMITER,
            "profile": profile,
            "requested_delimiter": None,
        }
    return {
        "status": str(profile["status"]),
        "delimiter": None,
        "origin": "profiled_proposal",
        "review_required": True,
        "profile": profile,
        "requested_delimiter": None,
    }


def _normalize_csv_record_terminators(source: Path, target: Path) -> None:
    """Normalize universal-newline byte forms without loading the file in memory."""

    pending_cr = False
    with source.open("rb") as source_handle, target.open("wb") as target_handle:
        while chunk := source_handle.read(CSV_TRANSPORT_CHUNK_BYTES):
            if pending_cr:
                target_handle.write(b"\n")
                if chunk.startswith(b"\n"):
                    chunk = chunk[1:]
                pending_cr = False
            if chunk.endswith(b"\r"):
                chunk = chunk[:-1]
                pending_cr = True
            target_handle.write(chunk.replace(b"\r\n", b"\n").replace(b"\r", b"\n"))
        if pending_cr:
            target_handle.write(b"\n")


def _read_csv_raw(path: Path, *, field_delimiter: str) -> pl.DataFrame:
    # Record terminators are transport syntax, not reviewed mapping intent.
    # A private streamed copy keeps CRLF split across chunks intact and lets
    # Polars own the full strict parse without duplicating a large CSV in RAM.
    with tempfile.TemporaryDirectory(prefix="jbr-csv-transport-") as temp_dir:
        normalized_path = Path(temp_dir) / "normalized.csv"
        _normalize_csv_record_terminators(path, normalized_path)
        delimiter = _validated_csv_field_delimiter(field_delimiter)
        expected_width: int | None = None
        try:
            with normalized_path.open(
                "r",
                encoding="utf-8-sig",
                newline="",
            ) as handle:
                for row in csv.reader(handle, delimiter=delimiter, strict=True):
                    if not any(_clean_text(value) for value in row):
                        continue
                    if expected_width is None:
                        expected_width = len(row)
                    elif len(row) != expected_width:
                        raise ValueError("CSV strict full parse found a ragged record")
        except csv.Error as exc:
            raise ValueError("CSV strict full parse failed") from exc
        return pl.read_csv(
            normalized_path,
            has_header=False,
            infer_schema=False,
            ignore_errors=False,
            truncate_ragged_lines=False,
            separator=delimiter,
        )


def _read_table_raw(
    path: Path,
    *,
    csv_field_delimiter: str | None = None,
) -> pl.DataFrame:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv_raw(
            path,
            field_delimiter=(
                csv_field_delimiter
                if csv_field_delimiter is not None
                else DEFAULT_CSV_FIELD_DELIMITER
            ),
        )
    if suffix in {".xls", ".xlsx", ".xlsm"}:
        return _read_excel_raw(path)
    raise ValueError(f"Unsupported tabular file: {path}")


def _drop_empty_columns(df: pl.DataFrame) -> pl.DataFrame:
    if df.is_empty() or df.width == 0:
        return df
    keep: list[str] = []
    for col in df.columns:
        if col in {SOURCE_ROW_COLUMN, SOURCE_SHEET_COLUMN}:
            keep.append(col)
            continue
        values = (
            df.get_column(col)
            .cast(pl.Utf8, strict=False)
            .fill_null("")
            .str.strip_chars()
        )
        if int((values != "").sum()) > 0:
            keep.append(col)
    return df.select(keep) if keep else pl.DataFrame()


def _drop_empty_rows(df: pl.DataFrame) -> pl.DataFrame:
    if df.is_empty() or df.width == 0:
        return df
    data_columns = [
        col for col in df.columns if col not in {SOURCE_ROW_COLUMN, SOURCE_SHEET_COLUMN}
    ]
    if not data_columns:
        return df.head(0)
    exprs = [
        pl.col(col).cast(pl.Utf8, strict=False).fill_null("").str.strip_chars() == ""
        for col in data_columns
    ]
    return df.filter(~pl.all_horizontal(exprs))


def _row_values(df: pl.DataFrame, idx: int) -> list[Any]:
    return list(df.row(idx))


def _suggest_header_rows(df: pl.DataFrame) -> list[int]:
    if df.is_empty():
        return [1]
    tokens = (
        "data",
        "date",
        "datum",
        "amount",
        "importo",
        "montant",
        "betrag",
        "dare",
        "avere",
        "debit",
        "credit",
        "soll",
        "haben",
        "descrizione",
        "description",
        "libelle",
        "libellé",
        "beschreibung",
        "beneficiario",
        "beneficiary",
        "payee",
        "iban",
        "reference",
        "riferimento",
        "movimento",
        "movement",
        "conto",
        "account",
    )
    best_idx = 0
    best_score = -1
    for idx in range(min(df.height, 30)):
        row = _row_values(df, idx)
        score = sum(1 for value in row if _clean_text(value))
        for value in row:
            label = _norm_label(value)
            if any(token in label for token in tokens):
                score += 3
            elif any(ch.isalpha() for ch in label):
                score += 1
        if score > best_score:
            best_score = score
            best_idx = idx
    return [best_idx + 1]


def _merge_header_rows(rows: Sequence[Sequence[Any]]) -> list[str]:
    width = max((len(row) for row in rows), default=0)
    labels: list[str] = []
    for idx in range(width):
        parts = []
        for row in rows:
            value = _clean_text(row[idx] if idx < len(row) else "")
            if value and value.lower() not in {"none", "nan"}:
                parts.append(value)
        labels.append(" ".join(parts))
    return _unique_names(labels)


def _apply_header(
    df: pl.DataFrame,
    rows_1_indexed: Sequence[int],
    *,
    source_sheet: str | None = None,
) -> pl.DataFrame:
    if df.is_empty():
        return df
    row_indexes = sorted({int(row) - 1 for row in rows_1_indexed})
    if not row_indexes or min(row_indexes) < 0:
        raise ValueError("Header rows must be 1-indexed positive integers.")
    if max(row_indexes) >= df.height:
        raise ValueError("Header row exceeds available rows.")
    labels = _merge_header_rows([_row_values(df, idx) for idx in row_indexes])
    body_start = max(row_indexes) + 1
    body = df.slice(body_start)
    if body.width != len(labels):
        labels = _unique_names(labels[: body.width])
    body.columns = labels
    body = body.with_columns(
        pl.Series(
            SOURCE_ROW_COLUMN,
            list(range(body_start + 1, body_start + 1 + body.height)),
            dtype=pl.Int64,
        ),
        pl.lit(source_sheet or "CSV").alias(SOURCE_SHEET_COLUMN),
    )
    # Preserve header-owned columns even when every body cell is blank. A
    # debit/credit export commonly leaves one side empty for the entire file,
    # but the explicit header still carries the monetary-role contract.
    return _drop_empty_rows(body)


def _source_sheet_names(path: Path) -> list[str]:
    if path.suffix.lower() == ".csv":
        return ["CSV"]
    workbook = fastexcel.read_excel(path)
    return [str(value) for value in workbook.sheet_names]


def _source_sheet_name(path: Path) -> str:
    names = _source_sheet_names(path)
    if len(names) != 1:
        raise ValueError(
            "Workbook must contain exactly one physical sheet for the bounded "
            "tabular adapter."
        )
    return names[0]


def _exact_mapping_from_columns(
    columns: Sequence[str],
) -> tuple[dict[str, str | None], list[str]]:
    """Resolve only exact, unique source-owned header labels.

    This bounded lookup is deterministic because it never profiles cell values,
    infers numeric position, or selects among duplicate semantic candidates.
    """

    available = [
        column
        for column in columns
        if column not in {SOURCE_ROW_COLUMN, SOURCE_SHEET_COLUMN}
    ]
    normalized = {column: _norm_label(column) for column in available}
    mapping: dict[str, str | None] = {}
    ambiguous: list[str] = []
    for field, aliases in EXACT_HEADER_ALIASES.items():
        matches = [column for column, label in normalized.items() if label in aliases]
        if len(matches) > 1:
            ambiguous.append(field)
            mapping[field] = None
        else:
            mapping[field] = matches[0] if matches else None
    monetary_fields = [
        field for field in ("amount", "debit", "credit") if mapping.get(field)
    ]
    if mapping.get("amount") and (mapping.get("debit") or mapping.get("credit")):
        ambiguous.append("monetary_role")
    if not monetary_fields:
        ambiguous.append("monetary_role")
    if not mapping.get("date"):
        ambiguous.append("date")
    return mapping, sorted(set(ambiguous))


def _exact_header_contract(
    raw: pl.DataFrame,
) -> tuple[list[int], dict[str, str | None]] | None:
    """Return one unambiguous bounded header contract, otherwise abstain."""

    candidates: list[tuple[list[int], dict[str, str | None]]] = []
    for index in range(min(raw.height, 30)):
        labels = _merge_header_rows([_row_values(raw, index)])
        mapping, ambiguous = _exact_mapping_from_columns(labels)
        if not ambiguous:
            candidates.append(([index + 1], mapping))
    if len(candidates) != 1:
        return None
    return candidates[0]


def _calendar_date(year: int, month: int, day: int) -> date | None:
    """Return one mechanically valid Gregorian date, otherwise abstain."""

    try:
        return date(year, month, day)
    except ValueError:
        return None


def _expanded_two_digit_year(value: str) -> int:
    """Mirror the documented POSIX/Python two-digit-year boundary exactly."""

    year = int(value)
    if len(value) == 4:
        return year
    return 2000 + year if year <= 68 else 1900 + year


def _date_parse_result(
    value: Any,
    *,
    date_convention: str | None = None,
    date_locale: str | None = None,
    allow_excel_serial: bool = False,
) -> tuple[str, date | None]:
    """Classify and parse a source date without list-order format guessing.

    Date syntax is mechanical, so a fixed parser is more auditable than model
    judgment here. Day/month strings are evaluated under both supported
    interpretations; a genuinely ambiguous value needs an explicit reviewed
    convention rather than whichever format happens to appear first.
    """

    if date_convention not in {None, *DATE_CONVENTIONS}:
        raise ValueError("date_convention must be day_first, month_first, or null")
    if date_locale not in {None, *DATE_LOCALES}:
        raise ValueError("date_locale must be it or null")
    if isinstance(value, datetime):
        return "parsed", value.date()
    if isinstance(value, date):
        return "parsed", value
    if isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
        try:
            serial = Decimal(str(value))
        except ArithmeticError:
            return "invalid", None
        if not serial.is_finite() or serial != serial.to_integral_value():
            return "invalid", None
        compact = str(int(serial))
        compact_match = COMPACT_ISO_DATE_RE.fullmatch(compact)
        if compact_match is not None:
            parsed = _calendar_date(
                int(compact_match["year"]),
                int(compact_match["month"]),
                int(compact_match["day"]),
            )
            return ("parsed", parsed) if parsed is not None else ("invalid", None)
        if allow_excel_serial and Decimal("20000") <= serial <= Decimal("60000"):
            return "parsed", date(1899, 12, 30) + timedelta(days=int(serial))
        return "invalid", None
    token = _clean_text(value)
    if not token:
        return "blank", None
    compact_match = COMPACT_ISO_DATE_RE.fullmatch(token)
    if compact_match is not None:
        parsed = _calendar_date(
            int(compact_match["year"]),
            int(compact_match["month"]),
            int(compact_match["day"]),
        )
        return ("parsed", parsed) if parsed is not None else ("invalid", None)
    if (
        allow_excel_serial
        and re.fullmatch(r"\d{5}", token) is not None
        and 20000 <= int(token) <= 60000
    ):
        return "parsed", date(1899, 12, 30) + timedelta(days=int(token))
    year_first_match = YEAR_FIRST_DATE_RE.fullmatch(token)
    if year_first_match is not None:
        parsed = _calendar_date(
            int(year_first_match["year"]),
            int(year_first_match["month"]),
            int(year_first_match["day"]),
        )
        return ("parsed", parsed) if parsed is not None else ("invalid", None)
    textual_month_match = TEXTUAL_MONTH_DATE_RE.fullmatch(token)
    if textual_month_match is not None:
        if date_locale is None:
            return "locale_required", None
        month_number = ITALIAN_MONTH_NUMBERS.get(
            _norm_label(textual_month_match["month"])
        )
        if month_number is None:
            return "invalid", None
        parsed = _calendar_date(
            int(textual_month_match["year"]),
            month_number,
            int(textual_month_match["day"]),
        )
        return ("parsed", parsed) if parsed is not None else ("invalid", None)
    day_month_match = DAY_MONTH_DATE_RE.fullmatch(token)
    if day_month_match is None:
        return "invalid", None
    year = _expanded_two_digit_year(day_month_match["year"])
    first = int(day_month_match["first"])
    second = int(day_month_match["second"])
    interpretations = {
        "day_first": _calendar_date(year, second, first),
        "month_first": _calendar_date(year, first, second),
    }
    if date_convention is not None:
        parsed = interpretations[date_convention]
        return ("parsed", parsed) if parsed is not None else ("invalid", None)
    valid = {parsed for parsed in interpretations.values() if parsed is not None}
    if not valid:
        return "invalid", None
    if len(valid) == 1:
        return "parsed", next(iter(valid))
    return "ambiguous", None


def _parse_date(
    value: Any,
    *,
    date_convention: str | None = None,
    date_locale: str | None = None,
    allow_excel_serial: bool = False,
) -> date | None:
    status, parsed = _date_parse_result(
        value,
        date_convention=date_convention,
        date_locale=date_locale,
        allow_excel_serial=allow_excel_serial,
    )
    return parsed if status == "parsed" else None


def _parse_number(
    value: Any,
    *,
    decimal_separator: str | None = None,
    thousands_separator: str | None = None,
) -> Decimal | None:
    """Parse source money exactly, rejecting ambiguous textual separators.

    Spreadsheet engines expose numeric cells as Python floats. ``allow_float`` is
    therefore explicit here and converts only the engine-provided display value
    through ``str``; text inputs remain subject to the stricter ambiguity rules.
    """

    if value is None:
        return None
    if isinstance(value, str) and not _clean_text(value):
        return None
    try:
        return parse_localized_decimal(
            value,
            label="source amount",
            decimal_separator=decimal_separator,
            thousands_separator=thousands_separator,
            allow_float=True,
        )
    except MoneyValidationError:
        return None


def _requires_numeric_separator_review(
    value: Any,
    *,
    decimal_separator: str | None,
    thousands_separator: str | None,
) -> bool:
    """Return whether two explicit separator conventions parse to different values."""

    if (
        decimal_separator is not None
        or thousands_separator is not None
        or not isinstance(value, str)
        or not _clean_text(value)
        or _parse_number(value) is not None
    ):
        return False
    candidates: set[Decimal] = set()
    for candidate_decimal_separator in (".", ","):
        try:
            candidates.add(
                parse_localized_decimal(
                    value,
                    label="source amount",
                    decimal_separator=candidate_decimal_separator,
                    thousands_separator=None,
                    allow_float=False,
                )
            )
        except MoneyValidationError:
            continue
    return len(candidates) > 1


def _date_ratio(series: pl.Series) -> float:
    values = [
        _date_parse_result(value)[0]
        for value in series.drop_nulls().head(100).to_list()
    ]
    if not values:
        return 0.0
    # Ambiguous numeric and localized textual-month strings remain useful
    # proposal evidence but never receive qualification authority until a
    # reviewer chooses the corresponding source-bound convention.
    return sum(
        value in {"parsed", "ambiguous", "locale_required"} for value in values
    ) / len(values)


def _amount_ratio(series: pl.Series) -> float:
    values = series.drop_nulls().head(100).to_list()
    if not values:
        return 0.0
    return sum(_parse_number(value) is not None for value in values) / len(values)


def _first_matching_column(df: pl.DataFrame, tokens: Sequence[str]) -> str | None:
    for col in df.columns:
        label = _norm_label(col)
        if any(token in label for token in tokens):
            return col
    return None


def infer_mapping(df: pl.DataFrame, side: str) -> dict[str, str | None]:
    """Propose a mapping for review without granting qualification authority."""

    mapping: dict[str, str | None] = {
        "date": _first_matching_column(df, ("data", "date", "datum", "booking")),
        "amount": _first_matching_column(
            df, ("amount", "importo", "montant", "betrag", "saldo", "total")
        ),
        "debit": _first_matching_column(df, ("dare", "debit", "addebito", "soll")),
        "credit": _first_matching_column(
            df, ("avere", "credit", "credito", "haben", "accredito")
        ),
        "description": _first_matching_column(
            df,
            (
                "descrizione",
                "description",
                "causale",
                "libelle",
                "libellé",
                "beschreibung",
                "narrative",
                "details",
            ),
        ),
        "beneficiary": _first_matching_column(
            df,
            (
                "beneficiario",
                "beneficiary",
                "payee",
                "payer",
                "cliente",
                "fornitore",
                "counterparty",
                "contrepartie",
                "beguenstigter",
                "begünstigter",
            ),
        ),
        "reference": _first_matching_column(
            df, ("reference", "riferimento", "document", "doc", "cro", "trn", "iban")
        ),
        "movement_number": _first_matching_column(
            df,
            (
                "movement",
                "movimento",
                "nr. reg",
                "n. reg",
                "registrazione",
                "journal",
                "beleg",
            ),
        ),
        "account": _first_matching_column(df, ("conto", "account", "iban", "konto")),
        "currency": _first_matching_column(df, ("currency", "valuta", "devise")),
        "unit": _first_matching_column(df, ("unit", "unita", "unità", "einheit")),
        "entity_ref": _first_matching_column(
            df, ("entity", "societa", "società", "company")
        ),
        "party_ref": _first_matching_column(
            df, ("party", "controparte", "counterparty id")
        ),
        "direction": _first_matching_column(
            df, ("direction", "segno", "verso", "richtung")
        ),
    }
    if mapping["date"] is None:
        candidates = [
            (col, _date_ratio(df.get_column(col)))
            for col in df.columns
            if col not in {SOURCE_ROW_COLUMN, SOURCE_SHEET_COLUMN}
        ]
        if candidates and max(score for _, score in candidates) >= 0.5:
            mapping["date"] = max(candidates, key=lambda item: item[1])[0]
    return mapping


def _mapping_for_file(
    recipe: dict[str, Any],
    side: str,
    path: Path,
    *,
    source_identity: str | None = None,
) -> dict[str, Any]:
    if side not in recipe:
        return {}
    raw_side_recipe = recipe.get(side)
    if not isinstance(raw_side_recipe, dict):
        return {
            _RECIPE_CONTAINER_ERROR_FIELD: (
                "recipe container must use object-valued side, files, and "
                "file entries"
            )
        }
    side_recipe = raw_side_recipe
    if "files" not in side_recipe:
        return side_recipe
    files = side_recipe.get("files")
    inherited = {name: value for name, value in side_recipe.items() if name != "files"}
    if not isinstance(files, dict) or any(
        not isinstance(key, str) or not isinstance(value, dict)
        for key, value in files.items()
    ):
        return {
            **inherited,
            _RECIPE_CONTAINER_ERROR_FIELD: (
                "recipe container must use object-valued side, files, and "
                "file entries"
            ),
        }
    if isinstance(files, dict):
        keys = [
            source_identity,
            path.name,
            path.as_posix(),
        ]
        for key in keys:
            file_recipe = files.get(key) if key else None
            if isinstance(file_recipe, dict):
                merged = dict(inherited)
                merged.update(file_recipe)
                return merged
    return inherited


def _normalized_mapping(mapping: dict[str, Any]) -> dict[str, str | None]:
    return {
        field: (_clean_text(mapping.get(field)) or None) for field in MAPPING_FIELDS
    }


def _normalized_direction_value_mapping(
    raw: object,
) -> tuple[dict[str, str], list[str]]:
    """Normalize a reviewed source vocabulary to canonical signed directions."""

    if raw is None:
        return {}, []
    if not isinstance(raw, dict):
        return {}, ["direction_value_mapping must be an object"]
    normalized: dict[str, str] = {}
    errors: list[str] = []
    for source_value, canonical_value in raw.items():
        source_label = _norm_label(source_value)
        target = _norm_label(canonical_value)
        if not source_label:
            errors.append("direction_value_mapping contains an empty source label")
            continue
        if target not in CANONICAL_DIRECTIONS:
            errors.append(
                "direction_value_mapping targets must be positive, negative, or zero"
            )
            continue
        if source_label in normalized:
            errors.append(
                "direction_value_mapping contains duplicate normalized source labels"
            )
            continue
        normalized[source_label] = target
    return dict(sorted(normalized.items())), sorted(set(errors))


def _normalized_date_convention(raw: object) -> tuple[str | None, list[str]]:
    if raw is None:
        return None, []
    if not isinstance(raw, str) or raw not in DATE_CONVENTIONS:
        return None, ["date_convention must be exactly day_first or month_first"]
    return raw, []


def _normalized_date_locale(raw: object) -> tuple[str | None, list[str]]:
    if raw is None:
        return None, []
    if not isinstance(raw, str) or raw not in DATE_LOCALES:
        return None, ["date_locale must be exactly it"]
    return raw, []


def _normalized_non_movement_summary_labels(
    raw: object,
) -> tuple[list[str], list[str]]:
    if raw is None:
        return [], []
    if not isinstance(raw, list) or any(not isinstance(value, str) for value in raw):
        return [], ["non_movement_summary_labels must be a list of strings"]
    normalized = [_norm_label(value) for value in raw]
    if (
        any(not value for value in normalized)
        or len(normalized) != len(set(normalized))
        or normalized != sorted(normalized)
    ):
        return [], [
            "non_movement_summary_labels must contain unique non-empty "
            "normalized labels in sorted order"
        ]
    return normalized, []


def _tabular_adapter_binding(
    date_locale: str | None,
    non_movement_summary_labels: Sequence[str] = (),
) -> tuple[str, str]:
    """Select v7 only for an explicit localized-date or summary authority."""

    if date_locale is not None or non_movement_summary_labels:
        return EXTENDED_TABULAR_ADAPTER_ID, EXTENDED_TABULAR_ADAPTER_VERSION
    return TABULAR_ADAPTER_ID, TABULAR_ADAPTER_VERSION


def _mapping_review_content(
    *,
    side: str,
    source_file: str,
    header_rows: Sequence[int],
    mapping: dict[str, Any],
    direction_value_mapping: object | None,
    potential_monetary_columns: Sequence[str],
    excluded_monetary_columns: Sequence[str],
    date_convention: object | None,
    date_locale: object | None,
    non_movement_summary_labels: object | None,
    csv_field_delimiter: object | None,
    decimal_separator: object | None,
    thousands_separator: object | None,
) -> dict[str, Any]:
    normalized_direction_values, direction_errors = _normalized_direction_value_mapping(
        direction_value_mapping
    )
    if direction_errors:
        raise ValueError("; ".join(direction_errors))
    normalized_date_convention, date_errors = _normalized_date_convention(
        date_convention
    )
    if date_errors:
        raise ValueError("; ".join(date_errors))
    normalized_date_locale, date_locale_errors = _normalized_date_locale(date_locale)
    if date_locale_errors:
        raise ValueError("; ".join(date_locale_errors))
    normalized_summary_labels, summary_label_errors = (
        _normalized_non_movement_summary_labels(non_movement_summary_labels)
    )
    if summary_label_errors:
        raise ValueError("; ".join(summary_label_errors))
    content = {
        "side": side,
        "source_file": source_file,
        "header_rows": [int(value) for value in header_rows],
        "mapping": _normalized_mapping(mapping),
        "direction_value_mapping": normalized_direction_values,
        "potential_monetary_columns": list(potential_monetary_columns),
        "excluded_monetary_columns": list(excluded_monetary_columns),
        "date_convention": normalized_date_convention,
        "csv_field_delimiter": _mapping_csv_field_delimiter(
            source_file,
            csv_field_delimiter,
        ),
        "decimal_separator": (
            _clean_text(decimal_separator) or None
            if decimal_separator is not None
            else None
        ),
        "thousands_separator": (
            _clean_text(thousands_separator) or None
            if thousands_separator is not None
            else None
        ),
    }
    if normalized_date_locale is not None:
        content["date_locale"] = normalized_date_locale
    if normalized_summary_labels:
        content["non_movement_summary_labels"] = normalized_summary_labels
    return content


def build_mapping_review_receipt(
    *,
    decision_id: str,
    reviewer_ref: str,
    reviewed_on: str,
    source_artifact_ref: str,
    side: str,
    source_file: str,
    header_rows: Sequence[int],
    mapping: dict[str, Any],
    potential_monetary_columns: Sequence[str],
    excluded_monetary_columns: Sequence[str],
    direction_value_mapping: object | None = None,
    date_convention: object | None = None,
    date_locale: object | None = None,
    non_movement_summary_labels: object | None = None,
    csv_field_delimiter: object | None = None,
    decimal_separator: object | None = None,
    thousands_separator: object | None = None,
) -> dict[str, Any]:
    """Seal a professional mapping decision against source bytes and adapter."""

    if isinstance(potential_monetary_columns, (str, bytes)) or isinstance(
        excluded_monetary_columns,
        (str, bytes),
    ):
        raise ValueError(
            "potential and excluded monetary columns must be explicit lists"
        )
    potential = [str(value).strip() for value in potential_monetary_columns]
    excluded = [str(value).strip() for value in excluded_monetary_columns]
    if (
        any(not value for value in potential)
        or len(potential) != len(set(potential))
        or any(not value for value in excluded)
        or len(excluded) != len(set(excluded))
    ):
        raise ValueError(
            "potential and excluded monetary columns must be unique non-empty names"
        )
    mapped_monetary = {
        str(mapping[field])
        for field in ("amount", "debit", "credit")
        if mapping.get(field)
    }
    if set(excluded) - set(potential):
        raise ValueError(
            "excluded_monetary_columns must be a subset of "
            "potential_monetary_columns"
        )
    if mapped_monetary & set(excluded):
        raise ValueError("a mapped monetary column cannot also be explicitly excluded")
    unresolved = set(potential) - mapped_monetary - set(excluded)
    if unresolved or mapped_monetary - set(potential):
        raise ValueError(
            "potential monetary columns require a complete mapped-or-excluded "
            "disposition"
        )
    normalized_date_locale, date_locale_errors = _normalized_date_locale(date_locale)
    if date_locale_errors:
        raise ValueError("; ".join(date_locale_errors))
    normalized_summary_labels, summary_label_errors = (
        _normalized_non_movement_summary_labels(non_movement_summary_labels)
    )
    if summary_label_errors:
        raise ValueError("; ".join(summary_label_errors))
    adapter_id, adapter_version = _tabular_adapter_binding(
        normalized_date_locale,
        normalized_summary_labels,
    )
    return build_reviewed_decision_receipt(
        decision_id=decision_id,
        decision_type="journal_bank_mapping",
        status="reviewed",
        reviewer_ref=reviewer_ref,
        reviewed_on=reviewed_on,
        adapter_id=adapter_id,
        adapter_version=adapter_version,
        source_artifact_refs=[source_artifact_ref],
        content=_mapping_review_content(
            side=side,
            source_file=source_file,
            header_rows=header_rows,
            mapping=mapping,
            direction_value_mapping=direction_value_mapping,
            potential_monetary_columns=potential,
            excluded_monetary_columns=excluded,
            date_convention=date_convention,
            date_locale=normalized_date_locale,
            non_movement_summary_labels=normalized_summary_labels,
            csv_field_delimiter=csv_field_delimiter,
            decimal_separator=decimal_separator,
            thousands_separator=thousands_separator,
        ),
    )


def _normalize_relationship_policy(policy: object) -> dict[str, Any]:
    if not isinstance(policy, dict) or set(policy) != RELATIONSHIP_POLICY_FIELDS:
        raise ValueError(
            "relationship policy must contain the exact reviewed perimeter fields"
        )
    shape = _clean_text(policy["relationship_shape"])
    if shape not in {"one_to_one", "one_to_many", "many_to_one", "many_to_many"}:
        raise ValueError("unsupported relationship shape")
    normalized: dict[str, Any] = {"relationship_shape": shape}
    for field in (
        "allow_evidence_reuse",
        "require_same_currency",
        "require_same_unit",
        "require_same_entity",
        "require_same_party",
    ):
        if not isinstance(policy[field], bool):
            raise ValueError(f"relationship policy {field} must be boolean")
        normalized[field] = policy[field]
    if normalized["allow_evidence_reuse"]:
        raise ValueError("journal-bank reconciliation cannot reuse evidence")
    if not normalized["require_same_currency"] or not normalized["require_same_unit"]:
        raise ValueError("currency and unit equality are mandatory")
    direction_policy = _clean_text(policy["direction_policy"])
    if direction_policy not in {"absolute_amount", "same_sign", "opposite_sign"}:
        raise ValueError("unsupported direction policy")
    normalized["direction_policy"] = direction_policy
    for field in (
        "default_currency",
        "default_unit",
        "default_entity_ref",
        "default_party_ref",
    ):
        value = _clean_text(policy[field]) if policy[field] is not None else ""
        normalized[field] = value or None
    tolerance, tolerance_text = _canonical_tolerance(policy["amount_tolerance"])
    if tolerance < ZERO:  # pragma: no cover - guarded by _canonical_tolerance
        raise ValueError("amount tolerance must not be negative")
    normalized["amount_tolerance"] = tolerance_text
    date_window = policy["date_window_days"]
    if (
        not isinstance(date_window, int)
        or isinstance(date_window, bool)
        or date_window < 0
    ):
        raise ValueError("date_window_days must be a non-negative integer")
    normalized["date_window_days"] = date_window
    return normalized


def build_relationship_review_receipt(
    *,
    decision_id: str,
    reviewer_ref: str,
    reviewed_on: str,
    source_artifact_refs: Sequence[str],
    policy: dict[str, Any],
) -> dict[str, Any]:
    """Seal the reviewed relationship perimeter, shape, and tolerance policy."""

    normalized = _normalize_relationship_policy(policy)
    return build_reviewed_decision_receipt(
        decision_id=decision_id,
        decision_type="journal_bank_relationship",
        status="reviewed",
        reviewer_ref=reviewer_ref,
        reviewed_on=reviewed_on,
        adapter_id=RELATIONSHIP_ADAPTER_ID,
        adapter_version=RELATIONSHIP_ADAPTER_VERSION,
        source_artifact_refs=list(source_artifact_refs),
        content={"policy": normalized},
    )


def _mapped(row: dict[str, Any], mapping: dict[str, Any], key: str) -> Any:
    col = mapping.get(key)
    return row.get(str(col)) if col else None


def _amount_from_row(
    row: dict[str, Any],
    mapping: dict[str, Any],
    *,
    decimal_separator: str | None = None,
    thousands_separator: str | None = None,
) -> Decimal | None:
    amount = _parse_number(
        _mapped(row, mapping, "amount"),
        decimal_separator=decimal_separator,
        thousands_separator=thousands_separator,
    )
    if amount is not None:
        return amount
    debit = _parse_number(
        _mapped(row, mapping, "debit"),
        decimal_separator=decimal_separator,
        thousands_separator=thousands_separator,
    )
    credit = _parse_number(
        _mapped(row, mapping, "credit"),
        decimal_separator=decimal_separator,
        thousands_separator=thousands_separator,
    )
    if debit is None and credit is None:
        return None
    return (debit or ZERO) - (credit or ZERO)


def _canonical_direction(
    explicit_value: object,
    amount: Decimal,
    direction_value_mapping: dict[str, str],
) -> tuple[str | None, str | None]:
    """Resolve reviewed labels and prove agreement with the signed amount."""

    amount_direction = (
        "positive" if amount > ZERO else "negative" if amount < ZERO else "zero"
    )
    source_label = _norm_label(explicit_value)
    if not source_label:
        return amount_direction, None
    direction = (
        source_label
        if source_label in CANONICAL_DIRECTIONS
        else direction_value_mapping.get(source_label)
    )
    if direction is None:
        return None, "unreviewed_direction_value"
    if direction != amount_direction:
        return None, "direction_amount_mismatch"
    return direction, None


def _generic_period_component_scan(token: str) -> tuple[str, bool, bool, bool]:
    """Strip leading generic period components from one normalized token."""

    remainder = token
    saw_period = False
    saw_year = False
    consumed = False
    period_words = sorted(GENERIC_PERIOD_WORDS, key=len, reverse=True)
    while remainder:
        code_match = GENERIC_PERIOD_CODE_RE.match(remainder)
        if code_match is not None:
            remainder = remainder[code_match.end() :]
            saw_period = True
            consumed = True
            continue
        word = next(
            (value for value in period_words if remainder.startswith(value)),
            None,
        )
        if word is not None:
            remainder = remainder[len(word) :]
            saw_period = True
            consumed = True
            continue
        year_match = re.match(r"(?:19|20)\d{2}", remainder)
        if year_match is not None:
            remainder = remainder[year_match.end() :]
            saw_year = True
            consumed = True
            continue
        short_year_match = re.match(r"\d{2}", remainder)
        if short_year_match is not None and (saw_period or len(remainder) == 2):
            remainder = remainder[short_year_match.end() :]
            saw_year = True
            consumed = True
            continue
        version_match = GENERIC_PERIOD_VERSION_RE.match(remainder)
        if version_match is not None:
            remainder = remainder[version_match.end() :]
            consumed = True
            continue
        break
    return remainder, saw_period, saw_year, consumed


def _is_generic_period_fragment(token: str) -> bool:
    remainder, saw_period, saw_year, consumed = _generic_period_component_scan(token)
    return consumed and remainder == "" and saw_period and saw_year


def _reference_tokens(*values: Any) -> set[str]:
    """Return stable explicit identifiers, excluding generic semantic words."""

    tokens: set[str] = set()
    for value in values:
        text = _norm_label(value)
        parts = re.findall(r"[a-z0-9]+", text)
        if not parts:
            continue
        compact = "".join(parts)
        compact_remainder, compact_has_period, compact_has_year, _ = (
            _generic_period_component_scan(compact)
        )
        part_scans = [_generic_period_component_scan(part) for part in parts]
        compact_period_match = GENERIC_PERIOD_REFERENCE_RE.match(compact)
        contains_year = any(
            re.fullmatch(r"(?:(?:19|20)\d{2}|\d{2})", part) is not None
            for part in parts
        )
        contains_period_word = any(
            part in GENERIC_PERIOD_WORDS or part in REFERENCE_GENERIC_TOKENS or scan[1]
            for part, scan in zip(parts, part_scans, strict=True)
        )
        period_signal = (
            compact_period_match is not None
            or (compact_has_period and compact_has_year)
            or (
                any(scan[1] for scan in part_scans)
                and (any(scan[2] for scan in part_scans) or contains_year)
            )
            or (contains_year and contains_period_word)
        )
        if period_signal:
            residual_parts: list[str] = []
            for part, scan in zip(parts, part_scans, strict=True):
                remainder, saw_period, saw_year, consumed = scan
                if (
                    part in GENERIC_PERIOD_WORDS
                    or part in REFERENCE_GENERIC_TOKENS
                    or re.fullmatch(r"(?:(?:19|20)\d{2}|\d{2})", part) is not None
                    or GENERIC_PERIOD_VERSION_RE.fullmatch(part) is not None
                ):
                    continue
                if consumed and (saw_period or saw_year):
                    if remainder:
                        residual_parts.append(remainder)
                    continue
                part_period_match = GENERIC_PERIOD_REFERENCE_RE.match(part)
                if part_period_match is not None:
                    suffix = part[part_period_match.end() :]
                    if suffix:
                        residual_parts.append(suffix)
                    continue
                residual_parts.append(part)
            residual_candidates = [
                *residual_parts,
                "".join(residual_parts),
            ]
            if compact_period_match is not None:
                suffix = compact[compact_period_match.end() :]
                if suffix:
                    residual_candidates.append(suffix)
            if compact_has_period and compact_has_year and compact_remainder:
                residual_candidates.append(compact_remainder)
            has_distinctive_suffix = any(
                len(candidate) >= 5
                and any(character.isdigit() for character in candidate)
                and GENERIC_PERIOD_VERSION_RE.fullmatch(candidate) is None
                for candidate in residual_candidates
            )
            if not has_distinctive_suffix:
                continue
        generic_period = GENERIC_PERIOD_REFERENCE_RE.fullmatch(compact) is not None
        generic_year_range = (
            len(parts) == 2
            and re.fullmatch(r"(?:19|20)\d{2}", parts[0]) is not None
            and re.fullmatch(r"(?:\d{2}|(?:19|20)\d{2})", parts[1]) is not None
        )
        if generic_period or generic_year_range:
            continue
        candidates = list(parts)
        if parts[0] not in REFERENCE_GENERIC_TOKENS:
            candidates.append(compact)
        for token in candidates:
            if token in REFERENCE_GENERIC_TOKENS or len(token) < 5:
                continue
            if not any(character.isdigit() for character in token):
                continue
            if GENERIC_PERIOD_REFERENCE_RE.fullmatch(token) is not None:
                continue
            if _is_generic_period_fragment(token):
                continue
            if token.isdigit():
                if len(token) < 5:
                    continue
                tokens.add(token)
                continue
            generic_prefix = next(
                (
                    prefix
                    for prefix in REFERENCE_GENERIC_TOKENS
                    if token.startswith(prefix) and token != prefix
                ),
                None,
            )
            if generic_prefix is not None:
                continue
            tokens.add(token)
    return tokens


def _transaction_frame(records: list[dict[str, Any]]) -> pl.DataFrame:
    if not records:
        return pl.DataFrame(schema={col: pl.Utf8 for col in TRANSACTION_COLUMNS})
    frame = pl.DataFrame(records, infer_schema_length=None)
    for col in TRANSACTION_COLUMNS:
        if col not in frame.columns:
            frame = frame.with_columns(pl.lit(None).alias(col))
    return frame.select(TRANSACTION_COLUMNS)


def _non_movement_frame(records: list[dict[str, Any]]) -> pl.DataFrame:
    if not records:
        return pl.DataFrame(schema={col: pl.Utf8 for col in NON_MOVEMENT_COLUMNS})
    frame = pl.DataFrame(records, infer_schema_length=None)
    for col in NON_MOVEMENT_COLUMNS:
        if col not in frame.columns:
            frame = frame.with_columns(pl.lit(None).alias(col))
    return frame.select(NON_MOVEMENT_COLUMNS)


def _non_movement_records(
    diagnostics: Sequence[dict[str, Any]],
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for diag in diagnostics:
        rows = diag.get("non_movement_rows")
        if isinstance(rows, list):
            records.extend(row for row in rows if isinstance(row, dict))
    return records


def _mapping_contract_errors(
    mapping: dict[str, Any], table_columns: Sequence[str]
) -> list[str]:
    errors = _missing_mapping(mapping)
    selected_columns = [
        value for value in _normalized_mapping(mapping).values() if value is not None
    ]
    missing_columns = sorted(set(selected_columns) - set(table_columns))
    errors.extend(f"missing column: {column}" for column in missing_columns)
    if mapping.get("amount") and (mapping.get("debit") or mapping.get("credit")):
        errors.append("amount cannot be combined with debit/credit")
    if mapping.get("debit") and mapping.get("credit"):
        if mapping.get("debit") == mapping.get("credit"):
            errors.append("debit and credit must be different columns")
    return sorted(set(errors))


def _potential_monetary_columns(
    table: pl.DataFrame,
    mapping: dict[str, Any],
    *,
    decimal_separator: str | None,
    thousands_separator: str | None,
) -> list[str]:
    """Return populated columns whose monetary role needs explicit disposition."""

    mapped_non_monetary = {
        column
        for field, column in mapping.items()
        if field not in {"amount", "debit", "credit"} and column
    }
    candidates: list[str] = []
    for column in table.columns:
        if column in {
            SOURCE_ROW_COLUMN,
            SOURCE_SHEET_COLUMN,
        }:
            continue
        explicit_header = any(
            token in _norm_label(column) for token in MONETARY_HEADER_TOKENS
        )
        if column in mapped_non_monetary and not explicit_header:
            continue
        numeric_values = any(
            _parse_number(
                value,
                decimal_separator=decimal_separator,
                thousands_separator=thousands_separator,
            )
            is not None
            for value in table.get_column(column).drop_nulls().to_list()
            if _clean_text(value)
        )
        if explicit_header or numeric_values:
            candidates.append(column)
    return candidates


def _excluded_monetary_columns(
    file_recipe: dict[str, Any],
) -> tuple[list[str], bool]:
    raw = file_recipe.get("excluded_monetary_columns", [])
    if not isinstance(raw, list) or any(
        not isinstance(value, str) or not value.strip() for value in raw
    ):
        return [], False
    normalized = [value.strip() for value in raw]
    return normalized, len(normalized) == len(set(normalized))


def _declared_potential_monetary_columns(
    file_recipe: dict[str, Any],
) -> tuple[list[str], bool]:
    raw = file_recipe.get("potential_monetary_columns", [])
    if not isinstance(raw, list) or any(
        not isinstance(value, str) or not value.strip() for value in raw
    ):
        return [], False
    normalized = [value.strip() for value in raw]
    return normalized, len(normalized) == len(set(normalized))


def _tabular_candidate_count(
    table: pl.DataFrame,
    mapping: dict[str, Any],
    *,
    non_movement_summary_labels: Sequence[str] = (),
) -> int:
    money_columns = [
        str(mapping[field])
        for field in ("amount", "debit", "credit")
        if mapping.get(field)
    ]
    if not money_columns:
        return 0
    return sum(
        (
            any(_clean_text(row.get(column)) for column in money_columns)
            and not _is_reviewed_non_movement_summary(
                row,
                mapping,
                non_movement_summary_labels,
            )
        )
        for row in table.iter_rows(named=True)
    )


def _is_reviewed_non_movement_summary(
    row: dict[str, Any],
    mapping: dict[str, Any],
    labels: Sequence[str],
) -> bool:
    """Apply an exact reviewed summary label only to structurally blank rows."""

    if not labels or _clean_text(_mapped(row, mapping, "date")):
        return False
    stable_reference_tokens = _reference_tokens(
        _clean_text(_mapped(row, mapping, "reference")),
        _clean_text(_mapped(row, mapping, "movement_number")),
    )
    if stable_reference_tokens:
        return False
    description = _norm_label(_mapped(row, mapping, "description"))
    return bool(description and description in labels)


def _date_contract_evidence(
    table: pl.DataFrame,
    mapping: dict[str, Any],
    *,
    date_convention: str | None,
    date_locale: str | None,
    non_movement_summary_labels: Sequence[str],
    allow_excel_serial: bool,
) -> dict[str, Any]:
    """Describe date syntax only for populated monetary candidate rows."""

    counts: Counter[str] = Counter()
    rows: dict[str, list[int]] = {
        "parsed": [],
        "blank": [],
        "ambiguous": [],
        "invalid": [],
    }
    for row in table.iter_rows(named=True):
        source_amount_values = (
            _mapped(row, mapping, "amount"),
            _mapped(row, mapping, "debit"),
            _mapped(row, mapping, "credit"),
        )
        if not any(_clean_text(value) for value in source_amount_values):
            continue
        if _is_reviewed_non_movement_summary(
            row,
            mapping,
            non_movement_summary_labels,
        ):
            continue
        status, _ = _date_parse_result(
            _mapped(row, mapping, "date"),
            date_convention=date_convention,
            date_locale=date_locale,
            allow_excel_serial=allow_excel_serial,
        )
        counts[status] += 1
        rows.setdefault(status, []).append(int(row[SOURCE_ROW_COLUMN]))
    evidence = {
        "date_convention": date_convention,
        "excel_serial_allowed": allow_excel_serial,
        "status_counts": {
            status: int(counts[status])
            for status in ("parsed", "blank", "ambiguous", "invalid")
        },
        "source_rows": rows,
    }
    if counts["locale_required"]:
        evidence["status_counts"]["locale_required"] = int(counts["locale_required"])
    if date_locale is not None:
        evidence["date_locale"] = date_locale
    return evidence


def _validated_mapping_decision(
    file_recipe: dict[str, Any],
    *,
    expected_content: dict[str, Any],
    source_artifact_ref: str,
    expected_adapter_id: str,
    expected_adapter_version: str,
) -> tuple[dict[str, Any] | None, str | None]:
    raw = file_recipe.get("mapping_decision")
    if not isinstance(raw, dict):
        return None, "A reviewed mapping receipt is required for this layout."
    try:
        receipt = validate_reviewed_decision_receipt(
            raw,
            expected_source_artifact_refs=[source_artifact_ref],
            expected_adapter_id=expected_adapter_id,
            expected_adapter_version=expected_adapter_version,
            require_reviewed=True,
        )
    except ValueError as exc:
        return None, f"Mapping receipt is invalid or stale: {exc}"
    if receipt["decision_type"] != "journal_bank_mapping":
        return None, "Mapping receipt has the wrong decision type."
    if receipt["content"] != expected_content:
        return None, "Mapping receipt content does not match the current recipe."
    return receipt, None


def _normalize_table(
    path: Path,
    side: str,
    recipe: dict[str, Any],
    *,
    source_identity: str,
    source_artifact_ref: str,
) -> tuple[pl.DataFrame, dict[str, Any]]:
    file_recipe = _mapping_for_file(
        recipe,
        side,
        path,
        source_identity=source_identity,
    )
    sheet_names = _source_sheet_names(path)
    if len(sheet_names) != 1:
        return _transaction_frame([]), {
            "source_file": source_identity,
            "source_sheet": None,
            "parser": "tabular",
            "adapter_id": TABULAR_ADAPTER_ID,
            "source_family": "tabular.multisheet.unqualified.v1",
            "qualification_status": "unsupported_source_layout",
            "failure_kind": "multiple_sheets_unsupported",
            "candidate_row_count": 0,
            "row_count": 0,
            "preview": [],
            "missing_required_mapping": [],
            "row_disposition_counts": {},
            "row_dispositions": [],
            "workbook_sheets": sheet_names,
            "csv_field_delimiter": None,
            "csv_field_delimiter_origin": "not_applicable",
            "csv_field_delimiter_profile": None,
            "limitations": [
                "The bounded tabular adapter requires exactly one physical sheet; "
                "no sheet was selected or silently omitted."
            ],
        }
    source_sheet = sheet_names[0]
    delimiter_resolution = _resolve_csv_field_delimiter(path, file_recipe)
    if delimiter_resolution["status"] not in {"resolved", "not_applicable"}:
        ambiguous = delimiter_resolution["status"] == "ambiguous"
        return _transaction_frame([]), {
            "source_file": source_identity,
            "source_sheet": source_sheet,
            "parser": "tabular",
            "adapter_id": TABULAR_ADAPTER_ID,
            "source_family": "tabular.csv_delimiter.v1",
            "qualification_status": (
                "needs_review" if ambiguous else "unsupported_source_layout"
            ),
            "failure_kind": (
                "ambiguous_csv_field_delimiter"
                if ambiguous
                else "unsupported_csv_field_delimiter"
            ),
            "csv_field_delimiter": None,
            "requested_csv_field_delimiter": delimiter_resolution[
                "requested_delimiter"
            ],
            "csv_field_delimiter_origin": delimiter_resolution["origin"],
            "csv_field_delimiter_profile": delimiter_resolution["profile"],
            "candidate_row_count": 0,
            "row_count": 0,
            "preview": [],
            "missing_required_mapping": (["csv_field_delimiter"] if ambiguous else []),
            "row_disposition_counts": {},
            "row_dispositions": [],
            "limitations": [
                (
                    "More than one supported CSV field delimiter produced a "
                    "consistent table; a reviewed delimiter is required."
                    if ambiguous
                    else "No supported one-byte CSV field delimiter was selected."
                )
            ],
        }
    csv_field_delimiter = delimiter_resolution["delimiter"]
    raw = _read_table_raw(
        path,
        csv_field_delimiter=(
            str(csv_field_delimiter) if csv_field_delimiter is not None else None
        ),
    )
    exact_contract = _exact_header_contract(raw)
    raw_header_rows = file_recipe.get("header_rows")
    header_rows_shape_valid = "header_rows" not in file_recipe or (
        isinstance(raw_header_rows, list)
        and bool(raw_header_rows)
        and all(
            isinstance(value, int) and not isinstance(value, bool) and value > 0
            for value in raw_header_rows
        )
        and len(raw_header_rows) == len(set(raw_header_rows))
    )
    provided_header_rows = raw_header_rows if header_rows_shape_valid else None
    header_rows = (
        list(provided_header_rows)
        if isinstance(provided_header_rows, list) and provided_header_rows
        else (
            exact_contract[0]
            if exact_contract is not None
            else _suggest_header_rows(raw)
        )
    )
    table = _apply_header(raw, header_rows, source_sheet=source_sheet)
    raw_mapping = file_recipe.get("mapping")
    mapping_shape_valid = "mapping" not in file_recipe or (
        isinstance(raw_mapping, dict)
        and set(raw_mapping) <= set(MAPPING_FIELDS)
        and all(
            value is None or isinstance(value, str) for value in raw_mapping.values()
        )
    )
    provided_mapping = raw_mapping if mapping_shape_valid else None
    exact_mapping = exact_contract[1] if exact_contract is not None else None
    exact_mapping_applies = (
        exact_contract is not None
        and header_rows == exact_contract[0]
        and (
            provided_mapping is None
            or _normalized_mapping(provided_mapping)
            == _normalized_mapping(exact_contract[1])
        )
    )
    mapping = (
        exact_mapping
        if exact_mapping_applies and exact_mapping is not None
        else (
            provided_mapping
            if provided_mapping is not None
            else infer_mapping(table, side)
        )
    )
    mapping = _normalized_mapping(mapping)
    direction_value_mapping, direction_mapping_errors = (
        _normalized_direction_value_mapping(file_recipe.get("direction_value_mapping"))
    )
    date_convention, date_convention_errors = _normalized_date_convention(
        file_recipe.get("date_convention")
    )
    date_locale, date_locale_errors = _normalized_date_locale(
        file_recipe.get("date_locale")
    )
    non_movement_summary_labels, summary_label_errors = (
        _normalized_non_movement_summary_labels(
            file_recipe.get("non_movement_summary_labels")
        )
    )
    tabular_adapter_id, tabular_adapter_version = _tabular_adapter_binding(
        date_locale,
        non_movement_summary_labels,
    )
    observed_direction_values = sorted(
        {
            _norm_label(_mapped(row, mapping, "direction"))
            for row in table.iter_rows(named=True)
            if _norm_label(_mapped(row, mapping, "direction"))
        }
    )
    noncanonical_direction_values = sorted(
        value
        for value in observed_direction_values
        if value not in CANONICAL_DIRECTIONS
    )
    if not mapping.get("direction") and direction_value_mapping:
        direction_mapping_errors.append(
            "direction_value_mapping requires a mapped direction column"
        )
    elif mapping.get("direction") and set(direction_value_mapping) != set(
        noncanonical_direction_values
    ):
        direction_mapping_errors.append(
            "direction_value_mapping must exactly cover every observed "
            "non-canonical direction label"
        )
    direction_mapping_errors = sorted(set(direction_mapping_errors))
    decimal_separator = file_recipe.get("decimal_separator")
    thousands_separator = file_recipe.get("thousands_separator")
    execution_decimal_separator = (
        _clean_text(decimal_separator) or None
        if decimal_separator is not None
        else None
    )
    execution_thousands_separator = (
        _clean_text(thousands_separator) or None
        if thousands_separator is not None
        else None
    )
    ambiguous_numeric_rows = sorted(
        {
            int(row[SOURCE_ROW_COLUMN])
            for row in table.iter_rows(named=True)
            if any(
                _requires_numeric_separator_review(
                    _mapped(row, mapping, field),
                    decimal_separator=execution_decimal_separator,
                    thousands_separator=execution_thousands_separator,
                )
                for field in ("amount", "debit", "credit")
                if mapping.get(field)
            )
        }
    )
    declared_potential_columns, potential_declaration_valid = (
        _declared_potential_monetary_columns(file_recipe)
    )
    excluded_monetary_columns, exclusions_valid = _excluded_monetary_columns(
        file_recipe
    )
    potential_monetary_columns = _potential_monetary_columns(
        table,
        mapping,
        decimal_separator=execution_decimal_separator,
        thousands_separator=execution_thousands_separator,
    )
    mapped_monetary_columns = {
        str(mapping[field])
        for field in ("amount", "debit", "credit")
        if mapping.get(field)
    }
    excluded_set = set(excluded_monetary_columns)
    exclusions_exist = excluded_set <= set(potential_monetary_columns)
    mapped_monetary_columns_have_evidence = mapped_monetary_columns <= set(
        potential_monetary_columns
    )
    mapped_and_excluded_columns = sorted(mapped_monetary_columns & excluded_set)
    unresolved_monetary_columns = sorted(
        set(potential_monetary_columns) - mapped_monetary_columns - excluded_set
    )
    date_evidence = _date_contract_evidence(
        table,
        mapping,
        date_convention=date_convention,
        date_locale=date_locale,
        non_movement_summary_labels=non_movement_summary_labels,
        allow_excel_serial=path.suffix.lower() in {".xls", ".xlsx", ".xlsm"},
    )
    ambiguous_date_rows = date_evidence["source_rows"]["ambiguous"]
    locale_required_date_rows = date_evidence["source_rows"].get(
        "locale_required",
        [],
    )
    review_content = _mapping_review_content(
        side=side,
        source_file=source_identity,
        header_rows=header_rows,
        mapping=mapping,
        direction_value_mapping=direction_value_mapping,
        potential_monetary_columns=potential_monetary_columns,
        excluded_monetary_columns=excluded_monetary_columns,
        date_convention=date_convention,
        date_locale=date_locale,
        non_movement_summary_labels=non_movement_summary_labels,
        csv_field_delimiter=csv_field_delimiter,
        decimal_separator=decimal_separator,
        thousands_separator=thousands_separator,
    )
    mapping_receipt: dict[str, Any] | None = None
    mapping_review_error: str | None = None
    monetary_declarations_match_automatic_evidence = (
        "potential_monetary_columns" not in file_recipe
        and "excluded_monetary_columns" not in file_recipe
    ) or (
        "potential_monetary_columns" in file_recipe
        and "excluded_monetary_columns" in file_recipe
        and potential_declaration_valid
        and exclusions_valid
        and declared_potential_columns == potential_monetary_columns
    )
    automatic_contract_applies = (
        exact_mapping_applies
        and not direction_value_mapping
        and date_convention is None
        and date_locale is None
        and not non_movement_summary_labels
        and not ambiguous_date_rows
        and not locale_required_date_rows
        and not ambiguous_numeric_rows
        and not excluded_monetary_columns
        and not unresolved_monetary_columns
        and not delimiter_resolution["review_required"]
        and decimal_separator is None
        and thousands_separator is None
        and monetary_declarations_match_automatic_evidence
        and file_recipe.get("mapping_decision") is None
    )
    if not automatic_contract_applies:
        mapping_receipt, mapping_review_error = _validated_mapping_decision(
            file_recipe,
            expected_content=review_content,
            source_artifact_ref=source_artifact_ref,
            expected_adapter_id=tabular_adapter_id,
            expected_adapter_version=tabular_adapter_version,
        )
    contract_errors = _mapping_contract_errors(mapping, table.columns)
    recipe_container_error = file_recipe.get(_RECIPE_CONTAINER_ERROR_FIELD)
    if recipe_container_error:
        contract_errors.append(str(recipe_container_error))
    if not header_rows_shape_valid:
        contract_errors.append(
            "header_rows must be a non-empty list of unique positive integers"
        )
    if not mapping_shape_valid:
        contract_errors.append(
            "mapping must be an object containing only supported fields with "
            "string or null column names"
        )
    contract_errors.extend(direction_mapping_errors)
    contract_errors.extend(date_convention_errors)
    contract_errors.extend(date_locale_errors)
    contract_errors.extend(summary_label_errors)
    if ambiguous_date_rows and date_convention is None:
        contract_errors.append(
            "ambiguous day/month dates require a reviewed date_convention"
        )
    if locale_required_date_rows and date_locale is None:
        contract_errors.append(
            "localized textual-month dates require a reviewed date_locale"
        )
    if ambiguous_numeric_rows:
        contract_errors.append(
            "ambiguous numeric separators require reviewed decimal_separator "
            "and thousands_separator values"
        )
    if not automatic_contract_applies:
        if "potential_monetary_columns" not in file_recipe:
            contract_errors.append(
                "potential_monetary_columns must be explicitly reviewed"
            )
        elif not potential_declaration_valid:
            contract_errors.append(
                "potential_monetary_columns must contain unique non-empty "
                "column names"
            )
        elif declared_potential_columns != potential_monetary_columns:
            contract_errors.append(
                "potential_monetary_columns does not match current source evidence"
            )
        if "excluded_monetary_columns" not in file_recipe:
            contract_errors.append(
                "excluded_monetary_columns must be explicitly reviewed, even "
                "when empty"
            )
    if not exclusions_valid:
        contract_errors.append(
            "excluded_monetary_columns must contain unique non-empty column names"
        )
    if not exclusions_exist:
        contract_errors.append(
            "excluded_monetary_columns contains a column without monetary evidence"
        )
    if not mapped_monetary_columns_have_evidence:
        contract_errors.append(
            "a mapped monetary column lacks current monetary evidence"
        )
    contract_errors.extend(
        f"mapped monetary column is also excluded: {column}"
        for column in mapped_and_excluded_columns
    )
    contract_errors.extend(
        f"unresolved monetary column: {column}"
        for column in unresolved_monetary_columns
    )
    contract_errors = sorted(set(contract_errors))
    if mapping_review_error is not None or contract_errors:
        limitations = [
            value
            for value in (
                mapping_review_error,
                *contract_errors,
            )
            if value
        ]
        return _transaction_frame([]), {
            "source_file": source_identity,
            "source_sheet": source_sheet,
            "parser": "tabular",
            "adapter_id": tabular_adapter_id,
            "source_family": "tabular.explicit_columns.v1",
            "qualification_status": "needs_review",
            "failure_kind": "mapping_review_required",
            "header_rows": header_rows,
            "mapping": mapping,
            "direction_value_mapping": direction_value_mapping,
            "observed_direction_values": observed_direction_values,
            "date_convention": date_convention,
            **({"date_locale": date_locale} if date_locale is not None else {}),
            **(
                {"non_movement_summary_labels": non_movement_summary_labels}
                if non_movement_summary_labels
                else {}
            ),
            "date_interpretation_evidence": date_evidence,
            "ambiguous_numeric_rows": ambiguous_numeric_rows,
            "mapping_origin": (
                "reviewed_recipe" if provided_mapping is not None else "proposal"
            ),
            "mapping_review_content": review_content,
            "mapping_review_content_sha256": canonical_json_sha256(review_content),
            "reviewed_mapping_ref": None,
            "mapping_decision": None,
            "csv_field_delimiter": csv_field_delimiter,
            "csv_field_delimiter_origin": delimiter_resolution["origin"],
            "csv_field_delimiter_profile": delimiter_resolution["profile"],
            "raw_columns": [
                column
                for column in table.columns
                if column not in {SOURCE_ROW_COLUMN, SOURCE_SHEET_COLUMN}
            ],
            "potential_monetary_columns": potential_monetary_columns,
            "excluded_monetary_columns": excluded_monetary_columns,
            "unresolved_monetary_columns": unresolved_monetary_columns,
            "candidate_row_count": _tabular_candidate_count(
                table,
                mapping,
                non_movement_summary_labels=non_movement_summary_labels,
            ),
            "row_count": 0,
            "preview": [],
            "missing_required_mapping": contract_errors,
            "row_disposition_counts": {},
            "row_dispositions": [],
            "limitations": limitations,
        }

    records: list[dict[str, Any]] = []
    row_dispositions: list[dict[str, Any]] = []
    disposition_counts: Counter[str] = Counter()
    invalid_candidate_rows: list[int] = []
    monetary_candidate_count = 0
    for row in table.iter_rows(named=True):
        source_row = int(row[SOURCE_ROW_COLUMN])
        amount = _amount_from_row(
            row,
            mapping,
            decimal_separator=execution_decimal_separator,
            thousands_separator=execution_thousands_separator,
        )
        date_status, parsed_date = _date_parse_result(
            _mapped(row, mapping, "date"),
            date_convention=date_convention,
            date_locale=date_locale,
            allow_excel_serial=path.suffix.lower() in {".xls", ".xlsx", ".xlsm"},
        )
        description = _clean_text(_mapped(row, mapping, "description"))
        beneficiary = _clean_text(_mapped(row, mapping, "beneficiary"))
        reference = _clean_text(_mapped(row, mapping, "reference"))
        movement = _clean_text(_mapped(row, mapping, "movement_number"))
        stable_reference_tokens = _reference_tokens(reference, movement)
        account = _clean_text(_mapped(row, mapping, "account"))
        currency = _clean_text(_mapped(row, mapping, "currency"))
        unit = _clean_text(_mapped(row, mapping, "unit"))
        entity_ref = _clean_text(_mapped(row, mapping, "entity_ref"))
        party_ref = _clean_text(_mapped(row, mapping, "party_ref"))
        explicit_direction = _clean_text(_mapped(row, mapping, "direction"))
        source_amount_values = (
            _mapped(row, mapping, "amount"),
            _mapped(row, mapping, "debit"),
            _mapped(row, mapping, "credit"),
        )
        has_monetary_value = any(_clean_text(value) for value in source_amount_values)
        locator = {"source_sheet": source_sheet, "source_row": source_row}
        if not has_monetary_value:
            disposition_counts["excluded_non_monetary"] += 1
            row_dispositions.append(
                {
                    **locator,
                    "status": "excluded_non_monetary",
                    "reason": "No mapped monetary cell is populated.",
                }
            )
            continue
        if _is_reviewed_non_movement_summary(
            row,
            mapping,
            non_movement_summary_labels,
        ):
            disposition_counts["excluded_reviewed_summary"] += 1
            row_dispositions.append(
                {
                    **locator,
                    "status": "excluded_reviewed_summary",
                    "reason": (
                        "A blank-date row with no stable reference matched an "
                        "exact reviewed non-movement summary label."
                    ),
                }
            )
            continue
        monetary_candidate_count += 1
        if amount is None:
            invalid_candidate_rows.append(source_row)
            disposition_counts["invalid_monetary_value"] += 1
            row_dispositions.append(
                {
                    **locator,
                    "status": "invalid_monetary_value",
                    "reason": "A populated mapped monetary cell did not parse exactly.",
                }
            )
            continue
        direction, direction_error = _canonical_direction(
            explicit_direction,
            amount,
            direction_value_mapping,
        )
        if direction_error is not None:
            invalid_candidate_rows.append(source_row)
            disposition_counts[direction_error] += 1
            row_dispositions.append(
                {
                    **locator,
                    "status": direction_error,
                    "reason": (
                        "The reviewed source direction does not agree with the "
                        "exact signed amount."
                        if direction_error == "direction_amount_mismatch"
                        else "A source direction label lacks a reviewed canonical mapping."
                    ),
                }
            )
            continue
        if date_status in {"ambiguous", "invalid", "locale_required"}:
            invalid_candidate_rows.append(source_row)
            date_disposition = {
                "ambiguous": "ambiguous_date_without_review",
                "locale_required": "localized_date_without_review",
                "invalid": "invalid_date_value",
            }[date_status]
            disposition_counts[date_disposition] += 1
            row_dispositions.append(
                {
                    **locator,
                    "status": date_disposition,
                    "reason": (
                        "A populated day/month date has more than one valid "
                        "interpretation and lacks a reviewed convention."
                        if date_status == "ambiguous"
                        else (
                            "A populated textual-month date lacks a reviewed locale."
                            if date_status == "locale_required"
                            else "A populated mapped date is not a valid date under "
                            "the current mechanical or reviewed convention."
                        )
                    ),
                }
            )
            continue
        if date_status == "blank" and not stable_reference_tokens:
            invalid_candidate_rows.append(source_row)
            disposition_counts["missing_date_without_stable_reference"] += 1
            row_dispositions.append(
                {
                    **locator,
                    "status": "missing_date_without_stable_reference",
                    "reason": (
                        "A monetary row without an actual date requires a stable "
                        "non-generic explicit identifier."
                    ),
                }
            )
            continue
        disposition = "emitted_reference_only" if date_status == "blank" else "emitted"
        disposition_counts[disposition] += 1
        row_dispositions.append(
            {
                **locator,
                "status": disposition,
                "reason": (
                    "Missing date; eligible only for explicit-reference matching."
                    if date_status == "blank"
                    else "Mapped monetary row emitted."
                ),
            }
        )
        transaction_id = (
            f"{side}:{_source_identity_fragment(source_identity)}:"
            f"{_identifier_fragment(source_sheet)}:{source_row}"
        )
        records.append(
            {
                "side": side,
                "transaction_id": transaction_id,
                "transaction_date": parsed_date.isoformat() if parsed_date else None,
                "amount_signed": decimal_text(amount) if amount is not None else None,
                "amount_abs": (
                    decimal_text(abs(amount)) if amount is not None else None
                ),
                "description": description or None,
                "beneficiary": beneficiary or None,
                "reference": reference or None,
                "movement_number": movement or None,
                "account": account or None,
                "currency": currency or None,
                "unit": unit or None,
                "entity_ref": entity_ref or None,
                "party_ref": party_ref or None,
                "direction": direction,
                "source_file": source_identity,
                "source_sheet": source_sheet,
                "source_row": source_row,
            }
        )
    if invalid_candidate_rows:
        return _transaction_frame([]), {
            "source_file": source_identity,
            "source_sheet": source_sheet,
            "parser": "tabular",
            "adapter_id": tabular_adapter_id,
            "source_family": "tabular.explicit_columns.v1",
            "qualification_status": "unsupported_source_layout",
            "failure_kind": "candidate_row_contract_failed",
            "header_rows": header_rows,
            "mapping": mapping,
            "direction_value_mapping": direction_value_mapping,
            "observed_direction_values": observed_direction_values,
            "date_convention": date_convention,
            **({"date_locale": date_locale} if date_locale is not None else {}),
            **(
                {"non_movement_summary_labels": non_movement_summary_labels}
                if non_movement_summary_labels
                else {}
            ),
            "date_interpretation_evidence": date_evidence,
            "ambiguous_numeric_rows": ambiguous_numeric_rows,
            "mapping_origin": (
                "bounded_exact_headers"
                if automatic_contract_applies
                else "reviewed_recipe"
            ),
            "mapping_review_content": review_content,
            "mapping_review_content_sha256": canonical_json_sha256(review_content),
            "reviewed_mapping_ref": (
                str(mapping_receipt["decision_id"]) if mapping_receipt else None
            ),
            "mapping_decision": mapping_receipt,
            "csv_field_delimiter": csv_field_delimiter,
            "csv_field_delimiter_origin": delimiter_resolution["origin"],
            "csv_field_delimiter_profile": delimiter_resolution["profile"],
            "raw_columns": [
                column
                for column in table.columns
                if column not in {SOURCE_ROW_COLUMN, SOURCE_SHEET_COLUMN}
            ],
            "potential_monetary_columns": potential_monetary_columns,
            "excluded_monetary_columns": excluded_monetary_columns,
            "unresolved_monetary_columns": unresolved_monetary_columns,
            "candidate_row_count": monetary_candidate_count,
            "row_count": 0,
            "preview": [],
            "missing_required_mapping": [],
            "invalid_candidate_rows": invalid_candidate_rows,
            "row_disposition_counts": dict(sorted(disposition_counts.items())),
            "row_dispositions": row_dispositions,
            "limitations": [
                "At least one monetary candidate failed exact amount, date, or "
                "reviewed-direction disposition; no rows were emitted."
            ],
        }
    frame = _transaction_frame(records)
    diagnostics = {
        "source_file": source_identity,
        "source_sheet": source_sheet,
        "parser": "tabular",
        "adapter_id": tabular_adapter_id,
        "source_family": "tabular.explicit_columns.v1",
        "qualification_status": "qualified",
        "failure_kind": None,
        "header_rows": header_rows,
        "mapping": mapping,
        "direction_value_mapping": direction_value_mapping,
        "observed_direction_values": observed_direction_values,
        "date_convention": date_convention,
        **({"date_locale": date_locale} if date_locale is not None else {}),
        **(
            {"non_movement_summary_labels": non_movement_summary_labels}
            if non_movement_summary_labels
            else {}
        ),
        "date_interpretation_evidence": date_evidence,
        "ambiguous_numeric_rows": ambiguous_numeric_rows,
        "mapping_origin": (
            "bounded_exact_headers" if automatic_contract_applies else "reviewed_recipe"
        ),
        "mapping_review_content": review_content,
        "mapping_review_content_sha256": canonical_json_sha256(review_content),
        "reviewed_mapping_ref": (
            str(mapping_receipt["decision_id"]) if mapping_receipt else None
        ),
        "mapping_decision": mapping_receipt,
        "csv_field_delimiter": csv_field_delimiter,
        "csv_field_delimiter_origin": delimiter_resolution["origin"],
        "csv_field_delimiter_profile": delimiter_resolution["profile"],
        "raw_columns": [
            column
            for column in table.columns
            if column not in {SOURCE_ROW_COLUMN, SOURCE_SHEET_COLUMN}
        ],
        "potential_monetary_columns": potential_monetary_columns,
        "excluded_monetary_columns": excluded_monetary_columns,
        "unresolved_monetary_columns": unresolved_monetary_columns,
        "candidate_row_count": monetary_candidate_count,
        "row_count": frame.height,
        "preview": frame.head(20).to_dicts(),
        "missing_required_mapping": [],
        "row_disposition_counts": dict(sorted(disposition_counts.items())),
        "row_dispositions": row_dispositions,
        "limitations": (
            [
                "Rows without actual dates are emitted only for explicit-reference matching."
            ]
            if disposition_counts["emitted_reference_only"]
            else []
        ),
    }
    return frame, diagnostics


def _extract_pdf_text(path: Path) -> str:
    import pdfplumber

    lines: list[str] = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            lines.extend(line.strip() for line in text.splitlines() if line.strip())
    return "\n".join(lines)


def _classify_bank_pdf_non_movement_line(line: str) -> tuple[str, str] | None:
    """Classify mechanically identifiable bank statement summary rows.

    These deterministic rules are intentionally narrow: they remove only explicit
    balance, total, scalare, and conditions lines that are statement metadata,
    while leaving ambiguous fee/payment descriptions in the matching population.
    """

    label = _norm_label(line)
    for classification, reason, pattern in BANK_PDF_NON_MOVEMENT_PATTERNS:
        if pattern.search(label):
            return classification, reason
    return None


def _count_classifications(rows: Sequence[dict[str, Any]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in rows:
        classification = _clean_text(row.get("classification"))
        if classification:
            counts[classification] = counts.get(classification, 0) + 1
    return counts


def _normalize_text_pdf(
    path: Path,
    side: str,
    recipe: dict[str, Any],
    *,
    source_identity: str,
) -> tuple[pl.DataFrame, dict[str, Any]]:
    """Inspect PDF text without emitting movements from an unbounded layout.

    A line containing a date and amount is not enough evidence of a transaction
    boundary, amount side, or statement layout. No generic PDF movement adapter
    is registered; future adapters must be source-family-specific and reviewed.
    The narrow non-movement classifications remain useful review evidence.
    """

    text = _extract_pdf_text(path)
    non_movement_rows: list[dict[str, Any]] = []
    candidate_movement_rows = 0
    for line_idx, line in enumerate(text.splitlines(), start=1):
        parsed_date = _parse_date(line)
        amounts = [
            _parse_number(match.group(0)) for match in AMOUNT_TOKEN_RE.finditer(line)
        ]
        amounts = [value for value in amounts if value is not None]
        if not amounts:
            continue
        amount = amounts[-1]
        classification = (
            _classify_bank_pdf_non_movement_line(line) if side == "bank" else None
        )
        if classification is not None:
            class_name, reason = classification
            non_movement_rows.append(
                {
                    "side": side,
                    "source_file": source_identity,
                    "source_sheet": "PDF",
                    "source_row": line_idx,
                    "classification": class_name,
                    "reason": reason,
                    "transaction_date": (
                        parsed_date.isoformat() if parsed_date else None
                    ),
                    "amount_signed": decimal_text(amount),
                    "amount_abs": decimal_text(abs(amount)),
                    "description": line,
                }
            )
            continue
        candidate_movement_rows += 1
    frame = _transaction_frame([])
    file_recipe = _mapping_for_file(
        recipe,
        side,
        path,
        source_identity=source_identity,
    )
    requested_adapter = _clean_text(file_recipe.get("pdf_adapter"))
    return frame, {
        "source_file": source_identity,
        "source_sheet": "PDF",
        "parser": "text_pdf_disabled",
        "adapter_id": TEXT_PDF_ADAPTER_ID,
        "source_family": "text_pdf.unqualified.v1",
        "qualification_status": "unsupported_source_layout",
        "candidate_row_count": candidate_movement_rows,
        "row_count": 0,
        "excluded_non_movement_row_count": len(non_movement_rows),
        "non_movement_classifications": _count_classifications(non_movement_rows),
        "non_movement_rows": non_movement_rows,
        "preview": [],
        "missing_required_mapping": ["supported reviewed PDF layout adapter"],
        "requested_pdf_adapter": requested_adapter or None,
        "limitations": [
            "Generic text-PDF movement extraction is disabled because row boundaries and amount-side meaning are unproven.",
            (
                f"Requested adapter {requested_adapter!r} is not registered."
                if requested_adapter
                else "No supported reviewed PDF layout adapter was supplied."
            ),
        ],
    }


def _parser_failure_diagnostic(
    path: Path,
    *,
    source_identity: str,
    error: BaseException,
) -> dict[str, Any]:
    return {
        "source_file": source_identity,
        "source_sheet": None,
        "parser": "failed",
        "adapter_id": TABULAR_ADAPTER_ID,
        "source_family": "parser_failure.v1",
        "qualification_status": "unsupported_source_layout",
        "failure_kind": "parser_failure",
        "candidate_row_count": 0,
        "row_count": 0,
        "preview": [],
        "missing_required_mapping": [],
        "row_disposition_counts": {},
        "row_dispositions": [],
        "csv_field_delimiter": None,
        "csv_field_delimiter_origin": "unresolved",
        "csv_field_delimiter_profile": None,
        "parser_error": f"{type(error).__name__}: {error}",
        "limitations": [
            "The source could not be parsed; parser failure is distinct from an unsupported readable layout."
        ],
    }


def _normalize_files(
    input_path: Path,
    side: str,
    recipe: dict[str, Any],
    *,
    source_refs: dict[tuple[str, str], str],
) -> tuple[pl.DataFrame, list[dict[str, Any]]]:
    frames: list[pl.DataFrame] = []
    diagnostics: list[dict[str, Any]] = []
    for file_path in supported_files(input_path):
        source_identity = _source_identity(input_path, file_path)
        try:
            if file_path.suffix.lower() == ".pdf":
                frame, diag = _normalize_text_pdf(
                    file_path,
                    side,
                    recipe,
                    source_identity=source_identity,
                )
            else:
                frame, diag = _normalize_table(
                    file_path,
                    side,
                    recipe,
                    source_identity=source_identity,
                    source_artifact_ref=source_refs[(side, source_identity)],
                )
        except (
            fastexcel.CalamineError,
            BadZipFile,
            InvalidFileException,
            pl.exceptions.PolarsError,
            UnicodeError,
            ValueError,
            RuntimeError,
            OSError,
        ) as exc:
            frame = _transaction_frame([])
            diag = _parser_failure_diagnostic(
                file_path,
                source_identity=source_identity,
                error=exc,
            )
        frames.append(frame)
        diagnostics.append(diag)
    if not frames:
        return _transaction_frame([]), diagnostics
    return pl.concat(frames, how="diagonal_relaxed"), diagnostics


def _source_artifact_receipts(
    inputs: Sequence[tuple[str, Path | None]],
) -> tuple[list[dict[str, Any]], dict[tuple[str, str], str]]:
    """Create stable source-byte receipts used by qualifications and lineage."""

    receipts: list[dict[str, Any]] = []
    refs: dict[tuple[str, str], str] = {}
    for side, input_path in inputs:
        if input_path is None:
            continue
        files = supported_files(input_path)
        root = _input_root(input_path)
        for index, source in enumerate(files, start=1):
            source_identity = _source_identity(input_path, source)
            _, source_sha256 = file_snapshot(source)
            artifact_id = f"source.{side}.{index}.{source_sha256}"
            receipt = artifact_receipt(
                root,
                source,
                artifact_id=artifact_id,
                root_id=f"source_{side}",
                role="source",
            )
            if receipt["sha256"] != source_sha256:
                raise ValueError(f"Source changed while receipt was built: {source}")
            receipts.append(receipt)
            refs[(side, source_identity)] = artifact_id
    return receipts, refs


def _source_snapshot_changes(
    before: Sequence[dict[str, Any]],
    after: Sequence[dict[str, Any]],
) -> set[tuple[str, str]]:
    def indexed(
        receipts: Sequence[dict[str, Any]],
    ) -> dict[tuple[str, str], tuple[int, str]]:
        return {
            (str(item["root_id"]), str(item["path"])): (
                int(item["byte_count"]),
                str(item["sha256"]),
            )
            for item in receipts
        }

    before_by_source = indexed(before)
    after_by_source = indexed(after)
    keys = set(before_by_source) | set(after_by_source)
    return {
        key for key in keys if before_by_source.get(key) != after_by_source.get(key)
    }


def _mark_changed_source_diagnostics(
    side: str,
    diagnostics: Sequence[dict[str, Any]],
    changed_sources: set[tuple[str, str]],
) -> list[dict[str, Any]]:
    source_root_changed = any(
        root_id == f"source_{side}" for root_id, _ in changed_sources
    )
    if not source_root_changed:
        return list(diagnostics)
    updated: list[dict[str, Any]] = []
    for diagnostic in diagnostics:
        updated.append(
            {
                **diagnostic,
                "qualification_status": "unsupported_source_layout",
                "failure_kind": "source_changed_during_run",
                "row_count": 0,
                "preview": [],
                "mapping_decision": None,
                "reviewed_mapping_ref": None,
                "limitations": [
                    *list(diagnostic.get("limitations") or []),
                    "Source bytes or membership within the affected source root "
                    "changed between the pre-parser and post-parser snapshots; "
                    "all prepared rows from that root were discarded.",
                ],
            }
        )
    return updated


def _mark_changed_sample_diagnostics(
    diagnostics: Sequence[dict[str, Any]],
    changed_sources: set[tuple[str, str]],
) -> list[dict[str, Any]]:
    changed_paths = {
        source_path
        for root_id, source_path in changed_sources
        if root_id == "source_sample"
    }
    updated: list[dict[str, Any]] = []
    for diagnostic in diagnostics:
        source_file = str(diagnostic["source_file"])
        if source_file not in changed_paths:
            updated.append(diagnostic)
            continue
        updated.append(
            {
                **diagnostic,
                "status": "invalid",
                "failure_kind": "source_changed_during_run",
                "movement_count": 0,
                "limitations": [
                    "Sample bytes or sample membership changed between the "
                    "pre-parser and post-parser snapshots."
                ],
            }
        )
    return updated


def _control(
    control_id: str,
    status: str,
    detail: str,
    *,
    evidence_refs: Sequence[str],
) -> dict[str, Any]:
    return {
        "control_id": control_id,
        "required": True,
        "status": status,
        "evidence_refs": list(evidence_refs),
        "detail": detail,
    }


def _source_qualifications(
    side: str,
    diagnostics: Sequence[dict[str, Any]],
    source_refs: dict[tuple[str, str], str],
) -> list[dict[str, Any]]:
    qualifications: list[dict[str, Any]] = []
    for index, diagnostic in enumerate(diagnostics, start=1):
        source_file = str(diagnostic["source_file"])
        artifact_ref = source_refs.get((side, source_file))
        if artifact_ref is None:
            continue
        status = str(
            diagnostic.get("qualification_status") or "unsupported_source_layout"
        )
        adapter_id = str(diagnostic.get("adapter_id") or TABULAR_ADAPTER_ID)
        if adapter_id == TABULAR_ADAPTER_ID:
            adapter_version = TABULAR_ADAPTER_VERSION
        elif adapter_id == EXTENDED_TABULAR_ADAPTER_ID:
            adapter_version = EXTENDED_TABULAR_ADAPTER_VERSION
        elif adapter_id == TEXT_PDF_ADAPTER_ID:
            adapter_version = TEXT_PDF_ADAPTER_VERSION
        else:
            raise ValueError(f"Unsupported Journal-Bank adapter: {adapter_id}")
        supported = status == "qualified"
        needs_review = status == "needs_review"
        control_status = (
            "passed" if supported else "not_assessed" if needs_review else "failed"
        )
        qualification = build_source_qualification(
            qualification_id=f"qualification.{side}.{index}",
            adapter_id=adapter_id,
            adapter_version=adapter_version,
            source_family=str(
                diagnostic.get("source_family") or "tabular.explicit_columns.v1"
            ),
            status=status,
            source_artifact_refs=[artifact_ref],
            candidate_row_count=int(diagnostic.get("candidate_row_count") or 0),
            emitted_row_count=int(diagnostic.get("row_count") or 0),
            controls=[
                _control(
                    "bounded_adapter",
                    control_status,
                    (
                        "A bounded tabular adapter identified explicit transaction columns."
                        if supported
                        else (
                            "A reviewed mapping is required before the bounded adapter can run."
                            if needs_review
                            else "No supported bounded adapter can emit movement rows from this layout."
                        )
                    ),
                    evidence_refs=[artifact_ref],
                ),
                _control(
                    "required_fields",
                    control_status,
                    (
                        "Date and exact monetary source fields are mapped."
                        if supported
                        else (
                            "Date and monetary ownership remain subject to mapping review."
                            if needs_review
                            else "Required transaction fields are not proven for this layout."
                        )
                    ),
                    evidence_refs=[artifact_ref],
                ),
                _control(
                    "monetary_field_disposition",
                    control_status,
                    (
                        "Every populated monetary candidate column is mapped or "
                        "explicitly excluded by the reviewed mapping."
                        if supported
                        else (
                            "One or more populated monetary candidate columns still "
                            "need an explicit mapped or excluded disposition."
                            if needs_review
                            else "Monetary column ownership is not proven for this layout."
                        )
                    ),
                    evidence_refs=[artifact_ref],
                ),
            ],
            reviewed_mapping_ref=(
                str(diagnostic["reviewed_mapping_ref"])
                if diagnostic.get("reviewed_mapping_ref")
                else None
            ),
            limitations=list(diagnostic.get("limitations") or []),
        )
        qualifications.append(validate_source_qualification(qualification))
    return qualifications


def _aggregate_qualification_status(
    qualifications: Sequence[dict[str, Any]],
) -> str:
    statuses = {str(item["status"]) for item in qualifications}
    if not qualifications or "unsupported_source_layout" in statuses:
        return "unsupported_source_layout"
    if "needs_review" in statuses:
        return "needs_review"
    return "qualified"


def _source_outcomes(
    bank_diagnostics: Sequence[dict[str, Any]],
    journal_diagnostics: Sequence[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Project the frozen source-outcome matrix without altering shared records."""

    outcomes: list[dict[str, Any]] = []
    for side, diagnostics in (
        ("bank", bank_diagnostics),
        ("journal", journal_diagnostics),
    ):
        statuses = [
            str(diagnostic.get("qualification_status") or "unsupported_source_layout")
            for diagnostic in diagnostics
        ]
        if "unsupported_source_layout" in statuses:
            qualification_status = "unsupported_source_layout"
        elif "needs_review" in statuses:
            qualification_status = "needs_review"
        else:
            qualification_status = "qualified"
        failure_kinds = sorted(
            {
                str(diagnostic["failure_kind"])
                for diagnostic in diagnostics
                if diagnostic.get("failure_kind")
            }
        )
        outcomes.append(
            {
                "side": side,
                "qualification_status": qualification_status,
                "failure_kind": (
                    "parser_failure"
                    if "parser_failure" in failure_kinds
                    else failure_kinds[0] if failure_kinds else None
                ),
                "emitted_row_count": sum(
                    int(diagnostic.get("row_count") or 0) for diagnostic in diagnostics
                ),
            }
        )
    return outcomes


def _mapping_decisions(
    diagnostics: Sequence[dict[str, Any]],
) -> list[dict[str, Any]]:
    decisions = [
        item["mapping_decision"]
        for item in diagnostics
        if isinstance(item.get("mapping_decision"), dict)
    ]
    decision_ids = [str(item["decision_id"]) for item in decisions]
    if len(decision_ids) != len(set(decision_ids)):
        raise ValueError("mapping decision IDs must be unique")
    return decisions


def _lineage_payload(
    bank: pl.DataFrame,
    journal: pl.DataFrame,
    source_refs: dict[tuple[str, str], str],
) -> dict[str, Any]:
    entries: list[dict[str, Any]] = []
    for side, frame in (("bank", bank), ("journal", journal)):
        for row in frame.to_dicts():
            source_file = str(row["source_file"])
            source_sheet = _clean_text(row.get("source_sheet"))
            locator = (
                f"sheet:{source_sheet};row:{row['source_row']}"
                if source_sheet and source_sheet != "CSV"
                else f"row:{row['source_row']}"
            )
            entries.append(
                {
                    "prepared_id": str(row["transaction_id"]),
                    "side": side,
                    "source_artifact_ref": source_refs[(side, source_file)],
                    "source_locator": locator,
                }
            )
    return {
        "schema_version": LINEAGE_SCHEMA_VERSION,
        "normalization_schema_version": NORMALIZATION_SCHEMA_VERSION,
        "entries": entries,
    }


def _gate_payload(
    *,
    source_status: str,
    source_refs: Sequence[str],
    preparation_status: str,
    reconciliation_status: str,
    reconciliation_ref: str | None = None,
    semantic_status: str = "not_assessed",
    reporting_status: str = "blocked",
    limitation: str | None = None,
) -> dict[str, Any]:
    limits = [limitation] if limitation else []
    return build_gate_register(
        {
            "source": {
                "status": source_status,
                "evidence_refs": list(source_refs) if source_status == "passed" else [],
                "limitations": [] if source_status == "passed" else limits,
            },
            "preparation": {
                "status": preparation_status,
                "evidence_refs": (
                    ["output.lineage_json"] if preparation_status == "passed" else []
                ),
                "limitations": [] if preparation_status == "passed" else limits,
            },
            "reconciliation": {
                "status": reconciliation_status,
                "evidence_refs": (
                    [reconciliation_ref or "output.relationship_ledger_json"]
                    if reconciliation_status == "passed"
                    else []
                ),
                "limitations": [] if reconciliation_status == "passed" else limits,
            },
            "semantic_review": {
                "status": semantic_status,
                "evidence_refs": (
                    ["review.applied"] if semantic_status == "passed" else []
                ),
                "limitations": [
                    "Professional review remains outside deterministic matching."
                ],
            },
            "reporting": {
                "status": reporting_status,
                "evidence_refs": (
                    ["artifact.receipts"] if reporting_status == "passed" else []
                ),
                "limitations": (
                    []
                    if reporting_status == "passed"
                    else ["Reporting is withheld pending semantic review."]
                ),
            },
            "publication": {
                "status": "blocked",
                "evidence_refs": [],
                "limitations": ["Publication is outside this component."],
            },
        }
    )


def _missing_mapping(mapping: dict[str, Any]) -> list[str]:
    missing = []
    if not mapping.get("date"):
        missing.append("date")
    if not (mapping.get("amount") or mapping.get("debit") or mapping.get("credit")):
        missing.append("amount or debit/credit")
    return missing


def _recipe_side(diagnostics: list[dict[str, Any]]) -> dict[str, Any]:
    files = {}
    for diag in diagnostics:
        entry: dict[str, Any] = {"parser": diag.get("parser")}
        if diag.get("header_rows"):
            entry["header_rows"] = diag["header_rows"]
        if diag.get("mapping"):
            entry["mapping"] = diag["mapping"]
        entry["direction_value_mapping"] = diag.get("direction_value_mapping", {})
        entry["observed_direction_values"] = diag.get("observed_direction_values", [])
        entry["date_convention"] = diag.get("date_convention")
        date_evidence = diag.get(
            "date_interpretation_evidence",
            {
                "date_convention": None,
                "status_counts": {},
                "source_rows": {},
            },
        )
        entry["date_interpretation_evidence"] = date_evidence
        if diag.get("date_locale") is not None or (
            isinstance(date_evidence, dict)
            and int(
                (date_evidence.get("status_counts") or {}).get(
                    "locale_required",
                    0,
                )
            )
            > 0
        ):
            entry["date_locale"] = diag.get("date_locale")
        if diag.get("non_movement_summary_labels"):
            entry["non_movement_summary_labels"] = diag["non_movement_summary_labels"]
        entry["mapping_origin"] = diag.get("mapping_origin")
        entry["mapping_review_content"] = diag.get("mapping_review_content")
        entry["mapping_review_content_sha256"] = diag.get(
            "mapping_review_content_sha256"
        )
        entry["mapping_decision"] = diag.get("mapping_decision")
        entry["csv_field_delimiter"] = diag.get("csv_field_delimiter")
        entry["csv_field_delimiter_origin"] = diag.get("csv_field_delimiter_origin")
        entry["csv_field_delimiter_profile"] = diag.get("csv_field_delimiter_profile")
        entry["potential_monetary_columns"] = diag.get("potential_monetary_columns", [])
        entry["excluded_monetary_columns"] = diag.get("excluded_monetary_columns", [])
        entry["unresolved_monetary_columns"] = diag.get(
            "unresolved_monetary_columns", []
        )
        files[str(diag["source_file"])] = entry
    return {"files": files}


def _read_sample_movements(
    sample_path: Path | None,
) -> tuple[set[str], list[dict[str, Any]]]:
    if sample_path is None:
        return set(), []
    values: set[str] = set()
    diagnostics: list[dict[str, Any]] = []
    for source in supported_files(sample_path):
        source_identity = _source_identity(sample_path, source)
        if source.suffix.lower() == ".pdf":
            diagnostics.append(
                {
                    "source_file": source_identity,
                    "status": "invalid",
                    "failure_kind": "unsupported_sample_layout",
                    "movement_count": 0,
                    "csv_field_delimiter": None,
                    "csv_field_delimiter_profile": None,
                }
            )
            continue
        delimiter_resolution: dict[str, Any] | None = None
        try:
            sheet_names = _source_sheet_names(source)
            if len(sheet_names) != 1:
                diagnostics.append(
                    {
                        "source_file": source_identity,
                        "status": "invalid",
                        "failure_kind": "multiple_sheets_unsupported",
                        "movement_count": 0,
                        "workbook_sheets": sheet_names,
                        "csv_field_delimiter": None,
                        "csv_field_delimiter_profile": None,
                    }
                )
                continue
            delimiter_resolution = _resolve_csv_field_delimiter(source, {})
            # A one-column sample has no field boundary to interpret. Using the
            # comma default is therefore transport-only and cannot promote
            # transaction rows or infer a professional mapping.
            single_column_default = (
                source.suffix.lower() == ".csv"
                and delimiter_resolution["status"] == "unsupported"
                and delimiter_resolution["profile"]["candidate_delimiters"] == []
            )
            if (
                delimiter_resolution["status"]
                not in {
                    "resolved",
                    "not_applicable",
                }
                and not single_column_default
            ):
                diagnostics.append(
                    {
                        "source_file": source_identity,
                        "status": "invalid",
                        "failure_kind": (
                            "ambiguous_csv_field_delimiter"
                            if delimiter_resolution["status"] == "ambiguous"
                            else "unsupported_csv_field_delimiter"
                        ),
                        "movement_count": 0,
                        "csv_field_delimiter": None,
                        "csv_field_delimiter_profile": delimiter_resolution["profile"],
                    }
                )
                continue
            sample_delimiter = (
                DEFAULT_CSV_FIELD_DELIMITER
                if single_column_default
                else delimiter_resolution["delimiter"]
            )
            raw = _read_table_raw(
                source,
                csv_field_delimiter=(
                    str(sample_delimiter) if sample_delimiter is not None else None
                ),
            )
            header_candidates: list[tuple[int, str]] = []
            sample_aliases = (
                EXACT_HEADER_ALIASES["movement_number"]
                | EXACT_HEADER_ALIASES["reference"]
            )
            for row_index in range(min(raw.height, 30)):
                labels = _merge_header_rows([_row_values(raw, row_index)])
                matches = [
                    label for label in labels if _norm_label(label) in sample_aliases
                ]
                if len(matches) == 1:
                    header_candidates.append((row_index + 1, matches[0]))
            if len(header_candidates) != 1:
                diagnostics.append(
                    {
                        "source_file": source_identity,
                        "status": "invalid",
                        "failure_kind": "ambiguous_sample_identifier_column",
                        "movement_count": 0,
                        "csv_field_delimiter": sample_delimiter,
                        "csv_field_delimiter_profile": delimiter_resolution["profile"],
                    }
                )
                continue
            header_row, identifier_column = header_candidates[0]
            table = _apply_header(
                raw,
                [header_row],
                source_sheet=_source_sheet_name(source),
            )
            file_values = {
                _clean_text(value)
                for value in table.get_column(identifier_column)
                .cast(pl.Utf8, strict=False)
                .to_list()
                if _clean_text(value)
            }
            values.update(file_values)
            diagnostics.append(
                {
                    "source_file": source_identity,
                    "status": "qualified" if file_values else "invalid",
                    "failure_kind": None if file_values else "empty_sample",
                    "movement_count": len(file_values),
                    "header_row": header_row,
                    "identifier_column": identifier_column,
                    "csv_field_delimiter": sample_delimiter,
                    "csv_field_delimiter_profile": delimiter_resolution["profile"],
                }
            )
        except (
            fastexcel.CalamineError,
            BadZipFile,
            InvalidFileException,
            pl.exceptions.PolarsError,
            UnicodeError,
            ValueError,
            RuntimeError,
            OSError,
        ) as exc:
            diagnostics.append(
                {
                    "source_file": source_identity,
                    "status": "invalid",
                    "failure_kind": "parser_failure",
                    "movement_count": 0,
                    "parser_error": f"{type(exc).__name__}: {exc}",
                    "csv_field_delimiter": (
                        delimiter_resolution["delimiter"]
                        if delimiter_resolution is not None
                        else None
                    ),
                    "csv_field_delimiter_profile": (
                        delimiter_resolution["profile"]
                        if delimiter_resolution is not None
                        else None
                    ),
                }
            )
    if diagnostics and any(item["status"] != "qualified" for item in diagnostics):
        return set(), diagnostics
    return values, diagnostics


def inspect_inputs(
    bank_path: Path,
    journal_path: Path,
    output_dir: Path,
    recipe_path: Path | None = None,
    *,
    sample_path: Path | None = None,
    language: object | None = None,
    document_language: object | None = None,
) -> InspectionResult:
    """Inspect input files and write deterministic recipe artifacts."""

    recipe = read_json(recipe_path)
    languages = language_assumptions(
        recipe, language=language, document_language=document_language
    )
    initial_source_receipts, source_refs = _source_artifact_receipts(
        (("bank", bank_path), ("journal", journal_path), ("sample", sample_path))
    )
    bank_frame, bank_diag = _normalize_files(
        bank_path,
        "bank",
        recipe,
        source_refs=source_refs,
    )
    journal_frame, journal_diag = _normalize_files(
        journal_path,
        "journal",
        recipe,
        source_refs=source_refs,
    )
    sample_movements, sample_diagnostics = _read_sample_movements(sample_path)
    source_receipts, current_source_refs = _source_artifact_receipts(
        (("bank", bank_path), ("journal", journal_path), ("sample", sample_path))
    )
    changed_sources = _source_snapshot_changes(
        initial_source_receipts,
        source_receipts,
    )
    source_refs = current_source_refs
    if changed_sources:
        bank_diag = _mark_changed_source_diagnostics("bank", bank_diag, changed_sources)
        journal_diag = _mark_changed_source_diagnostics(
            "journal", journal_diag, changed_sources
        )
        sample_diagnostics = _mark_changed_sample_diagnostics(
            sample_diagnostics,
            changed_sources,
        )
        if any(root_id == "source_bank" for root_id, _ in changed_sources):
            bank_frame = _transaction_frame([])
        if any(root_id == "source_journal" for root_id, _ in changed_sources):
            journal_frame = _transaction_frame([])
        if any(root_id == "source_sample" for root_id, _ in changed_sources):
            sample_movements = set()
    qualifications = [
        *_source_qualifications("bank", bank_diag, source_refs),
        *_source_qualifications("journal", journal_diag, source_refs),
    ]
    qualification_status = _aggregate_qualification_status(qualifications)
    if changed_sources:
        qualification_status = "unsupported_source_layout"
    sample_status = (
        "not_supplied"
        if sample_path is None
        else "qualified" if sample_movements else "invalid_or_empty"
    )
    proposed_relationship_policy = {
        "relationship_shape": "many_to_many",
        "allow_evidence_reuse": False,
        "require_same_currency": True,
        "require_same_unit": True,
        "require_same_entity": True,
        "require_same_party": False,
        "direction_policy": "absolute_amount",
        "default_currency": "EUR",
        "default_unit": "currency",
        "default_entity_ref": "entity.case",
        "default_party_ref": None,
        "amount_tolerance": "1",
        "date_window_days": 7,
    }
    normalized_relationship_policy = _normalize_relationship_policy(
        proposed_relationship_policy
    )
    suggested_recipe = {
        "version": 2,
        "description": "Bounded journal-bank reconciliation recipe for professional review.",
        **languages,
        "bank": _recipe_side(bank_diag),
        "journal": _recipe_side(journal_diag),
        "matching": {
            "amount_tolerance": "1",
            "date_window_days": 7,
            "use_absolute_amounts": True,
            "stages": list(MATCH_STAGE_ORDER),
        },
        "relationship": {
            "policy": normalized_relationship_policy,
            "review_content_sha256": canonical_json_sha256(
                {"policy": normalized_relationship_policy}
            ),
            "decision": None,
        },
    }
    output_dir.mkdir(parents=True, exist_ok=True)
    write_assurance_json(
        output_dir / "input_receipts.json",
        {
            "schema_version": "journal_bank.input_receipts.v1",
            "receipts": source_receipts,
        },
    )
    write_assurance_json(
        output_dir / "source_qualifications.json",
        {
            "schema_version": "journal_bank.source_qualifications.v1",
            "status": qualification_status,
            "source_outcomes": _source_outcomes(bank_diag, journal_diag),
            "qualifications": qualifications,
        },
    )
    write_assurance_json(
        output_dir / "reviewed_decisions.json",
        {
            "schema_version": "journal_bank.reviewed_decisions.v1",
            "decisions": _mapping_decisions([*bank_diag, *journal_diag]),
        },
    )
    write_assurance_json(
        output_dir / "lineage.json",
        _lineage_payload(bank_frame, journal_frame, source_refs),
    )
    write_json(
        output_dir / "inspection.json",
        {
            **languages,
            "qualification_status": qualification_status,
            "source_snapshot_changed": bool(changed_sources),
            "changed_sources": [
                {"root_id": root_id, "path": source_path}
                for root_id, source_path in sorted(changed_sources)
            ],
            "bank": {"row_count": bank_frame.height, "files": bank_diag},
            "bank_pdf_non_movement_row_count": len(_non_movement_records(bank_diag)),
            "journal": {"row_count": journal_frame.height, "files": journal_diag},
            "sample": {
                "path": sample_path.as_posix() if sample_path else None,
                "status": sample_status,
                "movement_count": len(sample_movements),
                "movements": sorted(sample_movements)[:100],
                "diagnostics": sample_diagnostics,
            },
        },
    )
    write_json(output_dir / "suggested_recipe.json", suggested_recipe)
    return InspectionResult(
        bank={"row_count": bank_frame.height, "files": bank_diag},
        journal={"row_count": journal_frame.height, "files": journal_diag},
        sample={"movement_count": len(sample_movements), "status": sample_status},
        suggested_recipe=suggested_recipe,
    )


def _date_diff_days(left: Any, right: Any) -> int | None:
    left_date = _parse_date(left)
    right_date = _parse_date(right)
    if left_date is None or right_date is None:
        return None
    return abs((left_date - right_date).days)


def _canonical_tolerance(value: object) -> tuple[Decimal, str]:
    """Return an exact non-negative tolerance and its canonical text.

    A reviewed relationship tolerance is authoritative policy, not localized
    source data. Accepting only canonical text, ``Decimal``, or integers avoids
    hidden locale inference and binary-float rounding at that boundary.
    """

    if isinstance(value, bool):
        raise MoneyValidationError(
            "tolerance must be canonical Decimal text, int, or Decimal"
        )
    if isinstance(value, str):
        tolerance = parse_canonical_decimal(value, label="tolerance")
    elif isinstance(value, Decimal):
        tolerance = value
    elif isinstance(value, int):
        tolerance = Decimal(value)
    else:
        raise MoneyValidationError(
            "tolerance must be canonical Decimal text, int, or Decimal"
        )
    if not tolerance.is_finite():
        raise MoneyValidationError("tolerance must be finite")
    if tolerance < ZERO:
        raise MoneyValidationError("tolerance must not be negative")
    return tolerance, decimal_text(tolerance)


def _validated_relationship_decision(
    recipe: dict[str, Any],
    *,
    source_artifact_refs: Sequence[str],
    tolerance_text: str,
    date_window_days: int,
) -> tuple[dict[str, Any] | None, dict[str, Any] | None, str | None]:
    relationship = recipe.get("relationship")
    if not isinstance(relationship, dict):
        return None, None, "A reviewed relationship and perimeter policy is required."
    try:
        policy = _normalize_relationship_policy(relationship.get("policy"))
    except ValueError as exc:
        return None, None, str(exc)
    if policy["amount_tolerance"] != tolerance_text:
        return None, None, "Execution tolerance differs from the reviewed policy."
    if policy["date_window_days"] != date_window_days:
        return None, None, "Execution date window differs from the reviewed policy."
    raw_decision = relationship.get("decision")
    if not isinstance(raw_decision, dict):
        return None, None, "A reviewed relationship decision receipt is required."
    try:
        decision = validate_reviewed_decision_receipt(
            raw_decision,
            expected_source_artifact_refs=list(source_artifact_refs),
            expected_adapter_id=RELATIONSHIP_ADAPTER_ID,
            expected_adapter_version=RELATIONSHIP_ADAPTER_VERSION,
            require_reviewed=True,
        )
    except ValueError as exc:
        return None, None, f"Relationship decision is invalid or stale: {exc}"
    if decision["decision_type"] != "journal_bank_relationship":
        return None, None, "Relationship decision has the wrong decision type."
    if decision["content"] != {"policy": policy}:
        return None, None, "Relationship decision content does not match the policy."
    return policy, decision, None


def _with_policy_defaults(frame: pl.DataFrame, policy: dict[str, Any]) -> pl.DataFrame:
    result = frame
    defaults = {
        "currency": policy["default_currency"],
        "unit": policy["default_unit"],
        "entity_ref": policy["default_entity_ref"],
        "party_ref": policy["default_party_ref"],
    }
    for field, default in defaults.items():
        if default is None:
            continue
        result = result.with_columns(
            pl.when(
                pl.col(field).is_null()
                | (pl.col(field).cast(pl.Utf8, strict=False).str.strip_chars() == "")
            )
            .then(pl.lit(default))
            .otherwise(pl.col(field))
            .alias(field)
        )
    return result


def _relationship_perimeter_error(
    bank: pl.DataFrame,
    journal: pl.DataFrame,
    policy: dict[str, Any],
) -> str | None:
    required = ["currency", "unit"]
    if policy["require_same_entity"]:
        required.append("entity_ref")
    if policy["require_same_party"]:
        required.append("party_ref")
    for side, frame in (("bank", bank), ("journal", journal)):
        for field in required:
            missing = frame.filter(
                pl.col(field).is_null()
                | (pl.col(field).cast(pl.Utf8, strict=False).str.strip_chars() == "")
            )
            if not missing.is_empty():
                return (
                    f"{side} rows lack required {field}; provide a source field or "
                    "a reviewed policy default."
                )
    return None


def _direction_compatible(
    bank_row: dict[str, Any],
    journal_row: dict[str, Any],
    policy: dict[str, Any],
) -> bool:
    direction_policy = str(policy["direction_policy"])
    if direction_policy == "absolute_amount":
        return True
    bank_direction = _clean_text(bank_row.get("direction"))
    journal_direction = _clean_text(journal_row.get("direction"))
    if bank_direction not in {"positive", "negative", "zero"}:
        return False
    if journal_direction not in {"positive", "negative", "zero"}:
        return False
    if direction_policy == "same_sign":
        return bank_direction == journal_direction
    return {bank_direction, journal_direction} == {"positive", "negative"}


def _same_required_perimeter(
    bank_row: dict[str, Any],
    journal_row: dict[str, Any],
    policy: dict[str, Any],
) -> bool:
    comparisons = [
        ("currency", policy["require_same_currency"]),
        ("unit", policy["require_same_unit"]),
        ("entity_ref", policy["require_same_entity"]),
        ("party_ref", policy["require_same_party"]),
    ]
    return all(
        not required
        or (
            _clean_text(bank_row.get(field))
            and _clean_text(bank_row.get(field)) == _clean_text(journal_row.get(field))
        )
        for field, required in comparisons
    ) and _direction_compatible(bank_row, journal_row, policy)


@dataclass(frozen=True)
class _JournalAmountIndex:
    """Exact Decimal range index for mechanically eligible journal amounts.

    The index only removes rows outside the authoritative reviewed-tolerance
    predicate. Existing perimeter, date, reference, and non-reuse checks remain
    authoritative for every row returned by the index.
    """

    amounts: tuple[Decimal, ...]
    positioned_rows: tuple[tuple[int, dict[str, Any]], ...]

    @classmethod
    def from_rows(cls, journal_rows: Sequence[dict[str, Any]]) -> _JournalAmountIndex:
        indexed_rows: list[tuple[Decimal, int, dict[str, Any]]] = []
        for position, journal_row in enumerate(journal_rows):
            journal_value = journal_row.get("amount_abs")
            if not isinstance(journal_value, str):
                continue
            indexed_rows.append(
                (
                    parse_canonical_decimal(journal_value, label="journal amount"),
                    position,
                    journal_row,
                )
            )
        indexed_rows.sort(key=lambda item: (item[0], item[1]))
        return cls(
            amounts=tuple(item[0] for item in indexed_rows),
            positioned_rows=tuple((item[1], item[2]) for item in indexed_rows),
        )

    def rows_within_tolerance(
        self, amount: Decimal, tolerance: Decimal
    ) -> list[dict[str, Any]]:
        """Return predicate-approved candidates in original journal order.

        Searching outward from the insertion point avoids calculating Decimal
        bounds that the ambient context could round inward. The authoritative
        tolerance predicate decides both inclusion and where each scan stops.
        """

        pivot = bisect_left(self.amounts, amount)
        positioned_candidates: list[tuple[int, dict[str, Any]]] = []

        left = pivot - 1
        while left >= 0:
            _, within_tolerance = difference_within_tolerance(
                amount, self.amounts[left], tolerance
            )
            if not within_tolerance:
                break
            positioned_candidates.append(self.positioned_rows[left])
            left -= 1

        right = pivot
        while right < len(self.amounts):
            _, within_tolerance = difference_within_tolerance(
                amount, self.amounts[right], tolerance
            )
            if not within_tolerance:
                break
            positioned_candidates.append(self.positioned_rows[right])
            right += 1

        positioned_candidates.sort(key=lambda item: item[0])
        return [row for _, row in positioned_candidates]


def _candidate_rows(
    bank_row: dict[str, Any],
    journal_index: _JournalAmountIndex,
    used_journal: set[str],
    *,
    tolerance: Decimal,
    date_window_days: int,
    relationship_policy: dict[str, Any],
) -> list[dict[str, Any]]:
    candidates = []
    bank_value = bank_row.get("amount_abs")
    bank_amount = (
        parse_canonical_decimal(bank_value, label="bank amount")
        if isinstance(bank_value, str)
        else None
    )
    if bank_amount is None:
        return candidates
    for journal_row in journal_index.rows_within_tolerance(bank_amount, tolerance):
        journal_id = str(journal_row["transaction_id"])
        if journal_id in used_journal:
            continue
        journal_value = journal_row.get("amount_abs")
        journal_amount = (
            parse_canonical_decimal(journal_value, label="journal amount")
            if isinstance(journal_value, str)
            else None
        )
        if journal_amount is None:
            continue
        if not _same_required_perimeter(bank_row, journal_row, relationship_policy):
            continue
        signed_delta, within_tolerance = difference_within_tolerance(
            bank_amount, journal_amount, tolerance
        )
        if not within_tolerance:
            continue
        date_diff = _date_diff_days(
            bank_row.get("transaction_date"), journal_row.get("transaction_date")
        )
        if date_diff is not None and date_diff > date_window_days:
            continue
        candidates.append(
            {
                "row": journal_row,
                "amount_delta": abs(signed_delta),
                "date_diff_days": date_diff,
                "shared_references": sorted(
                    _reference_tokens(
                        bank_row.get("reference"),
                        bank_row.get("movement_number"),
                    )
                    & _reference_tokens(
                        journal_row.get("reference"),
                        journal_row.get("movement_number"),
                    )
                ),
            }
        )
    return candidates


def _make_match_record(
    bank_row: dict[str, Any],
    journal_row: dict[str, Any],
    *,
    stage: str,
    amount_delta: Decimal,
    date_diff_days: int | None,
    shared_references: Sequence[str],
) -> dict[str, Any]:
    return {
        "status": "matched",
        "stage": stage,
        "bank_transaction_id": bank_row["transaction_id"],
        "journal_transaction_id": journal_row["transaction_id"],
        "bank_date": bank_row.get("transaction_date"),
        "journal_date": journal_row.get("transaction_date"),
        "date_diff_days": date_diff_days,
        "bank_amount": bank_row.get("amount_signed"),
        "journal_amount": journal_row.get("amount_signed"),
        "amount_delta": decimal_text(amount_delta),
        "bank_description": bank_row.get("description"),
        "journal_description": journal_row.get("description"),
        "shared_references": ",".join(shared_references),
        "review_note": "",
    }


def _unconflicted_singleton_batch(
    bank_rows: Sequence[dict[str, Any]],
    journal_index: _JournalAmountIndex,
    used_bank: set[str],
    used_journal: set[str],
    *,
    tolerance: Decimal,
    date_window_days: int,
    relationship_policy: dict[str, Any],
    require_shared_reference: bool,
) -> list[tuple[dict[str, Any], dict[str, Any]]]:
    """Return one order-independent batch of conflict-free singleton matches.

    Matching is a mechanical one-to-one allocation. Evaluating the complete
    batch before accepting any row prevents source order from granting a
    journal candidate to the first competing bank row.
    """

    singleton_candidates: list[tuple[dict[str, Any], dict[str, Any]]] = []
    for bank_row in bank_rows:
        if str(bank_row["transaction_id"]) in used_bank:
            continue
        candidates = _candidate_rows(
            bank_row,
            journal_index,
            used_journal,
            tolerance=tolerance,
            date_window_days=date_window_days,
            relationship_policy=relationship_policy,
        )
        if require_shared_reference:
            candidates = [
                candidate for candidate in candidates if candidate["shared_references"]
            ]
        else:
            candidates = [
                candidate
                for candidate in candidates
                if candidate["date_diff_days"] is not None
            ]
        if len(candidates) == 1:
            singleton_candidates.append((bank_row, candidates[0]))

    target_counts = Counter(
        str(candidate["row"]["transaction_id"]) for _, candidate in singleton_candidates
    )
    return sorted(
        (
            (bank_row, candidate)
            for bank_row, candidate in singleton_candidates
            if target_counts[str(candidate["row"]["transaction_id"])] == 1
        ),
        key=lambda item: (
            str(item[0]["transaction_id"]),
            str(item[1]["row"]["transaction_id"]),
        ),
    )


def _unconflicted_reference_group_batch(
    bank_rows: Sequence[dict[str, Any]],
    journal_rows: Sequence[dict[str, Any]],
    used_bank: set[str],
    used_journal: set[str],
    *,
    tolerance: Decimal,
    date_window_days: int,
    relationship_policy: dict[str, Any],
) -> list[dict[str, Any]]:
    """Return exact-sum reference groups with order-independent membership.

    Grouping is deterministic because a shared explicit identifier defines the
    population and exact Decimal conservation decides acceptance. Descriptions
    and beneficiary similarity never create a group.
    """

    shape = str(relationship_policy["relationship_shape"])
    if shape == "one_to_one":
        return []
    available_bank = [
        row for row in bank_rows if str(row["transaction_id"]) not in used_bank
    ]
    available_journal = [
        row for row in journal_rows if str(row["transaction_id"]) not in used_journal
    ]
    bank_by_token: dict[str, list[dict[str, Any]]] = {}
    journal_by_token: dict[str, list[dict[str, Any]]] = {}
    for row in available_bank:
        for token in _reference_tokens(
            row.get("reference"), row.get("movement_number")
        ):
            bank_by_token.setdefault(token, []).append(row)
    for row in available_journal:
        for token in _reference_tokens(
            row.get("reference"), row.get("movement_number")
        ):
            journal_by_token.setdefault(token, []).append(row)

    grouped: dict[tuple[tuple[str, ...], tuple[str, ...]], dict[str, Any]] = {}
    for token in sorted(set(bank_by_token) & set(journal_by_token)):
        bank_group = bank_by_token[token]
        journal_group = journal_by_token[token]
        bank_count = len(bank_group)
        journal_count = len(journal_group)
        one_to_many = bank_count == 1 and journal_count > 1
        many_to_one = bank_count > 1 and journal_count == 1
        if not (
            (one_to_many and shape in {"one_to_many", "many_to_many"})
            or (many_to_one and shape in {"many_to_one", "many_to_many"})
        ):
            continue
        if not all(
            _same_required_perimeter(bank_row, journal_row, relationship_policy)
            for bank_row in bank_group
            for journal_row in journal_group
        ):
            continue
        date_differences = [
            _date_diff_days(
                bank_row.get("transaction_date"),
                journal_row.get("transaction_date"),
            )
            for bank_row in bank_group
            for journal_row in journal_group
        ]
        if any(
            difference is not None and difference > date_window_days
            for difference in date_differences
        ):
            continue
        bank_total = sum(
            (parse_canonical_decimal(str(row["amount_abs"])) for row in bank_group),
            ZERO,
        )
        journal_total = sum(
            (parse_canonical_decimal(str(row["amount_abs"])) for row in journal_group),
            ZERO,
        )
        signed_delta, within_tolerance = difference_within_tolerance(
            bank_total,
            journal_total,
            tolerance,
        )
        if not within_tolerance:
            continue
        bank_ids = tuple(sorted(str(row["transaction_id"]) for row in bank_group))
        journal_ids = tuple(sorted(str(row["transaction_id"]) for row in journal_group))
        key = (bank_ids, journal_ids)
        candidate = grouped.setdefault(
            key,
            {
                "bank_rows": bank_group,
                "journal_rows": journal_group,
                "amount_delta": abs(signed_delta),
                "shared_references": [],
            },
        )
        candidate["shared_references"].append(token)

    bank_membership: Counter[str] = Counter()
    journal_membership: Counter[str] = Counter()
    for bank_ids, journal_ids in grouped:
        bank_membership.update(bank_ids)
        journal_membership.update(journal_ids)
    accepted = [
        candidate
        for (bank_ids, journal_ids), candidate in grouped.items()
        if all(bank_membership[item] == 1 for item in bank_ids)
        and all(journal_membership[item] == 1 for item in journal_ids)
    ]
    return sorted(
        accepted,
        key=lambda item: (
            tuple(str(row["transaction_id"]) for row in item["bank_rows"]),
            tuple(str(row["transaction_id"]) for row in item["journal_rows"]),
        ),
    )


def _match_transactions(
    bank: pl.DataFrame,
    journal: pl.DataFrame,
    *,
    tolerance: Decimal,
    date_window_days: int,
    relationship_policy: dict[str, Any],
) -> tuple[pl.DataFrame, pl.DataFrame, pl.DataFrame, dict[str, int]]:
    bank_rows = bank.to_dicts()
    journal_rows = journal.to_dicts()
    journal_index = _JournalAmountIndex.from_rows(journal_rows)
    used_bank: set[str] = set()
    used_journal: set[str] = set()
    matches: list[dict[str, Any]] = []
    stage_counts: dict[str, int] = {}
    (
        reference_stage,
        reference_group_stage,
        first_amount_date_stage,
        later_amount_date_stage,
    ) = MATCH_STAGE_ORDER

    def accept(bank_row: dict[str, Any], candidate: dict[str, Any], stage: str) -> None:
        matches.append(
            _make_match_record(
                bank_row,
                candidate["row"],
                stage=stage,
                amount_delta=candidate["amount_delta"],
                date_diff_days=candidate["date_diff_days"],
                shared_references=candidate["shared_references"],
            )
        )
        used_bank.add(str(bank_row["transaction_id"]))
        used_journal.add(str(candidate["row"]["transaction_id"]))
        stage_counts[stage] = stage_counts.get(stage, 0) + 1

    while True:
        reference_batch = _unconflicted_singleton_batch(
            bank_rows,
            journal_index,
            used_bank,
            used_journal,
            tolerance=tolerance,
            date_window_days=date_window_days,
            relationship_policy=relationship_policy,
            require_shared_reference=True,
        )
        if not reference_batch:
            break
        for bank_row, candidate in reference_batch:
            accept(bank_row, candidate, reference_stage)

    while True:
        reference_groups = _unconflicted_reference_group_batch(
            bank_rows,
            journal_rows,
            used_bank,
            used_journal,
            tolerance=tolerance,
            date_window_days=date_window_days,
            relationship_policy=relationship_policy,
        )
        if not reference_groups:
            break
        for group in reference_groups:
            bank_group = group["bank_rows"]
            journal_group = group["journal_rows"]
            if len(bank_group) == 1:
                pairs = [(bank_group[0], journal_row) for journal_row in journal_group]
            else:
                pairs = [(bank_row, journal_group[0]) for bank_row in bank_group]
            for bank_row, journal_row in pairs:
                matches.append(
                    _make_match_record(
                        bank_row,
                        journal_row,
                        stage=reference_group_stage,
                        amount_delta=group["amount_delta"],
                        date_diff_days=_date_diff_days(
                            bank_row.get("transaction_date"),
                            journal_row.get("transaction_date"),
                        ),
                        shared_references=group["shared_references"],
                    )
                )
            used_bank.update(str(row["transaction_id"]) for row in bank_group)
            used_journal.update(str(row["transaction_id"]) for row in journal_group)
            stage_counts[reference_group_stage] = stage_counts.get(
                reference_group_stage, 0
            ) + len(pairs)

    first_amount_date_batch = _unconflicted_singleton_batch(
        bank_rows,
        journal_index,
        used_bank,
        used_journal,
        tolerance=tolerance,
        date_window_days=date_window_days,
        relationship_policy=relationship_policy,
        require_shared_reference=False,
    )
    for bank_row, candidate in first_amount_date_batch:
        accept(bank_row, candidate, first_amount_date_stage)

    while first_amount_date_batch:
        later_amount_date_batch = _unconflicted_singleton_batch(
            bank_rows,
            journal_index,
            used_bank,
            used_journal,
            tolerance=tolerance,
            date_window_days=date_window_days,
            relationship_policy=relationship_policy,
            require_shared_reference=False,
        )
        if not later_amount_date_batch:
            break
        for bank_row, candidate in later_amount_date_batch:
            accept(bank_row, candidate, later_amount_date_stage)
        first_amount_date_batch = later_amount_date_batch

    unmatched_bank = _transaction_frame(
        [row for row in bank_rows if str(row["transaction_id"]) not in used_bank]
    )
    unmatched_journal = _transaction_frame(
        [row for row in journal_rows if str(row["transaction_id"]) not in used_journal]
    )
    match_frame = (
        pl.DataFrame(matches, infer_schema_length=None)
        if matches
        else pl.DataFrame(schema={col: pl.Utf8 for col in MATCH_COLUMNS})
    )
    for col in MATCH_COLUMNS:
        if col not in match_frame.columns:
            match_frame = match_frame.with_columns(pl.lit(None).alias(col))
    return (
        match_frame.select(MATCH_COLUMNS),
        unmatched_bank,
        unmatched_journal,
        stage_counts,
    )


def _relationship_ledger(
    bank: pl.DataFrame,
    journal: pl.DataFrame,
    matches: pl.DataFrame,
    *,
    policy: dict[str, Any],
) -> dict[str, Any]:
    bank_rows = bank.to_dicts()
    journal_rows = journal.to_dicts()
    bank_ids = {
        str(row["transaction_id"]): f"bank.{index}"
        for index, row in enumerate(bank_rows, start=1)
    }
    journal_ids = {
        str(row["transaction_id"]): f"journal.{index}"
        for index, row in enumerate(journal_rows, start=1)
    }

    def record(row: dict[str, Any], record_id: str) -> dict[str, str | None]:
        return {
            "record_id": record_id,
            "amount": str(row["amount_abs"]),
            "currency": _clean_text(row["currency"]),
            "unit": _clean_text(row["unit"]),
            "entity_ref": _clean_text(row.get("entity_ref")) or None,
            "party_ref": _clean_text(row.get("party_ref")) or None,
        }

    bank_by_id = {str(row["transaction_id"]): row for row in bank_rows}
    journal_by_id = {str(row["transaction_id"]): row for row in journal_rows}
    allocations: list[dict[str, Any]] = []
    for index, match in enumerate(matches.to_dicts(), start=1):
        bank_transaction_id = str(match["bank_transaction_id"])
        journal_transaction_id = str(match["journal_transaction_id"])
        bank_row = bank_by_id[bank_transaction_id]
        journal_row = journal_by_id[journal_transaction_id]
        amount = min(
            parse_canonical_decimal(str(bank_row["amount_abs"])),
            parse_canonical_decimal(str(journal_row["amount_abs"])),
        )
        allocations.append(
            {
                "allocation_id": f"allocation.{index}",
                "source_record_ref": bank_ids[bank_transaction_id],
                "target_record_ref": journal_ids[journal_transaction_id],
                "amount": decimal_text(amount),
                "currency": str(bank_row["currency"]),
                "unit": str(bank_row["unit"]),
                "evidence_refs": [f"match.{index}"],
            }
        )
    return build_allocation_ledger(
        ledger_id="journal_bank.relationship",
        policy={
            "relationship_shape": policy["relationship_shape"],
            "require_same_currency": policy["require_same_currency"],
            "require_same_unit": policy["require_same_unit"],
            "require_same_entity": policy["require_same_entity"],
            "require_same_party": policy["require_same_party"],
            "allow_evidence_reuse": policy["allow_evidence_reuse"],
            "tolerance": policy["amount_tolerance"],
        },
        source_records=[
            record(row, bank_ids[str(row["transaction_id"])]) for row in bank_rows
        ],
        target_records=[
            record(row, journal_ids[str(row["transaction_id"])]) for row in journal_rows
        ],
        allocations=allocations,
    )


def _blocked_relationship_ledger(
    *,
    block_code: str,
    block_detail: str,
) -> dict[str, Any]:
    """Return explicit zero-allocation evidence when reconciliation never ran."""

    content = {
        "schema_version": "journal_bank.blocked_relationship_ledger.v1",
        "ledger_id": "journal_bank.relationship",
        "status": "not_assessed",
        "block_code": block_code,
        "block_detail": block_detail,
        "source_records": [],
        "target_records": [],
        "allocations": [],
        "source_residuals": [],
        "target_residuals": [],
        "balanced": False,
    }
    return {**content, "content_sha256": canonical_json_sha256(content)}


def _relationship_residual_frame(
    ledger: dict[str, Any],
    bank: pl.DataFrame,
    journal: pl.DataFrame,
) -> pl.DataFrame:
    """Project allocation residuals without classifying or forcing them to zero."""

    validated = validate_allocation_ledger(ledger)
    records: list[dict[str, Any]] = []
    for side, population, ledger_records, residuals in (
        (
            "bank",
            bank,
            validated["source_records"],
            validated["source_residuals"],
        ),
        (
            "journal",
            journal,
            validated["target_records"],
            validated["target_residuals"],
        ),
    ):
        population_rows = population.to_dicts()
        if not (len(population_rows) == len(ledger_records) == len(residuals)):
            raise ValueError("Relationship residual population cardinality is stale.")
        for index, (population_row, record, residual_row) in enumerate(
            zip(population_rows, ledger_records, residuals, strict=True),
            start=1,
        ):
            expected_record_ref = f"{side}.{index}"
            record_ref = str(record["record_id"])
            if (
                record_ref != expected_record_ref
                or residual_row["record_ref"] != record_ref
                or str(record["amount"]) != str(population_row["amount_abs"])
            ):
                raise ValueError("Relationship residual row identity is stale.")
            record_amount = parse_canonical_decimal(
                str(record["amount"]),
                label=f"{record_ref}.record_amount",
            )
            residual = parse_canonical_decimal(
                str(residual_row["residual"]),
                label=f"{record_ref}.residual",
            )
            allocated_amount = record_amount - residual
            records.append(
                {
                    "side": side,
                    "record_ref": record_ref,
                    "transaction_id": str(population_row["transaction_id"]),
                    "record_amount": decimal_text(record_amount),
                    "allocated_amount": decimal_text(allocated_amount),
                    "residual": decimal_text(residual),
                    "currency": str(record["currency"]),
                    "unit": str(record["unit"]),
                    "entity_ref": record["entity_ref"],
                    "party_ref": record["party_ref"],
                }
            )
    if not records:
        return pl.DataFrame(schema={field: pl.Utf8 for field in RESIDUAL_COLUMNS})
    return pl.DataFrame(records, infer_schema_length=None).select(RESIDUAL_COLUMNS)


def _material_value(value: object, field: str) -> dict[str, str]:
    """Canonicalize one mechanically verifiable match or residual value."""

    if value is None or value == "":
        return {"kind": "empty", "value": ""}
    text = str(value)
    if field in MATERIAL_DECIMAL_FIELDS:
        parsed = parse_canonical_decimal(text, label=field)
        if decimal_text(parsed) != text:
            raise ValueError(f"{field} is not canonical Decimal text.")
        return {"kind": "decimal", "value": text}
    if field in MATERIAL_INTEGER_FIELDS:
        try:
            parsed_integer = int(text)
        except ValueError as exc:
            raise ValueError(f"{field} is not canonical integer text.") from exc
        if str(parsed_integer) != text:
            raise ValueError(f"{field} is not canonical integer text.")
        return {"kind": "integer", "value": text}
    return {"kind": "text", "value": text}


def _excel_cell(columns: Sequence[str], field: str, row_number: int) -> str:
    column_number = list(columns).index(field) + 1
    letters = ""
    while column_number:
        column_number, remainder = divmod(column_number - 1, 26)
        letters = chr(65 + remainder) + letters
    return f"{letters}{row_number}"


def _canonical_material_rows(
    frame: pl.DataFrame,
    fields: Sequence[str],
) -> list[dict[str, str]]:
    return [
        {field: _material_value(row.get(field), field)["value"] for field in fields}
        for row in frame.iter_rows(named=True)
    ]


def _read_material_csv(path: Path, columns: Sequence[str]) -> pl.DataFrame:
    frame = pl.read_csv(path, infer_schema=False)
    if frame.columns != list(columns):
        raise ValueError(f"{path.name} columns do not close to the material contract.")
    return frame


def _verify_material_native_outputs(
    output_dir: Path,
    matches: pl.DataFrame,
    residuals: pl.DataFrame,
) -> None:
    """Verify every declared material value in its CSV and XLSX addresses."""

    datasets = (
        (
            "matches",
            output_dir / "reconciliation_matches.csv",
            matches,
            MATCH_COLUMNS,
            MATCH_MATERIAL_FIELDS,
        ),
        (
            "relationship_residuals",
            output_dir / "relationship_residuals.csv",
            residuals,
            RESIDUAL_COLUMNS,
            RESIDUAL_MATERIAL_FIELDS,
        ),
    )
    csv_frames: dict[str, pl.DataFrame] = {}
    for worksheet, csv_path, prepared, columns, material_fields in datasets:
        csv_frame = _read_material_csv(csv_path, columns)
        csv_frames[worksheet] = csv_frame
        if _canonical_material_rows(csv_frame, material_fields) != (
            _canonical_material_rows(prepared, material_fields)
        ):
            raise ValueError(
                f"{csv_path.name} material values do not close to prepared {worksheet}."
            )

    workbook_path = output_dir / "journal_bank_reconciliation.xlsx"
    # A duplicate OOXML member makes a declared native address non-unique.
    # Rejecting it is a mechanically verifiable package-integrity check.
    with ZipFile(workbook_path, "r") as package:
        member_names = package.namelist()
        if len(member_names) != len(set(member_names)):
            raise ValueError("OOXML workbook contains duplicate member names.")

    workbook = openpyxl.load_workbook(
        workbook_path,
        read_only=True,
        data_only=False,
    )
    try:
        if workbook.sheetnames != list(WORKBOOK_SHEET_ORDER):
            raise ValueError("Reconciliation XLSX worksheet set does not close.")
        for worksheet, _, prepared, columns, material_fields in datasets:
            sheet = workbook[worksheet]
            header_values = next(
                sheet.iter_rows(
                    min_row=1,
                    max_row=1,
                    max_col=len(columns),
                    values_only=True,
                ),
                (),
            )
            headers = ["" if value is None else str(value) for value in header_values]
            if headers != list(columns):
                raise ValueError(
                    f"Reconciliation XLSX headers do not close for {worksheet}."
                )
            if sheet.max_row != prepared.height + 1 or sheet.max_column != len(columns):
                raise ValueError(
                    f"Reconciliation XLSX dimensions do not close for {worksheet}."
                )
            expected_rows = _canonical_material_rows(prepared, material_fields)
            workbook_rows = sheet.iter_rows(
                min_row=2,
                max_row=prepared.height + 1,
                max_col=len(columns),
                values_only=True,
            )
            field_positions = {
                field: list(columns).index(field) for field in material_fields
            }
            for row_number, (workbook_row, expected_row) in enumerate(
                zip(workbook_rows, expected_rows, strict=True),
                start=2,
            ):
                for field in material_fields:
                    actual = _material_value(
                        workbook_row[field_positions[field]],
                        field,
                    )["value"]
                    if actual != expected_row[field]:
                        raise ValueError(
                            "Reconciliation XLSX material value does not close at "
                            f"{worksheet}!{_excel_cell(columns, field, row_number)}."
                        )
            if (
                _canonical_material_rows(
                    csv_frames[worksheet],
                    material_fields,
                )
                != expected_rows
            ):
                raise ValueError(
                    f"Reconciliation CSV/XLSX material values diverge for {worksheet}."
                )
    finally:
        workbook.close()


def _build_material_value_ledger(
    output_dir: Path,
    matches: pl.DataFrame,
    residuals: pl.DataFrame,
) -> dict[str, Any]:
    """Build exact all-row prepared-to-CSV/XLSX material-value addresses."""

    _verify_material_native_outputs(output_dir, matches, residuals)
    entries: list[dict[str, Any]] = []
    dataset_specs = (
        {
            "dataset_id": "matches",
            "worksheet": "matches",
            "csv_path": "reconciliation_matches.csv",
            "prepared_artifact_ref": "prepared.reconciliation_matches",
            "csv_artifact_ref": "output.reconciliation_matches_csv",
            "frame": matches,
            "columns": MATCH_COLUMNS,
            "material_fields": MATCH_MATERIAL_FIELDS,
            "identity_fields": (
                "bank_transaction_id",
                "journal_transaction_id",
            ),
        },
        {
            "dataset_id": "relationship_residuals",
            "worksheet": "relationship_residuals",
            "csv_path": "relationship_residuals.csv",
            "prepared_artifact_ref": "prepared.relationship_residuals",
            "csv_artifact_ref": "output.relationship_residuals_csv",
            "frame": residuals,
            "columns": RESIDUAL_COLUMNS,
            "material_fields": RESIDUAL_MATERIAL_FIELDS,
            "identity_fields": ("side", "record_ref", "transaction_id"),
        },
    )
    datasets: list[dict[str, Any]] = []
    for spec in dataset_specs:
        frame = spec["frame"]
        material_fields = spec["material_fields"]
        datasets.append(
            {
                "dataset_id": spec["dataset_id"],
                "worksheet": spec["worksheet"],
                "csv_path": spec["csv_path"],
                "columns": list(spec["columns"]),
                "material_fields": list(material_fields),
                "identity_fields": list(spec["identity_fields"]),
                "row_count": frame.height,
                "entry_count": frame.height * len(material_fields),
            }
        )
        for prepared_row_number, row in enumerate(
            frame.iter_rows(named=True),
            start=1,
        ):
            output_row_number = prepared_row_number + 1
            identity = {
                field: _material_value(row.get(field), field)["value"]
                for field in spec["identity_fields"]
            }
            for field in material_fields:
                canonical = _material_value(row.get(field), field)
                entries.append(
                    {
                        "evidence_id": (
                            f"{spec['dataset_id']}.{prepared_row_number}.{field}"
                        ),
                        "dataset_id": spec["dataset_id"],
                        "prepared_row_number": prepared_row_number,
                        "field": field,
                        "value_kind": canonical["kind"],
                        "canonical_value": canonical["value"],
                        "row_identity": identity,
                        "prepared_locator": (
                            f"row={prepared_row_number};column={field}"
                        ),
                        "prepared": {
                            "artifact_ref": spec["prepared_artifact_ref"],
                            "locator": (f"row={prepared_row_number};column={field}"),
                            "value": canonical["value"],
                        },
                        "outputs": [
                            {
                                "artifact_ref": spec["csv_artifact_ref"],
                                "locator": (f"row={output_row_number};column={field}"),
                                "value": canonical["value"],
                            },
                            {
                                "artifact_ref": "output.workbook_xlsx",
                                "locator": (
                                    f"{spec['worksheet']}!"
                                    f"{_excel_cell(spec['columns'], field, output_row_number)}"
                                ),
                                "value": canonical["value"],
                            },
                        ],
                    }
                )
    content = {
        "schema_version": MATERIAL_VALUE_LEDGER_SCHEMA_VERSION,
        "ledger_id": "journal_bank.material_values",
        "datasets": datasets,
        "entry_count": len(entries),
        "entries": entries,
    }
    return {**content, "content_sha256": canonical_json_sha256(content)}


def _current_relationship_policy(output_dir: Path) -> dict[str, Any]:
    reviewed = read_json(output_dir / "reviewed_decisions.json")
    decisions = reviewed.get("decisions")
    if not isinstance(decisions, list):
        raise ValueError("Reviewed decisions do not contain a relationship policy.")
    relationship_decisions = [
        decision
        for decision in decisions
        if isinstance(decision, dict)
        and decision.get("decision_type") == "journal_bank_relationship"
    ]
    if len(relationship_decisions) != 1:
        raise ValueError("Exactly one reviewed relationship policy is required.")
    content = relationship_decisions[0].get("content")
    if not isinstance(content, dict):
        raise ValueError("Reviewed relationship policy content is malformed.")
    return _normalize_relationship_policy(content.get("policy"))


def validate_material_value_ledger(output_dir: Path) -> dict[str, Any]:
    """Freshly replay match/residual preparation and every native value address."""

    unresolved = output_dir.expanduser()
    if unresolved.is_symlink():
        raise ValueError("Reconciliation output directory cannot be a symlink.")
    output_dir = unresolved.resolve()
    payload = read_json(output_dir / "material_value_ledger.json")
    recorded_digest = payload.pop("content_sha256", None)
    if not isinstance(recorded_digest, str) or recorded_digest != canonical_json_sha256(
        payload
    ):
        raise ValueError("Material-value ledger content hash is stale.")
    if payload.get("schema_version") != MATERIAL_VALUE_LEDGER_SCHEMA_VERSION:
        raise ValueError("Unsupported material-value ledger schema.")

    bank = _read_material_csv(
        output_dir / "normalized_bank.csv",
        TRANSACTION_COLUMNS,
    )
    journal = _read_material_csv(
        output_dir / "normalized_journal.csv",
        TRANSACTION_COLUMNS,
    )
    policy = _current_relationship_policy(output_dir)
    audit = read_json(output_dir / "reconciliation_audit.json")
    tolerance, tolerance_text = _canonical_tolerance(audit.get("tolerance"))
    date_window_days = audit.get("date_window_days")
    if (
        not isinstance(date_window_days, int)
        or isinstance(date_window_days, bool)
        or date_window_days < 0
        or policy["amount_tolerance"] != tolerance_text
        or policy["date_window_days"] != date_window_days
    ):
        raise ValueError("Material replay relationship parameters are stale.")
    matches, _, _, _ = _match_transactions(
        bank,
        journal,
        tolerance=tolerance,
        date_window_days=date_window_days,
        relationship_policy=policy,
    )
    current_matches = _read_material_csv(
        output_dir / "reconciliation_matches.csv",
        MATCH_COLUMNS,
    )
    if _canonical_material_rows(
        current_matches,
        MATCH_MATERIAL_FIELDS,
    ) != _canonical_material_rows(matches, MATCH_MATERIAL_FIELDS):
        raise ValueError("Reconciliation match material values do not replay exactly.")
    ledger = validate_allocation_ledger(
        read_json(output_dir / "relationship_ledger.json")
    )
    replayed_ledger = _relationship_ledger(bank, journal, matches, policy=policy)
    if ledger != replayed_ledger:
        raise ValueError("Relationship ledger does not replay exactly.")
    residuals = _relationship_residual_frame(ledger, bank, journal)
    expected = _build_material_value_ledger(output_dir, matches, residuals)
    actual = {**payload, "content_sha256": recorded_digest}
    if actual != expected:
        raise ValueError("Material-value ledger does not replay exactly.")
    return actual


def _ledger_has_zero_residuals(ledger: dict[str, Any]) -> bool:
    return all(
        parse_canonical_decimal(str(item["residual"]), label="relationship residual")
        == ZERO
        for item in (
            *ledger["source_residuals"],
            *ledger["target_residuals"],
        )
    )


def _filter_journal_by_sample(
    journal: pl.DataFrame, movements: set[str]
) -> pl.DataFrame:
    if journal.is_empty():
        return journal
    if not movements:
        return journal.head(0)
    if "movement_number" not in journal.columns:
        return journal.head(0)
    return journal.filter(
        pl.col("movement_number").cast(pl.Utf8, strict=False).is_in(sorted(movements))
    )


def _stable_ooxml_member(name: str, payload: bytes) -> bytes:
    """Normalize package timestamps that do not describe workbook facts."""

    if name == "docProps/core.xml":
        return OOXML_CORE_TIMESTAMP_RE.sub(
            lambda match: match.group(1) + OOXML_TIMESTAMP + match.group(3),
            payload,
        )
    return payload


def _stabilize_ooxml_package(path: Path) -> None:
    """Canonicalize mechanical XLSX metadata and reject duplicate members."""

    original_mode = path.stat().st_mode & 0o7777
    with tempfile.NamedTemporaryFile(
        prefix=f".{path.name}.",
        suffix=".tmp",
        dir=path.parent,
        delete=False,
    ) as temporary:
        temporary_path = Path(temporary.name)
    try:
        with ZipFile(path, "r") as source:
            members = source.infolist()
            names = [member.filename for member in members]
            if len(names) != len(set(names)):
                raise ValueError("OOXML workbook contains duplicate member names.")
            with ZipFile(
                temporary_path,
                "w",
                compression=ZIP_DEFLATED,
                compresslevel=9,
            ) as target:
                for member in sorted(members, key=lambda value: value.filename):
                    info = ZipInfo(member.filename, date_time=OOXML_ZIP_TIMESTAMP)
                    info.compress_type = ZIP_DEFLATED
                    info.create_system = 0
                    info.external_attr = 0
                    info.flag_bits = 0
                    target.writestr(
                        info,
                        _stable_ooxml_member(
                            member.filename,
                            source.read(member),
                        ),
                    )
        temporary_path.chmod(original_mode)
        temporary_path.replace(path)
    finally:
        temporary_path.unlink(missing_ok=True)


def _write_workbook(path: Path, sheets: dict[str, pl.DataFrame]) -> None:
    workbook = openpyxl.Workbook()
    default = workbook.active
    workbook.remove(default)
    for title, frame in sheets.items():
        sheet = workbook.create_sheet(title[:31])
        sheet.append([excel_safe_value(value) for value in frame.columns])
        for row in frame.iter_rows():
            sheet.append([excel_safe_value(value) for value in row])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    _stabilize_ooxml_package(path)


def _write_review_notes(path: Path, audit: dict[str, Any]) -> None:
    if audit.get("language") == "es":
        lines = [
            "# Notas de revisión de la conciliación entre diario y banco",
            "",
            f"- Idioma: {audit['language']}",
            f"- Movimientos bancarios: {audit['bank_row_count']}",
            f"- Asientos del diario: {audit['journal_row_count']}",
            f"- Filas conciliadas: {audit['matched_count']}",
            f"- Movimientos bancarios sin conciliar: {audit['unmatched_bank_count']}",
            f"- Asientos del diario sin conciliar: {audit['unmatched_journal_count']}",
            "",
            "## Recuento por etapa",
        ]
        counts = audit.get("stage_counts", {})
        if counts:
            for stage, count in sorted(counts.items()):
                lines.append(f"- {stage}: {count}")
        else:
            lines.append("- ninguno")
        lines.extend(
            [
                "",
                "## Política de revisión",
                "Los scripts solo concilian evidencias deterministas. Claude debe explicar los casos no resueltos, inspeccionar las filas de origen cuando sea necesario y mantener explícito el juicio profesional.",
                "",
            ]
        )
        path.write_text("\n".join(lines), encoding="utf-8")
        return

    lines = [
        "# Journal-Bank Reconciliation Review Notes",
        "",
        f"- Language: {audit['language']}",
        f"- Bank rows: {audit['bank_row_count']}",
        f"- Journal rows: {audit['journal_row_count']}",
        f"- Matched rows: {audit['matched_count']}",
        f"- Unmatched bank rows: {audit['unmatched_bank_count']}",
        f"- Unmatched journal rows: {audit['unmatched_journal_count']}",
        "",
        "## Stage Counts",
    ]
    counts = audit.get("stage_counts", {})
    if counts:
        for stage, count in sorted(counts.items()):
            lines.append(f"- {stage}: {count}")
    else:
        lines.append("- none")
    lines.extend(
        [
            "",
            "## Review Policy",
            "The scripts only reconcile deterministic evidence. Claude must explain unresolved cases, inspect source rows where needed, and keep professional judgment explicit.",
            "",
        ]
    )
    path.write_text("\n".join(lines), encoding="utf-8")


def _write_output_receipts(
    output_dir: Path,
    paths: dict[str, Path],
    *,
    source_receipts: Sequence[dict[str, Any]],
) -> dict[str, Any]:
    covered = {
        "normalized_bank_csv",
        "normalized_journal_csv",
        "reconciliation_matches_csv",
        "relationship_residuals_csv",
        "unmatched_bank_csv",
        "unmatched_journal_csv",
        "bank_pdf_non_movement_rows_csv",
        "workbook_xlsx",
        "audit_json",
        "review_notes_md",
        "source_qualifications_json",
        "reviewed_decisions_json",
        "lineage_json",
        "relationship_ledger_json",
        "material_value_ledger_json",
        "assurance_gates_json",
        "review_payload_json",
        "ui_decisions_json",
        "final_artifacts_json",
        "run_intake_json",
        "assurance_envelope_json",
    }
    output_receipts = [
        artifact_receipt(
            output_dir,
            path,
            artifact_id=f"output.{key}",
            root_id="run",
            role="output",
        )
        for key, path in paths.items()
        if key in covered and path.is_file()
    ]
    payload = {
        "schema_version": "journal_bank.artifact_receipts.v1",
        "source_receipts": list(source_receipts),
        "output_receipts": output_receipts,
    }
    write_assurance_json(paths["artifact_receipts_json"], payload)
    return payload


def _final_artifact_closure_is_current(
    output_dir: Path,
    paths: dict[str, Path],
    receipt_bundle: dict[str, Any],
) -> bool:
    payload = read_json(paths["final_artifacts_json"])
    outputs = payload.get("outputs")
    if not isinstance(outputs, list):
        return False
    declared_paths: list[str] = []
    for output in outputs:
        if not isinstance(output, dict):
            return False
        relative = output.get("path")
        size_bytes = output.get("size_bytes")
        if not isinstance(relative, str) or not isinstance(size_bytes, int):
            return False
        relative_path = Path(relative)
        if (
            relative_path.is_absolute()
            or relative_path.as_posix() != relative
            or ".." in relative_path.parts
        ):
            return False
        declared_paths.append(relative)
        path = output_dir / relative
        if not path.is_file() or path.stat().st_size != size_bytes:
            return False
    current_paths = {
        path.relative_to(output_dir).as_posix()
        for path in output_dir.rglob("*")
        if path.is_file() and path.name not in FINAL_ARTIFACT_EXCLUDED_NAMES
    }
    if len(declared_paths) != len(set(declared_paths)):
        return False
    if set(declared_paths) != current_paths:
        return False

    final_receipt = next(
        (
            receipt
            for receipt in receipt_bundle.get("output_receipts", [])
            if isinstance(receipt, dict)
            and receipt.get("path") == paths["final_artifacts_json"].name
        ),
        None,
    )
    if final_receipt is None:
        return False
    try:
        validate_artifact_receipt(output_dir, final_receipt)
    except ValueError:
        return False
    return True


def _close_final_artifacts(
    output_dir: Path,
    paths: dict[str, Path],
    *,
    source_receipts: Sequence[dict[str, Any]],
) -> dict[str, Any]:
    """Reach a bounded size/receipt fixed point after all late output writes."""

    for _ in range(FINAL_ARTIFACT_CLOSURE_MAX_PASSES):
        refresh_final_artifacts(paths["final_artifacts_json"])
        refresh_review_execution_trace(
            paths["run_intake_json"],
            paths["final_artifacts_json"],
        )
        receipt_bundle = _write_output_receipts(
            output_dir,
            paths,
            source_receipts=source_receipts,
        )
        if _final_artifact_closure_is_current(output_dir, paths, receipt_bundle):
            return receipt_bundle
    raise RuntimeError("Final artifact manifest and receipts did not converge.")


def _artifact_roots(
    *,
    bank_path: Path,
    journal_path: Path,
    sample_path: Path | None,
    output_dir: Path,
) -> dict[str, Path]:
    roots = {
        "source_bank": _input_root(bank_path),
        "source_journal": _input_root(journal_path),
        "run": output_dir,
        **implementation_artifact_roots(),
    }
    if sample_path is not None:
        roots["source_sample"] = _input_root(sample_path)
    return roots


def implementation_artifact_roots() -> dict[str, Path]:
    """Return the two immutable code roots bound into every assurance envelope."""

    if _VERA_ASSURANCE_ROOT is None:  # pragma: no cover - import gate above
        raise RuntimeError("Vendored vera_assurance implementation is unavailable.")
    return {
        "implementation": _COMPONENT_ROOT,
        "shared_implementation": _VERA_ASSURANCE_ROOT,
    }


def _ordinary_single_link_implementation_file(
    root: Path,
    relative_path: str,
) -> Path:
    """Resolve one implementation leaf without accepting links or aliases."""

    relative = Path(relative_path)
    if relative.is_absolute() or ".." in relative.parts or "\\" in relative_path:
        raise ValueError("implementation receipt path must be canonical")
    absolute_root = Path(os.path.abspath(root))
    root_stat = absolute_root.lstat()
    if stat.S_ISLNK(root_stat.st_mode) or not stat.S_ISDIR(root_stat.st_mode):
        raise ValueError("implementation root must be an ordinary directory")
    current = absolute_root
    for component in relative.parts[:-1]:
        current /= component
        current_stat = current.lstat()
        if stat.S_ISLNK(current_stat.st_mode) or not stat.S_ISDIR(current_stat.st_mode):
            raise ValueError(
                "implementation receipt path must have ordinary directories"
            )
    candidate = absolute_root / relative
    candidate_stat = candidate.lstat()
    if stat.S_ISLNK(candidate_stat.st_mode) or not stat.S_ISREG(candidate_stat.st_mode):
        raise ValueError("implementation receipt path must be an ordinary file")
    if candidate_stat.st_nlink != 1:
        raise ValueError("implementation receipt path cannot have hardlink aliases")
    return candidate


def _validate_implementation_spec_coverage(roots: dict[str, Path]) -> None:
    """Keep the declared transitive implementation set complete and closed."""

    declared_contract = tuple(
        (
            "plugin" if root_id == "implementation" else "shared_assurance",
            relative_path,
        )
        for _, root_id, relative_path in IMPLEMENTATION_ARTIFACT_SPECS
    )
    if declared_contract != IMPLEMENTATION_CONTRACT:
        raise ValueError("implementation receipt specification is incomplete")
    validate_implementation_tree(
        str(roots["implementation"]),
        shared_assurance_root=str(roots["shared_implementation"]),
    )


def build_implementation_artifact_receipts() -> list[dict[str, Any]]:
    """Snapshot the exact ordered transitive implementation set."""

    roots = implementation_artifact_roots()
    _validate_implementation_spec_coverage(roots)
    receipts: list[dict[str, Any]] = []
    for artifact_id, root_id, relative_path in IMPLEMENTATION_ARTIFACT_SPECS:
        artifact_path = _ordinary_single_link_implementation_file(
            roots[root_id],
            relative_path,
        )
        receipts.append(
            artifact_receipt(
                roots[root_id],
                artifact_path,
                artifact_id=artifact_id,
                root_id=root_id,
                role="implementation",
            )
        )
    return receipts


def validate_exact_implementation_receipts(
    envelope: dict[str, Any],
    *,
    artifact_roots: dict[str, Path] | None = None,
) -> list[dict[str, Any]]:
    """Replay the exact ordered code receipts against ordinary current bytes."""

    expected = build_implementation_artifact_receipts()
    expected_refs = [receipt["artifact_id"] for receipt in expected]
    raw_refs = envelope.get("implementation_artifact_refs")
    raw_artifacts = envelope.get("artifact_receipts")
    if not isinstance(raw_refs, list) or raw_refs != expected_refs:
        raise ValueError("assurance implementation receipt set or order is invalid")
    if not isinstance(raw_artifacts, list):
        raise ValueError("assurance implementation receipts are unavailable")
    implementation_receipts = [
        receipt
        for receipt in raw_artifacts
        if isinstance(receipt, dict) and receipt.get("role") == "implementation"
    ]
    if implementation_receipts != expected:
        raise ValueError("assurance implementation receipts are not exact")
    roots = artifact_roots or implementation_artifact_roots()
    for receipt in implementation_receipts:
        validate_artifact_receipt(roots, receipt)
    return implementation_receipts


def _write_assurance_envelope(
    *,
    run_id: str,
    bank_path: Path,
    journal_path: Path,
    sample_path: Path | None,
    output_dir: Path,
    path: Path,
    receipt_bundle: dict[str, Any],
    reviewed_decisions: Sequence[dict[str, Any]],
    qualifications: Sequence[dict[str, Any]],
    gates: dict[str, Any],
    limitations: Sequence[str],
) -> dict[str, Any]:
    implementation_receipts = build_implementation_artifact_receipts()
    mutable_control_paths = {
        "artifact_receipts.json",
        "run_intake.json",
        "ui_decisions.json",
        "applied_decisions.json",
        "final_artifacts.json",
    }
    output_receipts = [
        receipt
        for receipt in receipt_bundle["output_receipts"]
        if str(receipt["path"]) not in mutable_control_paths
    ]
    roots = _artifact_roots(
        bank_path=bank_path,
        journal_path=journal_path,
        sample_path=sample_path,
        output_dir=output_dir,
    )
    envelope = build_assurance_envelope(
        run_id=run_id,
        workflow_id=ASSURANCE_WORKFLOW_ID,
        workflow_version=ASSURANCE_WORKFLOW_VERSION,
        artifact_receipts=[
            *receipt_bundle["source_receipts"],
            *implementation_receipts,
            *output_receipts,
        ],
        implementation_artifact_refs=[
            receipt["artifact_id"] for receipt in implementation_receipts
        ],
        reviewed_decisions=reviewed_decisions,
        source_qualifications=qualifications,
        allocation_ledgers=[],
        numeric_evidence_ledgers=[],
        gate_register=gates,
        limitations=list(limitations),
        artifact_roots=roots,
    )
    validate_exact_implementation_receipts(envelope, artifact_roots=roots)
    validate_assurance_envelope(envelope, artifact_roots=roots)
    write_assurance_json(path, envelope)
    return envelope


def run_reconciliation(
    bank_path: Path,
    journal_path: Path,
    output_dir: Path,
    recipe_path: Path | None = None,
    *,
    sample_path: Path | None = None,
    tolerance: object = "1",
    date_window_days: int = 7,
    language: object | None = None,
    document_language: object | None = None,
    client_run_id: str | None = None,
    client_run_root: Path | None = None,
) -> ReconciliationRunResult:
    """Run exact, policy-bound journal-to-bank reconciliation."""

    recipe = read_json(recipe_path)
    tolerance_value, tolerance_text = _canonical_tolerance(tolerance)
    if (
        not isinstance(date_window_days, int)
        or isinstance(date_window_days, bool)
        or date_window_days < 0
    ):
        raise ValueError("date_window_days must be a non-negative integer")
    languages = language_assumptions(
        recipe, language=language, document_language=document_language
    )
    output_dir.mkdir(parents=True, exist_ok=True)
    managed_run_root = (
        client_run_root.expanduser().resolve() if client_run_root is not None else None
    )

    def output_reference(path: Path) -> str:
        if managed_run_root is None:
            return path.as_posix()
        try:
            relative = path.expanduser().resolve().relative_to(managed_run_root)
        except ValueError as exc:
            raise ValueError(
                "Journal-Bank output leaves the managed customer run."
            ) from exc
        if not relative.parts:
            raise ValueError("Journal-Bank output must identify a run artifact.")
        return relative.as_posix()

    run_intake = write_run_intake(
        output_dir,
        bank_path=bank_path,
        journal_path=journal_path,
        recipe_path=recipe_path,
        sample_path=sample_path,
        language=languages["language"],
        document_language=languages["document_language"],
        tolerance=tolerance_text,
        date_window_days=date_window_days,
        client_run_id=client_run_id,
        client_run_root=client_run_root,
    )
    recorded_intake = read_json(run_intake.path)
    recorded_assumptions = recorded_intake.get("assumptions")
    if not isinstance(recorded_assumptions, dict):
        raise ValueError("Journal-Bank run intake assumptions are unavailable")
    recorded_bank_path = str(recorded_assumptions["bank_path"])
    recorded_journal_path = str(recorded_assumptions["journal_path"])
    recorded_sample_path = recorded_assumptions.get("sample_path")
    initial_source_receipts, source_refs = _source_artifact_receipts(
        (("bank", bank_path), ("journal", journal_path), ("sample", sample_path))
    )
    bank, bank_diag = _normalize_files(
        bank_path, "bank", recipe, source_refs=source_refs
    )
    journal, journal_diag = _normalize_files(
        journal_path, "journal", recipe, source_refs=source_refs
    )
    sample_movements, sample_diagnostics = _read_sample_movements(sample_path)
    source_receipts, current_source_refs = _source_artifact_receipts(
        (("bank", bank_path), ("journal", journal_path), ("sample", sample_path))
    )
    changed_sources = _source_snapshot_changes(
        initial_source_receipts,
        source_receipts,
    )
    source_refs = current_source_refs
    if changed_sources:
        bank_diag = _mark_changed_source_diagnostics("bank", bank_diag, changed_sources)
        journal_diag = _mark_changed_source_diagnostics(
            "journal", journal_diag, changed_sources
        )
        sample_diagnostics = _mark_changed_sample_diagnostics(
            sample_diagnostics,
            changed_sources,
        )
        if any(root_id == "source_bank" for root_id, _ in changed_sources):
            bank = _transaction_frame([])
        if any(root_id == "source_journal" for root_id, _ in changed_sources):
            journal = _transaction_frame([])
        if any(root_id == "source_sample" for root_id, _ in changed_sources):
            sample_movements = set()
    bank_pdf_non_movements = _non_movement_frame(_non_movement_records(bank_diag))
    qualifications = [
        *_source_qualifications("bank", bank_diag, source_refs),
        *_source_qualifications("journal", journal_diag, source_refs),
    ]
    qualification_status = _aggregate_qualification_status(qualifications)
    if changed_sources:
        qualification_status = "unsupported_source_layout"
    qualification_refs = [str(item["qualification_id"]) for item in qualifications]
    bank_journal_source_refs = [
        str(receipt["artifact_id"])
        for receipt in source_receipts
        if str(receipt["artifact_id"]).startswith(("source.bank.", "source.journal."))
    ]
    relationship_policy, relationship_decision, relationship_error = (
        _validated_relationship_decision(
            recipe,
            source_artifact_refs=bank_journal_source_refs,
            tolerance_text=tolerance_text,
            date_window_days=date_window_days,
        )
    )
    reviewed_decisions = _mapping_decisions([*bank_diag, *journal_diag])
    if relationship_decision is not None:
        reviewed_decisions.append(relationship_decision)

    paths = {
        "normalized_bank_csv": output_dir / "normalized_bank.csv",
        "normalized_journal_csv": output_dir / "normalized_journal.csv",
        "reconciliation_matches_csv": output_dir / "reconciliation_matches.csv",
        "relationship_residuals_csv": output_dir / "relationship_residuals.csv",
        "unmatched_bank_csv": output_dir / "unmatched_bank.csv",
        "unmatched_journal_csv": output_dir / "unmatched_journal.csv",
        "bank_pdf_non_movement_rows_csv": output_dir / "bank_pdf_non_movement_rows.csv",
        "workbook_xlsx": output_dir / "journal_bank_reconciliation.xlsx",
        "audit_json": output_dir / "reconciliation_audit.json",
        "review_notes_md": output_dir / "review_notes.md",
        "input_receipts_json": output_dir / "input_receipts.json",
        "source_qualifications_json": output_dir / "source_qualifications.json",
        "reviewed_decisions_json": output_dir / "reviewed_decisions.json",
        "lineage_json": output_dir / "lineage.json",
        "relationship_ledger_json": output_dir / "relationship_ledger.json",
        "material_value_ledger_json": output_dir / "material_value_ledger.json",
        "assurance_gates_json": output_dir / "assurance_gates.json",
        "artifact_receipts_json": output_dir / "artifact_receipts.json",
        "assurance_envelope_json": output_dir / "assurance_envelope.json",
        "run_intake_json": output_dir / "run_intake.json",
        "review_payload_json": output_dir / "review_payload.json",
        "ui_decisions_json": output_dir / "ui_decisions.json",
        "final_artifacts_json": output_dir / "final_artifacts.json",
    }
    write_assurance_json(
        paths["input_receipts_json"],
        {
            "schema_version": "journal_bank.input_receipts.v1",
            "receipts": source_receipts,
        },
    )
    write_assurance_json(
        paths["source_qualifications_json"],
        {
            "schema_version": "journal_bank.source_qualifications.v1",
            "status": qualification_status,
            "source_outcomes": _source_outcomes(bank_diag, journal_diag),
            "qualifications": qualifications,
        },
    )
    write_assurance_json(
        paths["reviewed_decisions_json"],
        {
            "schema_version": "journal_bank.reviewed_decisions.v1",
            "decisions": reviewed_decisions,
        },
    )
    bank_pdf_non_movements.write_csv(paths["bank_pdf_non_movement_rows_csv"])

    blocked_code: str | None = None
    blocked_detail: str | None = None
    source_gate_status = "passed"
    preparation_gate_status = "passed"
    if changed_sources:
        blocked_code = "source_changed_during_run"
        blocked_detail = (
            "Source bytes or source membership changed between the pre-parser and "
            "post-parser snapshots; all prepared rows from affected inputs were "
            "discarded."
        )
        source_gate_status = "failed"
        preparation_gate_status = "blocked"
    elif qualification_status == "unsupported_source_layout":
        parser_failed = any(
            item.get("failure_kind") == "parser_failure"
            for item in (*bank_diag, *journal_diag)
        )
        blocked_code = (
            "parser_failure" if parser_failed else "unsupported_source_layout"
        )
        blocked_detail = (
            "At least one source could not be parsed."
            if parser_failed
            else "At least one source lacks a supported bounded layout adapter."
        )
        source_gate_status = "failed"
        preparation_gate_status = "blocked"
    elif qualification_status == "needs_review":
        blocked_code = "mapping_review_required"
        blocked_detail = "At least one source requires a current reviewed mapping."
        source_gate_status = "blocked"
        preparation_gate_status = "blocked"
    elif sample_path is not None and not sample_movements:
        blocked_code = "invalid_or_empty_sample"
        blocked_detail = (
            "A sample was supplied but its exact identifier adapter did not yield "
            "a complete non-empty movement set."
        )
        preparation_gate_status = "blocked"
    elif bank.is_empty() or journal.is_empty():
        blocked_code = "empty_prepared_population"
        blocked_detail = (
            "A qualified source emitted no bank or journal monetary movements."
        )
        preparation_gate_status = "failed"
    elif relationship_error is not None:
        blocked_code = "relationship_review_required"
        blocked_detail = relationship_error

    filtered_journal = journal
    if blocked_code is None and sample_path is not None:
        filtered_journal = _filter_journal_by_sample(journal, sample_movements)
        if filtered_journal.is_empty():
            blocked_code = "sample_not_found_in_journal"
            blocked_detail = (
                "The supplied sample identifiers do not select any journal movement."
            )
            preparation_gate_status = "failed"

    if blocked_code is None and relationship_policy is not None:
        bank = _with_policy_defaults(bank, relationship_policy)
        filtered_journal = _with_policy_defaults(filtered_journal, relationship_policy)
        perimeter_error = _relationship_perimeter_error(
            bank, filtered_journal, relationship_policy
        )
        if perimeter_error is not None:
            blocked_code = "relationship_perimeter_incomplete"
            blocked_detail = perimeter_error

    write_assurance_json(
        paths["lineage_json"],
        _lineage_payload(bank, filtered_journal, source_refs),
    )
    if blocked_code is not None and blocked_detail is not None:
        gates = _gate_payload(
            source_status=source_gate_status,
            source_refs=qualification_refs,
            preparation_status=preparation_gate_status,
            reconciliation_status="blocked",
            limitation=blocked_detail,
        )
        write_assurance_json(paths["assurance_gates_json"], gates)
        blocked_matches = pl.DataFrame(
            schema={column: pl.Utf8 for column in MATCH_COLUMNS}
        )
        blocked_residuals = pl.DataFrame(
            schema={column: pl.Utf8 for column in RESIDUAL_COLUMNS}
        )
        unmatched_bank = bank
        unmatched_journal = filtered_journal
        bank.write_csv(paths["normalized_bank_csv"])
        filtered_journal.write_csv(paths["normalized_journal_csv"])
        blocked_matches.write_csv(paths["reconciliation_matches_csv"])
        blocked_residuals.write_csv(paths["relationship_residuals_csv"])
        unmatched_bank.write_csv(paths["unmatched_bank_csv"])
        unmatched_journal.write_csv(paths["unmatched_journal_csv"])
        _write_workbook(
            paths["workbook_xlsx"],
            {
                "matches": blocked_matches,
                "relationship_residuals": blocked_residuals,
                "unmatched_bank": unmatched_bank,
                "unmatched_journal": unmatched_journal,
                "bank_pdf_non_movements": bank_pdf_non_movements,
                "normalized_bank": bank,
                "normalized_journal": filtered_journal,
            },
        )
        write_assurance_json(
            paths["relationship_ledger_json"],
            _blocked_relationship_ledger(
                block_code=blocked_code,
                block_detail=blocked_detail,
            ),
        )
        blocked_audit = {
            **languages,
            "schema_version": "journal_bank.reconciliation_audit.v3",
            "status": "blocked",
            "block_code": blocked_code,
            "block_detail": blocked_detail,
            "bank_path": recorded_bank_path,
            "journal_path": recorded_journal_path,
            "sample_path": recorded_sample_path,
            "sample_movement_count": len(sample_movements),
            "sample_diagnostics": sample_diagnostics,
            "source_snapshot_changed": bool(changed_sources),
            "changed_sources": [
                {"root_id": root_id, "path": source_path}
                for root_id, source_path in sorted(changed_sources)
            ],
            "bank_row_count": bank.height,
            "journal_row_count": filtered_journal.height,
            "matched_count": 0,
            "unmatched_bank_count": unmatched_bank.height,
            "unmatched_journal_count": unmatched_journal.height,
            "bank_pdf_non_movement_row_count": bank_pdf_non_movements.height,
            "bank_pdf_non_movement_classifications": _count_classifications(
                bank_pdf_non_movements.to_dicts()
            ),
            "stage_counts": {},
            "tolerance": tolerance_text,
            "date_window_days": date_window_days,
            "source_qualification_status": qualification_status,
            "relationship_decision_ref": (
                str(relationship_decision["decision_id"])
                if relationship_decision is not None
                else None
            ),
            "relationship_balanced": False,
            "relationship_within_policy_tolerance": False,
            "relationship_residual_row_count": 0,
            "diagnostics": {"bank": bank_diag, "journal": journal_diag},
            "outputs": {
                key: output_reference(value)
                for key, value in paths.items()
                if key != "material_value_ledger_json"
            },
        }
        write_assurance_json(paths["audit_json"], blocked_audit)
        _write_review_notes(paths["review_notes_md"], blocked_audit)
        _write_output_receipts(
            output_dir,
            paths,
            source_receipts=source_receipts,
        )
        review_session = write_review_session_artifacts(
            output_dir,
            run_id=run_intake.run_id,
            run_intake_path=run_intake.path,
            matches=blocked_matches,
            unmatched_bank=unmatched_bank,
            unmatched_journal=unmatched_journal,
            bank_pdf_non_movements=bank_pdf_non_movements,
            relationship_residuals=blocked_residuals,
            audit=blocked_audit,
        )
        blocked_audit["review_session"] = {
            "run_id": review_session.run_id,
            "run_intake_path": output_reference(review_session.run_intake_path),
            "review_payload_path": output_reference(review_session.review_payload_path),
            "ui_decisions_path": output_reference(review_session.ui_decisions_path),
            "final_artifacts_path": output_reference(
                review_session.final_artifacts_path
            ),
            "review_item_count": review_session.review_item_count,
        }
        write_assurance_json(paths["audit_json"], blocked_audit)
        receipt_bundle = _write_output_receipts(
            output_dir,
            paths,
            source_receipts=source_receipts,
        )
        _write_assurance_envelope(
            run_id=run_intake.run_id,
            bank_path=bank_path,
            journal_path=journal_path,
            sample_path=sample_path,
            output_dir=output_dir,
            path=paths["assurance_envelope_json"],
            receipt_bundle=receipt_bundle,
            reviewed_decisions=reviewed_decisions,
            qualifications=qualifications,
            gates=gates,
            limitations=[blocked_detail],
        )
        _close_final_artifacts(
            output_dir,
            paths,
            source_receipts=source_receipts,
        )
        raise ReconciliationBlockedError(blocked_code, blocked_detail)

    if relationship_policy is None:  # pragma: no cover - guarded by block handling
        raise RuntimeError("Validated relationship policy unexpectedly unavailable.")
    matches, unmatched_bank, unmatched_journal, stage_counts = _match_transactions(
        bank,
        filtered_journal,
        tolerance=tolerance_value,
        date_window_days=date_window_days,
        relationship_policy=relationship_policy,
    )
    relationship_ledger = _relationship_ledger(
        bank,
        filtered_journal,
        matches,
        policy=relationship_policy,
    )
    write_assurance_json(paths["relationship_ledger_json"], relationship_ledger)
    relationship_residuals = _relationship_residual_frame(
        relationship_ledger,
        bank,
        filtered_journal,
    )
    relationship_exactly_balanced = _ledger_has_zero_residuals(relationship_ledger)
    reconciliation_status = (
        "passed"
        if relationship_ledger["balanced"]
        and relationship_exactly_balanced
        and unmatched_bank.is_empty()
        and unmatched_journal.is_empty()
        else "withheld"
    )
    unresolved_detail = (
        None
        if reconciliation_status == "passed"
        else "Unmatched rows or exact residuals remain unresolved."
    )
    gates = _gate_payload(
        source_status="passed",
        source_refs=qualification_refs,
        preparation_status="passed",
        reconciliation_status=reconciliation_status,
        reconciliation_ref="output.relationship_ledger_json",
        limitation=unresolved_detail,
    )
    write_assurance_json(paths["assurance_gates_json"], gates)

    bank.write_csv(paths["normalized_bank_csv"])
    filtered_journal.write_csv(paths["normalized_journal_csv"])
    matches.write_csv(paths["reconciliation_matches_csv"])
    relationship_residuals.write_csv(paths["relationship_residuals_csv"])
    unmatched_bank.write_csv(paths["unmatched_bank_csv"])
    unmatched_journal.write_csv(paths["unmatched_journal_csv"])
    workbook_frames = {
        "matches": matches,
        "relationship_residuals": relationship_residuals,
        "unmatched_bank": unmatched_bank,
        "unmatched_journal": unmatched_journal,
        "bank_pdf_non_movements": bank_pdf_non_movements,
        "normalized_bank": bank,
        "normalized_journal": filtered_journal,
    }
    _write_workbook(
        paths["workbook_xlsx"],
        {name: workbook_frames[name] for name in WORKBOOK_SHEET_ORDER},
    )
    audit = {
        **languages,
        "schema_version": "journal_bank.reconciliation_audit.v3",
        "status": (
            "completed_pending_review"
            if reconciliation_status == "passed"
            else "completed_with_unresolved_reconciliation"
        ),
        "bank_path": recorded_bank_path,
        "journal_path": recorded_journal_path,
        "sample_path": recorded_sample_path,
        "sample_movement_count": len(sample_movements),
        "sample_diagnostics": sample_diagnostics,
        "source_snapshot_changed": bool(changed_sources),
        "changed_sources": [
            {"root_id": root_id, "path": source_path}
            for root_id, source_path in sorted(changed_sources)
        ],
        "bank_row_count": bank.height,
        "journal_row_count": filtered_journal.height,
        "matched_count": matches.height,
        "unmatched_bank_count": unmatched_bank.height,
        "unmatched_journal_count": unmatched_journal.height,
        "bank_pdf_non_movement_row_count": bank_pdf_non_movements.height,
        "bank_pdf_non_movement_classifications": _count_classifications(
            bank_pdf_non_movements.to_dicts()
        ),
        "stage_counts": stage_counts,
        "tolerance": tolerance_text,
        "date_window_days": date_window_days,
        "source_qualification_status": qualification_status,
        "relationship_decision_ref": str(relationship_decision["decision_id"]),
        "relationship_balanced": relationship_exactly_balanced,
        "relationship_within_policy_tolerance": bool(relationship_ledger["balanced"]),
        "relationship_residual_row_count": relationship_residuals.height,
        "diagnostics": {"bank": bank_diag, "journal": journal_diag},
        "outputs": {key: output_reference(value) for key, value in paths.items()},
    }
    write_json(paths["audit_json"], audit)
    _write_review_notes(paths["review_notes_md"], audit)
    material_value_ledger = _build_material_value_ledger(
        output_dir,
        matches,
        relationship_residuals,
    )
    write_json(paths["material_value_ledger_json"], material_value_ledger)
    validate_material_value_ledger(output_dir)
    _write_output_receipts(
        output_dir,
        paths,
        source_receipts=source_receipts,
    )
    review_session = write_review_session_artifacts(
        output_dir,
        run_id=run_intake.run_id,
        run_intake_path=run_intake.path,
        matches=matches,
        unmatched_bank=unmatched_bank,
        unmatched_journal=unmatched_journal,
        bank_pdf_non_movements=bank_pdf_non_movements,
        relationship_residuals=relationship_residuals,
        audit=audit,
    )
    audit["review_session"] = {
        "run_id": review_session.run_id,
        "run_intake_path": output_reference(review_session.run_intake_path),
        "review_payload_path": output_reference(review_session.review_payload_path),
        "ui_decisions_path": output_reference(review_session.ui_decisions_path),
        "final_artifacts_path": output_reference(review_session.final_artifacts_path),
        "review_item_count": review_session.review_item_count,
    }
    write_json(paths["audit_json"], audit)
    receipt_bundle = _write_output_receipts(
        output_dir,
        paths,
        source_receipts=source_receipts,
    )
    _write_assurance_envelope(
        run_id=run_intake.run_id,
        bank_path=bank_path,
        journal_path=journal_path,
        sample_path=sample_path,
        output_dir=output_dir,
        path=paths["assurance_envelope_json"],
        receipt_bundle=receipt_bundle,
        reviewed_decisions=reviewed_decisions,
        qualifications=qualifications,
        gates=gates,
        limitations=(
            []
            if reconciliation_status == "passed"
            else [unresolved_detail or "Reconciliation assurance remains withheld."]
        ),
    )
    _close_final_artifacts(
        output_dir,
        paths,
        source_receipts=source_receipts,
    )
    return ReconciliationRunResult(
        matches=matches,
        unmatched_bank=unmatched_bank,
        unmatched_journal=unmatched_journal,
        audit=audit,
    )


def add_common_args(parser: argparse.ArgumentParser) -> None:
    parser.add_argument(
        "--language",
        help="Working/output language locale: it, en, fr, de, or es. Defaults to recipe or en.",
    )
    parser.add_argument(
        "--document-language",
        help="Source-document language locale: it, en, fr, de, es, or auto. Defaults to recipe or auto.",
    )
    parser.add_argument("--verbose", action="store_true", help="Enable debug logging.")
